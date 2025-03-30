import logging
import traceback
import sys
import svgwrite
import tempfile
from tempfile import NamedTemporaryFile
import base64
from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
from io import BytesIO
import io
from PyQt5.QtWidgets import (
    QDesktopWidget, QTableWidget, QTableWidgetItem, QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy, QMainWindow, QApplication, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit, QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog, QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox
)
from PyQt5.QtGui import QPixmap, QKeySequence, QImage, QPolygonF,QPainter, QColor, QFont, QKeySequence, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF,QUrl
import json
import os
import numpy as np
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font
# import ctypes

from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QSlider,QMenuBar, QMenu, QAction
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from skimage.restoration import rolling_ball 
from scipy.signal import find_peaks
from skimage import filters, morphology
from scipy.ndimage import gaussian_filter1d

# Configure logging to write errors to a log file
logging.basicConfig(
    filename="error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    # Log the exception
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))

    # Display a QMessageBox with the error details
    error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}"
    QMessageBox.critical(
        None,  # No parent window
        "Unexpected Error",  # Title of the message box
        error_message,  # Error message to display
        QMessageBox.Ok  # Button to close the dialog
    )
    

# # Set the custom exception handler
sys.excepthook = log_exception

class TableWindow(QDialog):
    def __init__(self, peak_areas, standard_dictionary, standard, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Table Export")
        self.setGeometry(100, 100, 600, 400)

        # Create a scroll area
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)  # Allow the table to resize within the scroll area

        # Create a table widget
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)  # Enable row selection
        self.table.setSelectionMode(QTableWidget.ContiguousSelection)  # Allow copying multiple rows

        # Populate the table with peak areas and percentages
        self.populate_table(peak_areas, standard_dictionary, standard)

        # Set the table as the widget for the scroll area
        self.scroll_area.setWidget(self.table)

        # Add an "Export to Excel" button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)  # Add the scroll area to the layout
        layout.addWidget(self.export_button)  # Add the export button
        self.setLayout(layout)

    def populate_table(self, peak_areas, standard_dictionary, standard):
        # Calculate the total sum of peak areas
        total_area = sum(peak_areas)

        # Set the number of rows in the table
        self.table.setRowCount(len(peak_areas))

        # Populate the table with band labels, peak areas, and percentages
        for row, area in enumerate(peak_areas):
            # Bang label (e.g., "Band 1", "Band 2", etc.)
            band_label = f"Band {row + 1}"
            self.table.setItem(row, 0, QTableWidgetItem(band_label))

            # Peak area (rounded to 3 decimal places)
            peak_area_rounded = round(area, 3)
            self.table.setItem(row, 1, QTableWidgetItem(str(peak_area_rounded)))

            # Percentage (rounded to 2 decimal places)
            if total_area != 0:  # Avoid division by zero
                percentage = (area / total_area) * 100
                percentage_rounded = round(percentage, 2)
            else:
                percentage_rounded = 0.0
            self.table.setItem(row, 2, QTableWidgetItem(f"{percentage_rounded:.2f}%"))

            # Quantity (if standard is True)
            if standard:
                std_peak_area = list(standard_dictionary.values())
                known_quantities = list(standard_dictionary.keys())
                coefficients = np.polyfit(std_peak_area, known_quantities, 1)
                unknown_quantity = np.polyval(coefficients, area)
                self.table.setItem(row, 3, QTableWidgetItem(f"{unknown_quantity:.2f}"))

        # Enable copying data from the table
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)  # Make the table read-only
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)  # Enable right-click context menu
        self.table.customContextMenuRequested.connect(self.copy_table_data)

    def copy_table_data(self):
        """Copy selected table data to the clipboard."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # Get the selected rows and columns
        rows = set(item.row() for item in selected_items)
        cols = set(item.column() for item in selected_items)

        # Prepare the data for copying
        data = []
        for row in sorted(rows):
            row_data = []
            for col in sorted(cols):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append("\t".join(row_data))

        # Copy the data to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(data))

    def export_to_excel(self):
        """Export the table data to an Excel file."""
        # Prompt the user to select a save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
        )
        if not file_path:
            return

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Peak Areas"

        # Write the table headers to the Excel sheet
        headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
            worksheet.cell(row=1, column=col).font = Font(bold=True)  # Make headers bold

        # Write the table data to the Excel sheet
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                worksheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

        # Save the Excel file
        workbook.save(file_path)
        QMessageBox.information(self, "Success", f"Table data exported to {file_path}")
                

class PeakAreaDialog(QDialog):
    """Interactive dialog to adjust peak regions and calculate peak areas."""
    def __init__(self,  cropped_image, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Adjust Peak Regions")
        self.setGeometry(100, 100, 1000, 800)  # Larger window for multiple sliders    
        # Store the original intensity profile
        self.cropped_image = cropped_image
        self.cropped_image = self.cropped_image.convert("L")  # Ensure grayscale mode
        self.profile = None
        self.background=None
        width, height = self.cropped_image.size
        self.intensity_array = np.array(self.cropped_image, dtype=np.uint8).reshape((height, width))
        self.rolling_ball_radius = 5000  # Default rolling ball radius
        self.peaks = []  
        self.peak_regions = []  
        self.peak_areas_rolling_ball = []  # Peak areas using rolling ball method
        self.peak_areas_straight_line = []  # Peak areas using straight-line method
        self.peak_sliders = []
        self.method = "Rolling Ball"  # Default method
        self.peak_height=0.1
        self.peak_distance=30
    
        # Create the main layout
        main_layout = QVBoxLayout(self)
    
        # Add the plot at the top
        self.fig, self.ax = plt.subplots(figsize=(10, 5))
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setMinimumSize(900, 400)
        main_layout.addWidget(self.canvas)
    
        # Add the rolling ball slider and OK button at the top
        top_layout = QHBoxLayout()
    
        # Band Estimation Combobox (FIX for AttributeError)
        self.band_estimation_combobox = QComboBox()
        self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
        self.band_estimation_combobox.setCurrentText("Mean")  
        self.band_estimation_combobox.currentIndexChanged.connect(self.update_peak_number)
        top_layout.addWidget(QLabel("Band Estimation Technique:"))
        top_layout.addWidget(self.band_estimation_combobox)

        # Rolling Ball Slider
        self.rolling_ball_label = QLabel("Rolling Ball Radius (0.000)")
        self.rolling_ball_label.setFixedWidth(180)
        self.rolling_ball_slider = QSlider(Qt.Horizontal)
        self.rolling_ball_slider.setRange(1, 50000)  
        self.rolling_ball_slider.setValue(self.rolling_ball_radius)
        self.rolling_ball_slider.valueChanged.connect(self.update_plot)
        top_layout.addWidget(self.rolling_ball_label)
        top_layout.addWidget(self.rolling_ball_slider)

        # Peak Area Calculation Method
        self.method_combobox = QComboBox()
        self.method_combobox.addItems(["Rolling Ball", "Straight Line"])
        self.method_combobox.currentIndexChanged.connect(self.update_plot)
        top_layout.addWidget(QLabel("Peak Area Calculation:"))
        top_layout.addWidget(self.method_combobox)
    
        # OK button
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        top_layout.addWidget(self.ok_button)
    
        main_layout.addLayout(top_layout)
    
        # Add a text box and update button for manual peak number adjustment
        peak_number_layout = QHBoxLayout()
        self.peak_number_label = QLabel("Number of Peaks:")
        self.peak_number_input = QLineEdit()
        self.peak_number_input.setPlaceholderText("Enter number of peaks")
        
        
        self.peak_height_slider_label = QLabel("Peak Height")
        self.peak_height_slider = QSlider(Qt.Horizontal)
        self.peak_height_slider.setRange(1, 200)  
        self.peak_height_slider.setValue(int(self.peak_height*100))
        self.peak_height_slider.valueChanged.connect(self.detect_peaks)
        
        self.peak_distance_slider_label = QLabel("Peak Distance")
        self.peak_distance_slider = QSlider(Qt.Horizontal)
        self.peak_distance_slider.setRange(1, 100)  
        self.peak_distance_slider.setValue(int(self.peak_distance))
        self.peak_distance_slider.valueChanged.connect(self.detect_peaks)
        
        self.update_peak_number_button = QPushButton("Update")
        self.update_peak_number_button.clicked.connect(self.update_peak_number)
    
        peak_number_layout.addWidget(self.peak_number_label)
        peak_number_layout.addWidget(self.peak_number_input)
        peak_number_layout.addWidget(self.update_peak_number_button)
        peak_number_layout.addWidget(self.peak_height_slider_label)
        peak_number_layout.addWidget(self.peak_height_slider)
        peak_number_layout.addWidget(self.peak_distance_slider_label)
        peak_number_layout.addWidget(self.peak_distance_slider)
        
        main_layout.addLayout(peak_number_layout)
    
        # Create a scrollable area for the peak sliders
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)  # Allow the widget to resize
    
        # Create a container widget for the scroll area
        self.container = QWidget()
        self.peak_sliders_layout = QVBoxLayout(self.container)
    
        # Get the selected band estimation technique
        technique = self.band_estimation_combobox.currentText()
        
        
        # **1. Apply Rolling Ball Background Subtraction (Like ImageJ)**
        # background = rolling_ball(intensity_array, radius=50)  # Radius is adjustable
        # corrected = intensity_array - background
        # corrected[corrected < 0] = 0  # Clip negative values
    
        if technique == "Mean":
            self.profile = np.mean(self.intensity_array, axis=1)
        elif technique == "Percentile:5%":
            self.profile = np.percentile(self.intensity_array, 5, axis=1)
        elif technique == "Percentile:10%":
            self.profile = np.percentile(self.intensity_array, 10, axis=1)
        elif technique == "Percentile:15%":
            self.profile = np.percentile(self.intensity_array, 15, axis=1)
        elif technique == "Percentile:30%":
            self.profile = np.percentile(self.intensity_array, 30, axis=1)
        else:
            self.profile = np.percentile(self.intensity_array, 5, axis=1)  # Default to Percentile:5%

        self.profile = (self.profile - np.min(self.profile)) /np.ptp(self.profile) * 255
        self.profile = 255 - self.profile  # Invert the profile for peak detection
        self.profile = gaussian_filter1d(self.profile, sigma=2)  # Smoothing filter
    
        self.detect_peaks()
    
        # Add sliders for each peak
        self.update_sliders()
        
        self.peak_sliders_layout.addStretch()
    
        # Set the container widget as the scroll area's widget
        scroll_area.setWidget(self.container)
    
        # Add the scroll area to the main layout
        main_layout.addWidget(scroll_area)
    
        self.update_plot()  # Initial plot rendering
    
    def update_peak_number(self):
        """Update the number of peaks based on user input."""
        # Get the selected band estimation technique
        technique = self.band_estimation_combobox.currentText()
        
        if technique == "Mean":
            self.profile = np.mean(self.intensity_array, axis=1)
        elif technique == "Percentile:5%":
            self.profile = np.percentile(self.intensity_array, 5, axis=1)
        elif technique == "Percentile:10%":
            self.profile = np.percentile(self.intensity_array, 10, axis=1)
        elif technique == "Percentile:15%":
            self.profile = np.percentile(self.intensity_array, 15, axis=1)
        elif technique == "Percentile:30%":
            self.profile = np.percentile(self.intensity_array, 30, axis=1)
        else:
            self.profile = np.percentile(self.intensity_array, 5, axis=1)  # Default to Percentile:5%

        self.profile = (self.profile - np.min(self.profile)) /np.ptp(self.profile) * 255
        self.profile = 255 - self.profile  # Invert the profile for peak detection
        self.profile = gaussian_filter1d(self.profile, sigma=2)  # Smoothing filter
        
        try:
            num_peaks = int(self.peak_number_input.text())
            if num_peaks <= 0:
                raise ValueError("Number of peaks must be positive.")
            
            # Convert self.peaks to a list if it's a NumPy array
            if isinstance(self.peaks, np.ndarray):
                self.peaks = self.peaks.tolist()
            
            # If the user enters fewer peaks than detected, truncate the list
            if num_peaks < len(self.peaks):
                self.peaks = self.peaks[:num_peaks]
                self.peak_regions = self.peak_regions[:num_peaks]
                # Update sliders
                self.update_sliders()
            # If the user enters more peaks, add dummy peaks at the end
            elif num_peaks > len(self.peaks):
                for _ in range(num_peaks - len(self.peaks)):
                    self.peaks.append(len(self.profile) // 2)  # Add a dummy peak in the middle
                    self.peak_regions.append((0, len(self.profile)))  # Default region
    
                # Update sliders
                self.update_sliders()
        except ValueError as e:
            QMessageBox.warning(self, "Error", f"Invalid input: {e}")
                    
        self.update_sliders()
        self.update_plot()
        
        
        
    
    def update_sliders(self):
        """Update the sliders based on the current peaks and peak regions."""
        # Clear existing sliders
        while self.peak_sliders_layout.count():
            item = self.peak_sliders_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()  # Properly delete widget
    
        self.peak_sliders.clear()
        
    
        # Add sliders for each peak
        for i, peak in enumerate(self.peaks):
            peak_group = QGroupBox(f"Peak {i + 1}")
            peak_layout = QVBoxLayout()
    
            # Start slider
            start_slider = QSlider(Qt.Horizontal)
            start_slider.setRange(0, len(self.profile))  # Range based on profile length
            start_slider.setValue(self.peak_regions[i][0])  # Set start position based on troughs
            start_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(QLabel("Start Position"))
            peak_layout.addWidget(start_slider)
    
            # End slider
            end_slider = QSlider(Qt.Horizontal)
            end_slider.setRange(0, len(self.profile))  # Range based on profile length
            end_slider.setValue(self.peak_regions[i][1])  # Set end position based on troughs
            end_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(QLabel("End Position"))
            peak_layout.addWidget(end_slider)
    
            peak_group.setLayout(peak_layout)
            self.peak_sliders_layout.addWidget(peak_group)
            self.peak_sliders.append((start_slider, end_slider))
        
        self.peak_sliders_layout.addStretch()

    def detect_peaks(self):
        """Detect peaks and troughs in the intensity profile."""
        self.peak_height=float(self.peak_height_slider.value()/100)
        self.peak_distance=int(self.peak_distance_slider.value())
        # Detect peaks with adjusted sensitivity
        peaks, _ = find_peaks(self.profile, height=np.mean(self.profile) * self.peak_height, distance=self.peak_distance)
        self.peaks = peaks
        
        # Calculate troughs between peaks
        troughs = []
        for i in range(len(peaks) - 1):
            start = peaks[i]
            end = peaks[i + 1]
            trough = start + np.argmin(self.profile[start:end])
            troughs.append(trough)
        
        # Set peak regions based on troughs
        self.peak_regions = []
        for i, peak in enumerate(peaks):
            if i == 0:
                start = max(0, peak - 50)  # Default start for the first peak
            else:
                start = troughs[i - 1]  # Start at the previous trough
        
            if i == len(peaks) - 1:
                end = min(len(self.profile), peak + 50)  # Default end for the last peak
            else:
                end = troughs[i]  # End at the next trough
        
            self.peak_regions.append((start, end))
        
        self.peak_areas = [0] * len(peaks)  # Initialize peak areas
        
        # Update sliders based on detected troughs
        for i, (start, end) in enumerate(self.peak_regions):
            if i < len(self.peak_sliders):
                start_slider, end_slider = self.peak_sliders[i]
                start_slider.setValue(start)
                end_slider.setValue(end)
        self.peak_number_input.setText(str(len(peaks)))
        self.peak_sliders.clear()
        self.update_plot()
        
    def update_plot(self):
        """Update the plot based on the selected method (Rolling Ball or Straight Line)."""
        if self.canvas is None:
            return  
        
    
        self.method = self.method_combobox.currentText()
    
        # Apply rolling ball background correction
        self.rolling_ball_radius = self.rolling_ball_slider.value() / 100
        label=f'Rolling Ball Radius ({self.rolling_ball_radius:.2f})'
        self.rolling_ball_label.setText(label)        
        background_fwd = rolling_ball(self.profile, radius=self.rolling_ball_radius)
        background_rev = rolling_ball(self.profile[::-1], radius=self.rolling_ball_radius)[::-1]
        self.background = np.minimum(background_fwd, background_rev)
    
        # Update peak regions and calculate peak areas
        self.peak_areas_rolling_ball.clear()
        self.peak_areas_straight_line.clear()
    
        self.fig.clf()
        
        # Create a grid layout for the plot and image
        grid = plt.GridSpec(2, 1, height_ratios=[3, 1])  # 3:1 ratio
    
        # Plot the intensity profile in the top subplot
        self.ax = self.fig.add_subplot(grid[0])
        self.ax.plot(self.profile, label="Raw Intensity", color="black", linestyle="-")
        self.ax.scatter(self.peaks, self.profile[self.peaks], color="red", label="Detected Peaks")
    
        for i, (start_slider, end_slider) in enumerate(self.peak_sliders):
            start = start_slider.value()
            end = end_slider.value()
            start = max(0, min(start, len(self.profile) - 1))
            end = max(0, min(end, len(self.profile) - 1))
    
            peak_region = self.profile[start:end + 1]
            baseline_region = self.background[start:end + 1]
    
            # Rolling Ball Area Calculation
            rolling_ball_area = np.trapz(peak_region - baseline_region)
            self.peak_areas_rolling_ball.append(rolling_ball_area)
    
            # Straight Line Baseline Calculation
            x_baseline = np.array([start, end])
            y_baseline = np.array([self.profile[start], self.profile[end]])
            x_fill = np.arange(start, end + 1)
            y_fill = np.interp(x_fill, x_baseline, y_baseline)
    
            straight_line_area = np.trapz(self.profile[start:end + 1] - y_fill)
            self.peak_areas_straight_line.append(straight_line_area)
    
            # Plot rolling ball baseline
            if self.method == "Rolling Ball":
                self.ax.plot(self.background, color="green", linestyle="--")
                self.ax.fill_between(range(start, end + 1), baseline_region, peak_region, color="yellow", alpha=0.3)
                self.ax.text((start + end) / 2, np.max(peak_region), f"Area: {rolling_ball_area:.2f}", ha="center", va="bottom")
                
            # Plot straight-line baseline
            elif self.method == "Straight Line":
                self.ax.plot(x_baseline, y_baseline, color="red", linestyle="--")
                self.ax.fill_between(x_fill, y_fill, self.profile[start:end + 1], color="cyan", alpha=0.3)
                self.ax.text((start + end) / 2, np.max(peak_region), f"Area: {straight_line_area:.2f}", ha="center", va="bottom")
            
            # Plot vertical markers
            self.ax.axvline(start, color="blue", linestyle="--")
            self.ax.axvline(end, color="orange", linestyle="--")
            
        # Add a subplot for the cropped image
        ax_image = self.fig.add_subplot(grid[1])
    
        # Align x-axis labels to match the image
        self.ax.xaxis.set_label_position('bottom')
        self.ax.xaxis.set_ticks_position('bottom')
        
        # Set the x-axis limits to match the extreme left and right of the image
        self.ax.set_xlim(0, len(self.profile))
    
        # Disable all ticks and labels for the image subplot
        ax_image.set_xticks([])
        ax_image.set_yticks([])
        ax_image.set_xticklabels([])
        ax_image.set_yticklabels([])
        
    
        # Display the cropped image if available
        if self.cropped_image:
            rotated_pil_image = self.cropped_image.rotate(90, expand=True)
            ax_image.imshow(rotated_pil_image, cmap='gray', aspect='auto')
            ax_image.set_xlabel('Pixel Row')
            
            
        self.ax.set_ylabel("Intensity")
        self.ax.legend()
        self.ax.set_title("Peak Area Calculation")
    
        self.canvas.draw()
        plt.close(self.fig)

    def get_final_peak_area(self):
        """Return the calculated peak areas based on selected method."""
        if self.method == "Rolling Ball":
            return self.peak_areas_rolling_ball
        else:
            return self.peak_areas_straight_line
    


class LiveViewLabel(QLabel):
    def __init__(self, font_type, font_size, marker_color, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)  # Enable mouse tracking
        self.preview_marker_enabled = False
        self.preview_marker_text = ""
        self.preview_marker_position = None
        self.marker_font_type = font_type
        self.marker_font_size = font_size
        self.marker_color = marker_color
        self.setFocusPolicy(Qt.StrongFocus)
        self.bounding_box_preview = []
        self.measure_quantity_mode = False
        self.counter = 0
        self.zoom_level = 1.0  # Initial zoom level
        self.pan_start = None  # Start position for panning
        self.pan_offset = QPointF(0, 0)  # Offset for panning
        self.quad_points = []  # Stores 4 points of the quadrilateral
        self.selected_point = -1  # Index of selected corner (-1 = none)
        self.drag_threshold = 10  # Pixel radius for selecting corners
        self.bounding_box_complete = False
        self.mode=None
        # Add rectangle-related attributes
        self.rectangle_start = None  # Start point of the rectangle
        self.rectangle_end = None    # End point of the rectangle
        self.rectangle_points = []   # Stores the rectangle points
        self.drag_start_pos = None  # For tracking drag operations
        self.draw_edges=True

    def mouseMoveEvent(self, event):
        if self.preview_marker_enabled:
            self.preview_marker_position = event.pos()
            self.update()  # Trigger repaint to show the preview
        if self.selected_point != -1 and self.measure_quantity_mode:# and self.mode=="quad":
            # Update dragged corner position
            self.quad_points[self.selected_point] = self.transform_point(event.pos())
            self.update()  # Show the bounding box preview
        if self.selected_point != -1 and self.mode=="move":
            self.drag_start_pos=event.pos()
        
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if self.preview_marker_enabled:
            # Place the marker text permanently at the clicked position
            parent = self.parent()
            parent.place_custom_marker(event, self.preview_marker_text)
            self.update()  # Clear the preview
        if self.measure_quantity_mode and self.mode=="quad":
            # Check if clicking near existing corner
            for i, p in enumerate(self.quad_points):
                if (self.transform_point(event.pos()) - p).manhattanLength() < self.drag_threshold:
                    self.selected_point = i
                    return
    
            # Add new point if < 4 corners
            if len(self.quad_points) < 4:
                self.quad_points.append(self.transform_point(event.pos()))
                self.selected_point = len(self.quad_points) - 1
    
            if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                # self.quad_points = [self.transform_point(p) for p in self.quad_points]
                self.bounding_box_complete = True
            self.update()  # Trigger repaint
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode=="quad":
            self.selected_point = -1
            self.update()
        super().mouseReleaseEvent(event)

    def transform_point(self, point):
        """Transform a point from widget coordinates to image coordinates."""
        if self.zoom_level != 1.0:
            return QPointF(
                (point.x() - self.pan_offset.x()) / self.zoom_level,
                (point.y() - self.pan_offset.y()) / self.zoom_level
            )
        return QPointF(point)

    def zoom_in(self):
        self.zoom_level *= 1.1
        self.update()

    def zoom_out(self):
        self.zoom_level /= 1.1
        if self.zoom_level < 1.0:
            self.zoom_level = 1.0
            self.pan_offset = QPointF(0, 0)  # Reset pan offset when zoom is reset
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)

        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)

        # Draw the preview marker if enabled
        if self.preview_marker_enabled and self.preview_marker_position:
            painter.setOpacity(0.5)  # Semi-transparent preview
            font = QFont(self.marker_font_type)
            font.setPointSize(self.marker_font_size)
            painter.setFont(font)
            painter.setPen(self.marker_color)
            text_width = painter.fontMetrics().horizontalAdvance(self.preview_marker_text)
            text_height = painter.fontMetrics().height()
            # Draw the text at the cursor's position
            x, y = self.preview_marker_position.x(), self.preview_marker_position.y()
            if self.zoom_level != 1.0:
                x = (x - self.pan_offset.x()) / self.zoom_level
                y = (y - self.pan_offset.y()) / self.zoom_level
            painter.drawText(int(x - text_width / 2), int(y + text_height / 4), self.preview_marker_text)
            
        if len(self.quad_points) > 0 and len(self.quad_points) <4 :
            for p in self.quad_points:
                painter.setPen(QPen(Qt.red, 2))
                painter.drawEllipse(p, 1, 1)
    
        if len(self.quad_points) == 4 and self.draw_edges==True:
            painter.setPen(QPen(Qt.red, 2))
            painter.drawPolygon(QPolygonF(self.quad_points))
            # Draw draggable corners
            for p in self.quad_points:
                painter.drawEllipse(p, self.drag_threshold, self.drag_threshold)
    
        # Draw the bounding box preview if it exists
        if self.bounding_box_preview:
            painter.setPen(QPen(Qt.red, 2))  # Use green color for the bounding box
            start_x, start_y, end_x, end_y = self.bounding_box_preview
            rect = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y))
            painter.drawRect(rect)
    
        painter.end()
        self.update()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            if self.preview_marker_enabled:
                self.preview_marker_enabled = False  # Turn off the preview
                self.update()  # Clear the overlay
            self.measure_quantity_mode = False
            self.counter = 0
            self.bounding_box_complete = False
            self.quad_points = []
            self.mode=None
            self.update()
        super().keyPressEvent(event)

class CombinedSDSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
        window_width = int(self.screen_width * 0.5)  # 60% of screen width
        window_height = int(self.screen_height * 0.75)  # 95% of screen height
        self.window_title="IMAGING ASSISTANT V5"
        self.setWindowTitle(self.window_title)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.undo_stack = []
        self.redo_stack = []
        self.quantities_peak_area_dict = {}
        self.image_path = None
        self.image = None
        self.image_master= None
        self.image_before_padding = None
        self.image_contrasted=None
        self.image_before_contrast=None
        self.contrast_applied=False
        self.image_padded=False
        self.predict_size=False
        self.warped_image=None
        self.left_markers = []
        self.right_markers = []
        self.top_markers = []
        self.custom_markers=[]
        self.current_marker_index = 0
        self.current_top_label_index = 0
        self.font_rotation=-45
        self.image_width=0
        self.image_height=0
        self.new_image_width=0
        self.new_image_height=0
        self.base_name="Image"
        self.image_path=""
        self.x_offset_s=0
        self.y_offset_s=0
        self.peak_area=[]
        # Variables to store bounding boxes and quantities
        self.bounding_boxes = []
        self.up_bounding_boxes = []
        self.standard_protein_areas=[]
        self.quantities = []
        self.protein_quantities = []
        self.measure_quantity_mode = False
        # Initialize self.marker_values to None initially
        self.marker_values_dict = {
            "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
        }
        self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"]
        self.top_label_dict={"Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"]}
        
        
        self.custom_marker_name = "Custom"
        # Load the config file if exists
        
        self.marker_mode = None
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_slider_range=[-1000,1000]
        self.right_slider_range=[-1000,1000]
        self.top_slider_range=[-1000,1000]
               
        
        self.top_padding = 0
        self.font_color = QColor(0, 0, 0)  # Default to black
        self.custom_marker_color = QColor(0, 0, 0)  # Default to black
        self.font_family = "Arial"  # Default font family
        self.font_size = 12  # Default font size
        self.image_array_backup= None
        self.run_predict_MW=False
        
        self.create_menu_bar()
        
        # Main container widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Upper section (Preview and buttons)
        upper_layout = QHBoxLayout()

        self.label_width=int(self.screen_width * 0.28)
    
        self.live_view_label = LiveViewLabel(
            font_type=QFont("Arial"),
            font_size=int(24),
            marker_color=QColor(0,0,0),
            parent=self,
        )
        # Image display
        self.live_view_label.setStyleSheet("border: 1px solid black;")
        # self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setFixedSize(self.label_width, self.label_width)
        # self.live_view_label.mousePressEvent = self.add_band()
        # self.live_view_label.mousePressEvent = self.add_band
        
        
       

        # Buttons for image loading and saving
        # Load, save, and crop buttons
        buttons_layout = QVBoxLayout()
        load_button = QPushButton("Load Image")
        load_button.setToolTip("Load an image or a previously saved file. Shortcut: Ctrl+O or CMD+O")
        load_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        load_button.clicked.connect(self.load_image)
        buttons_layout.addWidget(load_button)
        
        paste_button = QPushButton('Paste Image')
        paste_button.setToolTip("Paste an image from clipboard or folder. Shortcut: Ctrl+V or CMD+V")
        paste_button.clicked.connect(self.paste_image)  # Connect the button to the paste_image method
        paste_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(paste_button)
    
        
        reset_button = QPushButton("Reset Image")  # Add Reset Image button
        reset_button.setToolTip("Reset all image manipulations and marker placements. Shortcut: Ctrl+R or CMD+R")
        reset_button.clicked.connect(self.reset_image)  # Connect the reset functionality
        reset_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(reset_button)
        
        
        
        copy_button = QPushButton('Copy Image to Clipboard')
        copy_button.setToolTip("Copy the modified image to clipboard. Shortcut: Ctrl+C or CMD+C")
        copy_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        copy_button.clicked.connect(self.copy_to_clipboard)
        buttons_layout.addWidget(copy_button)
        
        save_button = QPushButton("Save Image with Configuration")
        save_button.setToolTip("Save the modified image with the configuration files. Shortcut: Ctrl+S or CMD+S")
        save_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_button.clicked.connect(self.save_image)
        buttons_layout.addWidget(save_button)
        
        #if platform.system() == "Windows": # "Darwin" for MacOS # "Windows" for Windows
        #    copy_svg_button = QPushButton('Copy SVG Image to Clipboard')
        #    copy_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        #    copy_svg_button.clicked.connect(self.copy_to_clipboard_SVG)
        #    buttons_layout.addWidget(copy_svg_button)
            
        
        save_svg_button = QPushButton("Save SVG Image (MS Word Import)")
        save_svg_button.setToolTip("Save the modified image as an SVG file so that it can be modified in MS Word or similar. Shortcut: Ctrl+M or CMD+M")
        save_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_svg_button.clicked.connect(self.save_image_svg)
        buttons_layout.addWidget(save_svg_button)
        
        undo_redo_layout=QHBoxLayout()
        
        undo_button = QPushButton("Undo")
        undo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        undo_button.clicked.connect(self.undo_action)
        undo_button.setToolTip("Undo settings related to image. Cannot Undo Marker Placement. Use remove last option. Shortcut: Ctrl+U or CMD+U")
        undo_redo_layout.addWidget(undo_button)
        
        redo_button = QPushButton("Redo")
        redo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        redo_button.clicked.connect(self.redo_action)
        redo_button.setToolTip("Redo settings related to image. Cannot Undo Marker Placement. Use remove last option.Shortcut: Ctrl+R or CMD+R")        
        undo_redo_layout.addWidget(redo_button)
        
        buttons_layout.addLayout(undo_redo_layout)
        
        # New Zoom buttons layout
        zoom_layout = QHBoxLayout()
        
        # Zoom In button
        zoom_in_button = QPushButton("Zoom In")
        zoom_in_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_in_button.clicked.connect(self.zoom_in)
        zoom_in_button.setToolTip("Increase zoom level. Click the display window and use arrow keys for moving")
        
        # Zoom Out button
        zoom_out_button = QPushButton("Zoom Out")
        zoom_out_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_out_button.clicked.connect(self.zoom_out)
        zoom_out_button.setToolTip("Decrease zoom level. Click the display window and use arrow keys for moving")
        
        # Add Zoom buttons to the zoom layout
        zoom_layout.addWidget(zoom_in_button)
        zoom_layout.addWidget(zoom_out_button)
        
        # Add Zoom layout below Undo/Redo
        buttons_layout.addLayout(zoom_layout)
        
        # buttons_layout.addStretch()
        
        
        upper_layout.addWidget(self.live_view_label, stretch=1)
        upper_layout.addLayout(buttons_layout, stretch=1)
        layout.addLayout(upper_layout)

        # Lower section (Tabbed interface)
        self.tab_widget = QTabWidget()
        # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
        self.tab_widget.addTab(self.font_and_image_tab(), "Font and Image Parameters")
        self.tab_widget.addTab(self.create_cropping_tab(), "Align, Crop and Skew Parameters")
        self.tab_widget.addTab(self.create_white_space_tab(), "White Space Parameters")
        self.tab_widget.addTab(self.create_markers_tab(), "Marker Parameters")
        self.tab_widget.addTab(self.combine_image_tab(), "Overlap Parameters")
        self.tab_widget.addTab(self.analysis_tab(), "Analysis")
        
        layout.addWidget(self.tab_widget)
        
        # Example: Shortcut for loading an image (Ctrl + O)
        self.load_shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        self.load_shortcut.activated.connect(self.load_image)
        
        # Example: Shortcut for saving an image (Ctrl + S)
        self.save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        self.save_shortcut.activated.connect(self.save_image)
        
        # Example: Shortcut for resetting the image (Ctrl + R)
        self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.reset_shortcut.activated.connect(self.reset_image)
        
        # Example: Shortcut for copying the image to clipboard (Ctrl + C)
        self.copy_shortcut = QShortcut(QKeySequence("Ctrl+C"), self)
        self.copy_shortcut.activated.connect(self.copy_to_clipboard)
        
        # Example: Shortcut for pasting an image from clipboard (Ctrl + V)
        self.paste_shortcut = QShortcut(QKeySequence("Ctrl+V"), self)
        self.paste_shortcut.activated.connect(self.paste_image)
        
        # Example: Shortcut for predicting molecular weight (Ctrl + P)
        self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        self.predict_shortcut.activated.connect(self.predict_molecular_weight)
        
        # Example: Shortcut for clearing the prediction marker (Ctrl + Shift + P)
        self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)
        
        # Example: Shortcut for saving SVG image (Ctrl + Shift + S)
        self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
        self.save_svg_shortcut.activated.connect(self.save_image_svg)
        
        # Example: Shortcut for enabling left marker mode (Ctrl + L)
        self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
        self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)
        
        # Example: Shortcut for enabling right marker mode (Ctrl + R)
        self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
        self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)
        
        # Example: Shortcut for enabling top marker mode (Ctrl + T)
        self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
        self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
        self.grid_shortcut.activated.connect(lambda: self.show_grid_checkbox.setChecked(not self.show_grid_checkbox.isChecked()))
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.guidelines_shortcut.activated.connect(lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked()))
        
        # Example: Shortcut for increasing grid size (Ctrl + Shift + Up)
        self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
        self.increase_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1))
        
        # Example: Shortcut for decreasing grid size (Ctrl + Shift + Down)
        self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
        self.decrease_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1))
        
        # # Example: Shortcut for increasing font size (Ctrl + Plus)
        # self.increase_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Up"), self)
        # self.increase_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() + 1))
        
        # # Example: Shortcut for decreasing font size (Ctrl + Minus)
        # self.decrease_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Down"), self)
        # self.decrease_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() - 1))
        
        # Example: Shortcut for custom marker left arrow (Ctrl + Left Arrow)
        self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker right arrow (Ctrl + Right Arrow)
        self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker top arrow (Ctrl + Up Arrow)
        self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker bottom arrow (Ctrl + Down Arrow)
        self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
        self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for inverting image (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self.invert_shortcut.activated.connect(self.invert_image)
        
        # Example: Shortcut for converting to grayscale (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
        self.invert_shortcut.activated.connect(self.convert_to_black_and_white)
        
        # Example: Move quickly between tabs (Ctrl + 1,2,3,4)
        self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
        self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
        
        self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
        self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
        
        self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
        self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
        
        self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
        self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
        
        self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
        self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
        
        self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
        self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
        
        self.undo_shortcut = QShortcut(QKeySequence("Ctrl+U"), self)
        self.undo_shortcut.activated.connect(self.undo_action)
        
        self.redo_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.redo_shortcut.activated.connect(self.redo_action)
        self.load_config()
        
    def quadrilateral_to_rect(self, image, quad_points):
        image = image.convertToFormat(QImage.Format_RGBA8888)
        # Get the dimensions of the live view label and the actual image
        label_width = self.live_view_label.width()
        label_height = self.live_view_label.height()
        image_width = image.width()
        image_height = image.height()
        
        # Calculate scaling factors and offsets
        scale_x = image_width / label_width
        scale_y = image_height / label_height
        try:
            if len(self.live_view_label.bounding_box_preview) == 4:
                start_x, start_y = self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]
                end_x, end_y = self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3]
                
                if self.live_view_label.zoom_level != 1.0:
                    start_x = int((start_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level)
                    start_y = int((start_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level)
                    end_x = int((end_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level)
                    end_y = int((end_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level)
                
                width, height = abs(end_x - start_x), abs(end_y - start_y)
                x, y = min(start_x, end_x), min(start_y, end_y)
                
                # Convert coordinates to image space
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
                image_width = self.image.width()
                image_height = self.image.height()
                scale = min(displayed_width / image_width, displayed_height / image_height)
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2
                
                x = max(0, min((x - x_offset) / scale, image_width))
                y = max(0, min((y - y_offset) / scale, image_height))
                w = min(width / scale, image_width - x)
                h = min(height / scale, image_height - y)
                
                if x + w > self.image.width():
                    w = self.image.width() - x
                if y + h > self.image.height():
                    h = self.image.height() - y
            
                if w <= 0 or h <= 0:
                    return 0  # Skip invalid regions
            
                cropped = self.image.copy(int(x), int(y), int(w), int(h))
                return cropped
        except:
            QMessageBox.warning(self, "Error", "No bands detected")
            
            
        
        if len(self.live_view_label.quad_points) == 4:
            
            # Transform the quadrilateral points from label coordinates to image coordinates
            src_points = []
            for point in quad_points:
                # Adjust for zoom and pan if applicable
                if self.live_view_label.zoom_level != 1.0:
                    x = (point.x() - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    y = (point.y() - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
                else:
                    x = point.x()
                    y = point.y()
                
                # Scale the points to match the image dimensions
                x_image = x * scale_x
                y_image = y * scale_y
                src_points.append([x_image, y_image])
            
            src_points = np.array(src_points, dtype=np.float32)
            
            # Calculate the bounding box of the quadrilateral in image coordinates
            min_x = np.min(src_points[:, 0])
            max_x = np.max(src_points[:, 0])
            min_y = np.min(src_points[:, 1])
            max_y = np.max(src_points[:, 1])
            
            # Ensure the bounding box is within the image dimensions
            min_x = np.maximum(0, min_x)
            max_x = np.minimum(image_width, max_x)
            min_y = np.maximum(0, min_y)
            max_y = np.minimum(image_height, max_y)
            
            # Define the destination points (the rectangle to which we want to map the quadrilateral)
            dst_points = np.array([
                [0, 0],
                [max_x - min_x, 0],
                [max_x - min_x, max_y - min_y],
                [0, max_y - min_y]
            ], dtype=np.float32)
            
            # Determine the image format (grayscale vs RGBA)
            if image.format() == image.Format_Grayscale8:
                channels = 1
                format_type = image.Format_Grayscale8
            else:
                channels = 4  # Assume RGBA
                format_type = image.Format_RGBA8888
            
            # Convert QImage to numpy
            ptr = image.constBits()
            ptr.setsize(image.byteCount())
            img_data = np.array(ptr).reshape(image_height, image_width, channels)
            
            if channels == 1:
                img = np.frombuffer(img_data, dtype=np.uint8).reshape((image_height, image_width))
            else:
                img = np.frombuffer(img_data, dtype=np.uint8).reshape((image_height, image_width, channels))
            
            # Compute the perspective transform matrix using numpy
            A = np.zeros((8, 8), dtype=np.float32)
            b = np.zeros((8,), dtype=np.float32)
            
            for i in range(4):
                x, y = src_points[i]
                u, v = dst_points[i]
                A[i*2] = [x, y, 1, 0, 0, 0, -x*u, -y*u]
                A[i*2+1] = [0, 0, 0, x, y, 1, -x*v, -y*v]
                b[i*2] = u
                b[i*2+1] = v
            
            # Solve the system of equations
            h = np.linalg.solve(A, b)
            
            # Correctly create 3x3 transformation matrix
            transform_matrix = np.zeros((3, 3), dtype=np.float32)
            transform_matrix[0, :] = [h[0], h[1], h[2]]
            transform_matrix[1, :] = [h[3], h[4], h[5]]
            transform_matrix[2, :] = [h[6], h[7], 1.0]
            
            # Create output image
            out_width = int(max_x - min_x)
            out_height = int(max_y - min_y)
            
            if channels == 1:
                warped_image = np.zeros((out_height, out_width), dtype=np.uint8)
            else:
                warped_image = np.zeros((out_height, out_width, channels), dtype=np.uint8)
            
            # Create mesh grid for destination coordinates
            y_coords, x_coords = np.mgrid[0:out_height, 0:out_width]
            coords = np.vstack((x_coords.flatten(), y_coords.flatten(), np.ones_like(x_coords.flatten())))
            
            # Apply inverse transformation
            transform_matrix_inv = np.linalg.inv(transform_matrix)
            source_coords = np.dot(transform_matrix_inv, coords)
            
            # Convert to homogeneous coordinates
            source_coords = source_coords / source_coords[2, :]
            source_x = source_coords[0, :].reshape(out_height, out_width)
            source_y = source_coords[1, :].reshape(out_height, out_width)
            
            # Find valid coordinates (within the source image)
            valid = (source_x >= 0) & (source_x < image_width) & (source_y >= 0) & (source_y < image_height)
            
            # Round to nearest pixel
            source_x_int = np.round(source_x).astype(np.int32)
            source_y_int = np.round(source_y).astype(np.int32)
            
            # Apply bounds
            source_x_int = np.clip(source_x_int, 0, image_width - 1)
            source_y_int = np.clip(source_y_int, 0, image_height - 1)
            
            # Copy pixels from source to destination
            if channels == 1:
                warped_image[y_coords[valid], x_coords[valid]] = img[source_y_int[valid], source_x_int[valid]]
            else:
                for i in range(channels):
                    warped_image[y_coords[valid], x_coords[valid], i] = img[source_y_int[valid], source_x_int[valid], i]
            
            # Convert the resulting warped image back to a QImage
            if channels == 1:
                warped_qimage = QImage(warped_image.data, out_width, out_height, out_width, format_type)
            else:
                warped_qimage = QImage(warped_image.data, out_width, out_height, out_width * channels, format_type)
        
        return warped_qimage
    
    

        
    def create_menu_bar(self):
        # Create the menu bar
        menubar = self.menuBar()

        # Create the "File" menu
        file_menu = menubar.addMenu("File")

        # Add "Load Image" action
        load_action = QAction("Load Image", self)
        load_action.triggered.connect(self.load_image)
        file_menu.addAction(load_action)

        # Add "Save Image" action
        save_action = QAction("Save Image", self)
        save_action.triggered.connect(self.save_image)
        file_menu.addAction(save_action)
        
        # Add "Save Image SVG" action
        save_action_svg = QAction("Save Image as SVG", self)
        save_action_svg.triggered.connect(self.save_image_svg)
        file_menu.addAction(save_action_svg)

        # Add "Reset Image" action
        reset_action = QAction("Reset Image", self)
        reset_action.triggered.connect(self.reset_image)
        file_menu.addAction(reset_action)

        # Add a separator
        file_menu.addSeparator()

        # Add "Exit" action
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Create the "About" menu
        about_menu = menubar.addMenu("About")

        # Add "GitHub" action
        github_action = QAction("GitHub", self)
        github_action.triggered.connect(self.open_github)
        about_menu.addAction(github_action)

    def open_github(self):
        # Open the GitHub link in the default web browser
        QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
        
    def zoom_in(self):
        self.live_view_label.zoom_in()
        self.update_live_view()

    def zoom_out(self):
        self.live_view_label.zoom_out()
        self.update_live_view()
    
    def enable_standard_protein_mode(self):
        """"Enable mode to define standard protein amounts for creating a standard curve."""
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        # Assign mouse event handlers for bounding box creation
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_measure_protein_mode(self):
        """Enable mode to measure protein quantity using the standard curve."""
        if len(self.quantities) < 2:
            QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)

    def call_live_view(self):
        self.update_live_view()      

    def analyze_bounding_box(self, image, standard):  
        # Prompt user for quantity input if standard
        if standard:
            quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known protein amount (e.g., 1.5 units):")
            if ok and quantity:
                try:
                    # Extract just the numeric part if unit is included
                    quantity_value = float(quantity.split()[0])
                    # self.quantities.append(quantity_value)
                    # Calculate the area under the peak and subtract the background
                    peak_area = self.calculate_peak_area(image)  
                    self.quantities_peak_area_dict[quantity_value] = round(sum(peak_area),3)
                    self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                except (ValueError, IndexError):
                    QMessageBox.warning(self, "Error", "Please enter a valid number for protein quantity.")
            self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
        else:
            # self.up_bounding_boxes=[]
            # self.up_bounding_boxes.append((image_x, image_y, image_w, image_h))
            self.peak_area = self.calculate_peak_area(image) 
            if len(list(self.quantities_peak_area_dict.keys()))>=2:
                self.calculate_unknown_quantity(list(self.quantities_peak_area_dict.values()), list(self.quantities_peak_area_dict.keys()), self.peak_area)
            try:
                self.target_protein_areas_text.setText(str([round(x, 3) for x in self.peak_area]))
            except:
                pass
        self.update_live_view()
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.bounding_box_start = None
    
    def calculate_peak_area(self, cropped):  
        cropped = cropped.convertToFormat(QImage.Format_Grayscale8)  # Ensure grayscale format
        
    
        
    
        # # Convert QImage to PIL Image
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        cropped.save(buffer, "PNG")
        cropped_image = Image.open(io.BytesIO(buffer.data()))
        buffer.close()
        
        # Open interactive adjustment window
        dialog = PeakAreaDialog(cropped_image, parent=self)
        if dialog.exec_() == QDialog.Accepted:
            peak_area = dialog.get_final_peak_area()
            return peak_area  # Return user-adjusted peak area
    
        return 0  # Return 0 if the user cancels the adjustment
    
    
    def calculate_unknown_quantity(self, peak_area_list, known_quantities, peak_area):
        coefficients = np.polyfit(peak_area_list, known_quantities, 1)
        unknown_quantity=[]
        for i in range(len(peak_area)):
            unknown_quantity.append(round(np.polyval(coefficients, peak_area[i])),2)
        self.protein_quantities=[]
        QMessageBox.information(self, "Protein Quantification", f"Predicted Quantity: {unknown_quantity} units")

        
    def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
        """Draw quantity text at the correct position."""
        text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
        painter.drawText(text_position, str(quantity))
    
    def update_standard_protein_quantities(self):
        self.standard_protein_values.text()
    
    def move_tab(self,tab):
        self.tab_widget.setCurrentIndex(tab)
        
    def save_state(self):
        """Save the current state of the image, markers, and other relevant data."""
        state = {
            "image": self.image.copy() if self.image else None,
            "left_markers": self.left_markers.copy(),
            "right_markers": self.right_markers.copy(),
            "top_markers": self.top_markers.copy(),
            "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
            "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
            "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
            "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
            "font_family": self.font_family,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "font_rotation": self.font_rotation,
            "left_marker_shift_added": self.left_marker_shift_added,
            "right_marker_shift_added": self.right_marker_shift_added,
            "top_marker_shift_added": self.top_marker_shift_added,
            "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
            "peak_area": self.peak_area.copy(),
            "quantities": self.quantities.copy(),
            "protein_quantities": self.protein_quantities.copy(),
            "standard_protein_areas": self.standard_protein_areas.copy(),
        }
        self.undo_stack.append(state)
        self.redo_stack.clear()
    
    def undo_action(self):
        """Undo the last action by restoring the previous state."""
        if self.undo_stack:
            # Save the current state to the redo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "peak_area": self.peak_area.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.redo_stack.append(current_state)
            
            # Restore the previous state from the undo stack
            previous_state = self.undo_stack.pop()
            self.image = previous_state["image"]
            self.left_markers = previous_state["left_markers"]
            self.right_markers = previous_state["right_markers"]
            self.top_markers = previous_state["top_markers"]
            self.custom_markers = previous_state["custom_markers"]
            self.image_before_padding = previous_state["image_before_padding"]
            self.image_contrasted = previous_state["image_contrasted"]
            self.image_before_contrast = previous_state["image_before_contrast"]
            self.font_family = previous_state["font_family"]
            self.font_size = previous_state["font_size"]
            self.font_color = previous_state["font_color"]
            self.font_rotation = previous_state["font_rotation"]
            self.left_marker_shift_added = previous_state["left_marker_shift_added"]
            self.right_marker_shift_added = previous_state["right_marker_shift_added"]
            self.top_marker_shift_added = previous_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
            self.peak_area = previous_state["peak_area"]
            self.quantities = previous_state["quantities"]
            self.protein_quantities = previous_state["protein_quantities"]
            self.standard_protein_areas = previous_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    
    def redo_action(self):
        """Redo the last undone action by restoring the next state."""
        if self.redo_stack:
            # Save the current state to the undo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "peak_area": self.peak_area.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.undo_stack.append(current_state)
            
            # Restore the next state from the redo stack
            next_state = self.redo_stack.pop()
            self.image = next_state["image"]
            self.left_markers = next_state["left_markers"]
            self.right_markers = next_state["right_markers"]
            self.top_markers = next_state["top_markers"]
            self.custom_markers = next_state["custom_markers"]
            self.image_before_padding = next_state["image_before_padding"]
            self.image_contrasted = next_state["image_contrasted"]
            self.image_before_contrast = next_state["image_before_contrast"]
            self.font_family = next_state["font_family"]
            self.font_size = next_state["font_size"]
            self.font_color = next_state["font_color"]
            self.font_rotation = next_state["font_rotation"]
            self.left_marker_shift_added = next_state["left_marker_shift_added"]
            self.right_marker_shift_added = next_state["right_marker_shift_added"]
            self.top_marker_shift_added = next_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
            self.peak_area = next_state["peak_area"]
            self.quantities = next_state["quantities"]
            self.protein_quantities = next_state["protein_quantities"]
            self.standard_protein_areas = next_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    def analysis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Group 1: Molecular Weight Prediction
        mw_group = QGroupBox("Molecular Weight Prediction")
        mw_layout = QVBoxLayout()
    
        self.predict_button = QPushButton("Predict Molecular Weight")
        self.predict_button.setToolTip("Predicts the size of the protein/DNA if the MW marker or ladder is labeled and puts a straight line marker on the image. Shortcut: Ctrl+P or CMD+P")
        self.predict_button.setEnabled(False)  # Initially disabled
        self.predict_button.clicked.connect(self.predict_molecular_weight)
    
        mw_layout.addWidget(self.predict_button)
        mw_group.setLayout(mw_layout)
        
        # Standard Curve Group
        std_group = QGroupBox("Peak Area / Sample Quantification")
        std_layout = QVBoxLayout()
        
        area_layout=QHBoxLayout()
        
        self.btn_define_quad = QPushButton("Define Quad Area")
        self.btn_define_quad.clicked.connect(self.enable_quad_mode)
        area_layout.addWidget(self.btn_define_quad)
        
        self.btn_define_rec = QPushButton("Define Rectangle Area")
        self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
        area_layout.addWidget(self.btn_define_rec)
        
        self.btn_sel_rec = QPushButton("Move Selected Area")
        self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
        area_layout.addWidget(self.btn_sel_rec)
        
        
        std_layout.addLayout(area_layout)
        
        self.btn_process_std = QPushButton("Process Standard Bands")
        self.btn_process_std.clicked.connect(self.process_standard)
        std_layout.addWidget(self.btn_process_std)
        
        self.standard_protein_values = QLineEdit()
        self.standard_protein_values.setPlaceholderText("Autopopulates with standard quantities")
        std_layout.addWidget(self.standard_protein_values)
    
        self.standard_protein_areas_text = QLineEdit()
        self.standard_protein_areas_text.setPlaceholderText("Autopopulates with peak areas")
        std_layout.addWidget(self.standard_protein_areas_text)
        
        
        self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
        self.btn_analyze_sample.clicked.connect(self.process_sample)
        std_layout.addWidget(self.btn_analyze_sample)
        
        self.target_protein_areas_text = QLineEdit()
        self.target_protein_areas_text.setPlaceholderText("Autopopulates with peak areas")
        std_layout.addWidget(self.target_protein_areas_text)
        
        # Add the "Table Export" button next to target_protein_areas_text
        self.table_export_button = QPushButton("Table Export")
        self.table_export_button.clicked.connect(self.open_table_window)
        std_layout.addWidget(self.table_export_button)
        
        self.clear_predict_button = QPushButton("Clear Markers")
        self.clear_predict_button.setToolTip("Clears all protein analysis markers. Shortcut: Ctrl+Shift+P or CMD+Shift+P")
        self.clear_predict_button.setEnabled(True)
        self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
        
        
        
        
        # Add to layout
        std_group.setLayout(std_layout)
        layout.addWidget(mw_group)
        layout.addWidget(std_group)
        layout.addWidget(self.clear_predict_button)
        layout.addStretch()
        return tab
    
    def enable_move_selection_mode(self):
        """Enable mode to move the selected quadrilateral or rectangle."""
        self.live_view_label.mode = "move"
        self.live_view_label.setCursor(Qt.SizeAllCursor)  # Change cursor to indicate move mode
        self.live_view_label.mousePressEvent = self.start_move_selection
        self.live_view_label.mouseReleaseEvent = self.end_move_selection
        self.update_live_view()
        
    def start_move_selection(self, event):
        """Start moving the selection when the mouse is pressed."""
        if self.live_view_label.mode == "move":
            self.live_view_label.draw_edges=False
            # Check if the mouse is near the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), self.live_view_label.quad_points)
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), [
                    QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1])
                ])
            self.live_view_label.drag_start_pos = event.pos()
            self.update_live_view()
        self.live_view_label.mouseMoveEvent = self.move_selection
    
    def move_selection(self, event):
        """Move the selection while the mouse is being dragged."""
        if self.live_view_label.mode == "move" and self.live_view_label.selected_point is not None:
            # Calculate the offset
            offset = event.pos() - self.live_view_label.drag_start_pos
            self.live_view_label.drag_start_pos = event.pos()
            
            # Move the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                for i in range(len(self.live_view_label.quad_points)):
                    self.live_view_label.quad_points[i] += offset
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.bounding_box_preview = (
                    self.live_view_label.bounding_box_preview[0] + offset.x(),
                    self.live_view_label.bounding_box_preview[1] + offset.y(),
                    self.live_view_label.bounding_box_preview[2] + offset.x(),
                    self.live_view_label.bounding_box_preview[3] + offset.y(),
                )
            self.update_live_view()
    
    def end_move_selection(self, event):
        """End moving the selection when the mouse is released."""
        if self.live_view_label.mode == "move":
            self.live_view_label.selected_point = -1            
            self.update_live_view()
            self.live_view_label.draw_edges=True
        self.live_view_label.mouseMoveEvent = None
        
        
             
    def get_nearest_point(self, mouse_pos, points):
        """Get the nearest point to the mouse position."""
        min_distance = float('inf')
        nearest_point = None
        for point in points:
            distance = (mouse_pos - point).manhattanLength()
            if distance < min_distance:
                min_distance = distance
                nearest_point = point
        return nearest_point
    
    def open_table_window(self):
        # Example list of peak areas
        if self.peak_area==[]:
            peak_areas = [0, 0, 0, 0]
            standard=False
            standard_dictionary={}
        else:
            peak_areas = self.peak_area
            standard_dictionary=self.quantities_peak_area_dict
            if len(self.quantities_peak_area_dict)>=2:
                standard=True
            else:
                standard=False
        
        # Open the table window with the peak areas
        self.table_window = TableWindow(peak_areas, standard_dictionary, standard,self)
        self.table_window.show()
    
    def enable_quad_mode(self):
        """Enable mode to define a quadrilateral area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "quad"
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # # Reset mouse event handlers
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        
        # Update the live view
        self.update_live_view()
    
    def enable_rectangle_mode(self):
        """Enable mode to define a rectangle area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "rectangle"
        self.live_view_label.rectangle_points = []  # Clear previous rectangle points
        self.live_view_label.rectangle_start = None  # Reset rectangle start
        self.live_view_label.rectangle_end = None    # Reset rectangle end
        self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Set mouse event handlers for rectangle mode
        self.live_view_label.mousePressEvent = self.start_rectangle
        self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
        
        # Update the live view
        self.update_live_view()

    
    def start_rectangle(self, event):
        """Record the start position of the rectangle."""
        if self.live_view_label.mode == "rectangle":
            self.live_view_label.rectangle_start = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start]
            self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
            self.update_live_view()
    
    def update_rectangle_preview(self, event):
        """Update the rectangle preview as the mouse moves."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Update bounding_box_preview with the rectangle coordinates
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            self.update_live_view()
    
    def finalize_rectangle(self, event):
        """Finalize the rectangle when the mouse is released."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Save the final bounding box preview
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            
            self.live_view_label.mode = None  # Exit rectangle mode
            self.live_view_label.setCursor(Qt.ArrowCursor)
            self.update_live_view()
        
    
    def process_standard(self,image):
        # if len(self.live_view_label.quad_points) != 4:
        #     QMessageBox.warning(self, "Error", "Please define quadrilateral area first")
        #     return
        try:
            if len(self.live_view_label.quad_points) == 4:
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                self.update_live_view()            
                warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.quad_points)            
                self.analyze_bounding_box(warped_image, standard=True) 
                return
            if len(self.live_view_label.bounding_box_preview) == 4:
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                self.update_live_view()            
                warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.rectangle_points)            
                self.analyze_bounding_box(warped_image, standard=True)
                return
        except:
            QMessageBox.warning(self, "Error", "No bands detected")
    
    def process_sample(self):
        # if len(self.live_view_label.quad_points) != 4:
        #     QMessageBox.warning(self, "Error", "Please define quadrilateral area first")
        #     return
        if len(self.live_view_label.quad_points) == 4:
            self.live_view_label.pan_offset = QPointF(0, 0)
            self.live_view_label.zoom_level=1.0
            self.update_live_view()            
            warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.quad_points)            
            self.analyze_bounding_box(warped_image, standard=False) 
            return
        if len(self.live_view_label.bounding_box_preview) == 4:
            self.live_view_label.pan_offset = QPointF(0, 0)
            self.live_view_label.zoom_level=1.0
            self.update_live_view()            
            warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.rectangle_points)            
            self.analyze_bounding_box(warped_image, standard=False)
            return                
        
        
            
    def combine_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        
        render_scale=3
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Group Box for Image 1
        image1_group = QGroupBox("Image 1")
        image1_layout = QVBoxLayout()
        
        load_layout_1=QHBoxLayout()
        # Save Image 1 Button
        save_image1_button = QPushButton("Copy Image")
        save_image1_button.clicked.connect(self.save_image1)
        load_layout_1.addWidget(save_image1_button)
    
        # Place Image 1 Button
        place_image1_button = QPushButton("Place Image")
        place_image1_button.clicked.connect(self.place_image1)
        load_layout_1.addWidget(place_image1_button)
    
        # Remove Image 1 Button
        remove_image1_button = QPushButton("Remove Image")
        remove_image1_button.clicked.connect(self.remove_image1)
        load_layout_1.addWidget(remove_image1_button)
        
        image1_layout.addLayout(load_layout_1)
        
    
        # Sliders for Image 1 Position
        self.image1_left_slider = QSlider(Qt.Horizontal)
        self.image1_left_slider.setRange(-int(render_width+100), int(render_width+100))
        self.image1_left_slider.setValue(0)
        self.image1_left_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(QLabel("Horizontal Position"))
        image1_layout.addWidget(self.image1_left_slider)
    
        self.image1_top_slider = QSlider(Qt.Horizontal)
        self.image1_top_slider.setRange(-int(render_height+100), int(render_height+100))
        self.image1_top_slider.setValue(0)
        self.image1_top_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(QLabel("Vertical Position"))
        image1_layout.addWidget(self.image1_top_slider)
    
        # Resize Slider for Image 1
        self.image1_resize_slider = QSlider(Qt.Horizontal)
        self.image1_resize_slider.setRange(10, 400)  # 10% to 200% scaling
        self.image1_resize_slider.setValue(100)  # Default to 100% (no scaling)
        self.image1_resize_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(QLabel("Resize Image (%)"))
        image1_layout.addWidget(self.image1_resize_slider)
    
        image1_group.setLayout(image1_layout)
        layout.addWidget(image1_group)
    
        # Group Box for Image 2
        image2_group = QGroupBox("Image 2")
        image2_layout = QVBoxLayout()
        
        load_layout_2=QHBoxLayout()
    
        # Save Image 2 Button
        save_image2_button = QPushButton("Copy Image")
        save_image2_button.clicked.connect(self.save_image2)
        load_layout_2.addWidget(save_image2_button)
    
        # Place Image 2 Button
        place_image2_button = QPushButton("Place Image")
        place_image2_button.clicked.connect(self.place_image2)
        load_layout_2.addWidget(place_image2_button)
    
        # Remove Image 2 Button
        remove_image2_button = QPushButton("Remove Image")
        remove_image2_button.clicked.connect(self.remove_image2)
        load_layout_2.addWidget(remove_image2_button)
        
        image2_layout.addLayout(load_layout_2)
    
        # Sliders for Image 2 Position
        self.image2_left_slider = QSlider(Qt.Horizontal)
        self.image2_left_slider.setRange(-int(render_width+100), int(render_width+100))
        self.image2_left_slider.setValue(0)
        self.image2_left_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(QLabel("Horizontal Position"))
        image2_layout.addWidget(self.image2_left_slider)
    
        self.image2_top_slider = QSlider(Qt.Horizontal)
        self.image2_top_slider.setRange(-int(render_height+100), int(render_height+100))
        self.image2_top_slider.setValue(0)
        self.image2_top_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(QLabel("Vertical Position"))
        image2_layout.addWidget(self.image2_top_slider)
    
        # Resize Slider for Image 2
        self.image2_resize_slider = QSlider(Qt.Horizontal)
        self.image2_resize_slider.setRange(10, 400)  # 10% to 200% scaling
        self.image2_resize_slider.setValue(100)  # Default to 100% (no scaling)
        self.image2_resize_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(QLabel("Resize Image(%)"))
        image2_layout.addWidget(self.image2_resize_slider)
    
        image2_group.setLayout(image2_layout)
        layout.addWidget(image2_group)
    
        # Finalize Image Button
        finalize_button = QPushButton("Rasterize the images")
        finalize_button.setToolTip("Will merge all images, markers and texts that is on the visible on the screen")
        finalize_button.clicked.connect(self.finalize_combined_image)
        layout.addWidget(finalize_button)
        layout.addStretch()
    
        return tab
    
    def remove_image1(self):
        """Remove Image 1 and reset its sliders."""
        if hasattr(self, 'image1'):
            # del self.image1
            # del self.image1_original
            try:
                del self.image1_position
            except:
                pass
            self.image1_left_slider.setValue(0)
            self.image1_top_slider.setValue(0)
            self.image1_resize_slider.setValue(100)
            self.update_live_view()
    
    def remove_image2(self):
        """Remove Image 2 and reset its sliders."""
        if hasattr(self, 'image2'):
            # del self.image2
            # del self.image2_original
            try:
                del self.image2_position
            except:
                pass
            self.image2_left_slider.setValue(0)
            self.image2_top_slider.setValue(0)
            self.image2_resize_slider.setValue(100)
            self.update_live_view()
    
    def save_image1(self):
        if self.image:
            self.image1 = self.image.copy()
            self.image1_original = self.image1.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 1 copied.")
    
    def save_image2(self):
        if self.image:
            self.image2 = self.image.copy()
            self.image2_original = self.image2.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 2 copied.")
    
    def place_image1(self):
        if hasattr(self, 'image1'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            self.update_live_view()
    
    def place_image2(self):
        if hasattr(self, 'image2'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            self.update_live_view()
    
    # def update_combined_image(self):
    #     if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
    #         return
    
    #     # Create a copy of the original image to avoid modifying it
    #     combined_image = self.image.copy()
    
    #     painter = QPainter(combined_image)
    
    #     # Draw Image 1 if it exists
    #     if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
    #         # Resize Image 1 based on the slider value
    #         scale_factor = self.image1_resize_slider.value() / 100.0
    #         width = int(self.image1_original.width() * scale_factor)
    #         height = int(self.image1_original.height() * scale_factor)
    #         resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image1_position[0], self.image1_position[1], resized_image1)
    
    #     # Draw Image 2 if it exists
    #     if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
    #         # Resize Image 2 based on the slider value
    #         scale_factor = self.image2_resize_slider.value() / 100.0
    #         width = int(self.image2_original.width() * scale_factor)
    #         height = int(self.image2_original.height() * scale_factor)
    #         resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image2_position[0], self.image2_position[1], resized_image2)
    
    #     painter.end()
    
    #     # Update the live view with the combined image
    #     self.live_view_label.setPixmap(QPixmap.fromImage(combined_image))
    
    def finalize_combined_image(self):
        """Overlap image1 and image2 on top of self.image and save the result as self.image."""
        # if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
        #     QMessageBox.warning(self, "Warning", "No images to overlap.")
        #     return
        
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
        # Create a high-resolution canvas for the final image
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Create a blank high-resolution image with a white background
        high_res_canvas = QImage(render_width, render_height, QImage.Format_RGB888)
        high_res_canvas.fill(Qt.white)
    
        # Create a QPainter to draw on the high-resolution canvas
        painter = QPainter(high_res_canvas)
    
        # Draw the base image (self.image) onto the high-resolution canvas
        scaled_base_image = self.image.scaled(render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        painter.drawImage(0, 0, scaled_base_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image1_position[0] + self.x_offset_s),
                int(self.image1_position[1] + self.y_offset_s),
                resized_image1
            )
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image2_position[0] + self.x_offset_s),
                int(self.image2_position[1] + self.y_offset_s),
                resized_image2
            )
    
        # End painting
        painter.end()
        
        
        self.render_image_on_canvas(
                high_res_canvas, scaled_base_image, x_start, y_start, render_scale, draw_guides=False
            )
        
        self.image=high_res_canvas.copy()
        self.image_before_padding=self.image.copy()
        self.image_before_contrast=self.image.copy()
        self.update_live_view()
    
        # Remove Image 1 and Image 2 after finalizing
        self.remove_image1()
        self.remove_image2()
        
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
    
        QMessageBox.information(self, "Success", "The images have been overlapped and saved in memory.")
    
        
    


    def font_and_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Group Box for Font Options
        font_options_group = QGroupBox("Font Options")
        font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        font_options_layout = QGridLayout()  # Use a grid layout for compact arrangement
    
        # Font type selection (QFontComboBox)
        font_type_label = QLabel("Font Type:")
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setEditable(False)
        self.font_combo_box.setCurrentFont(QFont("Arial"))
        font_options_layout.addWidget(font_type_label, 0, 0)  # Row 0, Column 0
        font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2)  # Row 0, Column 1-2
    
        # Font size selection (QSpinBox)
        font_size_label = QLabel("Font Size:")
        self.font_size_spinner = QSpinBox()
        self.font_size_spinner.setRange(2,150)  # Set a reasonable font size range
        self.font_size_spinner.setValue(12)  # Default font size
        font_options_layout.addWidget(font_size_label, 1, 0)  # Row 1, Column 0
        font_options_layout.addWidget(self.font_size_spinner, 1, 1)  # Row 1, Column 1
    
        # Font color selection (QPushButton to open QColorDialog)
        self.font_color_button = QPushButton("Font Color")
        self.font_color_button.clicked.connect(self.select_font_color)
        font_options_layout.addWidget(self.font_color_button, 1, 2)  # Row 1, Column 2
    
        # Font rotation input (QSpinBox)
        font_rotation_label = QLabel("Font Rotation (Top/Bottom Label):")
        self.font_rotation_input = QSpinBox()
        self.font_rotation_input.setRange(-180, 180)  # Set a reasonable font rotation range
        self.font_rotation_input.setValue(-45)  # Default rotation
        self.font_rotation_input.valueChanged.connect(self.update_font)
        font_options_layout.addWidget(font_rotation_label, 2, 0)  # Row 2, Column 0
        font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2)  # Row 2, Column 1-2
    
        # Set the layout for the font options group box
        font_options_group.setLayout(font_options_layout)
        layout.addWidget(font_options_group)
    
        # Group Box for Contrast and Gamma Adjustments
        contrast_gamma_group = QGroupBox("Contrast and Gamma Adjustments")
        contrast_gamma_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        contrast_gamma_layout = QGridLayout()  # Use a grid layout for a clean look
    
        # Bright Region Contrast Slider
        high_contrast_label = QLabel("Bright Region Contrast:")
        self.high_slider = QSlider(Qt.Horizontal)
        self.high_slider.setRange(0, 100)  # Range for contrast adjustment (0%-200%)
        self.high_slider.setValue(100)  # Default value (100% = no change)
        self.high_slider.valueChanged.connect(self.update_image_contrast)
        contrast_gamma_layout.addWidget(high_contrast_label, 0, 0)
        contrast_gamma_layout.addWidget(self.high_slider, 0, 1, 1, 2)
    
        # Dark Region Contrast Slider
        low_contrast_label = QLabel("Dark Region Contrast:")
        self.low_slider = QSlider(Qt.Horizontal)
        self.low_slider.setRange(0, 100)  # Range for contrast adjustment (0%-200%)
        self.low_slider.setValue(0)  # Default value (100% = no change)
        self.low_slider.valueChanged.connect(self.update_image_contrast)
        contrast_gamma_layout.addWidget(low_contrast_label, 1, 0)
        contrast_gamma_layout.addWidget(self.low_slider, 1, 1, 1, 2)
    
        # Gamma Adjustment Slider
        gamma_label = QLabel("Gamma Adjustment:")
        self.gamma_slider = QSlider(Qt.Horizontal)
        self.gamma_slider.setRange(0, 200)  # Range for gamma adjustment (0%-200%)
        self.gamma_slider.setValue(100)  # Default value (100% = no change)
        self.gamma_slider.valueChanged.connect(self.update_image_gamma)
        contrast_gamma_layout.addWidget(gamma_label, 2, 0)
        contrast_gamma_layout.addWidget(self.gamma_slider, 2, 1, 1, 2)
        
        # Black & White Button
        self.bw_button = QPushButton("Convert to Grayscale")
        self.bw_button.clicked.connect(self.convert_to_black_and_white)
        self.bw_button.setToolTip("Converts the image to grayscale. Shortcut: Ctrl+B or CMD+B")

        contrast_gamma_layout.addWidget(self.bw_button, 3, 0, 1, 3)
        
        #Invert the image
        invert_button = QPushButton("Invert Image")
        invert_button.clicked.connect(self.invert_image)
        invert_button.setToolTip("Inverts the image. Shortcut: Ctrl+I or CMD+I")
        contrast_gamma_layout.addWidget(invert_button, 4, 0, 1, 3)
        
    
        # Reset Button
        reset_button = QPushButton("Reset Contrast and Gamma")
        reset_button.clicked.connect(self.reset_gamma_contrast)
        reset_button.setToolTip("Resets the contrast and gamma options. Does not undo invert or grayscale functions! Please use reset image option")
        contrast_gamma_layout.addWidget(reset_button, 5, 0, 1, 3)  # Span all columns
        
        

        # Connect signals for dynamic updates
        self.font_combo_box.currentFontChanged.connect(self.update_font)
        self.font_size_spinner.valueChanged.connect(self.update_font)
    
        # Set the layout for the contrast and gamma group box
        contrast_gamma_group.setLayout(contrast_gamma_layout)
        layout.addWidget(contrast_gamma_group)
    
        # Add stretch at the end for proper spacing
        layout.addStretch()
    
        return tab
    
    def create_cropping_tab(self):
        """Create the Cropping tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
    
        # Group Box for Alignment Options
        alignment_params_group = QGroupBox("Alignment Options")
        alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        alignment_layout = QVBoxLayout()
    
        
        #Guide Lines
        self.show_guides_label = QLabel("Show Guide Lines")
        self.show_guides_checkbox = QCheckBox("", self)
        self.show_guides_checkbox.setChecked(False)
        self.show_guides_checkbox.stateChanged.connect(self.update_live_view)
        
        # Rotation Angle 
        rotation_layout = QHBoxLayout()
        self.orientation_label = QLabel("Rotation Angle (Degrees)")
        self.orientation_label.setFixedWidth(200)
        self.orientation_slider = QSlider(Qt.Horizontal)
        self.orientation_slider.setRange(-3600, 3600)  # Scale by 10 to allow decimals
        self.orientation_slider.setValue(0)
        self.orientation_slider.setSingleStep(1)
        self.orientation_slider.valueChanged.connect(self.update_live_view)
        
        rotation_layout.addWidget(self.show_guides_label)
        rotation_layout.addWidget(self.show_guides_checkbox)
        rotation_layout.addWidget(self.orientation_label)
        rotation_layout.addWidget(self.orientation_slider)
        
    
        # Align Button
        self.align_button = QPushButton("Align Image")
        self.align_button.clicked.connect(self.align_image)
        
        # Flip Vertical Button
        self.flip_vertical_button = QPushButton("Flip Vertical")
        self.flip_vertical_button.clicked.connect(self.flip_vertical)
    
        # Flip Horizontal Button
        self.flip_horizontal_button = QPushButton("Flip Horizontal")
        self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
    
        alignment_layout.addLayout(rotation_layout)
        alignment_layout.addWidget(self.align_button)
        alignment_layout.addWidget(self.flip_vertical_button)  
        alignment_layout.addWidget(self.flip_horizontal_button)  
        alignment_params_group.setLayout(alignment_layout)
        
        
       # Add Tapering Skew Fix Group
        taper_skew_group = QGroupBox("Skew Fix")
        taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        taper_skew_layout = QVBoxLayout()
    
        # Taper Skew Slider
        self.taper_skew_label = QLabel("Tapering Skew:")
        self.taper_skew_slider = QSlider(Qt.Horizontal)
        self.taper_skew_slider.setRange(-70, 70)  # Adjust as needed
        self.taper_skew_slider.setValue(0)
        self.taper_skew_slider.valueChanged.connect(self.update_live_view)
        
        # Align Button
        self.skew_button = QPushButton("Skew Image")
        self.skew_button.clicked.connect(self.update_skew)
    
        # Add widgets to taper skew layout
        taper_skew_layout.addWidget(self.taper_skew_label)
        taper_skew_layout.addWidget(self.taper_skew_slider)
        taper_skew_layout.addWidget(self.skew_button)
        taper_skew_group.setLayout(taper_skew_layout)
        
        # Group Box for Cropping Options
        cropping_params_group = QGroupBox("Cropping Options")
        cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        cropping_layout = QVBoxLayout()
    
        # Cropping Sliders
        crop_slider_layout = QGridLayout()
    
        crop_x_start_label = QLabel("Crop Left (%)")
        self.crop_x_start_slider = QSlider(Qt.Horizontal)
        self.crop_x_start_slider.setRange(0, 100)
        self.crop_x_start_slider.setValue(0)
        self.crop_x_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
        crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
    
        crop_x_end_label = QLabel("Crop Right (%)")
        self.crop_x_end_slider = QSlider(Qt.Horizontal)
        self.crop_x_end_slider.setRange(0, 100)
        self.crop_x_end_slider.setValue(100)
        self.crop_x_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_end_label, 0, 2)
        crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 3)
    
        crop_y_start_label = QLabel("Crop Top (%)")
        self.crop_y_start_slider = QSlider(Qt.Horizontal)
        self.crop_y_start_slider.setRange(0, 100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
        crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
    
        crop_y_end_label = QLabel("Crop Bottom (%)")
        self.crop_y_end_slider = QSlider(Qt.Horizontal)
        self.crop_y_end_slider.setRange(0, 100)
        self.crop_y_end_slider.setValue(100)
        self.crop_y_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_end_label, 1, 2)
        crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 3)
    
        cropping_layout.addLayout(crop_slider_layout)
    
        # Crop Update Button
        crop_button = QPushButton("Update Crop")
        crop_button.clicked.connect(self.update_crop)
        cropping_layout.addWidget(crop_button)
    
        cropping_params_group.setLayout(cropping_layout)
    
        # Add both group boxes to the main layout
        layout.addWidget(alignment_params_group)
        layout.addWidget(cropping_params_group)
        layout.addWidget(taper_skew_group)
        layout.addStretch()
    
        return tab
    
    def create_white_space_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
    
        # Group Box for Padding Parameters
        padding_params_group = QGroupBox("Add White Space to Image for Placing Markers")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout()  # Use a grid layout for better alignment
    
        # Left space input and label
        left_padding_label = QLabel("Left Padding (px):")
        self.left_padding_input = QLineEdit()
        self.left_padding_input.setText("100")  # Default value
        self.left_padding_input.setPlaceholderText("Enter padding for the left side")
        self.left_padding_input.setToolTip("Enter the number of pixels to add to the left of the image.")
        padding_layout.addWidget(left_padding_label, 0, 0)
        padding_layout.addWidget(self.left_padding_input, 0, 1)
    
        # Right space input and label
        right_padding_label = QLabel("Right Padding (px):")
        self.right_padding_input = QLineEdit()
        self.right_padding_input.setText("100")  # Default value
        self.right_padding_input.setPlaceholderText("Enter padding for the right side")
        self.right_padding_input.setToolTip("Enter the number of pixels to add to the right of the image.")
        padding_layout.addWidget(right_padding_label, 0, 2)
        padding_layout.addWidget(self.right_padding_input, 0, 3)
    
        # Top space input and label
        top_padding_label = QLabel("Top Padding (px):")
        self.top_padding_input = QLineEdit()
        self.top_padding_input.setText("100")  # Default value
        self.top_padding_input.setPlaceholderText("Enter padding for the top")
        self.top_padding_input.setToolTip("Enter the number of pixels to add to the top of the image.")
        padding_layout.addWidget(top_padding_label, 1, 0)
        padding_layout.addWidget(self.top_padding_input, 1, 1)
    
        # Bottom space input and label
        bottom_padding_label = QLabel("Bottom Padding (px):")
        self.bottom_padding_input = QLineEdit()
        self.bottom_padding_input.setText("0")  # Default value
        self.bottom_padding_input.setPlaceholderText("Enter padding for the bottom")
        self.bottom_padding_input.setToolTip("Enter the number of pixels to add to the bottom of the image.")
        padding_layout.addWidget(bottom_padding_label, 1, 2)
        padding_layout.addWidget(self.bottom_padding_input, 1, 3)
    
        # Buttons for Finalize and Reset
        button_layout = QHBoxLayout()
        self.finalize_button = QPushButton("Add Padding")
        self.finalize_button.clicked.connect(self.finalize_image)
        self.finalize_button.setToolTip("Click to apply the specified padding to the image.")
        button_layout.addWidget(self.finalize_button)
        
        self.recommend_button = QPushButton("Recommended Padding")
        self.recommend_button.clicked.connect(self.recommended_values)
        self.recommend_button.setToolTip("Autofill with recommended values")
        button_layout.addWidget(self.recommend_button)
        
        self.clear_padding_button = QPushButton("Clear Padding")
        self.clear_padding_button.clicked.connect(self.clear_padding_values)
        self.clear_padding_button.setToolTip("Clear Padding values")
        button_layout.addWidget(self.clear_padding_button)
        
    
        # self.reset_padding_button = QPushButton("Remove Padding")
        # self.reset_padding_button.clicked.connect(self.remove_padding)
        # self.reset_padding_button.setToolTip("Click to remove all added padding and revert the image.")
        # button_layout.addWidget(self.reset_padding_button)
    
        # Add padding layout and buttons to the group box
        padding_params_group.setLayout(padding_layout)
    
        # Add group box and buttons to the main layout
        layout.addWidget(padding_params_group)
        layout.addLayout(button_layout)
    
        # Add stretch for spacing
        layout.addStretch()
    
        return tab
    def clear_padding_values(self):
        self.bottom_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.left_padding_input.setText("0")
        
    def recommended_values(self):
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(render_width*0.1)))
        self.right_padding_input.setText(str(int(render_width*0.1)))
        self.top_padding_input.setText(str(int(render_height*0.15)))
        self.update_live_view()
        
        
        
    def create_markers_tab(self):
        """Create the Markers tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        marker_options_layout = QHBoxLayout()
    
        # Left/Right Marker Options
        left_right_marker_group = QGroupBox("Left/Right Marker Options")
        left_right_marker_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        left_right_marker_layout = QVBoxLayout()
        
        # Combo box for marker presets or custom option
        self.combo_box = QComboBox(self)
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
        self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
        
        # Textbox to allow modification of marker values (shown when "Custom" is selected)
        self.marker_values_textbox = QLineEdit(self)
        self.marker_values_textbox.setPlaceholderText("Enter custom values as comma-separated list")
        self.marker_values_textbox.setEnabled(False)  # Disable by default
        
        # Rename input for custom option
        self.rename_input = QLineEdit(self)
        self.rename_input.setPlaceholderText("Enter new name for Custom")
        self.rename_input.setEnabled(False)
        
        # Save button for the current configuration
        self.save_button = QPushButton("Save Config", self)
        self.save_button.clicked.connect(self.save_config)
        
        # Delete button 
        self.remove_config_button = QPushButton("Remove Config", self)
        self.remove_config_button.clicked.connect(self.remove_config)
        
        # Add widgets to the Left/Right Marker Options layout
        left_right_marker_layout.addWidget(self.combo_box)
        left_right_marker_layout.addWidget(self.marker_values_textbox)
        left_right_marker_layout.addWidget(self.rename_input)
        left_right_marker_layout.addWidget(self.save_button)
        left_right_marker_layout.addWidget(self.remove_config_button)
        
        # Set layout for the Left/Right Marker Options group
        left_right_marker_group.setLayout(left_right_marker_layout)
        
        # Top Marker Options
        top_marker_group = QGroupBox("Top/Bottom Marker Options")
        top_marker_group.setStyleSheet("QGroupBox { font-weight: bold;}")
        
        # Vertical layout for top marker group
        top_marker_layout = QVBoxLayout()
        
        # Text input for Top Marker Labels (multi-column support)
        self.top_marker_input = QTextEdit(self)
        self.top_marker_input.setText(", ".join(self.top_label))  # Populate with initial values
        self.top_marker_input.setMinimumHeight(50)  # Increase height for better visibility
        self.top_marker_input.setMaximumHeight(120)
        self.top_marker_input.setPlaceholderText("Enter labels for each column, separated by commas. Use new lines for multiple columns.")
        
        # # Button to add a new column
        # self.add_column_button = QPushButton("Add Column")
        # self.add_column_button.clicked.connect(self.add_column)
        
        # # Button to remove the last column
        # self.remove_column_button = QPushButton("Remove Last Column")
        # self.remove_column_button.clicked.connect(self.remove_column)
        
        # # Button to update all labels
        self.update_top_labels_button = QPushButton("Update All Labels")
        self.update_top_labels_button.clicked.connect(self.update_all_labels)
        
        # Layout for column management buttons
        # column_buttons_layout = QHBoxLayout()
        # column_buttons_layout.addWidget(self.add_column_button)
        # column_buttons_layout.addWidget(self.remove_column_button)
        
        # Add widgets to the top marker layout
        top_marker_layout.addWidget(self.top_marker_input)
        # top_marker_layout.addLayout(column_buttons_layout)
        top_marker_layout.addWidget(self.update_top_labels_button)
        
        # Set the layout for the Top Marker Group
        top_marker_group.setLayout(top_marker_layout)
        
        # Add both groups to the horizontal layout
        marker_options_layout.addWidget(left_right_marker_group)
        marker_options_layout.addWidget(top_marker_group)
        
        # Add the horizontal layout to the main layout
        layout.addLayout(marker_options_layout)
    
        # Marker padding sliders - Group box for marker distance adjustment
        padding_params_group = QGroupBox("Marker Placement and Distance")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        
        # Grid layout for the marker group box
        padding_layout = QGridLayout()
        
        # Left marker: Button, slider, reset, and duplicate in the same row
        left_marker_button = QPushButton("Left Markers")
        left_marker_button.setToolTip("Places the left markers at the exact location of the mouse pointer on the left. Shortcut: Ctrl+Shift+L or CMD+Shift+L")
        left_marker_button.clicked.connect(self.enable_left_marker_mode)
        self.left_padding_slider = QSlider(Qt.Horizontal)
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.left_padding_slider.setValue(0)
        self.left_padding_slider.valueChanged.connect(self.update_left_padding)
        
        remove_left_button = QPushButton("Remove Last")
        remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
        
        reset_left_button = QPushButton("Reset")
        reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
        
        duplicate_left_button = QPushButton("Duplicate Right")
        duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
        
        # Add left marker widgets to the grid layout
        padding_layout.addWidget(left_marker_button, 0, 0)
        padding_layout.addWidget(remove_left_button, 0, 1)
        padding_layout.addWidget(reset_left_button, 0, 2)
        padding_layout.addWidget(duplicate_left_button, 0, 3)
        padding_layout.addWidget(self.left_padding_slider, 0, 4,1,2)
        
        # Right marker: Button, slider, reset, and duplicate in the same row
        right_marker_button = QPushButton("Right Markers")
        right_marker_button.setToolTip("Places the right markers at the exact location of the mouse pointer on the right. Shortcut: Ctrl+Shift+R or CMD+Shift+R")
        right_marker_button.clicked.connect(self.enable_right_marker_mode)
        self.right_padding_slider = QSlider(Qt.Horizontal)
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.right_padding_slider.setValue(0)
        self.right_padding_slider.valueChanged.connect(self.update_right_padding)
        
        remove_right_button = QPushButton("Remove Last")
        remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
        
        reset_right_button = QPushButton("Reset")
        reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
        
        duplicate_right_button = QPushButton("Duplicate Left")
        duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
        
        # Add right marker widgets to the grid layout
        padding_layout.addWidget(right_marker_button, 1, 0)
        padding_layout.addWidget(remove_right_button, 1, 1)
        padding_layout.addWidget(reset_right_button, 1, 2)
        padding_layout.addWidget(duplicate_right_button, 1, 3)
        padding_layout.addWidget(self.right_padding_slider, 1, 4,1,2)
        
        # Top marker: Button, slider, and reset in the same row
        top_marker_button = QPushButton("Top Markers")
        top_marker_button.setToolTip("Places the top markers at the exact location of the mouse pointer on the top. Shortcut: Ctrl+Shift+T or CMD+Shift+T")
        top_marker_button.clicked.connect(self.enable_top_marker_mode)
        self.top_padding_slider = QSlider(Qt.Horizontal)
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.top_padding_slider.setValue(0)
        self.top_padding_slider.valueChanged.connect(self.update_top_padding)
        
        remove_top_button = QPushButton("Remove Last")
        remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
        
        reset_top_button = QPushButton("Reset")
        reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
        
        # Add top marker widgets to the grid layout
        padding_layout.addWidget(top_marker_button, 2, 0)
        padding_layout.addWidget(remove_top_button, 2, 1)
        padding_layout.addWidget(reset_top_button, 2, 2)
        padding_layout.addWidget(self.top_padding_slider, 2, 3, 1, 3)  # Slider spans 2 columns for better alignment
        
        for i in range(6):  # Assuming 6 columns in the grid
            padding_layout.setColumnStretch(i, 1)
        
        # Add button and QLineEdit for the custom marker
        self.custom_marker_button = QPushButton("Custom Marker", self)
        self.custom_marker_button.setToolTip("Places custom markers at the middle of the mouse pointer")
    
        self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
        
        self.custom_marker_button_left_arrow = QPushButton("←", self)
        
        self.custom_marker_button_right_arrow = QPushButton("→", self)
        
        self.custom_marker_button_top_arrow = QPushButton("↑", self)
        
        self.custom_marker_button_bottom_arrow = QPushButton("↓", self)
        
        self.custom_marker_text_entry = QLineEdit(self)        
        self.custom_marker_text_entry.setPlaceholderText("Enter custom marker text")
        
        self.remove_custom_marker_button = QPushButton("Remove Last", self)
        self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
        
        self.reset_custom_marker_button = QPushButton("Reset", self)
        self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
        
        # Add color selection button for custom markers
        self.custom_marker_color_button = QPushButton("Custom Marker Color")
        self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
        
        marker_buttons_layout = QHBoxLayout()
        
        # Add the arrow buttons with fixed sizes to the marker buttons layout
        self.custom_marker_button_left_arrow.setFixedSize(30, 30)
    
        marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow)
        
        self.custom_marker_button_right_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
        
        self.custom_marker_button_top_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow)
        
        self.custom_marker_button_bottom_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
        
        #Assign functions to the buttons
        self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))
    
        
        # Create a QWidget to hold the QHBoxLayout
        marker_buttons_widget = QWidget()
        marker_buttons_widget.setLayout(marker_buttons_layout)
        
        # Add the custom marker button
        padding_layout.addWidget(self.custom_marker_button, 3, 0)
        
        # Add the text entry for the custom marker
        padding_layout.addWidget(self.custom_marker_text_entry, 3, 1,1,1)
        
        # Add the marker buttons widget to the layout
        padding_layout.addWidget(marker_buttons_widget, 3, 2) 
        
        # Add the remove button
        padding_layout.addWidget(self.remove_custom_marker_button, 3, 3)
        
        # Add the reset button
        padding_layout.addWidget(self.reset_custom_marker_button, 3, 4)
        
        # Add the color button
        padding_layout.addWidget(self.custom_marker_color_button, 3, 5)
        
        self.custom_font_type_label = QLabel("Custom Marker Font:", self)
        self.custom_font_type_dropdown = QFontComboBox()
        self.custom_font_type_dropdown.setCurrentFont(QFont("Arial"))
        self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
        
        # Font size selector
        self.custom_font_size_label = QLabel("Custom Marker Size:", self)
        self.custom_font_size_spinbox = QSpinBox(self)
        self.custom_font_size_spinbox.setRange(2, 150)  # Allow font sizes from 8 to 72
        self.custom_font_size_spinbox.setValue(12)  # Default font size
        
        # Grid checkbox
        self.show_grid_checkbox = QCheckBox("Show Snap Grid", self)
        self.show_grid_checkbox.setToolTip("Places a snapping grid and the text or marker will be places at the center of the grid. Shortcut: Ctrl+Shift+G. To increase or decrease the grid size: Ctrl+Shift+Up or Down arrow or CMD+Shift+Up or Down arrow ")
        self.show_grid_checkbox.setChecked(False)  # Default: Grid is off
        self.show_grid_checkbox.stateChanged.connect(self.update_live_view)
        
        # Grid size input (optional
        self.grid_size_input = QSpinBox(self)
        self.grid_size_input.setRange(5, 100)  # Grid cell size in pixels
        self.grid_size_input.setValue(20)  # Default grid size
        self.grid_size_input.setPrefix("Grid Size (px): ")
        self.grid_size_input.valueChanged.connect(self.update_live_view)
        
        # Add font type and size widgets to the layout
        padding_layout.addWidget(self.custom_font_type_label, 4, 0,)
        padding_layout.addWidget(self.custom_font_type_dropdown, 4, 1,1,1)  # Span 2 columns
        
        padding_layout.addWidget(self.custom_font_size_label, 4, 2)
        padding_layout.addWidget(self.custom_font_size_spinbox, 4, 3,1,1)
        padding_layout.addWidget(self.show_grid_checkbox,4,4)
        padding_layout.addWidget(self.grid_size_input,4,5)
    
        # Set the layout for the marker group box
        padding_params_group.setLayout(padding_layout)
        
        # Add the font options group box to the main layout
        layout.addWidget(padding_params_group)
        
        layout.addStretch()
        return tab
    
    def add_column(self):
        """Add a new column to the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        if current_text.strip():
            self.top_marker_input.append("")  # Add a new line for a new column
        else:
            self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
    
    def remove_column(self):
        """Remove the last column from the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        lines = current_text.split("\n")
        if len(lines) > 1:
            lines.pop()  # Remove the last line
            self.top_marker_input.setPlainText("\n".join(lines))
        else:
            self.top_marker_input.clear()  # Clear the text if only one line exists

    
    def flip_vertical(self):
        self.save_state()
        """Flip the image vertically."""
        if self.image:
            self.image = self.image.mirrored(vertical=True, horizontal=False)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def flip_horizontal(self):
        self.save_state()
        """Flip the image horizontally."""
        if self.image:
            self.image = self.image.mirrored(vertical=False, horizontal=True)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def convert_to_black_and_white(self):
        self.save_state()
        """Convert the image to black and white."""
        if self.image:
            grayscale_image = self.image.convertToFormat(QImage.Format_Grayscale8)
            self.image = grayscale_image
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()


    def invert_image(self):
        self.save_state()
        if self.image:
            inverted_image = self.image.copy()
            inverted_image.invertPixels()
            self.image = inverted_image
            self.update_live_view()
        self.image_before_contrast=self.image.copy()
        self.image_before_padding=self.image.copy()
        self.image_contrasted=self.image.copy()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if self.live_view_label.preview_marker_enabled:
                self.live_view_label.preview_marker_enabled = False  # Turn off the preview
            self.live_view_label.measure_quantity_mode = False
            self.live_view_label.bounding_box_complete==False
            self.live_view_label.counter=0
            self.live_view_label.quad_points=[]
            self.live_view_label.bounding_box_preview = None
            self.live_view_label.rectangle_points = []
            self.live_view_label.mousePressEvent=None
            self.live_view_label.mode=None

        if self.live_view_label.zoom_level != 1.0:  # Only allow panning when zoomed in
            step = 20  # Adjust the panning step size as needed
            if event.key() == Qt.Key_Left:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - step)
            elif event.key() == Qt.Key_Right:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + step)
            elif event.key() == Qt.Key_Up:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - step)
            elif event.key() == Qt.Key_Down:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + step)
        self.update_live_view()
        super().keyPressEvent(event)
    
    def update_marker_text_font(self, font: QFont):
        """
        Updates the font of self.custom_marker_text_entry based on the selected font
        from self.custom_font_type_dropdown.
    
        :param font: QFont object representing the selected font from the combobox.
        """
        # Get the name of the selected font
        selected_font_name = font.family()
        
        # Set the font of the QLineEdit
        self.custom_marker_text_entry.setFont(QFont(selected_font_name))
    
    def arrow_marker(self, text: str):
        self.custom_marker_text_entry.clear()
        self.custom_marker_text_entry.setText(text)
    
    def enable_custom_marker_mode(self):
        """Enable the custom marker mode and set the mouse event."""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        custom_text = self.custom_marker_text_entry.text().strip()
    
        self.live_view_label.preview_marker_enabled = True
        self.live_view_label.preview_marker_text = custom_text
        self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
        self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
        self.live_view_label.marker_color=self.custom_marker_color
        
        self.live_view_label.setFocus()
        self.live_view_label.update()
        
        self.marker_mode = "custom"  # Indicate custom marker mode
        self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
        
    def remove_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers.pop()  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    def reset_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers=[]  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    
    def place_custom_marker(self, event, custom_text):
        """Place a custom marker at the cursor location."""
        self.save_state()
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            
        # Adjust for snapping to grid
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        image_x = (cursor_x - x_offset) / scale
        image_y = (cursor_y - y_offset) / scale
        
            
        # Store the custom marker's position and text
        self.custom_markers = getattr(self, "custom_markers", [])
        self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value()))
        # print("CUSTOM_MARKER: ",self.custom_markers)
        # Update the live view to render the custom marker
        self.update_live_view()
        
    def select_custom_marker_color(self):
        """Open a color picker dialog to select the color for custom markers."""
        color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
        if color.isValid():
            self.custom_marker_color = color  # Update the custom marker color
    
    def update_all_labels(self):
        """Update all labels from the QTextEdit, supporting multiple columns."""
        self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
        input_text = self.top_marker_input.toPlainText()
        
        # Split the text into a list by commas and strip whitespace
        self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
        

        # if len(self.top_label) < len(self.top_markers):
        #     self.top_markers = self.top_markers[:len(self.top_label)]
        for i in range(0, len(self.top_markers)):
            try:
                self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
            except:
                self.top_markers[i] = (self.top_markers[i][0], str(""))

        # if len(self.marker_values) < len(self.left_markers):
        #     self.left_markers = self.left_markers[:len(self.marker_values)]
        for i in range(0, len(self.left_markers)):
            try:
                self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
            except:
                self.left_markers[i] = (self.left_markers[i][0], str(""))
                

        # if len(self.marker_values) < len(self.right_markers):
        #     self.right_markers = self.right_markers[:len(self.marker_values)]
        for i in range(0, len(self.right_markers)):
            try:
                self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
            except:
                self.right_markers[i] = (self.right_markers[i][0], str(""))
                

        
        # Trigger a refresh of the live view
        self.update_live_view()
        
    def reset_marker(self, marker_type, param):       
        if marker_type == 'left':
            if param == 'remove' and len(self.left_markers)!=0:
                self.left_markers.pop()  
                if self.current_marker_index > 0:
                    self.current_marker_index -= 1
            elif param == 'reset':
                self.left_markers.clear()
                self.current_marker_index = 0  

             
        elif marker_type == 'right' and len(self.right_markers)!=0:
            if param == 'remove':
                self.right_markers.pop()  
                if self.current_marker_index > 0:
                    self.current_marker_index -= 1
            elif param == 'reset':
                self.right_markers.clear()
                self.current_marker_index = 0

        elif marker_type == 'top' and len(self.top_markers)!=0:
            if param == 'remove':
                self.top_markers.pop() 
                if self.current_top_label_index > 0:
                    self.current_top_label_index -= 1
            elif param == 'reset':
                self.top_markers.clear()
                self.current_top_label_index = 0
    
        # Call update live view after resetting markers
        self.update_live_view()
        
    def duplicate_marker(self, marker_type):
        if marker_type == 'left' and self.right_markers:
            self.left_markers = self.right_markers.copy()  # Copy right markers to left
        elif marker_type == 'right' and self.left_markers:
            self.right_markers = self.left_markers.copy()  # Copy left markers to right

            # self.right_padding = self.image_width*0.9
    
        # Call update live view after duplicating markers
        self.update_live_view()
        
    def on_combobox_changed(self):
        
        text=self.combo_box.currentText()
        """Handle the ComboBox change event to update marker values."""
        if text == "Custom":
            # Enable the textbox for custom values when "Custom" is selected
            self.marker_values_textbox.setEnabled(True)
            self.rename_input.setEnabled(True)
            
            # self.marker_values_textbox.setText()
        else:
            # Update marker values based on the preset option
            self.marker_values = self.marker_values_dict.get(text, [])
            self.marker_values_textbox.setEnabled(False)  # Disable the textbox
            self.rename_input.setEnabled(False)
            self.top_label = self.top_label_dict.get(text, [])
            self.top_label = [str(item) if not isinstance(item, str) else item for item in self.top_label]
            self.top_marker_input.setText(", ".join(self.top_label))
            try:
                
                # Ensure that the top_markers list only updates the top_label values serially
                for i in range(0, len(self.top_markers)):
                    self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                
                # # If self.top_label has more entries than current top_markers, add them
                # if len(self.top_label) > len(self.top_markers):
                #     additional_markers = [(self.top_markers[-1][0] + 50 * (i + 1), label) 
                #                           for i, label in enumerate(self.top_label[len(self.top_markers):])]
                #     self.top_markers.extend(additional_markers)
                
                # If self.top_label has fewer entries, truncate the list
                if len(self.top_label) < len(self.top_markers):
                    self.top_markers = self.top_markers[:len(self.top_label)]
                    
                
                
            except:
                pass
            try:
                self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
            except:
                pass

    # Functions for updating contrast and gamma
    
    def reset_gamma_contrast(self):
        try:
            if self.image_before_contrast==None:
                self.image_before_contrast=self.image_master.copy()
            self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
            self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
            self.high_slider.setValue(100)  # Reset contrast to default
            self.low_slider.setValue(0)  # Reset contrast to default
            self.gamma_slider.setValue(100)  # Reset gamma to default
            self.update_live_view()
        except:
            pass

    
    def update_image_contrast(self):
        self.save_state()
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
            
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                self.update_live_view()
        except:
            pass
    
    def update_image_gamma(self):
        self.save_state()
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
                
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                self.update_live_view()
        except:
            pass
    
    def apply_contrast_gamma(self, qimage, high, low, gamma):
        """
        Applies contrast and gamma adjustments to a QImage.
        Converts the QImage to RGBA format, performs the adjustments, and returns the modified QImage.
        """
        #Ensure the image is in the correct format (RGBA8888)
        if qimage.format() != QImage.Format_RGBA8888:
            qimage = qimage.convertToFormat(QImage.Format_RGBA8888)

        # Convert the QImage to a NumPy array
        width = qimage.width()
        height = qimage.height()

        # Get the raw byte data from QImage
        ptr = qimage.bits()
        ptr.setsize(height * width * 4)  # 4 bytes per pixel (RGBA)

        # Create a NumPy array from the pointer (for RGBA format)
        img_array = np.array(ptr).reshape(height, width, 4).astype(np.float32)

        # Normalize the image for contrast and gamma adjustments
        img_array = img_array / 255.0

        # Apply contrast adjustment
        img_array = np.clip((img_array - low) / (high - low), 0, 1)

        # Apply gamma correction
        img_array = np.power(img_array, gamma)

        # Scale back to [0, 255] range
        img_array = (img_array * 255).astype(np.uint8)

        # Convert back to QImage
        qimage = QImage(img_array.data, width, height, img_array.strides[0], QImage.Format_RGBA8888)

        return qimage
    

    def save_contrast_options(self):
        if self.image:
            self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
            self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
        else:
            QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

    def remove_config(self):
        try:
            # Get the currently selected marker label
            selected_marker = self.combo_box.currentText()
    
            # Ensure the selected marker is not "Custom" before deleting
            if selected_marker == "Custom":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' marker.")
                return
            
            elif selected_marker == "Precision Plus All Blue/Unstained":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
            
            elif selected_marker == "1 kB Plus":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
    
            # Remove the marker label and top_label if they exist
            if selected_marker in self.marker_values_dict:
                del self.marker_values_dict[selected_marker]
            if selected_marker in self.top_label_dict:
                del self.top_label_dict[selected_marker]
    
            # Save the updated configuration
            with open("Imaging_assistant_config.txt", "w") as f:
                config = {
                    "marker_values": self.marker_values_dict,
                    "top_label": self.top_label_dict
                }
                json.dump(config, f)
    
            # Remove from the ComboBox and reset UI
            self.combo_box.removeItem(self.combo_box.currentIndex())
            self.top_marker_input.clear()
    
            QMessageBox.information(self, "Success", f"Configuration '{selected_marker}' removed.")
    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error removing config: {e}")

    def save_config(self):
        """Rename the 'Custom' marker values option and save the configuration."""
        new_name = self.rename_input.text().strip()
        
        # Ensure that the correct dictionary (self.marker_values_dict) is being used
        if self.rename_input.text() != "Enter new name for Custom" and self.rename_input.text() != "":  # Correct condition
            # Save marker values list
            self.marker_values_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
            
            # Save top_label list under the new_name key
            self.top_label_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.top_marker_input.toPlainText().strip("[]").split(",")]

            try:
                # Save both the marker values and top label (under new_name) to the config file
                with open("Imaging_assistant_config.txt", "w") as f:
                    config = {
                        "marker_values": self.marker_values_dict,
                        "top_label": self.top_label_dict  # Save top_label_dict as a dictionary with new_name as key
                    }
                    json.dump(config, f)  # Save both marker_values and top_label_dict
            except Exception as e:
                print(f"Error saving config: {e}")

        self.combo_box.setCurrentText(new_name)
        self.load_config()  # Reload the configuration after saving
    
    
    def load_config(self):
        """Load the configuration from the file."""
        try:
            with open("Imaging_assistant_config.txt", "r") as f:
                config = json.load(f)
                
                # Load marker values and top label from the file
                self.marker_values_dict = config.get("marker_values", {})
                
                # Retrieve top_label list from the dictionary using the new_name key
                new_name = self.rename_input.text().strip()  # Assuming `new_name` is defined here; otherwise, set it manually
                self.top_label_dict = config.get("top_label", {})  # Default if not found
                # self.top_label = self.top_label_dict.get(new_name, ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])  # Default if not found
                
        except FileNotFoundError:
            # Set default marker values and top_label if config is not found
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }  # Default top labels
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
        except Exception as e:
            print(f"Error loading config: {e}")
            # Fallback to default values if there is an error
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
    
        # Update combo box options with loaded marker values
        self.combo_box.clear()
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
    
    def paste_image(self):
        """Handle pasting image from clipboard."""
        self.load_image_from_clipboard()
        self.update_live_view()
    
    def load_image_from_clipboard(self):
        """Load an image from the clipboard into self.image."""
        self.reset_image()
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
    
        # Check if the clipboard contains an image
        if mime_data.hasImage():
            image = clipboard.image()  # Get the image directly from the clipboard
            self.image = image  # Store the image in self.image
            self.original_image = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_master = self.image.copy()
            self.image_before_padding = None
    
        # Check if the clipboard contains URLs (file paths)
        if mime_data.hasUrls():
            urls = mime_data.urls()
            if urls:
                file_path = urls[0].toLocalFile()  # Get the first file path
                if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                    # Load the image from the file path
                    self.image = QImage(file_path)
                    self.original_image = self.image.copy()
                    self.image_contrasted = self.image.copy()
                    self.image_master = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_before_padding = None
                    
    
                    # Update the window title with the image path
                    self.setWindowTitle(f"{self.window_title}: {file_path}")
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
       
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(render_width*0.1)))
        self.right_padding_input.setText(str(int(render_width*0.1)))
        self.top_padding_input.setText(str(int(render_height*0.15)))
        self.update_live_view()
        self.save_state()
        
    def update_font(self):
        """Update the font settings based on UI inputs"""
        # Update font family from the combo box
        self.font_family = self.font_combo_box.currentFont().family()
        
        # Update font size from the spin box
        self.font_size = self.font_size_spinner.value()
        
        self.font_rotation = int(self.font_rotation_input.value())
        
    
        # Once font settings are updated, update the live view immediately
        self.update_live_view()
    
    def select_font_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.font_color = color  # Store the selected color   
            self.update_font()
            
    def load_image(self):
        self.undo_stack = []  # Reset the undo stack
        self.redo_stack = []  # Reset the redo stack
        self.reset_image()
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif)", options=options
        )
        if file_path:
            self.image_path = file_path
            self.original_image = QImage(self.image_path)  # Keep the original image
            self.image = self.original_image.copy()        # Start with a copy of the original
            self.image_master= self.original_image.copy() 
            self.image_before_padding=None
            self.image_before_contrast=self.original_image.copy() 
            self.image_contrasted= self.original_image.copy()  

            self.setWindowTitle(f"{self.window_title}:{self.image_path}")
    
            # Determine associated config file
            self.base_name = os.path.splitext(os.path.basename(file_path))[0]
            if self.base_name.endswith("_original"):
                config_name = self.base_name.replace("_original", "_config.txt")
            else:
                config_name = self.base_name + "_config.txt"
            config_path = os.path.join(os.path.dirname(file_path), config_name)
    
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r") as config_file:
                        config_data = json.load(config_file)
                    self.apply_config(config_data)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to load config file: {e}")
        # Preview window
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
        
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(render_width*0.1)))
        self.right_padding_input.setText(str(int(render_width*0.1)))
        self.top_padding_input.setText(str(int(render_height*0.15)))
        self.update_live_view()
        self.save_state()
    
    def apply_config(self, config_data):
        self.left_padding_input.setText(config_data["adding_white_space"]["left"])
        self.right_padding_input.setText(config_data["adding_white_space"]["right"])
        self.top_padding_input.setText(config_data["adding_white_space"]["top"])
        try:
            self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
        except:
            pass
    
        self.crop_x_start_slider.setValue(config_data["cropping_parameters"]["x_start_percent"])
        self.crop_x_end_slider.setValue(config_data["cropping_parameters"]["x_end_percent"])
        self.crop_y_start_slider.setValue(config_data["cropping_parameters"]["y_start_percent"])
        self.crop_y_end_slider.setValue(config_data["cropping_parameters"]["y_end_percent"])
    
        try:
            self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
            self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
            self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
        except (KeyError, ValueError) as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
            pass
        try:
            # print("TOP LABELS: ",config_data["marker_labels"]["top"])
            self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
            self.top_marker_input.setText(", ".join(self.top_label))
        except KeyError as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
            pass
    
        self.font_family = config_data["font_options"]["font_family"]
        self.font_size = config_data["font_options"]["font_size"]
        self.font_rotation = config_data["font_options"]["font_rotation"]
        self.font_color = QColor(config_data["font_options"]["font_color"])
    
        self.top_padding_slider.setValue(config_data["marker_padding"]["top"])
        self.left_padding_slider.setValue(config_data["marker_padding"]["left"])
        self.right_padding_slider.setValue(config_data["marker_padding"]["right"])
        
        try:
            try:
                x = list(map(int, config_data["marker_labels"]["left"]))
                self.marker_values_textbox.setText(str(x))
            except:
                x = list(map(int, config_data["marker_labels"]["right"]))
                self.marker_values_textbox.setText(str(x))
            if self.marker_values_textbox.text!="":
                self.combo_box.setCurrentText("Custom")
                self.marker_values_textbox.setEnabled(True)                
            else:
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.marker_values_textbox.setEnabled(False)
                
        except:
            # print("ERROR IN LEFT/RIGHT MARKER DATA")
            pass
    
        try:
            self.custom_markers = [
                (marker["x"], marker["y"], marker["text"], QColor(marker["color"]), marker["font"], marker["font_size"])
                for marker in config_data.get("custom_markers", [])
            ]
                
                
                
        except:
            pass
        # Set slider ranges from config_data
        try:
            self.left_padding_slider.setRange(
                int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
            )
            self.right_padding_slider.setRange(
                int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
            )
            self.top_padding_slider.setRange(
                int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
            )
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Slider ranges not found in config_data.")
            pass
        
        try:
            self.left_marker_shift_added=int(config_data["added_shift"]["left"])
            self.right_marker_shift_added=int(config_data["added_shift"]["right"])
            self.top_marker_shift_added=int(config_data["added_shift"]["top"])
            
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Added Shift not found in config_data.");
            pass
            
        # Apply font settings

        
        #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
        
        font_size_new=self.font_size
        font_rotation_new=self.font_rotation
        
        self.font_combo_box.setCurrentFont(QFont(self.font_family))
        self.font_size_spinner.setValue(font_size_new)
        self.font_rotation_input.setValue(font_rotation_new)

        self.update_live_view()
        
    def get_current_config(self):
        config = {
            "adding_white_space": {
                "left": self.left_padding_input.text(),
                "right": self.right_padding_input.text(),
                "top": self.top_padding_input.text(),
                "bottom": self.bottom_padding_input.text(),
            },
            "cropping_parameters": {
                "x_start_percent": self.crop_x_start_slider.value(),
                "x_end_percent": self.crop_x_end_slider.value(),
                "y_start_percent": self.crop_y_start_slider.value(),
                "y_end_percent": self.crop_y_end_slider.value(),
            },
            "marker_positions": {
                "left": self.left_markers,
                "right": self.right_markers,
                "top": self.top_markers,
            },
            "marker_labels": {
                "top": self.top_label,
                "left": [marker[1] for marker in self.left_markers],
                "right": [marker[1] for marker in self.right_markers],
            },
            "marker_padding": {
                "top": self.top_padding_slider.value(),
                "left": self.left_padding_slider.value(),
                "right": self.right_padding_slider.value(),
            },
            "font_options": {
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_rotation": self.font_rotation,
                "font_color": self.font_color.name(),
            },
        }
    
        try:
            # Add custom markers with font and font size
            config["custom_markers"] = [
                {"x": x, "y": y, "text": text, "color": color.name(), "font": font, "font_size": font_size}
                for x, y, text, color, font, font_size in self.custom_markers
            ]
        except AttributeError:
            # Handle the case where self.custom_markers is not defined or invalid
            config["custom_markers"] = []
        try:
            config["slider_ranges"] = {
                    "left": self.left_slider_range,
                    "right": self.right_slider_range,
                    "top": self.top_slider_range,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["slider_ranges"] = []
            
        try:
            config["added_shift"] = {
                    "left": self.left_marker_shift_added,
                    "right": self.right_marker_shift_added,
                    "top": self.top_marker_shift_added,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["added_shift"] = []
    
        return config
    
    def add_band(self, event):
        self.save_state()
        # Ensure there's an image loaded and marker mode is active
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        if not self.image or not self.marker_mode:
            return
    
        # Get the cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
        
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Get the dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Get the actual dimensions of the loaded image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate the scaling factor (assuming uniform scaling to maintain aspect ratio)
        scale = min(displayed_width / image_width, displayed_height / image_height)
        scale_x = displayed_width / image_width
        scale_y = displayed_height / image_height
    
        # Calculate offsets if the image is centered in the live_view_label
        offset_x = (displayed_width - image_width * scale_x) / 2
        offset_y = (displayed_height - image_height * scale_y) / 2
    
        # Transform cursor coordinates to the image coordinate space
        image_x = (cursor_x - offset_x) / scale_x
        image_y = (cursor_y - offset_y) / scale_y
        
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
		
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Validate that the transformed coordinates are within image bounds
        if not (0 <= image_y <= image_height):
            return  # Ignore clicks outside the image bounds
        try:
        # Add the band marker based on the active marker mode
            if self.marker_mode == "left" and self.current_marker_index < len(self.marker_values):
                if len(self.left_markers)!=0:
                    self.left_markers.append((image_y, self.marker_values[len(self.left_markers)]))                    
                    self.current_marker_index += 1
                else:
                    self.left_markers.append((image_y, self.marker_values[self.current_marker_index]))
                    self.current_marker_index += 1
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.left_padding_slider.setValue(0)
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.left_padding_slider.setValue(padding_value)
                    self.left_marker_shift_added = self.left_padding_slider.value()                    
            elif self.marker_mode == "right" and self.current_marker_index < len(self.marker_values):
                if len(self.right_markers)!=0:
                    self.right_markers.append((image_y, self.marker_values[len(self.right_markers)]))
                    self.current_marker_index += 1
                else:
                    self.right_markers.append((image_y, self.marker_values[self.current_marker_index]))
                    self.current_marker_index += 1
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.right_padding_slider.setValue(0)
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.right_padding_slider.setValue(padding_value)
                    self.right_marker_shift_added = self.right_padding_slider.value()
            elif self.marker_mode == "top" and self.current_top_label_index < len(self.top_label):
                if len(self.top_markers)!=0:
                    self.top_markers.append((image_x, self.top_label[len(self.top_markers)]))
                    self.current_top_label_index += 1
                else:
                    self.top_markers.append((image_x, self.top_label[self.current_top_label_index]))
                    self.current_top_label_index += 1
                    padding_value=int((image_y - y_start) * (render_height / self.image.height()))
                    self.top_padding_slider.setValue(0)
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
                    self.top_padding_slider.setValue(padding_value)
                    self.top_marker_shift_added = self.top_padding_slider.value()
        except:
            print("ERROR ADDING BANDS")
        self.update_live_view()
        

        
    def enable_left_marker_mode(self):
        self.marker_mode = "left"
        self.current_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)

    def enable_right_marker_mode(self):
        self.marker_mode = "right"
        self.current_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_top_marker_mode(self):
        self.marker_mode = "top"
        self.current_top_label_index
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
        
    # def remove_padding(self):
    #     if self.image_before_padding!=None:
    #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
    #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
    #     self.image_padded = False  # Reset the padding state
    #     w=self.image.width()
    #     h=self.image.height()
    #     # Preview window
    #     ratio=w/h
    #     self.label_width = 540
    #     label_height=int(self.label_width/ratio)
    #     if label_height>self.label_width:
    #         label_height=540
    #         self.label_width=ratio*label_height
    #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    #     self.update_live_view()
        
    def finalize_image(self):
        self.save_state()
        # Get the padding values from the text inputs
        try:
            padding_left = abs(int(self.left_padding_input.text()))
            padding_right = abs(int(self.right_padding_input.text()))
            padding_top = abs(int(self.top_padding_input.text()))
            padding_bottom = abs(int(self.bottom_padding_input.text()))
        except ValueError:
            # Handle invalid input (non-integer value)
            print("Please enter valid integers for padding.")
            return
    
        # Adjust marker positions based on padding
        self.adjust_markers_for_padding(padding_left, padding_right, padding_top, padding_bottom)
    
        # Ensure self.image_before_padding is initialized
        if self.image_before_padding is None:
            self.image_before_padding = self.image_contrasted.copy()
    
        # Reset to the original image if padding inputs have changed
        if self.image_padded:
            self.image = self.image_before_padding.copy()
            self.image_padded = False
    
        # Calculate the new dimensions with padding
        new_width = self.image.width() + padding_left + padding_right
        new_height = self.image.height() + padding_top + padding_bottom
    
        # Create a new blank image with white background
        padded_image = QImage(new_width, new_height, QImage.Format_RGB888)
        padded_image.fill(QColor(255, 255, 255))  # Fill with white
    
        # Copy the original image onto the padded image
        for y in range(self.image.height()):
            for x in range(self.image.width()):
                color = self.image.pixel(x, y)
                padded_image.setPixel(padding_left + x, padding_top + y, color)
    
        self.image = padded_image
        self.image_padded = True
        self.image_contrasted = self.image.copy()
        w = self.image.width()
        h = self.image.height()
        # Preview window
        ratio = w / h
        self.label_width = int(self.screen_width * 0.28)
        label_height = int(self.label_width / ratio)
        if label_height > self.label_width:
            label_height = self.label_width
            self.label_width = ratio * label_height
        self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range = [-100, int(render_width) + 100]
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.right_slider_range = [-100, int(render_width) + 100]
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.top_slider_range = [-100, int(render_height) + 100]
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.update_live_view()
    
    def adjust_markers_for_padding(self, padding_left, padding_right, padding_top, padding_bottom):
        """Adjust marker positions based on padding."""
        # Adjust left markers
        self.left_markers = [(y + padding_top, label) for y, label in self.left_markers]
        # Adjust right markers
        self.right_markers = [(y + padding_top, label) for y, label in self.right_markers]
        # Adjust top markers
        self.top_markers = [(x + padding_left, label) for x, label in self.top_markers]        
        
    
    def update_left_padding(self):
        # Update left padding when slider value changes
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.update_live_view()

    def update_right_padding(self):
        # Update right padding when slider value changes
        self.right_marker_shift_added = self.right_padding_slider.value()
        self.update_live_view()
        
    def update_top_padding(self):
        # Update top padding when slider value changes
        self.top_marker_shift_added = self.top_padding_slider.value()
        self.update_live_view()

    def update_live_view(self):
        if not self.image:
            return
    
        # Enable the "Predict Molecular Weight" button if markers are present
        if self.left_markers or self.right_markers:
            self.predict_button.setEnabled(True)
        else:
            self.predict_button.setEnabled(False)
    
        # Define a higher resolution for processing (e.g., 2x or 3x label size)
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Calculate scaling factors and offsets
        scale_x = self.image.width() / render_width
        scale_y = self.image.height() / render_height
    
        # Get the crop percentage values from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure the cropping boundaries are valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            self.crop_x_start_slider.setValue(0)
            self.crop_x_end_slider.setValue(100)
            self.crop_y_start_slider.setValue(0)
            self.crop_y_end_slider.setValue(100)
            return
    
        # Crop the image based on the defined boundaries
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
    
        # Get the orientation value from the slider
        orientation = float(self.orientation_slider.value() / 20)  # Orientation slider value
        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f} Degrees)")
    
        # Apply the rotation to the cropped image
        rotated_image = cropped_image.transformed(QTransform().rotate(orientation))
    
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1
        self.taper_skew_label.setText(f"Tapering Skew {taper_value:.2f} ")
    
        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        skewed_image = rotated_image.transformed(transform, Qt.SmoothTransformation)
    
        # Scale the rotated image to the rendering resolution
        scaled_image = skewed_image.scaled(
            render_width,
            render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        # Render on a high-resolution canvas
        canvas = QImage(render_width, render_height, QImage.Format_ARGB32)
        canvas.fill(Qt.transparent)  # Transparent background
    
        # Render the base image and overlays
        self.render_image_on_canvas(canvas, scaled_image, x_start, y_start, render_scale)
    
        # Scale the high-resolution canvas down to the label's size for display
        pixmap = QPixmap.fromImage(canvas).scaled(
            self.live_view_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        painter = QPainter(pixmap)
        pen = QPen(Qt.red)
        pen.setWidth(1)
        painter.setPen(pen)
    
        scale_x = self.live_view_label.width() / self.image.width()
        scale_y = self.live_view_label.height() / self.image.height()
    
        # Draw the quadrilateral or rectangle if in move mode
        if self.live_view_label.mode == "move":
            pen = QPen(Qt.red, 2)
            painter.setPen(pen)
            if self.live_view_label.quad_points:
                # Draw the quadrilateral
                painter.drawPolygon(QPolygonF(self.live_view_label.quad_points))
            elif self.live_view_label.bounding_box_preview:
                # Draw the rectangle
                start_x, start_y, end_x, end_y = self.live_view_label.bounding_box_preview
                rect = QRectF(start_x, start_y, end_x - start_x, end_y - start_y)
                painter.drawRect(rect)
    
        painter.end()
    
        # Apply zoom and pan transformations to the final pixmap
        if self.live_view_label.zoom_level != 1.0:
            # Create a new pixmap to apply zoom and pan
            zoomed_pixmap = QPixmap(pixmap.size())
            zoomed_pixmap.fill(Qt.transparent)
            zoom_painter = QPainter(zoomed_pixmap)
            zoom_painter.translate(self.live_view_label.pan_offset)
            zoom_painter.scale(self.live_view_label.zoom_level, self.live_view_label.zoom_level)
            zoom_painter.drawPixmap(0, 0, pixmap)
            zoom_painter.end()  # Properly end the QPainter
            pixmap = zoomed_pixmap
    
        # Set the final pixmap to the live view label
        self.live_view_label.setPixmap(pixmap)
    
    def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
        painter = QPainter(canvas)
        x_offset = (canvas.width() - scaled_image.width()) // 2
        y_offset = (canvas.height() - scaled_image.height()) // 2
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
    
        # Draw the base image
        painter.drawImage(x_offset, y_offset, scaled_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            # Resize Image 1 based on the slider value
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 1
            image1_x = x_offset + self.image1_position[0]
            image1_y = y_offset + self.image1_position[1]
            painter.drawImage(image1_x, image1_y, resized_image1)
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            # Resize Image 2 based on the slider value
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 2
            image2_x = x_offset + self.image2_position[0]
            image2_y = y_offset + self.image2_position[1]
            painter.drawImage(image2_x, image2_y, resized_image2)
    
        # Get the selected font settings
        font = QFont(self.font_combo_box.currentFont().family(), self.font_size_spinner.value() * render_scale)
        font_color = self.font_color if hasattr(self, 'font_color') else QColor(0, 0, 0)  # Default to black if no color selected
    
        painter.setFont(font)
        painter.setPen(font_color)  # Set the font color
    
        # Measure text height for vertical alignment
        font_metrics = painter.fontMetrics()
        text_height = font_metrics.height()
        line_padding = 5 * render_scale  # Space between text and line
        
        y_offset_global = text_height / 4
    
        # Draw the left markers (aligned right)
        for y_pos, marker_value in self.left_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f"{marker_value} ⎯ "  ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
        
        
        # Draw the right markers (aligned left)
        for y_pos, marker_value in self.right_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f" ⎯ {marker_value}" ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.right_marker_shift_added),# + line_padding),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
                
    
        # Draw the top markers (if needed)
        for x_pos, top_label in self.top_markers:
            x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width())
            if 0 <= x_pos_cropped <= scaled_image.width():
                text = f"{top_label}"
                painter.save()
                text_width = font_metrics.horizontalAdvance(text)
                text_height= font_metrics.height()
                label_x = x_offset + x_pos_cropped 
                label_y = y_offset + self.top_marker_shift + self.top_marker_shift_added 
                painter.translate(int(label_x), int(label_y))
                painter.rotate(self.font_rotation)
                painter.drawText(0,int(y_offset_global), f"{top_label}")
                painter.restore()
    
        # Draw guide lines
        if draw_guides and self.show_guides_checkbox.isChecked():
            pen = QPen(Qt.red, 2 * render_scale)
            painter.setPen(pen)
            center_x = canvas.width() // 2
            center_y = canvas.height() // 2
            painter.drawLine(center_x, 0, center_x, canvas.height())  # Vertical line
            painter.drawLine(0, center_y, canvas.width(), center_y)  # Horizontal line
            
        
        
        
        # Draw the protein location marker (*)
        if hasattr(self, "protein_location") and self.run_predict_MW == False:
            x, y = self.protein_location
            text = "⎯⎯"
            text_width = font_metrics.horizontalAdvance(text)
            text_height= font_metrics.height()
            painter.drawText(
                int(x * render_scale - text_width / 2),  #Currently left edge # FOR Center horizontally use int(x * render_scale - text_width / 2)
                int(y * render_scale + text_height / 4),  # Center vertically
                text
            )
        if hasattr(self, "custom_markers"):
            
            # Get default font type and size
            default_font_type = QFont(self.custom_font_type_dropdown.currentText())
            default_font_size = int(self.custom_font_size_spinbox.value())
        
            for x_pos, y_pos, marker_text, color, *optional in self.custom_markers:
                
                   
                # Use provided font type and size if available, otherwise use defaults
                marker_font_type = optional[0] if len(optional) > 0 else default_font_type
                marker_font_size = optional[1] if len(optional) > 1 else default_font_size
        
                # Ensure marker_font_type is a QFont instance
                font = QFont(marker_font_type) if isinstance(marker_font_type, str) else QFont(marker_font_type)
                font.setPointSize(marker_font_size * render_scale)  # Adjust font size for rendering scale
        
                # Apply the font and pen settings
                painter.setFont(font)
                painter.setPen(color)
        
                # Correct scaling and offsets
                x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width()) 
                y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height()) 
        
                # Only draw markers if they fall within the visible scaled image area
                if 0 <= x_pos_cropped <= scaled_image.width() and 0 <= y_pos_cropped <= scaled_image.height():
                    # Calculate text dimensions for alignment
                    font_metrics = painter.fontMetrics()
                    text_width = font_metrics.horizontalAdvance(marker_text)
                    text_height = font_metrics.height()
        
                    # Draw text centered horizontally and vertically
                    painter.drawText(
                        int(x_pos_cropped + x_offset- text_width / 2),  # Center horizontally
                        int(y_pos_cropped + y_offset+ text_height / 4),  # Center vertically
                        marker_text
                    )
            
            
    

        # Draw the grid (if enabled)
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value() * render_scale
            pen = QPen(Qt.red)
            pen.setStyle(Qt.DashLine)
            painter.setPen(pen)
    
            # Draw vertical grid lines
            for x in range(0, canvas.width(), grid_size):
                painter.drawLine(x, 0, x, canvas.height())
    
            # Draw horizontal grid lines
            for y in range(0, canvas.height(), grid_size):
                painter.drawLine(0, y, canvas.width(), y)
                
        painter.end()


     
    def crop_image(self):
        """Function to crop the current image."""
        if not self.image:
            return None
    
        # Get crop percentage from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate crop boundaries
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure cropping is valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            return None
    
        # Crop the image
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
        return cropped_image
    
    # Modify align_image and update_crop to preserve settings
    def align_image(self):
        self.save_state()
        """Align the image based on the orientation slider and keep high-resolution updates."""
        if not self.image:
            return
    
        self.draw_guides = False
        self.show_guides_checkbox.setChecked(False)
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
    
        # Get the orientation value from the slider
        angle = float(self.orientation_slider.value()/20)
    
        # Perform rotation
        transform = QTransform()
        transform.translate(self.image.width() / 2, self.image.height() / 2)  # Center rotation
        transform.rotate(angle)
        transform.translate(-self.image.width() / 2, -self.image.height() / 2)
    
        rotated_image = self.image.transformed(transform, Qt.SmoothTransformation)
    
        # Create a white canvas large enough to fit the rotated image
        canvas_width = rotated_image.width()
        canvas_height = rotated_image.height()
        high_res_canvas = QImage(canvas_width, canvas_height, QImage.Format_RGB888)
        high_res_canvas.fill(QColor(255, 255, 255))  # Fill with white background
    
        # Render the high-resolution canvas using `render_image_on_canvas`, without guides
        self.render_image_on_canvas(
            high_res_canvas,
            scaled_image=rotated_image,
            x_start=0,
            y_start=0,
            render_scale=1,  # Adjust scale if needed
            draw_guides=False  # Do not draw guides in this case
        )
    
        # Update the main high-resolution image
        self.image = high_res_canvas
        self.image_before_padding = self.image.copy()
        self.image_contrasted=self.image.copy()
        self.image_before_contrast=self.image.copy()
    
        # Create a low-resolution preview for display in `live_view_label`
        preview = high_res_canvas.scaled(
            self.live_view_label.width(),
            self.live_view_label.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.live_view_label.setPixmap(QPixmap.fromImage(preview))
    
        # Reset the orientation slider
        self.orientation_slider.setValue(0)
    
    
    def update_crop(self):
        self.save_state()
        """Update the image based on current crop sliders. First align, then crop the image."""
        # Save current configuration
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
    
        # Align the image first (rotate it)
        self.align_image()
    
        # Now apply cropping
        cropped_image = self.crop_image()
    
        if cropped_image:
            # Adjust marker positions based on cropping
            # self.adjust_markers_for_cropping(cropped_image)
            self.image = cropped_image
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
    
        # Reset sliders
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)
    
        try:
            w = cropped_image.width()
            h = cropped_image.height()
            # Preview window
            ratio = w / h
            self.label_width = int(self.screen_width * 0.28)
            label_height = int(self.label_width / ratio)
            if label_height > self.label_width:
                label_height = self.label_width
                self.label_width = ratio * label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
    
        self.update_live_view()
    
    # def adjust_markers_for_cropping(self, cropped_image):
    #     """Adjust marker positions based on cropping."""
    #     # Get crop percentage values from sliders
    #     x_start_percent = self.crop_x_start_slider.value() / 100
    #     x_end_percent = self.crop_x_end_slider.value() / 100
    #     y_start_percent = self.crop_y_start_slider.value() / 100
    #     y_end_percent = self.crop_y_end_slider.value() / 100
    
    #     # Calculate the crop boundaries based on the percentages
    #     x_start = int(self.image.width() * x_start_percent)
    #     x_end = int(self.image.width() * x_end_percent)
    #     y_start = int(self.image.height() * y_start_percent)
    #     y_end = int(self.image.height() * y_end_percent)
    
    #     # Adjust left markers
    #     self.left_markers = [(y - y_start, label) for y, label in self.left_markers if y_start <= y <= y_end]
    #     # Adjust right markers
    #     self.right_markers = [(y - y_start, label) for y, label in self.right_markers if y_start <= y <= y_end]
    #     # Adjust top markers
    #     self.top_markers = [(x - x_start, label) for x, label in self.top_markers if x_start <= x <= x_end]
    
    def update_skew(self):
        self.save_state()
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

 
        self.image = self.image.transformed(transform, Qt.SmoothTransformation)
        self.taper_skew_slider.setValue(0)

    
        
    def save_image(self):
        if not self.image:
            QMessageBox.warning(self, "Warning", "Please load an image first.")
            return
    
        options = QFileDialog.Options()
        save_path=""
        
        # Set default save directory and file name
        if hasattr(self, "image_path") and hasattr(self, "base_name"):
            default_file_name = os.path.join(self.image_path, f"{self.base_name}.png")
        else:
            default_file_name = ""
        
        base_save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image", default_file_name, "Image Files (*.png *.jpg *.bmp)", options=options
        )
        if base_save_path:
            # Check if the base name already contains "_original"
            if "_original" not in os.path.splitext(base_save_path)[0]:
                original_save_path = os.path.splitext(base_save_path)[0] + "_original.png"
                modified_save_path = os.path.splitext(base_save_path)[0] + "_modified.png"
                config_save_path = os.path.splitext(base_save_path)[0] + "_config.txt"
                save_path=original_save_path
                
            else:
                original_save_path = base_save_path
                modified_save_path = os.path.splitext(base_save_path)[0].replace("_original", "") + "_modified.png"
                config_save_path = os.path.splitext(base_save_path)[0].replace("_original", "") + "_config.txt"
                save_path=original_save_path
                # Check if "_modified" and "_config" already exist, and overwrite them if so
                if os.path.exists(modified_save_path):
                    os.remove(modified_save_path)
                if os.path.exists(config_save_path):
                    os.remove(config_save_path)
    
            # Save original image (cropped and aligned)
            self.original_image = self.image.copy()  # Save the current (cropped and aligned) image as original
            if not self.original_image.save(original_save_path):
                QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
    
            # Save modified image
            render_scale = 3
            high_res_canvas_width = self.live_view_label.width() * render_scale
            high_res_canvas_height = self.live_view_label.height() * render_scale
            high_res_canvas = QImage(
                high_res_canvas_width, high_res_canvas_height, QImage.Format_RGB888
            )
            high_res_canvas.fill(QColor(255, 255, 255))  # White background
    
            # Define cropping boundaries
            x_start_percent = self.crop_x_start_slider.value() / 100
            y_start_percent = self.crop_y_start_slider.value() / 100
            x_start = int(self.image.width() * x_start_percent)
            y_start = int(self.image.height() * y_start_percent)
    
            # Create a scaled version of the image
            scaled_image = self.image.scaled(
                high_res_canvas_width,
                high_res_canvas_height,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
            
            self.show_grid_checkbox.setChecked(False)
            self.update_live_view()
            # Render the high-resolution canvas without guides for saving
            self.render_image_on_canvas(
                high_res_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False
            )
    
            if not high_res_canvas.save(modified_save_path):
                QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
    
            # Save configuration to a .txt file
            config_data = self.get_current_config()
            try:
                with open(config_save_path, "w") as config_file:
                    json.dump(config_data, config_file, indent=4)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
    
            QMessageBox.information(self, "Saved", f"Files saved successfully.")
            
            self.setWindowTitle(f"{self.window_title}:{self.image_path}")
            
    def save_image_svg(self):
        """Save the processed image along with markers and labels in SVG format containing EMF data."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image as SVG for MS Word Image Editing", "", "SVG Files (*.svg)", options=options
        )
    
        if not file_path:
            return
    
        if not file_path.endswith(".svg"):
            file_path += ".svg"
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            font_metrics = QFontMetrics(QFont(font, font_size))
            text_width = (font_metrics.horizontalAdvance(text))  # Get text width
            text_height = (font_metrics.height())
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, y-text_height/4),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(self.left_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="end"  # Aligns text to the right
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()

    
            dwg.add(
                dwg.text(
                    f" ⎯ {text}",
                    insert=(self.right_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="start"  # Aligns text to the left
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text}"
            text_width = int(font_metrics.horizontalAdvance(final_text)) # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, self.top_marker_shift_added-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x-text_width/2}, {self.top_marker_shift_added+text_height/4})"
                )
            )
    
        # Save the SVG file
        dwg.save()
    
        QMessageBox.information(self, "Success", f"Image saved as SVG at {file_path}.")
    
    
    def copy_to_clipboard(self):
        """Copy the image from live view label to the clipboard."""
        if not self.image:
            print("No image to copy.")
            return
    
        # Define a high-resolution canvas for clipboard copy
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, QImage.Format_RGB888
        )
        high_res_canvas.fill(QColor(255, 255, 255))  # White background
    
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
    
        # Create a scaled version of the image
        scaled_image = self.image.scaled(
            high_res_canvas_width,
            high_res_canvas_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
        # Render the high-resolution canvas without guides for clipboard
        self.render_image_on_canvas(
            high_res_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False
        )
        
        # Copy the high-resolution image to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setImage(high_res_canvas)  # Copy the rendered image
        
    def copy_to_clipboard_SVG(self):
        """Create a temporary SVG file with EMF data, copy it to clipboard."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
    
        # Get scaling factors to match the live view window
        view_width = self.live_view_label.width()
        view_height = self.live_view_label.height()
        scale_x = self.image.width() / view_width
        scale_y = self.image.height() / view_height
    
        # Create a temporary file to store the SVG
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_file:
            temp_file_path = temp_file.name
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(temp_file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x * scale_x, y * scale_y),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(self.left_marker_shift_added / scale_x, y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=((self.image.width() / scale_x + self.right_marker_shift_added / scale_x), y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x, self.top_marker_shift_added / scale_y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x}, {self.top_marker_shift_added / scale_y})"
                )
            )
    
        # Save the SVG to the temporary file
        dwg.save()
    
        # Read the SVG content
        with open(temp_file_path, "r", encoding="utf-8") as temp_file:
            svg_content = temp_file.read()
    
        # Copy SVG content to clipboard (macOS-compatible approach)
        clipboard = QApplication.clipboard()
        clipboard.setText(svg_content, mode=clipboard.Clipboard)
    
        QMessageBox.information(self, "Success", "SVG content copied to clipboard.")

        
    def clear_predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        if hasattr(self, "protein_location"):
            del self.protein_location  # Clear the protein location marker
        self.predict_size=False
        self.bounding_boxes=[]
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.live_view_label.bounding_box_preview = None
        self.quantities=[]
        self.peak_area_list=[]
        self.protein_quantities=[]
        self.standard_protein_values.setText("")
        self.standard_protein_areas=[]
        self.standard_protein_areas_text.setText("")
        self.live_view_label.quad_points=[]
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.rectangle_points = []
        
        self.update_live_view()  # Update the display
        
    def predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Determine which markers to use (left or right)
        self.run_predict_MW=False
        if self.run_predict_MW!=True:
            markers_not_rounded = self.left_markers if self.left_markers else self.right_markers
            markers = [[round(float(value), 2) for value in sublist] for sublist in markers_not_rounded]
            if not markers:
                QMessageBox.warning(self, "Error", "No markers available for prediction.")
                return
        
            # Get marker positions and values
            marker_positions = np.array([pos for pos, _ in markers])
            marker_values = np.array([val for _, val in markers])
        
            # Ensure there are at least two markers for linear regression
            if len(marker_positions) < 2:
                QMessageBox.warning(self, "Error", "At least two markers are needed for prediction.")
                return
        
            # Normalize distances
            min_position = np.min(marker_positions)
            max_position = np.max(marker_positions)
            normalized_distances = (marker_positions - min_position) / (max_position - min_position)
            # normalized_distances = (marker_positions) / (min_position)
            # print("MARKERS: ", markers)
            
        
            # Allow the user to select a point for prediction
            QMessageBox.information(self, "Instruction", "Click on the protein location in the preview window.")
            self.live_view_label.mousePressEvent = lambda event: self.get_protein_location(
                event, normalized_distances, marker_values, min_position, max_position
            )

            


        
    def get_protein_location(self, event, normalized_distances, marker_values, min_position, max_position):
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
        
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        protein_position = round((cursor_y - y_offset) / scale,2)
        # protein_positio
        # print("PROTEIN POSITION: ", protein_position)
        
        
        
    
        # Normalize the protein position
        normalized_protein_position = (protein_position - min_position) / (max_position - min_position)
        # normalized_protein_position = (protein_position) / (min_position)
    
        # Perform linear regression on the log-transformed data
        log_marker_values = np.log10(marker_values)
        
        #LOG_FIT
        coefficients = np.polyfit(normalized_distances, log_marker_values, 1)
        #Polynomial Fit 6th order
        # coefficients = np.polyfit(normalized_distances, log_marker_values, 6)  # Fit a quadratic curve
        
        # Predict molecular weight
        predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
        predicted_weight = 10 ** predicted_log10_weight
    
        # Store the protein location in pixel coordinates
        self.protein_location = (cursor_x, cursor_y)

        
    
        # Update the live view to draw the *
        self.update_live_view()
    
        # Plot the graph with the clicked position
        self.plot_molecular_weight_graph(
            normalized_distances,
            marker_values,
            10 ** np.polyval(coefficients, normalized_distances),
            normalized_protein_position,
            predicted_weight,
            r_squared=1 - (np.sum((log_marker_values - np.polyval(coefficients, normalized_distances)) ** 2) /
                           np.sum((log_marker_values - np.mean(log_marker_values)) ** 2))
        )

        
        
    def plot_molecular_weight_graph(
        self, normalized_distances, marker_values, fitted_values, protein_position, predicted_weight, r_squared
    ):
        import matplotlib.pyplot as plt
        from io import BytesIO
        from PyQt5.QtGui import QPixmap
    
        plt.figure(figsize=(5, 3))
        plt.scatter(normalized_distances, marker_values, color="red", label="Marker Data")
        plt.plot(normalized_distances, fitted_values, color="blue", label=f"Fit (R²={r_squared:.3f})")
        plt.axvline(protein_position, color="green", linestyle="--", label=f"Protein Position\n({predicted_weight:.2f} units)")
        plt.xlabel("Normalized Relative Distance")
        plt.ylabel("Molecular Weight (units)")
        plt.yscale("log")  # Log scale for molecular weight
        plt.legend()
        plt.title("Molecular Weight Prediction")
    
        # Convert the plot to a pixmap
        buffer = BytesIO()
        plt.savefig(buffer, format="png", bbox_inches="tight")
        buffer.seek(0)
        pixmap = QPixmap()
        pixmap.loadFromData(buffer.read())
        buffer.close()
        plt.close()
    
        # Display the plot and results in a message box
        message_box = QMessageBox(self)
        message_box.setWindowTitle("Prediction Result")
        message_box.setText(
            f"The predicted molecular weight is approximately {predicted_weight:.2f} units.\n"
            f"R-squared value of the fit: {r_squared:.3f}"
        )
        label = QLabel()
        label.setPixmap(pixmap)
        message_box.layout().addWidget(label, 1, 0, 1, message_box.layout().columnCount())
        message_box.exec()
        # self.run_predict_MW=True

  
        
    def reset_image(self):
        # Reset the image to original
        if self.image != None:
            self.image = self.image_master.copy()
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy() # Update the contrasted image
            self.image_before_contrast= self.image.copy()
        else:
            self.image_before_padding = None
            self.image_contrasted = None  
            self.image_before_contrast=None
        self.warped_image=None
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
        self.update_live_view()  # Update the live view
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)
        self.orientation_slider.setValue(0)
        self.taper_skew_slider.setValue(0)
        self.high_slider.setValue(100)  # Reset contrast to default
        self.low_slider.setValue(0)  # Reset contrast to default
        self.gamma_slider.setValue(100)  # Reset gamma to default
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_padding_slider.setValue(0)
        self.right_padding_slider.setValue(0)
        self.top_padding_slider.setValue(0)
        self.marker_mode = None
        self.current_marker_index = 0
        self.current_top_label_index = 0
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
        self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
        self.update_live_view()


if __name__ == "__main__":
    try:
        app = QApplication([])
        app.setStyle("Fusion")
        app.setStyleSheet("""
        QSlider::handle:horizontal {
            width: 100px;
            height: 100px;
            margin: -5px 0;
            background: #FFFFFF;
            border: 2px solid #555;
            border-radius: 30px;
        }
    """)
        window = CombinedSDSApp()
        window.show()
        app.exec_()
    except Exception as e:
        logging.error("Application crashed", exc_info=True)
