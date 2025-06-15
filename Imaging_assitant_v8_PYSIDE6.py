import sys
import os
from PySide6.QtWidgets import (QApplication, QDialog, QLabel, QVBoxLayout,
                             QMessageBox) # QDesktopWidget removed here
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QScreen, QGuiApplication # QScreen added for desktop geometry
import logging
import traceback


# --- NEW Minimal Loading Dialog ---
class MinimalLoadingDialog(QDialog):
    """
    A very lightweight, minimal-resource loading dialog.
    Uses basic widgets and styling for fast appearance.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Loading...") # Simple title, might not show if frameless
        # Use flags for a clean, frameless look that stays on top
        self.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True) # Allows for rounded corners if stylesheet uses them

        # --- Styling (Keep it simple) ---
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(45, 45, 50, 240); /* Dark semi-transparent */
                border: 1px solid #777777;
                border-radius: 8px;
            }
            QLabel {
                color: #E0E0E0; /* Light text */
                padding: 25px; /* Add space around text */
                background-color: transparent; /* Ensure label bg is transparent */
            }
        """)

        # --- Layout and Label ---
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0) # No extra margins in layout

        self.label = QLabel("Loading Software,\nPlease Wait...") # Updated text
        font = QFont("Arial", 11) # Use a common system font
        font.setBold(True)
        self.label.setFont(font)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.setLayout(layout)

        # --- Size and Position ---
        self.setFixedSize(320, 120) # Adjust size as needed
        self.center_on_screen()

    def center_on_screen(self):
        """Centers the dialog on the primary screen."""
        try:
            # Use QGuiApplication to access screen information
            # QApplication.primaryScreen() also works if an app instance exists,
            # but QGuiApplication.primaryScreen() is safer if called very early.
            primary_screen = QGuiApplication.primaryScreen()
            if not primary_screen:
                screens = QGuiApplication.screens()
                if not screens:
                    print("Warning: No screens found to center loading dialog.")
                    self.move(100, 100)
                    return
                primary_screen = screens[0]

            screen_geo = primary_screen.availableGeometry()
            dialog_geo = self.frameGeometry()
            center_point = screen_geo.center()
            dialog_geo.moveCenter(center_point)
            self.move(dialog_geo.topLeft())
        except Exception as e:
            print(f"Warning: Could not center loading dialog: {e}")
            self.move(100, 100) # Fallback position


# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
script_dir = os.path.dirname(os.path.abspath(__file__))
# Construct the absolute path for the log file
log_file_path = os.path.join(script_dir, "error_log.txt")


# --- Configure logging ---
logging.basicConfig(
    filename=log_file_path, # Use the absolute path
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

try:
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not write initial log message: {e}") # Print error if immediate logging fails


def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    print("!!! log_exception called !!!") # Add print statement here too
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    # Log the exception
    try:
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
        # Optional: Force flush if you suspect buffering issues, though unlikely for ERROR level
        # logging.getLogger().handlers[0].flush()
    except Exception as log_err:
        print(f"ERROR: Failed to log exception to file: {log_err}") # Print logging specific errors

    # Display a QMessageBox with the error details
    try:
        error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}\n\n(Check error_log.txt for details)"
        QMessageBox.critical(
            None,
            "Unexpected Error",
            error_message,
            QMessageBox.Ok
        )
    except Exception as q_err:
         print(f"ERROR: Failed to show QMessageBox: {q_err}")


# Set the custom exception handler
sys.excepthook = log_exception



 



	
if __name__ == "__main__":
    app = None             # Initialize variable
    loading_dialog = None  # Initialize variable
    main_window = None     # Initialize variable

    try:
        # --- Try to get an existing QApplication instance FIRST ---
        # This helps in some environments or if the script is re-run partially
        app = QApplication.instance()
        if app is None:
            # --- Create QApplication ONLY if one doesn't exist ---
            # Enable High DPI Scaling FIRST if creating a new app
            try:
                # These attributes might only be settable *before* QApplication is instantiated
                # or on the class itself. Check PySide6 docs for exact timing.
                # For PySide6, it's often:
                # QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)
                # QApplication.setAttribute(Qt.AA_EnableHighDpiScaling) # PySide6 style
                # QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)   # PySide6 style
                # For now, keeping your original Qt5 style for direct porting,
                # but be aware these might need adjustment for PySide6 optimal behavior.
                QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
                QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
            except AttributeError:
                print("Warning: Could not set High DPI attributes (AttributeError).")
            except Exception as e_dpi: # Catch any other exception during DPI setting
                print(f"Warning: Error setting High DPI attributes: {e_dpi}")

            app = QApplication(sys.argv if hasattr(sys, 'argv') and len(sys.argv) > 0 else [])
        else:
            print("INFO: Using existing QApplication instance.")

        # --- Create and Show Minimal Loading Screen IMMEDIATELY ---
        try:
            loading_dialog = MinimalLoadingDialog()
            loading_dialog.show()
            if app: app.processEvents() # Crucial: Make the GUI update and show the dialog
        except Exception as e_load_dialog:
            print(f"ERROR: Could not create/show minimal loading dialog: {e_load_dialog}")
            # Proceed without loading screen if it fails, but log the error
            loading_dialog = None #

        import sys
        import svgwrite
        import tempfile
        from tempfile import NamedTemporaryFile
        import base64
        from PIL import ImageDraw, ImageFont, ImageGrab, Image, ImageQt, ImageOps  # Import Pillow's ImageGrab for clipboard access
        from io import BytesIO
        import io
        from PySide6.QtWidgets import (
            QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem,QToolBar,QStyle,
            QScrollArea, QInputDialog, QFrame, QApplication, QSizePolicy,
            QMainWindow, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit,
            QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog,
            QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox,
            QDialog, QHeaderView, QAbstractItemView, QMenu, QMenuBar, QFontDialog, QListWidget
        )
        from PySide6.QtGui import (
            QPixmap, QIcon, QPalette,QKeySequence, QImage, QPolygonF,QPainter, QBrush, QColor, QFont, QClipboard,
            QPen, QTransform,QFontMetrics,QDesktopServices, QAction, QShortcut, QIntValidator,QFocusEvent # <-- QIntValidator ADDED HERE
        )
        from PySide6.QtCore import (
            Qt, QBuffer, QPoint,QPointF, QRectF, QUrl, QSize, QSizeF, QMimeData, Signal
        )
        import json
        import os
        import numpy as np
        import matplotlib.pyplot as plt
        import matplotlib.lines as mlines # For draggable lines
        import matplotlib.patches as patches # For draggable handles
        import platform
        import openpyxl
        from openpyxl.styles import Font
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.gridspec import GridSpec
        from skimage.restoration import rolling_ball 
        from scipy.signal import find_peaks
        from scipy.ndimage import gaussian_filter1d
        from scipy.ndimage import grey_opening, grey_erosion, grey_dilation
        from scipy.interpolate import interp1d # Needed for interpolation
        import cv2
        import datetime

        def create_text_icon(font_type: QFont, icon_size: QSize, color: QColor, symbol: str) -> QIcon:
            """Creates a QIcon by drawing text/symbol onto a pixmap."""
            pixmap = QPixmap(icon_size)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing, True)
            painter.setRenderHint(QPainter.TextAntialiasing, True) # Good for text

            # Font settings (adjust as needed)
            font = QFont(font_type)
            # Make arrow slightly smaller than +/-, maybe not bold? Experiment.
            font.setPointSize(min(14, int(icon_size.height()*0.75)))
            # font.setBold(True) # Optional: Make arrows bold or not
            painter.setFont(font)
            painter.setPen(color)

            # Draw the symbol centered
            painter.drawText(pixmap.rect(), Qt.AlignCenter, symbol)
            painter.end()
            return QIcon(pixmap)

        # Set Style (can be done after app exists)
        app.setStyle("Fusion")
        app.setStyleSheet("""
            /* ... Your existing stylesheet ... */
            QSlider::handle:horizontal { /* Example */
                width: 10px; /* Example adjustment if needed */
                height: 20px;
                margin: -5px 0;
                background: #DDDDDD;
                border: 1px solid #555;
                border-radius: 5px;
            }
            QStatusBar QLabel {
                margin-left: 2px; margin-right: 5px; padding: 0px 0px; border: none;
            }
            QPushButton:checked {
                background-color: #a0d0a0; border: 1px solid #50a050;
            }
        """)
        class AutoLaneTuneDialog(QDialog):
            """
            Simplified dialog for tuning peak detection parameters for automatic lane markers.
            Shows the intensity profile and detected peaks. Allows adjustment of detection parameters.
            Does NOT handle area calculation or individual band boundary adjustments.

            This dialog operates on a rectangular PIL.Image object provided as `pil_image_data`.
            If this data comes from a warped quadrilateral region of an original image (i.e.,
            the `pil_image_data` itself is the rectangular result of a warp), the detected
            peak coordinates will be relative to this warped rectangular image. The calling
            code is responsible for any geometric transformations to map these coordinates
            back to the original image space if needed. The `is_from_quad_warp` parameter
            can be used to indicate this context, which will adjust the dialog's title.
            """
            def __init__(self, pil_image_data, initial_settings, parent=None, is_from_quad_warp=False): # Added is_from_quad_warp
                super().__init__(parent)

                # Set window title based on the source of the image data
                if is_from_quad_warp:
                    self.setWindowTitle("Tune Peaks (Warped Region)")
                else:
                    self.setWindowTitle("Tune Automatic Peak Detection")

                self.setGeometry(150, 150, 800, 700) # Adjusted height
                self.pil_image_for_display = pil_image_data
                self.selected_peak_index = -1 # Stores the X-coordinate of the peak selected for deletion
                self.deleted_peak_indices = set()
                self._all_initial_peaks = np.array([])
                self.add_peak_mode_active = False

                # --- Validate and Store Input Image ---
                if not isinstance(pil_image_data, Image.Image):
                    raise TypeError("Input 'pil_image_data' must be a PIL Image object")

                # Determine intensity range and create numpy array
                self.intensity_array_original_range = None
                self.original_max_value = 255.0 # Default
                pil_mode = pil_image_data.mode
                try:
                    # Handle common grayscale modes used in the main app
                    if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'L':
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        self.original_max_value = 255.0
                    elif pil_mode == 'F': # Handle float images
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                        self.original_max_value = max(1.0, max_in_float) # Use max value or 1.0
                    else: # Attempt conversion to grayscale 'L' as fallback
                        gray_img = pil_image_data.convert("L")
                        self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                        self.original_max_value = 255.0
                        print(f"AutoLaneTuneDialog: Converted input mode '{pil_mode}' to 'L'.")

                    if self.intensity_array_original_range is None:
                        raise ValueError("Failed to convert PIL image to NumPy array.")
                    if self.intensity_array_original_range.ndim != 2:
                        raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

                except Exception as e:
                    raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

                self.profile_original_inverted = None # Smoothed, inverted profile (original range)
                self.profile = None # Scaled (0-255), inverted, SMOOTHED profile for detection
                self.detected_peaks = np.array([]) # Store indices of detected peaks

                # --- Settings and State ---
                # Use settings passed from the main app (likely self.peak_dialog_settings)
                self.smoothing_sigma = initial_settings.get('smoothing_sigma', 2.0)
                self.peak_height_factor = initial_settings.get('peak_height_factor', 0.1)
                self.peak_distance = initial_settings.get('peak_distance', 10)
                self.peak_prominence_factor = initial_settings.get('peak_prominence_factor', 0.00)
                # Band estimation method is needed to generate the profile
                self.band_estimation_method = initial_settings.get('band_estimation_method', "Mean")
                self._final_settings = initial_settings.copy() # Store a copy to return modifications

                # Check dependencies needed for profile generation and peak detection
                if find_peaks is None or gaussian_filter1d is None:
                     QMessageBox.critical(self, "Dependency Error",
                                          "Missing SciPy library functions.\n"
                                          "Peak detection and smoothing require SciPy.\n"
                                          "Please install it (e.g., 'pip install scipy') and restart.")
                     # Don't call accept/reject here, let init fail or handle gracefully
                     # self.close() # Or close immediately

                # Build UI & Initial Setup
                self._setup_ui()
                self.run_peak_detection_and_plot() # Initial calculation and plot

            def _setup_ui(self):
                """Creates and arranges the UI elements, including image preview."""
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(10)

                # --- Matplotlib Plot Canvas Area ---
                plot_widget = QWidget() # Container for the plots
                plot_layout = QVBoxLayout(plot_widget)
                plot_layout.setContentsMargins(0, 0, 0, 0) # No margins for the layout

                # Use GridSpec for plot and image preview arrangement
                # More height for the profile plot
                self.fig = plt.figure(figsize=(7, 5)) # Adjusted height for gridspec
                gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig)
                self.fig.tight_layout(pad=0.5)
                self.ax_profile = self.fig.add_subplot(gs[0]) # Axis for the profile
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax_profile) # Axis for the image preview

                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                # --- Enable picking on the canvas ---
                self.canvas.mpl_connect('button_press_event', self.on_canvas_click)
                # --- End Enable picking ---
                plot_layout.addWidget(self.canvas)
                main_layout.addWidget(plot_widget, stretch=1)

                # --- Controls Group ---
                controls_group = QGroupBox("Peak Detection Parameters")
                controls_layout = QGridLayout(controls_group)
                controls_layout.setSpacing(8)

                # Band Estimation (Needed for profile generation)
                controls_layout.addWidget(QLabel("Profile Method:"), 0, 0)
                self.band_estimation_combobox = QComboBox()
                self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
                self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
                self.band_estimation_combobox.currentIndexChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2) # Span 2 columns

                # Smoothing Sigma
                self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
                self.smoothing_slider = QSlider(Qt.Horizontal)
                self.smoothing_slider.setRange(0, 100) # 0.0 to 10.0
                self.smoothing_slider.setValue(int(self.smoothing_sigma * 10))
                self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})"))
                self.smoothing_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.smoothing_label, 1, 0)
                controls_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2)

                # Peak Prominence
                self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
                self.peak_prominence_slider = QSlider(Qt.Horizontal)
                self.peak_prominence_slider.setRange(0, 100) # 0.0 to 1.0 factor
                self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
                self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})"))
                self.peak_prominence_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
                controls_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)

                # Peak Height
                self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})")
                self.peak_height_slider = QSlider(Qt.Horizontal)
                self.peak_height_slider.setRange(0, 100)
                self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
                self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f})"))
                self.peak_height_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_height_slider_label, 3, 0)
                controls_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)

                # Peak Distance
                self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
                self.peak_distance_slider = QSlider(Qt.Horizontal)
                self.peak_distance_slider.setRange(1, 200)
                self.peak_distance_slider.setValue(self.peak_distance)
                self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
                self.peak_distance_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_distance_slider_label, 4, 0)
                controls_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)


                # --- Add Delete Peak Button ---
                self.delete_peak_button = QPushButton("Delete Selected Peak")
                self.delete_peak_button.setEnabled(False)
                self.delete_peak_button.setToolTip("Click on a peak marker in the plot to select it, then click this.")
                self.delete_peak_button.clicked.connect(self.delete_selected_peak)
                controls_layout.addWidget(self.delete_peak_button, 5, 0, 1, 1) # Column 0, span 1

                # --- NEW: Add Peak Manually Button ---
                self.add_peak_button = QPushButton("Add Peak at Click")
                self.add_peak_button.setCheckable(True) # Make it a toggle button
                self.add_peak_button.setToolTip("Toggle: Click on the profile to add a new peak marker.")
                self.add_peak_button.clicked.connect(self.toggle_add_peak_mode)
                controls_layout.addWidget(self.add_peak_button, 5, 1, 1, 2) # Next to delete, span 2
                # --- End Add Delete Button ---

                controls_layout.setColumnStretch(1, 1) # Slider column stretch
                main_layout.addWidget(controls_group)

                # --- Bottom Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept_and_return_peaks)
                button_box.rejected.connect(self.reject)
                main_layout.addWidget(button_box)
                
            def toggle_add_peak_mode(self, checked):
                """Toggles the manual peak adding mode."""
                self.add_peak_mode_active = checked
                if checked:
                    self.canvas.setCursor(Qt.CrossCursor)
                    # The following lines were problematic as these attributes don't exist
                    # in AutoLaneTuneDialog. They are part of PeakAreaDialog's UI focusing.
                    # self.identify_peak_button.setChecked(False)
                    # self.manual_select_mode_active = False
                    # self.selected_peak_for_ui_focus = -1
                    # self._update_peak_group_box_styles()
                    
                    # Reset selection for deletion if add mode is activated
                    self.selected_peak_index = -1 
                    self.delete_peak_button.setEnabled(False)
                    self.update_plot_highlights() # Update plot to clear deletion selection highlights
                    QMessageBox.information(self, "Add Peak", "Add Peak Mode: ON. Click on the profile plot to add a peak.")
                else:
                    self.canvas.setCursor(Qt.ArrowCursor)

            def on_canvas_click(self, event):
                """Handles clicks on the canvas for adding or selecting peaks."""
                # Ignore clicks outside the profile axes or if no data
                if event.inaxes != self.ax_profile or self.profile_original_inverted is None:
                    return
                if event.button != 1: # Only process left-clicks
                    return

                clicked_x = int(round(event.xdata)) # Get x-coordinate of the click

                if self.add_peak_mode_active:
                    # --- Add Peak Mode ---
                    if 0 <= clicked_x < len(self.profile_original_inverted):
                        self.add_manual_peak(clicked_x)
                    else:
                        print(f"Clicked X ({clicked_x}) is outside profile bounds.")
                    # Deactivate add mode after one click for single additions
                    # self.add_peak_button.setChecked(False) # Optional: Toggle off after one add
                    # self.toggle_add_peak_mode(False)       # Or call the toggle function

                else:
                    # --- Select Peak Mode (for deletion) ---
                    # Find the closest *existing* peak marker to the click
                    if len(self.detected_peaks) > 0:
                        # Calculate distance from click to all active peak X-coordinates
                        distances = np.abs(self.detected_peaks - clicked_x)
                        min_dist_idx = np.argmin(distances)
                        # Define a click tolerance (e.g., 5 pixels on the x-axis)
                        click_tolerance_x = max(5, self.peak_distance / 4) # Heuristic based on peak distance

                        if distances[min_dist_idx] <= click_tolerance_x:
                            # A peak is close enough to the click
                            self.selected_peak_index = self.detected_peaks[min_dist_idx]
                            self.delete_peak_button.setEnabled(True)
                            print(f"Selected peak for deletion at index: {self.selected_peak_index}")
                        else:
                            # Click was not close enough to an existing peak
                            self.selected_peak_index = -1
                            self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights() # Update plot to show selection or clear it
                    else:
                        # No peaks to select
                        self.selected_peak_index = -1
                        self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()


            def add_manual_peak(self, x_coord):
                """Adds a new peak at the given x-coordinate if not already present or deleted."""
                if x_coord in self._all_initial_peaks or x_coord in self.deleted_peak_indices:
                    print(f"Peak at {x_coord} already exists or was previously deleted.")
                    # Optionally, undelete if it was in deleted_peak_indices
                    if x_coord in self.deleted_peak_indices:
                         reply = QMessageBox.question(self, "Undelete Peak?",
                                                      f"A peak at index {x_coord} was previously deleted. Undelete it?",
                                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                         if reply == QMessageBox.Yes:
                             self.deleted_peak_indices.discard(x_coord)
                             # Add to _all_initial_peaks if it wasn't there (unlikely if deleted, but for safety)
                             if x_coord not in self._all_initial_peaks:
                                  self._all_initial_peaks = np.sort(np.append(self._all_initial_peaks, x_coord))
                             # Update active peaks
                             self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])
                             print(f"Undeleted peak at {x_coord}.")
                             self.update_plot_highlights() # Update plot
                         return # Don't proceed to add again
                    else: # Already an initial peak
                         return

                # Add to both _all_initial_peaks and detected_peaks
                self._all_initial_peaks = np.sort(np.append(self._all_initial_peaks, x_coord))
                self.detected_peaks = np.sort(np.append(self.detected_peaks, x_coord))

                print(f"Manually added peak at index: {x_coord}")
                self.update_plot_highlights() # Update plot to show the new peak

            def run_peak_detection_and_plot(self):
                """Generates profile, detects peaks, updates plot and image preview."""
                # --- Reset deleted peaks if parameters change ---
                self.deleted_peak_indices.clear()
                self.selected_peak_index = -1
                self.delete_peak_button.setEnabled(False)
                # --- End Reset ---

                if self.intensity_array_original_range is None: return
                if find_peaks is None or gaussian_filter1d is None: return

                # ... (Keep parameter updates and profile generation logic) ...
                self.band_estimation_method = self.band_estimation_combobox.currentText()
                self.smoothing_sigma = self.smoothing_slider.value() / 10.0
                self.peak_height_factor = self.peak_height_slider.value() / 100.0
                self.peak_distance = self.peak_distance_slider.value()
                self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0

                profile_temp = None
                if self.band_estimation_method == "Mean":
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)
                elif self.band_estimation_method.startswith("Percentile"):
                    try:
                        percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                        profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
                    except:
                        profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1)
                        print("AutoLaneTuneDialog: Defaulting to 5th percentile for profile.")
                else:
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)

                if profile_temp is None or not np.all(np.isfinite(profile_temp)):
                    print("AutoLaneTuneDialog: Profile calculation failed or resulted in NaN/Inf.")
                    profile_temp = np.zeros(self.intensity_array_original_range.shape[0])

                profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
                min_inverted_raw = np.min(profile_original_inv_raw)
                profile_original_inv_raw -= min_inverted_raw

                self.profile_original_inverted = profile_original_inv_raw
                try:
                    current_sigma = self.smoothing_sigma
                    if current_sigma > 0.1 and len(self.profile_original_inverted) > int(3 * current_sigma) * 2 + 1:
                        self.profile_original_inverted = gaussian_filter1d(self.profile_original_inverted, sigma=current_sigma)
                except Exception as smooth_err:
                    print(f"AutoLaneTuneDialog: Error smoothing profile: {smooth_err}")

                prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
                if prof_max_inv > prof_min_inv + 1e-6:
                    self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
                else:
                    self.profile = np.zeros_like(self.profile_original_inverted)

                profile_range_detect = np.ptp(self.profile); min_val_profile_detect = np.min(self.profile)
                if profile_range_detect < 1e-6 : profile_range_detect = 1.0
                min_height_abs = min_val_profile_detect + profile_range_detect * self.peak_height_factor
                min_prominence_abs = profile_range_detect * self.peak_prominence_factor
                min_prominence_abs = max(1.0, min_prominence_abs)

                try:
                    peaks_indices, _ = find_peaks(
                        self.profile, height=min_height_abs, prominence=min_prominence_abs,
                        distance=self.peak_distance, width=1
                    )
                    # *** Store ALL initially detected peaks ***
                    self._all_initial_peaks = np.sort(peaks_indices)
                    # *** Filter out user-deleted peaks for display ***
                    self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])

                except Exception as e:
                    print(f"AutoLaneTuneDialog: Peak detection error: {e}")
                    self._all_initial_peaks = np.array([])
                    self.detected_peaks = np.array([])


                # --- Update Plot ---
                self.ax_profile.clear()
                self.ax_image.clear() # Clear image axis too

                if self.profile_original_inverted is not None and len(self.profile_original_inverted) > 0:
                    self.ax_profile.plot(self.profile_original_inverted, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.0)

                    # Plot *currently active* detected peaks
                    # These are the peaks that `on_canvas_click` will try to select for deletion
                    if len(self.detected_peaks) > 0:
                        valid_peaks = self.detected_peaks[(self.detected_peaks >= 0) & (self.detected_peaks < len(self.profile_original_inverted))]
                        if len(valid_peaks) > 0:
                            peak_y_values = self.profile_original_inverted[valid_peaks]
                            # No need for `picker=True` if on_canvas_click handles selection by coordinate
                            self.peak_plot_artist, = self.ax_profile.plot(
                                valid_peaks, peak_y_values, "rx", markersize=8,
                                label=f"Active Peaks ({len(valid_peaks)})") # Removed picker

                            # --- Highlight selected peak (logic remains the same) ---
                            if self.selected_peak_index != -1 and self.selected_peak_index in valid_peaks:
                                 idx_in_valid = np.where(valid_peaks == self.selected_peak_index)[0]
                                 if len(idx_in_valid) > 0:
                                      self.ax_profile.plot(self.selected_peak_index, peak_y_values[idx_in_valid[0]],
                                                         'o', markersize=12, markeredgecolor='blue', markerfacecolor='none',
                                                         label='Selected')
                            # --- End Highlight ---
                    
                    # Profile plot styling
                    self.ax_profile.set_ylabel("Intensity (Smoothed, Inverted)", fontsize=9)
                    self.ax_profile.legend(fontsize='x-small')
                    self.ax_profile.set_title("Intensity Profile and Detected Peaks", fontsize=10)
                    self.ax_profile.grid(True, linestyle=':', alpha=0.6)
                    # Remove x-axis labels/ticks for the top plot
                    self.ax_profile.tick_params(axis='x', labelbottom=False)
                    if np.max(self.profile_original_inverted) > 10000:
                        self.ax_profile.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)

                    # --- Plot Image Preview ---
                    try:
                        # Rotate the PIL image 90 degrees for horizontal display matching profile axis
                        rotated_pil_image = self.pil_image_for_display.rotate(90, expand=True)
                        im_array_disp = np.array(rotated_pil_image)

                        # Determine vmin/vmax for imshow based on original data range
                        im_vmin, im_vmax = 0, self.original_max_value
                        # Handle float case specifically if needed
                        if self.pil_image_for_display.mode == 'F':
                            im_vmin, im_vmax = 0.0, self.original_max_value # Or 0.0, 1.0 if normalized

                        profile_length = len(self.profile_original_inverted)
                        extent = [0, profile_length - 1 if profile_length > 0 else 0, 0, rotated_pil_image.height]

                        self.ax_image.imshow(im_array_disp, cmap='gray', aspect='auto',
                                             extent=extent, vmin=im_vmin, vmax=im_vmax)
                        self.ax_image.set_xlabel("Pixel Index along Profile Axis", fontsize=9)
                        self.ax_image.set_yticks([]) # Hide Y ticks for image preview
                        self.ax_image.set_ylabel("Lane Width", fontsize=9)
                    except Exception as img_e:
                        print(f"AutoLaneTuneDialog: Error displaying image preview: {img_e}")
                        self.ax_image.text(0.5, 0.5, 'Error displaying preview', ha='center', va='center', transform=self.ax_image.transAxes)
                        self.ax_image.set_xticks([]); self.ax_image.set_yticks([])

                else:
                    self.ax_profile.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax_profile.transAxes)
                    self.ax_image.text(0.5, 0.5, "No Image Data", ha='center', va='center', transform=self.ax_image.transAxes)

                # Use tight_layout or constrained_layout
                # try:
                #     self.fig.tight_layout(pad=0.5)
                # except ValueError: # Can happen with certain plot states
                #     try:
                #         self.fig.set_constrained_layout(True)
                #     except AttributeError: pass # Older matplotlib

                self.canvas.draw_idle()
                plt.close()

            def on_pick(self, event):
                """Handles clicking on a peak marker."""
                # Check if the picked artist is the one we stored for the peaks
                if event.artist != getattr(self, 'peak_plot_artist', None):
                    return

                # Get the index of the clicked data point within the plotted data
                ind = event.ind[0] # Get the first index if multiple points are close

                # Map this index back to the actual peak coordinate value (Y-index)
                # Ensure self.detected_peaks is valid and index is within bounds
                if ind < len(self.detected_peaks):
                    clicked_peak_coord = self.detected_peaks[ind]
                    # Check if this peak wasn't already deleted internally
                    if clicked_peak_coord in self._all_initial_peaks and clicked_peak_coord not in self.deleted_peak_indices:
                        self.selected_peak_index = clicked_peak_coord
                        self.delete_peak_button.setEnabled(True)
                        print(f"Selected peak at index: {self.selected_peak_index}") # Debug
                        # Re-plot to show selection highlight
                        self.update_plot_highlights() # Separate function for replotting highlights
                    else:
                        # Clicked on a potentially invalid/deleted point marker
                        self.selected_peak_index = -1
                        self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()
                else:
                     # Index out of bounds (shouldn't normally happen)
                     self.selected_peak_index = -1
                     self.delete_peak_button.setEnabled(False)
                     self.update_plot_highlights()

            def update_plot_highlights(self):
                """Redraws only the peak markers and highlights without full recalculation."""
                if not hasattr(self, 'ax_profile'): return

                # Clear only the peak markers and highlights from the profile axis
                # Keep the main profile line
                artists_to_remove = []
                for artist in self.ax_profile.lines + self.ax_profile.collections:
                    label = artist.get_label()
                    if label and ('Peaks' in label or 'Selected' in label):
                         artists_to_remove.append(artist)
                for artist in artists_to_remove:
                    artist.remove()

                # Re-plot the *currently active* detected peaks
                if len(self.detected_peaks) > 0:
                    valid_peaks = self.detected_peaks[(self.detected_peaks >= 0) & (self.detected_peaks < len(self.profile_original_inverted))]
                    if len(valid_peaks) > 0:
                        peak_y_values = self.profile_original_inverted[valid_peaks]
                        # Re-store the artist for picking
                        self.peak_plot_artist, = self.ax_profile.plot(
                            valid_peaks, peak_y_values, "rx", markersize=8,
                            label=f"Active Peaks ({len(valid_peaks)})", picker=True, pickradius=5)

                        # Re-apply highlight if a peak is selected
                        if self.selected_peak_index != -1 and self.selected_peak_index in valid_peaks:
                            idx_in_valid = np.where(valid_peaks == self.selected_peak_index)[0]
                            if len(idx_in_valid) > 0:
                                self.ax_profile.plot(self.selected_peak_index, peak_y_values[idx_in_valid[0]],
                                                     'o', markersize=12, markeredgecolor='blue', markerfacecolor='none',
                                                     label='Selected') # Visual indicator

                # Update the legend and redraw
                handles, labels = self.ax_profile.get_legend_handles_labels()
                # Filter out duplicate labels if any were added accidentally
                by_label = dict(zip(labels, handles))
                self.ax_profile.legend(by_label.values(), by_label.keys(), fontsize='x-small')
                self.canvas.draw_idle()


            def delete_selected_peak(self):
                """Marks the selected peak as deleted and updates the plot."""
                if self.selected_peak_index != -1:
                    # Add the index to the set of deleted peaks
                    self.deleted_peak_indices.add(self.selected_peak_index)

                    # Update the list of currently active peaks
                    self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])

                    print(f"Deleted peak at index: {self.selected_peak_index}") # Debug

                    # Reset selection and disable button
                    self.selected_peak_index = -1
                    self.delete_peak_button.setEnabled(False)

                    # Re-plot to remove the deleted peak visually
                    self.update_plot_highlights() # Use highlight update for efficiency

            # MODIFY this method
            def accept_and_return_peaks(self):
                """Store final settings and accept the dialog."""
                # *** Settings are updated when sliders change now ***
                # We just need to ensure the final state is stored if needed elsewhere
                self._final_settings = {
                    'smoothing_sigma': self.smoothing_slider.value() / 10.0,
                    'peak_height_factor': self.peak_height_slider.value() / 100.0,
                    'peak_distance': self.peak_distance_slider.value(),
                    'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
                    'band_estimation_method': self.band_estimation_combobox.currentText(),
                    'rolling_ball_radius': self._final_settings.get('rolling_ball_radius', 50),
                    'area_subtraction_method': self._final_settings.get('area_subtraction_method', "Rolling Ball"),
                }
                self.accept()

            # MODIFY this method
            def get_detected_peaks(self):
                """Returns the FINAL list/array of detected peak Y-coordinates (indices), excluding user-deleted ones."""
                # Return the filtered list
                return self.detected_peaks

            def get_final_settings(self):
                """Returns the final detection settings dictionary."""
                return self._final_settings
            
        
        
        class ModifyMarkersDialog(QDialog):
            """
            A dialog for modifying custom markers and shapes with global adjustments and
            individual item editing. This implementation uses a robust "repopulate-on-change"
            pattern to keep the UI view consistent with the data model, even after sorting.
            """
            shapes_adjusted_preview = Signal(list)
            global_markers_adjusted = Signal(list)
        
            def __init__(self, markers_list, shapes_list, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Modify Custom Markers and Shapes")
                self.setMinimumSize(950, 650)
        
                # The data model: two lists for markers to handle scaling correctly, one for shapes.
                self._original_markers_data = [tuple(m) for m in markers_list] # Pristine backup for scaling
                self.markers = [list(m) for m in markers_list]                 # Working copy for display
                self.shapes = [dict(s) for s in shapes_list]                   # Working copy
        
                self._block_signals = False
                self._current_image_width = parent.image.width() if parent and parent.image and not parent.image.isNull() else 1
                self._current_image_height = parent.image.height() if parent and parent.image and not parent.image.isNull() else 1
        
                self._setup_ui()
                self.populate_table()
        
            def _setup_ui(self):
                """Creates and arranges all UI elements for the dialog."""
                layout = QVBoxLayout(self)
        
                # --- Global Adjustment Controls ---
                global_adjust_group = QGroupBox("Global Adjustments for Markers")
                global_adjust_layout = QGridLayout(global_adjust_group)
                self.percent_precision_factor = 100.0
                self.scale_precision_factor = 10.0
        
                # Sliders for global adjustments (Shift, Scale, Font Scale)
                # Each slider's valueChanged signal triggers _update_global_adjustments
                # A second connection updates the percentage label for immediate feedback.
                # ... (Slider creation logic is unchanged and correct) ...
                # Absolute X Shift
                global_adjust_layout.addWidget(QLabel("Shift X (% Img W):"), 0, 0)
                self.abs_x_shift_slider = QSlider(Qt.Horizontal)
                self.abs_x_shift_slider.setRange(int(-100 * self.percent_precision_factor), int(100 * self.percent_precision_factor))
                self.abs_x_shift_slider.setValue(0)
                self.abs_x_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_x_shift_label = QLabel("0.00%"); self.abs_x_shift_label.setFixedSize(80, 20)
                self.abs_x_shift_slider.valueChanged.connect(lambda val: self.abs_x_shift_label.setText(f"{val / self.percent_precision_factor:.2f}%"))
                global_adjust_layout.addWidget(self.abs_x_shift_slider, 0, 1); global_adjust_layout.addWidget(self.abs_x_shift_label, 0, 2)
                # Absolute Y Shift
                global_adjust_layout.addWidget(QLabel("Shift Y (% Img H):"), 1, 0)
                self.abs_y_shift_slider = QSlider(Qt.Horizontal)
                self.abs_y_shift_slider.setRange(int(-100 * self.percent_precision_factor), int(100 * self.percent_precision_factor))
                self.abs_y_shift_slider.setValue(0)
                self.abs_y_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_y_shift_label = QLabel("0.00%"); self.abs_y_shift_label.setFixedSize(80, 20)
                self.abs_y_shift_slider.valueChanged.connect(lambda val: self.abs_y_shift_label.setText(f"{val / self.percent_precision_factor:.2f}%"))
                global_adjust_layout.addWidget(self.abs_y_shift_slider, 1, 1); global_adjust_layout.addWidget(self.abs_y_shift_label, 1, 2)
                # Relative X Scale
                global_adjust_layout.addWidget(QLabel("Scale X (%):"), 2, 0)
                self.rel_x_scale_slider = QSlider(Qt.Horizontal)
                self.rel_x_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor))
                self.rel_x_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.rel_x_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_x_scale_label = QLabel("100.0%"); self.rel_x_scale_label.setFixedSize(80, 20)
                self.rel_x_scale_slider.valueChanged.connect(lambda val: self.rel_x_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.rel_x_scale_slider, 2, 1); global_adjust_layout.addWidget(self.rel_x_scale_label, 2, 2)
                # Relative Y Scale
                global_adjust_layout.addWidget(QLabel("Scale Y (%):"), 3, 0)
                self.rel_y_scale_slider = QSlider(Qt.Horizontal)
                self.rel_y_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor))
                self.rel_y_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.rel_y_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_y_scale_label = QLabel("100.0%"); self.rel_y_scale_label.setFixedSize(80, 20)
                self.rel_y_scale_slider.valueChanged.connect(lambda val: self.rel_y_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.rel_y_scale_slider, 3, 1); global_adjust_layout.addWidget(self.rel_y_scale_label, 3, 2)
                # Font Scale
                global_adjust_layout.addWidget(QLabel("Font Scale (%):"), 4, 0)
                self.font_scale_slider = QSlider(Qt.Horizontal)
                self.font_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor))
                self.font_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.font_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.font_scale_label = QLabel("100.0%"); self.font_scale_label.setFixedSize(80, 20)
                self.font_scale_slider.valueChanged.connect(lambda val: self.font_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.font_scale_slider, 4, 1); global_adjust_layout.addWidget(self.font_scale_label, 4, 2)
        
                global_adjust_layout.setColumnStretch(1, 1)
                layout.addWidget(global_adjust_group)
        
                # --- Table Widget ---
                self.table_widget = QTableWidget()
                self.table_widget.setColumnCount(8)
                self.table_widget.setHorizontalHeaderLabels(["Type", "Text/Label", "Coordinates", "Style", "Bold", "Italic", "Color", "Actions"])
                self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
                self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
                self.table_widget.setEditTriggers(QAbstractItemView.AnyKeyPressed | QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
                self.table_widget.setSortingEnabled(True)
                self.table_widget.itemChanged.connect(self.handle_item_changed)
                self.table_widget.cellDoubleClicked.connect(self.handle_cell_double_clicked)
                layout.addWidget(self.table_widget)
                self.table_widget.resizeColumnsToContents()
                self.table_widget.horizontalHeader().setStretchLastSection(True)
        
                # --- Dialog Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept)
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box)
                self.setLayout(layout)
        
            def _update_global_adjustments(self):
                """Recalculates all marker positions and font sizes based on sliders."""
                if self._block_signals: return
        
                # Get values from all sliders
                abs_x_shift_percent = self.abs_x_shift_slider.value() / self.percent_precision_factor
                abs_y_shift_percent = self.abs_y_shift_slider.value() / self.percent_precision_factor
                rel_x_scale_factor = self.rel_x_scale_slider.value() / self.scale_precision_factor / 100.0
                rel_y_scale_factor = self.rel_y_scale_slider.value() / self.scale_precision_factor / 100.0
                font_size_scale_factor = self.font_scale_slider.value() / self.scale_precision_factor / 100.0
        
                abs_x_shift_pixels = (abs_x_shift_percent / 100.0) * self._current_image_width
                abs_y_shift_pixels = (abs_y_shift_percent / 100.0) * self._current_image_height
        
                # Rebuild the working `self.markers` list from the pristine `_original_markers_data`
                for i, original_marker_tuple in enumerate(self._original_markers_data):
                    orig_x, orig_y, text, qcolor, font_family, orig_font_size, is_bold, is_italic = original_marker_tuple
                    
                    final_x = (orig_x * rel_x_scale_factor) + abs_x_shift_pixels
                    final_y = (orig_y * rel_y_scale_factor) + abs_y_shift_pixels
                    final_font_size = max(1, int(round(orig_font_size * font_size_scale_factor)))
                    
                    # Rebuild the entire marker entry to ensure all properties are preserved
                    self.markers[i] = [final_x, final_y, text, qcolor, font_family, final_font_size, is_bold, is_italic]
        
                self.populate_table()
                self.global_markers_adjusted.emit(list(self.markers))
        
            def populate_table(self):
                """Rebuilds the entire table from the current state of the data model."""
                self._block_signals = True
                self.table_widget.setSortingEnabled(False)
                self.table_widget.clearContents()
                self.table_widget.setRowCount(len(self.markers) + len(self.shapes))

                # --- Define a standard font for the table display ---
                standard_table_font = QFont("Arial", 12) # A good, readable default

                # --- Populate Markers ---
                for i, marker_data in enumerate(self.markers):
                    x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_data
                    
                    # --- Create Table Items ---
                    type_item = QTableWidgetItem("Marker")
                    type_item.setData(Qt.UserRole, {'type': 'marker', 'original_index': i})
                    
                    text_item = QTableWidgetItem(text)
                    text_item.setFlags(text_item.flags() | Qt.ItemIsEditable)
                    
                    coord_item = QTableWidgetItem(f"{x:.1f},{y:.1f}")
                    coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable)
                    
                    # Display the marker's actual style as text, but don't apply it to the cell
                    style_item = QTableWidgetItem(f"{font_family} ({font_size}pt)")
                    style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)
                    
                    # --- Apply the standard font to all items ---
                    items_to_set = [type_item, text_item, coord_item, style_item]
                    for col, item in enumerate(items_to_set):
                        item.setFont(standard_table_font) # Use the standard font
                        self.table_widget.setItem(i, col, item)
                    
                    # --- Checkboxes ---
                    bold_checkbox = QCheckBox()
                    bold_checkbox.setChecked(is_bold)
                    bold_checkbox.stateChanged.connect(self.handle_marker_style_changed_from_checkbox)
                    cell_widget_bold = QWidget(); layout_bold = QHBoxLayout(cell_widget_bold); layout_bold.addWidget(bold_checkbox); layout_bold.setAlignment(Qt.AlignCenter); layout_bold.setContentsMargins(0,0,0,0)
                    self.table_widget.setCellWidget(i, 4, cell_widget_bold)

                    italic_checkbox = QCheckBox()
                    italic_checkbox.setChecked(is_italic)
                    italic_checkbox.stateChanged.connect(self.handle_marker_style_changed_from_checkbox)
                    cell_widget_italic = QWidget(); layout_italic = QHBoxLayout(cell_widget_italic); layout_italic.addWidget(italic_checkbox); layout_italic.setAlignment(Qt.AlignCenter); layout_italic.setContentsMargins(0,0,0,0)
                    self.table_widget.setCellWidget(i, 5, cell_widget_italic)

                    # --- Color cell ---
                    color_item = QTableWidgetItem(qcolor.name())
                    color_item.setBackground(QBrush(qcolor))
                    color_item.setForeground(QBrush(Qt.white if qcolor.lightness() < 128 else Qt.black))
                    color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
                    color_item.setFont(standard_table_font) # Use the standard font here too
                    self.table_widget.setItem(i, 6, color_item)
                    
                    # --- Delete button ---
                    delete_button = QPushButton("Delete")
                    delete_button.clicked.connect(self.delete_item)
                    self.table_widget.setCellWidget(i, 7, delete_button)

                # --- Populate Shapes (This part remains unchanged as it doesn't use variable fonts) ---
                marker_count = len(self.markers)
                for i, shape_data in enumerate(self.shapes):
                    row_idx = marker_count + i
                    shape_type = shape_data.get('type', 'Unknown').capitalize()
                    type_item = QTableWidgetItem(shape_type); type_item.setData(Qt.UserRole, {'type': 'shape', 'original_index': i})
                    text_item = QTableWidgetItem(""); text_item.setFlags(text_item.flags() & ~Qt.ItemIsEditable)
                    details_str, tooltip_str = "", ""
                    if shape_type == 'Line':
                        start, end = shape_data.get('start', (0,0)), shape_data.get('end', (0,0))
                        details_str = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"
                        tooltip_str = "Edit format: X1,Y1,X2,Y2"
                    elif shape_type == 'Rectangle':
                        x, y, w, h = shape_data.get('rect', (0,0,0,0))
                        details_str = f"{x:.1f},{y:.1f},{w:.1f},{h:.1f}"
                        tooltip_str = "Edit format: X,Y,Width,Height"
                    coord_item = QTableWidgetItem(details_str); coord_item.setToolTip(tooltip_str)
                    if shape_type in ['Line', 'Rectangle']: coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable)
                    thickness = int(shape_data.get('thickness', 1))
                    style_item = QTableWidgetItem(f"{thickness}px"); style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)
                    qcolor = QColor(shape_data.get('color', '#000000'))
                    color_item = QTableWidgetItem(qcolor.name()); color_item.setBackground(QBrush(qcolor))
                    color_item.setForeground(QBrush(Qt.white if qcolor.lightness() < 128 else Qt.black))
                    color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
                    
                    # Apply standard font to shape rows as well for consistency
                    shape_items_to_set = [type_item, text_item, coord_item, style_item, color_item]
                    for item in shape_items_to_set:
                        if item: item.setFont(standard_table_font)

                    self.table_widget.setItem(row_idx, 0, type_item)
                    self.table_widget.setItem(row_idx, 1, text_item)
                    self.table_widget.setItem(row_idx, 2, coord_item)
                    self.table_widget.setItem(row_idx, 3, style_item)
                    self.table_widget.setItem(row_idx, 6, color_item)
                    
                    delete_button = QPushButton("Delete"); delete_button.clicked.connect(self.delete_item)
                    self.table_widget.setCellWidget(row_idx, 7, delete_button)

                self.table_widget.resizeColumnsToContents()
                self.table_widget.setSortingEnabled(True)
                self._block_signals = False
        
            def handle_marker_style_changed_from_checkbox(self, state):
                """Updates the data model when a checkbox is clicked, then repopulates the table."""
                if self._block_signals: return
        
                sender_checkbox = self.sender()
                if not isinstance(sender_checkbox, QCheckBox): return
        
                # Find the row of the clicked checkbox
                for row in range(self.table_widget.rowCount()):
                    is_bold_checkbox = self.table_widget.cellWidget(row, 4) and sender_checkbox is self.table_widget.cellWidget(row, 4).findChild(QCheckBox)
                    is_italic_checkbox = self.table_widget.cellWidget(row, 5) and sender_checkbox is self.table_widget.cellWidget(row, 5).findChild(QCheckBox)
        
                    if is_bold_checkbox or is_italic_checkbox:
                        type_item = self.table_widget.item(row, 0)
                        if not type_item: return
        
                        item_data = type_item.data(Qt.UserRole)
                        if not item_data or item_data.get('type') != 'marker': return
        
                        original_marker_index = item_data.get('original_index')
                        if original_marker_index is None: return
        
                        # Update the data model
                        is_checked = (state == Qt.Checked)
                        idx_to_change = 6 if is_bold_checkbox else 7
                        self.markers[original_marker_index][idx_to_change] = is_checked
                        
                        temp_list = list(self._original_markers_data[original_marker_index])
                        temp_list[idx_to_change] = is_checked
                        self._original_markers_data[original_marker_index] = tuple(temp_list)
        
                        # Repopulate the table to reflect the change visually
                        self.populate_table()
                        self.global_markers_adjusted.emit(list(self.markers))
                        return
        
            def handle_item_changed(self, item):
                """Updates the data model when an editable cell's text is changed."""
                if self._block_signals: return
        
                row, col = item.row(), item.column()
                item_data = self.table_widget.item(row, 0).data(Qt.UserRole)
                if not item_data or item_data['type'] == 'error': return
                
                item_type, original_index = item_data['type'], item_data['original_index']
                new_value = item.text()
                
                # ... (Marker text/coord update logic is unchanged and correct) ...
                if item_type == 'marker':
                    if col == 1: # Text/Label
                        self.markers[original_index][2] = new_value
                        temp_list = list(self._original_markers_data[original_index]); temp_list[2] = new_value; self._original_markers_data[original_index] = tuple(temp_list)
                    elif col == 2: # Coordinates
                        try:
                            x_str, y_str = new_value.split(',')
                            new_x, new_y = float(x_str), float(y_str)
                            
                            self.markers[original_index][0] = new_x
                            self.markers[original_index][1] = new_y
                            
                            # Back-calculate to update the pristine data for global sliders
                            abs_x_shift = (self.abs_x_shift_slider.value() / self.percent_precision_factor / 100.0) * self._current_image_width
                            abs_y_shift = (self.abs_y_shift_slider.value() / self.percent_precision_factor / 100.0) * self._current_image_height
                            rel_x_scale = self.rel_x_scale_slider.value() / self.scale_precision_factor / 100.0
                            rel_y_scale = self.rel_y_scale_slider.value() / self.scale_precision_factor / 100.0
                            
                            base_x = (new_x - abs_x_shift) / rel_x_scale if rel_x_scale != 0 else new_x
                            base_y = (new_y - abs_y_shift) / rel_y_scale if rel_y_scale != 0 else new_y
        
                            temp_list = list(self._original_markers_data[original_index])
                            temp_list[0], temp_list[1] = base_x, base_y
                            self._original_markers_data[original_index] = tuple(temp_list)
        
                        except ValueError:
                            # Revert on error
                            self._block_signals = True
                            prev_x, prev_y = self.markers[original_index][0:2]
                            item.setText(f"{prev_x:.1f},{prev_y:.1f}")
                            self._block_signals = False
                            QMessageBox.warning(self, "Invalid Input", "Coordinates must be in 'X,Y' format (e.g., '100.5,250.2').")
                
                # --- NEW: Handle editable shape coordinates ---
                elif item_type == 'shape' and col == 2:
                    shape_data = self.shapes[original_index]
                    shape_type_internal = shape_data.get('type')
                    try:
                        coords = [float(c.strip()) for c in new_value.split(',')]
                        if shape_type_internal == 'line' and len(coords) == 4:
                            shape_data['start'], shape_data['end'] = (coords[0], coords[1]), (coords[2], coords[3])
                        elif shape_type_internal == 'rectangle' and len(coords) == 4:
                            if coords[2] < 0 or coords[3] < 0: raise ValueError("Width/Height must be non-negative.")
                            shape_data['rect'] = (coords[0], coords[1], coords[2], coords[3])
                        else: raise ValueError("Incorrect number of coordinates.")
                    except ValueError as e:
                        self._block_signals = True
                        self.populate_table() # Revert by repopulating
                        self._block_signals = False
                        QMessageBox.warning(self, "Invalid Input", f"Could not parse shape coordinates.\nError: {e}")
        
                self.shapes_adjusted_preview.emit(list(self.shapes))
                self.global_markers_adjusted.emit(list(self.markers))
        
            def handle_cell_double_clicked(self, row, column):
                """Handles dialogs for non-text edits (color, font, thickness)."""
                if column in [1, 2]: return # Let itemChanged handle these
        
                item_data = self.table_widget.item(row, 0).data(Qt.UserRole)
                if not item_data or item_data['type'] == 'error': return
                item_type, original_index = item_data['type'], item_data['original_index']
        
                if column == 6: # Color
                    current_color = QColor(self.markers[original_index][3] if item_type == 'marker' else self.shapes[original_index]['color'])
                    new_color = QColorDialog.getColor(current_color, self, "Select Color")
                    if new_color.isValid():
                        if item_type == 'marker':
                            self.markers[original_index][3] = new_color
                            temp_list = list(self._original_markers_data[original_index]); temp_list[3] = new_color; self._original_markers_data[original_index] = tuple(temp_list)
                        else: self.shapes[original_index]['color'] = new_color.name()
                        self.populate_table()
        
                elif column == 3 and item_type == 'marker': # Marker Font
                    _x, _y, _text, _qcolor, family, size, bold, italic = self.markers[original_index]
                    current_font = QFont(family, size); current_font.setBold(bold); current_font.setItalic(italic)
                    ok, new_font = QFontDialog.getFont(current_font, self)
                    if ok:
                        self.markers[original_index][4:8] = [new_font.family(), new_font.pointSize(), new_font.bold(), new_font.italic()]
                        temp_list = list(self._original_markers_data[original_index]); temp_list[4:8] = self.markers[original_index][4:8]; self._original_markers_data[original_index] = tuple(temp_list)
                        self.populate_table()
        
                elif column == 3 and item_type == 'shape': # Shape Thickness
                    current_thickness = self.shapes[original_index].get('thickness', 1)
                    new_thickness, ok = QInputDialog.getInt(self, "Set Thickness", "Enter thickness (pixels):", current_thickness, 1, 100)
                    if ok:
                        self.shapes[original_index]['thickness'] = new_thickness
                        self.populate_table()
                
                self.shapes_adjusted_preview.emit(list(self.shapes))
                self.global_markers_adjusted.emit(list(self.markers))
        
            def delete_item(self):
                """Deletes an item from the data model and repopulates the table."""
                clicked_button = self.sender()
                if not clicked_button: return
        
                # Find the visual row of the button that was clicked
                for row in range(self.table_widget.rowCount()):
                    if self.table_widget.cellWidget(row, 7) is clicked_button:
                        item_data = self.table_widget.item(row, 0).data(Qt.UserRole)
                        if not item_data: return
                        
                        item_type, original_index = item_data['type'], item_data['original_index']
                        
                        if item_type == 'marker' and 0 <= original_index < len(self.markers):
                            del self.markers[original_index]
                            del self._original_markers_data[original_index]
                        elif item_type == 'shape' and 0 <= original_index < len(self.shapes):
                            del self.shapes[original_index]
        
                        # Repopulate the table to reflect the deletion
                        self.populate_table()
                        self.global_markers_adjusted.emit(list(self.markers))
                        self.shapes_adjusted_preview.emit(list(self.shapes))
                        return
        
            def get_modified_markers_and_shapes(self):
                """Returns the final state of the data model."""
                return [tuple(m) for m in self.markers], self.shapes

        class TableWindow(QDialog):
            HISTORY_FILE_NAME = "analysis_history.json"
            current_lane_pil_images = {} 

            def __init__(self, current_peak_areas_data, current_standard_dictionary,
                         current_is_standard_mode, current_calculated_quantities_data,
                         parent_app_instance=None, peak_details_data=None):
                super().__init__(parent_app_instance)
                self.setWindowTitle("Analysis Results and History")
                # Increased width to accommodate band numbers next to lane + font slider
                self.setGeometry(100, 100, 1050, 800) 
                self.temp_clipboard_file_path = None
                self.parent_app = parent_app_instance
                self.current_results_data = {} 
                self.is_current_data_multi_lane = isinstance(current_peak_areas_data, dict)
                self.current_lane_pil_images = {}
                self.current_peak_details_data = peak_details_data if peak_details_data else {}

                # --- NEW: Attribute for band number font size ---
                self.band_number_font_size = 12 # Default
                self.band_number_font_slider = None # Will be created in UI

                # ... (rest of __init__ as before, populating self.current_results_data and self.current_lane_pil_images) ...
                if self.is_current_data_multi_lane:
                    for lane_id_key in current_peak_areas_data.keys(): # Iterate through keys of areas dict
                        self.current_results_data[lane_id_key] = {
                            'areas': current_peak_areas_data.get(lane_id_key, []),
                            'quantities': current_calculated_quantities_data.get(lane_id_key, []) if current_calculated_quantities_data else [],
                            'details': self.current_peak_details_data.get(lane_id_key, []) if peak_details_data else []
                        }
                        if self.parent_app and hasattr(self.parent_app, 'multi_lane_definitions'):
                            for lane_def in self.parent_app.multi_lane_definitions:
                                if lane_def['id'] == lane_id_key: 
                                    q_img_region = None
                                    if lane_def['type'] == 'quad': q_img_region = self.parent_app.quadrilateral_to_rect(self.parent_app.image, lane_def['points_label'])
                                    elif lane_def['type'] == 'rectangle':
                                        img_coords = self.parent_app._map_label_rect_to_image_rect(lane_def['points_label'][0])
                                        if img_coords and self.parent_app.image: q_img_region = self.parent_app.image.copy(*img_coords)
                                    if q_img_region and not q_img_region.isNull():
                                        pil_for_lane = self.parent_app.convert_qimage_to_grayscale_pil(q_img_region)
                                        if pil_for_lane: self.current_lane_pil_images[lane_id_key] = pil_for_lane 
                                    break
                elif current_peak_areas_data is not None: 
                     self.current_results_data[1] = {
                            'areas': current_peak_areas_data,
                            'quantities': current_calculated_quantities_data if current_calculated_quantities_data is not None else [],
                            'details': self.current_peak_details_data.get(1, []) if peak_details_data else [] 
                     }
                     if self.parent_app and self.parent_app.image: 
                        extracted_qimage_single = None
                        if self.parent_app.live_view_label.quad_points and len(self.parent_app.live_view_label.quad_points) == 4: extracted_qimage_single = self.parent_app.quadrilateral_to_rect(self.parent_app.image, self.parent_app.live_view_label.quad_points)
                        elif self.parent_app.live_view_label.bounding_box_preview:
                            img_coords_s = self.parent_app._map_label_rect_to_image_rect(QRectF(QPointF(self.parent_app.live_view_label.bounding_box_preview[0],self.parent_app.live_view_label.bounding_box_preview[1]),QPointF(self.parent_app.live_view_label.bounding_box_preview[2],self.parent_app.live_view_label.bounding_box_preview[3])).normalized())
                            if img_coords_s: extracted_qimage_single = self.parent_app.image.copy(*img_coords_s)
                        if extracted_qimage_single and not extracted_qimage_single.isNull():
                            pil_for_lane = self.parent_app.convert_qimage_to_grayscale_pil(extracted_qimage_single)
                            if pil_for_lane: self.current_lane_pil_images[1] = pil_for_lane

                self.current_standard_dictionary = current_standard_dictionary if current_standard_dictionary is not None else {}
                self.current_is_standard_mode = current_is_standard_mode
                self.source_image_name_current = "Unknown"
                if self.parent_app and hasattr(self.parent_app, 'image_path') and self.parent_app.image_path:
                    if isinstance(self.parent_app.image_path, str): self.source_image_name_current = os.path.basename(self.parent_app.image_path)
                self.current_analysis_custom_name = self.source_image_name_current
                self.analysis_name_input_widget = None
                self.analysis_history = [] 
                self.delete_entry_button = None; self.export_previous_button = None; self.previous_sessions_listwidget = None
                self.previous_results_table = None; self.previous_plot_placeholder_label = None
                self.previous_plot_groupbox_layout = None; self.previous_plot_canvas_widget = None
                self.previous_lane_image_preview_label = None 
                self.previous_lane_tab_widget = None 
                self._load_history() 
                main_layout = QVBoxLayout(self); self.tab_widget = QTabWidget(); main_layout.addWidget(self.tab_widget)
                self._create_current_results_tab(); self._create_previous_results_tab() 
                self.dialog_button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                self.dialog_button_box.accepted.connect(self._accept_dialog_and_save_current); self.dialog_button_box.rejected.connect(self.reject)
                main_layout.addWidget(self.dialog_button_box); self.setLayout(main_layout)
                if self.current_results_data: self.tab_widget.setCurrentIndex(0)
                elif self.analysis_history: self.tab_widget.setCurrentIndex(1)
                else: self.tab_widget.setCurrentIndex(0)


            def _create_lane_data_display_widget(self, lane_id, peak_areas, calculated_quantities, is_std_mode,
                                                 pil_lane_image=None, peak_details_for_lane=None, is_for_history=False):

                lane_widget = QWidget()
                if is_for_history or not pil_lane_image:
                    lane_layout = QVBoxLayout(lane_widget)
                else:
                    lane_layout = QHBoxLayout(lane_widget)

                table_scroll_area = QScrollArea(); table_scroll_area.setWidgetResizable(True)
                table_for_lane = QTableWidget(); table_for_lane.setColumnCount(4)
                table_for_lane.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
                table_for_lane.setEditTriggers(QTableWidget.NoEditTriggers); table_for_lane.setSelectionBehavior(QTableWidget.SelectRows)
                single_lane_data_for_populator = {1: {'areas': peak_areas, 'quantities': calculated_quantities}}
                self._populate_table_generic(table_for_lane, single_lane_data_for_populator, is_std_mode, is_multi_lane_data=False)
                table_scroll_area.setWidget(table_for_lane)
                # Give table more stretch factor if image is present
                lane_layout.addWidget(table_scroll_area, 3 if pil_lane_image and not is_for_history else 1) 

                if pil_lane_image and not is_for_history:
                    image_preview_group = QGroupBox(f"Lane {lane_id} Preview")
                    image_preview_layout = QVBoxLayout(image_preview_group)
                    
                    lane_image_label = QLabel("Image preview not available.")
                    lane_image_label.setMinimumSize(100, 200) # Keep a reasonable minimum
                    # Allow label to expand vertically, and take preferred width
                    lane_image_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding) 
                    lane_image_label.setAlignment(Qt.AlignCenter)
                    lane_image_label.setStyleSheet("border: 1px solid grey; background-color: #333;")
                    # Crucial for making the pixmap scale with the label:
                    lane_image_label.setScaledContents(True) 
                    lane_widget.setProperty("lane_image_label_ref", lane_image_label)

                    try:
                        if pil_lane_image.width > pil_lane_image.height:
                             display_pil_image_oriented = pil_lane_image.rotate(0, expand=True)
                        else:
                             display_pil_image_oriented = pil_lane_image.copy()
                        
                        display_pil_image_rgba = None
                        # ... (Normalization logic as before) ...
                        if display_pil_image_oriented.mode in ['I', 'I;16'] or display_pil_image_oriented.mode.startswith("I;16"):
                            img_array = np.array(display_pil_image_oriented, dtype=np.float32); 
                            min_val, max_val = np.percentile(img_array, 1), np.percentile(img_array, 99) # Wider percentiles for contrast
                            if max_val <= min_val: min_val, max_val = np.min(img_array), np.max(img_array)
                            normalized_array = (img_array - min_val) / (max_val - min_val + 1e-9) 
                            normalized_array = np.clip(normalized_array, 0.0, 1.0)
                            img_8bit_gray = (normalized_array * 255).astype(np.uint8); display_pil_image_rgba = Image.fromarray(img_8bit_gray, mode='L').convert("RGBA")
                        elif display_pil_image_oriented.mode == 'L': display_pil_image_rgba = ImageOps.autocontrast(display_pil_image_oriented, cutoff=1).convert("RGBA") # cutoff for autocontrast
                        elif display_pil_image_oriented.mode == 'F':
                            img_array = np.array(display_pil_image_oriented, dtype=np.float32); min_val, max_val = np.percentile(img_array, 1), np.percentile(img_array, 99)
                            if max_val <= min_val: min_val, max_val = np.min(img_array), np.max(img_array)
                            normalized_array = (img_array - min_val) / (max_val - min_val + 1e-9)
                            normalized_array = np.clip(normalized_array, 0.0, 1.0)
                            img_8bit_gray = (normalized_array * 255).astype(np.uint8); display_pil_image_rgba = Image.fromarray(img_8bit_gray, mode='L').convert("RGBA")
                        else: display_pil_image_rgba = display_pil_image_oriented.convert("RGBA")

                        if display_pil_image_rgba:
                            
                            draw = ImageDraw.Draw(display_pil_image_rgba)
                            
                            # --- INCREASED FONT SIZE for Band Numbers ---
                            # This size is relative to the full-resolution display_pil_image_rgba
                            band_number_pil_font_size = 12 # Try a larger size (e.g., 24, 28, 32)
                                                            # This will be scaled down by the QLabel if needed.
                            try: font_pil = ImageFont.truetype("arialbd.ttf", band_number_pil_font_size)
                            except IOError: 
                                try: font_pil = ImageFont.truetype("arial.ttf", band_number_pil_font_size)
                                except: font_pil = ImageFont.load_default() # Will be small
                            
                            text_color = (255, 0, 0, 255) # Bright Red

                            if peak_details_for_lane:
                                for i, peak_info in enumerate(peak_details_for_lane):
                                    y_pixel_on_oriented_image = int(peak_info['y_coord_in_lane_image'])
                                    y_pixel_on_oriented_image = max(0, min(y_pixel_on_oriented_image, display_pil_image_rgba.height -1))
                                    band_num_str = str(i + 1)
                                    text_bbox = draw.textbbox((0, 0), band_num_str, font=font_pil)
                                    text_width = text_bbox[2] - text_bbox[0]
                                    text_height = text_bbox[3] - text_bbox[1]
                                    
                                    # Position text to the very left of the lane image strip
                                    x_text_pos = 3 # Small padding from the left edge
                                    actual_y_draw_pos = y_pixel_on_oriented_image - (text_height // 2) - text_bbox[1] 
                                    
                                    draw.text((x_text_pos, actual_y_draw_pos), band_num_str, fill=text_color, font=font_pil)
                                    
                                    # Optional guide line
                                    line_start_x = x_text_pos + text_width + 2 
                                    line_end_x = x_text_pos + text_width + 10 # Make line a bit longer
                                    line_y = y_pixel_on_oriented_image 
                                    draw.line([(line_start_x, line_y), (line_end_x, line_y)], fill=(0,255,0,180), width=2) # Thicker, more visible guide

                            q_image_lane = ImageQt.ImageQt(display_pil_image_rgba)
                            if not q_image_lane.isNull():
                                pixmap_lane = QPixmap.fromImage(q_image_lane)
                                # Set the pixmap directly. QLabel with setScaledContents(True) will handle scaling.
                                lane_image_label.setPixmap(pixmap_lane) 
                        else:
                             lane_image_label.setText(f"Cannot display Lane {lane_id} (format error).")
                    except Exception as e_img: 
                        print(f"Error creating current lane image preview for lane {lane_id}: {e_img}"); traceback.print_exc()
                        lane_image_label.setText(f"Error displaying lane {lane_id} preview.")
                
                    image_preview_layout.addWidget(lane_image_label)
                    image_preview_group.setMinimumWidth(150) # Maintain a minimum width
                    image_preview_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding) # Allow vertical expansion
                    lane_layout.addWidget(image_preview_group, 1) # Image group takes less horizontal stretch
                return lane_widget
                

            def _get_config_dir(self):
                """Determines the directory for storing configuration/history files."""
                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()
                return application_path

            def _load_history(self):
                """Loads analysis history from the JSON file."""
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                if os.path.exists(history_file_path):
                    try:
                        with open(history_file_path, "r", encoding='utf-8') as f:
                            loaded_data = json.load(f)
                        if isinstance(loaded_data, list):
                            self.analysis_history = [entry for entry in loaded_data if isinstance(entry, dict)]
                        else:
                            self.analysis_history = []
                    except (json.JSONDecodeError, IOError) as e:
                        self.analysis_history = []
                else:
                    self.analysis_history = []

            def _save_history(self):
                """Saves the current analysis_history list to the JSON file."""
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                try:
                    with open(history_file_path, "w", encoding='utf-8') as f:
                        json.dump(self.analysis_history, f, indent=4)
                except IOError as e:
                    QMessageBox.critical(self, "Save History Error", f"Could not save analysis history to {history_file_path}: {e}")

            def _accept_dialog_and_save_current(self):
                if self.current_results_data: 
                    peak_dialog_settings_current = {}
                    if self.parent_app and hasattr(self.parent_app, 'peak_dialog_settings'):
                        peak_dialog_settings_current = self.parent_app.peak_dialog_settings.copy()
                    user_defined_analysis_name = self.analysis_name_input_widget.text().strip() if self.analysis_name_input_widget else ""
                    display_name_for_history = user_defined_analysis_name if user_defined_analysis_name else self.source_image_name_current
                    
                    # self.current_results_data already contains 'areas', 'quantities', and 'details' per lane
                    new_entry = {
                        "timestamp": datetime.datetime.now().isoformat(),
                        "user_defined_name": display_name_for_history,
                        "source_image_name": self.source_image_name_current,
                        "is_multi_lane": self.is_current_data_multi_lane,
                        "results_data": self.current_results_data, # This now includes 'details'
                        "standard_dictionary": self.current_standard_dictionary,
                        "analysis_settings": peak_dialog_settings_current
                    }
                    self.analysis_history.insert(0, new_entry)
                    self._save_history()
                    if self.previous_sessions_listwidget: self._populate_previous_sessions_list()
                self.accept()

            def _create_current_results_tab(self):
                current_tab_widget = QWidget()
                current_main_layout = QVBoxLayout(current_tab_widget) 

                name_layout = QHBoxLayout()
                name_label = QLabel("Analysis Name:")
                self.analysis_name_input_widget = QLineEdit(self.current_analysis_custom_name)
                self.analysis_name_input_widget.setPlaceholderText("Enter a name for this analysis...")
                self.analysis_name_input_widget.setToolTip("Used for history and export filenames.")
                name_layout.addWidget(name_label); name_layout.addWidget(self.analysis_name_input_widget)
                current_main_layout.addLayout(name_layout)

                # --- FONT SLIDER REMOVED ---
                # font_control_layout = QHBoxLayout()
                # ... (slider and label creation removed)
                # current_main_layout.addLayout(font_control_layout) 

                current_plot_widget = self._create_standard_curve_plot_generic(
                    self.current_standard_dictionary, self.current_is_standard_mode, for_history=False
                )
                if current_plot_widget:
                    plot_group_current = QGroupBox("Standard Curve (Current Analysis)")
                    plot_layout_current = QVBoxLayout(plot_group_current)
                    plot_layout_current.addWidget(current_plot_widget)
                    plot_group_current.setMaximumHeight(250) 
                    plot_group_current.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                    current_main_layout.addWidget(plot_group_current)

                results_display_area = QWidget()
                results_display_layout = QVBoxLayout(results_display_area); results_display_layout.setContentsMargins(0,0,0,0)

                if self.is_current_data_multi_lane and len(self.current_results_data) > 0:
                    self.current_lanes_tab_widget = QTabWidget() 
                    for lane_id_sorted in sorted(self.current_results_data.keys()):
                        lane_data = self.current_results_data[lane_id_sorted]
                        pil_image_for_lane = self.current_lane_pil_images.get(lane_id_sorted)
                        peak_details_for_this_lane = lane_data.get('details', []) 
                        lane_content_widget = self._create_lane_data_display_widget(
                            lane_id_sorted, lane_data['areas'], lane_data['quantities'], 
                            self.current_is_standard_mode, pil_image_for_lane, peak_details_for_this_lane,
                            is_for_history=False # Explicitly for current
                        )
                        self.current_lanes_tab_widget.addTab(lane_content_widget, f"Lane {lane_id_sorted}")
                    results_display_layout.addWidget(self.current_lanes_tab_widget)
                    # No need to connect currentChanged to font refresh anymore
                elif 1 in self.current_results_data: 
                    lane_data = self.current_results_data[1]; pil_image_for_lane = self.current_lane_pil_images.get(1); peak_details_for_this_lane = lane_data.get('details', [])
                    self.single_lane_content_widget_ref = self._create_lane_data_display_widget(
                        1, lane_data['areas'], lane_data['quantities'], self.current_is_standard_mode, 
                        pil_image_for_lane, peak_details_for_this_lane, is_for_history=False
                    )
                    results_display_layout.addWidget(self.single_lane_content_widget_ref)
                else: 
                    no_data_label = QLabel("No current analysis data to display."); no_data_label.setAlignment(Qt.AlignCenter); results_display_layout.addWidget(no_data_label)
                current_main_layout.addWidget(results_display_area)
                
                current_buttons_layout = QHBoxLayout(); copy_current_button = QPushButton("Copy Active Lane Table"); copy_current_button.clicked.connect(self._copy_active_lane_table_data)
                export_current_button = QPushButton("Export All Lanes to Excel")
                export_current_button.clicked.connect(lambda: self._export_to_excel_generic(self.current_results_data, self.analysis_name_input_widget.text() or self.source_image_name_current, self.current_standard_dictionary, is_multi_lane_data=self.is_current_data_multi_lane ))
                current_buttons_layout.addWidget(copy_current_button); current_buttons_layout.addStretch(); current_buttons_layout.addWidget(export_current_button); current_main_layout.addLayout(current_buttons_layout)
                self.tab_widget.addTab(current_tab_widget, "Current Analysis")
                
                       
            def _copy_active_lane_table_data(self):
                """Copies data from the table of the currently active lane tab (if multi-lane)."""
                table_to_copy = None
                if self.is_current_data_multi_lane and hasattr(self, 'current_lanes_tab_widget'):
                    current_lane_widget = self.current_lanes_tab_widget.currentWidget()
                    if current_lane_widget:
                        # Find the QTableWidget within the current lane's content widget
                        # This assumes _create_lane_data_display_widget places table in a predictable way
                        table_widgets_in_lane = current_lane_widget.findChildren(QTableWidget)
                        if table_widgets_in_lane:
                            table_to_copy = table_widgets_in_lane[0]
                elif hasattr(self, 'current_results_table_single_lane_ref'): # Fallback to old single table if it existed
                    # This assumes that if it's single lane, the table is directly accessible
                    # We might need a direct reference if _create_lane_data_display_widget is the only table creator
                    # For now, let's assume the single lane case also creates its table inside a widget
                    # that can be found similarly or by direct reference if we store it.
                    # For simplicity, if not multi-lane, let's assume there's a main table.
                    # The logic in _create_current_results_tab needs to ensure self.current_results_table
                    # refers to the single lane's table if not multi-lane.
                    # The previous version had self.current_results_table.
                    # If _create_lane_data_display_widget is used for single lanes too, that single table is the target.
                    # Let's find it generically.
                    current_tab_content = self.tab_widget.widget(0) # "Current Analysis" tab
                    if current_tab_content:
                         # If it's single lane, the table might be directly in current_main_layout or wrapped.
                         # Assuming _create_lane_data_display_widget is used, find it.
                         lane_widgets = current_tab_content.findChildren(QWidget) # Find the lane_widget
                         for lw in lane_widgets:
                             # A bit hacky way to identify our lane widget; could add objectName
                             if lw.layout() and isinstance(lw.layout(), QHBoxLayout) and lw.findChild(QScrollArea):
                                 table_widgets_in_lw = lw.findChildren(QTableWidget)
                                 if table_widgets_in_lw:
                                     table_to_copy = table_widgets_in_lw[0]
                                     break
                
                if table_to_copy:
                    self._copy_table_data_generic(table_to_copy)
                else:
                    QMessageBox.information(self, "Copy Error", "Could not find the active lane's table to copy.")

            def _create_previous_results_tab(self):
                previous_tab_widget = QWidget()
                previous_main_layout = QHBoxLayout(previous_tab_widget)

                # --- Left Pane (History List) ---
                left_pane_widget = QWidget()
                left_layout = QVBoxLayout(left_pane_widget); left_layout.setContentsMargins(0, 0, 5, 0)
                left_layout.addWidget(QLabel("Saved Analyses:"))
                self.previous_sessions_listwidget = QListWidget()
                self.previous_sessions_listwidget.itemSelectionChanged.connect(self._on_history_session_selected)
                left_layout.addWidget(self.previous_sessions_listwidget)
                history_buttons_layout = QHBoxLayout()
                self.delete_entry_button = QPushButton("Delete Selected"); self.delete_entry_button.clicked.connect(self._delete_selected_history_entry); self.delete_entry_button.setEnabled(False)
                history_buttons_layout.addWidget(self.delete_entry_button); history_buttons_layout.addStretch()
                self.clear_history_button = QPushButton("Clear All History"); self.clear_history_button.clicked.connect(self._clear_all_history)
                history_buttons_layout.addWidget(self.clear_history_button); left_layout.addLayout(history_buttons_layout)
                previous_main_layout.addWidget(left_pane_widget, 1)

                # --- Right Pane (Details of selected history) ---
                self.right_pane_history_widget = QWidget() # Main container for right side
                right_layout = QVBoxLayout(self.right_pane_history_widget) # Layout for this container

                self.previous_plot_groupbox = QGroupBox("Standard Curve (Selected History)")
                self.previous_plot_groupbox_layout = QVBoxLayout(self.previous_plot_groupbox)
                self.previous_plot_placeholder_label = QLabel("Select an analysis from the list to view details.")
                self.previous_plot_placeholder_label.setAlignment(Qt.AlignCenter)
                self.previous_plot_groupbox_layout.addWidget(self.previous_plot_placeholder_label)
                self.previous_plot_groupbox.setMaximumHeight(250) # Consistent height
                self.previous_plot_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                right_layout.addWidget(self.previous_plot_groupbox)

                # This widget will hold the lane tabs or single lane data for history
                self.history_results_display_container = QWidget() 
                self.history_results_display_layout = QVBoxLayout(self.history_results_display_container)
                self.history_results_display_layout.setContentsMargins(0,0,0,0)
                # Add a placeholder initially for the table/tabs area
                initial_hist_table_placeholder = QLabel("Lane data will appear here.")
                initial_hist_table_placeholder.setAlignment(Qt.AlignCenter)
                self.history_results_display_layout.addWidget(initial_hist_table_placeholder)
                right_layout.addWidget(self.history_results_display_container, 1) # Give it stretch factor 1

                previous_table_buttons_layout = QHBoxLayout()
                copy_previous_button = QPushButton("Copy Active History Lane Table") # Changed text
                copy_previous_button.clicked.connect(self._copy_active_history_lane_table_data)
                self.export_previous_button = QPushButton("Export Selected History to Excel") # Changed text
                self.export_previous_button.clicked.connect(self._export_selected_history_to_excel)
                self.export_previous_button.setEnabled(False)
                previous_table_buttons_layout.addWidget(copy_previous_button); previous_table_buttons_layout.addStretch()
                previous_table_buttons_layout.addWidget(self.export_previous_button)
                right_layout.addLayout(previous_table_buttons_layout)

                previous_main_layout.addWidget(self.right_pane_history_widget, 2) # right_pane takes more space
                self.tab_widget.addTab(previous_tab_widget, "Analysis History")
                self._populate_previous_sessions_list()
                
            def _copy_active_history_lane_table_data(self):
                """Copies data from the table of the currently active historical lane tab."""
                table_to_copy = None
                # Check if the active widget in history_results_display_container is a QTabWidget
                if hasattr(self, 'history_results_display_container'):
                    # The direct child should be the QTabWidget or the single_lane_display_widget
                    active_content_widget = self.history_results_display_layout.itemAt(0).widget() if self.history_results_display_layout.count() > 0 else None

                    if isinstance(active_content_widget, QTabWidget): # It's multi-lane history
                        current_lane_hist_widget = active_content_widget.currentWidget()
                        if current_lane_hist_widget:
                            table_widgets_in_lane = current_lane_hist_widget.findChildren(QTableWidget)
                            if table_widgets_in_lane:
                                table_to_copy = table_widgets_in_lane[0]
                    elif isinstance(active_content_widget, QWidget): # It's single-lane history widget
                         table_widgets_in_lane = active_content_widget.findChildren(QTableWidget)
                         if table_widgets_in_lane:
                             table_to_copy = table_widgets_in_lane[0]

                if table_to_copy:
                    self._copy_table_data_generic(table_to_copy)
                else:
                    QMessageBox.information(self, "Copy Error", "Could not find the active history lane's table to copy.")

            def _populate_previous_sessions_list(self):
                self.previous_sessions_listwidget.clear()
                if not self.analysis_history:
                    self.previous_sessions_listwidget.addItem("No history available.")
                    if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                    if self.export_previous_button: self.export_previous_button.setEnabled(False)
                    return

                for i, entry in enumerate(self.analysis_history):
                    ts_str = entry.get("timestamp", f"Entry {len(self.analysis_history) - i}")
                    
                    # --- Use user_defined_name if available, otherwise fallback ---
                    entry_display_name = entry.get("user_defined_name", "").strip()
                    source_img_name = entry.get("source_image_name", "Unknown Image")
                    
                    if not entry_display_name: # If user_defined_name is empty or missing
                        entry_display_name = source_img_name # Fallback to source image name

                    try:
                        dt_obj = datetime.datetime.fromisoformat(ts_str.split('.')[0])
                        # Display format: "YYYY-MM-DD HH:MM:SS (User Defined Name OR Source Image Name)"
                        final_display_string = f"{dt_obj.strftime('%Y-%m-%d %H:%M:%S')} ({entry_display_name})"
                    except ValueError:
                        final_display_string = f"{ts_str} ({entry_display_name})"
                    # --- End Name Logic ---
                    self.previous_sessions_listwidget.addItem(final_display_string)
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)

            def _clear_previous_details_view(self):
                # Clear content of history_results_display_layout
                if hasattr(self, 'history_results_display_layout'):
                    while self.history_results_display_layout.count() > 0:
                        item = self.history_results_display_layout.takeAt(0)
                        widget = item.widget()
                        if widget: widget.deleteLater()
                    # Add back a placeholder
                    initial_hist_table_placeholder = QLabel("Select an analysis to view details.")
                    initial_hist_table_placeholder.setAlignment(Qt.AlignCenter)
                    self.history_results_display_layout.addWidget(initial_hist_table_placeholder)


                if hasattr(self, 'previous_plot_canvas_widget') and self.previous_plot_canvas_widget:
                    if self.previous_plot_groupbox_layout: self.previous_plot_groupbox_layout.removeWidget(self.previous_plot_canvas_widget)
                    self.previous_plot_canvas_widget.deleteLater(); self.previous_plot_canvas_widget = None
                if self.previous_plot_placeholder_label: 
                    self.previous_plot_placeholder_label.setText("Select an analysis from the list to view details.")
                    if self.previous_plot_placeholder_label.isHidden(): self.previous_plot_placeholder_label.show()
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)

            def _on_history_session_selected(self):
                if not self.previous_sessions_listwidget: return
                selected_items = self.previous_sessions_listwidget.selectedItems()
                if not selected_items or "No history available." in selected_items[0].text():
                    self._clear_previous_details_view()
                    return

                selected_row_index = self.previous_sessions_listwidget.currentRow()
                if not (0 <= selected_row_index < len(self.analysis_history)):
                    self._clear_previous_details_view()
                    return

                entry = self.analysis_history[selected_row_index]
                if self.delete_entry_button: self.delete_entry_button.setEnabled(True)
                if self.export_previous_button: self.export_previous_button.setEnabled(True)

                hist_is_multi_lane_from_flag = entry.get("is_multi_lane", False)
                processed_hist_results = {} 

                if "results_data" in entry and isinstance(entry["results_data"], dict) and entry["results_data"]:
                    temp_results_data = entry["results_data"]
                    for lane_id_str, lane_content in temp_results_data.items():
                        try:
                            lane_id = int(lane_id_str)
                            processed_hist_results[lane_id] = {
                                'areas': lane_content.get('areas', []),
                                'quantities': lane_content.get('quantities', []),
                                'details': lane_content.get('details', []) 
                            }
                        except (ValueError, TypeError): pass # Skip invalid
                    if len(processed_hist_results) > 1: hist_is_multi_lane_from_flag = True 
                    elif len(processed_hist_results) == 1: hist_is_multi_lane_from_flag = False
                elif isinstance(entry.get("peak_areas"), list):
                    legacy_areas = entry.get("peak_areas", []); legacy_quantities = entry.get("calculated_quantities", [])
                    legacy_details_raw = entry.get("peak_details"); legacy_details_list = []
                    if isinstance(legacy_details_raw, dict) and 1 in legacy_details_raw: legacy_details_list = legacy_details_raw[1]
                    elif isinstance(legacy_details_raw, list): legacy_details_list = legacy_details_raw
                    if legacy_areas: processed_hist_results[1] = {'areas': legacy_areas, 'quantities': legacy_quantities, 'details': legacy_details_list}
                    hist_is_multi_lane_from_flag = False
                else: pass # processed_hist_results remains empty

                hist_std_dict = entry.get("standard_dictionary", {})
                hist_is_std_mode = bool(hist_std_dict)

                # --- Clear and Repopulate the History Details Pane ---
                if hasattr(self, 'previous_plot_canvas_widget') and self.previous_plot_canvas_widget:
                    if self.previous_plot_groupbox_layout: self.previous_plot_groupbox_layout.removeWidget(self.previous_plot_canvas_widget)
                    self.previous_plot_canvas_widget.deleteLater(); self.previous_plot_canvas_widget = None
                if self.previous_plot_placeholder_label and not self.previous_plot_placeholder_label.isHidden():
                    self.previous_plot_placeholder_label.hide()
                
                while self.history_results_display_layout.count() > 0: # Clear previous content
                    item = self.history_results_display_layout.takeAt(0)
                    widget = item.widget()
                    if widget: widget.deleteLater()
                
                display_as_multi_lane_hist = hist_is_multi_lane_from_flag and len(processed_hist_results) > 0
                display_as_single_lane_hist = (not hist_is_multi_lane_from_flag) and (1 in processed_hist_results) and len(processed_hist_results) == 1

                if display_as_multi_lane_hist:
                    hist_lane_tabs = QTabWidget()
                    for lane_id_hist_sorted in sorted(processed_hist_results.keys()):
                        lane_data_hist = processed_hist_results[lane_id_hist_sorted]
                        # Call _create_lane_data_display_widget with is_for_history=True
                        # It will only create the table part.
                        lane_content_hist_widget = self._create_lane_data_display_widget(
                            lane_id_hist_sorted, lane_data_hist.get('areas', []), lane_data_hist.get('quantities', []),
                            hist_is_std_mode, pil_lane_image=None, # No image for history
                            peak_details_for_lane=lane_data_hist.get('details', []), 
                            is_for_history=True
                        )
                        hist_lane_tabs.addTab(lane_content_hist_widget, f"Lane {lane_id_hist_sorted}")
                    self.history_results_display_layout.addWidget(hist_lane_tabs)
                elif display_as_single_lane_hist: 
                    lane_data_hist = processed_hist_results[1]
                    peak_details_hist = lane_data_hist.get('details', [])
                    single_hist_lane_widget = self._create_lane_data_display_widget(
                        1, lane_data_hist.get('areas', []), lane_data_hist.get('quantities', []), hist_is_std_mode, 
                        pil_lane_image=None, peak_details_for_lane=peak_details_hist,
                        is_for_history=True
                    )
                    self.history_results_display_layout.addWidget(single_hist_lane_widget)
                else: 
                    no_hist_data_label = QLabel("No valid lane results data found in this history entry.")
                    no_hist_data_label.setAlignment(Qt.AlignCenter)
                    self.history_results_display_layout.addWidget(no_hist_data_label)
                
                hist_plot_widget = self._create_standard_curve_plot_generic(hist_std_dict, hist_is_std_mode, for_history=True)
                self.previous_plot_canvas_widget = hist_plot_widget
                if self.previous_plot_groupbox_layout: self.previous_plot_groupbox_layout.addWidget(self.previous_plot_canvas_widget)


            def _delete_selected_history_entry(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    reply = QMessageBox.question(self, "Confirm Delete",
                                                 "Are you sure you want to delete this history entry?",
                                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        del self.analysis_history[current_row]
                        self._save_history()
                        self._populate_previous_sessions_list()
                        self._clear_previous_details_view()
                else:
                    QMessageBox.information(self, "No Selection", "Please select an entry to delete.")

            def _clear_all_history(self):
                reply = QMessageBox.question(self, "Confirm Clear All History",
                                             "Are you sure you want to delete ALL saved analysis entries?\nThis action cannot be undone.",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    self.analysis_history = []
                    self._save_history()
                    self._populate_previous_sessions_list() # This will show "No history available."
                    self._clear_previous_details_view()

            def _export_selected_history_to_excel(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    entry = self.analysis_history[current_row]
                    # ... (filename generation logic remains similar) ...
                    ts_str_raw = entry.get("timestamp", f"History_Entry_{current_row+1}")
                    try: dt_obj_export = datetime.datetime.fromisoformat(ts_str_raw.split('.')[0]); timestamp_str_for_file = dt_obj_export.strftime("%Y%m%d_%H%M%S")
                    except ValueError: timestamp_str_for_file = ts_str_raw.replace(":", "-").replace("T", "_").split('.')[0]
                    analysis_name_part = entry.get("user_defined_name", "").strip()
                    if not analysis_name_part: analysis_name_part = entry.get("source_image_name", "UnknownAnalysis")
                    analysis_name_part_sanitized = analysis_name_part.replace('.', '_').replace(' ', '_').replace(':', '-')
                    max_name_len = 50
                    if len(analysis_name_part_sanitized) > max_name_len: analysis_name_part_sanitized = analysis_name_part_sanitized[:max_name_len]
                    default_filename_base = f"Analysis_{timestamp_str_for_file}_{analysis_name_part_sanitized}"

                    hist_is_multi_lane = entry.get("is_multi_lane", False)
                    hist_results_data = entry.get("results_data", {})
                    
                    self._export_to_excel_generic(
                        hist_results_data, # Pass the raw data dict/list
                        default_filename_base,
                        entry.get("standard_dictionary", {}),
                        is_multi_lane_data=hist_is_multi_lane
                    )
                else:
                    QMessageBox.information(self, "No Selection", "Please select a history entry to export.")

            # --- Generic Helper Methods (Keep these as they were, they are fine) ---
            def _create_standard_curve_plot_generic(self, standard_dictionary, is_standard_mode, for_history=False):
                if not is_standard_mode or not standard_dictionary or len(standard_dictionary) < 2:
                    no_curve_label = QLabel("Standard curve requires at least 2 standard points." if not for_history else "No standard data for this historical entry.")
                    no_curve_label.setAlignment(Qt.AlignCenter)
                    return no_curve_label
                try:
                    quantities = np.array(list(standard_dictionary.keys()), dtype=float)
                    areas = np.array(list(standard_dictionary.values()), dtype=float)
                    if len(quantities) < 2: # Double check after potential type conversion
                         no_curve_label = QLabel("Insufficient valid standard points (less than 2).")
                         no_curve_label.setAlignment(Qt.AlignCenter)
                         return no_curve_label

                    coeffs = np.polyfit(areas, quantities, 1); slope, intercept = coeffs
                    predicted_quantities = np.polyval(coeffs, areas); residuals = quantities - predicted_quantities
                    ss_res = np.sum(residuals**2); ss_tot = np.sum((quantities - np.mean(quantities))**2)
                    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0 # Avoid division by zero if all quantities are same
                    
                    fig, ax = plt.subplots(figsize=(4.5, 3.2)) # Slightly adjusted for dialog
                    fig.set_dpi(90)

                    ax.scatter(areas, quantities, label='Standard Points', color='red', zorder=5, s=30)
                    
                    # Ensure x_line covers the range of areas, handle empty/single point arrays
                    x_min_plot = np.min(areas) * 0.9 if areas.size > 0 else 0
                    x_max_plot = np.max(areas) * 1.1 if areas.size > 0 else 1
                    if x_min_plot == x_max_plot: # Handle case where all areas are the same
                        x_min_plot -= 0.5
                        x_max_plot += 0.5

                    x_line = np.linspace(x_min_plot, x_max_plot, 100)
                    y_line = slope * x_line + intercept
                    
                    fit_label = (f'Qty = {slope:.3g}*Area + {intercept:.3g}\nR² = {r_squared:.3f}')
                    ax.plot(x_line, y_line, label=fit_label, color='blue', linewidth=1.2)
                    
                    ax.set_xlabel('Total Peak Area', fontsize=8)
                    ax.set_ylabel('Known Quantity', fontsize=8)
                    title_prefix = "Historical " if for_history else "" # Removed "Current" as it's implied
                    ax.set_title(f'{title_prefix}Standard Curve', fontsize=9, fontweight='bold')
                    
                    ax.legend(fontsize='xx-small', loc='best'); ax.grid(True, linestyle=':', alpha=0.7, linewidth=0.5)
                    ax.tick_params(axis='both', which='major', labelsize=7)
                    
                    if np.any(areas > 1e4) or (np.any(areas < 1e-2) and np.any(areas != 0)): ax.ticklabel_format(style='sci', axis='x', scilimits=(0,0), useMathText=True)
                    if np.any(quantities > 1e4) or (np.any(quantities < 1e-2) and np.any(quantities != 0)): ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                    
                    try: fig.set_constrained_layout(True)
                    except AttributeError: plt.tight_layout(pad=0.3)
                    
                    canvas = FigureCanvas(fig)
                    canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding); canvas.updateGeometry()
                    plt.close(fig)

                    return canvas
                except Exception as e:
                    traceback.print_exc()
                    error_label = QLabel(f"Error generating plot:\n{str(e)[:100]}...") # Show truncated error
                    error_label.setAlignment(Qt.AlignCenter); error_label.setStyleSheet("color: red;")
                    return error_label

            def _populate_table_generic(self, table_widget, results_data, is_standard_mode, is_multi_lane_data):
                table_widget.clearContents()
                table_widget.setRowCount(0) # Clear rows
                
                total_rows_to_add = 0
                if not results_data: # Empty dict or list
                    placeholder_text = "No data to display in table."
                    row_span = 1
                    col_span = table_widget.columnCount()
                    table_widget.setRowCount(1)
                    item = QTableWidgetItem(placeholder_text)
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(0, 0, item)
                    if col_span > 0 : table_widget.setSpan(0, 0, row_span, col_span)
                    table_widget.resizeColumnsToContents()
                    return

                current_row_idx = 0
                # results_data is dict {lane_id: {'areas': [], 'quantities': []}}
                for lane_id_sorted in sorted(results_data.keys()): 
                    lane_data = results_data[lane_id_sorted]
                    peak_areas = lane_data.get('areas', [])
                    calculated_quantities = lane_data.get('quantities', [])

                    if not peak_areas: continue # Skip empty lanes

                    total_area_this_lane = sum(peak_areas) if peak_areas else 0.0
                    
                    table_widget.setRowCount(current_row_idx + len(peak_areas)) # Ensure enough rows

                    for band_idx, area in enumerate(peak_areas):
                        col_offset = 0
                        if is_multi_lane_data:
                            table_widget.setItem(current_row_idx, 0, QTableWidgetItem(str(lane_id_sorted)))
                            col_offset = 1
                        
                        band_label = f"Band {band_idx + 1}"
                        table_widget.setItem(current_row_idx, col_offset + 0, QTableWidgetItem(band_label))
                        table_widget.setItem(current_row_idx, col_offset + 1, QTableWidgetItem(f"{area:.3f}"))
                        percentage_str = f"{(area / total_area_this_lane * 100):.2f}%" if total_area_this_lane != 0 else "0.00%"
                        table_widget.setItem(current_row_idx, col_offset + 2, QTableWidgetItem(percentage_str))
                        
                        quantity_str = ""
                        if is_standard_mode and calculated_quantities and band_idx < len(calculated_quantities):
                            quantity_str = f"{calculated_quantities[band_idx]:.2f}"
                        table_widget.setItem(current_row_idx, col_offset + 3, QTableWidgetItem(quantity_str))
                        current_row_idx += 1
                
                if current_row_idx == 0: # If all lanes were empty
                    self._populate_table_generic(table_widget, None, False, False) # Call with None to show placeholder
                else:
                    table_widget.resizeColumnsToContents()


            def _copy_table_data_generic(self, table_widget_source):
                if not table_widget_source: return
                selected_ranges = table_widget_source.selectedRanges()
                if not selected_ranges: return
                # ... (rest of copy logic is fine)
                selected_range = selected_ranges[0]
                start_row, end_row = selected_range.topRow(), selected_range.bottomRow()
                start_col, end_col = selected_range.leftColumn(), selected_range.rightColumn()
                clipboard_string = ""
                for r in range(start_row, end_row + 1):
                    row_data = []
                    for c in range(start_col, end_col + 1):
                        item = table_widget_source.item(r, c)
                        # Check if item is the placeholder span
                        if table_widget_source.rowSpan(r,c) > 1 or table_widget_source.columnSpan(r,c) > 1:
                            if "No data" in item.text() or "Select an analysis" in item.text():
                                row_data.append("") # Add empty string for placeholder cells
                                continue
                        row_data.append(item.text() if item else "")
                    if any(cell_text for cell_text in row_data): # Only add row if not all empty
                        clipboard_string += "\t".join(row_data) + "\n"
                QApplication.clipboard().setText(clipboard_string.strip())


            def _export_to_excel_generic(self, results_data_for_export, analysis_name_for_filename_base="Analysis_Results", 
                                         standard_dict_for_export=None, is_multi_lane_data=False):
                # ... (filename generation logic remains similar) ...
                safe_analysis_name = str(analysis_name_for_filename_base).replace('.', '_').replace(' ', '_').replace(':', '-')
                current_timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                final_default_filename = ""
                if "Analysis_" in safe_analysis_name and any(char.isdigit() for char in safe_analysis_name): 
                    final_default_filename = safe_analysis_name
                else: 
                    final_default_filename = f"Analysis_{current_timestamp_str}_{safe_analysis_name}"
                options = QFileDialog.Options(); file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Excel File", f"{final_default_filename}.xlsx", "Excel Files (*.xlsx)", options=options)
                if not file_path: return
                
                workbook = openpyxl.Workbook()
                # Remove default sheet, we will create our own
                if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])

                # --- Main Data Sheet ---
                if is_multi_lane_data and results_data_for_export: # results_data_for_export is a dict
                    worksheet_data = workbook.create_sheet("Multi-Lane Analysis")
                    headers = ["Lane ID", "Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"]
                    for col, header in enumerate(headers, start=1):
                        cell = worksheet_data.cell(row=1, column=col, value=header); cell.font = Font(bold=True)
                    
                    current_excel_row = 2
                    for lane_id_sorted in sorted(results_data_for_export.keys()):
                        lane_data = results_data_for_export[lane_id_sorted]
                        peak_areas = lane_data.get('areas', [])
                        calculated_quantities = lane_data.get('quantities', [])
                        if not peak_areas: continue
                        total_area_this_lane = sum(peak_areas)

                        for band_idx, area_val in enumerate(peak_areas):
                            worksheet_data.cell(row=current_excel_row, column=1, value=lane_id_sorted)
                            worksheet_data.cell(row=current_excel_row, column=2, value=f"Band {band_idx + 1}")
                            worksheet_data.cell(row=current_excel_row, column=3, value=float(f"{area_val:.3f}"))
                            perc_val_str = f"{(area_val / total_area_this_lane * 100):.2f}%" if total_area_this_lane != 0 else "0.00%"
                            try: # Store as number with % format
                                perc_num = float(perc_val_str.replace('%','')) / 100.0
                                cell_perc = worksheet_data.cell(row=current_excel_row, column=4, value=perc_num)
                                cell_perc.number_format = '0.00%'
                            except ValueError:
                                worksheet_data.cell(row=current_excel_row, column=4, value=perc_val_str)
                            
                            qty_str_val = ""
                            if self.current_is_standard_mode and calculated_quantities and band_idx < len(calculated_quantities):
                                qty_str_val = f"{calculated_quantities[band_idx]:.2f}"
                                try: worksheet_data.cell(row=current_excel_row, column=5, value=float(qty_str_val))
                                except ValueError: worksheet_data.cell(row=current_excel_row, column=5, value=qty_str_val)
                            else:
                                worksheet_data.cell(row=current_excel_row, column=5, value=qty_str_val)
                            current_excel_row += 1
                    
                    # Auto-size columns
                    for col_idx_letter in range(1, worksheet_data.max_column + 1):
                        column_letter = openpyxl.utils.get_column_letter(col_idx_letter)
                        max_length = 0; header_len = len(headers[col_idx_letter-1])
                        for cell in worksheet_data[column_letter]:
                            try: 
                                if cell.value: max_length = max(max_length, len(str(cell.value)) + (1 if cell.number_format == '0.00%' else 0) )
                            except: pass
                        adjusted_width = (max(max_length, header_len) + 2) * 1.1 
                        worksheet_data.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)


                elif not is_multi_lane_data and results_data_for_export and 1 in results_data_for_export: # Single lane data
                    worksheet_data_single = workbook.create_sheet("Single Lane Analysis")
                    headers_single = ["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"]
                    for col, header in enumerate(headers_single, start=1):
                        cell = worksheet_data_single.cell(row=1, column=col, value=header); cell.font = Font(bold=True)
                    
                    single_lane_actual_data = results_data_for_export[1] # Access the data for lane 1
                    peak_areas_single = single_lane_actual_data.get('areas', [])
                    calculated_quantities_single = single_lane_actual_data.get('quantities', [])
                    total_area_single = sum(peak_areas_single) if peak_areas_single else 0.0

                    for r_idx, area_val in enumerate(peak_areas_single):
                        excel_r = r_idx + 2
                        worksheet_data_single.cell(row=excel_r, column=1, value=f"Band {r_idx+1}")
                        worksheet_data_single.cell(row=excel_r, column=2, value=float(f"{area_val:.3f}"))
                        perc_val_str = f"{(area_val / total_area_single * 100):.2f}%" if total_area_single != 0 else "0.00%"
                        try: 
                            perc_num = float(perc_val_str.replace('%','')) / 100.0
                            cell_perc = worksheet_data_single.cell(row=excel_r, column=3, value=perc_num)
                            cell_perc.number_format = '0.00%'
                        except ValueError: worksheet_data_single.cell(row=excel_r, column=3, value=perc_val_str)
                        
                        qty_str_val = ""
                        if self.current_is_standard_mode and calculated_quantities_single and r_idx < len(calculated_quantities_single):
                            qty_str_val = f"{calculated_quantities_single[r_idx]:.2f}"
                            try: worksheet_data_single.cell(row=excel_r, column=4, value=float(qty_str_val))
                            except ValueError: worksheet_data_single.cell(row=excel_r, column=4, value=qty_str_val)
                        else:
                            worksheet_data_single.cell(row=excel_r, column=4, value=qty_str_val)

                    for col_idx_letter in range(1, worksheet_data_single.max_column + 1):
                        column_letter = openpyxl.utils.get_column_letter(col_idx_letter)
                        max_length = 0; header_len = len(headers_single[col_idx_letter-1])
                        for cell in worksheet_data_single[column_letter]:
                            try: 
                                if cell.value: max_length = max(max_length, len(str(cell.value)) + (1 if cell.number_format == '0.00%' else 0))
                            except: pass
                        adjusted_width = (max(max_length, header_len) + 2) * 1.1
                        worksheet_data_single.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)
                else: # No data to export
                    no_data_sheet = workbook.create_sheet("No Analysis Data")
                    no_data_sheet.cell(row=1, column=1, value="No analysis data was available for export.")


                # --- Standard Curve Data Sheet (remains the same) ---
                if standard_dict_for_export and len(standard_dict_for_export) >= 2:
                    worksheet_std = workbook.create_sheet("Standard Curve Data")
                    worksheet_std.cell(row=1, column=1, value="Known Quantity").font = Font(bold=True)
                    worksheet_std.cell(row=1, column=2, value="Total Peak Area").font = Font(bold=True)
                    current_row_std = 2
                    for qty_key, area_val in sorted(standard_dict_for_export.items()):
                        try:
                            worksheet_std.cell(row=current_row_std, column=1, value=float(qty_key))
                            worksheet_std.cell(row=current_row_std, column=2, value=float(area_val))
                            current_row_std += 1
                        except ValueError: print(f"Warning: Skipping invalid standard data for Excel: Qty={qty_key}, Area={area_val}")
                    for col_dim_std_letter in range(1, worksheet_std.max_column + 1): 
                         column_letter_std = openpyxl.utils.get_column_letter(col_dim_std_letter)
                         worksheet_std.column_dimensions[column_letter_std].auto_size = True
                try:
                    workbook.save(file_path)
                    QMessageBox.information(self, "Success", f"Table data exported to\n{file_path}")
                except Exception as e:
                     QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                        

        class PeakAreaDialog(QDialog):
            """
            Interactive dialog to adjust peak regions and calculate peak areas.
            Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
            Peak region boundaries are now manipulated directly on the image strip plot.
            Optimized for fast, interactive boundary adjustments using blitting.
            """
            HANDLE_SIZE = 8 # Pixel size for draggable handles on ax_image

            def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
                super().__init__(parent)
                self.parent_app = parent
                self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
                self.setGeometry(100, 100, 1000, 750)

                if not isinstance(cropped_data, Image.Image):
                     raise TypeError("Input 'cropped_data' must be a PIL Image object")

                self.original_pil_cropped_data = cropped_data
                self.enhanced_cropped_image_for_display = None

                self.original_max_value = 255.0
                pil_mode = self.original_pil_cropped_data.mode
                try:
                    if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'L':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        self.original_max_value = 255.0
                    elif pil_mode == 'F':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                        self.original_max_value = max(1.0, max_in_float)
                    else:
                        gray_img = self.original_pil_cropped_data.convert("L")
                        self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                        self.original_max_value = 255.0
                except Exception as e:
                    raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

                if self.intensity_array_original_range.ndim != 2:
                     raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

                self.profile_original_inverted = None; self.profile = None; self.background = None
                self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
                self.smoothing_sigma = current_settings.get('smoothing_sigma', 2.0)
                self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
                self.peak_distance = current_settings.get('peak_distance', 10)
                self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.00)
                self.peak_broadening_pixels = current_settings.get('valley_offset_pixels', 0) 
                self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
                self.area_subtraction_method = current_settings.get('area_subtraction_method', "Rolling Ball")
                self.auto_adjust_rb_radius = current_settings.get('auto_adjust_rb_radius', False)
                self.peaks = np.array([]); self.initial_valley_regions = []; self.peak_regions = []
                self.denoise_sigma = current_settings.get('denoise_sigma', 0.0)
                self.peak_areas_rolling_ball = []; self.peak_areas_straight_line = []; self.peak_areas_valley = []
                self._final_settings = {}; self._persist_enabled_on_exit = persist_checked
                
                self.manual_select_mode_active = False
                self.selected_peak_for_ui_focus = -1

                self.add_peak_mode_active = False; self.selected_peak_index_for_delete = -1

                self.background_blit = None
                self.dragging_handle_info = None
                self.interactive_artists = []
                
                if rolling_ball is None or find_peaks is None or gaussian_filter1d is None or interp1d is None or cv2 is None or ImageOps is None:
                     QMessageBox.critical(self, "Dependency Error","Missing required libraries...")
                
                self._setup_ui(persist_checked)
                self.regenerate_profile_and_detect()

            def _setup_ui(self, persist_checked_initial):
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(10)

                self.fig = plt.figure(figsize=(10, 6))
                gs = GridSpec(2, 1, height_ratios=[3, 1.5], hspace=0.1, figure=self.fig)
                self.ax = self.fig.add_subplot(gs[0])
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)
                
                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                
                self.canvas.mpl_connect('draw_event', self.on_draw)
                self.canvas.mpl_connect('button_press_event', self.on_canvas_press)
                self.canvas.mpl_connect('motion_notify_event', self.on_canvas_motion)
                self.canvas.mpl_connect('button_release_event', self.on_canvas_release)

                main_layout.addWidget(self.canvas, stretch=1)

                controls_main_hbox = QHBoxLayout()
                left_controls_vbox = QVBoxLayout()
                global_settings_group = QGroupBox("Global Settings & Area Method")
                global_settings_layout = QGridLayout(global_settings_group)
                global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
                self.band_estimation_combobox = QComboBox()
                self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
                self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
                self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
                global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2)
                global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
                self.method_combobox = QComboBox()
                self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
                self.method_combobox.setCurrentText(self.area_subtraction_method)
                self.method_combobox.currentIndexChanged.connect(self._on_method_changed)
                global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2)
                
                self.rolling_ball_label = QLabel(f"Rolling Ball Radius ({int(self.rolling_ball_radius)})")
                self.rolling_ball_label.setMinimumWidth(160)
                self.rolling_ball_slider = QSlider(Qt.Horizontal)
                self.rolling_ball_slider.setRange(1, 500)
                self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
                # FIX: Connect to the new handler for real-time region updates
                self.rolling_ball_slider.valueChanged.connect(self._on_rb_slider_changed)
                self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_label: lbl.setText(f"Rolling Ball Radius ({val})"))
                
                self.auto_adjust_checkbox = QCheckBox("Auto")
                self.auto_adjust_checkbox.setToolTip("Automatically calculate the optimal rolling ball radius based on detected peak widths.")
                self.auto_adjust_checkbox.setChecked(self.auto_adjust_rb_radius)
                self.auto_adjust_checkbox.stateChanged.connect(self.toggle_auto_adjust_rb)
                
                global_settings_layout.addWidget(self.rolling_ball_label, 2, 0)
                global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1)
                global_settings_layout.addWidget(self.auto_adjust_checkbox, 2, 2)
                
                left_controls_vbox.addWidget(global_settings_group)
                
                peak_detect_group = QGroupBox("Peak Detection & Manipulation")
                peak_detect_layout = QGridLayout(peak_detect_group)
                peak_detect_layout.addWidget(QLabel("Detected Peaks:"), 0, 0); self.peak_number_input = QLineEdit(); self.peak_number_input.setPlaceholderText("#"); self.peak_number_input.setMaximumWidth(60); self.update_peak_number_button = QPushButton("Set"); self.update_peak_number_button.clicked.connect(self.manual_peak_number_update); peak_detect_layout.addWidget(self.peak_number_input, 0, 1); peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2); self.denoise_sigma_label = QLabel(f"Denoise Sigma ({self.denoise_sigma:.1f})"); self.denoise_sigma_slider = QSlider(Qt.Horizontal); self.denoise_sigma_slider.setRange(0,50); self.denoise_sigma_slider.setValue(int(self.denoise_sigma*10)); self.denoise_sigma_slider.valueChanged.connect(lambda val,lbl=self.denoise_sigma_label: lbl.setText(f"Denoise Sigma ({val/10.0:.1f})")); self.denoise_sigma_slider.valueChanged.connect(self.regenerate_profile_and_detect); peak_detect_layout.addWidget(self.denoise_sigma_label,1,0); peak_detect_layout.addWidget(self.denoise_sigma_slider,1,1,1,2); self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})"); self.smoothing_slider = QSlider(Qt.Horizontal); self.smoothing_slider.setRange(0,100); self.smoothing_slider.setValue(int(self.smoothing_sigma*10)); self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})")); self.smoothing_slider.valueChanged.connect(self.regenerate_profile_and_detect); peak_detect_layout.addWidget(self.smoothing_label,2,0); peak_detect_layout.addWidget(self.smoothing_slider,2,1,1,2); self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})"); self.peak_prominence_slider = QSlider(Qt.Horizontal); self.peak_prominence_slider.setRange(0,100); self.peak_prominence_slider.setValue(int(self.peak_prominence_factor*100)); self.peak_prominence_slider.valueChanged.connect(self.detect_peaks); self.peak_prominence_slider.valueChanged.connect(lambda val,lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})")); peak_detect_layout.addWidget(self.peak_prominence_slider_label,3,0); peak_detect_layout.addWidget(self.peak_prominence_slider,3,1,1,2); self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})"); self.peak_height_slider = QSlider(Qt.Horizontal); self.peak_height_slider.setRange(0,100); self.peak_height_slider.setValue(int(self.peak_height_factor*100)); self.peak_height_slider.valueChanged.connect(self.detect_peaks); self.peak_height_slider.valueChanged.connect(lambda val,lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f})")); peak_detect_layout.addWidget(self.peak_height_slider_label,4,0); peak_detect_layout.addWidget(self.peak_height_slider,4,1,1,2); self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px"); self.peak_distance_slider = QSlider(Qt.Horizontal); self.peak_distance_slider.setRange(1,200); self.peak_distance_slider.setValue(self.peak_distance); self.peak_distance_slider.valueChanged.connect(self.detect_peaks); self.peak_distance_slider.valueChanged.connect(lambda val,lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px")); peak_detect_layout.addWidget(self.peak_distance_slider_label,5,0); peak_detect_layout.addWidget(self.peak_distance_slider,5,1,1,2);
                self.broadening_label = QLabel(f"Peak Broadening ({'+' if self.peak_broadening_pixels>=0 else ''}{self.peak_broadening_pixels} px)"); self.broadening_slider = QSlider(Qt.Horizontal); self.broadening_slider.setRange(-50, 50); self.broadening_slider.setValue(self.peak_broadening_pixels); self.broadening_slider.valueChanged.connect(lambda value, lbl=self.broadening_label: lbl.setText(f"Peak Broadening ({'+' if value>=0 else ''}{value} px)")); self.broadening_slider.sliderReleased.connect(self.apply_peak_broadening_on_release); peak_detect_layout.addWidget(self.broadening_label, 6, 0); peak_detect_layout.addWidget(self.broadening_slider, 6, 1, 1, 2); self.add_peak_manually_button = QPushButton("Add Peak"); self.add_peak_manually_button.setCheckable(True); self.add_peak_manually_button.clicked.connect(self.toggle_add_peak_mode); self.delete_selected_peak_button = QPushButton("Delete Peak"); self.delete_selected_peak_button.setEnabled(False); self.delete_selected_peak_button.clicked.connect(self.delete_selected_peak_action); self.identify_peak_button = QPushButton("Focus Peak"); self.identify_peak_button.setCheckable(True); self.identify_peak_button.clicked.connect(self.toggle_manual_select_mode); peak_detect_layout.addWidget(self.add_peak_manually_button, 7, 0, 1, 1); peak_detect_layout.addWidget(self.delete_selected_peak_button, 7, 1, 1, 1); peak_detect_layout.addWidget(self.identify_peak_button, 7, 2, 1, 1); self.copy_regions_button = QPushButton("Copy Regions"); self.copy_regions_button.clicked.connect(self.copy_peak_regions_to_app); self.paste_regions_button = QPushButton("Paste Regions"); self.paste_regions_button.clicked.connect(self.paste_peak_regions_from_app);
                if not (self.parent_app and self.parent_app.copied_peak_regions_data.get("regions")): self.paste_regions_button.setEnabled(False)
                peak_detect_layout.addWidget(self.copy_regions_button, 8,0,1,1); peak_detect_layout.addWidget(self.paste_regions_button, 8,1,1,2); left_controls_vbox.addWidget(peak_detect_group); left_controls_vbox.addStretch(1); controls_main_hbox.addLayout(left_controls_vbox, stretch=1); main_layout.addLayout(controls_main_hbox); bottom_button_layout = QHBoxLayout(); self.persist_settings_checkbox = QCheckBox("Persist Settings"); self.persist_settings_checkbox.setChecked(persist_checked_initial); bottom_button_layout.addWidget(self.persist_settings_checkbox); bottom_button_layout.addStretch(1); self.ok_button = QPushButton("OK"); self.ok_button.setDefault(True); self.ok_button.clicked.connect(self.accept_and_close); self.cancel_button = QPushButton("Cancel"); self.cancel_button.clicked.connect(self.reject); bottom_button_layout.addWidget(self.ok_button); bottom_button_layout.addWidget(self.cancel_button); main_layout.addLayout(bottom_button_layout); self.setLayout(main_layout)
                self.update_rb_controls_enabled_state()

            def _on_method_changed(self):
                """Handles area method changes by recalculating regions and then updating the plot."""
                self.update_rb_controls_enabled_state()
                self._recalculate_all_regions()
                self.update_plot()
            
            def _on_rb_slider_changed(self, value):
                """Handles real-time updates when the rolling ball slider is moved."""
                # This handler is only active when manual adjustment is possible.
                if not self.auto_adjust_checkbox.isChecked() and self.method_combobox.currentText() == "Rolling Ball":
                    # Update the internal radius value from the slider.
                    self.rolling_ball_radius = value
                    
                    # Recalculate the background and the regions based on it.
                    self._recalculate_all_regions()
                    
                    # Update the plot to show the new state.
                    self.update_plot()

            def update_rb_controls_enabled_state(self):
                """Enables/disables controls based on the selected area method."""
                is_rb_method = self.method_combobox.currentText() == "Rolling Ball"
                is_valley_method = self.method_combobox.currentText() == "Valley-to-Valley"
                
                self.rolling_ball_slider.setEnabled(is_rb_method and not self.auto_adjust_checkbox.isChecked())
                self.auto_adjust_checkbox.setEnabled(is_rb_method)
                self.rolling_ball_label.setEnabled(is_rb_method)
                self.broadening_slider.setEnabled(is_valley_method)
                self.broadening_label.setEnabled(is_valley_method)

            def toggle_auto_adjust_rb(self, state):
                """Handles the 'Auto' checkbox state change for rolling ball radius."""
                self.auto_adjust_rb_radius = bool(state)
                self.rolling_ball_slider.setEnabled(not self.auto_adjust_rb_radius)
                self.detect_peaks()

            def _calculate_optimal_rb_radius(self):
                if self.profile is None or len(self.peaks) == 0:
                    return 50 
                try:
                    profile_range = np.ptp(self.profile)
                    if profile_range < 1e-6: profile_range = 1.0
                    min_height_abs = np.min(self.profile) + profile_range * self.peak_height_factor
                    min_prominence_abs = max(1.0, profile_range * self.peak_prominence_factor)
                    _peaks_for_width, properties = find_peaks(
                        self.profile, height=min_height_abs, prominence=min_prominence_abs,
                        distance=self.peak_distance, width=1
                    )
                    if 'widths' in properties and len(properties['widths']) > 0:
                        avg_width = np.mean(properties['widths'])
                        optimal_radius = int(np.clip(2.5 * avg_width, 10, 500))
                        print(f"Auto-calculated optimal rolling ball radius: {optimal_radius} (from avg peak width: {avg_width:.2f})")
                        return optimal_radius
                    else:
                        return 50
                except Exception as e:
                    print(f"Error calculating optimal rolling ball radius: {e}")
                    return 50

            def _recalculate_all_regions(self):
                """The single source of truth for calculating peak_regions based on the current method."""
                self.method = self.method_combobox.currentText()
                if self.method == "Rolling Ball":
                    if rolling_ball and self.profile_original_inverted is not None:
                        try:
                            profile_float = self.profile_original_inverted.astype(np.float64)
                            safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
                            self.background = self._custom_rolling_ball(profile_float, safe_radius) if len(profile_float) > 1 else profile_float.copy()
                            self.background = np.maximum(self.background, 0)
                        except:
                            self.background = np.zeros_like(self.profile_original_inverted)
                    else:
                        self.background = np.zeros_like(self.profile_original_inverted) if self.profile_original_inverted is not None else np.array([])
                    
                    self._redefine_regions_from_background(self.background)
                else: 
                    self._redefine_all_valley_regions()

            def _redefine_regions_from_background(self, background):
                """Helper to define peak regions based on intersections with a given background."""
                self.peak_regions = []
                profile_to_analyze = self.profile_original_inverted
                if profile_to_analyze is None or len(profile_to_analyze) <= 1 or len(self.peaks) == 0:
                    return

                midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2 if len(self.peaks) > 1 else []
                search_boundaries_left = np.concatenate(([0], midpoints))
                search_boundaries_right = np.concatenate((midpoints, [len(profile_to_analyze) - 1]))

                for i, peak_idx in enumerate(self.peaks):
                    left_bound = int(search_boundaries_left[i])
                    right_bound = int(search_boundaries_right[i])
                    start, end = self._find_intersection_boundaries(
                        profile_to_analyze, background, peak_idx, left_bound, right_bound
                    )
                    self.peak_regions.append((start, end))
            
            def detect_peaks(self):
                if self.profile is None or len(self.profile) == 0:
                    self.peaks = np.array([])
                else:
                    self.peak_height_factor = self.peak_height_slider.value()/100.0
                    self.peak_distance = self.peak_distance_slider.value()
                    self.peak_prominence_factor = self.peak_prominence_slider.value()/100.0
                    profile_range = np.ptp(self.profile)
                    min_height_abs = np.min(self.profile) + profile_range * self.peak_height_factor
                    min_prominence_abs = max(1.0, profile_range * self.peak_prominence_factor)
                    try:
                        self.peaks, _ = find_peaks(self.profile, height=min_height_abs, prominence=min_prominence_abs, distance=self.peak_distance, width=1)
                    except Exception:
                        self.peaks = np.array([])
                
                if self.auto_adjust_rb_radius:
                    self.rolling_ball_radius = self._calculate_optimal_rb_radius()
                    if hasattr(self, 'rolling_ball_slider'):
                        self.rolling_ball_slider.blockSignals(True)
                        self.rolling_ball_slider.setValue(self.rolling_ball_radius)
                        self.rolling_ball_slider.blockSignals(False)
                    if hasattr(self, 'rolling_ball_label'):
                        self.rolling_ball_label.setText(f"Rolling Ball Radius ({self.rolling_ball_radius})")

                if hasattr(self, 'peak_number_input') and not self.peak_number_input.hasFocus():
                    self.peak_number_input.setText(str(len(self.peaks)))
                
                self._recalculate_all_regions()
                self.update_plot()

            def accept_and_close(self):
                self._final_settings = {
                    'rolling_ball_radius': self.rolling_ball_slider.value(),
                    'denoise_sigma': self.denoise_sigma_slider.value() / 10.0,
                    'peak_height_factor': self.peak_height_slider.value() / 100.0,
                    'peak_distance': self.peak_distance_slider.value(),
                    'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
                    'valley_offset_pixels': self.broadening_slider.value(),
                    'band_estimation_method': self.band_estimation_combobox.currentText(),
                    'area_subtraction_method': self.method_combobox.currentText(),
                    'smoothing_sigma': self.smoothing_slider.value() / 10.0,
                    'auto_adjust_rb_radius': self.auto_adjust_checkbox.isChecked()
                }
                self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
                self.accept()
                
            def update_plot(self):
                if self.canvas is None: return
                profile_to_plot_and_calc = self.profile_original_inverted
                if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
                     self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1.5], hspace=0.1, figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax); self.ax_image.set_xlabel("Pixel Index"); self.ax.tick_params(axis='x', labelbottom=False); self.ax_image.text(0.5, 0.5, 'No Profile Data', ha='center', va='center', transform=self.ax_image.transAxes); self.canvas.draw_idle(); return
                
                # The background is now assumed to be calculated and correct before calling update_plot.
                if not hasattr(self, 'background') or self.background is None or self.background.shape != profile_to_plot_and_calc.shape:
                    self.background = np.zeros_like(profile_to_plot_and_calc)
                
                self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1.5], hspace=0.1, figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)
                self.interactive_artists.clear()
                
                plot_label_with_denoise = f"Profile (Denoise σ={self.denoise_sigma:.1f}, Smooth σ={self.smoothing_sigma:.1f})"
                self.ax.plot(profile_to_plot_and_calc, label=plot_label_with_denoise, color="black", lw=1.2)
                
                if len(self.peaks) > 0:
                     valid_peaks_indices = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
                     if len(valid_peaks_indices) > 0:
                         peak_y_on_smoothed = profile_to_plot_and_calc[valid_peaks_indices]
                         self.ax.scatter(valid_peaks_indices, peak_y_on_smoothed, color="red", marker='x', s=50, label="Peaks", zorder=5) 
                         if self.selected_peak_for_ui_focus != -1 and 0 <= self.selected_peak_for_ui_focus < len(self.peaks):
                             focused_peak_x_val = self.peaks[self.selected_peak_for_ui_focus]
                             self.ax.plot(focused_peak_x_val, profile_to_plot_and_calc[focused_peak_x_val], 'o', markersize=12, markeredgecolor='orange', markerfacecolor='none', label='Focused', zorder=6)
                         if self.selected_peak_index_for_delete != -1:
                             self.ax.plot(self.selected_peak_index_for_delete, profile_to_plot_and_calc[self.selected_peak_index_for_delete], 's', markersize=14, markeredgecolor='blue', markerfacecolor='none', label='Selected for Delete', zorder=7)
                
                self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
                
                profile_range_plot = np.ptp(profile_to_plot_and_calc) if np.ptp(profile_to_plot_and_calc) > 0 else 1.0
                max_y_for_plot_limit = np.max(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 1

                for i in range(len(self.peak_regions)):
                    start_handle, end_handle = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                    if start_handle >= end_handle or i >= len(self.peaks): continue
                    peak_x = self.peaks[i]
                    baseline_rb = self.background
                    y_baseline_sl_points = np.interp([start_handle, end_handle], [start_handle, end_handle], [profile_to_plot_and_calc[start_handle], profile_to_plot_and_calc[end_handle]])
                    baseline_sl = np.interp(np.arange(len(profile_to_plot_and_calc)), [start_handle, end_handle], y_baseline_sl_points)
                    y_baseline_vv_level = max(profile_to_plot_and_calc[start_handle], profile_to_plot_and_calc[end_handle])
                    baseline_vv = np.full_like(profile_to_plot_and_calc, y_baseline_vv_level)
                    start_calc_rb, end_calc_rb = self._find_intersection_boundaries(profile_to_plot_and_calc, baseline_rb, peak_x, start_handle, end_handle)
                    start_calc_sl, end_calc_sl = self._find_intersection_boundaries(profile_to_plot_and_calc, baseline_sl, peak_x, start_handle, end_handle)
                    start_calc_vv, end_calc_vv = self._find_intersection_boundaries(profile_to_plot_and_calc, baseline_vv, peak_x, start_handle, end_handle)
                    area_rb = np.trapz(profile_to_plot_and_calc[start_calc_rb:end_calc_rb+1] - baseline_rb[start_calc_rb:end_calc_rb+1]) if start_calc_rb < end_calc_rb else 0.0
                    self.peak_areas_rolling_ball.append(max(0, area_rb))
                    area_sl = np.trapz(profile_to_plot_and_calc[start_calc_sl:end_calc_sl+1] - baseline_sl[start_calc_sl:end_calc_sl+1]) if start_calc_sl < end_calc_sl else 0.0
                    self.peak_areas_straight_line.append(max(0, area_sl))
                    area_vv = np.trapz(profile_to_plot_and_calc[start_calc_vv:end_calc_vv+1] - baseline_vv[start_calc_vv:end_calc_vv+1]) if start_calc_vv < end_calc_vv else 0.0
                    self.peak_areas_valley.append(max(0, area_vv))
                    if self.method == "Rolling Ball":
                        x_region = np.arange(start_calc_rb, end_calc_rb + 1)
                        self.ax.fill_between(x_region, baseline_rb[x_region], profile_to_plot_and_calc[x_region], color="yellow", alpha=0.4, interpolate=True)
                        if i == 0: self.ax.plot(np.arange(len(baseline_rb)), baseline_rb, color="green", ls="--", lw=1, label="Rolling Ball BG")
                        area_text = area_rb
                    elif self.method == "Straight Line":
                        x_region = np.arange(start_calc_sl, end_calc_sl + 1)
                        self.ax.fill_between(x_region, baseline_sl[x_region], profile_to_plot_and_calc[x_region], color="cyan", alpha=0.4, interpolate=True)
                        self.ax.plot([start_handle, end_handle], y_baseline_sl_points, color="purple", ls="--", lw=1, label="SL BG" if i == 0 else "")
                        area_text = area_sl
                    elif self.method == "Valley-to-Valley":
                        x_region = np.arange(start_calc_vv, end_calc_vv + 1)
                        self.ax.fill_between(x_region, baseline_vv[x_region], profile_to_plot_and_calc[x_region], color="lightblue", alpha=0.4, interpolate=True)
                        self.ax.plot([start_handle, end_handle], [y_baseline_vv_level, y_baseline_vv_level], color="orange", ls="--", lw=1, label="Valley BG" if i == 0 else "")
                        area_text = area_vv
                    text_x_pos = peak_x; text_y_pos = profile_to_plot_and_calc[peak_x] + profile_range_plot * 0.03
                    self.ax.text(text_x_pos, text_y_pos, f"{area_text:.0f}", ha="center", va="bottom", fontsize=7, color='black', bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="gray", alpha=0.7))
                    max_y_for_plot_limit = max(max_y_for_plot_limit, text_y_pos)

                self.ax.set_ylabel("Intensity"); self.ax.legend(fontsize='small', loc='upper right'); self.ax.set_title(f"Profile & Peak Regions"); self.ax.tick_params(axis='x', which='both', bottom=False, labelbottom=False)
                if len(profile_to_plot_and_calc) > 1: self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1); self.ax.set_ylim(bottom=min(0, np.min(profile_to_plot_and_calc)), top=max_y_for_plot_limit * 1.05)
                if np.max(profile_to_plot_and_calc) > 10000: self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                
                self.ax_image.clear() 
                if hasattr(self, 'enhanced_cropped_image_for_display') and self.enhanced_cropped_image_for_display:
                    rotated_pil_image_display = self.enhanced_cropped_image_for_display.rotate(90, expand=True)
                    image_extent = [0, len(profile_to_plot_and_calc) - 1, 0, rotated_pil_image_display.height]
                    self.ax_image.imshow(np.array(rotated_pil_image_display), cmap='gray', aspect='auto', extent=image_extent)
                    
                    self.ax_image.set_yticks([]); self.ax_image.set_ylabel("Lane Width", fontsize='small'); self.ax_image.set_xlabel("Pixel Index", fontsize='small')
                    for peak_idx, (start_px, end_px) in enumerate(self.peak_regions):
                        line_color, zorder_val, lw = ('orange', 11, 2.0) if peak_idx == self.selected_peak_for_ui_focus else ('blue', 10, 1.5)
                        start_line = mlines.Line2D([start_px, start_px], [0, rotated_pil_image_display.height], color=line_color, lw=lw, picker=self.HANDLE_SIZE, zorder=zorder_val); self.ax_image.add_line(start_line); self.interactive_artists.append((peak_idx, 'start_line', start_line))
                        end_line = mlines.Line2D([end_px, end_px], [0, rotated_pil_image_display.height], color=line_color, lw=lw, picker=self.HANDLE_SIZE, zorder=zorder_val); self.ax_image.add_line(end_line); self.interactive_artists.append((peak_idx, 'end_line', end_line))
                
                self.fig.tight_layout(pad=0.5); self.canvas.draw_idle(); plt.close(self.fig)

            def on_draw(self, event): self.background_blit = self.canvas.copy_from_bbox(self.fig.bbox)
            def on_canvas_press(self, event):
                if event.inaxes != self.ax_image or event.button != 1:
                    if event.inaxes == self.ax: self.handle_profile_plot_click(event)
                    return
                for i, (peak_idx, handle_type, artist) in enumerate(self.interactive_artists):
                    contains, _ = artist.contains(event)
                    if contains:
                        self.dragging_handle_info = {'peak_index': peak_idx, 'type': handle_type, 'artist': artist}
                        self.selected_peak_for_ui_focus = peak_idx
                        for a in self.interactive_artists: a[2].set_animated(False)
                        artist.set_animated(True)
                        self.canvas.draw()
                        return
                self.dragging_handle_info = None
            def on_canvas_motion(self, event):
                if not self.dragging_handle_info or event.inaxes != self.ax_image: return
                self.canvas.restore_region(self.background_blit)
                artist = self.dragging_handle_info['artist']
                new_x = int(round(event.xdata)) if event.xdata is not None else 0
                artist.set_xdata([new_x, new_x])
                self.ax_image.draw_artist(artist)
                self.canvas.blit(self.ax_image.bbox)
            def on_canvas_release(self, event):
                if not self.dragging_handle_info: return
                artist = self.dragging_handle_info['artist']
                artist.set_animated(False)
                self.background_blit = None
                peak_idx = self.dragging_handle_info['peak_index']
                handle_type = self.dragging_handle_info['type']
                new_x = int(round(event.xdata)) if event.xdata is not None else 0
                profile_len = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                new_x = max(0, min(new_x, profile_len - 1))
                current_start, current_end = self.peak_regions[peak_idx]
                if handle_type == 'start_line':
                    new_start = min(new_x, current_end - 1 if current_end > 0 else 0)
                    self.peak_regions[peak_idx] = (new_start, current_end)
                elif handle_type == 'end_line':
                    new_end = max(new_x, current_start + 1 if current_start < profile_len - 1 else profile_len - 1)
                    self.peak_regions[peak_idx] = (current_start, new_end)
                self.dragging_handle_info = None
                self.update_plot()
                self.canvas.draw_idle()
            def handle_profile_plot_click(self, event):
                if self.add_peak_mode_active:
                     if event.button == 1 and event.xdata is not None:
                        clicked_x = int(round(event.xdata))
                        if self.profile_original_inverted is not None and 0 <= clicked_x < len(self.profile_original_inverted): self.add_manual_peak(clicked_x)
                elif self.manual_select_mode_active:
                    if event.button == 1 and event.xdata is not None and self.peaks.any():
                        clicked_x = int(round(event.xdata)); distances = np.abs(self.peaks - clicked_x); closest_peak_idx = np.argmin(distances)
                        click_tolerance = self.peak_distance / 2.0 if self.peak_distance > 0 else 10.0
                        if distances[closest_peak_idx] <= click_tolerance: self.selected_peak_for_ui_focus = closest_peak_idx
                        else: self.selected_peak_for_ui_focus = -1
                        self.update_plot()
                else:
                    if event.button == 1 and event.xdata is not None and self.peaks.any():
                        clicked_x = int(round(event.xdata)); distances = np.abs(self.peaks - clicked_x); closest_peak_idx = np.argmin(distances)
                        click_tolerance_delete = max(5, self.peak_distance / 4.0)
                        if distances[closest_peak_idx] <= click_tolerance_delete:
                            self.selected_peak_index_for_delete = self.peaks[closest_peak_idx]
                            self.delete_selected_peak_button.setEnabled(True)
                        else:
                            self.selected_peak_index_for_delete = -1
                            self.delete_selected_peak_button.setEnabled(False)
                        self.update_plot()
            def _find_intersection_boundaries(self, profile, baseline, peak_x, search_start, search_end):
                diff = profile - baseline
                if not np.any(diff): return search_start, search_end
                above_indices = np.where(diff > 0)[0]
                if len(above_indices) == 0: return peak_x, peak_x
                sign_changes = np.where(np.diff(np.sign(diff)))[0]
                left_intersections = sign_changes[sign_changes < peak_x]
                start_calc = np.max(left_intersections) + 1 if left_intersections.size > 0 else np.min(above_indices)
                right_intersections = sign_changes[sign_changes > peak_x]
                end_calc = np.min(right_intersections) if right_intersections.size > 0 else np.max(above_indices)
                start_calc = max(search_start, start_calc)
                end_calc = min(search_end, end_calc)
                if start_calc >= end_calc: return search_start, search_end
                return int(start_calc), int(end_calc)
            def apply_peak_broadening(self, broadening_value):
                self.peak_broadening_pixels = broadening_value
                self.peak_regions = [] 
                profile_len = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                if profile_len == 0: return
                for i in range(len(self.initial_valley_regions)):
                    initial_start, initial_end = self.initial_valley_regions[i]
                    new_start = initial_start - self.peak_broadening_pixels
                    new_end = initial_end + self.peak_broadening_pixels
                    new_start_clamped = max(0, new_start)
                    new_end_clamped = min(profile_len - 1, new_end)
                    if new_start_clamped > new_end_clamped:
                        mid_valley = (initial_start + initial_end) // 2
                        new_start_clamped, new_end_clamped = mid_valley, mid_valley
                    self.peak_regions.append((new_start_clamped, new_end_clamped))
            def apply_peak_broadening_on_release(self):
                self.apply_peak_broadening(self.broadening_slider.value())
                self.update_plot()
            def get_final_peak_info(self):
                peak_info_list = []
                num_valid_peaks = len(self.peak_regions)
                current_area_list = []
                if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
                elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
                elif self.method == "Valley-to-Valley": current_area_list = self.peak_areas_valley
                num_peaks_to_process = min(num_valid_peaks, len(self.peaks), len(current_area_list))
                for i in range(num_peaks_to_process):
                    try:
                        original_peak_x_in_profile = int(self.peaks[i])
                        peak_info_list.append({'area': current_area_list[i],'y_coord_in_lane_image': original_peak_x_in_profile,'original_peak_index': original_peak_x_in_profile})
                    except IndexError: peak_info_list.append({'area': 0.0, 'y_coord_in_lane_image': 0, 'original_peak_index': -1})
                return peak_info_list
            def toggle_manual_select_mode(self, checked):
                self.manual_select_mode_active = checked
                if checked:
                    if self.add_peak_mode_active: self.add_peak_manually_button.setChecked(False); self.toggle_add_peak_mode(False) 
                    QMessageBox.information(self, "Identify Peak", "Click a peak in the profile plot to focus its handles.")
                else: self.selected_peak_for_ui_focus = -1; self.update_plot()
            def toggle_add_peak_mode(self, checked):
                self.add_peak_mode_active = checked
                if checked:
                    if self.manual_select_mode_active: self.identify_peak_button.setChecked(False); self.toggle_manual_select_mode(False)
                    self.selected_peak_index_for_delete = -1; self.delete_selected_peak_button.setEnabled(False)
                    self.update_plot(); QMessageBox.information(self, "Add Peak", "Click on the profile plot to add a peak.")
            def add_manual_peak(self, x_coord):
                if self.profile_original_inverted is None or x_coord in self.peaks: return
                self.peaks = np.array(sorted(self.peaks.tolist() + [x_coord]))
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self._recalculate_all_regions()
                self.update_plot()
            def _redefine_all_valley_regions(self):
                self.initial_valley_regions = []
                profile_to_analyze = self.profile_original_inverted
                if profile_to_analyze is None or len(profile_to_analyze) <= 1 or len(self.peaks) == 0:
                    self.peak_regions = []; return
                midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2 if len(self.peaks) > 1 else []
                search_boundaries_left = np.concatenate(([0], midpoints))
                search_boundaries_right = np.concatenate((midpoints, [len(profile_to_analyze) - 1]))
                for i, peak_idx in enumerate(self.peaks):
                    left_bound = int(search_boundaries_left[i]); right_bound = int(search_boundaries_right[i])
                    start, end = self._find_outward_troughs(profile_to_analyze, peak_idx, left_bound, right_bound)
                    self.initial_valley_regions.append((start, end))
                self.apply_peak_broadening(self.broadening_slider.value())
            def delete_selected_peak_action(self):
                if self.selected_peak_index_for_delete == -1 or self.selected_peak_index_for_delete not in self.peaks: return
                self.peaks = np.array([p for p in self.peaks if p != self.selected_peak_index_for_delete])
                self.selected_peak_index_for_delete = -1; self.delete_selected_peak_button.setEnabled(False)
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self._recalculate_all_regions()
                self.update_plot()
            def copy_peak_regions_to_app(self):
                if not self.parent_app: return
                if not self.peak_regions: QMessageBox.information(self, "No Regions", "No peak regions to copy."); return
                self.parent_app.copied_peak_regions_data["regions"] = [tuple(r) for r in self.peak_regions]
                self.parent_app.copied_peak_regions_data["profile_length"] = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                self.parent_app.copied_peak_regions_data["peaks"] = self.peaks.tolist()
                QMessageBox.information(self, "Regions Copied", f"{len(self.peak_regions)} regions copied.")
                if hasattr(self, 'paste_regions_button'): self.paste_regions_button.setEnabled(True)
            def paste_peak_regions_from_app(self):
                if not self.parent_app or not self.parent_app.copied_peak_regions_data.get("regions"): QMessageBox.information(self, "No Regions to Paste", "No peak regions have been copied yet."); return
                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0: QMessageBox.warning(self, "Error", "Current profile not available for pasting regions."); return
                copied_data = self.parent_app.copied_peak_regions_data; regions_to_paste = copied_data["regions"]; original_profile_len = copied_data["profile_length"]; copied_peaks_indices = np.array(copied_data.get("peaks", []))
                current_profile_len = len(self.profile_original_inverted); self.peak_regions = []
                scale_factor = 1.0
                if original_profile_len > 0 and current_profile_len > 0 and original_profile_len != current_profile_len: scale_factor = float(current_profile_len) / original_profile_len
                temp_derived_peaks = []
                if len(copied_peaks_indices) > 0 and len(copied_peaks_indices) == len(regions_to_paste): self.peaks = np.clip((copied_peaks_indices * scale_factor).round().astype(int), 0, current_profile_len - 1)
                for i, (start_orig, end_orig) in enumerate(regions_to_paste):
                    start_scaled = int(round(start_orig * scale_factor)); end_scaled = int(round(end_orig * scale_factor))
                    start_clamped = max(0, min(start_scaled, current_profile_len - 1)); end_clamped = max(0, min(end_scaled, current_profile_len - 1))
                    if start_clamped >= end_clamped: mid = (start_clamped + end_clamped) // 2; start_clamped, end_clamped = max(0, mid-1), min(current_profile_len-1, mid+1)
                    self.peak_regions.append((start_clamped, end_clamped))
                    if not (len(copied_peaks_indices) > 0 and len(copied_peaks_indices) == len(regions_to_paste)): temp_derived_peaks.append((start_clamped + end_clamped) // 2)
                if temp_derived_peaks: self.peaks = np.array(sorted(temp_derived_peaks))
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self.update_plot(); QMessageBox.information(self, "Regions Pasted", f"{len(self.peak_regions)} regions applied.")
            def get_current_settings(self): return self._final_settings
            def should_persist_settings(self): return self._persist_enabled_on_exit
            def get_final_peak_area(self): return [info['area'] for info in self.get_final_peak_info()]
            def regenerate_profile_and_detect(self):
                if gaussian_filter1d is None or cv2 is None or ImageOps is None: return
                self.band_estimation_method = self.band_estimation_combobox.currentText(); self.area_subtraction_method = self.method_combobox.currentText()
                self.smoothing_sigma = self.smoothing_slider.value() / 10.0; self.denoise_sigma = self.denoise_sigma_slider.value() / 10.0
                base_img = self.original_pil_cropped_data.copy()
                if self.denoise_sigma > 0.01:
                    try: base_img = Image.fromarray(cv2.GaussianBlur(np.array(base_img), (0,0), self.denoise_sigma).astype(np.array(base_img).dtype), mode=base_img.mode)
                    except: pass
                if base_img.mode.startswith('I') or base_img.mode == 'F': self.enhanced_cropped_image_for_display = Image.fromarray((np.clip((np.array(base_img, dtype=np.float32) - np.percentile(base_img, 2)) / (np.percentile(base_img, 98) - np.percentile(base_img, 2) + 1e-9), 0.0, 1.0) * 255).astype(np.uint8), mode='L')
                else: self.enhanced_cropped_image_for_display = ImageOps.autocontrast(base_img.convert('L'))
                if self.band_estimation_method == "Mean": profile_temp = np.mean(self.intensity_array_original_range, axis=1)
                else: profile_temp = np.percentile(self.intensity_array_original_range, int(self.band_estimation_method.split(":")[1].replace('%', '')), axis=1)
                profile_inv_raw = self.original_max_value - profile_temp.astype(np.float64); profile_inv_raw -= np.min(profile_inv_raw)
                profile_to_process = profile_inv_raw.copy()
                if self.denoise_sigma > 0.01: profile_to_process = gaussian_filter1d(profile_to_process, sigma=self.denoise_sigma)
                self.profile_original_inverted = profile_to_process
                if self.smoothing_sigma > 0.1: self.profile_original_inverted = gaussian_filter1d(self.profile_original_inverted, sigma=self.smoothing_sigma)
                prof_min, prof_max = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
                if prof_max > prof_min + 1e-6: self.profile = (self.profile_original_inverted - prof_min) / (prof_max - prof_min) * 255.0
                else: self.profile = np.zeros_like(self.profile_original_inverted)
                self.detect_peaks()
            def _find_outward_troughs(self, profile, peak_idx, left_bound, right_bound):
                profile_len = len(profile)
                if not (0 <= left_bound <= peak_idx <= right_bound < profile_len):
                    w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                    return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)
                valley_left_idx = peak_idx
                for idx in range(peak_idx - 1, left_bound - 1, -1):
                    if profile[idx] > profile[idx + 1]: valley_left_idx = idx + 1; break
                    valley_left_idx = idx
                else: valley_left_idx = left_bound
                valley_right_idx = peak_idx
                for idx in range(peak_idx + 1, right_bound + 1):
                    if profile[idx] > profile[idx - 1]: valley_right_idx = idx - 1; break
                    valley_right_idx = idx
                else: valley_right_idx = right_bound
                if valley_left_idx >= valley_right_idx:
                     w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                     return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)
                return valley_left_idx, valley_right_idx
            def manual_peak_number_update(self):
                if self.profile_original_inverted is None: return
                try:
                    num_peaks = int(self.peak_number_input.text())
                    if num_peaks == len(self.peaks): return
                    if num_peaks < len(self.peaks): self.peaks = self.peaks[:num_peaks]
                    else:
                        num_to_add = num_peaks - len(self.peaks)
                        new_peaks = np.linspace(0, len(self.profile)-1, num_peaks + 2)[1:-1].astype(int)
                        self.peaks = np.sort(np.unique(np.concatenate((self.peaks, new_peaks))))[:num_peaks]
                    self._recalculate_all_regions()
                    self.update_plot()
                except ValueError: self.peak_number_input.setText(str(len(self.peaks)))
            def _custom_rolling_ball(self, profile, radius):
                if grey_opening is None or profile is None or profile.ndim != 1 or profile.size == 0 or radius <= 0: return np.zeros_like(profile) if profile is not None else np.array([])
                structure_size = int(max(1, 2 * radius + 1));
                if structure_size > profile.shape[0]: structure_size = profile.shape[0]
                try: return grey_opening(profile, size=structure_size, mode='reflect')
                except Exception: return np.zeros_like(profile)
            

        class LiveViewLabel(QLabel):
            mouseMovedInLabel = Signal(QPointF, QPointF, bool)
            CORNER_HANDLE_BASE_RADIUS = 6.0
            def __init__(self, font_type, font_size, marker_color, app_instance, parent=None): # Added app_instance
                super().__init__(parent)
                self.app_instance = app_instance # Store the CombinedSDSApp instance
                self.setMouseTracking(True)
                self.preview_marker_enabled = False
                self.preview_marker_text = ""
                self.preview_marker_position = None
                self.marker_font_type = font_type
                self.marker_font_size = font_size
                self.marker_color = marker_color
                self.setFocusPolicy(Qt.StrongFocus)
                self.mw_predict_preview_enabled = False
                self.mw_predict_preview_position = None # QPointF in unzoomed label space
                self.bounding_box_preview = [] # Used for single rect, multi-lane rect, auto-lane rect previews
                self.measure_quantity_mode = False # App-level flag, LiveViewLabel checks it
                self.counter = 0
                self.zoom_level = 1.0
                self.pan_offset = QPointF(0, 0)
                self.is_panning = False
                self.pan_start_view_coords = None # Stores QPoint of mouse press in widget coords
                self.pan_offset_at_drag_start = None # Stores QPointF of pan_offset at drag start
                
                # self.quad_points is used for:
                # 1. Defining a NEW single analysis quad (when self.mode == "quad")
                # 2. Defining a NEW auto-lane quad (when self.mode == "auto_lane_quad")
                # 3. Displaying a FINALIZED single quad (when self.mode is None, and it's the active single shape)
                self.quad_points = [] 
                self.selected_point = -1 # For interactive quad point definition
                self.drag_threshold = 10 # Click radius for selecting quad points
                self.bounding_box_complete = False # For single quad definition completion
                
                self.mode=None # Tracks LiveViewLabel's current interaction mode (e.g., "quad", "rectangle", "auto_lane_quad", etc.)
                
                self.rectangle_start = None # Used for single analysis rect, auto-lane rect, multi-lane rect start point
                self.rectangle_end = None   # Used for the corresponding end point
                self.rectangle_points = []  # Stores [start, end] for single analysis rect after finalization
                
                self.drag_start_pos = None # Potentially for custom shape dragging if needed
                self.draw_edges=True # For selection highlights
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None

                # Custom hooks for CombinedSDSApp
                self._custom_left_click_handler_from_app = None
                self._custom_mouseMoveEvent_from_app = None
                self._custom_mouseReleaseEvent_from_app = None
                
            def clear_crop_preview(self):
                """Clears any existing crop rectangle preview."""
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None
                self.update() # Trigger repaint

            def start_crop_preview(self, start_point_view):
                """Initiates drawing the crop rectangle preview."""
                self.drawing_crop_rect = True
                self.crop_rect_start_view = start_point_view
                self.crop_rect_end_view = start_point_view # Start and end are same initially
                self.crop_rect_final_view = None # Clear any finalized rect
                self.update()

            def update_crop_preview(self, start_point_view, current_point_view):
                """Updates the crop rectangle preview while dragging."""
                if self.drawing_crop_rect:
                    # Keep start point fixed, update end point
                    self.crop_rect_start_view = start_point_view
                    self.crop_rect_end_view = current_point_view
                    self.update()

            def finalize_crop_preview(self, start_point_view, end_point_view):
                """Stores the final rectangle dimensions for persistent preview."""
                self.drawing_crop_rect = False
                # Store as a normalized QRectF for easier drawing
                self.crop_rect_final_view = QRectF(start_point_view, end_point_view).normalized()
                # Clear the temporary points used during drawing
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.update()
                
            def wheelEvent(self, event: 'QWheelEvent'): # Add type hint for clarity
                if not self.app_instance or not self.app_instance.image: # Ensure app and image exist
                    super().wheelEvent(event)
                    return
            
                # --- Determine Zoom Factor ---
                # event.angleDelta().y() is typically +/- 120 for one notch on a mouse wheel
                num_degrees = event.angleDelta().y() / 8
                num_steps = num_degrees / 15.0 # Standard step factor
            
                zoom_factor_increment = 1.1
                zoom_factor_decrement = 1 / 1.1
            
                # --- Calculate new zoom level ---
                old_zoom_level = self.zoom_level
                if num_steps > 0: # Zooming in
                    self.zoom_level *= (zoom_factor_increment ** abs(num_steps))
                elif num_steps < 0: # Zooming out
                    self.zoom_level *= (zoom_factor_decrement ** abs(num_steps))
            
                # --- Clamp Zoom Level (optional, but good practice) ---
                min_zoom = 0.1 # Example minimum zoom
                max_zoom = 20.0  # Example maximum zoom
                self.zoom_level = max(min_zoom, min(self.zoom_level, max_zoom))
            
                # --- Zoom Towards Mouse Cursor ---
                mouse_point_widget = event.position() 
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_point_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_point_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_point_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_point_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if self.zoom_level <= 1.0001: # Use a small tolerance for float comparison
                    self.zoom_level = 1.0
                    self.pan_offset = QPointF(0, 0)
            
                if not self.is_panning: # is_panning check is important here
                    if self.zoom_level > 1.001: # Use a small tolerance for float comparison
                        self.setCursor(Qt.OpenHandCursor)
                    else:
                        self.setCursor(Qt.ArrowCursor)
                
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()
                event.accept()
                
            def mouseMoveEvent(self, event: 'QMouseEvent'):
                current_mouse_pos_widget = event.position() 
                untransformed_label_pos = self.transform_point(current_mouse_pos_widget)
                image_coords_actual = None
                is_image_coord_valid = False

                if self.app_instance and self.app_instance.image and not self.app_instance.image.isNull():
                    img_w_orig = float(self.app_instance.image.width())
                    img_h_orig = float(self.app_instance.image.height())
                    label_w_widget = float(self.width())
                    label_h_widget = float(self.height())

                    if img_w_orig > 0 and img_h_orig > 0 and label_w_widget > 0 and label_h_widget > 0:
                        scale_factor_img_to_label = min(label_w_widget / img_w_orig, label_h_widget / img_h_orig)
                        if scale_factor_img_to_label > 1e-9: # Ensure scale_factor is not zero
                            displayed_img_w_in_label = img_w_orig * scale_factor_img_to_label
                            displayed_img_h_in_label = img_h_orig * scale_factor_img_to_label
                            offset_x_img_in_label = (label_w_widget - displayed_img_w_in_label) / 2.0
                            offset_y_img_in_label = (label_h_widget - displayed_img_h_in_label) / 2.0

                            if untransformed_label_pos.x() >= offset_x_img_in_label and \
                               untransformed_label_pos.x() <= offset_x_img_in_label + displayed_img_w_in_label and \
                               untransformed_label_pos.y() >= offset_y_img_in_label and \
                               untransformed_label_pos.y() <= offset_y_img_in_label + displayed_img_h_in_label:
                                
                                relative_x_in_display = untransformed_label_pos.x() - offset_x_img_in_label
                                relative_y_in_display = untransformed_label_pos.y() - offset_y_img_in_label
                                
                                img_x = relative_x_in_display / scale_factor_img_to_label
                                img_y = relative_y_in_display / scale_factor_img_to_label
                                image_coords_actual = QPointF(img_x, img_y)
                                is_image_coord_valid = True
                
                self.mouseMovedInLabel.emit(untransformed_label_pos, image_coords_actual if image_coords_actual else QPointF(), is_image_coord_valid)

                if self.is_panning and (event.buttons() & Qt.RightButton):
                    if self.pan_start_view_coords:
                        delta = event.position() - self.pan_start_view_coords
                        self.pan_offset = self.pan_offset_at_drag_start + delta
                        if self.app_instance: self.app_instance.update_live_view()
                        # QApplication.processEvents() # Usually not needed
                    event.accept()
                    return
            
                if hasattr(self, '_custom_mouseMoveEvent_from_app') and self._custom_mouseMoveEvent_from_app:
                    self._custom_mouseMoveEvent_from_app(event)
                    if event.isAccepted():
                        return

                if self.preview_marker_enabled: 
                    snapped_label_pos = untransformed_label_pos
                    if self.app_instance:
                        snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                    self.preview_marker_position = snapped_label_pos
                    self.update()
                    event.accept()
                    return
                
                if self.mw_predict_preview_enabled: 
                    if self.app_instance and hasattr(self.app_instance, 'update_mw_predict_preview'):
                         self.app_instance.update_mw_predict_preview(event)
                    else:
                        snapped_label_pos = untransformed_label_pos
                        if self.app_instance: snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                        self.mw_predict_preview_position = snapped_label_pos
                        self.update()
                    event.accept()
                    return
        
                if self.selected_point != -1 and self.measure_quantity_mode and self.mode=="quad":
                    snapped_mouse_label_space = untransformed_label_pos
                    if self.app_instance:
                        snapped_mouse_label_space = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                    self.quad_points[self.selected_point] = snapped_mouse_label_space
                    self.update()
                    event.accept()
                    return
                
                if not event.isAccepted():
                    super().mouseMoveEvent(event)
                
            def mousePressEvent(self, event: 'QMouseEvent'):
                if event.button() == Qt.RightButton and self.zoom_level > 1.001:
                    self.is_panning = True
                    self.pan_start_view_coords = event.position()
                    self.pan_offset_at_drag_start = QPointF(self.pan_offset)
                    self.setCursor(Qt.ClosedHandCursor)
                    event.accept() 
                    return
                
                if event.button() != Qt.RightButton and hasattr(self, '_custom_left_click_handler_from_app') and self._custom_left_click_handler_from_app:
                    self._custom_left_click_handler_from_app(event)
                    if event.isAccepted():
                        return
                
                if self.preview_marker_enabled and event.button() == Qt.LeftButton:
                    if self.app_instance and hasattr(self.app_instance, 'place_custom_marker'):
                        self.app_instance.place_custom_marker(event, self.preview_marker_text)
                    self.update()
                    event.accept()
                    return
        
                if self.measure_quantity_mode and self.mode == "quad" and event.button() == Qt.LeftButton:
                    clicked_label_point_transformed = self.transform_point(event.position())
                    snapped_click_point = clicked_label_point_transformed
                    if self.app_instance:
                        snapped_click_point = self.app_instance.snap_point_to_grid(clicked_label_point_transformed)
                    
                    point_interacted_with = False
                    for i, p in enumerate(self.quad_points):
                        if (snapped_click_point - p).manhattanLength() < self.drag_threshold:
                            self.selected_point = i
                            point_interacted_with = True
                            break
                    if not point_interacted_with and len(self.quad_points) < 4:
                        self.quad_points.append(snapped_click_point)
                        self.selected_point = len(self.quad_points) - 1
                        point_interacted_with = True
                    
                    if point_interacted_with:
                        if len(self.quad_points) == 4 and not self.bounding_box_complete: # For single quad
                             self.bounding_box_complete = True
                        self.update()
                        event.accept()
                        return
                
                if not event.isAccepted():
                    super().mousePressEvent(event)

            def mouseReleaseEvent(self, event: 'QMouseEvent'):
                if event.button() == Qt.RightButton and self.is_panning:
                    self.is_panning = False
                    self.pan_start_view_coords = None
                    self.pan_offset_at_drag_start = None
                    if self.zoom_level > 1.001: self.setCursor(Qt.OpenHandCursor)
                    else: self.setCursor(Qt.ArrowCursor)
                    event.accept()
                    return
            
                if event.button() != Qt.RightButton and hasattr(self, '_custom_mouseReleaseEvent_from_app') and self._custom_mouseReleaseEvent_from_app:
                    self._custom_mouseReleaseEvent_from_app(event)
                    if event.isAccepted():
                        return
                
                if self.mode == "quad" and event.button() == Qt.LeftButton: # For single quad definition
                    self.selected_point = -1 # Deselect point after release
                    self.update()
                    event.accept()
                    return 
                
                if not event.isAccepted():
                    super().mouseReleaseEvent(event)

            def transform_point(self, point):
                """Transform a point from widget coordinates to image coordinates."""
                if self.zoom_level != 1.0:
                    return QPointF(
                        (point.x() - self.pan_offset.x()) / self.zoom_level,
                        (point.y() - self.pan_offset.y()) / self.zoom_level
                    )
                return QPointF(point)
            
            
            def zoom_in(self): # Zooms towards view center by default
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level *= 1.1
                self.zoom_level = min(self.zoom_level, 20.0) # Max zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if not self.is_panning: self.setCursor(Qt.ArrowCursor)
                
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()

            def zoom_out(self): # Zooms towards view center
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level /= 1.1
                self.zoom_level = max(self.zoom_level, 0.1) # Min zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if self.zoom_level <= 1.0:
                    self.zoom_level = 1.0
                    self.pan_offset = QPointF(0, 0)
                    if not self.is_panning: self.setCursor(Qt.ArrowCursor)
                else:
                    if not self.is_panning: self.setCursor(Qt.ArrowCursor)
            
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()
                

            def paintEvent(self, event):
                super().paintEvent(event)
                painter = QPainter(self)
                # painter.setRenderHint(QPainter.Antialiasing, True)
                painter.setRenderHint(QPainter.TextAntialiasing, True)
                # painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
                
                painter.save()
                if self.zoom_level != 1.0:
                    painter.translate(self.pan_offset)
                    painter.scale(self.zoom_level, self.zoom_level)

                # --- Coordinate mapping helpers ---
                _image_to_label_space_valid = False
                _scale_factor_img_to_label = 1.0
                _img_w_orig_from_app = 1.0
                _img_h_orig_from_app = 1.0
                _offset_x_img_in_label = 0.0
                _offset_y_img_in_label = 0.0
                if self.app_instance and self.app_instance.image and not self.app_instance.image.isNull():
                    current_app_image = self.app_instance.image
                    label_w_widget = float(self.width())
                    label_h_widget = float(self.height())
                    _img_w_orig_from_app = float(current_app_image.width())
                    _img_h_orig_from_app = float(current_app_image.height())
                    if _img_w_orig_from_app > 0 and _img_h_orig_from_app > 0 and label_w_widget > 0 and label_h_widget > 0:
                        _scale_factor_img_to_label = min(label_w_widget / _img_w_orig_from_app, label_h_widget / _img_h_orig_from_app)
                        if _scale_factor_img_to_label > 1e-9: # Ensure scale factor is not zero
                            _displayed_img_w_in_label = _img_w_orig_from_app * _scale_factor_img_to_label
                            _displayed_img_h_in_label = _img_h_orig_from_app * _scale_factor_img_to_label
                            _offset_x_img_in_label = (label_w_widget - _displayed_img_w_in_label) / 2.0
                            _offset_y_img_in_label = (label_h_widget - _displayed_img_h_in_label) / 2.0
                            _image_to_label_space_valid = True
                
                def _app_image_coords_to_unzoomed_label_space(img_coords_tuple_or_qpointf):
                    if not _image_to_label_space_valid:
                        if isinstance(img_coords_tuple_or_qpointf, QPointF): return img_coords_tuple_or_qpointf
                        return QPointF(img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])
                    img_x, img_y = (img_coords_tuple_or_qpointf.x(), img_coords_tuple_or_qpointf.y()) if isinstance(img_coords_tuple_or_qpointf, QPointF) else (img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])
                    x_ls = _offset_x_img_in_label + img_x * _scale_factor_img_to_label; y_ls = _offset_y_img_in_label + img_y * _scale_factor_img_to_label
                    return QPointF(x_ls, y_ls)
                
                is_defining_single_quad_on_label = (self.mode == "quad") 
                is_defining_single_rect_on_label = (self.mode == "rectangle")

                is_selected_single_quad_for_move = (self.app_instance and
                                           self.app_instance.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"] and
                                           self.app_instance.moving_multi_lane_index == -2)
                is_selected_single_rect_for_move = (self.app_instance and
                                           self.app_instance.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"] and
                                           self.app_instance.moving_multi_lane_index == -3)

                # --- Draw standard markers, custom markers, custom shapes (as before) ---
                if _image_to_label_space_valid and self.app_instance: # Standard Markers
                    std_marker_font = QFont(self.app_instance.font_family, self.app_instance.font_size); std_marker_color = self.app_instance.font_color if hasattr(self.app_instance, 'font_color') else QColor(Qt.black)
                    painter.setFont(std_marker_font); painter.setPen(std_marker_color); font_metrics_std = QFontMetrics(std_marker_font)
                    text_height_std_label_space = font_metrics_std.height(); y_offset_text_baseline_std = text_height_std_label_space * 0.3
                    if hasattr(self.app_instance, 'left_markers'):
                        left_marker_offset_x_label_space = self.app_instance.left_marker_shift_added * _scale_factor_img_to_label
                        for y_pos_img, marker_text_val in self.app_instance.left_markers:
                            anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y(); text_to_draw = f"{marker_text_val} ⎯"; text_width_label_space = font_metrics_std.horizontalAdvance(text_to_draw)
                            draw_x_ls = _offset_x_img_in_label + left_marker_offset_x_label_space - text_width_label_space; draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                            painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    if hasattr(self.app_instance, 'right_markers'):
                            right_marker_start_x_label_space = _offset_x_img_in_label + (self.app_instance.right_marker_shift_added * _scale_factor_img_to_label)
                            for y_pos_img, marker_text_val in self.app_instance.right_markers:
                                anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y(); text_to_draw = f"⎯ {marker_text_val}"
                                draw_x_ls = right_marker_start_x_label_space; draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                                painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    if hasattr(self.app_instance, 'top_markers'):
                        top_marker_offset_y_label = self.app_instance.top_marker_shift_added * _scale_factor_img_to_label; rotation_angle = self.app_instance.font_rotation
                        for x_pos_img, marker_text_val in self.app_instance.top_markers:
                            anchor_x_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, 0)).x(); text_to_draw = str(marker_text_val); painter.save()
                            draw_baseline_y_ls = _offset_y_img_in_label + top_marker_offset_y_label + y_offset_text_baseline_std
                            painter.translate(anchor_x_label_space, draw_baseline_y_ls); painter.rotate(rotation_angle); painter.drawText(QPointF(0, 0), text_to_draw); painter.restore()
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_shapes'): # Custom Shapes
                    for shape_data in self.app_instance.custom_shapes:
                        try:
                            shape_type = shape_data.get('type'); color = QColor(shape_data.get('color', '#000000')); base_thickness = float(shape_data.get('thickness', 0.5)); thickness_on_label = base_thickness * _scale_factor_img_to_label
                            effective_pen_width = max(0.5, thickness_on_label / self.zoom_level if self.zoom_level > 0 else thickness_on_label); pen = QPen(color); pen.setWidthF(effective_pen_width); painter.setPen(pen)
                            if shape_type == 'line':
                                start_img = shape_data.get('start'); end_img = shape_data.get('end')
                                if start_img and end_img: start_label_space = _app_image_coords_to_unzoomed_label_space(start_img); end_label_space = _app_image_coords_to_unzoomed_label_space(end_img); painter.drawLine(start_label_space, end_label_space)
                            elif shape_type == 'rectangle':
                                rect_img = shape_data.get('rect')
                                if rect_img: x_img, y_img, w_img, h_img = rect_img; top_left_label_space = _app_image_coords_to_unzoomed_label_space((x_img, y_img)); w_label_space = w_img * _scale_factor_img_to_label; h_label_space = h_img * _scale_factor_img_to_label; painter.drawRect(QRectF(top_left_label_space, QSizeF(w_label_space, h_label_space)))
                        except Exception as e: print(f"Error drawing custom shape in LiveViewLabel.paintEvent: {shape_data}, {e}")
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_markers'): # Custom Markers
                    for marker_data_list in self.app_instance.custom_markers:
                        try:
                            x_pos_img, y_pos_img, marker_text_str, qcolor_obj, font_family_str, font_size_int, is_bold, is_italic = marker_data_list
                            anchor_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, y_pos_img)); current_marker_font = QFont(font_family_str, font_size_int); current_marker_font.setBold(is_bold); current_marker_font.setItalic(is_italic); painter.setFont(current_marker_font)
                            if not isinstance(qcolor_obj, QColor): qcolor_obj = QColor(qcolor_obj)
                            if not qcolor_obj.isValid(): qcolor_obj = Qt.black
                            painter.setPen(qcolor_obj); font_metrics_marker = QFontMetrics(current_marker_font); text_bounding_rect_marker = font_metrics_marker.boundingRect(marker_text_str)
                            draw_x_marker = anchor_label_space.x() - (text_bounding_rect_marker.left() + text_bounding_rect_marker.width() / 2.0); draw_y_marker = anchor_label_space.y() - (text_bounding_rect_marker.top() + text_bounding_rect_marker.height() / 2.0)
                            painter.drawText(QPointF(draw_x_marker, draw_y_marker), marker_text_str)
                        except Exception as e: print(f"Error drawing app_instance custom marker: {marker_data_list}, {e}")
                if self.preview_marker_enabled and self.preview_marker_position: # Preview Marker
                    painter.setOpacity(0.7); marker_preview_font = QFont(self.marker_font_type, self.marker_font_size); painter.setFont(marker_preview_font); painter.setPen(self.marker_color); font_metrics_preview = QFontMetrics(marker_preview_font)
                    preview_text_rect = font_metrics_preview.boundingRect(self.preview_marker_text); draw_x_preview = self.preview_marker_position.x() - (preview_text_rect.left() + preview_text_rect.width() / 2.0); draw_y_preview = self.preview_marker_position.y() - (preview_text_rect.top() + preview_text_rect.height() / 2.0)
                    painter.drawText(QPointF(draw_x_preview, draw_y_preview), self.preview_marker_text); painter.setOpacity(1.0)
                if self.app_instance and hasattr(self, 'moving_custom_item_info') and self.app_instance.moving_custom_item_info:
                    info = self.app_instance.moving_custom_item_info
                    is_resizing = self.app_instance.current_selection_mode == "resizing_custom_item"
                    selection_pen = QPen(Qt.magenta, max(0.5, 2.0 / self.zoom_level), Qt.DotLine)
                    handle_pen = QPen(Qt.red, max(0.5, 2.0 / self.zoom_level))
                    handle_brush = QBrush(Qt.red)
                    handle_radius = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS
                    if info['type'] == 'marker':
                        bbox_ls = self.app_instance._get_marker_bounding_box_in_label_space(info['index'])
                        if bbox_ls:
                            painter.setPen(selection_pen)
                            painter.drawRect(bbox_ls)
                    elif info['type'] == 'shape':
                        body_ls, handles_ls = self.app_instance._get_shape_bounding_box_and_handles_in_label_space(info['index'])
                        if body_ls and handles_ls:
                            painter.setPen(selection_pen)
                            painter.drawRect(body_ls.adjusted(-2,-2,2,2)) # Draw bounding box around shape
                            painter.setPen(handle_pen)
                            painter.setBrush(handle_brush)
                            for idx, handle_pt in enumerate(handles_ls):
                                if is_resizing and idx == self.app_instance.resizing_corner_index:
                                    painter.drawEllipse(handle_pt, handle_radius * 1.2, handle_radius * 1.2)
                                else:
                                    painter.drawEllipse(handle_pt, handle_radius, handle_radius)
                            painter.setBrush(Qt.NoBrush)

                # --- Draw Crop Rect, Shape Preview, MW Preview, Placed MW Marker (as before) ---
                preview_pen_crop = QPen(Qt.magenta); effective_pen_width_crop = max(0.5, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5); preview_pen_crop.setWidthF(effective_pen_width_crop)
                if self.drawing_crop_rect and self.crop_rect_start_view and self.crop_rect_end_view: preview_pen_crop.setStyle(Qt.DashLine); painter.setPen(preview_pen_crop); rect_to_draw = QRectF(self.crop_rect_start_view, self.crop_rect_end_view).normalized(); painter.drawRect(rect_to_draw)
                elif self.crop_rect_final_view: preview_pen_crop.setStyle(Qt.SolidLine); painter.setPen(preview_pen_crop); painter.drawRect(self.crop_rect_final_view)
                if self.app_instance and self.app_instance.drawing_mode in ['line', 'rectangle'] and self.app_instance.current_drawing_shape_preview: # Shape Preview
                    try:
                        start_pt_ls = self.app_instance.current_drawing_shape_preview['start']; end_pt_ls = self.app_instance.current_drawing_shape_preview['end']; preview_color = self.app_instance.custom_marker_color
                        base_preview_thickness = float(self.app_instance.custom_font_size_spinbox.value()); effective_preview_thickness = min(1.0, base_preview_thickness / self.zoom_level if self.zoom_level > 0 else base_preview_thickness)
                        preview_pen_shape = QPen(preview_color); preview_pen_shape.setWidthF(effective_preview_thickness); preview_pen_shape.setStyle(Qt.DotLine); painter.setPen(preview_pen_shape)
                        if self.app_instance.drawing_mode == 'line': painter.drawLine(start_pt_ls, end_pt_ls)
                        elif self.app_instance.drawing_mode == 'rectangle': painter.drawRect(QRectF(start_pt_ls, end_pt_ls).normalized())
                    except Exception as e: print(f"Error drawing live shape preview in paintEvent: {e}")
                if self.mw_predict_preview_enabled and self.mw_predict_preview_position: # MW Preview
                    try:
                        painter.setOpacity(0.7); mw_preview_font_size = self.app_instance.custom_font_size_spinbox.value() if self.app_instance and hasattr(self.app_instance, 'custom_font_size_spinbox') else 12
                        mw_preview_font_family = self.app_instance.custom_font_type_dropdown.currentText() if self.app_instance and hasattr(self.app_instance, 'custom_font_type_dropdown') else "Arial"
                        mw_preview_font = QFont(mw_preview_font_family, mw_preview_font_size); painter.setFont(mw_preview_font); painter.setPen(Qt.darkGreen); font_metrics_mw_preview = QFontMetrics(mw_preview_font)
                        text_mw_preview = "⎯⎯"; text_bounding_rect_mw_preview = font_metrics_mw_preview.boundingRect(text_mw_preview); preview_anchor_x_ls = self.mw_predict_preview_position.x(); preview_anchor_y_ls = self.mw_predict_preview_position.y()
                        draw_x_mw_preview_ls = preview_anchor_x_ls - (text_bounding_rect_mw_preview.left() + text_bounding_rect_mw_preview.width() / 2.0); draw_y_mw_preview_ls = preview_anchor_y_ls - (text_bounding_rect_mw_preview.top() + text_bounding_rect_mw_preview.height() / 2.0)
                        painter.drawText(QPointF(draw_x_mw_preview_ls, draw_y_mw_preview_ls), text_mw_preview); painter.setOpacity(1.0)
                    except Exception as e: print(f"Error drawing MW prediction *preview* marker in paintEvent: {e}")
                if self.app_instance and hasattr(self.app_instance, "protein_location") and self.app_instance.protein_location and not self.app_instance.run_predict_MW: # Placed MW Marker
                    try:
                        loc_x_ls, loc_y_ls = self.app_instance.protein_location; mw_marker_font = QFont(self.app_instance.custom_font_type_dropdown.currentText(), self.app_instance.custom_font_size_spinbox.value() + 2)
                        painter.setFont(mw_marker_font); painter.setPen(Qt.green); font_metrics_mw = QFontMetrics(mw_marker_font); text_mw = "⎯⎯"; text_bounding_rect_mw = font_metrics_mw.boundingRect(text_mw)
                        draw_x_mw_ls = loc_x_ls - (text_bounding_rect_mw.left() + text_bounding_rect_mw.width() / 2.0); draw_y_mw_ls = loc_y_ls - (text_bounding_rect_mw.top() + text_bounding_rect_mw.height() / 2.0)
                        painter.drawText(QPointF(draw_x_mw_ls, draw_y_mw_ls), text_mw)
                    except Exception as e: print(f"Error drawing placed MW prediction marker in paintEvent: {e}")

                # --- >>> MODIFICATION IS HERE <<< ---
                # --- Mode-Specific Previews and Finalized Shapes ---
                
                # Calculate pen width and handle radius based on zoom level.
                preview_pen_width = max(0.5, 1.5 / self.zoom_level if self.zoom_level > 0 else 1.5)
                handle_radius_view = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS

                # 1. Preview for Auto-Lane Quadrilateral (mode == 'auto_lane_quad')
                if self.mode == 'auto_lane_quad' and self.quad_points:
                    painter.setPen(QPen(QColor(0, 128, 128), preview_pen_width * 1.5, Qt.SolidLine)) # Teal for points
                    for p_ls in self.quad_points:
                        painter.drawEllipse(p_ls, handle_radius_view, handle_radius_view)
                    if 0 < len(self.quad_points) < 4:
                        painter.setPen(QPen(QColor(70, 130, 180), preview_pen_width, Qt.DotLine)) # SteelBlue for lines
                        painter.drawPolyline(QPolygonF(self.quad_points))
                
                # 2. Preview for Auto-Lane Rectangle (mode == 'auto_lane_rect')
                elif self.mode == 'auto_lane_rect' and self.bounding_box_preview:
                    painter.setPen(QPen(QColor(70, 130, 180), preview_pen_width, Qt.DotLine))
                    x1, y1, x2, y2 = self.bounding_box_preview
                    painter.drawRect(QRectF(QPointF(x1, y1), QPointF(x2, y2)).normalized())

                # 3. Preview for Single Analysis Quadrilateral (mode == 'quad')
                elif self.mode == "quad" and self.quad_points:
                    painter.setPen(QPen(Qt.red, preview_pen_width * 1.5, Qt.SolidLine))
                    for p_ls in self.quad_points:
                        painter.drawEllipse(p_ls, handle_radius_view, handle_radius_view)
                    if 0 < len(self.quad_points) < 4:
                        painter.setPen(QPen(Qt.blue, preview_pen_width, Qt.DotLine))
                        painter.drawPolyline(QPolygonF(self.quad_points))
                    elif len(self.quad_points) == 4:
                        painter.setPen(QPen(Qt.blue, preview_pen_width, Qt.DotLine))
                        painter.drawPolygon(QPolygonF(self.quad_points))
                
                # 4. Preview for Single Analysis Rectangle (mode == 'rectangle')
                elif self.mode == "rectangle" and self.bounding_box_preview:
                    painter.setPen(QPen(Qt.blue, preview_pen_width, Qt.DotLine))
                    x1, y1, x2, y2 = self.bounding_box_preview
                    painter.drawRect(QRectF(QPointF(x1, y1), QPointF(x2, y2)).normalized())
                
                # 5. Preview for a NEW Multi-Lane shape being defined
                elif self.app_instance and self.app_instance.multi_lane_mode_active and \
                     self.app_instance.current_selection_mode not in ["dragging_shape", "resizing_corner"] and \
                     self.app_instance.moving_multi_lane_index < 0 and \
                     self.mode not in ['auto_lane_quad', 'auto_lane_rect', 'quad', 'rectangle']:
                    
                    if self.app_instance.multi_lane_definition_type == 'quad' and self.app_instance.current_multi_lane_points:
                        points_to_draw = self.app_instance.current_multi_lane_points
                        painter.setPen(QPen(Qt.magenta, preview_pen_width * 1.5, Qt.SolidLine))
                        for p_ls in points_to_draw:
                            painter.drawEllipse(p_ls, handle_radius_view, handle_radius_view)
                        if 0 < len(points_to_draw) < 4:
                            painter.setPen(QPen(Qt.darkMagenta, preview_pen_width, Qt.DotLine))
                            painter.drawPolyline(QPolygonF(points_to_draw))
                        elif len(points_to_draw) == 4:
                            painter.setPen(QPen(Qt.darkMagenta, preview_pen_width, Qt.DotLine))
                            painter.drawPolygon(QPolygonF(points_to_draw))
                    elif self.app_instance.multi_lane_definition_type == 'rectangle' and self.bounding_box_preview:
                        painter.setPen(QPen(Qt.darkMagenta, preview_pen_width, Qt.DotLine))
                        x1, y1, x2, y2 = self.bounding_box_preview
                        painter.drawRect(QRectF(QPointF(x1, y1), QPointF(x2, y2)).normalized())

                # --- Draw Finalized Shapes (Single and Multi-lane) ---
                if self.quad_points and self.mode not in ["quad", "auto_lane_quad"] and \
                   (not (self.app_instance and self.app_instance.multi_lane_definitions) or is_selected_single_quad_for_move):
                    pen_color = Qt.magenta if is_selected_single_quad_for_move else Qt.darkYellow
                    pen_width_sf = 2.0 if is_selected_single_quad_for_move else 1.0
                    effective_pen_width_quad = max(0.5, pen_width_sf / self.zoom_level)
                    quad_pen = QPen(pen_color, effective_pen_width_quad, Qt.SolidLine)
                    painter.setPen(quad_pen)
                    if is_selected_single_quad_for_move or \
                       (self.app_instance and self.app_instance.current_selection_mode == "select_for_move" and self.app_instance.moving_multi_lane_index == -2):
                        point_pen_color = Qt.blue if self.app_instance.current_selection_mode == "select_for_move" else Qt.red
                        painter.setPen(QPen(point_pen_color, effective_pen_width_quad * 2))
                        for idx, p_label_space in enumerate(self.quad_points): 
                            if self.app_instance.current_selection_mode == "resizing_corner" and \
                               self.app_instance.moving_multi_lane_index == -2 and self.app_instance.resizing_corner_index == idx:
                                painter.setBrush(QBrush(Qt.red))
                                painter.drawEllipse(p_label_space, handle_radius_view * 1.2, handle_radius_view * 1.2) 
                                painter.setBrush(Qt.NoBrush)
                            else:
                                painter.drawEllipse(p_label_space, handle_radius_view, handle_radius_view)
                        painter.setPen(quad_pen)
                    if len(self.quad_points) == 4: painter.drawPolygon(QPolygonF(self.quad_points))

                elif self.bounding_box_preview and self.mode not in ["rectangle", "auto_lane_rect"] and \
                     (not (self.app_instance and self.app_instance.multi_lane_definitions) or is_selected_single_rect_for_move):
                    pen_color = Qt.magenta if is_selected_single_rect_for_move else Qt.darkYellow
                    pen_width_sf = 2.0 if is_selected_single_rect_for_move else 1.0
                    effective_pen_width_bbox = max(0.5, pen_width_sf / self.zoom_level)
                    rect_pen = QPen(pen_color, effective_pen_width_bbox, Qt.SolidLine)
                    painter.setPen(rect_pen)
                    start_x, start_y, end_x, end_y = self.bounding_box_preview
                    rect_label_space = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y)).normalized()
                    painter.drawRect(rect_label_space)
                    if is_selected_single_rect_for_move or \
                       (self.app_instance and self.app_instance.current_selection_mode == "select_for_move" and self.app_instance.moving_multi_lane_index == -3):
                        point_pen_color = Qt.blue if self.app_instance.current_selection_mode == "select_for_move" else Qt.red
                        painter.setPen(QPen(point_pen_color, effective_pen_width_bbox * 2))
                        rect_corners = [rect_label_space.topLeft(), rect_label_space.topRight(), rect_label_space.bottomRight(), rect_label_space.bottomLeft()]
                        for idx, corner_ls in enumerate(rect_corners):
                             if self.app_instance.current_selection_mode == "resizing_corner" and \
                                self.app_instance.moving_multi_lane_index == -3 and self.app_instance.resizing_corner_index == idx:
                                 painter.setBrush(QBrush(Qt.red))
                                 painter.drawEllipse(corner_ls, handle_radius_view * 1.2, handle_radius_view * 1.2)
                                 painter.setBrush(Qt.NoBrush)
                             else:
                                painter.drawEllipse(corner_ls, handle_radius_view, handle_radius_view)
                        painter.setPen(rect_pen)

                if self.app_instance and hasattr(self.app_instance, 'multi_lane_definitions') and self.app_instance.multi_lane_definitions:
                    lane_font_size = 10
                    lane_font = QFont("Arial", int(lane_font_size / self.zoom_level if self.zoom_level > 0 else lane_font_size)); lane_font.setBold(True)
                    for i, lane_def in enumerate(self.app_instance.multi_lane_definitions):
                        lane_id_str = str(lane_def['id']); center_point = QPointF()
                        is_selected_for_move_or_resize = (self.app_instance.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"] and self.app_instance.moving_multi_lane_index == i)
                        pen_width_multilane = max(0.5, (2.0 if is_selected_for_move_or_resize else 1.0) / self.zoom_level)
                        pen_defined_lane = QPen(Qt.magenta if is_selected_for_move_or_resize else Qt.darkYellow, pen_width_multilane, Qt.SolidLine); painter.setPen(pen_defined_lane)
                        current_lane_shape_points_label = []
                        if lane_def['type'] == 'rectangle':
                            rect_label_space = lane_def['points_label'][0]; painter.drawRect(rect_label_space); center_point = rect_label_space.center()
                            if is_selected_for_move_or_resize: current_lane_shape_points_label = [rect_label_space.topLeft(), rect_label_space.topRight(), rect_label_space.bottomRight(), rect_label_space.bottomLeft()]
                        elif lane_def['type'] == 'quad':
                            quad_points_label_space = lane_def['points_label']; poly = QPolygonF(quad_points_label_space); painter.drawPolygon(poly)
                            if is_selected_for_move_or_resize: current_lane_shape_points_label = quad_points_label_space
                            if len(quad_points_label_space) == 4: cx = sum(p.x() for p in quad_points_label_space) / 4.0; cy = sum(p.y() for p in quad_points_label_space) / 4.0; center_point = QPointF(cx, cy)
                        if is_selected_for_move_or_resize and current_lane_shape_points_label:
                            point_pen_color = Qt.blue if self.app_instance.current_selection_mode == "select_for_move" else Qt.red
                            painter.setPen(QPen(point_pen_color, pen_width_multilane * 1.5))
                            for idx, p_ls_multi in enumerate(current_lane_shape_points_label):
                                if self.app_instance.current_selection_mode == "resizing_corner" and \
                                   self.app_instance.moving_multi_lane_index == i and self.app_instance.resizing_corner_index == idx: 
                                    painter.setBrush(QBrush(Qt.red))
                                    painter.drawEllipse(p_ls_multi, handle_radius_view * 1.2, handle_radius_view * 1.2)
                                    painter.setBrush(Qt.NoBrush)
                                else:
                                    painter.drawEllipse(p_ls_multi, handle_radius_view, handle_radius_view)
                            painter.setPen(pen_defined_lane)
                        if not center_point.isNull(): 
                            painter.setFont(lane_font); painter.setPen(Qt.black)
                            fm_lane = QFontMetrics(lane_font); text_rect_lane = fm_lane.boundingRect(lane_id_str)
                            draw_x_lane = center_point.x() - text_rect_lane.width() / 2.0 - text_rect_lane.left(); draw_y_lane = center_point.y() - text_rect_lane.height() / 2.0 - text_rect_lane.top()
                            bg_rect_lane = QRectF(draw_x_lane - 2, draw_y_lane - 2 + text_rect_lane.top(), text_rect_lane.width() + 4, text_rect_lane.height() + 4)
                            painter.save(); painter.setBrush(QColor(255, 255, 255, 220 if is_selected_for_move_or_resize else 180)); painter.setPen(Qt.NoPen); painter.drawRoundedRect(bg_rect_lane, 3, 3); painter.restore()
                            painter.setPen(Qt.red if is_selected_for_move_or_resize else Qt.black); painter.drawText(QPointF(draw_x_lane, draw_y_lane), lane_id_str)

                # --- Draw Grid Lines (as before) ---
                if self.app_instance and hasattr(self.app_instance, 'grid_size_input') and \
                   hasattr(self.app_instance, 'show_grid_checkbox_x') and hasattr(self.app_instance, 'show_grid_checkbox_y'):
                    grid_size_label_space = self.app_instance.grid_size_input.value()
                    if grid_size_label_space > 0:
                        pen_grid_paint = QPen(Qt.red)
                        pen_grid_paint.setStyle(Qt.DashLine)
                        effective_pen_width_grid = max(0.5, 1.0 / self.zoom_level if self.zoom_level > 0 else 1.0)
                        pen_grid_paint.setWidthF(effective_pen_width_grid)
                        painter.setPen(pen_grid_paint)
                        label_width_unzoomed = self.width() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        label_height_unzoomed = self.height() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        view_origin_x_unzoomed = -self.pan_offset.x() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        view_origin_y_unzoomed = -self.pan_offset.y() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        if self.app_instance.show_grid_checkbox_x.isChecked():
                            start_x_grid = (int(view_origin_x_unzoomed / grid_size_label_space) -1) * grid_size_label_space
                            for x_grid_ls in range(start_x_grid, int(view_origin_x_unzoomed + label_width_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(x_grid_ls, view_origin_y_unzoomed), QPointF(x_grid_ls, view_origin_y_unzoomed + label_height_unzoomed))
                        if self.app_instance.show_grid_checkbox_y.isChecked():
                            start_y_grid = (int(view_origin_y_unzoomed / grid_size_label_space)-1) * grid_size_label_space
                            for y_grid_ls in range(start_y_grid, int(view_origin_y_unzoomed + label_height_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(view_origin_x_unzoomed, y_grid_ls), QPointF(view_origin_x_unzoomed + label_width_unzoomed, y_grid_ls))
                
                painter.restore()
            
            def leaveEvent(self, event):
                super().leaveEvent(event)
                # Emit signal with invalid label coordinates and False for image coordinate validity
                self.mouseMovedInLabel.emit(QPointF(-1,-1), QPointF(), False)

            def keyPressEvent(self, event):
                # This is LiveViewLabel's keyPressEvent.
                # It should primarily handle focus-related keys if needed.
                # Global shortcuts and mode cancellations are better handled in CombinedSDSApp.
                if event.key() == Qt.Key_Escape:
                    # If LiveViewLabel has a specific ESC action (like deselecting an internal point)
                    if self.selected_point != -1 and self.mode == "quad": # Example: deselect quad point
                        self.selected_point = -1
                        self.update()
                        event.accept()
                        return
                    # Propagate to parent (CombinedSDSApp) if not handled here
                    if self.app_instance and hasattr(self.app_instance, 'keyPressEvent'):
                        self.app_instance.keyPressEvent(event) # Let app handle global ESC
                        if event.isAccepted():
                            return
                
                super().keyPressEvent(event) # Default handling for other keys

        class CombinedSDSApp(QMainWindow):
            CONFIG_PRESET_FILE_NAME = "Imaging_assistant_preset_config.txt"
            MIME_TYPE_CUSTOM_ITEMS = "application/x-imaging-assistant.customitems+json"
            def __init__(self):
                super().__init__()
                primary_screen_obj = QApplication.primaryScreen()
                if primary_screen_obj:
                    self.screen = primary_screen_obj.geometry() # Or .availableGeometry()
                else: # Fallback if no primary screen
                    screens = QApplication.screens()
                    if screens:
                        self.screen = screens[0].geometry()
                    else:
                        # Further fallback or error if no screens detected
                        print("Warning: No screens detected. Using default screen dimensions.")
                        self.screen = QRect(0, 0, 1024, 768) # Arbitrary default
                self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
                # window_width = int(self.screen_width * 0.5)  # 60% of screen width
                # window_height = int(self.screen_height * 0.75)  # 95% of screen height
                self.preview_label_width_setting = int(self.screen_width * 0.35)
                self.preview_label_max_height_setting = int(self.screen_height * 0.30)
                self.label_size = self.preview_label_width_setting
                self.window_title="IMAGING ASSISTANT V8.0"
                # --- Initialize Status Bar Labels ---
                self.size_label = QLabel("Image Size: N/A")
                self.depth_label = QLabel("Bit Depth: N/A")
                self.location_label = QLabel("Source: N/A")
                self.mouse_coord_label = QLabel("X: --, Y: --")
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None  # 'quad' or 'rectangle'
                self.multi_lane_definitions = []  # List of dicts: {'type': 'quad'/'rect', 'points_label': [...], 'id': int}
                                                  # For rect: 'points_label' will be [QRectF_in_label_space]
                                                  # For quad: 'points_label' will be list of 4 QPointF_in_label_space
                self.current_multi_lane_points = []  # Temporary for defining current quad
                self.current_multi_lane_rect_start = None # For defining current rect
                self.latest_multi_lane_peak_areas = {} # Key: lane_id (int), Value: list of areas
                self.latest_multi_lane_peak_details = {} # Key: lane_id, Value: list of dicts with area & y_coord
                self.latest_multi_lane_calculated_quantities = {} # Key: lane_id (int), Value: list of quantities
                self.multi_lane_processing_finished = False # Flag
                self.moving_multi_lane_index = -1 
                self.moving_custom_item_info = None # Will be {'type': 'marker'/'shape', 'index': int}
                self.current_selection_mode = None # None, "select_for_move", "dragging_shape", "resizing_corner", "select_custom_item", ...
                self.resizing_corner_index = -1 # Index of the corner (0-3) being resized
                

                # --- Add Labels to Status Bar ---
                statusbar = self.statusBar() # Get the status bar instance
                statusbar.addWidget(self.size_label)
                statusbar.addWidget(self.depth_label)
                statusbar.addWidget(self.mouse_coord_label) # Add new label here
                statusbar.addWidget(self.location_label, 1) # location_label remains stretched
                self.setWindowTitle(self.window_title)
                self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
                self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
                self.undo_stack = []
                self.redo_stack = []
                self.custom_shapes = []  # NEW: List to store lines/rectangles
                self.drawing_mode = None # NEW: None, 'line', 'rectangle'
                self.current_drawing_shape_preview = None # NEW: For live preview data
                self.quantities_peak_area_dict = {}
                self.latest_peak_details = [] # List of dicts: [{'area':val, 'y_coord_in_lane_image':val}, ...]
                self.latest_peak_areas = []
                self.latest_calculated_quantities = []
                self.image_path = None
                self.image = None
                self.image_master= None
                self.image_before_padding = None
                self.image_contrasted=None
                self.image_before_contrast=None
                self.contrast_applied=False
                self.image_padded=False
                self.predict_size=False
                self.warped_image=None
                self.transparency=1
                self.left_markers = []
                self.right_markers = []
                self.top_markers = []
                self.custom_markers=[]
                self.current_left_marker_index = 0
                self.current_right_marker_index = 0
                self.current_top_label_index = 0
                self.font_rotation=-45
                self.image_width=0
                self.image_height=0
                self.new_image_width=0
                self.new_image_height=0
                self.base_name="Image"
                self.image_path=""
                self.x_offset_s=0
                self.y_offset_s=0
                self.peak_area=[]
                self.auto_marker_side = None # To store 'left' or 'right' for auto mode
                self.auto_lane_quad_points = [] # Store points defined for auto lane
                self.is_modified=False
                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None # Temp storage for starting point in view coords
                self.crop_rectangle_coords = None # Stores final (x, y, w, h) in *image* coordinates
                self.crop_offset_x = 0
                self.crop_offset_y = 0
                
                self.copied_peak_regions_data = {
                    "regions": None,        # List of (start, end) tuples
                    "profile_length": None  # Length of the profile from which regions were copied
                }
                self.peak_dialog_settings = {
                    'rolling_ball_radius': 50,
                    'peak_height_factor': 0.1,
                    'peak_distance': 10,
                    'peak_prominence_factor': 0.00,
                    'valley_offset_pixels': 0,  # Added to persist settings
                    'band_estimation_method': "Mean",
                    'area_subtraction_method': "Rolling Ball",
                    'smoothing_sigma': 2.0, # Added to persist settings
                }
                
                self.persist_peak_settings_enabled = True # State of the checkbox
                
                
                # Variables to store bounding boxes and quantities
                self.bounding_boxes = []
                self.up_bounding_boxes = []
                self.standard_protein_areas=[]
                self.quantities = []
                self.protein_quantities = []
                self.measure_quantity_mode = False
                # Initialize self.marker_values to None initially
                self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"] # Default internal top_label
                self.marker_values = [] # Default internal marker_values (for L/R) - will be populated by preset

                self.custom_marker_name = "Custom"
                self.presets_data = {}
                self.custom_marker_name = "Custom"
                # Load the config file if exists
                
                self.marker_mode = None
                self.left_marker_shift = 0   # Additional shift for marker text
                self.right_marker_shift = 0   # Additional shift for marker tex
                self.top_marker_shift=0 
                self.left_marker_shift_added=0
                self.right_marker_shift_added=0
                self.top_marker_shift_added= 0
                self.left_slider_range=[-1000,1000]
                self.right_slider_range=[-1000,1000]
                self.top_slider_range=[-1000,1000]
                       
                
                self.top_padding = 0
                self.font_color = QColor(0, 0, 0)  # Default to black
                self.custom_marker_color = QColor(0, 0, 0)  # Default to black
                self.font_family = "Arial"  # Default font family
                self.font_size = 12  # Default font size
                self.image_array_backup= None
                self.run_predict_MW=False
                
                
                
                # Main container widget
                main_widget = QWidget()
                self.setCentralWidget(main_widget)
                layout = QVBoxLayout(main_widget)

                # Upper section (Preview and buttons)
                upper_layout = QHBoxLayout()

                self.label_width=int(self.label_size)
            
                self.live_view_label = LiveViewLabel(
                    font_type=QFont("Arial"),
                    font_size=int(24),
                    marker_color=QColor(0,0,0),
                    app_instance=self, # Pass the CombinedSDSApp instance
                    parent=self,
                )
                # Image display
                self.live_view_label.setStyleSheet("background-color: white; border: 1px solid black;")
                self.live_view_label.mouseMovedInLabel.connect(self.update_mouse_coords_in_statusbar)
                
                self._create_actions()
                #self.create_menu_bar()
                self.create_tool_bar()
                
                self._update_preview_label_size()
                upper_layout.addWidget(self.live_view_label, stretch=1)
                layout.addLayout(upper_layout)
                
                # Lower section (Tabbed interface)
                self.tab_widget = QTabWidget()
                # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
                self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
                self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
                self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
                self.tab_widget.addTab(self.create_markers_tab(), "Markers")
                self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
                self.tab_widget.addTab(self.analysis_tab(), "Analysis")
                
                layout.addWidget(self.tab_widget)
                
                #self.load_shortcut = QShortcut(QKeySequence.Open, self) # Ctrl+O / Cmd+O
                #self.load_shortcut.activated.connect(self.load_action.trigger) # Trigger the action

                #self.save_shortcut = QShortcut(QKeySequence.Save, self) # Ctrl+S / Cmd+S
                #self.save_shortcut.activated.connect(self.save_action.trigger) # Trigger the action

                #self.copy_shortcut = QShortcut(QKeySequence.Copy, self) # Ctrl+C / Cmd+C
                #self.copy_shortcut.activated.connect(self.copy_action.trigger) # Trigger the action

                #self.paste_shortcut = QShortcut(QKeySequence.Paste, self) # Ctrl+V / Cmd+V
                #self.paste_shortcut.activated.connect(self.paste_action.trigger) # Trigger the action

                #self.undo_shortcut = QShortcut(QKeySequence.Undo, self) # Ctrl+Z / Cmd+Z
                #self.undo_shortcut.activated.connect(self.undo_action_m) # Connect directly to method

                #self.redo_shortcut = QShortcut(QKeySequence.Redo, self) # Ctrl+Y / Cmd+Shift+Z
                #self.redo_shortcut.activated.connect(self.redo_action_m) # Connect directly to method

                #self.zoom_out_shortcut = QShortcut(QKeySequence.ZoomOut, self) # Ctrl+- / Cmd+-
                #self.zoom_out_shortcut.activated.connect(self.zoom_out_action.trigger) # Trigger the action

                # Zoom In - Keep the standard one
                #self.zoom_in_shortcut = QShortcut(QKeySequence.ZoomIn, self) # Attempts Ctrl++ / Cmd++
                #self.zoom_in_shortcut.activated.connect(self.zoom_in_action.trigger) # Trigger the action
                # ======================================================


                # === KEEP QShortcut definitions for NON-STANDARD actions ===
                # self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
                # self.save_svg_shortcut.activated.connect(self.save_svg_action.trigger)

                self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
                self.reset_shortcut.activated.connect(self.reset_action.trigger)

                self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
                self.predict_shortcut.activated.connect(self.predict_molecular_weight)

                self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
                self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)

                self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
                self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)

                self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
                self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)

                self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
                self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)

                self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
                self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
                self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
                self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
                self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
                self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
                self.grid_shortcut.activated.connect(
                    lambda: (
                        # Determine the target state (True if both are currently False, False otherwise)
                        target_state := not (self.show_grid_checkbox_x.isChecked() or self.show_grid_checkbox_y.isChecked()),
                        # Set both checkboxes to the target state
                        self.show_grid_checkbox_x.setChecked(target_state),
                        self.show_grid_checkbox_y.setChecked(target_state)
                    ) if hasattr(self, 'show_grid_checkbox_x') and hasattr(self, 'show_grid_checkbox_y') else None
                )
                
                self.grid_shortcut_x = QShortcut(QKeySequence("Ctrl+Shift+X"), self) # Use uppercase X for consistency
                self.grid_shortcut_x.activated.connect(
                    lambda: self.show_grid_checkbox_x.setChecked(not self.show_grid_checkbox_x.isChecked())
                    if hasattr(self, 'show_grid_checkbox_x') else None
                )

                # Shortcut for Y Grid Snapping
                self.grid_shortcut_y = QShortcut(QKeySequence("Ctrl+Shift+Y"), self) # Use uppercase Y
                self.grid_shortcut_y.activated.connect(
                    lambda: self.show_grid_checkbox_y.setChecked(not self.show_grid_checkbox_y.isChecked())
                    if hasattr(self, 'show_grid_checkbox_y') else None
                )

                self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
                self.guidelines_shortcut.activated.connect(
                    lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked())
                    if hasattr(self, 'show_guides_checkbox') else None
                )

                self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
                self.increase_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1)
                    if hasattr(self, 'grid_size_input') else None
                )
                self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
                self.decrease_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1)
                    if hasattr(self, 'grid_size_input') else None
                )

                self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
                self.invert_shortcut.activated.connect(self.invert_image)

                self.bw_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
                self.bw_shortcut.activated.connect(self.convert_to_black_and_white)

                # Tab movement shortcuts
                self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
                self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
                self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
                self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
                self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
                self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
                self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
                self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
                self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
                self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
                self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
                self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
                
                self.load_config()
                
            def update_mouse_coords_in_statusbar(self, label_pos: QPointF, image_pos: QPointF = None):
                if image_pos is not None:
                    # Display image coordinates if available (mouse is over the image)
                    self.mouse_coord_label.setText(f"Img X: {image_pos.x():.0f}, Y: {image_pos.y():.0f}")
                elif label_pos is not None:
                    # Fallback to label coordinates if not over image or no image loaded
                    self.mouse_coord_label.setText(f"View X: {label_pos.x():.0f}, Y: {label_pos.y():.0f}")
                else:
                    # Default if no valid position data (e.g., mouse exited label)
                    self.mouse_coord_label.setText("X: --, Y: --")
                    
            def _find_image_content_borders(self):
                """
                Scans the current self.image to find the left and right borders of the
                non-padding content. Assumes padding is transparent.
                Returns a tuple (left_border_x, right_border_x), or (None, None) on failure.
                """
                if not self.image or self.image.isNull():
                    return None, None

                # Use a copy to avoid modifying the original
                qimg_to_scan = self.image.copy()
                # Ensure the image has an alpha channel for reliable detection
                if not qimg_to_scan.hasAlphaChannel():
                    qimg_to_scan = qimg_to_scan.convertToFormat(QImage.Format_ARGB32)

                np_img = self.qimage_to_numpy(qimg_to_scan)
                if np_img is None or np_img.ndim < 3 or np_img.shape[2] < 4:
                    print("Warning: Could not get image with alpha channel for border detection.")
                    return None, None

                height, width, _ = np_img.shape
                # The alpha channel is the 4th channel (index 3)
                alpha_channel = np_img[:, :, 3]

                # Scan from left to right to find the first non-transparent column
                left_border_x = None
                for x in range(width):
                    # np.any is fast for checking if any value in the column is non-zero
                    if np.any(alpha_channel[:, x] > 0):
                        left_border_x = x
                        break

                # Scan from right to left to find the first non-transparent column
                right_border_x = None
                for x in range(width - 1, -1, -1):
                    if np.any(alpha_channel[:, x] > 0):
                        right_border_x = x
                        break
                
                if left_border_x is None or right_border_x is None:
                    # This case happens if the image is fully transparent
                    return 0, width-1 # Fallback to image edges

                return left_border_x, right_border_x
                    
            def _reset_live_view_label_custom_handlers(self):
                """Helper to reset all custom mouse interaction hooks on LiveViewLabel."""
                if hasattr(self.live_view_label, '_custom_left_click_handler_from_app'):
                    self.live_view_label._custom_left_click_handler_from_app = None
                if hasattr(self.live_view_label, '_custom_mouseMoveEvent_from_app'):
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                if hasattr(self.live_view_label, '_custom_mouseReleaseEvent_from_app'):
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                # Ensure mouse tracking is on if it was turned off by a mode
                self.live_view_label.setMouseTracking(True)
                self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            
            def _create_actions(self):
                """Create QAction objects for menus and toolbars."""
                style = self.style() # Still useful for other icons
                icon_size = QSize(30, 30) # Match your toolbar size
                text_color = self.palette().color(QPalette.ButtonText) # Use theme color

                # # --- Create Zoom Icons (using previous method) ---
                # zoom_in_pixmap = QPixmap(icon_size)
                # zoom_in_pixmap.fill(Qt.transparent)
                # painter_in = QPainter(zoom_in_pixmap)
                # # ... (painter setup as before) ...
                # painter_in.drawText(zoom_in_pixmap.rect(), Qt.AlignCenter, "+")
                # painter_in.end()
                # zoom_in_icon = QIcon(zoom_in_pixmap)

                # zoom_out_pixmap = QPixmap(icon_size)
                # zoom_out_pixmap.fill(Qt.transparent)
                # painter_out = QPainter(zoom_out_pixmap)
                # # ... (painter setup as before) ...
                # painter_out.drawText(zoom_out_pixmap.rect(), Qt.AlignCenter, "-")
                # painter_out.end()
                # zoom_out_icon = QIcon(zoom_out_pixmap)
                # --- End Zoom Icons ---


                # --- Create Icons using the helper ---
                open_icon = create_text_icon("Wingdings",icon_size, text_color, "1") # Unicode Right Arrow
                save_icon = create_text_icon("Wingdings",icon_size, text_color, "=")
                save_svg_icon = create_text_icon("Wingdings",icon_size, text_color, "3")
                undo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "O")
                redo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "N")
                paste_icon = create_text_icon("Wingdings 2",icon_size, text_color, "2")
                copy_icon = create_text_icon("Wingdings",icon_size, text_color, "4")
                reset_icon = create_text_icon("Wingdings 3",icon_size, text_color, "Q")
                exit_icon = create_text_icon("Wingdings 2",icon_size, text_color, "V")
                
                zoom_in_icon = create_text_icon("Arial",icon_size, text_color, "+")
                zoom_out_icon = create_text_icon("Arial",icon_size, text_color, "-")
                pan_up_icon = create_text_icon("Arial",icon_size, text_color, "↑") # Unicode Up Arrow
                pan_down_icon = create_text_icon("Arial",icon_size, text_color, "↓") # Unicode Down Arrow
                pan_left_icon = create_text_icon("Arial",icon_size, text_color, "←") # Unicode Left Arrow
                pan_right_icon = create_text_icon("Arial",icon_size, text_color, "→") # Unicode Right Arrow
                bounding_box_icon = create_text_icon("Wingdings 2", icon_size, text_color, "0")
                draw_line_icon = create_text_icon("Arial", icon_size, text_color, "__")
                info_icon = create_text_icon("Wingdings", icon_size, text_color, "'")
                copy_custom_icon = create_text_icon("Wingdings", icon_size, text_color, "B") # Page with arrow out
                paste_custom_icon = create_text_icon("Wingdings", icon_size, text_color, "A") # Page with arrow in


                # --- File Actions ---
                self.load_action = QAction(open_icon, "&Load Image...", self)
                self.save_action = QAction(save_icon, "&Save with Config", self)
                # self.save_svg_action = QAction(save_svg_icon, "Save &SVG...", self)
                self.reset_action = QAction(reset_icon, "&Reset Image", self)
                self.exit_action = QAction(exit_icon, "E&xit", self)

                # --- Edit Actions ---
                self.undo_action = QAction(undo_icon, "&Undo", self) # Standard for now
                self.redo_action = QAction(redo_icon, "&Redo", self) # Standard for now
                self.copy_action = QAction(copy_icon, "&Copy Image", self)
                self.paste_action = QAction(paste_icon, "&Paste Image", self)

                # --- View Actions (using the created icons) ---
                self.zoom_in_action = QAction(zoom_in_icon, "Zoom &In", self)
                self.zoom_out_action = QAction(zoom_out_icon, "Zoom &Out", self)
                self.pan_left_action = QAction(pan_left_icon, "Pan Left", self)
                self.pan_right_action = QAction(pan_right_icon, "Pan Right", self)
                self.pan_up_action = QAction(pan_up_icon, "Pan Up", self)
                self.pan_down_action = QAction(pan_down_icon, "Pan Down", self)
                # --- END: Add Panning Actions ---
                self.auto_lane_action = QAction(create_text_icon("Arial", icon_size, text_color, "A"), "&Automatic Lane Markers", self)
                self.auto_lane_action.setToolTip("Automatically detect and place lane markers based on a defined region.")
                self.auto_lane_action.triggered.connect(self.start_auto_lane_marker)
                
                self.info_action = QAction(info_icon, "&Info/GitHub", self)
                self.info_action.setToolTip("Open Project GitHub Page")
                self.info_action.triggered.connect(self.open_github)

                # --- Set Shortcuts ---
                self.load_action.setShortcut(QKeySequence.Open)
                self.save_action.setShortcut(QKeySequence.Save)
                self.copy_action.setShortcut(QKeySequence.Copy)
                self.paste_action.setShortcut(QKeySequence.Paste)
                self.undo_action.setShortcut(QKeySequence.Undo)
                self.redo_action.setShortcut(QKeySequence.Redo)
                self.zoom_in_action.setShortcut(QKeySequence("Ctrl+="))
                self.zoom_out_action.setShortcut(QKeySequence.ZoomOut)
                # self.save_svg_action.setShortcut(QKeySequence("Ctrl+M"))
                self.reset_action.setShortcut(QKeySequence("Ctrl+R"))

                # --- Set Tooltips ---
                self.load_action.setToolTip("Load an image file (Ctrl+O)")
                self.save_action.setToolTip("Save image and configuration (Ctrl+S)")
                # self.save_svg_action.setToolTip("Save as SVG for Word/vector editing (Ctrl+M)")
                self.reset_action.setToolTip("Reset image and all annotations (Ctrl+R)")
                self.exit_action.setToolTip("Exit the application")
                self.undo_action.setToolTip("Undo last action (Default Shortcut: OS dependent")
                self.redo_action.setToolTip("Redo last undone action (Default Shortcut: OS dependent)")
                self.copy_action.setToolTip("Copy rendered image to clipboard (Ctrl+C)")
                self.paste_action.setToolTip("Paste image from clipboard (Ctrl+V)")
                self.zoom_in_action.setToolTip("Increase zoom level (Ctrl+= or mouse scroll bar)")
                self.zoom_out_action.setToolTip("Decrease zoom level (Ctrl+- or mouse scroll bar)). Auto resets the zoom when reaches zero.")
                # --- START: Set Tooltips for Panning Actions ---
                self.pan_left_action.setToolTip("Pan the view left (when zoomed) (Arrow key left or mouse right click)")
                self.pan_right_action.setToolTip("Pan the view right (when zoomed) (Arrow key right or mouse right click")
                self.pan_up_action.setToolTip("Pan the view up (when zoomed) (Arrow key up or mouse right click)")
                self.pan_down_action.setToolTip("Pan the view down (when zoomed) (Arrow key down or mouse right click")
                self.draw_bounding_box_action = QAction(bounding_box_icon, "Draw &Bounding Box", self)
                self.draw_bounding_box_action.setToolTip("Draw a bounding rectangle on the image. Use the marker tab, custom marker options for color and size")
                self.draw_bounding_box_action.triggered.connect(self.enable_rectangle_drawing_mode)
                
                self.draw_line_action = QAction(draw_line_icon, "Draw &a line", self)
                self.draw_line_action.setToolTip("Draw a line on the image. Use the marker tab, custom marker options for color and size")
                self.draw_line_action.triggered.connect(self.enable_line_drawing_mode)
                # --- END: Set Tooltips for Panning Actions ---

                # --- Connect signals ---
                self.load_action.triggered.connect(self.load_image)
                self.save_action.triggered.connect(self.save_image)
                # self.save_svg_action.triggered.connect(self.save_image_svg)
                self.reset_action.triggered.connect(self.reset_image)
                self.exit_action.triggered.connect(self.close)
                self.undo_action.triggered.connect(self.undo_action_m)
                self.redo_action.triggered.connect(self.redo_action_m)
                self.copy_action.triggered.connect(self.copy_to_clipboard)
                self.paste_action.triggered.connect(self.paste_image)
                self.zoom_in_action.triggered.connect(self.zoom_in)
                self.zoom_out_action.triggered.connect(self.zoom_out)
                # --- START: Connect Panning Action Signals ---
                # Use lambda functions to directly modify the offset
                pan_step = 30 # Define pan step size here or access from elsewhere if needed
                self.pan_right_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_right_action.triggered.connect(self.update_live_view) # Trigger update after offset change

                self.pan_left_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_left_action.triggered.connect(self.update_live_view)

                self.pan_down_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_down_action.triggered.connect(self.update_live_view)

                self.pan_up_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_up_action.triggered.connect(self.update_live_view)
                
                self.copy_custom_items_action = QAction(copy_custom_icon, "Copy Custom Markers/Shapes", self)
                self.copy_custom_items_action.setToolTip("Copy all custom markers and shapes to the clipboard.")
                self.copy_custom_items_action.triggered.connect(self.copy_custom_items)

                self.paste_custom_items_action = QAction(paste_custom_icon, "Paste Custom Markers/Shapes", self)
                self.paste_custom_items_action.setToolTip("Paste custom markers and shapes from the clipboard.\nItems will be added to existing ones.")
                self.paste_custom_items_action.triggered.connect(self.paste_custom_items)
                
                # --- END: Connect Panning Action Signals ---

                # --- START: Initially Disable Panning Actions ---
                self.pan_left_action.setEnabled(False)
                self.pan_right_action.setEnabled(False)
                self.pan_up_action.setEnabled(False)
                self.pan_down_action.setEnabled(False)
                
            def copy_custom_items(self):
                """Copies custom markers and shapes to the clipboard as JSON."""
                if not hasattr(self, "custom_markers"): self.custom_markers = []
                if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                if not self.custom_markers and not self.custom_shapes:
                    QMessageBox.information(self, "Nothing to Copy", "There are no custom markers or shapes to copy.")
                    return

                # Serialize markers and shapes
                # custom_markers are stored as lists, custom_shapes as dicts
                # For markers, we need to convert QColor to string for JSON
                serializable_markers = []
                for marker_data in self.custom_markers:
                    try:
                        # marker_data is a list: [x, y, text, qcolor, font_family, font_size, is_bold, is_italic]
                        m_copy = list(marker_data)
                        if isinstance(m_copy[3], QColor):
                            m_copy[3] = m_copy[3].name() # Convert QColor to hex string
                        serializable_markers.append(m_copy)
                    except IndexError:
                        print(f"Warning: Skipping marker during copy due to unexpected format: {marker_data}")
                
                items_to_copy = {
                    "custom_markers": serializable_markers,
                    "custom_shapes": self.custom_shapes # Already list of dicts
                }

                try:
                    json_data = json.dumps(items_to_copy)
                    mime_data = QMimeData()
                    mime_data.setData(self.MIME_TYPE_CUSTOM_ITEMS, json_data.encode('utf-8'))
                    
                    # Also set text for basic paste compatibility (e.g., into a text editor)
                    mime_data.setText(f"Imaging Assistant Custom Items (JSON):\n{json_data}")
                    
                    QApplication.clipboard().setMimeData(mime_data)
                    QMessageBox.information(self, "Items Copied", 
                                            f"{len(self.custom_markers)} markers and {len(self.custom_shapes)} shapes copied to clipboard.")
                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Could not copy custom items: {e}")
                    traceback.print_exc()

            def paste_custom_items(self):
                """Pastes custom markers and shapes from the clipboard."""
                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()

                if not mime_data.hasFormat(self.MIME_TYPE_CUSTOM_ITEMS):
                    QMessageBox.information(self, "Paste Error", 
                                            "Clipboard does not contain Imaging Assistant custom items in the expected format.")
                    return

                try:
                    json_bytes = mime_data.data(self.MIME_TYPE_CUSTOM_ITEMS)
                    json_data_str = json_bytes.data().decode('utf-8') # Convert QByteArray to bytes then to string
                    pasted_items = json.loads(json_data_str)

                    pasted_markers_data = pasted_items.get("custom_markers", [])
                    pasted_shapes_data = pasted_items.get("custom_shapes", [])

                    if not pasted_markers_data and not pasted_shapes_data:
                        QMessageBox.information(self, "Paste", "No custom markers or shapes found in clipboard data.")
                        return

                    self.save_state() # Save current state before adding new items

                    if not hasattr(self, "custom_markers"): self.custom_markers = []
                    if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                    num_pasted_markers = 0
                    for marker_data_serial in pasted_markers_data:
                        try:
                            # marker_data_serial is a list: [x, y, text, color_str, font_family, font_size, is_bold, is_italic]
                            m_copy = list(marker_data_serial)
                            m_copy[3] = QColor(m_copy[3]) # Convert color string back to QColor
                            if not m_copy[3].isValid(): m_copy[3] = QColor(Qt.black) # Fallback
                            self.custom_markers.append(m_copy)
                            num_pasted_markers += 1
                        except (IndexError, TypeError, ValueError) as e:
                            print(f"Warning: Skipping pasted marker due to format error: {marker_data_serial}, {e}")
                    
                    num_pasted_shapes = 0
                    for shape_data in pasted_shapes_data:
                        # Basic validation (can be extended)
                        if isinstance(shape_data, dict) and 'type' in shape_data:
                            self.custom_shapes.append(shape_data)
                            num_pasted_shapes +=1
                        else:
                             print(f"Warning: Skipping pasted shape due to format error: {shape_data}")


                    self.is_modified = True
                    self.update_live_view()
                    QMessageBox.information(self, "Items Pasted",
                                            f"{num_pasted_markers} markers and {num_pasted_shapes} shapes pasted.")

                except Exception as e:
                    QMessageBox.critical(self, "Paste Error", f"Could not paste custom items: {e}")
                    traceback.print_exc()
                
            def _update_preview_label_size(self):
                """
                Updates the fixed size of the live_view_label, respecting max width/height
                and maintaining aspect ratio. Prioritizes fixing height, but adjusts if
                width constraint is violated.
                """
                # --- Define Defaults and Minimums ---
                default_max_width = 600  # Fallback if width setting is missing/invalid
                default_max_height = 500 # Fallback if height setting is missing/invalid
                min_dim = 50             # Minimum allowed dimension for the label

                # --- Determine Maximum Constraints (with validation) ---
                max_w = default_max_width
                if hasattr(self, 'preview_label_width_setting'):
                    try:
                        setting_width = int(self.preview_label_width_setting)
                        if setting_width > 0:
                            max_w = setting_width
                        else:
                            print("Warning: preview_label_width_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_width_setting is invalid, using default.")
                else:
                    print("Warning: preview_label_width_setting attribute not found, using default.")
                max_w = max(min_dim, max_w) # Apply minimum constraint

                max_h = default_max_height
                if hasattr(self, 'preview_label_max_height_setting'):
                    try:
                        setting_height = int(self.preview_label_max_height_setting)
                        if setting_height > 0:
                            max_h = setting_height
                        else:
                            print("Warning: preview_label_max_height_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_max_height_setting is invalid, using default.")
                max_h = max(min_dim, max_h) # Apply minimum constraint


                # --- Initialize Final Dimensions ---
                final_w = max_w
                final_h = max_h

                # --- Calculate Dimensions Based on Image ---
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()

                    if w > 0 and h > 0:
                        img_ratio = w / h

                        # Attempt 1: Calculate width based on fixed max height
                        calc_w_based_on_h = max_h * img_ratio

                        if calc_w_based_on_h <= max_w:
                            # Width is within limits, height is fixed
                            final_w = calc_w_based_on_h
                            final_h = max_h
                        else:
                            # Calculated width exceeds max width, so fix width and calculate height
                            final_w = max_w
                            final_h = max_w / img_ratio

                        # Ensure calculated dimensions are not below minimum
                        final_w = max(min_dim, int(final_w))
                        final_h = max(min_dim, int(final_h))

                    else:
                        # Handle invalid image dimensions (0 width or height)
                        final_w = max_w # Already set above
                        final_h = max_h # Already set above

                else:
                    # No image loaded - Use max constraints as default size
                    final_w = max_w # Already set above
                    final_h = max_h # Already set above
                    # self.live_view_label.clear() # Optionally clear the label

                # --- Set the Calculated Fixed Size ---
                self.live_view_label.setFixedSize(final_w, final_h)
                
            def _update_status_bar(self):
                """Updates the status bar labels with current image information."""
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()
                    self.size_label.setText(f"Image Size: {w}x{h}")

                    # Determine bit depth/format string
                    img_format = self.image.format()
                    depth_str = "Unknown"
                    if img_format == QImage.Format_Grayscale8:
                        depth_str = "8-bit Grayscale"
                    elif img_format == QImage.Format_Grayscale16:
                        depth_str = "16-bit Grayscale"
                    elif img_format == QImage.Format_RGB888:
                         depth_str = "24-bit RGB"
                    elif img_format in (QImage.Format_RGB32, QImage.Format_ARGB32,
                                        QImage.Format_ARGB32_Premultiplied, QImage.Format_RGBA8888):
                         depth_str = "32-bit (A)RGB"
                    elif img_format == QImage.Format_Indexed8:
                         depth_str = "8-bit Indexed"
                    elif img_format == QImage.Format_Mono:
                         depth_str = "1-bit Mono"
                    else:
                        # Fallback to depth() if format is unusual
                        depth_val = self.image.depth()
                        depth_str = f"{depth_val}-bit Depth"
                    self.depth_label.setText(f"Bit Depth: {depth_str}")

                    # Update location
                    location = "N/A"
                    if hasattr(self, 'image_path') and self.image_path:
                        # Show only filename if it's a path, otherwise show the source info
                        if os.path.exists(self.image_path):
                            location = os.path.basename(self.image_path)
                        else: # Likely "Clipboard..." or similar
                            location = self.image_path
                    else:
                        location = "N/A"
                    self.location_label.setText(f"Source: {location}")

                else:
                    # No image loaded
                    self.size_label.setText("Image Size: N/A")
                    self.depth_label.setText("Bit Depth: N/A")
                    self.location_label.setText("Source: N/A")
                
            def prompt_save_if_needed(self):
                if not self.is_modified:
                    return True # No changes, proceed
                """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
                reply = QMessageBox.question(self, 'Unsaved Changes',
                                             "Do you want to save the file?",
                                             QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                             QMessageBox.Save) # Default button is Save

                if reply == QMessageBox.Save:
                    return self.save_image() # Returns True if saved, False if save cancelled
                elif reply == QMessageBox.Discard:
                    return True # Proceed without saving
                else: # Cancelled
                    return False # Abort the current action (close/load)

            def closeEvent(self, event):
                """Overrides the default close event to prompt for saving."""
                if self.prompt_save_if_needed():

                    event.accept() # Proceed with closing
                else:
                    event.ignore() # Abort closing
                    
            def enable_line_drawing_mode(self):
                self.save_state()
                self.drawing_mode = 'line'
                self.live_view_label.mode = 'draw_shape' 
                self.current_drawing_shape_preview = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                # Use custom hooks instead of direct overwrite
                self.live_view_label._custom_left_click_handler_from_app = self.start_shape_draw
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_shape_draw
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_shape_draw

            def enable_rectangle_drawing_mode(self):
                self.save_state()
                self.drawing_mode = 'rectangle'
                self.live_view_label.mode = 'draw_shape'
                self.current_drawing_shape_preview = None
                self.live_view_label.setCursor(Qt.CrossCursor)

                self.live_view_label._custom_left_click_handler_from_app = self.start_shape_draw
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_shape_draw
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_shape_draw

            def cancel_drawing_mode(self):
                """Resets drawing mode and cursor."""
                self.drawing_mode = None
                self.live_view_label.mode = None # App-level mode for LiveViewLabel's paintEvent logic
                self.current_drawing_shape_preview = None
                self._reset_live_view_label_custom_handlers() # Use helper
                self.update_live_view()
                
            def start_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle']:
                    start_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    self.current_drawing_shape_preview = {'start': snapped_start_point, 'end': snapped_start_point}
                    self.update_live_view()

            def update_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
                    
                    self.current_drawing_shape_preview['end'] = snapped_end_point
                    self.update_live_view()

            def finalize_shape_draw(self, event):
                """Finalizes the shape and adds it to the custom_shapes list."""
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    # start_point_label_space is already snapped from start_shape_draw
                    start_point_label_space = self.current_drawing_shape_preview['start']
                    
                    end_point_transformed_label_space = self.live_view_label.transform_point(event.position())
                    end_point_snapped_label_space = self.snap_point_to_grid(end_point_transformed_label_space) # Snap it
    
                    # ... (Get current style settings: color, thickness) ...
                    color = self.custom_marker_color
                    thickness = self.custom_font_size_spinbox.value()


                    displayed_width = float(self.live_view_label.width())
                    displayed_height = float(self.live_view_label.height())
                    current_image_width = float(self.image.width()) if self.image and self.image.width() > 0 else 1.0
                    current_image_height = float(self.image.height()) if self.image and self.image.height() > 0 else 1.0
                    if current_image_width <= 0 or current_image_height <= 0: # Safety check
                        scale_img_to_label = 1.0
                    else:
                        scale_x_fit = displayed_width / current_image_width
                        scale_y_fit = displayed_height / current_image_height
                        scale_img_to_label = min(scale_x_fit, scale_y_fit)
                    display_img_w = current_image_width * scale_img_to_label
                    display_img_h = current_image_height * scale_img_to_label
                    label_offset_x = (displayed_width - display_img_w) / 2.0
                    label_offset_y = (displayed_height - display_img_h) / 2.0
                    def label_to_image_coords(label_point):
                        if scale_img_to_label <= 1e-9: return (0.0, 0.0)
                        relative_x_in_display = label_point.x() - label_offset_x
                        relative_y_in_display = label_point.y() - label_offset_y
                        img_x = relative_x_in_display / scale_img_to_label
                        img_y = relative_y_in_display / scale_img_to_label
                        return (img_x, img_y)
                    start_img_coords = label_to_image_coords(start_point_label_space) # Use snapped start
                    end_img_coords = label_to_image_coords(end_point_snapped_label_space) # Use snapped end
    
                    shape_data = {
                        'type': self.drawing_mode,
                        'color': color.name(),
                        'thickness': thickness
                    }
                    valid_shape = False
                    if self.drawing_mode == 'line':
                        if abs(start_img_coords[0] - end_img_coords[0]) > 0.5 or abs(start_img_coords[1] - end_img_coords[1]) > 0.5:
                            shape_data['start'] = start_img_coords
                            shape_data['end'] = end_img_coords
                            valid_shape = True
                    elif self.drawing_mode == 'rectangle':
                        x_img = min(start_img_coords[0], end_img_coords[0])
                        y_img = min(start_img_coords[1], end_img_coords[1])
                        w_img = abs(end_img_coords[0] - start_img_coords[0])
                        h_img = abs(end_img_coords[1] - start_img_coords[1])
                        if w_img > 0.5 and h_img > 0.5:
                            shape_data['rect'] = (x_img, y_img, w_img, h_img)
                            valid_shape = True
                    if valid_shape:
                        self.custom_shapes.append(shape_data)
                        self.save_state()
                        self.is_modified = True
                    self.cancel_drawing_mode()
                else:
                    self.cancel_drawing_mode()
                    
            def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
                """Converts QImage to NumPy array, preserving format and handling row padding."""
                if qimage.isNull():
                    return None

                img_format = qimage.format()
                height = qimage.height()
                width = qimage.width()
                bytes_per_line = qimage.bytesPerLine() # Actual bytes per row (includes padding)


                ptr = qimage.constBits()
                if not ptr:
                    return None

                # Ensure ptr is treated as a bytes object of the correct size
                # ptr is a sip.voidptr, needs explicit size setting for buffer protocol
                try:
                    # PySide6/Qt6 uses sizeInBytes()
                    expected_total_bytes = qimage.sizeInBytes()
                    ptr.setsize(expected_total_bytes)
                except AttributeError: # Fallback for older Qt/PyQt where setsize might not be needed or byteCount was used
                    try:
                        expected_total_bytes = qimage.byteCount() # Try PyQt5 way if sizeInBytes fails
                        ptr.setsize(expected_total_bytes)
                    except AttributeError: # If both fail, proceed with caution
                        expected_total_bytes = height * bytes_per_line # Estimate
                        pass # Continue, hoping buffer protocol works

                buffer_data = bytes(ptr) # Create a bytes object from the pointer

                # Use sizeInBytes() for PySide6
                current_total_bytes = qimage.sizeInBytes()
                if len(buffer_data) != current_total_bytes:
                     print(f"qimage_to_numpy: Warning - Buffer size mismatch. Expected {current_total_bytes}, got {len(buffer_data)}.")
                     # Fallback or error? Let's try to continue but warn.

                # --- Grayscale Formats ---
                if img_format == QImage.Format_Grayscale16:
                    bytes_per_pixel = 2
                    dtype = np.uint16
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        # No padding, simple reshape
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                        return arr.copy() # Return a copy
                    else:
                        # Handle padding
                        arr = np.zeros((height, width), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype)
                            else:
                                print(f"qimage_to_numpy (Grayscale16): Row data length mismatch for row {y}. Skipping row.") # Error handling
                        return arr
            
                elif img_format == QImage.Format_Grayscale8:
                    bytes_per_pixel = 1
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype)
                            else:
                                print(f"qimage_to_numpy (Grayscale8): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # --- Color Formats ---
                # ARGB32 (often BGRA in memory) & RGBA8888 (often RGBA in memory)
                elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied):
                    bytes_per_pixel = 4
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 4), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                            else:
                                print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # RGB32 (often BGRX or RGBX in memory) & RGBX8888
                elif img_format in (QImage.Format_RGB32, QImage.Format_RGBX8888):
                    bytes_per_pixel = 4 # Stored with an ignored byte
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 4), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                            else:
                                print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # RGB888 (Tightly packed RGB)
                elif img_format == QImage.Format_RGB888:
                    bytes_per_pixel = 3
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 3)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 3), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 3)
                            else:
                                print(f"qimage_to_numpy (RGB888): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # --- Fallback / Conversion Attempt ---
                else:
                    try:
                        qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                        if qimage_conv.isNull():
                            print("qimage_to_numpy: Fallback conversion to ARGB32 failed.")
                            qimage_conv_gray = qimage.convertToFormat(QImage.Format_Grayscale8)
                            if qimage_conv_gray.isNull():
                                 return None
                            else:
                                return self.qimage_to_numpy(qimage_conv_gray)
            
                        return self.qimage_to_numpy(qimage_conv)
                    except Exception as e:
                        print(f"qimage_to_numpy: Error during fallback conversion: {e}")
                        return None

            def numpy_to_qimage(self, array: np.ndarray) -> QImage:
                """Converts NumPy array to QImage, selecting appropriate format."""
                if array is None:
                    return QImage() # Return invalid QImage
            
                if not isinstance(array, np.ndarray):
                    return QImage()
            
                try:
                    if array.ndim == 2: # Grayscale
                        height, width = array.shape
                        if array.dtype == np.uint16:
                            bytes_per_line = width * 2
                            # Ensure data is contiguous
                            contiguous_array = np.ascontiguousarray(array)
                            # Create QImage directly from contiguous data buffer
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                            # Important: QImage doesn't own the buffer by default. Return a copy.
                            return qimg.copy()
                        elif array.dtype == np.uint8:
                            bytes_per_line = width * 1
                            contiguous_array = np.ascontiguousarray(array)
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                            return qimg.copy()
                        elif np.issubdtype(array.dtype, np.floating):
                             # Assume float is in 0-1 range, scale to uint8
                             img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                             bytes_per_line = width * 1
                             contiguous_array = np.ascontiguousarray(img_norm)
                             qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                             return qimg.copy()
                        else:
                            raise TypeError(f"Unsupported grayscale NumPy dtype: {array.dtype}")
            
                    elif array.ndim == 3: # Color
                        height, width, channels = array.shape
                        if channels == 3 and array.dtype == np.uint8:
                            # Assume input is BGR (common from OpenCV), convert to RGB for QImage.Format_RGB888
                            # Make a contiguous copy before conversion
                            contiguous_array_bgr = np.ascontiguousarray(array)
                            rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                            # rgb_image is now contiguous RGB
                            bytes_per_line = width * 3
                            qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                            return qimg.copy()
                        elif channels == 4 and array.dtype == np.uint8:
                            # Assume input is BGRA (matches typical QImage.Format_ARGB32 memory layout)
                            contiguous_array_bgra = np.ascontiguousarray(array)
                            bytes_per_line = width * 4
                            qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                            return qimg.copy()
                        # Add handling for 16-bit color if needed (less common for display)
                        elif channels == 3 and array.dtype == np.uint16:
                             # Downscale to 8-bit for display format
                             array_8bit = (array / 257.0).astype(np.uint8)
                             contiguous_array_bgr = np.ascontiguousarray(array_8bit)
                             rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                             bytes_per_line = width * 3
                             qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                             return qimg.copy()
                        elif channels == 4 and array.dtype == np.uint16:
                             # Downscale color channels, keep alpha potentially?
                             array_8bit = (array / 257.0).astype(np.uint8) # Downscales all channels including alpha
                             contiguous_array_bgra = np.ascontiguousarray(array_8bit)
                             bytes_per_line = width * 4
                             qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                             return qimg.copy()
                        else:
                             raise TypeError(f"Unsupported color NumPy dtype/channel combination: {array.dtype} / {channels} channels")
                    else:
                        raise ValueError(f"Unsupported array dimension: {array.ndim}")
            
                except Exception as e:
                    traceback.print_exc()
                    return QImage() # Return invalid QImage on error

            def get_image_format(self, image=None):
                """Helper to safely get the format of self.image or a provided image."""
                img_to_check = image if image is not None else self.image
                if img_to_check and not img_to_check.isNull():
                    return img_to_check.format()
                return None

            def get_compatible_grayscale_format(self, image=None):
                """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
                current_format = self.get_image_format(image)
                if current_format == QImage.Format_Grayscale16:
                    return QImage.Format_Grayscale16
                else: # Treat 8-bit grayscale or color as needing 8-bit target
                    return QImage.Format_Grayscale8
            # --- END: Helper Functions ---
                
            def quadrilateral_to_rect(self, image, quad_points):
                """
                Warps the quadrilateral region defined by quad_points in the input image
                to a rectangular image using perspective transformation.
                Crucially, preserves the original bit depth (e.g., uint16) of the input image.
                """
                # --- Input Validation ---
                if not image or image.isNull():
                    QMessageBox.warning(self, "Warp Error", "Invalid input image provided for warping.")
                    return None
                if len(quad_points) != 4:
                     QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
                     return None
                if cv2 is None:
                     QMessageBox.critical(self, "Dependency Error", "OpenCV (cv2) is required for quadrilateral warping but is not installed.")
                     return None
             
                # --- Convert QImage to NumPy array (Preserves bit depth) ---
                try:
                    img_array = self.qimage_to_numpy(image) # Should return uint16 if input is Format_Grayscale16
                    if img_array is None: raise ValueError("NumPy Conversion returned None")
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
                    return None
             
                # --- !!! REMOVED INTERNAL GRAYSCALE CONVERSION !!! ---
                # # OLD code that caused the issue if input was color:
                # if img_array.ndim == 3:
                #      print("Warning: Warping input array is 3D. Converting to grayscale.")
                #      # THIS CONVERSION TO UINT8 WAS THE PROBLEM
                #      color_code = cv2.COLOR_BGR2GRAY if img_array.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
                #      img_array = cv2.cvtColor(img_array, color_code)
                #      # img_array = img_array.astype(np.uint8) # Explicit cast not needed, cvtColor usually returns uint8
                # --- End Removed Block ---
                # Now, warpPerspective will operate on the img_array with its original dtype (e.g., uint16 grayscale or color)
             
             
                # --- Transform points from LiveViewLabel space to Image space ---
                # (Coordinate transformation logic remains the same - assumes direct scaling for now)
                label_width = self.live_view_label.width()
                label_height = self.live_view_label.height()
                current_img_width = image.width()
                current_img_height = image.height()
                scale_x_disp = current_img_width / label_width if label_width > 0 else 1
                scale_y_disp = current_img_height / label_height if label_height > 0 else 1
             
                src_points_img = []
                for point in quad_points:
                    x_view, y_view = point.x(), point.y()
                    # Account for zoom/pan in LiveViewLabel coordinates
                    if self.live_view_label.zoom_level != 1.0:
                        x_view = (x_view - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                        y_view = (y_view - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
                    # Scale to image coordinates
                    x_image = x_view * scale_x_disp
                    y_image = y_view * scale_y_disp
                    src_points_img.append([x_image, y_image])
                src_np = np.array(src_points_img, dtype=np.float32)
             
                # --- Define Destination Rectangle ---
                # Calculate dimensions based on the source points
                width_a = np.linalg.norm(src_np[0] - src_np[1])
                width_b = np.linalg.norm(src_np[2] - src_np[3])
                height_a = np.linalg.norm(src_np[0] - src_np[3])
                height_b = np.linalg.norm(src_np[1] - src_np[2])
                max_width = int(max(width_a, width_b))
                max_height = int(max(height_a, height_b))
             
                if max_width <= 0 or max_height <= 0:
                     QMessageBox.warning(self, "Warp Error", f"Invalid destination rectangle size ({max_width}x{max_height}).")
                     return None
             
                dst_np = np.array([
                    [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
                ], dtype=np.float32)
             
                # --- Perform Perspective Warp using OpenCV ---
                try:
                    matrix = cv2.getPerspectiveTransform(src_np, dst_np)
                    # Determine border color based on input array type
                    if img_array.ndim == 3: # Color
                         # Use black (0,0,0) or white (255,255,255) depending on preference. Alpha needs 4 values if present.
                         border_val = (0, 0, 0, 0) if img_array.shape[2] == 4 else (0, 0, 0)
                    else: # Grayscale
                         border_val = 0 # Black for grayscale
             
                    # Warp the ORIGINAL NumPy array (could be uint8, uint16, color)
                    warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                                       flags=cv2.INTER_LINEAR, # Linear interpolation is usually good
                                                       borderMode=cv2.BORDER_CONSTANT,
                                                       borderValue=border_val) # Fill borders appropriately
                except Exception as e:
                     QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
                     traceback.print_exc() # Print full traceback for debugging
                     return None
             
                # --- Convert warped NumPy array back to QImage ---
                # numpy_to_qimage should handle uint16 correctly, creating Format_Grayscale16
                try:
                    warped_qimage = self.numpy_to_qimage(warped_array) # Handles uint8/uint16/color
                    if warped_qimage.isNull(): raise ValueError("numpy_to_qimage conversion failed.")
                    return warped_qimage
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array back to QImage: {e}")
                    return None
           # --- END: Modified Warping ---
            
            

                
            def create_menu_bar(self):
                menubar = self.menuBar()

                # --- File Menu ---
                file_menu = menubar.addMenu("&File")
                file_menu.addAction(self.load_action)
                file_menu.addAction(self.save_action)
                # file_menu.addAction(self.save_svg_action)
                file_menu.addSeparator()
                file_menu.addAction(self.reset_action)
                file_menu.addSeparator()
                file_menu.addAction(self.exit_action)

                # --- Edit Menu ---
                edit_menu = menubar.addMenu("&Edit")
                edit_menu.addAction(self.undo_action)
                edit_menu.addAction(self.redo_action)
                edit_menu.addSeparator()
                edit_menu.addAction(self.copy_action)
                edit_menu.addAction(self.paste_action)

                # --- View Menu ---
                view_menu = menubar.addMenu("&View")
                view_menu.addAction(self.zoom_in_action)
                view_menu.addAction(self.zoom_out_action)
                
                tools_menu = menubar.addMenu("&Tools")
                tools_menu.addAction(self.auto_lane_action)

                # --- About Menu ---
                about_menu = menubar.addMenu("&About")
                # Add "GitHub" action directly here or create it in _create_actions
                github_action = QAction("&GitHub", self)
                github_action.setToolTip("Open the project's GitHub page")
                github_action.triggered.connect(self.open_github)
                about_menu.addAction(github_action)
                
                # self.statusBar().showMessage("Ready")
            def open_github(self):
                # Open the GitHub link in the default web browser
                QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
            
            def create_tool_bar(self):
                """Create the main application toolbar."""
                self.tool_bar = QToolBar("Main Toolbar")
                self.tool_bar.setIconSize(QSize(30, 30))
                self.tool_bar.setMovable(False)
                self.tool_bar.setToolButtonStyle(Qt.ToolButtonIconOnly)

                # Add actions to the toolbar
                self.tool_bar.addAction(self.load_action)
                self.tool_bar.addAction(self.paste_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.save_action)
                self.tool_bar.addAction(self.copy_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.undo_action)
                self.tool_bar.addAction(self.redo_action)
                self.tool_bar.addSeparator()

                # --- Add Zoom and Pan Buttons ---
                self.tool_bar.addAction(self.zoom_in_action)
                self.tool_bar.addAction(self.zoom_out_action)
                self.tool_bar.addSeparator() # Optional separator

                self.tool_bar.addAction(self.pan_left_action)
                self.tool_bar.addAction(self.pan_up_action)   # Group visually
                self.tool_bar.addAction(self.pan_down_action)
                self.tool_bar.addAction(self.pan_right_action)
                # --- End Adding Pan Buttons ---

                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.auto_lane_action) # Add new button
                
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.draw_line_action) # Add the new action
                self.tool_bar.addAction(self.draw_bounding_box_action) # Add the new action
     
                
                # --- NEW: Add Copy/Paste Custom Items to Toolbar ---
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.copy_custom_items_action)
                self.tool_bar.addAction(self.paste_custom_items_action)
                # --- END NEW ---
                
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.reset_action)
                
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.info_action)

                # Add the toolbar to the main window
                self.addToolBar(Qt.TopToolBarArea, self.tool_bar)
                self.tool_bar.setContextMenuPolicy(Qt.PreventContextMenu)

                
            def start_auto_lane_marker(self):
                self._reset_live_view_label_custom_handlers()
                """Initiates the automatic lane marker placement process, allowing user to choose region type."""
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    return

                # --- 1. Ask User for Marker Side ---
                items_side = ["Left", "Right"]
                side, ok_side = QInputDialog.getItem(self, "Select Marker Side",
                                                     "Place markers on which side?", items_side, 0, False)
                if not ok_side or not side:
                    return  # User cancelled
                self.auto_marker_side = side.lower()

                # --- 2. Ask User for Region Definition Type ---
                items_region = ["Rectangle (for straight lanes)", "Quadrilateral (for skewed lanes)"]
                region_type_str, ok_region = QInputDialog.getItem(self, "Select Region Type",
                                                                 "How do you want to define the lane region?",
                                                                 items_region, 0, False)
                if not ok_region or not region_type_str:
                    return # User cancelled

                self.live_view_label.quad_points = [] # Clear any previous quad points
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True) # Ensure tracking is on

                if "Rectangle" in region_type_str:
                    self.live_view_label.mode = 'auto_lane_rect' # For paintEvent
                    self.live_view_label._custom_left_click_handler_from_app = self.start_rectangle 
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.update_rectangle_preview 
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_rectangle_for_auto_lane
                elif "Quadrilateral" in region_type_str:
                    self.live_view_label.mode = 'auto_lane_quad' # For paintEvent
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_auto_lane_quad_click

                self.update_live_view()

            def handle_auto_lane_quad_click(self, event):
                if self.live_view_label.mode != 'auto_lane_quad':
                    return
    
                if event.button() == Qt.LeftButton:
                    point_label_space_transformed = self.live_view_label.transform_point(event.position())
                    snapped_point_label_space = self.snap_point_to_grid(point_label_space_transformed) # Snap it
                    
                    self.live_view_label.quad_points.append(snapped_point_label_space)
                    self.update_live_view()
    
                    if len(self.live_view_label.quad_points) == 4:
                        self.finalize_quad_for_auto_lane(self.live_view_label.quad_points)

            def finalize_quad_for_auto_lane(self, quad_points_label_space):
                """
                Finalizes the quadrilateral using points from "label space",
                warps the corresponding region from self.image, and proceeds to processing.
                quad_points_label_space: List of 4 QPointF in "label space".
                """
                if len(quad_points_label_space) != 4:
                    QMessageBox.warning(self, "Error", "Quadrilateral definition incomplete for auto lane.")
                    self.live_view_label.mode = None # Reset mode
                    self.live_view_label.quad_points = []
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self._reset_live_view_label_custom_handlers() # Reset after finalization
                    self.live_view_label.mode = None # Reset LiveViewLabel's internal mode flag
                    self.update_live_view()
                    return
                
                warped_qimage_region = self.quadrilateral_to_rect(self.image, quad_points_label_space)

                if warped_qimage_region and not warped_qimage_region.isNull():
                    quad_points_image_space_for_marker_placement = []
                    try:
                        native_img_w = float(self.image.width())
                        native_img_h = float(self.image.height())
                        label_w_widget = float(self.live_view_label.width())
                        label_h_widget = float(self.live_view_label.height())
                        if native_img_w <= 0 or native_img_h <= 0 or label_w_widget <= 0 or label_h_widget <= 0:
                            raise ValueError("Invalid image_to_warp or label dimensions.")
                        scale_native_to_label = min(label_w_widget / native_img_w, label_h_widget / native_img_h)
                        displayed_w_in_label = native_img_w * scale_native_to_label
                        displayed_h_in_label = native_img_h * scale_native_to_label
                        offset_x_centering = (label_w_widget - displayed_w_in_label) / 2.0
                        offset_y_centering = (label_h_widget - displayed_h_in_label) / 2.0
                        for p_label in quad_points_label_space:
                            x_rel_display = p_label.x() - offset_x_centering
                            y_rel_display = p_label.y() - offset_y_centering
                            if scale_native_to_label < 1e-9: raise ValueError("Scale factor too small.")
                            img_x = x_rel_display / scale_native_to_label
                            img_y = y_rel_display / scale_native_to_label
                            quad_points_image_space_for_marker_placement.append(QPointF(img_x, img_y))
                        
                        # print(f"DEBUG AutoLaneQuad (finalize): Mapped image space points for marker placement: {[(p.x(), p.y()) for p in quad_points_image_space_for_marker_placement]}")

                        self.process_auto_lane_region(warped_qimage_region, 
                                                      quad_points_image_space_for_marker_placement, # Pass image-space points
                                                      is_quad_warp=True)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to map quad points for marker placement: {e}")
                        traceback.print_exc()
                else:
                    QMessageBox.warning(self, "Error", "Failed to warp quadrilateral region for auto lane.")

                # Reset state after processing or error
                self.live_view_label.mode = None
                self.live_view_label.quad_points = [] # Clear points from label
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self._reset_live_view_label_custom_handlers() # Important to reset
                self.update_live_view()


            def finalize_rectangle_for_auto_lane(self, event):
                if self.live_view_label.mode != 'auto_lane_rect' or not self.live_view_label.rectangle_start:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event)
                    return
    
                if event.button() == Qt.LeftButton:
                    end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
                    self.live_view_label.rectangle_end = snapped_end_point # Store snapped end point
    
                    rect_coords_img = None
                    try:
                        # self.live_view_label.rectangle_start is already snapped from start_rectangle
                        start_x_view, start_y_view = self.live_view_label.rectangle_start.x(), self.live_view_label.rectangle_start.y()
                        end_x_view, end_y_view = self.live_view_label.rectangle_end.x(), self.live_view_label.rectangle_end.y() # Use snapped end
    
                        # ... (rest of coordinate transformation logic remains the same) ...
                        zoom = self.live_view_label.zoom_level
                        # ... (as in your existing finalize_rectangle_for_auto_lane)
                        offset_x_pan, offset_y_pan = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                        start_x_unzoomed = (start_x_view - offset_x_pan) / zoom
                        start_y_unzoomed = (start_y_view - offset_y_pan) / zoom
                        end_x_unzoomed = (end_x_view - offset_x_pan) / zoom
                        end_y_unzoomed = (end_y_view - offset_y_pan) / zoom
    
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w, img_h = self.image.width(), self.image.height()
                        label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                        if img_w <= 0 or img_h <=0 or label_w <=0 or label_h <=0:
                            raise ValueError("Invalid image or label dimensions for coord conversion.")
    
                        scale_factor = min(label_w / img_w, label_h / img_h)
                        display_offset_x = (label_w - img_w * scale_factor) / 2
                        display_offset_y = (label_h - img_h * scale_factor) / 2
    
                        start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                        start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                        end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                        end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
    
                        rect_x = int(min(start_x_img, end_x_img))
                        rect_y = int(min(start_y_img, end_y_img))
                        rect_w = int(abs(end_x_img - start_x_img))
                        rect_h = int(abs(end_y_img - start_y_img))
    
                        rect_x = max(0, rect_x)
                        rect_y = max(0, rect_y)
                        rect_w = max(1, min(rect_w, img_w - rect_x))
                        rect_h = max(1, min(rect_h, img_h - rect_y))
                        rect_coords_img = (rect_x, rect_y, rect_w, rect_h)
                        # ...
    
                        extracted_qimage_region = self.image.copy(rect_x, rect_y, rect_w, rect_h)
                        if extracted_qimage_region.isNull():
                            raise ValueError("QImage copy failed for rectangle.")
    
                        self.process_auto_lane_region(extracted_qimage_region, rect_coords_img, is_quad_warp=False)
    
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to finalize rectangle for auto lane: {e}")
                        traceback.print_exc()
                    finally:
                        self._reset_live_view_label_custom_handlers() # Reset after finalization
                        self.live_view_label.mode = None # Reset LiveViewLabel's internal mode flag
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        self.live_view_label.bounding_box_preview = None
                        self.live_view_label.rectangle_start = None
                        self.update_live_view()

            def process_auto_lane_region(self, qimage_region, original_region_definition, is_quad_warp):
                """
                Processes the extracted/warped region, opens tuning dialog, and places markers.
                qimage_region: The QImage (rectangular) to be analyzed.
                original_region_definition: For rect, tuple (x,y,w,h) in original image space.
                                            For quad, list of 4 QPointF in original image space.
                is_quad_warp: Boolean indicating if qimage_region came from a warped quadrilateral.
                """
                if qimage_region.isNull():
                    QMessageBox.warning(self, "Error", "No valid image region provided for auto lane processing.")
                    return

                dialog_image_pil = self.convert_qimage_to_grayscale_pil(qimage_region)
                if not dialog_image_pil:
                    QMessageBox.warning(self, "Error", "Failed to convert extracted lane to grayscale for analysis.")
                    return

                # Store height AND WIDTH of the image passed to the dialog for later mapping
                dialog_image_height = dialog_image_pil.height
                dialog_image_width = dialog_image_pil.width # <<< --- ADDED

                dialog = AutoLaneTuneDialog(
                    pil_image_data=dialog_image_pil,
                    initial_settings=self.peak_dialog_settings,
                    parent=self,
                    is_from_quad_warp=is_quad_warp
                )

                if dialog.exec() == QDialog.Accepted:
                    detected_peak_coords_dialog = dialog.get_detected_peaks()
                    final_settings = dialog.get_final_settings()

                    if self.persist_peak_settings_enabled:
                        self.peak_dialog_settings.update(final_settings)

                    if detected_peak_coords_dialog is not None and len(detected_peak_coords_dialog) > 0:
                        self.place_markers_from_dialog(
                            original_region_definition,
                            detected_peak_coords_dialog,
                            self.auto_marker_side,
                            dialog_image_height,
                            dialog_image_width, # <<< --- PASS WIDTH
                            is_quad_warp
                        )
                    else:
                        QMessageBox.information(self, "Auto Lane", "No peaks were detected with the current settings.")
                else:
                    print("Automatic lane marker tuning cancelled.")

                self.live_view_label.bounding_box_preview = None
                self.live_view_label.quad_points = []
                self.update_live_view()


            def place_markers_from_dialog(self, original_region_definition, peak_coords_in_dialog, side,
                                          dialog_image_height, dialog_image_width, 
                                          is_from_quad_warp):
                if not peak_coords_in_dialog.any() or dialog_image_height <= 0 or dialog_image_width <= 0:
                    print("No peak coordinates or invalid dialog image dimensions for auto lane marker placement.")
                    return

                self.save_state()
                target_marker_list = None
                if side == 'left':
                    self.left_markers.clear(); self.current_left_marker_index = 0
                    target_marker_list = self.left_markers
                elif side == 'right':
                    self.right_markers.clear(); self.current_right_marker_index = 0
                    target_marker_list = self.right_markers
                else:
                    print(f"Invalid side '{side}' for auto lane marker placement."); return

                self.on_combobox_changed() 
                current_marker_values_for_labels = self.marker_values 

                if not current_marker_values_for_labels: 
                    current_marker_values_for_labels = [""] * len(peak_coords_in_dialog)
                elif len(current_marker_values_for_labels) < len(peak_coords_in_dialog):
                    current_marker_values_for_labels.extend([""] * (len(peak_coords_in_dialog) - len(current_marker_values_for_labels)))

                sorted_peaks_dialog = np.sort(peak_coords_in_dialog.astype(float)) 
                
                inv_perspective_matrix = None
                src_points_for_transform_np = None

                if is_from_quad_warp:
                    if len(original_region_definition) != 4:
                        print("Error: Quadrilateral definition requires 4 points for inverse mapping in auto lane.")
                        return
                    src_points_for_transform_list = []
                    for p in original_region_definition:
                        try: src_points_for_transform_list.append([float(p.x()), float(p.y())])
                        except AttributeError: src_points_for_transform_list.append([float(p[0]), float(p[1])])
                    src_points_for_transform_np = np.array(src_points_for_transform_list, dtype=np.float32)
                    
                    dst_points_for_dialog_transform_np = np.array([
                        [0.0, 0.0],
                        [float(dialog_image_width) - 1.0, 0.0],
                        [float(dialog_image_width) - 1.0, float(dialog_image_height) - 1.0],
                        [0.0, float(dialog_image_height) - 1.0]
                    ], dtype=np.float32)
                    try:
                        inv_perspective_matrix = cv2.getPerspectiveTransform(dst_points_for_dialog_transform_np, src_points_for_transform_np)
                    except Exception as e:
                        print(f"Error calculating inverse perspective matrix for auto lane: {e}")
                        return

                for i, peak_y_dialog in enumerate(sorted_peaks_dialog):
                    label = str(current_marker_values_for_labels[i]) if i < len(current_marker_values_for_labels) else ""
                    y_final_img_calc = 0.0
                    
                    if not is_from_quad_warp:
                        rect_x_img, rect_y_img, rect_w_img, rect_h_img = map(float, original_region_definition)
                        t = peak_y_dialog / (float(dialog_image_height) - 1.0) if dialog_image_height > 1 else 0.5
                        t = max(0.0, min(1.0, t))
                        y_final_img_calc = rect_y_img + t * rect_h_img
                    else:
                        if inv_perspective_matrix is None: continue
                        x_in_dialog_space = float(dialog_image_width) * 0.05 if side == 'left' else float(dialog_image_width) * 0.95
                        point_in_dialog_np = np.array([[[x_in_dialog_space, peak_y_dialog]]], dtype=np.float32)
                        try:
                            original_image_point_np = cv2.perspectiveTransform(point_in_dialog_np, inv_perspective_matrix)
                            y_final_img_calc = float(original_image_point_np[0,0,1])
                        except Exception as e_transform:
                            print(f"Error transforming point back for auto lane: {e_transform}"); continue
                    
                    final_y_to_store = float(y_final_img_calc)
                    target_marker_list.append((final_y_to_store, label))
                
                # --- MODIFICATION START (with pixel-perfect refinement) ---
                # Override X position logic to use detected content borders instead of region geometry.
                left_border, right_border = self._find_image_content_borders()
                
                target_x_in_image_space_for_slider = 0.0

                if side == 'left':
                    if left_border is not None:
                        # For the left marker, the text "value ⎯" is drawn to the left of the anchor.
                        # Setting the anchor exactly at the border ensures the line sits right on it.
                        target_x_in_image_space_for_slider = left_border
                        print(f"INFO: Auto lane marker X position set to detected left content border: {left_border}")
                    else:
                        print("WARNING: Could not detect image content border. Falling back to edge.")
                        target_x_in_image_space_for_slider = 0 
                elif side == 'right':
                    if right_border is not None:
                        # For the right marker, the text "⎯ value" is drawn starting at the anchor.
                        # Setting the anchor to the border + 1 ensures the line starts just outside the content.
                        target_x_in_image_space_for_slider = right_border + 1
                        print(f"INFO: Auto lane marker X position set to detected right content border + 1: {right_border + 1}")
                    else:
                        print("WARNING: Could not detect image content border. Falling back to edge.")
                        target_x_in_image_space_for_slider = self.image.width() - 1 if self.image else 0
                # --- MODIFICATION END ---
                
                slider_target_value_native_pixels = int(round(target_x_in_image_space_for_slider))
                self._update_marker_slider_ranges()

                if side == 'left':
                    if hasattr(self, 'left_padding_slider'):
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(
                            max(self.left_slider_range[0], min(slider_target_value_native_pixels, self.left_slider_range[1]))
                        )
                        self.left_padding_slider.blockSignals(False)
                        self.left_marker_shift_added = self.left_padding_slider.value()
                elif side == 'right':
                     if hasattr(self, 'right_padding_slider'):
                         self.right_padding_slider.blockSignals(True)
                         self.right_padding_slider.setValue(
                             max(self.right_slider_range[0], min(slider_target_value_native_pixels, self.right_slider_range[1]))
                         )
                         self.right_padding_slider.blockSignals(False)
                         self.right_marker_shift_added = self.right_padding_slider.value()

                self.is_modified = True
                self.update_live_view()
                
            def zoom_in(self):
                self.live_view_label.zoom_in()

            def zoom_out(self):
                self.live_view_label.zoom_out()
                
            
            def enable_standard_protein_mode(self):
                """"Enable mode to define standard protein amounts for creating a standard curve."""
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                # Assign mouse event handlers for bounding box creation
                self._reset_live_view_label_custom_handlers() # Good practice to reset first
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.start_bounding_box(event)
                self.live_view_label._custom_mouseReleaseEvent_from_app = lambda event: self.end_standard_bounding_box(event)
            
            def enable_measure_protein_mode(self):
                """Enable mode to measure protein quantity using the standard curve."""
                if len(self.quantities) < 2:
                    QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                self._reset_live_view_label_custom_handlers() # Good practice
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.start_bounding_box(event)
                self.live_view_label._custom_mouseReleaseEvent_from_app = lambda event: self.end_measure_bounding_box(event)

            def call_live_view(self):
                self.update_live_view()      

            def analyze_bounding_box(self, pil_image_for_dialog, standard):
                peak_info_result = None # Will hold list of dicts
                self.latest_peak_areas = [] # Clear for single lane mode
                self.latest_peak_details = [] # Clear for single lane mode
                self.latest_calculated_quantities = []

                if standard:
                    quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
                    if ok and quantity:
                        try:
                            quantity_value = float(quantity.split()[0])
                            peak_info_result = self.calculate_peak_area(pil_image_for_dialog)
                            
                            if peak_info_result and len(peak_info_result) > 0:
                                areas_for_standard = [info['area'] for info in peak_info_result]
                                total_area = sum(areas_for_standard)
                                self.quantities_peak_area_dict[quantity_value] = round(total_area, 3)
                                self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                                self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
                                print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                                self.latest_peak_areas = [round(a, 3) for a in areas_for_standard]
                                self.latest_peak_details = peak_info_result # Store full details
                            else:
                                 print("Peak area calculation cancelled or failed for standard.")
                        except (ValueError, IndexError) as e: # ... (error handling)
                            QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                        except Exception as e: # ...
                             QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                    else:
                        print("Standard quantity input cancelled.")
                else: # Analyze sample
                    peak_info_result = self.calculate_peak_area(pil_image_for_dialog)
                    if peak_info_result and len(peak_info_result) > 0:
                        self.latest_peak_areas = [round(info['area'], 3) for info in peak_info_result]
                        self.latest_peak_details = peak_info_result # Store full details
                        print(f"Sample Analysis: Calculated Peak Info = {self.latest_peak_details}")

                        if len(self.quantities_peak_area_dict) >= 2:
                            self.latest_calculated_quantities = self.calculate_unknown_quantity(
                                list(self.quantities_peak_area_dict.values()),
                                list(self.quantities_peak_area_dict.keys()),
                                self.latest_peak_areas 
                            )
                            print(f"Sample Analysis: Calculated Quantities = {self.latest_calculated_quantities}")
                        else: self.latest_calculated_quantities = []
                        try: self.target_protein_areas_text.setText(str(self.latest_peak_areas))
                        except Exception as e: print(f"Error displaying sample areas: {e}"); self.target_protein_areas_text.setText("Error")
                    else:
                         print("Peak area calculation cancelled or failed for sample.")
                         self.target_protein_areas_text.setText("N/A")
                self.update_live_view()
            
            def calculate_peak_area(self, pil_image_for_dialog):
                if pil_image_for_dialog is None: # ... (existing validation) ...
                    print("Error: No PIL Image data provided to calculate_peak_area.")
                    return None # Return None if no areas/info
                if not isinstance(pil_image_for_dialog, Image.Image):
                     QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
                     return None

                dialog = PeakAreaDialog(
                    cropped_data=pil_image_for_dialog,
                    current_settings=self.peak_dialog_settings,
                    persist_checked=self.persist_peak_settings_enabled,
                    parent=self
                )

                peak_info_list = None # Will be list of dicts
                if dialog.exec() == QDialog.Accepted:
                    peak_info_list = dialog.get_final_peak_info() # Get the detailed info
                    if dialog.should_persist_settings():
                        self.peak_dialog_settings = dialog.get_current_settings()
                        self.persist_peak_settings_enabled = True
                    else:
                        self.persist_peak_settings_enabled = False
                else:
                    print("PeakAreaDialog cancelled.")
                
                return peak_info_list if peak_info_list is not None else [] # Return list of dicts or empty list
            
            
            def calculate_unknown_quantity(self, standard_total_areas, known_quantities, sample_peak_areas):
                """Calculates unknown quantities based on standards and sample areas. Returns a list of quantities."""
                if not standard_total_areas or not known_quantities or len(standard_total_areas) != len(known_quantities):
                    print("Error: Invalid standard data for quantity calculation.")
                    return []
                if len(standard_total_areas) < 2:
                    print("Error: Need at least 2 standards for regression.")
                    return []
                if not sample_peak_areas:
                     print("Warning: No sample peak areas provided for quantity calculation.")
                     return []

                calculated_quantities = []
                try:
                    # Use total standard area vs known quantity for the standard curve
                    coefficients = np.polyfit(standard_total_areas, known_quantities, 1)

                    # Apply the standard curve to *each individual* sample peak area
                    for area in sample_peak_areas:
                        val = np.polyval(coefficients, area)
                        calculated_quantities.append(round(val, 2))

                    # Display the results in a message box (optional, but helpful feedback)
                    # QMessageBox.information(self, "Protein Quantification", f"Predicted Quantities: {calculated_quantities} units")

                except Exception as e:
                     # QMessageBox.warning(self, "Calculation Error", f"Could not calculate quantities: {e}")
                     return [] # Return empty list on error

                return calculated_quantities # <-- Return the list

                
            def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
                """Draw quantity text at the correct position."""
                text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
                painter.drawText(text_position, str(quantity))
            
            def update_standard_protein_quantities(self):
                self.standard_protein_values.text()
            
            def move_tab(self,tab):
                self.tab_widget.setCurrentIndex(tab)
                
            def save_state(self):
                """Save the current state of the image, markers, shapes, and relevant UI/font settings."""
                # Get current state of UI elements related to fonts/colors for custom markers
                custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                state = {
                    "image": self.image.copy() if self.image else None,
                    "left_markers": self.left_markers.copy(),
                    "right_markers": self.right_markers.copy(),
                    "top_markers": self.top_markers.copy(),
                    "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                    "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                    "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                    "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                    "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                    # Standard marker font settings
                    "font_family": self.font_family,
                    "font_size": self.font_size,
                    "font_color": self.font_color, # QColor object is fine for in-memory stack
                    "font_rotation": self.font_rotation,
                    # Marker positioning shifts
                    "left_marker_shift_added": self.left_marker_shift_added,
                    "right_marker_shift_added": self.right_marker_shift_added,
                    "top_marker_shift_added": self.top_marker_shift_added,
                    # Analysis data (keep as before)
                    "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                    "quantities": self.quantities.copy(),
                    "protein_quantities": self.protein_quantities.copy(),
                    "standard_protein_areas": self.standard_protein_areas.copy(),
                    # --- NEW: Custom marker font/color states ---
                    "custom_marker_color": self.custom_marker_color, # QColor object
                    "custom_font_family": custom_font_family, # String from dropdown
                    "custom_font_size": custom_font_size,     # Integer from spinbox
                    # --- END NEW ---
                }
                self.undo_stack.append(state)
                self.redo_stack.clear() # Clear redo stack on new action
            
            def undo_action_m(self):
                """Undo the last action by restoring the previous state, including font settings."""
                if self.undo_stack:
                    # Get current state of UI elements before overwriting self attributes
                    current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                    current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                    # Save the current state to the redo stack
                    current_state = {
                        "image": self.image.copy() if self.image else None,
                        "left_markers": self.left_markers.copy(),
                        "right_markers": self.right_markers.copy(),
                        "top_markers": self.top_markers.copy(),
                        "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                        "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                        "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                        "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                        "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                        # Standard marker font settings
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_color": self.font_color,
                        "font_rotation": self.font_rotation,
                        # Marker positioning shifts
                        "left_marker_shift_added": self.left_marker_shift_added,
                        "right_marker_shift_added": self.right_marker_shift_added,
                        "top_marker_shift_added": self.top_marker_shift_added,
                        # Analysis data
                        "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                        "quantities": self.quantities.copy(),
                        "protein_quantities": self.protein_quantities.copy(),
                        "standard_protein_areas": self.standard_protein_areas.copy(),
                        # --- NEW: Custom marker font/color states ---
                        "custom_marker_color": self.custom_marker_color,
                        "custom_font_family": current_custom_font_family, # Use value read from UI
                        "custom_font_size": current_custom_font_size,     # Use value read from UI
                        # --- END NEW ---
                    }
                    self.redo_stack.append(current_state)

                    # Restore the previous state from the undo stack
                    previous_state = self.undo_stack.pop()
                    self.image = previous_state["image"]
                    self.left_markers = previous_state["left_markers"]
                    self.right_markers = previous_state["right_markers"]
                    self.top_markers = previous_state["top_markers"]
                    self.custom_markers = previous_state.get("custom_markers", [])
                    self.custom_shapes = previous_state.get("custom_shapes", [])
                    self.image_before_padding = previous_state["image_before_padding"]
                    self.image_contrasted = previous_state["image_contrasted"]
                    self.image_before_contrast = previous_state["image_before_contrast"]
                    # Restore standard font settings
                    self.font_family = previous_state["font_family"]
                    self.font_size = previous_state["font_size"]
                    self.font_color = previous_state["font_color"]
                    self.font_rotation = previous_state["font_rotation"]
                    # Restore shifts
                    self.left_marker_shift_added = previous_state["left_marker_shift_added"]
                    self.right_marker_shift_added = previous_state["right_marker_shift_added"]
                    self.top_marker_shift_added = previous_state["top_marker_shift_added"]
                    # Restore analysis data
                    self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
                    self.quantities = previous_state["quantities"]
                    self.protein_quantities = previous_state["protein_quantities"]
                    self.standard_protein_areas = previous_state["standard_protein_areas"]
                    # --- NEW: Restore custom marker font/color states ---
                    self.custom_marker_color = previous_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
                    restored_custom_font_family = previous_state.get("custom_font_family", "Arial")
                    restored_custom_font_size = previous_state.get("custom_font_size", 12)
                    # --- END NEW RESTORE ---

                    # --- Update UI Elements to reflect restored state ---
                    # Standard font UI
                    if hasattr(self, 'font_combo_box'):
                        self.font_combo_box.blockSignals(True)
                        # Find the font; setCurrentText might be less reliable than setCurrentFont
                        found_font = False
                        for i in range(self.font_combo_box.count()):
                            if self.font_combo_box.itemText(i) == self.font_family:
                                self.font_combo_box.setCurrentIndex(i)
                                found_font = True
                                break
                        if not found_font: # Fallback if font name isn't exact match
                            self.font_combo_box.setCurrentFont(QFont(self.font_family))
                        self.font_combo_box.blockSignals(False)
                    if hasattr(self, 'font_size_spinner'):
                        self.font_size_spinner.blockSignals(True)
                        self.font_size_spinner.setValue(self.font_size)
                        self.font_size_spinner.blockSignals(False)
                    if hasattr(self, 'font_color_button'):
                        self._update_color_button_style(self.font_color_button, self.font_color)
                    if hasattr(self, 'font_rotation_input'):
                        self.font_rotation_input.blockSignals(True)
                        self.font_rotation_input.setValue(self.font_rotation)
                        self.font_rotation_input.blockSignals(False)

                    # Slider positions (reflecting restored shifts)
                    if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
                    if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
                    if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

                    # --- NEW: Update Custom Marker UI ---
                    if hasattr(self, 'custom_marker_color_button'):
                        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                    if hasattr(self, 'custom_font_type_dropdown'):
                         self.custom_font_type_dropdown.blockSignals(True)
                         # Find the font like standard font combo
                         found_custom_font = False
                         for i in range(self.custom_font_type_dropdown.count()):
                             if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                                 self.custom_font_type_dropdown.setCurrentIndex(i)
                                 found_custom_font = True
                                 break
                         if not found_custom_font: # Fallback
                             self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                         self.custom_font_type_dropdown.blockSignals(False)
                         # Also update the entry box font if needed (optional)
                         # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont())
                    if hasattr(self, 'custom_font_size_spinbox'):
                         self.custom_font_size_spinbox.blockSignals(True)
                         self.custom_font_size_spinbox.setValue(restored_custom_font_size)
                         self.custom_font_size_spinbox.blockSignals(False)
                    # --- END NEW UI UPDATE ---

                    # Update other UI if needed (e.g., analysis text boxes)

                    # Refresh display and status bar
                    try:
                        self._update_preview_label_size()
                    except Exception: pass # Ignore errors if label size fails temporarily
                    self._update_status_bar()
                    self.update_live_view()
                    self.is_modified = True # Undoing makes it modified again relative to last save
                    
            
            def redo_action_m(self):
                """Redo the last undone action by restoring the next state, including font settings."""
                if self.redo_stack:
                    # Get current state of UI elements before overwriting self attributes
                    current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                    current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                    # Save the current state to the undo stack
                    current_state = {
                        "image": self.image.copy() if self.image else None,
                        "left_markers": self.left_markers.copy(),
                        "right_markers": self.right_markers.copy(),
                        "top_markers": self.top_markers.copy(),
                        "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                        "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                        "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                        "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                        "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                        # Standard marker font settings
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_color": self.font_color,
                        "font_rotation": self.font_rotation,
                        # Marker positioning shifts
                        "left_marker_shift_added": self.left_marker_shift_added,
                        "right_marker_shift_added": self.right_marker_shift_added,
                        "top_marker_shift_added": self.top_marker_shift_added,
                        # Analysis data
                        "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                        "quantities": self.quantities.copy(),
                        "protein_quantities": self.protein_quantities.copy(),
                        "standard_protein_areas": self.standard_protein_areas.copy(),
                        # --- NEW: Custom marker font/color states ---
                        "custom_marker_color": self.custom_marker_color,
                        "custom_font_family": current_custom_font_family, # Use value read from UI
                        "custom_font_size": current_custom_font_size,     # Use value read from UI
                        # --- END NEW ---
                    }
                    self.undo_stack.append(current_state)

                    # Restore the next state from the redo stack
                    next_state = self.redo_stack.pop()
                    self.image = next_state["image"]
                    self.left_markers = next_state["left_markers"]
                    self.right_markers = next_state["right_markers"]
                    self.top_markers = next_state["top_markers"]
                    self.custom_markers = next_state.get("custom_markers", [])
                    self.custom_shapes = next_state.get("custom_shapes", [])
                    self.image_before_padding = next_state["image_before_padding"]
                    self.image_contrasted = next_state["image_contrasted"]
                    self.image_before_contrast = next_state["image_before_contrast"]
                    # Restore standard font settings
                    self.font_family = next_state["font_family"]
                    self.font_size = next_state["font_size"]
                    self.font_color = next_state["font_color"]
                    self.font_rotation = next_state["font_rotation"]
                    # Restore shifts
                    self.left_marker_shift_added = next_state["left_marker_shift_added"]
                    self.right_marker_shift_added = next_state["right_marker_shift_added"]
                    self.top_marker_shift_added = next_state["top_marker_shift_added"]
                    # Restore analysis data
                    self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
                    self.quantities = next_state["quantities"]
                    self.protein_quantities = next_state["protein_quantities"]
                    self.standard_protein_areas = next_state["standard_protein_areas"]
                    # --- NEW: Restore custom marker font/color states ---
                    self.custom_marker_color = next_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
                    restored_custom_font_family = next_state.get("custom_font_family", "Arial")
                    restored_custom_font_size = next_state.get("custom_font_size", 12)
                    # --- END NEW RESTORE ---

                    # --- Update UI Elements to reflect restored state ---
                    # (Identical UI update logic as in undo_action_m)
                    # Standard font UI
                    if hasattr(self, 'font_combo_box'):
                        self.font_combo_box.blockSignals(True)
                        found_font = False
                        for i in range(self.font_combo_box.count()):
                            if self.font_combo_box.itemText(i) == self.font_family:
                                self.font_combo_box.setCurrentIndex(i)
                                found_font = True
                                break
                        if not found_font: self.font_combo_box.setCurrentFont(QFont(self.font_family))
                        self.font_combo_box.blockSignals(False)
                    if hasattr(self, 'font_size_spinner'):
                        self.font_size_spinner.blockSignals(True); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.blockSignals(False)
                    if hasattr(self, 'font_color_button'):
                        self._update_color_button_style(self.font_color_button, self.font_color)
                    if hasattr(self, 'font_rotation_input'):
                        self.font_rotation_input.blockSignals(True); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.blockSignals(False)

                    # Slider positions
                    if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
                    if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
                    if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

                    # --- NEW: Update Custom Marker UI ---
                    if hasattr(self, 'custom_marker_color_button'):
                        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                    if hasattr(self, 'custom_font_type_dropdown'):
                         self.custom_font_type_dropdown.blockSignals(True)
                         found_custom_font = False
                         for i in range(self.custom_font_type_dropdown.count()):
                             if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                                 self.custom_font_type_dropdown.setCurrentIndex(i)
                                 found_custom_font = True
                                 break
                         if not found_custom_font: self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                         self.custom_font_type_dropdown.blockSignals(False)
                         # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont()) # Optional
                    if hasattr(self, 'custom_font_size_spinbox'):
                         self.custom_font_size_spinbox.blockSignals(True); self.custom_font_size_spinbox.setValue(restored_custom_font_size); self.custom_font_size_spinbox.blockSignals(False)
                    # --- END NEW UI UPDATE ---

                    # Update other UI if needed

                    # Refresh display and status bar
                    try:
                        self._update_preview_label_size()
                    except Exception: pass
                    self._update_status_bar()
                    self.update_live_view()
                    self.is_modified = True # Redoing makes it modified again relative to last save
                    
            def analysis_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                # layout.setSpacing(15) # Increase spacing in this tab

                # --- Molecular Weight Prediction ---
                mw_group = QGroupBox("Molecular Weight Prediction")
                mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                mw_layout = QHBoxLayout(mw_group) # Changed to QVBoxLayout for easier stacking
                # mw_layout.setSpacing(8)

                # --- NEW: Regression Model Selection ---
                mw_model_layout = QHBoxLayout()
                mw_model_layout.addWidget(QLabel("Regression Model:"))
                self.mw_regression_model_combo = QComboBox()
                self.mw_regression_model_combo.addItems([
                    "Log-Linear (Degree 1)",
                    "Log-Polynomial (Degree 2)",
                    "Log-Polynomial (Degree 3)"
                ])
                self.mw_regression_model_combo.setToolTip(
                    "Select the regression model for MW prediction.\n"
                    "Degree 1: Standard log-linear fit.\n"
                    "Degree 2/3: Polynomial fit on log(MW) vs. migration distance.\n"
                    "Higher degrees may better suit gradient gels or non-linear migration patterns."
                )
                mw_model_layout.addWidget(self.mw_regression_model_combo)
                self.predict_button = QPushButton("Predict Molecular Weight")
                self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
                self.predict_button.setEnabled(False)  # Initially disabled
                self.predict_button.clicked.connect(self.predict_molecular_weight)
                mw_layout.addLayout(mw_model_layout) # Add this layout to the main mw_layout
                mw_layout.addWidget(self.predict_button)
                
                layout.addWidget(mw_group)

                # --- Peak Area / Sample Quantification ---
                quant_group = QGroupBox("Peak Area and Sample Quantification")
                quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                quant_layout = QVBoxLayout(quant_group)
                # quant_layout.setSpacing(8)

                # Area Definition Buttons (Single Lane)
                area_def_layout = QHBoxLayout()
                self.btn_define_quad = QPushButton("Define Single Quad Area")
                self.btn_define_quad.setToolTip(
                    "Click 4 corner points to define a region. \n"
                    "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
                )
                self.btn_define_quad.clicked.connect(self.enable_quad_mode)
                self.btn_define_rec = QPushButton("Define Single Rectangle Area")
                self.btn_define_rec.setToolTip(
                    "Click and drag to define a rectangular region. \n"
                    "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
                )
                self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
                self.btn_sel_rec = QPushButton("Move Selected Area")
                self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
                self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
                
                self.btn_analyze_multiple_lanes = QPushButton("Define Multiple Lanes")
                self.btn_analyze_multiple_lanes.setToolTip("Define multiple lanes (quads or rectangles) sequentially.")
                self.btn_analyze_multiple_lanes.clicked.connect(self.start_analyze_multiple_lanes)
                
                self.btn_finish_multi_lane_def = QPushButton("Finish Defining Multiple Lanes")
                self.btn_finish_multi_lane_def.setToolTip("Signal that all multiple lanes have been defined and process them.")
                self.btn_finish_multi_lane_def.clicked.connect(self.finish_multi_lane_definition_and_process)
                self.btn_finish_multi_lane_def.setEnabled(False) # Enable when multi_lane_mode_active
                
                area_def_layout.addWidget(self.btn_define_quad)
                area_def_layout.addWidget(self.btn_define_rec)
                area_def_layout.addWidget(self.btn_sel_rec)
                area_def_layout.addWidget(self.btn_analyze_multiple_lanes)
                area_def_layout.addWidget(self.btn_finish_multi_lane_def)
                
                
                quant_layout.addLayout(area_def_layout)
                quant_layout.addWidget(self.create_separator())
                # --- END NEW ---


                # Standard Processing
                std_proc_layout = QHBoxLayout()
                self.btn_process_std = QPushButton("Process Standard Bands")
                self.btn_process_std.setToolTip(
                    "Analyze the defined area as a standard lane.\n"
                    "If 'Move/Resize Selected Area' was used to select a lane (single or multi-lane),\n"
                    "that selected lane will be processed. Otherwise, the current single quad/rect is used.\n"
                    "You will be prompted for the known quantity."
                )
                self.btn_process_std.clicked.connect(self.process_standard)
                std_proc_layout.addWidget(self.btn_process_std)
                quant_layout.addLayout(std_proc_layout)

                # Standard Info Display (Read-only makes more sense)
                std_info_layout = QGridLayout()
                std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
                self.standard_protein_values = QLineEdit()
                self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
                self.standard_protein_values.setReadOnly(True) # Make read-only
                std_info_layout.addWidget(self.standard_protein_values, 0, 1)

                std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
                self.standard_protein_areas_text = QLineEdit()
                self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
                self.standard_protein_areas_text.setReadOnly(True) # Make read-only
                std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
                quant_layout.addLayout(std_info_layout)

                quant_layout.addWidget(self.create_separator()) # Add visual separator

                # Sample Processing
                sample_proc_layout = QHBoxLayout()
                self.btn_analyze_sample = QPushButton("Analyze Sample Bands (Single/All Defined)")
                self.btn_analyze_sample.setToolTip("Analyze the defined area(s) as sample lane(s) using the standard curve.")
                self.btn_analyze_sample.clicked.connect(self.process_sample) # This will now check if multi-lane mode was used
                sample_proc_layout.addWidget(self.btn_analyze_sample)
                quant_layout.addLayout(sample_proc_layout)


                # Sample Info Display
                sample_info_layout = QGridLayout()
                sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
                self.target_protein_areas_text = QTextEdit() # Changed to QTextEdit for multi-lane
                self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
                self.target_protein_areas_text.setReadOnly(True)
                self.target_protein_areas_text.setFixedHeight(70) # Give it some initial height
                sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

                # Add Table Export button next to sample areas
                self.table_export_button = QPushButton("View/Export Results Table")
                self.table_export_button.setToolTip("View and export the analysis results (areas, percentages, quantities) to Excel.")
                self.table_export_button.clicked.connect(self.open_table_window)
                sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
                quant_layout.addLayout(sample_info_layout)

                layout.addWidget(quant_group)

                # --- Clear Button ---
                clear_layout = QHBoxLayout() 
                clear_layout.addStretch()
                self.clear_predict_button = QPushButton("Clear Analysis Markers & Regions")
                self.clear_predict_button.setToolTip("Clears MW prediction line and all analysis regions (single and multiple).\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
                self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
                layout.addWidget(self.clear_predict_button)
                layout.addLayout(clear_layout)


                layout.addStretch()
                return tab
            def start_analyze_multiple_lanes(self):
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.measure_quantity_mode = False
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.counter = 0
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_points = []
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self.live_view_label.selected_point = -1
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.mode = None
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.update_live_view()
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    return

                items_region = ["Rectangles (for straight lanes)", "Quadrilaterals (for skewed lanes)"]
                region_type_str, ok_region = QInputDialog.getItem(self, "Select Region Type for Multiple Lanes",
                                                                 "How do you want to define the lane regions?",
                                                                 items_region, 0, False)
                if not ok_region or not region_type_str:
                    self.cancel_multi_lane_mode()
                    return

                self.save_state() # Save state before starting multi-lane definition
                self.multi_lane_mode_active = True
                self.multi_lane_definitions = [] # Clear previous multi-lane definitions
                self.latest_multi_lane_peak_areas = {}
                self.latest_multi_lane_calculated_quantities = {}
                self.multi_lane_processing_finished = False
                self.target_protein_areas_text.clear()


                if "Rectangles" in region_type_str:
                    self.multi_lane_definition_type = 'rectangle'
                    self.live_view_label.mode = 'multi_lane_rect'
                    QMessageBox.information(self, "Define Multiple Lanes",
                                            f"Draw Lane 1 (Rectangle).\nPress ESC to cancel. Click 'Finish Defining' when done with all lanes.")
                    self.live_view_label._custom_left_click_handler_from_app = self.start_current_multi_lane_rect
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.update_current_multi_lane_rect_preview
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_current_multi_lane_definition
                elif "Quadrilaterals" in region_type_str:
                    self.multi_lane_definition_type = 'quad'
                    self.live_view_label.mode = 'multi_lane_quad'
                    self.current_multi_lane_points = []
                    QMessageBox.information(self, "Define Multiple Lanes",
                                            f"Click 4 corners for Lane 1 (Quadrilateral).\nPress ESC to cancel. Click 'Finish Defining' when done with all lanes.")
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_current_multi_lane_quad_click

                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)
                self.btn_finish_multi_lane_def.setEnabled(True)
                self.update_live_view()

            def start_current_multi_lane_rect(self, event):
                if self.multi_lane_mode_active and self.multi_lane_definition_type == 'rectangle':
                    if event.button() == Qt.LeftButton:
                        start_point_transformed = self.live_view_label.transform_point(event.position())
                        self.current_multi_lane_rect_start = self.snap_point_to_grid(start_point_transformed)
                        # Use bounding_box_preview for live drawing of the current rectangle
                        self.live_view_label.bounding_box_preview = (
                            self.current_multi_lane_rect_start.x(), self.current_multi_lane_rect_start.y(),
                            self.current_multi_lane_rect_start.x(), self.current_multi_lane_rect_start.y()
                        )
                        self.update_live_view()
            
            def update_current_multi_lane_rect_preview(self, event):
                if self.multi_lane_mode_active and self.multi_lane_definition_type == 'rectangle' and \
                   self.current_multi_lane_rect_start and (event.buttons() & Qt.LeftButton) :
                    current_end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(current_end_point_transformed)
                    self.live_view_label.bounding_box_preview = (
                        self.current_multi_lane_rect_start.x(), self.current_multi_lane_rect_start.y(),
                        snapped_end_point.x(), snapped_end_point.y()
                    )
                    self.update_live_view()

            def handle_current_multi_lane_quad_click(self, event):
                if self.multi_lane_mode_active and self.multi_lane_definition_type == 'quad':
                    if event.button() == Qt.LeftButton:
                        point_transformed = self.live_view_label.transform_point(event.position())
                        snapped_point = self.snap_point_to_grid(point_transformed)
                        self.current_multi_lane_points.append(snapped_point)
                        # Update live_view_label's quad_points for drawing the current quad being defined
                        self.live_view_label.quad_points = self.current_multi_lane_points[:]
                        self.update_live_view()

                        if len(self.current_multi_lane_points) == 4:
                            self.finalize_current_multi_lane_definition(event) # Pass event for consistency if needed


            def finalize_current_multi_lane_definition(self, event):
                lane_id = len(self.multi_lane_definitions) + 1
                definition_to_store = None

                if self.multi_lane_definition_type == 'rectangle':
                    if not self.current_multi_lane_rect_start or not self.live_view_label.bounding_box_preview:
                        return # Drag not completed
                    
                    # bounding_box_preview stores (x1, y1, x2, y2) in label space
                    x1, y1, x2, y2 = self.live_view_label.bounding_box_preview
                    rect_label_space = QRectF(QPointF(x1,y1), QPointF(x2,y2)).normalized()
                    
                    if rect_label_space.width() < 2 or rect_label_space.height() < 2: # Min size check
                        QMessageBox.warning(self, "Info", "Rectangle too small, please redraw.")
                        self.live_view_label.bounding_box_preview = None # Clear preview for redraw
                        self.current_multi_lane_rect_start = None
                        self.update_live_view()
                        return

                    definition_to_store = {'type': 'rectangle', 'points_label': [rect_label_space], 'id': lane_id}
                    self.multi_lane_definitions.append(definition_to_store)
                    # Reset for next rectangle
                    self.current_multi_lane_rect_start = None
                    self.live_view_label.bounding_box_preview = None # Clear current rect preview
                    QMessageBox.information(self, "Lane Defined", f"Lane {lane_id} (Rectangle) defined. Draw Lane {lane_id + 1} or click 'Finish Defining'.")

                elif self.multi_lane_definition_type == 'quad':
                    if len(self.current_multi_lane_points) == 4:
                        quad_points_label_space = [QPointF(p) for p in self.current_multi_lane_points]
                        definition_to_store = {'type': 'quad', 'points_label': quad_points_label_space, 'id': lane_id}
                        self.multi_lane_definitions.append(definition_to_store)
                        # Reset for next quad
                        self.current_multi_lane_points = []
                        self.live_view_label.quad_points = [] # Clear current quad preview from label
                        QMessageBox.information(self, "Lane Defined", f"Lane {lane_id} (Quadrilateral) defined. Click 4 points for Lane {lane_id + 1} or click 'Finish Defining'.")
                    else:
                        return # Not enough points for quad yet

                self.is_modified = True
                self.update_live_view() # Redraws all stored multi_lane_definitions

            def finish_multi_lane_definition_and_process(self):
                # ... (existing logic to discard incomplete definitions) ...
                if self.multi_lane_mode_active:
                    if self.multi_lane_definition_type == 'rectangle' and self.current_multi_lane_rect_start is not None:
                        print("INFO: Discarding incomplete rectangle definition upon finishing.")
                        self.current_multi_lane_rect_start = None
                        self.live_view_label.bounding_box_preview = None 
                    elif self.multi_lane_definition_type == 'quad' and self.current_multi_lane_points:
                        print("INFO: Discarding incomplete quadrilateral definition upon finishing.")
                        self.current_multi_lane_points = []
                        self.live_view_label.quad_points = [] 
                    self.update_live_view() 

                if not self.multi_lane_definitions:
                    # ... (message box) ...
                    self.cancel_multi_lane_mode() 
                    return

                QMessageBox.information(self, "Finished Processing Multiple Lanes", f"Finished Processing {len(self.multi_lane_definitions)} defined lanes. Click Analyze button to calculate the properties")
                
                self.multi_lane_mode_active = False 
                self.btn_finish_multi_lane_def.setEnabled(False)
                
                self.live_view_label.mode = None
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self._reset_live_view_label_custom_handlers()
                
                # === ADDED/UNCOMMENTED: Explicitly clear LiveViewLabel's preview buffers ===
                self.live_view_label.quad_points = [] 
                self.live_view_label.bounding_box_preview = None
                # === END ADDED/UNCOMMENTED ===

                self.multi_lane_processing_finished = True


            def cancel_multi_lane_mode(self):
                # ... (existing logic to clear multi-lane specific app state) ...
                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None
                self.current_multi_lane_points = []
                self.current_multi_lane_rect_start = None
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.quad_points = []
                self._reset_live_view_label_custom_handlers() # Use helper
                if hasattr(self, 'btn_finish_multi_lane_def'):
                    self.btn_finish_multi_lane_def.setEnabled(False)
                self.update_live_view()

            
            def enable_move_selection_mode(self):
                """Enables mode to select an area (single or multi-lane) for moving/resizing."""
                can_select_single_quad = bool(self.live_view_label.quad_points)
                can_select_single_rect = bool(self.live_view_label.bounding_box_preview)
                can_select_multi_lane = bool(self.multi_lane_definitions)

                if not (can_select_single_quad or can_select_single_rect or can_select_multi_lane):
                    QMessageBox.information(self, "Move/Resize Area", "No area is currently defined to select.")
                    self.cancel_selection_or_move_mode()
                    return

                self.current_selection_mode = "select_for_move"
                self.live_view_label.mode = "select_for_move" 
                self._reset_live_view_label_custom_handlers() # Clear previous
                self.live_view_label._custom_left_click_handler_from_app = self.handle_area_selection_click
                # Move and release for drag operation will be set within handle_area_selection_click
                
                QMessageBox.information(self, "Select Area to Move/Resize", "Click on a defined lane to select it. Click near a corner to resize, or inside to move the whole shape.")
                self.moving_multi_lane_index = -1 
                self.resizing_corner_index = -1
                self.update_live_view()
                
            def handle_area_selection_click(self, event):
                if self.current_selection_mode != "select_for_move" or event.button() != Qt.LeftButton:
                    if hasattr(self.live_view_label, '_original_mousePressEvent') and self.live_view_label._original_mousePressEvent:
                        self.live_view_label._original_mousePressEvent(event)
                    elif isinstance(self.live_view_label, QLabel):
                        QLabel.mousePressEvent(self.live_view_label, event)
                    return

                clicked_point_label_space = self.live_view_label.transform_point(event.position())
                
                self.moving_multi_lane_index = -1
                self.resizing_corner_index = -1
                selected_shape_for_interaction = False # Flag if any interaction (move or resize) is initiated

                # Click threshold for corners
                click_radius_threshold = LiveViewLabel.CORNER_HANDLE_BASE_RADIUS * 1.5

                # --- Stage 1: Check for Corner Clicks to Initiate RESIZE ---
                # Priority: Multi-lanes, then single quad, then single rect for corner clicks
                
                # Check Multi-Lane Corners
                if self.multi_lane_definitions:
                    for i, lane_def in reversed(list(enumerate(self.multi_lane_definitions))):
                        current_lane_corners_label = []
                        if lane_def['type'] == 'rectangle':
                            rect_ls = lane_def['points_label'][0]
                            current_lane_corners_label = [rect_ls.topLeft(), rect_ls.topRight(), rect_ls.bottomRight(), rect_ls.bottomLeft()]
                        elif lane_def['type'] == 'quad':
                            current_lane_corners_label = lane_def['points_label']
                        
                        for corner_idx, corner_pt in enumerate(current_lane_corners_label):
                            if (clicked_point_label_space - corner_pt).manhattanLength() < click_radius_threshold:
                                self.moving_multi_lane_index = i
                                self.resizing_corner_index = corner_idx
                                self.shape_points_at_drag_start_label = [QPointF(p) for p in current_lane_corners_label]
                                selected_shape_for_interaction = True
                                break
                        if selected_shape_for_interaction: break
                
                # Check Single Quad Corners (if no multi-lane corner selected)
                if not selected_shape_for_interaction and self.live_view_label.quad_points:
                    for corner_idx, corner_pt in enumerate(self.live_view_label.quad_points):
                        if (clicked_point_label_space - corner_pt).manhattanLength() < click_radius_threshold:
                            self.moving_multi_lane_index = -2 # Single quad
                            self.resizing_corner_index = corner_idx
                            self.shape_points_at_drag_start_label = [QPointF(p) for p in self.live_view_label.quad_points]
                            selected_shape_for_interaction = True
                            break
                
                # Check Single Rectangle Corners (if nothing else selected)
                if not selected_shape_for_interaction and self.live_view_label.bounding_box_preview:
                    x1, y1, x2, y2 = self.live_view_label.bounding_box_preview
                    rect_single = QRectF(QPointF(x1,y1), QPointF(x2,y2)).normalized()
                    single_rect_corners = [rect_single.topLeft(), rect_single.topRight(), rect_single.bottomRight(), rect_single.bottomLeft()]
                    for corner_idx, corner_pt in enumerate(single_rect_corners):
                        if (clicked_point_label_space - corner_pt).manhattanLength() < click_radius_threshold:
                            self.moving_multi_lane_index = -3 # Single rect
                            self.resizing_corner_index = corner_idx
                            self.shape_points_at_drag_start_label = [QPointF(p) for p in single_rect_corners]
                            selected_shape_for_interaction = True
                            break

                # --- Stage 2: If no corner was clicked, check for Whole Shape Click to Initiate MOVE ---
                if not selected_shape_for_interaction:
                    # Check Multi-Lane Bodies
                    if self.multi_lane_definitions:
                        for i, lane_def in reversed(list(enumerate(self.multi_lane_definitions))):
                            is_inside = False
                            current_lane_points_for_body_check = []
                            if lane_def['type'] == 'rectangle':
                                rect_ls = lane_def['points_label'][0]
                                if rect_ls.contains(clicked_point_label_space): is_inside = True
                                current_lane_points_for_body_check = [rect_ls.topLeft(), rect_ls.topRight(), rect_ls.bottomRight(), rect_ls.bottomLeft()]
                            elif lane_def['type'] == 'quad':
                                poly_ls = QPolygonF(lane_def['points_label'])
                                if poly_ls.containsPoint(clicked_point_label_space, Qt.OddEvenFill): is_inside = True
                                current_lane_points_for_body_check = lane_def['points_label']
                            
                            if is_inside:
                                self.moving_multi_lane_index = i
                                self.shape_points_at_drag_start_label = [QPointF(p) for p in current_lane_points_for_body_check]
                                selected_shape_for_interaction = True
                                break
                    
                    # Check Single Quad Body
                    if not selected_shape_for_interaction and self.live_view_label.quad_points:
                        poly_single_quad = QPolygonF(self.live_view_label.quad_points)
                        if poly_single_quad.containsPoint(clicked_point_label_space, Qt.OddEvenFill):
                            self.moving_multi_lane_index = -2
                            self.shape_points_at_drag_start_label = [QPointF(p) for p in self.live_view_label.quad_points]
                            selected_shape_for_interaction = True
                    
                    # Check Single Rectangle Body
                    if not selected_shape_for_interaction and self.live_view_label.bounding_box_preview:
                        x1, y1, x2, y2 = self.live_view_label.bounding_box_preview
                        rect_single = QRectF(QPointF(x1,y1), QPointF(x2,y2)).normalized()
                        if rect_single.contains(clicked_point_label_space):
                            self.moving_multi_lane_index = -3
                            self.shape_points_at_drag_start_label = [rect_single.topLeft(), rect_single.topRight(), rect_single.bottomRight(), rect_single.bottomLeft()]
                            selected_shape_for_interaction = True

                # --- Setup for Dragging/Resizing if a shape was selected ---
                if selected_shape_for_interaction:
                    self.initial_mouse_pos_for_shape_drag_label = clicked_point_label_space

                    if self.resizing_corner_index != -1: # A corner was hit in Stage 1
                        self.current_selection_mode = "resizing_corner"
                        self.live_view_label.mode = "resizing_corner"
                        self.live_view_label.setCursor(Qt.CrossCursor)
                    else: # A body was hit in Stage 2
                        self.current_selection_mode = "dragging_shape"
                        self.live_view_label.mode = "dragging_shape"
                        self.live_view_label.setCursor(Qt.CrossCursor)

                    self.live_view_label.draw_edges = False 
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_drag_operation 
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_drag_release
                else:
                    # Click was outside any known shape or its corners
                    self.moving_multi_lane_index = -1 
                    self.resizing_corner_index = -1
                    # current_selection_mode remains "select_for_move"
                
                self.update_live_view() # To show selection highlight
                
            def handle_drag_operation(self, event): # Renamed from move_selection
                if self.current_selection_mode not in ["dragging_shape", "resizing_corner"] or \
                   not self.shape_points_at_drag_start_label or \
                   not (event.buttons() & Qt.LeftButton):
                    if hasattr(self.live_view_label, '_original_mouseMoveEvent') and self.live_view_label._original_mouseMoveEvent:
                        self.live_view_label._original_mouseMoveEvent(event)
                    elif isinstance(self.live_view_label, QLabel):
                        QLabel.mouseMoveEvent(self.live_view_label, event)
                    return

                current_mouse_pos_label = self.live_view_label.transform_point(event.position())
                new_shape_points_label = [] # This will hold the 4 QPointF for the updated shape

                # Determine the type of the shape being manipulated
                is_multi_rect_resize = False
                is_single_rect_resize = False

                if self.moving_multi_lane_index >= 0 and \
                   self.multi_lane_definitions[self.moving_multi_lane_index]['type'] == 'rectangle' and \
                   self.current_selection_mode == "resizing_corner":
                    is_multi_rect_resize = True
                elif self.moving_multi_lane_index == -3 and \
                     self.current_selection_mode == "resizing_corner":
                    is_single_rect_resize = True

                is_rect_resize_mode = is_multi_rect_resize or is_single_rect_resize

                if self.current_selection_mode == "dragging_shape":
                    # --- Logic for dragging the whole shape (remains the same) ---
                    raw_mouse_delta_x = current_mouse_pos_label.x() - self.initial_mouse_pos_for_shape_drag_label.x()
                    raw_mouse_delta_y = current_mouse_pos_label.y() - self.initial_mouse_pos_for_shape_drag_label.y()
                    reference_point_orig_label = self.shape_points_at_drag_start_label[0]
                    raw_new_ref_point_x = reference_point_orig_label.x() + raw_mouse_delta_x
                    raw_new_ref_point_y = reference_point_orig_label.y() + raw_mouse_delta_y
                    snapped_new_ref_point_x = raw_new_ref_point_x; snapped_new_ref_point_y = raw_new_ref_point_y
                    grid_size = 0; snap_x_enabled = False; snap_y_enabled = False
                    if hasattr(self, 'grid_size_input'): grid_size = self.grid_size_input.value()
                    if hasattr(self, 'show_grid_checkbox_x'): snap_x_enabled = self.show_grid_checkbox_x.isChecked()
                    if hasattr(self, 'show_grid_checkbox_y'): snap_y_enabled = self.show_grid_checkbox_y.isChecked()
                    if grid_size > 0:
                        if snap_x_enabled: snapped_new_ref_point_x = round(raw_new_ref_point_x / grid_size) * grid_size
                        if snap_y_enabled: snapped_new_ref_point_y = round(raw_new_ref_point_y / grid_size) * grid_size
                    effective_delta_x = snapped_new_ref_point_x - reference_point_orig_label.x()
                    effective_delta_y = snapped_new_ref_point_y - reference_point_orig_label.y()
                    effective_delta_label = QPointF(effective_delta_x, effective_delta_y)
                    new_shape_points_label = [p_orig_label + effective_delta_label for p_orig_label in self.shape_points_at_drag_start_label]
                    # --- End whole shape drag logic ---

                elif self.current_selection_mode == "resizing_corner" and self.resizing_corner_index != -1:
                    snapped_mouse_pos_label = self.snap_point_to_grid(current_mouse_pos_label)
                    
                    if is_rect_resize_mode:
                        # --- Rectangle Resizing Logic ---
                        # Corners are: 0:TL, 1:TR, 2:BR, 3:BL (from QRectF.topLeft() etc.)
                        moved_corner_idx = self.resizing_corner_index
                        fixed_opposite_corner_idx = (moved_corner_idx + 2) % 4 # Diagonal opposite
                        
                        current_points = list(self.shape_points_at_drag_start_label) # Work with a copy
                        
                        # The corner being dragged moves to the snapped mouse position
                        current_points[moved_corner_idx] = snapped_mouse_pos_label
                        
                        # The diagonally opposite corner remains fixed from its start_drag position
                        fixed_corner = self.shape_points_at_drag_start_label[fixed_opposite_corner_idx]
                        current_points[fixed_opposite_corner_idx] = fixed_corner

                        # Determine the new X and Y for the other two corners
                        # to maintain the rectangle
                        new_x_coords = sorted([snapped_mouse_pos_label.x(), fixed_corner.x()])
                        new_y_coords = sorted([snapped_mouse_pos_label.y(), fixed_corner.y()])
                        
                        min_x, max_x = new_x_coords[0], new_x_coords[1]
                        min_y, max_y = new_y_coords[0], new_y_coords[1]

                        # Reconstruct all 4 corners for the new rectangle
                        new_shape_points_label = [
                            QPointF(min_x, min_y), # TopLeft
                            QPointF(max_x, min_y), # TopRight
                            QPointF(max_x, max_y), # BottomRight
                            QPointF(min_x, max_y)  # BottomLeft
                        ]
                        # --- End Rectangle Resizing ---
                    else: # Quadrilateral resizing (or other shape type if added later)
                        new_shape_points_label = list(self.shape_points_at_drag_start_label) 
                        new_shape_points_label[self.resizing_corner_index] = snapped_mouse_pos_label
                
                if not new_shape_points_label: return

                self.live_view_label.quad_points = [] 
                self.live_view_label.bounding_box_preview = None

                if self.moving_multi_lane_index >= 0: 
                    lane_to_update = self.multi_lane_definitions[self.moving_multi_lane_index]
                    if lane_to_update['type'] == 'quad': # Quads always store 4 points
                        lane_to_update['points_label'] = new_shape_points_label
                        self.live_view_label.quad_points = new_shape_points_label[:] # For paintEvent
                    elif lane_to_update['type'] == 'rectangle':
                        # After resize (or move), it's stored as a QRectF
                        all_x = [p.x() for p in new_shape_points_label]; all_y = [p.y() for p in new_shape_points_label]
                        min_x, max_x = min(all_x), max(all_x); min_y, max_y = min(all_y), max(all_y)
                        lane_to_update['points_label'] = [QRectF(QPointF(min_x, min_y), QPointF(max_x, max_y))]
                        rect_f = lane_to_update['points_label'][0]
                        self.live_view_label.bounding_box_preview = (rect_f.left(), rect_f.top(), rect_f.right(), rect_f.bottom())


                elif self.moving_multi_lane_index == -2: # Single Quad
                    self.live_view_label.quad_points = new_shape_points_label
                elif self.moving_multi_lane_index == -3: # Single Rect
                    # After resize or move, it's a rectangle
                    all_x = [p.x() for p in new_shape_points_label]; all_y = [p.y() for p in new_shape_points_label]
                    min_x, max_x = min(all_x), max(all_x); min_y, max_y = min(all_y), max(all_y)
                    self.live_view_label.bounding_box_preview = (min_x, min_y, max_x, max_y)
                    self.live_view_label.quad_points = [] # Clear quad representation if it was a rect
                
                self.update_live_view()
            
            def handle_drag_release(self, event): # Renamed from end_move_selection
                if self.current_selection_mode in ["dragging_shape", "resizing_corner"] and event.button() == Qt.LeftButton:
                    self.live_view_label.draw_edges = True
                    
                    # If a single rectangle was resized, its updated points were used to form
                    # self.live_view_label.bounding_box_preview in handle_drag_operation.
                    # If a multi-lane rectangle was resized, its 'points_label' (a QRectF) was updated.
                    # If a quad was resized, its 'points_label' or self.live_view_label.quad_points was updated.

                    if self.shape_points_at_drag_start_label: # Check if a drag actually happened
                        self.save_state()
                        self.is_modified = True

                    self.shape_points_at_drag_start_label = []
                    self.initial_mouse_pos_for_shape_drag_label = QPointF()
                    self.resizing_corner_index = -1 
                    
                    self.current_selection_mode = "select_for_move"
                    self.live_view_label.mode = "select_for_move"
                    
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_area_selection_click
                    
                    self.update_live_view()
            
            def cancel_selection_or_move_mode(self):
                self.current_selection_mode = None
                self.live_view_label.mode = None 
                self.moving_multi_lane_index = -1
                self.resizing_corner_index = -1
                self._reset_live_view_label_custom_handlers() # Use helper
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                self.live_view_label.draw_edges = True
                self.update_live_view()
                
                
                     
            def get_nearest_point(self, mouse_pos, points):
                """Get the nearest point to the mouse position."""
                min_distance = float('inf')
                nearest_point = None
                for point in points:
                    distance = (mouse_pos - point).manhattanLength()
                    if distance < min_distance:
                        min_distance = distance
                        nearest_point = point
                return nearest_point
            
            def open_table_window(self):
                is_multi_lane_results = bool(self.latest_multi_lane_peak_areas) and self.multi_lane_processing_finished

                peak_areas_data_for_table = None
                quantities_data_for_table = None
                peak_details_for_table = None # NEW

                if is_multi_lane_results:
                    peak_areas_data_for_table = self.latest_multi_lane_peak_areas 
                    quantities_data_for_table = self.latest_multi_lane_calculated_quantities
                    peak_details_for_table = self.latest_multi_lane_peak_details # Pass multi-lane details
                else: 
                    peak_areas_data_for_table = self.latest_peak_areas 
                    quantities_data_for_table = self.latest_calculated_quantities 
                    peak_details_for_table = {1: self.latest_peak_details} if self.latest_peak_details else {} # Wrap single in dict for TableWindow

                standard_dict_to_show_current = self.quantities_peak_area_dict
                is_standard_mode_current = len(standard_dict_to_show_current) >= 2
            
                self.table_window_instance = TableWindow(
                    peak_areas_data_for_table, 
                    standard_dict_to_show_current,
                    is_standard_mode_current,
                    quantities_data_for_table,
                    self, # parent_app_instance
                    peak_details_data=peak_details_for_table # NEW Pass peak details
                )
                self.table_window_instance.show()
            
            def enable_quad_mode(self):
                """Enable mode to define a single quadrilateral area, clearing multi-lane definitions."""
                # --- NEW: Clear multi-lane definitions if they exist ---
                
                if self.multi_lane_definitions:
                    print("INFO: Clearing existing multi-lane definitions to define a single quadrilateral.")
                    self.multi_lane_definitions.clear()
                    self.latest_multi_lane_peak_areas.clear()
                    self.latest_multi_lane_peak_details.clear()
                    self.latest_multi_lane_calculated_quantities.clear()
                    self.multi_lane_processing_finished = False
                    if hasattr(self, 'btn_finish_multi_lane_def'):
                        self.btn_finish_multi_lane_def.setEnabled(False)
                    # No need to clear self.current_multi_lane_points or self.current_multi_lane_rect_start
                    # as those are for the *current* multi-lane definition process, which is being superseded.
                # --- END NEW ---

                self.live_view_label.bounding_box_preview = [] # Clear single rect preview
                self.live_view_label.quad_points = []          # Clear single quad points for a fresh start
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.measure_quantity_mode = True # App flag
                self.live_view_label.mode = "quad" # For LiveViewLabel's paintEvent
                self._reset_live_view_label_custom_handlers() # Clear previous custom handlers
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                # Clearing single-lane results as well, as we are defining a new single area
                self.latest_peak_areas = []
                self.latest_peak_details = []
                self.latest_calculated_quantities = []
                self.target_protein_areas_text.clear() 

                # Ensure multi-lane UI elements are reset if they were active
                if hasattr(self, 'btn_finish_multi_lane_def'):
                    self.btn_finish_multi_lane_def.setEnabled(False)
                self.multi_lane_processing_finished = False
                self.multi_lane_mode_active = False # Ensure multi-lane mode is off
                
                self.update_live_view()
            
            def enable_rectangle_mode(self):
                """Enable mode to define a single rectangle area, clearing multi-lane definitions."""
                # --- NEW: Clear multi-lane definitions if they exist ---
                self.live_view_label.setCursor(Qt.CrossCursor)
                if self.multi_lane_definitions:
                    print("INFO: Clearing existing multi-lane definitions to define a single rectangle.")
                    self.multi_lane_definitions.clear()
                    self.latest_multi_lane_peak_areas.clear()
                    self.latest_multi_lane_peak_details.clear()
                    self.latest_multi_lane_calculated_quantities.clear()
                    self.multi_lane_processing_finished = False
                    if hasattr(self, 'btn_finish_multi_lane_def'):
                        self.btn_finish_multi_lane_def.setEnabled(False)
                # --- END NEW ---

                self.live_view_label.bounding_box_preview = None # Clear single rect preview for a fresh start
                self.live_view_label.quad_points = []            # Clear single quad points
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.measure_quantity_mode = True # App flag
                self.live_view_label.mode = "rectangle" # For LiveViewLabel's paintEvent
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                self.live_view_label._custom_left_click_handler_from_app = self.start_rectangle
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_rectangle_preview
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_rectangle

                # Clearing single-lane results as well
                self.latest_peak_areas = []
                self.latest_peak_details = []
                self.latest_calculated_quantities = []
                self.target_protein_areas_text.clear()

                # Ensure multi-lane UI elements are reset
                if hasattr(self, 'btn_finish_multi_lane_def'):
                    self.btn_finish_multi_lane_def.setEnabled(False)
                self.multi_lane_processing_finished = False
                self.multi_lane_mode_active = False # Ensure multi-lane mode is off
                

                
                self.update_live_view()
            
            def snap_point_to_grid(self, point: QPointF) -> QPointF:
                """Snaps a QPointF to the grid if enabled."""
                snapped_x = point.x()
                snapped_y = point.y()
    
                grid_size = 0
                snap_x_enabled = False
                snap_y_enabled = False
    
                if hasattr(self, 'grid_size_input'):
                    grid_size = self.grid_size_input.value()
                if hasattr(self, 'show_grid_checkbox_x'):
                    snap_x_enabled = self.show_grid_checkbox_x.isChecked()
                if hasattr(self, 'show_grid_checkbox_y'):
                    snap_y_enabled = self.show_grid_checkbox_y.isChecked()
    
                if grid_size > 0:
                    if snap_x_enabled:
                        snapped_x = round(point.x() / grid_size) * grid_size
                    if snap_y_enabled:
                        snapped_y = round(point.y() / grid_size) * grid_size
                
                return QPointF(snapped_x, snapped_y)
            
            def start_rectangle(self, event):
                """Record the start position of the rectangle (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"]:
                    start_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    self.live_view_label.rectangle_start = snapped_start_point
                    self.live_view_label.rectangle_points = [snapped_start_point]
                    self.live_view_label.bounding_box_preview = None
            
            def update_rectangle_preview(self, event):
                """Update the rectangle preview as the mouse moves (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"] and self.live_view_label.rectangle_start:
                    current_end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(current_end_point_transformed) # Snap it
    
                    self.live_view_label.rectangle_end = snapped_end_point
            
                    if self.live_view_label.rectangle_start and self.live_view_label.rectangle_end:
                        self.live_view_label.bounding_box_preview = (
                            self.live_view_label.rectangle_start.x(),
                            self.live_view_label.rectangle_start.y(),
                            self.live_view_label.rectangle_end.x(), # Use snapped end point
                            self.live_view_label.rectangle_end.y(), # Use snapped end point
                        )
                        self.update_live_view()
            
            def finalize_rectangle(self, event):
                """Finalize the rectangle when the mouse is released."""
                if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
                    end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
    
                    self.live_view_label.rectangle_end = snapped_end_point
                    # self.live_view_label.rectangle_start is already snapped
                    self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, snapped_end_point]
                    
                    self.live_view_label.bounding_box_preview = (
                        self.live_view_label.rectangle_start.x(),
                        self.live_view_label.rectangle_start.y(),
                        snapped_end_point.x(), # Use snapped end point
                        snapped_end_point.y(), # Use snapped end point
                    )
                    
                    self.live_view_label.mode = None
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self.update_live_view()
                
            
            def process_standard(self):
                extracted_qimage = None
                region_type_for_message = "" # For user feedback

                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                self.update_live_view()

                # Check if a multi-lane was selected via the "Move/Resize Selected Area" mode
                if hasattr(self, 'moving_multi_lane_index') and self.moving_multi_lane_index >= 0 and \
                   self.moving_multi_lane_index < len(self.multi_lane_definitions):
                    
                    selected_lane_def = self.multi_lane_definitions[self.moving_multi_lane_index]
                    lane_id = selected_lane_def['id']
                    region_type_for_message = f"Selected Multi-Lane {lane_id} ({selected_lane_def['type']})"
                    print(f"Processing Standard: {region_type_for_message}")

                    if selected_lane_def['type'] == 'quad':
                        quad_points_label_space = selected_lane_def['points_label']
                        extracted_qimage = self.quadrilateral_to_rect(self.image, quad_points_label_space)
                    elif selected_lane_def['type'] == 'rectangle':
                        rect_label_space = selected_lane_def['points_label'][0] # QRectF
                        img_coords_rect = self._map_label_rect_to_image_rect(rect_label_space)
                        if img_coords_rect:
                            x, y, w, h = img_coords_rect
                            extracted_qimage = self.image.copy(x, y, w, h)
                    
                    if not extracted_qimage or extracted_qimage.isNull():
                        QMessageBox.warning(self, "Error", f"Could not extract/warp {region_type_for_message}.")
                        return
                
                # Fallback to single defined quad on LiveViewLabel
                elif len(self.live_view_label.quad_points) == 4:
                    region_type_for_message = "Defined Single Quadrilateral"
                    print(f"Processing Standard: {region_type_for_message}")
                    extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
                    if not extracted_qimage or extracted_qimage.isNull():
                        QMessageBox.warning(self, "Error", "Single Quadrilateral warping failed.")
                        return

                # Fallback to single defined rectangle on LiveViewLabel
                elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
                    region_type_for_message = "Defined Single Rectangle"
                    print(f"Processing Standard: {region_type_for_message}")
                    try:
                        img_coords_rect = self._map_label_rect_to_image_rect(QRectF(
                            QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]),
                            QPointF(self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3])
                        ).normalized())
                        if img_coords_rect:
                            x, y, w, h = img_coords_rect
                            extracted_qimage = self.image.copy(x, y, w, h)
                        if not extracted_qimage or extracted_qimage.isNull():
                            raise ValueError("QImage.copy failed for single rectangle.")
                    except Exception as e:
                         print(f"Error processing single rectangle region for standard: {e}")
                         QMessageBox.warning(self, "Error", "Could not process single rectangular region.")
                         return
                else:
                    QMessageBox.warning(self, "Input Error", "Please define an area (Single Quad/Rect) or select a Multi-Lane area using 'Move/Resize Selected Area' first.")
                    return

                # --- Convert extracted region to Grayscale PIL for analysis ---
                if extracted_qimage and not extracted_qimage.isNull():
                    processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                    if processed_data_pil:
                        # analyze_bounding_box now expects standard=True/False
                        self.analyze_bounding_box(processed_data_pil, standard=True) 
                    else:
                        QMessageBox.warning(self, "Error", f"Could not convert {region_type_for_message} to grayscale for analysis.")
            
            def process_sample(self):
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                self.update_live_view()  
                
                self.latest_peak_areas = [] 
                self.latest_peak_details = []
                self.latest_calculated_quantities = []
                self.latest_multi_lane_peak_areas.clear()
                self.latest_multi_lane_peak_details.clear()
                self.latest_multi_lane_calculated_quantities.clear()

                extracted_regions_info = [] # List of dicts: {'pil': PIL, 'id': int, 'original_def': ...}

                if self.multi_lane_definitions and self.multi_lane_processing_finished: 
                    print(f"Processing {len(self.multi_lane_definitions)} multiple lanes as samples.")
                    for lane_def in self.multi_lane_definitions:
                        extracted_qimage = None
                        current_region_def_img_space = None 
                        pil_for_this_lane = None # Initialize here
                        
                        if lane_def['type'] == 'quad':
                            quad_points_label_space = lane_def['points_label']
                            extracted_qimage = self.quadrilateral_to_rect(self.image, quad_points_label_space)
                            current_region_def_img_space = self._map_label_points_to_image_points(quad_points_label_space)
                        elif lane_def['type'] == 'rectangle':
                            rect_label_space = lane_def['points_label'][0] 
                            img_coords_rect = self._map_label_rect_to_image_rect(rect_label_space)
                            if img_coords_rect and self.image: # Check if self.image is valid
                                x, y, w, h = img_coords_rect
                                extracted_qimage = self.image.copy(x, y, w, h)
                                current_region_def_img_space = img_coords_rect
                        
                        if extracted_qimage and not extracted_qimage.isNull():
                            pil_for_this_lane = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                            if pil_for_this_lane:
                                extracted_regions_info.append({'pil': pil_for_this_lane, 'id': lane_def['id'], 'original_def': current_region_def_img_space})
                            else:
                                QMessageBox.warning(self, "Error", f"Could not convert Lane {lane_def['id']} to PIL for analysis.")
                        else:
                             QMessageBox.warning(self, "Error", f"Could not extract/warp Lane {lane_def['id']}.")
                    
                    if not extracted_regions_info:
                        QMessageBox.warning(self, "Error", "No valid regions could be prepared for multi-lane sample analysis.")
                        return

                elif len(self.live_view_label.quad_points) == 4: 
                    print("Processing Sample: Single Quadrilateral")
                    extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
                    if extracted_qimage and not extracted_qimage.isNull():
                        pil_img = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                        if pil_img:
                             extracted_regions_info.append({'pil': pil_img, 'id': 1, 'original_def': self._map_label_points_to_image_points(self.live_view_label.quad_points)})
                        else: QMessageBox.warning(self, "Error", "Could not convert quad region to PIL.")
                    else: QMessageBox.warning(self, "Error", "Quadrilateral warping failed for sample.")

                elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4: 
                    print("Processing Sample: Single Rectangle")
                    try:
                        img_coords_rect = self._map_label_rect_to_image_rect(QRectF(
                            QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]),
                            QPointF(self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3])
                        ).normalized())
                        if img_coords_rect and self.image: # Check self.image
                            x, y, w, h = img_coords_rect
                            extracted_qimage = self.image.copy(x, y, w, h)
                            if extracted_qimage and not extracted_qimage.isNull():
                                pil_img = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                                if pil_img:
                                     extracted_regions_info.append({'pil': pil_img, 'id': 1, 'original_def': img_coords_rect})
                                else: QMessageBox.warning(self, "Error", "Could not convert rect region to PIL.")
                            else: raise ValueError("QImage.copy failed for rectangle.")
                        else: raise ValueError("Failed to map rectangle to image coords or image is missing.")
                    except Exception as e:
                         print(f"Error processing rectangle region for sample: {e}"); QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                else:
                    QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first, or finish defining multiple lanes.")
                    return

                # --- Analyze each extracted PIL image ---
                all_lanes_text_results = []
                if not extracted_regions_info: # Double check if list is empty before iterating
                    QMessageBox.warning(self, "Analysis Error", "No regions were successfully extracted for analysis.")
                    return

                for region_info in extracted_regions_info: # Iterate over the populated list
                    lane_id = region_info['id']
                    pil_image_for_dialog = region_info['pil'] # This is now correctly scoped
                    
                    peak_info_for_lane = self.calculate_peak_area(pil_image_for_dialog) 

                    if peak_info_for_lane and len(peak_info_for_lane) > 0:
                        areas_for_this_lane = [round(info['area'], 3) for info in peak_info_for_lane]
                        self.latest_multi_lane_peak_areas[lane_id] = areas_for_this_lane
                        self.latest_multi_lane_peak_details[lane_id] = peak_info_for_lane 

                        if len(self.quantities_peak_area_dict) >= 2:
                            quantities_for_lane = self.calculate_unknown_quantity(
                                list(self.quantities_peak_area_dict.values()),
                                list(self.quantities_peak_area_dict.keys()),
                                areas_for_this_lane 
                            )
                            self.latest_multi_lane_calculated_quantities[lane_id] = quantities_for_lane
                            all_lanes_text_results.append(f"Lane {lane_id}: Areas={areas_for_this_lane}, Qty={quantities_for_lane}")
                        else:
                            self.latest_multi_lane_calculated_quantities[lane_id] = []
                            all_lanes_text_results.append(f"Lane {lane_id}: Areas={self.latest_multi_lane_peak_areas[lane_id]} (No std curve for qty)")
                    else:
                        all_lanes_text_results.append(f"Lane {lane_id}: Analysis failed or no peaks.")
                        self.latest_multi_lane_peak_areas[lane_id] = []
                        self.latest_multi_lane_peak_details[lane_id] = [] # Ensure details is also cleared
                        self.latest_multi_lane_calculated_quantities[lane_id] = []
                
                if 1 in self.latest_multi_lane_peak_areas:
                     self.latest_peak_areas = self.latest_multi_lane_peak_areas[1]
                if 1 in self.latest_multi_lane_peak_details: 
                     self.latest_peak_details = self.latest_multi_lane_peak_details[1]
                if 1 in self.latest_multi_lane_calculated_quantities:
                     self.latest_calculated_quantities = self.latest_multi_lane_calculated_quantities[1]
                
                if all_lanes_text_results:
                    self.target_protein_areas_text.setText("\n".join(all_lanes_text_results))
                else:
                    self.target_protein_areas_text.setText("N/A or analysis failed for all lanes.")
                
                self.update_live_view()
                
            def _map_label_rect_to_image_rect(self, rect_label_space: QRectF):
                if not self.image or self.image.isNull(): return None
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                if not (img_w > 0 and img_h > 0 and label_w > 0 and label_h > 0): return None

                scale_factor = min(label_w / img_w, label_h / img_h)
                display_offset_x = (label_w - img_w * scale_factor) / 2.0
                display_offset_y = (label_h - img_h * scale_factor) / 2.0

                start_x_img = (rect_label_space.left() - display_offset_x) / scale_factor
                start_y_img = (rect_label_space.top() - display_offset_y) / scale_factor
                end_x_img = (rect_label_space.right() - display_offset_x) / scale_factor
                end_y_img = (rect_label_space.bottom() - display_offset_y) / scale_factor
                
                x_img = int(min(start_x_img, end_x_img))
                y_img = int(min(start_y_img, end_y_img))
                w_img = int(abs(end_x_img - start_x_img))
                h_img = int(abs(end_y_img - start_y_img))

                x_clamped = max(0, x_img)
                y_clamped = max(0, y_img)
                w_clamped = max(1, min(w_img, img_w - x_clamped))
                h_clamped = max(1, min(h_img, img_h - y_clamped))
                return (x_clamped, y_clamped, w_clamped, h_clamped)

            # Helper: _map_label_points_to_image_points
            def _map_label_points_to_image_points(self, points_label_space: list): # list of QPointF
                if not self.image or self.image.isNull() or not points_label_space: return None
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                if not (img_w > 0 and img_h > 0 and label_w > 0 and label_h > 0): return None

                scale_factor = min(label_w / img_w, label_h / img_h)
                display_offset_x = (label_w - img_w * scale_factor) / 2.0
                display_offset_y = (label_h - img_h * scale_factor) / 2.0
                
                image_points = []
                for p_label in points_label_space:
                    x_img = (p_label.x() - display_offset_x) / scale_factor
                    y_img = (p_label.y() - display_offset_y) / scale_factor
                    image_points.append(QPointF(x_img, y_img))
                return image_points


            def convert_qimage_to_grayscale_pil(self, qimg):
                """
                Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
                for use with PeakAreaDialog. Returns None on failure.
                """
                if not qimg or qimg.isNull():
                    return None

                fmt = qimg.format()
                pil_img = None

                try:
                    # Already grayscale? Convert directly if possible.
                    if fmt == QImage.Format_Grayscale16:
                        np_array = self.qimage_to_numpy(qimg)
                        if np_array is not None and np_array.dtype == np.uint16:
                            try: pil_img = Image.fromarray(np_array, mode='I;16')
                            except ValueError: pil_img = Image.fromarray(np_array, mode='I')
                        else: raise ValueError("Failed NumPy conversion for Grayscale16")
                    elif fmt == QImage.Format_Grayscale8:
                        # Try direct conversion first
                        try:
                            pil_img = ImageQt.fromqimage(qimg).convert('L')
                            if pil_img is None: raise ValueError("Direct QImage->PIL(L) failed.")
                        except Exception as e_direct:
                            np_array = self.qimage_to_numpy(qimg)
                            if np_array is not None and np_array.dtype == np.uint8:
                                pil_img = Image.fromarray(np_array, mode='L')
                            else: raise ValueError("Failed NumPy conversion for Grayscale8")
                    else: # Color or other format
                        # Use NumPy for robust conversion to 16-bit grayscale intermediate
                        np_img = self.qimage_to_numpy(qimg)
                        if np_img is None: raise ValueError("NumPy conversion failed for color.")
                        if np_img.ndim == 3:
                            gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) # Assume BGR/BGRA input
                            # Convert to 16-bit PIL
                            gray_np_16bit = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                            try: pil_img = Image.fromarray(gray_np_16bit, mode='I;16')
                            except ValueError: pil_img = Image.fromarray(gray_np_16bit, mode='I')
                        elif np_img.ndim == 2: # Should have been caught by Grayscale checks, but handle anyway
                             if np_img.dtype == np.uint16:
                                 try: pil_img = Image.fromarray(np_img, mode='I;16')
                                 except ValueError: pil_img = Image.fromarray(np_img, mode='I')
                             else: # Assume uint8 or other, convert to L
                                 pil_img = Image.fromarray(np_img).convert('L')
                        else:
                             raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

                    if pil_img is None:
                        raise ValueError("PIL Image creation failed.")

                    return pil_img

                except Exception as e:
                    traceback.print_exc()
                    return None            
                
                
                    
            def combine_image_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(10)
        
                # Initial placeholder ranges - will be updated by _update_overlay_slider_ranges()
                initial_dim_placeholder = 1000
                initial_x_range_min = -initial_dim_placeholder
                initial_x_range_max = initial_dim_placeholder * 2
                initial_y_range_min = -initial_dim_placeholder
                initial_y_range_max = initial_dim_placeholder * 2
        
        
                # --- Image 1 Group ---
                image1_group = QGroupBox("Image 1 Overlay")
                image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image1_layout = QGridLayout(image1_group)
                image1_layout.setSpacing(8)
        
                copy_image1_button = QPushButton("Copy Current Image")
                copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
                copy_image1_button.clicked.connect(self.save_image1)
                place_image1_button = QPushButton("Place Image 1")
                place_image1_button.setToolTip("Positions Image 1 based on sliders.")
                place_image1_button.clicked.connect(self.place_image1)
                remove_image1_button = QPushButton("Remove Image 1")
                remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
                remove_image1_button.clicked.connect(self.remove_image1)
        
                image1_layout.addWidget(copy_image1_button, 0, 0)
                image1_layout.addWidget(place_image1_button, 0, 1)
                image1_layout.addWidget(remove_image1_button, 0, 2)
        
                image1_layout.addWidget(QLabel("Horizontal Pos (px):"), 1, 0)
                self.image1_left_slider = QSlider(Qt.Horizontal)
                self.image1_left_slider.setRange(initial_x_range_min, initial_x_range_max)
                self.image1_left_slider.setValue(0)
                self.image1_left_slider.valueChanged.connect(self.place_image1) # Update position on value change
                image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2)
        
                image1_layout.addWidget(QLabel("Vertical Pos (px):"), 2, 0)
                self.image1_top_slider = QSlider(Qt.Horizontal)
                self.image1_top_slider.setRange(initial_y_range_min, initial_y_range_max)
                self.image1_top_slider.setValue(0)
                self.image1_top_slider.valueChanged.connect(self.place_image1) # Update position on value change
                image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2)
        
                image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
                self.image1_resize_slider = QSlider(Qt.Horizontal)
                self.image1_resize_slider.setRange(10, 300)
                self.image1_resize_slider.setValue(100)
                self.image1_resize_slider.valueChanged.connect(self.update_live_view) # Resize triggers full redraw
                self.image1_resize_label = QLabel("100%")
                self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
                image1_layout.addWidget(self.image1_resize_slider, 3, 1)
                image1_layout.addWidget(self.image1_resize_label, 3, 2)
                layout.addWidget(image1_group)
        
                # --- Image 2 Group ---
                image2_group = QGroupBox("Image 2 Overlay")
                image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image2_layout = QGridLayout(image2_group)
                # ... (similar setup for image2 sliders and connects) ...
                copy_image2_button = QPushButton("Copy Current Image")
                copy_image2_button.clicked.connect(self.save_image2)
                place_image2_button = QPushButton("Place Image 2")
                place_image2_button.clicked.connect(self.place_image2)
                remove_image2_button = QPushButton("Remove Image 2")
                remove_image2_button.clicked.connect(self.remove_image2)
                image2_layout.addWidget(copy_image2_button, 0, 0)
                image2_layout.addWidget(place_image2_button, 0, 1)
                image2_layout.addWidget(remove_image2_button, 0, 2)
        
                image2_layout.addWidget(QLabel("Horizontal Pos (px):"), 1, 0)
                self.image2_left_slider = QSlider(Qt.Horizontal)
                self.image2_left_slider.setRange(initial_x_range_min, initial_x_range_max)
                self.image2_left_slider.setValue(0)
                self.image2_left_slider.valueChanged.connect(self.place_image2)
                image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)
        
                image2_layout.addWidget(QLabel("Vertical Pos (px):"), 2, 0)
                self.image2_top_slider = QSlider(Qt.Horizontal)
                self.image2_top_slider.setRange(initial_y_range_min, initial_y_range_max)
                self.image2_top_slider.setValue(0)
                self.image2_top_slider.valueChanged.connect(self.place_image2)
                image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)
        
                image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
                self.image2_resize_slider = QSlider(Qt.Horizontal)
                self.image2_resize_slider.setRange(10, 300)
                self.image2_resize_slider.setValue(100)
                self.image2_resize_slider.valueChanged.connect(self.update_live_view)
                self.image2_resize_label = QLabel("100%")
                self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
                image2_layout.addWidget(self.image2_resize_slider, 3, 1)
                image2_layout.addWidget(self.image2_resize_label, 3, 2)
                layout.addWidget(image2_group)
        
        
                finalize_button = QPushButton("Rasterize Image")
                finalize_button.setToolTip("Permanently merges the placed overlays onto the main image.")
                finalize_button.clicked.connect(self.finalize_combined_image)
                layout.addWidget(finalize_button)
        
                layout.addStretch()
        
                # Call after UI elements are created, in case an image is already loaded
                self._update_overlay_slider_ranges()
        
                return tab
            
            def remove_image1(self):
                """Remove Image 1 and reset its sliders."""
                if hasattr(self, 'image1'):
                    # del self.image1
                    # del self.image1_original
                    try:
                        del self.image1_position
                    except:
                        pass
                    self.image1_left_slider.setValue(0)
                    self.image1_top_slider.setValue(0)
                    self.image1_resize_slider.setValue(100)
                    self.update_live_view()
            
            def remove_image2(self):
                """Remove Image 2 and reset its sliders."""
                if hasattr(self, 'image2'):
                    # del self.image2
                    # del self.image2_original
                    try:
                        del self.image2_position
                    except:
                        pass
                    self.image2_left_slider.setValue(0)
                    self.image2_top_slider.setValue(0)
                    self.image2_resize_slider.setValue(100)
                    self.update_live_view()
            
            def save_image1(self):
                if self.image:
                    self.image1 = self.image.copy()
                    self.image1_original = self.image1.copy()  # Save the original image for resizing
                    QMessageBox.information(self, "Success", "Image 1 copied.")
            
            def save_image2(self):
                if self.image:
                    self.image2 = self.image.copy()
                    self.image2_original = self.image2.copy()  # Save the original image for resizing
                    QMessageBox.information(self, "Success", "Image 2 copied.")
            
            def place_image1(self):
                if hasattr(self, 'image1'):
                    self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
                    self.update_live_view()
            
            def place_image2(self):
                if hasattr(self, 'image2'):
                    self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                    self.update_live_view()
            
            
            
            def finalize_combined_image(self):
                """
                Rasterizes overlays and annotations onto the image by generating
                a high-resolution canvas identical to the preview rendering process
                (but without guides/previews) and assigning it as the new self.image.
                This ensures visual consistency between preview and final result.
                The final image resolution might be higher than the original if render_scale > 1.
                """
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Warning", "No base image loaded.")
                    return

                # --- Save state before modifying ---
                self.save_state()

                # --- 1. Determine Render/Canvas Dimensions (Same as Preview) ---
                render_scale = 3 # Use the same scale factor as the preview
                try:
                    # Use current view dimensions to determine target render size
                    # Fallback if view is not ready
                    view_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 600
                    view_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 400
                    target_render_width = view_width * render_scale
                    target_render_height = view_height * render_scale
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Could not calculate render dimensions: {e}")
                    return

                # --- 2. Prepare the Base Image for Rendering ---
                # Scale the *current* self.image to fit within the target render dimensions,
                # just like it's done at the start of update_live_view before calling render_image_on_canvas.
                # Check if self.image is valid before scaling
                if self.image.isNull() or self.image.width() <= 0 or self.image.height() <= 0:
                     QMessageBox.critical(self, "Error", "Current base image is invalid before scaling.")
                     return

                scaled_image_for_render = self.image.scaled(
                    target_render_width,
                    target_render_height,
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation,
                )
                if scaled_image_for_render.isNull():
                    QMessageBox.critical(self, "Error", "Failed to scale base image for final rendering.")
                    return

                # --- 3. Create the High-Resolution Target Canvas ---
                # Use ARGB32_Premultiplied for safe drawing with alpha
                final_canvas = QImage(target_render_width, target_render_height, QImage.Format_ARGB32_Premultiplied)
                if final_canvas.isNull():
                    QMessageBox.critical(self, "Error", "Failed to create final high-resolution canvas.")
                    return
                final_canvas.fill(Qt.transparent) # Fill with transparent background

                # --- 4. Render onto the High-Res Canvas using the Preview Logic ---
                # Call the *exact same* rendering function used for the preview,
                # but disable drawing of guides and temporary previews.
                # Pass the scaled base image prepared in step 2.
                # Pass render_scale so elements inside render_image_on_canvas scale correctly.
                try:
                    # x_start and y_start are typically 0 when rendering the current self.image
                    # unless cropping offsets need to be handled differently, but render_image_on_canvas
                    # should already handle positioning based on the scaled_image provided.
                    self.render_image_on_canvas(
                        canvas=final_canvas,
                        scaled_image=scaled_image_for_render,
                        x_start=0, # Start position relative to the scaled_image itself
                        y_start=0,
                        render_scale=render_scale,
                        draw_guides=False # IMPORTANT: Disable guides for final output
                    )
                    # Inside render_image_on_canvas, ensure shape previews are also skipped if needed,
                    # although technically they shouldn't exist when finalize is called.
                    # Consider adding a flag to render_image_on_canvas like `is_finalizing=True`
                    # to explicitly skip previews if necessary.

                except Exception as e:
                     QMessageBox.critical(self, "Render Error", f"Failed during final rendering: {e}")
                     traceback.print_exc()
                     return

                # --- 5. Assign the High-Resolution Canvas as the New Image ---
                if final_canvas.isNull():
                     QMessageBox.critical(self, "Error", "Final rendered canvas became invalid.")
                     return

                self.image = final_canvas # The new self.image IS the high-res render
                self.is_modified = True

                # --- 6. Update Backups and State ---
                # The backups now store the high-resolution rendered image
                self.image_before_padding = self.image.copy() # New baseline includes baked-in elements
                self.image_contrasted = self.image.copy()
                self.image_before_contrast = self.image.copy() # Reset contrast baseline
                self.image_padded = True # Consider the result "padded" with overlays
                self._update_marker_slider_ranges() # Adjust ranges for new dimensions
                self._update_overlay_slider_ranges()
                # --- 7. Cleanup and UI Update ---
                

                self._update_preview_label_size() # Update label based on potentially larger image
                self._update_status_bar()         # Reflect new dimensions/depth
                self.remove_image1() # Remove buffer overlays
                self.remove_image2()
                
                self.update_live_view()           # Refresh display

                QMessageBox.information(self, "Success", "The overlays and annotations have been rasterized onto the image.\n(Note: Final image resolution might be higher than original)")
            
                
            


            def font_and_image_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(15) # Add spacing between groups

                # --- Font Options Group ---
                font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
                font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                font_options_layout = QGridLayout(font_options_group)
                font_options_layout.setSpacing(8)

                # Font type
                font_type_label = QLabel("Font Family:")
                self.font_combo_box = QFontComboBox()
                self.font_combo_box.setEditable(False)
                self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
                self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_type_label, 0, 0)
                font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

                # Font size
                font_size_label = QLabel("Font Size:")
                self.font_size_spinner = QSpinBox()
                self.font_size_spinner.setRange(6, 72)  # Adjusted range
                self.font_size_spinner.setValue(self.font_size) # Use initialized value
                self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_size_label, 1, 0)
                font_options_layout.addWidget(self.font_size_spinner, 1, 1)

                # Font color
                self.font_color_button = QPushButton("Font Color")
                self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
                self.font_color_button.clicked.connect(self.select_font_color)
                self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
                font_options_layout.addWidget(self.font_color_button, 1, 2)

                # Font rotation (Top/Bottom)
                font_rotation_label = QLabel("Top Label Rotation:") # Specific label
                self.font_rotation_input = QSpinBox()
                self.font_rotation_input.setRange(-180, 180)
                self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
                self.font_rotation_input.setSuffix(" °") # Add degree symbol
                self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_rotation_label, 2, 0)
                font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

                layout.addWidget(font_options_group)


                # --- Image Adjustments Group ---
                img_adjust_group = QGroupBox("Image Adjustments (Levels & Gamma)") # MODIFIED Group Title
                img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                img_adjust_layout = QGridLayout(img_adjust_group)
                img_adjust_layout.setSpacing(8)

                # Display Black Point Slider
                self.black_point_label = QLabel("Black Point:") # MODIFIED LABEL
                self.black_point_slider = QSlider(Qt.Horizontal)
                self.black_point_slider.setRange(0, 65535) # Initial default range, will be updated
                self.black_point_slider.setValue(0)      # Default black point
                self.black_point_slider.valueChanged.connect(self.update_image_levels_and_gamma) # Connect to new handler
                self.black_point_value_label = QLabel("0") # Display actual pixel value
                self.black_point_value_label.setMinimumWidth(50)
                self.black_point_slider.valueChanged.connect(lambda val, lbl=self.black_point_value_label: lbl.setText(f"{val}"))
                self.black_point_slider.setToolTip("Set the black point for display. Pixels below this value will be black.")
                img_adjust_layout.addWidget(self.black_point_label, 0, 0)
                img_adjust_layout.addWidget(self.black_point_slider, 0, 1)
                img_adjust_layout.addWidget(self.black_point_value_label, 0, 2)

                # Display White Point Slider
                self.white_point_label = QLabel("White Point:") # MODIFIED LABEL
                self.white_point_slider = QSlider(Qt.Horizontal)
                self.white_point_slider.setRange(0, 65535) # Initial default range
                self.white_point_slider.setValue(65535)    # Default white point
                self.white_point_slider.valueChanged.connect(self.update_image_levels_and_gamma) # Connect to new handler
                self.white_point_value_label = QLabel("65535") # Display actual pixel value
                self.white_point_value_label.setMinimumWidth(50)
                self.white_point_slider.valueChanged.connect(lambda val, lbl=self.white_point_value_label: lbl.setText(f"{val}"))
                self.white_point_slider.setToolTip("Set the white point for display. Pixels above this value will be white.")
                img_adjust_layout.addWidget(self.white_point_label, 1, 0)
                img_adjust_layout.addWidget(self.white_point_slider, 1, 1)
                img_adjust_layout.addWidget(self.white_point_value_label, 1, 2)


                # Gamma Adjustment Slider (remains similar)
                gamma_label = QLabel("Gamma:")
                self.gamma_slider = QSlider(Qt.Horizontal)
                self.gamma_slider.setRange(10, 500)  # Range 0.1 to 3.0 (factor = value / 100.0)
                self.gamma_slider.setValue(100)      # Default 1.0
                self.gamma_slider.valueChanged.connect(self.update_image_levels_and_gamma) # Connect to new handler
                self.gamma_value_label = QLabel("1.00") # Display factor
                self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
                self.gamma_slider.setToolTip("Adjust mid-tone brightness after levels adjustment.")
                img_adjust_layout.addWidget(gamma_label, 2, 0)
                img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
                img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


                # Separator
                img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

                # Action Buttons
                btn_layout = QHBoxLayout()
                self.bw_button = QPushButton("Grayscale")
                self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
                self.bw_button.clicked.connect(self.convert_to_black_and_white)
                invert_button = QPushButton("Invert")
                invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
                invert_button.clicked.connect(self.invert_image)
                reset_button = QPushButton("Reset Adjustments")
                reset_button.setToolTip("Reset Black/White Points and Gamma sliders to default.") # MODIFIED TOOLTIP
                reset_button.clicked.connect(self.reset_levels_and_gamma) # MODIFIED connection
                btn_layout.addWidget(self.bw_button)
                btn_layout.addWidget(invert_button)
                btn_layout.addStretch() # Push reset button to the right
                btn_layout.addWidget(reset_button)

                img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

                layout.addWidget(img_adjust_group)
                layout.addStretch() # Push groups up
                return tab
            
            def reset_levels_and_gamma(self): # Renamed from reset_gamma_contrast
                # This method resets the UI sliders and then triggers an update.
                # The actual image reset to "before contrast" state happens via update_image_levels_and_gamma
                # when it uses self.image_contrasted as its base.
                
                self._update_level_slider_ranges_and_defaults() # This sets sliders to 0 and max_val

                if hasattr(self, 'gamma_slider'):
                    self.gamma_slider.blockSignals(True)
                    self.gamma_slider.setValue(100) # Reset gamma to 1.0
                    self.gamma_slider.blockSignals(False)
                    if hasattr(self, 'gamma_value_label'): self.gamma_value_label.setText("1.00")
                
                # Crucially, after resetting sliders, call update_image_levels_and_gamma
                # to apply these defaults to self.image_contrasted (which should be the pristine version at this point if contrast_applied was false, or the version before last adjustment).
                # If contrast_applied was true, self.image_contrasted holds the "before current adjustment" state.
                # If contrast_applied was false, update_image_levels_and_gamma will set it up.
                
                # To ensure we are resetting from the true "before any contrast" state:
                if hasattr(self, 'image_before_contrast') and self.image_before_contrast and not self.image_before_contrast.isNull():
                     self.image_contrasted = self.image_before_contrast.copy()
                elif self.image_master and not self.image_master.isNull(): # Fallback to master
                    self.image_contrasted = self.image_master.copy()
                elif self.image and not self.image.isNull(): # Fallback to current if others are missing
                    self.image_contrasted = self.image.copy()
                # else: # No valid base image to reset from

                self.contrast_applied = True # Mark that an "adjustment" (the reset) has been made
                self.update_image_levels_and_gamma() # This will now apply the 0-max levels and gamma 1.0
                
            def apply_levels_gamma(self, qimage_base, black_point_ui, white_point_ui, gamma_ui_factor):
                if not qimage_base or qimage_base.isNull():
                    return qimage_base

                try:
                    img_array = self.qimage_to_numpy(qimage_base)
                    if img_array is None: raise ValueError("NumPy conversion failed.")

                    img_array_float = img_array.astype(np.float64) # Work with float for calculations
                    original_dtype = img_array.dtype # Preserve original dtype for output

                    # --- MODIFIED: Determine max_dtype_val based on the actual NumPy array's dtype ---
                    if original_dtype == np.uint16:
                        max_dtype_val = 65535.0
                    elif original_dtype == np.uint8:
                        max_dtype_val = 255.0
                    elif np.issubdtype(original_dtype, np.floating): # If it's already float (e.g. 0-1)
                        # If input is float, assume it's already normalized (0-1) or needs specific handling.
                        # For levels, we typically work with integer ranges. If a float image is passed here,
                        # it implies it might have been pre-processed.
                        # Let's assume if it's float, its range is 0-1 for this specific levels function,
                        # and we'll scale black/white points accordingly.
                        # Or, better, this function expects integer-like input for black/white points.
                        # If a float image (0-1) is passed, the black/white points (0-65535) are out of scale.
                        # This function should ideally operate on integer type arrays or scale them.
                        # For now, if original_dtype is float, we will convert it to uint8 for levels.
                        # This is a simplification. A more robust system might have different
                        # level adjustment logic for float images.
                        if np.max(img_array_float) <= 1.0 and np.min(img_array_float) >= 0.0:
                            print("Warning: Applying levels/gamma to float (0-1) image. Converting to 8-bit for levels.")
                            img_array_float = (img_array_float * 255.0).astype(np.float64)
                            original_dtype = np.uint8 # Treat as 8-bit for the rest of this function
                            max_dtype_val = 255.0
                        else: # For floats not in 0-1, try to guess max based on data.
                            max_dtype_val = np.max(img_array_float) if np.any(img_array_float) else 1.0
                    else: # Fallback for other integer types if any
                        # This assumes other integer types are like 8-bit for display purposes.
                        print(f"Warning: Unexpected original_dtype '{original_dtype}' in apply_levels_gamma. Assuming 8-bit range (0-255).")
                        max_dtype_val = 255.0

                    if max_dtype_val == 0: max_dtype_val = 1.0 # Avoid division by zero if image is all black

                    # Slider values are absolute (0-65535). We need to scale them if max_dtype_val is different.
                    # For example, if image is 8-bit (max_dtype_val=255) and slider is at 32768,
                    # the effective black point should be (32768/65535) * 255.
                    scale_factor_slider_to_img_range = max_dtype_val / 65535.0

                    current_black = float(black_point_ui) * scale_factor_slider_to_img_range
                    current_white = float(white_point_ui) * scale_factor_slider_to_img_range

                    # Ensure black < white for the current image's actual data range
                    if current_black >= current_white:
                        if current_black >= max_dtype_val -1 :
                            current_black = max_dtype_val - (2.0 if max_dtype_val > 1 else 0.1) # ensure black is less than white
                            current_white = max_dtype_val - (1.0 if max_dtype_val > 1 else 0.05)
                        else:
                            current_white = current_black + (1.0 if max_dtype_val > 1 else 0.05) # Ensure white is greater

                        if current_white > max_dtype_val: # cap white point
                            current_white = max_dtype_val
                        if current_black >= current_white: # final safety for black
                             current_black = current_white - (1.0 if max_dtype_val > 1 else 0.05)
                        current_black = max(0, current_black) # Black point cannot be less than 0


                    denominator = current_white - current_black
                    if abs(denominator) < 1e-9: # Increased tolerance for float comparisons
                        denominator = 1e-9 if denominator >= 0 else -1e-9


                    # Process each channel if color, or the single channel if grayscale
                    if img_array_float.ndim == 3: # Color image
                        processed_channels = []
                        num_channels = img_array_float.shape[2]
                        channels_to_adjust = min(num_channels, 3)

                        for i in range(channels_to_adjust):
                            channel_data = img_array_float[..., i]
                            channel_levels_adjusted = (channel_data - current_black) / denominator
                            channel_levels_adjusted = np.clip(channel_levels_adjusted, 0.0, 1.0)

                            safe_gamma = max(0.01, gamma_ui_factor)
                            channel_gamma_adjusted = np.power(channel_levels_adjusted, safe_gamma)
                            channel_gamma_adjusted = np.clip(channel_gamma_adjusted, 0.0, 1.0)
                            
                            processed_channels.append(channel_gamma_adjusted * max_dtype_val)

                        img_array_final_float = np.stack(processed_channels, axis=-1)

                        if num_channels == 4:
                            alpha_channel = img_array_float[..., 3]
                            img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

                    elif img_array_float.ndim == 2: # Grayscale image
                        img_levels_adjusted = (img_array_float - current_black) / denominator
                        img_levels_adjusted = np.clip(img_levels_adjusted, 0.0, 1.0)

                        safe_gamma = max(0.01, gamma_ui_factor)
                        img_gamma_adjusted = np.power(img_levels_adjusted, safe_gamma)
                        img_gamma_adjusted = np.clip(img_gamma_adjusted, 0.0, 1.0)

                        img_array_final_float = img_gamma_adjusted * max_dtype_val
                    else:
                        return qimage_base

                    img_array_final = img_array_final_float.astype(original_dtype)

                    result_qimage = self.numpy_to_qimage(img_array_final)
                    if result_qimage.isNull():
                        raise ValueError("Conversion back to QImage failed after levels/gamma.")
                    return result_qimage

                except Exception as e:
                    print(f"Error in apply_levels_gamma: {e}")
                    traceback.print_exc()
                    return qimage_base
            
            def update_image_levels_and_gamma(self):
                """Applies levels and gamma adjustments based on current slider values."""
                if not self.image or self.image.isNull():
                    return
                if not hasattr(self, 'image_contrasted') or not self.image_contrasted or self.image_contrasted.isNull():
                     if self.image_before_contrast and not self.image_before_contrast.isNull():
                         self.image_contrasted = self.image_before_contrast.copy()
                     elif self.image_master and not self.image_master.isNull():
                         self.image_contrasted = self.image_master.copy()
                     else: # Should not happen if self.image is valid
                          self.image_contrasted = self.image.copy()


                if self.contrast_applied == False: # Ensure base image for contrast is set
                    self.image_before_contrast = self.image.copy() # self.image is the one before any adjustments
                    self.image_contrasted = self.image_before_contrast.copy()
                    self.contrast_applied = True

                try:
                    black_point_val = 0
                    white_point_val = 255
                    gamma_val = 100 # Slider value for gamma (1.0 factor)

                    if hasattr(self, 'black_point_slider'): black_point_val = self.black_point_slider.value()
                    if hasattr(self, 'white_point_slider'): white_point_val = self.white_point_slider.value()
                    if hasattr(self, 'gamma_slider'): gamma_val = self.gamma_slider.value()
                    
                    gamma_factor = gamma_val / 100.0

                    # Ensure black point is less than white point for UI logic
                    if black_point_val >= white_point_val:
                        # If black >= white, clamp white to be slightly above black
                        # This prevents issues in apply_levels_gamma and provides a very high contrast visual
                        if self.black_point_slider.value() >= self.white_point_slider.value(): # Check slider values directly
                            self.white_point_slider.blockSignals(True)
                            self.white_point_slider.setValue(min(self.black_point_slider.value() + 1, self.white_point_slider.maximum()))
                            self.white_point_slider.blockSignals(False)
                            white_point_val = self.white_point_slider.value() # Get corrected value
                            if hasattr(self, 'white_point_value_label'): self.white_point_value_label.setText(str(white_point_val))

                    # Use self.image_contrasted as the base for adjustment
                    # self.image_contrasted holds the image state *before* current levels/gamma
                    # OR it holds the image state after a "permanent" operation like grayscale/invert/padding.
                    base_image_for_adjustment = self.image_contrasted
                    if not base_image_for_adjustment or base_image_for_adjustment.isNull():
                        # Fallback if image_contrasted is somehow invalid
                        base_image_for_adjustment = self.image_master.copy() if self.image_master else self.image.copy()


                    self.image = self.apply_levels_gamma(base_image_for_adjustment, black_point_val, white_point_val, gamma_factor)
                    self.update_live_view()
                except Exception as e:
                    print(f"Error in update_image_levels_and_gamma: {e}")
                    traceback.print_exc()
            
            def _update_level_slider_ranges_and_defaults(self):
                """
                Sets the Black/White Point slider default values.
                Black point defaults to 0.
                White point slider default value is set based on image format:
                - 255 for Format_Grayscale8 or Format_Mono.
                - 65535 for Format_Grayscale16 and all other formats (including color).
                The slider *ranges* are fixed at 0-65535 in the UI creation.
                """
                default_white_point_slider_value = 65535 # Default to 65535 for most cases

                if self.image and not self.image.isNull():
                    current_format = self.image.format()

                    if current_format == QImage.Format_Grayscale8:
                        default_white_point_slider_value = 255
                    elif current_format == QImage.Format_Mono:
                        # For a 1-bit image, max actual value is 1.
                        # Setting slider default to 255 gives more UI range.
                        # apply_levels_gamma will correctly scale from slider (0-65535) to image range (0-1).
                        default_white_point_slider_value = 255
                    # For QImage.Format_Grayscale16, and all color formats (RGB888, ARGB32, etc.),
                    # or any other/unknown format, we'll default the white point *slider value* to 65535.
                    # The apply_levels_gamma function handles the true data range of these images.

                # Black point slider: Range is 0-65535, Value defaults to 0
                if hasattr(self, 'black_point_slider'):
                    self.black_point_slider.blockSignals(True)
                    self.black_point_slider.setValue(0)
                    self.black_point_slider.blockSignals(False)
                    if hasattr(self, 'black_point_value_label'): self.black_point_value_label.setText("0")

                # White point slider: Range is 0-65535, Value defaults based on above logic
                if hasattr(self, 'white_point_slider'):
                    self.white_point_slider.blockSignals(True)
                    self.white_point_slider.setValue(default_white_point_slider_value)
                    self.white_point_slider.blockSignals(False)
                    if hasattr(self, 'white_point_value_label'): self.white_point_value_label.setText(str(default_white_point_slider_value))

            
            def _update_color_button_style(self, button, color):
                """ Helper to update button background color preview """
                if color.isValid():
                    button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
                else:
                    button.setStyleSheet("") # Reset to default stylesheet
            
            def create_separator(self):
                """Creates a horizontal separator line."""
                line = QFrame()
                line.setFrameShape(QFrame.HLine)
                line.setFrameShadow(QFrame.Sunken)
                return line
            
            
            def create_cropping_tab(self):
                """Create the Cropping tab with Rectangle Draw and Slider options."""
                tab = QWidget()
                layout = QVBoxLayout(tab)


                # --- Alignment Group (Keep as is - assume it's defined elsewhere) ---
                alignment_params_group = QGroupBox("Alignment Options")
                alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                # Use a QVBoxLayout for the main group layout
                alignment_group_layout = QVBoxLayout(alignment_params_group)
                alignment_group_layout.setSpacing(8) # Spacing between the two rows

                # --- Row 1: Guides, Rotation Controls ---
                rotation_controls_layout = QHBoxLayout()
                rotation_controls_layout.setSpacing(6) # Spacing within the row

                self.show_guides_label = QLabel("Show Guide Lines:")
                self.show_guides_checkbox = QCheckBox("", self)
                self.show_guides_checkbox.setChecked(False)
                self.show_guides_checkbox.setToolTip("Show a center linter to align/rotate the image properly. Shortcut: CTRL+G or CMD+G")
                self.show_guides_checkbox.stateChanged.connect(self.update_live_view)

                self.orientation_label = QLabel("Rotation Angle (0.00°)")
                # Make label width flexible but give it a minimum
                self.orientation_label.setMinimumWidth(150)
                # Allow label to shrink/grow slightly if needed:
                self.orientation_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)

                self.orientation_slider = QSlider(Qt.Horizontal)
                self.orientation_slider.setRange(-3600, 3600)
                self.orientation_slider.setValue(0)
                self.orientation_slider.setSingleStep(1)
                self.orientation_slider.valueChanged.connect(self._update_rotation_label)
                # Update preview only when slider is released for performance
                self.orientation_slider.valueChanged.connect(self.update_live_view)
                # Make slider expand to take available space
                self.orientation_slider.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                self.align_button = QPushButton("Apply Rotation")
                self.align_button.clicked.connect(self.align_image)

                self.reset_align_button = QPushButton("Reset Rotation")
                self.reset_align_button.clicked.connect(self.reset_align_image)

                # Add widgets to the first row layout
                rotation_controls_layout.addWidget(self.show_guides_label)
                rotation_controls_layout.addWidget(self.show_guides_checkbox)
                rotation_controls_layout.addSpacing(10) # Add a small visual gap
                rotation_controls_layout.addWidget(self.orientation_label)
                rotation_controls_layout.addWidget(self.orientation_slider) # Let slider expand
                rotation_controls_layout.addWidget(self.align_button)
                rotation_controls_layout.addWidget(self.reset_align_button)

                # Add the first row layout to the main group layout
                alignment_group_layout.addLayout(rotation_controls_layout)

                # --- Row 2: Flip Controls ---
                flip_controls_layout = QHBoxLayout()
                flip_controls_layout.setSpacing(6) # Spacing between flip buttons

                self.flip_vertical_button = QPushButton("Flip Vertical")
                self.flip_vertical_button.setToolTip("Flips the image in vertical direction")
                self.flip_vertical_button.clicked.connect(self.flip_vertical)
                # Make buttons expand equally to fill width
                self.flip_vertical_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                self.flip_horizontal_button = QPushButton("Flip Horizontal")
                self.flip_horizontal_button.setToolTip("Flips the image in horizontal direction")
                self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
                # Make buttons expand equally to fill width
                self.flip_horizontal_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                # Add widgets to the second row layout
                # No stretch needed here as buttons expand
                flip_controls_layout.addWidget(self.flip_vertical_button)
                flip_controls_layout.addWidget(self.flip_horizontal_button)

                # Add the second row layout to the main group layout
                alignment_group_layout.addLayout(flip_controls_layout)
                
                # guide_layout.addStretch()
                

                alignment_params_group.setLayout(alignment_group_layout)
                layout.addWidget(alignment_params_group)
                # --- End Alignment Group ---

                # --- Skew Fix Group (Keep as is - assume it's defined elsewhere) ---
                taper_skew_group = QGroupBox("Skew Fix")
                taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                # ... (Add skew widgets here) ...
                # Example structure:
                taper_skew_layout = QHBoxLayout()
                self.taper_skew_label = QLabel("Tapering Skew (0.00)")
                self.taper_skew_label.setToolTip("Can be used to fix gel distortion by adjusting the skewness/broadening of the gel at the top or bottom region")
                self.taper_skew_label.setFixedWidth(150)
                self.taper_skew_slider = QSlider(Qt.Horizontal)
                self.taper_skew_slider.setRange(-70, 70)
                self.taper_skew_slider.setValue(0)
                # self.taper_skew_slider.valueChanged.connect(self._update_skew_label)
                self.taper_skew_slider.valueChanged.connect(self.update_live_view)
                self.skew_button = QPushButton("Apply Skew")
                self.skew_button.clicked.connect(self.update_skew)
                taper_skew_layout.addWidget(self.taper_skew_label)
                taper_skew_layout.addWidget(self.taper_skew_slider)
                taper_skew_layout.addWidget(self.skew_button)
                taper_skew_group.setLayout(taper_skew_layout)
                layout.addWidget(taper_skew_group)
                # --- End Skew Fix Group ---


                # --- START: Modified Cropping Group ---
                cropping_params_group = QGroupBox("Cropping Options")
                cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                cropping_layout = QVBoxLayout(cropping_params_group)      
                

                # --- Draw Rectangle Button ---
                self.draw_crop_rect_button = QPushButton("Draw Crop Rectangle")
                self.draw_crop_rect_button.setToolTip("Click and drag on the image to define the crop area.\nThis will enable the sliders below for fine-tuning.")
                self.draw_crop_rect_button.setCheckable(True)
                self.draw_crop_rect_button.clicked.connect(self.toggle_rectangle_crop_mode)
                self.apply_crop_button = QPushButton("Apply Crop")
                self.apply_crop_button.setToolTip("Apply the defined crop (rectangle or sliders).")
                self.apply_crop_button.clicked.connect(self.update_crop)
                cropping_layout.addWidget(self.draw_crop_rect_button)
                cropping_layout.addWidget(self.apply_crop_button)

                # --- Separator ---
                cropping_layout.addWidget(self.create_separator())

                # --- Sliders for Fine-Tuning ---
                crop_slider_layout = QGridLayout()

                # --- Define Slider Range and Precision ---
                self.crop_slider_min = 0
                self.crop_slider_max = 10000 # Represents 0.00% to 100.00%
                self.crop_slider_precision_factor = 100.0 # Divide slider value by this

                # Helper to create label for slider value display
                def create_value_label(initial_value=0.0):
                    # Format to 2 decimal places for hundredths of percent
                    lbl = QLabel(f"{initial_value:.2f}%")
                    lbl.setMinimumWidth(50) # Ensure space for "100.00%"
                    lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    return lbl

                # --- Left Crop Slider (X Start) ---
                crop_x_start_label = QLabel("Crop Left:")
                self.crop_x_start_slider = QSlider(Qt.Horizontal)
                self.crop_x_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_x_start_slider.setValue(self.crop_slider_min)      # Default: Start at 0.00%
                self.crop_x_start_slider.setToolTip("Adjust the left edge of the crop area (enabled after drawing).")
                self.crop_x_start_value_label = create_value_label(0.00)
                self.crop_x_start_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_x_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_x_start_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_x_start_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
                crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
                crop_slider_layout.addWidget(self.crop_x_start_value_label, 0, 2)

                # --- Right Crop Slider (X End) ---
                crop_x_end_label = QLabel("Crop Right:")
                self.crop_x_end_slider = QSlider(Qt.Horizontal)
                self.crop_x_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_x_end_slider.setValue(self.crop_slider_max)    # Default: End at 100.00%
                self.crop_x_end_slider.setToolTip("Adjust the right edge of the crop area (enabled after drawing).")
                self.crop_x_end_value_label = create_value_label(100.00)
                self.crop_x_end_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_x_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_x_end_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_x_end_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_x_end_label, 0, 3)
                crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 4)
                crop_slider_layout.addWidget(self.crop_x_end_value_label, 0, 5)

                # --- Top Crop Slider (Y Start) ---
                crop_y_start_label = QLabel("Crop Top:")
                self.crop_y_start_slider = QSlider(Qt.Horizontal)
                self.crop_y_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_y_start_slider.setValue(self.crop_slider_min)
                self.crop_y_start_slider.setToolTip("Adjust the top edge of the crop area (enabled after drawing).")
                self.crop_y_start_value_label = create_value_label(0.00)
                self.crop_y_start_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_y_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_y_start_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_y_start_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
                crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
                crop_slider_layout.addWidget(self.crop_y_start_value_label, 1, 2)

                # --- Bottom Crop Slider (Y End) ---
                crop_y_end_label = QLabel("Crop Bottom:")
                self.crop_y_end_slider = QSlider(Qt.Horizontal)
                self.crop_y_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_y_end_slider.setValue(self.crop_slider_max)
                self.crop_y_end_slider.setToolTip("Adjust the bottom edge of the crop area (enabled after drawing).")
                self.crop_y_end_value_label = create_value_label(100.00)
                self.crop_y_end_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_y_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_y_end_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_y_end_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_y_end_label, 1, 3)
                crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 4)
                crop_slider_layout.addWidget(self.crop_y_end_value_label, 1, 5)

                # Make sliders expand horizontally
                crop_slider_layout.setColumnStretch(1, 1)
                crop_slider_layout.setColumnStretch(4, 1)
                cropping_layout.addLayout(crop_slider_layout)


                layout.addWidget(cropping_params_group)
                # --- END: Modified Cropping Group ---

                layout.addStretch()
                return tab
            
            def _update_rotation_label(self, value):
                """Updates the rotation label text."""
                if hasattr(self, 'orientation_label'):
                    orientation = value / 20.0
                    self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

            # def _update_skew_label(self, value):
            #     """Updates the skew label text."""
            #     if hasattr(self, 'taper_skew_label'):
            #         taper_value = value / 100.0
            #         self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

            def _update_crop_from_sliders(self):
                """Updates crop rectangle based on slider percentages and ensures labels update."""
                if not self.image or self.image.isNull():
                    return
            
                # Read integer slider values
                try:
                    x_start_val = self.crop_x_start_slider.value()
                    x_end_val = self.crop_x_end_slider.value()
                    y_start_val = self.crop_y_start_slider.value()
                    y_end_val = self.crop_y_end_slider.value()
                except AttributeError:
                    print("Error: Crop sliders not fully initialized.")
                    return
            
                # Convert to percentage
                x_start_perc = x_start_val / self.crop_slider_precision_factor
                x_end_perc = x_end_val / self.crop_slider_precision_factor
                y_start_perc = y_start_val / self.crop_slider_precision_factor
                y_end_perc = y_end_val / self.crop_slider_precision_factor
            
                # Validation: Ensure Start <= End
                x_start = min(x_start_perc, x_end_perc)
                x_end = max(x_start_perc, x_end_perc)
                y_start = min(y_start_perc, y_end_perc)
                y_end = max(y_start_perc, y_end_perc)
            
                # Convert corrected percentages back to integer slider values
                x_start_val_corr = int(round(x_start * self.crop_slider_precision_factor))
                x_end_val_corr = int(round(x_end * self.crop_slider_precision_factor))
                y_start_val_corr = int(round(y_start * self.crop_slider_precision_factor))
                y_end_val_corr = int(round(y_end * self.crop_slider_precision_factor))
            
                # Reflect potentially swapped values back in sliders (only if different)
                # Use a flag to track if signals were blocked to avoid redundant label updates later
                signals_were_blocked = False
                sliders_to_check = [
                    (self.crop_x_start_slider, x_start_val_corr), (self.crop_x_end_slider, x_end_val_corr),
                    (self.crop_y_start_slider, y_start_val_corr), (self.crop_y_end_slider, y_end_val_corr)
                ]
                for slider, correct_value in sliders_to_check:
                    if slider.value() != correct_value:
                        slider.blockSignals(True)
                        slider.setValue(correct_value)
                        slider.blockSignals(False)
                        signals_were_blocked = True # Record that we manually set a value
            
                # --- Explicitly Update Labels AFTER potential corrections ---
                # This ensures labels always reflect the validated slider state, even if
                # the correction logic blocked the initial valueChanged signal for the label lambda.
                # Use the corrected integer values derived above.
                try:
                    self.crop_x_start_value_label.setText(f"{x_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_x_end_value_label.setText(f"{x_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_start_value_label.setText(f"{y_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_end_value_label.setText(f"{y_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                except AttributeError:
                    print("Error: Crop value labels not fully initialized for explicit update.")
                    # Handle case where labels might not exist yet if called too early
                    pass
                # --- End Explicit Label Update ---
            
            
                # Deactivate drawing mode if sliders are used
                if self.crop_rectangle_mode:
                    self.cancel_rectangle_crop_mode()
            
                # Calculate Image Coordinates using corrected percentages
                img_w = float(self.image.width())
                img_h = float(self.image.height())
                if img_w <= 0 or img_h <= 0: return
            
                img_x = img_w * (x_start / 100.0)
                img_y = img_h * (y_start / 100.0)
                img_crop_w = img_w * ((x_end - x_start) / 100.0)
                img_crop_h = img_h * ((y_end - y_start) / 100.0)
            
                # Store the calculated *image* coordinates
                self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))
            
                # Calculate View Coordinates for Preview
                label_w = float(self.live_view_label.width())
                label_h = float(self.live_view_label.height())
                if label_w <= 0 or label_h <= 0: return
            
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                if scale_factor <= 1e-9: scale_factor = 1
            
                display_img_w = img_w * scale_factor
                display_img_h = img_h * scale_factor
                display_offset_x = (label_w - display_img_w) / 2.0
                display_offset_y = (label_h - display_img_h) / 2.0
            
                view_x_start = display_offset_x + img_x * scale_factor
                view_y_start = display_offset_y + img_y * scale_factor
                view_x_end = view_x_start + img_crop_w * scale_factor
                view_y_end = view_y_start + img_crop_h * scale_factor
            
                # Update the visual preview rectangle
                self.live_view_label.crop_rect_final_view = QRectF(
                    QPointF(view_x_start, view_y_start),
                    QPointF(view_x_end, view_y_end)
                ).normalized()
                self.live_view_label.drawing_crop_rect = False
            
                self.update_live_view()
            
            def toggle_rectangle_crop_mode(self, checked):
                if checked:
                    self.enable_rectangle_crop_mode()
                else:
                    self.cancel_rectangle_crop_mode()

            def enable_rectangle_crop_mode(self):
                if self.crop_rectangle_mode: return
                self.marker_mode = None # Deactivate other modes
                self.crop_rectangle_mode = True
                self.crop_rectangle_coords = None
                self.live_view_label.clear_crop_preview()
                self.draw_crop_rect_button.setChecked(True)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label._custom_left_click_handler_from_app = self.start_crop_rectangle
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_crop_rectangle_preview
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_crop_rectangle

            def cancel_rectangle_crop_mode(self):
                """Deactivates the rectangle drawing mode."""
                if not self.crop_rectangle_mode: return

                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None
                if hasattr(self, 'draw_crop_rect_button'):
                     self.draw_crop_rect_button.setChecked(False)
                self._reset_live_view_label_custom_handlers() # Use helper
                self.update_live_view()

            def start_crop_rectangle(self, event):
                """Handles mouse press to start drawing the crop rectangle."""
                if not self.crop_rectangle_mode:
                    # If not in crop mode, pass event to the base class or intended handler
                    super(LiveViewLabel, self.live_view_label.__class__).mousePressEvent(self.live_view_label, event)
                    return

                if event.button() == Qt.LeftButton:
                    # Get start point in *view* coordinates (unzoomed)
                    self.crop_rect_start_view = self.live_view_label.transform_point(event.position())
                    # Tell LiveViewLabel to start drawing the preview
                    self.live_view_label.start_crop_preview(self.crop_rect_start_view)

            def update_crop_rectangle_preview(self, event):
                """Handles mouse move to update the crop rectangle preview."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                     # If not in crop mode or not dragging, pass event along
                     super(LiveViewLabel, self.live_view_label.__class__).mouseMoveEvent(self.live_view_label, event)
                     return

                # Check if left button is held down (QApplication.mouseButtons() might be needed)
                if event.buttons() & Qt.LeftButton:
                    current_point_view = self.live_view_label.transform_point(event.position())
                    # Tell LiveViewLabel to update the preview
                    self.live_view_label.update_crop_preview(self.crop_rect_start_view, current_point_view)

            def finalize_crop_rectangle(self, event):
                """Handles mouse release to finalize the crop rectangle and updates sliders/labels."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                    # If not in crop mode or drag never started, pass event along
                    # Use super() with the correct class hierarchy if LiveViewLabel inherits directly
                    # If LiveViewLabel is just an instance variable, this super() call is incorrect.
                    # Assuming LiveViewLabel is a custom class inheriting QLabel:
                    # super(LiveViewLabel, self.live_view_label).mouseReleaseEvent(event)
                    # If LiveViewLabel is just a QLabel instance, default behaviour is likely sufficient:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event) # Call base QLabel handler if needed
                    return

                if event.button() == Qt.LeftButton:
                    end_point_view = self.live_view_label.transform_point(event.position())
                    start_point_view = self.crop_rect_start_view

                    # Finalize the visual preview in LiveViewLabel
                    self.live_view_label.finalize_crop_preview(start_point_view, end_point_view)

                    try:
                        # --- Coordinate Conversion (same as before) ---
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w = float(self.image.width())
                        img_h = float(self.image.height())
                        label_w = float(self.live_view_label.width())
                        label_h = float(self.live_view_label.height())

                        if img_w <= 0 or img_h <= 0 or label_w <= 0 or label_h <= 0:
                            raise ValueError("Invalid image or label dimensions.")

                        scale_factor = min(label_w / img_w, label_h / img_h)
                        if scale_factor <= 1e-9: raise ValueError("Scale factor too small.")

                        display_img_w = img_w * scale_factor
                        display_img_h = img_h * scale_factor
                        display_offset_x = (label_w - display_img_w) / 2.0
                        display_offset_y = (label_h - display_img_h) / 2.0

                        start_x_img = (start_point_view.x() - display_offset_x) / scale_factor
                        start_y_img = (start_point_view.y() - display_offset_y) / scale_factor
                        end_x_img = (end_point_view.x() - display_offset_x) / scale_factor
                        end_y_img = (end_point_view.y() - display_offset_y) / scale_factor

                        img_x = min(start_x_img, end_x_img)
                        img_y = min(start_y_img, end_y_img)
                        img_crop_w = abs(end_x_img - start_x_img)
                        img_crop_h = abs(end_y_img - start_y_img)
                        # --- End Coordinate Conversion ---

                        if img_crop_w < 1 or img_crop_h < 1:
                             # Handle case where drawn rectangle is too small
                             self.crop_rectangle_coords = None
                             self.live_view_label.clear_crop_preview()
                             # Keep sliders disabled
                             self.crop_x_start_slider.setEnabled(False)
                             self.crop_x_end_slider.setEnabled(False)
                             self.crop_y_start_slider.setEnabled(False)
                             self.crop_y_end_slider.setEnabled(False)
                        else:
                            # Store valid image coordinates
                            self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))

                            # Calculate percentages from valid image coordinates
                            x_start_perc = (img_x / img_w) * 100.0 if img_w > 0 else 0
                            y_start_perc = (img_y / img_h) * 100.0 if img_h > 0 else 0
                            x_end_perc = ((img_x + img_crop_w) / img_w) * 100.0 if img_w > 0 else 100
                            y_end_perc = ((img_y + img_crop_h) / img_h) * 100.0 if img_h > 0 else 100

                            # Convert percentages to integer slider values
                            x_start_val = int(round(x_start_perc * self.crop_slider_precision_factor))
                            y_start_val = int(round(y_start_perc * self.crop_slider_precision_factor))
                            x_end_val = int(round(x_end_perc * self.crop_slider_precision_factor))
                            y_end_val = int(round(y_end_perc * self.crop_slider_precision_factor))

                            # --- Update Sliders (Block signals) ---
                            self.crop_x_start_slider.blockSignals(True); self.crop_x_start_slider.setValue(x_start_val); self.crop_x_start_slider.blockSignals(False)
                            self.crop_x_end_slider.blockSignals(True); self.crop_x_end_slider.setValue(x_end_val); self.crop_x_end_slider.blockSignals(False)
                            self.crop_y_start_slider.blockSignals(True); self.crop_y_start_slider.setValue(y_start_val); self.crop_y_start_slider.blockSignals(False)
                            self.crop_y_end_slider.blockSignals(True); self.crop_y_end_slider.setValue(y_end_val); self.crop_y_end_slider.blockSignals(False)

                            # --- Explicitly Update Labels AFTER setting sliders ---
                            self.crop_x_start_value_label.setText(f"{x_start_perc:.2f}%")
                            self.crop_x_end_value_label.setText(f"{x_end_perc:.2f}%")
                            self.crop_y_start_value_label.setText(f"{y_start_perc:.2f}%")
                            self.crop_y_end_value_label.setText(f"{y_end_perc:.2f}%")
                            # --- End Explicit Label Update ---

                            # --- Enable the sliders ---
                            self.crop_x_start_slider.setEnabled(True)
                            self.crop_x_end_slider.setEnabled(True)
                            self.crop_y_start_slider.setEnabled(True)
                            self.crop_y_end_slider.setEnabled(True)

                    except Exception as e:
                         # Handle errors during coordinate calculation or UI update
                         QMessageBox.warning(self, "Coordinate Error", f"Could not calculate/update crop: {e}")
                         traceback.print_exc()
                         self.crop_rectangle_coords = None
                         self.live_view_label.clear_crop_preview()
                         # Ensure sliders remain disabled on error
                         self.crop_x_start_slider.setEnabled(False)
                         self.crop_x_end_slider.setEnabled(False)
                         self.crop_y_start_slider.setEnabled(False)
                         self.crop_y_end_slider.setEnabled(False)

                    # Reset the temporary start point used for drawing
                    self.crop_rect_start_view = None
            
            def create_white_space_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(15)

                # --- Padding Group ---
                padding_params_group = QGroupBox("Add White Space (Padding)")
                padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                padding_layout = QGridLayout(padding_params_group)
                padding_layout.setSpacing(8)


                # Input fields with validation (optional but good)
                int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

                # Left Padding
                left_padding_label = QLabel("Left Padding (px):")
                self.left_padding_input = QLineEdit("100") # Default
                self.left_padding_input.setValidator(int_validator)
                self.left_padding_input.setToolTip("Pixels to add to the left.")
                padding_layout.addWidget(left_padding_label, 0, 0)
                padding_layout.addWidget(self.left_padding_input, 0, 1)

                # Right Padding
                right_padding_label = QLabel("Right Padding (px):")
                self.right_padding_input = QLineEdit("100") # Default
                self.right_padding_input.setValidator(int_validator)
                self.right_padding_input.setToolTip("Pixels to add to the right.")
                padding_layout.addWidget(right_padding_label, 0, 2)
                padding_layout.addWidget(self.right_padding_input, 0, 3)

                # Top Padding
                top_padding_label = QLabel("Top Padding (px):")
                self.top_padding_input = QLineEdit("100") # Default
                self.top_padding_input.setValidator(int_validator)
                self.top_padding_input.setToolTip("Pixels to add to the top.")
                padding_layout.addWidget(top_padding_label, 1, 0)
                padding_layout.addWidget(self.top_padding_input, 1, 1)

                # Bottom Padding
                bottom_padding_label = QLabel("Bottom Padding (px):")
                self.bottom_padding_input = QLineEdit("0") # Default
                self.bottom_padding_input.setValidator(int_validator)
                self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
                padding_layout.addWidget(bottom_padding_label, 1, 2)
                padding_layout.addWidget(self.bottom_padding_input, 1, 3)

                layout.addWidget(padding_params_group)

                # --- Buttons Layout ---
                button_layout = QHBoxLayout()
                self.recommend_button = QPushButton("Set Recommended Values")
                self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
                self.recommend_button.clicked.connect(self.recommended_values)

                self.clear_padding_button = QPushButton("Clear Values")
                self.clear_padding_button.setToolTip("Set all padding values to 0.")
                self.clear_padding_button.clicked.connect(self.clear_padding_values)

                self.finalize_button = QPushButton("Apply Padding")
                self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
                self.finalize_button.clicked.connect(self.finalize_image)

                button_layout.addWidget(self.recommend_button)
                button_layout.addWidget(self.clear_padding_button)
                button_layout.addStretch()
                button_layout.addWidget(self.finalize_button)

                layout.addLayout(button_layout)
                layout.addStretch() # Push content up

                return tab
            def clear_padding_values(self):
                self.bottom_padding_input.setText("0")
                self.top_padding_input.setText("0")
                self.right_padding_input.setText("0")
                self.left_padding_input.setText("0")
                
            def recommended_values(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to set recommended padding.")
                    return

                # This function should ONLY set the text in the QLineEdit fields for padding.
                # It should NOT directly change slider ranges or _marker_shift_added values.
                # The marker offsets are independent of padding until "Apply Padding" is clicked.

                try:
                    native_width = self.image.width()
                    native_height = self.image.height()

                    if native_width <= 0 or native_height <= 0:
                        QMessageBox.warning(self, "Error", "Current image has invalid dimensions.")
                        return

                    # Set recommended padding values in the QLineEdit fields
                    self.left_padding_input.setText(str(int(native_width * 0.1)))
                    self.right_padding_input.setText(str(int(native_width * 0.1)))
                    self.top_padding_input.setText(str(int(native_height * 0.15)))
                    self.bottom_padding_input.setText(str(int(0))) # Or another sensible default

                    # DO NOT update slider ranges or _marker_shift_added here.
                    # Slider ranges are updated by _update_marker_slider_ranges when self.image changes.
                    # _marker_shift_added are updated by sliders or by crop/pad operations.
                    
                    # No need to call self.update_live_view() here either, as only
                    # text fields for a future operation were changed.
                    # QMessageBox.information(self, "Recommended Values Set", 
                    #                         "Recommended padding values have been entered into the fields.\n"
                    #                         "Click 'Apply Padding' to use them.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not set recommended values: {e}")
                
                
                
            # In class CombinedSDSApp:

            def create_markers_tab(self):
                """Create the Markers tab with a more compact and organized layout."""
                tab = QWidget()
                # Main vertical layout for the tab
                main_layout = QVBoxLayout(tab)
                main_layout.setSpacing(10) # Consistent spacing between major sections

                # --- Top Row: Presets and Top Labels ---
                top_row_layout = QHBoxLayout()
                presets_group = QGroupBox("Left/Right Marker Presets")
                presets_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                presets_layout = QGridLayout(presets_group)
                presets_layout.setSpacing(5)
                presets_layout.addWidget(QLabel("Preset:"), 0, 0)
                self.combo_box = QComboBox(self)
                if hasattr(self, 'presets_data') and self.presets_data:
                    self.combo_box.addItems(sorted(self.presets_data.keys()))
                self.combo_box.addItem("Custom")
                self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
                presets_layout.addWidget(self.combo_box, 0, 1)
                self.marker_values_textbox = QLineEdit(self)
                self.marker_values_textbox.setPlaceholderText("Custom L/R values (comma-sep)")
                self.marker_values_textbox.setEnabled(False)
                presets_layout.addWidget(self.marker_values_textbox, 1, 0, 1, 2)
                self.rename_input = QLineEdit(self)
                self.rename_input.setPlaceholderText("New name for Custom preset")
                self.rename_input.setEnabled(False)
                presets_layout.addWidget(self.rename_input, 2, 0, 1, 2)
                preset_buttons_layout = QHBoxLayout()
                preset_buttons_layout.setContentsMargins(0, 0, 0, 0)
                self.save_button = QPushButton("Save Preset", self)
                self.save_button.setToolTip("Saves the current L/R, Top, Custom Markers/Shapes to the selected/new preset name.")
                self.save_button.clicked.connect(self.save_config)
                self.remove_config_button = QPushButton("Remove Preset", self)
                self.remove_config_button.clicked.connect(self.remove_config)
                preset_buttons_layout.addWidget(self.save_button)
                preset_buttons_layout.addWidget(self.remove_config_button)
                preset_buttons_layout.addStretch()
                presets_layout.addLayout(preset_buttons_layout, 3, 0, 1, 2)
                presets_layout.setColumnStretch(1, 1)
                top_row_layout.addWidget(presets_group, 1)

                top_labels_group = QGroupBox("Top Marker Labels")
                top_labels_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                top_labels_layout = QVBoxLayout(top_labels_group)
                self.top_marker_input = QTextEdit(self)
                current_top_label_text = ", ".join(map(str, getattr(self, 'top_label', [])))
                self.top_marker_input.setText(current_top_label_text)
                self.top_marker_input.setMinimumHeight(40)
                self.top_marker_input.setMaximumHeight(100)
                self.top_marker_input.setPlaceholderText("Top labels (comma-separated)")
                top_labels_layout.addWidget(self.top_marker_input)
                self.update_top_labels_button = QPushButton("Update All L/R/Top Labels")
                self.update_top_labels_button.setToolTip("Apply values from text boxes to current markers on the image.")
                self.update_top_labels_button.clicked.connect(self.update_all_labels)
                top_labels_layout.addWidget(self.update_top_labels_button)
                top_row_layout.addWidget(top_labels_group, 1)
                main_layout.addLayout(top_row_layout)

                # --- Middle Section: Marker Placement and Offsets ---
                placement_group = QGroupBox("Marker Placement and Offsets")
                placement_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                placement_layout = QGridLayout(placement_group)
                placement_layout.setColumnStretch(4, 1)
                if not hasattr(self, 'left_slider_range'): self.left_slider_range = [-100, 1000]
                if not hasattr(self, 'right_slider_range'): self.right_slider_range = [-100, 1000]
                if not hasattr(self, 'top_slider_range'): self.top_slider_range = [-100, 1000]
                if not hasattr(self, 'left_marker_shift_added'): self.left_marker_shift_added = 0
                if not hasattr(self, 'right_marker_shift_added'): self.right_marker_shift_added = 0
                if not hasattr(self, 'top_marker_shift_added'): self.top_marker_shift_added = 0
                left_marker_button = QPushButton("Place Left"); left_marker_button.setToolTip("Ctrl+Shift+L")
                left_marker_button.clicked.connect(self.enable_left_marker_mode)
                remove_left_button = QPushButton("Remove Last")
                remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
                reset_left_button = QPushButton("Reset All"); reset_left_button.setToolTip("Reset All Left Markers")
                reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
                self.left_padding_slider = QSlider(Qt.Horizontal)
                self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
                self.left_padding_slider.setValue(self.left_marker_shift_added)
                self.left_padding_slider.valueChanged.connect(self.update_left_padding)
                duplicate_left_button = QPushButton("Copy →") 
                duplicate_left_button.setToolTip("Copy Right Markers & Offset to Left")
                duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
                placement_layout.addWidget(left_marker_button, 0, 0)
                placement_layout.addWidget(remove_left_button, 0, 1)
                placement_layout.addWidget(reset_left_button, 0, 2)
                placement_layout.addWidget(QLabel("Offset Left:"), 0, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.left_padding_slider, 0, 4)
                placement_layout.addWidget(duplicate_left_button, 0, 5)
                right_marker_button = QPushButton("Place Right"); right_marker_button.setToolTip("Ctrl+Shift+R")
                right_marker_button.clicked.connect(self.enable_right_marker_mode)
                remove_right_button = QPushButton("Remove Last")
                remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
                reset_right_button = QPushButton("Reset All"); reset_right_button.setToolTip("Reset All Right Markers")
                reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
                self.right_padding_slider = QSlider(Qt.Horizontal)
                self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
                self.right_padding_slider.setValue(self.right_marker_shift_added)
                self.right_padding_slider.valueChanged.connect(self.update_right_padding)
                duplicate_right_button = QPushButton("← Copy") 
                duplicate_right_button.setToolTip("Copy Left Markers & Offset to Right")
                duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
                placement_layout.addWidget(right_marker_button, 1, 0)
                placement_layout.addWidget(remove_right_button, 1, 1)
                placement_layout.addWidget(reset_right_button, 1, 2)
                placement_layout.addWidget(QLabel("Offset Right:"), 1, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.right_padding_slider, 1, 4)
                placement_layout.addWidget(duplicate_right_button, 1, 5)
                top_marker_button = QPushButton("Place Top"); top_marker_button.setToolTip("Ctrl+Shift+T")
                top_marker_button.clicked.connect(self.enable_top_marker_mode)
                remove_top_button = QPushButton("Remove Last")
                remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
                reset_top_button = QPushButton("Reset All"); reset_top_button.setToolTip("Reset All Top Markers")
                reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
                self.top_padding_slider = QSlider(Qt.Horizontal)
                self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
                self.top_padding_slider.setValue(self.top_marker_shift_added)
                self.top_padding_slider.valueChanged.connect(self.update_top_padding)
                placement_layout.addWidget(top_marker_button, 2, 0)
                placement_layout.addWidget(remove_top_button, 2, 1)
                placement_layout.addWidget(reset_top_button, 2, 2)
                placement_layout.addWidget(QLabel("Offset Top:"), 2, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.top_padding_slider, 2, 4)
                main_layout.addWidget(placement_group)


                # --- Bottom Section: Custom Markers / Shapes / Grid (COMPACT 2-ROW LAYOUT) ---
                custom_group = QGroupBox("Custom Markers, Shapes, and Grid")
                custom_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                custom_layout = QVBoxLayout(custom_group)
                custom_layout.setSpacing(6)

                # --- Row 1: Placement and Styling ---
                row1_layout = QHBoxLayout()
                row1_layout.setSpacing(6)

                self.custom_marker_button = QPushButton("Place Custom", self)
                self.custom_marker_button.setToolTip("Click to activate, then click on image to place text/arrow")
                self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)

                self.custom_marker_text_entry = QLineEdit(self)
                self.custom_marker_text_entry.setPlaceholderText("Custom text...")

                arrow_buttons_layout = QHBoxLayout()
                arrow_buttons_layout.setContentsMargins(0, 0, 0, 0); arrow_buttons_layout.setSpacing(2)
                arrow_size = 25
                self.custom_marker_button_left_arrow = QPushButton("←"); self.custom_marker_button_left_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_left_arrow.setToolTip("Ctrl+Left")
                self.custom_marker_button_right_arrow = QPushButton("→"); self.custom_marker_button_right_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_right_arrow.setToolTip("Ctrl+Right")
                self.custom_marker_button_top_arrow = QPushButton("↑"); self.custom_marker_button_top_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_top_arrow.setToolTip("Ctrl+Up")
                self.custom_marker_button_bottom_arrow = QPushButton("↓"); self.custom_marker_button_bottom_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_bottom_arrow.setToolTip("Ctrl+Down")
                arrow_buttons_layout.addWidget(self.custom_marker_button_left_arrow); arrow_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
                arrow_buttons_layout.addWidget(self.custom_marker_button_top_arrow); arrow_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
                self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))
                
                self.custom_font_type_dropdown = QFontComboBox()
                self.custom_font_type_dropdown.setCurrentFont(QFont("Arial"))
                self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
                
                self.custom_font_size_spinbox = QSpinBox()
                self.custom_font_size_spinbox.setRange(1, 150); self.custom_font_size_spinbox.setValue(12); self.custom_font_size_spinbox.setPrefix("Size: ")

                self.custom_marker_color_button = QPushButton("Color")
                self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
                if not hasattr(self, 'custom_marker_color'): self.custom_marker_color = QColor(0,0,0)
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                
                row1_layout.addWidget(self.custom_marker_button)
                row1_layout.addWidget(self.custom_marker_text_entry)
                row1_layout.addLayout(arrow_buttons_layout)
                row1_layout.addWidget(self.custom_font_type_dropdown, 1) # Give font dropdown stretch factor
                row1_layout.addWidget(self.custom_font_size_spinbox)
                row1_layout.addWidget(self.custom_marker_color_button)
                custom_layout.addLayout(row1_layout)

                # --- Row 2: Management and Tools ---
                row2_layout = QHBoxLayout()
                row2_layout.setSpacing(6)

                self.remove_custom_marker_button = QPushButton("Remove Last"); self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
                self.reset_custom_marker_button = QPushButton("Reset All"); self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
                
                shape_size = 25
                self.draw_line_button = QPushButton("L"); self.draw_line_button.setToolTip("Draw Line"); self.draw_line_button.setFixedSize(shape_size, shape_size); self.draw_line_button.clicked.connect(self.enable_line_drawing_mode)
                self.draw_rect_button = QPushButton("R"); self.draw_rect_button.setToolTip("Draw Rectangle"); self.draw_rect_button.setFixedSize(shape_size, shape_size); self.draw_rect_button.clicked.connect(self.enable_rectangle_drawing_mode)
                self.remove_shape_button = QPushButton("X"); self.remove_shape_button.setToolTip("Remove Last Shape"); self.remove_shape_button.setFixedSize(shape_size, shape_size); self.remove_shape_button.clicked.connect(self.remove_last_custom_shape)
                
                self.show_grid_checkbox_x = QCheckBox("Snap X"); self.show_grid_checkbox_x.setToolTip("Snap horizontally. Ctrl+Shift+X or CMD+Shift+X toggles X and Ctrl+Shift+G or CMD+Shift+G for both X and Y.")
                self.show_grid_checkbox_x.stateChanged.connect(self.update_live_view)
                self.show_grid_checkbox_y = QCheckBox("Snap Y"); self.show_grid_checkbox_y.setToolTip("Snap vertically. Ctrl+Shift+Y or CMD+Shift+Y  toggles Y and Ctrl+Shift+G or CMD+Shift+G for both X and Y.")
                self.show_grid_checkbox_y.stateChanged.connect(self.update_live_view)
                self.grid_size_input = QSpinBox(); self.grid_size_input.setRange(5, 100); self.grid_size_input.setValue(20); self.grid_size_input.setPrefix("Grid (px): ")
                self.grid_size_input.valueChanged.connect(self.update_live_view)
                self.grid_size_input.setToolTip("Can increase or decrease grid pixel size by CTRL+Shift+Up or CTRL+Shift+Down")

                self.move_resize_button = QPushButton("Move/Resize"); self.move_resize_button.setToolTip("Toggle mode to move/resize custom markers and shapes on the image."); self.move_resize_button.setCheckable(True); self.move_resize_button.clicked.connect(self.toggle_custom_item_interaction_mode)
                self.modify_custom_marker_button = QPushButton("Modify All"); self.modify_custom_marker_button.setToolTip("Modify/Delete Custom Markers & Shapes"); self.modify_custom_marker_button.clicked.connect(self.open_modify_markers_dialog)

                row2_layout.addWidget(self.remove_custom_marker_button)
                row2_layout.addWidget(self.reset_custom_marker_button)
                row2_layout.addSpacing(10)
                row2_layout.addWidget(QLabel("Shapes:"))
                row2_layout.addWidget(self.draw_line_button); row2_layout.addWidget(self.draw_rect_button); row2_layout.addWidget(self.remove_shape_button)
                row2_layout.addSpacing(10)
                row2_layout.addWidget(self.show_grid_checkbox_x)
                row2_layout.addWidget(self.show_grid_checkbox_y)
                row2_layout.addWidget(self.grid_size_input)
                row2_layout.addStretch(1) # Main stretch to push final buttons to the right
                row2_layout.addWidget(self.move_resize_button)
                row2_layout.addWidget(self.modify_custom_marker_button)
                
                custom_layout.addLayout(row2_layout)

                main_layout.addWidget(custom_group)
                main_layout.addStretch()
                return tab

            def open_modify_markers_dialog(self):
                self._backup_custom_markers_before_modify_dialog = [list(m) for m in self.custom_markers]
                self._backup_custom_shapes_before_modify_dialog = [dict(s) for s in self.custom_shapes] # Backup shapes
                
                if not hasattr(self, "custom_markers") or not isinstance(self.custom_markers, list):
                    self.custom_markers = []
                if not hasattr(self, "custom_shapes") or not isinstance(self.custom_shapes, list):
                    self.custom_shapes = []
            
                if not self.custom_markers and not self.custom_shapes:
                    QMessageBox.information(self, "No Items", "There are no custom markers or shapes to modify.")
                    return
            
                dialog = ModifyMarkersDialog(
                    [list(m) for m in self.custom_markers],
                    [dict(s) for s in self.custom_shapes],
                    self
                )
            
                dialog.global_markers_adjusted.connect(self.handle_live_marker_adjustment_preview)
                dialog.shapes_adjusted_preview.connect(self.handle_live_shape_adjustment_preview)
            
                if dialog.exec() == QDialog.Accepted:
                    modified_markers_tuples, modified_shapes_dicts = dialog.get_modified_markers_and_shapes()
                    modified_markers_lists = [list(m) for m in modified_markers_tuples]
            
                    markers_changed = (modified_markers_lists != self._backup_custom_markers_before_modify_dialog)
                    shapes_changed = (modified_shapes_dicts != self._backup_custom_shapes_before_modify_dialog) # FIX: Compare to shape backup
            
                    if markers_changed or shapes_changed:
                        self.save_state()
                        self.custom_markers = modified_markers_lists
                        self.custom_shapes = modified_shapes_dicts
                        self.is_modified = True
                        self.update_live_view()
                    else:
                        # If OK was clicked but no changes were made (e.g., user undid their changes in dialog)
                        self.custom_markers = self._backup_custom_markers_before_modify_dialog
                        # FIX: Also revert shapes if no changes were accepted.
                        self.custom_shapes = self._backup_custom_shapes_before_modify_dialog
                        self.update_live_view()
            
                else: # Dialog was cancelled or closed
                    # Revert both markers and shapes to their state before the dialog was opened.
                    self.custom_markers = self._backup_custom_markers_before_modify_dialog
                    # FIX: Add the missing line to revert shapes on cancel.
                    self.custom_shapes = self._backup_custom_shapes_before_modify_dialog
                    self.update_live_view() # Refresh to show the original state
            
                # Clean up the backup attributes
                if hasattr(self, '_backup_custom_markers_before_modify_dialog'):
                    del self._backup_custom_markers_before_modify_dialog
                if hasattr(self, '_backup_custom_shapes_before_modify_dialog'):
                    del self._backup_custom_shapes_before_modify_dialog
                
                # Disconnect the signals
                try:
                    dialog.global_markers_adjusted.disconnect(self.handle_live_marker_adjustment_preview)
                    dialog.shapes_adjusted_preview.disconnect(self.handle_live_shape_adjustment_preview)
                except (TypeError, RuntimeError): 
                    pass
            
            
            def handle_live_marker_adjustment_preview(self, temporarily_adjusted_markers):
                """
                Temporarily updates self.custom_markers and refreshes the live_view_label
                to show the effect of global adjustments from ModifyMarkersDialog.
                This does NOT mark the image as permanently modified or save to undo stack.
                """
                # We are directly modifying self.custom_markers here for the preview.
                # The backup in open_modify_markers_dialog handles reverting on cancel.
                self.custom_markers = temporarily_adjusted_markers # Update with the live adjusted list
                self.update_live_view()
                
            def handle_live_shape_adjustment_preview(self, temporarily_adjusted_shapes):
                """
                Temporarily updates self.custom_shapes and refreshes the live_view_label
                to show the effect of shape edits from ModifyMarkersDialog.
                """
                self.custom_shapes = temporarily_adjusted_shapes # Update with the live adjusted list
                self.update_live_view()
                         
            def _serialize_custom_markers(self, custom_markers_list):
                """Converts a list of custom marker tuples (with QColor) to a list of serializable dicts."""
                serialized_list = []
                for marker_tuple in custom_markers_list:
                    try:
                        # Ensure the tuple has the expected 8 elements
                        if len(marker_tuple) == 8:
                            x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_tuple
                            serialized_list.append({
                                "x": x, "y": y, "text": text, "color": qcolor.name(), # Store color name
                                "font_family": font_family, "font_size": font_size,
                                "bold": is_bold, "italic": is_italic
                            })
                        else:
                            # Handle older format or malformed data gracefully if necessary
                            # For now, we'll skip markers not matching the 8-element structure
                            print(f"Warning: Skipping marker during serialization due to unexpected format: {marker_tuple}")
                    except (ValueError, TypeError, IndexError, AttributeError) as e:
                        print(f"Warning: Skipping marker during serialization: {marker_tuple}, Error: {e}")
                return serialized_list

            def _deserialize_custom_markers(self, custom_markers_config_list):
                """Converts a list of serializable dicts back to custom marker tuples (with QColor)."""
                deserialized_list = []
                if not isinstance(custom_markers_config_list, list):
                    print("Warning: custom_markers_config_list is not a list, cannot deserialize.")
                    return []
                    
                for marker_conf in custom_markers_config_list:
                    if not isinstance(marker_conf, dict):
                        print(f"Warning: Skipping non-dictionary item in custom_markers_config_list: {marker_conf}")
                        continue
                    try:
                        color_str = marker_conf.get("color", "#000000") # Default to black if missing
                        qcolor = QColor(color_str)
                        if not qcolor.isValid():
                            print(f"Warning: Invalid color string '{color_str}' for marker, using black.")
                            qcolor = QColor(0,0,0) # Fallback to black

                        deserialized_list.append((
                            float(marker_conf.get("x", 0.0)),
                            float(marker_conf.get("y", 0.0)),
                            str(marker_conf.get("text", "")),
                            qcolor,
                            str(marker_conf.get("font_family", "Arial")),
                            int(marker_conf.get("font_size", 12)),
                            bool(marker_conf.get("bold", False)),
                            bool(marker_conf.get("italic", False))
                        ))
                    except (ValueError, TypeError, KeyError) as e:
                        print(f"Warning: Skipping marker during deserialization: {marker_conf}, Error: {e}")
                return deserialized_list
            
            def add_column(self):
                """Add a new column to the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                if current_text.strip():
                    self.top_marker_input.append("")  # Add a new line for a new column
                else:
                    self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
            
            def remove_column(self):
                """Remove the last column from the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                lines = current_text.split("\n")
                if len(lines) > 1:
                    lines.pop()  # Remove the last line
                    self.top_marker_input.setPlainText("\n".join(lines))
                else:
                    self.top_marker_input.clear()  # Clear the text if only one line exists

            
            def flip_vertical(self):
                self.save_state()
                """Flip the image vertically."""
                if self.image and not self.image.isNull():
                    transform = QTransform()
                    # Scale by -1 in Y, then translate back because scaling is around (0,0)
                    transform.scale(1, -1)
                    transform.translate(0, -self.image.height())
                    self.image = self.image.transformed(transform) # Get a new transformed image

                    # Update backups after transformation
                    if not self.image.isNull(): # Check if transform was successful
                        self.image_before_contrast = self.image.copy()
                        self.image_before_padding = self.image.copy() # Or None if padding invalidates this
                        self.image_contrasted = self.image.copy()
                        self.update_live_view()
                    else:
                        print("Warning: Vertical flip resulted in a null image.")
            
            def flip_horizontal(self):
                self.save_state()
                """Flip the image horizontally."""
                if self.image and not self.image.isNull():
                    transform = QTransform()
                    # Scale by -1 in X, then translate back
                    transform.scale(-1, 1)
                    transform.translate(-self.image.width(), 0)
                    self.image = self.image.transformed(transform) # Get a new transformed image

                    # Update backups after transformation
                    if not self.image.isNull(): # Check if transform was successful
                        self.image_before_contrast = self.image.copy()
                        self.image_before_padding = self.image.copy() # Or None
                        self.image_contrasted = self.image.copy()
                        self.update_live_view()
                    else:
                        print("Warning: Horizontal flip resulted in a null image.")
            
            def convert_to_black_and_white(self):
                """
                Converts the current self.image to grayscale, preserving alpha channel
                if present by creating an ARGB image where R=G=B=GrayValue.
                Aims for 16-bit grayscale precision when converting from color.
                """
                self.save_state()
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded.")
                    return

                original_format = self.image.format()
                original_has_alpha = self.image.hasAlphaChannel() # Check QImage's report

                # Check if already effectively grayscale (but maybe not the specific format)
                if original_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16, QImage.Format_Mono]:
                    # It's already a standard grayscale format (without explicit alpha in format enum)
                    QMessageBox.information(self, "Info", f"Image is already grayscale (Format: {original_format}).")
                    return

                converted_image = None
                try:
                    np_img = self.qimage_to_numpy(self.image)
                    if np_img is None: raise ValueError("NumPy conversion failed.")

                    print(f"Original NumPy shape: {np_img.shape}, dtype: {np_img.dtype}") # Debug

                    target_dtype = np.uint16 # Prefer 16-bit precision for grayscale conversion
                    target_max_val = 65535.0

                    if np_img.ndim == 3 and np_img.shape[2] == 4: # Color with Alpha (BGRA or RGBA)
                        print("Processing image with Alpha channel.")
                        # Assume input order from qimage_to_numpy is BGRA for ARGB32/RGB32
                        # or RGBA for RGBA8888. cvtColor expects BGR.
                        # Let's explicitly convert the color part assuming BGR input slice
                        color_part = np_img[..., :3] # Slice BGR or RGB
                        alpha_part = np_img[..., 3]  # Extract alpha channel

                        # Convert color to grayscale using cv2 (expects BGR)
                        # If the input was RGBA, this might be slightly off, but often acceptable.
                        # A more robust way might check the original_format if needed.
                        gray_np_8bit = cv2.cvtColor(color_part, cv2.COLOR_BGR2GRAY)

                        # Scale grayscale to target bit depth (16-bit)
                        gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)

                        # Create a new 4-channel array (BGRA format for QImage.Format_ARGB32)
                        # Replicate grayscale channel for B, G, R
                        bgra_target = np.zeros((np_img.shape[0], np_img.shape[1], 4), dtype=target_dtype)
                        bgra_target[..., 0] = gray_np_target # Blue = Gray
                        bgra_target[..., 1] = gray_np_target # Green = Gray
                        bgra_target[..., 2] = gray_np_target # Red = Gray

                        # Handle Alpha channel: ensure it's the correct dtype for combining
                        if alpha_part.dtype != target_dtype:
                             # Scale alpha if necessary (e.g., if source was 16-bit RGBA)
                             if alpha_part.dtype == np.uint16:
                                 alpha_scaled = (alpha_part / 65535.0 * target_max_val).astype(target_dtype)
                             elif alpha_part.dtype == np.uint8:
                                 alpha_scaled = (alpha_part / 255.0 * target_max_val).astype(target_dtype)
                             else: # Fallback: assume it doesn't need scaling or use 8-bit alpha
                                  print(f"Warning: Unexpected alpha dtype {alpha_part.dtype}. Attempting direct use or scaling.")
                                  try: # Try scaling assuming it's 0-max range
                                      alpha_scaled = (alpha_part / np.max(alpha_part) * target_max_val).astype(target_dtype) if np.max(alpha_part) > 0 else np.zeros_like(alpha_part, dtype=target_dtype)
                                  except: # Final fallback
                                       alpha_scaled = (alpha_part.astype(np.float64) / 255.0 * target_max_val).astype(target_dtype) if np.max(alpha_part)>0 else np.zeros_like(alpha_part,dtype=target_dtype) #Assume 8 bit if failsafe
                             bgra_target[..., 3] = alpha_scaled
                        else:
                            bgra_target[..., 3] = alpha_part # Alpha dtype already matches

                        # Convert back to QImage - should become Format_ARGB32 due to 4 channels
                        # Note: numpy_to_qimage needs to handle 16-bit 4-channel input if we want ARGB64
                        # Currently, it scales down 16-bit color to 8-bit ARGB32. Let's adapt that.
                        # --- MODIFICATION NEEDED in numpy_to_qimage for 16-bit ARGB ---
                        # For now, let's convert bgra_target to uint8 for Format_ARGB32 output
                        bgra_target_uint8 = (bgra_target / target_max_val * 255.0).astype(np.uint8)
                        converted_image = self.numpy_to_qimage(bgra_target_uint8)
                        if converted_image.isNull(): raise ValueError("Conversion of BGRA Grayscale to QImage failed.")
                        print(f"Converted to Grayscale + Alpha. QImage format: {converted_image.format()}") # Debug


                    elif np_img.ndim == 3 and np_img.shape[2] == 3: # Color without Alpha (BGR or RGB)
                        print("Processing image without Alpha channel.")
                        # Convert color to grayscale (standard procedure)
                        gray_np_8bit = cv2.cvtColor(np_img[..., :3], cv2.COLOR_BGR2GRAY)
                         # Scale grayscale to target bit depth (16-bit)
                        gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)
                        # Convert back to standard grayscale QImage
                        converted_image = self.numpy_to_qimage(gray_np_target) # Should yield Format_Grayscale16
                        if converted_image.isNull(): raise ValueError("Conversion of BGR Grayscale to QImage failed.")
                        print(f"Converted to Grayscale. QImage format: {converted_image.format()}") # Debug


                    elif np_img.ndim == 2: # Already grayscale (but QImage didn't report standard format)
                         # This case might occur for unusual grayscale formats QImage loads but doesn't map to standard enums.
                         QMessageBox.information(self, "Info", "Image is already grayscale (based on channel analysis).")
                         # Optionally convert to a standard format like Grayscale16
                         if np_img.dtype != target_dtype:
                             gray_np_target = (np_img.astype(np.float64) / np.max(np_img) * target_max_val).astype(target_dtype) if np.max(np_img)>0 else np.zeros_like(np_img,dtype=target_dtype)
                             converted_image = self.numpy_to_qimage(gray_np_target)
                         else:
                             converted_image = self.image.copy() # No conversion needed
                    else:
                        raise ValueError(f"Unsupported NumPy array dimension: {np_img.ndim}")


                except Exception as e:
                     QMessageBox.critical(self, "Conversion Error", f"Could not convert image to grayscale: {e}")
                     traceback.print_exc()
                     return # Keep original image

                if converted_image and not converted_image.isNull():
                     self.image = converted_image
                     # Update backups consistently
                     self.image_before_contrast = self.image.copy()
                     self.image_contrasted = self.image.copy()
                     # Reset padding state if format changes implicitly? Let's assume padding needs re-application.
                     if self.image_padded:
                         self.image_before_padding = None # Invalidate padding backup
                         self.image_padded = False
                     else:
                         self.image_before_padding = self.image.copy() if self.image else None # Ensure image exists before copying

                     # Reset contrast/gamma sliders as appearance changed significantly
                     self.reset_gamma_contrast() # Resets sliders and updates view
                     self._update_status_bar()
                     self.update_live_view() # Ensure view updates even if reset_gamma_contrast fails
                else:
                     # This case should be less likely now with better error handling
                     QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


            def invert_image(self):
                self.save_state()
                if self.image:
                    inverted_image = self.image.copy()
                    inverted_image.invertPixels()
                    self.image = inverted_image
                    self.update_live_view()
                self.image_before_contrast=self.image.copy()
                self.image_before_padding=self.image.copy()
                self.image_contrasted=self.image.copy()

            def keyPressEvent(self, event):
                # --- Escape Key Handling ---
                if event.key() == Qt.Key_Escape:
                    a_mode_was_cancelled_or_view_reset = False # Flag if any action taken by Esc

                    if self.current_selection_mode in ["select_custom_item", "dragging_custom_item", "resizing_custom_item"]:
                        self.cancel_custom_item_interaction_mode()
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"]:
                        # If actively dragging/resizing, revert to "select_for_move"
                        # or fully cancel the selection mode.
                        if self.current_selection_mode in ["dragging_shape", "resizing_corner"]:
                            # Attempt to revert the in-progress drag by restoring original points.
                            # This part relies on shape_points_at_drag_start_label being correctly set.
                            if self.shape_points_at_drag_start_label:
                                original_points_restored = False
                                if self.moving_multi_lane_index >= 0 and self.multi_lane_definitions and \
                                   self.moving_multi_lane_index < len(self.multi_lane_definitions):
                                    lane_def = self.multi_lane_definitions[self.moving_multi_lane_index]
                                    if lane_def['type'] == 'quad':
                                        lane_def['points_label'] = [QPointF(p) for p in self.shape_points_at_drag_start_label]
                                        original_points_restored = True
                                    elif lane_def['type'] == 'rectangle':
                                        orig_rect_points = self.shape_points_at_drag_start_label
                                        if len(orig_rect_points) == 4:
                                            all_x = [p.x() for p in orig_rect_points]; all_y = [p.y() for p in orig_rect_points]
                                            lane_def['points_label'] = [QRectF(QPointF(min(all_x), min(all_y)), QPointF(max(all_x), max(all_y)))]
                                            original_points_restored = True
                                elif self.moving_multi_lane_index == -2: # Single Quad
                                    self.live_view_label.quad_points = [QPointF(p) for p in self.shape_points_at_drag_start_label]
                                    original_points_restored = True
                                elif self.moving_multi_lane_index == -3: # Single Rect
                                    orig_rect_points = self.shape_points_at_drag_start_label
                                    if len(orig_rect_points) == 4:
                                        all_x = [p.x() for p in orig_rect_points]; all_y = [p.y() for p in orig_rect_points]
                                        self.live_view_label.bounding_box_preview = (min(all_x), min(all_y), max(all_x), max(all_y))
                                        self.live_view_label.quad_points = []
                                        original_points_restored = True
                                if original_points_restored:
                                    pass
                            # Transition back to selection mode or cancel fully
                            self.current_selection_mode = "select_for_move"
                            self.live_view_label.mode = "select_for_move"
                            # Keep custom_left_click_handler_from_app for re-selection
                            self.live_view_label._custom_mouseMoveEvent_from_app = None
                            self.live_view_label._custom_mouseReleaseEvent_from_app = None
                        else: # Was in "select_for_move", so cancel the whole mode
                            self.cancel_selection_or_move_mode()
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.multi_lane_mode_active:
                        self.cancel_multi_lane_mode() # This calls _reset_live_view_label_custom_handlers
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.mode == 'auto_lane_quad':
                        self.live_view_label.mode = None
                        self.live_view_label.quad_points = []
                        self.live_view_label.selected_point = -1
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.crop_rectangle_mode or self.live_view_label.mode == 'auto_lane_rect':
                        self.cancel_rectangle_crop_mode() # This calls _reset_live_view_label_custom_handlers
                        if self.live_view_label.mode == 'auto_lane_rect': self.live_view_label.mode = None
                        self.live_view_label.clear_crop_preview()
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.drawing_mode in ['line', 'rectangle']:
                        self.cancel_drawing_mode() # This calls _reset_live_view_label_custom_handlers
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.preview_marker_enabled: # Custom marker preview
                        self.live_view_label.preview_marker_enabled = False
                        self.live_view_label.preview_marker_position = None # Crucial
                        # The marker_mode might still be "custom", _reset_live_view_label_custom_handlers will clear its specific handlers
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.marker_mode is not None: # Standard L/R/Top marker placement
                        self.marker_mode = None
                        # _reset_live_view_label_custom_handlers will clear its specific handlers
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.measure_quantity_mode or \
                         self.live_view_label.mode in ["quad", "rectangle", "move"]: # Analysis area definition
                        self.live_view_label.measure_quantity_mode = False
                        self.live_view_label.bounding_box_complete = False
                        self.live_view_label.counter = 0
                        self.live_view_label.quad_points = []
                        self.live_view_label.bounding_box_preview = None
                        self.live_view_label.rectangle_points = []
                        self.live_view_label.rectangle_start = None
                        self.live_view_label.rectangle_end = None
                        self.live_view_label.selected_point = -1
                        self.live_view_label.mode = None
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.mw_predict_preview_enabled:
                        self.live_view_label.mw_predict_preview_enabled = False
                        self.live_view_label.mw_predict_preview_position = None # Crucial
                        self.live_view_label.setMouseTracking(False) # Turn off if only for this mode
                        a_mode_was_cancelled_or_view_reset = True

                    # 2. ALWAYS reset LiveViewLabel's custom handlers and cursor state
                    # This is the most important step to ensure a clean slate for LiveViewLabel.
                    self._reset_live_view_label_custom_handlers()

                    # 3. Reset zoom if it's not 1.0, OR if any specific mode was cancelled by Esc
                    if self.live_view_label.zoom_level != 1.0 or a_mode_was_cancelled_or_view_reset:
                        self.live_view_label.zoom_level = 1.0
                        self.live_view_label.pan_offset = QPointF(0, 0)
                        # _reset_live_view_label_custom_handlers would have set cursor to Arrow
                        # if zoom became 1.0 and no panning.
                        a_mode_was_cancelled_or_view_reset = True


                    # 4. Update the view if any state change occurred due to Escape
                    if a_mode_was_cancelled_or_view_reset:
                        self.update_live_view()

                    event.accept()  # Consume the Escape key event
                    return # IMPORTANT: Ensure Escape key processing finishes here

                # --- Panning with Arrow Keys (only if zoom_level != 1.0 and Esc was not processed above) ---
                if self.live_view_label.zoom_level != 1.0:
                    step = 20 # Pan step
                    offset_changed = False
                    current_x = self.live_view_label.pan_offset.x()
                    current_y = self.live_view_label.pan_offset.y()

                    if event.key() == Qt.Key_Left:
                        self.live_view_label.pan_offset.setX(current_x - step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Right:
                        self.live_view_label.pan_offset.setX(current_x + step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Up:
                        self.live_view_label.pan_offset.setY(current_y - step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Down:
                        self.live_view_label.pan_offset.setY(current_y + step)
                        offset_changed = True

                    if offset_changed:
                        self.update_live_view()
                        event.accept() # Consume arrow key if panning occurred
                        return

                # If not Escape and not a panning arrow key (while zoomed), pass to super
                super().keyPressEvent(event)
                
            def toggle_custom_item_interaction_mode(self, checked):
                if checked:
                    # Cancel any other conflicting modes
                    self.cancel_drawing_mode()
                    self.cancel_rectangle_crop_mode()
                    self.cancel_selection_or_move_mode() # Cancel analysis area selection
                    if hasattr(self, 'marker_mode'): self.marker_mode = None

                    self.current_selection_mode = "select_custom_item"
                    self.live_view_label.mode = "select_custom_item"
                    self.moving_custom_item_info = None
                    self.resizing_corner_index = -1

                    self._reset_live_view_label_custom_handlers()
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_custom_item_selection_click

                    QMessageBox.information(self, "Move/Resize Custom Items",
                                            "Click on a marker or shape to select it.\n"
                                            "Drag its body to move, or a corner handle to resize (shapes only).\n"
                                            "Press ESC or un-check the button to exit this mode.")
                    self.update_live_view()
                else:
                    self.cancel_custom_item_interaction_mode()

            def cancel_custom_item_interaction_mode(self):
                if self.current_selection_mode not in ["select_custom_item", "dragging_custom_item", "resizing_custom_item"]:
                    return

                self.current_selection_mode = None
                self.live_view_label.mode = None
                self.moving_custom_item_info = None
                self.resizing_corner_index = -1
                self.shape_points_at_drag_start_label = []

                self._reset_live_view_label_custom_handlers()

                if hasattr(self, 'move_resize_button') and self.move_resize_button.isChecked():
                    self.move_resize_button.setChecked(False)

                self.update_live_view()

            def handle_custom_item_selection_click(self, event):
                if self.current_selection_mode != "select_custom_item" or event.button() != Qt.LeftButton:
                    return

                clicked_point_ls = self.live_view_label.transform_point(event.position())
                click_radius_threshold = self.live_view_label.CORNER_HANDLE_BASE_RADIUS * 1.5
                item_selected = False

                # Iterate in reverse to select topmost items first
                # Check shape handles first (higher priority than body)
                for i in range(len(self.custom_shapes) - 1, -1, -1):
                    _body, handles_ls = self._get_shape_bounding_box_and_handles_in_label_space(i)
                    if handles_ls:
                        for corner_idx, handle_pt in enumerate(handles_ls):
                            if (clicked_point_ls - handle_pt).manhattanLength() < click_radius_threshold:
                                self.moving_custom_item_info = {'type': 'shape', 'index': i}
                                self.resizing_corner_index = corner_idx
                                self.shape_points_at_drag_start_label = handles_ls
                                item_selected = True
                                break
                    if item_selected: break

                # Check shape bodies if no handle was clicked
                if not item_selected:
                    for i in range(len(self.custom_shapes) - 1, -1, -1):
                        body_ls, handles_ls = self._get_shape_bounding_box_and_handles_in_label_space(i)
                        if body_ls and body_ls.contains(clicked_point_ls):
                            self.moving_custom_item_info = {'type': 'shape', 'index': i}
                            self.resizing_corner_index = -1 # Body move
                            self.shape_points_at_drag_start_label = handles_ls # Store corners for move calculation
                            item_selected = True
                            break

                # Check marker bodies if no shape was clicked
                if not item_selected:
                    for i in range(len(self.custom_markers) - 1, -1, -1):
                        bbox_ls = self._get_marker_bounding_box_in_label_space(i)
                        if bbox_ls and bbox_ls.contains(clicked_point_ls):
                            self.moving_custom_item_info = {'type': 'marker', 'index': i}
                            self.resizing_corner_index = -1 # Markers can't be resized this way
                            self.shape_points_at_drag_start_label = [bbox_ls.center()]
                            item_selected = True
                            break

                if item_selected:
                    self.initial_mouse_pos_for_shape_drag_label = clicked_point_ls
                    if self.resizing_corner_index != -1:
                        self.current_selection_mode = "resizing_custom_item"
                        self.live_view_label.mode = "resizing_custom_item"
                    else:
                        self.current_selection_mode = "dragging_custom_item"
                        self.live_view_label.mode = "dragging_custom_item"

                    self.live_view_label.setCursor(Qt.CrossCursor)
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_custom_item_drag
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_custom_item_drag_release
                else:
                    self.moving_custom_item_info = None
                    self.resizing_corner_index = -1

                self.update_live_view()

            def handle_custom_item_drag(self, event):
                if not self.moving_custom_item_info or not (event.buttons() & Qt.LeftButton):
                    return

                current_mouse_pos_ls = self.live_view_label.transform_point(event.position())
                info = self.moving_custom_item_info

                # Coordinate transformation helper
                label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                img_w = float(self.image.width()); img_h = float(self.image.height())
                if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return
                scale = min(label_w / img_w, label_h / img_h)
                offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0
                def label_to_image(p_ls):
                    if scale == 0: return QPointF(0,0)
                    return QPointF((p_ls.x() - offset_x) / scale, (p_ls.y() - offset_y) / scale)

                # --- DRAG LOGIC ---
                if self.current_selection_mode == "dragging_custom_item":
                    raw_delta_ls = current_mouse_pos_ls - self.initial_mouse_pos_for_shape_drag_label
                    snapped_new_ref_point_ls = self.snap_point_to_grid(self.shape_points_at_drag_start_label[0] + raw_delta_ls)
                    effective_delta_ls = snapped_new_ref_point_ls - self.shape_points_at_drag_start_label[0]

                    if info['type'] == 'marker':
                        new_center_ls = self.shape_points_at_drag_start_label[0] + effective_delta_ls
                        new_center_img = label_to_image(new_center_ls)
                        self.custom_markers[info['index']][0] = new_center_img.x()
                        self.custom_markers[info['index']][1] = new_center_img.y()
                    elif info['type'] == 'shape':
                        new_points_ls = [p + effective_delta_ls for p in self.shape_points_at_drag_start_label]
                        new_points_img = [label_to_image(p) for p in new_points_ls]
                        shape_data = self.custom_shapes[info['index']]
                        if shape_data['type'] == 'rectangle':
                            xs = [p.x() for p in new_points_img]; ys = [p.y() for p in new_points_img]
                            shape_data['rect'] = (min(xs), min(ys), max(xs) - min(xs), max(ys) - min(ys))
                        elif shape_data['type'] == 'line':
                            shape_data['start'] = (new_points_img[0].x(), new_points_img[0].y())
                            shape_data['end'] = (new_points_img[1].x(), new_points_img[1].y())

                # --- RESIZE LOGIC ---
                elif self.current_selection_mode == "resizing_custom_item":
                    snapped_mouse_ls = self.snap_point_to_grid(current_mouse_pos_ls)
                    shape_data = self.custom_shapes[info['index']]
                    if shape_data['type'] == 'rectangle':
                        fixed_corner_ls = self.shape_points_at_drag_start_label[(self.resizing_corner_index + 2) % 4]
                        p1_img = label_to_image(fixed_corner_ls); p2_img = label_to_image(snapped_mouse_ls)
                        xs = sorted([p1_img.x(), p2_img.x()]); ys = sorted([p1_img.y(), p2_img.y()])
                        shape_data['rect'] = (xs[0], ys[0], xs[1] - xs[0], ys[1] - ys[0])
                    elif shape_data['type'] == 'line':
                        other_endpoint_ls = self.shape_points_at_drag_start_label[(self.resizing_corner_index + 1) % 2]
                        p1_img = label_to_image(other_endpoint_ls); p2_img = label_to_image(snapped_mouse_ls)
                        if self.resizing_corner_index == 0:
                            shape_data['start'], shape_data['end'] = (p2_img.x(), p2_img.y()), (p1_img.x(), p1_img.y())
                        else:
                            shape_data['start'], shape_data['end'] = (p1_img.x(), p1_img.y()), (p2_img.x(), p2_img.y())
                
                self.update_live_view()

            def handle_custom_item_drag_release(self, event):
                if self.current_selection_mode in ["dragging_custom_item", "resizing_custom_item"] and event.button() == Qt.LeftButton:
                    self.save_state()
                    self.is_modified = True
                    self.shape_points_at_drag_start_label = []
                    self.initial_mouse_pos_for_shape_drag_label = QPointF()
                    self.resizing_corner_index = -1
                    self.current_selection_mode = "select_custom_item"
                    self.live_view_label.mode = "select_custom_item"
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                    self.update_live_view()
            
            def _get_marker_bounding_box_in_label_space(self, marker_index):
                if not (self.image and not self.image.isNull()): return None
                try:
                    marker_data = self.custom_markers[marker_index]
                    x_img, y_img, text, _color, font_family, font_size, is_bold, is_italic = marker_data
                    label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                    img_w = float(self.image.width()); img_h = float(self.image.height())
                    if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return None
                    scale = min(label_w / img_w, label_h / img_h)
                    offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0
                    anchor_x_ls = x_img * scale + offset_x; anchor_y_ls = y_img * scale + offset_y
                    
                    font = QFont(font_family, int(font_size)); font.setBold(is_bold); font.setItalic(is_italic)
                    fm = QFontMetrics(font)
                    # Add a small padding to the bounding box to make it easier to click
                    padding = 4
                    text_rect = fm.boundingRect(text).adjusted(-padding, -padding, padding, padding)
                    
                    return QRectF(
                        anchor_x_ls - text_rect.width() / 2.0,
                        anchor_y_ls - text_rect.height() / 2.0,
                        text_rect.width(),
                        text_rect.height()
                    )
                except (IndexError, Exception):
                    return None

            def _get_shape_bounding_box_and_handles_in_label_space(self, shape_index):
                if not (self.image and not self.image.isNull()): return None, None
                try:
                    shape_data = self.custom_shapes[shape_index]
                    label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                    img_w = float(self.image.width()); img_h = float(self.image.height())
                    if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return None, None
                    scale = min(label_w / img_w, label_h / img_h)
                    offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0

                    def img_to_label(p): return QPointF(p[0] * scale + offset_x, p[1] * scale + offset_y)
                        
                    shape_type = shape_data.get('type')
                    body = None; handles = []
                    if shape_type == 'rectangle':
                        x, y, w, h = shape_data['rect']
                        p1 = img_to_label((x, y)); p2 = img_to_label((x + w, y + h))
                        body = QRectF(p1, p2).normalized()
                        handles = [body.topLeft(), body.topRight(), body.bottomRight(), body.bottomLeft()]
                    elif shape_type == 'line':
                        p1 = img_to_label(shape_data['start']); p2 = img_to_label(shape_data['end'])
                        # The "body" for clicking a line needs tolerance
                        body = QRectF(p1, p2).normalized().adjusted(-5, -5, 5, 5) 
                        handles = [p1, p2]
                    return body, handles
                except (IndexError, KeyError, Exception):
                    return None, None
            
            def update_marker_text_font(self, font: QFont):
                """
                Updates the font of self.custom_marker_text_entry based on the selected font
                from self.custom_font_type_dropdown.
            
                :param font: QFont object representing the selected font from the combobox.
                """
                # Get the name of the selected font
                selected_font_name = font.family()
                
                # Set the font of the QLineEdit
                self.custom_marker_text_entry.setFont(QFont(selected_font_name))
            
            def arrow_marker(self, text: str):
                self.custom_marker_text_entry.clear()
                self.custom_marker_text_entry.setText(text)
            
            def enable_custom_marker_mode(self):
                """Enable the custom marker mode and set the mouse event."""
                self.live_view_label.setCursor(Qt.ArrowCursor) # Often Arrow for text placement
                custom_text = self.custom_marker_text_entry.text().strip()
                self._reset_live_view_label_custom_handlers()
            
                self.live_view_label.preview_marker_enabled = True
                self.live_view_label.preview_marker_text = custom_text
                self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
                self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
                self.live_view_label.marker_color=self.custom_marker_color
                
                self.live_view_label.setFocus()
                self.live_view_label.update()
                
                self.marker_mode = "custom"  # Indicate custom marker mode
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.place_custom_marker(event, custom_text)
                # self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
                
            def remove_custom_marker_mode(self):
                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    self.custom_markers.pop()  # Remove the last entry from the list           
                self.update_live_view()  # Update the display
                
            def remove_last_custom_shape(self):
                """Removes the last added custom SHAPE (line or rectangle)."""
                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    self.save_state() # Save state before removal for undo
                    removed_shape = self.custom_shapes.pop()  # Remove the last shape
                    print(f"Removed last custom shape: {removed_shape}") # Optional debug
                    self.is_modified = True
                    self.update_live_view()  # Update the display
                    # Optionally, provide user feedback via status bar or QMessageBox
                    # self.statusBar().showMessage("Last custom shape removed.", 2000)
                else:
                    print("No custom shapes to remove.") # Info message
                    # Optionally, provide user feedback
                    # self.statusBar().showMessage("No custom shapes to remove.", 2000)
                
            def reset_custom_marker_mode(self):
                """Resets (clears) all custom MARKERS and custom SHAPES."""
                markers_cleared = False
                shapes_cleared = False
                needs_save = False # Flag to check if state needs saving

                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    if not needs_save: # Save state only once before the first clear operation
                         self.save_state()
                         needs_save = True
                    self.custom_markers.clear()
                    markers_cleared = True
                    print("Cleared all custom markers.")

                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    if not needs_save: # Save state only once if markers weren't cleared but shapes are
                         self.save_state()
                         needs_save = True
                    self.custom_shapes.clear()
                    shapes_cleared = True
                    print("Cleared all custom shapes.")

                if markers_cleared or shapes_cleared:
                    self.is_modified = True
                    self.update_live_view()
                else:
                    print("No custom markers or shapes to reset.")
                
            
            def place_custom_marker(self, event, custom_text):
                """Place a custom marker at the cursor location."""
                self.save_state()
                # Get cursor position from the event
                pos = event.position()
                cursor_x, cursor_y = pos.x(), pos.y()
                
                
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                grid_size = self.grid_size_input.value() # Get grid size once

                if self.show_grid_checkbox_x.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_x = round(cursor_x / grid_size) * grid_size

                if self.show_grid_checkbox_y.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_y = round(cursor_y / grid_size) * grid_size
            
                # Dimensions of the displayed image
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
            
                # Dimensions of the actual image
                image_width = self.image.width()
                image_height = self.image.height()
            
                # Calculate scaling factors
                scale = min(displayed_width / image_width, displayed_height / image_height)
            
                # Calculate offsets
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2
            
                # Transform cursor position to image space
                image_x = (cursor_x - x_offset) / scale
                image_y = (cursor_y - y_offset) / scale
                
                    
                # Store the custom marker's position and text
                self.custom_markers = getattr(self, "custom_markers", [])
                self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value(),False,False))
                self.update_live_view()
                
            def select_custom_marker_color(self):
                """Open a color picker dialog to select the color for custom markers."""
                color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
                if color.isValid():
                    self.custom_marker_color = color  # Update the custom marker color
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
            
            def update_all_labels(self):
                """Update all labels from the QTextEdit, supporting multiple columns."""
                self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
                input_text = self.top_marker_input.toPlainText()
                
                # Split the text into a list by commas and strip whitespace
                self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
                

                # if len(self.top_label) < len(self.top_markers):
                #     self.top_markers = self.top_markers[:len(self.top_label)]
                for i in range(0, len(self.top_markers)):
                    try:
                        self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                    except:
                        self.top_markers[i] = (self.top_markers[i][0], str(""))

                # if len(self.marker_values) < len(self.left_markers):
                #     self.left_markers = self.left_markers[:len(self.marker_values)]
                for i in range(0, len(self.left_markers)):
                    try:
                        self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
                    except:
                        self.left_markers[i] = (self.left_markers[i][0], str(""))
                        

                # if len(self.marker_values) < len(self.right_markers):
                #     self.right_markers = self.right_markers[:len(self.marker_values)]
                for i in range(0, len(self.right_markers)):
                    try:
                        self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
                    except:
                        self.right_markers[i] = (self.right_markers[i][0], str(""))
                        

                
                # Trigger a refresh of the live view
                self.update_live_view()
                
            def reset_marker(self, marker_type, param):    
                self.save_state()
                if marker_type == 'left':
                    if param == 'remove' and len(self.left_markers)!=0:
                        self.left_markers.pop()  
                        if self.current_left_marker_index > 0:
                            self.current_left_marker_index -= 1
                    elif param == 'reset':
                        self.left_markers.clear()
                        self.current_left_marker_index = 0  

                     
                elif marker_type == 'right' and len(self.right_markers)!=0:
                    if param == 'remove':
                        self.right_markers.pop()  
                        if self.current_right_marker_index > 0:
                            self.current_right_marker_index -= 1
                    elif param == 'reset':
                        self.right_markers.clear()
                        self.current_right_marker_index = 0

                elif marker_type == 'top' and len(self.top_markers)!=0:
                    if param == 'remove':
                        self.top_markers.pop() 
                        if self.current_top_label_index > 0:
                            self.current_top_label_index -= 1
                    elif param == 'reset':
                        self.top_markers.clear()
                        self.current_top_label_index = 0
            
                # Call update live view after resetting markers
                self.update_live_view()
                
            def duplicate_marker(self, marker_type):
                self.save_state() # Save state before making changes

                # --- NEW: Detect the actual content borders ---
                left_border, right_border = self._find_image_content_borders()

                if left_border is None or right_border is None:
                    QMessageBox.warning(self, "Border Detection Failed", 
                                        "Could not detect image content borders. Please ensure image has padding or clear boundaries.")
                    # Revert state if we can't proceed
                    if self.undo_stack: self.undo_stack.pop()
                    return

                if marker_type == 'left' and self.right_markers:
                    # Copy Right Markers TO Left
                    self.left_markers = self.right_markers.copy()
                    
                    # --- FIX: Set offset to the detected LEFT border ---
                    self.left_marker_shift_added = left_border

                    # Update Left Slider position to match the detected border
                    if hasattr(self, 'left_padding_slider'):
                        min_val, max_val = self.left_padding_slider.minimum(), self.left_padding_slider.maximum()
                        clamped_value = max(min_val, min(self.left_marker_shift_added, max_val))
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(clamped_value)
                        self.left_padding_slider.blockSignals(False)
                        # Re-sync internal variable with the actual slider value after potential clamping
                        self.left_marker_shift_added = self.left_padding_slider.value()

                elif marker_type == 'right' and self.left_markers:
                    # Copy Left Markers TO Right
                    self.right_markers = self.left_markers.copy()

                    # --- FIX: Set offset to the detected RIGHT border ---
                    self.right_marker_shift_added = right_border

                    # Update Right Slider position to match the detected border
                    if hasattr(self, 'right_padding_slider'):
                        min_val, max_val = self.right_padding_slider.minimum(), self.right_padding_slider.maximum()
                        clamped_value = max(min_val, min(self.right_marker_shift_added, max_val))
                        self.right_padding_slider.blockSignals(True)
                        self.right_padding_slider.setValue(clamped_value)
                        self.right_padding_slider.blockSignals(False)
                        # Re-sync internal variable
                        self.right_marker_shift_added = self.right_padding_slider.value()

                # Call update live view after duplicating markers and updating offsets/sliders
                self.update_live_view()
                
            def on_combobox_changed(self):
                preset_name = self.combo_box.currentText()
                
                # Save state if switching *from* a modified "Custom" state or another preset
                # This allows undoing the application of a new preset template.
                if self.is_modified or preset_name != "Custom": # More precise condition
                    self.save_state()

                if preset_name == "Custom":
                    self.marker_values_textbox.setEnabled(True)
                    self.rename_input.setEnabled(True)
                    self.rename_input.clear() # Clear rename field, ready for new name
                    self.marker_values_textbox.setPlaceholderText("Current L/R values, edit to save as new")
                    # Update L/R marker values textbox from current internal self.marker_values
                    current_lr_display_values = [str(v) if isinstance(v, (int, float)) else v for v in getattr(self, 'marker_values', [])]
                    self.marker_values_textbox.setText(", ".join(current_lr_display_values))

                    # Update Top labels textedit from current internal self.top_label
                    current_top_display_labels = [str(v) for v in getattr(self, 'top_label', [])]
                    self.top_marker_input.setText(", ".join(current_top_display_labels))

                elif preset_name in self.presets_data:
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.rename_input.clear()

                    preset_config = self.presets_data[preset_name]
                    
                    # Load L/R marker values for the preset into the internal list and textbox
                    self.marker_values = list(preset_config.get("marker_values", []))
                    display_marker_values = [str(v) if isinstance(v, (int, float)) else v for v in self.marker_values]
                    self.marker_values_textbox.setText(", ".join(display_marker_values))

                    # Load Top labels for the preset into the internal list and textedit
                    self.top_label = list(preset_config.get("top_labels", []))
                    self.top_marker_input.setText(", ".join(map(str, self.top_label)))

                    # --- LOAD CUSTOM MARKERS FROM PRESET ---
                    custom_markers_config_from_preset = preset_config.get("custom_markers_config", [])
                    if not isinstance(custom_markers_config_from_preset, list):
                        print(f"Warning: 'custom_markers_config' for preset '{preset_name}' is not a list.")
                        custom_markers_config_from_preset = []
                    deserialized_markers = self._deserialize_custom_markers(custom_markers_config_from_preset)
                    self.custom_markers = [list(m) for m in deserialized_markers] # Populate self.custom_markers

                    # --- LOAD CUSTOM SHAPES FROM PRESET ---
                    custom_shapes_config_from_preset = preset_config.get("custom_shapes_config", [])
                    if not isinstance(custom_shapes_config_from_preset, list):
                        print(f"Warning: 'custom_shapes_config' for preset '{preset_name}' is not a list.")
                        custom_shapes_config_from_preset = []
                    self.custom_shapes = [dict(s) for s in custom_shapes_config_from_preset if isinstance(s, dict)] # Populate self.custom_shapes
                    
                    # Update standard L/R/Top markers *on the image* if any are already placed.
                    # This also calls update_live_view() at the end.
                    self.update_all_labels() 
                    
                else: # Preset name not found in data (should be rare)
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.marker_values_textbox.clear()
                    self.top_marker_input.clear()
                    self.custom_markers.clear() 
                    self.custom_shapes.clear()  
                    self.marker_values = []
                    self.top_label = []
                    self.update_live_view()

            
            def reset_gamma_contrast(self):
                try:
                    if self.image_before_contrast==None:
                        self.image_before_contrast=self.image_master.copy()
                    self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
                    self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
                    self.high_slider.setValue(100)  # Reset contrast to default
                    self.low_slider.setValue(100)  # Reset contrast to default
                    self.gamma_slider.setValue(100)  # Reset gamma to default
                    self.update_live_view()
                except:
                    pass

            
            def update_image_contrast(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                    
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                        self.update_live_view()
                except:
                    pass
            
            def update_image_gamma(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                        
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                        self.update_live_view()
                except:
                    pass
            
            def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
                """
                Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
                preserving the original format (including color and bit depth) where possible.
                Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
                """
                if not qimage or qimage.isNull():
                    return qimage

                original_format = qimage.format()
                try:
                    img_array = self.qimage_to_numpy(qimage)
                    if img_array is None: raise ValueError("NumPy conversion failed.")

                    # Work with float64 for calculations to avoid precision issues
                    img_array_float = img_array.astype(np.float64)

                    # Determine max value based on original data type
                    if img_array.dtype == np.uint16:
                        max_val = 65535.0
                    elif img_array.dtype == np.uint8:
                        max_val = 255.0
                    else: # Default for unexpected types (e.g., float input?)
                         max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0

                    # --- Apply adjustments ---
                    if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                        num_channels = img_array.shape[2]
                        adjusted_channels = []

                        # Process only the color channels (first 3 usually)
                        channels_to_process = min(num_channels, 3)
                        for i in range(channels_to_process):
                            channel = img_array_float[:, :, i]
                            # Normalize to 0-1 range
                            channel_norm = channel / max_val

                            # Apply brightness (high_factor): Multiply
                            channel_norm = channel_norm * high_factor

                            # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                            mid_grey = 0.5
                            contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                            channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                            # Clip to 0-1 range after contrast/brightness
                            channel_norm = np.clip(channel_norm, 0.0, 1.0)

                            # Apply gamma correction
                            safe_gamma = max(0.01, gamma)
                            channel_norm = np.power(channel_norm, safe_gamma)

                            # Clip again after gamma
                            channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                            # Scale back to original range
                            adjusted_channels.append(channel_norm_clipped * max_val)

                        # Reconstruct the image array
                        img_array_final_float = np.stack(adjusted_channels, axis=2)

                        # Keep the alpha channel (if present) untouched
                        if num_channels == 4:
                            alpha_channel = img_array_float[:, :, 3] # Get original alpha
                            img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

                    elif img_array.ndim == 2: # Grayscale Image
                        # Normalize to 0-1 range
                        img_array_norm = img_array_float / max_val

                        # Apply brightness
                        img_array_norm = img_array_norm * high_factor

                        # Apply contrast
                        mid_grey = 0.5
                        contrast_factor = max(0.01, low_factor)
                        img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                        # Clip
                        img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                        # Apply gamma
                        safe_gamma = max(0.01, gamma)
                        img_array_norm = np.power(img_array_norm, safe_gamma)

                        # Clip again
                        img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                        # Scale back
                        img_array_final_float = img_array_norm_clipped * max_val
                    else:
                        return qimage # Return original if unsupported dimensions

                    # Convert back to original data type
                    img_array_final = img_array_final_float.astype(img_array.dtype)

                    # Convert back to QImage using the helper function
                    result_qimage = self.numpy_to_qimage(img_array_final)
                    if result_qimage.isNull():
                        raise ValueError("Conversion back to QImage failed.")

                    # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
                    return result_qimage

                except Exception as e:
                    traceback.print_exc() # Print detailed traceback
                    return qimage # Return original QImage on error
            

            def save_contrast_options(self):
                if self.image:
                    self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
                    self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
                else:
                    QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

            def remove_config(self): # This is for "Remove Preset" button
                selected_preset_name = self.combo_box.currentText()

                if selected_preset_name == "Custom":
                    QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' option.")
                    return
                
                # Prevent deletion of essential default presets if desired (example)
                # essential_defaults = ["Precision Plus Protein All Blue Prestained (Bio-Rad)", "1 kb Plus DNA Ladder (Thermo 10787018)"]
                # if selected_preset_name in essential_defaults:
                #     QMessageBox.warning(self, "Error", f"Cannot remove the built-in preset: '{selected_preset_name}'.")
                #     return

                if selected_preset_name not in self.presets_data:
                    QMessageBox.warning(self, "Error", f"Preset '{selected_preset_name}' not found in data.")
                    return

                reply = QMessageBox.question(self, "Confirm Remove",
                                             f"Are you sure you want to remove the preset '{selected_preset_name}'?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return

                try:
                    del self.presets_data[selected_preset_name]

                    # Save the updated configuration
                    config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)

                    # Update ComboBox
                    self.combo_box.blockSignals(True)
                    current_idx = self.combo_box.findText(selected_preset_name)
                    if current_idx != -1:
                        self.combo_box.removeItem(current_idx)
                    # Select "Custom" or the first available preset after removal
                    if self.combo_box.count() > 1: # More than just "Custom" left
                        custom_idx = self.combo_box.findText("Custom")
                        if self.combo_box.currentIndex() == custom_idx and custom_idx > 0: # If custom was selected and not first
                             self.combo_box.setCurrentIndex(custom_idx -1) # Select item before custom
                        elif self.combo_box.currentIndex() == custom_idx and custom_idx == 0 and self.combo_box.count() > 1: # custom is first, but others exist
                            self.combo_box.setCurrentIndex(1)
                        # else keep current selection if it's not the removed one
                    elif self.combo_box.count() == 1 and self.combo_box.itemText(0) == "Custom":
                        self.combo_box.setCurrentIndex(0)

                    self.combo_box.blockSignals(False)
                    self.on_combobox_changed() # Refresh UI based on new selection

                    QMessageBox.information(self, "Success", f"Preset '{selected_preset_name}' removed.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Error removing preset: {e}")
                    traceback.print_exc()

            def save_config(self): # This is for "Save Preset" button
                """Saves the current settings (L/R/Top values, custom markers/shapes) to the selected/new preset name."""
                
                current_combo_text = self.combo_box.currentText()
                new_name_from_input = self.rename_input.text().strip()
                
                name_to_save = ""
                is_new_preset_creation = False

                if new_name_from_input: # User entered a name in rename_input
                    name_to_save = new_name_from_input
                    is_new_preset_creation = True
                    if name_to_save == "Custom":
                        QMessageBox.warning(self, "Invalid Name", "Cannot save a preset with the name 'Custom'. Please choose another name or clear the rename field to update the selected preset.")
                        return
                elif current_combo_text != "Custom": # Updating an existing preset
                    name_to_save = current_combo_text
                else: # "Custom" is selected, and rename_input is empty
                    QMessageBox.information(self, "Save Preset", "Please enter a new name in the 'New name for Custom preset' field to save the current settings as a new preset.")
                    return

                # Confirmation for overwriting
                if name_to_save in self.presets_data and is_new_preset_creation : # Only ask if it's a new name that already exists
                     reply = QMessageBox.question(self, "Overwrite Preset?",
                                                      f"A preset named '{name_to_save}' already exists. Overwrite it?",
                                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                     if reply == QMessageBox.No:
                         return
                
                # --- Gather data to save ---
                # L/R Marker Values: Always take from the textbox.
                try:
                    raw_markers_text = self.marker_values_textbox.text().strip("[]")
                    if raw_markers_text:
                        marker_values_to_save = [
                            int(num.strip()) if num.strip().isdigit() else float(num.strip()) if '.' in num.strip() else num.strip()
                            for num in raw_markers_text.split(",") if num.strip()
                        ]
                    else:
                        marker_values_to_save = []
                except ValueError:
                    QMessageBox.warning(self, "Input Error", "Invalid format for L/R marker values. Please use comma-separated numbers or text.")
                    return

                # Top Labels: Always take from the textedit.
                top_labels_to_save = [label.strip() for label in self.top_marker_input.toPlainText().split(",") if label.strip()]
                
                # Custom Markers: Serialize current self.custom_markers
                custom_markers_config_to_save = self._serialize_custom_markers(getattr(self, 'custom_markers', []))
                
                # Custom Shapes: Are already serializable
                custom_shapes_config_to_save = [dict(s) for s in getattr(self, 'custom_shapes', [])] # Ensure they are dicts

                # --- Update self.presets_data in memory ---
                self.presets_data[name_to_save] = {
                    "marker_values": marker_values_to_save,
                    "top_labels": top_labels_to_save,
                    "custom_markers_config": custom_markers_config_to_save,
                    "custom_shapes_config": custom_shapes_config_to_save
                }

                # --- Save to file ---
                config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                try:
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)
                    
                    # --- Update ComboBox if a new preset was created ---
                    if is_new_preset_creation and self.combo_box.findText(name_to_save) == -1:
                        self.combo_box.blockSignals(True)
                        custom_idx = self.combo_box.findText("Custom")
                        if custom_idx != -1:
                            self.combo_box.insertItem(custom_idx, name_to_save)
                        else:
                            self.combo_box.addItem(name_to_save) # Fallback
                        self.combo_box.setCurrentText(name_to_save) # Select the new/updated preset
                        self.combo_box.blockSignals(False)
                        # self.on_combobox_changed() # Triggered by setCurrentText if not blocked

                    self.rename_input.clear() # Clear rename field after successful save
                    self.marker_values_textbox.setEnabled(False) # Disable textbox after saving to a named preset
                    self.rename_input.setEnabled(False)

                    QMessageBox.information(self, "Preset Saved", f"Preset '{name_to_save}' saved successfully.")

                except Exception as e:
                    QMessageBox.critical(self, "Save Preset Error", f"Could not save preset configuration:\n{e}")
                    traceback.print_exc()
            
            
            def load_config(self):
                """
                Load preset configuration from file. If the file doesn't exist,
                create it with default popular marker standards and empty custom templates.
                """
                config_loaded_successfully = False
                self.presets_data.clear() # Start with an empty dictionary

                # --- Determine Application Path (same as your existing load_config) ---
                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                     application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()
                # --- End Application Path Detection ---

                config_filepath = os.path.join(application_path, self.CONFIG_PRESET_FILE_NAME)
                print(f"INFO: Attempting to load/create preset config at: {config_filepath}")

                # --- Define Default Marker Data ---
                # Use kDa for proteins, bp for DNA for clarity in keys, but values are just numbers
                default_marker_standards = {
                    # Proteins (kDa)
                    "Precision Plus Protein All Blue Prestained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Unstained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Dual Color (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "PageRuler Prestained Protein Ladder (Thermo)": [250, 130, 100, 70, 55, 35, 25, 15, 10],
                    "PageRuler Unstained Protein Ladder (Thermo)": [200, 150, 100, 70, 50, 40, 30, 20, 10],
                    "Spectra Multicolor Broad Range Protein Ladder (Thermo)": [260, 140, 100, 75, 60, 45, 35, 25, 15, 10],
                    # DNA (bp)
                    "1 kb DNA Ladder (NEB N3232)": [10000, 8000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 500],
                    "1 kb Plus DNA Ladder (Thermo 10787018)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 75],
                    "TrackIt 1 Kb Plus DNA Ladder (Invitrogen 10488085)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
                    "Lambda DNA/HindIII Marker (NEB N3012)": [23130, 9416, 6557, 4361, 2322, 2027, 564],
                }
                # Generic default top labels
                default_top_labels_generic = ["MWM", "L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L13", "L14", "L15"]
                # default_top_label_dict = {name: default_top_labels_generic[:] for name in default_marker_values.keys()} # Assign generic to all defaults
                default_presets_init = {}
                for name, markers in default_marker_standards.items():
                    default_presets_init[name] = {
                        "marker_values": list(markers), # Ensure it's a list
                        "top_labels": list(default_top_labels_generic), # Ensure it's a list
                        "custom_markers_config": [], # Empty by default
                        "custom_shapes_config": []   # Empty by default
                    }
                # --- End Default Marker Data ---

                # --- Initialize internal dictionaries BEFORE trying to load/create ---
                # Start with defaults, then overwrite if load is successful.
                if os.path.exists(config_filepath):
                    try:
                        with open(config_filepath, "r", encoding='utf-8') as f:
                            loaded_json = json.load(f)
                        
                        # Check for new "presets" key
                        if "presets" in loaded_json and isinstance(loaded_json["presets"], dict):
                            self.presets_data = loaded_json["presets"]
                            # Ensure all default fields exist for loaded presets
                            for preset_name, data in self.presets_data.items():
                                if "marker_values" not in data: data["marker_values"] = []
                                if "top_labels" not in data: data["top_labels"] = []
                                if "custom_markers_config" not in data: data["custom_markers_config"] = []
                                if "custom_shapes_config" not in data: data["custom_shapes_config"] = []
                            config_loaded_successfully = True
                        else: # Try to migrate old format
                            print("INFO: Old config format detected. Attempting migration...")
                            migrated_data = {}
                            old_marker_values = loaded_json.get("marker_values", {})
                            old_top_labels = loaded_json.get("top_label", {})
                            all_preset_names = set(old_marker_values.keys()) | set(old_top_labels.keys())

                            for name in all_preset_names:
                                migrated_data[name] = {
                                    "marker_values": list(old_marker_values.get(name, [])),
                                    "top_labels": list(old_top_labels.get(name, default_top_labels_generic)),
                                    "custom_markers_config": [], # No custom markers in old format
                                    "custom_shapes_config": []   # No custom shapes in old format
                                }
                            self.presets_data = migrated_data
                            # Save immediately in new format after migration
                            with open(config_filepath, "w", encoding='utf-8') as f_new:
                                json.dump({"presets": self.presets_data}, f_new, indent=4)
                            print("INFO: Config migrated to new format and saved.")
                            config_loaded_successfully = True

                    except (json.JSONDecodeError, IOError, TypeError) as e:
                        QMessageBox.warning(self, "Preset Config Load Error",
                                            f"Could not load '{self.CONFIG_PRESET_FILE_NAME}':\n{e}\n\nUsing default presets.")
                        self.presets_data = default_presets_init.copy() # Fallback to defaults
                    except Exception as e:
                        traceback.print_exc()
                        QMessageBox.warning(self, "Preset Config Load Error",
                                            f"Unexpected error loading '{self.CONFIG_PRESET_FILE_NAME}'.\n\nUsing default presets.")
                        self.presets_data = default_presets_init.copy() # Fallback
                else:
                    # Config file NOT found, create it with defaults
                    self.presets_data = default_presets_init.copy()
                    try:
                        with open(config_filepath, "w", encoding='utf-8') as f:
                            json.dump({"presets": self.presets_data}, f, indent=4)
                        config_loaded_successfully = True # Created successfully
                        print(f"INFO: Default preset config created at: {config_filepath}")
                    except Exception as e:
                        QMessageBox.critical(self, "Preset Config Creation Error",
                                             f"Could not create default preset config file:\n{e}\n\nDefault presets will be used for this session only.")
                        # self.presets_data remains default_presets_init

                # --- Update UI (ComboBox) ---
                try:
                    if hasattr(self, 'combo_box'):
                        self.combo_box.blockSignals(True)
                        self.combo_box.clear()
                        sorted_preset_names = sorted(self.presets_data.keys())
                        self.combo_box.addItems(sorted_preset_names)
                        self.combo_box.addItem("Custom")

                        # Set current item (same logic as your existing load_config)
                        default_biorad = "Precision Plus Protein All Blue Prestained (Bio-Rad)"
                        idx_to_select = self.combo_box.findText(default_biorad)
                        if idx_to_select == -1 and sorted_preset_names:
                            idx_to_select = 0 # Select first actual preset if BioRad not found
                        elif idx_to_select == -1: # No actual presets, only "Custom"
                            idx_to_select = self.combo_box.findText("Custom")
                        
                        if idx_to_select != -1:
                             self.combo_box.setCurrentIndex(idx_to_select)
                        
                        self.combo_box.blockSignals(False)
                        self.on_combobox_changed() # Trigger update based on selection
                    else:
                        print("Warning: combo_box UI element not found during config load.")
                except Exception as e_ui:
                    traceback.print_exc()
                    QMessageBox.warning(self, "UI Error", f"Error populating preset combobox: {e_ui}")
            
            def paste_image(self):
                """Handle pasting image from clipboard."""
                self.is_modified = True
                self.reset_image() # Clear previous state first
                self.load_image_from_clipboard()
                self._update_overlay_slider_ranges()
                # UI updates (label size, sliders) should happen within load_image_from_clipboard
                self.update_live_view()
                self.save_state() # Save state after pasting
            
            def load_image_from_clipboard(self):
                """
                Load an image from the clipboard into self.image, preserving format.
                Prioritizes file paths over raw image data to handle macOS file copying correctly.
                If loaded from a file path, attempts to load an associated config file.
                """
                # Check for unsaved changes before proceeding
                # if not self.prompt_save_if_needed(): # Optional: uncomment if needed
                #     return # Abort paste if user cancels save

                self.reset_image() # Clear previous state first

                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()
                loaded_image = None
                source_info = "Clipboard" # Default source
                config_loaded_from_paste = False # Flag to track if config was loaded
                

                # --- PRIORITY 1: Check for File URLs ---
                if mime_data.hasUrls():
                    urls = mime_data.urls()
                    if urls:
                        file_path = urls[0].toLocalFile()
                        if os.path.isfile(file_path) and file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                            loaded_image = QImage(file_path)
                            source_info = file_path

                            if loaded_image.isNull():
                                try:
                                    pil_image = Image.open(file_path)
                                    loaded_image = ImageQt.ImageQt(pil_image)
                                    if loaded_image.isNull():
                                        loaded_image = None
                                    else:
                                        source_info = f"{file_path} (Pillow)"
                                except Exception as e:
                                    QMessageBox.warning(self, "File Load Error", f"Could not load image from file '{os.path.basename(file_path)}':\n{e}")
                                    loaded_image = None

                            # --- *** ADDED: CONFIG FILE LOADING FOR PASTED FILE *** ---
                            if loaded_image and not loaded_image.isNull():
                                self.image_path = file_path # Store the path early for config lookup
                                base_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                                # Construct potential config file name (remove _original/_modified if present)
                                config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                                config_path = os.path.join(os.path.dirname(file_path), f"{config_base}_config.txt")

                                if os.path.exists(config_path):
                                    try:
                                        with open(config_path, "r") as config_file:
                                            config_data = json.load(config_file)
                                        self.apply_config(config_data) # Apply loaded settings
                                        config_loaded_from_paste = True # Set flag
                                    except Exception as e:
                                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply associated config file '{os.path.basename(config_path)}': {e}")

                            # --- *** END OF ADDED CONFIG LOADING *** ---


                # --- PRIORITY 2: Check for Raw Image Data (if no valid file was loaded) ---
                if loaded_image is None and mime_data.hasImage():
                    loaded_image = clipboard.image()
                    if not loaded_image.isNull():
                        source_info = "Clipboard Image Data"
                    else:
                        try:
                            pil_image = ImageGrab.grabclipboard()
                            if isinstance(pil_image, Image.Image):
                                loaded_image = ImageQt.ImageQt(pil_image)
                                if loaded_image.isNull():
                                     loaded_image = None
                                else:
                                    source_info = "Clipboard Image Data (Pillow Grab)"
                            else:
                                loaded_image = None
                        except Exception: # Keep quiet on Pillow grab errors
                            loaded_image = None

                # --- Process the successfully loaded image ---
                if loaded_image and not loaded_image.isNull():
                    self.is_modified = True
                    self.image = loaded_image

                    # Initialize backups
                    self.original_image = self.image.copy()
                    self.image_master = self.image.copy()
                    self.image_before_padding = None
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False
                    # self.image_path is set earlier if loaded from file

                    # --- Update UI Elements ---
                    try:
                        w = self.image.width()
                        h = self.image.height()
                        if w <= 0 or h <= 0: raise ValueError("Invalid image dimensions after load.")

                        # Update preview label size
                        ratio=w/h if h > 0 else 1
                        target_width = int(self.label_size) # Use current label_size setting
                        target_height = int(target_width / ratio)
                        # Optional: Apply max height constraint
                        # if hasattr(self, 'preview_label_max_height_setting') and target_height > self.preview_label_max_height_setting:
                        #     target_height = self.preview_label_max_height_setting
                        #     target_width = int(target_height * ratio)
                        self._update_preview_label_size()

                        # Update slider ranges based on *new* label size
                        render_scale = 3
                        render_width = self.live_view_label.width() * render_scale
                        render_height = self.live_view_label.height() * render_scale
                        self._update_marker_slider_ranges()

                        # Set recommended padding only if config wasn't loaded
                        if not config_loaded_from_paste:
                            self.left_padding_input.setText(str(int(w * 0.1)))
                            self.right_padding_input.setText(str(int(w * 0.1)))
                            self.top_padding_input.setText(str(int(h * 0.15)))
                            self.bottom_padding_input.setText("0")

                        # Update window title and status bar
                        display_source = os.path.basename(source_info.split(" (")[0]) # Show filename or simplified source
                        title_suffix = " (+ Config)" if config_loaded_from_paste else ""
                        self.setWindowTitle(f"{self.window_title}::{display_source}{title_suffix}")
                        self._update_status_bar()

                    except Exception as e:
                        QMessageBox.warning(self, "UI Update Error", f"Could not update UI elements after pasting: {e}")
                        self._update_preview_label_size()
                        
                    enable_pan = self.live_view_label.zoom_level > 1.0
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                    self.update_live_view()
                    self.save_state()

                else:
                    # If no image was successfully loaded
                    QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
                    # Clean up state
                    self.image = None
                    self.original_image = None
                    self.image_master = None
                    self.image_path = None
                    self.live_view_label.clear()
                    self._update_preview_label_size()
                    self.setWindowTitle(self.window_title)
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                    self._update_status_bar()
                    self.update_live_view()
                
            def update_font(self):
                self.save_state()
                """Update the font settings based on UI inputs"""
                # Update font family from the combo box
                self.font_family = self.font_combo_box.currentFont().family()
                
                # Update font size from the spin box
                self.font_size = self.font_size_spinner.value()
                
                self.font_rotation = int(self.font_rotation_input.value())
                
            
                # Once font settings are updated, update the live view immediately
                self.update_live_view()
            
            def select_font_color(self):
                self.save_state()
                color = QColorDialog.getColor()
                if color.isValid():
                    self.font_color = color  # Store the selected color   
                    self.update_font()
                self._update_color_button_style(self.font_color_button, self.font_color)
                    
            def load_image(self):
                self.prompt_save_if_needed()
                # self.undo_stack = []
                # self.redo_stack = []
                self.reset_image() # Clear previous state

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getOpenFileName(
                    self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
                )
                if file_path:
                    self.image_path = file_path
                    loaded_image = QImage(self.image_path)                    

                    if loaded_image.isNull():
                        # Try loading with Pillow as fallback
                        try:
                            pil_image = Image.open(self.image_path)
                            # Use ImageQt for reliable conversion, preserves most formats
                            loaded_image = ImageQt.ImageQt(pil_image)
                            if loaded_image.isNull():
                                raise ValueError("Pillow could not convert to QImage.")

                        except Exception as e:
                            QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}': {e}")
                            self.image_path = None
                            return

                    # --- Keep the loaded image format ---
                    self.image = loaded_image
                    self._update_overlay_slider_ranges()

                    # --- Initialize backups with the loaded format ---
                    if not self.image.isNull():
                        self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                        self.image_master = self.image.copy()   # Master copy for resets
                        self.image_before_padding = None        # Reset padding state
                        self.image_contrasted = self.image.copy() # Backup for contrast
                        self.image_before_contrast = self.image.copy() # Backup for contrast
                        self.image_padded = False               # Reset flag

                        self.setWindowTitle(f"{self.window_title}::{self.image_path}")

                        # --- Load Associated Config File (Logic remains the same) ---
                        self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                        config_name = ""
                        if self.base_name.endswith("_original"):
                            config_name = self.base_name.replace("_original", "_config.txt")
                        else:
                            config_name = self.base_name + "_config.txt"

                        config_path = os.path.join(os.path.dirname(file_path), config_name)

                        if os.path.exists(config_path):
                            try:
                                with open(config_path, "r") as config_file:
                                    config_data = json.load(config_file)
                                self.apply_config(config_data) # Apply loaded settings
                            except Exception as e:
                                QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                        # --- End Config File Loading ---
                        self.is_modified = True # Mark as modified when loading new image
                    else:
                         QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                         return

                    # --- Update UI Elements (Label size, sliders) ---
                    if self.image and not self.image.isNull():
                        try:
                            # (UI update logic remains the same as before)
                            self._update_preview_label_size()

                            render_scale = 3
                            render_width = self.live_view_label.width() * render_scale
                            render_height = self.live_view_label.height() * render_scale
                            self._update_marker_slider_ranges()


                            if not os.path.exists(config_path):
                                self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                                self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                                self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                                self.bottom_padding_input.setText("0")

                        except Exception as e:
                            pass
                    # --- End UI Element Update ---
                    
                    
                    self._update_status_bar() # <--- Add this
                    enable_pan = self.live_view_label.zoom_level > 1.0
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                    self.update_live_view() # Render the loaded image
                    self.save_state() # Save initial loaded state
            
            def apply_config(self, config_data):
                # --- 1. Load data from config_data into self attributes ---

                # Padding text inputs
                adding_white_space_data = config_data.get("adding_white_space", {})
                self.left_padding_input.setText(str(adding_white_space_data.get("left", "0")))
                self.right_padding_input.setText(str(adding_white_space_data.get("right", "0")))
                self.top_padding_input.setText(str(adding_white_space_data.get("top", "0")))
                self.bottom_padding_input.setText(str(adding_white_space_data.get("bottom", "0")))
                try:
                    self.transparency = int(adding_white_space_data.get("transparency", 1))
                except (ValueError, TypeError):
                    self.transparency = 1

                # Standard marker positions
                marker_positions_data = config_data.get("marker_positions", {})
                self.left_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("left", [])]
                self.right_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("right", [])]
                self.top_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("top", [])]

                # Top labels (internal list)
                marker_labels_data = config_data.get("marker_labels", {})
                self.top_label = [str(label).strip() for label in marker_labels_data.get("top", []) if str(label).strip()]

                # L/R marker values (internal list) - Determine if they are "custom" from this image config
                custom_lr_marker_values_from_config = []
                config_suggests_custom_lr_for_combobox = False
                lr_source_key = None
                # Check "left" first, then "right" as the source for L/R values in an image-specific config
                if "left" in marker_labels_data and isinstance(marker_labels_data["left"], list):
                    lr_source_key = "left"
                elif "right" in marker_labels_data and isinstance(marker_labels_data["right"], list):
                    lr_source_key = "right"
                
                if lr_source_key:
                    raw_values = marker_labels_data[lr_source_key]
                    for v_str in raw_values:
                        s_val = str(v_str).strip()
                        if not s_val: continue # Skip empty strings
                        try: custom_lr_marker_values_from_config.append(int(s_val))
                        except ValueError:
                            try: custom_lr_marker_values_from_config.append(float(s_val))
                            except ValueError: custom_lr_marker_values_from_config.append(s_val) # Keep as string if not number
                    if custom_lr_marker_values_from_config:
                         config_suggests_custom_lr_for_combobox = True
                
                if config_suggests_custom_lr_for_combobox:
                    self.marker_values = custom_lr_marker_values_from_config # Set internal list
                else:
                    # If no L/R values from this specific image config,
                    # self.marker_values will be populated by on_combobox_changed based on the selected preset.
                    self.marker_values = []


                # Font options for standard markers
                font_options_data = config_data.get("font_options", {})
                self.font_family = font_options_data.get("font_family", "Arial")
                self.font_size = int(font_options_data.get("font_size", 12))
                self.font_rotation = int(font_options_data.get("font_rotation", -45))
                font_color_str = font_options_data.get("font_color", "#000000")
                self.font_color = QColor(font_color_str)
                if not self.font_color.isValid(): self.font_color = QColor(0,0,0) # Fallback


                # --- Marker Shifts and Slider Value Logic ---
                marker_padding_data = config_data.get("marker_padding", {})
                added_shift_data = config_data.get("added_shift", {})

                # 1. Determine the values from "marker_padding" (these will directly set the sliders)
                try: padding_val_left = int(marker_padding_data.get("left", 0))
                except (ValueError, TypeError): padding_val_left = 0
                try: padding_val_right = int(marker_padding_data.get("right", 0))
                except (ValueError, TypeError): padding_val_right = 0
                try: padding_val_top = int(marker_padding_data.get("top", 0))
                except (ValueError, TypeError): padding_val_top = 0

                # 2. Load "added_shift" values from config. These are the true internal state.
                #    If "added_shift" is missing, fall back to "marker_padding" values for the internal state.
                self.left_marker_shift_added = int(added_shift_data.get("left", padding_val_left))
                self.right_marker_shift_added = int(added_shift_data.get("right", padding_val_right))
                self.top_marker_shift_added = int(added_shift_data.get("top", padding_val_top))

                # 3. Load and set slider RANGES.
                slider_ranges_data = config_data.get("slider_ranges", {})
                lr_range_conf = slider_ranges_data.get("left", [])
                rr_range_conf = slider_ranges_data.get("right", [])
                tr_range_conf = slider_ranges_data.get("top", [])
                default_min_range, default_max_range_w, default_max_range_h = -1000, 2000, 2000 # Adjusted defaults

                if hasattr(self, 'left_padding_slider'):
                    try:
                        lr_min, lr_max = (int(lr_range_conf[0]), int(lr_range_conf[1])) if len(lr_range_conf) == 2 else (default_min_range, default_max_range_w)
                        self.left_padding_slider.setRange(lr_min, lr_max); self.left_slider_range = [lr_min, lr_max]
                    except (IndexError, TypeError, ValueError): self.left_padding_slider.setRange(default_min_range, default_max_range_w)
                if hasattr(self, 'right_padding_slider'):
                    try:
                        rr_min, rr_max = (int(rr_range_conf[0]), int(rr_range_conf[1])) if len(rr_range_conf) == 2 else (default_min_range, default_max_range_w)
                        self.right_padding_slider.setRange(rr_min, rr_max); self.right_slider_range = [rr_min, rr_max]
                    except (IndexError, TypeError, ValueError): self.right_padding_slider.setRange(default_min_range, default_max_range_w)
                if hasattr(self, 'top_padding_slider'):
                    try:
                        tr_min, tr_max = (int(tr_range_conf[0]), int(tr_range_conf[1])) if len(tr_range_conf) == 2 else (default_min_range, default_max_range_h)
                        self.top_padding_slider.setRange(tr_min, tr_max); self.top_slider_range = [tr_min, tr_max]
                    except (IndexError, TypeError, ValueError): self.top_padding_slider.setRange(default_min_range, default_max_range_h)

                # 4. Set slider VALUES using self.xxx_marker_shift_added (which now holds the true state)
                #    Signals are blocked. Qt will clamp these if they are outside the range just set.
                if hasattr(self, 'left_padding_slider'):
                    self.left_padding_slider.blockSignals(True); self.left_padding_slider.setValue(self.left_marker_shift_added); self.left_padding_slider.blockSignals(False)
                if hasattr(self, 'right_padding_slider'):
                    self.right_padding_slider.blockSignals(True); self.right_padding_slider.setValue(self.right_marker_shift_added); self.right_padding_slider.blockSignals(False)
                if hasattr(self, 'top_padding_slider'):
                    self.top_padding_slider.blockSignals(True); self.top_padding_slider.setValue(self.top_marker_shift_added); self.top_padding_slider.blockSignals(False)

                # 5. CRITICAL RE-SYNC: Ensure internal `_marker_shift_added` variables
                #    MATCH the SLIDER'S ACTUAL VALUE after potential clamping by setValue.
                if hasattr(self, 'left_padding_slider'): self.left_marker_shift_added = self.left_padding_slider.value()
                if hasattr(self, 'right_padding_slider'): self.right_marker_shift_added = self.right_padding_slider.value()
                if hasattr(self, 'top_padding_slider'): self.top_marker_shift_added = self.top_padding_slider.value()
                # --- End Marker Shifts and Slider Logic ---


                # --- ComboBox and UI Update for L/R values and Top Labels ---
                if hasattr(self, 'combo_box'):
                    self.combo_box.blockSignals(True)
                    if config_suggests_custom_lr_for_combobox: # Implies self.marker_values is set from image config
                        custom_idx = self.combo_box.findText("Custom")
                        if custom_idx != -1: self.combo_box.setCurrentIndex(custom_idx)
                    else:
                        # If no L/R in image config, try to load saved preset name from config_data itself.
                        # This 'selected_preset_name' would have been saved if the image was last saved with a specific preset selected.
                        selected_preset_name_from_config = config_data.get("selected_preset_name", None)
                        if selected_preset_name_from_config and self.combo_box.findText(selected_preset_name_from_config) != -1:
                            self.combo_box.setCurrentText(selected_preset_name_from_config)
                        else: # Fallback to a default preset if nothing specific is found.
                            default_biorad = "Precision Plus Protein All Blue Prestained (Bio-Rad)"
                            idx_biorad = self.combo_box.findText(default_biorad)
                            if idx_biorad != -1: self.combo_box.setCurrentIndex(idx_biorad)
                            elif self.combo_box.count() > 0: # If BioRad not found, select first available or "Custom"
                                custom_idx_fallback = self.combo_box.findText("Custom")
                                if custom_idx_fallback == 0 and self.combo_box.count() > 1: self.combo_box.setCurrentIndex(1)
                                elif custom_idx_fallback != -1: self.combo_box.setCurrentIndex(custom_idx_fallback)
                                else: self.combo_box.setCurrentIndex(0)
                    self.combo_box.blockSignals(False)
                    self.on_combobox_changed() # Updates text boxes based on current combobox selection and internal lists

                # Load custom markers and shapes
                custom_markers_data_from_image_config = config_data.get("custom_markers_config", config_data.get("custom_markers", [])) # Check both keys
                self.custom_markers = [list(m) for m in self._deserialize_custom_markers(custom_markers_data_from_image_config)]
                
                self.custom_shapes = []
                loaded_shapes_from_config = config_data.get("custom_shapes_config", config_data.get("custom_shapes", [])) # Check both keys
                if isinstance(loaded_shapes_from_config, list):
                    for shape_item in loaded_shapes_from_config:
                        if isinstance(shape_item, dict):
                            self.custom_shapes.append(dict(shape_item))

                # Update standard font UI elements
                if hasattr(self, 'font_combo_box'):
                    self.font_combo_box.blockSignals(True); self.font_combo_box.setCurrentFont(QFont(self.font_family)); self.font_combo_box.blockSignals(False)
                if hasattr(self, 'font_size_spinner'):
                    self.font_size_spinner.blockSignals(True); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.blockSignals(False)
                if hasattr(self, 'font_rotation_input'):
                    self.font_rotation_input.blockSignals(True); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.blockSignals(False)
                if hasattr(self, 'font_color_button'):
                    self._update_color_button_style(self.font_color_button, self.font_color)
                
                # Apply UI state for custom marker creation tools (if saved in config, e.g. from a full preset)
                ui_custom_settings = config_data.get("ui_custom_marker_settings", {})
                
                custom_marker_color_str = ui_custom_settings.get("color", self.custom_marker_color.name() if hasattr(self, 'custom_marker_color') else "#000000")
                temp_custom_color = QColor(custom_marker_color_str)
                self.custom_marker_color = temp_custom_color if temp_custom_color.isValid() else QColor(0,0,0)

                if hasattr(self, 'custom_font_type_dropdown'):
                    default_custom_family = self.custom_font_type_dropdown.currentFont().family() if self.custom_font_type_dropdown.count() > 0 else "Arial"
                    custom_font_family_str = ui_custom_settings.get("font_family", default_custom_family)
                    self.custom_font_type_dropdown.blockSignals(True)
                    self.custom_font_type_dropdown.setCurrentFont(QFont(custom_font_family_str))
                    self.custom_font_type_dropdown.blockSignals(False)
                    self.update_marker_text_font(self.custom_font_type_dropdown.currentFont())

                if hasattr(self, 'custom_font_size_spinbox'):
                    default_custom_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12
                    custom_font_size_val = int(ui_custom_settings.get("font_size", default_custom_size))
                    self.custom_font_size_spinbox.blockSignals(True)
                    self.custom_font_size_spinbox.setValue(custom_font_size_val)
                    self.custom_font_size_spinbox.blockSignals(False)

                if hasattr(self, 'custom_marker_color_button'):
                    self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                
            def get_current_config(self):
                """Gathers the current application state into a dictionary for saving."""
                def make_json_serializable(value):
                    if isinstance(value, (np.float16, np.float32, np.float64, np.floating)):
                        return float(value)
                    if isinstance(value, (np.int8, np.int16, np.int32, np.int64, np.integer)):
                        return int(value)
                    if isinstance(value, np.bool_):
                        return bool(value)
                    if isinstance(value, list):
                        return [make_json_serializable(v) for v in value]
                    if isinstance(value, tuple):
                        # For JSON, lists are generally preferred over tuples if modification isn't an issue
                        # For coordinates like (pos, label), we want [float(pos), str(label)]
                        if len(value) == 2 and isinstance(value[1], str): # Heuristic for (pos, label)
                            return [make_json_serializable(value[0]), str(value[1])]
                        return [make_json_serializable(v) for v in value] # General tuple to list
                    if isinstance(value, dict):
                        return {make_json_serializable(k): make_json_serializable(v) for k, v in value.items()}
                    return value
                
                config = {
                    "adding_white_space": {
                        "left": self.left_padding_input.text(),
                        "right": self.right_padding_input.text(),
                        "top": self.top_padding_input.text(),
                        "bottom": self.bottom_padding_input.text(),
                        "transparency": self.transparency, # Assuming self.transparency exists
                    },
                    "marker_positions": {
                        # Store positions only for standard markers
                        "left": [(pos, label) for pos, label in getattr(self, 'left_markers', [])],
                        "right": [(pos, label) for pos, label in getattr(self, 'right_markers', [])],
                        "top": [(pos, label) for pos, label in getattr(self, 'top_markers', [])],
                    },
                    "marker_labels": { # Storing labels separately might be redundant if already in positions
                        "top": getattr(self, 'top_label', []),
                        "left": [marker[1] for marker in getattr(self, 'left_markers', [])],
                        "right": [marker[1] for marker in getattr(self, 'right_markers', [])],
                    },
                    "marker_padding": { # Current slider values
                        "top": self.top_padding_slider.value() if hasattr(self, 'top_padding_slider') else 0,
                        "left": self.left_padding_slider.value() if hasattr(self, 'left_padding_slider') else 0,
                        "right": self.right_padding_slider.value() if hasattr(self, 'right_padding_slider') else 0,
                    },
                    "font_options": { # Default font options for standard markers
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_rotation": self.font_rotation,
                        "font_color": self.font_color.name(),
                    },
                    "slider_ranges": { # Current slider ranges
                         "left": getattr(self, 'left_slider_range', [-100, 1000]),
                         "right": getattr(self, 'right_slider_range', [-100, 1000]),
                         "top": getattr(self, 'top_slider_range', [-100, 1000]),
                     },
                    "added_shift": { # Store current added shifts
                         "left": getattr(self, 'left_marker_shift_added', 0),
                         "right": getattr(self, 'right_marker_shift_added', 0),
                         "top": getattr(self, 'top_marker_shift_added', 0),
                    }
                }

                # --- Updated Custom Markers Section ---
                custom_markers_data = []
                # Use getattr for safety, default to empty list
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        # Unpack 8 elements
                        x, y, text, color, font, font_size, is_bold, is_italic = marker_tuple
                        custom_markers_data.append({
                            "x": x,
                            "y": y,
                            "text": text,
                            "color": color.name(), # Save color name/hex
                            "font": font,
                            "font_size": font_size,
                            "bold": is_bold,       # Save bold flag
                            "italic": is_italic    # Save italic flag
                        })
                    except (ValueError, TypeError, IndexError) as e:
                        pass
                config["custom_markers"] = custom_markers_data
                config["custom_shapes"] = [dict(s) for s in getattr(self, "custom_shapes", [])]
                # --- End Updated Custom Markers ---

                return config
            
            def add_band(self, event):
                self.update_all_labels() 
                self.save_state()
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                
                if not self.image or self.image.isNull() or not self.marker_mode:
                    return

                # --- 1. Get Click Position in Unzoomed Label Space (handles zoom/pan) ---
                click_pos_unzoomed_label_space = self.live_view_label.transform_point(event.position())
                
                # --- 2. Apply Grid Snapping in Unzoomed Label Space ---
                snapped_click_pos_label_space = self.snap_point_to_grid(click_pos_unzoomed_label_space)
                cursor_x_ls = snapped_click_pos_label_space.x() 
                cursor_y_ls = snapped_click_pos_label_space.y()

                # --- 3. Transform Snapped Label Space Coords to Native Image Coords ---
                if not (self.image and not self.image.isNull()): # Corrected: use self directly
                    return 

                current_app_image = self.image
 
                label_w_widget = float(self.live_view_label.width())
                label_h_widget = float(self.live_view_label.height())
                img_w_native = float(current_app_image.width())
                img_h_native = float(current_app_image.height())

                if img_w_native <= 0 or img_h_native <= 0 or label_w_widget <= 0 or label_h_widget <= 0:
                    return

                scale_native_to_label = min(label_w_widget / img_w_native, label_h_widget / img_h_native)
                if scale_native_to_label <= 1e-9: 
                    return
                
                displayed_img_w_in_label = img_w_native * scale_native_to_label
                displayed_img_h_in_label = img_h_native * scale_native_to_label
                offset_x_img_in_label = (label_w_widget - displayed_img_w_in_label) / 2.0
                offset_y_img_in_label = (label_h_widget - displayed_img_h_in_label) / 2.0

                image_x = (cursor_x_ls - offset_x_img_in_label) / scale_native_to_label
                image_y = (cursor_y_ls - offset_y_img_in_label) / scale_native_to_label
                
                # Clamp to native image dimensions
                image_x = max(0.0, min(image_x, img_w_native))
                image_y = max(0.0, min(image_y, img_h_native))
                
                try:
                    if self.marker_mode == "left":
                        current_marker_count = len(self.left_markers)
                        is_first_marker = (current_marker_count == 0)
                        marker_value_to_add = self.marker_values[current_marker_count] if current_marker_count < len(self.marker_values) else ""
                        
                        self.left_markers.append((image_y, marker_value_to_add)) 
                        self.current_left_marker_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_x))
                            self._update_marker_slider_ranges() 
                            
                            self.left_padding_slider.blockSignals(True)
                            self.left_padding_slider.setValue(
                                max(self.left_slider_range[0], min(slider_target_value_native_pixels, self.left_slider_range[1]))
                            )
                            self.left_padding_slider.blockSignals(False)
                            # Update internal shift variable to match the actual (possibly clamped) slider value.
                            self.left_marker_shift_added = self.left_padding_slider.value()


                    elif self.marker_mode == "right":
                        current_marker_count = len(self.right_markers)
                        is_first_marker = (current_marker_count == 0)
                        marker_value_to_add = self.marker_values[current_marker_count] if current_marker_count < len(self.marker_values) else ""
                        
                        self.right_markers.append((image_y, marker_value_to_add))
                        self.current_right_marker_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_x))
                            self._update_marker_slider_ranges()
                            
                            self.right_padding_slider.blockSignals(True)
                            self.right_padding_slider.setValue(
                                max(self.right_slider_range[0], min(slider_target_value_native_pixels, self.right_slider_range[1]))
                            )
                            self.right_padding_slider.blockSignals(False)
                            self.right_marker_shift_added = self.right_padding_slider.value()


                    elif self.marker_mode == "top":
                        current_marker_count = len(self.top_markers)
                        is_first_marker = (current_marker_count == 0)
                        label_to_add = self.top_label[current_marker_count] if current_marker_count < len(self.top_label) else ""
                        
                        self.top_markers.append((image_x, label_to_add))
                        self.current_top_label_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_y)) # For Top marker, Y-coord determines offset
                            self._update_marker_slider_ranges()

                            self.top_padding_slider.blockSignals(True)
                            self.top_padding_slider.setValue(
                                max(self.top_slider_range[0], min(slider_target_value_native_pixels, self.top_slider_range[1]))
                            )
                            self.top_padding_slider.blockSignals(False)
                            self.top_marker_shift_added = self.top_padding_slider.value()
                            
                except Exception as e:
                     import traceback
                     traceback.print_exc()
                     QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

                self.update_live_view()
                

                
            def enable_left_marker_mode(self):
                self.marker_mode = "left"
                self.current_left_marker_index = 0 # Reset index for this mode
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)

            def enable_right_marker_mode(self):
                self.marker_mode = "right"
                self.current_right_marker_index = 0
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)
            
            def enable_top_marker_mode(self):
                self.marker_mode = "top"
                self.current_top_label_index = 0 # Assuming this is the correct index variable
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)
                
            # def remove_padding(self):
            #     if self.image_before_padding!=None:
            #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
            #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
            #     self.image_padded = False  # Reset the padding state
            #     w=self.image.width()
            #     h=self.image.height()
            #     # Preview window
            #     ratio=w/h
            #     self.label_width = 540
            #     label_height=int(self.label_width/ratio)
            #     if label_height>self.label_width:
            #         label_height=540
            #         self.label_width=ratio*label_height
            #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            #     self.update_live_view()
                
            def finalize_image(self): # Padding
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
                    return
                try:
                    padding_left = max(0, int(self.left_padding_input.text()))
                    padding_right = max(0, int(self.right_padding_input.text()))
                    padding_top = max(0, int(self.top_padding_input.text()))
                    padding_bottom = max(0, int(self.bottom_padding_input.text()))
                    self.left_padding_input.setText(str(padding_left))
                    self.right_padding_input.setText(str(padding_right))
                    self.top_padding_input.setText(str(padding_top))
                    self.bottom_padding_input.setText(str(padding_bottom))
                except ValueError:
                    QMessageBox.warning(self, "Error", "Please enter valid non-negative integers for padding.")
                    return

                if padding_left == 0 and padding_right == 0 and padding_top == 0 and padding_bottom == 0:
                    QMessageBox.information(self, "Info", "No padding specified. Image remains unchanged.")
                    return

                self.save_state()

                try:
                    # --- 1. Adjust element coordinates and _marker_shift_added values ---
                    self.adjust_elements_for_padding(padding_left, padding_top)

                    # --- 2. Pad the image using NumPy for a pixel-perfect copy ---
                    # >>> MODIFICATION START: Replaced QPainter with NumPy <<<
                    
                    # Convert the source image to a NumPy array
                    np_img = self.qimage_to_numpy(self.image)
                    if np_img is None:
                        raise ValueError("Failed to convert source image to NumPy array for padding.")

                    original_height, original_width = np_img.shape[:2]
                    new_width = original_width + padding_left + padding_right
                    new_height = original_height + padding_top + padding_bottom
                    
                    # Determine the fill value for "white" based on the image's data type
                    # and the correct shape for the new canvas.
                    if np_img.ndim == 2: # Grayscale
                        fill_value = 65535 if np_img.dtype == np.uint16 else 255
                        padded_shape = (new_height, new_width)
                    elif np_img.ndim == 3: # Color
                        channels = np_img.shape[2]
                        fill_value = [255] * channels # e.g., (255, 255, 255) or (255, 255, 255, 255)
                        padded_shape = (new_height, new_width, channels)
                    else:
                        raise ValueError(f"Unsupported image dimension for padding: {np_img.ndim}")

                    # Create a new, larger NumPy array filled with the white value
                    padded_np = np.full(padded_shape, fill_value, dtype=np_img.dtype)
                    
                    # Copy the original image data into the correct slice of the new array.
                    # This is a direct data copy, no interpolation or anti-aliasing will occur.
                    padded_np[padding_top:padding_top + original_height, padding_left:padding_left + original_width] = np_img

                    # Convert the padded NumPy array back to a QImage
                    padded_image = self.numpy_to_qimage(padded_np)
                    if padded_image.isNull():
                        raise ValueError("Conversion back to QImage failed after padding with NumPy.")
                    # >>> MODIFICATION END <<<


                    # --- 3. Update main image and backups ---
                    self.image = padded_image
                    self.image_padded = True 
                    self.is_modified = True
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()

                    # --- 4. Update UI ---
                    self._update_preview_label_size() 
                    self._update_status_bar()         
                    
                    self._update_marker_slider_ranges() 
                    
                    sliders_to_set = [
                        (getattr(self, 'left_padding_slider', None), self.left_marker_shift_added),
                        (getattr(self, 'right_padding_slider', None), self.right_marker_shift_added),
                        (getattr(self, 'top_padding_slider', None), self.top_marker_shift_added)
                    ]
                    for slider, value_to_set in sliders_to_set:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(value_to_set)
                            slider.blockSignals(False)
                            if slider == self.left_padding_slider: self.left_marker_shift_added = slider.value()
                            elif slider == self.right_padding_slider: self.right_marker_shift_added = slider.value()
                            elif slider == self.top_padding_slider: self.top_marker_shift_added = slider.value()

                    self.update_live_view() 

                except Exception as e:
                    QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
                    traceback.print_exc()
                    if self.undo_stack:
                        try: self.undo_action_m()
                        except: pass
                    

            def adjust_elements_for_padding(self, padding_left, padding_top):
                """
                Adjusts coordinates of all markers and custom shapes for added padding.
                Also adjusts the _marker_shift_added values.
                This function is called BEFORE the image is actually padded.
                """

                # Adjust standard markers (Y for L/R, X for Top)
                self.left_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'left_markers', [])]
                self.right_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'right_markers', [])]
                self.top_markers = [(x_pos_img + padding_left, label) for x_pos_img, label in getattr(self, 'top_markers', [])]

                # Adjust custom markers (both X and Y)
                new_custom_markers = []
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        m_list = list(marker_data) # Ensure mutable
                        m_list[0] = float(m_list[0]) + padding_left  # Adjust X
                        m_list[1] = float(m_list[1]) + padding_top   # Adjust Y
                        new_custom_markers.append(m_list)
                    except (IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom marker during padding adjustment: {marker_data}")
                         new_custom_markers.append(marker_data) # Keep original if error
                self.custom_markers = new_custom_markers

                # Adjust Custom Shapes
                new_custom_shapes = []
                for shape_data_orig in getattr(self, "custom_shapes", []):
                    shape_data = dict(shape_data_orig) # Work on a copy
                    try:
                        shape_type = shape_data.get('type')
                        if shape_type == 'line':
                            sx, sy = shape_data['start']
                            ex, ey = shape_data['end']
                            shape_data['start'] = (float(sx) + padding_left, float(sy) + padding_top)
                            shape_data['end'] = (float(ex) + padding_left, float(ey) + padding_top)
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            shape_data['rect'] = (float(x) + padding_left, float(y) + padding_top, w, h)
                        # Add other shape types here if needed
                        new_custom_shapes.append(shape_data)
                    except (KeyError, IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom shape during padding adjustment: {shape_data_orig}")
                         new_custom_shapes.append(shape_data_orig)
                self.custom_shapes = new_custom_shapes

                # Adjust the absolute marker shift variables (these are in native image pixels)
                self.left_marker_shift_added += padding_left
                self.right_marker_shift_added += padding_left
                self.top_marker_shift_added += padding_top      
                self._update_overlay_slider_ranges
                
            
            def update_left_padding(self):
                # Update left padding when slider value changes
                self.left_marker_shift_added = self.left_padding_slider.value()
                self.update_live_view()

            def update_right_padding(self):
                # Update right padding when slider value changes
                new_value = self.right_padding_slider.value()
                # --- Add check to prevent redundant updates ---
                if new_value != self.right_marker_shift_added:
                    self.right_marker_shift_added = new_value
                    self.update_live_view()
                
            def update_top_padding(self):
                # Update top padding when slider value changes
                self.top_marker_shift_added = self.top_padding_slider.value()
                self.update_live_view()

            def update_live_view(self):
                # --- Initial check for a valid base image ---
                if hasattr(self, 'live_view_label') and hasattr(self, 'pan_left_action'): # Check if attributes exist
                    enable_pan_actions = self.live_view_label.zoom_level > 1.0
                    self.pan_left_action.setEnabled(enable_pan_actions)
                    self.pan_right_action.setEnabled(enable_pan_actions)
                    self.pan_up_action.setEnabled(enable_pan_actions)
                    self.pan_down_action.setEnabled(enable_pan_actions)
                if not self.image or self.image.isNull(): 
                    if hasattr(self, 'live_view_label'):
                        self.live_view_label.clear()
                        self.live_view_label.update() # Ensure paintEvent clears overlays if needed
                    # Disable UI elements that depend on an image
                    if hasattr(self, 'predict_button'): self.predict_button.setEnabled(False)
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(False)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(False)
                    # ... (potentially other UI elements) ...
                    return # Exit if no valid image
                else:
                    # Enable UI elements if image is valid
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(True)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(True)
                    if hasattr(self, 'predict_button'):
                        left_m = getattr(self, 'left_markers', [])
                        right_m = getattr(self, 'right_markers', [])
                        self.predict_button.setEnabled(bool(left_m or right_m))

                # --- Capture current pan and zoom state from LiveViewLabel ---
                # This ensures the entire render cycle uses a consistent state.
                current_zoom_level = 1.0
                current_pan_offset = QPointF(0, 0)
                if hasattr(self, 'live_view_label'): # Ensure live_view_label exists
                    current_zoom_level = self.live_view_label.zoom_level
                    current_pan_offset = QPointF(self.live_view_label.pan_offset) # Create a copy

                # --- Define rendering parameters ---
                render_scale = 3 # For high-resolution intermediate canvas
                try: 
                    view_width = self.live_view_label.width()
                    view_height = self.live_view_label.height()
                    if view_width <= 0: view_width = 600 # Fallback width
                    if view_height <= 0: view_height = 400 # Fallback height
                    render_width = view_width * render_scale
                    render_height = view_height * render_scale
                except AttributeError: # If live_view_label doesn't exist yet (early call)
                     render_width = 1800 # Default if view not ready
                     render_height = 1200
                except Exception: # Catch any other error during dimension access
                     render_width = 1800
                     render_height = 1200
                
                # --- Prepare image for transformations (Start with current self.image) ---
                if not self.image or self.image.isNull(): # Re-check, though earlier check should catch this
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                image_to_transform = self.image.copy() # Work on a copy

                # --- Apply Rotation (if UI elements exist) ---
                orientation = 0.0
                if hasattr(self, 'orientation_slider') and self.orientation_slider:
                    orientation = float(self.orientation_slider.value() / 20)
                    # Update the label text only if the label exists
                    if hasattr(self, 'orientation_label') and self.orientation_label:
                        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

                    if abs(orientation) > 0.01: 
                         if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                             transform_rotate = QTransform()
                             w_rot, h_rot = image_to_transform.width(), image_to_transform.height()
                             transform_rotate.translate(w_rot / 2.0, h_rot / 2.0) # Use float division
                             transform_rotate.rotate(orientation)
                             transform_rotate.translate(-w_rot / 2.0, -h_rot / 2.0)
                             temp_rotated = image_to_transform.transformed(transform_rotate, Qt.SmoothTransformation)
                             if not temp_rotated.isNull(): 
                                 image_to_transform = temp_rotated
                             else: 
                                 print("Warning: Rotation transform resulted in an invalid image.")
                
                # --- Apply Skew / Taper (if UI elements exist) ---
                taper_value = 0.0
                if hasattr(self, 'taper_skew_slider') and self.taper_skew_slider:
                    taper_value = self.taper_skew_slider.value() / 100.0
                    if hasattr(self, 'taper_skew_label') and self.taper_skew_label:
                        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

                    if abs(taper_value) > 0.01: 
                        if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                            try:
                                # Convert QImage to NumPy array for OpenCV processing
                                np_image = self.qimage_to_numpy(image_to_transform)
                                if np_image is None:
                                    raise ValueError("Failed to convert QImage to NumPy array.")

                                height, width = np_image.shape[:2]
                                
                                # Define source and destination points for perspective transform
                                source_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                                destination_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])

                                if taper_value > 0: # Narrower at top
                                    destination_np[0][0] = width * taper_value / 2.0
                                    destination_np[1][0] = width * (1 - taper_value / 2.0)
                                elif taper_value < 0: # Narrower at bottom
                                    destination_np[3][0] = width * (-taper_value / 2.0)
                                    destination_np[2][0] = width * (1 + taper_value / 2.0)
                                
                                # Get the perspective transformation matrix
                                matrix = cv2.getPerspectiveTransform(source_np, destination_np)
                                # Apply the transformation
                                skewed_np_image = cv2.warpPerspective(np_image, matrix, (width, height))
                                
                                # Convert the resulting NumPy array back to a QImage
                                temp_skewed_qimage = self.numpy_to_qimage(skewed_np_image)
                                if not temp_skewed_qimage.isNull():
                                    image_to_transform = temp_skewed_qimage
                                else:
                                    print("Warning: OpenCV skew preview failed to convert back to QImage.")
                            except Exception as e:
                                print(f"Error during OpenCV skew preview: {e}")

                # --- Final check on image_to_transform before proceeding ---
                if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    print("Error: image_to_transform became invalid before scaling for render.")
                    return

                # --- Scale the final transformed image for rendering on high-res canvas ---
                scaled_image_for_render_canvas = image_to_transform.scaled(
                    render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                if scaled_image_for_render_canvas.isNull():
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    print("Error: Final scaling for high-res render_canvas failed.")
                    return

                # --- Create high-resolution intermediate canvas ---
                canvas_format = QImage.Format_ARGB32_Premultiplied if image_to_transform.hasAlphaChannel() else QImage.Format_RGB888
                render_canvas = QImage(render_width, render_height, canvas_format)
                if render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to create high-res render_canvas QImage.")
                     return
                render_canvas.fill(Qt.white if canvas_format == QImage.Format_RGB888 else Qt.transparent)

                # Get crop offsets (these are 0 if image is not from a crop operation or crop was reset)
                current_crop_offset_x = getattr(self, 'crop_offset_x', 0)
                current_crop_offset_y = getattr(self, 'crop_offset_y', 0)

                # Render base image content (transformed image, raster overlays, guides) onto render_canvas
                # This function DOES NOT use pan/zoom itself.
                self.render_image_on_canvas(render_canvas, scaled_image_for_render_canvas,
                                            x_start=current_crop_offset_x, 
                                            y_start=current_crop_offset_y, 
                                            render_scale=render_scale,
                                            draw_guides=True) # Guides are part of the "view" rendering

                # --- Scale high-res canvas down for display on LiveViewLabel ---
                if render_canvas.isNull(): # Should not happen if previous checks passed
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                pixmap_from_render_canvas = QPixmap.fromImage(render_canvas)
                if pixmap_from_render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to create QPixmap from render_canvas.")
                     return

                # This is the pixmap representing the 100% zoom, unpanned view, scaled to fit the label
                scaled_pixmap_for_label_fit = pixmap_from_render_canvas.scaled(
                    self.live_view_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation # Use current label size
                )
                if scaled_pixmap_for_label_fit.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to scale final pixmap for label display.")
                     return

                # This will be the pixmap ultimately set on the label
                final_pixmap_to_set_on_label = scaled_pixmap_for_label_fit

                # --- Apply Zoom and Pan for display if zoom_level is not 1.0 ---
                # Uses the locally captured current_zoom_level and current_pan_offset
                if current_zoom_level != 1.0:
                    # Target pixmap for zoomed/panned view, same size as the label
                    zoomed_display_pixmap = QPixmap(self.live_view_label.size()) 
                    if zoomed_display_pixmap.isNull():
                        print("Error: Failed to create zoomed_display_pixmap.")
                        # Fallback to showing the unzoomed pixmap
                    else:
                        zoomed_display_pixmap.fill(Qt.transparent) 
                        zoom_painter = QPainter(zoomed_display_pixmap)
                        if not zoom_painter.isActive():
                            print("Error: Failed to create QPainter for zooming in update_live_view.")
                        else:
                            zoom_painter.translate(current_pan_offset) # Apply pan
                            zoom_painter.scale(current_zoom_level, current_zoom_level) # Apply zoom
                            # Draw the 100% view (scaled_pixmap_for_label_fit) onto the transformed painter
                            zoom_painter.drawPixmap(0, 0, scaled_pixmap_for_label_fit) 
                            zoom_painter.end()

                            if zoomed_display_pixmap.isNull(): # Check after drawing
                                print("Error: zoomed_display_pixmap became invalid after drawing.")
                            else:
                                 final_pixmap_to_set_on_label = zoomed_display_pixmap # Use the panned/zoomed version
                
                # --- Set the final pixmap on LiveViewLabel ---
                if final_pixmap_to_set_on_label.isNull():
                    self.live_view_label.clear()
                    print("Error: final_pixmap_to_set_on_label is invalid before setting.")
                else:
                     self.live_view_label.setPixmap(final_pixmap_to_set_on_label)

                # Trigger LiveViewLabel.paintEvent to draw vector overlays (markers, shapes, grid, previews)
                # on top of the newly set base pixmap.
                self.live_view_label.update()
                
            def _update_overlay_slider_ranges(self):
                """
                Updates the ranges of overlay position sliders.
                The ranges are based on the NATIVE dimensions of the current self.image.
                Slider values will represent native pixel offsets.
                """
                native_width = 1000  # Fallback
                native_height = 1000 # Fallback
        
                if self.image and not self.image.isNull() and self.image.width() > 0 and self.image.height() > 0:
                    native_width = self.image.width()
                    native_height = self.image.height()
        
                # Allow positioning overlay's top-left from -1*dimension to +2*dimension
                # relative to the main image's top-left in NATIVE pixels.
                x_range_min = -native_width
                x_range_max = native_width * 2
                y_range_min = -native_height
                y_range_max = native_height * 2
                
                
                sliders_to_update = [
                    (getattr(self, 'image1_left_slider', None), x_range_min, x_range_max, getattr(self, 'image1_position', (0,0))[0]),
                    (getattr(self, 'image1_top_slider', None), y_range_min, y_range_max, getattr(self, 'image1_position', (0,0))[1]),
                    (getattr(self, 'image2_left_slider', None), x_range_min, x_range_max, getattr(self, 'image2_position', (0,0))[0]),
                    (getattr(self, 'image2_top_slider', None), y_range_min, y_range_max, getattr(self, 'image2_position', (0,0))[1]),
                ]
        
                for slider, min_val, max_val, current_pos_val in sliders_to_update:
                    if slider:
                        slider.blockSignals(True)
                        slider.setRange(min_val, max_val)
                        # Try to set the slider to the current stored native position, clamped by new range
                        clamped_val = max(min_val, min(current_pos_val, max_val))
                        slider.setValue(clamped_val)
                        slider.blockSignals(False)
            
            def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
                painter = QPainter(canvas)
                # Calculate centering offsets for the scaled_image on the canvas
                x_offset = (canvas.width() - scaled_image.width()) // 2
                y_offset = (canvas.height() - scaled_image.height()) // 2
                
                self.x_offset_s=x_offset # Store for other potential uses if needed
                self.y_offset_s=y_offset
            
                # Draw the base image
                painter.drawImage(x_offset, y_offset, scaled_image)
                
                # --- NO CUSTOM SHAPES DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent

                # --- NO CUSTOM MARKERS DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent

                # --- NO STANDARD L/R/TOP MARKERS DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent
                
                # Native dimensions of the image that scaled_image was derived from
                # This assumes self.image is the source for scaled_image
                native_width = self.image.width() if self.image and self.image.width() > 0 else 1
                native_height = self.image.height() if self.image and self.image.height() > 0 else 1

                # Scale factor mapping native coordinates to the scaled_image dimensions
                scale_native_to_scaled_x = scaled_image.width() / native_width
                scale_native_to_scaled_y = scaled_image.height() / native_height

                # --- Draw Image 1 Overlay (Combined Image Feature) ---
                if hasattr(self, 'image1_original') and hasattr(self, 'image1_position') and not self.image1_original.isNull():
                    try:
                        scale_factor_overlay = self.image1_resize_slider.value() / 100.0
                        target_overlay_w = int(self.image1_original.width() * scale_factor_overlay)
                        target_overlay_h = int(self.image1_original.height() * scale_factor_overlay)
                        self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())

                        if target_overlay_w > 0 and target_overlay_h > 0:
                            # Resize the *original* overlay image
                            resized_overlay1 = self.image1_original.scaled(
                                target_overlay_w, target_overlay_h,
                                Qt.KeepAspectRatio, Qt.SmoothTransformation
                            )
                            if not resized_overlay1.isNull():
                                # Get NATIVE offsets stored
                                native_offset_x1 = self.image1_position[0]
                                native_offset_y1 = self.image1_position[1]

                                # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                                # Scale the NATIVE offset according to how the base image was scaled
                                # Then add the centering offset of the base image on the canvas
                                canvas_draw_x = x_offset + native_offset_x1 * scale_native_to_scaled_x
                                canvas_draw_y = y_offset + native_offset_y1 * scale_native_to_scaled_y

                                painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay1)
                    except Exception as e:
                        print(f"Error rendering overlay image 1: {e}")


                # --- Draw Image 2 Overlay ---
                if hasattr(self, 'image2_original') and hasattr(self, 'image2_position') and not self.image2_original.isNull():
                    try:
                        scale_factor_overlay = self.image2_resize_slider.value() / 100.0
                        self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                        target_overlay_w = int(self.image2_original.width() * scale_factor_overlay)
                        target_overlay_h = int(self.image2_original.height() * scale_factor_overlay)

                        if target_overlay_w > 0 and target_overlay_h > 0:
                            resized_overlay2 = self.image2_original.scaled(
                                target_overlay_w, target_overlay_h,
                                Qt.KeepAspectRatio, Qt.SmoothTransformation
                            )
                            if not resized_overlay2.isNull():
                                native_offset_x2 = self.image2_position[0]
                                native_offset_y2 = self.image2_position[1]

                                # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                                canvas_draw_x = x_offset + native_offset_x2 * scale_native_to_scaled_x
                                canvas_draw_y = y_offset + native_offset_y2 * scale_native_to_scaled_y

                                painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay2)
                    except Exception as e:
                        print(f"Error rendering overlay image 2: {e}")
                
            
            
                # Draw guide lines (These are view-dependent, not content, so okay here)
                if draw_guides and hasattr(self, 'show_guides_checkbox') and self.show_guides_checkbox.isChecked():
                    pen_guides = QPen(Qt.red, 2 * render_scale) # Keep guides scaled by render_scale for visibility on hi-res canvas
                    painter.setPen(pen_guides)
                    center_x_canvas = canvas.width() // 2
                    center_y_canvas = canvas.height() // 2
                    painter.drawLine(center_x_canvas, 0, center_x_canvas, canvas.height())
                    painter.drawLine(0, center_y_canvas, canvas.width(), center_y_canvas)
            
                painter.end()


             
            def crop_image(self):
                """Function to crop the current image."""
                if not self.image:
                    return None
            
                # Get crop percentage from sliders
                x_start_percent = 0
                x_end_percent = 100
                y_start_percent = 0
                y_end_percent = 100
            
                # Calculate crop boundaries
                x_start = int(self.image.width() * x_start_percent)
                x_end = int(self.image.width() * x_end_percent)
                y_start = int(self.image.height() * y_start_percent)
                y_end = int(self.image.height() * y_end_percent)
            
                # Ensure cropping is valid
                if x_start >= x_end or y_start >= y_end:
                    QMessageBox.warning(self, "Warning", "Invalid cropping values.")
                    return None
            
                # Crop the image
                cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
                return cropped_image
            
            # Modify align_image and update_crop to preserve settings
            def reset_align_image(self):
                self.orientation_slider.setValue(0)
                try:
                    if self.image: # Check if image exists after cropping
                        self._update_preview_label_size()
                except Exception as e:
                    # Fallback size?
                    self._update_preview_label_size()


                self.update_live_view() # Final update with corrected markers and image
                
                
            def align_image(self):
                """
                Aligns (rotates) the main self.image based on the orientation slider value.
                This modifies the high-resolution self.image directly.
                """
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Rotation Error", "No image loaded to rotate.")
                    return

                # Get the orientation value from the slider
                angle = float(self.orientation_slider.value() / 20.0) # Use float division

                if abs(angle) < 0.01: # Threshold to avoid unnecessary processing for tiny angles
                    # Reset slider if needed, but don't modify image
                    self.orientation_slider.setValue(0)
                    return

                # --- Save state BEFORE modifying the image ---
                self.save_state()

                # Disable temporary guides before applying permanent change
                self.draw_guides = False
                if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)
                # Keep grid settings as they are independent of rotation


                try:
                    # Perform rotation calculation
                    transform = QTransform()
                    # Rotate around the center of the *current* image
                    current_width = self.image.width()
                    current_height = self.image.height()
                    transform.translate(current_width / 2.0, current_height / 2.0)
                    transform.rotate(angle)
                    transform.translate(-current_width / 2.0, -current_height / 2.0)

                    # Apply the transformation to get the high-resolution rotated image
                    # Qt.SmoothTransformation provides better quality
                    # The resulting image might be larger to accommodate rotated corners
                    rotated_image_high_res = self.image.transformed(transform, Qt.SmoothTransformation)

                    if rotated_image_high_res.isNull():
                        raise ValueError("Image transformation resulted in an invalid (null) image.")

                    # --- Directly update the main image object ---
                    self.image = rotated_image_high_res
                    self.is_modified = True # Mark as modified

                    # --- Update backup images to reflect the new state ---
                    # Rotation invalidates previous padding
                    self.image_before_padding = None # Or self.image.copy() if padding should persist conceptually
                    self.image_padded = False
                    # Contrast backups should also reflect the rotated image
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()

                    # --- Reset the orientation slider ---
                    self.orientation_slider.setValue(0)

                    # --- Update UI elements based on the new image dimensions ---
                    # This will adjust the preview label size and marker slider ranges
                    self._update_preview_label_size()
                    self._update_marker_slider_ranges() # Important after potential size change
                    self._update_status_bar()

                    # --- Let update_live_view handle the rendering ---
                    # This will correctly scale the new high-res self.image for display
                    # and draw markers/overlays relative to the new image.
                    self.update_live_view()

                except Exception as e:
                    QMessageBox.critical(self, "Rotation Error", f"Failed to rotate image: {e}")
                    traceback.print_exc()
                    # Attempt to revert state if possible (though save_state was called)
                    # self.undo_action_m() # Reverting might be complex if backups were already updated
                
                self.crop_x_start_slider.setEnabled(False) 
                self.crop_x_end_slider.setEnabled(False) 
                self.crop_y_start_slider.setEnabled(False) 
                self.crop_y_end_slider.setEnabled(False) 
            
            def _update_marker_slider_ranges(self):
                if not self.image or self.image.isNull() or not self.live_view_label:
                    min_abs, max_abs_w, max_abs_h = -100, 1000, 800 # Default absolute pixel values
                else:
                    try:
                        # Ranges are based on the NATIVE dimensions of the CURRENT self.image
                        native_img_width = self.image.width()
                        native_img_height = self.image.height()

                        margin = 100 # Allow markers to be set slightly outside the image bounds

                        min_abs_x = -margin
                        max_abs_x = native_img_width + margin
                        min_abs_y = -margin
                        max_abs_y = native_img_height + margin

                        # Fallbacks if dimensions are invalid
                        if native_img_width <= 0: max_abs_x = 1000; min_abs_x = -100
                        if native_img_height <= 0: max_abs_y = 800; min_abs_y = -100
                        
                        max_abs_w = max_abs_x # For horizontal sliders (left/right markers)
                        max_abs_h = max_abs_y # For Y-offsets (top markers)
                        min_abs = min(min_abs_x, min_abs_y) # Common minimum

                    except Exception as e:
                        print(f"Warning: Error calculating slider ranges: {e}. Using defaults.")
                        min_abs, max_abs_w, max_abs_h = -100, 1000, 800

                # Store the calculated ranges internally (optional, but good for reference)
                self.left_slider_range = [min_abs, max_abs_w]
                self.right_slider_range = [min_abs, max_abs_w]
                self.top_slider_range = [min_abs, max_abs_h]

                # Update sliders: setRange, then re-apply current value (which might get clamped)
                sliders_and_ranges = [
                    (getattr(self, 'left_padding_slider', None), self.left_slider_range),
                    (getattr(self, 'right_padding_slider', None), self.right_slider_range),
                    (getattr(self, 'top_padding_slider', None), self.top_slider_range)
                ]

                for slider, new_range in sliders_and_ranges:
                    if slider:
                        current_val = slider.value() # Get current value BEFORE changing range
                        slider.blockSignals(True)
                        slider.setRange(new_range[0], new_range[1])
                        slider.setValue(current_val) # Re-apply; Qt will clamp if current_val is outside new_range
                        slider.blockSignals(False)
                    
                
                    
            def update_crop(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Crop Error", "No image loaded to crop.")
                    return

                if not self.crop_rectangle_coords:
                    # ... (message box logic) ...
                    return

                try:
                    img_x_intent, img_y_intent, img_w_intent, img_h_intent = self.crop_rectangle_coords
                    original_image_width_before_crop = self.image.width()
                    original_image_height_before_crop = self.image.height()

                    crop_x_start = max(0, int(round(img_x_intent)))
                    crop_y_start = max(0, int(round(img_y_intent)))
                    crop_width = max(1, min(int(round(img_w_intent)), original_image_width_before_crop - crop_x_start))
                    crop_height = max(1, min(int(round(img_h_intent)), original_image_height_before_crop - crop_y_start))
                                        
                    if crop_width <= 0 or crop_height <= 0:
                         QMessageBox.warning(self, "Crop Error", "Calculated crop area has invalid dimensions. Aborting.")
                         self.crop_rectangle_coords = None; self.live_view_label.clear_crop_preview(); self.cancel_rectangle_crop_mode() 
                         return

                    self.save_state()

                    # --- 1. Adjust marker and shape coordinates ---
                    # (This part from your previous good version, ensuring elements are kept if their
                    #  anchor/significant part is within the new bounds, and coords are made relative)
                    new_left_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'left_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.left_markers = new_left_markers
                    new_right_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'right_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.right_markers = new_right_markers
                    new_top_markers = [(x_old - crop_x_start, label) for x_old, label in getattr(self, 'top_markers', []) if crop_x_start <= x_old < crop_x_start + crop_width]
                    self.top_markers = new_top_markers
                    
                    new_custom_markers = [] # ... (full adjustment logic as before) ...
                    if hasattr(self, "custom_markers"):
                        for marker_data in self.custom_markers:
                            try:
                                m_list = list(marker_data); x_old, y_old = float(m_list[0]), float(m_list[1])
                                x_new, y_new = x_old - crop_x_start, y_old - crop_y_start
                                if 0 <= x_new < crop_width and 0 <= y_new < crop_height:
                                    m_list[0] = x_new; m_list[1] = y_new
                                    new_custom_markers.append(m_list)
                            except: pass # Skip malformed
                    self.custom_markers = new_custom_markers

                    new_custom_shapes = [] # ... (full adjustment logic as before, including clipping for rects) ...
                    if hasattr(self, "custom_shapes"):
                        for shape_data_orig in self.custom_shapes:
                            shape_data = dict(shape_data_orig); adjusted_shape_data = None
                            try:
                                stype = shape_data.get('type')
                                if stype == 'line':
                                    sx_old, sy_old = map(float, shape_data['start']); ex_old, ey_old = map(float, shape_data['end'])
                                    adjusted_shape_data = shape_data.copy()
                                    adjusted_shape_data['start'] = (sx_old - crop_x_start, sy_old - crop_y_start)
                                    adjusted_shape_data['end'] = (ex_old - crop_x_start, ey_old - crop_y_start)
                                elif stype == 'rectangle':
                                    rx_old, ry_old, rw_old, rh_old = map(float, shape_data['rect'])
                                    overlap_x1 = max(crop_x_start, rx_old); overlap_y1 = max(crop_y_start, ry_old)
                                    overlap_x2 = min(crop_x_start + crop_width, rx_old + rw_old)
                                    overlap_y2 = min(crop_y_start + crop_height, ry_old + rh_old)
                                    if overlap_x2 > overlap_x1 and overlap_y2 > overlap_y1:
                                        adjusted_shape_data = shape_data.copy()
                                        adjusted_shape_data['rect'] = (overlap_x1 - crop_x_start, overlap_y1 - crop_y_start, overlap_x2 - overlap_x1, overlap_y2 - overlap_y1)
                                if adjusted_shape_data: new_custom_shapes.append(adjusted_shape_data)
                            except: pass # Skip malformed
                    self.custom_shapes = new_custom_shapes
                    
                    # --- 2. Adjust the _marker_shift_added variables ---
                    # These are absolute X/Y positions in the *pre-crop* image.
                    # Make them relative to the new cropped image's origin.
                    self.left_marker_shift_added = self.left_marker_shift_added - crop_x_start
                    self.right_marker_shift_added = self.right_marker_shift_added - crop_x_start
                    self.top_marker_shift_added = self.top_marker_shift_added - crop_y_start

                    # Clamp them to be valid within the new cropped image dimensions [0, new_dim -1]
                    self.left_marker_shift_added = max(0, min(self.left_marker_shift_added, crop_width - 1 if crop_width > 0 else 0))
                    self.right_marker_shift_added = max(0, min(self.right_marker_shift_added, crop_width - 1 if crop_width > 0 else 0))
                    self.top_marker_shift_added = max(0, min(self.top_marker_shift_added, crop_height - 1 if crop_height > 0 else 0))

                    # --- 3. Perform the actual crop on self.image ---
                    cropped_qimage = self.image.copy(crop_x_start, crop_y_start, crop_width, crop_height)
                    if cropped_qimage.isNull():
                        raise ValueError("QImage.copy failed for cropping.")

                    self.image = cropped_qimage # self.image is now the cropped image
                    self.is_modified = True
                    self.image_before_padding = self.image.copy() 
                    self.image_contrasted = self.image.copy()      
                    self.image_before_contrast = self.image.copy() 
                    self.image_padded = False 

                    self.crop_rectangle_coords = None 
                    self.live_view_label.clear_crop_preview() 
                    self.cancel_rectangle_crop_mode() 

                    # --- 4. Update UI ---
                    self._update_preview_label_size() 
                    self._update_status_bar()         
                    
                    # Update slider RANGES based on new image dimensions.
                    self._update_marker_slider_ranges() 
                    
                    # NOW, set the slider values to the *transformed and clamped* _marker_shift_added values.
                    # This ensures the UI accurately reflects the state.
                    sliders_to_set = [
                        (getattr(self, 'left_padding_slider', None), self.left_marker_shift_added),
                        (getattr(self, 'right_padding_slider', None), self.right_marker_shift_added),
                        (getattr(self, 'top_padding_slider', None), self.top_marker_shift_added)
                    ]
                    for slider, value_to_set in sliders_to_set:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(value_to_set) # This will be clamped by the new range if needed
                            slider.blockSignals(False)
                            # Re-sync internal var with actual slider value after potential clamping by setValue
                            if slider == self.left_padding_slider: self.left_marker_shift_added = slider.value()
                            elif slider == self.right_padding_slider: self.right_marker_shift_added = slider.value()
                            elif slider == self.top_padding_slider: self.top_marker_shift_added = slider.value()
                    
                    self.update_live_view()

                except Exception as e:
                    # ... (error handling) ...
                    QMessageBox.critical(self, "Crop Error", f"An error occurred during cropping: {e}")
                    traceback.print_exc()
                    if self.undo_stack: 
                        try: self.undo_action_m() 
                        except: pass 
                    self.crop_rectangle_coords = None
                    self.live_view_label.clear_crop_preview()
                    self.cancel_rectangle_crop_mode()
                
            def update_skew(self):

                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Skew Error", "No master image loaded to apply skew.")
                    return

                self.save_state()
                taper_value = self.taper_skew_slider.value() / 100.0

                try:
                    source_image = self.image.copy()
                    
                    # Convert QImage to NumPy for OpenCV processing
                    np_image = self.qimage_to_numpy(source_image)
                    if np_image is None:
                        raise ValueError("Failed to convert master image to NumPy array.")

                    height, width = np_image.shape[:2]
                
                    # Define corner points for perspective transformation
                    source_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                    destination_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                
                    # Adjust perspective based on taper value
                    if taper_value > 0: # Narrower at the top
                        destination_np[0][0] = width * taper_value / 2.0
                        destination_np[1][0] = width * (1 - taper_value / 2.0)
                    elif taper_value < 0: # Narrower at the bottom
                        destination_np[3][0] = width * (-taper_value / 2.0)
                        destination_np[2][0] = width * (1 + taper_value / 2.0)
                
                    # Get the perspective transformation matrix and apply it
                    matrix = cv2.getPerspectiveTransform(source_np, destination_np)
                    skewed_np_image = cv2.warpPerspective(np_image, matrix, (width, height))
                    
                    # Convert back to QImage
                    skewed_image = self.numpy_to_qimage(skewed_np_image)
                    if skewed_image.isNull():
                        raise ValueError("Skew transformation resulted in an invalid image after conversion.")

                    # Update all relevant image states
                    self.image = skewed_image
                    self.image_before_contrast = self.image.copy()
                    self.image_contrasted = self.image.copy()
                    self.image_before_padding = None
                    self.image_padded = False
                    self.is_modified = True

                    self.taper_skew_slider.setValue(0)
                    
                    self._update_preview_label_size()
                    self._update_marker_slider_ranges()
                    self._update_status_bar()
                    self.update_live_view()

                except Exception as e:
                    QMessageBox.critical(self, "Skew Error", f"Failed to apply skew: {e}")
                    traceback.print_exc()

            
                
            def save_image(self):
                self.draw_guides = False
                if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)

                if not self.image or self.image.isNull():
                     QMessageBox.warning(self, "Error", "No image data to save.")
                     return False

                # (File dialog logic remains the same...)
                options = QFileDialog.Options()
                suggested_name = os.path.splitext(os.path.basename(self.image_path))[0] if self.image_path else "untitled_image"
                base_name_clean = suggested_name.replace("_original", "").replace("_modified", "")
                save_dir = os.path.dirname(self.image_path) if self.image_path else ""
                base_save_path, selected_filter = QFileDialog.getSaveFileName(
                    self, "Save Image Base Name", os.path.join(save_dir, base_name_clean),
                    "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
                    options=options
                )
                if not base_save_path: return False
                base_name_nosuffix = os.path.splitext(base_save_path)[0]
                suffix = os.path.splitext(base_save_path)[1].lower() or ".png"
                if "tif" in selected_filter.lower(): suffix = ".tif"
                elif "jpg" in selected_filter.lower(): suffix = ".jpg"
                elif "bmp" in selected_filter.lower(): suffix = ".bmp"
                original_save_path = f"{base_name_nosuffix}_original{suffix}"
                modified_save_path = f"{base_name_nosuffix}_modified{suffix}"
                config_save_path = f"{base_name_nosuffix}_config.txt"

                # --- Save _original image (current state of self.image) ---
                if self.image and not self.image.isNull():
                    save_format = suffix.replace(".", "").upper()
                    if save_format == "TIF": save_format = "TIFF"
                    quality = 95 if save_format in ["JPG", "JPEG"] else -1
                    if not self.image.save(original_save_path, format=save_format if save_format else None, quality=quality):
                        QMessageBox.warning(self, "Error", f"Failed to save original image.")

                # --- Create and save _modified image with annotations (SHARP VERSION) ---
                render_scale = 3
                native_width = self.image.width(); native_height = self.image.height()
                if native_width <= 0 or native_height <= 0: return False

                canvas_width = native_width * render_scale; canvas_height = native_height * render_scale
                modified_canvas = QImage(canvas_width, canvas_height, QImage.Format_ARGB32_Premultiplied)
                modified_canvas.fill(Qt.transparent)
                
                painter = QPainter(modified_canvas)
                
                # ** THE FIX for SHARPNESS **
                painter.setRenderHint(QPainter.SmoothPixmapTransform, False)
                painter.drawImage(QRectF(0.0, 0.0, float(canvas_width), float(canvas_height)), self.image, QRectF(self.image.rect()))

                # ** THE FIX for TEXT SCALING and SHARPNESS (Part 2) **
                label_width = float(self.live_view_label.width()); label_height = float(self.live_view_label.height())
                scale_native_to_view = min(label_width / native_width, label_height / native_height) if label_width > 0 and label_height > 0 else 1.0
                font_scale_factor = render_scale / scale_native_to_view if scale_native_to_view > 1e-6 else render_scale
                
                painter.setRenderHint(QPainter.Antialiasing, True); painter.setRenderHint(QPainter.TextAntialiasing, True)
                
                # --- Draw annotations (logic is identical to the corrected copy_to_clipboard) ---
                def map_img_coords_to_canvas(img_x, img_y):
                    return QPointF(img_x * render_scale, img_y * render_scale)

                # A. Standard L/R/Top Markers
                std_font = QFont(self.font_family, int(self.font_size * font_scale_factor))
                painter.setFont(std_font); painter.setPen(self.font_color)
                fm_std = QFontMetrics(std_font); y_offset_baseline = fm_std.height() * 0.3
                for y_img, text in self.left_markers:
                    anchor_x = self.left_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    full_text = f"{text} ⎯"; painter.drawText(QPointF(anchor_x - fm_std.horizontalAdvance(full_text), anchor_y + y_offset_baseline), full_text)
                for y_img, text in self.right_markers:
                    anchor_x = self.right_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    painter.drawText(QPointF(anchor_x, anchor_y + y_offset_baseline), f"⎯ {text}")
                for x_img, text in self.top_markers:
                    painter.save(); anchor_x = x_img * render_scale; anchor_y = self.top_marker_shift_added * render_scale
                    painter.translate(anchor_x, anchor_y + y_offset_baseline); painter.rotate(self.font_rotation)
                    painter.drawText(QPointF(0, 0), str(text)); painter.restore()
                # B. Custom Markers & C. Custom Shapes
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        x, y, text, color, font, size, is_bold, is_italic = marker_data
                        custom_font = QFont(font, int(size * font_scale_factor)); custom_font.setBold(is_bold); custom_font.setItalic(is_italic)
                        painter.setFont(custom_font); painter.setPen(QColor(color))
                        fm = QFontMetrics(custom_font); rect = fm.boundingRect(text)
                        draw_pos = QPointF(x * render_scale - rect.center().x(), y * render_scale - rect.center().y())
                        painter.drawText(draw_pos, text)
                    except Exception: pass
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type, color_str, thickness = shape_data.get('type'), shape_data.get('color'), shape_data.get('thickness')
                        pen = QPen(QColor(color_str), max(1.0, thickness * render_scale)); painter.setPen(pen)
                        if shape_type == 'line':
                            painter.drawLine(map_img_coords_to_canvas(*shape_data['start']), map_img_coords_to_canvas(*shape_data['end']))
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            painter.drawRect(QRectF(map_img_coords_to_canvas(x, y), QSizeF(w * render_scale, h * render_scale)))
                    except Exception: pass
                
                painter.end()

                # Save the final modified canvas
                save_format = suffix.replace(".", "").upper()
                if save_format == "TIF": save_format = "TIFF"
                if not modified_canvas.save(modified_save_path, format=save_format if save_format else None, quality=-1):
                    QMessageBox.warning(self, "Error", f"Failed to save modified image.")
                    return False
                
                # --- Save Config File (Unchanged) ---
                config_data = self.get_current_config()
                try:
                    with open(config_save_path, "w", encoding='utf-8') as config_file:
                        json.dump(config_data, config_file, indent=4)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
                    return False

                self.is_modified = False
                QMessageBox.information(self, "Saved", f"Files saved successfully to '{os.path.dirname(base_save_path)}'")
                self.setWindowTitle(f"{self.window_title}::{base_name_clean}")
                return True
                    
            def save_image_svg(self):
                """Save the processed image along with markers, labels, and custom shapes in SVG format."""
                if not self.image or self.image.isNull(): # Check current image validity
                    QMessageBox.warning(self, "Warning", "No image to save.")
                    return

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Image as SVG for MS Word/Vector Editing", "", "SVG Files (*.svg)", options=options
                )

                if not file_path:
                    return

                if not file_path.lower().endswith(".svg"): # Ensure .svg extension
                    file_path += ".svg"

                # --- Determine Render/Canvas Dimensions ---
                # Use the current image dimensions directly as the base SVG size
                # This avoids scaling issues if the view label size is different.
                # Markers and shapes will be positioned relative to this size.
                svg_width = self.image.width()
                svg_height = self.image.height()

                if svg_width <= 0 or svg_height <= 0:
                     QMessageBox.warning(self, "Warning", "Invalid image dimensions for SVG.")
                     return

                # --- Create SVG Drawing Object ---
                dwg = svgwrite.Drawing(file_path, profile='tiny', size=(f"{svg_width}px", f"{svg_height}px"))
                # Set viewbox to match image dimensions for correct coordinate system
                dwg.viewbox(0, 0, svg_width, svg_height)

                # --- Embed the Base Image ---
                try:
                    # Convert the QImage to a base64-encoded PNG for embedding
                    buffer = QBuffer()
                    buffer.open(QBuffer.ReadWrite)
                    # Save the *current* self.image (which might be cropped/transformed)
                    if not self.image.save(buffer, "PNG"):
                        raise IOError("Failed to save image to PNG buffer for SVG.")
                    image_data = base64.b64encode(buffer.data()).decode('utf-8')
                    buffer.close()

                    # Embed the image at position (0, 0) with original dimensions
                    dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}",
                                      insert=(0, 0),
                                      size=(f"{svg_width}px", f"{svg_height}px")))
                except Exception as e:
                    QMessageBox.critical(self, "SVG Error", f"Failed to embed base image: {e}")
                    return # Stop if base image fails

                # --- Define marker/label font style for SVG ---
                # Use the standard marker font settings
                svg_font_family = self.font_family
                svg_font_size_px = f"{self.font_size}px" # Use pixels for SVG consistency
                svg_font_color = self.font_color.name() if self.font_color else "#000000"

                # Calculate horizontal offset for left/right markers based on font size
                # Use QFontMetrics for accurate width calculation
                try:
                     qfont_for_metrics = QFont(svg_font_family, self.font_size)
                     font_metrics = QFontMetrics(qfont_for_metrics)
                     # Use a representative character like 'm' or average width if needed
                     avg_char_width = font_metrics.averageCharWidth()
                     horizontal_offset = avg_char_width * 0.5 # Small offset from edge
                     vertical_offset_adjust = font_metrics.ascent() * 0.75 # Adjustment for vertical alignment
                except Exception:
                     horizontal_offset = 5 # Fallback offset
                     vertical_offset_adjust = self.font_size * 0.75 # Fallback adjustment


                # --- Add Left Markers ---
                left_marker_x_pos = self.left_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "left_markers", []):
                    final_text = f"{text}" # Remove the line "⎯" for cleaner SVG text
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(left_marker_x_pos - horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="end" # Align text right, ending at the x position
                        )
                    )

                # --- Add Right Markers ---
                right_marker_x_pos = self.right_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "right_markers", []):
                    final_text = f"{text}" # Remove the line "⎯"
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(right_marker_x_pos + horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="start" # Align text left, starting at the x position
                        )
                    )

                # --- Add Top Markers ---
                top_marker_y_pos = self.top_marker_shift_added # Use the absolute offset
                for x_pos, text in getattr(self, "top_markers", []):
                    dwg.add(
                        dwg.text(
                            text,
                            insert=(x_pos, top_marker_y_pos + vertical_offset_adjust), # Apply vertical adjust
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="middle", # Center text horizontally over the x position
                            # Apply rotation around the insertion point
                            transform=f"rotate({self.font_rotation}, {x_pos}, {top_marker_y_pos + vertical_offset_adjust})"
                        )
                    )

                # --- Add Custom Markers ---
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        # Default values for optional elements
                        is_bold = False
                        is_italic = False

                        # Unpack based on length for backward compatibility
                        if len(marker_tuple) == 8:
                            x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple
                        elif len(marker_tuple) == 6:
                            x_pos, y_pos, marker_text, color, font_family, font_size = marker_tuple
                        else:
                            continue # Skip invalid marker data

                        # Prepare SVG attributes
                        text_content = str(marker_text)
                        fill_color = QColor(color).name() if isinstance(color, QColor) else str(color) # Ensure hex/name
                        font_family_svg = str(font_family)
                        font_size_svg = f"{int(font_size)}px" if isinstance(font_size, (int, float)) else "12px"
                        font_weight_svg = "bold" if bool(is_bold) else "normal"
                        font_style_svg = "italic" if bool(is_italic) else "normal"

                        # Adjust vertical position slightly for better alignment if needed
                        # This might require font metrics specific to the marker's font
                        # For simplicity, using the standard marker offset for now
                        y_pos_adjusted = y_pos + vertical_offset_adjust

                        # Add SVG text element, centered at the marker's coordinates
                        dwg.add(
                            dwg.text(
                                text_content,
                                insert=(x_pos, y_pos_adjusted), # Position text anchor at the coordinate
                                fill=fill_color,
                                font_family=font_family_svg,
                                font_size=font_size_svg,
                                font_weight=font_weight_svg,
                                font_style=font_style_svg,
                                text_anchor="middle", # Center horizontally
                                dominant_baseline="central" # Attempt vertical centering (support varies)
                            )
                        )
                    except Exception as e:
                         print(f"Warning: Skipping invalid custom marker during SVG export: {e}")
                         # import traceback; traceback.print_exc() # Uncomment for detailed debug

                # --- START: Add Custom Shapes (Lines and Rectangles) ---
                for shape_data in getattr(self, "custom_shapes", []):
                     try:
                         shape_type = shape_data.get('type')
                         color_str = shape_data.get('color', '#000000') # Default black
                         thickness = int(shape_data.get('thickness', 1))
                         if thickness < 1: thickness = 1

                         if shape_type == 'line':
                             start_coords = shape_data.get('start')
                             end_coords = shape_data.get('end')
                             if start_coords and end_coords:
                                 dwg.add(dwg.line(start=start_coords,
                                                  end=end_coords,
                                                  stroke=color_str,
                                                  stroke_width=thickness))
                         elif shape_type == 'rectangle':
                             rect_coords = shape_data.get('rect') # (x, y, w, h)
                             if rect_coords:
                                 x, y, w, h = rect_coords
                                 dwg.add(dwg.rect(insert=(x, y),
                                                  size=(f"{w}px", f"{h}px"),
                                                  stroke=color_str,
                                                  stroke_width=thickness,
                                                  fill="none")) # No fill for outline rectangle
                     except Exception as e:
                         print(f"Warning: Skipping invalid custom shape during SVG export: {e}")
                # --- END: Add Custom Shapes ---

                # --- Save the SVG file ---
                try:
                    dwg.save()
                    QMessageBox.information(self, "Success", f"Image and annotations saved as SVG at\n{file_path}")
                    self.is_modified = False # Mark as saved if successful
                except Exception as e:
                    QMessageBox.critical(self, "SVG Save Error", f"Failed to save SVG file:\n{e}")
            
            
            def copy_to_clipboard(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Warning", "No image to copy.")
                    return

                # --- 1. Setup High-Resolution Canvas Based on Native Image Size ---
                render_scale = 3
                native_width = self.image.width()
                native_height = self.image.height()
                if native_width <= 0 or native_height <= 0: return

                canvas_width = native_width * render_scale
                canvas_height = native_height * render_scale

                render_canvas = QImage(canvas_width, canvas_height, QImage.Format_ARGB32_Premultiplied)
                render_canvas.fill(Qt.transparent)
                
                painter = QPainter(render_canvas)
                if not painter.isActive(): return

                # --- 2. THE FIX for SHARPNESS: Draw the base image without smooth scaling ---
                painter.setRenderHint(QPainter.SmoothPixmapTransform, False)
                # Use QRectF for precision
                painter.drawImage(QRectF(0.0, 0.0, float(canvas_width), float(canvas_height)), self.image, QRectF(self.image.rect()))

                # --- 3. THE FIX for TEXT SCALING: Calculate the correct font scale factor ---
                # This factor compensates for the image being scaled down in the live view.
                label_width = float(self.live_view_label.width())
                label_height = float(self.live_view_label.height())
                scale_native_to_view = 1.0
                if label_width > 0 and label_height > 0:
                    scale_native_to_view = min(label_width / native_width, label_height / native_height)
                
                # Prevent division by zero if scale is tiny
                font_scale_factor = render_scale / scale_native_to_view if scale_native_to_view > 1e-6 else render_scale

                # --- 4. Re-enable anti-aliasing for drawing smooth vector annotations ---
                painter.setRenderHint(QPainter.Antialiasing, True)
                painter.setRenderHint(QPainter.TextAntialiasing, True)

                # --- 5. Draw all annotations using the new font_scale_factor ---
                def map_img_coords_to_canvas(img_x, img_y):
                    return QPointF(img_x * render_scale, img_y * render_scale)

                # A. Standard L/R/Top Markers
                std_font = QFont(self.font_family, int(self.font_size * font_scale_factor))
                painter.setFont(std_font); painter.setPen(self.font_color)
                fm_std = QFontMetrics(std_font); y_offset_baseline = fm_std.height() * 0.3
                for y_img, text in self.left_markers:
                    anchor_x = self.left_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    full_text = f"{text} ⎯"; painter.drawText(QPointF(anchor_x - fm_std.horizontalAdvance(full_text), anchor_y + y_offset_baseline), full_text)
                for y_img, text in self.right_markers:
                    anchor_x = self.right_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    painter.drawText(QPointF(anchor_x, anchor_y + y_offset_baseline), f"⎯ {text}")
                for x_img, text in self.top_markers:
                    painter.save(); anchor_x = x_img * render_scale; anchor_y = self.top_marker_shift_added * render_scale
                    painter.translate(anchor_x, anchor_y + y_offset_baseline); painter.rotate(self.font_rotation)
                    painter.drawText(QPointF(0, 0), str(text)); painter.restore()

                # B. Custom Markers
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        x, y, text, color, font, size, is_bold, is_italic = marker_data
                        custom_font = QFont(font, int(size * font_scale_factor)); custom_font.setBold(is_bold); custom_font.setItalic(is_italic)
                        painter.setFont(custom_font); painter.setPen(QColor(color))
                        fm = QFontMetrics(custom_font); rect = fm.boundingRect(text)
                        draw_pos = QPointF(x * render_scale - rect.center().x(), y * render_scale - rect.center().y())
                        painter.drawText(draw_pos, text)
                    except Exception: pass

                # C. Custom Shapes
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type, color_str, thickness = shape_data.get('type'), shape_data.get('color'), shape_data.get('thickness')
                        pen = QPen(QColor(color_str), max(1.0, thickness * render_scale)); painter.setPen(pen)
                        if shape_type == 'line':
                            painter.drawLine(map_img_coords_to_canvas(*shape_data['start']), map_img_coords_to_canvas(*shape_data['end']))
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            painter.drawRect(QRectF(map_img_coords_to_canvas(x, y), QSizeF(w * render_scale, h * render_scale)))
                    except Exception: pass
                
                painter.end()
                
                # --- 6. Set clipboard data ---
                self.cleanup_temp_clipboard_file()
                mime_data = QMimeData()
                mime_data.setImageData(render_canvas)
                # (Optional: Add other formats like PNG buffer or file URL as before)
                QApplication.clipboard().setMimeData(mime_data)
                QMessageBox.information(self, "Copied", "High-resolution rendered image copied to clipboard.")


            def cleanup_temp_clipboard_file(self):
                """Deletes the temporary clipboard file if it exists."""
                path_to_delete = getattr(self, 'temp_clipboard_file_path', None)
                if path_to_delete:
                    # print(f"INFO: Attempting to clean up temp file: {path_to_delete}")
                    if os.path.exists(path_to_delete):
                        try:
                            os.remove(path_to_delete)
                            # print(f"INFO: Cleaned up temp file: {path_to_delete}")
                        except PermissionError: # Specifically catch permission errors
                            # This can happen if another app still has the file locked.
                            # Schedule for later deletion if possible, or just log it.
                            print(f"WARNING: PermissionError deleting {path_to_delete}. Will try again if app quits cleanly.")
                            # For a GUI app, trying again on quit is the best we can do.
                            # If this is called ON quit, then there's not much more to do.
                        except OSError as e:
                            print(f"WARNING: Could not delete temp clipboard file {path_to_delete}: {e}")
                    self.temp_clipboard_file_path = None # Clear the path regardless of success
                
                
            def clear_predict_molecular_weight(self):
                # ... (existing clear logic for MW prediction, single bounding boxes, etc.) ...
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                self.live_view_label.setCursor(Qt.ArrowCursor)
                if hasattr(self, "protein_location"):
                    del self.protein_location 
                self.predict_size=False
                self.bounding_boxes=[]
                self.bounding_box_start = None
                # self.live_view_label.bounding_box_start = None # This is not an attribute of LiveViewLabel
                # self.live_view_label.bounding_box_preview = None # Will be cleared below
                self.quantities=[]
                self.peak_area_list=[]
                self.protein_quantities=[]
                self.standard_protein_values.setText("")
                self.standard_protein_areas=[]
                self.standard_protein_areas_text.setText("")
                # self.live_view_label.quad_points=[] # Will be cleared below
                # self.live_view_label.bounding_box_preview = None # Will be cleared below
                # self.live_view_label.rectangle_points = [] # This is likely for single rect, also clear
                self.latest_calculated_quantities = []
                self.quantities_peak_area_dict={}
                
                # --- MODIFIED/ADDED: Clear multi-lane and LiveViewLabel preview states ---
                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None
                self.multi_lane_definitions = []
                self.current_multi_lane_points = []      # Clears app's buffer for current multi-lane quad
                self.current_multi_lane_rect_start = None # Clears app's buffer for current multi-lane rect
                self.latest_multi_lane_peak_areas = {}
                self.latest_multi_lane_calculated_quantities = {}
                self.latest_multi_lane_peak_details = {} # Also clear this
                self.target_protein_areas_text.clear()
                if hasattr(self, 'btn_finish_multi_lane_def'):
                    self.btn_finish_multi_lane_def.setEnabled(False)
                self.multi_lane_processing_finished = False

                # Explicitly clear LiveViewLabel's general-purpose preview buffers
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_points = [] # Also clear this, it's used for single rect
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self._reset_live_view_label_custom_handlers()
                self.update_live_view()
                
            def predict_molecular_weight(self):
                # ... (existing initial checks for markers) ...
                
                self.live_view_label.preview_marker_enabled = False # Disable custom marker preview
                self.live_view_label.mw_predict_preview_enabled = True # Enable MW preview
                self.live_view_label.mw_predict_preview_position = None # Clear old position
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True) # Ensure it's on
    
                
                # Determine which markers to use (left or right)
                markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
                if not markers_raw_tuples:
                    QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
                    self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                    return

                # --- Crucial Step: Sort markers by Y-position (migration distance) ---
                # This ensures the order reflects the actual separation on the gel/blot
                try:
                    # Ensure marker values are numeric for sorting and processing
                    numeric_markers = []
                    for pos, val_str in markers_raw_tuples:
                        try:
                            numeric_markers.append((float(pos), float(val_str)))
                        except (ValueError, TypeError):
                            # Skip markers with non-numeric values for prediction
                            continue

                    if len(numeric_markers) < 2:
                         QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                         self.live_view_label.setCursor(Qt.ArrowCursor)
                         return

                    sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
                    # Separate sorted positions and values
                    sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
                    sorted_marker_values = np.array([val for pos, val in sorted_markers])

                except Exception as e:
                    QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    return

                # Ensure there are still at least two markers after potential filtering
                if len(sorted_marker_positions) < 2:
                    QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    return

                # Prompt user and set up the click handler
                QMessageBox.information(self, "Instruction",
                                        "Click on the target protein location in the preview window.\n"
                                        "The closest standard set (Gel or WB) will be used for calculation.")
                # Pass the *sorted* full marker data to the click handler
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.get_protein_location_and_clear_preview(
                    event, sorted_marker_positions, sorted_marker_values
                )
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_mw_predict_preview
                
            def get_protein_location(self, event, all_marker_positions, all_marker_values):
                """
                Handles the mouse click event for protein selection.
                Determines the relevant standard set based on click proximity,
                performs regression on that set, predicts MW, and plots the results.
                """
                # --- 1. Get Protein Click Position (Image Coordinates) ---
                pos = event.position()
                cursor_x, cursor_y = pos.x(), pos.y()

                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
                image_width = self.image.width() if self.image and self.image.width() > 0 else 1
                image_height = self.image.height() if self.image and self.image.height() > 0 else 1

                scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1.0
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2

                protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
                self.protein_location = (cursor_x, cursor_y)

                # --- 2. Identify Potential Standard Sets (Partitioning) ---
                transition_index = -1
                initial_decrease = False
                for k in range(1, len(all_marker_values)):
                    if all_marker_values[k] < all_marker_values[k-1]:
                        initial_decrease = True
                    if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                        transition_index = k
                        break

                # --- 3. Select the Active Standard Set ---
                active_marker_positions = None
                active_marker_values = None
                set_name = "Full Set"

                if transition_index != -1:
                    set1_positions = all_marker_positions[:transition_index]
                    set1_values = all_marker_values[:transition_index]
                    set2_positions = all_marker_positions[transition_index:]
                    set2_values = all_marker_values[transition_index:]
                    valid_set1 = len(set1_positions) >= 2
                    valid_set2 = len(set2_positions) >= 2

                    if valid_set1 and valid_set2:
                        mean_y_set1 = np.mean(set1_positions)
                        mean_y_set2 = np.mean(set2_positions)
                        if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                            active_marker_positions = set1_positions
                            active_marker_values = set1_values
                            set_name = "Set 1 (e.g., Gel)"
                        else:
                            active_marker_positions = set2_positions
                            active_marker_values = set2_values
                            set_name = "Set 2 (e.g., WB)"
                    elif valid_set1:
                        active_marker_positions = set1_positions
                        active_marker_values = set1_values
                        set_name = "Set 1 (e.g., Gel)"
                    elif valid_set2:
                        active_marker_positions = set2_positions
                        active_marker_values = set2_values
                        set_name = "Set 2 (e.g., WB)"
                    else:
                        QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning (need >= 2 points per set).")
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        if hasattr(self, "protein_location"): del self.protein_location
                        return
                else:
                    if len(all_marker_positions) >= 2:
                        active_marker_positions = all_marker_positions
                        active_marker_values = all_marker_values
                        set_name = "Single Set"
                    else:
                        QMessageBox.warning(self, "Error", "Not enough markers in the single set (need >= 2 points).")
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        if hasattr(self, "protein_location"): del self.protein_location
                        return

                if active_marker_positions is None or len(active_marker_positions) < 2:
                    QMessageBox.warning(self, "Error", f"Selected active set '{set_name}' has insufficient points ({len(active_marker_positions) if active_marker_positions is not None else 0}).")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    if hasattr(self, "protein_location"): del self.protein_location
                    return

                # --- 4. Perform Regression on the Active Set ---
                # THESE ARE THE CORRECT DEFINITIONS:
                min_pos_active = np.min(active_marker_positions)
                max_pos_active = np.max(active_marker_positions)

                if max_pos_active == min_pos_active:
                    QMessageBox.warning(self, "Error", "All markers in the selected active set have the same position.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    if hasattr(self, "protein_location"): del self.protein_location
                    return

                normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)
                try:
                    log_marker_values_active = np.log10(active_marker_values)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not log-transform marker values for active set (are they all positive?): {e}")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    if hasattr(self, "protein_location"): del self.protein_location
                    return

                selected_model_text = "Log-Linear (Degree 1)"
                if hasattr(self, 'mw_regression_model_combo'):
                    selected_model_text = self.mw_regression_model_combo.currentText()

                poly_degree = 1
                if "Degree 2" in selected_model_text: poly_degree = 2
                elif "Degree 3" in selected_model_text: poly_degree = 3

                if len(normalized_distances_active) <= poly_degree:
                    QMessageBox.warning(self, "Regression Error",
                                        f"Not enough data points ({len(normalized_distances_active)}) in the active set for a polynomial of degree {poly_degree}.\n"
                                        "Please select a lower degree or add more markers to the active set.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    if hasattr(self, "protein_location"): del self.protein_location
                    return

                coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, poly_degree)
                residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
                ss_res = np.sum(residuals**2)
                ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
                r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0

                # --- 5. Predict MW for the Clicked Protein ---
                normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)
                predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
                predicted_weight = 10 ** predicted_log10_weight

                # --- 6. Update View and Plot ---
                self.update_live_view()

                fit_line_x_dense_norm = np.linspace(np.min(normalized_distances_active), np.max(normalized_distances_active), 200)
                fit_line_y_log_dense = np.polyval(coefficients, fit_line_x_dense_norm)
                fit_line_y_mw_dense = 10**fit_line_y_log_dense

                # Normalize all marker positions using the active set's min/max for plotting
                # THIS IS WHERE min_pos_active and max_pos_active ARE USED:
                if (max_pos_active - min_pos_active) == 0: # Should have been caught by earlier check
                    all_norm_distances_plot = np.zeros_like(all_marker_positions.astype(float))
                else:
                    all_norm_distances_plot = (all_marker_positions.astype(float) - min_pos_active) / (max_pos_active - min_pos_active)

                self.plot_molecular_weight_graph(
                    all_norm_distances_plot=all_norm_distances_plot,
                    all_marker_values_plot=all_marker_values.astype(float),
                    active_norm_distances_points=normalized_distances_active,
                    active_marker_values_points=active_marker_values.astype(float),
                    fit_line_x_dense=fit_line_x_dense_norm,
                    fit_line_y_dense=fit_line_y_mw_dense,
                    predicted_norm_position_plot=normalized_protein_position,
                    predicted_mw_value=predicted_weight,
                    r_squared=r_squared,
                    active_set_name=set_name,
                    regression_model_name=selected_model_text
                )

                self._reset_live_view_label_custom_handlers()
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.run_predict_MW = True

            def get_protein_location_and_clear_preview(self, event, all_marker_positions, all_marker_values):
                self.get_protein_location(event, all_marker_positions, all_marker_values)
                self.live_view_label.mw_predict_preview_enabled = False
                self.live_view_label.mw_predict_preview_position = None
                self.live_view_label.setMouseTracking(False)
                # Reset the custom handlers after the action is complete
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.update()
                
            def update_mw_predict_preview(self, event):
                if self.live_view_label.mw_predict_preview_enabled:
                    untransformed_label_pos = self.live_view_label.transform_point(event.position())
                    # Snapping for MW preview is optional, but can be nice
                    snapped_label_pos = self.snap_point_to_grid(untransformed_label_pos) 
                    self.live_view_label.mw_predict_preview_position = snapped_label_pos
                    self.live_view_label.update()
                elif hasattr(self.live_view_label, '_original_mouseMoveEvent'): # Pass to original if exists
                     self.live_view_label._original_mouseMoveEvent(event)
                
            def plot_molecular_weight_graph(
                self,
                all_norm_distances_plot, all_marker_values_plot,
                active_norm_distances_points, active_marker_values_points,
                fit_line_x_dense, fit_line_y_dense,
                predicted_norm_position_plot, predicted_mw_value,
                r_squared, active_set_name, regression_model_name
            ):
                """
                Plots the molecular weight prediction graph and displays it in a custom dialog.
                This function now expects `all_norm_distances_plot` to be pre-calculated.
                """
                fig, ax = plt.subplots(figsize=(4.5, 3.5))

                ax.scatter(all_norm_distances_plot, all_marker_values_plot, color="grey", alpha=0.5, label="All Markers (Context)", s=25)
                ax.scatter(active_norm_distances_points, active_marker_values_points, color="red", label=f"Active Set Data", s=40, marker='o')

                simple_model_name_for_legend = regression_model_name.split('(')[0].strip()
                fit_line_label = f"Fit ({simple_model_name_for_legend})"
                ax.plot(fit_line_x_dense, fit_line_y_dense, color="blue", label=fit_line_label, linewidth=1.5)

                target_line_label = f"Target Protein ({predicted_mw_value:.1f} units)"
                ax.axvline(predicted_norm_position_plot, color="green", linestyle="--",
                            label=target_line_label, linewidth=1.5)

                ax.set_xlabel(f"Normalized Distance (Relative to {active_set_name})", fontsize=9)
                ax.set_ylabel("Molecular Weight (units)", fontsize=9)
                ax.set_yscale("log")
                ax.legend(fontsize='x-small', loc='best')
                ax.set_title(f"MW Prediction (Using: {regression_model_name}, Active Set: {active_set_name})\nFit R² on active set: {r_squared:.3f}", fontsize=9, wrap=True)
                ax.grid(True, which='both', linestyle=':', linewidth=0.5)
                ax.tick_params(axis='both', which='major', labelsize=8)

                plt.tight_layout(pad=0.5)

                pixmap = None
                try:
                    buffer = BytesIO()
                    plt.savefig(buffer, format="png", bbox_inches="tight", dpi=100)
                    buffer.seek(0)
                    pixmap = QPixmap()
                    pixmap.loadFromData(buffer.read())
                    buffer.close()
                except Exception as plot_err:
                    QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
                finally:
                     plt.close(fig)

                if pixmap:
                    dialog = QDialog(self)
                    dialog.setWindowTitle("Prediction Result")
                    layout_diag = QVBoxLayout(dialog)

                    results_text = (
                        f"<b>Prediction using {active_set_name} and {regression_model_name}:</b><br>"
                        f"Predicted MW: <b>{predicted_mw_value:.2f}</b> units<br>"
                        f"Fit R² (on active set): {r_squared:.3f}"
                    )
                    info_label_diag = QLabel(results_text)
                    info_label_diag.setTextFormat(Qt.RichText)
                    info_label_diag.setWordWrap(True)
                    layout_diag.addWidget(info_label_diag)

                    plot_label_diag = QLabel()
                    plot_label_diag.setPixmap(pixmap)
                    plot_label_diag.setAlignment(Qt.AlignCenter)
                    layout_diag.addWidget(plot_label_diag)

                    button_box_diag = QDialogButtonBox(QDialogButtonBox.Ok)
                    button_box_diag.accepted.connect(dialog.accept)
                    layout_diag.addWidget(button_box_diag)

                    dialog.setLayout(layout_diag)
                    dialog.exec()
                else:
                     QMessageBox.information(self, "Prediction Result (No Plot)",
                        f"Used {active_set_name} and {regression_model_name} for prediction.\n"
                        f"Predicted MW: {predicted_mw_value:.2f} units\n"
                        f"Fit R² (on active set): {r_squared:.3f}"
                     )

                
            def reset_image(self):
                self.cancel_rectangle_crop_mode()
                self.crop_rectangle_coords = None
                self.live_view_label.clear_crop_preview()

                # --- Reset and Disable Crop Sliders ---
                slider_info = [
                    (getattr(self, 'crop_x_start_slider', None), self.crop_slider_min),
                    (getattr(self, 'crop_x_end_slider', None), self.crop_slider_max),
                    (getattr(self, 'crop_y_start_slider', None), self.crop_slider_min),
                    (getattr(self, 'crop_y_end_slider', None), self.crop_slider_max)
                ]
                for slider, default_value in slider_info:
                    if slider:
                        slider.blockSignals(True)
                        slider.setValue(default_value)
                        slider.setEnabled(False) # Disable sliders on reset
                        slider.blockSignals(False)
                # --- End Reset Crop ---

                # ... (rest of reset_image method: clearing image, other sliders, markers, etc.) ...
                if hasattr(self, 'image_master') and self.image_master and not self.image_master.isNull():
                    self.image = self.image_master.copy()
                    self.image_before_padding = None # Will be self.image if padding is first op
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False
                    self.contrast_applied = False # Reset contrast flag
                else:
                    self.image = None
                    self.original_image = None
                    self.image_master = None
                    self.image_before_padding = None
                    self.image_contrasted = None
                    self.image_before_contrast = None
                    self.image_padded = False
                    self.contrast_applied = False

                if hasattr(self, 'show_grid_checkbox_x'): self.show_grid_checkbox_x.setChecked(False)
                if hasattr(self, 'show_grid_checkbox_y'): self.show_grid_checkbox_y.setChecked(False)
                if hasattr(self, 'grid_size_input'): self.grid_size_input.setValue(20)

                self.warped_image=None
                if hasattr(self, 'left_markers'): self.left_markers.clear()
                if hasattr(self, 'right_markers'): self.right_markers.clear()
                if hasattr(self, 'top_markers'): self.top_markers.clear()
                if hasattr(self, 'custom_markers'): self.custom_markers.clear()
                if hasattr(self, 'custom_shapes'): self.custom_shapes.clear()
                self.cancel_drawing_mode()
                self.clear_predict_molecular_weight()

                if hasattr(self, 'orientation_slider'): self.orientation_slider.setValue(0)
                if hasattr(self, 'taper_skew_slider'): self.taper_skew_slider.setValue(0)
                self._update_level_slider_ranges_and_defaults() # This sets black to 0, white to max
                if hasattr(self, 'gamma_slider'):
                    self.gamma_slider.blockSignals(True)
                    self.gamma_slider.setValue(100)
                    self.gamma_slider.blockSignals(False)
                if hasattr(self, 'gamma_value_label'): self.gamma_value_label.setText("1.00")
                if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(0)
                if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(0)
                if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(0)
                if hasattr(self, 'left_padding_input'): self.left_padding_input.setText("0")
                if hasattr(self, 'right_padding_input'): self.right_padding_input.setText("0")
                if hasattr(self, 'top_padding_input'): self.top_padding_input.setText("0")
                if hasattr(self, 'bottom_padding_input'): self.bottom_padding_input.setText("0")
                if hasattr(self, "custom_markers"): self.custom_markers.clear()
                if hasattr(self, "custom_shapes"): self.custom_shapes.clear()

                self.marker_mode = None
                self.current_left_marker_index = 0
                self.current_right_marker_index = 0
                self.current_top_label_index = 0
                self.left_marker_shift_added = 0
                self.right_marker_shift_added = 0
                self.top_marker_shift_added = 0
                self.live_view_label.mode = None
                self.live_view_label.quad_points = []
                self.live_view_label.setCursor(Qt.ArrowCursor)

                try:
                    if hasattr(self, 'combo_box'):
                        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                        self.on_combobox_changed()
                except Exception as e: pass

                if self.image and not self.image.isNull():
                    try:
                        self._update_preview_label_size()
                        if hasattr(self, 'left_padding_input'): self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'right_padding_input'): self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'top_padding_input'): self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                    except Exception as e: pass
                else:
                    self.live_view_label.clear()
                    self._update_preview_label_size()

                self.live_view_label.zoom_level = 1.0
                self.live_view_label.pan_offset = QPointF(0, 0)
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                                
                # Reset overlay sliders to 0 after their ranges are updated
                if hasattr(self, 'image1_left_slider'): self.image1_left_slider.setValue(0)
                if hasattr(self, 'image1_top_slider'): self.image1_top_slider.setValue(0)
                if hasattr(self, 'image1_resize_slider'): self.image1_resize_slider.setValue(100)
                if hasattr(self, 'image2_left_slider'): self.image2_left_slider.setValue(0)
                if hasattr(self, 'image2_top_slider'): self.image2_top_slider.setValue(0)
                if hasattr(self, 'image2_resize_slider'): self.image2_resize_slider.setValue(100)
                
                self._update_overlay_slider_ranges()

                self._update_marker_slider_ranges()
                self.update_live_view()
                self._update_status_bar()

        if app: # Ensure app exists before trying to set style
            app.setStyle("Fusion")
            # ... (your stylesheet) ...

        main_window = CombinedSDSApp()

        # --- Close Loading Screen and Show Main Window ---
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # YOU ALREADY HAVE THIS, WHICH IS GOOD!
            loading_dialog = None      # <<--- ADD THIS LINE HERE

        main_window.show()

        # Connect cleanup
        if main_window and app: # Check both exist
            app.aboutToQuit.connect(main_window.cleanup_temp_clipboard_file)

        # --- Start Main Event Loop ---
        if app:
            exit_code = app.exec()
            sys.exit(exit_code)
        else:
            print("FATAL ERROR: QApplication could not be initialized.")
            sys.exit(1)

    except ImportError as e_imp:
        print(f"FATAL ERROR: Missing required library: {e_imp}")
        traceback.print_exc()
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # Good
            loading_dialog = None

        # Use existing app instance for QMessageBox if possible, or create one carefully
        error_app_instance = QApplication.instance()
        temp_app_created = False
        if not error_app_instance:
            try:
                error_app_instance = QApplication([])
                temp_app_created = True
            except RuntimeError: # Catch "already exists" if some other part created it
                error_app_instance = QApplication.instance() # Try to grab it again

        if error_app_instance:
            QMessageBox.critical(None, "Import Error", f"A required library is missing: {e_imp}\nPlease install it and restart the application.")
            if temp_app_created:
                # If we created a temporary app just for the message box,
                # we don't call exec on it.
                pass
        else:
            print("Could not display import error message box: No QApplication.")
        sys.exit(1)

    except Exception as e_startup:
        print(f"FATAL ERROR during application startup: {e_startup}")
        traceback.print_exc()
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # Good
            loading_dialog = None

        # Log exception using your global handler if sys.excepthook is set
        if sys.excepthook is not sys.__excepthook__: # Check if it's your custom hook
            try:
                sys.excepthook(type(e_startup), e_startup, e_startup.__traceback__)
            except Exception as log_hook_err:
                print(f"ERROR calling custom excepthook: {log_hook_err}")
        else: # Fallback to direct logging if hook wasn't set or failed
            try:
                logging.error("Uncaught startup exception", exc_info=(type(e_startup), e_startup, e_startup.__traceback__))
            except Exception as log_err:
                print(f"Failed to log startup exception directly: {log_err}")

        # Show critical error message box using existing app or a temporary one
        error_app_instance = QApplication.instance()
        temp_app_created = False
        if not error_app_instance:
            try:
                error_app_instance = QApplication([])
                temp_app_created = True
            except RuntimeError:
                error_app_instance = QApplication.instance()

        if error_app_instance:
            QMessageBox.critical(None, "Application Startup Error", f"An unexpected error occurred during startup:\n{e_startup}\n\nCheck error_log.txt for details.")
            if temp_app_created:
                pass
        else:
            print("Could not display startup error message box: No QApplication.")
        sys.exit(1)

    finally:
        # Ensure the loading dialog is closed and cleaned up
        if loading_dialog is not None: # Check if the Python variable still holds an object
            try:
                if loading_dialog.isVisible():
                    loading_dialog.close()
                # Regardless of visibility, if it's not None, schedule for deletion
                # to handle cases where it might exist but wasn't properly closed.
                loading_dialog.deleteLater()
            except RuntimeError:
                # This catches the "Internal C++ object already deleted" specifically
                # print("INFO: loading_dialog was already deleted (RuntimeError in finally).")
                pass # Object is already gone, nothing more to do with it
            except Exception as e_finally_close:
                # Catch any other unexpected error during cleanup
                print(f"ERROR in finally block trying to clean up loading_dialog: {e_finally_close}")
            finally:
                loading_dialog = None # Ensure the Python variable is cleared