import sys
import os
from PyQt5.QtWidgets import (QApplication, QDialog, QLabel, QVBoxLayout,
                             QDesktopWidget, QMessageBox) # Minimal needed for loading + error
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont # For setting font
import logging
import traceback


# --- NEW Minimal Loading Dialog ---
class MinimalLoadingDialog(QDialog):
    """
    A very lightweight, minimal-resource loading dialog.
    Uses basic widgets and styling for fast appearance.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Loading...") # Simple title, might not show if frameless
        # Use flags for a clean, frameless look that stays on top
        self.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True) # Allows for rounded corners if stylesheet uses them

        # --- Styling (Keep it simple) ---
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(45, 45, 50, 240); /* Dark semi-transparent */
                border: 1px solid #777777;
                border-radius: 8px;
            }
            QLabel {
                color: #E0E0E0; /* Light text */
                padding: 25px; /* Add space around text */
                background-color: transparent; /* Ensure label bg is transparent */
            }
        """)

        # --- Layout and Label ---
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0) # No extra margins in layout

        self.label = QLabel("Loading Software,\nPlease Wait...") # Updated text
        font = QFont("Segoe UI", 11) # Use a common system font
        font.setBold(True)
        self.label.setFont(font)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.setLayout(layout)

        # --- Size and Position ---
        self.setFixedSize(320, 120) # Adjust size as needed
        self.center_on_screen()

    def center_on_screen(self):
        """Centers the dialog on the primary screen."""
        try:
            screen_geo = QDesktopWidget().availableGeometry()
            dialog_geo = self.frameGeometry()
            center_point = screen_geo.center()
            dialog_geo.moveCenter(center_point)
            self.move(dialog_geo.topLeft())
        except Exception as e:
            print(f"Warning: Could not center loading dialog: {e}")
            self.move(100, 100) # Fallback position


# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
script_dir = os.path.dirname(os.path.abspath(__file__))
# Construct the absolute path for the log file
log_file_path = os.path.join(script_dir, "error_log.txt")


# --- Configure logging ---
logging.basicConfig(
    filename=log_file_path, # Use the absolute path
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

try:
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not write initial log message: {e}") # Print error if immediate logging fails


def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    print("!!! log_exception called !!!") # Add print statement here too
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    # Log the exception
    try:
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
        # Optional: Force flush if you suspect buffering issues, though unlikely for ERROR level
        # logging.getLogger().handlers[0].flush()
    except Exception as log_err:
        print(f"ERROR: Failed to log exception to file: {log_err}") # Print logging specific errors

    # Display a QMessageBox with the error details
    try:
        error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}\n\n(Check error_log.txt for details)"
        QMessageBox.critical(
            None,
            "Unexpected Error",
            error_message,
            QMessageBox.Ok
        )
    except Exception as q_err:
         print(f"ERROR: Failed to show QMessageBox: {q_err}")


# Set the custom exception handler
sys.excepthook = log_exception



 



	
if __name__ == "__main__":
    loading_dialog = None  # Initialize variable
    main_window = None     # Initialize variable
    app = None             # Initialize variable

    try:
        # --- Enable High DPI Scaling FIRST ---
        try:
            QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
            QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
        except AttributeError:
            print("Warning: Could not set High DPI attributes.")
        except Exception as e:
            print(f"Warning: Error setting High DPI attributes: {e}")

        # --- Create QApplication ---
        # Use sys.argv for command line arguments if needed, otherwise empty list
        app = QApplication(sys.argv if len(sys.argv) > 0 else [])

        # --- Create and Show Minimal Loading Screen IMMEDIATELY ---
        try:
            loading_dialog = MinimalLoadingDialog()
            loading_dialog.show()
            app.processEvents() # Crucial: Make the GUI update and show the dialog
        except Exception as e:
            print(f"ERROR: Could not create/show minimal loading dialog: {e}")
            # Proceed without loading screen if it fails, but log the error
            loading_dialog = None # Ensure it's None if creation failed

        import sys
        import svgwrite
        import tempfile
        from tempfile import NamedTemporaryFile
        import base64
        from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
        from io import BytesIO
        import io
        from PyQt5.QtWidgets import (
            QDesktopWidget, QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem,QToolBar,QStyle,
            QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy,
            QMainWindow, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit,
            QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog,
            QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox,
            QDialog, QHeaderView, QAbstractItemView, QMenu, QAction, QMenuBar, QFontDialog, QListWidget
        )
        from PyQt5.QtGui import QPixmap, QIcon, QPalette,QKeySequence, QImage, QPolygonF,QPainter, QBrush, QColor, QFont, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
        from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF, QUrl, QSize, QSizeF, QMimeData, QUrl, pyqtSignal
        import json
        import os
        import numpy as np
        import matplotlib.pyplot as plt
        import platform
        import openpyxl
        from openpyxl.styles import Font
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.gridspec import GridSpec
        from skimage.restoration import rolling_ball 
        from scipy.signal import find_peaks
        from scipy.ndimage import gaussian_filter1d
        from scipy.ndimage import grey_opening, grey_erosion, grey_dilation
        from scipy.interpolate import interp1d # Needed for interpolation
        import cv2
        import datetime

        def create_text_icon(font_type: QFont, icon_size: QSize, color: QColor, symbol: str) -> QIcon:
            """Creates a QIcon by drawing text/symbol onto a pixmap."""
            pixmap = QPixmap(icon_size)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing, True)
            painter.setRenderHint(QPainter.TextAntialiasing, True) # Good for text

            # Font settings (adjust as needed)
            font = QFont(font_type)
            # Make arrow slightly smaller than +/-, maybe not bold? Experiment.
            font.setPointSize(min(14, int(icon_size.height()*0.75)))
            # font.setBold(True) # Optional: Make arrows bold or not
            painter.setFont(font)
            painter.setPen(color)

            # Draw the symbol centered
            painter.drawText(pixmap.rect(), Qt.AlignCenter, symbol)
            painter.end()
            return QIcon(pixmap)

        # Set Style (can be done after app exists)
        app.setStyle("Fusion")
        app.setStyleSheet("""
            /* ... Your existing stylesheet ... */
            QSlider::handle:horizontal { /* Example */
                width: 10px; /* Example adjustment if needed */
                height: 20px;
                margin: -5px 0;
                background: #DDDDDD;
                border: 1px solid #555;
                border-radius: 5px;
            }
            QStatusBar QLabel {
                margin-left: 2px; margin-right: 5px; padding: 0px 0px; border: none;
            }
            QPushButton:checked {
                background-color: #a0d0a0; border: 1px solid #50a050;
            }
        """)
        class AutoLaneTuneDialog(QDialog):
            """
            Simplified dialog for tuning peak detection parameters for automatic lane markers.
            Shows the intensity profile and detected peaks. Allows adjustment of detection parameters.
            Does NOT handle area calculation or individual band boundary adjustments.

            This dialog operates on a rectangular PIL.Image object provided as `pil_image_data`.
            If this data comes from a warped quadrilateral region of an original image (i.e.,
            the `pil_image_data` itself is the rectangular result of a warp), the detected
            peak coordinates will be relative to this warped rectangular image. The calling
            code is responsible for any geometric transformations to map these coordinates
            back to the original image space if needed. The `is_from_quad_warp` parameter
            can be used to indicate this context, which will adjust the dialog's title.
            """
            def __init__(self, pil_image_data, initial_settings, parent=None, is_from_quad_warp=False): # Added is_from_quad_warp
                super().__init__(parent)

                # Set window title based on the source of the image data
                if is_from_quad_warp:
                    self.setWindowTitle("Tune Peaks (Warped Region)")
                else:
                    self.setWindowTitle("Tune Automatic Peak Detection")

                self.setGeometry(150, 150, 800, 700) # Adjusted height
                self.pil_image_for_display = pil_image_data
                self.selected_peak_index = -1
                self.deleted_peak_indices = set()
                self._all_initial_peaks = np.array([])
                self.add_peak_mode_active = False

                # --- Validate and Store Input Image ---
                if not isinstance(pil_image_data, Image.Image):
                    raise TypeError("Input 'pil_image_data' must be a PIL Image object")

                # Determine intensity range and create numpy array
                self.intensity_array_original_range = None
                self.original_max_value = 255.0 # Default
                pil_mode = pil_image_data.mode
                try:
                    # Handle common grayscale modes used in the main app
                    if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'L':
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        self.original_max_value = 255.0
                    elif pil_mode == 'F': # Handle float images
                        self.intensity_array_original_range = np.array(pil_image_data, dtype=np.float64)
                        max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                        self.original_max_value = max(1.0, max_in_float) # Use max value or 1.0
                    else: # Attempt conversion to grayscale 'L' as fallback
                        gray_img = pil_image_data.convert("L")
                        self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                        self.original_max_value = 255.0
                        print(f"AutoLaneTuneDialog: Converted input mode '{pil_mode}' to 'L'.")

                    if self.intensity_array_original_range is None:
                        raise ValueError("Failed to convert PIL image to NumPy array.")
                    if self.intensity_array_original_range.ndim != 2:
                        raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

                except Exception as e:
                    raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

                self.profile_original_inverted = None # Smoothed, inverted profile (original range)
                self.profile = None # Scaled (0-255), inverted, SMOOTHED profile for detection
                self.detected_peaks = np.array([]) # Store indices of detected peaks

                # --- Settings and State ---
                # Use settings passed from the main app (likely self.peak_dialog_settings)
                self.smoothing_sigma = initial_settings.get('smoothing_sigma', 2.0)
                self.peak_height_factor = initial_settings.get('peak_height_factor', 0.1)
                self.peak_distance = initial_settings.get('peak_distance', 10)
                self.peak_prominence_factor = initial_settings.get('peak_prominence_factor', 0.02)
                # Band estimation method is needed to generate the profile
                self.band_estimation_method = initial_settings.get('band_estimation_method', "Mean")
                self._final_settings = initial_settings.copy() # Store a copy to return modifications

                # Check dependencies needed for profile generation and peak detection
                if find_peaks is None or gaussian_filter1d is None:
                     QMessageBox.critical(self, "Dependency Error",
                                          "Missing SciPy library functions.\n"
                                          "Peak detection and smoothing require SciPy.\n"
                                          "Please install it (e.g., 'pip install scipy') and restart.")
                     # Don't call accept/reject here, let init fail or handle gracefully
                     # self.close() # Or close immediately

                # Build UI & Initial Setup
                self._setup_ui()
                self.run_peak_detection_and_plot() # Initial calculation and plot

            def _setup_ui(self):
                """Creates and arranges the UI elements, including image preview."""
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(10)

                # --- Matplotlib Plot Canvas Area ---
                plot_widget = QWidget() # Container for the plots
                plot_layout = QVBoxLayout(plot_widget)
                plot_layout.setContentsMargins(0, 0, 0, 0) # No margins for the layout

                # Use GridSpec for plot and image preview arrangement
                # More height for the profile plot
                self.fig = plt.figure(figsize=(7, 5)) # Adjusted height for gridspec
                gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig)
                self.fig.tight_layout(pad=0.5)
                self.ax_profile = self.fig.add_subplot(gs[0]) # Axis for the profile
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax_profile) # Axis for the image preview

                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                # --- Enable picking on the canvas ---
                self.canvas.mpl_connect('button_press_event', self.on_canvas_click)
                # --- End Enable picking ---
                plot_layout.addWidget(self.canvas)
                main_layout.addWidget(plot_widget, stretch=1)

                # --- Controls Group ---
                controls_group = QGroupBox("Peak Detection Parameters")
                controls_layout = QGridLayout(controls_group)
                controls_layout.setSpacing(8)

                # Band Estimation (Needed for profile generation)
                controls_layout.addWidget(QLabel("Profile Method:"), 0, 0)
                self.band_estimation_combobox = QComboBox()
                self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
                self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
                self.band_estimation_combobox.currentIndexChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2) # Span 2 columns

                # Smoothing Sigma
                self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
                self.smoothing_slider = QSlider(Qt.Horizontal)
                self.smoothing_slider.setRange(0, 100) # 0.0 to 10.0
                self.smoothing_slider.setValue(int(self.smoothing_sigma * 10))
                self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})"))
                self.smoothing_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.smoothing_label, 1, 0)
                controls_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2)

                # Peak Prominence
                self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
                self.peak_prominence_slider = QSlider(Qt.Horizontal)
                self.peak_prominence_slider.setRange(0, 100) # 0.0 to 1.0 factor
                self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
                self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})"))
                self.peak_prominence_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
                controls_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)

                # Peak Height
                self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})")
                self.peak_height_slider = QSlider(Qt.Horizontal)
                self.peak_height_slider.setRange(0, 100)
                self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
                self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f})"))
                self.peak_height_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_height_slider_label, 3, 0)
                controls_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)

                # Peak Distance
                self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
                self.peak_distance_slider = QSlider(Qt.Horizontal)
                self.peak_distance_slider.setRange(1, 200)
                self.peak_distance_slider.setValue(self.peak_distance)
                self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
                self.peak_distance_slider.valueChanged.connect(self.run_peak_detection_and_plot) # Re-run on change
                controls_layout.addWidget(self.peak_distance_slider_label, 4, 0)
                controls_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)


                # --- Add Delete Peak Button ---
                self.delete_peak_button = QPushButton("Delete Selected Peak")
                self.delete_peak_button.setEnabled(False)
                self.delete_peak_button.setToolTip("Click on a peak marker in the plot to select it, then click this.")
                self.delete_peak_button.clicked.connect(self.delete_selected_peak)
                controls_layout.addWidget(self.delete_peak_button, 5, 0, 1, 1) # Column 0, span 1

                # --- NEW: Add Peak Manually Button ---
                self.add_peak_button = QPushButton("Add Peak at Click")
                self.add_peak_button.setCheckable(True) # Make it a toggle button
                self.add_peak_button.setToolTip("Toggle: Click on the profile to add a new peak marker.")
                self.add_peak_button.clicked.connect(self.toggle_add_peak_mode)
                controls_layout.addWidget(self.add_peak_button, 5, 1, 1, 2) # Next to delete, span 2
                # --- End Add Delete Button ---

                controls_layout.setColumnStretch(1, 1) # Slider column stretch
                main_layout.addWidget(controls_group)

                # --- Bottom Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept_and_return_peaks)
                button_box.rejected.connect(self.reject)
                main_layout.addWidget(button_box)
                
            def toggle_add_peak_mode(self, checked):
                """Toggles the manual peak adding mode."""
                self.add_peak_mode_active = checked
                if checked:
                    self.canvas.setCursor(Qt.CrossCursor)
                    # Deactivate other selection modes if necessary
                    self.identify_peak_button.setChecked(False)
                    self.manual_select_mode_active = False
                    self.selected_peak_for_ui_focus = -1
                    self.selected_peak_index_for_delete = -1
                    self.delete_selected_peak_button.setEnabled(False)
                    self._update_peak_group_box_styles()
                    self.update_plot() # To clear any selection highlights
                    QMessageBox.information(self, "Add Peak", "Add Peak Mode: ON. Click on the profile plot to add a peak.")
                else:
                    self.canvas.setCursor(Qt.ArrowCursor)

            def on_canvas_click(self, event):
                """Handles clicks on the canvas for adding or selecting peaks."""
                # Ignore clicks outside the profile axes or if no data
                if event.inaxes != self.ax_profile or self.profile_original_inverted is None:
                    return
                if event.button != 1: # Only process left-clicks
                    return

                clicked_x = int(round(event.xdata)) # Get x-coordinate of the click

                if self.add_peak_mode_active:
                    # --- Add Peak Mode ---
                    if 0 <= clicked_x < len(self.profile_original_inverted):
                        self.add_manual_peak(clicked_x)
                    else:
                        print(f"Clicked X ({clicked_x}) is outside profile bounds.")
                    # Deactivate add mode after one click for single additions
                    # self.add_peak_button.setChecked(False) # Optional: Toggle off after one add
                    # self.toggle_add_peak_mode(False)       # Or call the toggle function

                else:
                    # --- Select Peak Mode (for deletion) ---
                    # Find the closest *existing* peak marker to the click
                    if len(self.detected_peaks) > 0:
                        # Calculate distance from click to all active peak X-coordinates
                        distances = np.abs(self.detected_peaks - clicked_x)
                        min_dist_idx = np.argmin(distances)
                        # Define a click tolerance (e.g., 5 pixels on the x-axis)
                        click_tolerance_x = max(5, self.peak_distance / 4) # Heuristic based on peak distance

                        if distances[min_dist_idx] <= click_tolerance_x:
                            # A peak is close enough to the click
                            self.selected_peak_index = self.detected_peaks[min_dist_idx]
                            self.delete_peak_button.setEnabled(True)
                            print(f"Selected peak for deletion at index: {self.selected_peak_index}")
                        else:
                            # Click was not close enough to an existing peak
                            self.selected_peak_index = -1
                            self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights() # Update plot to show selection or clear it
                    else:
                        # No peaks to select
                        self.selected_peak_index = -1
                        self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()


            def add_manual_peak(self, x_coord):
                """Adds a new peak at the given x-coordinate if not already present or deleted."""
                if x_coord in self._all_initial_peaks or x_coord in self.deleted_peak_indices:
                    print(f"Peak at {x_coord} already exists or was previously deleted.")
                    # Optionally, undelete if it was in deleted_peak_indices
                    if x_coord in self.deleted_peak_indices:
                         reply = QMessageBox.question(self, "Undelete Peak?",
                                                      f"A peak at index {x_coord} was previously deleted. Undelete it?",
                                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                         if reply == QMessageBox.Yes:
                             self.deleted_peak_indices.discard(x_coord)
                             # Add to _all_initial_peaks if it wasn't there (unlikely if deleted, but for safety)
                             if x_coord not in self._all_initial_peaks:
                                  self._all_initial_peaks = np.sort(np.append(self._all_initial_peaks, x_coord))
                             # Update active peaks
                             self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])
                             print(f"Undeleted peak at {x_coord}.")
                             self.update_plot_highlights() # Update plot
                         return # Don't proceed to add again
                    else: # Already an initial peak
                         return

                # Add to both _all_initial_peaks and detected_peaks
                self._all_initial_peaks = np.sort(np.append(self._all_initial_peaks, x_coord))
                self.detected_peaks = np.sort(np.append(self.detected_peaks, x_coord))

                print(f"Manually added peak at index: {x_coord}")
                self.update_plot_highlights() # Update plot to show the new peak

            def run_peak_detection_and_plot(self):
                """Generates profile, detects peaks, updates plot and image preview."""
                # --- Reset deleted peaks if parameters change ---
                self.deleted_peak_indices.clear()
                self.selected_peak_index = -1
                self.delete_peak_button.setEnabled(False)
                # --- End Reset ---

                if self.intensity_array_original_range is None: return
                if find_peaks is None or gaussian_filter1d is None: return

                # ... (Keep parameter updates and profile generation logic) ...
                self.band_estimation_method = self.band_estimation_combobox.currentText()
                self.smoothing_sigma = self.smoothing_slider.value() / 10.0
                self.peak_height_factor = self.peak_height_slider.value() / 100.0
                self.peak_distance = self.peak_distance_slider.value()
                self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0

                profile_temp = None
                if self.band_estimation_method == "Mean":
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)
                elif self.band_estimation_method.startswith("Percentile"):
                    try:
                        percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                        profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
                    except:
                        profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1)
                        print("AutoLaneTuneDialog: Defaulting to 5th percentile for profile.")
                else:
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)

                if profile_temp is None or not np.all(np.isfinite(profile_temp)):
                    print("AutoLaneTuneDialog: Profile calculation failed or resulted in NaN/Inf.")
                    profile_temp = np.zeros(self.intensity_array_original_range.shape[0])

                profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
                min_inverted_raw = np.min(profile_original_inv_raw)
                profile_original_inv_raw -= min_inverted_raw

                self.profile_original_inverted = profile_original_inv_raw
                try:
                    current_sigma = self.smoothing_sigma
                    if current_sigma > 0.1 and len(self.profile_original_inverted) > int(3 * current_sigma) * 2 + 1:
                        self.profile_original_inverted = gaussian_filter1d(self.profile_original_inverted, sigma=current_sigma)
                except Exception as smooth_err:
                    print(f"AutoLaneTuneDialog: Error smoothing profile: {smooth_err}")

                prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
                if prof_max_inv > prof_min_inv + 1e-6:
                    self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
                else:
                    self.profile = np.zeros_like(self.profile_original_inverted)

                profile_range_detect = np.ptp(self.profile); min_val_profile_detect = np.min(self.profile)
                if profile_range_detect < 1e-6 : profile_range_detect = 1.0
                min_height_abs = min_val_profile_detect + profile_range_detect * self.peak_height_factor
                min_prominence_abs = profile_range_detect * self.peak_prominence_factor
                min_prominence_abs = max(1.0, min_prominence_abs)

                try:
                    peaks_indices, _ = find_peaks(
                        self.profile, height=min_height_abs, prominence=min_prominence_abs,
                        distance=self.peak_distance, width=1
                    )
                    # *** Store ALL initially detected peaks ***
                    self._all_initial_peaks = np.sort(peaks_indices)
                    # *** Filter out user-deleted peaks for display ***
                    self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])

                except Exception as e:
                    print(f"AutoLaneTuneDialog: Peak detection error: {e}")
                    self._all_initial_peaks = np.array([])
                    self.detected_peaks = np.array([])


                # --- Update Plot ---
                self.ax_profile.clear()
                self.ax_image.clear() # Clear image axis too

                if self.profile_original_inverted is not None and len(self.profile_original_inverted) > 0:
                    self.ax_profile.plot(self.profile_original_inverted, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.0)

                    # Plot *currently active* detected peaks
                    # These are the peaks that `on_canvas_click` will try to select for deletion
                    if len(self.detected_peaks) > 0:
                        valid_peaks = self.detected_peaks[(self.detected_peaks >= 0) & (self.detected_peaks < len(self.profile_original_inverted))]
                        if len(valid_peaks) > 0:
                            peak_y_values = self.profile_original_inverted[valid_peaks]
                            # No need for `picker=True` if on_canvas_click handles selection by coordinate
                            self.peak_plot_artist, = self.ax_profile.plot(
                                valid_peaks, peak_y_values, "rx", markersize=8,
                                label=f"Active Peaks ({len(valid_peaks)})") # Removed picker

                            # --- Highlight selected peak (logic remains the same) ---
                            if self.selected_peak_index != -1 and self.selected_peak_index in valid_peaks:
                                 idx_in_valid = np.where(valid_peaks == self.selected_peak_index)[0]
                                 if len(idx_in_valid) > 0:
                                      self.ax_profile.plot(self.selected_peak_index, peak_y_values[idx_in_valid[0]],
                                                         'o', markersize=12, markeredgecolor='blue', markerfacecolor='none',
                                                         label='Selected')
                            # --- End Highlight ---
                    
                    # Profile plot styling
                    self.ax_profile.set_ylabel("Intensity (Smoothed, Inverted)", fontsize=9)
                    self.ax_profile.legend(fontsize='x-small')
                    self.ax_profile.set_title("Intensity Profile and Detected Peaks", fontsize=10)
                    self.ax_profile.grid(True, linestyle=':', alpha=0.6)
                    # Remove x-axis labels/ticks for the top plot
                    self.ax_profile.tick_params(axis='x', labelbottom=False)
                    if np.max(self.profile_original_inverted) > 10000:
                        self.ax_profile.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)

                    # --- Plot Image Preview ---
                    try:
                        # Rotate the PIL image 90 degrees for horizontal display matching profile axis
                        rotated_pil_image = self.pil_image_for_display.rotate(90, expand=True)
                        im_array_disp = np.array(rotated_pil_image)

                        # Determine vmin/vmax for imshow based on original data range
                        im_vmin, im_vmax = 0, self.original_max_value
                        # Handle float case specifically if needed
                        if self.pil_image_for_display.mode == 'F':
                            im_vmin, im_vmax = 0.0, self.original_max_value # Or 0.0, 1.0 if normalized

                        profile_length = len(self.profile_original_inverted)
                        extent = [0, profile_length - 1 if profile_length > 0 else 0, 0, rotated_pil_image.height]

                        self.ax_image.imshow(im_array_disp, cmap='gray', aspect='auto',
                                             extent=extent, vmin=im_vmin, vmax=im_vmax)
                        self.ax_image.set_xlabel("Pixel Index along Profile Axis", fontsize=9)
                        self.ax_image.set_yticks([]) # Hide Y ticks for image preview
                        self.ax_image.set_ylabel("Lane Width", fontsize=9)
                    except Exception as img_e:
                        print(f"AutoLaneTuneDialog: Error displaying image preview: {img_e}")
                        self.ax_image.text(0.5, 0.5, 'Error displaying preview', ha='center', va='center', transform=self.ax_image.transAxes)
                        self.ax_image.set_xticks([]); self.ax_image.set_yticks([])

                else:
                    self.ax_profile.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax_profile.transAxes)
                    self.ax_image.text(0.5, 0.5, "No Image Data", ha='center', va='center', transform=self.ax_image.transAxes)

                # Use tight_layout or constrained_layout
                # try:
                #     self.fig.tight_layout(pad=0.5)
                # except ValueError: # Can happen with certain plot states
                #     try:
                #         self.fig.set_constrained_layout(True)
                #     except AttributeError: pass # Older matplotlib

                self.canvas.draw_idle()
                plt.close()

            def on_pick(self, event):
                """Handles clicking on a peak marker."""
                # Check if the picked artist is the one we stored for the peaks
                if event.artist != getattr(self, 'peak_plot_artist', None):
                    return

                # Get the index of the clicked data point within the plotted data
                ind = event.ind[0] # Get the first index if multiple points are close

                # Map this index back to the actual peak coordinate value (Y-index)
                # Ensure self.detected_peaks is valid and index is within bounds
                if ind < len(self.detected_peaks):
                    clicked_peak_coord = self.detected_peaks[ind]
                    # Check if this peak wasn't already deleted internally
                    if clicked_peak_coord in self._all_initial_peaks and clicked_peak_coord not in self.deleted_peak_indices:
                        self.selected_peak_index = clicked_peak_coord
                        self.delete_peak_button.setEnabled(True)
                        print(f"Selected peak at index: {self.selected_peak_index}") # Debug
                        # Re-plot to show selection highlight
                        self.update_plot_highlights() # Separate function for replotting highlights
                    else:
                        # Clicked on a potentially invalid/deleted point marker
                        self.selected_peak_index = -1
                        self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()
                else:
                     # Index out of bounds (shouldn't normally happen)
                     self.selected_peak_index = -1
                     self.delete_peak_button.setEnabled(False)
                     self.update_plot_highlights()

            def update_plot_highlights(self):
                """Redraws only the peak markers and highlights without full recalculation."""
                if not hasattr(self, 'ax_profile'): return

                # Clear only the peak markers and highlights from the profile axis
                # Keep the main profile line
                artists_to_remove = []
                for artist in self.ax_profile.lines + self.ax_profile.collections:
                    label = artist.get_label()
                    if label and ('Peaks' in label or 'Selected' in label):
                         artists_to_remove.append(artist)
                for artist in artists_to_remove:
                    artist.remove()

                # Re-plot the *currently active* detected peaks
                if len(self.detected_peaks) > 0:
                    valid_peaks = self.detected_peaks[(self.detected_peaks >= 0) & (self.detected_peaks < len(self.profile_original_inverted))]
                    if len(valid_peaks) > 0:
                        peak_y_values = self.profile_original_inverted[valid_peaks]
                        # Re-store the artist for picking
                        self.peak_plot_artist, = self.ax_profile.plot(
                            valid_peaks, peak_y_values, "rx", markersize=8,
                            label=f"Active Peaks ({len(valid_peaks)})", picker=True, pickradius=5)

                        # Re-apply highlight if a peak is selected
                        if self.selected_peak_index != -1 and self.selected_peak_index in valid_peaks:
                            idx_in_valid = np.where(valid_peaks == self.selected_peak_index)[0]
                            if len(idx_in_valid) > 0:
                                self.ax_profile.plot(self.selected_peak_index, peak_y_values[idx_in_valid[0]],
                                                     'o', markersize=12, markeredgecolor='blue', markerfacecolor='none',
                                                     label='Selected') # Visual indicator

                # Update the legend and redraw
                handles, labels = self.ax_profile.get_legend_handles_labels()
                # Filter out duplicate labels if any were added accidentally
                by_label = dict(zip(labels, handles))
                self.ax_profile.legend(by_label.values(), by_label.keys(), fontsize='x-small')
                self.canvas.draw_idle()


            def delete_selected_peak(self):
                """Marks the selected peak as deleted and updates the plot."""
                if self.selected_peak_index != -1:
                    # Add the index to the set of deleted peaks
                    self.deleted_peak_indices.add(self.selected_peak_index)

                    # Update the list of currently active peaks
                    self.detected_peaks = np.array([p for p in self._all_initial_peaks if p not in self.deleted_peak_indices])

                    print(f"Deleted peak at index: {self.selected_peak_index}") # Debug

                    # Reset selection and disable button
                    self.selected_peak_index = -1
                    self.delete_peak_button.setEnabled(False)

                    # Re-plot to remove the deleted peak visually
                    self.update_plot_highlights() # Use highlight update for efficiency

            # MODIFY this method
            def accept_and_return_peaks(self):
                """Store final settings and accept the dialog."""
                # *** Settings are updated when sliders change now ***
                # We just need to ensure the final state is stored if needed elsewhere
                self._final_settings = {
                    'smoothing_sigma': self.smoothing_slider.value() / 10.0,
                    'peak_height_factor': self.peak_height_slider.value() / 100.0,
                    'peak_distance': self.peak_distance_slider.value(),
                    'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
                    'band_estimation_method': self.band_estimation_combobox.currentText(),
                    'rolling_ball_radius': self._final_settings.get('rolling_ball_radius', 50),
                    'area_subtraction_method': self._final_settings.get('area_subtraction_method', "Rolling Ball"),
                }
                self.accept()

            # MODIFY this method
            def get_detected_peaks(self):
                """Returns the FINAL list/array of detected peak Y-coordinates (indices), excluding user-deleted ones."""
                # Return the filtered list
                return self.detected_peaks

            def get_final_settings(self):
                """Returns the final detection settings dictionary."""
                return self._final_settings

            
        class ModifyMarkersDialog(QDialog):
            shapes_adjusted_preview = pyqtSignal(list) 
            global_markers_adjusted = pyqtSignal(list)
            def __init__(self, markers_list, shapes_list, parent=None): # shapes_list remains for future use
                super().__init__(parent)
                self.setWindowTitle("Modify Custom Markers and Shapes")
                self.setMinimumSize(950, 650) # Increased height for new controls
        
                self._original_markers_data = [] # Store pristine original data for scaling
                self.markers = [] # This will hold the currently modified data
        
                for marker_data in markers_list:
                    try:
                        marker_copy = list(marker_data) # x, y, text, qcolor, font_family, font_size, is_bold, is_italic
                        if len(marker_copy) == 6: marker_copy.extend([False, False]) # Old format
                        if len(marker_copy) == 8:
                            self._original_markers_data.append(tuple(marker_copy)) # Store original as tuple
                            self.markers.append(list(marker_copy)) # Working copy
                        else:
                            print(f"Warning: Skipping marker with invalid data length: {len(marker_copy)}")
                    except Exception as e:
                        print(f"Warning: Error processing marker data: {e}")
        
                # Shapes are not globally adjusted by these sliders for now, but keep for consistency
                self.shapes = []
                for shape_data in shapes_list:
                     if isinstance(shape_data, dict): self.shapes.append(dict(shape_data))
                     else: print(f"Warning: Skipping shape with invalid data type: {type(shape_data)}")
        
                self._block_signals = False
                self._current_image_width = parent.image.width() if parent and parent.image and not parent.image.isNull() else 1
                self._current_image_height = parent.image.height() if parent and parent.image and not parent.image.isNull() else 1
        
        
                # --- Main Layout ---
                layout = QVBoxLayout(self)
        
                # --- START: Global Adjustment Controls ---
                global_adjust_group = QGroupBox("Global Adjustments for Markers")
                global_adjust_layout = QGridLayout(global_adjust_group)
        
                # Define precision for percentage sliders (e.g., 100 means 2 decimal places for percent)
                self.percent_precision_factor = 100.0 # For displaying XX.YY%
                self.scale_precision_factor = 10.0    # For displaying XXX.Y% for scale sliders
        
                # Absolute X Shift Slider
                global_adjust_layout.addWidget(QLabel("Shift X (% Img W):"), 0, 0)
                self.abs_x_shift_slider = QSlider(Qt.Horizontal)
                # Range: -100.00% to +100.00% -> -10000 to +10000 if precision_factor is 100
                self.abs_x_shift_slider.setRange(int(-100 * self.percent_precision_factor), 
                                                 int(100 * self.percent_precision_factor)) 
                self.abs_x_shift_slider.setValue(0)
                self.abs_x_shift_slider.setToolTip("Shift all markers horizontally by a percentage of image width.")
                self.abs_x_shift_label = QLabel("0.00%")
                self.abs_x_shift_label.setFixedSize(80, 20)
                self.abs_x_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_x_shift_slider.valueChanged.connect(
                    lambda val, lbl=self.abs_x_shift_label: lbl.setText(f"{val / self.percent_precision_factor:.2f}%")
                )
                global_adjust_layout.addWidget(self.abs_x_shift_slider, 0, 1)
                global_adjust_layout.addWidget(self.abs_x_shift_label, 0, 2)
        
                # Absolute Y Shift Slider
                global_adjust_layout.addWidget(QLabel("Shift Y (% Img H):"), 1, 0)
                self.abs_y_shift_slider = QSlider(Qt.Horizontal)
                self.abs_y_shift_slider.setRange(int(-100 * self.percent_precision_factor), 
                                                 int(100 * self.percent_precision_factor))
                self.abs_y_shift_slider.setValue(0)
                self.abs_y_shift_slider.setToolTip("Shift all markers vertically by a percentage of image height.")
                self.abs_y_shift_label = QLabel("0.00%")
                self.abs_y_shift_label.setFixedSize(80, 20)
                self.abs_y_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_y_shift_slider.valueChanged.connect(
                    lambda val, lbl=self.abs_y_shift_label: lbl.setText(f"{val / self.percent_precision_factor:.2f}%")
                )
                global_adjust_layout.addWidget(self.abs_y_shift_slider, 1, 1)
                global_adjust_layout.addWidget(self.abs_y_shift_label, 1, 2)
        
                # Relative X Scale Slider
                global_adjust_layout.addWidget(QLabel("Scale X (%):"), 2, 0)
                self.rel_x_scale_slider = QSlider(Qt.Horizontal)
                # Range: 10.0% to 300.0% -> 100 to 3000 if scale_precision_factor is 10
                self.rel_x_scale_slider.setRange(int(10 * self.scale_precision_factor), 
                                                 int(300 * self.scale_precision_factor)) 
                self.rel_x_scale_slider.setValue(int(100 * self.scale_precision_factor)) # Default 100.0%
                self.rel_x_scale_slider.setToolTip("Scale markers' X positions relative to their original layout.")
                self.rel_x_scale_label = QLabel("100.0%")
                self.rel_x_scale_label.setFixedSize(80, 20)
                self.rel_x_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_x_scale_slider.valueChanged.connect(
                    lambda val, lbl=self.rel_x_scale_label: lbl.setText(f"{val / self.scale_precision_factor:.1f}%")
                )
                global_adjust_layout.addWidget(self.rel_x_scale_slider, 2, 1)
                global_adjust_layout.addWidget(self.rel_x_scale_label, 2, 2)
        
                # Relative Y Scale Slider
                global_adjust_layout.addWidget(QLabel("Scale Y (%):"), 3, 0)
                self.rel_y_scale_slider = QSlider(Qt.Horizontal)
                self.rel_y_scale_slider.setRange(int(10 * self.scale_precision_factor), 
                                                 int(300 * self.scale_precision_factor))
                self.rel_y_scale_slider.setValue(int(100 * self.scale_precision_factor)) # Default 100.0%
                self.rel_y_scale_slider.setToolTip("Scale markers' Y positions relative to their original layout.")
                self.rel_y_scale_label = QLabel("100.0%")
                self.rel_y_scale_label.setFixedSize(80, 20)
                self.rel_y_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_y_scale_slider.valueChanged.connect(
                    lambda val, lbl=self.rel_y_scale_label: lbl.setText(f"{val / self.scale_precision_factor:.1f}%")
                )
                global_adjust_layout.addWidget(self.rel_y_scale_slider, 3, 1)
                global_adjust_layout.addWidget(self.rel_y_scale_label, 3, 2)
        
                # Font Size Scale Slider
                global_adjust_layout.addWidget(QLabel("Font Scale (%):"), 4, 0)
                self.font_scale_slider = QSlider(Qt.Horizontal)
                # Range 10.0% to 300.0% -> 100 to 3000 if scale_precision_factor is 10
                self.font_scale_slider.setRange(int(10 * self.scale_precision_factor), 
                                                int(300 * self.scale_precision_factor))
                self.font_scale_slider.setValue(int(100 * self.scale_precision_factor)) # Default 100.0%
                self.font_scale_slider.setToolTip("Scale font size of all markers.")
                self.font_scale_label = QLabel("100.0%")
                self.font_scale_label.setFixedSize(80, 20)
                self.font_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.font_scale_slider.valueChanged.connect(
                    lambda val, lbl=self.font_scale_label: lbl.setText(f"{val / self.scale_precision_factor:.1f}%")
                )
                global_adjust_layout.addWidget(self.font_scale_slider, 4, 1)
                global_adjust_layout.addWidget(self.font_scale_label, 4, 2)
                
                global_adjust_layout.setColumnStretch(1, 1)
                layout.addWidget(global_adjust_group)
                # --- END: Global Adjustment Controls ---
        
                # --- Table Widget ---
                self.table_widget = QTableWidget()
                self.table_widget.setColumnCount(8)
                self.table_widget.setHorizontalHeaderLabels([
                    "Type", "Text/Label", "Coordinates", "Style", "Bold", "Italic", "Color", "Actions"
                ])
                self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
                self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
                self.table_widget.setEditTriggers(QAbstractItemView.AnyKeyPressed | QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
                self.table_widget.setSortingEnabled(True) # Enable after initial population if needed
        
                self.table_widget.itemChanged.connect(self.handle_item_changed)
                self.table_widget.cellDoubleClicked.connect(self.handle_cell_double_clicked)
        
                layout.addWidget(self.table_widget) # Table below global controls
                self.populate_table() # Initial population
                self.table_widget.resizeColumnsToContents()
                self.table_widget.setColumnWidth(0, 70)
                self.table_widget.setColumnWidth(1, 180)
                self.table_widget.setColumnWidth(2, 180)
                self.table_widget.setColumnWidth(3, 150)
                self.table_widget.setColumnWidth(4, 40)
                self.table_widget.setColumnWidth(5, 40)
                self.table_widget.setColumnWidth(6, 80)
                self.table_widget.setColumnWidth(7, 80)
                self.table_widget.horizontalHeader().setStretchLastSection(False)
        
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept) # Default accept is fine now
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box)
                self.setLayout(layout)
        
            def _update_global_adjustments(self):
                if self._block_signals or not self._original_markers_data:
                    return
        
                # Get values from all sliders and divide by precision factors
                abs_x_shift_percent = self.abs_x_shift_slider.value() / self.percent_precision_factor
                abs_y_shift_percent = self.abs_y_shift_slider.value() / self.percent_precision_factor
                rel_x_scale_percent = self.rel_x_scale_slider.value() / self.scale_precision_factor
                rel_y_scale_percent = self.rel_y_scale_slider.value() / self.scale_precision_factor
                font_scale_percent = self.font_scale_slider.value() / self.scale_precision_factor # Using scale_precision_factor
        
                # Calculate pixel shifts for absolute translation
                img_width = self._current_image_width if self._current_image_width > 0 else 1
                img_height = self._current_image_height if self._current_image_height > 0 else 1
                
                abs_x_shift_pixels = (abs_x_shift_percent / 100.0) * img_width
                abs_y_shift_pixels = (abs_y_shift_percent / 100.0) * img_height
        
                # Calculate scale factors (percentage / 100)
                rel_x_scale_factor = rel_x_scale_percent / 100.0
                rel_y_scale_factor = rel_y_scale_percent / 100.0
                font_size_scale_factor = font_scale_percent / 100.0
        
                for i, original_marker_tuple in enumerate(self._original_markers_data):
                    if i >= len(self.markers):
                        continue
        
                    orig_x, orig_y, text, qcolor, font_family, orig_font_size, is_bold, is_italic = original_marker_tuple
        
                    scaled_x = orig_x * rel_x_scale_factor
                    scaled_y = orig_y * rel_y_scale_factor
                    final_x = scaled_x + abs_x_shift_pixels
                    final_y = scaled_y + abs_y_shift_pixels
                    final_font_size = max(1, int(round(orig_font_size * font_size_scale_factor)))
        
                    self.markers[i][0] = final_x
                    self.markers[i][1] = final_y
                    self.markers[i][5] = final_font_size
        
                self.populate_table()
                self.global_markers_adjusted.emit(list(self.markers))
        
        
            def populate_table(self):
                self._block_signals = True
                self.table_widget.setSortingEnabled(False) # Disable sorting during population
                self.table_widget.setRowCount(0)
                
                total_items = len(self.markers) + len(self.shapes) # markers is now the working copy
                self.table_widget.setRowCount(total_items)
                
                current_row_idx = 0
        
                # --- Populate Markers (using self.markers, which reflects global adjustments) ---
                for marker_idx, marker_data_working in enumerate(self.markers):
                    row_idx = current_row_idx
                    try:
                        # marker_data_working is already a list: [x, y, text, qcolor, font_family, font_size, is_bold, is_italic]
                        x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_data_working
                        
                        if not isinstance(qcolor, QColor): qcolor = QColor(qcolor)
                        if not qcolor.isValid(): raise ValueError("Invalid marker color")
        
                        type_item = QTableWidgetItem("Marker")
                        # Store the index relative to the self.markers list (and _original_markers_data)
                        type_item.setData(Qt.UserRole, {'type': 'marker', 'original_index': marker_idx})
        
                        text_item = QTableWidgetItem(str(text))
                        text_item.setFlags(text_item.flags() | Qt.ItemIsEditable)
        
                        coord_str = f"{x:.1f},{y:.1f}" # Display the currently calculated coordinates
                        coord_item = QTableWidgetItem(coord_str)
                        coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable)
                        coord_item.setToolTip("Edit format: X,Y (e.g., 100.5,250.2)")
        
                        style_item = QTableWidgetItem(f"{font_family} ({font_size}pt)") # Display current font size
                        style_item.setToolTip("Double-click to change font/size for this specific marker")
                        style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)
        
                        color_item = QTableWidgetItem(qcolor.name())
                        color_item.setBackground(QBrush(qcolor))
                        text_color = Qt.white if qcolor.lightness() < 128 else Qt.black
                        color_item.setForeground(QBrush(text_color))
                        color_item.setToolTip("Double-click to change color for this specific marker")
                        color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
        
                        self.table_widget.setItem(row_idx, 0, type_item)
                        self.table_widget.setItem(row_idx, 1, text_item)
                        self.table_widget.setItem(row_idx, 2, coord_item)
                        self.table_widget.setItem(row_idx, 3, style_item)
                        self.table_widget.setItem(row_idx, 6, color_item)
        
                        bold_checkbox = QCheckBox(); bold_checkbox.setChecked(bool(is_bold))
                        bold_checkbox.stateChanged.connect(lambda state, r=row_idx: self.handle_marker_style_changed(state, r, "bold"))
                        cell_widget_bold = QWidget(); layout_bold = QHBoxLayout(cell_widget_bold); layout_bold.addWidget(bold_checkbox); layout_bold.setAlignment(Qt.AlignCenter); layout_bold.setContentsMargins(0,0,0,0); cell_widget_bold.setLayout(layout_bold)
                        self.table_widget.setCellWidget(row_idx, 4, cell_widget_bold)
        
                        italic_checkbox = QCheckBox(); italic_checkbox.setChecked(bool(is_italic))
                        italic_checkbox.stateChanged.connect(lambda state, r=row_idx: self.handle_marker_style_changed(state, r, "italic"))
                        cell_widget_italic = QWidget(); layout_italic = QHBoxLayout(cell_widget_italic); layout_italic.addWidget(italic_checkbox); layout_italic.setAlignment(Qt.AlignCenter); layout_italic.setContentsMargins(0,0,0,0); cell_widget_italic.setLayout(layout_italic)
                        self.table_widget.setCellWidget(row_idx, 5, cell_widget_italic)
        
                        delete_button = QPushButton("Delete"); delete_button.setStyleSheet("QPushButton { padding: 2px 5px; }")
                        # Connect delete to a lambda that uses the current row_idx,
                        # the delete_item function will then map this row to original_index if needed.
                        delete_button.clicked.connect(lambda checked, r=row_idx: self.delete_item(r))
                        self.table_widget.setCellWidget(row_idx, 7, delete_button)
        
                        current_row_idx += 1
                    except (ValueError, IndexError, TypeError) as e:
                        print(f"Error populating table for marker at original index {marker_idx}: {e}")
                        error_item = QTableWidgetItem("Marker Error")
                        error_item.setData(Qt.UserRole, {'type': 'error', 'original_index': -1}) # Use original_index
                        self.table_widget.setItem(row_idx, 0, error_item)
                        for col in range(1, self.table_widget.columnCount()):
                            placeholder = QTableWidgetItem("---"); placeholder.setFlags(placeholder.flags() & ~Qt.ItemIsEditable)
                            self.table_widget.setItem(row_idx, col, placeholder)
                        current_row_idx += 1
                
                # --- Populate Shapes (No global adjustments for shapes in this iteration) ---
                for shape_idx, shape_data in enumerate(self.shapes):
                    # ... (rest of shape population logic remains the same, using original_index for shapes) ...
                    row_idx = current_row_idx
                    try:
                        shape_type = shape_data.get('type', 'Unknown').capitalize()
                        color_name = shape_data.get('color', '#000000')
                        thickness = int(shape_data.get('thickness', 1))
                        qcolor = QColor(color_name);
                        if not qcolor.isValid(): raise ValueError(f"Invalid shape color: {color_name}")
                        if thickness < 1: raise ValueError("Thickness must be >= 1")
        
                        type_item = QTableWidgetItem(shape_type)
                        # Store index relative to self.shapes list
                        type_item.setData(Qt.UserRole, {'type': 'shape', 'original_index': shape_idx})
        
        
                        text_item = QTableWidgetItem("") 
                        text_item.setFlags(text_item.flags() & ~Qt.ItemIsEditable)
        
                        details_str = ""
                        tooltip_str = "Edit format: "
                        if shape_type == 'Line':
                            start = shape_data.get('start', (0,0)); end = shape_data.get('end', (0,0))
                            details_str = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"
                            tooltip_str += "X1,Y1,X2,Y2"
                        elif shape_type == 'Rectangle':
                            rect = shape_data.get('rect', (0,0,0,0)) 
                            details_str = f"{rect[0]:.1f},{rect[1]:.1f},{rect[2]:.1f},{rect[3]:.1f}"
                            tooltip_str += "X,Y,W,H"
                        else:
                            details_str = "N/A"
                            tooltip_str = "Cannot edit coordinates for this shape type."
        
                        coord_item = QTableWidgetItem(details_str)
                        coord_item.setToolTip(tooltip_str)
                        if shape_type in ['Line', 'Rectangle']:
                            coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable)
                        else:
                            coord_item.setFlags(coord_item.flags() & ~Qt.ItemIsEditable)
        
                        style_item = QTableWidgetItem(f"{thickness}px")
                        style_item.setToolTip("Double-click to change thickness")
                        style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)
        
                        color_item = QTableWidgetItem(qcolor.name())
                        color_item.setBackground(QBrush(qcolor)); text_color_shape = Qt.white if qcolor.lightness() < 128 else Qt.black
                        color_item.setForeground(QBrush(text_color_shape)); color_item.setToolTip("Double-click to change color")
                        color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
        
                        self.table_widget.setItem(row_idx, 0, type_item)
                        self.table_widget.setItem(row_idx, 1, text_item)    
                        self.table_widget.setItem(row_idx, 2, coord_item)   
                        self.table_widget.setItem(row_idx, 3, style_item)   
                        self.table_widget.setItem(row_idx, 6, color_item)   
        
                        delete_button_shape = QPushButton("Delete"); delete_button_shape.setStyleSheet("QPushButton { padding: 2px 5px; }")
                        delete_button_shape.clicked.connect(lambda checked, r=row_idx: self.delete_item(r))
                        self.table_widget.setCellWidget(row_idx, 7, delete_button_shape) 
        
                        current_row_idx += 1
                    except (ValueError, IndexError, TypeError, KeyError) as e:
                        print(f"Error processing shape data at original index {shape_idx}: {e}")
                        error_item_shape = QTableWidgetItem("Shape Error")
                        error_item_shape.setData(Qt.UserRole, {'type': 'error', 'original_index': -1})
                        self.table_widget.setItem(row_idx, 0, error_item_shape)
                        for col in range(1, self.table_widget.columnCount()):
                            placeholder_shape = QTableWidgetItem("---"); placeholder_shape.setFlags(placeholder_shape.flags() & ~Qt.ItemIsEditable)
                            self.table_widget.setItem(row_idx, col, placeholder_shape)
                        current_row_idx += 1
        
        
                if current_row_idx < total_items:
                    self.table_widget.setRowCount(current_row_idx)
                
                self.table_widget.setSortingEnabled(True) # Re-enable sorting
                self._block_signals = False
        
            def handle_item_changed(self, item):
                """Update internal lists when an editable cell (Text/Label or Coords) changes.
                   If coordinates are edited manually, this will update the 'current' state,
                   and subsequent global slider adjustments will be based on this new 'current' state
                   as if it were the original for that specific item."""
                if self._block_signals: return
        
                row = item.row()
                col = item.column()
        
                type_item = self.table_widget.item(row, 0)
                if not type_item: return
                item_data = type_item.data(Qt.UserRole)
                if not item_data or item_data.get('type') == 'error': return
        
                item_type = item_data['type']
                original_item_index = item_data.get('original_index') # This is key, index into _original_markers_data or self.shapes
                
                new_value_str = item.text()
                revert_needed = False
                error_message = ""
                revert_text = ""
        
                # --- Handle Marker Text Changes (Column 1) ---
                if col == 1 and item_type == 'marker':
                    if not (0 <= original_item_index < len(self.markers)): return
                    prev_text = self.markers[original_item_index][2]
                    try:
                        new_text = new_value_str.strip()
                        self.markers[original_item_index][2] = new_text
                        # Also update the corresponding _original_markers_data to reflect this manual edit
                        # This makes the manual edit the new "base" for future global scaling for THIS item
                        if 0 <= original_item_index < len(self._original_markers_data):
                            temp_list = list(self._original_markers_data[original_item_index])
                            temp_list[2] = new_text
                            self._original_markers_data[original_item_index] = tuple(temp_list)
                    except Exception as e:
                        revert_needed = True; revert_text = prev_text; error_message = f"Error updating text: {e}"
        
                # --- Handle Coordinate Changes (Column 2) ---
                elif col == 2:
                    if item_type == 'marker':
                        if not (0 <= original_item_index < len(self.markers)): return
                        
                        prev_x_display, prev_y_display = self.markers[original_item_index][0:2]
                        revert_text = f"{prev_x_display:.1f},{prev_y_display:.1f}"
    
                        try:
                            parts = new_value_str.split(',')
                            if len(parts) != 2: raise ValueError("Expected X,Y format")
                            new_x_manual = float(parts[0].strip())
                            new_y_manual = float(parts[1].strip())
    
                            # Update the working copy first (what's displayed)
                            self.markers[original_item_index][0] = new_x_manual
                            self.markers[original_item_index][1] = new_y_manual
    
                            # Now, back-calculate the "true original" for _original_markers_data
                            # new_x_manual = (original_x * rel_x_scale_factor) + abs_x_shift_pixels
                            # So, original_x = (new_x_manual - abs_x_shift_pixels) / rel_x_scale_factor
                            if 0 <= original_item_index < len(self._original_markers_data):
                                temp_list = list(self._original_markers_data[original_item_index])
                                
                                # Get current global slider values and adjust by precision
                                abs_x_shift_percent_curr = self.abs_x_shift_slider.value() / self.percent_precision_factor
                                abs_y_shift_percent_curr = self.abs_y_shift_slider.value() / self.percent_precision_factor
                                rel_x_scale_percent_curr = self.rel_x_scale_slider.value() / self.scale_precision_factor
                                rel_y_scale_percent_curr = self.rel_y_scale_slider.value() / self.scale_precision_factor
    
                                img_width_curr = self._current_image_width if self._current_image_width > 0 else 1
                                img_height_curr = self._current_image_height if self._current_image_height > 0 else 1
    
                                abs_x_shift_pixels_curr = (abs_x_shift_percent_curr / 100.0) * img_width_curr
                                abs_y_shift_pixels_curr = (abs_y_shift_percent_curr / 100.0) * img_height_curr
                                
                                rel_x_scale_factor_curr = rel_x_scale_percent_curr / 100.0
                                rel_y_scale_factor_curr = rel_y_scale_percent_curr / 100.0
                                
                                base_x = (new_x_manual - abs_x_shift_pixels_curr) / rel_x_scale_factor_curr if rel_x_scale_factor_curr != 0 else (new_x_manual - abs_x_shift_pixels_curr) # Avoid division by zero, if scale is 0, offset is just from the shifted val
                                base_y = (new_y_manual - abs_y_shift_pixels_curr) / rel_y_scale_factor_curr if rel_y_scale_factor_curr != 0 else (new_y_manual - abs_y_shift_pixels_curr)
                                
                                temp_list[0] = base_x
                                temp_list[1] = base_y
                                self._original_markers_data[original_item_index] = tuple(temp_list)
                            
                        except (ValueError, IndexError) as e:
                            revert_needed = True; error_message = f"Could not parse marker coordinates:\n{e}\n\nExpected format: X,Y"
                    
                    elif item_type == 'shape': # Shapes are not affected by global sliders, direct update
                        if not (0 <= original_item_index < len(self.shapes)): return
                        shape_data = self.shapes[original_item_index]
                        shape_type_internal = shape_data.get('type')
                        # (Shape coordinate update logic remains the same as your existing code)
                        if shape_type_internal == 'line':
                            start = shape_data.get('start', (0,0)); end = shape_data.get('end', (0,0))
                            revert_text = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"
                        elif shape_type_internal == 'rectangle':
                            rect = shape_data.get('rect', (0,0,0,0))
                            revert_text = f"{rect[0]:.1f},{rect[1]:.1f},{rect[2]:.1f},{rect[3]:.1f}"
                        else: revert_text = "N/A"
                        try:
                            coords_str = new_value_str.split(',')
                            coords_float = [float(c.strip()) for c in coords_str]
                            coords_updated = False
                            if shape_type_internal == 'line' and len(coords_float) == 4:
                                x1, y1, x2, y2 = coords_float
                                shape_data['start'] = (x1, y1); shape_data['end'] = (x2, y2)
                                coords_updated = True
                            elif shape_type_internal == 'rectangle' and len(coords_float) == 4:
                                x, y, w, h = coords_float
                                if w >= 0 and h >= 0: shape_data['rect'] = (x, y, w, h); coords_updated = True
                                else: raise ValueError("Rectangle width/height must be non-negative.")
                            else: raise ValueError(f"Incorrect number of coordinates for {shape_type_internal.capitalize()} (expected 4).")
                            if not coords_updated: raise ValueError("Coordinate update failed.")
                        except (ValueError, IndexError) as e:
                            revert_needed = True
                            expected_format = "X1,Y1,X2,Y2" if shape_type_internal == 'line' else "X,Y,W,H"
                            error_message = (f"Could not parse shape coordinates:\n{e}\n\nExpected format for {shape_type_internal.capitalize()}: {expected_format}")
        
        
                if revert_needed:
                    QMessageBox.warning(self, "Invalid Input", error_message)
                    self._block_signals = True
                    item.setText(revert_text)
                    self._block_signals = False
                    
                if not revert_needed: # Only emit if the change was valid
                    self.shapes_adjusted_preview.emit(list(self.shapes))
                    self.global_markers_adjusted.emit(list(self.markers))
        
            def handle_marker_style_changed(self, state, row, style_type):
                """Update the bold/italic flag for a marker. Updates both working and original copy."""
                if self._block_signals: return
                type_item = self.table_widget.item(row, 0)
                if not type_item: return
                item_data = type_item.data(Qt.UserRole)
                if not item_data or item_data.get('type') != 'marker': return
                
                original_marker_index = item_data.get('original_index')
                if not isinstance(original_marker_index, int) or not (0 <= original_marker_index < len(self.markers)):
                    return
        
                is_checked = (state == Qt.Checked)
                try:
                    if style_type == "bold":
                        self.markers[original_marker_index][6] = is_checked
                        if 0 <= original_marker_index < len(self._original_markers_data):
                            temp_list = list(self._original_markers_data[original_marker_index])
                            temp_list[6] = is_checked
                            self._original_markers_data[original_marker_index] = tuple(temp_list)
                    elif style_type == "italic":
                        self.markers[original_marker_index][7] = is_checked
                        if 0 <= original_marker_index < len(self._original_markers_data):
                            temp_list = list(self._original_markers_data[original_marker_index])
                            temp_list[7] = is_checked
                            self._original_markers_data[original_marker_index] = tuple(temp_list)
                    self.global_markers_adjusted.emit(list(self.markers))
                    self.shapes_adjusted_preview.emit(list(self.shapes))
                    
                except IndexError:
                    print(f"Error: Index mismatch updating style for marker original index {original_marker_index}.")
        
        
            def handle_cell_double_clicked(self, row, column):
                if column in [1, 2]: return # Ignore Text/Coords columns for double-click non-edit actions
        
                type_item = self.table_widget.item(row, 0)
                if not type_item: return
                item_data = type_item.data(Qt.UserRole)
                if not item_data or item_data.get('type') == 'error': return
        
                item_type = item_data['type']
                original_item_index = item_data.get('original_index') # Index into self.markers / self._original_markers_data
        
                # Color Change (Column 6)
                if column == 6:
                    current_qcolor_to_edit = None
                    if item_type == 'marker':
                        if not (0 <= original_item_index < len(self.markers)): return
                        current_qcolor_to_edit = self.markers[original_item_index][3] # Get from working copy
                    elif item_type == 'shape':
                        if not (0 <= original_item_index < len(self.shapes)): return
                        current_qcolor_to_edit = QColor(self.shapes[original_item_index].get('color', '#000000'))
                    
                    if current_qcolor_to_edit:
                        new_color = QColorDialog.getColor(current_qcolor_to_edit, self, f"Select {item_type.capitalize()} Color")
                        if new_color.isValid():
                            if item_type == 'marker':
                                self.markers[original_item_index][3] = new_color
                                if 0 <= original_item_index < len(self._original_markers_data): # Update original too
                                    temp_list = list(self._original_markers_data[original_item_index])
                                    temp_list[3] = new_color
                                    self._original_markers_data[original_item_index] = tuple(temp_list)
                            elif item_type == 'shape':
                                self.shapes[original_item_index]['color'] = new_color.name()
                            
                            self._block_signals = True # Update table display
                            color_item = self.table_widget.item(row, 6); color_item.setText(new_color.name())
                            color_item.setBackground(QBrush(new_color));
                            text_color = Qt.white if new_color.lightness() < 128 else Qt.black
                            color_item.setForeground(QBrush(text_color))
                            self._block_signals = False
                            self.shapes_adjusted_preview.emit(list(self.shapes))
                            self.global_markers_adjusted.emit(list(self.markers))
        
                # Font/Size Change for Markers (Column 3)
                elif column == 3 and item_type == 'marker':
                    if not (0 <= original_item_index < len(self.markers)): return
                    _, _, _, _, font_family, current_display_font_size, is_bold, is_italic = self.markers[original_item_index]
                    initial_qfont = QFont(font_family, current_display_font_size); initial_qfont.setBold(is_bold); initial_qfont.setItalic(is_italic)
                    
                    selected_font, ok = QFontDialog.getFont(initial_qfont, self, "Select Marker Font")
                    if ok:
                        # Update working copy with the new font selected by the user
                        self.markers[original_item_index][4] = selected_font.family()
                        self.markers[original_item_index][5] = selected_font.pointSize() # This is the new displayed size
                        self.markers[original_item_index][6] = selected_font.bold()
                        self.markers[original_item_index][7] = selected_font.italic()
    
                        # Update the _original_markers_data to reflect this manual edit
                        # The "original" font size should be the one that, when scaled by the global
                        # font_scale_slider, results in selected_font.pointSize().
                        current_font_scale_factor = self.font_scale_slider.value() / 100.0
                        
                        if 0 <= original_item_index < len(self._original_markers_data):
                            temp_list = list(self._original_markers_data[original_item_index])
                            temp_list[4] = selected_font.family()
                            # Store the "unscaled" base font size
                            base_font_size = int(round(selected_font.pointSize() / current_font_scale_factor)) if current_font_scale_factor > 0 else selected_font.pointSize()
                            temp_list[5] = max(1, base_font_size) # Ensure original font size is at least 1
                            temp_list[6] = selected_font.bold()
                            temp_list[7] = selected_font.italic()
                            self._original_markers_data[original_item_index] = tuple(temp_list)
    
                        # Update table display (which shows self.markers values)
                        self._block_signals = True
                        self.table_widget.item(row, 3).setText(f"{selected_font.family()} ({selected_font.pointSize()}pt)")
                        bold_widget = self.table_widget.cellWidget(row, 4); italic_widget = self.table_widget.cellWidget(row, 5)
                        if bold_widget: bold_widget.findChild(QCheckBox).setChecked(selected_font.bold())
                        if italic_widget: italic_widget.findChild(QCheckBox).setChecked(selected_font.italic())
                        self._block_signals = False
                        self.global_markers_adjusted.emit(list(self.markers))
        
                # Thickness Change for Shapes (Column 3)
                elif column == 3 and item_type == 'shape':
                     if not (0 <= original_item_index < len(self.shapes)): return
                     current_thickness = self.shapes[original_item_index].get('thickness', 1)
                     new_thickness, ok = QInputDialog.getInt(self, "Set Thickness", "Enter line/border thickness (pixels):", current_thickness, 1, 100, 1)
                     if ok:
                         self.shapes[original_item_index]['thickness'] = new_thickness
                         self._block_signals = True
                         self.table_widget.item(row, 3).setText(f"{new_thickness}px")
                         self._block_signals = False
                         self.shapes_adjusted_preview.emit(list(self.shapes))
        
            def delete_item(self, row_to_delete):
                if not (0 <= row_to_delete < self.table_widget.rowCount()): return
                type_item = self.table_widget.item(row_to_delete, 0)
                if not type_item: return
                item_data = type_item.data(Qt.UserRole)
                if not item_data or item_data.get('type') == 'error': return
                
                item_type = item_data['type']
                # original_index refers to the index in self.markers or self.shapes
                # (and correspondingly self._original_markers_data)
                original_index_in_list = item_data['original_index'] 
                
                sort_col = self.table_widget.horizontalHeader().sortIndicatorSection()
                sort_order = self.table_widget.horizontalHeader().sortIndicatorOrder()
                self.table_widget.setSortingEnabled(False)
        
                item_deleted_from_list = False
                if item_type == 'marker':
                    if 0 <= original_index_in_list < len(self.markers):
                        del self.markers[original_index_in_list]
                        # Also delete from _original_markers_data to keep them in sync
                        if 0 <= original_index_in_list < len(self._original_markers_data):
                            del self._original_markers_data[original_index_in_list]
                        item_deleted_from_list = True
                    else: print(f"Warning: Stale original index {original_index_in_list} for marker deletion.")
                elif item_type == 'shape':
                    if 0 <= original_index_in_list < len(self.shapes):
                        del self.shapes[original_index_in_list]
                        item_deleted_from_list = True
                    else: print(f"Warning: Stale original index {original_index_in_list} for shape deletion.")
        
                if item_deleted_from_list:
                    self.table_widget.removeRow(row_to_delete)
                    # --- Re-index UserRole data ('original_index') and reconnect signals ---
                    for current_row_in_table in range(self.table_widget.rowCount()):
                        current_type_item = self.table_widget.item(current_row_in_table, 0)
                        if not current_type_item: continue
                        current_item_data = current_type_item.data(Qt.UserRole)
                        if not current_item_data or current_item_data.get('type') == 'error': continue
                        
                        current_item_type = current_item_data['type']
                        current_original_idx_stored = current_item_data['original_index']
                        
                        new_original_idx_to_store = -1
                        if current_item_type == item_type: # If it's the same type as the one deleted
                            if current_original_idx_stored > original_index_in_list:
                                new_original_idx_to_store = current_original_idx_stored - 1
                            else:
                                new_original_idx_to_store = current_original_idx_stored
                        else: # Different type, index is unaffected by this deletion
                            new_original_idx_to_store = current_original_idx_stored
                        
                        current_type_item.setData(Qt.UserRole, {'type': current_item_type, 'original_index': new_original_idx_to_store})
        
                        # Reconnect signals for the widgets in the current row
                        delete_button_widget = self.table_widget.cellWidget(current_row_in_table, 7)
                        if isinstance(delete_button_widget, QPushButton):
                            try: delete_button_widget.clicked.disconnect()
                            except TypeError: pass
                            delete_button_widget.clicked.connect(lambda checked, r=current_row_in_table: self.delete_item(r))
        
                        if current_item_type == 'marker':
                            bold_cell_widget = self.table_widget.cellWidget(current_row_in_table, 4)
                            if bold_cell_widget:
                                bold_checkbox = bold_cell_widget.findChild(QCheckBox)
                                if bold_checkbox:
                                    try: bold_checkbox.stateChanged.disconnect()
                                    except TypeError: pass
                                    bold_checkbox.stateChanged.connect(lambda state, r=current_row_in_table: self.handle_marker_style_changed(state, r, "bold"))
                            italic_cell_widget = self.table_widget.cellWidget(current_row_in_table, 5)
                            if italic_cell_widget:
                                italic_checkbox = italic_cell_widget.findChild(QCheckBox)
                                if italic_checkbox:
                                    try: italic_checkbox.stateChanged.disconnect()
                                    except TypeError: pass
                                    italic_checkbox.stateChanged.connect(lambda state, r=current_row_in_table: self.handle_marker_style_changed(state, r, "italic"))
                    # --- End Re-indexing/Reconnecting ---
                else: print(f"Warning: Could not delete item from internal list for row {row_to_delete}.")
        
                self.table_widget.setSortingEnabled(True)
                if sort_col >= 0: self.table_widget.sortByColumn(sort_col, sort_order)
        
            def get_modified_markers_and_shapes(self):
                """Returns the modified lists of markers (already reflecting global adjustments) and shapes."""
                # self.markers already contains the final state after global adjustments and individual edits
                final_markers = [tuple(m) for m in self.markers] # Convert back to list of tuples
                return final_markers, self.shapes # self.shapes is not globally adjusted by these sliders

        class TableWindow(QDialog):
            """
            A dialog window to display peak analysis results in a table,
            manage a history of previous analyses, and compare results.
            Includes functionality to copy selected data and export to Excel.
            """
            HISTORY_FILE_NAME = "analysis_history.json"

            def __init__(self, current_peak_areas, current_standard_dictionary, current_is_standard_mode, current_calculated_quantities, parent_app_instance=None):
                super().__init__(parent_app_instance)
                self.setWindowTitle("Analysis Results and History")
                self.setGeometry(100, 100, 850, 700)

                self.parent_app = parent_app_instance

                self.current_peak_areas = current_peak_areas if current_peak_areas is not None else []
                self.current_standard_dictionary = current_standard_dictionary if current_standard_dictionary is not None else {}
                self.current_is_standard_mode = current_is_standard_mode
                self.current_calculated_quantities = current_calculated_quantities if current_calculated_quantities is not None else []
                self.source_image_name_current = "Unknown"
                if self.parent_app and hasattr(self.parent_app, 'image_path') and self.parent_app.image_path:
                    # Ensure image_path is a string before calling os.path.basename
                    if isinstance(self.parent_app.image_path, str):
                        self.source_image_name_current = os.path.basename(self.parent_app.image_path)
                    else:
                        print(f"Warning: parent_app.image_path is not a string: {self.parent_app.image_path}")

                self.current_analysis_custom_name = self.source_image_name_current # Initialize with image name
                self.analysis_name_input_widget = None
                
                self.analysis_history = []
                self.delete_entry_button = None
                self.export_previous_button = None
                self.previous_sessions_listwidget = None
                self.previous_results_table = None
                self.previous_plot_placeholder_label = None
                self.previous_plot_groupbox_layout = None
                self.previous_plot_canvas_widget = None

                self._load_history() # Load history first

                main_layout = QVBoxLayout(self)
                self.tab_widget = QTabWidget()
                main_layout.addWidget(self.tab_widget)

                self._create_current_results_tab()
                self._create_previous_results_tab() # This also calls _populate_previous_sessions_list

                self.dialog_button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                self.dialog_button_box.accepted.connect(self._accept_dialog_and_save_current) # Correct connection
                self.dialog_button_box.rejected.connect(self.reject)
                main_layout.addWidget(self.dialog_button_box)

                self.setLayout(main_layout)

                # Determine initial tab based on whether *new* current data was passed
                if self.current_peak_areas: # If there's fresh data for "Current Analysis"
                    self.tab_widget.setCurrentIndex(0)
                elif self.analysis_history: # Otherwise, if history exists, show it
                    self.tab_widget.setCurrentIndex(1)
                else: # Default to current tab if no new data and no history
                    self.tab_widget.setCurrentIndex(0)
                
                

            def _get_config_dir(self):
                """Determines the directory for storing configuration/history files."""
                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()
                return application_path

            def _load_history(self):
                """Loads analysis history from the JSON file."""
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                if os.path.exists(history_file_path):
                    try:
                        with open(history_file_path, "r", encoding='utf-8') as f:
                            loaded_data = json.load(f)
                        if isinstance(loaded_data, list):
                            self.analysis_history = [entry for entry in loaded_data if isinstance(entry, dict)]
                        else:
                            self.analysis_history = []
                    except (json.JSONDecodeError, IOError) as e:
                        self.analysis_history = []
                else:
                    self.analysis_history = []

            def _save_history(self):
                """Saves the current analysis_history list to the JSON file."""
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                try:
                    with open(history_file_path, "w", encoding='utf-8') as f:
                        json.dump(self.analysis_history, f, indent=4)
                except IOError as e:
                    QMessageBox.critical(self, "Save History Error", f"Could not save analysis history to {history_file_path}: {e}")

            def _accept_dialog_and_save_current(self):
                """Saves the current analysis results to history if they are new, then accepts the dialog."""
                if self.current_peak_areas:
                    peak_dialog_settings_current = {}
                    if self.parent_app and hasattr(self.parent_app, 'peak_dialog_settings'):
                        peak_dialog_settings_current = self.parent_app.peak_dialog_settings.copy()

                    # --- Get analysis name from the input widget ---
                    user_defined_analysis_name = ""
                    if self.analysis_name_input_widget: # Check if widget exists
                        user_defined_analysis_name = self.analysis_name_input_widget.text().strip()
                    
                    # Use user-defined name if provided, otherwise fallback to source image name
                    display_name_for_history = user_defined_analysis_name if user_defined_analysis_name else self.source_image_name_current
                    # --- End Get analysis name ---

                    new_entry = {
                        "timestamp": datetime.datetime.now().isoformat(),
                        "user_defined_name": display_name_for_history, # Store the potentially custom name
                        "source_image_name": self.source_image_name_current, # Still store original source
                        "peak_areas": self.current_peak_areas,
                        "calculated_quantities": self.current_calculated_quantities,
                        "standard_dictionary": self.current_standard_dictionary,
                        "analysis_settings": peak_dialog_settings_current
                    }
                    self.analysis_history.insert(0, new_entry)
                    self._save_history()
                    if self.previous_sessions_listwidget:
                        self._populate_previous_sessions_list()
                self.accept()

            def _create_current_results_tab(self):
                """Creates the tab for displaying current analysis results."""
                current_tab_widget = QWidget()
                current_layout = QVBoxLayout(current_tab_widget)

                # --- NEW: Analysis Name Input ---
                name_layout = QHBoxLayout()
                name_label = QLabel("Analysis Name:")
                self.analysis_name_input_widget = QLineEdit(self.current_analysis_custom_name)
                self.analysis_name_input_widget.setPlaceholderText("Enter a name for this analysis...")
                self.analysis_name_input_widget.setToolTip("This name will be used when saving to history and for export filenames.")
                # Optionally, connect textChanged if you want to update self.current_analysis_custom_name live
                # self.analysis_name_input_widget.textChanged.connect(lambda text: setattr(self, 'current_analysis_custom_name', text))
                name_layout.addWidget(name_label)
                name_layout.addWidget(self.analysis_name_input_widget)
                current_layout.addLayout(name_layout)

                current_plot_widget = self._create_standard_curve_plot_generic(
                    self.current_standard_dictionary,
                    self.current_is_standard_mode,
                    for_history=False
                )
                if current_plot_widget:
                    plot_group_current = QGroupBox("Standard Curve (Current Analysis)")
                    plot_layout_current = QVBoxLayout(plot_group_current)
                    plot_layout_current.addWidget(current_plot_widget)
                    plot_group_current.setMaximumHeight(300)
                    plot_group_current.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                    current_layout.addWidget(plot_group_current)

                scroll_area_current = QScrollArea(current_tab_widget)
                scroll_area_current.setWidgetResizable(True)
                self.current_results_table = QTableWidget()
                self.current_results_table.setColumnCount(4)
                self.current_results_table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
                self.current_results_table.setEditTriggers(QTableWidget.NoEditTriggers)
                self.current_results_table.setSelectionBehavior(QTableWidget.SelectRows)
                self.current_results_table.setSelectionMode(QTableWidget.ContiguousSelection)

                # Populate with current data *if it exists*
                if self.current_peak_areas:
                     self._populate_table_generic(self.current_results_table, self.current_peak_areas, self.current_is_standard_mode, self.current_calculated_quantities)
                else: # Display placeholder if no current data
                    self.current_results_table.setRowCount(1)
                    placeholder_item = QTableWidgetItem("No current analysis data to display.")
                    placeholder_item.setTextAlignment(Qt.AlignCenter)
                    self.current_results_table.setItem(0,0, placeholder_item)
                    self.current_results_table.setSpan(0,0,1,4)


                scroll_area_current.setWidget(self.current_results_table)
                current_layout.addWidget(scroll_area_current)

                current_buttons_layout = QHBoxLayout()
                copy_current_button = QPushButton("Copy Current Table")
                copy_current_button.clicked.connect(lambda: self._copy_table_data_generic(self.current_results_table))
                export_current_button = QPushButton("Export Current Table to Excel")
                # Prepare default filename for current export
                current_timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                default_filename_current = f"Analysis_{current_timestamp_str}_{self.source_image_name_current.replace('.', '_')}"

                export_current_button.clicked.connect(
                    lambda: self._export_to_excel_generic(
                        self.current_results_table,
                        self.analysis_name_input_widget.text() or self.source_image_name_current, # Pass current name
                        self.current_standard_dictionary
                    )
                )

                current_buttons_layout.addWidget(copy_current_button)
                current_buttons_layout.addStretch()
                current_buttons_layout.addWidget(export_current_button)
                current_layout.addLayout(current_buttons_layout)

                self.tab_widget.addTab(current_tab_widget, "Current Analysis")

            def _create_previous_results_tab(self):
                """Creates the tab for browsing and managing analysis history."""
                previous_tab_widget = QWidget()
                previous_main_layout = QHBoxLayout(previous_tab_widget)

                left_pane_widget = QWidget()
                left_layout = QVBoxLayout(left_pane_widget)
                left_layout.setContentsMargins(0, 0, 5, 0)

                left_layout.addWidget(QLabel("Saved Analyses:"))
                self.previous_sessions_listwidget = QListWidget()
                self.previous_sessions_listwidget.itemSelectionChanged.connect(self._on_history_session_selected)
                # _populate_previous_sessions_list is called after UI creation now in __init__
                left_layout.addWidget(self.previous_sessions_listwidget)

                history_buttons_layout = QHBoxLayout()
                self.delete_entry_button = QPushButton("Delete Selected")
                self.delete_entry_button.clicked.connect(self._delete_selected_history_entry)
                self.delete_entry_button.setEnabled(False)
                history_buttons_layout.addWidget(self.delete_entry_button)
                history_buttons_layout.addStretch()
                self.clear_history_button = QPushButton("Clear All History")
                self.clear_history_button.clicked.connect(self._clear_all_history)
                history_buttons_layout.addWidget(self.clear_history_button)
                left_layout.addLayout(history_buttons_layout)

                previous_main_layout.addWidget(left_pane_widget, 1)

                right_pane_widget = QWidget()
                right_layout = QVBoxLayout(right_pane_widget)

                self.previous_plot_groupbox = QGroupBox("Standard Curve (Selected History)")
                self.previous_plot_groupbox_layout = QVBoxLayout(self.previous_plot_groupbox)
                self.previous_plot_placeholder_label = QLabel("Select an analysis from the list to view details.")
                self.previous_plot_placeholder_label.setAlignment(Qt.AlignCenter)
                self.previous_plot_groupbox_layout.addWidget(self.previous_plot_placeholder_label)
                self.previous_plot_groupbox.setMaximumHeight(300)
                self.previous_plot_groupbox.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
                right_layout.addWidget(self.previous_plot_groupbox)

                self.previous_results_table = QTableWidget()
                self.previous_results_table.setColumnCount(4)
                self.previous_results_table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
                self.previous_results_table.setEditTriggers(QTableWidget.NoEditTriggers)
                right_layout.addWidget(self.previous_results_table)

                previous_table_buttons_layout = QHBoxLayout()
                copy_previous_button = QPushButton("Copy Selected History Table")
                copy_previous_button.clicked.connect(lambda: self._copy_table_data_generic(self.previous_results_table))
                self.export_previous_button = QPushButton("Export Selected History Table")
                self.export_previous_button.clicked.connect(self._export_selected_history_to_excel)
                self.export_previous_button.setEnabled(False)
                previous_table_buttons_layout.addWidget(copy_previous_button)
                previous_table_buttons_layout.addStretch()
                previous_table_buttons_layout.addWidget(self.export_previous_button)
                right_layout.addLayout(previous_table_buttons_layout)

                previous_main_layout.addWidget(right_pane_widget, 2)
                self.tab_widget.addTab(previous_tab_widget, "Analysis History")

                # Now that previous_sessions_listwidget is created, populate it
                self._populate_previous_sessions_list()

            def _populate_previous_sessions_list(self):
                self.previous_sessions_listwidget.clear()
                if not self.analysis_history:
                    self.previous_sessions_listwidget.addItem("No history available.")
                    if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                    if self.export_previous_button: self.export_previous_button.setEnabled(False)
                    return

                for i, entry in enumerate(self.analysis_history):
                    ts_str = entry.get("timestamp", f"Entry {len(self.analysis_history) - i}")
                    
                    # --- Use user_defined_name if available, otherwise fallback ---
                    entry_display_name = entry.get("user_defined_name", "").strip()
                    source_img_name = entry.get("source_image_name", "Unknown Image")
                    
                    if not entry_display_name: # If user_defined_name is empty or missing
                        entry_display_name = source_img_name # Fallback to source image name

                    try:
                        dt_obj = datetime.datetime.fromisoformat(ts_str.split('.')[0])
                        # Display format: "YYYY-MM-DD HH:MM:SS (User Defined Name OR Source Image Name)"
                        final_display_string = f"{dt_obj.strftime('%Y-%m-%d %H:%M:%S')} ({entry_display_name})"
                    except ValueError:
                        final_display_string = f"{ts_str} ({entry_display_name})"
                    # --- End Name Logic ---
                    self.previous_sessions_listwidget.addItem(final_display_string)
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)

            def _clear_previous_details_view(self):
                """Clears the right pane of the history tab."""
                if self.previous_results_table:
                    self.previous_results_table.setRowCount(0) # Clear table
                    # Add placeholder if empty
                    placeholder_item = QTableWidgetItem("Select an analysis to view details.")
                    placeholder_item.setTextAlignment(Qt.AlignCenter)
                    self.previous_results_table.setRowCount(1)
                    self.previous_results_table.setItem(0,0, placeholder_item)
                    self.previous_results_table.setSpan(0,0,1,4)


                if hasattr(self, 'previous_plot_canvas_widget') and self.previous_plot_canvas_widget and self.previous_plot_groupbox_layout:
                    self.previous_plot_groupbox_layout.removeWidget(self.previous_plot_canvas_widget)
                    self.previous_plot_canvas_widget.deleteLater()
                    self.previous_plot_canvas_widget = None
                
                if self.previous_plot_placeholder_label: # Check if it exists
                    self.previous_plot_placeholder_label.setText("Select an analysis from the list to view details.")
                    if self.previous_plot_placeholder_label.isHidden(): # Check if it's hidden
                        self.previous_plot_placeholder_label.show()
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)

            def _on_history_session_selected(self):
                """Handles selection change in the previous_sessions_listwidget."""
                if not self.previous_sessions_listwidget: return

                selected_items = self.previous_sessions_listwidget.selectedItems()
                if not selected_items or "No history available." in selected_items[0].text():
                    self._clear_previous_details_view()
                    return

                selected_row_index = self.previous_sessions_listwidget.currentRow()
                if 0 <= selected_row_index < len(self.analysis_history):
                    entry = self.analysis_history[selected_row_index]
                    if self.delete_entry_button: self.delete_entry_button.setEnabled(True)
                    if self.export_previous_button: self.export_previous_button.setEnabled(True)

                    hist_peak_areas = entry.get("peak_areas", [])
                    hist_std_dict = entry.get("standard_dictionary", {})
                    hist_is_std_mode = bool(hist_std_dict)
                    hist_calc_qty = entry.get("calculated_quantities", [])
                    self._populate_table_generic(self.previous_results_table, hist_peak_areas, hist_is_std_mode, hist_calc_qty)

                    if hasattr(self, 'previous_plot_canvas_widget') and self.previous_plot_canvas_widget and self.previous_plot_groupbox_layout:
                        self.previous_plot_groupbox_layout.removeWidget(self.previous_plot_canvas_widget)
                        self.previous_plot_canvas_widget.deleteLater()
                        self.previous_plot_canvas_widget = None
                    
                    if self.previous_plot_placeholder_label and not self.previous_plot_placeholder_label.isHidden():
                         self.previous_plot_placeholder_label.hide()

                    hist_plot_widget = self._create_standard_curve_plot_generic(hist_std_dict, hist_is_std_mode, for_history=True)
                    self.previous_plot_canvas_widget = hist_plot_widget
                    if self.previous_plot_groupbox_layout: # Ensure layout exists
                        self.previous_plot_groupbox_layout.addWidget(self.previous_plot_canvas_widget)
                else:
                    self._clear_previous_details_view()


            def _delete_selected_history_entry(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    reply = QMessageBox.question(self, "Confirm Delete",
                                                 "Are you sure you want to delete this history entry?",
                                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        del self.analysis_history[current_row]
                        self._save_history()
                        self._populate_previous_sessions_list()
                        self._clear_previous_details_view()
                else:
                    QMessageBox.information(self, "No Selection", "Please select an entry to delete.")

            def _clear_all_history(self):
                reply = QMessageBox.question(self, "Confirm Clear All History",
                                             "Are you sure you want to delete ALL saved analysis entries?\nThis action cannot be undone.",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    self.analysis_history = []
                    self._save_history()
                    self._populate_previous_sessions_list() # This will show "No history available."
                    self._clear_previous_details_view()

            def _export_selected_history_to_excel(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    entry = self.analysis_history[current_row]
                    ts_str_raw = entry.get("timestamp", f"History_Entry_{current_row+1}")
                    try:
                        dt_obj_export = datetime.datetime.fromisoformat(ts_str_raw.split('.')[0])
                        timestamp_str_for_file = dt_obj_export.strftime("%Y%m%d_%H%M%S")
                    except ValueError:
                        timestamp_str_for_file = ts_str_raw.replace(":", "-").replace("T", "_").split('.')[0]

                    # --- Use user_defined_name for filename if available ---
                    analysis_name_part = entry.get("user_defined_name", "").strip()
                    if not analysis_name_part: # Fallback if empty
                        analysis_name_part = entry.get("source_image_name", "UnknownAnalysis")
                    
                    # Sanitize the name part for filename
                    analysis_name_part_sanitized = analysis_name_part.replace('.', '_').replace(' ', '_').replace(':', '-')
                    # Limit length to avoid overly long filenames
                    max_name_len = 50
                    if len(analysis_name_part_sanitized) > max_name_len:
                         analysis_name_part_sanitized = analysis_name_part_sanitized[:max_name_len]

                    default_filename_base = f"Analysis_{timestamp_str_for_file}_{analysis_name_part_sanitized}"
                    # --- End Filename Logic ---

                    if self.previous_results_table:
                        self._export_to_excel_generic(
                            self.previous_results_table,
                            default_filename_base, # Pass the constructed base name
                            entry.get("standard_dictionary", {})
                        )
                    else:
                         QMessageBox.warning(self, "Error", "History table UI element not found.")
                else:
                    QMessageBox.information(self, "No Selection", "Please select a history entry to export.")

            # --- Generic Helper Methods (Keep these as they were, they are fine) ---
            def _create_standard_curve_plot_generic(self, standard_dictionary, is_standard_mode, for_history=False):
                if not is_standard_mode or not standard_dictionary or len(standard_dictionary) < 2:
                    no_curve_label = QLabel("Standard curve requires at least 2 standard points." if not for_history else "No standard data for this historical entry.")
                    no_curve_label.setAlignment(Qt.AlignCenter)
                    return no_curve_label
                try:
                    quantities = np.array(list(standard_dictionary.keys()), dtype=float)
                    areas = np.array(list(standard_dictionary.values()), dtype=float)
                    if len(quantities) < 2: # Double check after potential type conversion
                         no_curve_label = QLabel("Insufficient valid standard points (less than 2).")
                         no_curve_label.setAlignment(Qt.AlignCenter)
                         return no_curve_label

                    coeffs = np.polyfit(areas, quantities, 1); slope, intercept = coeffs
                    predicted_quantities = np.polyval(coeffs, areas); residuals = quantities - predicted_quantities
                    ss_res = np.sum(residuals**2); ss_tot = np.sum((quantities - np.mean(quantities))**2)
                    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0 # Avoid division by zero if all quantities are same
                    
                    fig, ax = plt.subplots(figsize=(4.5, 3.2)) # Slightly adjusted for dialog
                    fig.set_dpi(90)

                    ax.scatter(areas, quantities, label='Standard Points', color='red', zorder=5, s=30)
                    
                    # Ensure x_line covers the range of areas, handle empty/single point arrays
                    x_min_plot = np.min(areas) * 0.9 if areas.size > 0 else 0
                    x_max_plot = np.max(areas) * 1.1 if areas.size > 0 else 1
                    if x_min_plot == x_max_plot: # Handle case where all areas are the same
                        x_min_plot -= 0.5
                        x_max_plot += 0.5

                    x_line = np.linspace(x_min_plot, x_max_plot, 100)
                    y_line = slope * x_line + intercept
                    
                    fit_label = (f'Qty = {slope:.3g}*Area + {intercept:.3g}\nR² = {r_squared:.3f}')
                    ax.plot(x_line, y_line, label=fit_label, color='blue', linewidth=1.2)
                    
                    ax.set_xlabel('Total Peak Area', fontsize=8)
                    ax.set_ylabel('Known Quantity', fontsize=8)
                    title_prefix = "Historical " if for_history else "" # Removed "Current" as it's implied
                    ax.set_title(f'{title_prefix}Standard Curve', fontsize=9, fontweight='bold')
                    
                    ax.legend(fontsize='xx-small', loc='best'); ax.grid(True, linestyle=':', alpha=0.7, linewidth=0.5)
                    ax.tick_params(axis='both', which='major', labelsize=7)
                    
                    if np.any(areas > 1e4) or (np.any(areas < 1e-2) and np.any(areas != 0)): ax.ticklabel_format(style='sci', axis='x', scilimits=(0,0), useMathText=True)
                    if np.any(quantities > 1e4) or (np.any(quantities < 1e-2) and np.any(quantities != 0)): ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                    
                    try: fig.set_constrained_layout(True)
                    except AttributeError: plt.tight_layout(pad=0.3)
                    
                    canvas = FigureCanvas(fig)
                    canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding); canvas.updateGeometry()
                    plt.close(fig)

                    return canvas
                except Exception as e:
                    traceback.print_exc()
                    error_label = QLabel(f"Error generating plot:\n{str(e)[:100]}...") # Show truncated error
                    error_label.setAlignment(Qt.AlignCenter); error_label.setStyleSheet("color: red;")
                    return error_label

            def _populate_table_generic(self, table_widget, peak_areas, is_standard_mode, calculated_quantities):
                table_widget.clearContents()
                if not peak_areas: # If peak_areas is empty or None
                    table_widget.setRowCount(1)
                    placeholder_item = QTableWidgetItem("No data to display in table.")
                    placeholder_item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(0,0, placeholder_item)
                    table_widget.setSpan(0,0,1,table_widget.columnCount()) # Span across all columns
                    table_widget.resizeColumnsToContents()
                    return

                total_area = sum(peak_areas) if peak_areas else 0.0
                table_widget.setRowCount(len(peak_areas))
                for row, area in enumerate(peak_areas):
                    band_label = f"Band {row + 1}"
                    table_widget.setItem(row, 0, QTableWidgetItem(band_label))
                    table_widget.setItem(row, 1, QTableWidgetItem(f"{area:.3f}"))
                    percentage_str = f"{(area / total_area * 100):.2f}%" if total_area != 0 else "0.00%"
                    table_widget.setItem(row, 2, QTableWidgetItem(percentage_str))
                    quantity_str = ""
                    if is_standard_mode and calculated_quantities and row < len(calculated_quantities):
                        quantity_str = f"{calculated_quantities[row]:.2f}"
                    table_widget.setItem(row, 3, QTableWidgetItem(quantity_str))
                table_widget.resizeColumnsToContents()


            def _copy_table_data_generic(self, table_widget_source):
                if not table_widget_source: return
                selected_ranges = table_widget_source.selectedRanges()
                if not selected_ranges: return
                # ... (rest of copy logic is fine)
                selected_range = selected_ranges[0]
                start_row, end_row = selected_range.topRow(), selected_range.bottomRow()
                start_col, end_col = selected_range.leftColumn(), selected_range.rightColumn()
                clipboard_string = ""
                for r in range(start_row, end_row + 1):
                    row_data = []
                    for c in range(start_col, end_col + 1):
                        item = table_widget_source.item(r, c)
                        # Check if item is the placeholder span
                        if table_widget_source.rowSpan(r,c) > 1 or table_widget_source.columnSpan(r,c) > 1:
                            if "No data" in item.text() or "Select an analysis" in item.text():
                                row_data.append("") # Add empty string for placeholder cells
                                continue
                        row_data.append(item.text() if item else "")
                    if any(cell_text for cell_text in row_data): # Only add row if not all empty
                        clipboard_string += "\t".join(row_data) + "\n"
                QApplication.clipboard().setText(clipboard_string.strip())


            def _export_to_excel_generic(self, table_widget_source, analysis_name_for_filename_base="Analysis_Results", standard_dict_for_export=None):
                safe_analysis_name = str(analysis_name_for_filename_base).replace('.', '_').replace(' ', '_').replace(':', '-')
                current_timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Construct default filename using the (potentially custom) analysis name
                # If analysis_name_for_filename_base was already a full base name (like from history export), this is fine.
                # If it was just the analysis name part (from current export), prepend "Analysis_" and append timestamp.
                
                final_default_filename = ""
                if "Analysis_" in safe_analysis_name and any(char.isdigit() for char in safe_analysis_name): # Heuristic: likely already a full base name
                    final_default_filename = safe_analysis_name
                else: # Just the name part, construct full base
                    final_default_filename = f"Analysis_{current_timestamp_str}_{safe_analysis_name}"


                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Excel File", f"{final_default_filename}.xlsx", "Excel Files (*.xlsx)", options=options
                )
                if not file_path: return
                # ... (rest of excel export logic is fine) ...
                workbook = openpyxl.Workbook()
                worksheet_data = workbook.active
                worksheet_data.title = "Peak Analysis"
                headers = [table_widget_source.horizontalHeaderItem(col).text() for col in range(table_widget_source.columnCount())]
                for col, header in enumerate(headers, start=1):
                    cell = worksheet_data.cell(row=1, column=col, value=header); cell.font = Font(bold=True)
                
                r_offset = 0 # Offset for writing data if header is present
                for r_idx in range(table_widget_source.rowCount()):
                    # Skip placeholder rows
                    first_item_in_row = table_widget_source.item(r_idx, 0)
                    if first_item_in_row and ("No data" in first_item_in_row.text() or "Select an analysis" in first_item_in_row.text()):
                        continue

                    for c_idx in range(table_widget_source.columnCount()):
                        item = table_widget_source.item(r_idx, c_idx)
                        value = item.text() if item else ""
                        try:
                            if '%' in value:
                                numeric_value = float(value.replace('%', '')) / 100.0
                                cell = worksheet_data.cell(row=r_idx + 2 - r_offset, column=c_idx + 1, value=numeric_value); cell.number_format = '0.00%'
                            else:
                                numeric_value = float(value)
                                worksheet_data.cell(row=r_idx + 2 - r_offset, column=c_idx + 1, value=numeric_value)
                        except ValueError:
                            worksheet_data.cell(row=r_idx + 2 - r_offset, column=c_idx + 1, value=value)
                
                # Auto-size columns after writing data
                for col_idx_letter in range(1, worksheet_data.max_column + 1):
                    column_letter = openpyxl.utils.get_column_letter(col_idx_letter)
                    max_length = 0
                    for cell in worksheet_data[column_letter]:
                        try:
                            if cell.value:
                                cell_len = len(str(cell.value))
                                if '%' in str(cell.value) and cell.number_format == '0.00%': # Adjust for % display
                                    cell_len += 1 # approx for '%'
                                max_length = max(max_length, cell_len)
                        except: pass
                    adjusted_width = (max_length + 2) * 1.1 # A bit more padding
                    worksheet_data.column_dimensions[column_letter].width = min(max(adjusted_width, len(headers[col_idx_letter-1])+2 if col_idx_letter-1 < len(headers) else 0), 50) # Max width 50

                if standard_dict_for_export and len(standard_dict_for_export) >= 2:
                    worksheet_std = workbook.create_sheet("Standard Curve Data")
                    worksheet_std.cell(row=1, column=1, value="Known Quantity").font = Font(bold=True)
                    worksheet_std.cell(row=1, column=2, value="Total Peak Area").font = Font(bold=True)
                    current_row_std = 2
                    for qty_key, area_val in sorted(standard_dict_for_export.items()):
                        try:
                            worksheet_std.cell(row=current_row_std, column=1, value=float(qty_key))
                            worksheet_std.cell(row=current_row_std, column=2, value=float(area_val))
                            current_row_std += 1
                        except ValueError: print(f"Warning: Skipping invalid standard data for Excel: Qty={qty_key}, Area={area_val}")
                    for col_dim_std_letter in range(1, worksheet_std.max_column + 1): # Auto-size
                         column_letter_std = openpyxl.utils.get_column_letter(col_dim_std_letter)
                         worksheet_std.column_dimensions[column_letter_std].auto_size = True
                try:
                    workbook.save(file_path)
                    QMessageBox.information(self, "Success", f"Table data exported to\n{file_path}")
                except Exception as e:
                     QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                        

        class PeakAreaDialog(QDialog):
            """
            Interactive dialog to adjust peak regions and calculate peak areas.
            Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
            Area calculations use original data range (inverted).
            Peak detection uses a 0-255 scaled profile (inverted).
            Initial regions defined by outward trough search. V-V baseline uses adjacent troughs.
            """

            def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
                super().__init__(parent)
                self.parent_app = parent # Store parent_app reference
                self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
                self.setGeometry(100, 100, 1100, 850)

                if not isinstance(cropped_data, Image.Image):
                     raise TypeError("Input 'cropped_data' must be a PIL Image object")
                self.cropped_image_for_display = cropped_data

                self.original_max_value = 255.0
                pil_mode = cropped_data.mode
                try:
                    if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                        self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'L':
                        self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                        self.original_max_value = 255.0
                    elif pil_mode == 'F':
                        self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                        max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                        self.original_max_value = max(1.0, max_in_float)
                        scaled_for_display = np.clip(self.intensity_array_original_range * 255.0 / self.original_max_value, 0, 255).astype(np.uint8)
                        self.cropped_image_for_display = Image.fromarray(scaled_for_display, mode='L')
                    else: 
                        gray_img = cropped_data.convert("L")
                        self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                        self.original_max_value = 255.0
                        self.cropped_image_for_display = gray_img
                except Exception as e:
                    raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

                if self.intensity_array_original_range.ndim != 2:
                     raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

                self.profile_original_inverted = None 
                self.profile = None 
                self.background = None 

                self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
                self.smoothing_sigma = current_settings.get('smoothing_sigma', 2.0)
                self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
                self.peak_distance = current_settings.get('peak_distance', 10)
                self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.02)
                self.valley_offset_pixels = current_settings.get('valley_offset_pixels', 0)
                self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
                self.area_subtraction_method = current_settings.get('area_subtraction_method', "Rolling Ball")
                self.peaks = np.array([])
                self.initial_valley_regions = [] 
                self.peak_regions = [] 
                self.peak_areas_rolling_ball = []
                self.peak_areas_straight_line = []
                self.peak_areas_valley = []
                self.peak_sliders = [] 
                self._final_settings = {}
                self._persist_enabled_on_exit = persist_checked 

                self.manual_select_mode_active = False
                self.selected_peak_for_ui_focus = -1 
                self.peak_group_boxes = [] 
                
                self.add_peak_mode_active = False
                self.selected_peak_index_for_delete = -1

                if rolling_ball is None or find_peaks is None or gaussian_filter1d is None or interp1d is None:
                     QMessageBox.critical(self, "Dependency Error",
                                          "Missing SciPy or scikit-image library functions.\n"
                                          "Peak detection, smoothing, rolling ball, and interpolation require these libraries.\n"
                                          "Please install them (e.g., 'pip install scipy scikit-image') and restart.")

                self._setup_ui(persist_checked) 
                self.regenerate_profile_and_detect() 

            def _setup_ui(self, persist_checked_initial):
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(15)

                self.fig = plt.figure(figsize=(10, 5))
                self.fig.clf() 
                gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig) # type: ignore
                self.ax = self.fig.add_subplot(gs[0]) # type: ignore
                self.fig.canvas.mpl_connect('button_press_event', self.on_plot_click_for_selection)
                self.fig.tight_layout(pad=2) # type: ignore
                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                main_layout.addWidget(self.canvas, stretch=3)

                controls_hbox = QHBoxLayout()
                controls_hbox.setSpacing(15)

                left_controls_vbox = QVBoxLayout()
                fm_dialog = QFontMetrics(self.font()) 
                def get_slider_label_min_width(example_text_prefix, max_val_digits, suffix=""):
                    example_full_text = f"{example_text_prefix} ({'9' * max_val_digits}{suffix})  "
                    return fm_dialog.horizontalAdvance(example_full_text)

                global_settings_group = QGroupBox("Global Settings")
                global_settings_layout = QGridLayout(global_settings_group)
                global_settings_layout.setSpacing(8)
                self.band_estimation_combobox = QComboBox()
                self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
                self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
                self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
                global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
                global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2)
                self.method_combobox = QComboBox()
                self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
                self.method_combobox.setCurrentText(self.area_subtraction_method)
                self.method_combobox.currentIndexChanged.connect(self.update_plot)
                global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
                global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2)
                self.rolling_ball_label = QLabel(f"Rolling Ball Radius ({int(self.rolling_ball_radius)})")
                self.rolling_ball_label.setMinimumWidth(get_slider_label_min_width("Rolling Ball Radius", 3))
                self.rolling_ball_slider = QSlider(Qt.Horizontal)
                self.rolling_ball_slider.setRange(1, 500)
                self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
                self.rolling_ball_slider.setEnabled(self.area_subtraction_method == "Rolling Ball")
                self.rolling_ball_slider.valueChanged.connect(self.update_plot)
                self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_label: lbl.setText(f"Rolling Ball Radius ({val})"))
                global_settings_layout.addWidget(self.rolling_ball_label, 2, 0)
                global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1, 1, 2)
                self.method_combobox.currentIndexChanged.connect(lambda: self.rolling_ball_slider.setEnabled(self.method_combobox.currentText() == "Rolling Ball"))
                left_controls_vbox.addWidget(global_settings_group)

                peak_detect_group = QGroupBox("Peak Detection Parameters")
                peak_detect_layout = QGridLayout(peak_detect_group)
                peak_detect_layout.setSpacing(8)
                self.peak_number_label = QLabel("Detected Peaks:")
                self.peak_number_input = QLineEdit()
                self.peak_number_input.setPlaceholderText("#")
                self.peak_number_input.setMaximumWidth(60)
                self.update_peak_number_button = QPushButton("Set")
                self.update_peak_number_button.setToolTip("Manually override the number of peaks detected.")
                self.update_peak_number_button.clicked.connect(self.manual_peak_number_update)
                peak_detect_layout.addWidget(self.peak_number_label, 0, 0)
                peak_detect_layout.addWidget(self.peak_number_input, 0, 1)
                peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)
                self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
                self.smoothing_label.setMinimumWidth(get_slider_label_min_width("Smoothing Sigma", 1, ".0"))
                self.smoothing_slider = QSlider(Qt.Horizontal)
                self.smoothing_slider.setRange(0, 100)
                self.smoothing_slider.setValue(int(self.smoothing_sigma * 10))
                self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})"))
                self.smoothing_slider.valueChanged.connect(self.regenerate_profile_and_detect)
                peak_detect_layout.addWidget(self.smoothing_label, 1, 0)
                peak_detect_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2)
                self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
                self.peak_prominence_slider_label.setMinimumWidth(get_slider_label_min_width("Min Prominence", 1, ".00"))
                self.peak_prominence_slider = QSlider(Qt.Horizontal)
                self.peak_prominence_slider.setRange(0, 100)
                self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
                self.peak_prominence_slider.valueChanged.connect(self.detect_peaks)
                self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})"))
                peak_detect_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
                peak_detect_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)
                self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})")
                self.peak_height_slider_label.setMinimumWidth(get_slider_label_min_width("Min Height", 1, ".00"))
                self.peak_height_slider = QSlider(Qt.Horizontal)
                self.peak_height_slider.setRange(0, 100)
                self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
                self.peak_height_slider.valueChanged.connect(self.detect_peaks)
                self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f})"))
                peak_detect_layout.addWidget(self.peak_height_slider_label, 3, 0)
                peak_detect_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)
                self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
                self.peak_distance_slider_label.setMinimumWidth(get_slider_label_min_width("Min Distance", 3, " px"))
                self.peak_distance_slider = QSlider(Qt.Horizontal)
                self.peak_distance_slider.setRange(1, 200)
                self.peak_distance_slider.setValue(self.peak_distance)
                self.peak_distance_slider.valueChanged.connect(self.detect_peaks)
                self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
                peak_detect_layout.addWidget(self.peak_distance_slider_label, 4, 0)
                peak_detect_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)
                
                self.add_peak_manually_button = QPushButton("Add Peak at Click")
                self.add_peak_manually_button.setCheckable(True)
                self.add_peak_manually_button.setToolTip("Toggle: Click on the profile plot to add a new peak marker.")
                self.add_peak_manually_button.clicked.connect(self.toggle_add_peak_mode)
                peak_detect_layout.addWidget(self.add_peak_manually_button, 5, 0, 1, 1)
                self.delete_selected_peak_button = QPushButton("Delete Selected Peak")
                self.delete_selected_peak_button.setToolTip("Click a peak marker on the plot, then click this to delete it.")
                self.delete_selected_peak_button.setEnabled(False)
                self.delete_selected_peak_button.clicked.connect(self.delete_selected_peak_action)
                peak_detect_layout.addWidget(self.delete_selected_peak_button, 5, 1, 1, 2)
                left_controls_vbox.addWidget(peak_detect_group)
                left_controls_vbox.addStretch(1)
                controls_hbox.addLayout(left_controls_vbox, stretch=1)

                right_controls_vbox = QVBoxLayout()
                peak_spread_group = QGroupBox("Peak Region Adjustments")
                peak_spread_layout = QGridLayout(peak_spread_group)
                peak_spread_layout.setSpacing(8)
                self.valley_offset_label = QLabel(f"Valley Offset ({'+/-' if self.valley_offset_pixels>=0 else ''}{self.valley_offset_pixels} px)")
                self.valley_offset_label.setMinimumWidth(get_slider_label_min_width("Valley Offset (+/-", 3, " px"))
                self.valley_offset_slider = QSlider(Qt.Horizontal)
                self.valley_offset_slider.setRange(-20, 100)
                self.valley_offset_slider.setValue(self.valley_offset_pixels)
                self.valley_offset_slider.setToolTip("Applies an offset to the automatically detected valley boundaries.")
                self.valley_offset_slider.valueChanged.connect(self.apply_valley_offset)
                self.valley_offset_slider.valueChanged.connect(lambda value, lbl=self.valley_offset_label: lbl.setText(f"Valley Offset ({'+/-' if value>=0 else ''}{value} px)"))
                peak_spread_layout.addWidget(self.valley_offset_label, 0, 0)
                peak_spread_layout.addWidget(self.valley_offset_slider, 0, 1)
                region_actions_layout = QHBoxLayout()
                self.copy_regions_button = QPushButton("Copy Peak Regions")
                self.copy_regions_button.setToolTip("Copy current peak start/end boundaries.")
                self.copy_regions_button.clicked.connect(self.copy_peak_regions_to_app)
                region_actions_layout.addWidget(self.copy_regions_button)
                self.paste_regions_button = QPushButton("Paste Peak Regions")
                self.paste_regions_button.setToolTip("Paste previously copied peak boundaries.\nRegions will be scaled if profile length differs.")
                self.paste_regions_button.clicked.connect(self.paste_peak_regions_from_app)
                if not (self.parent_app and self.parent_app.copied_peak_regions_data.get("regions")):
                    self.paste_regions_button.setEnabled(False)
                region_actions_layout.addWidget(self.paste_regions_button)
                peak_spread_layout.addLayout(region_actions_layout, 1, 0, 1, 2)
                self.identify_peak_button = QPushButton("Identify/Select Peak")
                self.identify_peak_button.setCheckable(True)
                self.identify_peak_button.setToolTip("Click to activate, then click on a peak in the plot above to focus its sliders below.")
                self.identify_peak_button.clicked.connect(self.toggle_manual_select_mode)
                peak_spread_layout.addWidget(self.identify_peak_button, 2, 0, 1, 2)
                right_controls_vbox.addWidget(peak_spread_group)
                self.scroll_area_peaks = QScrollArea()
                self.scroll_area_peaks.setWidgetResizable(True)
                self.scroll_area_peaks.setMinimumHeight(250)
                self.scroll_area_peaks.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.container = QWidget()
                self.peak_sliders_layout = QVBoxLayout(self.container)
                self.peak_sliders_layout.setSpacing(10)
                self.scroll_area_peaks.setWidget(self.container)
                right_controls_vbox.addWidget(self.scroll_area_peaks, stretch=1)
                controls_hbox.addLayout(right_controls_vbox, stretch=2)
                main_layout.addLayout(controls_hbox)

                bottom_button_layout = QHBoxLayout()
                self.persist_settings_checkbox = QCheckBox("Persist Settings")
                self.persist_settings_checkbox.setChecked(persist_checked_initial)
                self.persist_settings_checkbox.setToolTip("Save current detection parameters for the next time this dialog opens.")
                bottom_button_layout.addWidget(self.persist_settings_checkbox)
                bottom_button_layout.addStretch(1)
                self.ok_button = QPushButton("OK")
                self.ok_button.setMinimumWidth(100)
                self.ok_button.setDefault(True)
                self.ok_button.clicked.connect(self.accept_and_close)
                bottom_button_layout.addWidget(self.ok_button)
                main_layout.addLayout(bottom_button_layout)
                self.setLayout(main_layout)

            def toggle_manual_select_mode(self, checked):
                self.manual_select_mode_active = checked
                if checked:
                    self.canvas.setCursor(Qt.PointingHandCursor)
                    if self.add_peak_mode_active: # Deactivate add mode if active
                        self.add_peak_manually_button.setChecked(False)
                        self.toggle_add_peak_mode(False) 
                    QMessageBox.information(self, "Identify Peak", "Manual selection mode is ON.\nClick on a peak in the profile plot above.")
                else:
                    self.canvas.setCursor(Qt.ArrowCursor)
                    self.selected_peak_for_ui_focus = -1
                    self._update_peak_group_box_styles()
                    self.update_plot()

            def toggle_add_peak_mode(self, checked):
                self.add_peak_mode_active = checked
                if checked:
                    self.canvas.setCursor(Qt.CrossCursor)
                    if self.manual_select_mode_active: # Deactivate identify mode if active
                        self.identify_peak_button.setChecked(False)
                        self.toggle_manual_select_mode(False)
                    self.selected_peak_index_for_delete = -1 # Clear deletion selection
                    self.delete_selected_peak_button.setEnabled(False)
                    self.update_plot() 
                    QMessageBox.information(self, "Add Peak", "Add Peak Mode: ON. Click on the profile plot to add a peak.")
                else:
                    self.canvas.setCursor(Qt.ArrowCursor)

            def on_plot_click_for_selection(self, event):
                # Prioritize add mode if active
                if self.add_peak_mode_active:
                    if event.inaxes == self.ax and event.button == 1 and event.xdata is not None:
                        clicked_x = int(round(event.xdata))
                        if self.profile_original_inverted is not None and 0 <= clicked_x < len(self.profile_original_inverted):
                            self.add_manual_peak(clicked_x)
                    return 

                # Then prioritize manual UI focus selection mode
                if self.manual_select_mode_active:
                    if event.inaxes == self.ax and event.button == 1 and event.xdata is not None and self.peaks.any():
                        clicked_x = int(round(event.xdata))
                        distances = np.abs(self.peaks - clicked_x)
                        closest_peak_idx_in_self_peaks = np.argmin(distances)
                        click_tolerance = self.peak_distance / 2.0 if self.peak_distance > 0 else 10.0
                        if distances[closest_peak_idx_in_self_peaks] <= click_tolerance:
                            self.selected_peak_for_ui_focus = closest_peak_idx_in_self_peaks
                            if 0 <= self.selected_peak_for_ui_focus < len(self.peak_group_boxes):
                                group_box_to_focus = self.peak_group_boxes[self.selected_peak_for_ui_focus]
                                self.scroll_area_peaks.ensureWidgetVisible(group_box_to_focus, yMargin=10)
                            self._update_peak_group_box_styles()
                            self.update_plot()
                        else:
                            self.selected_peak_for_ui_focus = -1; self._update_peak_group_box_styles(); self.update_plot()
                    elif event.inaxes != self.ax : 
                         self.identify_peak_button.setChecked(False)
                         self.toggle_manual_select_mode(False)
                    return 

                # Default: Select for deletion (if no other mode is active)
                if event.inaxes == self.ax and event.button == 1 and event.xdata is not None and self.peaks.any():
                    clicked_x = int(round(event.xdata))
                    distances = np.abs(self.peaks - clicked_x)
                    closest_peak_idx_in_self_peaks = np.argmin(distances)
                    click_tolerance_delete = max(5, self.peak_distance / 4.0) 
                    if distances[closest_peak_idx_in_self_peaks] <= click_tolerance_delete:
                        self.selected_peak_index_for_delete = self.peaks[closest_peak_idx_in_self_peaks]
                        self.delete_selected_peak_button.setEnabled(True)
                        self.update_plot()
                        print(f"Peak at x={self.selected_peak_index_for_delete} selected for deletion.")
                    else:
                        self.selected_peak_index_for_delete = -1
                        self.delete_selected_peak_button.setEnabled(False)
                        self.update_plot()

            def add_manual_peak(self, x_coord):
                if self.profile_original_inverted is None or x_coord in self.peaks:
                    print(f"Manual Add: Peak at {x_coord} already exists or no profile.")
                    return
                new_peaks_list = self.peaks.tolist()
                new_peaks_list.append(x_coord)
                self.peaks = np.array(sorted(new_peaks_list))
                print(f"Manually added peak at index: {x_coord}. New peaks: {self.peaks}")
                self._redefine_all_valley_regions() # This correctly updates sliders and plot
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))

            def _redefine_all_valley_regions(self):
                self.initial_valley_regions = []
                profile_to_analyze = self.profile_original_inverted
                if profile_to_analyze is None: # Guard against None profile
                    self.update_sliders(); self.update_plot()
                    return
                profile_len = len(profile_to_analyze)

                if profile_len <= 1 or len(self.peaks) == 0:
                    self.initial_valley_regions = []
                    self.peak_regions = []
                    self.update_sliders()
                    self.update_plot()
                    return

                if len(self.peaks) > 1:
                    midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2
                    search_boundaries_left = np.concatenate(([0], midpoints))
                    search_boundaries_right = np.concatenate((midpoints, [profile_len - 1]))
                elif len(self.peaks) == 1:
                     search_boundaries_left = np.array([0])
                     search_boundaries_right = np.array([profile_len - 1])
                else: # No peaks
                     search_boundaries_left = np.array([])
                     search_boundaries_right = np.array([])

                for i, peak_idx in enumerate(self.peaks):
                     left_bound = search_boundaries_left[i]
                     right_bound = search_boundaries_right[i]
                     try:
                         valley_start, valley_end = self._find_outward_troughs(
                             profile_to_analyze, peak_idx, int(left_bound), int(right_bound)
                         )
                         self.initial_valley_regions.append((valley_start, valley_end))
                     except Exception as e_trough:
                         print(f"Error finding troughs for peak {i} (idx {peak_idx}) after modification: {e_trough}")
                         fallback_width = max(2, self.peak_distance // 4 if hasattr(self, 'peak_distance') else 5)
                         fb_start = max(0, peak_idx - fallback_width)
                         fb_end = min(profile_len - 1, peak_idx + fallback_width)
                         if fb_start >= fb_end: fb_end = min(profile_len - 1, fb_start + 1)
                         self.initial_valley_regions.append((fb_start, fb_end))
                
                self.apply_valley_offset(self.valley_offset_slider.value()) # This calls update_sliders & update_plot

            def delete_selected_peak_action(self):
                if self.selected_peak_index_for_delete == -1 or self.selected_peak_index_for_delete not in self.peaks:
                    QMessageBox.warning(self, "No Peak Selected", "Please click on a peak marker in the plot to select it for deletion.")
                    return
                
                new_peaks_list = self.peaks.tolist()
                try: new_peaks_list.remove(self.selected_peak_index_for_delete)
                except ValueError: # Should not happen if selection logic is correct
                    print(f"Delete: Peak {self.selected_peak_index_for_delete} not found in current peaks list.")
                    self.selected_peak_index_for_delete = -1; self.delete_selected_peak_button.setEnabled(False); self.update_plot(); return

                self.peaks = np.array(sorted(new_peaks_list))
                print(f"Deleted peak at x={self.selected_peak_index_for_delete}. New peaks: {self.peaks}")

                self.selected_peak_index_for_delete = -1 
                self.delete_selected_peak_button.setEnabled(False)
                
                # Reset UI focus if the deleted peak was the one focused for UI
                if self.selected_peak_for_ui_focus != -1:
                    # Check if the focused peak index is still valid for the new self.peaks array
                    if self.selected_peak_for_ui_focus >= len(self.peaks) or \
                       (0 <= self.selected_peak_for_ui_focus < len(self.peaks) and \
                        self.peaks[self.selected_peak_for_ui_focus] != self.selected_peak_for_ui_focus): # Simplified check; direct index might shift
                        # A more robust check would be to see if the *value* of the focused peak is still in self.peaks
                        # For now, a simple reset if the index becomes invalid or points to a different peak value
                         self.selected_peak_for_ui_focus = -1 
                         self._update_peak_group_box_styles()

                self._redefine_all_valley_regions() # This correctly updates sliders and plot

                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))

            def _update_peak_group_box_styles(self):
                if not hasattr(self, 'peak_group_boxes'): return
                for i, group_box in enumerate(self.peak_group_boxes):
                    if group_box:
                        if i == self.selected_peak_for_ui_focus:
                            group_box.setStyleSheet("QGroupBox { border: 2px solid #FFA500; margin-top: 1ex; font-weight: bold; } QGroupBox::title { subcontrol-origin: margin; left: 7px; padding: 0 3px 0 3px; }")
                        else:
                            group_box.setStyleSheet("")

            def update_sliders(self):
                while self.peak_sliders_layout.count():
                    item = self.peak_sliders_layout.takeAt(0)
                    widget = item.widget()
                    if widget: widget.deleteLater()
                self.peak_sliders.clear()
                self.peak_group_boxes.clear()

                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
                    return

                profile_len = len(self.profile_original_inverted)
                fm = QFontMetrics(self.font())
                max_val_str = str(profile_len -1 if profile_len > 0 else 0)
                label_min_width = max(fm.horizontalAdvance(f"Start: {max_val_str}"), fm.horizontalAdvance(f"End: {max_val_str}")) + 10

                num_items = len(self.peak_regions)
                num_to_display = min(len(self.peaks), num_items)

                for i in range(num_to_display):
                    try:
                        start_val, end_val = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                        peak_index_val = int(self.peaks[i])
                    except (IndexError, ValueError, TypeError): continue

                    peak_group = QGroupBox(f"Peak {i + 1} (Idx: {peak_index_val})")
                    self.peak_group_boxes.append(peak_group)
                    peak_layout = QGridLayout(peak_group)
                    peak_layout.setSpacing(5)

                    start_slider = QSlider(Qt.Horizontal); start_slider.setRange(0, profile_len - 1)
                    start_val_clamped = max(0, min(profile_len - 1, start_val))
                    start_slider.setValue(start_val_clamped)
                    start_label = QLabel(f"Start: {start_val_clamped}")
                    start_label.setMinimumWidth(label_min_width); start_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    start_slider.valueChanged.connect(lambda val, lbl=start_label, idx=i, sl=start_slider: self._update_region_from_slider(idx, 'start', val, lbl, sl))
                    start_slider.valueChanged.connect(self.update_plot)
                    peak_layout.addWidget(start_label, 0, 0); peak_layout.addWidget(start_slider, 0, 1)

                    end_slider = QSlider(Qt.Horizontal); end_slider.setRange(0, profile_len - 1)
                    end_val_clamped = max(start_val_clamped, min(profile_len - 1, end_val))
                    end_slider.setValue(end_val_clamped)
                    end_label = QLabel(f"End: {end_val_clamped}")
                    end_label.setMinimumWidth(label_min_width); end_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    end_slider.valueChanged.connect(lambda val, lbl=end_label, idx=i, sl=end_slider: self._update_region_from_slider(idx, 'end', val, lbl, sl))
                    end_slider.valueChanged.connect(self.update_plot)
                    peak_layout.addWidget(end_label, 1, 0); peak_layout.addWidget(end_slider, 1, 1)

                    self.peak_sliders_layout.addWidget(peak_group)
                    self.peak_sliders.append((start_slider, end_slider))
                
                if num_to_display > 0: self.peak_sliders_layout.addStretch(1)
                self._update_peak_group_box_styles()
                if hasattr(self, 'container') and self.container:
                    self.container.adjustSize(); self.container.update()
            
            def _update_region_from_slider(self, index, boundary_type, value, label_widget, slider_widget):
                if not (0 <= index < len(self.peak_regions)): return
                current_start, current_end = self.peak_regions[index]
                peer_slider_widget = None
                if 0 <= index < len(self.peak_sliders):
                    all_sliders_for_peak = self.peak_sliders[index]
                    peer_slider_widget = all_sliders_for_peak[1] if boundary_type == 'start' else all_sliders_for_peak[0]

                if boundary_type == 'start':
                    new_start = value
                    if peer_slider_widget: new_start = min(new_start, peer_slider_widget.value())
                    else: new_start = min(new_start, current_end)
                    self.peak_regions[index] = (new_start, current_end if not peer_slider_widget else peer_slider_widget.value())
                    label_widget.setText(f"Start: {new_start}")
                    if slider_widget.value() != new_start:
                        slider_widget.blockSignals(True); slider_widget.setValue(new_start); slider_widget.blockSignals(False)
                elif boundary_type == 'end':
                    new_end = value
                    if peer_slider_widget: new_end = max(new_end, peer_slider_widget.value())
                    else: new_end = max(new_end, current_start)
                    self.peak_regions[index] = (current_start if not peer_slider_widget else peer_slider_widget.value(), new_end)
                    label_widget.setText(f"End: {new_end}")
                    if slider_widget.value() != new_end:
                        slider_widget.blockSignals(True); slider_widget.setValue(new_end); slider_widget.blockSignals(False)

            def copy_peak_regions_to_app(self):
                if not self.parent_app:
                    QMessageBox.warning(self, "Error", "Parent application reference not found.")
                    return
                if not self.peak_regions:
                    QMessageBox.information(self, "No Regions", "No peak regions defined to copy.")
                    return

                # Store current peak regions and profile length
                self.parent_app.copied_peak_regions_data["regions"] = [tuple(r) for r in self.peak_regions]
                self.parent_app.copied_peak_regions_data["profile_length"] = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                
                # --- NEW: Store the actual peak indices ---
                self.parent_app.copied_peak_regions_data["peaks"] = self.peaks.tolist() # Convert numpy array to list for storage

                QMessageBox.information(self, "Regions Copied", f"{len(self.peak_regions)} peak regions and {len(self.peaks)} peak locations copied.")
                if hasattr(self, 'paste_regions_button'): # Check if button exists (it should)
                    self.paste_regions_button.setEnabled(True)

            def paste_peak_regions_from_app(self):
                if not self.parent_app or not self.parent_app.copied_peak_regions_data.get("regions"):
                    QMessageBox.information(self, "No Regions to Paste", "No peak regions have been copied yet.")
                    return
                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
                    QMessageBox.warning(self, "Error", "Current profile not available for pasting regions.")
                    return

                copied_data = self.parent_app.copied_peak_regions_data
                regions_to_paste = copied_data["regions"]
                original_profile_len = copied_data["profile_length"]
                # --- NEW: Get copied peak indices ---
                copied_peaks_indices = np.array(copied_data.get("peaks", [])) # Default to empty if not found

                current_profile_len = len(self.profile_original_inverted)
                
                self.peak_regions = [] # Clear current regions
                # new_peaks_for_self_dot_peaks = [] # Will be populated either directly or from region centers

                if original_profile_len > 0 and current_profile_len > 0 and \
                   original_profile_len == current_profile_len and len(copied_peaks_indices) > 0 and \
                   len(copied_peaks_indices) == len(regions_to_paste): # Ensure counts match for direct application
                    # --- DIRECT PEAK AND REGION RESTORATION (NO SCALING) ---
                    print(f"Pasting: Profile length MATCH ({current_profile_len}). Restoring {len(copied_peaks_indices)} peaks and regions directly.")
                    self.peaks = np.array(sorted(copied_peaks_indices)) # Use copied peaks directly

                    for i, (start_orig, end_orig) in enumerate(regions_to_paste):
                        # Basic clamping for safety, though they should be valid if from same profile
                        start_clamped = max(0, min(int(start_orig), current_profile_len - 1))
                        end_clamped = max(0, min(int(end_orig), current_profile_len - 1))
                        if start_clamped > end_clamped: # Ensure start <= end
                            # This case should be rare if regions were valid when copied
                            mid_point = (start_clamped + end_clamped) // 2
                            start_clamped = mid_point
                            end_clamped = min(current_profile_len -1, mid_point + 1)
                            if start_clamped >= end_clamped and start_clamped > 0 : start_clamped -=1
                        self.peak_regions.append((start_clamped, end_clamped))
                    
                    # Ensure self.peaks and self.peak_regions are consistent in length
                    if len(self.peaks) != len(self.peak_regions):
                        print(f"Warning: Pasted peaks ({len(self.peaks)}) and regions ({len(self.peak_regions)}) count mismatch. Re-deriving peaks from regions.")
                        # Fallback to deriving peaks from regions if counts don't match (should not happen if copied correctly)
                        temp_derived_peaks = []
                        for sr, er in self.peak_regions:
                            temp_derived_peaks.append((sr + er) // 2)
                        self.peaks = np.array(sorted(temp_derived_peaks))
                else:
                    # --- ORIGINAL SCALING LOGIC (for regions, and derive peaks from centers) ---
                    if original_profile_len != current_profile_len:
                        print(f"Pasting: Profile length MISMATCH. Original: {original_profile_len}, Current: {current_profile_len}. Scaling regions and peaks.")
                    elif not (len(copied_peaks_indices) > 0 and len(copied_peaks_indices) == len(regions_to_paste)):
                         print(f"Pasting: Copied peak data insufficient or mismatched with regions. Deriving peaks from region centers.")


                    temp_derived_peaks_from_scaled_regions = []
                    scaled_copied_peak_indices = []

                    scale_factor = 1.0
                    if original_profile_len > 0 and current_profile_len > 0 and original_profile_len != current_profile_len:
                        scale_factor = float(current_profile_len) / original_profile_len

                    # Scale copied peak indices if profile length differs and peaks were copied
                    if len(copied_peaks_indices) > 0 and scale_factor != 1.0:
                        for peak_idx_orig in copied_peaks_indices:
                            scaled_peak_idx = int(round(float(peak_idx_orig) * scale_factor))
                            scaled_copied_peak_indices.append(max(0, min(scaled_peak_idx, current_profile_len - 1)))
                        # Use scaled peaks if available and counts match regions
                        if len(scaled_copied_peak_indices) == len(regions_to_paste):
                             self.peaks = np.array(sorted(scaled_copied_peak_indices))
                             print(f"  Used SCALED original peak indices for {len(self.peaks)} peaks.")
                        else: # Fallback to deriving from region centers
                            self.peaks = np.array([]) # Mark for recalculation below

                    for i, (start_orig, end_orig) in enumerate(regions_to_paste):
                        start_scaled, end_scaled = start_orig, end_orig
                        if scale_factor != 1.0: # Apply scaling if factor is not 1
                            start_scaled = int(round(float(start_orig) * scale_factor))
                            end_scaled = int(round(float(end_orig) * scale_factor))
                        
                        start_clamped = max(0, min(start_scaled, current_profile_len - 1))
                        end_clamped = max(0, min(end_scaled, current_profile_len - 1))
                        if start_clamped > end_clamped:
                            mid_point = (start_clamped + end_clamped) // 2
                            start_clamped = mid_point
                            end_clamped = min(current_profile_len -1, mid_point + 1)
                            if start_clamped >= end_clamped and start_clamped > 0 : start_clamped -=1
                        self.peak_regions.append((start_clamped, end_clamped))
                        temp_derived_peaks_from_scaled_regions.append((start_clamped + end_clamped) // 2)

                    # If self.peaks was not set from scaled_copied_peak_indices, derive from region centers
                    if len(self.peaks) == 0 and temp_derived_peaks_from_scaled_regions:
                        self.peaks = np.array(sorted(temp_derived_peaks_from_scaled_regions))
                        print(f"  Derived {len(self.peaks)} peaks from SCALED region centers.")
                    elif len(self.peaks) == 0: # No regions pasted or all invalid
                        self.peaks = np.array([])
                        print("  No peaks derived as no valid regions were processed.")


                if hasattr(self, 'peak_number_input'):
                    self.peak_number_input.setText(str(len(self.peaks)))
                
                # Temporarily disconnect signals from sliders that call detect_peaks or apply_valley_offset
                # to prevent them from re-triggering detection based on old peak numbers or default values
                # during the UI update process.
                sliders_to_disconnect = [
                    (getattr(self, 'peak_prominence_slider', None), self.detect_peaks),
                    (getattr(self, 'peak_height_slider', None), self.detect_peaks),
                    (getattr(self, 'peak_distance_slider', None), self.detect_peaks),
                    (getattr(self, 'valley_offset_slider', None), self.apply_valley_offset)
                ]
                for slider, method_to_disconnect in sliders_to_disconnect:
                    if slider:
                        try:
                            slider.valueChanged.disconnect(method_to_disconnect)
                        except TypeError: # Signal not connected or already disconnected
                            pass 

                self.update_sliders() # This will create/update sliders based on self.peak_regions
                self.update_plot()    # This will draw based on self.peaks and self.peak_regions

                # Reconnect signals
                for slider, method_to_reconnect in sliders_to_disconnect:
                    if slider:
                        try:
                            slider.valueChanged.connect(method_to_reconnect)
                        except TypeError: # Should not happen if disconnected correctly, but for safety
                            pass 

                QMessageBox.information(self, "Regions Pasted", f"{len(self.peak_regions)} peak regions and {len(self.peaks)} peak locations applied.")

            def accept_and_close(self):
                self._final_settings = {
                    'rolling_ball_radius': self.rolling_ball_slider.value(),
                    'peak_height_factor': self.peak_height_slider.value() / 100.0,
                    'peak_distance': self.peak_distance_slider.value(),
                    'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
                    'valley_offset_pixels': self.valley_offset_slider.value(),
                    'band_estimation_method': self.band_estimation_combobox.currentText(),
                    'area_subtraction_method': self.method_combobox.currentText(),
                    'smoothing_sigma': self.smoothing_slider.value() / 10.0,
                }
                self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
                self.accept()
                
            def get_current_settings(self):
                return self._final_settings

            def should_persist_settings(self):
                return self._persist_enabled_on_exit

            def get_final_peak_area(self):
                num_valid_peaks = len(self.peak_regions)
                current_area_list = []
                if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
                elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
                elif self.method == "Valley-to-Valley": current_area_list = self.peak_areas_valley
                else: return []
                
                if len(current_area_list) != num_valid_peaks:
                    return current_area_list[:num_valid_peaks] 
                else:
                    return current_area_list
            
            def regenerate_profile_and_detect(self):
                if gaussian_filter1d is None: return
                self.band_estimation_method = self.band_estimation_combobox.currentText()
                self.area_subtraction_method = self.method_combobox.currentText()
                if hasattr(self, 'smoothing_slider'): self.smoothing_sigma = self.smoothing_slider.value() / 10.0
                else: self.smoothing_sigma = 2.0

                profile_temp = None
                if self.band_estimation_method == "Mean":
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)
                elif self.band_estimation_method.startswith("Percentile"):
                    try:
                        percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                        profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
                    except:
                        profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1)
                        print("Warning: Defaulting to 5th percentile for profile.")
                else:
                    profile_temp = np.mean(self.intensity_array_original_range, axis=1)

                if profile_temp is None or not np.all(np.isfinite(profile_temp)):
                    print("Warning: Profile calculation failed or resulted in NaN/Inf. Using zeros.")
                    profile_temp = np.zeros(self.intensity_array_original_range.shape[0])

                profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
                min_inverted_raw = np.min(profile_original_inv_raw)
                profile_original_inv_raw -= min_inverted_raw

                self.profile_original_inverted = profile_original_inv_raw
                try:
                    current_sigma = self.smoothing_sigma
                    if current_sigma > 0.1 and len(self.profile_original_inverted) > int(3 * current_sigma) * 2 + 1:
                        self.profile_original_inverted = gaussian_filter1d(self.profile_original_inverted, sigma=current_sigma)
                except Exception as smooth_err:
                    print(f"Error smoothing main profile: {smooth_err}")

                prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
                if prof_max_inv > prof_min_inv + 1e-6:
                    self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
                else:
                    self.profile = np.zeros_like(self.profile_original_inverted)
                self.detect_peaks()

            def _find_outward_troughs(self, profile, peak_idx, left_bound, right_bound):
                profile_len = len(profile)
                if not (0 <= left_bound <= peak_idx <= right_bound < profile_len):
                    w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                    return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)

                valley_left_idx = peak_idx 
                min_val_left = profile[peak_idx]
                found_trough_left = False
                for idx in range(peak_idx - 1, left_bound - 1, -1):
                    current_val = profile[idx]
                    if current_val <= profile[idx + 1]:
                        is_local_min = True
                        if idx > 0 and profile[idx - 1] < current_val: is_local_min = False
                        if is_local_min: valley_left_idx = idx; found_trough_left = True; break
                        if current_val < min_val_left: min_val_left = current_val; valley_left_idx = idx
                    else: 
                        if valley_left_idx == peak_idx: valley_left_idx = idx + 1
                        found_trough_left = True; break
                if not found_trough_left and valley_left_idx == peak_idx:
                    if left_bound < peak_idx and profile[left_bound] <= min_val_left: valley_left_idx = left_bound
                    else: valley_left_idx = max(0, peak_idx - 1)

                valley_right_idx = peak_idx
                min_val_right = profile[peak_idx]
                found_trough_right = False
                for idx in range(peak_idx + 1, right_bound + 1, 1):
                    current_val = profile[idx]
                    if current_val <= profile[idx - 1]:
                        is_local_min = True
                        if idx > peak_idx + 1 and profile[idx - 1] < current_val: is_local_min = False
                        if idx < profile_len - 1 and profile[idx + 1] < current_val: is_local_min = False
                        if is_local_min: valley_right_idx = idx; found_trough_right = True; break
                        if current_val < min_val_right: min_val_right = current_val; valley_right_idx = idx
                    else: 
                        if valley_right_idx == peak_idx: valley_right_idx = idx - 1
                        found_trough_right = True; break
                if not found_trough_right and valley_right_idx == peak_idx:
                    if right_bound > peak_idx and profile[right_bound] <= min_val_right: valley_right_idx = right_bound
                    else: valley_right_idx = min(profile_len - 1, peak_idx + 1)

                valley_left_idx = max(0, min(peak_idx, valley_left_idx))
                valley_right_idx = min(profile_len - 1, max(peak_idx, valley_right_idx))
                if valley_left_idx >= valley_right_idx:
                     w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                     valley_left_idx = max(0, peak_idx - w); valley_right_idx = min(profile_len - 1, peak_idx + w)
                     if valley_left_idx >= valley_right_idx:
                         if valley_right_idx < profile_len - 1: valley_right_idx += 1
                         elif valley_left_idx > 0: valley_left_idx -= 1
                return valley_left_idx, valley_right_idx

            def detect_peaks(self):
                if self.profile is None or len(self.profile) == 0 or find_peaks is None:
                    self.peaks, self.initial_valley_regions, self.peak_regions = np.array([]), [], []
                    if hasattr(self, 'peak_number_input'): self.peak_number_input.setText("0")
                    self.update_sliders(); self.update_plot(); return

                self.peak_height_factor = self.peak_height_slider.value() / 100.0
                self.peak_distance = self.peak_distance_slider.value()
                self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
                # self.valley_offset_pixels = self.valley_offset_slider.value() # Applied in apply_valley_offset
                # self.rolling_ball_radius = self.rolling_ball_slider.value() # Applied in update_plot

                profile_range = np.ptp(self.profile); min_val_profile = np.min(self.profile)
                if profile_range < 1e-6 : profile_range = 1.0 
                min_height_abs = min_val_profile + profile_range * self.peak_height_factor
                min_prominence_abs = profile_range * self.peak_prominence_factor
                min_prominence_abs = max(1.0, min_prominence_abs)

                try:
                    peaks_indices, _ = find_peaks(
                        self.profile, height=min_height_abs, prominence=min_prominence_abs,
                        distance=self.peak_distance, width=1
                    )
                    self.peaks = np.sort(peaks_indices)
                except Exception as e:
                    QMessageBox.warning(self, "Peak Detection Error", f"Error finding peak locations:\n{e}")
                    self.peaks = np.array([]); self.initial_valley_regions = []; self.peak_regions = []
                    if hasattr(self, 'peak_number_input') and not self.peak_number_input.hasFocus(): self.peak_number_input.setText("0")
                    self.update_sliders(); self.update_plot()
                    return

                self._redefine_all_valley_regions() 

                if hasattr(self, 'peak_number_input') and (not self.peak_number_input.hasFocus() or self.peak_number_input.text() == ""):
                     self.peak_number_input.setText(str(len(self.peaks)))


            def apply_valley_offset(self, offset_value):
                self.valley_offset_pixels = offset_value
                self.peak_regions = [] 

                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
                    self.update_sliders(); self.update_plot(); return

                profile_len = len(self.profile_original_inverted)
                
                num_initial = min(len(self.peaks), len(self.initial_valley_regions))
                if len(self.peaks) != len(self.initial_valley_regions):
                    print(f"Warning (apply_valley_offset): Peak ({len(self.peaks)}) / initial valley region ({len(self.initial_valley_regions)}) mismatch. Recalculating valleys.")
                    self._redefine_all_valley_regions() 
                    num_initial = min(len(self.peaks), len(self.initial_valley_regions))


                for i in range(num_initial):
                    try:
                        valley_start, valley_end = self.initial_valley_regions[i]
                        new_start = valley_start - self.valley_offset_pixels
                        new_end = valley_end + self.valley_offset_pixels
                        new_start_clamped = max(0, new_start)
                        new_end_clamped = min(profile_len - 1, new_end)
                        if new_start_clamped > new_end_clamped:
                            mid_valley = (valley_start + valley_end) // 2
                            new_start_clamped = mid_valley
                            new_end_clamped = mid_valley
                        self.peak_regions.append((new_start_clamped, new_end_clamped))
                    except IndexError:
                         print(f"Error accessing initial valley region at index {i}")
                         continue
                if len(self.peak_regions) != num_initial:
                     print(f"Warning: Final peak_regions length ({len(self.peak_regions)}) mismatch after offset application.")
                self.update_sliders(); self.update_plot()

            def manual_peak_number_update(self):
                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
                    QMessageBox.warning(self, "Error", "Profile must be generated first."); return
                profile_len = len(self.profile_original_inverted)
                try:
                    num_peaks_manual = int(self.peak_number_input.text())
                    if num_peaks_manual < 0: raise ValueError("Negative number")
                    current_num_peaks = len(self.peaks)
                    if num_peaks_manual == current_num_peaks: return
                    peaks_list = self.peaks.tolist()
                    if num_peaks_manual == 0: self.peaks = np.array([])
                    elif num_peaks_manual < current_num_peaks: self.peaks = self.peaks[:num_peaks_manual]
                    else: 
                        num_to_add = num_peaks_manual - current_num_peaks; profile_center = profile_len // 2
                        current_peaks_set = set(self.peaks)
                        for _ in range(num_to_add):
                            new_peak_pos = profile_center; offset = 0
                            while new_peak_pos in current_peaks_set or new_peak_pos < 0 or new_peak_pos >= profile_len:
                                offset += 5; new_peak_pos = profile_center + np.random.choice([-offset, offset]) # type: ignore
                                if offset > profile_len // 2:
                                    new_peak_pos = np.random.randint(0, profile_len)
                                    if new_peak_pos in current_peaks_set: continue; break
                            peaks_list.append(new_peak_pos); current_peaks_set.add(new_peak_pos)
                        peaks_list.sort(); self.peaks = np.array(peaks_list)
                    
                    self._redefine_all_valley_regions() 

                except ValueError: self.peak_number_input.setText(str(len(self.peaks))); QMessageBox.warning(self, "Input Error", "Please enter a valid non-negative integer.")
                except Exception as e: print(f"Error during manual peak number update: {e}"); QMessageBox.critical(self, "Error", f"Manual peak update error:\n{e}"); self.peak_number_input.setText(str(len(self.peaks)))

            def _find_adjacent_trough(self, profile, start_point_in_region, direction, window=15):
                profile_len = len(profile)
                if not (0 <= start_point_in_region < profile_len): return start_point_in_region

                for i in range(window + 1): 
                    current_idx = start_point_in_region + i * direction
                    if not (0 <= current_idx < profile_len):
                        return max(0, min(start_point_in_region + (i-1) * direction, profile_len -1)) if i > 0 else start_point_in_region

                    is_local_min = True
                    if current_idx > 0 and profile[current_idx - 1] < profile[current_idx]: is_local_min = False
                    if current_idx < profile_len - 1 and profile[current_idx + 1] < profile[current_idx]: is_local_min = False
                    if current_idx == 0 and profile_len > 1 and profile[current_idx + 1] >= profile[current_idx]: is_local_min = True
                    elif current_idx == 0 and profile_len == 1: is_local_min = True
                    if current_idx == profile_len - 1 and profile_len > 1 and profile[current_idx -1] >= profile[current_idx]: is_local_min = True
                    elif current_idx == profile_len - 1 and profile_len == 1: is_local_min = True
                    if is_local_min: return current_idx
                return start_point_in_region

            def update_plot(self):
                if self.canvas is None: return
                profile_to_plot_and_calc = self.profile_original_inverted
                if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
                     try:
                         self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); # type: ignore
                         # Add an x-axis to the cleared top plot, it will be shared by the bottom one
                         ax_image_temp = self.fig.add_subplot(gs[1], sharex=self.ax) # type: ignore
                         ax_image_temp.set_xlabel("Pixel Index Along Profile Axis") # Set xlabel on the bottom plot
                         self.ax.tick_params(axis='x', labelbottom=False) # Hide x-labels on top plot
                         self.canvas.draw_idle()
                     except Exception as e: print(f"Error clearing plot: {e}")
                     return

                self.method = self.method_combobox.currentText()
                self.rolling_ball_radius = self.rolling_ball_slider.value()

                if rolling_ball:
                    try:
                        profile_float = profile_to_plot_and_calc.astype(np.float64)
                        safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
                        if len(profile_float) > 1 :
                            self.background = self._custom_rolling_ball(profile_float, safe_radius)
                            self.background = np.maximum(self.background, 0)
                        else: self.background = profile_float.copy()
                    except ImportError: self.background = np.zeros_like(profile_to_plot_and_calc)
                    except Exception as e: print(f"Error rolling ball: {e}."); self.background = np.zeros_like(profile_to_plot_and_calc)
                else: 
                     self.background = np.zeros_like(profile_to_plot_and_calc)

                self.fig.clf()
                gs = GridSpec(2, 1, height_ratios=[3.5, 1], hspace=0.05, figure=self.fig) # Slightly more space for top, reduce hspace # type: ignore
                self.ax = self.fig.add_subplot(gs[0]) # type: ignore
                ax_image = self.fig.add_subplot(gs[1], sharex=self.ax) # type: ignore

                self.ax.plot(profile_to_plot_and_calc, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.2)

                if len(self.peaks) > 0:
                     valid_peaks_indices = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
                     if len(valid_peaks_indices) > 0:
                         peak_y_on_smoothed = profile_to_plot_and_calc[valid_peaks_indices]
                         self.ax.scatter(valid_peaks_indices, peak_y_on_smoothed, color="red", marker='x', s=50, label="Detected Peaks", zorder=5)
                         
                         if self.selected_peak_for_ui_focus != -1 and 0 <= self.selected_peak_for_ui_focus < len(self.peaks):
                             if self.selected_peak_for_ui_focus < len(self.peaks): 
                                 focused_peak_x = self.peaks[self.selected_peak_for_ui_focus]
                                 if 0 <= focused_peak_x < len(profile_to_plot_and_calc):
                                     focused_peak_y = profile_to_plot_and_calc[focused_peak_x]
                                     self.ax.plot(focused_peak_x, focused_peak_y, 'o', markersize=12, markeredgecolor='orange', markerfacecolor='none', label='UI Focused Peak', zorder=6, linewidth=2) # type: ignore
                         
                         if self.selected_peak_index_for_delete != -1 and self.selected_peak_index_for_delete in valid_peaks_indices:
                             del_peak_y = profile_to_plot_and_calc[self.selected_peak_index_for_delete]
                             self.ax.plot(self.selected_peak_index_for_delete, del_peak_y, 's', markersize=14, 
                                          markeredgecolor='blue', markerfacecolor='none',
                                          label='Selected for Delete', zorder=7, linewidth=2) # type: ignore

                self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
                num_items_to_plot = len(self.peak_regions)
                profile_range_plot = np.ptp(profile_to_plot_and_calc) if np.ptp(profile_to_plot_and_calc) > 0 else 1.0
                
                # --- Variables for text positioning ---
                min_overall_y_for_text = np.min(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 0
                max_overall_y_for_plot_limit = np.max(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 1
                # --- End ---

                for i in range(num_items_to_plot):
                    start, end = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                    if start >= end: self.peak_areas_rolling_ball.append(0.0); self.peak_areas_straight_line.append(0.0); self.peak_areas_valley.append(0.0); continue
                    x_region = np.arange(start, end + 1); profile_region_smoothed = profile_to_plot_and_calc[start : end + 1]
                    
                    # ... (area calculation logic for area_rb, area_sl, area_vv - REMAINS THE SAME) ...
                    bg_start = max(0, min(start, len(self.background)-1)); bg_end = max(0, min(end + 1, len(self.background)))
                    background_region = np.zeros_like(profile_region_smoothed)
                    if bg_start < bg_end and len(self.background) > 0 and interp1d:
                         raw_bg_region = self.background[bg_start:bg_end]
                         if len(raw_bg_region) == len(profile_region_smoothed): background_region = raw_bg_region
                         elif len(self.background) > 1:
                             try: x_full_bg = np.arange(len(self.background)); interp_func_bg = interp1d(x_full_bg, self.background, kind='linear', bounds_error=False, fill_value=(self.background[0], self.background[-1])); background_region = interp_func_bg(x_region)
                             except Exception as interp_err_bg: print(f"Warning: BG interp failed peak {i+1}: {interp_err_bg}")
                    area_rb = max(0, np.trapz(profile_region_smoothed - background_region, x=x_region)) if len(x_region) > 1 else 0.0; self.peak_areas_rolling_ball.append(area_rb)
                    area_sl = 0.0; y_baseline_pts_sl = np.array([0,0]); y_baseline_interp_sl = np.zeros_like(x_region)
                    if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                        y_baseline_pts_sl = np.array([profile_to_plot_and_calc[start], profile_to_plot_and_calc[end]]); y_baseline_interp_sl = np.interp(x_region, [start, end], y_baseline_pts_sl)
                        area_sl = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_sl, x=x_region)) if len(x_region) > 1 else 0.0
                    self.peak_areas_straight_line.append(area_sl)
                    area_vv = 0.0
                    valley_start_anchor_idx, valley_end_anchor_idx = start, end 
                    valley_plot_x_coords = [start, end]
                    y_baseline_pts_vv = np.array([profile_to_plot_and_calc[start], profile_to_plot_and_calc[end]])
                    if i < len(self.initial_valley_regions): 
                        valley_start_anchor_idx, valley_end_anchor_idx = self.initial_valley_regions[i]
                        valley_start_anchor_idx = max(0, min(valley_start_anchor_idx, len(profile_to_plot_and_calc) - 1))
                        valley_end_anchor_idx = max(0, min(valley_end_anchor_idx, len(profile_to_plot_and_calc) - 1))
                        if valley_end_anchor_idx > valley_start_anchor_idx:
                             y_baseline_pts_vv = np.array([profile_to_plot_and_calc[valley_start_anchor_idx], profile_to_plot_and_calc[valley_end_anchor_idx]])
                             valley_plot_x_coords = [valley_start_anchor_idx, valley_end_anchor_idx]
                    y_baseline_interp_vv = np.interp(x_region, valley_plot_x_coords, y_baseline_pts_vv)
                    area_vv = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_vv, x=x_region)) if len(x_region) > 1 else 0.0
                    self.peak_areas_valley.append(area_vv)


                    current_area = 0.0; text_x_pos = (start + end) / 2.0
                    # Determine y_position for text (above the peak or fill area)
                    peak_top_y_in_region = np.max(profile_region_smoothed) if len(profile_region_smoothed) > 0 else 0

                    if self.method == "Rolling Ball":
                        if i == 0: self.ax.plot(np.arange(len(self.background)), self.background, color="green", ls="--", lw=1, label="Rolling Ball BG")
                        self.ax.fill_between(x_region, background_region, profile_region_smoothed, where=profile_region_smoothed >= background_region, color="yellow", alpha=0.4, interpolate=True)
                        current_area = area_rb
                    elif self.method == "Straight Line":
                        if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                            self.ax.plot([start, end], y_baseline_pts_sl, color="purple", ls="--", lw=1, label="SL BG" if i == 0 else "")
                            self.ax.fill_between(x_region, y_baseline_interp_sl, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_sl, color="cyan", alpha=0.4, interpolate=True)
                            current_area = area_sl
                    elif self.method == "Valley-to-Valley":
                        self.ax.plot(valley_plot_x_coords, y_baseline_pts_vv, color="orange", ls="--", lw=1, label="Valley BG" if i == 0 else "")
                        self.ax.fill_between(x_region, y_baseline_interp_vv, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_vv, color="lightblue", alpha=0.4, interpolate=True)
                        current_area = area_vv
                    
                    # --- Improved Text Placement ---
                    area_text_format = "{:.0f}" # Keep area as integer for display
                    # combined_text = f"Peak {i + 1}\n{area_text_format.format(current_area)}" # Show peak number and area
                    combined_text = f"{area_text_format.format(current_area)}" # Just area for cleaner look

                    # Place text slightly above the peak's highest point within its region
                    text_y_offset_above_peak = profile_range_plot * 0.03 # Small offset above the peak
                    text_y_pos = peak_top_y_in_region + text_y_offset_above_peak
                    
                    self.ax.text(text_x_pos, text_y_pos, combined_text, 
                                 ha="center", va="bottom", # Anchor text at bottom-center
                                 fontsize=7, color='black', zorder=6,
                                 bbox=dict(boxstyle="round,pad=0.2", fc="white", ec="gray", alpha=0.7)) # Optional: add a light background box
                    
                    # Update overall Y limits tracking
                    min_overall_y_for_text = min(min_overall_y_for_text, np.min(profile_to_plot_and_calc[start:end+1]) if end+1 <= len(profile_to_plot_and_calc) else 0) # Consider baseline for min
                    max_overall_y_for_plot_limit = max(max_overall_y_for_plot_limit, text_y_pos + profile_range_plot * 0.05) # Ensure text is visible
                    # --- End Improved Text Placement ---

                    self.ax.axvline(start, color="gray", ls=":", lw=1.0, alpha=0.8); self.ax.axvline(end, color="gray", ls=":", lw=1.0, alpha=0.8)
                
                self.ax.set_ylabel("Intensity (Smoothed, Inverted)")
                self.ax.legend(fontsize='small', loc='upper right')
                self.ax.set_title(f"Smoothed Intensity Profile (σ={self.smoothing_sigma:.1f}) and Peak Regions")
                
                # --- X-axis tick and label management ---
                self.ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False) # Hide x-ticks on TOP plot
                ax_image.set_xlabel("Pixel Index Along Profile Axis") # Set xlabel on BOTTOM plot
                # --- End ---

                if len(profile_to_plot_and_calc) > 1: self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1)
                
                # --- Y-axis limit adjustment ---
                # Use overall min/max determined during text placement for robust limits
                y_padding_factor = 0.05 # 5% padding
                y_range_for_padding = max_overall_y_for_plot_limit - min_overall_y_for_text
                if y_range_for_padding <= 0 : y_range_for_padding = 1.0 # Avoid zero or negative range for padding

                y_max_limit = max_overall_y_for_plot_limit + y_range_for_padding * y_padding_factor
                y_min_limit = min_overall_y_for_text - y_range_for_padding * y_padding_factor
                
                if y_max_limit <= y_min_limit: y_max_limit = y_min_limit + 1.0 
                self.ax.set_ylim(y_min_limit, y_max_limit)
                # --- End Y-axis Limit Adjustment ---

                if np.max(profile_to_plot_and_calc) > 10000: self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                
                ax_image.clear() 
                if hasattr(self, 'cropped_image_for_display') and isinstance(self.cropped_image_for_display, Image.Image):
                     try:
                         rotated_pil_image = self.cropped_image_for_display.rotate(90, expand=True); im_array_disp = np.array(rotated_pil_image)
                         if self.original_max_value == 1.0 and np.issubdtype(self.intensity_array_original_range.dtype, np.floating): # type: ignore
                              im_vmin, im_vmax = 0.0, 1.0
                         else: im_vmin, im_vmax = 0, self.original_max_value
                         ax_image.imshow(im_array_disp, cmap='gray', aspect='auto', 
                                         extent=[0, len(profile_to_plot_and_calc)-1 if len(profile_to_plot_and_calc)>0 else 0, 0, rotated_pil_image.height], 
                                         vmin=im_vmin, vmax=im_vmax) 
                         ax_image.set_yticks([]); ax_image.set_ylabel("Lane Width", fontsize='small') 
                     except Exception as img_e: print(f"Error displaying cropped image preview: {img_e}"); ax_image.text(0.5, 0.5, 'Error loading preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([]) 
                else: ax_image.text(0.5, 0.5, 'No Image Preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([]) 
                
                # self.fig.tight_layout(pad=0.5) # May not be needed with GridSpec hspace
                try: self.canvas.draw_idle()
                except Exception as draw_e: print(f"Error drawing canvas: {draw_e}")
                plt.close(self.fig) # type: ignore

            def _custom_rolling_ball(self, profile, radius):
                if grey_opening is None: return np.zeros_like(profile)
                if profile is None or profile.ndim != 1 or profile.size == 0: return np.zeros_like(profile) if profile is not None else np.array([])
                if radius <= 0: return np.zeros_like(profile)
                profile_len = profile.shape[0]; structure_size = int(max(1, 2 * radius + 1))
                if structure_size > profile_len: structure_size = profile_len
                try: background = grey_opening(profile, size=structure_size, mode='reflect')
                except Exception as e: print(f"Error during morphological opening: {e}"); traceback.print_exc(); background = np.zeros_like(profile)
                return background
            
                
            


        class LiveViewLabel(QLabel):
            def __init__(self, font_type, font_size, marker_color, app_instance, parent=None): # Added app_instance
                super().__init__(parent)
                self.app_instance = app_instance # Store the CombinedSDSApp instance
                self.setMouseTracking(True)
                self.preview_marker_enabled = False
                self.preview_marker_text = ""
                self.preview_marker_position = None
                self.marker_font_type = font_type
                self.marker_font_size = font_size
                self.marker_color = marker_color
                self.setFocusPolicy(Qt.StrongFocus)
                self.mw_predict_preview_enabled = False
                self.mw_predict_preview_position = None # QPointF in unzoomed label space
                self.bounding_box_preview = []
                self.measure_quantity_mode = False
                self.counter = 0
                self.zoom_level = 1.0
                self.pan_offset = QPointF(0, 0)
                self.is_panning = False
                self.pan_start_view_coords = None # Stores QPoint of mouse press in widget coords
                self.pan_offset_at_drag_start = None # Stores QPointF of pan_offset at drag start
                self.quad_points = []
                self.selected_point = -1
                self.drag_threshold = 10
                self.bounding_box_complete = False
                self.mode=None
                self.rectangle_start = None
                self.rectangle_end = None
                self.rectangle_points = []
                self.drag_start_pos = None
                self.draw_edges=True
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None
            def clear_crop_preview(self):
                """Clears any existing crop rectangle preview."""
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None
                self.update() # Trigger repaint

            def start_crop_preview(self, start_point_view):
                """Initiates drawing the crop rectangle preview."""
                self.drawing_crop_rect = True
                self.crop_rect_start_view = start_point_view
                self.crop_rect_end_view = start_point_view # Start and end are same initially
                self.crop_rect_final_view = None # Clear any finalized rect
                self.update()

            def update_crop_preview(self, start_point_view, current_point_view):
                """Updates the crop rectangle preview while dragging."""
                if self.drawing_crop_rect:
                    # Keep start point fixed, update end point
                    self.crop_rect_start_view = start_point_view
                    self.crop_rect_end_view = current_point_view
                    self.update()

            def finalize_crop_preview(self, start_point_view, end_point_view):
                """Stores the final rectangle dimensions for persistent preview."""
                self.drawing_crop_rect = False
                # Store as a normalized QRectF for easier drawing
                self.crop_rect_final_view = QRectF(start_point_view, end_point_view).normalized()
                # Clear the temporary points used during drawing
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.update()
                
            def wheelEvent(self, event: 'QWheelEvent'): # Add type hint for clarity
                if not self.app_instance or not self.app_instance.image: # Ensure app and image exist
                    super().wheelEvent(event)
                    return
            
                # --- Determine Zoom Factor ---
                # event.angleDelta().y() is typically +/- 120 for one notch on a mouse wheel
                num_degrees = event.angleDelta().y() / 8
                num_steps = num_degrees / 15.0 # Standard step factor
            
                zoom_factor_increment = 1.1
                zoom_factor_decrement = 1 / 1.1
            
                # --- Calculate new zoom level ---
                old_zoom_level = self.zoom_level
                if num_steps > 0: # Zooming in
                    self.zoom_level *= (zoom_factor_increment ** abs(num_steps))
                elif num_steps < 0: # Zooming out
                    self.zoom_level *= (zoom_factor_decrement ** abs(num_steps))
            
                # --- Clamp Zoom Level (optional, but good practice) ---
                min_zoom = 0.1 # Example minimum zoom
                max_zoom = 20.0  # Example maximum zoom
                self.zoom_level = max(min_zoom, min(self.zoom_level, max_zoom))
            
                # --- Zoom Towards Mouse Cursor ---
                # If we are zooming, the point under the mouse cursor should ideally stay in the same place
                # in the view. This requires adjusting the pan_offset.
                
                # Mouse position in widget coordinates
                mouse_point_widget = event.pos() 
            
                # Point in "unzoomed label space" before this zoom operation
                # This uses the OLD zoom level and OLD pan_offset
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_point_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_point_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
            
                # After zoom, this same unzoomed_label_space point, if it were to stay under the cursor,
                # would satisfy: mouse_point_widget.x() = new_pan_offset.x() + point_before_zoom_unzoomed_label.x() * new_zoom_level
                # So, new_pan_offset.x() = mouse_point_widget.x() - point_before_zoom_unzoomed_label.x() * new_zoom_level
                
                new_pan_x = mouse_point_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_point_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                # --- Reset pan if zoom is back to (or below) 1.0 ---
                if self.zoom_level <= 1.0: # Use <= to catch cases where it might go slightly below due to float math
                    self.zoom_level = 1.0 # Clamp to 1.0 exactly
                    self.pan_offset = QPointF(0, 0) # Reset pan
            
                # --- Update Cursor and View ---
                if not self.is_panning: # Don't change cursor if a pan operation is also in progress (unlikely with wheel)
                    if self.zoom_level > 1.0:
                        self.setCursor(Qt.OpenHandCursor)
                    else:
                        self.setCursor(Qt.ArrowCursor)
                
                if self.app_instance and hasattr(self.app_instance, 'update_live_view'):
                    self.app_instance.update_live_view()
                    QApplication.processEvents() # May still be needed for smooth visual update
                else:
                    self.update()
            
                event.accept()
                
            def mouseMoveEvent(self, event: 'QMouseEvent'):
                if self.is_panning and (event.buttons() & Qt.RightButton):
                    if self.pan_start_view_coords:
                        delta = event.pos() - self.pan_start_view_coords
                        self.pan_offset = self.pan_offset_at_drag_start + delta
                        if self.app_instance:
                            self.app_instance.update_live_view()
                            QApplication.processEvents() 
                    event.accept()
                    return
            
                # Handle left-button drags or general mouse moves for other modes
                if hasattr(self, '_custom_mouseMoveEvent_from_app') and self._custom_mouseMoveEvent_from_app:
                    self._custom_mouseMoveEvent_from_app(event)
                    if event.isAccepted():
                        return
                if self.preview_marker_enabled: # Custom marker preview
                    untransformed_label_pos = self.transform_point(event.pos())
                    snapped_label_pos = untransformed_label_pos
                    if self.app_instance and isinstance(self.app_instance, CombinedSDSApp):
                        snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                    self.preview_marker_position = snapped_label_pos
                    self.update()
                
                elif self.mw_predict_preview_enabled: # <<<< NEW MW PREDICT PREVIEW
                    # Forward to app_instance's handler if it exists (it should)
                    if self.app_instance and hasattr(self.app_instance, 'update_mw_predict_preview'):
                        self.app_instance.update_mw_predict_preview(event) # Pass the event
                    else: # Fallback if handler is missing (should not happen)
                        untransformed_label_pos = self.transform_point(event.pos())
                        snapped_label_pos = untransformed_label_pos
                        if self.app_instance: snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                        self.mw_predict_preview_position = snapped_label_pos
                        self.update()
        
                elif self.selected_point != -1 and self.measure_quantity_mode and self.mode=="quad":
                    current_mouse_label_space = self.transform_point(event.pos())
                    snapped_mouse_label_space = current_mouse_label_space
                    if self.app_instance and isinstance(self.app_instance, CombinedSDSApp):
                        snapped_mouse_label_space = self.app_instance.snap_point_to_grid(current_mouse_label_space)
                    self.quad_points[self.selected_point] = snapped_mouse_label_space
                    self.update()
                
                # Add other drawing mode previews here if needed (e.g., shape drawing)
                elif self.app_instance and self.app_instance.drawing_mode in ['line', 'rectangle'] and \
                     self.app_instance.current_drawing_shape_preview:
                     if hasattr(self.app_instance, 'update_shape_draw'): # Check if app_instance handles it
                         self.app_instance.update_shape_draw(event)
                
                elif self.mode == "move" and self.app_instance and hasattr(self.app_instance, 'move_selection'): # For moving quad/rect
                     self.app_instance.move_selection(event)
                
                super().mouseMoveEvent(event)
                
            def mousePressEvent(self, event: 'QMouseEvent'): # Add type hint
                # --- PRIORITY 1: Right-Click Panning ---
                if event.button() == Qt.RightButton and self.zoom_level > 1.0:
                    self.is_panning = True
                    self.pan_start_view_coords = event.pos()
                    self.pan_offset_at_drag_start = QPointF(self.pan_offset)
                    self.setCursor(Qt.ClosedHandCursor)
                    event.accept() 
                    return #
                if hasattr(self, '_custom_left_click_handler_from_app') and self._custom_left_click_handler_from_app:
                    # This assumes CombinedSDSApp sets self.live_view_label._custom_left_click_handler_from_app
                    # when it enables a specific mode (like add_band, start_rectangle etc.)
                    self._custom_left_click_handler_from_app(event) # Call the app's designated handler
                    if event.isAccepted():
                        return
                if self.preview_marker_enabled:
                    # Use the stored app_instance to call the method
                    if self.app_instance and hasattr(self.app_instance, 'place_custom_marker'):
                        self.app_instance.place_custom_marker(event, self.preview_marker_text)
                    else:
                        print("ERROR: LiveViewLabel cannot call place_custom_marker. app_instance not set or method missing.")
                    self.update()  # Clear the preview (or refresh)
        
                elif self.measure_quantity_mode and self.mode=="quad":
                    # parent_app = self.parent() # Old
                    clicked_label_point_transformed = self.transform_point(event.pos())
                    snapped_click_point = clicked_label_point_transformed
                    if self.app_instance and isinstance(self.app_instance, CombinedSDSApp): # Use app_instance
                        snapped_click_point = self.app_instance.snap_point_to_grid(clicked_label_point_transformed)
        
                    for i, p in enumerate(self.quad_points):
                        if (snapped_click_point - p).manhattanLength() < self.drag_threshold:
                            self.selected_point = i
                            super().mousePressEvent(event) # Call super only if we don't handle it fully
                            return # Return after handling
        
                    if len(self.quad_points) < 4:
                        self.quad_points.append(snapped_click_point)
                        self.selected_point = len(self.quad_points) - 1
        
                    if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                        self.bounding_box_complete = True
                    self.update()
                    
                if not self.preview_marker_enabled and not (self.measure_quantity_mode and self.mode=="quad"):
                    super().mousePressEvent(event)
                else: # If not preview_marker_enabled and not (quad_mode...)
                    super().mousePressEvent(event)
                if not event.isAccepted():
                    super().mousePressEvent(event)

            def mouseReleaseEvent(self, event: 'QMouseEvent'):
                if event.button() == Qt.RightButton and self.is_panning:
                    self.is_panning = False
                    self.pan_start_view_coords = None
                    self.pan_offset_at_drag_start = None
                    if self.zoom_level > 1.0: self.setCursor(Qt.OpenHandCursor)
                    else: self.setCursor(Qt.ArrowCursor)
                    event.accept()
                    return
            
                # Handle left-button release for other modes
                if hasattr(self, '_custom_mouseReleaseEvent_from_app') and self._custom_mouseReleaseEvent_from_app:
                    self._custom_mouseReleaseEvent_from_app(event)
                    if event.isAccepted():
                        return
                if self.mode=="quad":
                    self.selected_point = -1
                    self.update()
                super().mouseReleaseEvent(event)

            def transform_point(self, point):
                """Transform a point from widget coordinates to image coordinates."""
                if self.zoom_level != 1.0:
                    return QPointF(
                        (point.x() - self.pan_offset.x()) / self.zoom_level,
                        (point.y() - self.pan_offset.y()) / self.zoom_level
                    )
                return QPointF(point)
            
            
            def zoom_in(self): # Zooms towards view center by default
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level *= 1.1
                self.zoom_level = min(self.zoom_level, 20.0) # Max zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if not self.is_panning: self.setCursor(Qt.OpenHandCursor)
                
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()

            def zoom_out(self): # Zooms towards view center
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level /= 1.1
                self.zoom_level = max(self.zoom_level, 0.1) # Min zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if self.zoom_level <= 1.0:
                    self.zoom_level = 1.0
                    self.pan_offset = QPointF(0, 0)
                    if not self.is_panning: self.setCursor(Qt.ArrowCursor)
                else:
                    if not self.is_panning: self.setCursor(Qt.OpenHandCursor)
            
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()
                

            def paintEvent(self, event):
                super().paintEvent(event) # Draws the base pixmap set by setPixmap()
                
                painter = QPainter(self)
                painter.setRenderHint(QPainter.Antialiasing, True)
                painter.setRenderHint(QPainter.TextAntialiasing, True)
                painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

                # === Single Transformation Block for all content affected by zoom/pan ===
                painter.save()
                if self.zoom_level != 1.0:
                    painter.translate(self.pan_offset)
                    painter.scale(self.zoom_level, self.zoom_level)

                # --- Common variables for transforming app_instance items to unzoomed label space ---
                _image_to_label_space_valid = False
                _scale_factor_img_to_label = 1.0
                _img_w_orig_from_app = 1.0 # Fallback
                _img_h_orig_from_app = 1.0 # Fallback
                _offset_x_img_in_label = 0.0
                _offset_y_img_in_label = 0.0

                # This context is for elements originating from self.app_instance.image
                if self.app_instance and self.app_instance.image and not self.app_instance.image.isNull():
                    current_app_image = self.app_instance.image
                    label_w_widget = float(self.width())
                    label_h_widget = float(self.height())
                    
                    _img_w_orig_from_app = float(current_app_image.width())
                    _img_h_orig_from_app = float(current_app_image.height())

                    if _img_w_orig_from_app > 0 and _img_h_orig_from_app > 0 and \
                       label_w_widget > 0 and label_h_widget > 0:
                        
                        _scale_factor_img_to_label = min(label_w_widget / _img_w_orig_from_app, 
                                                         label_h_widget / _img_h_orig_from_app)
                        
                        _displayed_img_w_in_label = _img_w_orig_from_app * _scale_factor_img_to_label
                        _displayed_img_h_in_label = _img_h_orig_from_app * _scale_factor_img_to_label
                        
                        _offset_x_img_in_label = (label_w_widget - _displayed_img_w_in_label) / 2.0
                        _offset_y_img_in_label = (label_h_widget - _displayed_img_h_in_label) / 2.0
                        _image_to_label_space_valid = True

                def _app_image_coords_to_unzoomed_label_space(img_coords_tuple_or_qpointf):
                    if not _image_to_label_space_valid:
                        if isinstance(img_coords_tuple_or_qpointf, QPointF): return img_coords_tuple_or_qpointf 
                        return QPointF(img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])

                    img_x, img_y = (img_coords_tuple_or_qpointf.x(), img_coords_tuple_or_qpointf.y()) \
                                   if isinstance(img_coords_tuple_or_qpointf, QPointF) \
                                   else (img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])
                        
                    x_ls = _offset_x_img_in_label + img_x * _scale_factor_img_to_label
                    y_ls = _offset_y_img_in_label + img_y * _scale_factor_img_to_label
                    return QPointF(x_ls, y_ls)

                # --- 0. Draw Standard L/R/Top Markers (from self.app_instance) ---
                if _image_to_label_space_valid and self.app_instance:
                    # Get font settings from app_instance for standard markers
                    std_marker_font = QFont(self.app_instance.font_family, self.app_instance.font_size)
                    std_marker_color = self.app_instance.font_color if hasattr(self.app_instance, 'font_color') else QColor(Qt.black)
                    painter.setFont(std_marker_font)
                    painter.setPen(std_marker_color)
                    
                    font_metrics_std = QFontMetrics(std_marker_font)
                    text_height_std_label_space = font_metrics_std.height() # Height in label space (before zoom)
                    # Effective y_offset for text baseline alignment relative to anchor point
                    # This tries to make the anchor point (y_pos_label_space) near the vertical middle of the text.
                    y_offset_text_baseline_std = text_height_std_label_space * 0.3 # Heuristic, adjust as needed

                    # Left Markers
                    if hasattr(self.app_instance, 'left_markers'):
                        # Offset for left markers is from app_instance, scaled to label space
                        left_marker_offset_x_label_space = self.app_instance.left_marker_shift_added * _scale_factor_img_to_label
                        
                        for y_pos_img, marker_text_val in self.app_instance.left_markers:
                            anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y() # X doesn't matter for Y anchor
                            text_to_draw = f"{marker_text_val} ⎯"
                            text_width_label_space = font_metrics_std.horizontalAdvance(text_to_draw)
                            draw_x_ls = _offset_x_img_in_label + left_marker_offset_x_label_space - text_width_label_space
                            draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                            painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    
                    # Right Markers
                    if hasattr(self.app_instance, 'right_markers'):
                            # right_marker_offset_x_label_space IS the X-coordinate in unzoomed label space
                            # where the text should begin.
                            right_marker_start_x_label_space = _offset_x_img_in_label + \
                                                                (self.app_instance.right_marker_shift_added * _scale_factor_img_to_label)
            
                            for y_pos_img, marker_text_val in self.app_instance.right_markers:
                                anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y()
                                text_to_draw = f"⎯ {marker_text_val}"
                                # No need to subtract text_width here for right markers, as the offset is the start
                                draw_x_ls = right_marker_start_x_label_space
                                draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                                painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)

                    # Top Markers
                    if hasattr(self.app_instance, 'top_markers'):
                        top_marker_offset_y_label = self.app_instance.top_marker_shift_added * _scale_factor_img_to_label
                        rotation_angle = self.app_instance.font_rotation
        
                        for x_pos_img, marker_text_val in self.app_instance.top_markers:
                            anchor_x_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, 0)).x()
                            text_to_draw = str(marker_text_val)
                            painter.save()
                            draw_baseline_y_ls = _offset_y_img_in_label + top_marker_offset_y_label + y_offset_text_baseline_std
                            painter.translate(anchor_x_label_space, draw_baseline_y_ls)
                            painter.rotate(rotation_angle)
                            painter.drawText(QPointF(0, 0), text_to_draw) # Draw at the new origin
                            painter.restore()


                # --- 1. Draw Custom Shapes (from self.app_instance.custom_shapes) ---
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_shapes'):
                    for shape_data in self.app_instance.custom_shapes:
                        try:
                            shape_type = shape_data.get('type')
                            color = QColor(shape_data.get('color', '#000000'))
                            base_thickness = float(shape_data.get('thickness', 0.5))
                            
                            thickness_on_label = base_thickness * _scale_factor_img_to_label 
                            # Pen width is in device pixels. To make it appear `thickness_on_label` wide
                            # in the unzoomed view, and scale with zoom:
                            effective_pen_width = max(0.5, thickness_on_label / self.zoom_level if self.zoom_level > 0 else thickness_on_label)
                                                        
                            pen = QPen(color)
                            pen.setWidthF(effective_pen_width) 
                            painter.setPen(pen)

                            if shape_type == 'line':
                                start_img = shape_data.get('start') 
                                end_img = shape_data.get('end')   
                                if start_img and end_img:
                                    start_label_space = _app_image_coords_to_unzoomed_label_space(start_img)
                                    end_label_space = _app_image_coords_to_unzoomed_label_space(end_img)
                                    painter.drawLine(start_label_space, end_label_space)
                            elif shape_type == 'rectangle':
                                rect_img = shape_data.get('rect') 
                                if rect_img:
                                    x_img, y_img, w_img, h_img = rect_img
                                    top_left_label_space = _app_image_coords_to_unzoomed_label_space((x_img, y_img))
                                    w_label_space = w_img * _scale_factor_img_to_label
                                    h_label_space = h_img * _scale_factor_img_to_label
                                    painter.drawRect(QRectF(top_left_label_space, QSizeF(w_label_space, h_label_space)))
                        except Exception as e:
                            print(f"Error drawing custom shape in LiveViewLabel.paintEvent: {shape_data}, {e}")
                
                # --- Draw Custom Markers (from self.app_instance.custom_markers) ---
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_markers'):
                    for marker_data_list in self.app_instance.custom_markers:
                        try:
                            x_pos_img, y_pos_img, marker_text_str, qcolor_obj, \
                            font_family_str, font_size_int, is_bold, is_italic = marker_data_list
                            
                            anchor_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, y_pos_img))

                            current_marker_font = QFont(font_family_str, font_size_int) # Point size
                            current_marker_font.setBold(is_bold)
                            current_marker_font.setItalic(is_italic)
                            painter.setFont(current_marker_font)
                            
                            if not isinstance(qcolor_obj, QColor): qcolor_obj = QColor(str(qcolor_obj))
                            if not qcolor_obj.isValid(): qcolor_obj = Qt.black 
                            painter.setPen(qcolor_obj)

                            font_metrics_marker = QFontMetrics(current_marker_font)
                            text_bounding_rect_marker = font_metrics_marker.boundingRect(marker_text_str)

                            draw_x_marker = anchor_label_space.x() - (text_bounding_rect_marker.left() + text_bounding_rect_marker.width() / 2.0)
                            draw_y_marker = anchor_label_space.y() - (text_bounding_rect_marker.top() + text_bounding_rect_marker.height() / 2.0)
                            
                            painter.drawText(QPointF(draw_x_marker, draw_y_marker), marker_text_str)
                        except Exception as e:
                            print(f"Error drawing app_instance custom marker in LiveViewLabel.paintEvent: {marker_data_list}, {e}")

                # --- 2. Draw Live Preview Marker (if enabled) ---
                if self.preview_marker_enabled and self.preview_marker_position:
                    painter.setOpacity(0.7) 
                    # Font size for preview marker should be point size, Qt handles scaling
                    marker_preview_font = QFont(self.marker_font_type, self.marker_font_size) 
                    painter.setFont(marker_preview_font)
                    painter.setPen(self.marker_color)
                    
                    font_metrics_preview = QFontMetrics(marker_preview_font)
                    preview_text_rect = font_metrics_preview.boundingRect(self.preview_marker_text)
                    
                    draw_x_preview = self.preview_marker_position.x() - (preview_text_rect.left() + preview_text_rect.width() / 2.0)
                    draw_y_preview = self.preview_marker_position.y() - (preview_text_rect.top() + preview_text_rect.height() / 2.0)
                    
                    painter.drawText(QPointF(draw_x_preview, draw_y_preview), self.preview_marker_text)
                    painter.setOpacity(1.0)

                # --- 3. Draw Quadrilateral Points (if any) ---
                if self.quad_points: 
                    # Pen width should be small in device pixels, so it scales with zoom
                    effective_pen_width_quad = max(0.5, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5)
                    painter.setPen(QPen(Qt.red, effective_pen_width_quad)) 
                    ellipse_radius_view = self.drag_threshold # This is in unzoomed label space
                    for p_label_space in self.quad_points:
                        painter.drawEllipse(p_label_space, ellipse_radius_view, ellipse_radius_view)
                
                    if len(self.quad_points) == 4 and self.draw_edges:
                        effective_pen_width_poly = max(0.5, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5)
                        painter.setPen(QPen(Qt.blue, effective_pen_width_poly))
                        painter.drawPolygon(QPolygonF(self.quad_points))

                # --- 4. Draw Bounding Box Preview ---
                if self.bounding_box_preview: 
                    effective_pen_width_bbox = max(0.5, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5)
                    painter.setPen(QPen(Qt.blue, effective_pen_width_bbox))
                    start_x, start_y, end_x, end_y = self.bounding_box_preview
                    rect_label_space = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y)).normalized()
                    painter.drawRect(rect_label_space)

                # --- 5. Draw Crop Rectangle Preview ---
                preview_pen_crop = QPen(Qt.magenta) 
                effective_pen_width_crop = max(0.5, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5)
                preview_pen_crop.setWidthF(effective_pen_width_crop) 

                if self.drawing_crop_rect and self.crop_rect_start_view and self.crop_rect_end_view:
                    preview_pen_crop.setStyle(Qt.DashLine)
                    painter.setPen(preview_pen_crop)
                    rect_to_draw = QRectF(self.crop_rect_start_view, self.crop_rect_end_view).normalized()
                    painter.drawRect(rect_to_draw)
                elif self.crop_rect_final_view: 
                    preview_pen_crop.setStyle(Qt.SolidLine)
                    painter.setPen(preview_pen_crop)
                    painter.drawRect(self.crop_rect_final_view)
                
                # --- Draw Shape Preview (for line/rectangle drawing mode) ---
                if self.app_instance and self.app_instance.drawing_mode in ['line', 'rectangle'] and \
                   self.app_instance.current_drawing_shape_preview:
                    try:
                        start_pt_ls = self.app_instance.current_drawing_shape_preview['start'] # Already in unzoomed label space
                        end_pt_ls = self.app_instance.current_drawing_shape_preview['end']     # Already in unzoomed label space

                        preview_color = self.app_instance.custom_marker_color
                        # Thickness from spinbox is in points, for vector drawing, this should be fine.
                        # Or, if it's intended as image pixels, scale it by _scale_factor_img_to_label.
                        # Let's assume it's meant as a visual thickness on the label.
                        base_preview_thickness = float(self.app_instance.custom_font_size_spinbox.value())
                        # Scale with zoom
                        effective_preview_thickness = min(1.0, base_preview_thickness / self.zoom_level if self.zoom_level > 0 else base_preview_thickness)

                        preview_pen_shape = QPen(preview_color)
                        preview_pen_shape.setWidthF(effective_preview_thickness)
                        preview_pen_shape.setStyle(Qt.DotLine)
                        painter.setPen(preview_pen_shape)

                        if self.app_instance.drawing_mode == 'line':
                            painter.drawLine(start_pt_ls, end_pt_ls)
                        elif self.app_instance.drawing_mode == 'rectangle':
                            painter.drawRect(QRectF(start_pt_ls, end_pt_ls).normalized())
                    except Exception as e:
                        print(f"Error drawing live shape preview in paintEvent: {e}")

                if self.mw_predict_preview_enabled and self.mw_predict_preview_position:
                    try:
                        # Use similar styling to the actual placed MW marker, but maybe slightly different
                        painter.setOpacity(0.7) # Preview opacity
                        # Font settings could be from app_instance's custom marker settings or fixed
                        mw_preview_font_size = self.app_instance.custom_font_size_spinbox.value() \
                                              if self.app_instance and hasattr(self.app_instance, 'custom_font_size_spinbox') \
                                              else 12
                        mw_preview_font_family = self.app_instance.custom_font_type_dropdown.currentText() \
                                                if self.app_instance and hasattr(self.app_instance, 'custom_font_type_dropdown') \
                                                else "Arial"
                        
                        mw_preview_font = QFont(mw_preview_font_family, mw_preview_font_size)
                        painter.setFont(mw_preview_font)
                        painter.setPen(Qt.darkGreen) # Slightly different color for preview
    
                        font_metrics_mw_preview = QFontMetrics(mw_preview_font)
                        text_mw_preview = "⎯⎯" 
                        text_bounding_rect_mw_preview = font_metrics_mw_preview.boundingRect(text_mw_preview)
                        
                        # self.mw_predict_preview_position is already in unzoomed label space
                        preview_anchor_x_ls = self.mw_predict_preview_position.x()
                        preview_anchor_y_ls = self.mw_predict_preview_position.y()
    
                        draw_x_mw_preview_ls = preview_anchor_x_ls - (text_bounding_rect_mw_preview.left() + text_bounding_rect_mw_preview.width() / 2.0)
                        draw_y_mw_preview_ls = preview_anchor_y_ls - (text_bounding_rect_mw_preview.top() + text_bounding_rect_mw_preview.height() / 2.0)
    
                        painter.drawText(QPointF(draw_x_mw_preview_ls, draw_y_mw_preview_ls), text_mw_preview)
                        painter.setOpacity(1.0)
                    except Exception as e:
                        print(f"Error drawing MW prediction *preview* marker in paintEvent: {e}")
    
    
                # --- Draw Molecular Weight Prediction *FINAL PLACED* Marker ---
                # (Keep the existing logic for the final placed marker)
                if self.app_instance and hasattr(self.app_instance, "protein_location") and \
                   self.app_instance.protein_location and not self.app_instance.run_predict_MW: # Only if not yet finalized
                    try:
                        loc_x_ls, loc_y_ls = self.app_instance.protein_location # Unzoomed label space
                        mw_marker_font = QFont(self.app_instance.custom_font_type_dropdown.currentText(), 
                                               self.app_instance.custom_font_size_spinbox.value() + 2)
                        painter.setFont(mw_marker_font)
                        painter.setPen(Qt.green) 
                        font_metrics_mw = QFontMetrics(mw_marker_font)
                        text_mw = "⎯⎯"
                        text_bounding_rect_mw = font_metrics_mw.boundingRect(text_mw)
                        draw_x_mw_ls = loc_x_ls - (text_bounding_rect_mw.left() + text_bounding_rect_mw.width() / 2.0)
                        draw_y_mw_ls = loc_y_ls - (text_bounding_rect_mw.top() + text_bounding_rect_mw.height() / 2.0)
                        painter.drawText(QPointF(draw_x_mw_ls, draw_y_mw_ls), text_mw)
                    except Exception as e:
                        print(f"Error drawing placed MW prediction marker in paintEvent: {e}")

                # --- 8. Draw Grid Lines ---
                if self.app_instance and \
                   hasattr(self.app_instance, 'grid_size_input') and \
                   hasattr(self.app_instance, 'show_grid_checkbox_x') and \
                   hasattr(self.app_instance, 'show_grid_checkbox_y'):
                    
                    # Grid size from app_instance is in *unzoomed label space pixels*
                    grid_size_label_space = self.app_instance.grid_size_input.value()
                    
                    if grid_size_label_space > 0:
                        pen_grid_paint = QPen(Qt.red) # Color for grid
                        pen_grid_paint.setStyle(Qt.DashLine)
                        # Pen width should be thin in device pixels (after zoom)
                        effective_pen_width_grid = max(0.5, 1.0 / self.zoom_level if self.zoom_level > 0 else 1.0) # Ensure it's visible
                        pen_grid_paint.setWidthF(effective_pen_width_grid)
                        painter.setPen(pen_grid_paint)
    
                        # Grid lines should span the entire *visible* area of the label widget.
                        # The painter is already transformed by zoom/pan, so we draw in the
                        # unzoomed label space from (0,0) to (self.width(), self.height()).
                        label_width_unzoomed = self.width() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        label_height_unzoomed = self.height() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        
                        # Determine the drawing boundaries in unzoomed label space that correspond
                        # to the current viewport after panning.
                        view_origin_x_unzoomed = -self.pan_offset.x() / (self.zoom_level if self.zoom_level > 0 else 1.0)
                        view_origin_y_unzoomed = -self.pan_offset.y() / (self.zoom_level if self.zoom_level > 0 else 1.0)

                        if self.app_instance.show_grid_checkbox_x.isChecked():
                            # Start drawing from the first grid line visible in the viewport
                            start_x_grid = (int(view_origin_x_unzoomed / grid_size_label_space) -1) * grid_size_label_space 
                            for x_grid_ls in range(start_x_grid, int(view_origin_x_unzoomed + label_width_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(x_grid_ls, view_origin_y_unzoomed), 
                                                 QPointF(x_grid_ls, view_origin_y_unzoomed + label_height_unzoomed))
    
                        if self.app_instance.show_grid_checkbox_y.isChecked():
                            start_y_grid = (int(view_origin_y_unzoomed / grid_size_label_space)-1) * grid_size_label_space
                            for y_grid_ls in range(start_y_grid, int(view_origin_y_unzoomed + label_height_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(view_origin_x_unzoomed, y_grid_ls), 
                                                 QPointF(view_origin_x_unzoomed + label_width_unzoomed, y_grid_ls))
                
                painter.restore()

            def keyPressEvent(self, event):
                if event.key() == Qt.Key_Escape:  
                    parent_app = self.parent()
                    if isinstance(parent_app, CombinedSDSApp) and parent_app.crop_rectangle_mode:
                        parent_app.cancel_rectangle_crop_mode() # Call the app's cancel method
                        return # Prevent further processing in this case
                    if self.preview_marker_enabled:
                        self.preview_marker_enabled = False  # Turn off the preview
                        self.update()  # Clear the overlay                    
                    self.measure_quantity_mode = False
                    self.counter = 0
                    self.bounding_box_complete = False
                    self.quad_points = []
                    self.mode=None
                    self.clear_crop_preview()
                    self.update()
                super().keyPressEvent(event)

        class CombinedSDSApp(QMainWindow):
            CONFIG_PRESET_FILE_NAME = "Imaging_assistant_preset_config.txt"
            MIME_TYPE_CUSTOM_ITEMS = "application/x-imaging-assistant.customitems+json"
            def __init__(self):
                super().__init__()
                self.screen = QDesktopWidget().screenGeometry()
                self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
                # window_width = int(self.screen_width * 0.5)  # 60% of screen width
                # window_height = int(self.screen_height * 0.75)  # 95% of screen height
                self.preview_label_width_setting = int(self.screen_width * 0.35)
                self.preview_label_max_height_setting = int(self.screen_height * 0.30)
                self.label_size = self.preview_label_width_setting
                self.window_title="IMAGING ASSISTANT V8.0"
                # --- Initialize Status Bar Labels ---
                self.size_label = QLabel("Image Size: N/A")
                self.depth_label = QLabel("Bit Depth: N/A")
                self.location_label = QLabel("Source: N/A")
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                

                # --- Add Labels to Status Bar ---
                statusbar = self.statusBar() # Get the status bar instance
                statusbar.addWidget(self.size_label)
                statusbar.addWidget(self.depth_label)
                # Add location label with stretch factor 1 to push it to the right
                statusbar.addWidget(self.location_label, 1)
                self.setWindowTitle(self.window_title)
                self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
                self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
                self.undo_stack = []
                self.redo_stack = []
                self.custom_shapes = []  # NEW: List to store lines/rectangles
                self.drawing_mode = None # NEW: None, 'line', 'rectangle'
                self.current_drawing_shape_preview = None # NEW: For live preview data
                self.quantities_peak_area_dict = {}
                self.latest_peak_areas = []
                self.latest_calculated_quantities = []
                self.image_path = None
                self.image = None
                self.image_master= None
                self.image_before_padding = None
                self.image_contrasted=None
                self.image_before_contrast=None
                self.contrast_applied=False
                self.image_padded=False
                self.predict_size=False
                self.warped_image=None
                self.transparency=1
                self.left_markers = []
                self.right_markers = []
                self.top_markers = []
                self.custom_markers=[]
                self.current_left_marker_index = 0
                self.current_right_marker_index = 0
                self.current_top_label_index = 0
                self.font_rotation=-45
                self.image_width=0
                self.image_height=0
                self.new_image_width=0
                self.new_image_height=0
                self.base_name="Image"
                self.image_path=""
                self.x_offset_s=0
                self.y_offset_s=0
                self.peak_area=[]
                self.auto_marker_side = None # To store 'left' or 'right' for auto mode
                self.auto_lane_quad_points = [] # Store points defined for auto lane
                self.is_modified=False
                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None # Temp storage for starting point in view coords
                self.crop_rectangle_coords = None # Stores final (x, y, w, h) in *image* coordinates
                self.crop_offset_x = 0
                self.crop_offset_y = 0
                
                self.copied_peak_regions_data = {
                    "regions": None,        # List of (start, end) tuples
                    "profile_length": None  # Length of the profile from which regions were copied
                }
                self.peak_dialog_settings = {
                    'rolling_ball_radius': 50,
                    'peak_height_factor': 0.1,
                    'peak_distance': 10,
                    'peak_prominence_factor': 0.02,
                    'valley_offset_pixels': 0,  # Added to persist settings
                    'band_estimation_method': "Mean",
                    'area_subtraction_method': "Rolling Ball",
                    'smoothing_sigma': 2.0, # Added to persist settings
                }
                self.persist_peak_settings_enabled = True # State of the checkbox
                
                
                # Variables to store bounding boxes and quantities
                self.bounding_boxes = []
                self.up_bounding_boxes = []
                self.standard_protein_areas=[]
                self.quantities = []
                self.protein_quantities = []
                self.measure_quantity_mode = False
                # Initialize self.marker_values to None initially
                self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"] # Default internal top_label
                self.marker_values = [] # Default internal marker_values (for L/R) - will be populated by preset

                self.custom_marker_name = "Custom"
                self.presets_data = {}
                self.custom_marker_name = "Custom"
                # Load the config file if exists
                
                self.marker_mode = None
                self.left_marker_shift = 0   # Additional shift for marker text
                self.right_marker_shift = 0   # Additional shift for marker tex
                self.top_marker_shift=0 
                self.left_marker_shift_added=0
                self.right_marker_shift_added=0
                self.top_marker_shift_added= 0
                self.left_slider_range=[-1000,1000]
                self.right_slider_range=[-1000,1000]
                self.top_slider_range=[-1000,1000]
                       
                
                self.top_padding = 0
                self.font_color = QColor(0, 0, 0)  # Default to black
                self.custom_marker_color = QColor(0, 0, 0)  # Default to black
                self.font_family = "Arial"  # Default font family
                self.font_size = 12  # Default font size
                self.image_array_backup= None
                self.run_predict_MW=False
                
                
                
                # Main container widget
                main_widget = QWidget()
                self.setCentralWidget(main_widget)
                layout = QVBoxLayout(main_widget)

                # Upper section (Preview and buttons)
                upper_layout = QHBoxLayout()

                self.label_width=int(self.label_size)
            
                self.live_view_label = LiveViewLabel(
                    font_type=QFont("Arial"),
                    font_size=int(24),
                    marker_color=QColor(0,0,0),
                    app_instance=self, # Pass the CombinedSDSApp instance
                    parent=self,
                )
                # Image display
                self.live_view_label.setStyleSheet("background-color: white; border: 1px solid black;")
                
                self._create_actions()
                #self.create_menu_bar()
                self.create_tool_bar()
                
                self._update_preview_label_size()
                upper_layout.addWidget(self.live_view_label, stretch=1)
                layout.addLayout(upper_layout)
                
                # Lower section (Tabbed interface)
                self.tab_widget = QTabWidget()
                # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
                self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
                self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
                self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
                self.tab_widget.addTab(self.create_markers_tab(), "Markers")
                self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
                self.tab_widget.addTab(self.analysis_tab(), "Analysis")
                
                layout.addWidget(self.tab_widget)
                
                #self.load_shortcut = QShortcut(QKeySequence.Open, self) # Ctrl+O / Cmd+O
                #self.load_shortcut.activated.connect(self.load_action.trigger) # Trigger the action

                #self.save_shortcut = QShortcut(QKeySequence.Save, self) # Ctrl+S / Cmd+S
                #self.save_shortcut.activated.connect(self.save_action.trigger) # Trigger the action

                #self.copy_shortcut = QShortcut(QKeySequence.Copy, self) # Ctrl+C / Cmd+C
                #self.copy_shortcut.activated.connect(self.copy_action.trigger) # Trigger the action

                #self.paste_shortcut = QShortcut(QKeySequence.Paste, self) # Ctrl+V / Cmd+V
                #self.paste_shortcut.activated.connect(self.paste_action.trigger) # Trigger the action

                #self.undo_shortcut = QShortcut(QKeySequence.Undo, self) # Ctrl+Z / Cmd+Z
                #self.undo_shortcut.activated.connect(self.undo_action_m) # Connect directly to method

                #self.redo_shortcut = QShortcut(QKeySequence.Redo, self) # Ctrl+Y / Cmd+Shift+Z
                #self.redo_shortcut.activated.connect(self.redo_action_m) # Connect directly to method

                #self.zoom_out_shortcut = QShortcut(QKeySequence.ZoomOut, self) # Ctrl+- / Cmd+-
                #self.zoom_out_shortcut.activated.connect(self.zoom_out_action.trigger) # Trigger the action

                # Zoom In - Keep the standard one
                #self.zoom_in_shortcut = QShortcut(QKeySequence.ZoomIn, self) # Attempts Ctrl++ / Cmd++
                #self.zoom_in_shortcut.activated.connect(self.zoom_in_action.trigger) # Trigger the action
                # ======================================================


                # === KEEP QShortcut definitions for NON-STANDARD actions ===
                # self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
                # self.save_svg_shortcut.activated.connect(self.save_svg_action.trigger)

                self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
                self.reset_shortcut.activated.connect(self.reset_action.trigger)

                self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
                self.predict_shortcut.activated.connect(self.predict_molecular_weight)

                self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
                self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)

                self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
                self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)

                self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
                self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)

                self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
                self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)

                self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
                self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
                self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
                self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
                self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
                self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
                self.grid_shortcut.activated.connect(
                    lambda: (
                        # Determine the target state (True if both are currently False, False otherwise)
                        target_state := not (self.show_grid_checkbox_x.isChecked() or self.show_grid_checkbox_y.isChecked()),
                        # Set both checkboxes to the target state
                        self.show_grid_checkbox_x.setChecked(target_state),
                        self.show_grid_checkbox_y.setChecked(target_state)
                    ) if hasattr(self, 'show_grid_checkbox_x') and hasattr(self, 'show_grid_checkbox_y') else None
                )
                
                self.grid_shortcut_x = QShortcut(QKeySequence("Ctrl+Shift+X"), self) # Use uppercase X for consistency
                self.grid_shortcut_x.activated.connect(
                    lambda: self.show_grid_checkbox_x.setChecked(not self.show_grid_checkbox_x.isChecked())
                    if hasattr(self, 'show_grid_checkbox_x') else None
                )

                # Shortcut for Y Grid Snapping
                self.grid_shortcut_y = QShortcut(QKeySequence("Ctrl+Shift+Y"), self) # Use uppercase Y
                self.grid_shortcut_y.activated.connect(
                    lambda: self.show_grid_checkbox_y.setChecked(not self.show_grid_checkbox_y.isChecked())
                    if hasattr(self, 'show_grid_checkbox_y') else None
                )

                self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
                self.guidelines_shortcut.activated.connect(
                    lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked())
                    if hasattr(self, 'show_guides_checkbox') else None
                )

                self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
                self.increase_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1)
                    if hasattr(self, 'grid_size_input') else None
                )
                self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
                self.decrease_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1)
                    if hasattr(self, 'grid_size_input') else None
                )

                self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
                self.invert_shortcut.activated.connect(self.invert_image)

                self.bw_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
                self.bw_shortcut.activated.connect(self.convert_to_black_and_white)

                # Tab movement shortcuts
                self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
                self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
                self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
                self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
                self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
                self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
                self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
                self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
                self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
                self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
                self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
                self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
                
                self.load_config()
            
            def _create_actions(self):
                """Create QAction objects for menus and toolbars."""
                style = self.style() # Still useful for other icons
                icon_size = QSize(30, 30) # Match your toolbar size
                text_color = self.palette().color(QPalette.ButtonText) # Use theme color

                # # --- Create Zoom Icons (using previous method) ---
                # zoom_in_pixmap = QPixmap(icon_size)
                # zoom_in_pixmap.fill(Qt.transparent)
                # painter_in = QPainter(zoom_in_pixmap)
                # # ... (painter setup as before) ...
                # painter_in.drawText(zoom_in_pixmap.rect(), Qt.AlignCenter, "+")
                # painter_in.end()
                # zoom_in_icon = QIcon(zoom_in_pixmap)

                # zoom_out_pixmap = QPixmap(icon_size)
                # zoom_out_pixmap.fill(Qt.transparent)
                # painter_out = QPainter(zoom_out_pixmap)
                # # ... (painter setup as before) ...
                # painter_out.drawText(zoom_out_pixmap.rect(), Qt.AlignCenter, "-")
                # painter_out.end()
                # zoom_out_icon = QIcon(zoom_out_pixmap)
                # --- End Zoom Icons ---


                # --- Create Icons using the helper ---
                open_icon = create_text_icon("Wingdings",icon_size, text_color, "1") # Unicode Right Arrow
                save_icon = create_text_icon("Wingdings",icon_size, text_color, "=")
                save_svg_icon = create_text_icon("Wingdings",icon_size, text_color, "3")
                undo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "O")
                redo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "N")
                paste_icon = create_text_icon("Wingdings 2",icon_size, text_color, "2")
                copy_icon = create_text_icon("Wingdings",icon_size, text_color, "4")
                reset_icon = create_text_icon("Wingdings 3",icon_size, text_color, "Q")
                exit_icon = create_text_icon("Wingdings 2",icon_size, text_color, "V")
                
                zoom_in_icon = create_text_icon("Arial",icon_size, text_color, "+")
                zoom_out_icon = create_text_icon("Arial",icon_size, text_color, "-")
                pan_up_icon = create_text_icon("Arial",icon_size, text_color, "↑") # Unicode Up Arrow
                pan_down_icon = create_text_icon("Arial",icon_size, text_color, "↓") # Unicode Down Arrow
                pan_left_icon = create_text_icon("Arial",icon_size, text_color, "←") # Unicode Left Arrow
                pan_right_icon = create_text_icon("Arial",icon_size, text_color, "→") # Unicode Right Arrow
                bounding_box_icon = create_text_icon("Wingdings 2", icon_size, text_color, "0")
                draw_line_icon = create_text_icon("Arial", icon_size, text_color, "__")
                info_icon = create_text_icon("Wingdings", icon_size, text_color, "'")
                copy_custom_icon = create_text_icon("Wingdings", icon_size, text_color, "B") # Page with arrow out
                paste_custom_icon = create_text_icon("Wingdings", icon_size, text_color, "A") # Page with arrow in


                # --- File Actions ---
                self.load_action = QAction(open_icon, "&Load Image...", self)
                self.save_action = QAction(save_icon, "&Save with Config", self)
                # self.save_svg_action = QAction(save_svg_icon, "Save &SVG...", self)
                self.reset_action = QAction(reset_icon, "&Reset Image", self)
                self.exit_action = QAction(exit_icon, "E&xit", self)

                # --- Edit Actions ---
                self.undo_action = QAction(undo_icon, "&Undo", self) # Standard for now
                self.redo_action = QAction(redo_icon, "&Redo", self) # Standard for now
                self.copy_action = QAction(copy_icon, "&Copy Image", self)
                self.paste_action = QAction(paste_icon, "&Paste Image", self)

                # --- View Actions (using the created icons) ---
                self.zoom_in_action = QAction(zoom_in_icon, "Zoom &In", self)
                self.zoom_out_action = QAction(zoom_out_icon, "Zoom &Out", self)
                self.pan_left_action = QAction(pan_left_icon, "Pan Left", self)
                self.pan_right_action = QAction(pan_right_icon, "Pan Right", self)
                self.pan_up_action = QAction(pan_up_icon, "Pan Up", self)
                self.pan_down_action = QAction(pan_down_icon, "Pan Down", self)
                # --- END: Add Panning Actions ---
                self.auto_lane_action = QAction(create_text_icon("Arial", icon_size, text_color, "A"), "&Automatic Lane Markers", self)
                self.auto_lane_action.setToolTip("Automatically detect and place lane markers based on a defined region.")
                self.auto_lane_action.triggered.connect(self.start_auto_lane_marker)
                
                self.info_action = QAction(info_icon, "&Info/GitHub", self)
                self.info_action.setToolTip("Open Project GitHub Page")
                self.info_action.triggered.connect(self.open_github)

                # --- Set Shortcuts ---
                self.load_action.setShortcut(QKeySequence.Open)
                self.save_action.setShortcut(QKeySequence.Save)
                self.copy_action.setShortcut(QKeySequence.Copy)
                self.paste_action.setShortcut(QKeySequence.Paste)
                self.undo_action.setShortcut(QKeySequence.Undo)
                self.redo_action.setShortcut(QKeySequence.Redo)
                self.zoom_in_action.setShortcut(QKeySequence("Ctrl+="))
                self.zoom_out_action.setShortcut(QKeySequence.ZoomOut)
                # self.save_svg_action.setShortcut(QKeySequence("Ctrl+M"))
                self.reset_action.setShortcut(QKeySequence("Ctrl+R"))

                # --- Set Tooltips ---
                self.load_action.setToolTip("Load an image file (Ctrl+O)")
                self.save_action.setToolTip("Save image and configuration (Ctrl+S)")
                # self.save_svg_action.setToolTip("Save as SVG for Word/vector editing (Ctrl+M)")
                self.reset_action.setToolTip("Reset image and all annotations (Ctrl+R)")
                self.exit_action.setToolTip("Exit the application")
                self.undo_action.setToolTip("Undo last action (Default Shortcut: OS dependent")
                self.redo_action.setToolTip("Redo last undone action (Default Shortcut: OS dependent)")
                self.copy_action.setToolTip("Copy rendered image to clipboard (Ctrl+C)")
                self.paste_action.setToolTip("Paste image from clipboard (Ctrl+V)")
                self.zoom_in_action.setToolTip("Increase zoom level (Ctrl+= or mouse scroll bar)")
                self.zoom_out_action.setToolTip("Decrease zoom level (Ctrl+- or mouse scroll bar)). Auto resets the zoom when reaches zero.")
                # --- START: Set Tooltips for Panning Actions ---
                self.pan_left_action.setToolTip("Pan the view left (when zoomed) (Arrow key left or mouse right click)")
                self.pan_right_action.setToolTip("Pan the view right (when zoomed) (Arrow key right or mouse right click")
                self.pan_up_action.setToolTip("Pan the view up (when zoomed) (Arrow key up or mouse right click)")
                self.pan_down_action.setToolTip("Pan the view down (when zoomed) (Arrow key down or mouse right click")
                self.draw_bounding_box_action = QAction(bounding_box_icon, "Draw &Bounding Box", self)
                self.draw_bounding_box_action.setToolTip("Draw a bounding rectangle on the image. Use the marker tab, custom marker options for color and size")
                self.draw_bounding_box_action.triggered.connect(self.enable_rectangle_drawing_mode)
                
                self.draw_line_action = QAction(draw_line_icon, "Draw &a line", self)
                self.draw_line_action.setToolTip("Draw a line on the image. Use the marker tab, custom marker options for color and size")
                self.draw_line_action.triggered.connect(self.enable_line_drawing_mode)
                # --- END: Set Tooltips for Panning Actions ---

                # --- Connect signals ---
                self.load_action.triggered.connect(self.load_image)
                self.save_action.triggered.connect(self.save_image)
                # self.save_svg_action.triggered.connect(self.save_image_svg)
                self.reset_action.triggered.connect(self.reset_image)
                self.exit_action.triggered.connect(self.close)
                self.undo_action.triggered.connect(self.undo_action_m)
                self.redo_action.triggered.connect(self.redo_action_m)
                self.copy_action.triggered.connect(self.copy_to_clipboard)
                self.paste_action.triggered.connect(self.paste_image)
                self.zoom_in_action.triggered.connect(self.zoom_in)
                self.zoom_out_action.triggered.connect(self.zoom_out)
                # --- START: Connect Panning Action Signals ---
                # Use lambda functions to directly modify the offset
                pan_step = 30 # Define pan step size here or access from elsewhere if needed
                self.pan_right_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_right_action.triggered.connect(self.update_live_view) # Trigger update after offset change

                self.pan_left_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_left_action.triggered.connect(self.update_live_view)

                self.pan_down_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_down_action.triggered.connect(self.update_live_view)

                self.pan_up_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_up_action.triggered.connect(self.update_live_view)
                
                self.copy_custom_items_action = QAction(copy_custom_icon, "Copy Custom Markers/Shapes", self)
                self.copy_custom_items_action.setToolTip("Copy all custom markers and shapes to the clipboard.")
                self.copy_custom_items_action.triggered.connect(self.copy_custom_items)

                self.paste_custom_items_action = QAction(paste_custom_icon, "Paste Custom Markers/Shapes", self)
                self.paste_custom_items_action.setToolTip("Paste custom markers and shapes from the clipboard.\nItems will be added to existing ones.")
                self.paste_custom_items_action.triggered.connect(self.paste_custom_items)
                
                # --- END: Connect Panning Action Signals ---

                # --- START: Initially Disable Panning Actions ---
                self.pan_left_action.setEnabled(False)
                self.pan_right_action.setEnabled(False)
                self.pan_up_action.setEnabled(False)
                self.pan_down_action.setEnabled(False)
                
            def copy_custom_items(self):
                """Copies custom markers and shapes to the clipboard as JSON."""
                if not hasattr(self, "custom_markers"): self.custom_markers = []
                if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                if not self.custom_markers and not self.custom_shapes:
                    QMessageBox.information(self, "Nothing to Copy", "There are no custom markers or shapes to copy.")
                    return

                # Serialize markers and shapes
                # custom_markers are stored as lists, custom_shapes as dicts
                # For markers, we need to convert QColor to string for JSON
                serializable_markers = []
                for marker_data in self.custom_markers:
                    try:
                        # marker_data is a list: [x, y, text, qcolor, font_family, font_size, is_bold, is_italic]
                        m_copy = list(marker_data)
                        if isinstance(m_copy[3], QColor):
                            m_copy[3] = m_copy[3].name() # Convert QColor to hex string
                        serializable_markers.append(m_copy)
                    except IndexError:
                        print(f"Warning: Skipping marker during copy due to unexpected format: {marker_data}")
                
                items_to_copy = {
                    "custom_markers": serializable_markers,
                    "custom_shapes": self.custom_shapes # Already list of dicts
                }

                try:
                    json_data = json.dumps(items_to_copy)
                    mime_data = QMimeData()
                    mime_data.setData(self.MIME_TYPE_CUSTOM_ITEMS, json_data.encode('utf-8'))
                    
                    # Also set text for basic paste compatibility (e.g., into a text editor)
                    mime_data.setText(f"Imaging Assistant Custom Items (JSON):\n{json_data}")
                    
                    QApplication.clipboard().setMimeData(mime_data)
                    QMessageBox.information(self, "Items Copied", 
                                            f"{len(self.custom_markers)} markers and {len(self.custom_shapes)} shapes copied to clipboard.")
                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Could not copy custom items: {e}")
                    traceback.print_exc()

            def paste_custom_items(self):
                """Pastes custom markers and shapes from the clipboard."""
                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()

                if not mime_data.hasFormat(self.MIME_TYPE_CUSTOM_ITEMS):
                    QMessageBox.information(self, "Paste Error", 
                                            "Clipboard does not contain Imaging Assistant custom items in the expected format.")
                    return

                try:
                    json_bytes = mime_data.data(self.MIME_TYPE_CUSTOM_ITEMS)
                    json_data_str = json_bytes.data().decode('utf-8') # Convert QByteArray to bytes then to string
                    pasted_items = json.loads(json_data_str)

                    pasted_markers_data = pasted_items.get("custom_markers", [])
                    pasted_shapes_data = pasted_items.get("custom_shapes", [])

                    if not pasted_markers_data and not pasted_shapes_data:
                        QMessageBox.information(self, "Paste", "No custom markers or shapes found in clipboard data.")
                        return

                    self.save_state() # Save current state before adding new items

                    if not hasattr(self, "custom_markers"): self.custom_markers = []
                    if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                    num_pasted_markers = 0
                    for marker_data_serial in pasted_markers_data:
                        try:
                            # marker_data_serial is a list: [x, y, text, color_str, font_family, font_size, is_bold, is_italic]
                            m_copy = list(marker_data_serial)
                            m_copy[3] = QColor(m_copy[3]) # Convert color string back to QColor
                            if not m_copy[3].isValid(): m_copy[3] = QColor(Qt.black) # Fallback
                            self.custom_markers.append(m_copy)
                            num_pasted_markers += 1
                        except (IndexError, TypeError, ValueError) as e:
                            print(f"Warning: Skipping pasted marker due to format error: {marker_data_serial}, {e}")
                    
                    num_pasted_shapes = 0
                    for shape_data in pasted_shapes_data:
                        # Basic validation (can be extended)
                        if isinstance(shape_data, dict) and 'type' in shape_data:
                            self.custom_shapes.append(shape_data)
                            num_pasted_shapes +=1
                        else:
                             print(f"Warning: Skipping pasted shape due to format error: {shape_data}")


                    self.is_modified = True
                    self.update_live_view()
                    QMessageBox.information(self, "Items Pasted",
                                            f"{num_pasted_markers} markers and {num_pasted_shapes} shapes pasted.")

                except Exception as e:
                    QMessageBox.critical(self, "Paste Error", f"Could not paste custom items: {e}")
                    traceback.print_exc()
                
            def _update_preview_label_size(self):
                """
                Updates the fixed size of the live_view_label, respecting max width/height
                and maintaining aspect ratio. Prioritizes fixing height, but adjusts if
                width constraint is violated.
                """
                # --- Define Defaults and Minimums ---
                default_max_width = 600  # Fallback if width setting is missing/invalid
                default_max_height = 500 # Fallback if height setting is missing/invalid
                min_dim = 50             # Minimum allowed dimension for the label

                # --- Determine Maximum Constraints (with validation) ---
                max_w = default_max_width
                if hasattr(self, 'preview_label_width_setting'):
                    try:
                        setting_width = int(self.preview_label_width_setting)
                        if setting_width > 0:
                            max_w = setting_width
                        else:
                            print("Warning: preview_label_width_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_width_setting is invalid, using default.")
                else:
                    print("Warning: preview_label_width_setting attribute not found, using default.")
                max_w = max(min_dim, max_w) # Apply minimum constraint

                max_h = default_max_height
                if hasattr(self, 'preview_label_max_height_setting'):
                    try:
                        setting_height = int(self.preview_label_max_height_setting)
                        if setting_height > 0:
                            max_h = setting_height
                        else:
                            print("Warning: preview_label_max_height_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_max_height_setting is invalid, using default.")
                max_h = max(min_dim, max_h) # Apply minimum constraint


                # --- Initialize Final Dimensions ---
                final_w = max_w
                final_h = max_h

                # --- Calculate Dimensions Based on Image ---
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()

                    if w > 0 and h > 0:
                        img_ratio = w / h

                        # Attempt 1: Calculate width based on fixed max height
                        calc_w_based_on_h = max_h * img_ratio

                        if calc_w_based_on_h <= max_w:
                            # Width is within limits, height is fixed
                            final_w = calc_w_based_on_h
                            final_h = max_h
                        else:
                            # Calculated width exceeds max width, so fix width and calculate height
                            final_w = max_w
                            final_h = max_w / img_ratio

                        # Ensure calculated dimensions are not below minimum
                        final_w = max(min_dim, int(final_w))
                        final_h = max(min_dim, int(final_h))

                    else:
                        # Handle invalid image dimensions (0 width or height)
                        final_w = max_w # Already set above
                        final_h = max_h # Already set above

                else:
                    # No image loaded - Use max constraints as default size
                    final_w = max_w # Already set above
                    final_h = max_h # Already set above
                    # self.live_view_label.clear() # Optionally clear the label

                # --- Set the Calculated Fixed Size ---
                self.live_view_label.setFixedSize(final_w, final_h)
                
            def _update_status_bar(self):
                """Updates the status bar labels with current image information."""
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()
                    self.size_label.setText(f"Image Size: {w}x{h}")

                    # Determine bit depth/format string
                    img_format = self.image.format()
                    depth_str = "Unknown"
                    if img_format == QImage.Format_Grayscale8:
                        depth_str = "8-bit Grayscale"
                    elif img_format == QImage.Format_Grayscale16:
                        depth_str = "16-bit Grayscale"
                    elif img_format == QImage.Format_RGB888:
                         depth_str = "24-bit RGB"
                    elif img_format in (QImage.Format_RGB32, QImage.Format_ARGB32,
                                        QImage.Format_ARGB32_Premultiplied, QImage.Format_RGBA8888):
                         depth_str = "32-bit (A)RGB"
                    elif img_format == QImage.Format_Indexed8:
                         depth_str = "8-bit Indexed"
                    elif img_format == QImage.Format_Mono:
                         depth_str = "1-bit Mono"
                    else:
                        # Fallback to depth() if format is unusual
                        depth_val = self.image.depth()
                        depth_str = f"{depth_val}-bit Depth"
                    self.depth_label.setText(f"Bit Depth: {depth_str}")

                    # Update location
                    location = "N/A"
                    if hasattr(self, 'image_path') and self.image_path:
                        # Show only filename if it's a path, otherwise show the source info
                        if os.path.exists(self.image_path):
                            location = os.path.basename(self.image_path)
                        else: # Likely "Clipboard..." or similar
                            location = self.image_path
                    else:
                        location = "N/A"
                    self.location_label.setText(f"Source: {location}")

                else:
                    # No image loaded
                    self.size_label.setText("Image Size: N/A")
                    self.depth_label.setText("Bit Depth: N/A")
                    self.location_label.setText("Source: N/A")
                
            def prompt_save_if_needed(self):
                if not self.is_modified:
                    return True # No changes, proceed
                """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
                reply = QMessageBox.question(self, 'Unsaved Changes',
                                             "Do you want to save the file?",
                                             QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                             QMessageBox.Save) # Default button is Save

                if reply == QMessageBox.Save:
                    return self.save_image() # Returns True if saved, False if save cancelled
                elif reply == QMessageBox.Discard:
                    return True # Proceed without saving
                else: # Cancelled
                    return False # Abort the current action (close/load)

            def closeEvent(self, event):
                """Overrides the default close event to prompt for saving."""
                if self.prompt_save_if_needed():

                    event.accept() # Proceed with closing
                else:
                    event.ignore() # Abort closing
                    
            def enable_line_drawing_mode(self):
                self.save_state()
                """Sets up the application to draw a line."""
                self.drawing_mode = 'line'
                self.live_view_label.mode = 'draw_shape' # Generic drawing mode for label
                self.current_drawing_shape_preview = None # Clear previous preview
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.mousePressEvent = self.start_shape_draw
                self.live_view_label.mouseMoveEvent = self.update_shape_draw
                self.live_view_label.mouseReleaseEvent = self.finalize_shape_draw
                print("Line drawing mode enabled.") # Debug

            def enable_rectangle_drawing_mode(self):
                self.save_state()
                """Sets up the application to draw a rectangle."""
                self.drawing_mode = 'rectangle'
                self.live_view_label.mode = 'draw_shape' # Generic drawing mode for label
                self.current_drawing_shape_preview = None # Clear previous preview
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.mousePressEvent = self.start_shape_draw
                self.live_view_label.mouseMoveEvent = self.update_shape_draw
                self.live_view_label.mouseReleaseEvent = self.finalize_shape_draw
                print("Rectangle drawing mode enabled.") # Debug

            def cancel_drawing_mode(self):
                """Resets drawing mode and cursor."""
                self.drawing_mode = None
                self.live_view_label.mode = None
                self.current_drawing_shape_preview = None
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.live_view_label.mousePressEvent = None # Reset handler
                self.live_view_label.mouseMoveEvent = None # Reset handler
                self.live_view_label.mouseReleaseEvent = None # Reset handler
                self.update_live_view() # Refresh to clear any preview
                
            def start_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle']:
                    start_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    self.current_drawing_shape_preview = {'start': snapped_start_point, 'end': snapped_start_point}
                    self.update_live_view()

            def update_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    end_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
                    
                    self.current_drawing_shape_preview['end'] = snapped_end_point
                    self.update_live_view()

            def finalize_shape_draw(self, event):
                """Finalizes the shape and adds it to the custom_shapes list."""
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    # start_point_label_space is already snapped from start_shape_draw
                    start_point_label_space = self.current_drawing_shape_preview['start']
                    
                    end_point_transformed_label_space = self.live_view_label.transform_point(event.pos())
                    end_point_snapped_label_space = self.snap_point_to_grid(end_point_transformed_label_space) # Snap it
    
                    # ... (Get current style settings: color, thickness) ...
                    color = self.custom_marker_color
                    thickness = self.custom_font_size_spinbox.value()


                    displayed_width = float(self.live_view_label.width())
                    displayed_height = float(self.live_view_label.height())
                    current_image_width = float(self.image.width()) if self.image and self.image.width() > 0 else 1.0
                    current_image_height = float(self.image.height()) if self.image and self.image.height() > 0 else 1.0
                    if current_image_width <= 0 or current_image_height <= 0: # Safety check
                        scale_img_to_label = 1.0
                    else:
                        scale_x_fit = displayed_width / current_image_width
                        scale_y_fit = displayed_height / current_image_height
                        scale_img_to_label = min(scale_x_fit, scale_y_fit)
                    display_img_w = current_image_width * scale_img_to_label
                    display_img_h = current_image_height * scale_img_to_label
                    label_offset_x = (displayed_width - display_img_w) / 2.0
                    label_offset_y = (displayed_height - display_img_h) / 2.0
                    def label_to_image_coords(label_point):
                        if scale_img_to_label <= 1e-9: return (0.0, 0.0)
                        relative_x_in_display = label_point.x() - label_offset_x
                        relative_y_in_display = label_point.y() - label_offset_y
                        img_x = relative_x_in_display / scale_img_to_label
                        img_y = relative_y_in_display / scale_img_to_label
                        return (img_x, img_y)
                    start_img_coords = label_to_image_coords(start_point_label_space) # Use snapped start
                    end_img_coords = label_to_image_coords(end_point_snapped_label_space) # Use snapped end
    
                    shape_data = {
                        'type': self.drawing_mode,
                        'color': color.name(),
                        'thickness': thickness
                    }
                    valid_shape = False
                    if self.drawing_mode == 'line':
                        if abs(start_img_coords[0] - end_img_coords[0]) > 0.5 or abs(start_img_coords[1] - end_img_coords[1]) > 0.5:
                            shape_data['start'] = start_img_coords
                            shape_data['end'] = end_img_coords
                            valid_shape = True
                    elif self.drawing_mode == 'rectangle':
                        x_img = min(start_img_coords[0], end_img_coords[0])
                        y_img = min(start_img_coords[1], end_img_coords[1])
                        w_img = abs(end_img_coords[0] - start_img_coords[0])
                        h_img = abs(end_img_coords[1] - start_img_coords[1])
                        if w_img > 0.5 and h_img > 0.5:
                            shape_data['rect'] = (x_img, y_img, w_img, h_img)
                            valid_shape = True
                    if valid_shape:
                        self.custom_shapes.append(shape_data)
                        self.save_state()
                        self.is_modified = True
                    self.cancel_drawing_mode()
                else:
                    self.cancel_drawing_mode()
                    
            def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
                """Converts QImage to NumPy array, preserving format and handling row padding."""
                if qimage.isNull():
                    return None
            
                img_format = qimage.format()
                height = qimage.height()
                width = qimage.width()
                bytes_per_line = qimage.bytesPerLine() # Actual bytes per row (includes padding)
            
            
                ptr = qimage.constBits()
                if not ptr:
                    return None
            
                # Ensure ptr is treated as a bytes object of the correct size
                # ptr is a sip.voidptr, needs explicit size setting for buffer protocol
                try:
                    ptr.setsize(qimage.byteCount())
                except AttributeError: # Handle cases where setsize might not be needed or available depending on Qt/Python version
                    pass # Continue, hoping buffer protocol works based on qimage.byteCount()
            
                buffer_data = bytes(ptr) # Create a bytes object from the pointer
                if len(buffer_data) != qimage.byteCount():
                     print(f"qimage_to_numpy: Warning - Buffer size mismatch. Expected {qimage.byteCount()}, got {len(buffer_data)}.")
                     # Fallback or error? Let's try to continue but warn.
            
                # --- Grayscale Formats ---
                if img_format == QImage.Format_Grayscale16:
                    bytes_per_pixel = 2
                    dtype = np.uint16
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        # No padding, simple reshape
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                        return arr.copy() # Return a copy
                    else:
                        # Handle padding
                        arr = np.zeros((height, width), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype)
                            else:
                                print(f"qimage_to_numpy (Grayscale16): Row data length mismatch for row {y}. Skipping row.") # Error handling
                        return arr
            
                elif img_format == QImage.Format_Grayscale8:
                    bytes_per_pixel = 1
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype)
                            else:
                                print(f"qimage_to_numpy (Grayscale8): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # --- Color Formats ---
                # ARGB32 (often BGRA in memory) & RGBA8888 (often RGBA in memory)
                elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied):
                    bytes_per_pixel = 4
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                        # QImage ARGB32 is typically BGRA in memory order for numpy
                        # QImage RGBA8888 is typically RGBA in memory order for numpy
                        # Let's return the raw order for now. The caller might need to swap if needed.
                        # If Format_ARGB32, the array is likely BGRA.
                        # If Format_RGBA8888, the array is likely RGBA.
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 4), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                            else:
                                print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # RGB32 (often BGRX or RGBX in memory) & RGBX8888
                elif img_format in (QImage.Format_RGB32, QImage.Format_RGBX8888):
                    bytes_per_pixel = 4 # Stored with an ignored byte
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                        # Memory order is often BGRX (Blue, Green, Red, Ignored Alpha/Padding) for RGB32
                        # Let's return the 4 channels for now.
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 4), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                            else:
                                print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # RGB888 (Tightly packed RGB)
                elif img_format == QImage.Format_RGB888:
                    bytes_per_pixel = 3
                    dtype = np.uint8
                    expected_bytes_per_line = width * bytes_per_pixel
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 3)
                        # QImage RGB888 is RGB order in memory
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width, 3), dtype=dtype)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            if len(row_data) == expected_bytes_per_line:
                                arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 3)
                            else:
                                print(f"qimage_to_numpy (RGB888): Row data length mismatch for row {y}. Skipping row.")
                        return arr
            
                # --- Fallback / Conversion Attempt ---
                else:
                    # For other formats, try converting to a known format first
                    try:
                        # Convert to ARGB32 as a robust intermediate format
                        qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                        if qimage_conv.isNull():
                            print("qimage_to_numpy: Fallback conversion to ARGB32 failed.")
                            # Last resort: try Grayscale8
                            qimage_conv_gray = qimage.convertToFormat(QImage.Format_Grayscale8)
                            if qimage_conv_gray.isNull():
                                 return None
                            else:
                                return self.qimage_to_numpy(qimage_conv_gray) # Recursive call with Grayscale8
            
                        return self.qimage_to_numpy(qimage_conv) # Recursive call with ARGB32
                    except Exception as e:
                        print(f"qimage_to_numpy: Error during fallback conversion: {e}")
                        return None

            def numpy_to_qimage(self, array: np.ndarray) -> QImage:
                """Converts NumPy array to QImage, selecting appropriate format."""
                if array is None:
                    return QImage() # Return invalid QImage
            
                if not isinstance(array, np.ndarray):
                    return QImage()
            
                try:
                    if array.ndim == 2: # Grayscale
                        height, width = array.shape
                        if array.dtype == np.uint16:
                            bytes_per_line = width * 2
                            # Ensure data is contiguous
                            contiguous_array = np.ascontiguousarray(array)
                            # Create QImage directly from contiguous data buffer
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                            # Important: QImage doesn't own the buffer by default. Return a copy.
                            return qimg.copy()
                        elif array.dtype == np.uint8:
                            bytes_per_line = width * 1
                            contiguous_array = np.ascontiguousarray(array)
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                            return qimg.copy()
                        elif np.issubdtype(array.dtype, np.floating):
                             # Assume float is in 0-1 range, scale to uint8
                             img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                             bytes_per_line = width * 1
                             contiguous_array = np.ascontiguousarray(img_norm)
                             qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                             return qimg.copy()
                        else:
                            raise TypeError(f"Unsupported grayscale NumPy dtype: {array.dtype}")
            
                    elif array.ndim == 3: # Color
                        height, width, channels = array.shape
                        if channels == 3 and array.dtype == np.uint8:
                            # Assume input is BGR (common from OpenCV), convert to RGB for QImage.Format_RGB888
                            # Make a contiguous copy before conversion
                            contiguous_array_bgr = np.ascontiguousarray(array)
                            rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                            # rgb_image is now contiguous RGB
                            bytes_per_line = width * 3
                            qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                            return qimg.copy()
                        elif channels == 4 and array.dtype == np.uint8:
                            # Assume input is BGRA (matches typical QImage.Format_ARGB32 memory layout)
                            contiguous_array_bgra = np.ascontiguousarray(array)
                            bytes_per_line = width * 4
                            qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                            return qimg.copy()
                        # Add handling for 16-bit color if needed (less common for display)
                        elif channels == 3 and array.dtype == np.uint16:
                             # Downscale to 8-bit for display format
                             array_8bit = (array / 257.0).astype(np.uint8)
                             contiguous_array_bgr = np.ascontiguousarray(array_8bit)
                             rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                             bytes_per_line = width * 3
                             qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                             return qimg.copy()
                        elif channels == 4 and array.dtype == np.uint16:
                             # Downscale color channels, keep alpha potentially?
                             array_8bit = (array / 257.0).astype(np.uint8) # Downscales all channels including alpha
                             contiguous_array_bgra = np.ascontiguousarray(array_8bit)
                             bytes_per_line = width * 4
                             qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                             return qimg.copy()
                        else:
                             raise TypeError(f"Unsupported color NumPy dtype/channel combination: {array.dtype} / {channels} channels")
                    else:
                        raise ValueError(f"Unsupported array dimension: {array.ndim}")
            
                except Exception as e:
                    traceback.print_exc()
                    return QImage() # Return invalid QImage on error

            def get_image_format(self, image=None):
                """Helper to safely get the format of self.image or a provided image."""
                img_to_check = image if image is not None else self.image
                if img_to_check and not img_to_check.isNull():
                    return img_to_check.format()
                return None

            def get_compatible_grayscale_format(self, image=None):
                """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
                current_format = self.get_image_format(image)
                if current_format == QImage.Format_Grayscale16:
                    return QImage.Format_Grayscale16
                else: # Treat 8-bit grayscale or color as needing 8-bit target
                    return QImage.Format_Grayscale8
            # --- END: Helper Functions ---
                
            def quadrilateral_to_rect(self, image, quad_points):
                """
                Warps the quadrilateral region defined by quad_points in the input image
                to a rectangular image using perspective transformation.
                Crucially, preserves the original bit depth (e.g., uint16) of the input image.
                """
                # --- Input Validation ---
                if not image or image.isNull():
                    QMessageBox.warning(self, "Warp Error", "Invalid input image provided for warping.")
                    return None
                if len(quad_points) != 4:
                     QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
                     return None
                if cv2 is None:
                     QMessageBox.critical(self, "Dependency Error", "OpenCV (cv2) is required for quadrilateral warping but is not installed.")
                     return None
             
                # --- Convert QImage to NumPy array (Preserves bit depth) ---
                try:
                    img_array = self.qimage_to_numpy(image) # Should return uint16 if input is Format_Grayscale16
                    if img_array is None: raise ValueError("NumPy Conversion returned None")
                    print(f"quad_to_rect: Input image dtype = {img_array.dtype}, shape = {img_array.shape}") # Debug
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
                    return None
             
                # --- !!! REMOVED INTERNAL GRAYSCALE CONVERSION !!! ---
                # # OLD code that caused the issue if input was color:
                # if img_array.ndim == 3:
                #      print("Warning: Warping input array is 3D. Converting to grayscale.")
                #      # THIS CONVERSION TO UINT8 WAS THE PROBLEM
                #      color_code = cv2.COLOR_BGR2GRAY if img_array.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
                #      img_array = cv2.cvtColor(img_array, color_code)
                #      # img_array = img_array.astype(np.uint8) # Explicit cast not needed, cvtColor usually returns uint8
                # --- End Removed Block ---
                # Now, warpPerspective will operate on the img_array with its original dtype (e.g., uint16 grayscale or color)
             
             
                # --- Transform points from LiveViewLabel space to Image space ---
                # (Coordinate transformation logic remains the same - assumes direct scaling for now)
                label_width = self.live_view_label.width()
                label_height = self.live_view_label.height()
                current_img_width = image.width()
                current_img_height = image.height()
                scale_x_disp = current_img_width / label_width if label_width > 0 else 1
                scale_y_disp = current_img_height / label_height if label_height > 0 else 1
             
                src_points_img = []
                for point in quad_points:
                    x_view, y_view = point.x(), point.y()
                    # Account for zoom/pan in LiveViewLabel coordinates
                    if self.live_view_label.zoom_level != 1.0:
                        x_view = (x_view - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                        y_view = (y_view - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
                    # Scale to image coordinates
                    x_image = x_view * scale_x_disp
                    y_image = y_view * scale_y_disp
                    src_points_img.append([x_image, y_image])
                src_np = np.array(src_points_img, dtype=np.float32)
             
                # --- Define Destination Rectangle ---
                # Calculate dimensions based on the source points
                width_a = np.linalg.norm(src_np[0] - src_np[1])
                width_b = np.linalg.norm(src_np[2] - src_np[3])
                height_a = np.linalg.norm(src_np[0] - src_np[3])
                height_b = np.linalg.norm(src_np[1] - src_np[2])
                max_width = int(max(width_a, width_b))
                max_height = int(max(height_a, height_b))
             
                if max_width <= 0 or max_height <= 0:
                     QMessageBox.warning(self, "Warp Error", f"Invalid destination rectangle size ({max_width}x{max_height}).")
                     return None
             
                dst_np = np.array([
                    [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
                ], dtype=np.float32)
             
                # --- Perform Perspective Warp using OpenCV ---
                try:
                    matrix = cv2.getPerspectiveTransform(src_np, dst_np)
                    # Determine border color based on input array type
                    if img_array.ndim == 3: # Color
                         # Use black (0,0,0) or white (255,255,255) depending on preference. Alpha needs 4 values if present.
                         border_val = (0, 0, 0, 0) if img_array.shape[2] == 4 else (0, 0, 0)
                    else: # Grayscale
                         border_val = 0 # Black for grayscale
             
                    # Warp the ORIGINAL NumPy array (could be uint8, uint16, color)
                    warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                                       flags=cv2.INTER_LINEAR, # Linear interpolation is usually good
                                                       borderMode=cv2.BORDER_CONSTANT,
                                                       borderValue=border_val) # Fill borders appropriately
                    print(f"quad_to_rect: Warped array dtype = {warped_array.dtype}, shape = {warped_array.shape}") # Debug
                except Exception as e:
                     QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
                     traceback.print_exc() # Print full traceback for debugging
                     return None
             
                # --- Convert warped NumPy array back to QImage ---
                # numpy_to_qimage should handle uint16 correctly, creating Format_Grayscale16
                try:
                    warped_qimage = self.numpy_to_qimage(warped_array) # Handles uint8/uint16/color
                    if warped_qimage.isNull(): raise ValueError("numpy_to_qimage conversion failed.")
                    print(f"quad_to_rect: Returned QImage format = {warped_qimage.format()}") # Debug
                    return warped_qimage
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array back to QImage: {e}")
                    return None
           # --- END: Modified Warping ---
            
            

                
            def create_menu_bar(self):
                menubar = self.menuBar()

                # --- File Menu ---
                file_menu = menubar.addMenu("&File")
                file_menu.addAction(self.load_action)
                file_menu.addAction(self.save_action)
                # file_menu.addAction(self.save_svg_action)
                file_menu.addSeparator()
                file_menu.addAction(self.reset_action)
                file_menu.addSeparator()
                file_menu.addAction(self.exit_action)

                # --- Edit Menu ---
                edit_menu = menubar.addMenu("&Edit")
                edit_menu.addAction(self.undo_action)
                edit_menu.addAction(self.redo_action)
                edit_menu.addSeparator()
                edit_menu.addAction(self.copy_action)
                edit_menu.addAction(self.paste_action)

                # --- View Menu ---
                view_menu = menubar.addMenu("&View")
                view_menu.addAction(self.zoom_in_action)
                view_menu.addAction(self.zoom_out_action)
                
                tools_menu = menubar.addMenu("&Tools")
                tools_menu.addAction(self.auto_lane_action)

                # --- About Menu ---
                about_menu = menubar.addMenu("&About")
                # Add "GitHub" action directly here or create it in _create_actions
                github_action = QAction("&GitHub", self)
                github_action.setToolTip("Open the project's GitHub page")
                github_action.triggered.connect(self.open_github)
                about_menu.addAction(github_action)
                
                # self.statusBar().showMessage("Ready")
            def open_github(self):
                # Open the GitHub link in the default web browser
                QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
            
            def create_tool_bar(self):
                """Create the main application toolbar."""
                self.tool_bar = QToolBar("Main Toolbar")
                self.tool_bar.setIconSize(QSize(30, 30))
                self.tool_bar.setMovable(False)
                self.tool_bar.setToolButtonStyle(Qt.ToolButtonIconOnly)

                # Add actions to the toolbar
                self.tool_bar.addAction(self.load_action)
                self.tool_bar.addAction(self.paste_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.save_action)
                self.tool_bar.addAction(self.copy_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.undo_action)
                self.tool_bar.addAction(self.redo_action)
                self.tool_bar.addSeparator()

                # --- Add Zoom and Pan Buttons ---
                self.tool_bar.addAction(self.zoom_in_action)
                self.tool_bar.addAction(self.zoom_out_action)
                self.tool_bar.addSeparator() # Optional separator

                self.tool_bar.addAction(self.pan_left_action)
                self.tool_bar.addAction(self.pan_up_action)   # Group visually
                self.tool_bar.addAction(self.pan_down_action)
                self.tool_bar.addAction(self.pan_right_action)
                # --- End Adding Pan Buttons ---

                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.auto_lane_action) # Add new button
                
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.draw_line_action) # Add the new action
                self.tool_bar.addAction(self.draw_bounding_box_action) # Add the new action
     
                
                # --- NEW: Add Copy/Paste Custom Items to Toolbar ---
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.copy_custom_items_action)
                self.tool_bar.addAction(self.paste_custom_items_action)
                # --- END NEW ---
                
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.reset_action)
                
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.info_action)

                # Add the toolbar to the main window
                self.addToolBar(Qt.TopToolBarArea, self.tool_bar)
                self.tool_bar.setContextMenuPolicy(Qt.PreventContextMenu)

                
            def start_auto_lane_marker(self):
                """Initiates the automatic lane marker placement process, allowing user to choose region type."""
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    return

                # --- 1. Ask User for Marker Side ---
                items_side = ["Left", "Right"]
                side, ok_side = QInputDialog.getItem(self, "Select Marker Side",
                                                     "Place markers on which side?", items_side, 0, False)
                if not ok_side or not side:
                    return  # User cancelled
                self.auto_marker_side = side.lower()

                # --- 2. Ask User for Region Definition Type ---
                items_region = ["Rectangle (for straight lanes)", "Quadrilateral (for skewed lanes)"]
                region_type_str, ok_region = QInputDialog.getItem(self, "Select Region Type",
                                                                 "How do you want to define the lane region?",
                                                                 items_region, 0, False)
                if not ok_region or not region_type_str:
                    return # User cancelled

                self.live_view_label.quad_points = [] # Clear any previous quad points
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True) # Ensure tracking is on

                if "Rectangle" in region_type_str:
                    QMessageBox.information(self, "Define Lane Region",
                                            "Please click and drag on the image preview\n"
                                            "to define the rectangular lane region for automatic marker detection.\n"
                                            "Press ESC to cancel.")
                    self.live_view_label.mode = 'auto_lane_rect'
                    self.live_view_label.mousePressEvent = self.start_rectangle
                    self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
                    self.live_view_label.mouseReleaseEvent = self.finalize_rectangle_for_auto_lane
                elif "Quadrilateral" in region_type_str:
                    QMessageBox.information(self, "Define Lane Region",
                                            "Please click 4 corner points on the image preview\n"
                                            "to define the quadrilateral lane region for automatic marker detection.\n"
                                            "Press ESC to cancel.")
                    self.live_view_label.mode = 'auto_lane_quad'
                    self.live_view_label.mousePressEvent = self.handle_auto_lane_quad_click
                    self.live_view_label.mouseMoveEvent = None # No drag for quad point placement
                    self.live_view_label.mouseReleaseEvent = None # Not needed for single clicks

                self.update_live_view()

            def handle_auto_lane_quad_click(self, event):
                if self.live_view_label.mode != 'auto_lane_quad':
                    return
    
                if event.button() == Qt.LeftButton:
                    point_label_space_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_point_label_space = self.snap_point_to_grid(point_label_space_transformed) # Snap it
                    
                    self.live_view_label.quad_points.append(snapped_point_label_space)
                    self.update_live_view()
    
                    if len(self.live_view_label.quad_points) == 4:
                        self.finalize_quad_for_auto_lane(self.live_view_label.quad_points)

            def finalize_quad_for_auto_lane(self, quad_points_label_space):
                """
                Finalizes the quadrilateral using points from "label space",
                warps the corresponding region from self.image, and proceeds to processing.
                quad_points_label_space: List of 4 QPointF in "label space".
                """
                if len(quad_points_label_space) != 4:
                    QMessageBox.warning(self, "Error", "Quadrilateral definition incomplete for auto lane.")
                    self.live_view_label.mode = None # Reset mode
                    self.live_view_label.quad_points = []
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self.update_live_view()
                    return
                
                warped_qimage_region = self.quadrilateral_to_rect(self.image, quad_points_label_space)

                if warped_qimage_region and not warped_qimage_region.isNull():
                    quad_points_image_space_for_marker_placement = []
                    try:
                        native_img_w = float(self.image.width())
                        native_img_h = float(self.image.height())
                        label_w_widget = float(self.live_view_label.width())
                        label_h_widget = float(self.live_view_label.height())
                        if native_img_w <= 0 or native_img_h <= 0 or label_w_widget <= 0 or label_h_widget <= 0:
                            raise ValueError("Invalid image_to_warp or label dimensions.")
                        scale_native_to_label = min(label_w_widget / native_img_w, label_h_widget / native_img_h)
                        displayed_w_in_label = native_img_w * scale_native_to_label
                        displayed_h_in_label = native_img_h * scale_native_to_label
                        offset_x_centering = (label_w_widget - displayed_w_in_label) / 2.0
                        offset_y_centering = (label_h_widget - displayed_h_in_label) / 2.0
                        for p_label in quad_points_label_space:
                            x_rel_display = p_label.x() - offset_x_centering
                            y_rel_display = p_label.y() - offset_y_centering
                            if scale_native_to_label < 1e-9: raise ValueError("Scale factor too small.")
                            img_x = x_rel_display / scale_native_to_label
                            img_y = y_rel_display / scale_native_to_label
                            quad_points_image_space_for_marker_placement.append(QPointF(img_x, img_y))
                        
                        # print(f"DEBUG AutoLaneQuad (finalize): Mapped image space points for marker placement: {[(p.x(), p.y()) for p in quad_points_image_space_for_marker_placement]}")

                        self.process_auto_lane_region(warped_qimage_region, 
                                                      quad_points_image_space_for_marker_placement, # Pass image-space points
                                                      is_quad_warp=True)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to map quad points for marker placement: {e}")
                        traceback.print_exc()
                else:
                    QMessageBox.warning(self, "Error", "Failed to warp quadrilateral region for auto lane.")

                # Reset state after processing or error
                self.live_view_label.mode = None
                self.live_view_label.quad_points = [] # Clear points from label
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.live_view_label.mousePressEvent = None # Important to reset
                self.live_view_label.mouseMoveEvent = None
                self.live_view_label.mouseReleaseEvent = None
                self.update_live_view()


            def finalize_rectangle_for_auto_lane(self, event):
                if self.live_view_label.mode != 'auto_lane_rect' or not self.live_view_label.rectangle_start:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event)
                    return
    
                if event.button() == Qt.LeftButton:
                    end_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
                    self.live_view_label.rectangle_end = snapped_end_point # Store snapped end point
    
                    rect_coords_img = None
                    try:
                        # self.live_view_label.rectangle_start is already snapped from start_rectangle
                        start_x_view, start_y_view = self.live_view_label.rectangle_start.x(), self.live_view_label.rectangle_start.y()
                        end_x_view, end_y_view = self.live_view_label.rectangle_end.x(), self.live_view_label.rectangle_end.y() # Use snapped end
    
                        # ... (rest of coordinate transformation logic remains the same) ...
                        zoom = self.live_view_label.zoom_level
                        # ... (as in your existing finalize_rectangle_for_auto_lane)
                        offset_x_pan, offset_y_pan = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                        start_x_unzoomed = (start_x_view - offset_x_pan) / zoom
                        start_y_unzoomed = (start_y_view - offset_y_pan) / zoom
                        end_x_unzoomed = (end_x_view - offset_x_pan) / zoom
                        end_y_unzoomed = (end_y_view - offset_y_pan) / zoom
    
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w, img_h = self.image.width(), self.image.height()
                        label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                        if img_w <= 0 or img_h <=0 or label_w <=0 or label_h <=0:
                            raise ValueError("Invalid image or label dimensions for coord conversion.")
    
                        scale_factor = min(label_w / img_w, label_h / img_h)
                        display_offset_x = (label_w - img_w * scale_factor) / 2
                        display_offset_y = (label_h - img_h * scale_factor) / 2
    
                        start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                        start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                        end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                        end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
    
                        rect_x = int(min(start_x_img, end_x_img))
                        rect_y = int(min(start_y_img, end_y_img))
                        rect_w = int(abs(end_x_img - start_x_img))
                        rect_h = int(abs(end_y_img - start_y_img))
    
                        rect_x = max(0, rect_x)
                        rect_y = max(0, rect_y)
                        rect_w = max(1, min(rect_w, img_w - rect_x))
                        rect_h = max(1, min(rect_h, img_h - rect_y))
                        rect_coords_img = (rect_x, rect_y, rect_w, rect_h)
                        # ...
    
                        extracted_qimage_region = self.image.copy(rect_x, rect_y, rect_w, rect_h)
                        if extracted_qimage_region.isNull():
                            raise ValueError("QImage copy failed for rectangle.")
    
                        self.process_auto_lane_region(extracted_qimage_region, rect_coords_img, is_quad_warp=False)
    
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to finalize rectangle for auto lane: {e}")
                        traceback.print_exc()
                    finally:
                        self.live_view_label.mode = None
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        self.live_view_label.mousePressEvent = None
                        self.live_view_label.mouseMoveEvent = None
                        self.live_view_label.mouseReleaseEvent = None
                        self.live_view_label.bounding_box_preview = None
                        self.live_view_label.rectangle_start = None
                        self.update_live_view()

            def process_auto_lane_region(self, qimage_region, original_region_definition, is_quad_warp):
                """
                Processes the extracted/warped region, opens tuning dialog, and places markers.
                qimage_region: The QImage (rectangular) to be analyzed.
                original_region_definition: For rect, tuple (x,y,w,h) in original image space.
                                            For quad, list of 4 QPointF in original image space.
                is_quad_warp: Boolean indicating if qimage_region came from a warped quadrilateral.
                """
                if qimage_region.isNull():
                    QMessageBox.warning(self, "Error", "No valid image region provided for auto lane processing.")
                    return

                dialog_image_pil = self.convert_qimage_to_grayscale_pil(qimage_region)
                if not dialog_image_pil:
                    QMessageBox.warning(self, "Error", "Failed to convert extracted lane to grayscale for analysis.")
                    return

                # Store height AND WIDTH of the image passed to the dialog for later mapping
                dialog_image_height = dialog_image_pil.height
                dialog_image_width = dialog_image_pil.width # <<< --- ADDED

                dialog = AutoLaneTuneDialog(
                    pil_image_data=dialog_image_pil,
                    initial_settings=self.peak_dialog_settings,
                    parent=self,
                    is_from_quad_warp=is_quad_warp
                )

                if dialog.exec_() == QDialog.Accepted:
                    detected_peak_coords_dialog = dialog.get_detected_peaks()
                    final_settings = dialog.get_final_settings()

                    if self.persist_peak_settings_enabled:
                        self.peak_dialog_settings.update(final_settings)

                    if detected_peak_coords_dialog is not None and len(detected_peak_coords_dialog) > 0:
                        self.place_markers_from_dialog(
                            original_region_definition,
                            detected_peak_coords_dialog,
                            self.auto_marker_side,
                            dialog_image_height,
                            dialog_image_width, # <<< --- PASS WIDTH
                            is_quad_warp
                        )
                    else:
                        QMessageBox.information(self, "Auto Lane", "No peaks were detected with the current settings.")
                else:
                    print("Automatic lane marker tuning cancelled.")

                self.live_view_label.bounding_box_preview = None
                self.live_view_label.quad_points = []
                self.update_live_view()


            def place_markers_from_dialog(self, original_region_definition, peak_coords_in_dialog, side,
                                          dialog_image_height, dialog_image_width, 
                                          is_from_quad_warp):
                if not peak_coords_in_dialog.any() or dialog_image_height <= 0 or dialog_image_width <= 0:
                    print("No peak coordinates or invalid dialog image dimensions for auto lane marker placement.")
                    return

                self.save_state()
                target_marker_list = None
                if side == 'left':
                    self.left_markers.clear(); self.current_left_marker_index = 0
                    target_marker_list = self.left_markers
                elif side == 'right':
                    self.right_markers.clear(); self.current_right_marker_index = 0
                    target_marker_list = self.right_markers
                else:
                    print(f"Invalid side '{side}' for auto lane marker placement."); return

                self.on_combobox_changed() 
                current_marker_values_for_labels = self.marker_values 

                if not current_marker_values_for_labels: 
                    current_marker_values_for_labels = [""] * len(peak_coords_in_dialog)
                elif len(current_marker_values_for_labels) < len(peak_coords_in_dialog):
                    current_marker_values_for_labels.extend([""] * (len(peak_coords_in_dialog) - len(current_marker_values_for_labels)))

                sorted_peaks_dialog = np.sort(peak_coords_in_dialog.astype(float)) 
                
                inv_perspective_matrix = None
                src_points_for_transform_np = None
                # dst_points_for_transform_np = None # Not needed here, only for creating the matrix

                if is_from_quad_warp:
                    if len(original_region_definition) != 4:
                        print("Error: Quadrilateral definition requires 4 points for inverse mapping in auto lane.")
                        return
                    src_points_for_transform_list = []
                    for p in original_region_definition: # original_region_definition is list of QPointF in IMAGE space
                        try: src_points_for_transform_list.append([float(p.x()), float(p.y())])
                        except AttributeError: src_points_for_transform_list.append([float(p[0]), float(p[1])]) # If it was a tuple
                    src_points_for_transform_np = np.array(src_points_for_transform_list, dtype=np.float32)
                    
                    # Destination points are in the DIALOG's (warped rectangle) image space
                    dst_points_for_dialog_transform_np = np.array([
                        [0.0, 0.0],
                        [float(dialog_image_width) - 1.0, 0.0],
                        [float(dialog_image_width) - 1.0, float(dialog_image_height) - 1.0],
                        [0.0, float(dialog_image_height) - 1.0]
                    ], dtype=np.float32)
                    try:
                        # We need to map from dialog space (dst) back to original image space (src)
                        inv_perspective_matrix = cv2.getPerspectiveTransform(dst_points_for_dialog_transform_np, src_points_for_transform_np)
                    except Exception as e:
                        print(f"Error calculating inverse perspective matrix for auto lane: {e}")
                        return

                sum_x_coords_img_space = 0.0
                count_x_coords = 0

                for i, peak_y_dialog in enumerate(sorted_peaks_dialog): # peak_y_dialog is already float
                    label = str(current_marker_values_for_labels[i]) if i < len(current_marker_values_for_labels) else ""
                    y_final_img_calc = 0.0
                    x_final_img_for_marker_calc = 0.0 # This X is for the alignment line

                    if not is_from_quad_warp: # Rectangle case
                        # original_region_definition is (rect_x_img, rect_y_img, rect_w_img, rect_h_img) in NATIVE IMAGE space
                        rect_x_img, rect_y_img, rect_w_img, rect_h_img = map(float, original_region_definition)
                        
                        # peak_y_dialog is a Y-coordinate within the rectangular dialog_image_pil (height: dialog_image_height)
                        # We need to map this Y back to the original image's Y scale within the defined rectangle
                        t = peak_y_dialog / (float(dialog_image_height) - 1.0) if dialog_image_height > 1 else 0.5
                        t = max(0.0, min(1.0, t)) # Clamp t to [0,1]
                        y_final_img_calc = rect_y_img + t * rect_h_img

                        if side == 'left':
                            # For left markers, the alignment line is typically at the left edge of the rectangle
                            x_final_img_for_marker_calc = rect_x_img 
                        else: # side == 'right'
                            # For right markers, the alignment line is at the right edge of the rectangle
                            x_final_img_for_marker_calc = rect_x_img + rect_w_img
                    else: # Quadrilateral warp case
                        if inv_perspective_matrix is None: continue
                        
                        # Determine an X-coordinate in the *dialog's (warped)* space to represent the lane edge.
                        # E.g., 5% from left/right edge of the dialog image.
                        x_in_dialog_space = 0.0
                        if side == 'left':
                            x_in_dialog_space = float(dialog_image_width) * 0.05 
                        elif side == 'right':
                            x_in_dialog_space = float(dialog_image_width) * 0.95 
                        
                        point_in_dialog_np = np.array([[[x_in_dialog_space, peak_y_dialog]]], dtype=np.float32)
                        try:
                            original_image_point_np = cv2.perspectiveTransform(point_in_dialog_np, inv_perspective_matrix)
                            x_final_img_for_marker_calc = float(original_image_point_np[0,0,0])
                            y_final_img_calc = float(original_image_point_np[0,0,1])
                        except Exception as e_transform:
                            print(f"Error transforming point back for auto lane: {e_transform}")
                            # Fallback if transform fails (less accurate)
                            t = peak_y_dialog / (float(dialog_image_height) - 1.0) if dialog_image_height > 1 else 0.5
                            t = max(0.0, min(1.0, t))
                            p0_img = QPointF(src_points_for_transform_np[0,0], src_points_for_transform_np[0,1])
                            p1_img = QPointF(src_points_for_transform_np[1,0], src_points_for_transform_np[1,1])
                            p2_img = QPointF(src_points_for_transform_np[2,0], src_points_for_transform_np[2,1])
                            p3_img = QPointF(src_points_for_transform_np[3,0], src_points_for_transform_np[3,1])
                            if side == 'left': # Interpolate along the left edge of the original quad
                                y_final_img_calc = float(p0_img.y() * (1 - t) + p3_img.y() * t)
                                x_final_img_for_marker_calc = float(p0_img.x() * (1 - t) + p3_img.x() * t)
                            else: # Interpolate along the right edge of the original quad
                                y_final_img_calc = float(p1_img.y() * (1 - t) + p2_img.y() * t)
                                x_final_img_for_marker_calc = float(p1_img.x() * (1 - t) + p2_img.x() * t)
                    
                    final_y_to_store = float(y_final_img_calc)
                    target_marker_list.append((final_y_to_store, label))
                    
                    sum_x_coords_img_space += float(x_final_img_for_marker_calc)
                    count_x_coords += 1

                # Calculate the average X position for the alignment line in native image space
                target_x_in_image_space_for_slider = 0.0
                if count_x_coords > 0:
                    target_x_in_image_space_for_slider = sum_x_coords_img_space / count_x_coords
                
                # --- Set the appropriate offset slider ---
                # This target_x is in NATIVE IMAGE PIXELS
                slider_target_value_native_pixels = int(round(target_x_in_image_space_for_slider))

                self._update_marker_slider_ranges() # Ensure slider ranges are current

                if side == 'left':
                    if hasattr(self, 'left_padding_slider'):
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(
                            max(self.left_slider_range[0], min(slider_target_value_native_pixels, self.left_slider_range[1]))
                        )
                        self.left_padding_slider.blockSignals(False)
                        self.left_marker_shift_added = self.left_padding_slider.value() # Sync internal var
                elif side == 'right':
                     if hasattr(self, 'right_padding_slider'):
                         self.right_padding_slider.blockSignals(True)
                         self.right_padding_slider.setValue(
                             max(self.right_slider_range[0], min(slider_target_value_native_pixels, self.right_slider_range[1]))
                         )
                         self.right_padding_slider.blockSignals(False)
                         self.right_marker_shift_added = self.right_padding_slider.value() # Sync internal var

                self.is_modified = True
                self.update_live_view()
                
            def zoom_in(self):
                self.live_view_label.zoom_in() # LiveViewLabel handles its own update
                # --- START: Update Pan Button State ---
                enable_pan = self.live_view_label.zoom_level > 1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                self.update_live_view()
                # --- END: Update Pan Button State ---

            def zoom_out(self):
                self.live_view_label.zoom_out() # LiveViewLabel handles its own update
                # --- START: Update Pan Button State ---
                enable_pan = self.live_view_label.zoom_level > 1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                self.update_live_view()
                # --- END: Update Pan Button State ---
            
            def enable_standard_protein_mode(self):
                """"Enable mode to define standard protein amounts for creating a standard curve."""
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                # Assign mouse event handlers for bounding box creation
                self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
                self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
                self.live_view_label.setCursor(Qt.CrossCursor)
            
            def enable_measure_protein_mode(self):
                """Enable mode to measure protein quantity using the standard curve."""
                if len(self.quantities) < 2:
                    QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
                self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
                self.live_view_label.setCursor(Qt.CrossCursor)

            def call_live_view(self):
                self.update_live_view()      

            def analyze_bounding_box(self, pil_image_for_dialog, standard): # Input is PIL Image
            
                peak_area = None # Initialize
                # Clear previous results before new analysis
                self.latest_peak_areas = []
                self.latest_calculated_quantities = []

                if standard:
                    quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
                    if ok and quantity:
                        try:
                            quantity_value = float(quantity.split()[0])
                            # Calculate peak area using the PIL data
                            peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL
                            if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                                total_area = sum(peak_area)
                                self.quantities_peak_area_dict[quantity_value] = round(total_area, 3)
                                self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                                self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
                                print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                                # Store the areas for potential later export if needed (though usually total area is used for standards)
                                self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store individual areas too
                            else:
                                 print("Peak area calculation cancelled or failed for standard.")
                                 self.latest_peak_areas = [] # Ensure it's cleared

                        except (ValueError, IndexError) as e:
                            QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                            self.latest_peak_areas = []
                        except Exception as e:
                             QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                             self.latest_peak_areas = []
                    else:
                        print("Standard quantity input cancelled.")
                        self.latest_peak_areas = []

                else: # Analyze sample
                    # Calculate peak area using the PIL data
                    peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL

                    if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                        self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store latest areas
                        print(f"Sample Analysis: Calculated Areas = {self.latest_peak_areas}")

                        if len(self.quantities_peak_area_dict) >= 2:
                            # Calculate quantities and store them
                            self.latest_calculated_quantities = self.calculate_unknown_quantity(
                                list(self.quantities_peak_area_dict.values()), # Standard Areas (total)
                                list(self.quantities_peak_area_dict.keys()),   # Standard Quantities
                                self.latest_peak_areas                         # Sample Peak Areas (individual)
                            )
                            print(f"Sample Analysis: Calculated Quantities = {self.latest_calculated_quantities}")

                        else:
                            self.latest_calculated_quantities = [] # No quantities calculated

                        try:
                            # Display the areas in the text box
                            self.target_protein_areas_text.setText(str(self.latest_peak_areas))
                        except Exception as e:
                            print(f"Error displaying sample areas: {e}")
                            self.target_protein_areas_text.setText("Error")
                    else:
                         print("Peak area calculation cancelled or failed for sample.")
                         self.target_protein_areas_text.setText("N/A")
                         self.latest_peak_areas = []
                         self.latest_calculated_quantities = []


                # --- UI updates after analysis ---
                self.update_live_view() # Update display
            
            def calculate_peak_area(self, pil_image_for_dialog): # Input is EXPECTED to be PIL Image
                """Opens the PeakAreaDialog for interactive adjustment and area calculation."""

                # --- Validate Input ---
                if pil_image_for_dialog is None:
                    print("Error: No PIL Image data provided to calculate_peak_area.")
                    return None
                if not isinstance(pil_image_for_dialog, Image.Image):
                     # This indicates an error in the calling function (process_sample/standard)
                     QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
                     return None
                # --- End Validation ---


                # --- Call PeakAreaDialog passing the PIL Image with the 'cropped_data' keyword ---
                # Assumes PeakAreaDialog __init__ expects 'cropped_data' as the first arg (PIL Image type)
                dialog = PeakAreaDialog(
                    cropped_data=pil_image_for_dialog, # Pass the received PIL Image
                    current_settings=self.peak_dialog_settings,
                    persist_checked=self.persist_peak_settings_enabled,
                    parent=self
                )

                peak_areas = None
                if dialog.exec_() == QDialog.Accepted:
                    peak_areas = dialog.get_final_peak_area()
                    if dialog.should_persist_settings():
                        self.peak_dialog_settings = dialog.get_current_settings()
                        self.persist_peak_settings_enabled = True
                    else:
                        self.persist_peak_settings_enabled = False
                else:
                    print("PeakAreaDialog cancelled.")

                return peak_areas if peak_areas is not None else []
            
            
            def calculate_unknown_quantity(self, standard_total_areas, known_quantities, sample_peak_areas):
                """Calculates unknown quantities based on standards and sample areas. Returns a list of quantities."""
                if not standard_total_areas or not known_quantities or len(standard_total_areas) != len(known_quantities):
                    print("Error: Invalid standard data for quantity calculation.")
                    return []
                if len(standard_total_areas) < 2:
                    print("Error: Need at least 2 standards for regression.")
                    return []
                if not sample_peak_areas:
                     print("Warning: No sample peak areas provided for quantity calculation.")
                     return []

                calculated_quantities = []
                try:
                    # Use total standard area vs known quantity for the standard curve
                    coefficients = np.polyfit(standard_total_areas, known_quantities, 1)

                    # Apply the standard curve to *each individual* sample peak area
                    for area in sample_peak_areas:
                        val = np.polyval(coefficients, area)
                        calculated_quantities.append(round(val, 2))

                    # Display the results in a message box (optional, but helpful feedback)
                    # QMessageBox.information(self, "Protein Quantification", f"Predicted Quantities: {calculated_quantities} units")

                except Exception as e:
                     # QMessageBox.warning(self, "Calculation Error", f"Could not calculate quantities: {e}")
                     return [] # Return empty list on error

                return calculated_quantities # <-- Return the list

                
            def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
                """Draw quantity text at the correct position."""
                text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
                painter.drawText(text_position, str(quantity))
            
            def update_standard_protein_quantities(self):
                self.standard_protein_values.text()
            
            def move_tab(self,tab):
                self.tab_widget.setCurrentIndex(tab)
                
            def save_state(self):
                """Save the current state of the image, markers, shapes, and relevant UI/font settings."""
                # Get current state of UI elements related to fonts/colors for custom markers
                custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                state = {
                    "image": self.image.copy() if self.image else None,
                    "left_markers": self.left_markers.copy(),
                    "right_markers": self.right_markers.copy(),
                    "top_markers": self.top_markers.copy(),
                    "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                    "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                    "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                    "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                    "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                    # Standard marker font settings
                    "font_family": self.font_family,
                    "font_size": self.font_size,
                    "font_color": self.font_color, # QColor object is fine for in-memory stack
                    "font_rotation": self.font_rotation,
                    # Marker positioning shifts
                    "left_marker_shift_added": self.left_marker_shift_added,
                    "right_marker_shift_added": self.right_marker_shift_added,
                    "top_marker_shift_added": self.top_marker_shift_added,
                    # Analysis data (keep as before)
                    "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                    "quantities": self.quantities.copy(),
                    "protein_quantities": self.protein_quantities.copy(),
                    "standard_protein_areas": self.standard_protein_areas.copy(),
                    # --- NEW: Custom marker font/color states ---
                    "custom_marker_color": self.custom_marker_color, # QColor object
                    "custom_font_family": custom_font_family, # String from dropdown
                    "custom_font_size": custom_font_size,     # Integer from spinbox
                    # --- END NEW ---
                }
                self.undo_stack.append(state)
                self.redo_stack.clear() # Clear redo stack on new action
            
            def undo_action_m(self):
                """Undo the last action by restoring the previous state, including font settings."""
                if self.undo_stack:
                    # Get current state of UI elements before overwriting self attributes
                    current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                    current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                    # Save the current state to the redo stack
                    current_state = {
                        "image": self.image.copy() if self.image else None,
                        "left_markers": self.left_markers.copy(),
                        "right_markers": self.right_markers.copy(),
                        "top_markers": self.top_markers.copy(),
                        "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                        "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                        "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                        "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                        "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                        # Standard marker font settings
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_color": self.font_color,
                        "font_rotation": self.font_rotation,
                        # Marker positioning shifts
                        "left_marker_shift_added": self.left_marker_shift_added,
                        "right_marker_shift_added": self.right_marker_shift_added,
                        "top_marker_shift_added": self.top_marker_shift_added,
                        # Analysis data
                        "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                        "quantities": self.quantities.copy(),
                        "protein_quantities": self.protein_quantities.copy(),
                        "standard_protein_areas": self.standard_protein_areas.copy(),
                        # --- NEW: Custom marker font/color states ---
                        "custom_marker_color": self.custom_marker_color,
                        "custom_font_family": current_custom_font_family, # Use value read from UI
                        "custom_font_size": current_custom_font_size,     # Use value read from UI
                        # --- END NEW ---
                    }
                    self.redo_stack.append(current_state)

                    # Restore the previous state from the undo stack
                    previous_state = self.undo_stack.pop()
                    self.image = previous_state["image"]
                    self.left_markers = previous_state["left_markers"]
                    self.right_markers = previous_state["right_markers"]
                    self.top_markers = previous_state["top_markers"]
                    self.custom_markers = previous_state.get("custom_markers", [])
                    self.custom_shapes = previous_state.get("custom_shapes", [])
                    self.image_before_padding = previous_state["image_before_padding"]
                    self.image_contrasted = previous_state["image_contrasted"]
                    self.image_before_contrast = previous_state["image_before_contrast"]
                    # Restore standard font settings
                    self.font_family = previous_state["font_family"]
                    self.font_size = previous_state["font_size"]
                    self.font_color = previous_state["font_color"]
                    self.font_rotation = previous_state["font_rotation"]
                    # Restore shifts
                    self.left_marker_shift_added = previous_state["left_marker_shift_added"]
                    self.right_marker_shift_added = previous_state["right_marker_shift_added"]
                    self.top_marker_shift_added = previous_state["top_marker_shift_added"]
                    # Restore analysis data
                    self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
                    self.quantities = previous_state["quantities"]
                    self.protein_quantities = previous_state["protein_quantities"]
                    self.standard_protein_areas = previous_state["standard_protein_areas"]
                    # --- NEW: Restore custom marker font/color states ---
                    self.custom_marker_color = previous_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
                    restored_custom_font_family = previous_state.get("custom_font_family", "Arial")
                    restored_custom_font_size = previous_state.get("custom_font_size", 12)
                    # --- END NEW RESTORE ---

                    # --- Update UI Elements to reflect restored state ---
                    # Standard font UI
                    if hasattr(self, 'font_combo_box'):
                        self.font_combo_box.blockSignals(True)
                        # Find the font; setCurrentText might be less reliable than setCurrentFont
                        found_font = False
                        for i in range(self.font_combo_box.count()):
                            if self.font_combo_box.itemText(i) == self.font_family:
                                self.font_combo_box.setCurrentIndex(i)
                                found_font = True
                                break
                        if not found_font: # Fallback if font name isn't exact match
                            self.font_combo_box.setCurrentFont(QFont(self.font_family))
                        self.font_combo_box.blockSignals(False)
                    if hasattr(self, 'font_size_spinner'):
                        self.font_size_spinner.blockSignals(True)
                        self.font_size_spinner.setValue(self.font_size)
                        self.font_size_spinner.blockSignals(False)
                    if hasattr(self, 'font_color_button'):
                        self._update_color_button_style(self.font_color_button, self.font_color)
                    if hasattr(self, 'font_rotation_input'):
                        self.font_rotation_input.blockSignals(True)
                        self.font_rotation_input.setValue(self.font_rotation)
                        self.font_rotation_input.blockSignals(False)

                    # Slider positions (reflecting restored shifts)
                    if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
                    if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
                    if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

                    # --- NEW: Update Custom Marker UI ---
                    if hasattr(self, 'custom_marker_color_button'):
                        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                    if hasattr(self, 'custom_font_type_dropdown'):
                         self.custom_font_type_dropdown.blockSignals(True)
                         # Find the font like standard font combo
                         found_custom_font = False
                         for i in range(self.custom_font_type_dropdown.count()):
                             if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                                 self.custom_font_type_dropdown.setCurrentIndex(i)
                                 found_custom_font = True
                                 break
                         if not found_custom_font: # Fallback
                             self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                         self.custom_font_type_dropdown.blockSignals(False)
                         # Also update the entry box font if needed (optional)
                         # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont())
                    if hasattr(self, 'custom_font_size_spinbox'):
                         self.custom_font_size_spinbox.blockSignals(True)
                         self.custom_font_size_spinbox.setValue(restored_custom_font_size)
                         self.custom_font_size_spinbox.blockSignals(False)
                    # --- END NEW UI UPDATE ---

                    # Update other UI if needed (e.g., analysis text boxes)

                    # Refresh display and status bar
                    try:
                        self._update_preview_label_size()
                    except Exception: pass # Ignore errors if label size fails temporarily
                    self._update_status_bar()
                    self.update_live_view()
                    self.is_modified = True # Undoing makes it modified again relative to last save
                    
            
            def redo_action_m(self):
                """Redo the last undone action by restoring the next state, including font settings."""
                if self.redo_stack:
                    # Get current state of UI elements before overwriting self attributes
                    current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
                    current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

                    # Save the current state to the undo stack
                    current_state = {
                        "image": self.image.copy() if self.image else None,
                        "left_markers": self.left_markers.copy(),
                        "right_markers": self.right_markers.copy(),
                        "top_markers": self.top_markers.copy(),
                        "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                        "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                        "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                        "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                        "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                        # Standard marker font settings
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_color": self.font_color,
                        "font_rotation": self.font_rotation,
                        # Marker positioning shifts
                        "left_marker_shift_added": self.left_marker_shift_added,
                        "right_marker_shift_added": self.right_marker_shift_added,
                        "top_marker_shift_added": self.top_marker_shift_added,
                        # Analysis data
                        "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                        "quantities": self.quantities.copy(),
                        "protein_quantities": self.protein_quantities.copy(),
                        "standard_protein_areas": self.standard_protein_areas.copy(),
                        # --- NEW: Custom marker font/color states ---
                        "custom_marker_color": self.custom_marker_color,
                        "custom_font_family": current_custom_font_family, # Use value read from UI
                        "custom_font_size": current_custom_font_size,     # Use value read from UI
                        # --- END NEW ---
                    }
                    self.undo_stack.append(current_state)

                    # Restore the next state from the redo stack
                    next_state = self.redo_stack.pop()
                    self.image = next_state["image"]
                    self.left_markers = next_state["left_markers"]
                    self.right_markers = next_state["right_markers"]
                    self.top_markers = next_state["top_markers"]
                    self.custom_markers = next_state.get("custom_markers", [])
                    self.custom_shapes = next_state.get("custom_shapes", [])
                    self.image_before_padding = next_state["image_before_padding"]
                    self.image_contrasted = next_state["image_contrasted"]
                    self.image_before_contrast = next_state["image_before_contrast"]
                    # Restore standard font settings
                    self.font_family = next_state["font_family"]
                    self.font_size = next_state["font_size"]
                    self.font_color = next_state["font_color"]
                    self.font_rotation = next_state["font_rotation"]
                    # Restore shifts
                    self.left_marker_shift_added = next_state["left_marker_shift_added"]
                    self.right_marker_shift_added = next_state["right_marker_shift_added"]
                    self.top_marker_shift_added = next_state["top_marker_shift_added"]
                    # Restore analysis data
                    self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
                    self.quantities = next_state["quantities"]
                    self.protein_quantities = next_state["protein_quantities"]
                    self.standard_protein_areas = next_state["standard_protein_areas"]
                    # --- NEW: Restore custom marker font/color states ---
                    self.custom_marker_color = next_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
                    restored_custom_font_family = next_state.get("custom_font_family", "Arial")
                    restored_custom_font_size = next_state.get("custom_font_size", 12)
                    # --- END NEW RESTORE ---

                    # --- Update UI Elements to reflect restored state ---
                    # (Identical UI update logic as in undo_action_m)
                    # Standard font UI
                    if hasattr(self, 'font_combo_box'):
                        self.font_combo_box.blockSignals(True)
                        found_font = False
                        for i in range(self.font_combo_box.count()):
                            if self.font_combo_box.itemText(i) == self.font_family:
                                self.font_combo_box.setCurrentIndex(i)
                                found_font = True
                                break
                        if not found_font: self.font_combo_box.setCurrentFont(QFont(self.font_family))
                        self.font_combo_box.blockSignals(False)
                    if hasattr(self, 'font_size_spinner'):
                        self.font_size_spinner.blockSignals(True); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.blockSignals(False)
                    if hasattr(self, 'font_color_button'):
                        self._update_color_button_style(self.font_color_button, self.font_color)
                    if hasattr(self, 'font_rotation_input'):
                        self.font_rotation_input.blockSignals(True); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.blockSignals(False)

                    # Slider positions
                    if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
                    if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
                    if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

                    # --- NEW: Update Custom Marker UI ---
                    if hasattr(self, 'custom_marker_color_button'):
                        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                    if hasattr(self, 'custom_font_type_dropdown'):
                         self.custom_font_type_dropdown.blockSignals(True)
                         found_custom_font = False
                         for i in range(self.custom_font_type_dropdown.count()):
                             if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                                 self.custom_font_type_dropdown.setCurrentIndex(i)
                                 found_custom_font = True
                                 break
                         if not found_custom_font: self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                         self.custom_font_type_dropdown.blockSignals(False)
                         # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont()) # Optional
                    if hasattr(self, 'custom_font_size_spinbox'):
                         self.custom_font_size_spinbox.blockSignals(True); self.custom_font_size_spinbox.setValue(restored_custom_font_size); self.custom_font_size_spinbox.blockSignals(False)
                    # --- END NEW UI UPDATE ---

                    # Update other UI if needed

                    # Refresh display and status bar
                    try:
                        self._update_preview_label_size()
                    except Exception: pass
                    self._update_status_bar()
                    self.update_live_view()
                    self.is_modified = True # Redoing makes it modified again relative to last save
                    
            def analysis_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                # layout.setSpacing(15) # Increase spacing in this tab

                # --- Molecular Weight Prediction ---
                mw_group = QGroupBox("Molecular Weight Prediction")
                mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                mw_layout = QVBoxLayout(mw_group)
                # mw_layout.setSpacing(8)

                self.predict_button = QPushButton("Predict Molecular Weight")
                self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
                self.predict_button.setEnabled(False)  # Initially disabled
                self.predict_button.clicked.connect(self.predict_molecular_weight)
                mw_layout.addWidget(self.predict_button)

                layout.addWidget(mw_group)

                # --- Peak Area / Sample Quantification ---
                quant_group = QGroupBox("Peak Area and Sample Quantification")
                quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                quant_layout = QVBoxLayout(quant_group)
                # quant_layout.setSpacing(8)

                # Area Definition Buttons
                area_def_layout=QHBoxLayout()
                self.btn_define_quad = QPushButton("Define Quad Area")
                self.btn_define_quad.setToolTip(
                    "Click 4 corner points to define a region. \n"
                    "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
                    "Results may differ from Rectangle selection due to warping."
                )
                self.btn_define_quad.clicked.connect(self.enable_quad_mode)
                self.btn_define_rec = QPushButton("Define Rectangle Area")
                self.btn_define_rec.setToolTip(
                    "Click and drag to define a rectangular region. \n"
                    "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
                    "Does not correct for skew."
                )
                self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
                self.btn_sel_rec = QPushButton("Move Selected Area")
                self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
                self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
 
                area_def_layout.addWidget(self.btn_define_quad)
                area_def_layout.addWidget(self.btn_define_rec)
                area_def_layout.addWidget(self.btn_sel_rec)
                quant_layout.addLayout(area_def_layout)

                # Standard Processing
                std_proc_layout = QHBoxLayout()
                self.btn_process_std = QPushButton("Process Standard Bands")
                self.btn_process_std.setToolTip("Analyze the defined area as a standard lane.\nYou will be prompted for the known quantity.")
                self.btn_process_std.clicked.connect(self.process_standard)
                std_proc_layout.addWidget(self.btn_process_std)
                quant_layout.addLayout(std_proc_layout)

                # Standard Info Display (Read-only makes more sense)
                std_info_layout = QGridLayout()
                std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
                self.standard_protein_values = QLineEdit()
                self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
                self.standard_protein_values.setReadOnly(True) # Make read-only
                std_info_layout.addWidget(self.standard_protein_values, 0, 1)

                std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
                self.standard_protein_areas_text = QLineEdit()
                self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
                self.standard_protein_areas_text.setReadOnly(True) # Make read-only
                std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
                quant_layout.addLayout(std_info_layout)

                quant_layout.addWidget(self.create_separator()) # Add visual separator

                # Sample Processing
                sample_proc_layout = QHBoxLayout()
                self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
                self.btn_analyze_sample.setToolTip("Analyze the defined area as a sample lane using the standard curve.")
                self.btn_analyze_sample.clicked.connect(self.process_sample)
                sample_proc_layout.addWidget(self.btn_analyze_sample)
                quant_layout.addLayout(sample_proc_layout)


                # Sample Info Display
                sample_info_layout = QGridLayout()
                sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
                self.target_protein_areas_text = QLineEdit()
                self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
                self.target_protein_areas_text.setReadOnly(True)
                sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

                # Add Table Export button next to sample areas
                self.table_export_button = QPushButton("Export Results Table")
                self.table_export_button.setToolTip("Export the analysis results (areas, percentages, quantities) to Excel.")
                self.table_export_button.clicked.connect(self.open_table_window)
                sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
                quant_layout.addLayout(sample_info_layout)

                layout.addWidget(quant_group)

                # --- Clear Button ---
                clear_layout = QHBoxLayout() # Layout to center button maybe
                clear_layout.addStretch()
                self.clear_predict_button = QPushButton("Clear Analysis Markers")
                self.clear_predict_button.setToolTip("Clears MW prediction line and analysis regions.\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
                self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
                layout.addWidget(self.clear_predict_button)
                # clear_layout.addStretch()
                layout.addLayout(clear_layout)


                layout.addStretch()
                return tab
            
            def enable_move_selection_mode(self):
                """Enable mode to move the selected quadrilateral or rectangle."""
                if not self.live_view_label.quad_points and not self.live_view_label.bounding_box_preview:
                    QMessageBox.information(self, "Move Area", "No area (quadrilateral or rectangle) is currently defined to move.")
                    self.live_view_label.mode = None # Ensure mode is reset if no area
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    # Explicitly reset handlers if no area to move
                    self.live_view_label.mousePressEvent = None
                    self.live_view_label.mouseMoveEvent = None
                    self.live_view_label.mouseReleaseEvent = None
                    return

                self.live_view_label.mode = "move"
                self.live_view_label.setCursor(Qt.SizeAllCursor)
                
                # Explicitly unbind any previous specific handlers from other modes
                # to avoid conflicts before binding new ones.
                self.live_view_label.mousePressEvent = None
                self.live_view_label.mouseMoveEvent = None
                self.live_view_label.mouseReleaseEvent = None

                # Bind to CombinedSDSApp's handlers for the "move" mode
                self.live_view_label.mousePressEvent = self.start_move_selection
                # self.live_view_label.mouseMoveEvent will be set by start_move_selection
                self.live_view_label.mouseReleaseEvent = self.end_move_selection
                
                self.update_live_view()
                
            def start_move_selection(self, event):
                """Start moving the selection when the mouse is pressed. Stores initial state for snapped dragging."""
                if self.live_view_label.mode == "move" and event.button() == Qt.LeftButton:
                    # Get initial mouse position in label coordinates (unzoomed, unpanned view)
                    self.initial_mouse_pos_for_shape_drag_label = self.live_view_label.transform_point(event.pos())
                    
                    self.shape_points_at_drag_start_label = [] # Clear previous
                    if self.live_view_label.quad_points:
                        # quad_points are already in label coordinates
                        self.shape_points_at_drag_start_label = [QPointF(p) for p in self.live_view_label.quad_points]
                    elif self.live_view_label.bounding_box_preview:
                        # bounding_box_preview stores (x1, y1, x2, y2) in label coordinates
                        x1, y1, x2, y2 = self.live_view_label.bounding_box_preview
                        # Ensure normalization for consistent corner order for the drag operation
                        rect = QRectF(QPointF(x1,y1), QPointF(x2,y2)).normalized()
                        self.shape_points_at_drag_start_label = [
                            rect.topLeft(), rect.topRight(), rect.bottomRight(), rect.bottomLeft()
                        ]
                    
                    if not self.shape_points_at_drag_start_label: # Should not happen if enable_move_selection_mode checked
                        return

                    self.live_view_label.draw_edges = False # Optional: Hide edges during drag for smoother visual
                    self.live_view_label.mouseMoveEvent = self.move_selection # Connect move handler for the duration of the drag
                    # self.update_live_view() # No immediate update needed, wait for move
            
            def move_selection(self, event):
                """Move the selection while the mouse is being dragged, snapping the movement delta."""
                if self.live_view_label.mode == "move" and self.shape_points_at_drag_start_label and (event.buttons() & Qt.LeftButton):
                    current_mouse_pos_label = self.live_view_label.transform_point(event.pos())
                    
                    raw_delta_x_label = current_mouse_pos_label.x() - self.initial_mouse_pos_for_shape_drag_label.x()
                    raw_delta_y_label = current_mouse_pos_label.y() - self.initial_mouse_pos_for_shape_drag_label.y()

                    snapped_delta_x = raw_delta_x_label
                    snapped_delta_y = raw_delta_y_label

                    grid_size = 0
                    snap_x_enabled = False
                    snap_y_enabled = False

                    if hasattr(self, 'grid_size_input'):
                        grid_size = self.grid_size_input.value()
                    if hasattr(self, 'show_grid_checkbox_x'):
                        snap_x_enabled = self.show_grid_checkbox_x.isChecked()
                    if hasattr(self, 'show_grid_checkbox_y'):
                        snap_y_enabled = self.show_grid_checkbox_y.isChecked()

                    if grid_size > 0:
                        if snap_x_enabled:
                            snapped_delta_x = round(raw_delta_x_label / grid_size) * grid_size
                        if snap_y_enabled:
                            snapped_delta_y = round(raw_delta_y_label / grid_size) * grid_size
                    
                    snapped_delta_label = QPointF(snapped_delta_x, snapped_delta_y)

                    new_shape_points_label = []
                    for p_orig_label in self.shape_points_at_drag_start_label:
                        new_shape_points_label.append(p_orig_label + snapped_delta_label)

                    if self.live_view_label.quad_points: # If it was a quad initially
                        self.live_view_label.quad_points = new_shape_points_label
                        # self.live_view_label.bounding_box_preview = None # Keep this if we want to visually show quad points
                    elif self.live_view_label.bounding_box_preview: # If it was a rect initially
                        if len(new_shape_points_label) == 4:
                            # Reconstruct the bounding box (min x, min y, max x, max y) from the moved points
                            all_x = [p.x() for p in new_shape_points_label]
                            all_y = [p.y() for p in new_shape_points_label]
                            min_x, max_x = min(all_x), max(all_x)
                            min_y, max_y = min(all_y), max(all_y)
                            self.live_view_label.bounding_box_preview = (min_x, min_y, max_x, max_y)
                        # self.live_view_label.quad_points = [] # Keep this if we want to visually show rect
                    self.update_live_view()
            
            def end_move_selection(self, event):
                """End moving the selection when the mouse is released."""
                if self.live_view_label.mode == "move" and event.button() == Qt.LeftButton:
                    self.live_view_label.draw_edges = True # Restore edge drawing
                    
                    # Clear helper attributes used during drag
                    self.shape_points_at_drag_start_label = []
                    self.initial_mouse_pos_for_shape_drag_label = QPointF()

                    # Unbind the specific mouseMoveEvent handler
                    self.live_view_label.mouseMoveEvent = None 
                    
                    # After movement, if it was a rectangle represented by bounding_box_preview,
                    # ensure rectangle_points (if used by other logic) is also updated.
                    # For now, bounding_box_preview is the primary representation for rectangles.
                    
                    self.update_live_view()
                
                
                     
            def get_nearest_point(self, mouse_pos, points):
                """Get the nearest point to the mouse position."""
                min_distance = float('inf')
                nearest_point = None
                for point in points:
                    distance = (mouse_pos - point).manhattanLength()
                    if distance < min_distance:
                        min_distance = distance
                        nearest_point = point
                return nearest_point
            
            def open_table_window(self):
                peak_areas_to_show_current = self.latest_peak_areas
                calculated_quantities_to_show_current = self.latest_calculated_quantities
                standard_dict_to_show_current = self.quantities_peak_area_dict
                is_standard_mode_current = len(standard_dict_to_show_current) >= 2
            
                self.table_window_instance = TableWindow(
                    peak_areas_to_show_current,
                    standard_dict_to_show_current,
                    is_standard_mode_current,
                    calculated_quantities_to_show_current,
                    self
                )
                self.table_window_instance.show()
            
            def enable_quad_mode(self):
                """Enable mode to define a quadrilateral area."""
                self.live_view_label.bounding_box_preview = []
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.mode = "quad"
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                # # Reset mouse event handlers
                self.live_view_label.mousePressEvent = None
                self.live_view_label.mouseMoveEvent = None
                self.live_view_label.mouseReleaseEvent = None
                
                # Update the live view
                self.update_live_view()
            
            def enable_rectangle_mode(self):
                """Enable mode to define a rectangle area."""
                self.live_view_label.bounding_box_preview = []
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.mode = "rectangle"
                self.live_view_label.rectangle_points = []  # Clear previous rectangle points
                self.live_view_label.rectangle_start = None  # Reset rectangle start
                self.live_view_label.rectangle_end = None    # Reset rectangle end
                self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                # Set mouse event handlers for rectangle mode
                self.live_view_label.mousePressEvent = self.start_rectangle
                self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
                self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
                
                # Update the live view
                self.update_live_view()
            
            def snap_point_to_grid(self, point: QPointF) -> QPointF:
                """Snaps a QPointF to the grid if enabled."""
                snapped_x = point.x()
                snapped_y = point.y()
    
                grid_size = 0
                snap_x_enabled = False
                snap_y_enabled = False
    
                if hasattr(self, 'grid_size_input'):
                    grid_size = self.grid_size_input.value()
                if hasattr(self, 'show_grid_checkbox_x'):
                    snap_x_enabled = self.show_grid_checkbox_x.isChecked()
                if hasattr(self, 'show_grid_checkbox_y'):
                    snap_y_enabled = self.show_grid_checkbox_y.isChecked()
    
                if grid_size > 0:
                    if snap_x_enabled:
                        snapped_x = round(point.x() / grid_size) * grid_size
                    if snap_y_enabled:
                        snapped_y = round(point.y() / grid_size) * grid_size
                
                return QPointF(snapped_x, snapped_y)
            
            def start_rectangle(self, event):
                """Record the start position of the rectangle (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"]:
                    start_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    self.live_view_label.rectangle_start = snapped_start_point
                    self.live_view_label.rectangle_points = [snapped_start_point]
                    self.live_view_label.bounding_box_preview = None
            
            def update_rectangle_preview(self, event):
                """Update the rectangle preview as the mouse moves (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"] and self.live_view_label.rectangle_start:
                    current_end_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_end_point = self.snap_point_to_grid(current_end_point_transformed) # Snap it
    
                    self.live_view_label.rectangle_end = snapped_end_point
            
                    if self.live_view_label.rectangle_start and self.live_view_label.rectangle_end:
                        self.live_view_label.bounding_box_preview = (
                            self.live_view_label.rectangle_start.x(),
                            self.live_view_label.rectangle_start.y(),
                            self.live_view_label.rectangle_end.x(), # Use snapped end point
                            self.live_view_label.rectangle_end.y(), # Use snapped end point
                        )
                        self.update_live_view()
            
            def finalize_rectangle(self, event):
                """Finalize the rectangle when the mouse is released."""
                if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
                    end_point_transformed = self.live_view_label.transform_point(event.pos())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
    
                    self.live_view_label.rectangle_end = snapped_end_point
                    # self.live_view_label.rectangle_start is already snapped
                    self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, snapped_end_point]
                    
                    self.live_view_label.bounding_box_preview = (
                        self.live_view_label.rectangle_start.x(),
                        self.live_view_label.rectangle_start.y(),
                        snapped_end_point.x(), # Use snapped end point
                        snapped_end_point.y(), # Use snapped end point
                    )
                    
                    self.live_view_label.mode = None
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self.update_live_view()
                
            
            def process_standard(self):
                """Processes the defined region (quad or rect) as a standard."""
                extracted_qimage = None # Will hold the QImage of the extracted region
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                self.update_live_view()       

                if len(self.live_view_label.quad_points) == 4:
                    # --- Quadrilateral Logic ---
                    print("Processing Standard: Quadrilateral")
                    # Pass the *current* self.image (could be color or gray)
                    extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
                    if not extracted_qimage or extracted_qimage.isNull():
                        QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                        return

                elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
                    # --- Rectangle Logic ---
                    print("Processing Standard: Rectangle")
                    try:
                        # (Coordinate Transformation Logic - KEEP AS IS from previous fix)
                        start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                        # ... (rest of coordinate transformation logic) ...
                        zoom = self.live_view_label.zoom_level
                        offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                        start_x_unzoomed = (start_x_view - offset_x) / zoom
                        start_y_unzoomed = (start_y_view - offset_y) / zoom
                        end_x_unzoomed = (end_x_view - offset_x) / zoom
                        end_y_unzoomed = (end_y_view - offset_y) / zoom
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w, img_h = self.image.width(), self.image.height()
                        label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                        scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                        display_offset_x = (label_w - img_w * scale_factor) / 2
                        display_offset_y = (label_h - img_h * scale_factor) / 2
                        start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                        start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                        end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                        end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                        x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                        w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                        if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                        x_clamped, y_clamped = max(0, x), max(0, y)
                        w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                        if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                        # --- End Coordinate Transformation ---

                        extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                        if extracted_qimage.isNull():
                            raise ValueError("QImage.copy failed for rectangle.")

                    except Exception as e:
                         print(f"Error processing rectangle region for standard: {e}")
                         QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                         return
                else:
                    QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
                    return

                # --- Convert extracted region to Grayscale PIL for analysis ---
                if extracted_qimage and not extracted_qimage.isNull():
                    processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                    if processed_data_pil:
                        # Call analyze_bounding_box with the Grayscale PIL image
                        self.analyze_bounding_box(processed_data_pil, standard=True)
                    else:
                        QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")
                # No else needed, errors handled above
            
            def process_sample(self):
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                self.update_live_view()  
                """Processes the defined region (quad or rect) as a sample."""
                extracted_qimage = None # Will hold the QImage of the extracted region

                if len(self.live_view_label.quad_points) == 4:
                    # --- Quadrilateral Logic ---
                    print("Processing Sample: Quadrilateral")
                    extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
                    if not extracted_qimage or extracted_qimage.isNull():
                        QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                        return

                elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
                    # --- Rectangle Logic ---
                    print("Processing Sample: Rectangle")
                    try:
                        # (Coordinate Transformation Logic - KEEP AS IS)
                        start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                        # ... (rest of coordinate transformation logic) ...
                        zoom = self.live_view_label.zoom_level
                        offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                        start_x_unzoomed = (start_x_view - offset_x) / zoom
                        start_y_unzoomed = (start_y_view - offset_y) / zoom
                        end_x_unzoomed = (end_x_view - offset_x) / zoom
                        end_y_unzoomed = (end_y_view - offset_y) / zoom
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w, img_h = self.image.width(), self.image.height()
                        label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                        scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                        display_offset_x = (label_w - img_w * scale_factor) / 2
                        display_offset_y = (label_h - img_h * scale_factor) / 2
                        start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                        start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                        end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                        end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                        x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                        w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                        if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                        x_clamped, y_clamped = max(0, x), max(0, y)
                        w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                        if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                        # --- End Coordinate Transformation ---

                        extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                        if extracted_qimage.isNull():
                            raise ValueError("QImage.copy failed for rectangle.")

                    except Exception as e:
                         print(f"Error processing rectangle region for sample: {e}")
                         QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                         return
                else:
                    QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
                    return

                # --- Convert extracted region to Grayscale PIL for analysis ---
                if extracted_qimage and not extracted_qimage.isNull():
                    processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                    if processed_data_pil:
                        # Call analyze_bounding_box with the Grayscale PIL image
                        self.analyze_bounding_box(processed_data_pil, standard=False)
                    else:
                        QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")


            def convert_qimage_to_grayscale_pil(self, qimg):
                """
                Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
                for use with PeakAreaDialog. Returns None on failure.
                """
                if not qimg or qimg.isNull():
                    return None

                fmt = qimg.format()
                pil_img = None

                try:
                    # Already grayscale? Convert directly if possible.
                    if fmt == QImage.Format_Grayscale16:
                        print("Converting Grayscale16 QImage to PIL 'I;16'")
                        np_array = self.qimage_to_numpy(qimg)
                        if np_array is not None and np_array.dtype == np.uint16:
                            try: pil_img = Image.fromarray(np_array, mode='I;16')
                            except ValueError: pil_img = Image.fromarray(np_array, mode='I')
                        else: raise ValueError("Failed NumPy conversion for Grayscale16")
                    elif fmt == QImage.Format_Grayscale8:
                        print("Converting Grayscale8 QImage to PIL 'L'")
                        # Try direct conversion first
                        try:
                            pil_img = ImageQt.fromqimage(qimg).convert('L')
                            if pil_img is None: raise ValueError("Direct QImage->PIL(L) failed.")
                        except Exception as e_direct:
                            print(f"Direct QImage->PIL(L) failed ({e_direct}), falling back via NumPy.")
                            np_array = self.qimage_to_numpy(qimg)
                            if np_array is not None and np_array.dtype == np.uint8:
                                pil_img = Image.fromarray(np_array, mode='L')
                            else: raise ValueError("Failed NumPy conversion for Grayscale8")
                    else: # Color or other format
                        print(f"Converting format {fmt} QImage to Grayscale PIL")
                        # Use NumPy for robust conversion to 16-bit grayscale intermediate
                        np_img = self.qimage_to_numpy(qimg)
                        if np_img is None: raise ValueError("NumPy conversion failed for color.")
                        if np_img.ndim == 3:
                            gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) # Assume BGR/BGRA input
                            # Convert to 16-bit PIL
                            gray_np_16bit = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                            try: pil_img = Image.fromarray(gray_np_16bit, mode='I;16')
                            except ValueError: pil_img = Image.fromarray(gray_np_16bit, mode='I')
                        elif np_img.ndim == 2: # Should have been caught by Grayscale checks, but handle anyway
                             if np_img.dtype == np.uint16:
                                 try: pil_img = Image.fromarray(np_img, mode='I;16')
                                 except ValueError: pil_img = Image.fromarray(np_img, mode='I')
                             else: # Assume uint8 or other, convert to L
                                 pil_img = Image.fromarray(np_img).convert('L')
                        else:
                             raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

                    if pil_img is None:
                        raise ValueError("PIL Image creation failed.")

                    print(f"  Successfully converted to PIL Image: mode={pil_img.mode}, size={pil_img.size}")
                    return pil_img

                except Exception as e:
                    print(f"Error in convert_qimage_to_grayscale_pil: {e}")
                    traceback.print_exc()
                    return None            
                
                
                    
            def combine_image_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(10)
        
                # Initial placeholder ranges - will be updated by _update_overlay_slider_ranges()
                initial_dim_placeholder = 1000
                initial_x_range_min = -initial_dim_placeholder
                initial_x_range_max = initial_dim_placeholder * 2
                initial_y_range_min = -initial_dim_placeholder
                initial_y_range_max = initial_dim_placeholder * 2
        
        
                # --- Image 1 Group ---
                image1_group = QGroupBox("Image 1 Overlay")
                image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image1_layout = QGridLayout(image1_group)
                image1_layout.setSpacing(8)
        
                copy_image1_button = QPushButton("Copy Current Image")
                copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
                copy_image1_button.clicked.connect(self.save_image1)
                place_image1_button = QPushButton("Place Image 1")
                place_image1_button.setToolTip("Positions Image 1 based on sliders.")
                place_image1_button.clicked.connect(self.place_image1)
                remove_image1_button = QPushButton("Remove Image 1")
                remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
                remove_image1_button.clicked.connect(self.remove_image1)
        
                image1_layout.addWidget(copy_image1_button, 0, 0)
                image1_layout.addWidget(place_image1_button, 0, 1)
                image1_layout.addWidget(remove_image1_button, 0, 2)
        
                image1_layout.addWidget(QLabel("Horizontal Pos (px):"), 1, 0)
                self.image1_left_slider = QSlider(Qt.Horizontal)
                self.image1_left_slider.setRange(initial_x_range_min, initial_x_range_max)
                self.image1_left_slider.setValue(0)
                self.image1_left_slider.valueChanged.connect(self.place_image1) # Update position on value change
                image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2)
        
                image1_layout.addWidget(QLabel("Vertical Pos (px):"), 2, 0)
                self.image1_top_slider = QSlider(Qt.Horizontal)
                self.image1_top_slider.setRange(initial_y_range_min, initial_y_range_max)
                self.image1_top_slider.setValue(0)
                self.image1_top_slider.valueChanged.connect(self.place_image1) # Update position on value change
                image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2)
        
                image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
                self.image1_resize_slider = QSlider(Qt.Horizontal)
                self.image1_resize_slider.setRange(10, 300)
                self.image1_resize_slider.setValue(100)
                self.image1_resize_slider.valueChanged.connect(self.update_live_view) # Resize triggers full redraw
                self.image1_resize_label = QLabel("100%")
                self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
                image1_layout.addWidget(self.image1_resize_slider, 3, 1)
                image1_layout.addWidget(self.image1_resize_label, 3, 2)
                layout.addWidget(image1_group)
        
                # --- Image 2 Group ---
                image2_group = QGroupBox("Image 2 Overlay")
                image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image2_layout = QGridLayout(image2_group)
                # ... (similar setup for image2 sliders and connects) ...
                copy_image2_button = QPushButton("Copy Current Image")
                copy_image2_button.clicked.connect(self.save_image2)
                place_image2_button = QPushButton("Place Image 2")
                place_image2_button.clicked.connect(self.place_image2)
                remove_image2_button = QPushButton("Remove Image 2")
                remove_image2_button.clicked.connect(self.remove_image2)
                image2_layout.addWidget(copy_image2_button, 0, 0)
                image2_layout.addWidget(place_image2_button, 0, 1)
                image2_layout.addWidget(remove_image2_button, 0, 2)
        
                image2_layout.addWidget(QLabel("Horizontal Pos (px):"), 1, 0)
                self.image2_left_slider = QSlider(Qt.Horizontal)
                self.image2_left_slider.setRange(initial_x_range_min, initial_x_range_max)
                self.image2_left_slider.setValue(0)
                self.image2_left_slider.valueChanged.connect(self.place_image2)
                image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)
        
                image2_layout.addWidget(QLabel("Vertical Pos (px):"), 2, 0)
                self.image2_top_slider = QSlider(Qt.Horizontal)
                self.image2_top_slider.setRange(initial_y_range_min, initial_y_range_max)
                self.image2_top_slider.setValue(0)
                self.image2_top_slider.valueChanged.connect(self.place_image2)
                image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)
        
                image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
                self.image2_resize_slider = QSlider(Qt.Horizontal)
                self.image2_resize_slider.setRange(10, 300)
                self.image2_resize_slider.setValue(100)
                self.image2_resize_slider.valueChanged.connect(self.update_live_view)
                self.image2_resize_label = QLabel("100%")
                self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
                image2_layout.addWidget(self.image2_resize_slider, 3, 1)
                image2_layout.addWidget(self.image2_resize_label, 3, 2)
                layout.addWidget(image2_group)
        
        
                finalize_button = QPushButton("Rasterize Image")
                finalize_button.setToolTip("Permanently merges the placed overlays onto the main image.")
                finalize_button.clicked.connect(self.finalize_combined_image)
                layout.addWidget(finalize_button)
        
                layout.addStretch()
        
                # Call after UI elements are created, in case an image is already loaded
                self._update_overlay_slider_ranges()
        
                return tab
            
            def remove_image1(self):
                """Remove Image 1 and reset its sliders."""
                if hasattr(self, 'image1'):
                    # del self.image1
                    # del self.image1_original
                    try:
                        del self.image1_position
                    except:
                        pass
                    self.image1_left_slider.setValue(0)
                    self.image1_top_slider.setValue(0)
                    self.image1_resize_slider.setValue(100)
                    self.update_live_view()
            
            def remove_image2(self):
                """Remove Image 2 and reset its sliders."""
                if hasattr(self, 'image2'):
                    # del self.image2
                    # del self.image2_original
                    try:
                        del self.image2_position
                    except:
                        pass
                    self.image2_left_slider.setValue(0)
                    self.image2_top_slider.setValue(0)
                    self.image2_resize_slider.setValue(100)
                    self.update_live_view()
            
            def save_image1(self):
                if self.image:
                    self.image1 = self.image.copy()
                    self.image1_original = self.image1.copy()  # Save the original image for resizing
                    QMessageBox.information(self, "Success", "Image 1 copied.")
            
            def save_image2(self):
                if self.image:
                    self.image2 = self.image.copy()
                    self.image2_original = self.image2.copy()  # Save the original image for resizing
                    QMessageBox.information(self, "Success", "Image 2 copied.")
            
            def place_image1(self):
                if hasattr(self, 'image1'):
                    self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
                    self.update_live_view()
            
            def place_image2(self):
                if hasattr(self, 'image2'):
                    self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                    self.update_live_view()
            
            
            
            def finalize_combined_image(self):
                """
                Rasterizes overlays and annotations onto the image by generating
                a high-resolution canvas identical to the preview rendering process
                (but without guides/previews) and assigning it as the new self.image.
                This ensures visual consistency between preview and final result.
                The final image resolution might be higher than the original if render_scale > 1.
                """
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Warning", "No base image loaded.")
                    return

                # --- Save state before modifying ---
                self.save_state()

                # --- 1. Determine Render/Canvas Dimensions (Same as Preview) ---
                render_scale = 3 # Use the same scale factor as the preview
                try:
                    # Use current view dimensions to determine target render size
                    # Fallback if view is not ready
                    view_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 600
                    view_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 400
                    target_render_width = view_width * render_scale
                    target_render_height = view_height * render_scale
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Could not calculate render dimensions: {e}")
                    return

                # --- 2. Prepare the Base Image for Rendering ---
                # Scale the *current* self.image to fit within the target render dimensions,
                # just like it's done at the start of update_live_view before calling render_image_on_canvas.
                # Check if self.image is valid before scaling
                if self.image.isNull() or self.image.width() <= 0 or self.image.height() <= 0:
                     QMessageBox.critical(self, "Error", "Current base image is invalid before scaling.")
                     return

                scaled_image_for_render = self.image.scaled(
                    target_render_width,
                    target_render_height,
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation,
                )
                if scaled_image_for_render.isNull():
                    QMessageBox.critical(self, "Error", "Failed to scale base image for final rendering.")
                    return

                # --- 3. Create the High-Resolution Target Canvas ---
                # Use ARGB32_Premultiplied for safe drawing with alpha
                final_canvas = QImage(target_render_width, target_render_height, QImage.Format_ARGB32_Premultiplied)
                if final_canvas.isNull():
                    QMessageBox.critical(self, "Error", "Failed to create final high-resolution canvas.")
                    return
                final_canvas.fill(Qt.transparent) # Fill with transparent background

                # --- 4. Render onto the High-Res Canvas using the Preview Logic ---
                # Call the *exact same* rendering function used for the preview,
                # but disable drawing of guides and temporary previews.
                # Pass the scaled base image prepared in step 2.
                # Pass render_scale so elements inside render_image_on_canvas scale correctly.
                try:
                    # x_start and y_start are typically 0 when rendering the current self.image
                    # unless cropping offsets need to be handled differently, but render_image_on_canvas
                    # should already handle positioning based on the scaled_image provided.
                    self.render_image_on_canvas(
                        canvas=final_canvas,
                        scaled_image=scaled_image_for_render,
                        x_start=0, # Start position relative to the scaled_image itself
                        y_start=0,
                        render_scale=render_scale,
                        draw_guides=False # IMPORTANT: Disable guides for final output
                    )
                    # Inside render_image_on_canvas, ensure shape previews are also skipped if needed,
                    # although technically they shouldn't exist when finalize is called.
                    # Consider adding a flag to render_image_on_canvas like `is_finalizing=True`
                    # to explicitly skip previews if necessary.

                except Exception as e:
                     QMessageBox.critical(self, "Render Error", f"Failed during final rendering: {e}")
                     traceback.print_exc()
                     return

                # --- 5. Assign the High-Resolution Canvas as the New Image ---
                if final_canvas.isNull():
                     QMessageBox.critical(self, "Error", "Final rendered canvas became invalid.")
                     return

                self.image = final_canvas # The new self.image IS the high-res render
                self.is_modified = True

                # --- 6. Update Backups and State ---
                # The backups now store the high-resolution rendered image
                self.image_before_padding = self.image.copy() # New baseline includes baked-in elements
                self.image_contrasted = self.image.copy()
                self.image_before_contrast = self.image.copy() # Reset contrast baseline
                self.image_padded = True # Consider the result "padded" with overlays
                self._update_marker_slider_ranges() # Adjust ranges for new dimensions
                self._update_overlay_slider_ranges()
                # --- 7. Cleanup and UI Update ---
                self.remove_image1() # Remove buffer overlays
                self.remove_image2()

                self._update_preview_label_size() # Update label based on potentially larger image
                self._update_status_bar()         # Reflect new dimensions/depth
                
                self.update_live_view()           # Refresh display

                QMessageBox.information(self, "Success", "The overlays and annotations have been rasterized onto the image.\n(Note: Final image resolution might be higher than original)")
            
                
            


            def font_and_image_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(15) # Add spacing between groups

                # --- Font Options Group ---
                font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
                font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                font_options_layout = QGridLayout(font_options_group)
                font_options_layout.setSpacing(8)

                # Font type
                font_type_label = QLabel("Font Family:")
                self.font_combo_box = QFontComboBox()
                self.font_combo_box.setEditable(False)
                self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
                self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_type_label, 0, 0)
                font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

                # Font size
                font_size_label = QLabel("Font Size:")
                self.font_size_spinner = QSpinBox()
                self.font_size_spinner.setRange(6, 72)  # Adjusted range
                self.font_size_spinner.setValue(self.font_size) # Use initialized value
                self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_size_label, 1, 0)
                font_options_layout.addWidget(self.font_size_spinner, 1, 1)

                # Font color
                self.font_color_button = QPushButton("Font Color")
                self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
                self.font_color_button.clicked.connect(self.select_font_color)
                self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
                font_options_layout.addWidget(self.font_color_button, 1, 2)

                # Font rotation (Top/Bottom)
                font_rotation_label = QLabel("Top Label Rotation:") # Specific label
                self.font_rotation_input = QSpinBox()
                self.font_rotation_input.setRange(-180, 180)
                self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
                self.font_rotation_input.setSuffix(" °") # Add degree symbol
                self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
                font_options_layout.addWidget(font_rotation_label, 2, 0)
                font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

                layout.addWidget(font_options_group)


                # --- Image Adjustments Group ---
                img_adjust_group = QGroupBox("Image Adjustments")
                img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                img_adjust_layout = QGridLayout(img_adjust_group)
                img_adjust_layout.setSpacing(8)

                # Contrast Sliders with Value Labels
                high_contrast_label = QLabel("Brightness:") # Renamed for clarity
                self.high_slider = QSlider(Qt.Horizontal)
                self.high_slider.setRange(0, 200) # Range 0 to 200%
                self.high_slider.setValue(100) # Default 100%
                self.high_slider.valueChanged.connect(self.update_image_contrast)
                self.high_value_label = QLabel("1.00") # Display factor
                self.high_slider.valueChanged.connect(lambda val, lbl=self.high_value_label: lbl.setText(f"{val/100.0:.2f}"))
                img_adjust_layout.addWidget(high_contrast_label, 0, 0)
                img_adjust_layout.addWidget(self.high_slider, 0, 1)
                img_adjust_layout.addWidget(self.high_value_label, 0, 2)

                low_contrast_label = QLabel("Contrast:") # Renamed for clarity
                self.low_slider = QSlider(Qt.Horizontal)
                self.low_slider.setRange(0, 200) # Range 0 to 100%
                self.low_slider.setValue(100) # Default 100%
                self.low_slider.valueChanged.connect(self.update_image_contrast)
                self.low_value_label = QLabel("1.00") # Display factor
                self.low_slider.valueChanged.connect(lambda val, lbl=self.low_value_label: lbl.setText(f"{val/100.0:.2f}"))
                img_adjust_layout.addWidget(low_contrast_label, 1, 0)
                img_adjust_layout.addWidget(self.low_slider, 1, 1)
                img_adjust_layout.addWidget(self.low_value_label, 1, 2)


                # Gamma Adjustment Slider
                gamma_label = QLabel("Gamma:")
                self.gamma_slider = QSlider(Qt.Horizontal)
                self.gamma_slider.setRange(10, 300)  # Range 0.1 to 3.0
                self.gamma_slider.setValue(100)  # Default 1.0 (100%)
                self.gamma_slider.valueChanged.connect(self.update_image_gamma)
                self.gamma_value_label = QLabel("1.00") # Display factor
                self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
                img_adjust_layout.addWidget(gamma_label, 2, 0)
                img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
                img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


                # Separator
                img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

                # Action Buttons
                btn_layout = QHBoxLayout()
                self.bw_button = QPushButton("Grayscale")
                self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
                self.bw_button.clicked.connect(self.convert_to_black_and_white)
                invert_button = QPushButton("Invert")
                invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
                invert_button.clicked.connect(self.invert_image)
                reset_button = QPushButton("Reset Adjustments")
                reset_button.setToolTip("Reset Brightness, Contrast, and Gamma sliders to default.\n(Does not undo Grayscale or Invert)")
                reset_button.clicked.connect(self.reset_gamma_contrast)
                btn_layout.addWidget(self.bw_button)
                btn_layout.addWidget(invert_button)
                btn_layout.addStretch() # Push reset button to the right
                btn_layout.addWidget(reset_button)

                img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

                layout.addWidget(img_adjust_group)
                layout.addStretch() # Push groups up
                return tab

            
            def _update_color_button_style(self, button, color):
                """ Helper to update button background color preview """
                if color.isValid():
                    button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
                else:
                    button.setStyleSheet("") # Reset to default stylesheet
            
            def create_separator(self):
                """Creates a horizontal separator line."""
                line = QFrame()
                line.setFrameShape(QFrame.HLine)
                line.setFrameShadow(QFrame.Sunken)
                return line
            
            
            def create_cropping_tab(self):
                """Create the Cropping tab with Rectangle Draw and Slider options."""
                tab = QWidget()
                layout = QVBoxLayout(tab)


                # --- Alignment Group (Keep as is - assume it's defined elsewhere) ---
                alignment_params_group = QGroupBox("Alignment Options")
                alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                # Use a QVBoxLayout for the main group layout
                alignment_group_layout = QVBoxLayout(alignment_params_group)
                alignment_group_layout.setSpacing(8) # Spacing between the two rows

                # --- Row 1: Guides, Rotation Controls ---
                rotation_controls_layout = QHBoxLayout()
                rotation_controls_layout.setSpacing(6) # Spacing within the row

                self.show_guides_label = QLabel("Show Guide Lines:")
                self.show_guides_checkbox = QCheckBox("", self)
                self.show_guides_checkbox.setChecked(False)
                self.show_guides_checkbox.stateChanged.connect(self.update_live_view)

                self.orientation_label = QLabel("Rotation Angle (0.00°)")
                # Make label width flexible but give it a minimum
                self.orientation_label.setMinimumWidth(150)
                # Allow label to shrink/grow slightly if needed:
                self.orientation_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)

                self.orientation_slider = QSlider(Qt.Horizontal)
                self.orientation_slider.setRange(-3600, 3600)
                self.orientation_slider.setValue(0)
                self.orientation_slider.setSingleStep(1)
                self.orientation_slider.valueChanged.connect(self._update_rotation_label)
                # Update preview only when slider is released for performance
                self.orientation_slider.valueChanged.connect(self.update_live_view)
                # Make slider expand to take available space
                self.orientation_slider.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                self.align_button = QPushButton("Apply Rotation")
                self.align_button.clicked.connect(self.align_image)

                self.reset_align_button = QPushButton("Reset Rotation")
                self.reset_align_button.clicked.connect(self.reset_align_image)

                # Add widgets to the first row layout
                rotation_controls_layout.addWidget(self.show_guides_label)
                rotation_controls_layout.addWidget(self.show_guides_checkbox)
                rotation_controls_layout.addSpacing(10) # Add a small visual gap
                rotation_controls_layout.addWidget(self.orientation_label)
                rotation_controls_layout.addWidget(self.orientation_slider) # Let slider expand
                rotation_controls_layout.addWidget(self.align_button)
                rotation_controls_layout.addWidget(self.reset_align_button)

                # Add the first row layout to the main group layout
                alignment_group_layout.addLayout(rotation_controls_layout)

                # --- Row 2: Flip Controls ---
                flip_controls_layout = QHBoxLayout()
                flip_controls_layout.setSpacing(6) # Spacing between flip buttons

                self.flip_vertical_button = QPushButton("Flip Vertical")
                self.flip_vertical_button.clicked.connect(self.flip_vertical)
                # Make buttons expand equally to fill width
                self.flip_vertical_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                self.flip_horizontal_button = QPushButton("Flip Horizontal")
                self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
                # Make buttons expand equally to fill width
                self.flip_horizontal_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

                # Add widgets to the second row layout
                # No stretch needed here as buttons expand
                flip_controls_layout.addWidget(self.flip_vertical_button)
                flip_controls_layout.addWidget(self.flip_horizontal_button)

                # Add the second row layout to the main group layout
                alignment_group_layout.addLayout(flip_controls_layout)
                
                # guide_layout.addStretch()
                

                alignment_params_group.setLayout(alignment_group_layout)
                layout.addWidget(alignment_params_group)
                # --- End Alignment Group ---

                # --- Skew Fix Group (Keep as is - assume it's defined elsewhere) ---
                taper_skew_group = QGroupBox("Skew Fix")
                taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                # ... (Add skew widgets here) ...
                # Example structure:
                taper_skew_layout = QHBoxLayout()
                self.taper_skew_label = QLabel("Tapering Skew (0.00)")
                self.taper_skew_label.setFixedWidth(150)
                self.taper_skew_slider = QSlider(Qt.Horizontal)
                self.taper_skew_slider.setRange(-70, 70)
                self.taper_skew_slider.setValue(0)
                self.taper_skew_slider.valueChanged.connect(self._update_skew_label)
                self.taper_skew_slider.valueChanged.connect(self.update_live_view)
                self.skew_button = QPushButton("Apply Skew")
                self.skew_button.clicked.connect(self.update_skew)
                taper_skew_layout.addWidget(self.taper_skew_label)
                taper_skew_layout.addWidget(self.taper_skew_slider)
                taper_skew_layout.addWidget(self.skew_button)
                taper_skew_group.setLayout(taper_skew_layout)
                layout.addWidget(taper_skew_group)
                # --- End Skew Fix Group ---


                # --- START: Modified Cropping Group ---
                cropping_params_group = QGroupBox("Cropping Options")
                cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                cropping_layout = QVBoxLayout(cropping_params_group)      
                

                # --- Draw Rectangle Button ---
                self.draw_crop_rect_button = QPushButton("Draw Crop Rectangle")
                self.draw_crop_rect_button.setToolTip("Click and drag on the image to define the crop area.\nThis will enable the sliders below for fine-tuning.")
                self.draw_crop_rect_button.setCheckable(True)
                self.draw_crop_rect_button.clicked.connect(self.toggle_rectangle_crop_mode)
                self.apply_crop_button = QPushButton("Apply Crop")
                self.apply_crop_button.setToolTip("Apply the defined crop (rectangle or sliders).")
                self.apply_crop_button.clicked.connect(self.update_crop)
                cropping_layout.addWidget(self.draw_crop_rect_button)
                cropping_layout.addWidget(self.apply_crop_button)

                # --- Separator ---
                cropping_layout.addWidget(self.create_separator())

                # --- Sliders for Fine-Tuning ---
                crop_slider_layout = QGridLayout()

                # --- Define Slider Range and Precision ---
                self.crop_slider_min = 0
                self.crop_slider_max = 10000 # Represents 0.00% to 100.00%
                self.crop_slider_precision_factor = 100.0 # Divide slider value by this

                # Helper to create label for slider value display
                def create_value_label(initial_value=0.0):
                    # Format to 2 decimal places for hundredths of percent
                    lbl = QLabel(f"{initial_value:.2f}%")
                    lbl.setMinimumWidth(50) # Ensure space for "100.00%"
                    lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    return lbl

                # --- Left Crop Slider (X Start) ---
                crop_x_start_label = QLabel("Crop Left:")
                self.crop_x_start_slider = QSlider(Qt.Horizontal)
                self.crop_x_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_x_start_slider.setValue(self.crop_slider_min)      # Default: Start at 0.00%
                self.crop_x_start_slider.setToolTip("Adjust the left edge of the crop area (enabled after drawing).")
                self.crop_x_start_value_label = create_value_label(0.00)
                self.crop_x_start_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_x_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_x_start_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_x_start_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
                crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
                crop_slider_layout.addWidget(self.crop_x_start_value_label, 0, 2)

                # --- Right Crop Slider (X End) ---
                crop_x_end_label = QLabel("Crop Right:")
                self.crop_x_end_slider = QSlider(Qt.Horizontal)
                self.crop_x_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_x_end_slider.setValue(self.crop_slider_max)    # Default: End at 100.00%
                self.crop_x_end_slider.setToolTip("Adjust the right edge of the crop area (enabled after drawing).")
                self.crop_x_end_value_label = create_value_label(100.00)
                self.crop_x_end_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_x_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_x_end_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_x_end_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_x_end_label, 0, 3)
                crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 4)
                crop_slider_layout.addWidget(self.crop_x_end_value_label, 0, 5)

                # --- Top Crop Slider (Y Start) ---
                crop_y_start_label = QLabel("Crop Top:")
                self.crop_y_start_slider = QSlider(Qt.Horizontal)
                self.crop_y_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_y_start_slider.setValue(self.crop_slider_min)
                self.crop_y_start_slider.setToolTip("Adjust the top edge of the crop area (enabled after drawing).")
                self.crop_y_start_value_label = create_value_label(0.00)
                self.crop_y_start_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_y_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_y_start_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_y_start_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
                crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
                crop_slider_layout.addWidget(self.crop_y_start_value_label, 1, 2)

                # --- Bottom Crop Slider (Y End) ---
                crop_y_end_label = QLabel("Crop Bottom:")
                self.crop_y_end_slider = QSlider(Qt.Horizontal)
                self.crop_y_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
                self.crop_y_end_slider.setValue(self.crop_slider_max)
                self.crop_y_end_slider.setToolTip("Adjust the bottom edge of the crop area (enabled after drawing).")
                self.crop_y_end_value_label = create_value_label(100.00)
                self.crop_y_end_slider.valueChanged.connect(
                    lambda val, lbl=self.crop_y_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
                )
                self.crop_y_end_slider.valueChanged.connect(self._update_crop_from_sliders)
                self.crop_y_end_slider.setEnabled(False) # Initially disabled
                crop_slider_layout.addWidget(crop_y_end_label, 1, 3)
                crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 4)
                crop_slider_layout.addWidget(self.crop_y_end_value_label, 1, 5)

                # Make sliders expand horizontally
                crop_slider_layout.setColumnStretch(1, 1)
                crop_slider_layout.setColumnStretch(4, 1)
                cropping_layout.addLayout(crop_slider_layout)


                layout.addWidget(cropping_params_group)
                # --- END: Modified Cropping Group ---

                layout.addStretch()
                return tab
            
            def _update_rotation_label(self, value):
                """Updates the rotation label text."""
                if hasattr(self, 'orientation_label'):
                    orientation = value / 20.0
                    self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

            def _update_skew_label(self, value):
                """Updates the skew label text."""
                if hasattr(self, 'taper_skew_label'):
                    taper_value = value / 100.0
                    self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

            def _update_crop_from_sliders(self):
                """Updates crop rectangle based on slider percentages and ensures labels update."""
                if not self.image or self.image.isNull():
                    return
            
                # Read integer slider values
                try:
                    x_start_val = self.crop_x_start_slider.value()
                    x_end_val = self.crop_x_end_slider.value()
                    y_start_val = self.crop_y_start_slider.value()
                    y_end_val = self.crop_y_end_slider.value()
                except AttributeError:
                    print("Error: Crop sliders not fully initialized.")
                    return
            
                # Convert to percentage
                x_start_perc = x_start_val / self.crop_slider_precision_factor
                x_end_perc = x_end_val / self.crop_slider_precision_factor
                y_start_perc = y_start_val / self.crop_slider_precision_factor
                y_end_perc = y_end_val / self.crop_slider_precision_factor
            
                # Validation: Ensure Start <= End
                x_start = min(x_start_perc, x_end_perc)
                x_end = max(x_start_perc, x_end_perc)
                y_start = min(y_start_perc, y_end_perc)
                y_end = max(y_start_perc, y_end_perc)
            
                # Convert corrected percentages back to integer slider values
                x_start_val_corr = int(round(x_start * self.crop_slider_precision_factor))
                x_end_val_corr = int(round(x_end * self.crop_slider_precision_factor))
                y_start_val_corr = int(round(y_start * self.crop_slider_precision_factor))
                y_end_val_corr = int(round(y_end * self.crop_slider_precision_factor))
            
                # Reflect potentially swapped values back in sliders (only if different)
                # Use a flag to track if signals were blocked to avoid redundant label updates later
                signals_were_blocked = False
                sliders_to_check = [
                    (self.crop_x_start_slider, x_start_val_corr), (self.crop_x_end_slider, x_end_val_corr),
                    (self.crop_y_start_slider, y_start_val_corr), (self.crop_y_end_slider, y_end_val_corr)
                ]
                for slider, correct_value in sliders_to_check:
                    if slider.value() != correct_value:
                        slider.blockSignals(True)
                        slider.setValue(correct_value)
                        slider.blockSignals(False)
                        signals_were_blocked = True # Record that we manually set a value
            
                # --- Explicitly Update Labels AFTER potential corrections ---
                # This ensures labels always reflect the validated slider state, even if
                # the correction logic blocked the initial valueChanged signal for the label lambda.
                # Use the corrected integer values derived above.
                try:
                    self.crop_x_start_value_label.setText(f"{x_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_x_end_value_label.setText(f"{x_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_start_value_label.setText(f"{y_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_end_value_label.setText(f"{y_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                except AttributeError:
                    print("Error: Crop value labels not fully initialized for explicit update.")
                    # Handle case where labels might not exist yet if called too early
                    pass
                # --- End Explicit Label Update ---
            
            
                # Deactivate drawing mode if sliders are used
                if self.crop_rectangle_mode:
                    self.cancel_rectangle_crop_mode()
            
                # Calculate Image Coordinates using corrected percentages
                img_w = float(self.image.width())
                img_h = float(self.image.height())
                if img_w <= 0 or img_h <= 0: return
            
                img_x = img_w * (x_start / 100.0)
                img_y = img_h * (y_start / 100.0)
                img_crop_w = img_w * ((x_end - x_start) / 100.0)
                img_crop_h = img_h * ((y_end - y_start) / 100.0)
            
                # Store the calculated *image* coordinates
                self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))
            
                # Calculate View Coordinates for Preview
                label_w = float(self.live_view_label.width())
                label_h = float(self.live_view_label.height())
                if label_w <= 0 or label_h <= 0: return
            
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                if scale_factor <= 1e-9: scale_factor = 1
            
                display_img_w = img_w * scale_factor
                display_img_h = img_h * scale_factor
                display_offset_x = (label_w - display_img_w) / 2.0
                display_offset_y = (label_h - display_img_h) / 2.0
            
                view_x_start = display_offset_x + img_x * scale_factor
                view_y_start = display_offset_y + img_y * scale_factor
                view_x_end = view_x_start + img_crop_w * scale_factor
                view_y_end = view_y_start + img_crop_h * scale_factor
            
                # Update the visual preview rectangle
                self.live_view_label.crop_rect_final_view = QRectF(
                    QPointF(view_x_start, view_y_start),
                    QPointF(view_x_end, view_y_end)
                ).normalized()
                self.live_view_label.drawing_crop_rect = False
            
                self.update_live_view()
            
            def toggle_rectangle_crop_mode(self, checked):
                """Toggles the rectangle crop drawing mode based on button state."""
                if checked:
                    self.enable_rectangle_crop_mode()
                else:
                    self.cancel_rectangle_crop_mode()

            def enable_rectangle_crop_mode(self):
                """Activates the rectangle drawing mode for cropping."""
                if self.crop_rectangle_mode: return # Already active

                # Deactivate other interactive modes if necessary (e.g., marker placement)
                self.marker_mode = None
                # self.live_view_label.setCursor(Qt.ArrowCursor) # Set cursor below

                self.crop_rectangle_mode = True
                self.crop_rectangle_coords = None # Clear previous coords
                self.live_view_label.clear_crop_preview() # Clear visual preview

                # Update button state (ensure it's checked)
                self.draw_crop_rect_button.setChecked(True)

                # Setup LiveViewLabel for drawing
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.mousePressEvent = self.start_crop_rectangle
                self.live_view_label.mouseMoveEvent = self.update_crop_rectangle_preview
                self.live_view_label.mouseReleaseEvent = self.finalize_crop_rectangle
                # print("Rectangle Crop Mode Enabled") # Debug

            def cancel_rectangle_crop_mode(self):
                """Deactivates the rectangle drawing mode."""
                if not self.crop_rectangle_mode: return # Already inactive

                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None # Clear temp start point
                # Keep self.crop_rectangle_coords if user might still apply it
                # Keep self.live_view_label.crop_rect_final_view for visual feedback

                # Update button state (ensure it's unchecked)
                if hasattr(self, 'draw_crop_rect_button'): # Check if button exists
                     self.draw_crop_rect_button.setChecked(False)

                # Reset LiveViewLabel
                self.live_view_label.setCursor(Qt.ArrowCursor)

                # --- CORRECTED RESET ---
                # Reset mouse handlers by setting them to None, allowing default behavior
                self.live_view_label.mousePressEvent = None
                self.live_view_label.mouseMoveEvent = None
                self.live_view_label.mouseReleaseEvent = None
                # --- END CORRECTION ---


                # Optionally clear the visual preview immediately on cancel
                # self.live_view_label.clear_crop_preview()
                # self.crop_rectangle_coords = None # Also clear coords if cancel means discard

                # print("Rectangle Crop Mode Disabled") # Debug
                self.update_live_view() # Refresh display (might clear preview if cleared above)

            def start_crop_rectangle(self, event):
                """Handles mouse press to start drawing the crop rectangle."""
                if not self.crop_rectangle_mode:
                    # If not in crop mode, pass event to the base class or intended handler
                    super(LiveViewLabel, self.live_view_label.__class__).mousePressEvent(self.live_view_label, event)
                    return

                if event.button() == Qt.LeftButton:
                    # Get start point in *view* coordinates (unzoomed)
                    self.crop_rect_start_view = self.live_view_label.transform_point(event.pos())
                    # Tell LiveViewLabel to start drawing the preview
                    self.live_view_label.start_crop_preview(self.crop_rect_start_view)

            def update_crop_rectangle_preview(self, event):
                """Handles mouse move to update the crop rectangle preview."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                     # If not in crop mode or not dragging, pass event along
                     super(LiveViewLabel, self.live_view_label.__class__).mouseMoveEvent(self.live_view_label, event)
                     return

                # Check if left button is held down (QApplication.mouseButtons() might be needed)
                if event.buttons() & Qt.LeftButton:
                    current_point_view = self.live_view_label.transform_point(event.pos())
                    # Tell LiveViewLabel to update the preview
                    self.live_view_label.update_crop_preview(self.crop_rect_start_view, current_point_view)

            def finalize_crop_rectangle(self, event):
                """Handles mouse release to finalize the crop rectangle and updates sliders/labels."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                    # If not in crop mode or drag never started, pass event along
                    # Use super() with the correct class hierarchy if LiveViewLabel inherits directly
                    # If LiveViewLabel is just an instance variable, this super() call is incorrect.
                    # Assuming LiveViewLabel is a custom class inheriting QLabel:
                    # super(LiveViewLabel, self.live_view_label).mouseReleaseEvent(event)
                    # If LiveViewLabel is just a QLabel instance, default behaviour is likely sufficient:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event) # Call base QLabel handler if needed
                    return

                if event.button() == Qt.LeftButton:
                    end_point_view = self.live_view_label.transform_point(event.pos())
                    start_point_view = self.crop_rect_start_view

                    # Finalize the visual preview in LiveViewLabel
                    self.live_view_label.finalize_crop_preview(start_point_view, end_point_view)

                    try:
                        # --- Coordinate Conversion (same as before) ---
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w = float(self.image.width())
                        img_h = float(self.image.height())
                        label_w = float(self.live_view_label.width())
                        label_h = float(self.live_view_label.height())

                        if img_w <= 0 or img_h <= 0 or label_w <= 0 or label_h <= 0:
                            raise ValueError("Invalid image or label dimensions.")

                        scale_factor = min(label_w / img_w, label_h / img_h)
                        if scale_factor <= 1e-9: raise ValueError("Scale factor too small.")

                        display_img_w = img_w * scale_factor
                        display_img_h = img_h * scale_factor
                        display_offset_x = (label_w - display_img_w) / 2.0
                        display_offset_y = (label_h - display_img_h) / 2.0

                        start_x_img = (start_point_view.x() - display_offset_x) / scale_factor
                        start_y_img = (start_point_view.y() - display_offset_y) / scale_factor
                        end_x_img = (end_point_view.x() - display_offset_x) / scale_factor
                        end_y_img = (end_point_view.y() - display_offset_y) / scale_factor

                        img_x = min(start_x_img, end_x_img)
                        img_y = min(start_y_img, end_y_img)
                        img_crop_w = abs(end_x_img - start_x_img)
                        img_crop_h = abs(end_y_img - start_y_img)
                        # --- End Coordinate Conversion ---

                        if img_crop_w < 1 or img_crop_h < 1:
                             # Handle case where drawn rectangle is too small
                             self.crop_rectangle_coords = None
                             self.live_view_label.clear_crop_preview()
                             # Keep sliders disabled
                             self.crop_x_start_slider.setEnabled(False)
                             self.crop_x_end_slider.setEnabled(False)
                             self.crop_y_start_slider.setEnabled(False)
                             self.crop_y_end_slider.setEnabled(False)
                        else:
                            # Store valid image coordinates
                            self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))

                            # Calculate percentages from valid image coordinates
                            x_start_perc = (img_x / img_w) * 100.0 if img_w > 0 else 0
                            y_start_perc = (img_y / img_h) * 100.0 if img_h > 0 else 0
                            x_end_perc = ((img_x + img_crop_w) / img_w) * 100.0 if img_w > 0 else 100
                            y_end_perc = ((img_y + img_crop_h) / img_h) * 100.0 if img_h > 0 else 100

                            # Convert percentages to integer slider values
                            x_start_val = int(round(x_start_perc * self.crop_slider_precision_factor))
                            y_start_val = int(round(y_start_perc * self.crop_slider_precision_factor))
                            x_end_val = int(round(x_end_perc * self.crop_slider_precision_factor))
                            y_end_val = int(round(y_end_perc * self.crop_slider_precision_factor))

                            # --- Update Sliders (Block signals) ---
                            self.crop_x_start_slider.blockSignals(True); self.crop_x_start_slider.setValue(x_start_val); self.crop_x_start_slider.blockSignals(False)
                            self.crop_x_end_slider.blockSignals(True); self.crop_x_end_slider.setValue(x_end_val); self.crop_x_end_slider.blockSignals(False)
                            self.crop_y_start_slider.blockSignals(True); self.crop_y_start_slider.setValue(y_start_val); self.crop_y_start_slider.blockSignals(False)
                            self.crop_y_end_slider.blockSignals(True); self.crop_y_end_slider.setValue(y_end_val); self.crop_y_end_slider.blockSignals(False)

                            # --- Explicitly Update Labels AFTER setting sliders ---
                            self.crop_x_start_value_label.setText(f"{x_start_perc:.2f}%")
                            self.crop_x_end_value_label.setText(f"{x_end_perc:.2f}%")
                            self.crop_y_start_value_label.setText(f"{y_start_perc:.2f}%")
                            self.crop_y_end_value_label.setText(f"{y_end_perc:.2f}%")
                            # --- End Explicit Label Update ---

                            # --- Enable the sliders ---
                            self.crop_x_start_slider.setEnabled(True)
                            self.crop_x_end_slider.setEnabled(True)
                            self.crop_y_start_slider.setEnabled(True)
                            self.crop_y_end_slider.setEnabled(True)

                    except Exception as e:
                         # Handle errors during coordinate calculation or UI update
                         QMessageBox.warning(self, "Coordinate Error", f"Could not calculate/update crop: {e}")
                         traceback.print_exc()
                         self.crop_rectangle_coords = None
                         self.live_view_label.clear_crop_preview()
                         # Ensure sliders remain disabled on error
                         self.crop_x_start_slider.setEnabled(False)
                         self.crop_x_end_slider.setEnabled(False)
                         self.crop_y_start_slider.setEnabled(False)
                         self.crop_y_end_slider.setEnabled(False)

                    # Reset the temporary start point used for drawing
                    self.crop_rect_start_view = None
            
            def create_white_space_tab(self):
                tab = QWidget()
                layout = QVBoxLayout(tab)
                layout.setSpacing(15)

                # --- Padding Group ---
                padding_params_group = QGroupBox("Add White Space (Padding)")
                padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                padding_layout = QGridLayout(padding_params_group)
                padding_layout.setSpacing(8)


                # Input fields with validation (optional but good)
                from PyQt5.QtGui import QIntValidator
                int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

                # Left Padding
                left_padding_label = QLabel("Left Padding (px):")
                self.left_padding_input = QLineEdit("100") # Default
                self.left_padding_input.setValidator(int_validator)
                self.left_padding_input.setToolTip("Pixels to add to the left.")
                padding_layout.addWidget(left_padding_label, 0, 0)
                padding_layout.addWidget(self.left_padding_input, 0, 1)

                # Right Padding
                right_padding_label = QLabel("Right Padding (px):")
                self.right_padding_input = QLineEdit("100") # Default
                self.right_padding_input.setValidator(int_validator)
                self.right_padding_input.setToolTip("Pixels to add to the right.")
                padding_layout.addWidget(right_padding_label, 0, 2)
                padding_layout.addWidget(self.right_padding_input, 0, 3)

                # Top Padding
                top_padding_label = QLabel("Top Padding (px):")
                self.top_padding_input = QLineEdit("100") # Default
                self.top_padding_input.setValidator(int_validator)
                self.top_padding_input.setToolTip("Pixels to add to the top.")
                padding_layout.addWidget(top_padding_label, 1, 0)
                padding_layout.addWidget(self.top_padding_input, 1, 1)

                # Bottom Padding
                bottom_padding_label = QLabel("Bottom Padding (px):")
                self.bottom_padding_input = QLineEdit("0") # Default
                self.bottom_padding_input.setValidator(int_validator)
                self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
                padding_layout.addWidget(bottom_padding_label, 1, 2)
                padding_layout.addWidget(self.bottom_padding_input, 1, 3)

                layout.addWidget(padding_params_group)

                # --- Buttons Layout ---
                button_layout = QHBoxLayout()
                self.recommend_button = QPushButton("Set Recommended Values")
                self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
                self.recommend_button.clicked.connect(self.recommended_values)

                self.clear_padding_button = QPushButton("Clear Values")
                self.clear_padding_button.setToolTip("Set all padding values to 0.")
                self.clear_padding_button.clicked.connect(self.clear_padding_values)

                self.finalize_button = QPushButton("Apply Padding")
                self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
                self.finalize_button.clicked.connect(self.finalize_image)

                button_layout.addWidget(self.recommend_button)
                button_layout.addWidget(self.clear_padding_button)
                button_layout.addStretch()
                button_layout.addWidget(self.finalize_button)

                layout.addLayout(button_layout)
                layout.addStretch() # Push content up

                return tab
            def clear_padding_values(self):
                self.bottom_padding_input.setText("0")
                self.top_padding_input.setText("0")
                self.right_padding_input.setText("0")
                self.left_padding_input.setText("0")
                
            def recommended_values(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to set recommended padding.")
                    return

                # This function should ONLY set the text in the QLineEdit fields for padding.
                # It should NOT directly change slider ranges or _marker_shift_added values.
                # The marker offsets are independent of padding until "Apply Padding" is clicked.

                try:
                    native_width = self.image.width()
                    native_height = self.image.height()

                    if native_width <= 0 or native_height <= 0:
                        QMessageBox.warning(self, "Error", "Current image has invalid dimensions.")
                        return

                    # Set recommended padding values in the QLineEdit fields
                    self.left_padding_input.setText(str(int(native_width * 0.1)))
                    self.right_padding_input.setText(str(int(native_width * 0.1)))
                    self.top_padding_input.setText(str(int(native_height * 0.15)))
                    self.bottom_padding_input.setText(str(int(0))) # Or another sensible default

                    # DO NOT update slider ranges or _marker_shift_added here.
                    # Slider ranges are updated by _update_marker_slider_ranges when self.image changes.
                    # _marker_shift_added are updated by sliders or by crop/pad operations.
                    
                    # No need to call self.update_live_view() here either, as only
                    # text fields for a future operation were changed.
                    QMessageBox.information(self, "Recommended Values Set", 
                                            "Recommended padding values have been entered into the fields.\n"
                                            "Click 'Apply Padding' to use them.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not set recommended values: {e}")
                
                
                
            # In class CombinedSDSApp:

            def create_markers_tab(self):
                """Create the Markers tab with a more compact and organized layout."""
                tab = QWidget()
                # Main vertical layout for the tab
                main_layout = QVBoxLayout(tab)
                main_layout.setSpacing(10) # Consistent spacing between major sections

                # --- Top Row: Presets and Top Labels ---
                top_row_layout = QHBoxLayout()

                # Group Box for Presets (Left/Right Markers)
                presets_group = QGroupBox("Left/Right Marker Presets")
                presets_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                presets_layout = QGridLayout(presets_group)
                presets_layout.setSpacing(5) 

                presets_layout.addWidget(QLabel("Preset:"), 0, 0)
                self.combo_box = QComboBox(self)
                
                # --- MODIFIED SECTION for ComboBox Population ---
                # self.presets_data should be populated by self.load_config() before this method is called
                if hasattr(self, 'presets_data') and self.presets_data:
                    sorted_preset_names = sorted(self.presets_data.keys())
                    self.combo_box.addItems(sorted_preset_names)
                else:
                    # Fallback if presets_data is empty or not found (should ideally not happen if load_config is robust)
                    print("Warning: self.presets_data not populated or empty when creating markers tab.")
                self.combo_box.addItem("Custom") # "Custom" is always an option
                
                # Attempt to set a sensible default after items are added.
                # Default selection logic moved to load_config's UI update section for robustness.
                # Here, we just ensure on_combobox_changed is connected.
                # --- END MODIFIED SECTION ---

                self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
                presets_layout.addWidget(self.combo_box, 0, 1)

                self.marker_values_textbox = QLineEdit(self)
                self.marker_values_textbox.setPlaceholderText("Custom L/R values (comma-sep)")
                self.marker_values_textbox.setEnabled(False) # Initial state, on_combobox_changed will manage
                presets_layout.addWidget(self.marker_values_textbox, 1, 0, 1, 2)

                self.rename_input = QLineEdit(self)
                self.rename_input.setPlaceholderText("New name for Custom preset")
                self.rename_input.setEnabled(False) # Initial state
                presets_layout.addWidget(self.rename_input, 2, 0, 1, 2)

                preset_buttons_layout = QHBoxLayout()
                preset_buttons_layout.setContentsMargins(0, 0, 0, 0)
                self.save_button = QPushButton("Save Preset", self)
                self.save_button.setToolTip("Saves the current L/R, Top, Custom Markers/Shapes to the selected/new preset name.")
                self.save_button.clicked.connect(self.save_config) # save_config is the preset saver now
                self.remove_config_button = QPushButton("Remove Preset", self)
                self.remove_config_button.clicked.connect(self.remove_config)
                preset_buttons_layout.addWidget(self.save_button)
                preset_buttons_layout.addWidget(self.remove_config_button)
                preset_buttons_layout.addStretch()
                presets_layout.addLayout(preset_buttons_layout, 3, 0, 1, 2)

                presets_layout.setColumnStretch(1, 1)
                top_row_layout.addWidget(presets_group, 1)

                # Group Box for Top Labels
                top_labels_group = QGroupBox("Top Marker Labels")
                top_labels_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                top_labels_layout = QVBoxLayout(top_labels_group)

                self.top_marker_input = QTextEdit(self)
                # self.top_label should be initialized or loaded by load_config -> on_combobox_changed
                current_top_label_text = ", ".join(map(str, getattr(self, 'top_label', [])))
                self.top_marker_input.setText(current_top_label_text)
                self.top_marker_input.setMinimumHeight(40)
                self.top_marker_input.setMaximumHeight(100)
                self.top_marker_input.setPlaceholderText("Top labels (comma-separated)")
                top_labels_layout.addWidget(self.top_marker_input)

                self.update_top_labels_button = QPushButton("Update All L/R/Top Labels")
                self.update_top_labels_button.setToolTip("Apply values from text boxes to current markers on the image.")
                self.update_top_labels_button.clicked.connect(self.update_all_labels)
                top_labels_layout.addWidget(self.update_top_labels_button)

                top_row_layout.addWidget(top_labels_group, 1)
                main_layout.addLayout(top_row_layout)

                # --- Middle Section: Marker Placement and Offsets ---
                # (Rest of this section remains the same: placement_group, sliders, etc.)
                # ...
                placement_group = QGroupBox("Marker Placement and Offsets")
                placement_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                placement_layout = QGridLayout(placement_group)
                placement_layout.setColumnStretch(0, 0) 
                placement_layout.setColumnStretch(1, 0) 
                placement_layout.setColumnStretch(2, 0) 
                placement_layout.setColumnStretch(3, 0) 
                placement_layout.setColumnStretch(4, 1) 
                placement_layout.setColumnStretch(5, 0) 

                if not hasattr(self, 'left_slider_range'): self.left_slider_range = [-100, 1000]
                if not hasattr(self, 'right_slider_range'): self.right_slider_range = [-100, 1000]
                if not hasattr(self, 'top_slider_range'): self.top_slider_range = [-100, 1000]
                if not hasattr(self, 'left_marker_shift_added'): self.left_marker_shift_added = 0
                if not hasattr(self, 'right_marker_shift_added'): self.right_marker_shift_added = 0
                if not hasattr(self, 'top_marker_shift_added'): self.top_marker_shift_added = 0

                left_marker_button = QPushButton("Place Left"); left_marker_button.setToolTip("Ctrl+Shift+L")
                left_marker_button.clicked.connect(self.enable_left_marker_mode)
                remove_left_button = QPushButton("Remove Last")
                remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
                reset_left_button = QPushButton("Reset All"); reset_left_button.setToolTip("Reset All Left Markers")
                reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
                self.left_padding_slider = QSlider(Qt.Horizontal)
                self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
                self.left_padding_slider.setValue(self.left_marker_shift_added)
                self.left_padding_slider.valueChanged.connect(self.update_left_padding)
                duplicate_left_button = QPushButton("Copy →") 
                duplicate_left_button.setToolTip("Copy Right Markers & Offset to Left")
                duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
                placement_layout.addWidget(left_marker_button, 0, 0)
                placement_layout.addWidget(remove_left_button, 0, 1)
                placement_layout.addWidget(reset_left_button, 0, 2)
                placement_layout.addWidget(QLabel("Offset Left:"), 0, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.left_padding_slider, 0, 4)
                placement_layout.addWidget(duplicate_left_button, 0, 5)

                right_marker_button = QPushButton("Place Right"); right_marker_button.setToolTip("Ctrl+Shift+R")
                right_marker_button.clicked.connect(self.enable_right_marker_mode)
                remove_right_button = QPushButton("Remove Last")
                remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
                reset_right_button = QPushButton("Reset All"); reset_right_button.setToolTip("Reset All Right Markers")
                reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
                self.right_padding_slider = QSlider(Qt.Horizontal)
                self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
                self.right_padding_slider.setValue(self.right_marker_shift_added)
                self.right_padding_slider.valueChanged.connect(self.update_right_padding)
                duplicate_right_button = QPushButton("← Copy") 
                duplicate_right_button.setToolTip("Copy Left Markers & Offset to Right")
                duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
                placement_layout.addWidget(right_marker_button, 1, 0)
                placement_layout.addWidget(remove_right_button, 1, 1)
                placement_layout.addWidget(reset_right_button, 1, 2)
                placement_layout.addWidget(QLabel("Offset Right:"), 1, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.right_padding_slider, 1, 4)
                placement_layout.addWidget(duplicate_right_button, 1, 5)

                top_marker_button = QPushButton("Place Top"); top_marker_button.setToolTip("Ctrl+Shift+T")
                top_marker_button.clicked.connect(self.enable_top_marker_mode)
                remove_top_button = QPushButton("Remove Last")
                remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
                reset_top_button = QPushButton("Reset All"); reset_top_button.setToolTip("Reset All Top Markers")
                reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
                self.top_padding_slider = QSlider(Qt.Horizontal)
                self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
                self.top_padding_slider.setValue(self.top_marker_shift_added)
                self.top_padding_slider.valueChanged.connect(self.update_top_padding)
                placement_layout.addWidget(top_marker_button, 2, 0)
                placement_layout.addWidget(remove_top_button, 2, 1)
                placement_layout.addWidget(reset_top_button, 2, 2)
                placement_layout.addWidget(QLabel("Offset Top:"), 2, 3, Qt.AlignRight | Qt.AlignVCenter)
                placement_layout.addWidget(self.top_padding_slider, 2, 4)
                
                main_layout.addWidget(placement_group)


                # --- Bottom Section: Custom Markers / Shapes / Grid ---
                # (Rest of this section remains the same: custom_group, etc.)
                # ...
                custom_group = QGroupBox("Custom Markers, Shapes, and Grid")
                custom_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                custom_layout = QGridLayout(custom_group)
                custom_layout.setColumnStretch(1, 1) 
                custom_layout.setColumnStretch(4, 1) 

                self.custom_marker_button = QPushButton("Place Custom", self)
                self.custom_marker_button.setToolTip("Click to activate, then click on image to place text/arrow")
                self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
                self.custom_marker_text_entry = QLineEdit(self)
                self.custom_marker_text_entry.setPlaceholderText("Custom text or use arrows →")

                marker_buttons_layout = QHBoxLayout()
                marker_buttons_layout.setContentsMargins(0, 0, 0, 0); marker_buttons_layout.setSpacing(2)
                arrow_size = 25
                self.custom_marker_button_left_arrow = QPushButton("←"); self.custom_marker_button_left_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_left_arrow.setToolTip("Ctrl+Left")
                self.custom_marker_button_right_arrow = QPushButton("→"); self.custom_marker_button_right_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_right_arrow.setToolTip("Ctrl+Right")
                self.custom_marker_button_top_arrow = QPushButton("↑"); self.custom_marker_button_top_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_top_arrow.setToolTip("Ctrl+Up")
                self.custom_marker_button_bottom_arrow = QPushButton("↓"); self.custom_marker_button_bottom_arrow.setFixedSize(arrow_size, arrow_size); self.custom_marker_button_bottom_arrow.setToolTip("Ctrl+Down")
                marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow); marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
                marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow); marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
                marker_buttons_layout.addStretch()
                self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))

                custom_manage_layout = QHBoxLayout(); custom_manage_layout.setContentsMargins(0,0,0,0); custom_manage_layout.setSpacing(4)
                self.remove_custom_marker_button = QPushButton("Remove Last"); self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
                self.reset_custom_marker_button = QPushButton("Reset All"); self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
                self.modify_custom_marker_button = QPushButton("Modify All..."); self.modify_custom_marker_button.setToolTip("Modify/Delete Custom Markers & Shapes")
                self.modify_custom_marker_button.clicked.connect(self.open_modify_markers_dialog)
                custom_manage_layout.addWidget(self.remove_custom_marker_button); custom_manage_layout.addWidget(self.reset_custom_marker_button); custom_manage_layout.addWidget(self.modify_custom_marker_button)

                custom_layout.addWidget(self.custom_marker_button, 0, 0)
                custom_layout.addWidget(self.custom_marker_text_entry, 0, 1)
                custom_layout.addLayout(marker_buttons_layout, 0, 2)
                custom_layout.addLayout(custom_manage_layout, 0, 3, 1, 3)

                custom_style_layout = QHBoxLayout(); custom_style_layout.setContentsMargins(0,0,0,0); custom_style_layout.setSpacing(5)
                self.custom_font_type_label = QLabel("Font:")
                self.custom_font_type_dropdown = QFontComboBox()
                initial_font = QFont("Arial"); self.custom_font_type_dropdown.setCurrentFont(initial_font)
                self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
                self.custom_font_size_label = QLabel("Size:")
                self.custom_font_size_spinbox = QSpinBox()
                self.custom_font_size_spinbox.setRange(1, 150); self.custom_font_size_spinbox.setValue(12)
                self.custom_marker_color_button = QPushButton("Color")
                self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
                if not hasattr(self, 'custom_marker_color'): self.custom_marker_color = QColor(0,0,0)
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                custom_style_layout.addWidget(self.custom_font_type_label); custom_style_layout.addWidget(self.custom_font_type_dropdown)
                custom_style_layout.addWidget(self.custom_font_size_label); custom_style_layout.addWidget(self.custom_font_size_spinbox)
                custom_style_layout.addWidget(self.custom_marker_color_button)
                custom_style_layout.addStretch()

                shape_button_layout = QHBoxLayout(); shape_button_layout.setContentsMargins(0,0,0,0); shape_button_layout.setSpacing(2)
                shape_size = 25
                self.draw_line_button = QPushButton("L"); self.draw_line_button.setToolTip("Draw Line"); self.draw_line_button.setFixedSize(shape_size, shape_size); self.draw_line_button.clicked.connect(self.enable_line_drawing_mode)
                self.draw_rect_button = QPushButton("R"); self.draw_rect_button.setToolTip("Draw Rectangle"); self.draw_rect_button.setFixedSize(shape_size, shape_size); self.draw_rect_button.clicked.connect(self.enable_rectangle_drawing_mode)
                self.remove_shape_button = QPushButton("X"); self.remove_shape_button.setToolTip("Remove Last Shape"); self.remove_shape_button.setFixedSize(shape_size, shape_size); self.remove_shape_button.clicked.connect(self.remove_last_custom_shape)
                shape_button_layout.addWidget(QLabel("Shapes:"))
                shape_button_layout.addWidget(self.draw_line_button); shape_button_layout.addWidget(self.draw_rect_button); shape_button_layout.addWidget(self.remove_shape_button)
                shape_button_layout.addStretch()

                grid_layout = QHBoxLayout(); grid_layout.setContentsMargins(0,0,0,0); grid_layout.setSpacing(5)
                self.show_grid_checkbox_x = QCheckBox("Snap X"); self.show_grid_checkbox_x.setToolTip("Snap horizontally. Ctrl+Shift+X toggles X and Ctrl+Shift+G for both X and Y.")
                self.show_grid_checkbox_x.stateChanged.connect(self.update_live_view)
                self.show_grid_checkbox_y = QCheckBox("Snap Y"); self.show_grid_checkbox_y.setToolTip("Snap vertically. Ctrl+Shift+Y toggles Y and Ctrl+Shift+G for both X and Y.")
                self.show_grid_checkbox_y.stateChanged.connect(self.update_live_view)
                self.grid_size_input = QSpinBox(); self.grid_size_input.setRange(5, 100); self.grid_size_input.setValue(20); self.grid_size_input.setPrefix("Grid (px): ")
                self.grid_size_input.valueChanged.connect(self.update_live_view)
                grid_layout.addWidget(self.show_grid_checkbox_x); grid_layout.addWidget(self.show_grid_checkbox_y); grid_layout.addWidget(self.grid_size_input)
                grid_layout.addStretch()

                custom_layout.addLayout(custom_style_layout, 1, 0, 1, 2)
                custom_layout.addLayout(shape_button_layout, 1, 2, 1, 2)
                custom_layout.addLayout(grid_layout, 1, 4, 1, 2)

                main_layout.addWidget(custom_group)
                main_layout.addStretch()
                return tab

            def open_modify_markers_dialog(self):
                self._backup_custom_markers_before_modify_dialog = [list(m) for m in self.custom_markers]
                self._backup_custom_shapes_before_modify_dialog = [dict(s) for s in self.custom_shapes] # Backup shapes
                
                if not hasattr(self, "custom_markers") or not isinstance(self.custom_markers, list):
                    self.custom_markers = []
                if not hasattr(self, "custom_shapes") or not isinstance(self.custom_shapes, list):
                    self.custom_shapes = []
            
                if not self.custom_markers and not self.custom_shapes: # Only check markers for now if shapes aren't globally adjusted
                    QMessageBox.information(self, "No Markers", "There are no custom markers to modify with global adjustments.")
                    return
            
                # Store a backup of the current custom_markers before opening the dialog
                # This allows reverting if the dialog is cancelled.
                self._backup_custom_markers_before_modify_dialog = [list(m) for m in self.custom_markers]
            
                dialog = ModifyMarkersDialog(
                    [list(m) for m in self.custom_markers],
                    [dict(s) for s in self.custom_shapes],
                    self
                )
            
                # --- CONNECT the dialog's signal to a handler in CombinedSDSApp ---
                dialog.global_markers_adjusted.connect(self.handle_live_marker_adjustment_preview)
                dialog.shapes_adjusted_preview.connect(self.handle_live_shape_adjustment_preview)
            
                if dialog.exec_() == QDialog.Accepted:
                    modified_markers_tuples, modified_shapes_dicts = dialog.get_modified_markers_and_shapes()
                    modified_markers_lists = [list(m) for m in modified_markers_tuples]
            
                    markers_changed = (modified_markers_lists != self._backup_custom_markers_before_modify_dialog) # Compare with backup
                    shapes_changed = (modified_shapes_dicts != self.custom_shapes) # self.custom_shapes wasn't backed up for live preview here
            
                    if markers_changed or shapes_changed: # Only save state if actual accepted changes occurred
                        self.save_state()
                        self.custom_markers = modified_markers_lists
                        self.custom_shapes = modified_shapes_dicts # Shapes are only updated on OK
                        self.is_modified = True
                        self.update_live_view() # Final update with accepted changes
                    else:
                        # If no changes were made (e.g., OK clicked without modifications, or changes were undone in dialog)
                        # ensure the live preview is reverted to the state before dialog.
                        self.custom_markers = self._backup_custom_markers_before_modify_dialog
                        self.update_live_view()
            
            
                else: # Dialog was cancelled or closed
                    # Revert self.custom_markers to the state before the dialog was opened
                    self.custom_markers = self._backup_custom_markers_before_modify_dialog
                    self.update_live_view() # Refresh to show the original state
            
                # Clean up the backup
                del self._backup_custom_markers_before_modify_dialog
                del self._backup_custom_shapes_before_modify_dialog
                # Disconnect the signal to avoid issues if the dialog object is somehow reused (though usually not)
                try:
                    dialog.global_markers_adjusted.disconnect(self.handle_live_marker_adjustment_preview)
                    dialog.shapes_adjusted_preview.disconnect(self.handle_live_shape_adjustment_preview) # Disconnect new signal
                except TypeError: 
                    pass
            
            
            def handle_live_marker_adjustment_preview(self, temporarily_adjusted_markers):
                """
                Temporarily updates self.custom_markers and refreshes the live_view_label
                to show the effect of global adjustments from ModifyMarkersDialog.
                This does NOT mark the image as permanently modified or save to undo stack.
                """
                # We are directly modifying self.custom_markers here for the preview.
                # The backup in open_modify_markers_dialog handles reverting on cancel.
                self.custom_markers = temporarily_adjusted_markers # Update with the live adjusted list
                self.update_live_view()
                
            def handle_live_shape_adjustment_preview(self, temporarily_adjusted_shapes):
                """
                Temporarily updates self.custom_shapes and refreshes the live_view_label
                to show the effect of shape edits from ModifyMarkersDialog.
                """
                self.custom_shapes = temporarily_adjusted_shapes # Update with the live adjusted list
                self.update_live_view()
                         
            def _serialize_custom_markers(self, custom_markers_list):
                """Converts a list of custom marker tuples (with QColor) to a list of serializable dicts."""
                serialized_list = []
                for marker_tuple in custom_markers_list:
                    try:
                        # Ensure the tuple has the expected 8 elements
                        if len(marker_tuple) == 8:
                            x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_tuple
                            serialized_list.append({
                                "x": x, "y": y, "text": text, "color": qcolor.name(), # Store color name
                                "font_family": font_family, "font_size": font_size,
                                "bold": is_bold, "italic": is_italic
                            })
                        else:
                            # Handle older format or malformed data gracefully if necessary
                            # For now, we'll skip markers not matching the 8-element structure
                            print(f"Warning: Skipping marker during serialization due to unexpected format: {marker_tuple}")
                    except (ValueError, TypeError, IndexError, AttributeError) as e:
                        print(f"Warning: Skipping marker during serialization: {marker_tuple}, Error: {e}")
                return serialized_list

            def _deserialize_custom_markers(self, custom_markers_config_list):
                """Converts a list of serializable dicts back to custom marker tuples (with QColor)."""
                deserialized_list = []
                if not isinstance(custom_markers_config_list, list):
                    print("Warning: custom_markers_config_list is not a list, cannot deserialize.")
                    return []
                    
                for marker_conf in custom_markers_config_list:
                    if not isinstance(marker_conf, dict):
                        print(f"Warning: Skipping non-dictionary item in custom_markers_config_list: {marker_conf}")
                        continue
                    try:
                        color_str = marker_conf.get("color", "#000000") # Default to black if missing
                        qcolor = QColor(color_str)
                        if not qcolor.isValid():
                            print(f"Warning: Invalid color string '{color_str}' for marker, using black.")
                            qcolor = QColor(0,0,0) # Fallback to black

                        deserialized_list.append((
                            float(marker_conf.get("x", 0.0)),
                            float(marker_conf.get("y", 0.0)),
                            str(marker_conf.get("text", "")),
                            qcolor,
                            str(marker_conf.get("font_family", "Arial")),
                            int(marker_conf.get("font_size", 12)),
                            bool(marker_conf.get("bold", False)),
                            bool(marker_conf.get("italic", False))
                        ))
                    except (ValueError, TypeError, KeyError) as e:
                        print(f"Warning: Skipping marker during deserialization: {marker_conf}, Error: {e}")
                return deserialized_list
            
            def add_column(self):
                """Add a new column to the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                if current_text.strip():
                    self.top_marker_input.append("")  # Add a new line for a new column
                else:
                    self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
            
            def remove_column(self):
                """Remove the last column from the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                lines = current_text.split("\n")
                if len(lines) > 1:
                    lines.pop()  # Remove the last line
                    self.top_marker_input.setPlainText("\n".join(lines))
                else:
                    self.top_marker_input.clear()  # Clear the text if only one line exists

            
            def flip_vertical(self):
                self.save_state()
                """Flip the image vertically."""
                if self.image:
                    self.image = self.image.mirrored(vertical=True, horizontal=False)
                    self.image_before_contrast=self.image.copy()
                    self.image_before_padding=self.image.copy()
                    self.image_contrasted=self.image.copy()
                    self.update_live_view()
            
            def flip_horizontal(self):
                self.save_state()
                """Flip the image horizontally."""
                if self.image:
                    self.image = self.image.mirrored(vertical=False, horizontal=True)
                    self.image_before_contrast=self.image.copy()
                    self.image_before_padding=self.image.copy()
                    self.image_contrasted=self.image.copy()
                    self.update_live_view()
            
            def convert_to_black_and_white(self):
                """
                Converts the current self.image to grayscale, preserving alpha channel
                if present by creating an ARGB image where R=G=B=GrayValue.
                Aims for 16-bit grayscale precision when converting from color.
                """
                self.save_state()
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded.")
                    return

                original_format = self.image.format()
                original_has_alpha = self.image.hasAlphaChannel() # Check QImage's report

                # Check if already effectively grayscale (but maybe not the specific format)
                if original_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16, QImage.Format_Mono]:
                    # It's already a standard grayscale format (without explicit alpha in format enum)
                    QMessageBox.information(self, "Info", f"Image is already grayscale (Format: {original_format}).")
                    return

                converted_image = None
                try:
                    np_img = self.qimage_to_numpy(self.image)
                    if np_img is None: raise ValueError("NumPy conversion failed.")

                    print(f"Original NumPy shape: {np_img.shape}, dtype: {np_img.dtype}") # Debug

                    target_dtype = np.uint16 # Prefer 16-bit precision for grayscale conversion
                    target_max_val = 65535.0

                    if np_img.ndim == 3 and np_img.shape[2] == 4: # Color with Alpha (BGRA or RGBA)
                        print("Processing image with Alpha channel.")
                        # Assume input order from qimage_to_numpy is BGRA for ARGB32/RGB32
                        # or RGBA for RGBA8888. cvtColor expects BGR.
                        # Let's explicitly convert the color part assuming BGR input slice
                        color_part = np_img[..., :3] # Slice BGR or RGB
                        alpha_part = np_img[..., 3]  # Extract alpha channel

                        # Convert color to grayscale using cv2 (expects BGR)
                        # If the input was RGBA, this might be slightly off, but often acceptable.
                        # A more robust way might check the original_format if needed.
                        gray_np_8bit = cv2.cvtColor(color_part, cv2.COLOR_BGR2GRAY)

                        # Scale grayscale to target bit depth (16-bit)
                        gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)

                        # Create a new 4-channel array (BGRA format for QImage.Format_ARGB32)
                        # Replicate grayscale channel for B, G, R
                        bgra_target = np.zeros((np_img.shape[0], np_img.shape[1], 4), dtype=target_dtype)
                        bgra_target[..., 0] = gray_np_target # Blue = Gray
                        bgra_target[..., 1] = gray_np_target # Green = Gray
                        bgra_target[..., 2] = gray_np_target # Red = Gray

                        # Handle Alpha channel: ensure it's the correct dtype for combining
                        if alpha_part.dtype != target_dtype:
                             # Scale alpha if necessary (e.g., if source was 16-bit RGBA)
                             if alpha_part.dtype == np.uint16:
                                 alpha_scaled = (alpha_part / 65535.0 * target_max_val).astype(target_dtype)
                             elif alpha_part.dtype == np.uint8:
                                 alpha_scaled = (alpha_part / 255.0 * target_max_val).astype(target_dtype)
                             else: # Fallback: assume it doesn't need scaling or use 8-bit alpha
                                  print(f"Warning: Unexpected alpha dtype {alpha_part.dtype}. Attempting direct use or scaling.")
                                  try: # Try scaling assuming it's 0-max range
                                      alpha_scaled = (alpha_part / np.max(alpha_part) * target_max_val).astype(target_dtype) if np.max(alpha_part) > 0 else np.zeros_like(alpha_part, dtype=target_dtype)
                                  except: # Final fallback
                                       alpha_scaled = (alpha_part.astype(np.float64) / 255.0 * target_max_val).astype(target_dtype) if np.max(alpha_part)>0 else np.zeros_like(alpha_part,dtype=target_dtype) #Assume 8 bit if failsafe
                             bgra_target[..., 3] = alpha_scaled
                        else:
                            bgra_target[..., 3] = alpha_part # Alpha dtype already matches

                        # Convert back to QImage - should become Format_ARGB32 due to 4 channels
                        # Note: numpy_to_qimage needs to handle 16-bit 4-channel input if we want ARGB64
                        # Currently, it scales down 16-bit color to 8-bit ARGB32. Let's adapt that.
                        # --- MODIFICATION NEEDED in numpy_to_qimage for 16-bit ARGB ---
                        # For now, let's convert bgra_target to uint8 for Format_ARGB32 output
                        bgra_target_uint8 = (bgra_target / target_max_val * 255.0).astype(np.uint8)
                        converted_image = self.numpy_to_qimage(bgra_target_uint8)
                        if converted_image.isNull(): raise ValueError("Conversion of BGRA Grayscale to QImage failed.")
                        print(f"Converted to Grayscale + Alpha. QImage format: {converted_image.format()}") # Debug


                    elif np_img.ndim == 3 and np_img.shape[2] == 3: # Color without Alpha (BGR or RGB)
                        print("Processing image without Alpha channel.")
                        # Convert color to grayscale (standard procedure)
                        gray_np_8bit = cv2.cvtColor(np_img[..., :3], cv2.COLOR_BGR2GRAY)
                         # Scale grayscale to target bit depth (16-bit)
                        gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)
                        # Convert back to standard grayscale QImage
                        converted_image = self.numpy_to_qimage(gray_np_target) # Should yield Format_Grayscale16
                        if converted_image.isNull(): raise ValueError("Conversion of BGR Grayscale to QImage failed.")
                        print(f"Converted to Grayscale. QImage format: {converted_image.format()}") # Debug


                    elif np_img.ndim == 2: # Already grayscale (but QImage didn't report standard format)
                         # This case might occur for unusual grayscale formats QImage loads but doesn't map to standard enums.
                         QMessageBox.information(self, "Info", "Image is already grayscale (based on channel analysis).")
                         # Optionally convert to a standard format like Grayscale16
                         if np_img.dtype != target_dtype:
                             gray_np_target = (np_img.astype(np.float64) / np.max(np_img) * target_max_val).astype(target_dtype) if np.max(np_img)>0 else np.zeros_like(np_img,dtype=target_dtype)
                             converted_image = self.numpy_to_qimage(gray_np_target)
                         else:
                             converted_image = self.image.copy() # No conversion needed
                    else:
                        raise ValueError(f"Unsupported NumPy array dimension: {np_img.ndim}")


                except Exception as e:
                     QMessageBox.critical(self, "Conversion Error", f"Could not convert image to grayscale: {e}")
                     traceback.print_exc()
                     return # Keep original image

                if converted_image and not converted_image.isNull():
                     self.image = converted_image
                     # Update backups consistently
                     self.image_before_contrast = self.image.copy()
                     self.image_contrasted = self.image.copy()
                     # Reset padding state if format changes implicitly? Let's assume padding needs re-application.
                     if self.image_padded:
                         self.image_before_padding = None # Invalidate padding backup
                         self.image_padded = False
                     else:
                         self.image_before_padding = self.image.copy() if self.image else None # Ensure image exists before copying

                     # Reset contrast/gamma sliders as appearance changed significantly
                     self.reset_gamma_contrast() # Resets sliders and updates view
                     self._update_status_bar()
                     self.update_live_view() # Ensure view updates even if reset_gamma_contrast fails
                else:
                     # This case should be less likely now with better error handling
                     QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


            def invert_image(self):
                self.save_state()
                if self.image:
                    inverted_image = self.image.copy()
                    inverted_image.invertPixels()
                    self.image = inverted_image
                    self.update_live_view()
                self.image_before_contrast=self.image.copy()
                self.image_before_padding=self.image.copy()
                self.image_contrasted=self.image.copy()

            def keyPressEvent(self, event):
                # --- Escape Key Handling ---
                if event.key() == Qt.Key_Escape:
                    self.live_view_label.mousePressEvent = None # Reset handler
                    self.live_view_label.mouseMoveEvent = None  # Reset handler
                    self.live_view_label.mouseReleaseEvent = None
                    self.live_view_label.zoom_level=1.0
                    self.update_live_view()
                    # 0. NEW: Cancel Auto Lane Quadrilateral Definition
                    if self.live_view_label.mode == 'auto_lane_quad':
                        print("Debug: Escape pressed, cancelling auto lane quad definition.") # Optional Debug
                        self.live_view_label.mode = None
                        self.live_view_label.quad_points = [] # Clear points from label
                        self.live_view_label.selected_point = -1 # Not really used here, but good practice
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        self.live_view_label.mousePressEvent = None # Reset handlers
                        self.live_view_label.mouseMoveEvent = None
                        self.live_view_label.mouseReleaseEvent = None
                        self.update_live_view() # Clear potential quad drawing
                        return # Consume the event
                    # 1. PRIORITIZE CANCELLING CROP MODE if active (Rectangle or Auto Lane Rectangle)
                    if self.crop_rectangle_mode or self.live_view_label.mode == 'auto_lane_rect':
                        print("Debug: Escape pressed, cancelling crop/auto_lane_rect mode.") # Optional Debug
                        self.cancel_rectangle_crop_mode() # This handles both now if auto_lane_rect is part of crop_rectangle_mode logic
                        if self.live_view_label.mode == 'auto_lane_rect': # Specific reset for auto_lane_rect if not covered
                            self.live_view_label.mode = None
                            self.live_view_label.setCursor(Qt.ArrowCursor)
                            self.live_view_label.mousePressEvent = None
                            self.live_view_label.mouseMoveEvent = None
                            self.live_view_label.mouseReleaseEvent = None
                        self.live_view_label.clear_crop_preview()
                        self.update_live_view()
                        return # Consume the event

                    # 2. Cancel Line/Rectangle Drawing Mode
                    if self.drawing_mode in ['line', 'rectangle']:
                        # print("Debug: Escape pressed, cancelling shape drawing.") # Optional Debug
                        self.cancel_drawing_mode()
                        self.update_live_view()
                        return # Consume the event

                    # 3. Cancel Marker Preview/Placement Modes
                    if self.live_view_label.preview_marker_enabled:
                        # print("Debug: Escape pressed, cancelling marker preview.") # Optional Debug
                        self.live_view_label.preview_marker_enabled = False
                        self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                        self.update_live_view() # Update to clear preview
                        return # Consume the event

                    if self.marker_mode is not None: # If any marker placement mode was active
                        # print(f"Debug: Escape pressed, cancelling marker mode: {self.marker_mode}") # Optional Debug
                        self.marker_mode = None
                        self.live_view_label.mousePressEvent = None # Reset handler
                        self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                        self.update_live_view()
                        return # Consume the event

                    # 4. Cancel Analysis Quadrilateral/Rectangle Definition/Move
                    if self.live_view_label.measure_quantity_mode or self.live_view_label.mode in ["quad", "rectangle", "move"]:
                        # print("Debug: Escape pressed, cancelling analysis area mode.") # Optional Debug
                        self.live_view_label.measure_quantity_mode = False
                        self.live_view_label.bounding_box_complete = False
                        self.live_view_label.counter = 0
                        self.live_view_label.quad_points = []
                        self.live_view_label.bounding_box_preview = None
                        self.live_view_label.rectangle_points = []
                        self.live_view_label.rectangle_start = None
                        self.live_view_label.rectangle_end = None
                        self.live_view_label.selected_point = -1
                        self.live_view_label.mousePressEvent = None
                        self.live_view_label.mouseMoveEvent = None
                        self.live_view_label.mouseReleaseEvent = None
                        self.live_view_label.mode = None
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        self.update_live_view()
                        return # Consume the event
                    
                    if self.live_view_label.mw_predict_preview_enabled:
                        self.live_view_label.mw_predict_preview_enabled = False
                        self.live_view_label.mw_predict_preview_position = None
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        self.live_view_label.mousePressEvent = None # Reset handler
                        self.live_view_label.mouseMoveEvent = None  # Reset handler
                        self.live_view_label.setMouseTracking(False) # Can turn off
                        self.update_live_view()
                # ... (rest of keyPressEvent for panning) ...
                if self.live_view_label.zoom_level != 1.0:
                    step = 20
                    offset_changed = False # Flag to update view only if changed
                    current_x = self.live_view_label.pan_offset.x()
                    current_y = self.live_view_label.pan_offset.y()

                    if event.key() == Qt.Key_Left:
                        self.live_view_label.pan_offset.setX(current_x - step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Right:
                        self.live_view_label.pan_offset.setX(current_x + step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Up:
                        self.live_view_label.pan_offset.setY(current_y - step)
                        offset_changed = True
                    elif event.key() == Qt.Key_Down:
                        self.live_view_label.pan_offset.setY(current_y + step)
                        offset_changed = True

                    if offset_changed:
                        self.update_live_view()
                        return # Consume arrow keys if panning occurred
                super().keyPressEvent(event)
            
            def update_marker_text_font(self, font: QFont):
                """
                Updates the font of self.custom_marker_text_entry based on the selected font
                from self.custom_font_type_dropdown.
            
                :param font: QFont object representing the selected font from the combobox.
                """
                # Get the name of the selected font
                selected_font_name = font.family()
                
                # Set the font of the QLineEdit
                self.custom_marker_text_entry.setFont(QFont(selected_font_name))
            
            def arrow_marker(self, text: str):
                self.custom_marker_text_entry.clear()
                self.custom_marker_text_entry.setText(text)
            
            def enable_custom_marker_mode(self):
                """Enable the custom marker mode and set the mouse event."""
                self.live_view_label.setCursor(Qt.ArrowCursor)
                custom_text = self.custom_marker_text_entry.text().strip()
            
                self.live_view_label.preview_marker_enabled = True
                self.live_view_label.preview_marker_text = custom_text
                self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
                self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
                self.live_view_label.marker_color=self.custom_marker_color
                
                self.live_view_label.setFocus()
                self.live_view_label.update()
                
                self.marker_mode = "custom"  # Indicate custom marker mode
                self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
                
            def remove_custom_marker_mode(self):
                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    self.custom_markers.pop()  # Remove the last entry from the list           
                self.update_live_view()  # Update the display
                
            def remove_last_custom_shape(self):
                """Removes the last added custom SHAPE (line or rectangle)."""
                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    self.save_state() # Save state before removal for undo
                    removed_shape = self.custom_shapes.pop()  # Remove the last shape
                    print(f"Removed last custom shape: {removed_shape}") # Optional debug
                    self.is_modified = True
                    self.update_live_view()  # Update the display
                    # Optionally, provide user feedback via status bar or QMessageBox
                    # self.statusBar().showMessage("Last custom shape removed.", 2000)
                else:
                    print("No custom shapes to remove.") # Info message
                    # Optionally, provide user feedback
                    # self.statusBar().showMessage("No custom shapes to remove.", 2000)
                
            def reset_custom_marker_mode(self):
                """Resets (clears) all custom MARKERS and custom SHAPES."""
                markers_cleared = False
                shapes_cleared = False
                needs_save = False # Flag to check if state needs saving

                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    if not needs_save: # Save state only once before the first clear operation
                         self.save_state()
                         needs_save = True
                    self.custom_markers.clear()
                    markers_cleared = True
                    print("Cleared all custom markers.")

                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    if not needs_save: # Save state only once if markers weren't cleared but shapes are
                         self.save_state()
                         needs_save = True
                    self.custom_shapes.clear()
                    shapes_cleared = True
                    print("Cleared all custom shapes.")

                if markers_cleared or shapes_cleared:
                    self.is_modified = True
                    self.update_live_view()
                else:
                    print("No custom markers or shapes to reset.")
                
            
            def place_custom_marker(self, event, custom_text):
                """Place a custom marker at the cursor location."""
                self.save_state()
                # Get cursor position from the event
                pos = event.pos()
                cursor_x, cursor_y = pos.x(), pos.y()
                
                
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                grid_size = self.grid_size_input.value() # Get grid size once

                if self.show_grid_checkbox_x.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_x = round(cursor_x / grid_size) * grid_size

                if self.show_grid_checkbox_y.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_y = round(cursor_y / grid_size) * grid_size
            
                # Dimensions of the displayed image
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
            
                # Dimensions of the actual image
                image_width = self.image.width()
                image_height = self.image.height()
            
                # Calculate scaling factors
                scale = min(displayed_width / image_width, displayed_height / image_height)
            
                # Calculate offsets
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2
            
                # Transform cursor position to image space
                image_x = (cursor_x - x_offset) / scale
                image_y = (cursor_y - y_offset) / scale
                
                    
                # Store the custom marker's position and text
                self.custom_markers = getattr(self, "custom_markers", [])
                self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value(),False,False))
                self.update_live_view()
                
            def select_custom_marker_color(self):
                """Open a color picker dialog to select the color for custom markers."""
                color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
                if color.isValid():
                    self.custom_marker_color = color  # Update the custom marker color
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
            
            def update_all_labels(self):
                """Update all labels from the QTextEdit, supporting multiple columns."""
                self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
                input_text = self.top_marker_input.toPlainText()
                
                # Split the text into a list by commas and strip whitespace
                self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
                

                # if len(self.top_label) < len(self.top_markers):
                #     self.top_markers = self.top_markers[:len(self.top_label)]
                for i in range(0, len(self.top_markers)):
                    try:
                        self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                    except:
                        self.top_markers[i] = (self.top_markers[i][0], str(""))

                # if len(self.marker_values) < len(self.left_markers):
                #     self.left_markers = self.left_markers[:len(self.marker_values)]
                for i in range(0, len(self.left_markers)):
                    try:
                        self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
                    except:
                        self.left_markers[i] = (self.left_markers[i][0], str(""))
                        

                # if len(self.marker_values) < len(self.right_markers):
                #     self.right_markers = self.right_markers[:len(self.marker_values)]
                for i in range(0, len(self.right_markers)):
                    try:
                        self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
                    except:
                        self.right_markers[i] = (self.right_markers[i][0], str(""))
                        

                
                # Trigger a refresh of the live view
                self.update_live_view()
                
            def reset_marker(self, marker_type, param):    
                self.save_state()
                if marker_type == 'left':
                    if param == 'remove' and len(self.left_markers)!=0:
                        self.left_markers.pop()  
                        if self.current_left_marker_index > 0:
                            self.current_left_marker_index -= 1
                    elif param == 'reset':
                        self.left_markers.clear()
                        self.current_left_marker_index = 0  

                     
                elif marker_type == 'right' and len(self.right_markers)!=0:
                    if param == 'remove':
                        self.right_markers.pop()  
                        if self.current_right_marker_index > 0:
                            self.current_right_marker_index -= 1
                    elif param == 'reset':
                        self.right_markers.clear()
                        self.current_right_marker_index = 0

                elif marker_type == 'top' and len(self.top_markers)!=0:
                    if param == 'remove':
                        self.top_markers.pop() 
                        if self.current_top_label_index > 0:
                            self.current_top_label_index -= 1
                    elif param == 'reset':
                        self.top_markers.clear()
                        self.current_top_label_index = 0
            
                # Call update live view after resetting markers
                self.update_live_view()
                
            def duplicate_marker(self, marker_type):
                self.save_state() # Save state before making changes

                if marker_type == 'left' and self.right_markers:
                    # Copy Right Markers TO Left
                    self.left_markers = self.right_markers.copy()
                    # Copy Right Offset value TO Left Offset variable
                    self.left_marker_shift_added = self.right_marker_shift_added

                    # Update Left Slider position to match the copied offset
                    if hasattr(self, 'left_padding_slider'):
                        min_val, max_val = self.left_padding_slider.minimum(), self.left_padding_slider.maximum()
                        # Clamp the value to the slider's current range
                        clamped_value = max(min_val, min(self.left_marker_shift_added, max_val))
                        # Update slider visually, preventing signal emission temporarily
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(clamped_value)
                        self.left_padding_slider.blockSignals(False)
                        # Ensure the internal variable matches the potentially clamped value
                        self.left_marker_shift_added = clamped_value

                elif marker_type == 'right' and self.left_markers:
                    # Copy Left Markers TO Right
                    self.right_markers = self.left_markers.copy()
                    # Copy Left Offset value TO Right Offset variable
                    self.right_marker_shift_added = self.left_marker_shift_added

                    # Update Right Slider position to match the copied offset
                    if hasattr(self, 'right_padding_slider'):
                        min_val, max_val = self.right_padding_slider.minimum(), self.right_padding_slider.maximum()
                        # Clamp the value to the slider's current range
                        clamped_value = max(min_val, min(self.right_marker_shift_added, max_val))
                        # Update slider visually, preventing signal emission temporarily
                        self.right_padding_slider.blockSignals(True)
                        self.right_padding_slider.setValue(clamped_value)
                        self.right_padding_slider.blockSignals(False)
                        # Ensure the internal variable matches the potentially clamped value
                        self.right_marker_shift_added = clamped_value


                # Call update live view after duplicating markers and updating offsets/sliders
                self.update_live_view()
                
            def on_combobox_changed(self):
                preset_name = self.combo_box.currentText()
                
                # Save state if switching *from* a modified "Custom" state or another preset
                # This allows undoing the application of a new preset template.
                if self.is_modified or preset_name != "Custom": # More precise condition
                    self.save_state()

                if preset_name == "Custom":
                    self.marker_values_textbox.setEnabled(True)
                    self.rename_input.setEnabled(True)
                    self.rename_input.clear() # Clear rename field, ready for new name
                    self.marker_values_textbox.setPlaceholderText("Current L/R values, edit to save as new")
                    # Update L/R marker values textbox from current internal self.marker_values
                    current_lr_display_values = [str(v) if isinstance(v, (int, float)) else v for v in getattr(self, 'marker_values', [])]
                    self.marker_values_textbox.setText(", ".join(current_lr_display_values))

                    # Update Top labels textedit from current internal self.top_label
                    current_top_display_labels = [str(v) for v in getattr(self, 'top_label', [])]
                    self.top_marker_input.setText(", ".join(current_top_display_labels))

                elif preset_name in self.presets_data:
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.rename_input.clear()

                    preset_config = self.presets_data[preset_name]
                    
                    # Load L/R marker values for the preset into the internal list and textbox
                    self.marker_values = list(preset_config.get("marker_values", []))
                    display_marker_values = [str(v) if isinstance(v, (int, float)) else v for v in self.marker_values]
                    self.marker_values_textbox.setText(", ".join(display_marker_values))

                    # Load Top labels for the preset into the internal list and textedit
                    self.top_label = list(preset_config.get("top_labels", []))
                    self.top_marker_input.setText(", ".join(map(str, self.top_label)))

                    # --- LOAD CUSTOM MARKERS FROM PRESET ---
                    custom_markers_config_from_preset = preset_config.get("custom_markers_config", [])
                    if not isinstance(custom_markers_config_from_preset, list):
                        print(f"Warning: 'custom_markers_config' for preset '{preset_name}' is not a list.")
                        custom_markers_config_from_preset = []
                    deserialized_markers = self._deserialize_custom_markers(custom_markers_config_from_preset)
                    self.custom_markers = [list(m) for m in deserialized_markers] # Populate self.custom_markers

                    # --- LOAD CUSTOM SHAPES FROM PRESET ---
                    custom_shapes_config_from_preset = preset_config.get("custom_shapes_config", [])
                    if not isinstance(custom_shapes_config_from_preset, list):
                        print(f"Warning: 'custom_shapes_config' for preset '{preset_name}' is not a list.")
                        custom_shapes_config_from_preset = []
                    self.custom_shapes = [dict(s) for s in custom_shapes_config_from_preset if isinstance(s, dict)] # Populate self.custom_shapes
                    
                    # Update standard L/R/Top markers *on the image* if any are already placed.
                    # This also calls update_live_view() at the end.
                    self.update_all_labels() 
                    
                else: # Preset name not found in data (should be rare)
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.marker_values_textbox.clear()
                    self.top_marker_input.clear()
                    self.custom_markers.clear() 
                    self.custom_shapes.clear()  
                    self.marker_values = []
                    self.top_label = []
                    self.update_live_view()

            
            def reset_gamma_contrast(self):
                try:
                    if self.image_before_contrast==None:
                        self.image_before_contrast=self.image_master.copy()
                    self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
                    self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
                    self.high_slider.setValue(100)  # Reset contrast to default
                    self.low_slider.setValue(100)  # Reset contrast to default
                    self.gamma_slider.setValue(100)  # Reset gamma to default
                    self.update_live_view()
                except:
                    pass

            
            def update_image_contrast(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                    
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                        self.update_live_view()
                except:
                    pass
            
            def update_image_gamma(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                        
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                        self.update_live_view()
                except:
                    pass
            
            def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
                """
                Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
                preserving the original format (including color and bit depth) where possible.
                Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
                """
                if not qimage or qimage.isNull():
                    return qimage

                original_format = qimage.format()
                try:
                    img_array = self.qimage_to_numpy(qimage)
                    if img_array is None: raise ValueError("NumPy conversion failed.")

                    # Work with float64 for calculations to avoid precision issues
                    img_array_float = img_array.astype(np.float64)

                    # Determine max value based on original data type
                    if img_array.dtype == np.uint16:
                        max_val = 65535.0
                    elif img_array.dtype == np.uint8:
                        max_val = 255.0
                    else: # Default for unexpected types (e.g., float input?)
                         max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0

                    # --- Apply adjustments ---
                    if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                        num_channels = img_array.shape[2]
                        adjusted_channels = []

                        # Process only the color channels (first 3 usually)
                        channels_to_process = min(num_channels, 3)
                        for i in range(channels_to_process):
                            channel = img_array_float[:, :, i]
                            # Normalize to 0-1 range
                            channel_norm = channel / max_val

                            # Apply brightness (high_factor): Multiply
                            channel_norm = channel_norm * high_factor

                            # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                            mid_grey = 0.5
                            contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                            channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                            # Clip to 0-1 range after contrast/brightness
                            channel_norm = np.clip(channel_norm, 0.0, 1.0)

                            # Apply gamma correction
                            safe_gamma = max(0.01, gamma)
                            channel_norm = np.power(channel_norm, safe_gamma)

                            # Clip again after gamma
                            channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                            # Scale back to original range
                            adjusted_channels.append(channel_norm_clipped * max_val)

                        # Reconstruct the image array
                        img_array_final_float = np.stack(adjusted_channels, axis=2)

                        # Keep the alpha channel (if present) untouched
                        if num_channels == 4:
                            alpha_channel = img_array_float[:, :, 3] # Get original alpha
                            img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

                    elif img_array.ndim == 2: # Grayscale Image
                        # Normalize to 0-1 range
                        img_array_norm = img_array_float / max_val

                        # Apply brightness
                        img_array_norm = img_array_norm * high_factor

                        # Apply contrast
                        mid_grey = 0.5
                        contrast_factor = max(0.01, low_factor)
                        img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                        # Clip
                        img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                        # Apply gamma
                        safe_gamma = max(0.01, gamma)
                        img_array_norm = np.power(img_array_norm, safe_gamma)

                        # Clip again
                        img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                        # Scale back
                        img_array_final_float = img_array_norm_clipped * max_val
                    else:
                        return qimage # Return original if unsupported dimensions

                    # Convert back to original data type
                    img_array_final = img_array_final_float.astype(img_array.dtype)

                    # Convert back to QImage using the helper function
                    result_qimage = self.numpy_to_qimage(img_array_final)
                    if result_qimage.isNull():
                        raise ValueError("Conversion back to QImage failed.")

                    # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
                    return result_qimage

                except Exception as e:
                    traceback.print_exc() # Print detailed traceback
                    return qimage # Return original QImage on error
            

            def save_contrast_options(self):
                if self.image:
                    self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
                    self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
                else:
                    QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

            def remove_config(self): # This is for "Remove Preset" button
                selected_preset_name = self.combo_box.currentText()

                if selected_preset_name == "Custom":
                    QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' option.")
                    return
                
                # Prevent deletion of essential default presets if desired (example)
                # essential_defaults = ["Precision Plus Protein All Blue Prestained (Bio-Rad)", "1 kb Plus DNA Ladder (Thermo 10787018)"]
                # if selected_preset_name in essential_defaults:
                #     QMessageBox.warning(self, "Error", f"Cannot remove the built-in preset: '{selected_preset_name}'.")
                #     return

                if selected_preset_name not in self.presets_data:
                    QMessageBox.warning(self, "Error", f"Preset '{selected_preset_name}' not found in data.")
                    return

                reply = QMessageBox.question(self, "Confirm Remove",
                                             f"Are you sure you want to remove the preset '{selected_preset_name}'?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return

                try:
                    del self.presets_data[selected_preset_name]

                    # Save the updated configuration
                    config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)

                    # Update ComboBox
                    self.combo_box.blockSignals(True)
                    current_idx = self.combo_box.findText(selected_preset_name)
                    if current_idx != -1:
                        self.combo_box.removeItem(current_idx)
                    # Select "Custom" or the first available preset after removal
                    if self.combo_box.count() > 1: # More than just "Custom" left
                        custom_idx = self.combo_box.findText("Custom")
                        if self.combo_box.currentIndex() == custom_idx and custom_idx > 0: # If custom was selected and not first
                             self.combo_box.setCurrentIndex(custom_idx -1) # Select item before custom
                        elif self.combo_box.currentIndex() == custom_idx and custom_idx == 0 and self.combo_box.count() > 1: # custom is first, but others exist
                            self.combo_box.setCurrentIndex(1)
                        # else keep current selection if it's not the removed one
                    elif self.combo_box.count() == 1 and self.combo_box.itemText(0) == "Custom":
                        self.combo_box.setCurrentIndex(0)

                    self.combo_box.blockSignals(False)
                    self.on_combobox_changed() # Refresh UI based on new selection

                    QMessageBox.information(self, "Success", f"Preset '{selected_preset_name}' removed.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Error removing preset: {e}")
                    traceback.print_exc()

            def save_config(self): # This is for "Save Preset" button
                """Saves the current settings (L/R/Top values, custom markers/shapes) to the selected/new preset name."""
                
                current_combo_text = self.combo_box.currentText()
                new_name_from_input = self.rename_input.text().strip()
                
                name_to_save = ""
                is_new_preset_creation = False

                if new_name_from_input: # User entered a name in rename_input
                    name_to_save = new_name_from_input
                    is_new_preset_creation = True
                    if name_to_save == "Custom":
                        QMessageBox.warning(self, "Invalid Name", "Cannot save a preset with the name 'Custom'. Please choose another name or clear the rename field to update the selected preset.")
                        return
                elif current_combo_text != "Custom": # Updating an existing preset
                    name_to_save = current_combo_text
                else: # "Custom" is selected, and rename_input is empty
                    QMessageBox.information(self, "Save Preset", "Please enter a new name in the 'New name for Custom preset' field to save the current settings as a new preset.")
                    return

                # Confirmation for overwriting
                if name_to_save in self.presets_data and is_new_preset_creation : # Only ask if it's a new name that already exists
                     reply = QMessageBox.question(self, "Overwrite Preset?",
                                                      f"A preset named '{name_to_save}' already exists. Overwrite it?",
                                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                     if reply == QMessageBox.No:
                         return
                
                # --- Gather data to save ---
                # L/R Marker Values: Always take from the textbox.
                try:
                    raw_markers_text = self.marker_values_textbox.text().strip("[]")
                    if raw_markers_text:
                        marker_values_to_save = [
                            int(num.strip()) if num.strip().isdigit() else float(num.strip()) if '.' in num.strip() else num.strip()
                            for num in raw_markers_text.split(",") if num.strip()
                        ]
                    else:
                        marker_values_to_save = []
                except ValueError:
                    QMessageBox.warning(self, "Input Error", "Invalid format for L/R marker values. Please use comma-separated numbers or text.")
                    return

                # Top Labels: Always take from the textedit.
                top_labels_to_save = [label.strip() for label in self.top_marker_input.toPlainText().split(",") if label.strip()]
                
                # Custom Markers: Serialize current self.custom_markers
                custom_markers_config_to_save = self._serialize_custom_markers(getattr(self, 'custom_markers', []))
                
                # Custom Shapes: Are already serializable
                custom_shapes_config_to_save = [dict(s) for s in getattr(self, 'custom_shapes', [])] # Ensure they are dicts

                # --- Update self.presets_data in memory ---
                self.presets_data[name_to_save] = {
                    "marker_values": marker_values_to_save,
                    "top_labels": top_labels_to_save,
                    "custom_markers_config": custom_markers_config_to_save,
                    "custom_shapes_config": custom_shapes_config_to_save
                }

                # --- Save to file ---
                config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                try:
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)
                    
                    # --- Update ComboBox if a new preset was created ---
                    if is_new_preset_creation and self.combo_box.findText(name_to_save) == -1:
                        self.combo_box.blockSignals(True)
                        custom_idx = self.combo_box.findText("Custom")
                        if custom_idx != -1:
                            self.combo_box.insertItem(custom_idx, name_to_save)
                        else:
                            self.combo_box.addItem(name_to_save) # Fallback
                        self.combo_box.setCurrentText(name_to_save) # Select the new/updated preset
                        self.combo_box.blockSignals(False)
                        # self.on_combobox_changed() # Triggered by setCurrentText if not blocked

                    self.rename_input.clear() # Clear rename field after successful save
                    self.marker_values_textbox.setEnabled(False) # Disable textbox after saving to a named preset
                    self.rename_input.setEnabled(False)

                    QMessageBox.information(self, "Preset Saved", f"Preset '{name_to_save}' saved successfully.")

                except Exception as e:
                    QMessageBox.critical(self, "Save Preset Error", f"Could not save preset configuration:\n{e}")
                    traceback.print_exc()
            
            
            def load_config(self):
                """
                Load preset configuration from file. If the file doesn't exist,
                create it with default popular marker standards and empty custom templates.
                """
                config_loaded_successfully = False
                self.presets_data.clear() # Start with an empty dictionary

                # --- Determine Application Path (same as your existing load_config) ---
                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                     application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()
                # --- End Application Path Detection ---

                config_filepath = os.path.join(application_path, self.CONFIG_PRESET_FILE_NAME)
                print(f"INFO: Attempting to load/create preset config at: {config_filepath}")

                # --- Define Default Marker Data ---
                # Use kDa for proteins, bp for DNA for clarity in keys, but values are just numbers
                default_marker_standards = {
                    # Proteins (kDa)
                    "Precision Plus Protein All Blue Prestained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Unstained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Dual Color (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "PageRuler Prestained Protein Ladder (Thermo)": [250, 130, 100, 70, 55, 35, 25, 15, 10],
                    "PageRuler Unstained Protein Ladder (Thermo)": [200, 150, 100, 70, 50, 40, 30, 20, 10],
                    "Spectra Multicolor Broad Range Protein Ladder (Thermo)": [260, 140, 100, 75, 60, 45, 35, 25, 15, 10],
                    # DNA (bp)
                    "1 kb DNA Ladder (NEB N3232)": [10000, 8000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 500],
                    "1 kb Plus DNA Ladder (Thermo 10787018)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 75],
                    "TrackIt 1 Kb Plus DNA Ladder (Invitrogen 10488085)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
                    "Lambda DNA/HindIII Marker (NEB N3012)": [23130, 9416, 6557, 4361, 2322, 2027, 564],
                }
                # Generic default top labels
                default_top_labels_generic = ["MWM", "L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L13", "L14", "L15"]
                # default_top_label_dict = {name: default_top_labels_generic[:] for name in default_marker_values.keys()} # Assign generic to all defaults
                default_presets_init = {}
                for name, markers in default_marker_standards.items():
                    default_presets_init[name] = {
                        "marker_values": list(markers), # Ensure it's a list
                        "top_labels": list(default_top_labels_generic), # Ensure it's a list
                        "custom_markers_config": [], # Empty by default
                        "custom_shapes_config": []   # Empty by default
                    }
                # --- End Default Marker Data ---

                # --- Initialize internal dictionaries BEFORE trying to load/create ---
                # Start with defaults, then overwrite if load is successful.
                if os.path.exists(config_filepath):
                    try:
                        with open(config_filepath, "r", encoding='utf-8') as f:
                            loaded_json = json.load(f)
                        
                        # Check for new "presets" key
                        if "presets" in loaded_json and isinstance(loaded_json["presets"], dict):
                            self.presets_data = loaded_json["presets"]
                            # Ensure all default fields exist for loaded presets
                            for preset_name, data in self.presets_data.items():
                                if "marker_values" not in data: data["marker_values"] = []
                                if "top_labels" not in data: data["top_labels"] = []
                                if "custom_markers_config" not in data: data["custom_markers_config"] = []
                                if "custom_shapes_config" not in data: data["custom_shapes_config"] = []
                            config_loaded_successfully = True
                        else: # Try to migrate old format
                            print("INFO: Old config format detected. Attempting migration...")
                            migrated_data = {}
                            old_marker_values = loaded_json.get("marker_values", {})
                            old_top_labels = loaded_json.get("top_label", {})
                            all_preset_names = set(old_marker_values.keys()) | set(old_top_labels.keys())

                            for name in all_preset_names:
                                migrated_data[name] = {
                                    "marker_values": list(old_marker_values.get(name, [])),
                                    "top_labels": list(old_top_labels.get(name, default_top_labels_generic)),
                                    "custom_markers_config": [], # No custom markers in old format
                                    "custom_shapes_config": []   # No custom shapes in old format
                                }
                            self.presets_data = migrated_data
                            # Save immediately in new format after migration
                            with open(config_filepath, "w", encoding='utf-8') as f_new:
                                json.dump({"presets": self.presets_data}, f_new, indent=4)
                            print("INFO: Config migrated to new format and saved.")
                            config_loaded_successfully = True

                    except (json.JSONDecodeError, IOError, TypeError) as e:
                        QMessageBox.warning(self, "Preset Config Load Error",
                                            f"Could not load '{self.CONFIG_PRESET_FILE_NAME}':\n{e}\n\nUsing default presets.")
                        self.presets_data = default_presets_init.copy() # Fallback to defaults
                    except Exception as e:
                        traceback.print_exc()
                        QMessageBox.warning(self, "Preset Config Load Error",
                                            f"Unexpected error loading '{self.CONFIG_PRESET_FILE_NAME}'.\n\nUsing default presets.")
                        self.presets_data = default_presets_init.copy() # Fallback
                else:
                    # Config file NOT found, create it with defaults
                    self.presets_data = default_presets_init.copy()
                    try:
                        with open(config_filepath, "w", encoding='utf-8') as f:
                            json.dump({"presets": self.presets_data}, f, indent=4)
                        config_loaded_successfully = True # Created successfully
                        print(f"INFO: Default preset config created at: {config_filepath}")
                    except Exception as e:
                        QMessageBox.critical(self, "Preset Config Creation Error",
                                             f"Could not create default preset config file:\n{e}\n\nDefault presets will be used for this session only.")
                        # self.presets_data remains default_presets_init

                # --- Update UI (ComboBox) ---
                try:
                    if hasattr(self, 'combo_box'):
                        self.combo_box.blockSignals(True)
                        self.combo_box.clear()
                        sorted_preset_names = sorted(self.presets_data.keys())
                        self.combo_box.addItems(sorted_preset_names)
                        self.combo_box.addItem("Custom")

                        # Set current item (same logic as your existing load_config)
                        default_biorad = "Precision Plus Protein All Blue Prestained (Bio-Rad)"
                        idx_to_select = self.combo_box.findText(default_biorad)
                        if idx_to_select == -1 and sorted_preset_names:
                            idx_to_select = 0 # Select first actual preset if BioRad not found
                        elif idx_to_select == -1: # No actual presets, only "Custom"
                            idx_to_select = self.combo_box.findText("Custom")
                        
                        if idx_to_select != -1:
                             self.combo_box.setCurrentIndex(idx_to_select)
                        
                        self.combo_box.blockSignals(False)
                        self.on_combobox_changed() # Trigger update based on selection
                    else:
                        print("Warning: combo_box UI element not found during config load.")
                except Exception as e_ui:
                    traceback.print_exc()
                    QMessageBox.warning(self, "UI Error", f"Error populating preset combobox: {e_ui}")
            
            def paste_image(self):
                """Handle pasting image from clipboard."""
                self.is_modified = True
                self.reset_image() # Clear previous state first
                self.load_image_from_clipboard()
                self._update_overlay_slider_ranges()
                # UI updates (label size, sliders) should happen within load_image_from_clipboard
                self.update_live_view()
                self.save_state() # Save state after pasting
            
            def load_image_from_clipboard(self):
                """
                Load an image from the clipboard into self.image, preserving format.
                Prioritizes file paths over raw image data to handle macOS file copying correctly.
                If loaded from a file path, attempts to load an associated config file.
                """
                # Check for unsaved changes before proceeding
                # if not self.prompt_save_if_needed(): # Optional: uncomment if needed
                #     return # Abort paste if user cancels save

                self.reset_image() # Clear previous state first

                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()
                loaded_image = None
                source_info = "Clipboard" # Default source
                config_loaded_from_paste = False # Flag to track if config was loaded
                

                # --- PRIORITY 1: Check for File URLs ---
                if mime_data.hasUrls():
                    urls = mime_data.urls()
                    if urls:
                        file_path = urls[0].toLocalFile()
                        if os.path.isfile(file_path) and file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                            loaded_image = QImage(file_path)
                            source_info = file_path

                            if loaded_image.isNull():
                                try:
                                    pil_image = Image.open(file_path)
                                    loaded_image = ImageQt.ImageQt(pil_image)
                                    if loaded_image.isNull():
                                        loaded_image = None
                                    else:
                                        source_info = f"{file_path} (Pillow)"
                                except Exception as e:
                                    QMessageBox.warning(self, "File Load Error", f"Could not load image from file '{os.path.basename(file_path)}':\n{e}")
                                    loaded_image = None

                            # --- *** ADDED: CONFIG FILE LOADING FOR PASTED FILE *** ---
                            if loaded_image and not loaded_image.isNull():
                                self.image_path = file_path # Store the path early for config lookup
                                base_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                                # Construct potential config file name (remove _original/_modified if present)
                                config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                                config_path = os.path.join(os.path.dirname(file_path), f"{config_base}_config.txt")

                                if os.path.exists(config_path):
                                    try:
                                        with open(config_path, "r") as config_file:
                                            config_data = json.load(config_file)
                                        self.apply_config(config_data) # Apply loaded settings
                                        config_loaded_from_paste = True # Set flag
                                    except Exception as e:
                                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply associated config file '{os.path.basename(config_path)}': {e}")

                            # --- *** END OF ADDED CONFIG LOADING *** ---


                # --- PRIORITY 2: Check for Raw Image Data (if no valid file was loaded) ---
                if loaded_image is None and mime_data.hasImage():
                    loaded_image = clipboard.image()
                    if not loaded_image.isNull():
                        source_info = "Clipboard Image Data"
                    else:
                        try:
                            pil_image = ImageGrab.grabclipboard()
                            if isinstance(pil_image, Image.Image):
                                loaded_image = ImageQt.ImageQt(pil_image)
                                if loaded_image.isNull():
                                     loaded_image = None
                                else:
                                    source_info = "Clipboard Image Data (Pillow Grab)"
                            else:
                                loaded_image = None
                        except Exception: # Keep quiet on Pillow grab errors
                            loaded_image = None

                # --- Process the successfully loaded image ---
                if loaded_image and not loaded_image.isNull():
                    self.is_modified = True
                    self.image = loaded_image

                    # Initialize backups
                    self.original_image = self.image.copy()
                    self.image_master = self.image.copy()
                    self.image_before_padding = None
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False
                    # self.image_path is set earlier if loaded from file

                    # --- Update UI Elements ---
                    try:
                        w = self.image.width()
                        h = self.image.height()
                        if w <= 0 or h <= 0: raise ValueError("Invalid image dimensions after load.")

                        # Update preview label size
                        ratio=w/h if h > 0 else 1
                        target_width = int(self.label_size) # Use current label_size setting
                        target_height = int(target_width / ratio)
                        # Optional: Apply max height constraint
                        # if hasattr(self, 'preview_label_max_height_setting') and target_height > self.preview_label_max_height_setting:
                        #     target_height = self.preview_label_max_height_setting
                        #     target_width = int(target_height * ratio)
                        self._update_preview_label_size()

                        # Update slider ranges based on *new* label size
                        render_scale = 3
                        render_width = self.live_view_label.width() * render_scale
                        render_height = self.live_view_label.height() * render_scale
                        self._update_marker_slider_ranges()

                        # Set recommended padding only if config wasn't loaded
                        if not config_loaded_from_paste:
                            self.left_padding_input.setText(str(int(w * 0.1)))
                            self.right_padding_input.setText(str(int(w * 0.1)))
                            self.top_padding_input.setText(str(int(h * 0.15)))
                            self.bottom_padding_input.setText("0")

                        # Update window title and status bar
                        display_source = os.path.basename(source_info.split(" (")[0]) # Show filename or simplified source
                        title_suffix = " (+ Config)" if config_loaded_from_paste else ""
                        self.setWindowTitle(f"{self.window_title}::{display_source}{title_suffix}")
                        self._update_status_bar()

                    except Exception as e:
                        QMessageBox.warning(self, "UI Update Error", f"Could not update UI elements after pasting: {e}")
                        self._update_preview_label_size()
                        
                    enable_pan = self.live_view_label.zoom_level > 1.0
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                    self.update_live_view()
                    self.save_state()

                else:
                    # If no image was successfully loaded
                    QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
                    # Clean up state
                    self.image = None
                    self.original_image = None
                    self.image_master = None
                    self.image_path = None
                    self.live_view_label.clear()
                    self._update_preview_label_size()
                    self.setWindowTitle(self.window_title)
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                    self._update_status_bar()
                    self.update_live_view()
                
            def update_font(self):
                self.save_state()
                """Update the font settings based on UI inputs"""
                # Update font family from the combo box
                self.font_family = self.font_combo_box.currentFont().family()
                
                # Update font size from the spin box
                self.font_size = self.font_size_spinner.value()
                
                self.font_rotation = int(self.font_rotation_input.value())
                
            
                # Once font settings are updated, update the live view immediately
                self.update_live_view()
            
            def select_font_color(self):
                self.save_state()
                color = QColorDialog.getColor()
                if color.isValid():
                    self.font_color = color  # Store the selected color   
                    self.update_font()
                self._update_color_button_style(self.font_color_button, self.font_color)
                    
            def load_image(self):
                self.prompt_save_if_needed()
                # self.undo_stack = []
                # self.redo_stack = []
                self.reset_image() # Clear previous state

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getOpenFileName(
                    self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
                )
                if file_path:
                    self.image_path = file_path
                    loaded_image = QImage(self.image_path)                    

                    if loaded_image.isNull():
                        # Try loading with Pillow as fallback
                        try:
                            pil_image = Image.open(self.image_path)
                            # Use ImageQt for reliable conversion, preserves most formats
                            loaded_image = ImageQt.ImageQt(pil_image)
                            if loaded_image.isNull():
                                raise ValueError("Pillow could not convert to QImage.")

                        except Exception as e:
                            QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}': {e}")
                            self.image_path = None
                            return

                    # --- Keep the loaded image format ---
                    self.image = loaded_image
                    self._update_overlay_slider_ranges()

                    # --- Initialize backups with the loaded format ---
                    if not self.image.isNull():
                        self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                        self.image_master = self.image.copy()   # Master copy for resets
                        self.image_before_padding = None        # Reset padding state
                        self.image_contrasted = self.image.copy() # Backup for contrast
                        self.image_before_contrast = self.image.copy() # Backup for contrast
                        self.image_padded = False               # Reset flag

                        self.setWindowTitle(f"{self.window_title}::{self.image_path}")

                        # --- Load Associated Config File (Logic remains the same) ---
                        self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                        config_name = ""
                        if self.base_name.endswith("_original"):
                            config_name = self.base_name.replace("_original", "_config.txt")
                        else:
                            config_name = self.base_name + "_config.txt"

                        config_path = os.path.join(os.path.dirname(file_path), config_name)

                        if os.path.exists(config_path):
                            try:
                                with open(config_path, "r") as config_file:
                                    config_data = json.load(config_file)
                                self.apply_config(config_data) # Apply loaded settings
                            except Exception as e:
                                QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                        # --- End Config File Loading ---
                        self.is_modified = True # Mark as modified when loading new image
                    else:
                         QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                         return

                    # --- Update UI Elements (Label size, sliders) ---
                    if self.image and not self.image.isNull():
                        try:
                            # (UI update logic remains the same as before)
                            self._update_preview_label_size()

                            render_scale = 3
                            render_width = self.live_view_label.width() * render_scale
                            render_height = self.live_view_label.height() * render_scale
                            self._update_marker_slider_ranges()


                            if not os.path.exists(config_path):
                                self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                                self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                                self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                                self.bottom_padding_input.setText("0")

                        except Exception as e:
                            pass
                    # --- End UI Element Update ---
                    
                    
                    self._update_status_bar() # <--- Add this
                    enable_pan = self.live_view_label.zoom_level > 1.0
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                    self.update_live_view() # Render the loaded image
                    self.save_state() # Save initial loaded state
            
            def apply_config(self, config_data):
                self.left_padding_input.setText(config_data["adding_white_space"]["left"])
                self.right_padding_input.setText(config_data["adding_white_space"]["right"])
                self.top_padding_input.setText(config_data["adding_white_space"]["top"])
                
                try:
                    self.transparency=int(config_data["adding_white_space"]["transparency"])
                except:
                    pass
                
                    
                try:
                    self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
                except:
                    pass
                
            
                try:
                    self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
                    self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
                    self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
                except (KeyError, ValueError) as e:
                    # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
                    pass
                try:
                    self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
                    self.top_marker_input.setText(", ".join(self.top_label))
                except KeyError as e:
                    # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
                    pass
            
                self.font_family = config_data["font_options"]["font_family"]
                self.font_size = config_data["font_options"]["font_size"]
                self.font_rotation = config_data["font_options"]["font_rotation"]
                self.font_color = QColor(config_data["font_options"]["font_color"])
            
                try:
                    top_padding_val = int(config_data["marker_padding"]["top"])
                    left_padding_val = int(config_data["marker_padding"]["left"])
                    right_padding_val = int(config_data["marker_padding"]["right"])
                except (KeyError, ValueError, TypeError) as e:
                    print(f"Warning: Invalid 'marker_padding' data in config: {e}. Using defaults.")
                    top_padding_val, left_padding_val, right_padding_val = 0, 0, 0

                # Load added shifts (internal variables used for rendering)
                try:
                    self.left_marker_shift_added = int(config_data["added_shift"]["left"])
                    self.right_marker_shift_added = int(config_data["added_shift"]["right"])
                    self.top_marker_shift_added = int(config_data["added_shift"]["top"])
                    # Validation: Ensure loaded shift matches padding value if both exist and are valid
                    if self.left_marker_shift_added != left_padding_val or \
                       self.right_marker_shift_added != right_padding_val or \
                       self.top_marker_shift_added != top_padding_val:
                         print("Warning: Config 'added_shift' differs from 'marker_padding'. Prioritizing 'added_shift'.")
                         # Optionally update padding_val to match shift for consistency if desired:
                         # left_padding_val = self.left_marker_shift_added
                         # right_padding_val = self.right_marker_shift_added
                         # top_padding_val = self.top_marker_shift_added

                except KeyError:
                    # If added_shift is missing (older config?), use marker_padding as the source
                    print("Info: 'added_shift' not found in config, using 'marker_padding'.")
                    self.left_marker_shift_added = left_padding_val
                    self.right_marker_shift_added = right_padding_val
                    self.top_marker_shift_added = top_padding_val
                except (ValueError, TypeError) as e:
                     # Handle potential conversion errors if data is invalid
                     print(f"Warning: Invalid 'added_shift' data in config: {e}. Resetting shifts.")
                     self.left_marker_shift_added, left_padding_val = 0, 0
                     self.right_marker_shift_added, right_padding_val = 0, 0
                     self.top_marker_shift_added, top_padding_val = 0, 0

                # Load slider ranges (set range BEFORE value for Qt)
                try:
                    lr_min, lr_max = map(int, config_data["slider_ranges"]["left"])
                    self.left_padding_slider.setRange(lr_min, lr_max)
                    self.left_slider_range = [lr_min, lr_max] # Store internally

                    rr_min, rr_max = map(int, config_data["slider_ranges"]["right"])
                    self.right_padding_slider.setRange(rr_min, rr_max)
                    self.right_slider_range = [rr_min, rr_max] # Store internally

                    tr_min, tr_max = map(int, config_data["slider_ranges"]["top"])
                    self.top_padding_slider.setRange(tr_min, tr_max)
                    self.top_slider_range = [tr_min, tr_max] # Store internally

                except (KeyError, ValueError, TypeError) as e:
                    print(f"Warning: Slider ranges missing or invalid in config: {e}. Ranges will be updated.")
                    # Don't set default ranges here, let _update_marker_slider_ranges handle it later
                    pass

                # Now set the slider values using the determined padding_val
                # Qt might clamp these values if they are outside the range just set.
                self.top_padding_slider.setValue(top_padding_val)
                self.left_padding_slider.setValue(left_padding_val)
                self.right_padding_slider.setValue(right_padding_val)

                # --- CRITICAL RE-SYNC ---
                # After setting slider values (which might have been clamped by setRange),
                # ensure the internal _shift_added variables MATCH the final slider values.
                # This makes the internal state consistent with the UI state *after* loading.
                self.left_marker_shift_added = self.left_padding_slider.value()
                self.right_marker_shift_added = self.right_padding_slider.value()
                self.top_marker_shift_added = self.top_padding_slider.value()
                
                try:
                    try:
                        x = list(map(int, config_data["marker_labels"]["left"]))
                        self.marker_values_textbox.setText(str(x))
                    except:
                        x = list(map(int, config_data["marker_labels"]["right"]))
                        self.marker_values_textbox.setText(str(x))
                    if self.marker_values_textbox.text!="":
                        self.combo_box.setCurrentText("Custom")
                        self.marker_values_textbox.setEnabled(True)                
                    else:
                        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                        self.marker_values_textbox.setEnabled(False)
                        
                except:
                    pass
            
                try:
                    loaded_custom_markers = []
                    for marker_dict in config_data.get("custom_markers", []):
                        try:
                            x = float(marker_dict["x"])
                            y = float(marker_dict["y"])
                            text = str(marker_dict["text"])
                            color = QColor(marker_dict["color"])
                            if not color.isValid(): color = QColor(0,0,0) # Fallback if invalid string
                            loaded_custom_markers.append((
                                float(marker_dict["x"]), float(marker_dict["y"]), str(marker_dict["text"]),
                                color, # Store the QColor object
                                str(marker_dict["font"]), int(marker_dict["font_size"]),
                                bool(marker_dict.get("bold", False)), bool(marker_dict.get("italic", False))
                            ))
                        except (KeyError, ValueError, TypeError) as e:
                            print(f"Warning: Skipping invalid custom marker in config: {marker_dict}, Error: {e}")
                    self.custom_markers = [list(m) for m in loaded_custom_markers] 
                        
                        
                        
                except:
                    pass
                
                try:
                    loaded_custom_shapes = []
                    for shape_dict in config_data.get("custom_shapes", []):
                        try:
                            shape_type = shape_dict.get('type')
                            color_str = shape_dict.get('color')
                            thickness = int(shape_dict.get('thickness'))
            
                            # Validate basic fields
                            if shape_type not in ['line', 'rectangle'] or not color_str or thickness < 1:
                                raise ValueError("Missing or invalid basic shape data")
            
                            # Validate and convert coordinate data
                            parsed_shape = {'type': shape_type, 'color': color_str, 'thickness': thickness}
                            if shape_type == 'line':
                                start = tuple(map(float, shape_dict.get('start')))
                                end = tuple(map(float, shape_dict.get('end')))
                                if len(start) != 2 or len(end) != 2: raise ValueError("Invalid line coordinates")
                                parsed_shape['start'] = start
                                parsed_shape['end'] = end
                            elif shape_type == 'rectangle':
                                rect = tuple(map(float, shape_dict.get('rect'))) # x,y,w,h
                                if len(rect) != 4 or rect[2] < 0 or rect[3] < 0: raise ValueError("Invalid rectangle coordinates/size")
                                parsed_shape['rect'] = rect
                            else:
                                continue # Skip unknown types
            
                            loaded_custom_shapes.append(parsed_shape) # Add the validated shape dict
            
                        except (KeyError, ValueError, TypeError, AttributeError) as e:
                             print(f"Warning: Skipping invalid custom shape in config: {shape_dict}, Error: {e}")
                    self.custom_shapes = loaded_custom_shapes # Assign the loaded list
                except:
                    pass
                
                custom_markers_data_from_image_config = config_data.get("custom_markers", [])
                self.custom_markers = [list(m) for m in self._deserialize_custom_markers(custom_markers_data_from_image_config)]
                self.custom_shapes = list(config_data.get("custom_shapes", []))
                # Set slider ranges from config_data
                try:
                    self.left_padding_slider.setRange(
                        int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
                    )
                    self.right_padding_slider.setRange(
                        int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
                    )
                    self.top_padding_slider.setRange(
                        int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
                    )
                except KeyError:
                    # Handle missing or incomplete slider_ranges data
                    # print("Error: Slider ranges not found in config_data.")
                    pass
                
                try:
                    self.left_marker_shift_added=int(config_data["added_shift"]["left"])
                    self.right_marker_shift_added=int(config_data["added_shift"]["right"])
                    self.top_marker_shift_added=int(config_data["added_shift"]["top"])
                    
                except KeyError:
                    # Handle missing or incomplete slider_ranges data
                    # print("Error: Added Shift not found in config_data.");
                    pass
                    
                # Apply font settings

                
                #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
                
                font_size_new=self.font_size
                font_rotation_new=self.font_rotation
                
                self.font_combo_box.setCurrentFont(QFont(self.font_family))
                self.font_size_spinner.setValue(font_size_new)
                self.font_rotation_input.setValue(font_rotation_new)

                self.update_live_view()
                
            def get_current_config(self):
                """Gathers the current application state into a dictionary for saving."""
                def make_json_serializable(value):
                    if isinstance(value, (np.float16, np.float32, np.float64, np.floating)):
                        return float(value)
                    if isinstance(value, (np.int8, np.int16, np.int32, np.int64, np.integer)):
                        return int(value)
                    if isinstance(value, np.bool_):
                        return bool(value)
                    if isinstance(value, list):
                        return [make_json_serializable(v) for v in value]
                    if isinstance(value, tuple):
                        # For JSON, lists are generally preferred over tuples if modification isn't an issue
                        # For coordinates like (pos, label), we want [float(pos), str(label)]
                        if len(value) == 2 and isinstance(value[1], str): # Heuristic for (pos, label)
                            return [make_json_serializable(value[0]), str(value[1])]
                        return [make_json_serializable(v) for v in value] # General tuple to list
                    if isinstance(value, dict):
                        return {make_json_serializable(k): make_json_serializable(v) for k, v in value.items()}
                    return value
                
                config = {
                    "adding_white_space": {
                        "left": self.left_padding_input.text(),
                        "right": self.right_padding_input.text(),
                        "top": self.top_padding_input.text(),
                        "bottom": self.bottom_padding_input.text(),
                        "transparency": self.transparency, # Assuming self.transparency exists
                    },
                    "marker_positions": {
                        # Store positions only for standard markers
                        "left": [(pos, label) for pos, label in getattr(self, 'left_markers', [])],
                        "right": [(pos, label) for pos, label in getattr(self, 'right_markers', [])],
                        "top": [(pos, label) for pos, label in getattr(self, 'top_markers', [])],
                    },
                    "marker_labels": { # Storing labels separately might be redundant if already in positions
                        "top": getattr(self, 'top_label', []),
                        "left": [marker[1] for marker in getattr(self, 'left_markers', [])],
                        "right": [marker[1] for marker in getattr(self, 'right_markers', [])],
                    },
                    "marker_padding": { # Current slider values
                        "top": self.top_padding_slider.value() if hasattr(self, 'top_padding_slider') else 0,
                        "left": self.left_padding_slider.value() if hasattr(self, 'left_padding_slider') else 0,
                        "right": self.right_padding_slider.value() if hasattr(self, 'right_padding_slider') else 0,
                    },
                    "font_options": { # Default font options for standard markers
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_rotation": self.font_rotation,
                        "font_color": self.font_color.name(),
                    },
                    "slider_ranges": { # Current slider ranges
                         "left": getattr(self, 'left_slider_range', [-100, 1000]),
                         "right": getattr(self, 'right_slider_range', [-100, 1000]),
                         "top": getattr(self, 'top_slider_range', [-100, 1000]),
                     },
                    "added_shift": { # Store current added shifts
                         "left": getattr(self, 'left_marker_shift_added', 0),
                         "right": getattr(self, 'right_marker_shift_added', 0),
                         "top": getattr(self, 'top_marker_shift_added', 0),
                    }
                }

                # --- Updated Custom Markers Section ---
                custom_markers_data = []
                # Use getattr for safety, default to empty list
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        # Unpack 8 elements
                        x, y, text, color, font, font_size, is_bold, is_italic = marker_tuple
                        custom_markers_data.append({
                            "x": x,
                            "y": y,
                            "text": text,
                            "color": color.name(), # Save color name/hex
                            "font": font,
                            "font_size": font_size,
                            "bold": is_bold,       # Save bold flag
                            "italic": is_italic    # Save italic flag
                        })
                    except (ValueError, TypeError, IndexError) as e:
                        pass
                config["custom_markers"] = custom_markers_data
                config["custom_shapes"] = [dict(s) for s in getattr(self, "custom_shapes", [])]
                # --- End Updated Custom Markers ---

                return config
            
            def add_band(self, event):
                self.update_all_labels() 
                self.save_state()
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                
                if not self.image or self.image.isNull() or not self.marker_mode:
                    return

                # --- 1. Get Click Position in Unzoomed Label Space (handles zoom/pan) ---
                click_pos_unzoomed_label_space = self.live_view_label.transform_point(event.pos())
                
                # --- 2. Apply Grid Snapping in Unzoomed Label Space ---
                snapped_click_pos_label_space = self.snap_point_to_grid(click_pos_unzoomed_label_space)
                cursor_x_ls = snapped_click_pos_label_space.x() 
                cursor_y_ls = snapped_click_pos_label_space.y()

                # --- 3. Transform Snapped Label Space Coords to Native Image Coords ---
                if not (self.image and not self.image.isNull()): # Corrected: use self directly
                    return 

                current_app_image = self.image
 
                label_w_widget = float(self.live_view_label.width())
                label_h_widget = float(self.live_view_label.height())
                img_w_native = float(current_app_image.width())
                img_h_native = float(current_app_image.height())

                if img_w_native <= 0 or img_h_native <= 0 or label_w_widget <= 0 or label_h_widget <= 0:
                    return

                scale_native_to_label = min(label_w_widget / img_w_native, label_h_widget / img_h_native)
                if scale_native_to_label <= 1e-9: 
                    return
                
                displayed_img_w_in_label = img_w_native * scale_native_to_label
                displayed_img_h_in_label = img_h_native * scale_native_to_label
                offset_x_img_in_label = (label_w_widget - displayed_img_w_in_label) / 2.0
                offset_y_img_in_label = (label_h_widget - displayed_img_h_in_label) / 2.0

                image_x = (cursor_x_ls - offset_x_img_in_label) / scale_native_to_label
                image_y = (cursor_y_ls - offset_y_img_in_label) / scale_native_to_label
                
                # Clamp to native image dimensions
                image_x = max(0.0, min(image_x, img_w_native))
                image_y = max(0.0, min(image_y, img_h_native))
                
                try:
                    if self.marker_mode == "left":
                        current_marker_count = len(self.left_markers)
                        is_first_marker = (current_marker_count == 0)
                        marker_value_to_add = self.marker_values[current_marker_count] if current_marker_count < len(self.marker_values) else ""
                        
                        self.left_markers.append((image_y, marker_value_to_add)) 
                        self.current_left_marker_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_x))
                            self._update_marker_slider_ranges() 
                            
                            self.left_padding_slider.blockSignals(True)
                            self.left_padding_slider.setValue(
                                max(self.left_slider_range[0], min(slider_target_value_native_pixels, self.left_slider_range[1]))
                            )
                            self.left_padding_slider.blockSignals(False)
                            # Update internal shift variable to match the actual (possibly clamped) slider value.
                            self.left_marker_shift_added = self.left_padding_slider.value()


                    elif self.marker_mode == "right":
                        current_marker_count = len(self.right_markers)
                        is_first_marker = (current_marker_count == 0)
                        marker_value_to_add = self.marker_values[current_marker_count] if current_marker_count < len(self.marker_values) else ""
                        
                        self.right_markers.append((image_y, marker_value_to_add))
                        self.current_right_marker_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_x))
                            self._update_marker_slider_ranges()
                            
                            self.right_padding_slider.blockSignals(True)
                            self.right_padding_slider.setValue(
                                max(self.right_slider_range[0], min(slider_target_value_native_pixels, self.right_slider_range[1]))
                            )
                            self.right_padding_slider.blockSignals(False)
                            self.right_marker_shift_added = self.right_padding_slider.value()


                    elif self.marker_mode == "top":
                        current_marker_count = len(self.top_markers)
                        is_first_marker = (current_marker_count == 0)
                        label_to_add = self.top_label[current_marker_count] if current_marker_count < len(self.top_label) else ""
                        
                        self.top_markers.append((image_x, label_to_add))
                        self.current_top_label_index += 1

                        if is_first_marker:
                            slider_target_value_native_pixels = int(round(image_y)) # For Top marker, Y-coord determines offset
                            self._update_marker_slider_ranges()

                            self.top_padding_slider.blockSignals(True)
                            self.top_padding_slider.setValue(
                                max(self.top_slider_range[0], min(slider_target_value_native_pixels, self.top_slider_range[1]))
                            )
                            self.top_padding_slider.blockSignals(False)
                            self.top_marker_shift_added = self.top_padding_slider.value()
                            
                except Exception as e:
                     import traceback
                     traceback.print_exc()
                     QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

                self.update_live_view()
                

                
            def enable_left_marker_mode(self):
                self.marker_mode = "left"
                self.current_left_marker_index = 0
                self.live_view_label.mousePressEvent = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)

            def enable_right_marker_mode(self):
                self.marker_mode = "right"
                self.current_right_marker_index = 0
                self.live_view_label.mousePressEvent = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)
            
            def enable_top_marker_mode(self):
                self.marker_mode = "top"
                self.current_top_label_index
                self.live_view_label.mousePressEvent = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)
                
            # def remove_padding(self):
            #     if self.image_before_padding!=None:
            #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
            #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
            #     self.image_padded = False  # Reset the padding state
            #     w=self.image.width()
            #     h=self.image.height()
            #     # Preview window
            #     ratio=w/h
            #     self.label_width = 540
            #     label_height=int(self.label_width/ratio)
            #     if label_height>self.label_width:
            #         label_height=540
            #         self.label_width=ratio*label_height
            #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            #     self.update_live_view()
                
            def finalize_image(self): # Padding
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
                    return
                try:
                    padding_left = max(0, int(self.left_padding_input.text()))
                    padding_right = max(0, int(self.right_padding_input.text()))
                    padding_top = max(0, int(self.top_padding_input.text()))
                    padding_bottom = max(0, int(self.bottom_padding_input.text()))
                    self.left_padding_input.setText(str(padding_left))
                    self.right_padding_input.setText(str(padding_right))
                    self.top_padding_input.setText(str(padding_top))
                    self.bottom_padding_input.setText(str(padding_bottom))
                except ValueError:
                    QMessageBox.warning(self, "Error", "Please enter valid non-negative integers for padding.")
                    return

                if padding_left == 0 and padding_right == 0 and padding_top == 0 and padding_bottom == 0:
                    QMessageBox.information(self, "Info", "No padding specified. Image remains unchanged.")
                    return

                self.save_state()

                try:
                    # --- 1. Adjust element coordinates and _marker_shift_added values ---
                    # This updates them based on padding_left and padding_top.
                    self.adjust_elements_for_padding(padding_left, padding_top)

                    # --- 2. Pad the image ---
                    # (Existing NumPy/OpenCV padding logic is good)
                    source_has_alpha = self.image.hasAlphaChannel()
                    np_img = self.qimage_to_numpy(self.image)
                    if np_img is None: raise ValueError("NumPy conversion failed.")
                    numpy_indicates_alpha = (np_img.ndim == 3 and np_img.shape[2] == 4)
                    has_alpha = source_has_alpha or numpy_indicates_alpha
                    fill_value = None
                    if has_alpha: fill_value = (0, 0, 0, 0) 
                    elif np_img.ndim == 3: fill_value = (255, 255, 255) 
                    elif np_img.ndim == 2: fill_value = 65535 if np_img.dtype == np.uint16 else 255
                    else: raise ValueError(f"Unsupported image dimensions for padding: {np_img.ndim}")
                    padded_np = cv2.copyMakeBorder(np_img, padding_top, padding_bottom,
                                                   padding_left, padding_right,
                                                   cv2.BORDER_CONSTANT, value=fill_value)
                    padded_image = self.numpy_to_qimage(padded_np)
                    if padded_image.isNull():
                         raise ValueError("Conversion back to QImage failed after padding.")

                    # --- 3. Update main image and backups ---
                    self.image = padded_image # self.image is now the padded image
                    self.image_padded = True 
                    self.is_modified = True
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    # self.image_before_padding is implicitly handled by save_state before this operation

                    # --- 4. Update UI ---
                    self._update_preview_label_size() 
                    self._update_status_bar()         
                    
                    # Update slider RANGES based on new image dimensions.
                    self._update_marker_slider_ranges() 
                    
                    # Set slider values to the _marker_shift_added values (which were updated by adjust_elements_for_padding).
                    sliders_to_set = [
                        (getattr(self, 'left_padding_slider', None), self.left_marker_shift_added),
                        (getattr(self, 'right_padding_slider', None), self.right_marker_shift_added),
                        (getattr(self, 'top_padding_slider', None), self.top_marker_shift_added)
                    ]
                    for slider, value_to_set in sliders_to_set:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(value_to_set) # Will be clamped by new range if needed
                            slider.blockSignals(False)
                            # Re-sync internal var with actual slider value
                            if slider == self.left_padding_slider: self.left_marker_shift_added = slider.value()
                            elif slider == self.right_padding_slider: self.right_marker_shift_added = slider.value()
                            elif slider == self.top_padding_slider: self.top_marker_shift_added = slider.value()

                    self.update_live_view() 

                except Exception as e:
                    QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
                    traceback.print_exc()
                    if self.undo_stack:
                        try: self.undo_action_m()
                        except: pass
                    

            def adjust_elements_for_padding(self, padding_left, padding_top):
                """
                Adjusts coordinates of all markers and custom shapes for added padding.
                Also adjusts the _marker_shift_added values.
                This function is called BEFORE the image is actually padded.
                """

                # Adjust standard markers (Y for L/R, X for Top)
                self.left_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'left_markers', [])]
                self.right_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'right_markers', [])]
                self.top_markers = [(x_pos_img + padding_left, label) for x_pos_img, label in getattr(self, 'top_markers', [])]

                # Adjust custom markers (both X and Y)
                new_custom_markers = []
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        m_list = list(marker_data) # Ensure mutable
                        m_list[0] = float(m_list[0]) + padding_left  # Adjust X
                        m_list[1] = float(m_list[1]) + padding_top   # Adjust Y
                        new_custom_markers.append(m_list)
                    except (IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom marker during padding adjustment: {marker_data}")
                         new_custom_markers.append(marker_data) # Keep original if error
                self.custom_markers = new_custom_markers

                # Adjust Custom Shapes
                new_custom_shapes = []
                for shape_data_orig in getattr(self, "custom_shapes", []):
                    shape_data = dict(shape_data_orig) # Work on a copy
                    try:
                        shape_type = shape_data.get('type')
                        if shape_type == 'line':
                            sx, sy = shape_data['start']
                            ex, ey = shape_data['end']
                            shape_data['start'] = (float(sx) + padding_left, float(sy) + padding_top)
                            shape_data['end'] = (float(ex) + padding_left, float(ey) + padding_top)
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            shape_data['rect'] = (float(x) + padding_left, float(y) + padding_top, w, h)
                        # Add other shape types here if needed
                        new_custom_shapes.append(shape_data)
                    except (KeyError, IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom shape during padding adjustment: {shape_data_orig}")
                         new_custom_shapes.append(shape_data_orig)
                self.custom_shapes = new_custom_shapes

                # Adjust the absolute marker shift variables (these are in native image pixels)
                self.left_marker_shift_added += padding_left
                self.right_marker_shift_added += padding_left
                self.top_marker_shift_added += padding_top      
                self._update_overlay_slider_ranges
                
            
            def update_left_padding(self):
                # Update left padding when slider value changes
                self.left_marker_shift_added = self.left_padding_slider.value()
                self.update_live_view()

            def update_right_padding(self):
                # Update right padding when slider value changes
                new_value = self.right_padding_slider.value()
                # --- Add check to prevent redundant updates ---
                if new_value != self.right_marker_shift_added:
                    self.right_marker_shift_added = new_value
                    self.update_live_view()
                
            def update_top_padding(self):
                # Update top padding when slider value changes
                self.top_marker_shift_added = self.top_padding_slider.value()
                self.update_live_view()

            def update_live_view(self):
                # --- Initial check for a valid base image ---
                if hasattr(self, 'live_view_label') and hasattr(self, 'pan_left_action'): # Check if attributes exist
                    enable_pan_actions = self.live_view_label.zoom_level > 1.0
                    self.pan_left_action.setEnabled(enable_pan_actions)
                    self.pan_right_action.setEnabled(enable_pan_actions)
                    self.pan_up_action.setEnabled(enable_pan_actions)
                    self.pan_down_action.setEnabled(enable_pan_actions)
                if not self.image or self.image.isNull(): 
                    if hasattr(self, 'live_view_label'):
                        self.live_view_label.clear()
                        self.live_view_label.update() # Ensure paintEvent clears overlays if needed
                    # Disable UI elements that depend on an image
                    if hasattr(self, 'predict_button'): self.predict_button.setEnabled(False)
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(False)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(False)
                    # ... (potentially other UI elements) ...
                    return # Exit if no valid image
                else:
                    # Enable UI elements if image is valid
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(True)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(True)
                    if hasattr(self, 'predict_button'):
                        left_m = getattr(self, 'left_markers', [])
                        right_m = getattr(self, 'right_markers', [])
                        self.predict_button.setEnabled(bool(left_m or right_m))

                # --- Capture current pan and zoom state from LiveViewLabel ---
                # This ensures the entire render cycle uses a consistent state.
                current_zoom_level = 1.0
                current_pan_offset = QPointF(0, 0)
                if hasattr(self, 'live_view_label'): # Ensure live_view_label exists
                    current_zoom_level = self.live_view_label.zoom_level
                    current_pan_offset = QPointF(self.live_view_label.pan_offset) # Create a copy

                # --- Define rendering parameters ---
                render_scale = 3 # For high-resolution intermediate canvas
                try: 
                    view_width = self.live_view_label.width()
                    view_height = self.live_view_label.height()
                    if view_width <= 0: view_width = 600 # Fallback width
                    if view_height <= 0: view_height = 400 # Fallback height
                    render_width = view_width * render_scale
                    render_height = view_height * render_scale
                except AttributeError: # If live_view_label doesn't exist yet (early call)
                     render_width = 1800 # Default if view not ready
                     render_height = 1200
                except Exception: # Catch any other error during dimension access
                     render_width = 1800
                     render_height = 1200
                
                # --- Prepare image for transformations (Start with current self.image) ---
                if not self.image or self.image.isNull(): # Re-check, though earlier check should catch this
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                image_to_transform = self.image.copy() # Work on a copy

                # --- Apply Rotation (if UI elements exist) ---
                orientation = 0.0
                if hasattr(self, 'orientation_slider') and self.orientation_slider:
                    orientation = float(self.orientation_slider.value() / 20)
                    # Update the label text only if the label exists
                    if hasattr(self, 'orientation_label') and self.orientation_label:
                        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

                    if abs(orientation) > 0.01: 
                         if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                             transform_rotate = QTransform()
                             w_rot, h_rot = image_to_transform.width(), image_to_transform.height()
                             transform_rotate.translate(w_rot / 2.0, h_rot / 2.0) # Use float division
                             transform_rotate.rotate(orientation)
                             transform_rotate.translate(-w_rot / 2.0, -h_rot / 2.0)
                             temp_rotated = image_to_transform.transformed(transform_rotate, Qt.SmoothTransformation)
                             if not temp_rotated.isNull(): 
                                 image_to_transform = temp_rotated
                             else: 
                                 print("Warning: Rotation transform resulted in an invalid image.")
                
                # --- Apply Skew / Taper (if UI elements exist) ---
                taper_value = 0.0
                if hasattr(self, 'taper_skew_slider') and self.taper_skew_slider:
                    taper_value = self.taper_skew_slider.value() / 100.0
                    if hasattr(self, 'taper_skew_label') and self.taper_skew_label:
                        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

                    if abs(taper_value) > 0.01: 
                        if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                            width = image_to_transform.width(); height = image_to_transform.height()
                            source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
                            destination_corners = QPolygonF(source_corners)
                            if taper_value > 0:
                                destination_corners[0].setX(width * taper_value / 2.0)
                                destination_corners[1].setX(width * (1 - taper_value / 2.0))
                            elif taper_value < 0:
                                destination_corners[2].setX(width * (-taper_value / 2.0))
                                destination_corners[3].setX(width * (1 + taper_value / 2.0))
                            transform_skew = QTransform()
                            if QTransform.quadToQuad(source_corners, destination_corners, transform_skew):
                                temp_skewed = image_to_transform.transformed(transform_skew, Qt.SmoothTransformation)
                                if not temp_skewed.isNull(): 
                                    image_to_transform = temp_skewed
                                else: 
                                    print("Warning: Skew transform resulted in an invalid image.")
                            else: 
                                print("Warning: Failed to calculate skew transformation matrix.")

                # --- Final check on image_to_transform before proceeding ---
                if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    print("Error: image_to_transform became invalid before scaling for render.")
                    return

                # --- Scale the final transformed image for rendering on high-res canvas ---
                scaled_image_for_render_canvas = image_to_transform.scaled(
                    render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                if scaled_image_for_render_canvas.isNull():
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    print("Error: Final scaling for high-res render_canvas failed.")
                    return

                # --- Create high-resolution intermediate canvas ---
                canvas_format = QImage.Format_ARGB32_Premultiplied if image_to_transform.hasAlphaChannel() else QImage.Format_RGB888
                render_canvas = QImage(render_width, render_height, canvas_format)
                if render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to create high-res render_canvas QImage.")
                     return
                render_canvas.fill(Qt.white if canvas_format == QImage.Format_RGB888 else Qt.transparent)

                # Get crop offsets (these are 0 if image is not from a crop operation or crop was reset)
                current_crop_offset_x = getattr(self, 'crop_offset_x', 0)
                current_crop_offset_y = getattr(self, 'crop_offset_y', 0)

                # Render base image content (transformed image, raster overlays, guides) onto render_canvas
                # This function DOES NOT use pan/zoom itself.
                self.render_image_on_canvas(render_canvas, scaled_image_for_render_canvas,
                                            x_start=current_crop_offset_x, 
                                            y_start=current_crop_offset_y, 
                                            render_scale=render_scale,
                                            draw_guides=True) # Guides are part of the "view" rendering

                # --- Scale high-res canvas down for display on LiveViewLabel ---
                if render_canvas.isNull(): # Should not happen if previous checks passed
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                pixmap_from_render_canvas = QPixmap.fromImage(render_canvas)
                if pixmap_from_render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to create QPixmap from render_canvas.")
                     return

                # This is the pixmap representing the 100% zoom, unpanned view, scaled to fit the label
                scaled_pixmap_for_label_fit = pixmap_from_render_canvas.scaled(
                    self.live_view_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation # Use current label size
                )
                if scaled_pixmap_for_label_fit.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     print("Error: Failed to scale final pixmap for label display.")
                     return

                # This will be the pixmap ultimately set on the label
                final_pixmap_to_set_on_label = scaled_pixmap_for_label_fit

                # --- Apply Zoom and Pan for display if zoom_level is not 1.0 ---
                # Uses the locally captured current_zoom_level and current_pan_offset
                if current_zoom_level != 1.0:
                    # Target pixmap for zoomed/panned view, same size as the label
                    zoomed_display_pixmap = QPixmap(self.live_view_label.size()) 
                    if zoomed_display_pixmap.isNull():
                        print("Error: Failed to create zoomed_display_pixmap.")
                        # Fallback to showing the unzoomed pixmap
                    else:
                        zoomed_display_pixmap.fill(Qt.transparent) 
                        zoom_painter = QPainter(zoomed_display_pixmap)
                        if not zoom_painter.isActive():
                            print("Error: Failed to create QPainter for zooming in update_live_view.")
                        else:
                            zoom_painter.translate(current_pan_offset) # Apply pan
                            zoom_painter.scale(current_zoom_level, current_zoom_level) # Apply zoom
                            # Draw the 100% view (scaled_pixmap_for_label_fit) onto the transformed painter
                            zoom_painter.drawPixmap(0, 0, scaled_pixmap_for_label_fit) 
                            zoom_painter.end()

                            if zoomed_display_pixmap.isNull(): # Check after drawing
                                print("Error: zoomed_display_pixmap became invalid after drawing.")
                            else:
                                 final_pixmap_to_set_on_label = zoomed_display_pixmap # Use the panned/zoomed version
                
                # --- Set the final pixmap on LiveViewLabel ---
                if final_pixmap_to_set_on_label.isNull():
                    self.live_view_label.clear()
                    print("Error: final_pixmap_to_set_on_label is invalid before setting.")
                else:
                     self.live_view_label.setPixmap(final_pixmap_to_set_on_label)

                # Trigger LiveViewLabel.paintEvent to draw vector overlays (markers, shapes, grid, previews)
                # on top of the newly set base pixmap.
                self.live_view_label.update()
                
            def _update_overlay_slider_ranges(self):
                """
                Updates the ranges of overlay position sliders.
                The ranges are based on the NATIVE dimensions of the current self.image.
                Slider values will represent native pixel offsets.
                """
                native_width = 1000  # Fallback
                native_height = 1000 # Fallback
        
                if self.image and not self.image.isNull() and self.image.width() > 0 and self.image.height() > 0:
                    native_width = self.image.width()
                    native_height = self.image.height()
        
                # Allow positioning overlay's top-left from -1*dimension to +2*dimension
                # relative to the main image's top-left in NATIVE pixels.
                x_range_min = -native_width
                x_range_max = native_width * 2
                y_range_min = -native_height
                y_range_max = native_height * 2
                
                
                sliders_to_update = [
                    (getattr(self, 'image1_left_slider', None), x_range_min, x_range_max, getattr(self, 'image1_position', (0,0))[0]),
                    (getattr(self, 'image1_top_slider', None), y_range_min, y_range_max, getattr(self, 'image1_position', (0,0))[1]),
                    (getattr(self, 'image2_left_slider', None), x_range_min, x_range_max, getattr(self, 'image2_position', (0,0))[0]),
                    (getattr(self, 'image2_top_slider', None), y_range_min, y_range_max, getattr(self, 'image2_position', (0,0))[1]),
                ]
        
                for slider, min_val, max_val, current_pos_val in sliders_to_update:
                    if slider:
                        slider.blockSignals(True)
                        slider.setRange(min_val, max_val)
                        # Try to set the slider to the current stored native position, clamped by new range
                        clamped_val = max(min_val, min(current_pos_val, max_val))
                        slider.setValue(clamped_val)
                        slider.blockSignals(False)
            
            def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
                painter = QPainter(canvas)
                # Calculate centering offsets for the scaled_image on the canvas
                x_offset = (canvas.width() - scaled_image.width()) // 2
                y_offset = (canvas.height() - scaled_image.height()) // 2
                
                self.x_offset_s=x_offset # Store for other potential uses if needed
                self.y_offset_s=y_offset
            
                # Draw the base image
                painter.drawImage(x_offset, y_offset, scaled_image)
                
                # --- NO CUSTOM SHAPES DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent

                # --- NO CUSTOM MARKERS DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent

                # --- NO STANDARD L/R/TOP MARKERS DRAWN HERE ANYMORE ---
                # They will be drawn by LiveViewLabel.paintEvent
                
                # Native dimensions of the image that scaled_image was derived from
                # This assumes self.image is the source for scaled_image
                native_width = self.image.width() if self.image and self.image.width() > 0 else 1
                native_height = self.image.height() if self.image and self.image.height() > 0 else 1

                # Scale factor mapping native coordinates to the scaled_image dimensions
                scale_native_to_scaled_x = scaled_image.width() / native_width
                scale_native_to_scaled_y = scaled_image.height() / native_height

                # --- Draw Image 1 Overlay (Combined Image Feature) ---
                if hasattr(self, 'image1_original') and hasattr(self, 'image1_position') and not self.image1_original.isNull():
                    try:
                        scale_factor_overlay = self.image1_resize_slider.value() / 100.0
                        target_overlay_w = int(self.image1_original.width() * scale_factor_overlay)
                        target_overlay_h = int(self.image1_original.height() * scale_factor_overlay)
                        self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())

                        if target_overlay_w > 0 and target_overlay_h > 0:
                            # Resize the *original* overlay image
                            resized_overlay1 = self.image1_original.scaled(
                                target_overlay_w, target_overlay_h,
                                Qt.KeepAspectRatio, Qt.SmoothTransformation
                            )
                            if not resized_overlay1.isNull():
                                # Get NATIVE offsets stored
                                native_offset_x1 = self.image1_position[0]
                                native_offset_y1 = self.image1_position[1]

                                # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                                # Scale the NATIVE offset according to how the base image was scaled
                                # Then add the centering offset of the base image on the canvas
                                canvas_draw_x = x_offset + native_offset_x1 * scale_native_to_scaled_x
                                canvas_draw_y = y_offset + native_offset_y1 * scale_native_to_scaled_y

                                painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay1)
                    except Exception as e:
                        print(f"Error rendering overlay image 1: {e}")


                # --- Draw Image 2 Overlay ---
                if hasattr(self, 'image2_original') and hasattr(self, 'image2_position') and not self.image2_original.isNull():
                    try:
                        scale_factor_overlay = self.image2_resize_slider.value() / 100.0
                        self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                        target_overlay_w = int(self.image2_original.width() * scale_factor_overlay)
                        target_overlay_h = int(self.image2_original.height() * scale_factor_overlay)

                        if target_overlay_w > 0 and target_overlay_h > 0:
                            resized_overlay2 = self.image2_original.scaled(
                                target_overlay_w, target_overlay_h,
                                Qt.KeepAspectRatio, Qt.SmoothTransformation
                            )
                            if not resized_overlay2.isNull():
                                native_offset_x2 = self.image2_position[0]
                                native_offset_y2 = self.image2_position[1]

                                # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                                canvas_draw_x = x_offset + native_offset_x2 * scale_native_to_scaled_x
                                canvas_draw_y = y_offset + native_offset_y2 * scale_native_to_scaled_y

                                painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay2)
                    except Exception as e:
                        print(f"Error rendering overlay image 2: {e}")
                
            
            
                # Draw guide lines (These are view-dependent, not content, so okay here)
                if draw_guides and hasattr(self, 'show_guides_checkbox') and self.show_guides_checkbox.isChecked():
                    pen_guides = QPen(Qt.red, 2 * render_scale) # Keep guides scaled by render_scale for visibility on hi-res canvas
                    painter.setPen(pen_guides)
                    center_x_canvas = canvas.width() // 2
                    center_y_canvas = canvas.height() // 2
                    painter.drawLine(center_x_canvas, 0, center_x_canvas, canvas.height())
                    painter.drawLine(0, center_y_canvas, canvas.width(), center_y_canvas)
            
                painter.end()


             
            def crop_image(self):
                """Function to crop the current image."""
                if not self.image:
                    return None
            
                # Get crop percentage from sliders
                x_start_percent = 0
                x_end_percent = 100
                y_start_percent = 0
                y_end_percent = 100
            
                # Calculate crop boundaries
                x_start = int(self.image.width() * x_start_percent)
                x_end = int(self.image.width() * x_end_percent)
                y_start = int(self.image.height() * y_start_percent)
                y_end = int(self.image.height() * y_end_percent)
            
                # Ensure cropping is valid
                if x_start >= x_end or y_start >= y_end:
                    QMessageBox.warning(self, "Warning", "Invalid cropping values.")
                    return None
            
                # Crop the image
                cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
                return cropped_image
            
            # Modify align_image and update_crop to preserve settings
            def reset_align_image(self):
                self.orientation_slider.setValue(0)
                try:
                    if self.image: # Check if image exists after cropping
                        self._update_preview_label_size()
                except Exception as e:
                    # Fallback size?
                    self._update_preview_label_size()


                self.update_live_view() # Final update with corrected markers and image
                
                
            def align_image(self):
                """
                Aligns (rotates) the main self.image based on the orientation slider value.
                This modifies the high-resolution self.image directly.
                """
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Rotation Error", "No image loaded to rotate.")
                    return

                # Get the orientation value from the slider
                angle = float(self.orientation_slider.value() / 20.0) # Use float division

                if abs(angle) < 0.01: # Threshold to avoid unnecessary processing for tiny angles
                    # print("Debug: Rotation angle negligible, skipping.") # Optional debug
                    # Reset slider if needed, but don't modify image
                    self.orientation_slider.setValue(0)
                    return

                # --- Save state BEFORE modifying the image ---
                self.save_state()

                # Disable temporary guides before applying permanent change
                self.draw_guides = False
                if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)
                # Keep grid settings as they are independent of rotation


                try:
                    # Perform rotation calculation
                    transform = QTransform()
                    # Rotate around the center of the *current* image
                    current_width = self.image.width()
                    current_height = self.image.height()
                    transform.translate(current_width / 2.0, current_height / 2.0)
                    transform.rotate(angle)
                    transform.translate(-current_width / 2.0, -current_height / 2.0)

                    # Apply the transformation to get the high-resolution rotated image
                    # Qt.SmoothTransformation provides better quality
                    # The resulting image might be larger to accommodate rotated corners
                    rotated_image_high_res = self.image.transformed(transform, Qt.SmoothTransformation)

                    if rotated_image_high_res.isNull():
                        raise ValueError("Image transformation resulted in an invalid (null) image.")

                    # --- Directly update the main image object ---
                    self.image = rotated_image_high_res
                    self.is_modified = True # Mark as modified

                    # --- Update backup images to reflect the new state ---
                    # Rotation invalidates previous padding
                    self.image_before_padding = None # Or self.image.copy() if padding should persist conceptually
                    self.image_padded = False
                    # Contrast backups should also reflect the rotated image
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()

                    # --- Reset the orientation slider ---
                    self.orientation_slider.setValue(0)

                    # --- Update UI elements based on the new image dimensions ---
                    # This will adjust the preview label size and marker slider ranges
                    self._update_preview_label_size()
                    self._update_marker_slider_ranges() # Important after potential size change
                    self._update_status_bar()

                    # --- Let update_live_view handle the rendering ---
                    # This will correctly scale the new high-res self.image for display
                    # and draw markers/overlays relative to the new image.
                    self.update_live_view()

                except Exception as e:
                    QMessageBox.critical(self, "Rotation Error", f"Failed to rotate image: {e}")
                    traceback.print_exc()
                    # Attempt to revert state if possible (though save_state was called)
                    # self.undo_action_m() # Reverting might be complex if backups were already updated
                
                self.crop_x_start_slider.setEnabled(False) 
                self.crop_x_end_slider.setEnabled(False) 
                self.crop_y_start_slider.setEnabled(False) 
                self.crop_y_end_slider.setEnabled(False) 
            
            def _update_marker_slider_ranges(self):
                if not self.image or self.image.isNull() or not self.live_view_label:
                    min_abs, max_abs_w, max_abs_h = -100, 1000, 800 # Default absolute pixel values
                else:
                    try:
                        # Ranges are based on the NATIVE dimensions of the CURRENT self.image
                        native_img_width = self.image.width()
                        native_img_height = self.image.height()

                        margin = 100 # Allow markers to be set slightly outside the image bounds

                        min_abs_x = -margin
                        max_abs_x = native_img_width + margin
                        min_abs_y = -margin
                        max_abs_y = native_img_height + margin

                        # Fallbacks if dimensions are invalid
                        if native_img_width <= 0: max_abs_x = 1000; min_abs_x = -100
                        if native_img_height <= 0: max_abs_y = 800; min_abs_y = -100
                        
                        max_abs_w = max_abs_x # For horizontal sliders (left/right markers)
                        max_abs_h = max_abs_y # For Y-offsets (top markers)
                        min_abs = min(min_abs_x, min_abs_y) # Common minimum

                    except Exception as e:
                        print(f"Warning: Error calculating slider ranges: {e}. Using defaults.")
                        min_abs, max_abs_w, max_abs_h = -100, 1000, 800

                # Store the calculated ranges internally (optional, but good for reference)
                self.left_slider_range = [min_abs, max_abs_w]
                self.right_slider_range = [min_abs, max_abs_w]
                self.top_slider_range = [min_abs, max_abs_h]

                # Update sliders: setRange, then re-apply current value (which might get clamped)
                sliders_and_ranges = [
                    (getattr(self, 'left_padding_slider', None), self.left_slider_range),
                    (getattr(self, 'right_padding_slider', None), self.right_slider_range),
                    (getattr(self, 'top_padding_slider', None), self.top_slider_range)
                ]

                for slider, new_range in sliders_and_ranges:
                    if slider:
                        current_val = slider.value() # Get current value BEFORE changing range
                        slider.blockSignals(True)
                        slider.setRange(new_range[0], new_range[1])
                        slider.setValue(current_val) # Re-apply; Qt will clamp if current_val is outside new_range
                        slider.blockSignals(False)
                    
                
                    
            def update_crop(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Crop Error", "No image loaded to crop.")
                    return

                if not self.crop_rectangle_coords:
                    # ... (message box logic) ...
                    return

                try:
                    img_x_intent, img_y_intent, img_w_intent, img_h_intent = self.crop_rectangle_coords
                    original_image_width_before_crop = self.image.width()
                    original_image_height_before_crop = self.image.height()

                    crop_x_start = max(0, int(round(img_x_intent)))
                    crop_y_start = max(0, int(round(img_y_intent)))
                    crop_width = max(1, min(int(round(img_w_intent)), original_image_width_before_crop - crop_x_start))
                    crop_height = max(1, min(int(round(img_h_intent)), original_image_height_before_crop - crop_y_start))
                                        
                    if crop_width <= 0 or crop_height <= 0:
                         QMessageBox.warning(self, "Crop Error", "Calculated crop area has invalid dimensions. Aborting.")
                         self.crop_rectangle_coords = None; self.live_view_label.clear_crop_preview(); self.cancel_rectangle_crop_mode() 
                         return

                    self.save_state()

                    # --- 1. Adjust marker and shape coordinates ---
                    # (This part from your previous good version, ensuring elements are kept if their
                    #  anchor/significant part is within the new bounds, and coords are made relative)
                    new_left_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'left_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.left_markers = new_left_markers
                    new_right_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'right_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.right_markers = new_right_markers
                    new_top_markers = [(x_old - crop_x_start, label) for x_old, label in getattr(self, 'top_markers', []) if crop_x_start <= x_old < crop_x_start + crop_width]
                    self.top_markers = new_top_markers
                    
                    new_custom_markers = [] # ... (full adjustment logic as before) ...
                    if hasattr(self, "custom_markers"):
                        for marker_data in self.custom_markers:
                            try:
                                m_list = list(marker_data); x_old, y_old = float(m_list[0]), float(m_list[1])
                                x_new, y_new = x_old - crop_x_start, y_old - crop_y_start
                                if 0 <= x_new < crop_width and 0 <= y_new < crop_height:
                                    m_list[0] = x_new; m_list[1] = y_new
                                    new_custom_markers.append(m_list)
                            except: pass # Skip malformed
                    self.custom_markers = new_custom_markers

                    new_custom_shapes = [] # ... (full adjustment logic as before, including clipping for rects) ...
                    if hasattr(self, "custom_shapes"):
                        for shape_data_orig in self.custom_shapes:
                            shape_data = dict(shape_data_orig); adjusted_shape_data = None
                            try:
                                stype = shape_data.get('type')
                                if stype == 'line':
                                    sx_old, sy_old = map(float, shape_data['start']); ex_old, ey_old = map(float, shape_data['end'])
                                    adjusted_shape_data = shape_data.copy()
                                    adjusted_shape_data['start'] = (sx_old - crop_x_start, sy_old - crop_y_start)
                                    adjusted_shape_data['end'] = (ex_old - crop_x_start, ey_old - crop_y_start)
                                elif stype == 'rectangle':
                                    rx_old, ry_old, rw_old, rh_old = map(float, shape_data['rect'])
                                    overlap_x1 = max(crop_x_start, rx_old); overlap_y1 = max(crop_y_start, ry_old)
                                    overlap_x2 = min(crop_x_start + crop_width, rx_old + rw_old)
                                    overlap_y2 = min(crop_y_start + crop_height, ry_old + rh_old)
                                    if overlap_x2 > overlap_x1 and overlap_y2 > overlap_y1:
                                        adjusted_shape_data = shape_data.copy()
                                        adjusted_shape_data['rect'] = (overlap_x1 - crop_x_start, overlap_y1 - crop_y_start, overlap_x2 - overlap_x1, overlap_y2 - overlap_y1)
                                if adjusted_shape_data: new_custom_shapes.append(adjusted_shape_data)
                            except: pass # Skip malformed
                    self.custom_shapes = new_custom_shapes
                    
                    # --- 2. Adjust the _marker_shift_added variables ---
                    # These are absolute X/Y positions in the *pre-crop* image.
                    # Make them relative to the new cropped image's origin.
                    self.left_marker_shift_added = self.left_marker_shift_added - crop_x_start
                    self.right_marker_shift_added = self.right_marker_shift_added - crop_x_start
                    self.top_marker_shift_added = self.top_marker_shift_added - crop_y_start

                    # Clamp them to be valid within the new cropped image dimensions [0, new_dim -1]
                    self.left_marker_shift_added = max(0, min(self.left_marker_shift_added, crop_width - 1 if crop_width > 0 else 0))
                    self.right_marker_shift_added = max(0, min(self.right_marker_shift_added, crop_width - 1 if crop_width > 0 else 0))
                    self.top_marker_shift_added = max(0, min(self.top_marker_shift_added, crop_height - 1 if crop_height > 0 else 0))

                    # --- 3. Perform the actual crop on self.image ---
                    cropped_qimage = self.image.copy(crop_x_start, crop_y_start, crop_width, crop_height)
                    if cropped_qimage.isNull():
                        raise ValueError("QImage.copy failed for cropping.")

                    self.image = cropped_qimage # self.image is now the cropped image
                    self.is_modified = True
                    self.image_before_padding = self.image.copy() 
                    self.image_contrasted = self.image.copy()      
                    self.image_before_contrast = self.image.copy() 
                    self.image_padded = False 

                    self.crop_rectangle_coords = None 
                    self.live_view_label.clear_crop_preview() 
                    self.cancel_rectangle_crop_mode() 

                    # --- 4. Update UI ---
                    self._update_preview_label_size() 
                    self._update_status_bar()         
                    
                    # Update slider RANGES based on new image dimensions.
                    self._update_marker_slider_ranges() 
                    
                    # NOW, set the slider values to the *transformed and clamped* _marker_shift_added values.
                    # This ensures the UI accurately reflects the state.
                    sliders_to_set = [
                        (getattr(self, 'left_padding_slider', None), self.left_marker_shift_added),
                        (getattr(self, 'right_padding_slider', None), self.right_marker_shift_added),
                        (getattr(self, 'top_padding_slider', None), self.top_marker_shift_added)
                    ]
                    for slider, value_to_set in sliders_to_set:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(value_to_set) # This will be clamped by the new range if needed
                            slider.blockSignals(False)
                            # Re-sync internal var with actual slider value after potential clamping by setValue
                            if slider == self.left_padding_slider: self.left_marker_shift_added = slider.value()
                            elif slider == self.right_padding_slider: self.right_marker_shift_added = slider.value()
                            elif slider == self.top_padding_slider: self.top_marker_shift_added = slider.value()
                    
                    self.update_live_view()

                except Exception as e:
                    # ... (error handling) ...
                    QMessageBox.critical(self, "Crop Error", f"An error occurred during cropping: {e}")
                    traceback.print_exc()
                    if self.undo_stack: 
                        try: self.undo_action_m() 
                        except: pass 
                    self.crop_rectangle_coords = None
                    self.live_view_label.clear_crop_preview()
                    self.cancel_rectangle_crop_mode()
                
            def update_skew(self):
                self.save_state()
                taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

                width = self.image.width()
                height = self.image.height()
            
                # Define corner points for perspective transformation
                source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
            
                # Initialize destination corners as a copy of source corners
                destination_corners = QPolygonF(source_corners)
            
                # Adjust perspective based on taper value
                if taper_value > 0:
                    # Narrower at the top, wider at the bottom
                    destination_corners[0].setX(width * taper_value / 2)  # Top-left
                    destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
                elif taper_value < 0:
                    # Wider at the top, narrower at the bottom
                    destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
                    destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
            
                # Create a perspective transformation using quadToQuad
                transform = QTransform()
                if not QTransform.quadToQuad(source_corners, destination_corners, transform):
                    return
            
                # Apply the transformation
                # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

         
                self.image = self.image.transformed(transform, Qt.SmoothTransformation)
                self.taper_skew_slider.setValue(0)

            
                
            def save_image(self):
                self.draw_guides = False 
                if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)
                
                if not self.image or self.image.isNull(): 
                     QMessageBox.warning(self, "Error", "No image data to save.")
                     return False

                options = QFileDialog.Options()
                suggested_name = ""
                if self.image_path:
                    base = os.path.splitext(os.path.basename(self.image_path))[0]
                    base = base.replace("_original", "").replace("_modified", "")
                    suggested_name = f"{base}"
                else:
                    suggested_name = "untitled_image"
                save_dir = os.path.dirname(self.image_path) if self.image_path else ""

                base_save_path, selected_filter = QFileDialog.getSaveFileName(
                    self, "Save Image Base Name", os.path.join(save_dir, suggested_name),
                    "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
                    options=options
                )

                if not base_save_path:
                    return False 

                base_name_nosuffix = os.path.splitext(base_save_path)[0].replace("_original", "").replace("_modified", "")
                suffix = ".png" 
                if "tif" in selected_filter.lower(): suffix = ".tif"
                elif "jpg" in selected_filter.lower() or "jpeg" in selected_filter.lower(): suffix = ".jpg"
                elif "bmp" in selected_filter.lower(): suffix = ".bmp"
                else: 
                    user_suffix = os.path.splitext(base_save_path)[1]
                    if user_suffix: suffix = user_suffix
                
                original_save_path = f"{base_name_nosuffix}_original{suffix}"
                modified_save_path = f"{base_name_nosuffix}_modified{suffix}"
                config_save_path = f"{base_name_nosuffix}_config.txt"

                img_to_save_as_original = self.image.copy() 
                if img_to_save_as_original and not img_to_save_as_original.isNull():
                    save_format_orig_str = suffix.replace(".", "").upper()
                    if save_format_orig_str == "TIF": save_format_orig_str = "TIFF"
                    elif save_format_orig_str == "JPEG": save_format_orig_str = "JPG"
                    quality_orig = 95 if save_format_orig_str in ["JPG", "JPEG"] else -1
                    if save_format_orig_str in ["JPG", "JPEG", "BMP"] and img_to_save_as_original.hasAlphaChannel():
                        temp_opaque_canvas = QImage(img_to_save_as_original.size(), QImage.Format_RGB888) 
                        temp_opaque_canvas.fill(Qt.white)
                        painter_opaque = QPainter(temp_opaque_canvas)
                        painter_opaque.drawImage(0,0, img_to_save_as_original)
                        painter_opaque.end()
                        if not temp_opaque_canvas.save(original_save_path, format=save_format_orig_str if save_format_orig_str else None, quality=quality_orig):
                            QMessageBox.warning(self, "Error", f"Failed to save original (composited) image to {original_save_path}.")
                    else: 
                        if not img_to_save_as_original.save(original_save_path, format=save_format_orig_str if save_format_orig_str else None, quality=quality_orig):
                            QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
                else:
                     QMessageBox.warning(self, "Error", "Current working image (for '_original') is missing or invalid.")

                render_scale = 3 
                if not self.image or self.image.isNull() or self.image.width() <= 0 or self.image.height() <= 0:
                    QMessageBox.critical(self, "Save Error", "Cannot create canvas for modified image; base image invalid.")
                    return False
                    
                high_res_canvas_width = self.image.width() * render_scale
                high_res_canvas_height = self.image.height() * render_scale
                
                save_format_mod_str = suffix.replace(".", "").upper()
                canvas_format_mod = QImage.Format_ARGB32_Premultiplied 
                fill_color_mod = Qt.transparent 
                if save_format_mod_str in ["JPG", "JPEG", "BMP"]:
                    fill_color_mod = Qt.white 

                high_res_canvas_mod = QImage(high_res_canvas_width, high_res_canvas_height, canvas_format_mod)
                if high_res_canvas_mod.isNull():
                    QMessageBox.critical(self, "Save Error", "Failed to create high-resolution canvas for modified image.")
                    return False
                high_res_canvas_mod.fill(fill_color_mod)

                scaled_image_mod = self.image.scaled(
                    high_res_canvas_width, high_res_canvas_height,
                    Qt.IgnoreAspectRatio, 
                    Qt.SmoothTransformation)
                if scaled_image_mod.isNull():
                    QMessageBox.critical(self, "Save Error", "Failed to scale current image for saving modified version.")
                    return False

                painter_mod = QPainter(high_res_canvas_mod)
                if not painter_mod.isActive(): 
                    QMessageBox.critical(self, "Save Error", "Failed to create painter for modified image canvas.")
                    return False
                painter_mod.drawImage(0, 0, scaled_image_mod) 
                
                painter_mod.setRenderHint(QPainter.Antialiasing, True)
                painter_mod.setRenderHint(QPainter.TextAntialiasing, True)

                # --- Define Local Coordinate Mapping Helper for Modified Canvas ---
                # This maps coordinates from self.image (native) to high_res_canvas_mod
                img_w_current_save = self.image.width() if self.image.width() > 0 else 1
                img_h_current_save = self.image.height() if self.image.height() > 0 else 1
                
                # scaled_image_mod is self.image scaled directly to high_res_canvas_mod dimensions.
                # So, canvas_center_x/y_offset will be 0 as scaled_image_mod fills high_res_canvas_mod.
                canvas_center_x_offset_mod = (high_res_canvas_mod.width() - scaled_image_mod.width()) // 2 # Should be 0
                canvas_center_y_offset_mod = (high_res_canvas_mod.height() - scaled_image_mod.height()) // 2 # Should be 0

                # Scale factor from self.image coordinates to coordinates on scaled_image_mod (which is on high_res_canvas_mod)
                # This is simply render_scale if IgnoreAspectRatio was used for scaling.
                scale_factor_x_img_to_canvas_mod = scaled_image_mod.width() / img_w_current_save # Should be render_scale
                scale_factor_y_img_to_canvas_mod = scaled_image_mod.height() / img_h_current_save # Should be render_scale

                def map_img_coords_to_save_canvas(img_x, img_y): # Specific name for this function's scope
                    # Since scaled_image_mod fills high_res_canvas_mod, mapping is direct scaling
                    canvas_x = img_x * scale_factor_x_img_to_canvas_mod 
                    canvas_y = img_y * scale_factor_y_img_to_canvas_mod
                    return QPointF(canvas_x, canvas_y)
                # --- End Local Coordinate Mapping Helper ---

                # A. Draw Standard L/R/Top Markers
                std_font_size_on_canvas_save = int(self.font_size * render_scale) 
                std_marker_font_config_save = QFont(self.font_family, std_font_size_on_canvas_save)
                painter_mod.setFont(std_marker_font_config_save)
                painter_mod.setPen(self.font_color)
                font_metrics_std_mod_save = QFontMetrics(std_marker_font_config_save)
                text_height_std_mod_canvas_save = font_metrics_std_mod_save.height()
                y_offset_text_baseline_std_canvas_save = text_height_std_mod_canvas_save * 0.3 

                # Left Markers
                left_marker_offset_x_on_canvas_save = self.left_marker_shift_added * scale_factor_x_img_to_canvas_mod
                for y_pos_img, marker_text_val in self.left_markers:
                    anchor_on_canvas = map_img_coords_to_save_canvas(0, y_pos_img) 
                    text_to_draw = f"{marker_text_val} ⎯"
                    text_width_canvas = font_metrics_std_mod_save.horizontalAdvance(text_to_draw)
                    # No canvas_center_x_offset_mod needed here if scaled_image_mod fills the canvas
                    draw_x_c = left_marker_offset_x_on_canvas_save - text_width_canvas 
                    draw_y_c = anchor_on_canvas.y() + y_offset_text_baseline_std_canvas_save
                    painter_mod.drawText(QPointF(draw_x_c, draw_y_c), text_to_draw)
                
                # Right Markers
                right_marker_offset_x_on_canvas_save = self.right_marker_shift_added * scale_factor_x_img_to_canvas_mod
                for y_pos_img, marker_text_val in self.right_markers:
                    anchor_on_canvas = map_img_coords_to_save_canvas(0, y_pos_img)
                    text_to_draw = f"⎯ {marker_text_val}"
                    draw_x_c = right_marker_offset_x_on_canvas_save
                    draw_y_c = anchor_on_canvas.y() + y_offset_text_baseline_std_canvas_save
                    painter_mod.drawText(QPointF(draw_x_c, draw_y_c), text_to_draw)

                # Top Markers
                top_marker_offset_y_on_canvas_save = self.top_marker_shift_added * scale_factor_y_img_to_canvas_mod
                rotation_angle_save = self.font_rotation
                for x_pos_img, marker_text_val in self.top_markers:
                    anchor_on_canvas = map_img_coords_to_save_canvas(x_pos_img, 0) 
                    text_to_draw = str(marker_text_val)
                    painter_mod.save()
                    translate_x_c = anchor_on_canvas.x()
                    # No canvas_center_y_offset_mod if scaled_image_mod fills the canvas
                    translate_y_c = top_marker_offset_y_on_canvas_save + y_offset_text_baseline_std_canvas_save
                    painter_mod.translate(translate_x_c, translate_y_c)
                    painter_mod.rotate(rotation_angle_save)
                    painter_mod.drawText(QPointF(0, 0), text_to_draw) 
                    painter_mod.restore()

                # B. Draw Custom Markers
                for marker_data_list in getattr(self, "custom_markers", []):
                    try:
                        x_pos_img, y_pos_img, marker_text_str, qcolor_obj, \
                        font_family_str, font_size_int, is_bold, is_italic = marker_data_list
                        anchor_on_canvas = map_img_coords_to_save_canvas(x_pos_img, y_pos_img) # Use local helper
                        custom_font_size_on_canvas_save = int(font_size_int * render_scale)
                        custom_font_save = QFont(font_family_str, custom_font_size_on_canvas_save)
                        custom_font_save.setBold(is_bold)
                        custom_font_save.setItalic(is_italic)
                        painter_mod.setFont(custom_font_save)
                        current_color = QColor(qcolor_obj) if isinstance(qcolor_obj, QColor) else QColor(str(qcolor_obj))
                        if not current_color.isValid(): current_color = Qt.black
                        painter_mod.setPen(current_color)
                        font_metrics_custom_mod_save = QFontMetrics(custom_font_save)
                        text_bounding_rect_custom_save = font_metrics_custom_mod_save.boundingRect(marker_text_str)
                        draw_x_c = anchor_on_canvas.x() - (text_bounding_rect_custom_save.left() + text_bounding_rect_custom_save.width() / 2.0)
                        draw_y_c = anchor_on_canvas.y() - (text_bounding_rect_custom_save.top() + text_bounding_rect_custom_save.height() / 2.0)
                        painter_mod.drawText(QPointF(draw_x_c, draw_y_c), marker_text_str)
                    except Exception as e_cm_save:
                        print(f"Error drawing custom marker during save: {marker_data_list}, {e_cm_save}")

                # C. Draw Custom Shapes
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type = shape_data.get('type')
                        color = QColor(shape_data.get('color', '#000000'))
                        base_thickness_img_pixels = float(shape_data.get('thickness', 1.0))
                        thickness_on_canvas_save = max(1.0, base_thickness_img_pixels * scale_factor_x_img_to_canvas_mod) # Use consistent scale factor
                        pen = QPen(color)
                        pen.setWidthF(thickness_on_canvas_save) 
                        painter_mod.setPen(pen)
                        if shape_type == 'line':
                            start_img_coords = shape_data.get('start') 
                            end_img_coords = shape_data.get('end')   
                            if start_img_coords and end_img_coords:
                                start_canvas = map_img_coords_to_save_canvas(start_img_coords[0], start_img_coords[1])
                                end_canvas = map_img_coords_to_save_canvas(end_img_coords[0], end_img_coords[1])       
                                painter_mod.drawLine(start_canvas, end_canvas)
                        elif shape_type == 'rectangle':
                            rect_img_coords = shape_data.get('rect') 
                            if rect_img_coords:
                                x_img, y_img, w_img, h_img = rect_img_coords
                                top_left_canvas = map_img_coords_to_save_canvas(x_img, y_img)
                                w_on_canvas = w_img * scale_factor_x_img_to_canvas_mod # Scale width
                                h_on_canvas = h_img * scale_factor_y_img_to_canvas_mod # Scale height
                                painter_mod.drawRect(QRectF(top_left_canvas, QSizeF(w_on_canvas, h_on_canvas)))
                    except Exception as e_cs_save:
                         print(f"Error drawing custom shape during save: {shape_data}, {e_cs_save}")
                
                painter_mod.end() 

                quality_mod = 95 if save_format_mod_str in ["JPG", "JPEG"] else -1
                if not high_res_canvas_mod.save(modified_save_path, format=save_format_mod_str if save_format_mod_str else None, quality=quality_mod):
                    QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
                
                config_data = self.get_current_config()
                try:
                    with open(config_save_path, "w", encoding='utf-8') as config_file:
                        json.dump(config_data, config_file, indent=4)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
                    return False

                self.is_modified = False 
                QMessageBox.information(self, "Saved", f"Files saved successfully:\n- {os.path.basename(original_save_path)}\n- {os.path.basename(modified_save_path)}\n- {os.path.basename(config_save_path)}")
                self.setWindowTitle(f"{self.window_title}::{base_name_nosuffix}")
                return True
                    
            def save_image_svg(self):
                """Save the processed image along with markers, labels, and custom shapes in SVG format."""
                if not self.image or self.image.isNull(): # Check current image validity
                    QMessageBox.warning(self, "Warning", "No image to save.")
                    return

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Image as SVG for MS Word/Vector Editing", "", "SVG Files (*.svg)", options=options
                )

                if not file_path:
                    return

                if not file_path.lower().endswith(".svg"): # Ensure .svg extension
                    file_path += ".svg"

                # --- Determine Render/Canvas Dimensions ---
                # Use the current image dimensions directly as the base SVG size
                # This avoids scaling issues if the view label size is different.
                # Markers and shapes will be positioned relative to this size.
                svg_width = self.image.width()
                svg_height = self.image.height()

                if svg_width <= 0 or svg_height <= 0:
                     QMessageBox.warning(self, "Warning", "Invalid image dimensions for SVG.")
                     return

                # --- Create SVG Drawing Object ---
                dwg = svgwrite.Drawing(file_path, profile='tiny', size=(f"{svg_width}px", f"{svg_height}px"))
                # Set viewbox to match image dimensions for correct coordinate system
                dwg.viewbox(0, 0, svg_width, svg_height)

                # --- Embed the Base Image ---
                try:
                    # Convert the QImage to a base64-encoded PNG for embedding
                    buffer = QBuffer()
                    buffer.open(QBuffer.ReadWrite)
                    # Save the *current* self.image (which might be cropped/transformed)
                    if not self.image.save(buffer, "PNG"):
                        raise IOError("Failed to save image to PNG buffer for SVG.")
                    image_data = base64.b64encode(buffer.data()).decode('utf-8')
                    buffer.close()

                    # Embed the image at position (0, 0) with original dimensions
                    dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}",
                                      insert=(0, 0),
                                      size=(f"{svg_width}px", f"{svg_height}px")))
                except Exception as e:
                    QMessageBox.critical(self, "SVG Error", f"Failed to embed base image: {e}")
                    return # Stop if base image fails

                # --- Define marker/label font style for SVG ---
                # Use the standard marker font settings
                svg_font_family = self.font_family
                svg_font_size_px = f"{self.font_size}px" # Use pixels for SVG consistency
                svg_font_color = self.font_color.name() if self.font_color else "#000000"

                # Calculate horizontal offset for left/right markers based on font size
                # Use QFontMetrics for accurate width calculation
                try:
                     qfont_for_metrics = QFont(svg_font_family, self.font_size)
                     font_metrics = QFontMetrics(qfont_for_metrics)
                     # Use a representative character like 'm' or average width if needed
                     avg_char_width = font_metrics.averageCharWidth()
                     horizontal_offset = avg_char_width * 0.5 # Small offset from edge
                     vertical_offset_adjust = font_metrics.ascent() * 0.75 # Adjustment for vertical alignment
                except Exception:
                     horizontal_offset = 5 # Fallback offset
                     vertical_offset_adjust = self.font_size * 0.75 # Fallback adjustment


                # --- Add Left Markers ---
                left_marker_x_pos = self.left_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "left_markers", []):
                    final_text = f"{text}" # Remove the line "⎯" for cleaner SVG text
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(left_marker_x_pos - horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="end" # Align text right, ending at the x position
                        )
                    )

                # --- Add Right Markers ---
                right_marker_x_pos = self.right_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "right_markers", []):
                    final_text = f"{text}" # Remove the line "⎯"
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(right_marker_x_pos + horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="start" # Align text left, starting at the x position
                        )
                    )

                # --- Add Top Markers ---
                top_marker_y_pos = self.top_marker_shift_added # Use the absolute offset
                for x_pos, text in getattr(self, "top_markers", []):
                    dwg.add(
                        dwg.text(
                            text,
                            insert=(x_pos, top_marker_y_pos + vertical_offset_adjust), # Apply vertical adjust
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="middle", # Center text horizontally over the x position
                            # Apply rotation around the insertion point
                            transform=f"rotate({self.font_rotation}, {x_pos}, {top_marker_y_pos + vertical_offset_adjust})"
                        )
                    )

                # --- Add Custom Markers ---
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        # Default values for optional elements
                        is_bold = False
                        is_italic = False

                        # Unpack based on length for backward compatibility
                        if len(marker_tuple) == 8:
                            x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple
                        elif len(marker_tuple) == 6:
                            x_pos, y_pos, marker_text, color, font_family, font_size = marker_tuple
                        else:
                            continue # Skip invalid marker data

                        # Prepare SVG attributes
                        text_content = str(marker_text)
                        fill_color = QColor(color).name() if isinstance(color, QColor) else str(color) # Ensure hex/name
                        font_family_svg = str(font_family)
                        font_size_svg = f"{int(font_size)}px" if isinstance(font_size, (int, float)) else "12px"
                        font_weight_svg = "bold" if bool(is_bold) else "normal"
                        font_style_svg = "italic" if bool(is_italic) else "normal"

                        # Adjust vertical position slightly for better alignment if needed
                        # This might require font metrics specific to the marker's font
                        # For simplicity, using the standard marker offset for now
                        y_pos_adjusted = y_pos + vertical_offset_adjust

                        # Add SVG text element, centered at the marker's coordinates
                        dwg.add(
                            dwg.text(
                                text_content,
                                insert=(x_pos, y_pos_adjusted), # Position text anchor at the coordinate
                                fill=fill_color,
                                font_family=font_family_svg,
                                font_size=font_size_svg,
                                font_weight=font_weight_svg,
                                font_style=font_style_svg,
                                text_anchor="middle", # Center horizontally
                                dominant_baseline="central" # Attempt vertical centering (support varies)
                            )
                        )
                    except Exception as e:
                         print(f"Warning: Skipping invalid custom marker during SVG export: {e}")
                         # import traceback; traceback.print_exc() # Uncomment for detailed debug

                # --- START: Add Custom Shapes (Lines and Rectangles) ---
                for shape_data in getattr(self, "custom_shapes", []):
                     try:
                         shape_type = shape_data.get('type')
                         color_str = shape_data.get('color', '#000000') # Default black
                         thickness = int(shape_data.get('thickness', 1))
                         if thickness < 1: thickness = 1

                         if shape_type == 'line':
                             start_coords = shape_data.get('start')
                             end_coords = shape_data.get('end')
                             if start_coords and end_coords:
                                 dwg.add(dwg.line(start=start_coords,
                                                  end=end_coords,
                                                  stroke=color_str,
                                                  stroke_width=thickness))
                         elif shape_type == 'rectangle':
                             rect_coords = shape_data.get('rect') # (x, y, w, h)
                             if rect_coords:
                                 x, y, w, h = rect_coords
                                 dwg.add(dwg.rect(insert=(x, y),
                                                  size=(f"{w}px", f"{h}px"),
                                                  stroke=color_str,
                                                  stroke_width=thickness,
                                                  fill="none")) # No fill for outline rectangle
                     except Exception as e:
                         print(f"Warning: Skipping invalid custom shape during SVG export: {e}")
                # --- END: Add Custom Shapes ---

                # --- Save the SVG file ---
                try:
                    dwg.save()
                    QMessageBox.information(self, "Success", f"Image and annotations saved as SVG at\n{file_path}")
                    self.is_modified = False # Mark as saved if successful
                except Exception as e:
                    QMessageBox.critical(self, "SVG Save Error", f"Failed to save SVG file:\n{e}")
            
            
            def copy_to_clipboard(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Warning", "No image to copy.")
                    return

                render_scale = 3
                try:
                    view_width = self.live_view_label.width(); view_height = self.live_view_label.height()
                    if view_width <= 0: view_width = 600
                    if view_height <= 0: view_height = 400
                    high_res_canvas_width = view_width * render_scale
                    high_res_canvas_height = view_height * render_scale
                    if high_res_canvas_width <= 0 or high_res_canvas_height <= 0:
                        raise ValueError("Invalid canvas size for clipboard.")
                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Could not calculate render dimensions: {e}")
                    return

                render_canvas_for_clip = QImage(high_res_canvas_width, high_res_canvas_height, QImage.Format_ARGB32_Premultiplied)
                if render_canvas_for_clip.isNull():
                    QMessageBox.critical(self, "Copy Error", "Failed to create canvas for clipboard.")
                    return
                render_canvas_for_clip.fill(Qt.transparent)

                scaled_image_for_clip = None # Initialize
                if self.image and not self.image.isNull():
                    try:
                        scaled_image_for_clip = self.image.scaled( # This is self.image scaled to fit render_canvas_for_clip
                            high_res_canvas_width, high_res_canvas_height,
                            Qt.KeepAspectRatio, Qt.SmoothTransformation
                        )
                        if scaled_image_for_clip.isNull():
                            raise ValueError("Scaling image for clipboard failed.")
                        
                        self.render_image_on_canvas(
                            render_canvas_for_clip, scaled_image_for_clip,
                            x_start=0, y_start=0, 
                            render_scale=render_scale,
                            draw_guides=False
                        )
                    except Exception as e:
                        QMessageBox.critical(self, "Copy Error", f"Failed to render base image for clipboard: {e}")
                        self.cleanup_temp_clipboard_file() 
                        return
                else:
                    QMessageBox.warning(self, "Copy Error", "No valid current image to render for copying.")
                    return
                
                # Ensure scaled_image_for_clip is valid before proceeding
                if not scaled_image_for_clip or scaled_image_for_clip.isNull():
                    QMessageBox.critical(self, "Copy Error", "Base image for clipboard is invalid after scaling/rendering attempt.")
                    return


                painter_clip = QPainter(render_canvas_for_clip)
                if not painter_clip.isActive():
                    QMessageBox.critical(self, "Copy Error", "Failed to create painter for clipboard image.")
                    self.cleanup_temp_clipboard_file()
                    return
                
                painter_clip.setRenderHint(QPainter.Antialiasing, True)
                painter_clip.setRenderHint(QPainter.TextAntialiasing, True)

                # --- Define Local Coordinate Mapping Helper for Clipboard Canvas ---
                # This maps coordinates from self.image (native) to render_canvas_for_clip
                img_w_current = self.image.width() if self.image.width() > 0 else 1
                img_h_current = self.image.height() if self.image.height() > 0 else 1
                
                # scaled_image_for_clip is self.image scaled and centered on render_canvas_for_clip
                canvas_center_x_offset_clip = (render_canvas_for_clip.width() - scaled_image_for_clip.width()) // 2
                canvas_center_y_offset_clip = (render_canvas_for_clip.height() - scaled_image_for_clip.height()) // 2

                # Scale factor from self.image coordinates to coordinates on scaled_image_for_clip
                scale_factor_x_img_to_scaled_clip = scaled_image_for_clip.width() / img_w_current
                scale_factor_y_img_to_scaled_clip = scaled_image_for_clip.height() / img_h_current

                def map_img_coords_to_clipboard_canvas(img_x, img_y): # Renamed for clarity within this function
                    scaled_img_x_on_canvas = img_x * scale_factor_x_img_to_scaled_clip
                    scaled_img_y_on_canvas = img_y * scale_factor_y_img_to_scaled_clip
                    # Add the centering offset of scaled_image_for_clip on render_canvas_for_clip
                    final_canvas_x = canvas_center_x_offset_clip + scaled_img_x_on_canvas
                    final_canvas_y = canvas_center_y_offset_clip + scaled_img_y_on_canvas
                    return QPointF(final_canvas_x, final_canvas_y)
                # --- End Local Coordinate Mapping Helper ---


                # A. Draw Standard L/R/Top Markers
                std_font_size_on_canvas_clip = int(self.font_size * render_scale)
                std_marker_font_clip = QFont(self.font_family, std_font_size_on_canvas_clip)
                painter_clip.setFont(std_marker_font_clip)
                painter_clip.setPen(self.font_color)
                font_metrics_std_clip = QFontMetrics(std_marker_font_clip)
                text_height_std_clip_canvas = font_metrics_std_clip.height()
                y_offset_text_baseline_std_canvas_clip = text_height_std_clip_canvas * 0.3

                # Left Markers
                # self.left_marker_shift_added is in NATIVE IMAGE PIXELS
                left_marker_offset_x_on_canvas_clip = self.left_marker_shift_added * scale_factor_x_img_to_scaled_clip
                for y_pos_img, marker_text_val in self.left_markers:
                    # map_img_coords_to_clipboard_canvas gives the point on the canvas corresponding to (0, y_pos_img)
                    # The X part of its return isn't used for Y-anchor, only Y.
                    anchor_on_canvas = map_img_coords_to_clipboard_canvas(0, y_pos_img) 
                    text_to_draw = f"{marker_text_val} ⎯"
                    text_width_canvas = font_metrics_std_clip.horizontalAdvance(text_to_draw)
                    # Draw relative to the centered image's left edge + scaled offset
                    draw_x_c = canvas_center_x_offset_clip + left_marker_offset_x_on_canvas_clip - text_width_canvas
                    draw_y_c = anchor_on_canvas.y() + y_offset_text_baseline_std_canvas_clip
                    painter_clip.drawText(QPointF(draw_x_c, draw_y_c), text_to_draw)
                
                # Right Markers
                right_marker_offset_x_on_canvas_clip = self.right_marker_shift_added * scale_factor_x_img_to_scaled_clip
                for y_pos_img, marker_text_val in self.right_markers:
                    anchor_on_canvas = map_img_coords_to_clipboard_canvas(0, y_pos_img)
                    text_to_draw = f"⎯ {marker_text_val}"
                    draw_x_c = canvas_center_x_offset_clip + right_marker_offset_x_on_canvas_clip
                    draw_y_c = anchor_on_canvas.y() + y_offset_text_baseline_std_canvas_clip
                    painter_clip.drawText(QPointF(draw_x_c, draw_y_c), text_to_draw)

                # Top Markers
                top_marker_offset_y_on_canvas_clip = self.top_marker_shift_added * scale_factor_y_img_to_scaled_clip
                rotation_angle = self.font_rotation # Use self.font_rotation

                for x_pos_img, marker_text_val in self.top_markers:
                    anchor_on_canvas = map_img_coords_to_clipboard_canvas(x_pos_img, 0) 
                    text_to_draw = str(marker_text_val)
                    painter_clip.save()
                    translate_x_c = anchor_on_canvas.x()
                    translate_y_c = canvas_center_y_offset_clip + top_marker_offset_y_on_canvas_clip + y_offset_text_baseline_std_canvas_clip
                    painter_clip.translate(translate_x_c, translate_y_c)
                    painter_clip.rotate(rotation_angle)
                    painter_clip.drawText(QPointF(0, 0), text_to_draw) 
                    painter_clip.restore()

                # B. Draw Custom Markers
                for marker_data_list in getattr(self, "custom_markers", []):
                    try:
                        x_pos_img, y_pos_img, marker_text_str, qcolor_obj, \
                        font_family_str, font_size_int, is_bold, is_italic = marker_data_list
                        anchor_on_canvas = map_img_coords_to_clipboard_canvas(x_pos_img, y_pos_img) # Use the local helper
                        custom_font_size_on_canvas_clip = int(font_size_int * render_scale)
                        custom_font_clip = QFont(font_family_str, custom_font_size_on_canvas_clip)
                        custom_font_clip.setBold(is_bold)
                        custom_font_clip.setItalic(is_italic)
                        painter_clip.setFont(custom_font_clip)
                        current_color = QColor(qcolor_obj) if isinstance(qcolor_obj, QColor) else QColor(str(qcolor_obj))
                        if not current_color.isValid(): current_color = Qt.black
                        painter_clip.setPen(current_color)
                        font_metrics_custom_clip = QFontMetrics(custom_font_clip)
                        text_bounding_rect_custom = font_metrics_custom_clip.boundingRect(marker_text_str)
                        draw_x_c = anchor_on_canvas.x() - (text_bounding_rect_custom.left() + text_bounding_rect_custom.width() / 2.0)
                        draw_y_c = anchor_on_canvas.y() - (text_bounding_rect_custom.top() + text_bounding_rect_custom.height() / 2.0)
                        painter_clip.drawText(QPointF(draw_x_c, draw_y_c), marker_text_str)
                    except Exception as e_cm_clip:
                        print(f"Error drawing custom marker for clipboard: {marker_data_list}, {e_cm_clip}")

                # C. Draw Custom Shapes
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type = shape_data.get('type')
                        color = QColor(shape_data.get('color', '#000000'))
                        base_thickness_img_pixels = float(shape_data.get('thickness', 1.0))
                        # Scale thickness from image pixels to pixels on scaled_image_for_clip
                        thickness_on_canvas_clip = max(1.0, base_thickness_img_pixels * scale_factor_x_img_to_scaled_clip) 
                        pen = QPen(color)
                        pen.setWidthF(thickness_on_canvas_clip)
                        painter_clip.setPen(pen)
                        if shape_type == 'line':
                            start_img_coords = shape_data.get('start')
                            end_img_coords = shape_data.get('end')
                            if start_img_coords and end_img_coords:
                                start_canvas = map_img_coords_to_clipboard_canvas(start_img_coords[0], start_img_coords[1]) # Use local helper
                                end_canvas = map_img_coords_to_clipboard_canvas(end_img_coords[0], end_img_coords[1])       # Use local helper
                                painter_clip.drawLine(start_canvas, end_canvas)
                        elif shape_type == 'rectangle':
                            rect_img_coords = shape_data.get('rect')
                            if rect_img_coords:
                                x_img, y_img, w_img, h_img = rect_img_coords
                                top_left_canvas = map_img_coords_to_clipboard_canvas(x_img, y_img) # Use local helper
                                w_on_canvas = w_img * scale_factor_x_img_to_scaled_clip
                                h_on_canvas = h_img * scale_factor_y_img_to_scaled_clip
                                painter_clip.drawRect(QRectF(top_left_canvas, QSizeF(w_on_canvas, h_on_canvas)))
                    except Exception as e_cs_clip:
                         print(f"Error drawing custom shape for clipboard: {shape_data}, {e_cs_clip}")
                
                painter_clip.end() 
                # --- End drawing vector annotations ---

                final_canvas_for_clipboard_data = render_canvas_for_clip
                if render_canvas_for_clip.hasAlphaChannel():
                    straight_alpha_canvas_clip = render_canvas_for_clip.convertToFormat(QImage.Format_ARGB32)
                    if not straight_alpha_canvas_clip.isNull():
                        final_canvas_for_clipboard_data = straight_alpha_canvas_clip
                
                try:
                    self.cleanup_temp_clipboard_file() 
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False, mode='wb') as temp_file:
                        self.temp_clipboard_file_path = temp_file.name
                        png_save_buffer = QBuffer()
                        png_save_buffer.open(QBuffer.WriteOnly)
                        if not final_canvas_for_clipboard_data.save(png_save_buffer, "PNG"):
                             raise IOError("Failed to save final canvas to PNG buffer for clipboard.")
                        temp_file.write(png_save_buffer.data())
                        png_save_buffer.close()
                    if not os.path.exists(self.temp_clipboard_file_path):
                        raise IOError("Temporary file for clipboard not created.")
                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Failed to create/save temporary file for clipboard: {e}")
                    self.temp_clipboard_file_path = None
                    return
                
                clipboard_app = QApplication.clipboard() # Use different variable name for clarity
                mime_data = QMimeData()
                try:
                    file_url = QUrl.fromLocalFile(self.temp_clipboard_file_path)
                    if file_url.isValid() and file_url.isLocalFile():
                        mime_data.setUrls([file_url])
                    else:
                        print("Warning: Could not create valid file URL for clipboard.")
                except Exception as e_url:
                     print(f"Warning: Error creating QUrl for clipboard: {e_url}")

                mime_data.setImageData(final_canvas_for_clipboard_data)
                
                png_buffer_direct = QBuffer()
                png_buffer_direct.open(QBuffer.WriteOnly)
                if final_canvas_for_clipboard_data.save(png_buffer_direct, "PNG"):
                    mime_data.setData("image/png", png_buffer_direct.data())
                png_buffer_direct.close()

                clipboard_app.setMimeData(mime_data)
                
            def cleanup_temp_clipboard_file(self):
                """Deletes the temporary clipboard file if it exists."""
                path_to_delete = getattr(self, 'temp_clipboard_file_path', None)
                if path_to_delete:
                    # print(f"Attempting to clean up temp file on exit: {path_to_delete}") # Debug
                    # No need to clear self.temp_clipboard_file_path immediately here,
                    # as the object is likely being destroyed soon anyway.
                    if os.path.exists(path_to_delete):
                        try:
                            os.remove(path_to_delete)
                            # print(f"Cleaned up temp file: {path_to_delete}") # Debug
                        # Don't retry with QTimer here, the app is quitting.
                        except OSError as e:
                            # Log it, but can't do much else at this point.
                            print(f"Warning: Could not delete temp clipboard file {path_to_delete} on exit: {e}")
                # Clear the attribute *after* attempting deletion (optional on exit)
                self.temp_clipboard_file_path = None
                
                
            def clear_predict_molecular_weight(self):
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                self.live_view_label.setCursor(Qt.ArrowCursor)
                if hasattr(self, "protein_location"):
                    del self.protein_location  # Clear the protein location marker
                self.predict_size=False
                self.bounding_boxes=[]
                self.bounding_box_start = None
                self.live_view_label.bounding_box_start = None
                self.live_view_label.bounding_box_preview = None
                self.quantities=[]
                self.peak_area_list=[]
                self.protein_quantities=[]
                self.standard_protein_values.setText("")
                self.standard_protein_areas=[]
                self.standard_protein_areas_text.setText("")
                self.live_view_label.quad_points=[]
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_points = []
                self.latest_calculated_quantities = []
                self.quantities_peak_area_dict={}
                
                self.update_live_view()  # Update the display
                
            def predict_molecular_weight(self):
                # ... (existing initial checks for markers) ...
                
                self.live_view_label.preview_marker_enabled = False # Disable custom marker preview
                self.live_view_label.mw_predict_preview_enabled = True # Enable MW preview
                self.live_view_label.mw_predict_preview_position = None # Clear old position
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True) # Ensure it's on
    
                
                # Determine which markers to use (left or right)
                markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
                if not markers_raw_tuples:
                    QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
                    self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                    return

                # --- Crucial Step: Sort markers by Y-position (migration distance) ---
                # This ensures the order reflects the actual separation on the gel/blot
                try:
                    # Ensure marker values are numeric for sorting and processing
                    numeric_markers = []
                    for pos, val_str in markers_raw_tuples:
                        try:
                            numeric_markers.append((float(pos), float(val_str)))
                        except (ValueError, TypeError):
                            # Skip markers with non-numeric values for prediction
                            continue

                    if len(numeric_markers) < 2:
                         QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                         self.live_view_label.setCursor(Qt.ArrowCursor)
                         return

                    sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
                    # Separate sorted positions and values
                    sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
                    sorted_marker_values = np.array([val for pos, val in sorted_markers])

                except Exception as e:
                    QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    return

                # Ensure there are still at least two markers after potential filtering
                if len(sorted_marker_positions) < 2:
                    QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    return

                # Prompt user and set up the click handler
                QMessageBox.information(self, "Instruction",
                                        "Click on the target protein location in the preview window.\n"
                                        "The closest standard set (Gel or WB) will be used for calculation.")
                # Pass the *sorted* full marker data to the click handler
                self.live_view_label.mousePressEvent = lambda event: self.get_protein_location_and_clear_preview(
                    event, sorted_marker_positions, sorted_marker_values
                )
                # Add mouseMoveEvent for the preview
                self.live_view_label.mouseMoveEvent = self.update_mw_predict_preview

            def get_protein_location(self, event, all_marker_positions, all_marker_values):
                """
                Handles the mouse click event for protein selection.
                Determines the relevant standard set based on click proximity,
                performs regression on that set, predicts MW, and plots the results.
                """
                # --- 1. Get Protein Click Position (Image Coordinates) ---
                pos = event.pos()
                cursor_x, cursor_y = pos.x(), pos.y()

                # Account for zoom and pan
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                # Transform cursor position from view coordinates to image coordinates
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
                image_width = self.image.width() if self.image and self.image.width() > 0 else 1
                image_height = self.image.height() if self.image and self.image.height() > 0 else 1

                scale = min(displayed_width / image_width, displayed_height / image_height)
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2

                protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
                # protein_x_image = (cursor_x - x_offset) / scale if scale != 0 else 0 # Keep X if needed later

                # Store the clicked location in *view* coordinates for drawing the marker later
                self.protein_location = (cursor_x, cursor_y) # Use view coordinates for the marker

                # --- 2. Identify Potential Standard Sets (Partitioning) ---
                transition_index = -1
                # Find the first index where the molecular weight INCREASES after initially decreasing
                # (indicates a likely switch from Gel high->low MW to WB high->low MW)
                initial_decrease = False
                for k in range(1, len(all_marker_values)):
                     if all_marker_values[k] < all_marker_values[k-1]:
                         initial_decrease = True # Confirm we've started migrating down
                     # Check for increase *after* we've seen a decrease
                     if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                         transition_index = k
                         break # Found the likely transition

                # --- 3. Select the Active Standard Set based on Click Proximity ---
                active_marker_positions = None
                active_marker_values = None
                set_name = "Full Set" # Default name

                if transition_index != -1:
                    # We have two potential sets
                    set1_positions = all_marker_positions[:transition_index]
                    set1_values = all_marker_values[:transition_index]
                    set2_positions = all_marker_positions[transition_index:]
                    set2_values = all_marker_values[transition_index:]

                    # Check if both sets are valid (at least 2 points)
                    valid_set1 = len(set1_positions) >= 2
                    valid_set2 = len(set2_positions) >= 2

                    if valid_set1 and valid_set2:
                        # Calculate the mean Y position for each set
                        mean_y_set1 = np.mean(set1_positions)
                        mean_y_set2 = np.mean(set2_positions)

                        # Assign the click to the set whose mean Y is closer
                        if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                            active_marker_positions = set1_positions
                            active_marker_values = set1_values
                            set_name = "Set 1 (Gel?)" # Tentative name
                        else:
                            active_marker_positions = set2_positions
                            active_marker_values = set2_values
                            set_name = "Set 2 (WB?)" # Tentative name
                    elif valid_set1: # Only set 1 is valid
                         active_marker_positions = set1_positions
                         active_marker_values = set1_values
                         set_name = "Set 1 (Gel?)"
                    elif valid_set2: # Only set 2 is valid
                         active_marker_positions = set2_positions
                         active_marker_values = set2_values
                         set_name = "Set 2 (WB?)"
                    else: # Neither set is valid after splitting
                         QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning.")
                         self.live_view_label.setCursor(Qt.ArrowCursor)
                         if hasattr(self, "protein_location"): del self.protein_location
                         return
                else:
                    # Only one set detected, use all markers
                    if len(all_marker_positions) >= 2:
                        active_marker_positions = all_marker_positions
                        active_marker_values = all_marker_values
                        set_name = "Single Set"
                    else: # Should have been caught earlier, but double-check
                         QMessageBox.warning(self, "Error", "Not enough markers in the single set.")
                         self.live_view_label.setCursor(Qt.ArrowCursor)
                         if hasattr(self, "protein_location"): del self.protein_location
                         return

                # --- 4. Perform Regression on the Active Set ---
                # Normalize distances *within the active set*
                min_pos_active = np.min(active_marker_positions)
                max_pos_active = np.max(active_marker_positions)
                if max_pos_active == min_pos_active: # Avoid division by zero if all points are the same
                    QMessageBox.warning(self, "Error", "All markers in the selected set have the same position.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    if hasattr(self, "protein_location"): del self.protein_location
                    return

                normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)

                # Log transform marker values
                try:
                    log_marker_values_active = np.log10(active_marker_values)
                except Exception as e:
                     QMessageBox.warning(self, "Error", f"Could not log-transform marker values (are they all positive?): {e}")
                     self.live_view_label.setCursor(Qt.ArrowCursor)
                     if hasattr(self, "protein_location"): del self.protein_location
                     return

                # Perform linear regression (log-linear fit)
                coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, 1)

                # Calculate R-squared for the fit on the active set
                residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
                ss_res = np.sum(residuals**2)
                ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
                r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1 # Handle case of perfect fit or single point mean

                # --- 5. Predict MW for the Clicked Protein ---
                # Normalize the protein's Y position *using the active set's min/max*
                normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)

                # Predict using the derived coefficients
                predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
                predicted_weight = 10 ** predicted_log10_weight

                # --- 6. Update View and Plot ---
                self.update_live_view() # Draw the '*' marker at self.protein_location

                # Plot the results, passing both active and all data for context
                self.plot_molecular_weight_graph(
                    # Active set data for fitting and line plot:
                    normalized_distances_active,
                    active_marker_values,
                    10 ** np.polyval(coefficients, normalized_distances_active), # Fitted values for active set
                    # Full set data for context (plotting all points):
                    all_marker_positions, # Pass original positions
                    all_marker_values,
                    min_pos_active,       # Pass min/max of *active* set for normalization context
                    max_pos_active,
                    # Prediction results:
                    normalized_protein_position, # Position relative to active set scale
                    predicted_weight,
                    r_squared,
                    set_name # Name of the set used
                )

                # Reset mouse event handler after prediction
                self.live_view_label.mousePressEvent = None
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.run_predict_MW = True # Indicate prediction was attempted/completed

            def get_protein_location_and_clear_preview(self, event, all_marker_positions, all_marker_values):
                # First, call the original logic
                self.get_protein_location(event, all_marker_positions, all_marker_values)
                
                # Then, disable the MW preview
                self.live_view_label.mw_predict_preview_enabled = False
                self.live_view_label.mw_predict_preview_position = None
                self.live_view_label.setMouseTracking(False) # Can turn off if only for this mode
                self.live_view_label.mouseMoveEvent = None # Unbind move handler
                # live_view_label.mousePressEvent is already reset by get_protein_location
                self.live_view_label.update() # Refresh to clear preview
                
            def update_mw_predict_preview(self, event):
                if self.live_view_label.mw_predict_preview_enabled:
                    untransformed_label_pos = self.live_view_label.transform_point(event.pos())
                    # Snapping for MW preview is optional, but can be nice
                    snapped_label_pos = self.snap_point_to_grid(untransformed_label_pos) 
                    self.live_view_label.mw_predict_preview_position = snapped_label_pos
                    self.live_view_label.update()
                elif hasattr(self.live_view_label, '_original_mouseMoveEvent'): # Pass to original if exists
                     self.live_view_label._original_mouseMoveEvent(event)
                
            def plot_molecular_weight_graph(
                self,
                # Data for the active set (used for the fit line and highlighted points)
                active_norm_distances,  # Normalized distances for the active set points
                active_marker_values,   # Original MW values of the active set points
                active_fitted_values,   # Fitted MW values corresponding to active_norm_distances

                # Data for *all* markers (for plotting all points for context)
                all_marker_positions,   # *Original* Y positions of all markers
                all_marker_values,      # *Original* MW values of all markers
                active_min_pos,         # Min Y position of the *active* set (for normalizing all points)
                active_max_pos,         # Max Y position of the *active* set (for normalizing all points)

                # Prediction results
                predicted_norm_position,# Predicted protein position normalized relative to active set scale
                predicted_weight,       # Predicted MW value

                # Fit quality and set info
                r_squared,              # R-squared of the fit on the active set
                set_name                # Name of the set used (e.g., "Set 1", "Set 2", "Single Set")
            ):
                """
                Plots the molecular weight prediction graph and displays it in a custom dialog.
                """
                # --- Normalize *all* marker positions using the *active* set's scale ---
                if active_max_pos == active_min_pos: # Avoid division by zero
                     all_norm_distances_for_plot = np.zeros_like(all_marker_positions)
                else:
                    all_norm_distances_for_plot = (all_marker_positions - active_min_pos) / (active_max_pos - active_min_pos)

                # --- Create Plot ---
                # Adjust figsize and DPI for a more compact plot suitable for a dialog
                fig, ax = plt.subplots(figsize=(4.5, 3.5)) # Smaller figure size

                # 1. Plot *all* marker points lightly for context
                ax.scatter(all_norm_distances_for_plot, all_marker_values, color="grey", alpha=0.5, label="All Markers", s=25) # Slightly smaller

                # 2. Plot the *active* marker points prominently
                ax.scatter(active_norm_distances, active_marker_values, color="red", label=f"Active Set ({set_name})", s=40, marker='o') # Slightly smaller

                # 3. Plot the fitted line for the *active* set
                sort_indices = np.argsort(active_norm_distances)
                # Move R^2 out of the legend
                ax.plot(active_norm_distances[sort_indices], active_fitted_values[sort_indices],
                         color="blue", label="Fit Line", linewidth=1.5)

                # 4. Plot the predicted protein position
                # Move predicted value out of the legend
                ax.axvline(predicted_norm_position, color="green", linestyle="--",
                            label="Target Protein", linewidth=1.5)

                # --- Configure Plot ---
                ax.set_xlabel(f"Normalized Distance (Relative to {set_name})", fontsize=9)
                ax.set_ylabel("Molecular Weight (units)", fontsize=9)
                ax.set_yscale("log")
                # Smaller legend font size
                ax.legend(fontsize='x-small', loc='best') # Use 'best' location
                ax.set_title(f"Molecular Weight Prediction", fontsize=10) # Simpler title
                ax.grid(True, which='both', linestyle=':', linewidth=0.5)
                ax.tick_params(axis='both', which='major', labelsize=8) # Smaller tick labels

                plt.tight_layout(pad=0.5) # Adjust layout to prevent labels overlapping

                # --- Save Plot to Buffer ---
                pixmap = None
                try:
                    buffer = BytesIO()
                    # Use a moderate DPI suitable for screen display in a dialog
                    plt.savefig(buffer, format="png", bbox_inches="tight", dpi=100)
                    buffer.seek(0)
                    pixmap = QPixmap()
                    pixmap.loadFromData(buffer.read())
                    buffer.close()
                except Exception as plot_err:
                    QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
                    # pixmap remains None
                finally:
                     plt.close(fig) # Ensure figure is always closed using the fig object

                # --- Display Results in a Custom Dialog ---
                if pixmap:
                    # Create a custom dialog instead of modifying QMessageBox layout
                    dialog = QDialog(self)
                    dialog.setWindowTitle("Prediction Result")

                    # Layout for the dialog
                    layout = QVBoxLayout(dialog)

                    # Text label for results
                    results_text = (
                        f"<b>Prediction using {set_name}:</b><br>"
                        f"Predicted MW: <b>{predicted_weight:.2f}</b> units<br>"
                        f"Fit R²: {r_squared:.3f}"
                    )
                    info_label = QLabel(results_text)
                    info_label.setTextFormat(Qt.RichText) # Allow basic HTML like <b>
                    info_label.setWordWrap(True)
                    layout.addWidget(info_label)

                    # Label to display the plot
                    plot_label = QLabel()
                    plot_label.setPixmap(pixmap)
                    plot_label.setAlignment(Qt.AlignCenter) # Center the plot
                    layout.addWidget(plot_label)

                    # OK button
                    button_box = QDialogButtonBox(QDialogButtonBox.Ok)
                    button_box.accepted.connect(dialog.accept)
                    layout.addWidget(button_box)

                    dialog.setLayout(layout)
                    dialog.exec_() # Show the custom dialog modally
                else:
                     # If plot generation failed, show a simpler message box
                     QMessageBox.information(self, "Prediction Result (No Plot)",
                        f"Used {set_name} for prediction.\n"
                        f"Predicted MW: {predicted_weight:.2f} units\n"
                        f"Fit R²: {r_squared:.3f}"
                     )

          
                
            def reset_image(self):
                self.cancel_rectangle_crop_mode()
                self.crop_rectangle_coords = None
                self.live_view_label.clear_crop_preview()

                # --- Reset and Disable Crop Sliders ---
                slider_info = [
                    (getattr(self, 'crop_x_start_slider', None), self.crop_slider_min),
                    (getattr(self, 'crop_x_end_slider', None), self.crop_slider_max),
                    (getattr(self, 'crop_y_start_slider', None), self.crop_slider_min),
                    (getattr(self, 'crop_y_end_slider', None), self.crop_slider_max)
                ]
                for slider, default_value in slider_info:
                    if slider:
                        slider.blockSignals(True)
                        slider.setValue(default_value)
                        slider.setEnabled(False) # Disable sliders on reset
                        slider.blockSignals(False)
                # --- End Reset Crop ---

                # ... (rest of reset_image method: clearing image, other sliders, markers, etc.) ...
                if hasattr(self, 'image_master') and self.image_master and not self.image_master.isNull():
                    self.image = self.image_master.copy()
                    self.image_before_padding = None
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False
                else:
                    self.image = None
                    self.original_image = None
                    self.image_master = None
                    self.image_before_padding = None
                    self.image_contrasted = None
                    self.image_before_contrast = None
                    self.image_padded = False

                if hasattr(self, 'show_grid_checkbox_x'): self.show_grid_checkbox_x.setChecked(False)
                if hasattr(self, 'show_grid_checkbox_y'): self.show_grid_checkbox_y.setChecked(False)
                if hasattr(self, 'grid_size_input'): self.grid_size_input.setValue(20)

                self.warped_image=None
                if hasattr(self, 'left_markers'): self.left_markers.clear()
                if hasattr(self, 'right_markers'): self.right_markers.clear()
                if hasattr(self, 'top_markers'): self.top_markers.clear()
                if hasattr(self, 'custom_markers'): self.custom_markers.clear()
                if hasattr(self, 'custom_shapes'): self.custom_shapes.clear()
                self.cancel_drawing_mode()
                self.clear_predict_molecular_weight()

                if hasattr(self, 'orientation_slider'): self.orientation_slider.setValue(0)
                if hasattr(self, 'taper_skew_slider'): self.taper_skew_slider.setValue(0)
                if hasattr(self, 'high_slider'): self.high_slider.setValue(100)
                if hasattr(self, 'low_slider'): self.low_slider.setValue(100)
                if hasattr(self, 'gamma_slider'): self.gamma_slider.setValue(100)
                if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(0)
                if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(0)
                if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(0)
                if hasattr(self, 'left_padding_input'): self.left_padding_input.setText("0")
                if hasattr(self, 'right_padding_input'): self.right_padding_input.setText("0")
                if hasattr(self, 'top_padding_input'): self.top_padding_input.setText("0")
                if hasattr(self, 'bottom_padding_input'): self.bottom_padding_input.setText("0")
                if hasattr(self, "custom_markers"): self.custom_markers.clear()
                if hasattr(self, "custom_shapes"): self.custom_shapes.clear()

                self.marker_mode = None
                self.current_left_marker_index = 0
                self.current_right_marker_index = 0
                self.current_top_label_index = 0
                self.left_marker_shift_added = 0
                self.right_marker_shift_added = 0
                self.top_marker_shift_added = 0
                self.live_view_label.mode = None
                self.live_view_label.quad_points = []
                self.live_view_label.setCursor(Qt.ArrowCursor)

                try:
                    if hasattr(self, 'combo_box'):
                        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                        self.on_combobox_changed()
                except Exception as e: pass

                if self.image and not self.image.isNull():
                    try:
                        self._update_preview_label_size()
                        if hasattr(self, 'left_padding_input'): self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'right_padding_input'): self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'top_padding_input'): self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                    except Exception as e: pass
                else:
                    self.live_view_label.clear()
                    self._update_preview_label_size()

                self.live_view_label.zoom_level = 1.0
                self.live_view_label.pan_offset = QPointF(0, 0)
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                                
                # Reset overlay sliders to 0 after their ranges are updated
                if hasattr(self, 'image1_left_slider'): self.image1_left_slider.setValue(0)
                if hasattr(self, 'image1_top_slider'): self.image1_top_slider.setValue(0)
                if hasattr(self, 'image1_resize_slider'): self.image1_resize_slider.setValue(100)
                if hasattr(self, 'image2_left_slider'): self.image2_left_slider.setValue(0)
                if hasattr(self, 'image2_top_slider'): self.image2_top_slider.setValue(0)
                if hasattr(self, 'image2_resize_slider'): self.image2_resize_slider.setValue(100)
                
                self._update_overlay_slider_ranges()

                self._update_marker_slider_ranges()
                self.update_live_view()
                self._update_status_bar()

        main_window = CombinedSDSApp()

        # --- Close Loading Screen and Show Main Window ---
        if loading_dialog:
            loading_dialog.close()

        main_window.show()

        # Connect cleanup
        if main_window:
            app.aboutToQuit.connect(main_window.cleanup_temp_clipboard_file)

        # --- Start Main Event Loop ---
        exit_code = app.exec_()
        sys.exit(exit_code)

    except ImportError as e:
        # Handle missing libraries critically
        print(f"FATAL ERROR: Missing required library: {e}")
        traceback.print_exc()
        if loading_dialog: loading_dialog.close() # Ensure loading screen is closed
        # Show error message using basic QMessageBox (app might not be fully functional)
        try:
            error_app = QApplication.instance() # Get existing app instance if possible
            if not error_app: error_app = QApplication([]) # Create minimal app for msgbox
            QMessageBox.critical(None, "Import Error", f"A required library is missing: {e}\nPlease install it and restart the application.")
        except Exception as msg_err:
            print(f"Could not display import error message box: {msg_err}")
        sys.exit(1) # Exit with error code

    except Exception as e:
        # Catch any other exception during startup
        print(f"FATAL ERROR during application startup: {e}")
        traceback.print_exc()
        if loading_dialog: loading_dialog.close() # Ensure loading screen is closed
        # Try to log using your defined function
        try:
            # Make sure log_exception is defined globally or imported
            log_exception(type(e), e, e.__traceback__)
        except NameError:
             print("Logging function 'log_exception' not found.")
        except Exception as log_err:
            print(f"Failed to log startup exception: {log_err}")
        # Show critical error message box
        try:
            error_app = QApplication.instance()
            if not error_app: error_app = QApplication([])
            QMessageBox.critical(None, "Application Startup Error", f"An unexpected error occurred during startup:\n{e}\n\nCheck error_log.txt for details.")
        except Exception as msg_err:
            print(f"Could not display startup error message box: {msg_err}")
        sys.exit(1) # Exit with error code

    finally:
        # Ensure the loading dialog is closed in case of unhandled exit paths
        if loading_dialog and loading_dialog.isVisible():
            loading_dialog.close()