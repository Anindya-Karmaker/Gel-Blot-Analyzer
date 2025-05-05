import logging
import traceback
import sys
import svgwrite
import tempfile
from tempfile import NamedTemporaryFile
import base64
from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
from io import BytesIO
import io
from PyQt5.QtWidgets import (
    QDesktopWidget, QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem,QToolBar,QStyle,
    QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy,
    QMainWindow, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit,
    QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog,
    QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox,
    QDialog, QHeaderView, QAbstractItemView, QMenu, QAction, QMenuBar, QFontDialog
)
from PyQt5.QtGui import QPixmap, QIcon, QPalette,QKeySequence, QImage, QPolygonF,QPainter, QBrush, QColor, QFont, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF, QUrl, QSize, QSizeF, QMimeData, QUrl
import json
import os
import numpy as np
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.gridspec import GridSpec
from skimage.restoration import rolling_ball 
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from scipy.ndimage import grey_opening, grey_erosion, grey_dilation
from scipy.interpolate import interp1d # Needed for interpolation
import cv2


# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
script_dir = os.path.dirname(os.path.abspath(__file__))
# Construct the absolute path for the log file
log_file_path = os.path.join(script_dir, "error_log.txt")


# --- Configure logging ---
logging.basicConfig(
    filename=log_file_path, # Use the absolute path
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

try:
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not write initial log message: {e}") # Print error if immediate logging fails


def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    print("!!! log_exception called !!!") # Add print statement here too
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    # Log the exception
    try:
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
        # Optional: Force flush if you suspect buffering issues, though unlikely for ERROR level
        # logging.getLogger().handlers[0].flush()
    except Exception as log_err:
        print(f"ERROR: Failed to log exception to file: {log_err}") # Print logging specific errors

    # Display a QMessageBox with the error details
    try:
        error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}\n\n(Check error_log.txt for details)"
        QMessageBox.critical(
            None,
            "Unexpected Error",
            error_message,
            QMessageBox.Ok
        )
    except Exception as q_err:
         print(f"ERROR: Failed to show QMessageBox: {q_err}")


# Set the custom exception handler
sys.excepthook = log_exception

def create_text_icon(font_type: QFont, icon_size: QSize, color: QColor, symbol: str) -> QIcon:
    """Creates a QIcon by drawing text/symbol onto a pixmap."""
    pixmap = QPixmap(icon_size)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setRenderHint(QPainter.TextAntialiasing, True) # Good for text

    # Font settings (adjust as needed)
    font = QFont(font_type)
    # Make arrow slightly smaller than +/-, maybe not bold? Experiment.
    font.setPointSize(max(12, int(icon_size.width()*0.85)))
    # font.setBold(True) # Optional: Make arrows bold or not
    painter.setFont(font)
    painter.setPen(color)

    # Draw the symbol centered
    painter.drawText(pixmap.rect(), Qt.AlignCenter, symbol)
    painter.end()
    return QIcon(pixmap)

class LoadingDialog(QDialog):
    """A simple modal dialog to show while the main application loads."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint) # Frameless splash screen style
        self.setAttribute(Qt.WA_TranslucentBackground) # Optional: Allows for shaped windows if using masks
        self.setModal(True) # Block interaction with other windows (though none exist yet)

        # --- Styling ---
        self.setStyleSheet("""
            QDialog {
                background-color: rgba(50, 50, 60, 230); /* Semi-transparent dark background */
                border: 1px solid #AAAAAA;
                border-radius: 8px;
            }
            QLabel {
                color: #E0E0E0; /* Light text color */
                padding: 20px; /* Add padding around text */
            }
        """)

        # --- Layout and Label ---
        layout = QVBoxLayout(self)
        self.label = QLabel("Initializing Application...\nPlease Wait")
        font = QFont()
        font.setPointSize(14)
        font.setBold(True)
        self.label.setFont(font)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        self.setLayout(layout)

        # --- Size and Position ---
        self.setFixedSize(350, 150) # Adjust size as needed
        self.center_on_screen()

    def center_on_screen(self):
        """Centers the dialog on the primary screen."""
        try:
            screen_geo = QDesktopWidget().availableGeometry() # Get available geometry
            dialog_geo = self.frameGeometry()
            center_point = screen_geo.center()
            dialog_geo.moveCenter(center_point)
            self.move(dialog_geo.topLeft())
        except Exception as e:
            print(f"Warning: Could not center loading dialog: {e}")
            # Fallback position if centering fails
            self.move(100, 100)

    def set_message(self, message):
        """Updates the message displayed on the loading screen."""
        self.label.setText(message)
        QApplication.processEvents() # Process events to ensure the label updates
    
class ModifyMarkersDialog(QDialog):
    """
    Dialog to view, edit, delete custom markers (text) and shapes (lines/rectangles).
    Handles different properties for each type in a unified table.
    Separates Text/Label and Coordinates into distinct editable columns.
    """
    def __init__(self, markers_list, shapes_list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Modify Custom Markers and Shapes")
        self.setMinimumSize(950, 500) # Increased width slightly for new column

        # --- Store Working Copies (Same as before) ---
        self.markers = []
        for marker_data in markers_list:
            try:
                marker_copy = list(marker_data)
                if len(marker_copy) == 6: marker_copy.extend([False, False])
                if len(marker_copy) == 8: self.markers.append(marker_copy)
                else: print(f"Warning: Skipping marker with invalid data length: {len(marker_copy)}")
            except Exception as e: print(f"Warning: Error processing marker data: {e}")

        self.shapes = []
        for shape_data in shapes_list:
             if isinstance(shape_data, dict): self.shapes.append(dict(shape_data))
             else: print(f"Warning: Skipping shape with invalid data type: {type(shape_data)}")

        self._block_signals = False

        # --- Main Layout ---
        layout = QVBoxLayout(self)

        # --- Table Widget ---
        self.table_widget = QTableWidget()
        # --- MODIFIED: Column Count and Headers ---
        self.table_widget.setColumnCount(8) # Increased count
        self.table_widget.setHorizontalHeaderLabels([
            "Type", "Text/Label", "Coordinates", "Style", "Bold", "Italic", "Color", "Actions"
        ])
        # --- END MODIFICATION ---
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table_widget.setEditTriggers(QAbstractItemView.AnyKeyPressed | QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        self.table_widget.setSortingEnabled(True)

        self.table_widget.itemChanged.connect(self.handle_item_changed)
        self.table_widget.cellDoubleClicked.connect(self.handle_cell_double_clicked)

        layout.addWidget(self.table_widget)
        self.populate_table()
        self.table_widget.resizeColumnsToContents()
        # --- Adjust Column Widths (Adjust for new columns) ---
        self.table_widget.setColumnWidth(0, 70)  # Type
        self.table_widget.setColumnWidth(1, 180) # Text/Label
        self.table_widget.setColumnWidth(2, 180) # Coordinates
        self.table_widget.setColumnWidth(3, 150) # Style
        self.table_widget.setColumnWidth(4, 40)  # Bold
        self.table_widget.setColumnWidth(5, 40)  # Italic
        self.table_widget.setColumnWidth(6, 80)  # Color
        self.table_widget.setColumnWidth(7, 80)  # Actions
        self.table_widget.horizontalHeader().setStretchLastSection(False)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        self.setLayout(layout)

    def populate_table(self):
        """Fills the table, placing Text and Coordinates in separate columns."""
        self._block_signals = True
        self.table_widget.setRowCount(0)
        total_items = len(self.markers) + len(self.shapes)
        self.table_widget.setRowCount(total_items)
        self.table_widget.setSortingEnabled(False)
        current_row_idx = 0

        # --- Populate Markers ---
        for marker_idx, marker_data in enumerate(self.markers):
            row_idx = current_row_idx
            try:
                x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_data
                if not isinstance(qcolor, QColor): qcolor = QColor(qcolor)
                if not qcolor.isValid(): raise ValueError("Invalid marker color")

                type_item = QTableWidgetItem("Marker")
                type_item.setData(Qt.UserRole, {'type': 'marker', 'index': marker_idx})

                # --- MODIFIED: Split Text and Coordinates ---
                text_item = QTableWidgetItem(str(text))
                text_item.setFlags(text_item.flags() | Qt.ItemIsEditable) # Text is editable

                coord_str = f"{x:.1f},{y:.1f}"
                coord_item = QTableWidgetItem(coord_str)
                coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable) # Coordinates are editable
                coord_item.setToolTip("Edit format: X,Y (e.g., 100.5,250.2)")
                # --- END MODIFICATION ---

                style_item = QTableWidgetItem(f"{font_family} ({font_size}pt)")
                style_item.setToolTip("Double-click to change font/size")
                style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)

                color_item = QTableWidgetItem(qcolor.name())
                color_item.setBackground(QBrush(qcolor)); text_color = Qt.white if qcolor.lightness() < 128 else Qt.black
                color_item.setForeground(QBrush(text_color)); color_item.setToolTip("Double-click to change color")
                color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)

                # --- MODIFIED: Place items in new columns ---
                self.table_widget.setItem(row_idx, 0, type_item)
                self.table_widget.setItem(row_idx, 1, text_item)    # Text/Label in Col 1
                self.table_widget.setItem(row_idx, 2, coord_item)   # Coordinates in Col 2
                self.table_widget.setItem(row_idx, 3, style_item)   # Style in Col 3
                self.table_widget.setItem(row_idx, 6, color_item)   # Color in Col 6
                # --- END MODIFICATION ---

                # --- MODIFIED: Place widgets in new columns ---
                bold_checkbox = QCheckBox(); bold_checkbox.setChecked(bool(is_bold))
                bold_checkbox.stateChanged.connect(lambda state, r=row_idx: self.handle_marker_style_changed(state, r, "bold"))
                cell_widget_bold = QWidget(); layout_bold = QHBoxLayout(cell_widget_bold); layout_bold.addWidget(bold_checkbox); layout_bold.setAlignment(Qt.AlignCenter); layout_bold.setContentsMargins(0,0,0,0); cell_widget_bold.setLayout(layout_bold)
                self.table_widget.setCellWidget(row_idx, 4, cell_widget_bold) # Bold in Col 4

                italic_checkbox = QCheckBox(); italic_checkbox.setChecked(bool(is_italic))
                italic_checkbox.stateChanged.connect(lambda state, r=row_idx: self.handle_marker_style_changed(state, r, "italic"))
                cell_widget_italic = QWidget(); layout_italic = QHBoxLayout(cell_widget_italic); layout_italic.addWidget(italic_checkbox); layout_italic.setAlignment(Qt.AlignCenter); layout_italic.setContentsMargins(0,0,0,0); cell_widget_italic.setLayout(layout_italic)
                self.table_widget.setCellWidget(row_idx, 5, cell_widget_italic) # Italic in Col 5

                delete_button = QPushButton("Delete"); delete_button.setStyleSheet("QPushButton { padding: 2px 5px; }")
                delete_button.clicked.connect(lambda checked, row=row_idx: self.delete_item(row))
                self.table_widget.setCellWidget(row_idx, 7, delete_button) # Delete in Col 7
                # --- END MODIFICATION ---

                current_row_idx += 1
            except (ValueError, IndexError, TypeError) as e:
                # Error Row Handling (adapt column range)
                print(f"Error processing marker data at original index {marker_idx}: {e}")
                error_item = QTableWidgetItem("Marker Error")
                error_item.setData(Qt.UserRole, {'type': 'error', 'index': -1})
                self.table_widget.setItem(row_idx, 0, error_item)
                for col in range(1, self.table_widget.columnCount()): # Use columnCount
                    placeholder = QTableWidgetItem("---"); placeholder.setFlags(placeholder.flags() & ~Qt.ItemIsEditable)
                    self.table_widget.setItem(row_idx, col, placeholder)
                current_row_idx += 1

        # --- Populate Shapes ---
        for shape_idx, shape_data in enumerate(self.shapes):
            row_idx = current_row_idx
            try:
                shape_type = shape_data.get('type', 'Unknown').capitalize()
                color_name = shape_data.get('color', '#000000')
                thickness = int(shape_data.get('thickness', 1))
                qcolor = QColor(color_name);
                if not qcolor.isValid(): raise ValueError(f"Invalid shape color: {color_name}")
                if thickness < 1: raise ValueError("Thickness must be >= 1")

                type_item = QTableWidgetItem(shape_type)
                type_item.setData(Qt.UserRole, {'type': 'shape', 'index': shape_idx})

                # --- MODIFIED: Split Text (Blank) and Coordinates ---
                text_item = QTableWidgetItem("") # Shapes have no text label
                text_item.setFlags(text_item.flags() & ~Qt.ItemIsEditable) # Not editable

                details_str = ""
                tooltip_str = "Edit format: "
                if shape_type == 'Line':
                    start = shape_data.get('start', (0,0)); end = shape_data.get('end', (0,0))
                    details_str = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"
                    tooltip_str += "X1,Y1,X2,Y2"
                elif shape_type == 'Rectangle':
                    rect = shape_data.get('rect', (0,0,0,0)) # x,y,w,h
                    details_str = f"{rect[0]:.1f},{rect[1]:.1f},{rect[2]:.1f},{rect[3]:.1f}"
                    tooltip_str += "X,Y,W,H"
                else:
                    details_str = "N/A"
                    tooltip_str = "Cannot edit coordinates for this shape type."

                coord_item = QTableWidgetItem(details_str)
                coord_item.setToolTip(tooltip_str)
                if shape_type in ['Line', 'Rectangle']:
                    coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable) # Make editable
                else:
                    coord_item.setFlags(coord_item.flags() & ~Qt.ItemIsEditable)
                # --- END MODIFICATION ---

                style_item = QTableWidgetItem(f"{thickness}px")
                style_item.setToolTip("Double-click to change thickness")
                style_item.setFlags(style_item.flags() & ~Qt.ItemIsEditable)

                color_item = QTableWidgetItem(qcolor.name())
                color_item.setBackground(QBrush(qcolor)); text_color = Qt.white if qcolor.lightness() < 128 else Qt.black
                color_item.setForeground(QBrush(text_color)); color_item.setToolTip("Double-click to change color")
                color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)

                # --- MODIFIED: Place items in new columns ---
                self.table_widget.setItem(row_idx, 0, type_item)
                self.table_widget.setItem(row_idx, 1, text_item)    # Text/Label (Blank) in Col 1
                self.table_widget.setItem(row_idx, 2, coord_item)   # Coordinates in Col 2
                self.table_widget.setItem(row_idx, 3, style_item)   # Style in Col 3
                self.table_widget.setItem(row_idx, 6, color_item)   # Color in Col 6
                # --- END MODIFICATION ---

                # Delete Button (Adjusted Column)
                delete_button = QPushButton("Delete"); delete_button.setStyleSheet("QPushButton { padding: 2px 5px; }")
                delete_button.clicked.connect(lambda checked, row=row_idx: self.delete_item(row))
                self.table_widget.setCellWidget(row_idx, 7, delete_button) # Delete in Col 7

                current_row_idx += 1
            except (ValueError, IndexError, TypeError, KeyError) as e:
                # Error Row Handling (adapt column range)
                print(f"Error processing shape data at original index {shape_idx}: {e}")
                error_item = QTableWidgetItem("Shape Error")
                error_item.setData(Qt.UserRole, {'type': 'error', 'index': -1})
                self.table_widget.setItem(row_idx, 0, error_item)
                for col in range(1, self.table_widget.columnCount()): # Use columnCount
                    placeholder = QTableWidgetItem("---"); placeholder.setFlags(placeholder.flags() & ~Qt.ItemIsEditable)
                    self.table_widget.setItem(row_idx, col, placeholder)
                current_row_idx += 1

        if current_row_idx < total_items:
            self.table_widget.setRowCount(current_row_idx)
        self.table_widget.setSortingEnabled(True)
        self._block_signals = False
        
    def handle_marker_style_changed(self, state, row, style_type):
        """Update the bold/italic flag for a marker when its checkbox changes."""
        if self._block_signals: return

        type_item = self.table_widget.item(row, 0)
        if not type_item: return
        item_data = type_item.data(Qt.UserRole)
        # Ensure it's a marker and get its original index
        if not item_data or item_data.get('type') != 'marker': return
        original_marker_index = item_data.get('index')
        if not isinstance(original_marker_index, int) or not (0 <= original_marker_index < len(self.markers)):
            print(f"Warning: Invalid original marker index {original_marker_index} for style change at row {row}")
            return

        is_checked = (state == Qt.Checked)
        try:
            if style_type == "bold":
                self.markers[original_marker_index][6] = is_checked # Update bold flag (index 6)
            elif style_type == "italic":
                self.markers[original_marker_index][7] = is_checked # Update italic flag (index 7)
            # print(f"Updated marker original index {original_marker_index} {style_type} to: {is_checked}") # Debug
        except IndexError:
            print(f"Error: Index mismatch updating style for marker original index {original_marker_index}.")

    def handle_item_changed(self, item):
        """Update internal lists when an editable cell (Text/Label or Coords) changes."""
        if self._block_signals: return

        row = item.row()
        col = item.column()

        type_item = self.table_widget.item(row, 0)
        if not type_item: return
        item_data = type_item.data(Qt.UserRole)
        if not item_data or item_data.get('type') == 'error': return

        item_type = item_data['type']
        original_index = item_data['index']
        new_value_str = item.text()
        revert_needed = False
        error_message = ""
        revert_text = "" # Store text to revert to

        # --- Handle Marker Text Changes (Column 1) ---
        if col == 1 and item_type == 'marker':
            if not (0 <= original_index < len(self.markers)): return
            prev_text = self.markers[original_index][2] # Store previous text
            try:
                new_text = new_value_str.strip()
                self.markers[original_index][2] = new_text
                # print(f"Updated marker {original_index} text: '{new_text}'") # Debug
            except Exception as e: # Should be unlikely for text, but include for safety
                revert_needed = True
                revert_text = prev_text
                error_message = f"Error updating text: {e}"

        # --- Handle Coordinate Changes (Column 2) ---
        elif col == 2:
            # --- Marker Coordinate Handling (Col 2) ---
            if item_type == 'marker':
                if not (0 <= original_index < len(self.markers)): return
                prev_x, prev_y = self.markers[original_index][0:2]
                revert_text = f"{prev_x:.1f},{prev_y:.1f}" # Text to revert to on error
                try:
                    parts = new_value_str.split(',')
                    if len(parts) != 2: raise ValueError("Expected X,Y format")
                    new_x = float(parts[0].strip())
                    new_y = float(parts[1].strip())
                    self.markers[original_index][0] = new_x
                    self.markers[original_index][1] = new_y
                    # print(f"Updated marker {original_index} coords: X={new_x}, Y={new_y}") # Debug
                except (ValueError, IndexError) as e:
                    revert_needed = True
                    error_message = (f"Could not parse marker coordinates:\n{e}\n\n"
                                     f"Expected format: X,Y")
                    # No need to revert internal data, wasn't updated on error

            # --- Shape Coordinate Handling (Col 2) ---
            elif item_type == 'shape':
                if not (0 <= original_index < len(self.shapes)): return

                shape_data = self.shapes[original_index]
                shape_type_internal = shape_data.get('type')

                # Generate previous string representation for revert
                if shape_type_internal == 'line':
                    start = shape_data.get('start', (0,0)); end = shape_data.get('end', (0,0))
                    revert_text = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"
                elif shape_type_internal == 'rectangle':
                    rect = shape_data.get('rect', (0,0,0,0))
                    revert_text = f"{rect[0]:.1f},{rect[1]:.1f},{rect[2]:.1f},{rect[3]:.1f}"
                else: revert_text = "N/A" # Cannot revert if type is unknown

                try:
                    coords_str = new_value_str.split(',')
                    coords_float = [float(c.strip()) for c in coords_str]
                    coords_updated = False
                    if shape_type_internal == 'line' and len(coords_float) == 4:
                        x1, y1, x2, y2 = coords_float
                        shape_data['start'] = (x1, y1); shape_data['end'] = (x2, y2)
                        coords_updated = True
                    elif shape_type_internal == 'rectangle' and len(coords_float) == 4:
                        x, y, w, h = coords_float
                        if w >= 0 and h >= 0: shape_data['rect'] = (x, y, w, h); coords_updated = True
                        else: raise ValueError("Rectangle width/height must be non-negative.")
                    else: raise ValueError(f"Incorrect number of coordinates for {shape_type_internal.capitalize()} (expected 4).")
                    if not coords_updated: raise ValueError("Coordinate update failed.")
                    # print(f"Updated shape {original_index} coords.") # Debug
                except (ValueError, IndexError) as e:
                    revert_needed = True
                    expected_format = "X1,Y1,X2,Y2" if shape_type_internal == 'line' else "X,Y,W,H"
                    error_message = (f"Could not parse shape coordinates:\n{e}\n\n"
                                     f"Expected format for {shape_type_internal.capitalize()}: {expected_format}")
                    # No need to revert internal data, wasn't updated on error

        # --- Revert Cell Text if Necessary ---
        if revert_needed:
            QMessageBox.warning(self, "Invalid Input", error_message)
            self._block_signals = True
            item.setText(revert_text) # Use the generated revert text
            self._block_signals = False


    def handle_cell_double_clicked(self, row, column):
        """Handle double-clicks for Style (col 3), Color (col 6). Ignore cols 1, 2."""
        # --->> Ignore double-clicks on Text/Label (1) and Coordinates (2) <<---
        if column in [1, 2]:
            return

        type_item = self.table_widget.item(row, 0)
        if not type_item: return
        item_data = type_item.data(Qt.UserRole)
        if not item_data or item_data.get('type') == 'error': return

        item_type = item_data['type']
        original_index = item_data['index']

        # --- Color Change (Column 6 - Adjusted Index) ---
        if column == 6:
            # ... (Color change logic remains the same, but use col 6) ...
            current_color = None
            if item_type == 'marker':
                if 0 <= original_index < len(self.markers): current_color = self.markers[original_index][3]
                else: print(f"Error: Stale marker index {original_index} for color."); return
            elif item_type == 'shape':
                if 0 <= original_index < len(self.shapes): current_color = QColor(self.shapes[original_index].get('color', '#000000'))
                else: print(f"Error: Stale shape index {original_index} for color."); return

            if current_color:
                new_color = QColorDialog.getColor(current_color, self, f"Select {item_type.capitalize()} Color")
                if new_color.isValid():
                    if item_type == 'marker': self.markers[original_index][3] = new_color
                    elif item_type == 'shape': self.shapes[original_index]['color'] = new_color.name()
                    # Update table view
                    self._block_signals = True
                    color_item = self.table_widget.item(row, 6); color_item.setText(new_color.name()) # Update col 6
                    color_item.setBackground(QBrush(new_color)); text_color = Qt.white if new_color.lightness() < 128 else Qt.black
                    color_item.setForeground(QBrush(text_color))
                    self._block_signals = False


        # --- Font/Size Change for Markers (Column 3 - Adjusted Index) ---
        elif column == 3 and item_type == 'marker':
            # ... (Font dialog logic remains the same, but use col 3 for style item, cols 4/5 for checkboxes) ...
            if not (0 <= original_index < len(self.markers)): print(f"Error: Stale marker index {original_index} for font."); return
            _, _, _, _, font_family, font_size, is_bold, is_italic = self.markers[original_index]
            initial_qfont = QFont(font_family, font_size); initial_qfont.setBold(is_bold); initial_qfont.setItalic(is_italic)
            selected_font, ok = QFontDialog.getFont(initial_qfont, self, "Select Marker Font")
            if ok:
                self.markers[original_index][4] = selected_font.family()
                self.markers[original_index][5] = selected_font.pointSize()
                self.markers[original_index][6] = selected_font.bold()
                self.markers[original_index][7] = selected_font.italic()
                # Update table view
                self._block_signals = True
                self.table_widget.item(row, 3).setText(f"{selected_font.family()} ({selected_font.pointSize()}pt)") # Update col 3
                bold_widget = self.table_widget.cellWidget(row, 4); italic_widget = self.table_widget.cellWidget(row, 5) # Use cols 4/5
                if bold_widget: bold_widget.findChild(QCheckBox).setChecked(selected_font.bold())
                if italic_widget: italic_widget.findChild(QCheckBox).setChecked(selected_font.italic())
                self._block_signals = False


        # --- Thickness Change for Shapes (Column 3 - Adjusted Index) ---
        elif column == 3 and item_type == 'shape':
             # ... (Thickness dialog logic remains the same, but use col 3 for style item) ...
             if not (0 <= original_index < len(self.shapes)): print(f"Error: Stale shape index {original_index} for thickness."); return
             current_thickness = self.shapes[original_index].get('thickness', 1)
             new_thickness, ok = QInputDialog.getInt(self, "Set Thickness", "Enter line/border thickness (pixels):", current_thickness, 1, 100, 1)
             if ok:
                 self.shapes[original_index]['thickness'] = new_thickness
                 # Update table view
                 self._block_signals = True
                 self.table_widget.item(row, 3).setText(f"{new_thickness}px") # Update col 3
                 self._block_signals = False

    def delete_item(self, row_to_delete):
        """Deletes the item and re-indexes/reconnects remaining items."""
        # --- (Get item info - same as before) ---
        if not (0 <= row_to_delete < self.table_widget.rowCount()): return
        type_item = self.table_widget.item(row_to_delete, 0)
        if not type_item: return
        item_data = type_item.data(Qt.UserRole)
        if not item_data or item_data.get('type') == 'error': return
        item_type = item_data['type']; original_index = item_data['index']
        sort_col = self.table_widget.horizontalHeader().sortIndicatorSection()
        sort_order = self.table_widget.horizontalHeader().sortIndicatorOrder()
        self.table_widget.setSortingEnabled(False)

        # --- (Delete from internal list - same as before) ---
        item_deleted_from_list = False
        if item_type == 'marker':
            if 0 <= original_index < len(self.markers): del self.markers[original_index]; item_deleted_from_list = True
            else: print(f"Warning: Stale original index {original_index} for marker deletion.")
        elif item_type == 'shape':
            if 0 <= original_index < len(self.shapes): del self.shapes[original_index]; item_deleted_from_list = True
            else: print(f"Warning: Stale original index {original_index} for shape deletion.")

        if item_deleted_from_list:
            self.table_widget.removeRow(row_to_delete)
            # --- Re-index UserRole data and reconnect signals (Adjusted Columns) ---
            for current_row in range(self.table_widget.rowCount()):
                # (Re-indexing logic for UserRole is the same)
                current_type_item = self.table_widget.item(current_row, 0)
                if not current_type_item: continue
                current_item_data = current_type_item.data(Qt.UserRole)
                if not current_item_data or current_item_data.get('type') == 'error': continue
                current_item_type = current_item_data['type']; current_original_index = current_item_data['index']
                new_original_index = -1
                if current_item_type == item_type:
                    if current_original_index > original_index: new_original_index = current_original_index - 1
                    else: new_original_index = current_original_index
                else: new_original_index = current_original_index
                current_type_item.setData(Qt.UserRole, {'type': current_item_type, 'index': new_original_index})

                # --- MODIFIED: Reconnect Delete Button (Col 7) ---
                delete_button_widget = self.table_widget.cellWidget(current_row, 7)
                if isinstance(delete_button_widget, QPushButton):
                    try: delete_button_widget.clicked.disconnect()
                    except TypeError: pass
                    delete_button_widget.clicked.connect(lambda checked, row=current_row: self.delete_item(row))

                # --- MODIFIED: Reconnect Checkboxes (Cols 4 & 5) ---
                if current_item_type == 'marker':
                    bold_cell_widget = self.table_widget.cellWidget(current_row, 4) # Col 4
                    if bold_cell_widget:
                        bold_checkbox = bold_cell_widget.findChild(QCheckBox)
                        if bold_checkbox:
                            try: bold_checkbox.stateChanged.disconnect()
                            except TypeError: pass
                            bold_checkbox.stateChanged.connect(lambda state, r=current_row: self.handle_marker_style_changed(state, r, "bold"))
                    italic_cell_widget = self.table_widget.cellWidget(current_row, 5) # Col 5
                    if italic_cell_widget:
                        italic_checkbox = italic_cell_widget.findChild(QCheckBox)
                        if italic_checkbox:
                            try: italic_checkbox.stateChanged.disconnect()
                            except TypeError: pass
                            italic_checkbox.stateChanged.connect(lambda state, r=current_row: self.handle_marker_style_changed(state, r, "italic"))
            # --- End Re-indexing/Reconnecting ---
        else: print(f"Warning: Could not delete item from internal list for row {row_to_delete}.")

        self.table_widget.setSortingEnabled(True)
        if sort_col >= 0: self.table_widget.sortByColumn(sort_col, sort_order)

    def get_modified_markers_and_shapes(self):
        """Returns the modified lists of markers and shapes."""
        final_markers = [tuple(m) for m in self.markers]
        return final_markers, self.shapes


class TableWindow(QDialog):
    """
    A dialog window to display peak analysis results in a table.
    Includes functionality to copy selected data to the clipboard
    in a format pasteable into Excel (tab-separated).
    Also includes Excel export functionality and standard curve plot.
    """
    # *** MODIFIED: Added standard_dictionary to the constructor ***
    def __init__(self, peak_areas, standard_dictionary, standard, calculated_quantities, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Analysis Results and Standard Curve") # Updated title
        self.setGeometry(100, 100, 700, 650) # Increased height for plot
        self.calculated_quantities = calculated_quantities
        self.standard_dictionary = standard_dictionary # Store standard data

        # --- Main Layout ---
        main_layout = QVBoxLayout(self)

        # --- Standard Curve Plot Area (Conditional) ---
        plot_widget = self._create_standard_curve_plot() # Create plot or placeholder
        if plot_widget:
            plot_group = QGroupBox("Standard Curve")
            plot_layout = QVBoxLayout(plot_group)
            plot_layout.addWidget(plot_widget)
            # Set a fixed or maximum height for the plot group if needed
            plot_group.setMaximumHeight(300) # Example max height
            plot_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            main_layout.addWidget(plot_group) # Add plot at the top


        # --- Table Setup ---
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers) # Data is read-only
        self.table.setSelectionBehavior(QTableWidget.SelectRows) # Select whole rows
        self.table.setSelectionMode(QTableWidget.ContiguousSelection) # Allow selecting multiple adjacent rows

        # --- Populate Table ---
        self.populate_table(peak_areas, standard_dictionary, standard)

        self.scroll_area.setWidget(self.table)
        main_layout.addWidget(self.scroll_area) # Add table below plot


        # --- Buttons ---
        self.copy_button = QPushButton("Copy Selected to Clipboard")
        self.copy_button.setToolTip("Copy selected rows to clipboard (tab-separated)")
        self.copy_button.clicked.connect(self.copy_table_data) # Connect copy function

        self.export_button = QPushButton("Export Table to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        # --- Layout Buttons ---
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.copy_button)
        button_layout.addStretch()
        button_layout.addWidget(self.export_button)

        main_layout.addLayout(button_layout) # Add buttons at the bottom
        self.setLayout(main_layout)

        # Enable keyboard copy (Ctrl+C / Cmd+C)
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.table)
        self.copy_shortcut.activated.connect(self.copy_table_data)
        # Make table focusable to receive shortcut events
        self.table.setFocusPolicy(Qt.StrongFocus)


    def _create_standard_curve_plot(self):
        """Creates and returns the Matplotlib canvas for the standard curve, or None."""
        if not self.standard_dictionary or len(self.standard_dictionary) < 2:
            # Return a placeholder label if no/insufficient standard data
            no_curve_label = QLabel("Standard curve requires at least 2 standard points.")
            no_curve_label.setAlignment(Qt.AlignCenter)
            return no_curve_label

        try:
            # Extract data
            quantities = np.array(list(self.standard_dictionary.keys()), dtype=float)
            areas = np.array(list(self.standard_dictionary.values()), dtype=float)

            # Perform linear regression (Area vs Quantity)
            coeffs = np.polyfit(areas, quantities, 1)
            slope, intercept = coeffs

            # Calculate R-squared
            predicted_quantities = np.polyval(coeffs, areas)
            residuals = quantities - predicted_quantities
            ss_res = np.sum(residuals**2)
            ss_tot = np.sum((quantities - np.mean(quantities))**2)
            r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1.0

            # Create Matplotlib figure and axes
            fig, ax = plt.subplots(figsize=(5, 3)) # Keep adjusted size
            fig.set_dpi(90)

            # Plot data points
            ax.scatter(areas, quantities, label='Standard Points', color='red', zorder=5)

            # Plot regression line
            x_line = np.linspace(min(areas) * 0.95, max(areas) * 1.05, 100)
            y_line = slope * x_line + intercept

            # *** MODIFICATION: Create multi-line label for the legend ***
            fit_label = (
                f'Linear Fit\n'
                f'Qty = {slope}*Area + {intercept}\n'
                f'R² = {r_squared:.3f}'
            )
            ax.plot(x_line, y_line, label=fit_label, color='blue') # Use the combined label

            # *** REMOVED: Separate text annotation is no longer needed ***
            # eq_text = f'Qty = {slope:.2f} * Area + {intercept:.2f}\nR² = {r_squared:.3f}'
            # ax.text(0.05, 0.05, eq_text, transform=ax.transAxes, fontsize=8,
            #         verticalalignment='bottom', # Align text bottom to position
            #         horizontalalignment='left', # Align text left to position
            #         bbox=dict(boxstyle='round,pad=0.3', fc='wheat', alpha=0.5))

            # Customize plot
            ax.set_xlabel('Total Peak Area (Arbitrary Units)', fontsize=9)
            ax.set_ylabel('Known Quantity (Unit)', fontsize=9)
            ax.set_title('Standard Curve', fontsize=10, fontweight='bold')
            # Keep legend font size small and specify location
            ax.legend(fontsize=8, loc='best') # Use 'best' or 'upper right', etc.
            ax.grid(True, linestyle='--', alpha=0.6)
            ax.tick_params(axis='both', which='major', labelsize=8)
            ax.ticklabel_format(style='sci', axis='x', scilimits=(0,0), useMathText=True)

            # Use constrained_layout for potentially better automatic spacing
            try:
                fig.set_constrained_layout(True)
            except AttributeError: # Fallback for older Matplotlib
                plt.tight_layout(pad=0.5)

            # Create canvas
            canvas = FigureCanvas(fig)
            canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            canvas.updateGeometry()
            return canvas

        except Exception as e:
            print(f"Error creating standard curve plot: {e}")
            error_label = QLabel(f"Error generating plot:\n{e}")
            error_label.setAlignment(Qt.AlignCenter)
            error_label.setStyleSheet("color: red;")
            return error_label


    def populate_table(self, peak_areas, standard_dictionary, standard):
        """Populates the table with the provided peak analysis data."""
        total_area = sum(peak_areas) if peak_areas else 0.0 # Handle empty list
        self.table.setRowCount(len(peak_areas))

        for row, area in enumerate(peak_areas):
            band_label = f"Band {row + 1}"
            self.table.setItem(row, 0, QTableWidgetItem(band_label))

            peak_area_rounded = round(area, 3)
            self.table.setItem(row, 1, QTableWidgetItem(str(peak_area_rounded)))

            if total_area != 0:
                percentage = (area / total_area) * 100
                percentage_rounded = round(percentage, 2)
            else:
                percentage_rounded = 0.0
            self.table.setItem(row, 2, QTableWidgetItem(f"{percentage_rounded:.2f}%"))

            # --- Use pre-calculated quantities if available ---
            quantity_str = "" # Default empty string
            if standard and row < len(self.calculated_quantities): # Check if standard mode and index is valid
                quantity_str = f"{self.calculated_quantities[row]:.2f}" # Format as needed
            # Always set the item, even if empty
            self.table.setItem(row, 3, QTableWidgetItem(quantity_str))
            # --- End quantity handling ---

        self.table.resizeColumnsToContents() # Adjust column widths


    def copy_table_data(self):
        """Copy selected table data to the clipboard in a tab-separated format."""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return # Nothing selected

        # We copy contiguous blocks as selected. If multiple non-contiguous blocks are selected,
        # this standard implementation will only copy the first one.
        # For pasting into Excel, copying a single contiguous block is usually expected.
        selected_range = selected_ranges[0]
        start_row = selected_range.topRow()
        end_row = selected_range.bottomRow()
        start_col = selected_range.leftColumn()
        end_col = selected_range.rightColumn()

        clipboard_string = ""
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            clipboard_string += "\t".join(row_data) + "\n" # Tab separated, newline at end of row

        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_string.strip()) # Remove trailing newline


    def export_to_excel(self):
        """Export the table data to an Excel file."""
        # Prompt the user to select a save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
        )
        if not file_path:
            return

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Peak Analysis Results" # More descriptive title

        # Write the table headers to the Excel sheet
        headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)  # Make headers bold

        # Write the table data to the Excel sheet
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                value = item.text() if item else ""
                # Attempt to convert numeric strings back to numbers for Excel
                try:
                    if '%' in value: # Handle percentages
                        numeric_value = float(value.replace('%', '')) / 100.0
                        cell = worksheet.cell(row=row + 2, column=col + 1, value=numeric_value)
                        cell.number_format = '0.00%' # Apply percentage format
                    else:
                        numeric_value = float(value)
                        worksheet.cell(row=row + 2, column=col + 1, value=numeric_value)
                except ValueError:
                    # Keep as text if conversion fails
                    worksheet.cell(row=row + 2, column=col + 1, value=value)

        # Auto-adjust column widths (optional, but nice)
        for col_idx, column_letter in enumerate(worksheet.columns, start=1):
            max_length = 0
            column = column_letter[0].column_letter # Get the column letter ('A', 'B', ...)
            for cell in worksheet[column]:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # Add padding and factor
            worksheet.column_dimensions[column].width = adjusted_width


        # Save the Excel file
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Success", f"Table data exported to\n{file_path}")
        except Exception as e:
             QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                

class PeakAreaDialog(QDialog):
    """
    Interactive dialog to adjust peak regions and calculate peak areas.
    Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
    Area calculations use original data range (inverted).
    Peak detection uses a 0-255 scaled profile (inverted).
    Initial regions defined by outward trough search. V-V baseline uses adjacent troughs.
    """

    def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
        """
        Initializes the dialog.
        Args:
            cropped_data (PIL.Image.Image): The cropped grayscale image data.
            current_settings (dict): Dictionary of previous settings.
            persist_checked (bool): Initial state of the 'persist settings' checkbox.
            parent (QWidget): Parent widget.
        """
        super().__init__(parent)
        self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
        self.setGeometry(100, 100, 1100, 850) # Keep original size for now

        # --- Validate and Store Input Image ---
        if not isinstance(cropped_data, Image.Image):
             raise TypeError("Input 'cropped_data' must be a PIL Image object")
        self.cropped_image_for_display = cropped_data # Keep original PIL for display

        self.original_max_value = 255.0 # Default assumption
        pil_mode = cropped_data.mode

        # Determine intensity range and create numpy array
        try:
            if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                self.original_max_value = 65535.0
            elif pil_mode == 'L':
                self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                self.original_max_value = 255.0
            elif pil_mode == 'F': # Handle float images
                self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
                max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                self.original_max_value = max(1.0, max_in_float) # Use max value or 1.0
                # Scale float for display (assuming 0-max range)
                scaled_for_display = np.clip(self.intensity_array_original_range * 255.0 / self.original_max_value, 0, 255).astype(np.uint8)
                self.cropped_image_for_display = Image.fromarray(scaled_for_display, mode='L')
            else: # Attempt conversion to grayscale 'L' as fallback
                gray_img = cropped_data.convert("L")
                self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                self.original_max_value = 255.0
                self.cropped_image_for_display = gray_img # Use the converted image for display
        except Exception as e:
            raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

        if self.intensity_array_original_range.ndim != 2:
             raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

        self.profile_original_inverted = None # Smoothed, inverted profile (original range)
        self.profile = None # Scaled (0-255), inverted, SMOOTHED profile for detection
        self.background = None # Rolling ball background estimate

        # --- Settings and State ---
        self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
        self.smoothing_sigma = current_settings.get('smoothing_sigma', 2.0)
        self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
        self.peak_distance = current_settings.get('peak_distance', 30)
        self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.02)
        # --- Use Valley Offset ---
        self.valley_offset_pixels = current_settings.get('valley_offset_pixels', 0) # Default to 0 offset from TROUGH
        # --- End ---
        self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
        self.area_subtraction_method = current_settings.get('area_subtraction_method', "Valley-to-Valley")
        self.peaks = np.array([])
        # --- Store initial TROUGH regions ---
        self.initial_valley_regions = [] # Store (valley_start, valley_end) tuples
        # --- End ---
        self.peak_regions = [] # Store final (potentially offset) regions (start, end)
        self.peak_areas_rolling_ball = []
        self.peak_areas_straight_line = []
        self.peak_areas_valley = []
        self.peak_sliders = []
        self._final_settings = {}
        self._persist_enabled_on_exit = persist_checked

        # Check dependencies
        if rolling_ball is None or find_peaks is None or gaussian_filter1d is None or interp1d is None:
             QMessageBox.critical(self, "Dependency Error",
                                  "Missing SciPy or scikit-image library.\n"
                                  "Peak detection, smoothing, and rolling ball background features require these libraries.\n"
                                  "Please install them (e.g., 'pip install scipy scikit-image') and restart.")
             # Optionally disable controls that depend on these libraries
             # self.close() # Or close the dialog immediately

        # Build UI & Initial Setup
        self._setup_ui(persist_checked)
        self.regenerate_profile_and_detect() # Initial calculation

    def _setup_ui(self, persist_checked_initial):
        """Creates and arranges the UI elements."""
        # --- Main Layout ---
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # --- Matplotlib Plot Canvas ---
        self.fig = plt.figure(figsize=(10, 5))
        self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]);
        self.fig.tight_layout(pad=2)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.canvas, stretch=3)

        # --- Controls Layout (Horizontal Box for side-by-side groups) ---
        controls_hbox = QHBoxLayout()
        controls_hbox.setSpacing(15)

        # --- Left Controls Column (Global & Detection) ---
        left_controls_vbox = QVBoxLayout()

        # Group 1: Global Settings
        global_settings_group = QGroupBox("Global Settings")
        global_settings_layout = QGridLayout(global_settings_group)
        global_settings_layout.setSpacing(8)
        # Band Estimation
        self.band_estimation_combobox = QComboBox()
        self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
        self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
        self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
        global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
        global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2)
        # Area Subtraction Method
        self.method_combobox = QComboBox()
        self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
        self.method_combobox.setCurrentText(self.area_subtraction_method)
        self.method_combobox.currentIndexChanged.connect(self.update_plot)
        global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
        global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2)
        # Rolling Ball Radius
        global_settings_layout.addWidget(QLabel("Rolling Ball Radius:"), 2, 0)
        self.rolling_ball_slider = QSlider(Qt.Horizontal)
        self.rolling_ball_slider.setRange(1, 500)
        self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
        self.rolling_ball_slider.setEnabled(False)
        self.rolling_ball_slider.valueChanged.connect(self.update_plot) # Only needs plot update
        self.rolling_ball_value_label = QLabel(f"({int(self.rolling_ball_radius)})")
        fm = QFontMetrics(self.rolling_ball_value_label.font())
        self.rolling_ball_value_label.setMinimumWidth(fm.horizontalAdvance("(500) "))
        self.rolling_ball_value_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_value_label: lbl.setText(f"({val})"))
        global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1)
        global_settings_layout.addWidget(self.rolling_ball_value_label, 2, 2)
        left_controls_vbox.addWidget(global_settings_group)

        # Group 2: Peak Detection Parameters
        peak_detect_group = QGroupBox("Peak Detection Parameters")
        peak_detect_layout = QGridLayout(peak_detect_group)
        peak_detect_layout.setSpacing(8)
        # Manual Peak Number
        self.peak_number_label = QLabel("Detected Peaks:")
        self.peak_number_input = QLineEdit()
        self.peak_number_input.setPlaceholderText("#")
        self.peak_number_input.setMaximumWidth(60)
        self.update_peak_number_button = QPushButton("Set")
        self.update_peak_number_button.setToolTip("Manually override the number of peaks detected.")
        self.update_peak_number_button.clicked.connect(self.manual_peak_number_update)
        peak_detect_layout.addWidget(self.peak_number_label, 0, 0)
        peak_detect_layout.addWidget(self.peak_number_input, 0, 1)
        peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)
        # Smoothing Sigma
        self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.smoothing_slider = QSlider(Qt.Horizontal)
        self.smoothing_slider.setRange(0, 100) # 0.0 to 10.0
        self.smoothing_slider.setValue(int(self.smoothing_sigma * 10))
        self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})"))
        self.smoothing_slider.valueChanged.connect(self.regenerate_profile_and_detect)
        peak_detect_layout.addWidget(self.smoothing_label, 1, 0)
        peak_detect_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2)
        # Peak Prominence
        self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        self.peak_prominence_slider = QSlider(Qt.Horizontal)
        self.peak_prominence_slider.setRange(0, 100) # 0.0 to 1.0 factor
        self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
        self.peak_prominence_slider.valueChanged.connect(self.detect_peaks)
        self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})"))
        peak_detect_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
        peak_detect_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)
        # Peak Height
        self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})")
        self.peak_height_slider = QSlider(Qt.Horizontal)
        self.peak_height_slider.setRange(0, 100)
        self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
        self.peak_height_slider.valueChanged.connect(self.detect_peaks)
        self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f})"))
        peak_detect_layout.addWidget(self.peak_height_slider_label, 3, 0)
        peak_detect_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)
        # Peak Distance
        self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
        self.peak_distance_slider = QSlider(Qt.Horizontal)
        self.peak_distance_slider.setRange(1, 200)
        self.peak_distance_slider.setValue(self.peak_distance)
        self.peak_distance_slider.valueChanged.connect(self.detect_peaks)
        self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
        peak_detect_layout.addWidget(self.peak_distance_slider_label, 4, 0)
        peak_detect_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)
        left_controls_vbox.addWidget(peak_detect_group)
        left_controls_vbox.addStretch(1)
        controls_hbox.addLayout(left_controls_vbox, stretch=1)

        # --- Right Controls Column (Peak Region Adjustments) ---
        right_controls_vbox = QVBoxLayout()
        peak_spread_group = QGroupBox("Peak Region Adjustments")
        peak_spread_layout = QGridLayout(peak_spread_group)
        peak_spread_layout.setSpacing(8)

        # --- Global Valley Offset Slider ---
        self.valley_offset_label = QLabel(f"Valley Offset (+/- {self.valley_offset_pixels} px)") # Use valley name
        self.valley_offset_slider = QSlider(Qt.Horizontal)
        self.valley_offset_slider.setRange(-20, 100) # Allow inward and outward offset
        self.valley_offset_slider.setValue(self.valley_offset_pixels)
        self.valley_offset_slider.setToolTip(
            "Applies an offset to the automatically detected valley boundaries.\n"
            "0 = Use exact valley positions. Positive expands outwards, Negative contracts inwards."
        )
        self.valley_offset_slider.valueChanged.connect(self.apply_valley_offset) # Connect to valley offset function
        self.valley_offset_slider.valueChanged.connect(
            lambda value, lbl=self.valley_offset_label: lbl.setText(f"Valley Offset ({'+/-' if value>=0 else ''}{value} px)") # Show +/- correctly
        )
        peak_spread_layout.addWidget(self.valley_offset_label, 0, 0)
        peak_spread_layout.addWidget(self.valley_offset_slider, 0, 1)
        # --- END MODIFICATION ---

        right_controls_vbox.addWidget(peak_spread_group)

        # Scroll area for individual peak sliders
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(250)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.container = QWidget()
        self.peak_sliders_layout = QVBoxLayout(self.container)
        self.peak_sliders_layout.setSpacing(10)
        scroll_area.setWidget(self.container)
        right_controls_vbox.addWidget(scroll_area, stretch=1)
        controls_hbox.addLayout(right_controls_vbox, stretch=2)
        main_layout.addLayout(controls_hbox)

        # --- Bottom Button Layout ---
        bottom_button_layout = QHBoxLayout()
        self.persist_settings_checkbox = QCheckBox("Persist Settings")
        self.persist_settings_checkbox.setChecked(persist_checked_initial)
        self.persist_settings_checkbox.setToolTip("Save current detection parameters for the next time this dialog opens.")
        bottom_button_layout.addWidget(self.persist_settings_checkbox)
        bottom_button_layout.addStretch(1)
        self.ok_button = QPushButton("OK")
        self.ok_button.setMinimumWidth(100)
        self.ok_button.setDefault(True)
        self.ok_button.clicked.connect(self.accept_and_close)
        bottom_button_layout.addWidget(self.ok_button)
        main_layout.addLayout(bottom_button_layout)

    def accept_and_close(self):
        """Save settings and close the dialog."""
        self._final_settings = {
            'rolling_ball_radius': self.rolling_ball_slider.value(),
            'peak_height_factor': self.peak_height_slider.value() / 100.0,
            'peak_distance': self.peak_distance_slider.value(),
            'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
            'valley_offset_pixels': self.valley_offset_slider.value(), # Save valley offset
            'band_estimation_method': self.band_estimation_combobox.currentText(),
            'area_subtraction_method': self.method_combobox.currentText(),
            'smoothing_sigma': self.smoothing_slider.value() / 10.0,
        }
        self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
        self.accept()
        
    def _custom_rolling_ball(self, profile, radius):
        """
        Calculates background using morphological opening (erosion followed by dilation).
        This simulates a 'ball' rolling under the inverted profile.

        Args:
            profile (np.ndarray): The 1D intensity profile (inverted, peaks are high).
            radius (int): The radius of the conceptual rolling ball.

        Returns:
            np.ndarray: The calculated background profile.
        """
        # Check if SciPy functions are available
        if grey_opening is None:
            print("Error: SciPy (grey_opening) is required for custom rolling ball but not found.")
            return np.zeros_like(profile) # Return zeros if function missing

        # --- Input Validation ---
        if profile is None or profile.ndim != 1 or profile.size == 0:
            print("Warning (_custom_rolling_ball): Invalid profile provided.")
            return np.zeros_like(profile) if profile is not None else np.array([])
        if radius <= 0:
            print("Warning (_custom_rolling_ball): Radius must be positive.")
            # Return profile itself or zeros? Returning zeros is safer for background.
            return np.zeros_like(profile)

        profile_len = profile.shape[0]
        # --- Structure Size ---
        # The structure size should represent the diameter of the ball influence.
        # Size = 2 * radius + 1 (for center pixel)
        structure_size = int(max(1, 2 * radius + 1)) # Ensure at least size 1

        # Ensure structure size is not larger than the profile itself
        if structure_size > profile_len:
            # print(f"Warning (_custom_rolling_ball): Structure size ({structure_size}) larger than profile ({profile_len}). Clamping.")
            structure_size = profile_len

        # --- Morphological Opening ---
        # grey_opening = erosion followed by dilation. This removes features smaller
        # than the structure (peaks) and gives the lower envelope.
        try:
            # Use 'reflect' mode for handling edges reasonably well.
            background = grey_opening(profile, size=structure_size, mode='reflect')

            # --- Optional Refinement (Not standard opening, but closer to ball touching troughs) ---
            # Sometimes, opening can slightly lower the background in wide troughs.
            # A dilation of the original profile followed by taking the minimum
            # with the opened profile can sometimes help pull the background up
            # in wider valleys. This is closer to how some rolling ball implementations work.
            # dilated_profile = grey_dilation(profile, size=structure_size, mode='reflect')
            # background = np.minimum(background, dilated_profile) # Take the lower of the two results? NO - opening is already lower env.
            # Let's stick to standard opening for simplicity and performance unless refinement is strictly needed.

        except Exception as e:
            print(f"Error during morphological opening: {e}")
            traceback.print_exc()
            background = np.zeros_like(profile) # Fallback on error

        return background

    def get_current_settings(self):
        """Returns the final settings dictionary."""
        return self._final_settings

    def should_persist_settings(self):
        """Returns whether the persist checkbox was checked."""
        return self._persist_enabled_on_exit

    def regenerate_profile_and_detect(self):
        """
        Calculates the raw inverted profile, applies smoothing, stores this as the
        main profile, scales a copy for peak detection, and detects peaks/valleys.
        """
        if gaussian_filter1d is None: return # Cannot proceed without scipy

        self.band_estimation_method = self.band_estimation_combobox.currentText()
        self.area_subtraction_method = self.method_combobox.currentText()
        if hasattr(self, 'smoothing_slider'):
             self.smoothing_sigma = self.smoothing_slider.value() / 10.0
        else: self.smoothing_sigma = 2.0 # Default if slider not ready

        # --- Calculate profile from ORIGINAL intensity data ---
        profile_temp = None
        if self.band_estimation_method == "Mean":
            profile_temp = np.mean(self.intensity_array_original_range, axis=1)
        elif self.band_estimation_method.startswith("Percentile"):
            try:
                percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
            except:
                profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1)
                print("Warning: Defaulting to 5th percentile for profile.")
        else:
            profile_temp = np.mean(self.intensity_array_original_range, axis=1) # Default to mean

        if profile_temp is None or not np.all(np.isfinite(profile_temp)):
            print("Warning: Profile calculation failed or resulted in NaN/Inf. Using zeros.")
            profile_temp = np.zeros(self.intensity_array_original_range.shape[0])

        # --- Create INVERTED Original Profile and shift baseline to zero ---
        profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
        min_inverted_raw = np.min(profile_original_inv_raw)
        profile_original_inv_raw -= min_inverted_raw

        # --- Apply Smoothing to the INVERTED ORIGINAL profile ---
        self.profile_original_inverted = profile_original_inv_raw
        try:
            current_sigma = self.smoothing_sigma
            if current_sigma > 0.1 and len(self.profile_original_inverted) > int(3 * current_sigma) * 2 + 1: # Ensure profile long enough for filter kernel
                self.profile_original_inverted = gaussian_filter1d(self.profile_original_inverted, sigma=current_sigma)
            # else: print("Skipping smoothing (sigma too small or profile too short)") # Debug
        except Exception as smooth_err:
            print(f"Error smoothing main profile: {smooth_err}")
            # Keep raw profile if smoothing fails

        # --- Create the SCALED (0-255) version FOR PEAK DETECTION ONLY ---
        prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
        if prof_max_inv > prof_min_inv + 1e-6: # Avoid division by zero for flat profiles
            self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
        else:
            self.profile = np.zeros_like(self.profile_original_inverted) # Handle flat profile

        # Detect peaks using the scaled profile, then find troughs on original
        self.detect_peaks()


    def _find_outward_troughs(self, profile, peak_idx, left_bound, right_bound):
        """
        Finds the nearest local minima outwards from peak_idx within bounds.
        Stops searching if the profile starts increasing significantly.
        Returns indices (valley_left_idx, valley_right_idx).
        """
        profile_len = len(profile)
        # Basic validation
        if not (0 <= left_bound <= peak_idx <= right_bound < profile_len):
            print(f"Warning (_find_outward_troughs): Invalid bounds [{left_bound},{right_bound}] or peak {peak_idx}")
            w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
            return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)

        # --- Search Left ---
        valley_left_idx = peak_idx # Default
        min_val_left = profile[peak_idx]
        found_trough_left = False
        # Search from peak-1 down to left_bound
        for idx in range(peak_idx - 1, left_bound - 1, -1):
            current_val = profile[idx]
            # Potential Trough Check: Is it lower than the point closer to the peak?
            if current_val <= profile[idx + 1]:
                # Check if it's a local minimum (lower than both neighbors)
                is_local_min = True
                if idx > 0 and profile[idx - 1] < current_val:
                    is_local_min = False
                # We already know profile[idx+1] >= current_val from above check

                if is_local_min:
                    valley_left_idx = idx
                    found_trough_left = True
                    break # Found first local minimum going left

                # If not local min, but lower than current lowest, update potential index
                if current_val < min_val_left:
                     min_val_left = current_val
                     valley_left_idx = idx

            else: # Profile started increasing significantly going left
                # If we had already found a point lower than the peak, use that index.
                # Otherwise, the point just before the rise started is the effective edge.
                if valley_left_idx == peak_idx: # Haven't found a lower point yet
                    valley_left_idx = idx + 1 # Point before the rise
                found_trough_left = True # Consider the point before rise as the boundary
                break

        # If loop finished without finding a rise or clear minimum, use the lowest point found
        if not found_trough_left and valley_left_idx == peak_idx:
             # If still at peak_idx, maybe the profile was flat or always decreasing?
             # Check boundary value
             if left_bound < peak_idx and profile[left_bound] <= min_val_left:
                 valley_left_idx = left_bound
             else: # Fallback to a small offset if no lower point found
                  valley_left_idx = max(0, peak_idx - 1)


        # --- Search Right ---
        valley_right_idx = peak_idx # Default
        min_val_right = profile[peak_idx]
        found_trough_right = False
        # Search from peak+1 up to right_bound
        for idx in range(peak_idx + 1, right_bound + 1, 1):
            current_val = profile[idx]
            # Potential Trough Check: Is it lower than the point closer to the peak?
            if current_val <= profile[idx - 1]:
                is_local_min = True
                if idx > peak_idx + 1 and profile[idx - 1] < current_val: # Check left neighbor (towards peak)
                     is_local_min = False
                if idx < profile_len - 1 and profile[idx + 1] < current_val: # Check right neighbor
                    is_local_min = False

                if is_local_min:
                    valley_right_idx = idx
                    found_trough_right = True
                    break

                if current_val < min_val_right:
                    min_val_right = current_val
                    valley_right_idx = idx

            else: # Profile started increasing significantly going right
                if valley_right_idx == peak_idx:
                    valley_right_idx = idx - 1 # Point before the rise
                found_trough_right = True
                break

        if not found_trough_right and valley_right_idx == peak_idx:
            if right_bound > peak_idx and profile[right_bound] <= min_val_right:
                 valley_right_idx = right_bound
            else:
                 valley_right_idx = min(profile_len - 1, peak_idx + 1)


        # Final safety checks
        valley_left_idx = max(0, min(peak_idx, valley_left_idx))
        valley_right_idx = min(profile_len - 1, max(peak_idx, valley_right_idx))
        if valley_left_idx >= valley_right_idx: # Handle overlap
             w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
             valley_left_idx = max(0, peak_idx - w)
             valley_right_idx = min(profile_len - 1, peak_idx + w)
             if valley_left_idx >= valley_right_idx: # Ensure distinct points
                 if valley_right_idx < profile_len - 1: valley_right_idx += 1
                 elif valley_left_idx > 0: valley_left_idx -= 1


        return valley_left_idx, valley_right_idx


    def detect_peaks(self):
        """
        Detect peak locations using the scaled profile, then find the initial
        valley regions around each peak by searching outwards on the original profile.
        Finally, applies the current valley offset.
        """
        if self.profile is None or len(self.profile) == 0 or find_peaks is None:
            self.peaks, self.initial_valley_regions, self.peak_regions = np.array([]), [], []
            if hasattr(self, 'peak_number_input'): self.peak_number_input.setText("0")
            self.update_sliders(); self.update_plot()
            return

        # --- Update parameters ---
        self.peak_height_factor = self.peak_height_slider.value() / 100.0
        self.peak_distance = self.peak_distance_slider.value()
        self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
        self.valley_offset_pixels = self.valley_offset_slider.value() # Use valley name
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Update UI Labels ---
        if hasattr(self, 'smoothing_label'): self.smoothing_label.setText(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.peak_height_slider_label.setText(f"Min Height ({self.peak_height_factor:.2f})")
        self.peak_distance_slider_label.setText(f"Min Distance ({self.peak_distance}) px")
        self.peak_prominence_slider_label.setText(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        if hasattr(self, 'valley_offset_label'): self.valley_offset_label.setText(f"Valley Offset ({'+/-' if self.valley_offset_pixels>=0 else ''}{self.valley_offset_pixels} px)")

        # --- Thresholds ---
        profile_range = np.ptp(self.profile); min_val_profile = np.min(self.profile)
        if profile_range < 1e-6 : profile_range = 1.0
        min_height_abs = min_val_profile + profile_range * self.peak_height_factor
        min_prominence_abs = profile_range * self.peak_prominence_factor
        min_prominence_abs = max(1.0, min_prominence_abs)

        # --- Find Peak Indices ---
        try:
            peaks_indices, properties = find_peaks(
                self.profile, height=min_height_abs, prominence=min_prominence_abs,
                distance=self.peak_distance, width=1
            )
            self.peaks = np.sort(peaks_indices) # Ensure sorted
        except Exception as e:
            QMessageBox.warning(self, "Peak Detection Error", f"Error finding peak locations:\n{e}")
            self.peaks = np.array([]); self.initial_valley_regions = []; self.peak_regions = []
            if hasattr(self, 'peak_number_input') and not self.peak_number_input.hasFocus(): self.peak_number_input.setText("0")
            self.update_sliders(); self.update_plot()
            return

        # --- Find initial VALLEY regions using outward search ---
        self.initial_valley_regions = [] # Use valley name
        profile_to_analyze = self.profile_original_inverted
        profile_len = len(profile_to_analyze)

        if profile_len <= 1 or len(self.peaks) == 0:
             self.peaks = np.array([]); self.initial_valley_regions = []; self.peak_regions = []
             if hasattr(self, 'peak_number_input') and not self.peak_number_input.hasFocus(): self.peak_number_input.setText("0")
             self.update_sliders(); self.update_plot()
             return

        # Define search boundaries using midpoints between peaks
        if len(self.peaks) > 1:
            midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2
            search_boundaries_left = np.concatenate(([0], midpoints))
            search_boundaries_right = np.concatenate((midpoints, [profile_len - 1]))
        elif len(self.peaks) == 1: # Handle single peak case
             search_boundaries_left = np.array([0])
             search_boundaries_right = np.array([profile_len - 1])
        else: # No peaks found, should have returned earlier
             search_boundaries_left = np.array([])
             search_boundaries_right = np.array([])


        for i, peak_idx in enumerate(self.peaks):
             left_bound = search_boundaries_left[i]
             right_bound = search_boundaries_right[i]
             try:
                 valley_start, valley_end = self._find_outward_troughs(
                     profile_to_analyze, peak_idx, int(left_bound), int(right_bound) # Ensure bounds are int
                 )
                 self.initial_valley_regions.append((valley_start, valley_end))
             except Exception as e_trough:
                 print(f"Error finding troughs for peak {i} (idx {peak_idx}): {e_trough}")
                 # Fallback
                 fallback_width = max(2, self.peak_distance // 4)
                 fb_start = max(0, peak_idx - fallback_width)
                 fb_end = min(profile_len - 1, peak_idx + fallback_width)
                 if fb_start >= fb_end: fb_end = min(profile_len - 1, fb_start + 1)
                 self.initial_valley_regions.append((fb_start, fb_end))


        # Update peak number input
        if hasattr(self, 'peak_number_input') and (not self.peak_number_input.hasFocus() or self.peak_number_input.text() == ""):
             self.peak_number_input.setText(str(len(self.peaks)))

        # Apply the current valley offset to generate the initial peak_regions
        self.apply_valley_offset(self.valley_offset_pixels) # Calls update_sliders & update_plot


    # --- Renamed back apply_valley_offset ---
    def apply_valley_offset(self, offset_value):
        """
        Applies the global valley offset value to the initial valley regions
        to calculate the final peak_regions used for integration and sliders.
        """
        self.valley_offset_pixels = offset_value # Store the current offset
        self.peak_regions = [] # Clear previous final regions

        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            self.update_sliders(); self.update_plot(); return

        profile_len = len(self.profile_original_inverted)
        # Use initial_valley_regions now
        num_initial = min(len(self.peaks), len(self.initial_valley_regions))
        if len(self.peaks) != len(self.initial_valley_regions):
            print(f"Warning: Peak ({len(self.peaks)}) / initial valley region ({len(self.initial_valley_regions)}) mismatch.")

        for i in range(num_initial):
            try:
                valley_start, valley_end = self.initial_valley_regions[i]

                # Apply the offset (positive expands, negative contracts)
                new_start = valley_start - self.valley_offset_pixels
                new_end = valley_end + self.valley_offset_pixels

                # Clamp to profile boundaries
                new_start_clamped = max(0, new_start)
                new_end_clamped = min(profile_len - 1, new_end)

                # Ensure start is not after end
                if new_start_clamped > new_end_clamped:
                    # If offset causes overlap, use the midpoint of the original valley as a single point
                    mid_valley = (valley_start + valley_end) // 2
                    new_start_clamped = mid_valley
                    new_end_clamped = mid_valley


                self.peak_regions.append((new_start_clamped, new_end_clamped))

            except IndexError:
                 print(f"Error accessing initial valley region at index {i}")
                 continue

        if len(self.peak_regions) != num_initial:
             print(f"Warning: Final peak_regions length ({len(self.peak_regions)}) mismatch after offset application.")

        self.update_sliders()
        self.update_plot()


    def manual_peak_number_update(self):
        """Handles manual changes to the number of peaks, then re-finds valleys."""
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            QMessageBox.warning(self, "Error", "Profile must be generated first."); return

        profile_len = len(self.profile_original_inverted)
        try:
            num_peaks_manual = int(self.peak_number_input.text())
            if num_peaks_manual < 0: raise ValueError("Negative number")
            current_num_peaks = len(self.peaks)
            if num_peaks_manual == current_num_peaks: return

            peaks_list = self.peaks.tolist()
            if num_peaks_manual == 0:
                self.peaks = np.array([])
            elif num_peaks_manual < current_num_peaks:
                # Simple truncation - might need smarter selection based on peak properties
                self.peaks = self.peaks[:num_peaks_manual]
            else: # Add dummy peaks
                num_to_add = num_peaks_manual - current_num_peaks
                profile_center = profile_len // 2
                current_peaks_set = set(self.peaks)
                for _ in range(num_to_add):
                    new_peak_pos = profile_center; offset = 0
                    while new_peak_pos in current_peaks_set or new_peak_pos < 0 or new_peak_pos >= profile_len:
                        offset += 5
                        new_peak_pos = profile_center + np.random.choice([-offset, offset])
                        if offset > profile_len // 2:
                            new_peak_pos = np.random.randint(0, profile_len)
                            if new_peak_pos in current_peaks_set: continue
                            break
                    peaks_list.append(new_peak_pos)
                    current_peaks_set.add(new_peak_pos)
                peaks_list.sort()
                self.peaks = np.array(peaks_list)

            # --- Re-find VALLEY regions for the new set of peaks ---
            self.initial_valley_regions = [] # Use correct name
            profile_to_analyze = self.profile_original_inverted
            if profile_len > 1 and len(self.peaks) > 0:
                 # Define search boundaries using midpoints between peaks
                if len(self.peaks) > 1:
                    midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2
                    search_boundaries_left = np.concatenate(([0], midpoints))
                    search_boundaries_right = np.concatenate((midpoints, [profile_len - 1]))
                else: # Single peak
                     search_boundaries_left = np.array([0])
                     search_boundaries_right = np.array([profile_len - 1])

                for i, peak_idx in enumerate(self.peaks):
                     left_bound = search_boundaries_left[i]
                     right_bound = search_boundaries_right[i]
                     try:
                         valley_start, valley_end = self._find_outward_troughs(
                             profile_to_analyze, peak_idx, int(left_bound), int(right_bound)
                         )
                         self.initial_valley_regions.append((valley_start, valley_end))
                     except Exception as e_trough:
                         print(f"Error finding troughs after manual peak update for peak {i}: {e_trough}")
                         fallback_width = max(2, self.peak_distance // 4)
                         fb_start = max(0, peak_idx - fallback_width); fb_end = min(profile_len - 1, peak_idx + fallback_width)
                         if fb_start >= fb_end: fb_end = min(profile_len - 1, fb_start + 1)
                         self.initial_valley_regions.append((fb_start, fb_end))
            # --- END VALLEY RE-FIND ---

            # Apply the current valley offset based on the new peaks/valleys
            self.apply_valley_offset(self.valley_offset_slider.value())

        except ValueError:
            self.peak_number_input.setText(str(len(self.peaks)))
            QMessageBox.warning(self, "Input Error", "Please enter a valid non-negative integer.")
        except Exception as e:
            print(f"Error during manual peak number update: {e}")
            QMessageBox.critical(self, "Error", f"Manual peak update error:\n{e}")
            self.peak_number_input.setText(str(len(self.peaks)))


    def update_sliders(self):
        """Update individual peak sliders based on the current self.peak_regions."""
        # Clear existing sliders
        while self.peak_sliders_layout.count():
            item = self.peak_sliders_layout.takeAt(0)
            widget = item.widget()
            if widget: widget.deleteLater()

        self.peak_sliders.clear()

        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            return

        profile_len = len(self.profile_original_inverted)
        num_items = len(self.peak_regions)
        num_to_display = min(len(self.peaks), num_items)

        if len(self.peaks) != num_items:
            print(f"Warning: Peak count ({len(self.peaks)}) / regions count ({num_items}) mismatch in update_sliders.")

        for i in range(num_to_display):
            try:
                start_val, end_val = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                peak_index = int(self.peaks[i])
            except (IndexError, ValueError, TypeError) as e:
                print(f"Warning: Invalid data for slider index {i}: {e}")
                continue

            peak_group = QGroupBox(f"Peak {i + 1} (Idx: {peak_index})")
            peak_layout = QGridLayout(peak_group)
            peak_layout.setSpacing(5)

            start_slider = QSlider(Qt.Horizontal); start_slider.setRange(0, profile_len - 1)
            start_val_clamped = max(0, min(profile_len - 1, start_val))
            start_slider.setValue(start_val_clamped)
            start_label = QLabel(f"Start: {start_val_clamped}")
            start_slider.valueChanged.connect(lambda val, lbl=start_label, idx=i: self._update_region_from_slider(idx, 'start', val, lbl))
            start_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(start_label, 0, 0); peak_layout.addWidget(start_slider, 0, 1)

            end_slider = QSlider(Qt.Horizontal); end_slider.setRange(0, profile_len - 1)
            end_val_clamped = max(start_val_clamped, min(profile_len - 1, end_val))
            end_slider.setValue(end_val_clamped)
            end_label = QLabel(f"End: {end_val_clamped}")
            end_slider.valueChanged.connect(lambda val, lbl=end_label, idx=i: self._update_region_from_slider(idx, 'end', val, lbl))
            end_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(end_label, 1, 0); peak_layout.addWidget(end_slider, 1, 1)

            self.peak_sliders_layout.addWidget(peak_group)
            self.peak_sliders.append((start_slider, end_slider))

        if num_to_display > 0:
            self.peak_sliders_layout.addStretch(1)

        if hasattr(self, 'container') and self.container:
            self.container.adjustSize()
            self.container.update()


    def _update_region_from_slider(self, index, boundary_type, value, label_widget):
        """
        Helper to update self.peak_regions when an individual start/end slider is moved.
        Ensures start <= end.
        """
        if not (0 <= index < len(self.peak_regions)): return

        current_start, current_end = self.peak_regions[index]
        start_slider_widget, end_slider_widget = (self.peak_sliders[index]
                                                  if 0 <= index < len(self.peak_sliders)
                                                  else (None, None))

        if boundary_type == 'start':
            new_start = min(value, current_end)
            self.peak_regions[index] = (new_start, current_end)
            label_widget.setText(f"Start: {new_start}")
            if start_slider_widget and start_slider_widget.value() != new_start:
                start_slider_widget.blockSignals(True); start_slider_widget.setValue(new_start); start_slider_widget.blockSignals(False)
            # Sync end slider if start pushed it
            if end_slider_widget and current_end < new_start:
                 self._update_region_from_slider(index, 'end', new_start, end_slider_widget.parent().findChild(QLabel, "EndLabelName")) # Need label ref

        elif boundary_type == 'end':
            new_end = max(value, current_start)
            self.peak_regions[index] = (current_start, new_end)
            label_widget.setText(f"End: {new_end}")
            if end_slider_widget and end_slider_widget.value() != new_end:
                end_slider_widget.blockSignals(True); end_slider_widget.setValue(new_end); end_slider_widget.blockSignals(False)
            # Sync start slider if end pulled it
            if start_slider_widget and current_start > new_end:
                 self._update_region_from_slider(index, 'start', new_end, start_slider_widget.parent().findChild(QLabel, "StartLabelName")) # Need label ref


    def _find_adjacent_trough(self, profile, start_index, direction, window=15):
        """
        Helper to find the first local minimum adjacent to start_index.
        Direction: -1 for left, +1 for right.
        Window: Max number of steps to search.
        Returns the index of the minimum or start_index if none found/better.
        """
        current_index = start_index
        profile_len = len(profile)
        # Ensure start_index is valid
        if not (0 <= start_index < profile_len): return start_index

        min_val_found = profile[start_index]
        min_idx_found = start_index
        last_val = profile[start_index]

        for i in range(1, window + 1):
            next_index = start_index + i * direction
            if not (0 <= next_index < profile_len): break # Stop at boundary

            current_val = profile[next_index]

            # Check if it's a local minimum
            is_local_min = True
            # Check left neighbor relative to current point
            prev_neighbor_idx = next_index - direction
            if (0 <= prev_neighbor_idx < profile_len) and profile[prev_neighbor_idx] < current_val:
                is_local_min = False
            # Check right neighbor relative to current point
            next_neighbor_idx = next_index + direction
            if (0 <= next_neighbor_idx < profile_len) and profile[next_neighbor_idx] < current_val:
                is_local_min = False

            # If it's a local minimum and lower than starting point, return it
            if is_local_min and current_val < profile[start_index]:
                 return next_index

            # If profile starts increasing after decreasing, return the lowest point found before increase
            if current_val > last_val and last_val < profile[start_index]:
                 return min_idx_found # Return the lowest point before the rise

            # Keep track of the lowest value encountered so far
            if current_val < min_val_found:
                min_val_found = current_val
                min_idx_found = next_index

            last_val = current_val # Update last value for rise detection

        # If loop finishes, return the lowest index found (might be start_index)
        return min_idx_found


    def update_plot(self):
        """
        Update plot using the smoothed original inverted profile.
        Valley-to-Valley baseline now connects immediate adjacent troughs.
        """
        if self.canvas is None: return
        profile_to_plot_and_calc = self.profile_original_inverted
        if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
             try:
                 self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]);
                 # self.ax.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax.transAxes)
                 self.canvas.draw_idle()
             except Exception as e: print(f"Error clearing plot: {e}")
             return

        self.method = self.method_combobox.currentText()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Calculate Rolling Ball Background ---
        if rolling_ball:
            try:
                profile_float = profile_to_plot_and_calc.astype(np.float64)
                safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
                if len(profile_float) > 1 :
                    self.background = self._custom_rolling_ball(profile_float, safe_radius)
                    self.background = np.maximum(self.background, 0)
                else: self.background = profile_float.copy()
            except ImportError: self.background = np.zeros_like(profile_to_plot_and_calc); # Handled by check
            except Exception as e: print(f"Error rolling ball: {e}."); self.background = np.zeros_like(profile_to_plot_and_calc)
        else: # rolling_ball not available
             self.background = np.zeros_like(profile_to_plot_and_calc)

        # --- Setup Plot ---
        self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)

        # --- Plot Smoothed Profile ---
        self.ax.plot(profile_to_plot_and_calc, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.2)

        # --- Plot Detected Peak Markers ---
        if len(self.peaks) > 0:
             valid_peaks = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
             if len(valid_peaks) > 0:
                 peak_y_on_smoothed = profile_to_plot_and_calc[valid_peaks]
                 self.ax.scatter(valid_peaks, peak_y_on_smoothed, color="red", marker='x', s=50, label="Detected Peaks", zorder=5)

        # --- Process Peak Regions ---
        self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
        num_items_to_plot = len(self.peak_regions)
        profile_range_plot = np.ptp(profile_to_plot_and_calc) if np.ptp(profile_to_plot_and_calc) > 0 else 1.0
        min_text_y_position = np.min(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 0
        max_text_y_position_above = min_text_y_position

        for i in range(num_items_to_plot):
            start, end = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
            if start >= end:
                 self.peak_areas_rolling_ball.append(0.0); self.peak_areas_straight_line.append(0.0); self.peak_areas_valley.append(0.0); continue

            x_region = np.arange(start, end + 1)
            profile_region_smoothed = profile_to_plot_and_calc[start : end + 1] # Slice using final region boundaries

            # Get background region
            bg_start = max(0, min(start, len(self.background)-1)); bg_end = max(0, min(end + 1, len(self.background)))
            background_region = np.zeros_like(profile_region_smoothed)
            if bg_start < bg_end and len(self.background) > 0 and interp1d:
                 raw_bg_region = self.background[bg_start:bg_end]
                 if len(raw_bg_region) == len(profile_region_smoothed): background_region = raw_bg_region
                 elif len(self.background) > 1:
                     try:
                        x_full_bg = np.arange(len(self.background))
                        interp_func_bg = interp1d(x_full_bg, self.background, kind='linear', bounds_error=False, fill_value=(self.background[0], self.background[-1]))
                        background_region = interp_func_bg(x_region)
                     except Exception as interp_err_bg: print(f"Warning: BG interp failed peak {i+1}: {interp_err_bg}")

            # --- Area Calculations ---
            area_rb = max(0, np.trapz(profile_region_smoothed - background_region, x=x_region)) if len(x_region) > 1 else 0.0
            self.peak_areas_rolling_ball.append(area_rb)

            area_sl = 0.0; y_baseline_pts_sl = np.array([0,0]); y_baseline_interp_sl = np.zeros_like(x_region)
            if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                y_baseline_pts_sl = np.array([profile_to_plot_and_calc[start], profile_to_plot_and_calc[end]])
                y_baseline_interp_sl = np.interp(x_region, [start, end], y_baseline_pts_sl)
                area_sl = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_sl, x=x_region)) if len(x_region) > 1 else 0.0
            self.peak_areas_straight_line.append(area_sl)

            # --- Valley-to-Valley using adjacent troughs ---
            area_vv = 0.0; y_baseline_pts_vv = np.array([0,0]); y_baseline_interp_vv = np.zeros_like(x_region)
            valley_start_idx_vv = start; valley_end_idx_vv = end
            try:
                 valley_start_idx_vv = self._find_adjacent_trough(profile_to_plot_and_calc, start, direction=-1, window=15)
                 valley_end_idx_vv = self._find_adjacent_trough(profile_to_plot_and_calc, end, direction=1, window=15)
                 valley_start_idx_vv = max(0, valley_start_idx_vv)
                 valley_end_idx_vv = min(len(profile_to_plot_and_calc) - 1, valley_end_idx_vv)
                 if valley_end_idx_vv <= valley_start_idx_vv: valley_start_idx_vv, valley_end_idx_vv = start, end # Fallback

                 y_baseline_pts_vv = np.array([profile_to_plot_and_calc[valley_start_idx_vv], profile_to_plot_and_calc[valley_end_idx_vv]])
                 y_baseline_interp_vv = np.interp(x_region, [valley_start_idx_vv, valley_end_idx_vv], y_baseline_pts_vv)
                 area_vv = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_vv, x=x_region)) if len(x_region) > 1 else 0.0
            except Exception as e_vv: print(f"Error VV calc peak {i+1}: {e_vv}."); area_vv = 0.0
            self.peak_areas_valley.append(area_vv)

            # --- Plot Baselines and Fills ---
            current_area = 0.0; baseline_y_at_center = 0.0
            text_x_pos = (start + end) / 2.0

            if self.method == "Rolling Ball":
                self.rolling_ball_slider.setEnabled(True)
                if i == 0: self.ax.plot(np.arange(len(self.background)), self.background, color="green", ls="--", lw=1, label="Rolling Ball BG")
                self.ax.fill_between(x_region, background_region, profile_region_smoothed, where=profile_region_smoothed >= background_region, color="yellow", alpha=0.4, interpolate=True)
                current_area = area_rb
                if len(self.background) > 1 and interp1d:
                    try:
                        interp_func_bg = interp1d(np.arange(len(self.background)), self.background, kind='linear', fill_value="extrapolate")
                        baseline_y_at_center = interp_func_bg(text_x_pos)
                    except: baseline_y_at_center = np.min(background_region) if len(background_region)>0 else 0
                else: baseline_y_at_center = np.min(background_region) if len(background_region)>0 else 0

            elif self.method == "Straight Line":
                self.rolling_ball_slider.setEnabled(False)
                if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                    self.ax.plot([start, end], y_baseline_pts_sl, color="purple", ls="--", lw=1, label="SL BG" if i == 0 else "")
                    self.ax.fill_between(x_region, y_baseline_interp_sl, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_sl, color="cyan", alpha=0.4, interpolate=True)
                    current_area = area_sl
                    baseline_y_at_center = np.interp(text_x_pos, [start, end], y_baseline_pts_sl)
                else: current_area = 0.0; baseline_y_at_center = 0.0

            elif self.method == "Valley-to-Valley":
                self.rolling_ball_slider.setEnabled(False)
                if valley_start_idx_vv < len(profile_to_plot_and_calc) and valley_end_idx_vv < len(profile_to_plot_and_calc):
                    self.ax.plot([valley_start_idx_vv, valley_end_idx_vv], y_baseline_pts_vv, color="orange", ls="--", lw=1, label="Adj. Valley BG" if i == 0 else "")
                    self.ax.fill_between(x_region, y_baseline_interp_vv, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_vv, color="lightblue", alpha=0.4, interpolate=True)
                    current_area = area_vv
                    baseline_y_at_center = np.interp(text_x_pos, [valley_start_idx_vv, valley_end_idx_vv], y_baseline_pts_vv)
                else: current_area = 0.0; baseline_y_at_center = 0.0

            # --- Plot Area Text BELOW the Baseline ---
            area_text_format = "{:.0f}"
            combined_text = f"Peak {i + 1}\n{area_text_format.format(current_area)}"
            text_y_offset = profile_range_plot * 0.01
            text_y_pos = baseline_y_at_center - text_y_offset
            self.ax.text(text_x_pos, text_y_pos, combined_text, ha="center", va="top", fontsize=7, color='black', zorder=6)
            min_text_y_position = min(min_text_y_position, text_y_pos)
            if len(profile_region_smoothed) > 0: max_text_y_position_above = max(max_text_y_position_above, np.max(profile_region_smoothed) + profile_range_plot*0.03)

            self.ax.axvline(start, color="gray", ls=":", lw=1.0, alpha=0.8)
            self.ax.axvline(end, color="gray", ls=":", lw=1.0, alpha=0.8)

        # --- Final Plot Configuration ---
        self.ax.set_ylabel("Intensity (Smoothed, Inverted)")
        self.ax.legend(fontsize='small', loc='upper right')
        self.ax.set_title(f"Smoothed Intensity Profile (σ={self.smoothing_sigma:.1f}) and Peak Regions")
        self.ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

        if len(profile_to_plot_and_calc) > 1: self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1)
        prof_min_smooth, prof_max_smooth = (np.min(profile_to_plot_and_calc), np.max(profile_to_plot_and_calc)) if len(profile_to_plot_and_calc) > 0 else (0, 1)
        y_max_limit = max(prof_max_smooth, max_text_y_position_above) + profile_range_plot * 0.05
        y_min_limit = min(prof_min_smooth, min_text_y_position) - profile_range_plot * 0.05
        if y_max_limit <= y_min_limit: y_max_limit = y_min_limit + 1.0
        self.ax.set_ylim(y_min_limit, y_max_limit)
        if prof_max_smooth > 10000: self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)

        # --- Display Cropped Image ---
        ax_image.clear()
        if hasattr(self, 'cropped_image_for_display') and isinstance(self.cropped_image_for_display, Image.Image):
             try:
                 rotated_pil_image = self.cropped_image_for_display.rotate(90, expand=True)
                 im_array_disp = np.array(rotated_pil_image)
                 if self.original_max_value == 1.0 and np.issubdtype(self.intensity_array_original_range.dtype, np.floating):
                      im_vmin, im_vmax = 0.0, 1.0
                 else: im_vmin, im_vmax = 0, self.original_max_value
                 ax_image.imshow(im_array_disp, cmap='gray', aspect='auto',
                                 extent=[0, len(profile_to_plot_and_calc)-1 if len(profile_to_plot_and_calc)>0 else 0, 0, rotated_pil_image.height],
                                 vmin=im_vmin, vmax=im_vmax)
                 ax_image.set_xlabel("Pixel Index Along Profile Axis")
                 ax_image.set_yticks([]); ax_image.set_ylabel("Lane Width", fontsize='small')
             except Exception as img_e:
                 print(f"Error displaying cropped image preview: {img_e}")
                 ax_image.text(0.5, 0.5, 'Error loading preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])
        else:
             ax_image.text(0.5, 0.5, 'No Image Preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])

        # --- Adjust and Draw ---
        # try: self.fig.tight_layout(pad=0.5) # Use tight_layout for better spacing
        # except Exception as layout_e: print(f"Error adjusting layout: {layout_e}")
        try: self.canvas.draw_idle()
        except Exception as draw_e: print(f"Error drawing canvas: {draw_e}")
        
        plt.close(self.fig)

    def get_final_peak_area(self):
        """Return the list of calculated peak areas based on the selected method."""
        num_valid_peaks = len(self.peak_regions)
        current_area_list = []
        if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
        elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
        elif self.method == "Valley-to-Valley": current_area_list = self.peak_areas_valley
        else: return []

        # Ensure list length matches number of regions processed
        if len(current_area_list) != num_valid_peaks:
            print(f"Warning: Area list length ({len(current_area_list)}) mismatch with regions ({num_valid_peaks}) for method '{self.method}'. Truncating.")
            # Return a list matching the length, padding with 0 if needed, or truncating
            # Truncating is safer if some peaks failed calculation
            return current_area_list[:num_valid_peaks]
        else:
            return current_area_list
    
        
    


class LiveViewLabel(QLabel):
    def __init__(self, font_type, font_size, marker_color, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)  # Enable mouse tracking
        self.preview_marker_enabled = False
        self.preview_marker_text = ""
        self.preview_marker_position = None
        self.marker_font_type = font_type
        self.marker_font_size = font_size
        self.marker_color = marker_color
        self.setFocusPolicy(Qt.StrongFocus)
        self.bounding_box_preview = []
        self.measure_quantity_mode = False
        self.counter = 0
        self.zoom_level = 1.0  # Initial zoom level
        self.pan_start = None  # Start position for panning
        self.pan_offset = QPointF(0, 0)  # Offset for panning
        self.quad_points = []  # Stores 4 points of the quadrilateral
        self.selected_point = -1  # Index of selected corner (-1 = none)
        self.drag_threshold = 10  # Pixel radius for selecting corners
        self.bounding_box_complete = False
        self.mode=None
        # Add rectangle-related attributes
        self.rectangle_start = None  # Start point of the rectangle
        self.rectangle_end = None    # End point of the rectangle
        self.rectangle_points = []   # Stores the rectangle points
        self.drag_start_pos = None  # For tracking drag operations
        self.draw_edges=True
        self.drawing_crop_rect = False # Is the user currently dragging to draw?
        self.crop_rect_start_view = None # Start point in view coordinates
        self.crop_rect_end_view = None   # Current/End point in view coordinates
        self.crop_rect_final_view = None # The finalized rectangle (QRectF) in view coords, for persistent display

    def clear_crop_preview(self):
        """Clears any existing crop rectangle preview."""
        self.drawing_crop_rect = False
        self.crop_rect_start_view = None
        self.crop_rect_end_view = None
        self.crop_rect_final_view = None
        self.update() # Trigger repaint

    def start_crop_preview(self, start_point_view):
        """Initiates drawing the crop rectangle preview."""
        self.drawing_crop_rect = True
        self.crop_rect_start_view = start_point_view
        self.crop_rect_end_view = start_point_view # Start and end are same initially
        self.crop_rect_final_view = None # Clear any finalized rect
        self.update()

    def update_crop_preview(self, start_point_view, current_point_view):
        """Updates the crop rectangle preview while dragging."""
        if self.drawing_crop_rect:
            # Keep start point fixed, update end point
            self.crop_rect_start_view = start_point_view
            self.crop_rect_end_view = current_point_view
            self.update()

    def finalize_crop_preview(self, start_point_view, end_point_view):
        """Stores the final rectangle dimensions for persistent preview."""
        self.drawing_crop_rect = False
        # Store as a normalized QRectF for easier drawing
        self.crop_rect_final_view = QRectF(start_point_view, end_point_view).normalized()
        # Clear the temporary points used during drawing
        self.crop_rect_start_view = None
        self.crop_rect_end_view = None
        self.update()
        
    def mouseMoveEvent(self, event):
        if self.preview_marker_enabled:
            self.preview_marker_position = event.pos()
            self.update()  # Trigger repaint to show the preview
        if self.selected_point != -1 and self.measure_quantity_mode:# and self.mode=="quad":
            # Update dragged corner position
            self.quad_points[self.selected_point] = self.transform_point(event.pos())
            self.update()  # Show the bounding box preview
        if self.selected_point != -1 and self.mode=="move":
            self.drag_start_pos=event.pos()
        
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if self.preview_marker_enabled:
            # Place the marker text permanently at the clicked position
            parent = self.parent()
            parent.place_custom_marker(event, self.preview_marker_text)
            self.update()  # Clear the preview
        if self.measure_quantity_mode and self.mode=="quad":
            # Check if clicking near existing corner
            for i, p in enumerate(self.quad_points):
                if (self.transform_point(event.pos()) - p).manhattanLength() < self.drag_threshold:
                    self.selected_point = i
                    return
    
            # Add new point if < 4 corners
            if len(self.quad_points) < 4:
                self.quad_points.append(self.transform_point(event.pos()))
                self.selected_point = len(self.quad_points) - 1
    
            if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                # self.quad_points = [self.transform_point(p) for p in self.quad_points]
                self.bounding_box_complete = True
            self.update()  # Trigger repaint
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode=="quad":
            self.selected_point = -1
            self.update()
        super().mouseReleaseEvent(event)

    def transform_point(self, point):
        """Transform a point from widget coordinates to image coordinates."""
        if self.zoom_level != 1.0:
            return QPointF(
                (point.x() - self.pan_offset.x()) / self.zoom_level,
                (point.y() - self.pan_offset.y()) / self.zoom_level
            )
        return QPointF(point)
    
    
    def zoom_in(self):
        self.zoom_level *= 1.1
        self.update()

    def zoom_out(self):
        self.zoom_level /= 1.1
        if self.zoom_level < 1.0:
            self.zoom_level = 1.0
            self.pan_offset = QPointF(0, 0)  # Reset pan offset when zoom is reset
        self.update()
        

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)


        painter.save()
        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)
            
        painter.restore()
        
        painter.save()

        if self.zoom_level != 1.0:
             painter.translate(self.pan_offset)
             painter.scale(self.zoom_level, self.zoom_level)

        preview_pen = QPen(Qt.red) # Use a distinct color
        preview_pen.setWidth(1) # Thin line

        if self.drawing_crop_rect and self.crop_rect_start_view and self.crop_rect_end_view:
            # Draw dashed line while actively dragging
            preview_pen.setStyle(Qt.DashLine)
            painter.setPen(preview_pen)
            rect_to_draw = QRectF(self.crop_rect_start_view, self.crop_rect_end_view).normalized()
            painter.drawRect(rect_to_draw)
        elif self.crop_rect_final_view:
            # Draw solid line for the finalized rectangle
            preview_pen.setStyle(Qt.SolidLine)
            painter.setPen(preview_pen)
            painter.drawRect(self.crop_rect_final_view) # Already a QRectF

        painter.restore()
        
        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)

        # Draw the preview marker if enabled
        if self.preview_marker_enabled and self.preview_marker_position:
            painter.setOpacity(0.5)  # Semi-transparent preview
            font = QFont(self.marker_font_type)
            font.setPointSize(self.marker_font_size)
            painter.setFont(font)
            painter.setPen(self.marker_color)
            text_width = painter.fontMetrics().horizontalAdvance(self.preview_marker_text)
            text_height = painter.fontMetrics().height()
            # Draw the text at the cursor's position
            x, y = self.preview_marker_position.x(), self.preview_marker_position.y()
            if self.zoom_level != 1.0:
                x = (x - self.pan_offset.x()) / self.zoom_level
                y = (y - self.pan_offset.y()) / self.zoom_level
            painter.drawText(int(x - text_width / 2), int(y + text_height / 4), self.preview_marker_text)
            
        if len(self.quad_points) > 0 and len(self.quad_points) <4 :
            for p in self.quad_points:
                painter.setPen(QPen(Qt.red, 1))
                painter.drawEllipse(p, 1, 1)
    
        if len(self.quad_points) == 4 and self.draw_edges==True:
            painter.setPen(QPen(Qt.blue, 1))
            painter.drawPolygon(QPolygonF(self.quad_points))
            # Draw draggable corners
            for p in self.quad_points:
                painter.drawEllipse(p, self.drag_threshold, self.drag_threshold)
    
        # Draw the bounding box preview if it exists
        if self.bounding_box_preview:
            painter.setPen(QPen(Qt.blue, 1))  
            start_x, start_y, end_x, end_y = self.bounding_box_preview
            rect = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y))
            painter.drawRect(rect)
        
        
    
        painter.end()
        self.update()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:  
            parent_app = self.parent()
            if isinstance(parent_app, CombinedSDSApp) and parent_app.crop_rectangle_mode:
                parent_app.cancel_rectangle_crop_mode() # Call the app's cancel method
                return # Prevent further processing in this case
            if self.preview_marker_enabled:
                self.preview_marker_enabled = False  # Turn off the preview
                self.update()  # Clear the overlay
            self.measure_quantity_mode = False
            self.counter = 0
            self.bounding_box_complete = False
            self.quad_points = []
            self.mode=None
            self.clear_crop_preview()
            self.update()
        super().keyPressEvent(event)

class CombinedSDSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
        # window_width = int(self.screen_width * 0.5)  # 60% of screen width
        # window_height = int(self.screen_height * 0.75)  # 95% of screen height
        self.preview_label_width_setting = int(self.screen_width * 0.45)
        self.preview_label_max_height_setting = int(self.screen_height * 0.35)
        self.label_size = self.preview_label_width_setting
        self.window_title="IMAGING ASSISTANT V7.0"
        # --- Initialize Status Bar Labels ---
        self.size_label = QLabel("Image Size: N/A")
        self.depth_label = QLabel("Bit Depth: N/A")
        self.location_label = QLabel("Source: N/A")

        # --- Add Labels to Status Bar ---
        statusbar = self.statusBar() # Get the status bar instance
        statusbar.addWidget(self.size_label)
        statusbar.addWidget(self.depth_label)
        # Add location label with stretch factor 1 to push it to the right
        statusbar.addWidget(self.location_label, 1)
        self.setWindowTitle(self.window_title)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.undo_stack = []
        self.redo_stack = []
        self.custom_shapes = []  # NEW: List to store lines/rectangles
        self.drawing_mode = None # NEW: None, 'line', 'rectangle'
        self.current_drawing_shape_preview = None # NEW: For live preview data
        self.quantities_peak_area_dict = {}
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []
        self.image_path = None
        self.image = None
        self.image_master= None
        self.image_before_padding = None
        self.image_contrasted=None
        self.image_before_contrast=None
        self.contrast_applied=False
        self.image_padded=False
        self.predict_size=False
        self.warped_image=None
        self.transparency=1
        self.left_markers = []
        self.right_markers = []
        self.top_markers = []
        self.custom_markers=[]
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.font_rotation=-45
        self.image_width=0
        self.image_height=0
        self.new_image_width=0
        self.new_image_height=0
        self.base_name="Image"
        self.image_path=""
        self.x_offset_s=0
        self.y_offset_s=0
        self.peak_area=[]
        self.is_modified=False
        self.crop_rectangle_mode = False
        self.crop_rect_start_view = None # Temp storage for starting point in view coords
        self.crop_rectangle_coords = None # Stores final (x, y, w, h) in *image* coordinates
        self.crop_offset_x = 0
        self.crop_offset_y = 0
        
        self.peak_dialog_settings = {
            'rolling_ball_radius': 50,
            'peak_height_factor': 0.1,
            'peak_distance': 30,
            'peak_prominence_factor': 0.02,
            'peak_spread_pixels': 10,
            'band_estimation_method': "Mean",
            'area_subtraction_method': "Valley-to-Valley"
        }
        self.persist_peak_settings_enabled = True # State of the checkbox
        
        # Variables to store bounding boxes and quantities
        self.bounding_boxes = []
        self.up_bounding_boxes = []
        self.standard_protein_areas=[]
        self.quantities = []
        self.protein_quantities = []
        self.measure_quantity_mode = False
        # Initialize self.marker_values to None initially
        self.marker_values_dict = {
            "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
        }
        self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"]
        self.top_label_dict={"Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"]}
        
        
        self.custom_marker_name = "Custom"
        # Load the config file if exists
        
        self.marker_mode = None
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_slider_range=[-1000,1000]
        self.right_slider_range=[-1000,1000]
        self.top_slider_range=[-1000,1000]
               
        
        self.top_padding = 0
        self.font_color = QColor(0, 0, 0)  # Default to black
        self.custom_marker_color = QColor(0, 0, 0)  # Default to black
        self.font_family = "Arial"  # Default font family
        self.font_size = 12  # Default font size
        self.image_array_backup= None
        self.run_predict_MW=False
        
        
        
        # Main container widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Upper section (Preview and buttons)
        upper_layout = QHBoxLayout()

        self.label_width=int(self.label_size)
    
        self.live_view_label = LiveViewLabel(
            font_type=QFont("Arial"),
            font_size=int(24),
            marker_color=QColor(0,0,0),
            parent=self,
        )
        # Image display
        self.live_view_label.setStyleSheet("background-color: white; border: 1px solid black;")
        
        self._create_actions()
        self.create_menu_bar()
        self.create_tool_bar()
        
        self._update_preview_label_size()
        upper_layout.addWidget(self.live_view_label, stretch=1)
        layout.addLayout(upper_layout)
        
        # Lower section (Tabbed interface)
        self.tab_widget = QTabWidget()
        # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
        self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
        self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
        self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
        self.tab_widget.addTab(self.create_markers_tab(), "Markers")
        self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
        self.tab_widget.addTab(self.analysis_tab(), "Analysis")
        
        layout.addWidget(self.tab_widget)
        
        #self.load_shortcut = QShortcut(QKeySequence.Open, self) # Ctrl+O / Cmd+O
        #self.load_shortcut.activated.connect(self.load_action.trigger) # Trigger the action

        #self.save_shortcut = QShortcut(QKeySequence.Save, self) # Ctrl+S / Cmd+S
        #self.save_shortcut.activated.connect(self.save_action.trigger) # Trigger the action

        #self.copy_shortcut = QShortcut(QKeySequence.Copy, self) # Ctrl+C / Cmd+C
        #self.copy_shortcut.activated.connect(self.copy_action.trigger) # Trigger the action

        #self.paste_shortcut = QShortcut(QKeySequence.Paste, self) # Ctrl+V / Cmd+V
        #self.paste_shortcut.activated.connect(self.paste_action.trigger) # Trigger the action

        #self.undo_shortcut = QShortcut(QKeySequence.Undo, self) # Ctrl+Z / Cmd+Z
        #self.undo_shortcut.activated.connect(self.undo_action_m) # Connect directly to method

        #self.redo_shortcut = QShortcut(QKeySequence.Redo, self) # Ctrl+Y / Cmd+Shift+Z
        #self.redo_shortcut.activated.connect(self.redo_action_m) # Connect directly to method

        #self.zoom_out_shortcut = QShortcut(QKeySequence.ZoomOut, self) # Ctrl+- / Cmd+-
        #self.zoom_out_shortcut.activated.connect(self.zoom_out_action.trigger) # Trigger the action

        # Zoom In - Keep the standard one
        #self.zoom_in_shortcut = QShortcut(QKeySequence.ZoomIn, self) # Attempts Ctrl++ / Cmd++
        #self.zoom_in_shortcut.activated.connect(self.zoom_in_action.trigger) # Trigger the action
        # ======================================================


        # === KEEP QShortcut definitions for NON-STANDARD actions ===
        self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
        self.save_svg_shortcut.activated.connect(self.save_svg_action.trigger)

        self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.reset_shortcut.activated.connect(self.reset_action.trigger)

        self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        self.predict_shortcut.activated.connect(self.predict_molecular_weight)

        self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)

        self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
        self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)

        self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
        self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)

        self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
        self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)

        self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
        self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
        self.grid_shortcut.activated.connect(
            lambda: (
                # Determine the target state (True if both are currently False, False otherwise)
                target_state := not (self.show_grid_checkbox_x.isChecked() or self.show_grid_checkbox_y.isChecked()),
                # Set both checkboxes to the target state
                self.show_grid_checkbox_x.setChecked(target_state),
                self.show_grid_checkbox_y.setChecked(target_state)
            ) if hasattr(self, 'show_grid_checkbox_x') and hasattr(self, 'show_grid_checkbox_y') else None
        )

        self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.guidelines_shortcut.activated.connect(
            lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked())
            if hasattr(self, 'show_guides_checkbox') else None
        )

        self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
        self.increase_grid_size_shortcut.activated.connect(
            lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1)
            if hasattr(self, 'grid_size_input') else None
        )
        self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
        self.decrease_grid_size_shortcut.activated.connect(
            lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1)
            if hasattr(self, 'grid_size_input') else None
        )

        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self.invert_shortcut.activated.connect(self.invert_image)

        self.bw_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
        self.bw_shortcut.activated.connect(self.convert_to_black_and_white)

        # Tab movement shortcuts
        self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
        self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
        self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
        self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
        self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
        self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
        self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
        self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
        self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
        self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
        self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
        self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
        
        self.load_config()
    
    def _create_actions(self):
        """Create QAction objects for menus and toolbars."""
        style = self.style() # Still useful for other icons
        icon_size = QSize(30, 30) # Match your toolbar size
        text_color = self.palette().color(QPalette.ButtonText) # Use theme color

        # # --- Create Zoom Icons (using previous method) ---
        # zoom_in_pixmap = QPixmap(icon_size)
        # zoom_in_pixmap.fill(Qt.transparent)
        # painter_in = QPainter(zoom_in_pixmap)
        # # ... (painter setup as before) ...
        # painter_in.drawText(zoom_in_pixmap.rect(), Qt.AlignCenter, "+")
        # painter_in.end()
        # zoom_in_icon = QIcon(zoom_in_pixmap)

        # zoom_out_pixmap = QPixmap(icon_size)
        # zoom_out_pixmap.fill(Qt.transparent)
        # painter_out = QPainter(zoom_out_pixmap)
        # # ... (painter setup as before) ...
        # painter_out.drawText(zoom_out_pixmap.rect(), Qt.AlignCenter, "-")
        # painter_out.end()
        # zoom_out_icon = QIcon(zoom_out_pixmap)
        # --- End Zoom Icons ---


        # --- Create Icons using the helper ---
        open_icon = create_text_icon("Wingdings",icon_size, text_color, "1") # Unicode Right Arrow
        save_icon = create_text_icon("Wingdings",icon_size, text_color, "=")
        save_svg_icon = create_text_icon("Wingdings",icon_size, text_color, "3")
        undo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "O")
        redo_icon = create_text_icon("Wingdings 3",icon_size, text_color, "N")
        paste_icon = create_text_icon("Wingdings",icon_size, text_color, "2")
        copy_icon = create_text_icon("Wingdings",icon_size, text_color, "4")
        reset_icon = create_text_icon("Wingdings 3",icon_size, text_color, "Q")
        exit_icon = create_text_icon("Wingdings 2",icon_size, text_color, "V")
        
        zoom_in_icon = create_text_icon("Arial",icon_size, text_color, "+")
        zoom_out_icon = create_text_icon("Arial",icon_size, text_color, "-")
        pan_up_icon = create_text_icon("Arial",icon_size, text_color, "↑") # Unicode Up Arrow
        pan_down_icon = create_text_icon("Arial",icon_size, text_color, "↓") # Unicode Down Arrow
        pan_left_icon = create_text_icon("Arial",icon_size, text_color, "←") # Unicode Left Arrow
        pan_right_icon = create_text_icon("Arial",icon_size, text_color, "→") # Unicode Right Arrow
        # open_icon = create_text_icon(icon_size, text_color, "") # Unicode Right Arrow
        # paste_icon = create_text_icon(icon_size, text_color, "→") # Unicode Right Arrow
        # copy_icon = create_text_icon(icon_size, text_color, "→") # Unicode Right Arrow
        # --- End Pan Icons ---


        # --- File Actions ---
        self.load_action = QAction(open_icon, "&Load Image...", self)
        self.save_action = QAction(save_icon, "&Save with Config", self)
        self.save_svg_action = QAction(save_svg_icon, "Save &SVG...", self)
        self.reset_action = QAction(reset_icon, "&Reset Image", self)
        self.exit_action = QAction(exit_icon, "E&xit", self)

        # --- Edit Actions ---
        self.undo_action = QAction(undo_icon, "&Undo", self) # Standard for now
        self.redo_action = QAction(redo_icon, "&Redo", self) # Standard for now
        self.copy_action = QAction(copy_icon, "&Copy Image", self)
        self.paste_action = QAction(paste_icon, "&Paste Image", self)

        # --- View Actions (using the created icons) ---
        self.zoom_in_action = QAction(zoom_in_icon, "Zoom &In", self)
        self.zoom_out_action = QAction(zoom_out_icon, "Zoom &Out", self)
        self.pan_left_action = QAction(pan_left_icon, "Pan Left", self)
        self.pan_right_action = QAction(pan_right_icon, "Pan Right", self)
        self.pan_up_action = QAction(pan_up_icon, "Pan Up", self)
        self.pan_down_action = QAction(pan_down_icon, "Pan Down", self)
        # --- END: Add Panning Actions ---

        # --- Set Shortcuts ---
        self.load_action.setShortcut(QKeySequence.Open)
        self.save_action.setShortcut(QKeySequence.Save)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.paste_action.setShortcut(QKeySequence.Paste)
        self.undo_action.setShortcut(QKeySequence.Undo)
        self.redo_action.setShortcut(QKeySequence.Redo)
        self.zoom_in_action.setShortcut(QKeySequence("Ctrl+="))
        self.zoom_out_action.setShortcut(QKeySequence.ZoomOut)
        self.save_svg_action.setShortcut(QKeySequence("Ctrl+M"))
        self.reset_action.setShortcut(QKeySequence("Ctrl+R"))

        # --- Set Tooltips ---
        self.load_action.setToolTip("Load an image file (Ctrl+O)")
        self.save_action.setToolTip("Save image and configuration (Ctrl+S)")
        self.save_svg_action.setToolTip("Save as SVG for Word/vector editing (Ctrl+M)")
        self.reset_action.setToolTip("Reset image and all annotations (Ctrl+R)")
        self.exit_action.setToolTip("Exit the application")
        self.undo_action.setToolTip("Undo last action (Ctrl+Z)")
        self.redo_action.setToolTip("Redo last undone action (Ctrl+Y)")
        self.copy_action.setToolTip("Copy rendered image to clipboard (Ctrl+C)")
        self.paste_action.setToolTip("Paste image from clipboard (Ctrl+V)")
        self.zoom_in_action.setToolTip("Increase zoom level (Ctrl+=)")
        self.zoom_out_action.setToolTip("Decrease zoom level (Ctrl+-)")
        # --- START: Set Tooltips for Panning Actions ---
        self.pan_left_action.setToolTip("Pan the view left (when zoomed) (Arrow key left)")
        self.pan_right_action.setToolTip("Pan the view right (when zoomed) (Arrow key right")
        self.pan_up_action.setToolTip("Pan the view up (when zoomed) (Arrow key up)")
        self.pan_down_action.setToolTip("Pan the view down (when zoomed) (Arrow key down")
        # --- END: Set Tooltips for Panning Actions ---

        # --- Connect signals ---
        self.load_action.triggered.connect(self.load_image)
        self.save_action.triggered.connect(self.save_image)
        self.save_svg_action.triggered.connect(self.save_image_svg)
        self.reset_action.triggered.connect(self.reset_image)
        self.exit_action.triggered.connect(self.close)
        self.undo_action.triggered.connect(self.undo_action_m)
        self.redo_action.triggered.connect(self.redo_action_m)
        self.copy_action.triggered.connect(self.copy_to_clipboard)
        self.paste_action.triggered.connect(self.paste_image)
        self.zoom_in_action.triggered.connect(self.zoom_in)
        self.zoom_out_action.triggered.connect(self.zoom_out)
        # --- START: Connect Panning Action Signals ---
        # Use lambda functions to directly modify the offset
        pan_step = 30 # Define pan step size here or access from elsewhere if needed
        self.pan_right_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_right_action.triggered.connect(self.update_live_view) # Trigger update after offset change

        self.pan_left_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_left_action.triggered.connect(self.update_live_view)

        self.pan_down_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_down_action.triggered.connect(self.update_live_view)

        self.pan_up_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_up_action.triggered.connect(self.update_live_view)
        # --- END: Connect Panning Action Signals ---

        # --- START: Initially Disable Panning Actions ---
        self.pan_left_action.setEnabled(False)
        self.pan_right_action.setEnabled(False)
        self.pan_up_action.setEnabled(False)
        self.pan_down_action.setEnabled(False)
        
    def _update_preview_label_size(self):
        """
        Updates the fixed size of the live_view_label, respecting max width/height
        and maintaining aspect ratio. Prioritizes fixing height, but adjusts if
        width constraint is violated.
        """
        # --- Define Defaults and Minimums ---
        default_max_width = 600  # Fallback if width setting is missing/invalid
        default_max_height = 500 # Fallback if height setting is missing/invalid
        min_dim = 50             # Minimum allowed dimension for the label

        # --- Determine Maximum Constraints (with validation) ---
        max_w = default_max_width
        if hasattr(self, 'preview_label_width_setting'):
            try:
                setting_width = int(self.preview_label_width_setting)
                if setting_width > 0:
                    max_w = setting_width
                else:
                    print("Warning: preview_label_width_setting is not positive, using default.")
            except (TypeError, ValueError):
                print("Warning: preview_label_width_setting is invalid, using default.")
        else:
            print("Warning: preview_label_width_setting attribute not found, using default.")
        max_w = max(min_dim, max_w) # Apply minimum constraint

        max_h = default_max_height
        if hasattr(self, 'preview_label_max_height_setting'):
            try:
                setting_height = int(self.preview_label_max_height_setting)
                if setting_height > 0:
                    max_h = setting_height
                else:
                    print("Warning: preview_label_max_height_setting is not positive, using default.")
            except (TypeError, ValueError):
                print("Warning: preview_label_max_height_setting is invalid, using default.")
        max_h = max(min_dim, max_h) # Apply minimum constraint


        # --- Initialize Final Dimensions ---
        final_w = max_w
        final_h = max_h

        # --- Calculate Dimensions Based on Image ---
        if self.image and not self.image.isNull():
            w = self.image.width()
            h = self.image.height()

            if w > 0 and h > 0:
                img_ratio = w / h

                # Attempt 1: Calculate width based on fixed max height
                calc_w_based_on_h = max_h * img_ratio

                if calc_w_based_on_h <= max_w:
                    # Width is within limits, height is fixed
                    final_w = calc_w_based_on_h
                    final_h = max_h
                else:
                    # Calculated width exceeds max width, so fix width and calculate height
                    final_w = max_w
                    final_h = max_w / img_ratio

                # Ensure calculated dimensions are not below minimum
                final_w = max(min_dim, int(final_w))
                final_h = max(min_dim, int(final_h))

            else:
                # Handle invalid image dimensions (0 width or height)
                final_w = max_w # Already set above
                final_h = max_h # Already set above

        else:
            # No image loaded - Use max constraints as default size
            final_w = max_w # Already set above
            final_h = max_h # Already set above
            # self.live_view_label.clear() # Optionally clear the label

        # --- Set the Calculated Fixed Size ---
        self.live_view_label.setFixedSize(final_w, final_h)
        
    def _update_status_bar(self):
        """Updates the status bar labels with current image information."""
        if self.image and not self.image.isNull():
            w = self.image.width()
            h = self.image.height()
            self.size_label.setText(f"Image Size: {w}x{h}")

            # Determine bit depth/format string
            img_format = self.image.format()
            depth_str = "Unknown"
            if img_format == QImage.Format_Grayscale8:
                depth_str = "8-bit Grayscale"
            elif img_format == QImage.Format_Grayscale16:
                depth_str = "16-bit Grayscale"
            elif img_format == QImage.Format_RGB888:
                 depth_str = "24-bit RGB"
            elif img_format in (QImage.Format_RGB32, QImage.Format_ARGB32,
                                QImage.Format_ARGB32_Premultiplied, QImage.Format_RGBA8888):
                 depth_str = "32-bit (A)RGB"
            elif img_format == QImage.Format_Indexed8:
                 depth_str = "8-bit Indexed"
            elif img_format == QImage.Format_Mono:
                 depth_str = "1-bit Mono"
            else:
                # Fallback to depth() if format is unusual
                depth_val = self.image.depth()
                depth_str = f"{depth_val}-bit Depth"
            self.depth_label.setText(f"Bit Depth: {depth_str}")

            # Update location
            location = "N/A"
            if hasattr(self, 'image_path') and self.image_path:
                # Show only filename if it's a path, otherwise show the source info
                if os.path.exists(self.image_path):
                    location = os.path.basename(self.image_path)
                else: # Likely "Clipboard..." or similar
                    location = self.image_path
            else:
                location = "N/A"
            self.location_label.setText(f"Source: {location}")

        else:
            # No image loaded
            self.size_label.setText("Image Size: N/A")
            self.depth_label.setText("Bit Depth: N/A")
            self.location_label.setText("Source: N/A")
        
    def prompt_save_if_needed(self):
        if not self.is_modified:
            return True # No changes, proceed
        """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
        reply = QMessageBox.question(self, 'Unsaved Changes',
                                     "Do you want to save the file?",
                                     QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                     QMessageBox.Save) # Default button is Save

        if reply == QMessageBox.Save:
            return self.save_image() # Returns True if saved, False if save cancelled
        elif reply == QMessageBox.Discard:
            return True # Proceed without saving
        else: # Cancelled
            return False # Abort the current action (close/load)

    def closeEvent(self, event):
        """Overrides the default close event to prompt for saving."""
        if self.prompt_save_if_needed():

            event.accept() # Proceed with closing
        else:
            event.ignore() # Abort closing
            
    def enable_line_drawing_mode(self):
        self.save_state()
        """Sets up the application to draw a line."""
        self.drawing_mode = 'line'
        self.live_view_label.mode = 'draw_shape' # Generic drawing mode for label
        self.current_drawing_shape_preview = None # Clear previous preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.mousePressEvent = self.start_shape_draw
        self.live_view_label.mouseMoveEvent = self.update_shape_draw
        self.live_view_label.mouseReleaseEvent = self.finalize_shape_draw
        print("Line drawing mode enabled.") # Debug

    def enable_rectangle_drawing_mode(self):
        self.save_state()
        """Sets up the application to draw a rectangle."""
        self.drawing_mode = 'rectangle'
        self.live_view_label.mode = 'draw_shape' # Generic drawing mode for label
        self.current_drawing_shape_preview = None # Clear previous preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.mousePressEvent = self.start_shape_draw
        self.live_view_label.mouseMoveEvent = self.update_shape_draw
        self.live_view_label.mouseReleaseEvent = self.finalize_shape_draw
        print("Rectangle drawing mode enabled.") # Debug

    def cancel_drawing_mode(self):
        """Resets drawing mode and cursor."""
        self.drawing_mode = None
        self.live_view_label.mode = None
        self.current_drawing_shape_preview = None
        self.live_view_label.setCursor(Qt.ArrowCursor)
        self.live_view_label.mousePressEvent = None # Reset handler
        self.live_view_label.mouseMoveEvent = None # Reset handler
        self.live_view_label.mouseReleaseEvent = None # Reset handler
        self.update_live_view() # Refresh to clear any preview
        
    def start_shape_draw(self, event):
        """Starts drawing a line or rectangle."""
        if self.drawing_mode in ['line', 'rectangle']:
            start_point = self.live_view_label.transform_point(event.pos())
            # Apply snapping if enabled
            grid_size = self.grid_size_input.value()
            if self.show_grid_checkbox_x.isChecked() and grid_size > 0:
                start_point.setX(round(start_point.x() / grid_size) * grid_size)
            if self.show_grid_checkbox_y.isChecked() and grid_size > 0:
                start_point.setY(round(start_point.y() / grid_size) * grid_size)

            self.current_drawing_shape_preview = {'start': start_point, 'end': start_point}
            self.update_live_view() # Trigger repaint for preview

    def update_shape_draw(self, event):
        """Updates the preview of the line or rectangle while dragging."""
        if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
            end_point = self.live_view_label.transform_point(event.pos())
             # Apply snapping if enabled
            grid_size = self.grid_size_input.value()
            if self.show_grid_checkbox_x.isChecked() and grid_size > 0:
                end_point.setX(round(end_point.x() / grid_size) * grid_size)
            if self.show_grid_checkbox_y.isChecked() and grid_size > 0:
                end_point.setY(round(end_point.y() / grid_size) * grid_size)

            self.current_drawing_shape_preview['end'] = end_point
            self.update_live_view() # Trigger repaint for preview

    def finalize_shape_draw(self, event):
        """Finalizes the shape and adds it to the custom_shapes list."""
        if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
            start_point = self.current_drawing_shape_preview['start'] # Label coords (unzoomed)
            end_point = self.live_view_label.transform_point(event.pos()) # Get final point (unzoomed)

            # Apply snapping to the END point if enabled
            grid_size = self.grid_size_input.value()
            if self.show_grid_checkbox_x.isChecked() and grid_size > 0:
                end_point.setX(round(end_point.x() / grid_size) * grid_size)
            if self.show_grid_checkbox_y.isChecked() and grid_size > 0:
                end_point.setY(round(end_point.y() / grid_size) * grid_size)

            # Get current style settings
            color = self.custom_marker_color
            thickness = self.custom_font_size_spinbox.value()

            # --- Convert Label Space points to Current Image Space points ---
            # This transformation needs to map coordinates from the label widget
            # (where the image is potentially scaled and centered) to the
            # coordinate system of the actual self.image object.

            # Get necessary dimensions using floats for precision
            displayed_width = float(self.live_view_label.width())
            displayed_height = float(self.live_view_label.height())
            current_image_width = float(self.image.width()) if self.image and self.image.width() > 0 else 1.0
            current_image_height = float(self.image.height()) if self.image and self.image.height() > 0 else 1.0

            # Determine the scale factor used to fit the image into the label
            if current_image_width <= 0 or current_image_height <= 0: # Safety check
                scale_img_to_label = 1.0
            else:
                scale_x_fit = displayed_width / current_image_width
                scale_y_fit = displayed_height / current_image_height
                scale_img_to_label = min(scale_x_fit, scale_y_fit)

            # Calculate the size of the image as it's actually displayed in the label
            display_img_w = current_image_width * scale_img_to_label
            display_img_h = current_image_height * scale_img_to_label

            # Calculate the centering offset (blank space around the scaled image)
            label_offset_x = (displayed_width - display_img_w) / 2.0
            label_offset_y = (displayed_height - display_img_h) / 2.0

            # Define the inverse transformation function: Label Coords -> Image Coords
            def label_to_image_coords(label_point):
                # label_point is from transform_point (already accounts for zoom/pan)
                # It represents coordinates relative to the label widget's top-left (0,0)
                if scale_img_to_label <= 1e-9: # Avoid division by zero or near-zero
                    return (0.0, 0.0)

                # 1. Calculate position relative to the *displayed image's* top-left corner within the label
                relative_x_in_display = label_point.x() - label_offset_x
                relative_y_in_display = label_point.y() - label_offset_y

                # 2. Scale these relative coordinates back up to the original image's size
                img_x = relative_x_in_display / scale_img_to_label
                img_y = relative_y_in_display / scale_img_to_label

                # 3. Return the calculated image coordinates (clamping can be done by renderer)
                return (img_x, img_y)
            # --- End Coordinate Conversion Function ---

            start_img_coords = label_to_image_coords(start_point)
            end_img_coords = label_to_image_coords(end_point)

            # --- Store Shape Data with Calculated Image Coordinates ---
            shape_data = {
                'type': self.drawing_mode,
                'color': color.name(), # Store color as name/hex string for saving
                'thickness': thickness
            }

            valid_shape = False # Flag to check if shape is worth adding
            if self.drawing_mode == 'line':
                 # Check if start and end points are meaningfully different
                 if abs(start_img_coords[0] - end_img_coords[0]) > 0.5 or abs(start_img_coords[1] - end_img_coords[1]) > 0.5: # Threshold for min length
                     shape_data['start'] = start_img_coords
                     shape_data['end'] = end_img_coords
                     valid_shape = True
                 else:
                     print("Skipping zero-length line after coordinate conversion.")

            elif self.drawing_mode == 'rectangle':
                # Store as top-left corner (x, y) and dimensions (width, height) in image coords
                x_img = min(start_img_coords[0], end_img_coords[0])
                y_img = min(start_img_coords[1], end_img_coords[1])
                w_img = abs(end_img_coords[0] - start_img_coords[0])
                h_img = abs(end_img_coords[1] - start_img_coords[1])
                # Check if width and height are meaningfully positive
                if w_img > 0.5 and h_img > 0.5: # Threshold for min size
                    shape_data['rect'] = (x_img, y_img, w_img, h_img)
                    valid_shape = True
                else:
                    print("Skipping zero-area rectangle after coordinate conversion.")

            # Append to list only if the shape is valid
            if valid_shape:
                self.custom_shapes.append(shape_data)
                self.save_state() # Save state after adding a shape
                self.is_modified = True
                print(f"Added shape: {shape_data}") # Debug: Confirm addition and coords
            else:
                 # If shape wasn't valid, no need to save state or mark modified
                 pass

            # Reset drawing state - calls update_live_view(), which will use the updated custom_shapes list
            self.cancel_drawing_mode()

        else: # Not in drawing mode or no preview available (shouldn't normally happen on release)
             self.cancel_drawing_mode() # Ensure mode is reset anyway
            
    def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
        """Converts QImage to NumPy array, preserving format and handling row padding."""
        if qimage.isNull():
            return None
    
        img_format = qimage.format()
        height = qimage.height()
        width = qimage.width()
        bytes_per_line = qimage.bytesPerLine() # Actual bytes per row (includes padding)
    
    
        ptr = qimage.constBits()
        if not ptr:
            return None
    
        # Ensure ptr is treated as a bytes object of the correct size
        # ptr is a sip.voidptr, needs explicit size setting for buffer protocol
        try:
            ptr.setsize(qimage.byteCount())
        except AttributeError: # Handle cases where setsize might not be needed or available depending on Qt/Python version
            pass # Continue, hoping buffer protocol works based on qimage.byteCount()
    
        buffer_data = bytes(ptr) # Create a bytes object from the pointer
        if len(buffer_data) != qimage.byteCount():
             print(f"qimage_to_numpy: Warning - Buffer size mismatch. Expected {qimage.byteCount()}, got {len(buffer_data)}.")
             # Fallback or error? Let's try to continue but warn.
    
        # --- Grayscale Formats ---
        if img_format == QImage.Format_Grayscale16:
            bytes_per_pixel = 2
            dtype = np.uint16
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                # No padding, simple reshape
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy() # Return a copy
            else:
                # Handle padding
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale16): Row data length mismatch for row {y}. Skipping row.") # Error handling
                return arr
    
        elif img_format == QImage.Format_Grayscale8:
            bytes_per_pixel = 1
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy()
            else:
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale8): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Color Formats ---
        # ARGB32 (often BGRA in memory) & RGBA8888 (often RGBA in memory)
        elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied):
            bytes_per_pixel = 4
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # QImage ARGB32 is typically BGRA in memory order for numpy
                # QImage RGBA8888 is typically RGBA in memory order for numpy
                # Let's return the raw order for now. The caller might need to swap if needed.
                # If Format_ARGB32, the array is likely BGRA.
                # If Format_RGBA8888, the array is likely RGBA.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB32 (often BGRX or RGBX in memory) & RGBX8888
        elif img_format in (QImage.Format_RGB32, QImage.Format_RGBX8888):
            bytes_per_pixel = 4 # Stored with an ignored byte
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # Memory order is often BGRX (Blue, Green, Red, Ignored Alpha/Padding) for RGB32
                # Let's return the 4 channels for now.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB888 (Tightly packed RGB)
        elif img_format == QImage.Format_RGB888:
            bytes_per_pixel = 3
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 3)
                # QImage RGB888 is RGB order in memory
                return arr.copy()
            else:
                arr = np.zeros((height, width, 3), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 3)
                    else:
                        print(f"qimage_to_numpy (RGB888): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Fallback / Conversion Attempt ---
        else:
            # For other formats, try converting to a known format first
            try:
                # Convert to ARGB32 as a robust intermediate format
                qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                if qimage_conv.isNull():
                    print("qimage_to_numpy: Fallback conversion to ARGB32 failed.")
                    # Last resort: try Grayscale8
                    qimage_conv_gray = qimage.convertToFormat(QImage.Format_Grayscale8)
                    if qimage_conv_gray.isNull():
                         return None
                    else:
                        return self.qimage_to_numpy(qimage_conv_gray) # Recursive call with Grayscale8
    
                return self.qimage_to_numpy(qimage_conv) # Recursive call with ARGB32
            except Exception as e:
                print(f"qimage_to_numpy: Error during fallback conversion: {e}")
                return None

    def numpy_to_qimage(self, array: np.ndarray) -> QImage:
        """Converts NumPy array to QImage, selecting appropriate format."""
        if array is None:
            return QImage() # Return invalid QImage
    
        if not isinstance(array, np.ndarray):
            return QImage()
    
        try:
            if array.ndim == 2: # Grayscale
                height, width = array.shape
                if array.dtype == np.uint16:
                    bytes_per_line = width * 2
                    # Ensure data is contiguous
                    contiguous_array = np.ascontiguousarray(array)
                    # Create QImage directly from contiguous data buffer
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                    # Important: QImage doesn't own the buffer by default. Return a copy.
                    return qimg.copy()
                elif array.dtype == np.uint8:
                    bytes_per_line = width * 1
                    contiguous_array = np.ascontiguousarray(array)
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                    return qimg.copy()
                elif np.issubdtype(array.dtype, np.floating):
                     # Assume float is in 0-1 range, scale to uint8
                     img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                     bytes_per_line = width * 1
                     contiguous_array = np.ascontiguousarray(img_norm)
                     qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                     return qimg.copy()
                else:
                    raise TypeError(f"Unsupported grayscale NumPy dtype: {array.dtype}")
    
            elif array.ndim == 3: # Color
                height, width, channels = array.shape
                if channels == 3 and array.dtype == np.uint8:
                    # Assume input is BGR (common from OpenCV), convert to RGB for QImage.Format_RGB888
                    # Make a contiguous copy before conversion
                    contiguous_array_bgr = np.ascontiguousarray(array)
                    rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                    # rgb_image is now contiguous RGB
                    bytes_per_line = width * 3
                    qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                    return qimg.copy()
                elif channels == 4 and array.dtype == np.uint8:
                    # Assume input is BGRA (matches typical QImage.Format_ARGB32 memory layout)
                    contiguous_array_bgra = np.ascontiguousarray(array)
                    bytes_per_line = width * 4
                    qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                    return qimg.copy()
                # Add handling for 16-bit color if needed (less common for display)
                elif channels == 3 and array.dtype == np.uint16:
                     # Downscale to 8-bit for display format
                     array_8bit = (array / 257.0).astype(np.uint8)
                     contiguous_array_bgr = np.ascontiguousarray(array_8bit)
                     rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                     bytes_per_line = width * 3
                     qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                     return qimg.copy()
                elif channels == 4 and array.dtype == np.uint16:
                     # Downscale color channels, keep alpha potentially?
                     array_8bit = (array / 257.0).astype(np.uint8) # Downscales all channels including alpha
                     contiguous_array_bgra = np.ascontiguousarray(array_8bit)
                     bytes_per_line = width * 4
                     qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                     return qimg.copy()
                else:
                     raise TypeError(f"Unsupported color NumPy dtype/channel combination: {array.dtype} / {channels} channels")
            else:
                raise ValueError(f"Unsupported array dimension: {array.ndim}")
    
        except Exception as e:
            traceback.print_exc()
            return QImage() # Return invalid QImage on error

    def get_image_format(self, image=None):
        """Helper to safely get the format of self.image or a provided image."""
        img_to_check = image if image is not None else self.image
        if img_to_check and not img_to_check.isNull():
            return img_to_check.format()
        return None

    def get_compatible_grayscale_format(self, image=None):
        """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
        current_format = self.get_image_format(image)
        if current_format == QImage.Format_Grayscale16:
            return QImage.Format_Grayscale16
        else: # Treat 8-bit grayscale or color as needing 8-bit target
            return QImage.Format_Grayscale8
    # --- END: Helper Functions ---
        
    def quadrilateral_to_rect(self, image, quad_points):
        """
        Warps the quadrilateral region defined by quad_points in the input image
        to a rectangular image using perspective transformation.
        Crucially, preserves the original bit depth (e.g., uint16) of the input image.
        """
        # --- Input Validation ---
        if not image or image.isNull():
            QMessageBox.warning(self, "Warp Error", "Invalid input image provided for warping.")
            return None
        if len(quad_points) != 4:
             QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
             return None
        if cv2 is None:
             QMessageBox.critical(self, "Dependency Error", "OpenCV (cv2) is required for quadrilateral warping but is not installed.")
             return None
     
        # --- Convert QImage to NumPy array (Preserves bit depth) ---
        try:
            img_array = self.qimage_to_numpy(image) # Should return uint16 if input is Format_Grayscale16
            if img_array is None: raise ValueError("NumPy Conversion returned None")
            print(f"quad_to_rect: Input image dtype = {img_array.dtype}, shape = {img_array.shape}") # Debug
        except Exception as e:
            QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
            return None
     
        # --- !!! REMOVED INTERNAL GRAYSCALE CONVERSION !!! ---
        # # OLD code that caused the issue if input was color:
        # if img_array.ndim == 3:
        #      print("Warning: Warping input array is 3D. Converting to grayscale.")
        #      # THIS CONVERSION TO UINT8 WAS THE PROBLEM
        #      color_code = cv2.COLOR_BGR2GRAY if img_array.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
        #      img_array = cv2.cvtColor(img_array, color_code)
        #      # img_array = img_array.astype(np.uint8) # Explicit cast not needed, cvtColor usually returns uint8
        # --- End Removed Block ---
        # Now, warpPerspective will operate on the img_array with its original dtype (e.g., uint16 grayscale or color)
     
     
        # --- Transform points from LiveViewLabel space to Image space ---
        # (Coordinate transformation logic remains the same - assumes direct scaling for now)
        label_width = self.live_view_label.width()
        label_height = self.live_view_label.height()
        current_img_width = image.width()
        current_img_height = image.height()
        scale_x_disp = current_img_width / label_width if label_width > 0 else 1
        scale_y_disp = current_img_height / label_height if label_height > 0 else 1
     
        src_points_img = []
        for point in quad_points:
            x_view, y_view = point.x(), point.y()
            # Account for zoom/pan in LiveViewLabel coordinates
            if self.live_view_label.zoom_level != 1.0:
                x_view = (x_view - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                y_view = (y_view - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            # Scale to image coordinates
            x_image = x_view * scale_x_disp
            y_image = y_view * scale_y_disp
            src_points_img.append([x_image, y_image])
        src_np = np.array(src_points_img, dtype=np.float32)
     
        # --- Define Destination Rectangle ---
        # Calculate dimensions based on the source points
        width_a = np.linalg.norm(src_np[0] - src_np[1])
        width_b = np.linalg.norm(src_np[2] - src_np[3])
        height_a = np.linalg.norm(src_np[0] - src_np[3])
        height_b = np.linalg.norm(src_np[1] - src_np[2])
        max_width = int(max(width_a, width_b))
        max_height = int(max(height_a, height_b))
     
        if max_width <= 0 or max_height <= 0:
             QMessageBox.warning(self, "Warp Error", f"Invalid destination rectangle size ({max_width}x{max_height}).")
             return None
     
        dst_np = np.array([
            [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
        ], dtype=np.float32)
     
        # --- Perform Perspective Warp using OpenCV ---
        try:
            matrix = cv2.getPerspectiveTransform(src_np, dst_np)
            # Determine border color based on input array type
            if img_array.ndim == 3: # Color
                 # Use black (0,0,0) or white (255,255,255) depending on preference. Alpha needs 4 values if present.
                 border_val = (0, 0, 0, 0) if img_array.shape[2] == 4 else (0, 0, 0)
            else: # Grayscale
                 border_val = 0 # Black for grayscale
     
            # Warp the ORIGINAL NumPy array (could be uint8, uint16, color)
            warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                               flags=cv2.INTER_LINEAR, # Linear interpolation is usually good
                                               borderMode=cv2.BORDER_CONSTANT,
                                               borderValue=border_val) # Fill borders appropriately
            print(f"quad_to_rect: Warped array dtype = {warped_array.dtype}, shape = {warped_array.shape}") # Debug
        except Exception as e:
             QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
             traceback.print_exc() # Print full traceback for debugging
             return None
     
        # --- Convert warped NumPy array back to QImage ---
        # numpy_to_qimage should handle uint16 correctly, creating Format_Grayscale16
        try:
            warped_qimage = self.numpy_to_qimage(warped_array) # Handles uint8/uint16/color
            if warped_qimage.isNull(): raise ValueError("numpy_to_qimage conversion failed.")
            print(f"quad_to_rect: Returned QImage format = {warped_qimage.format()}") # Debug
            return warped_qimage
        except Exception as e:
            QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array back to QImage: {e}")
            return None
   # --- END: Modified Warping ---
    
    

        
    def create_menu_bar(self):
        menubar = self.menuBar()

        # --- File Menu ---
        file_menu = menubar.addMenu("&File")
        file_menu.addAction(self.load_action)
        file_menu.addAction(self.save_action)
        file_menu.addAction(self.save_svg_action)
        file_menu.addSeparator()
        file_menu.addAction(self.reset_action)
        file_menu.addSeparator()
        file_menu.addAction(self.exit_action)

        # --- Edit Menu ---
        edit_menu = menubar.addMenu("&Edit")
        edit_menu.addAction(self.undo_action)
        edit_menu.addAction(self.redo_action)
        edit_menu.addSeparator()
        edit_menu.addAction(self.copy_action)
        edit_menu.addAction(self.paste_action)

        # --- View Menu ---
        view_menu = menubar.addMenu("&View")
        view_menu.addAction(self.zoom_in_action)
        view_menu.addAction(self.zoom_out_action)

        # --- About Menu ---
        about_menu = menubar.addMenu("&About")
        # Add "GitHub" action directly here or create it in _create_actions
        github_action = QAction("&GitHub", self)
        github_action.setToolTip("Open the project's GitHub page")
        github_action.triggered.connect(self.open_github)
        about_menu.addAction(github_action)
        
        # self.statusBar().showMessage("Ready")
    def open_github(self):
        # Open the GitHub link in the default web browser
        QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
    
    def create_tool_bar(self):
        """Create the main application toolbar."""
        self.tool_bar = QToolBar("Main Toolbar")
        self.tool_bar.setIconSize(QSize(24, 24))
        self.tool_bar.setMovable(False)
        self.tool_bar.setToolButtonStyle(Qt.ToolButtonIconOnly)

        # Add actions to the toolbar
        self.tool_bar.addAction(self.load_action)
        self.tool_bar.addAction(self.paste_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.save_action)
        self.tool_bar.addAction(self.copy_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.undo_action)
        self.tool_bar.addAction(self.redo_action)
        self.tool_bar.addSeparator()

        # --- Add Zoom and Pan Buttons ---
        self.tool_bar.addAction(self.zoom_in_action)
        self.tool_bar.addAction(self.zoom_out_action)
        self.tool_bar.addSeparator() # Optional separator

        self.tool_bar.addAction(self.pan_left_action)
        self.tool_bar.addAction(self.pan_up_action)   # Group visually
        self.tool_bar.addAction(self.pan_down_action)
        self.tool_bar.addAction(self.pan_right_action)
        # --- End Adding Pan Buttons ---

        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.reset_action)

        # Add the toolbar to the main window
        self.addToolBar(Qt.TopToolBarArea, self.tool_bar)
        
    def zoom_in(self):
        self.live_view_label.zoom_in() # LiveViewLabel handles its own update
        # --- START: Update Pan Button State ---
        enable_pan = self.live_view_label.zoom_level > 1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
        self.update_live_view()
        # --- END: Update Pan Button State ---

    def zoom_out(self):
        self.live_view_label.zoom_out() # LiveViewLabel handles its own update
        # --- START: Update Pan Button State ---
        enable_pan = self.live_view_label.zoom_level > 1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
        self.update_live_view()
        # --- END: Update Pan Button State ---
    
    def enable_standard_protein_mode(self):
        """"Enable mode to define standard protein amounts for creating a standard curve."""
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        # Assign mouse event handlers for bounding box creation
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_measure_protein_mode(self):
        """Enable mode to measure protein quantity using the standard curve."""
        if len(self.quantities) < 2:
            QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)

    def call_live_view(self):
        self.update_live_view()      

    def analyze_bounding_box(self, pil_image_for_dialog, standard): # Input is PIL Image
    
        peak_area = None # Initialize
        # Clear previous results before new analysis
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []

        if standard:
            quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
            if ok and quantity:
                try:
                    quantity_value = float(quantity.split()[0])
                    # Calculate peak area using the PIL data
                    peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL
                    if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                        total_area = sum(peak_area)
                        self.quantities_peak_area_dict[quantity_value] = round(total_area, 3)
                        self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                        self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
                        print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                        # Store the areas for potential later export if needed (though usually total area is used for standards)
                        self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store individual areas too
                    else:
                         print("Peak area calculation cancelled or failed for standard.")
                         self.latest_peak_areas = [] # Ensure it's cleared

                except (ValueError, IndexError) as e:
                    QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                    self.latest_peak_areas = []
                except Exception as e:
                     QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                     self.latest_peak_areas = []
            else:
                print("Standard quantity input cancelled.")
                self.latest_peak_areas = []

        else: # Analyze sample
            # Calculate peak area using the PIL data
            peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL

            if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store latest areas
                print(f"Sample Analysis: Calculated Areas = {self.latest_peak_areas}")

                if len(self.quantities_peak_area_dict) >= 2:
                    # Calculate quantities and store them
                    self.latest_calculated_quantities = self.calculate_unknown_quantity(
                        list(self.quantities_peak_area_dict.values()), # Standard Areas (total)
                        list(self.quantities_peak_area_dict.keys()),   # Standard Quantities
                        self.latest_peak_areas                         # Sample Peak Areas (individual)
                    )
                    print(f"Sample Analysis: Calculated Quantities = {self.latest_calculated_quantities}")

                else:
                    self.latest_calculated_quantities = [] # No quantities calculated

                try:
                    # Display the areas in the text box
                    self.target_protein_areas_text.setText(str(self.latest_peak_areas))
                except Exception as e:
                    print(f"Error displaying sample areas: {e}")
                    self.target_protein_areas_text.setText("Error")
            else:
                 print("Peak area calculation cancelled or failed for sample.")
                 self.target_protein_areas_text.setText("N/A")
                 self.latest_peak_areas = []
                 self.latest_calculated_quantities = []


        # --- UI updates after analysis ---
        self.update_live_view() # Update display
    
    def calculate_peak_area(self, pil_image_for_dialog): # Input is EXPECTED to be PIL Image
        """Opens the PeakAreaDialog for interactive adjustment and area calculation."""

        # --- Validate Input ---
        if pil_image_for_dialog is None:
            print("Error: No PIL Image data provided to calculate_peak_area.")
            return None
        if not isinstance(pil_image_for_dialog, Image.Image):
             # This indicates an error in the calling function (process_sample/standard)
             QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
             return None
        # --- End Validation ---


        # --- Call PeakAreaDialog passing the PIL Image with the 'cropped_data' keyword ---
        # Assumes PeakAreaDialog __init__ expects 'cropped_data' as the first arg (PIL Image type)
        dialog = PeakAreaDialog(
            cropped_data=pil_image_for_dialog, # Pass the received PIL Image
            current_settings=self.peak_dialog_settings,
            persist_checked=self.persist_peak_settings_enabled,
            parent=self
        )

        peak_areas = None
        if dialog.exec_() == QDialog.Accepted:
            peak_areas = dialog.get_final_peak_area()
            if dialog.should_persist_settings():
                self.peak_dialog_settings = dialog.get_current_settings()
                self.persist_peak_settings_enabled = True
            else:
                self.persist_peak_settings_enabled = False
        else:
            print("PeakAreaDialog cancelled.")

        return peak_areas if peak_areas is not None else []
    
    
    def calculate_unknown_quantity(self, standard_total_areas, known_quantities, sample_peak_areas):
        """Calculates unknown quantities based on standards and sample areas. Returns a list of quantities."""
        if not standard_total_areas or not known_quantities or len(standard_total_areas) != len(known_quantities):
            print("Error: Invalid standard data for quantity calculation.")
            return []
        if len(standard_total_areas) < 2:
            print("Error: Need at least 2 standards for regression.")
            return []
        if not sample_peak_areas:
             print("Warning: No sample peak areas provided for quantity calculation.")
             return []

        calculated_quantities = []
        try:
            # Use total standard area vs known quantity for the standard curve
            coefficients = np.polyfit(standard_total_areas, known_quantities, 1)

            # Apply the standard curve to *each individual* sample peak area
            for area in sample_peak_areas:
                val = np.polyval(coefficients, area)
                calculated_quantities.append(round(val, 2))

            # Display the results in a message box (optional, but helpful feedback)
            # QMessageBox.information(self, "Protein Quantification", f"Predicted Quantities: {calculated_quantities} units")

        except Exception as e:
             # QMessageBox.warning(self, "Calculation Error", f"Could not calculate quantities: {e}")
             return [] # Return empty list on error

        return calculated_quantities # <-- Return the list

        
    def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
        """Draw quantity text at the correct position."""
        text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
        painter.drawText(text_position, str(quantity))
    
    def update_standard_protein_quantities(self):
        self.standard_protein_values.text()
    
    def move_tab(self,tab):
        self.tab_widget.setCurrentIndex(tab)
        
    def save_state(self):
        """Save the current state of the image, markers, shapes, and relevant UI/font settings."""
        # Get current state of UI elements related to fonts/colors for custom markers
        custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
        custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

        state = {
            "image": self.image.copy() if self.image else None,
            "left_markers": self.left_markers.copy(),
            "right_markers": self.right_markers.copy(),
            "top_markers": self.top_markers.copy(),
            "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
            "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
            "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
            "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
            "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
            # Standard marker font settings
            "font_family": self.font_family,
            "font_size": self.font_size,
            "font_color": self.font_color, # QColor object is fine for in-memory stack
            "font_rotation": self.font_rotation,
            # Marker positioning shifts
            "left_marker_shift_added": self.left_marker_shift_added,
            "right_marker_shift_added": self.right_marker_shift_added,
            "top_marker_shift_added": self.top_marker_shift_added,
            # Analysis data (keep as before)
            "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
            "quantities": self.quantities.copy(),
            "protein_quantities": self.protein_quantities.copy(),
            "standard_protein_areas": self.standard_protein_areas.copy(),
            # --- NEW: Custom marker font/color states ---
            "custom_marker_color": self.custom_marker_color, # QColor object
            "custom_font_family": custom_font_family, # String from dropdown
            "custom_font_size": custom_font_size,     # Integer from spinbox
            # --- END NEW ---
        }
        self.undo_stack.append(state)
        self.redo_stack.clear() # Clear redo stack on new action
    
    def undo_action_m(self):
        """Undo the last action by restoring the previous state, including font settings."""
        if self.undo_stack:
            # Get current state of UI elements before overwriting self attributes
            current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
            current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

            # Save the current state to the redo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                # Standard marker font settings
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                # Marker positioning shifts
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                # Analysis data
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
                # --- NEW: Custom marker font/color states ---
                "custom_marker_color": self.custom_marker_color,
                "custom_font_family": current_custom_font_family, # Use value read from UI
                "custom_font_size": current_custom_font_size,     # Use value read from UI
                # --- END NEW ---
            }
            self.redo_stack.append(current_state)

            # Restore the previous state from the undo stack
            previous_state = self.undo_stack.pop()
            self.image = previous_state["image"]
            self.left_markers = previous_state["left_markers"]
            self.right_markers = previous_state["right_markers"]
            self.top_markers = previous_state["top_markers"]
            self.custom_markers = previous_state.get("custom_markers", [])
            self.custom_shapes = previous_state.get("custom_shapes", [])
            self.image_before_padding = previous_state["image_before_padding"]
            self.image_contrasted = previous_state["image_contrasted"]
            self.image_before_contrast = previous_state["image_before_contrast"]
            # Restore standard font settings
            self.font_family = previous_state["font_family"]
            self.font_size = previous_state["font_size"]
            self.font_color = previous_state["font_color"]
            self.font_rotation = previous_state["font_rotation"]
            # Restore shifts
            self.left_marker_shift_added = previous_state["left_marker_shift_added"]
            self.right_marker_shift_added = previous_state["right_marker_shift_added"]
            self.top_marker_shift_added = previous_state["top_marker_shift_added"]
            # Restore analysis data
            self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
            self.quantities = previous_state["quantities"]
            self.protein_quantities = previous_state["protein_quantities"]
            self.standard_protein_areas = previous_state["standard_protein_areas"]
            # --- NEW: Restore custom marker font/color states ---
            self.custom_marker_color = previous_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
            restored_custom_font_family = previous_state.get("custom_font_family", "Arial")
            restored_custom_font_size = previous_state.get("custom_font_size", 12)
            # --- END NEW RESTORE ---

            # --- Update UI Elements to reflect restored state ---
            # Standard font UI
            if hasattr(self, 'font_combo_box'):
                self.font_combo_box.blockSignals(True)
                # Find the font; setCurrentText might be less reliable than setCurrentFont
                found_font = False
                for i in range(self.font_combo_box.count()):
                    if self.font_combo_box.itemText(i) == self.font_family:
                        self.font_combo_box.setCurrentIndex(i)
                        found_font = True
                        break
                if not found_font: # Fallback if font name isn't exact match
                    self.font_combo_box.setCurrentFont(QFont(self.font_family))
                self.font_combo_box.blockSignals(False)
            if hasattr(self, 'font_size_spinner'):
                self.font_size_spinner.blockSignals(True)
                self.font_size_spinner.setValue(self.font_size)
                self.font_size_spinner.blockSignals(False)
            if hasattr(self, 'font_color_button'):
                self._update_color_button_style(self.font_color_button, self.font_color)
            if hasattr(self, 'font_rotation_input'):
                self.font_rotation_input.blockSignals(True)
                self.font_rotation_input.setValue(self.font_rotation)
                self.font_rotation_input.blockSignals(False)

            # Slider positions (reflecting restored shifts)
            if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
            if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
            if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

            # --- NEW: Update Custom Marker UI ---
            if hasattr(self, 'custom_marker_color_button'):
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
            if hasattr(self, 'custom_font_type_dropdown'):
                 self.custom_font_type_dropdown.blockSignals(True)
                 # Find the font like standard font combo
                 found_custom_font = False
                 for i in range(self.custom_font_type_dropdown.count()):
                     if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                         self.custom_font_type_dropdown.setCurrentIndex(i)
                         found_custom_font = True
                         break
                 if not found_custom_font: # Fallback
                     self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                 self.custom_font_type_dropdown.blockSignals(False)
                 # Also update the entry box font if needed (optional)
                 # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont())
            if hasattr(self, 'custom_font_size_spinbox'):
                 self.custom_font_size_spinbox.blockSignals(True)
                 self.custom_font_size_spinbox.setValue(restored_custom_font_size)
                 self.custom_font_size_spinbox.blockSignals(False)
            # --- END NEW UI UPDATE ---

            # Update other UI if needed (e.g., analysis text boxes)

            # Refresh display and status bar
            try:
                self._update_preview_label_size()
            except Exception: pass # Ignore errors if label size fails temporarily
            self._update_status_bar()
            self.update_live_view()
            self.is_modified = True # Undoing makes it modified again relative to last save
            
    
    def redo_action_m(self):
        """Redo the last undone action by restoring the next state, including font settings."""
        if self.redo_stack:
            # Get current state of UI elements before overwriting self attributes
            current_custom_font_family = self.custom_font_type_dropdown.currentText() if hasattr(self, 'custom_font_type_dropdown') else "Arial"
            current_custom_font_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12

            # Save the current state to the undo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])],
                "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                # Standard marker font settings
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                # Marker positioning shifts
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                # Analysis data
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
                # --- NEW: Custom marker font/color states ---
                "custom_marker_color": self.custom_marker_color,
                "custom_font_family": current_custom_font_family, # Use value read from UI
                "custom_font_size": current_custom_font_size,     # Use value read from UI
                # --- END NEW ---
            }
            self.undo_stack.append(current_state)

            # Restore the next state from the redo stack
            next_state = self.redo_stack.pop()
            self.image = next_state["image"]
            self.left_markers = next_state["left_markers"]
            self.right_markers = next_state["right_markers"]
            self.top_markers = next_state["top_markers"]
            self.custom_markers = next_state.get("custom_markers", [])
            self.custom_shapes = next_state.get("custom_shapes", [])
            self.image_before_padding = next_state["image_before_padding"]
            self.image_contrasted = next_state["image_contrasted"]
            self.image_before_contrast = next_state["image_before_contrast"]
            # Restore standard font settings
            self.font_family = next_state["font_family"]
            self.font_size = next_state["font_size"]
            self.font_color = next_state["font_color"]
            self.font_rotation = next_state["font_rotation"]
            # Restore shifts
            self.left_marker_shift_added = next_state["left_marker_shift_added"]
            self.right_marker_shift_added = next_state["right_marker_shift_added"]
            self.top_marker_shift_added = next_state["top_marker_shift_added"]
            # Restore analysis data
            self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
            self.quantities = next_state["quantities"]
            self.protein_quantities = next_state["protein_quantities"]
            self.standard_protein_areas = next_state["standard_protein_areas"]
            # --- NEW: Restore custom marker font/color states ---
            self.custom_marker_color = next_state.get("custom_marker_color", QColor(0,0,0)) # Default if missing
            restored_custom_font_family = next_state.get("custom_font_family", "Arial")
            restored_custom_font_size = next_state.get("custom_font_size", 12)
            # --- END NEW RESTORE ---

            # --- Update UI Elements to reflect restored state ---
            # (Identical UI update logic as in undo_action_m)
            # Standard font UI
            if hasattr(self, 'font_combo_box'):
                self.font_combo_box.blockSignals(True)
                found_font = False
                for i in range(self.font_combo_box.count()):
                    if self.font_combo_box.itemText(i) == self.font_family:
                        self.font_combo_box.setCurrentIndex(i)
                        found_font = True
                        break
                if not found_font: self.font_combo_box.setCurrentFont(QFont(self.font_family))
                self.font_combo_box.blockSignals(False)
            if hasattr(self, 'font_size_spinner'):
                self.font_size_spinner.blockSignals(True); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.blockSignals(False)
            if hasattr(self, 'font_color_button'):
                self._update_color_button_style(self.font_color_button, self.font_color)
            if hasattr(self, 'font_rotation_input'):
                self.font_rotation_input.blockSignals(True); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.blockSignals(False)

            # Slider positions
            if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(self.left_marker_shift_added)
            if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(self.right_marker_shift_added)
            if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(self.top_marker_shift_added)

            # --- NEW: Update Custom Marker UI ---
            if hasattr(self, 'custom_marker_color_button'):
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
            if hasattr(self, 'custom_font_type_dropdown'):
                 self.custom_font_type_dropdown.blockSignals(True)
                 found_custom_font = False
                 for i in range(self.custom_font_type_dropdown.count()):
                     if self.custom_font_type_dropdown.itemText(i) == restored_custom_font_family:
                         self.custom_font_type_dropdown.setCurrentIndex(i)
                         found_custom_font = True
                         break
                 if not found_custom_font: self.custom_font_type_dropdown.setCurrentFont(QFont(restored_custom_font_family))
                 self.custom_font_type_dropdown.blockSignals(False)
                 # self.update_marker_text_font(self.custom_font_type_dropdown.currentFont()) # Optional
            if hasattr(self, 'custom_font_size_spinbox'):
                 self.custom_font_size_spinbox.blockSignals(True); self.custom_font_size_spinbox.setValue(restored_custom_font_size); self.custom_font_size_spinbox.blockSignals(False)
            # --- END NEW UI UPDATE ---

            # Update other UI if needed

            # Refresh display and status bar
            try:
                self._update_preview_label_size()
            except Exception: pass
            self._update_status_bar()
            self.update_live_view()
            self.is_modified = True # Redoing makes it modified again relative to last save
            
    def analysis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        # layout.setSpacing(15) # Increase spacing in this tab

        # --- Molecular Weight Prediction ---
        mw_group = QGroupBox("Molecular Weight Prediction")
        mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        mw_layout = QVBoxLayout(mw_group)
        # mw_layout.setSpacing(8)

        self.predict_button = QPushButton("Predict Molecular Weight")
        self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
        self.predict_button.setEnabled(False)  # Initially disabled
        self.predict_button.clicked.connect(self.predict_molecular_weight)
        mw_layout.addWidget(self.predict_button)

        layout.addWidget(mw_group)

        # --- Peak Area / Sample Quantification ---
        quant_group = QGroupBox("Peak Area and Sample Quantification")
        quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        quant_layout = QVBoxLayout(quant_group)
        # quant_layout.setSpacing(8)

        # Area Definition Buttons
        area_def_layout=QHBoxLayout()
        self.btn_define_quad = QPushButton("Define Quad Area")
        self.btn_define_quad.setToolTip(
            "Click 4 corner points to define a region. \n"
            "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
            "Results may differ from Rectangle selection due to warping."
        )
        self.btn_define_quad.clicked.connect(self.enable_quad_mode)
        self.btn_define_rec = QPushButton("Define Rectangle Area")
        self.btn_define_rec.setToolTip(
            "Click and drag to define a rectangular region. \n"
            "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
            "Does not correct for skew."
        )
        self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
        self.btn_sel_rec = QPushButton("Move Selected Area")
        self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
        self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
        area_def_layout.addWidget(self.btn_define_quad)
        area_def_layout.addWidget(self.btn_define_rec)
        area_def_layout.addWidget(self.btn_sel_rec)
        quant_layout.addLayout(area_def_layout)

        # Standard Processing
        std_proc_layout = QHBoxLayout()
        self.btn_process_std = QPushButton("Process Standard Bands")
        self.btn_process_std.setToolTip("Analyze the defined area as a standard lane.\nYou will be prompted for the known quantity.")
        self.btn_process_std.clicked.connect(self.process_standard)
        std_proc_layout.addWidget(self.btn_process_std)
        quant_layout.addLayout(std_proc_layout)

        # Standard Info Display (Read-only makes more sense)
        std_info_layout = QGridLayout()
        std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
        self.standard_protein_values = QLineEdit()
        self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
        self.standard_protein_values.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_values, 0, 1)

        std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
        self.standard_protein_areas_text = QLineEdit()
        self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
        self.standard_protein_areas_text.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
        quant_layout.addLayout(std_info_layout)

        quant_layout.addWidget(self.create_separator()) # Add visual separator

        # Sample Processing
        sample_proc_layout = QHBoxLayout()
        self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
        self.btn_analyze_sample.setToolTip("Analyze the defined area as a sample lane using the standard curve.")
        self.btn_analyze_sample.clicked.connect(self.process_sample)
        sample_proc_layout.addWidget(self.btn_analyze_sample)
        quant_layout.addLayout(sample_proc_layout)


        # Sample Info Display
        sample_info_layout = QGridLayout()
        sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
        self.target_protein_areas_text = QLineEdit()
        self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
        self.target_protein_areas_text.setReadOnly(True)
        sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

        # Add Table Export button next to sample areas
        self.table_export_button = QPushButton("Export Results Table")
        self.table_export_button.setToolTip("Export the analysis results (areas, percentages, quantities) to Excel.")
        self.table_export_button.clicked.connect(self.open_table_window)
        sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
        quant_layout.addLayout(sample_info_layout)

        layout.addWidget(quant_group)

        # --- Clear Button ---
        clear_layout = QHBoxLayout() # Layout to center button maybe
        clear_layout.addStretch()
        self.clear_predict_button = QPushButton("Clear Analysis Markers")
        self.clear_predict_button.setToolTip("Clears MW prediction line and analysis regions.\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
        self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
        layout.addWidget(self.clear_predict_button)
        # clear_layout.addStretch()
        layout.addLayout(clear_layout)


        layout.addStretch()
        return tab
    
    def enable_move_selection_mode(self):
        """Enable mode to move the selected quadrilateral or rectangle."""
        self.live_view_label.mode = "move"
        self.live_view_label.setCursor(Qt.SizeAllCursor)  # Change cursor to indicate move mode
        self.live_view_label.mousePressEvent = self.start_move_selection
        self.live_view_label.mouseReleaseEvent = self.end_move_selection
        self.update_live_view()
        
    def start_move_selection(self, event):
        """Start moving the selection when the mouse is pressed."""
        if self.live_view_label.mode == "move":
            self.live_view_label.draw_edges=False
            # Check if the mouse is near the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), self.live_view_label.quad_points)
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), [
                    QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1])
                ])
            self.live_view_label.drag_start_pos = event.pos()
            self.update_live_view()
        self.live_view_label.mouseMoveEvent = self.move_selection
    
    def move_selection(self, event):
        """Move the selection while the mouse is being dragged."""
        if self.live_view_label.mode == "move" and self.live_view_label.selected_point is not None:
            # Calculate the offset
            offset = event.pos() - self.live_view_label.drag_start_pos
            self.live_view_label.drag_start_pos = event.pos()
            
            # Move the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                for i in range(len(self.live_view_label.quad_points)):
                    self.live_view_label.quad_points[i] += offset
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.bounding_box_preview = (
                    self.live_view_label.bounding_box_preview[0] + offset.x(),
                    self.live_view_label.bounding_box_preview[1] + offset.y(),
                    self.live_view_label.bounding_box_preview[2] + offset.x(),
                    self.live_view_label.bounding_box_preview[3] + offset.y(),
                )
            self.update_live_view()
    
    def end_move_selection(self, event):
        """End moving the selection when the mouse is released."""
        if self.live_view_label.mode == "move":
            self.live_view_label.selected_point = -1            
            self.update_live_view()
            self.live_view_label.draw_edges=True
        self.live_view_label.mouseMoveEvent = None
        
        
             
    def get_nearest_point(self, mouse_pos, points):
        """Get the nearest point to the mouse position."""
        min_distance = float('inf')
        nearest_point = None
        for point in points:
            distance = (mouse_pos - point).manhattanLength()
            if distance < min_distance:
                min_distance = distance
                nearest_point = point
        return nearest_point
    
    def open_table_window(self):
        # Use the latest calculated peak areas
        peak_areas_to_export = self.latest_peak_areas
        calculated_quantities_to_export = self.latest_calculated_quantities

        if not peak_areas_to_export:
            QMessageBox.information(self, "No Data", "No analysis results available to export.")
            return

        standard_dict = self.quantities_peak_area_dict
        # Determine if quantities should be calculated/displayed based on standards
        standard_flag = len(standard_dict) >= 2

        # Open the table window with the latest data
        # Pass the pre-calculated quantities if available
        self.table_window = TableWindow(
            peak_areas_to_export,
            standard_dict,
            standard_flag,
            calculated_quantities_to_export, # Pass the calculated quantities
            self
        )
        self.table_window.show()
    
    def enable_quad_mode(self):
        """Enable mode to define a quadrilateral area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "quad"
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # # Reset mouse event handlers
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        
        # Update the live view
        self.update_live_view()
    
    def enable_rectangle_mode(self):
        """Enable mode to define a rectangle area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "rectangle"
        self.live_view_label.rectangle_points = []  # Clear previous rectangle points
        self.live_view_label.rectangle_start = None  # Reset rectangle start
        self.live_view_label.rectangle_end = None    # Reset rectangle end
        self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Set mouse event handlers for rectangle mode
        self.live_view_label.mousePressEvent = self.start_rectangle
        self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
        
        # Update the live view
        self.update_live_view()

    
    def start_rectangle(self, event):
        """Record the start position of the rectangle."""
        if self.live_view_label.mode == "rectangle":
            self.live_view_label.rectangle_start = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start]
            self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
            self.update_live_view()
    
    def update_rectangle_preview(self, event):
        """Update the rectangle preview as the mouse moves."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Update bounding_box_preview with the rectangle coordinates
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            self.update_live_view()
    
    def finalize_rectangle(self, event):
        """Finalize the rectangle when the mouse is released."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Save the final bounding box preview
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            
            self.live_view_label.mode = None  # Exit rectangle mode
            self.live_view_label.setCursor(Qt.ArrowCursor)
            self.update_live_view()
        
    
    def process_standard(self):
        """Processes the defined region (quad or rect) as a standard."""
        extracted_qimage = None # Will hold the QImage of the extracted region
        self.live_view_label.pan_offset = QPointF(0, 0)
        self.live_view_label.zoom_level=1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
        self.update_live_view()       

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Standard: Quadrilateral")
            # Pass the *current* self.image (could be color or gray)
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Standard: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS from previous fix)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for standard: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=True)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")
        # No else needed, errors handled above
    
    def process_sample(self):
        self.live_view_label.pan_offset = QPointF(0, 0)
        self.live_view_label.zoom_level=1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
        self.update_live_view()  
        """Processes the defined region (quad or rect) as a sample."""
        extracted_qimage = None # Will hold the QImage of the extracted region

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Sample: Quadrilateral")
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Sample: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for sample: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=False)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")


    def convert_qimage_to_grayscale_pil(self, qimg):
        """
        Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
        for use with PeakAreaDialog. Returns None on failure.
        """
        if not qimg or qimg.isNull():
            return None

        fmt = qimg.format()
        pil_img = None

        try:
            # Already grayscale? Convert directly if possible.
            if fmt == QImage.Format_Grayscale16:
                print("Converting Grayscale16 QImage to PIL 'I;16'")
                np_array = self.qimage_to_numpy(qimg)
                if np_array is not None and np_array.dtype == np.uint16:
                    try: pil_img = Image.fromarray(np_array, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(np_array, mode='I')
                else: raise ValueError("Failed NumPy conversion for Grayscale16")
            elif fmt == QImage.Format_Grayscale8:
                print("Converting Grayscale8 QImage to PIL 'L'")
                # Try direct conversion first
                try:
                    pil_img = ImageQt.fromqimage(qimg).convert('L')
                    if pil_img is None: raise ValueError("Direct QImage->PIL(L) failed.")
                except Exception as e_direct:
                    print(f"Direct QImage->PIL(L) failed ({e_direct}), falling back via NumPy.")
                    np_array = self.qimage_to_numpy(qimg)
                    if np_array is not None and np_array.dtype == np.uint8:
                        pil_img = Image.fromarray(np_array, mode='L')
                    else: raise ValueError("Failed NumPy conversion for Grayscale8")
            else: # Color or other format
                print(f"Converting format {fmt} QImage to Grayscale PIL")
                # Use NumPy for robust conversion to 16-bit grayscale intermediate
                np_img = self.qimage_to_numpy(qimg)
                if np_img is None: raise ValueError("NumPy conversion failed for color.")
                if np_img.ndim == 3:
                    gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) # Assume BGR/BGRA input
                    # Convert to 16-bit PIL
                    gray_np_16bit = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                    try: pil_img = Image.fromarray(gray_np_16bit, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(gray_np_16bit, mode='I')
                elif np_img.ndim == 2: # Should have been caught by Grayscale checks, but handle anyway
                     if np_img.dtype == np.uint16:
                         try: pil_img = Image.fromarray(np_img, mode='I;16')
                         except ValueError: pil_img = Image.fromarray(np_img, mode='I')
                     else: # Assume uint8 or other, convert to L
                         pil_img = Image.fromarray(np_img).convert('L')
                else:
                     raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

            if pil_img is None:
                raise ValueError("PIL Image creation failed.")

            print(f"  Successfully converted to PIL Image: mode={pil_img.mode}, size={pil_img.size}")
            return pil_img

        except Exception as e:
            print(f"Error in convert_qimage_to_grayscale_pil: {e}")
            traceback.print_exc()
            return None            
        
        
            
    def combine_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        # Calculate initial slider ranges based on current view size
        render_scale=3
        # Use a sensible default or calculate from screen if view isn't ready
        initial_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 500
        initial_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 500
        render_width = initial_width * render_scale
        render_height = initial_height * render_scale

        # --- Image 1 Group ---
        image1_group = QGroupBox("Image 1 Overlay")
        image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image1_layout = QGridLayout(image1_group) # Use grid for better control layout
        image1_layout.setSpacing(8)

        # Load/Place/Remove buttons
        copy_image1_button = QPushButton("Copy Current Image")
        copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
        copy_image1_button.clicked.connect(self.save_image1) # Keep original name save_image1
        place_image1_button = QPushButton("Place Image 1")
        place_image1_button.setToolTip("Positions Image 1 based on sliders.")
        place_image1_button.clicked.connect(self.place_image1)
        remove_image1_button = QPushButton("Remove Image 1")
        remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
        remove_image1_button.clicked.connect(self.remove_image1)

        image1_layout.addWidget(copy_image1_button, 0, 0)
        image1_layout.addWidget(place_image1_button, 0, 1)
        image1_layout.addWidget(remove_image1_button, 0, 2)

        # Position Sliders
        image1_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image1_left_slider = QSlider(Qt.Horizontal)
        self.image1_left_slider.setRange(-render_width, render_width) # Wider range
        self.image1_left_slider.setValue(0)
        self.image1_left_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2) # Span 2 columns

        image1_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image1_top_slider = QSlider(Qt.Horizontal)
        self.image1_top_slider.setRange(-render_height, render_height) # Wider range
        self.image1_top_slider.setValue(0)
        self.image1_top_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2) # Span 2 columns

        # Resize Slider
        image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image1_resize_slider = QSlider(Qt.Horizontal)
        self.image1_resize_slider.setRange(10, 300)  # Range 10% to 300%
        self.image1_resize_slider.setValue(100)
        self.image1_resize_slider.valueChanged.connect(self.update_live_view)
        self.image1_resize_label = QLabel("100%") # Show current percentage
        self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
        image1_layout.addWidget(self.image1_resize_slider, 3, 1)
        image1_layout.addWidget(self.image1_resize_label, 3, 2)


        layout.addWidget(image1_group)

        # --- Image 2 Group --- (Similar structure to Image 1)
        image2_group = QGroupBox("Image 2 Overlay")
        image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image2_layout = QGridLayout(image2_group)
        image2_layout.setSpacing(8)

        copy_image2_button = QPushButton("Copy Current Image")
        copy_image2_button.clicked.connect(self.save_image2)
        place_image2_button = QPushButton("Place Image 2")
        place_image2_button.clicked.connect(self.place_image2)
        remove_image2_button = QPushButton("Remove Image 2")
        remove_image2_button.clicked.connect(self.remove_image2)
        image2_layout.addWidget(copy_image2_button, 0, 0)
        image2_layout.addWidget(place_image2_button, 0, 1)
        image2_layout.addWidget(remove_image2_button, 0, 2)

        image2_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image2_left_slider = QSlider(Qt.Horizontal)
        self.image2_left_slider.setRange(-render_width, render_width)
        self.image2_left_slider.setValue(0)
        self.image2_left_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)

        image2_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image2_top_slider = QSlider(Qt.Horizontal)
        self.image2_top_slider.setRange(-render_height, render_height)
        self.image2_top_slider.setValue(0)
        self.image2_top_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)

        image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image2_resize_slider = QSlider(Qt.Horizontal)
        self.image2_resize_slider.setRange(10, 300)
        self.image2_resize_slider.setValue(100)
        self.image2_resize_slider.valueChanged.connect(self.update_live_view)
        self.image2_resize_label = QLabel("100%")
        self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
        image2_layout.addWidget(self.image2_resize_slider, 3, 1)
        image2_layout.addWidget(self.image2_resize_label, 3, 2)

        layout.addWidget(image2_group)

        # --- Finalize Button ---
        finalize_layout = QHBoxLayout()
        finalize_layout.addStretch()
        finalize_button = QPushButton("Rasterize Image")
        finalize_button.setToolTip("Permanently merges the placed overlays onto the main image. This action cannot be undone easily.")
        finalize_button.clicked.connect(self.finalize_combined_image)
        layout.addWidget(finalize_button)
        finalize_layout.addStretch()
        layout.addLayout(finalize_layout)

        layout.addStretch() # Push content up
        return tab
    
    def remove_image1(self):
        """Remove Image 1 and reset its sliders."""
        if hasattr(self, 'image1'):
            # del self.image1
            # del self.image1_original
            try:
                del self.image1_position
            except:
                pass
            self.image1_left_slider.setValue(0)
            self.image1_top_slider.setValue(0)
            self.image1_resize_slider.setValue(100)
            self.update_live_view()
    
    def remove_image2(self):
        """Remove Image 2 and reset its sliders."""
        if hasattr(self, 'image2'):
            # del self.image2
            # del self.image2_original
            try:
                del self.image2_position
            except:
                pass
            self.image2_left_slider.setValue(0)
            self.image2_top_slider.setValue(0)
            self.image2_resize_slider.setValue(100)
            self.update_live_view()
    
    def save_image1(self):
        if self.image:
            self.image1 = self.image.copy()
            self.image1_original = self.image1.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 1 copied.")
    
    def save_image2(self):
        if self.image:
            self.image2 = self.image.copy()
            self.image2_original = self.image2.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 2 copied.")
    
    def place_image1(self):
        if hasattr(self, 'image1'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            self.update_live_view()
    
    def place_image2(self):
        if hasattr(self, 'image2'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            self.update_live_view()
    
    
    
    def finalize_combined_image(self):
        """
        Rasterizes overlays and annotations onto the image by generating
        a high-resolution canvas identical to the preview rendering process
        (but without guides/previews) and assigning it as the new self.image.
        This ensures visual consistency between preview and final result.
        The final image resolution might be higher than the original if render_scale > 1.
        """
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Warning", "No base image loaded.")
            return

        # --- Save state before modifying ---
        self.save_state()

        # --- 1. Determine Render/Canvas Dimensions (Same as Preview) ---
        render_scale = 3 # Use the same scale factor as the preview
        try:
            # Use current view dimensions to determine target render size
            # Fallback if view is not ready
            view_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 600
            view_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 400
            target_render_width = view_width * render_scale
            target_render_height = view_height * render_scale
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Could not calculate render dimensions: {e}")
            return

        # --- 2. Prepare the Base Image for Rendering ---
        # Scale the *current* self.image to fit within the target render dimensions,
        # just like it's done at the start of update_live_view before calling render_image_on_canvas.
        # Check if self.image is valid before scaling
        if self.image.isNull() or self.image.width() <= 0 or self.image.height() <= 0:
             QMessageBox.critical(self, "Error", "Current base image is invalid before scaling.")
             return

        scaled_image_for_render = self.image.scaled(
            target_render_width,
            target_render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        if scaled_image_for_render.isNull():
            QMessageBox.critical(self, "Error", "Failed to scale base image for final rendering.")
            return

        # --- 3. Create the High-Resolution Target Canvas ---
        # Use ARGB32_Premultiplied for safe drawing with alpha
        final_canvas = QImage(target_render_width, target_render_height, QImage.Format_ARGB32_Premultiplied)
        if final_canvas.isNull():
            QMessageBox.critical(self, "Error", "Failed to create final high-resolution canvas.")
            return
        final_canvas.fill(Qt.transparent) # Fill with transparent background

        # --- 4. Render onto the High-Res Canvas using the Preview Logic ---
        # Call the *exact same* rendering function used for the preview,
        # but disable drawing of guides and temporary previews.
        # Pass the scaled base image prepared in step 2.
        # Pass render_scale so elements inside render_image_on_canvas scale correctly.
        try:
            # x_start and y_start are typically 0 when rendering the current self.image
            # unless cropping offsets need to be handled differently, but render_image_on_canvas
            # should already handle positioning based on the scaled_image provided.
            self.render_image_on_canvas(
                canvas=final_canvas,
                scaled_image=scaled_image_for_render,
                x_start=0, # Start position relative to the scaled_image itself
                y_start=0,
                render_scale=render_scale,
                draw_guides=False # IMPORTANT: Disable guides for final output
            )
            # Inside render_image_on_canvas, ensure shape previews are also skipped if needed,
            # although technically they shouldn't exist when finalize is called.
            # Consider adding a flag to render_image_on_canvas like `is_finalizing=True`
            # to explicitly skip previews if necessary.

        except Exception as e:
             QMessageBox.critical(self, "Render Error", f"Failed during final rendering: {e}")
             traceback.print_exc()
             return

        # --- 5. Assign the High-Resolution Canvas as the New Image ---
        if final_canvas.isNull():
             QMessageBox.critical(self, "Error", "Final rendered canvas became invalid.")
             return

        self.image = final_canvas # The new self.image IS the high-res render
        self.is_modified = True

        # --- 6. Update Backups and State ---
        # The backups now store the high-resolution rendered image
        self.image_before_padding = self.image.copy() # New baseline includes baked-in elements
        self.image_contrasted = self.image.copy()
        self.image_before_contrast = self.image.copy() # Reset contrast baseline
        self.image_padded = True # Consider the result "padded" with overlays

        # --- 7. Cleanup and UI Update ---
        self.remove_image1() # Remove buffer overlays
        self.remove_image2()

        self._update_preview_label_size() # Update label based on potentially larger image
        self._update_status_bar()         # Reflect new dimensions/depth
        self._update_marker_slider_ranges() # Adjust ranges for new dimensions
        self.update_live_view()           # Refresh display

        QMessageBox.information(self, "Success", "The overlays and annotations have been rasterized onto the image.\n(Note: Final image resolution might be higher than original)")
    
        
    


    def font_and_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Add spacing between groups

        # --- Font Options Group ---
        font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
        font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        font_options_layout = QGridLayout(font_options_group)
        font_options_layout.setSpacing(8)

        # Font type
        font_type_label = QLabel("Font Family:")
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setEditable(False)
        self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
        self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_type_label, 0, 0)
        font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

        # Font size
        font_size_label = QLabel("Font Size:")
        self.font_size_spinner = QSpinBox()
        self.font_size_spinner.setRange(6, 72)  # Adjusted range
        self.font_size_spinner.setValue(self.font_size) # Use initialized value
        self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_size_label, 1, 0)
        font_options_layout.addWidget(self.font_size_spinner, 1, 1)

        # Font color
        self.font_color_button = QPushButton("Font Color")
        self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
        self.font_color_button.clicked.connect(self.select_font_color)
        self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
        font_options_layout.addWidget(self.font_color_button, 1, 2)

        # Font rotation (Top/Bottom)
        font_rotation_label = QLabel("Top Label Rotation:") # Specific label
        self.font_rotation_input = QSpinBox()
        self.font_rotation_input.setRange(-180, 180)
        self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
        self.font_rotation_input.setSuffix(" °") # Add degree symbol
        self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_rotation_label, 2, 0)
        font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

        layout.addWidget(font_options_group)


        # --- Image Adjustments Group ---
        img_adjust_group = QGroupBox("Image Adjustments")
        img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        img_adjust_layout = QGridLayout(img_adjust_group)
        img_adjust_layout.setSpacing(8)

        # Contrast Sliders with Value Labels
        high_contrast_label = QLabel("Brightness:") # Renamed for clarity
        self.high_slider = QSlider(Qt.Horizontal)
        self.high_slider.setRange(0, 200) # Range 0 to 200%
        self.high_slider.setValue(100) # Default 100%
        self.high_slider.valueChanged.connect(self.update_image_contrast)
        self.high_value_label = QLabel("1.00") # Display factor
        self.high_slider.valueChanged.connect(lambda val, lbl=self.high_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(high_contrast_label, 0, 0)
        img_adjust_layout.addWidget(self.high_slider, 0, 1)
        img_adjust_layout.addWidget(self.high_value_label, 0, 2)

        low_contrast_label = QLabel("Contrast:") # Renamed for clarity
        self.low_slider = QSlider(Qt.Horizontal)
        self.low_slider.setRange(0, 200) # Range 0 to 100%
        self.low_slider.setValue(100) # Default 100%
        self.low_slider.valueChanged.connect(self.update_image_contrast)
        self.low_value_label = QLabel("1.00") # Display factor
        self.low_slider.valueChanged.connect(lambda val, lbl=self.low_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(low_contrast_label, 1, 0)
        img_adjust_layout.addWidget(self.low_slider, 1, 1)
        img_adjust_layout.addWidget(self.low_value_label, 1, 2)


        # Gamma Adjustment Slider
        gamma_label = QLabel("Gamma:")
        self.gamma_slider = QSlider(Qt.Horizontal)
        self.gamma_slider.setRange(10, 300)  # Range 0.1 to 3.0
        self.gamma_slider.setValue(100)  # Default 1.0 (100%)
        self.gamma_slider.valueChanged.connect(self.update_image_gamma)
        self.gamma_value_label = QLabel("1.00") # Display factor
        self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(gamma_label, 2, 0)
        img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
        img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


        # Separator
        img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

        # Action Buttons
        btn_layout = QHBoxLayout()
        self.bw_button = QPushButton("Grayscale")
        self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
        self.bw_button.clicked.connect(self.convert_to_black_and_white)
        invert_button = QPushButton("Invert")
        invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
        invert_button.clicked.connect(self.invert_image)
        reset_button = QPushButton("Reset Adjustments")
        reset_button.setToolTip("Reset Brightness, Contrast, and Gamma sliders to default.\n(Does not undo Grayscale or Invert)")
        reset_button.clicked.connect(self.reset_gamma_contrast)
        btn_layout.addWidget(self.bw_button)
        btn_layout.addWidget(invert_button)
        btn_layout.addStretch() # Push reset button to the right
        btn_layout.addWidget(reset_button)

        img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

        layout.addWidget(img_adjust_group)
        layout.addStretch() # Push groups up
        return tab

    
    def _update_color_button_style(self, button, color):
        """ Helper to update button background color preview """
        if color.isValid():
            button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
        else:
            button.setStyleSheet("") # Reset to default stylesheet
    
    def create_separator(self):
        """Creates a horizontal separator line."""
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        return line
    
    
    def create_cropping_tab(self):
        """Create the Cropping tab with Rectangle Draw and Slider options."""
        tab = QWidget()
        layout = QVBoxLayout(tab)


        # --- Alignment Group (Keep as is - assume it's defined elsewhere) ---
        alignment_params_group = QGroupBox("Alignment Options")
        alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        # Use a QVBoxLayout for the main group layout
        alignment_group_layout = QVBoxLayout(alignment_params_group)
        alignment_group_layout.setSpacing(8) # Spacing between the two rows

        # --- Row 1: Guides, Rotation Controls ---
        rotation_controls_layout = QHBoxLayout()
        rotation_controls_layout.setSpacing(6) # Spacing within the row

        self.show_guides_label = QLabel("Show Guide Lines:")
        self.show_guides_checkbox = QCheckBox("", self)
        self.show_guides_checkbox.setChecked(False)
        self.show_guides_checkbox.stateChanged.connect(self.update_live_view)

        self.orientation_label = QLabel("Rotation Angle (0.00°)")
        # Make label width flexible but give it a minimum
        self.orientation_label.setMinimumWidth(130)
        # Allow label to shrink/grow slightly if needed:
        self.orientation_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)

        self.orientation_slider = QSlider(Qt.Horizontal)
        self.orientation_slider.setRange(-3600, 3600)
        self.orientation_slider.setValue(0)
        self.orientation_slider.setSingleStep(1)
        self.orientation_slider.valueChanged.connect(self._update_rotation_label)
        # Update preview only when slider is released for performance
        self.orientation_slider.valueChanged.connect(self.update_live_view)
        # Make slider expand to take available space
        self.orientation_slider.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.align_button = QPushButton("Apply Rotation")
        self.align_button.clicked.connect(self.align_image)

        self.reset_align_button = QPushButton("Reset Rotation")
        self.reset_align_button.clicked.connect(self.reset_align_image)

        # Add widgets to the first row layout
        rotation_controls_layout.addWidget(self.show_guides_label)
        rotation_controls_layout.addWidget(self.show_guides_checkbox)
        rotation_controls_layout.addSpacing(10) # Add a small visual gap
        rotation_controls_layout.addWidget(self.orientation_label)
        rotation_controls_layout.addWidget(self.orientation_slider) # Let slider expand
        rotation_controls_layout.addWidget(self.align_button)
        rotation_controls_layout.addWidget(self.reset_align_button)

        # Add the first row layout to the main group layout
        alignment_group_layout.addLayout(rotation_controls_layout)

        # --- Row 2: Flip Controls ---
        flip_controls_layout = QHBoxLayout()
        flip_controls_layout.setSpacing(6) # Spacing between flip buttons

        self.flip_vertical_button = QPushButton("Flip Vertical")
        self.flip_vertical_button.clicked.connect(self.flip_vertical)
        # Make buttons expand equally to fill width
        self.flip_vertical_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.flip_horizontal_button = QPushButton("Flip Horizontal")
        self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
        # Make buttons expand equally to fill width
        self.flip_horizontal_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        # Add widgets to the second row layout
        # No stretch needed here as buttons expand
        flip_controls_layout.addWidget(self.flip_vertical_button)
        flip_controls_layout.addWidget(self.flip_horizontal_button)

        # Add the second row layout to the main group layout
        alignment_group_layout.addLayout(flip_controls_layout)
        
        # guide_layout.addStretch()
        

        alignment_params_group.setLayout(alignment_group_layout)
        layout.addWidget(alignment_params_group)
        # --- End Alignment Group ---

        # --- Skew Fix Group (Keep as is - assume it's defined elsewhere) ---
        taper_skew_group = QGroupBox("Skew Fix")
        taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        # ... (Add skew widgets here) ...
        # Example structure:
        taper_skew_layout = QHBoxLayout()
        self.taper_skew_label = QLabel("Tapering Skew (0.00)")
        self.taper_skew_label.setFixedWidth(150)
        self.taper_skew_slider = QSlider(Qt.Horizontal)
        self.taper_skew_slider.setRange(-70, 70)
        self.taper_skew_slider.setValue(0)
        self.taper_skew_slider.valueChanged.connect(self._update_skew_label)
        self.taper_skew_slider.sliderReleased.connect(self.update_live_view)
        self.skew_button = QPushButton("Apply Skew")
        self.skew_button.clicked.connect(self.update_skew)
        taper_skew_layout.addWidget(self.taper_skew_label)
        taper_skew_layout.addWidget(self.taper_skew_slider)
        taper_skew_layout.addWidget(self.skew_button)
        taper_skew_group.setLayout(taper_skew_layout)
        layout.addWidget(taper_skew_group)
        # --- End Skew Fix Group ---


        # --- START: Modified Cropping Group ---
        cropping_params_group = QGroupBox("Cropping Options")
        cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        cropping_layout = QVBoxLayout(cropping_params_group)      
        

        # --- Draw Rectangle Button ---
        self.draw_crop_rect_button = QPushButton("Draw Crop Rectangle")
        self.draw_crop_rect_button.setToolTip("Click and drag on the image to define the crop area.\nThis will enable the sliders below for fine-tuning.")
        self.draw_crop_rect_button.setCheckable(True)
        self.draw_crop_rect_button.clicked.connect(self.toggle_rectangle_crop_mode)
        self.apply_crop_button = QPushButton("Apply Crop")
        self.apply_crop_button.setToolTip("Apply the defined crop (rectangle or sliders).")
        self.apply_crop_button.clicked.connect(self.update_crop)
        cropping_layout.addWidget(self.draw_crop_rect_button)
        cropping_layout.addWidget(self.apply_crop_button)

        # --- Separator ---
        cropping_layout.addWidget(self.create_separator())

        # --- Sliders for Fine-Tuning ---
        crop_slider_layout = QGridLayout()

        # --- Define Slider Range and Precision ---
        self.crop_slider_min = 0
        self.crop_slider_max = 10000 # Represents 0.00% to 100.00%
        self.crop_slider_precision_factor = 100.0 # Divide slider value by this

        # Helper to create label for slider value display
        def create_value_label(initial_value=0.0):
            # Format to 2 decimal places for hundredths of percent
            lbl = QLabel(f"{initial_value:.2f}%")
            lbl.setMinimumWidth(50) # Ensure space for "100.00%"
            lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return lbl

        # --- Left Crop Slider (X Start) ---
        crop_x_start_label = QLabel("Crop Left:")
        self.crop_x_start_slider = QSlider(Qt.Horizontal)
        self.crop_x_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
        self.crop_x_start_slider.setValue(self.crop_slider_min)      # Default: Start at 0.00%
        self.crop_x_start_slider.setToolTip("Adjust the left edge of the crop area (enabled after drawing).")
        self.crop_x_start_value_label = create_value_label(0.00)
        self.crop_x_start_slider.valueChanged.connect(
            lambda val, lbl=self.crop_x_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
        )
        self.crop_x_start_slider.valueChanged.connect(self._update_crop_from_sliders)
        self.crop_x_start_slider.setEnabled(False) # Initially disabled
        crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
        crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
        crop_slider_layout.addWidget(self.crop_x_start_value_label, 0, 2)

        # --- Right Crop Slider (X End) ---
        crop_x_end_label = QLabel("Crop Right:")
        self.crop_x_end_slider = QSlider(Qt.Horizontal)
        self.crop_x_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
        self.crop_x_end_slider.setValue(self.crop_slider_max)    # Default: End at 100.00%
        self.crop_x_end_slider.setToolTip("Adjust the right edge of the crop area (enabled after drawing).")
        self.crop_x_end_value_label = create_value_label(100.00)
        self.crop_x_end_slider.valueChanged.connect(
            lambda val, lbl=self.crop_x_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
        )
        self.crop_x_end_slider.valueChanged.connect(self._update_crop_from_sliders)
        self.crop_x_end_slider.setEnabled(False) # Initially disabled
        crop_slider_layout.addWidget(crop_x_end_label, 0, 3)
        crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 4)
        crop_slider_layout.addWidget(self.crop_x_end_value_label, 0, 5)

        # --- Top Crop Slider (Y Start) ---
        crop_y_start_label = QLabel("Crop Top:")
        self.crop_y_start_slider = QSlider(Qt.Horizontal)
        self.crop_y_start_slider.setRange(self.crop_slider_min, self.crop_slider_max)
        self.crop_y_start_slider.setValue(self.crop_slider_min)
        self.crop_y_start_slider.setToolTip("Adjust the top edge of the crop area (enabled after drawing).")
        self.crop_y_start_value_label = create_value_label(0.00)
        self.crop_y_start_slider.valueChanged.connect(
            lambda val, lbl=self.crop_y_start_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
        )
        self.crop_y_start_slider.valueChanged.connect(self._update_crop_from_sliders)
        self.crop_y_start_slider.setEnabled(False) # Initially disabled
        crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
        crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
        crop_slider_layout.addWidget(self.crop_y_start_value_label, 1, 2)

        # --- Bottom Crop Slider (Y End) ---
        crop_y_end_label = QLabel("Crop Bottom:")
        self.crop_y_end_slider = QSlider(Qt.Horizontal)
        self.crop_y_end_slider.setRange(self.crop_slider_min, self.crop_slider_max)
        self.crop_y_end_slider.setValue(self.crop_slider_max)
        self.crop_y_end_slider.setToolTip("Adjust the bottom edge of the crop area (enabled after drawing).")
        self.crop_y_end_value_label = create_value_label(100.00)
        self.crop_y_end_slider.valueChanged.connect(
            lambda val, lbl=self.crop_y_end_value_label: lbl.setText(f"{val / self.crop_slider_precision_factor:.2f}%")
        )
        self.crop_y_end_slider.valueChanged.connect(self._update_crop_from_sliders)
        self.crop_y_end_slider.setEnabled(False) # Initially disabled
        crop_slider_layout.addWidget(crop_y_end_label, 1, 3)
        crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 4)
        crop_slider_layout.addWidget(self.crop_y_end_value_label, 1, 5)

        # Make sliders expand horizontally
        crop_slider_layout.setColumnStretch(1, 1)
        crop_slider_layout.setColumnStretch(4, 1)
        cropping_layout.addLayout(crop_slider_layout)


        layout.addWidget(cropping_params_group)
        # --- END: Modified Cropping Group ---

        layout.addStretch()
        return tab
    
    def _update_rotation_label(self, value):
        """Updates the rotation label text."""
        if hasattr(self, 'orientation_label'):
            orientation = value / 20.0
            self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

    def _update_skew_label(self, value):
        """Updates the skew label text."""
        if hasattr(self, 'taper_skew_label'):
            taper_value = value / 100.0
            self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

    def _update_crop_from_sliders(self):
        """Updates crop rectangle based on slider percentages and ensures labels update."""
        if not self.image or self.image.isNull():
            return
    
        # Read integer slider values
        try:
            x_start_val = self.crop_x_start_slider.value()
            x_end_val = self.crop_x_end_slider.value()
            y_start_val = self.crop_y_start_slider.value()
            y_end_val = self.crop_y_end_slider.value()
        except AttributeError:
            print("Error: Crop sliders not fully initialized.")
            return
    
        # Convert to percentage
        x_start_perc = x_start_val / self.crop_slider_precision_factor
        x_end_perc = x_end_val / self.crop_slider_precision_factor
        y_start_perc = y_start_val / self.crop_slider_precision_factor
        y_end_perc = y_end_val / self.crop_slider_precision_factor
    
        # Validation: Ensure Start <= End
        x_start = min(x_start_perc, x_end_perc)
        x_end = max(x_start_perc, x_end_perc)
        y_start = min(y_start_perc, y_end_perc)
        y_end = max(y_start_perc, y_end_perc)
    
        # Convert corrected percentages back to integer slider values
        x_start_val_corr = int(round(x_start * self.crop_slider_precision_factor))
        x_end_val_corr = int(round(x_end * self.crop_slider_precision_factor))
        y_start_val_corr = int(round(y_start * self.crop_slider_precision_factor))
        y_end_val_corr = int(round(y_end * self.crop_slider_precision_factor))
    
        # Reflect potentially swapped values back in sliders (only if different)
        # Use a flag to track if signals were blocked to avoid redundant label updates later
        signals_were_blocked = False
        sliders_to_check = [
            (self.crop_x_start_slider, x_start_val_corr), (self.crop_x_end_slider, x_end_val_corr),
            (self.crop_y_start_slider, y_start_val_corr), (self.crop_y_end_slider, y_end_val_corr)
        ]
        for slider, correct_value in sliders_to_check:
            if slider.value() != correct_value:
                slider.blockSignals(True)
                slider.setValue(correct_value)
                slider.blockSignals(False)
                signals_were_blocked = True # Record that we manually set a value
    
        # --- Explicitly Update Labels AFTER potential corrections ---
        # This ensures labels always reflect the validated slider state, even if
        # the correction logic blocked the initial valueChanged signal for the label lambda.
        # Use the corrected integer values derived above.
        try:
            self.crop_x_start_value_label.setText(f"{x_start_val_corr / self.crop_slider_precision_factor:.2f}%")
            self.crop_x_end_value_label.setText(f"{x_end_val_corr / self.crop_slider_precision_factor:.2f}%")
            self.crop_y_start_value_label.setText(f"{y_start_val_corr / self.crop_slider_precision_factor:.2f}%")
            self.crop_y_end_value_label.setText(f"{y_end_val_corr / self.crop_slider_precision_factor:.2f}%")
        except AttributeError:
            print("Error: Crop value labels not fully initialized for explicit update.")
            # Handle case where labels might not exist yet if called too early
            pass
        # --- End Explicit Label Update ---
    
    
        # Deactivate drawing mode if sliders are used
        if self.crop_rectangle_mode:
            self.cancel_rectangle_crop_mode()
    
        # Calculate Image Coordinates using corrected percentages
        img_w = float(self.image.width())
        img_h = float(self.image.height())
        if img_w <= 0 or img_h <= 0: return
    
        img_x = img_w * (x_start / 100.0)
        img_y = img_h * (y_start / 100.0)
        img_crop_w = img_w * ((x_end - x_start) / 100.0)
        img_crop_h = img_h * ((y_end - y_start) / 100.0)
    
        # Store the calculated *image* coordinates
        self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))
    
        # Calculate View Coordinates for Preview
        label_w = float(self.live_view_label.width())
        label_h = float(self.live_view_label.height())
        if label_w <= 0 or label_h <= 0: return
    
        scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
        if scale_factor <= 1e-9: scale_factor = 1
    
        display_img_w = img_w * scale_factor
        display_img_h = img_h * scale_factor
        display_offset_x = (label_w - display_img_w) / 2.0
        display_offset_y = (label_h - display_img_h) / 2.0
    
        view_x_start = display_offset_x + img_x * scale_factor
        view_y_start = display_offset_y + img_y * scale_factor
        view_x_end = view_x_start + img_crop_w * scale_factor
        view_y_end = view_y_start + img_crop_h * scale_factor
    
        # Update the visual preview rectangle
        self.live_view_label.crop_rect_final_view = QRectF(
            QPointF(view_x_start, view_y_start),
            QPointF(view_x_end, view_y_end)
        ).normalized()
        self.live_view_label.drawing_crop_rect = False
    
        self.update_live_view()
    
    def toggle_rectangle_crop_mode(self, checked):
        """Toggles the rectangle crop drawing mode based on button state."""
        if checked:
            self.enable_rectangle_crop_mode()
        else:
            self.cancel_rectangle_crop_mode()

    def enable_rectangle_crop_mode(self):
        """Activates the rectangle drawing mode for cropping."""
        if self.crop_rectangle_mode: return # Already active

        # Deactivate other interactive modes if necessary (e.g., marker placement)
        self.marker_mode = None
        # self.live_view_label.setCursor(Qt.ArrowCursor) # Set cursor below

        self.crop_rectangle_mode = True
        self.crop_rectangle_coords = None # Clear previous coords
        self.live_view_label.clear_crop_preview() # Clear visual preview

        # Update button state (ensure it's checked)
        self.draw_crop_rect_button.setChecked(True)

        # Setup LiveViewLabel for drawing
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.mousePressEvent = self.start_crop_rectangle
        self.live_view_label.mouseMoveEvent = self.update_crop_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_crop_rectangle
        # print("Rectangle Crop Mode Enabled") # Debug

    def cancel_rectangle_crop_mode(self):
        """Deactivates the rectangle drawing mode."""
        if not self.crop_rectangle_mode: return # Already inactive

        self.crop_rectangle_mode = False
        self.crop_rect_start_view = None # Clear temp start point
        # Keep self.crop_rectangle_coords if user might still apply it
        # Keep self.live_view_label.crop_rect_final_view for visual feedback

        # Update button state (ensure it's unchecked)
        if hasattr(self, 'draw_crop_rect_button'): # Check if button exists
             self.draw_crop_rect_button.setChecked(False)

        # Reset LiveViewLabel
        self.live_view_label.setCursor(Qt.ArrowCursor)

        # --- CORRECTED RESET ---
        # Reset mouse handlers by setting them to None, allowing default behavior
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        # --- END CORRECTION ---


        # Optionally clear the visual preview immediately on cancel
        # self.live_view_label.clear_crop_preview()
        # self.crop_rectangle_coords = None # Also clear coords if cancel means discard

        # print("Rectangle Crop Mode Disabled") # Debug
        self.update_live_view() # Refresh display (might clear preview if cleared above)

    def start_crop_rectangle(self, event):
        """Handles mouse press to start drawing the crop rectangle."""
        if not self.crop_rectangle_mode:
            # If not in crop mode, pass event to the base class or intended handler
            super(LiveViewLabel, self.live_view_label.__class__).mousePressEvent(self.live_view_label, event)
            return

        if event.button() == Qt.LeftButton:
            # Get start point in *view* coordinates (unzoomed)
            self.crop_rect_start_view = self.live_view_label.transform_point(event.pos())
            # Tell LiveViewLabel to start drawing the preview
            self.live_view_label.start_crop_preview(self.crop_rect_start_view)

    def update_crop_rectangle_preview(self, event):
        """Handles mouse move to update the crop rectangle preview."""
        if not self.crop_rectangle_mode or not self.crop_rect_start_view:
             # If not in crop mode or not dragging, pass event along
             super(LiveViewLabel, self.live_view_label.__class__).mouseMoveEvent(self.live_view_label, event)
             return

        # Check if left button is held down (QApplication.mouseButtons() might be needed)
        if event.buttons() & Qt.LeftButton:
            current_point_view = self.live_view_label.transform_point(event.pos())
            # Tell LiveViewLabel to update the preview
            self.live_view_label.update_crop_preview(self.crop_rect_start_view, current_point_view)

    def finalize_crop_rectangle(self, event):
        """Handles mouse release to finalize the crop rectangle and updates sliders/labels."""
        if not self.crop_rectangle_mode or not self.crop_rect_start_view:
            # If not in crop mode or drag never started, pass event along
            # Use super() with the correct class hierarchy if LiveViewLabel inherits directly
            # If LiveViewLabel is just an instance variable, this super() call is incorrect.
            # Assuming LiveViewLabel is a custom class inheriting QLabel:
            # super(LiveViewLabel, self.live_view_label).mouseReleaseEvent(event)
            # If LiveViewLabel is just a QLabel instance, default behaviour is likely sufficient:
            if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                 QLabel.mouseReleaseEvent(self.live_view_label, event) # Call base QLabel handler if needed
            return

        if event.button() == Qt.LeftButton:
            end_point_view = self.live_view_label.transform_point(event.pos())
            start_point_view = self.crop_rect_start_view

            # Finalize the visual preview in LiveViewLabel
            self.live_view_label.finalize_crop_preview(start_point_view, end_point_view)

            try:
                # --- Coordinate Conversion (same as before) ---
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w = float(self.image.width())
                img_h = float(self.image.height())
                label_w = float(self.live_view_label.width())
                label_h = float(self.live_view_label.height())

                if img_w <= 0 or img_h <= 0 or label_w <= 0 or label_h <= 0:
                    raise ValueError("Invalid image or label dimensions.")

                scale_factor = min(label_w / img_w, label_h / img_h)
                if scale_factor <= 1e-9: raise ValueError("Scale factor too small.")

                display_img_w = img_w * scale_factor
                display_img_h = img_h * scale_factor
                display_offset_x = (label_w - display_img_w) / 2.0
                display_offset_y = (label_h - display_img_h) / 2.0

                start_x_img = (start_point_view.x() - display_offset_x) / scale_factor
                start_y_img = (start_point_view.y() - display_offset_y) / scale_factor
                end_x_img = (end_point_view.x() - display_offset_x) / scale_factor
                end_y_img = (end_point_view.y() - display_offset_y) / scale_factor

                img_x = min(start_x_img, end_x_img)
                img_y = min(start_y_img, end_y_img)
                img_crop_w = abs(end_x_img - start_x_img)
                img_crop_h = abs(end_y_img - start_y_img)
                # --- End Coordinate Conversion ---

                if img_crop_w < 1 or img_crop_h < 1:
                     # Handle case where drawn rectangle is too small
                     self.crop_rectangle_coords = None
                     self.live_view_label.clear_crop_preview()
                     # Keep sliders disabled
                     self.crop_x_start_slider.setEnabled(False)
                     self.crop_x_end_slider.setEnabled(False)
                     self.crop_y_start_slider.setEnabled(False)
                     self.crop_y_end_slider.setEnabled(False)
                else:
                    # Store valid image coordinates
                    self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))

                    # Calculate percentages from valid image coordinates
                    x_start_perc = (img_x / img_w) * 100.0 if img_w > 0 else 0
                    y_start_perc = (img_y / img_h) * 100.0 if img_h > 0 else 0
                    x_end_perc = ((img_x + img_crop_w) / img_w) * 100.0 if img_w > 0 else 100
                    y_end_perc = ((img_y + img_crop_h) / img_h) * 100.0 if img_h > 0 else 100

                    # Convert percentages to integer slider values
                    x_start_val = int(round(x_start_perc * self.crop_slider_precision_factor))
                    y_start_val = int(round(y_start_perc * self.crop_slider_precision_factor))
                    x_end_val = int(round(x_end_perc * self.crop_slider_precision_factor))
                    y_end_val = int(round(y_end_perc * self.crop_slider_precision_factor))

                    # --- Update Sliders (Block signals) ---
                    self.crop_x_start_slider.blockSignals(True); self.crop_x_start_slider.setValue(x_start_val); self.crop_x_start_slider.blockSignals(False)
                    self.crop_x_end_slider.blockSignals(True); self.crop_x_end_slider.setValue(x_end_val); self.crop_x_end_slider.blockSignals(False)
                    self.crop_y_start_slider.blockSignals(True); self.crop_y_start_slider.setValue(y_start_val); self.crop_y_start_slider.blockSignals(False)
                    self.crop_y_end_slider.blockSignals(True); self.crop_y_end_slider.setValue(y_end_val); self.crop_y_end_slider.blockSignals(False)

                    # --- Explicitly Update Labels AFTER setting sliders ---
                    self.crop_x_start_value_label.setText(f"{x_start_perc:.2f}%")
                    self.crop_x_end_value_label.setText(f"{x_end_perc:.2f}%")
                    self.crop_y_start_value_label.setText(f"{y_start_perc:.2f}%")
                    self.crop_y_end_value_label.setText(f"{y_end_perc:.2f}%")
                    # --- End Explicit Label Update ---

                    # --- Enable the sliders ---
                    self.crop_x_start_slider.setEnabled(True)
                    self.crop_x_end_slider.setEnabled(True)
                    self.crop_y_start_slider.setEnabled(True)
                    self.crop_y_end_slider.setEnabled(True)

            except Exception as e:
                 # Handle errors during coordinate calculation or UI update
                 QMessageBox.warning(self, "Coordinate Error", f"Could not calculate/update crop: {e}")
                 traceback.print_exc()
                 self.crop_rectangle_coords = None
                 self.live_view_label.clear_crop_preview()
                 # Ensure sliders remain disabled on error
                 self.crop_x_start_slider.setEnabled(False)
                 self.crop_x_end_slider.setEnabled(False)
                 self.crop_y_start_slider.setEnabled(False)
                 self.crop_y_end_slider.setEnabled(False)

            # Reset the temporary start point used for drawing
            self.crop_rect_start_view = None
    
    def create_white_space_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # --- Padding Group ---
        padding_params_group = QGroupBox("Add White Space (Padding)")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setSpacing(8)


        # Input fields with validation (optional but good)
        from PyQt5.QtGui import QIntValidator
        int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

        # Left Padding
        left_padding_label = QLabel("Left Padding (px):")
        self.left_padding_input = QLineEdit("100") # Default
        self.left_padding_input.setValidator(int_validator)
        self.left_padding_input.setToolTip("Pixels to add to the left.")
        padding_layout.addWidget(left_padding_label, 0, 0)
        padding_layout.addWidget(self.left_padding_input, 0, 1)

        # Right Padding
        right_padding_label = QLabel("Right Padding (px):")
        self.right_padding_input = QLineEdit("100") # Default
        self.right_padding_input.setValidator(int_validator)
        self.right_padding_input.setToolTip("Pixels to add to the right.")
        padding_layout.addWidget(right_padding_label, 0, 2)
        padding_layout.addWidget(self.right_padding_input, 0, 3)

        # Top Padding
        top_padding_label = QLabel("Top Padding (px):")
        self.top_padding_input = QLineEdit("100") # Default
        self.top_padding_input.setValidator(int_validator)
        self.top_padding_input.setToolTip("Pixels to add to the top.")
        padding_layout.addWidget(top_padding_label, 1, 0)
        padding_layout.addWidget(self.top_padding_input, 1, 1)

        # Bottom Padding
        bottom_padding_label = QLabel("Bottom Padding (px):")
        self.bottom_padding_input = QLineEdit("0") # Default
        self.bottom_padding_input.setValidator(int_validator)
        self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
        padding_layout.addWidget(bottom_padding_label, 1, 2)
        padding_layout.addWidget(self.bottom_padding_input, 1, 3)

        layout.addWidget(padding_params_group)

        # --- Buttons Layout ---
        button_layout = QHBoxLayout()
        self.recommend_button = QPushButton("Set Recommended Values")
        self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
        self.recommend_button.clicked.connect(self.recommended_values)

        self.clear_padding_button = QPushButton("Clear Values")
        self.clear_padding_button.setToolTip("Set all padding values to 0.")
        self.clear_padding_button.clicked.connect(self.clear_padding_values)

        self.finalize_button = QPushButton("Apply Padding")
        self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
        self.finalize_button.clicked.connect(self.finalize_image)

        button_layout.addWidget(self.recommend_button)
        button_layout.addWidget(self.clear_padding_button)
        button_layout.addStretch()
        button_layout.addWidget(self.finalize_button)

        layout.addLayout(button_layout)
        layout.addStretch() # Push content up

        return tab
    def clear_padding_values(self):
        self.bottom_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.left_padding_input.setText("0")
        
    def recommended_values(self):
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        
        
        
    def create_markers_tab(self):
        """Create the Markers tab with preset, label, placement, offset, and custom controls."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)  # Reduce main vertical spacing

        # --- Combined Group Box for Presets and Labels ---
        presets_labels_group = QGroupBox("Marker Presets and Labels")
        presets_labels_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        presets_labels_layout = QGridLayout(presets_labels_group)
        presets_labels_layout.setSpacing(5) # Reduce spacing within the grid

        # -- Left/Right Marker Options (Columns 0 & 1) --
        presets_labels_layout.addWidget(QLabel("Preset:"), 0, 0) # Label for combo
        self.combo_box = QComboBox(self)
        if not hasattr(self, 'marker_values_dict'):
             self.load_config() # Load config attempts to initialize it
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained") # Default or load last used?
        self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
        presets_labels_layout.addWidget(self.combo_box, 0, 1) # Combo box in col 1

        self.marker_values_textbox = QLineEdit(self)
        self.marker_values_textbox.setPlaceholderText("Custom values (comma-separated)") # Shorter text
        self.marker_values_textbox.setEnabled(False)
        presets_labels_layout.addWidget(self.marker_values_textbox, 1, 0, 1, 2) # Span text box across cols 0 & 1

        self.rename_input = QLineEdit(self)
        self.rename_input.setPlaceholderText("New name for Custom")
        self.rename_input.setEnabled(False)
        presets_labels_layout.addWidget(self.rename_input, 2, 0, 1, 2) # Span rename input

        preset_buttons_layout = QHBoxLayout()
        self.save_button = QPushButton("Save Config", self)
        self.save_button.clicked.connect(self.save_config)
        self.remove_config_button = QPushButton("Remove Config", self)
        self.remove_config_button.clicked.connect(self.remove_config)
        preset_buttons_layout.addWidget(self.save_button)
        preset_buttons_layout.addWidget(self.remove_config_button)
        preset_buttons_layout.addStretch() # Push buttons left
        presets_labels_layout.addLayout(preset_buttons_layout, 3, 0, 1, 2) # Add button layout, spanning

        # -- Top Marker Options (Columns 2 & 3) --
        presets_labels_layout.addWidget(QLabel("Top Labels:"), 0, 2) # Label for text edit
        self.top_marker_input = QTextEdit(self)
        if not hasattr(self, 'top_label'):
            self.top_label = ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"] # Default
        self.top_marker_input.setText(", ".join(map(str, self.top_label))) # Ensure all elements are strings for join
        self.top_marker_input.setMinimumHeight(40)
        self.top_marker_input.setMaximumHeight(80) # Reduced max height
        self.top_marker_input.setPlaceholderText("Labels (comma-separated)") # Shorter text
        presets_labels_layout.addWidget(self.top_marker_input, 0, 3, 3, 1) # Row 0, Col 3, Span 3 rows, 1 col

        self.update_top_labels_button = QPushButton("Update All Labels")
        self.update_top_labels_button.clicked.connect(self.update_all_labels)
        presets_labels_layout.addWidget(self.update_top_labels_button, 3, 3) # Place below text edit in col 3

        presets_labels_layout.setColumnStretch(1, 1)
        presets_labels_layout.setColumnStretch(3, 1)

        layout.addWidget(presets_labels_group)

        # --- Marker Placement and Offsets Group ---
        padding_params_group = QGroupBox("Marker Placement and Offsets")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setVerticalSpacing(5)
        padding_layout.setHorizontalSpacing(8)

        # Initialize slider ranges if not already done (e.g., by loading config)
        if not hasattr(self, 'left_slider_range'): self.left_slider_range = [-100, 1000]
        if not hasattr(self, 'right_slider_range'): self.right_slider_range = [-100, 1000]
        if not hasattr(self, 'top_slider_range'): self.top_slider_range = [-100, 1000]
        # Initialize shift values if not already done
        if not hasattr(self, 'left_marker_shift_added'): self.left_marker_shift_added = 0
        if not hasattr(self, 'right_marker_shift_added'): self.right_marker_shift_added = 0
        if not hasattr(self, 'top_marker_shift_added'): self.top_marker_shift_added = 0

        # --- Row 0: Left Marker ---
        left_marker_button = QPushButton("Place Left")
        left_marker_button.setToolTip("Place left markers. Shortcut: Ctrl+Shift+L")
        left_marker_button.clicked.connect(self.enable_left_marker_mode)
        self.left_padding_slider = QSlider(Qt.Horizontal)
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.left_padding_slider.setValue(self.left_marker_shift_added) # Use initialized/loaded value
        self.left_padding_slider.valueChanged.connect(self.update_left_padding)
        remove_left_button = QPushButton("Remove Last")
        remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
        reset_left_button = QPushButton("Reset")
        reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
        duplicate_left_button = QPushButton("Copy Right")
        duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
        padding_layout.addWidget(QLabel("Offset Left:"), 0, 3)
        padding_layout.addWidget(left_marker_button, 0, 0)
        padding_layout.addWidget(remove_left_button, 0, 1)
        padding_layout.addWidget(reset_left_button, 0, 2)
        padding_layout.addWidget(self.left_padding_slider, 0, 4, 1, 3) # Span 3 columns
        padding_layout.addWidget(duplicate_left_button, 0, 7)

        # --- Row 1: Right Marker ---
        right_marker_button = QPushButton("Place Right")
        right_marker_button.setToolTip("Place right markers. Shortcut: Ctrl+Shift+R")
        right_marker_button.clicked.connect(self.enable_right_marker_mode)
        self.right_padding_slider = QSlider(Qt.Horizontal)
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.right_padding_slider.setValue(self.right_marker_shift_added)
        self.right_padding_slider.valueChanged.connect(self.update_right_padding)
        remove_right_button = QPushButton("Remove Last")
        remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
        reset_right_button = QPushButton("Reset")
        reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
        duplicate_right_button = QPushButton("Copy Left")
        duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
        padding_layout.addWidget(QLabel("Offset Right:"), 1, 3)
        padding_layout.addWidget(right_marker_button, 1, 0)
        padding_layout.addWidget(remove_right_button, 1, 1)
        padding_layout.addWidget(reset_right_button, 1, 2)
        padding_layout.addWidget(self.right_padding_slider, 1, 4, 1, 3) # Span 3 columns
        padding_layout.addWidget(duplicate_right_button, 1, 7)

        # --- Row 2: Top Marker ---
        top_marker_button = QPushButton("Place Top")
        top_marker_button.setToolTip("Place top markers. Shortcut: Ctrl+Shift+T")
        top_marker_button.clicked.connect(self.enable_top_marker_mode)
        self.top_padding_slider = QSlider(Qt.Horizontal)
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.top_padding_slider.setValue(self.top_marker_shift_added)
        self.top_padding_slider.valueChanged.connect(self.update_top_padding)
        remove_top_button = QPushButton("Remove Last")
        remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
        reset_top_button = QPushButton("Reset")
        reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
        padding_layout.addWidget(QLabel("Offset Top:"), 2, 3)
        padding_layout.addWidget(top_marker_button, 2, 0)
        padding_layout.addWidget(remove_top_button, 2, 1)
        padding_layout.addWidget(reset_top_button, 2, 2)
        padding_layout.addWidget(self.top_padding_slider, 2, 4, 1, 4)  # Span 4 columns (takes space of duplicate btn too)

        # --- Row 3: Custom Marker Main Controls ---
        self.custom_marker_button = QPushButton("Place Custom", self)
        self.custom_marker_button.setToolTip("Click to activate, then click on image to place the text/arrow")
        self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
        self.custom_marker_text_entry = QLineEdit(self)
        self.custom_marker_text_entry.setPlaceholderText("Custom text")

        marker_buttons_layout = QHBoxLayout()
        marker_buttons_layout.setContentsMargins(0, 0, 0, 0)
        marker_buttons_layout.setSpacing(2)
        self.custom_marker_button_left_arrow = QPushButton("←", self)
        self.custom_marker_button_right_arrow = QPushButton("→", self)
        self.custom_marker_button_top_arrow = QPushButton("↑", self)
        self.custom_marker_button_bottom_arrow = QPushButton("↓", self)
        arrow_size = 25
        self.custom_marker_button_left_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_right_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_top_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_bottom_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_left_arrow.setToolTip("Ctrl+Left: Add ← and activate placement")
        self.custom_marker_button_right_arrow.setToolTip("Ctrl+Right: Add → and activate placement")
        self.custom_marker_button_top_arrow.setToolTip("Ctrl+Up: Add ↑ and activate placement")
        self.custom_marker_button_bottom_arrow.setToolTip("Ctrl+Down: Add ↓ and activate placement")
        marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
        marker_buttons_layout.addStretch()
        self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))

        self.remove_custom_marker_button = QPushButton("Remove Last", self)
        self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
        self.reset_custom_marker_button = QPushButton("Reset", self)
        self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
        self.modify_custom_marker_button = QPushButton("Modify", self) # << THE NEW BUTTON
        self.modify_custom_marker_button.setToolTip("Open a dialog to manage custom markers")
        self.modify_custom_marker_button.clicked.connect(self.open_modify_markers_dialog) # << CONNECTED
        self.custom_marker_color_button = QPushButton("Color")
        self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
        if not hasattr(self, 'custom_marker_color'): self.custom_marker_color = QColor(0,0,0)
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)

        padding_layout.addWidget(self.custom_marker_button, 3, 0)
        padding_layout.addWidget(self.custom_marker_text_entry, 3, 1, 1, 2)
        padding_layout.addLayout(marker_buttons_layout, 3, 3)
        padding_layout.addWidget(self.remove_custom_marker_button, 3, 4)
        padding_layout.addWidget(self.reset_custom_marker_button, 3, 5)
        padding_layout.addWidget(self.modify_custom_marker_button, 3, 6) # Added Modify Button
        padding_layout.addWidget(self.custom_marker_color_button, 3, 7)   # Color Button shifted

        # --- Row 4: Custom Marker Font and Grid ---
        self.custom_font_type_label = QLabel("Custom Font:", self)
        self.custom_font_type_dropdown = QFontComboBox()
        # Initialize font: Check if 'Arial' exists, otherwise use first available
        initial_font = QFont("Arial")
        # if initial_font.family() not in [f.family() for f in self.custom_font_type_dropdown.availableFonts()]:
        #      initial_font = self.custom_font_type_dropdown.availableFonts()[0] if self.custom_font_type_dropdown.availableFonts() else QFont()
        self.custom_font_type_dropdown.setCurrentFont(initial_font)
        self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)

        self.custom_font_size_label = QLabel("Size:", self)
        self.custom_font_size_spinbox = QSpinBox(self)
        self.custom_font_size_spinbox.setRange(1, 150)
        self.custom_font_size_spinbox.setValue(12) # Default

        # self.show_grid_checkbox = QCheckBox("Snap Grid", self)
        # self.show_grid_checkbox.setToolTip("Places a snapping grid. Shortcut: Ctrl+Shift+G")
        # self.show_grid_checkbox.setChecked(False)
        # self.show_grid_checkbox.stateChanged.connect(self.update_live_view)
        
        self.show_grid_checkbox_x = QCheckBox("Snap X", self)
        self.show_grid_checkbox_x.setToolTip("Snaps marker placement horizontally to the grid. Ctrl+Shift+G toggles X & Y.")
        self.show_grid_checkbox_x.setChecked(False)
        self.show_grid_checkbox_x.stateChanged.connect(self.update_live_view) # Trigger redraw if grid lines depend on this

        self.show_grid_checkbox_y = QCheckBox("Snap Y", self)
        self.show_grid_checkbox_y.setToolTip("Snaps marker placement vertically to the grid. Ctrl+Shift+G toggles X & Y.")
        self.show_grid_checkbox_y.setChecked(False)
        self.show_grid_checkbox_y.stateChanged.connect(self.update_live_view) # Trigger redraw if grid lines depend on this

        self.grid_size_input = QSpinBox(self)
        self.grid_size_input.setRange(5, 100)
        self.grid_size_input.setValue(20)
        self.grid_size_input.setPrefix("Grid (px): ")
        self.grid_size_input.valueChanged.connect(self.update_live_view)
        
        shape_button_layout = QHBoxLayout()
        shape_button_layout.setContentsMargins(0,0,0,0)
        shape_button_layout.setSpacing(2)

        self.draw_line_button = QPushButton("L")
        self.draw_line_button.setToolTip("Draw Line: Click and drag on the image.\nUses current Custom Color and Size/Thickness.")
        self.draw_line_button.setFixedSize(25, 25) # Small square button
        self.draw_line_button.clicked.connect(self.enable_line_drawing_mode)

        self.draw_rect_button = QPushButton("R")
        self.draw_rect_button.setToolTip("Draw Rectangle: Click and drag on the image.\nUses current Custom Color and Size/Thickness.")
        self.draw_rect_button.setFixedSize(25, 25) # Small square button
        self.draw_rect_button.clicked.connect(self.enable_rectangle_drawing_mode)
        
        self.remove_shape_button = QPushButton("X")
        self.remove_shape_button.setToolTip("Remove Last Drawn Shape (Line/Rectangle)")
        self.remove_shape_button.setFixedSize(25, 25)
        self.remove_shape_button.clicked.connect(self.remove_last_custom_shape) # Connect to specific method
        
        shape_button_layout.addWidget(self.draw_line_button)
        shape_button_layout.addWidget(self.draw_rect_button)
        shape_button_layout.addWidget(self.remove_shape_button)
        shape_button_layout.addStretch()

        padding_layout.addWidget(self.custom_font_type_label, 4, 0)
        padding_layout.addWidget(self.custom_font_type_dropdown, 4, 1)
        padding_layout.addWidget(self.custom_font_size_label, 4, 2)
        padding_layout.addWidget(self.custom_font_size_spinbox, 4, 3)
        padding_layout.addWidget(self.show_grid_checkbox_x, 4, 4)
        padding_layout.addWidget(self.show_grid_checkbox_y, 4, 5)
        padding_layout.addWidget(self.grid_size_input, 4, 6, 1, 1) # Span 3 columns
        padding_layout.addLayout(shape_button_layout, 4, 7)

        # Set column stretches - Now 8 columns (0-7)
        padding_layout.setColumnStretch(0, 1) # Place buttons
        padding_layout.setColumnStretch(1, 1) # Remove/Custom Text/Font Combo
        padding_layout.setColumnStretch(2, 1) # Reset/Custom Text/Font Size Label
        padding_layout.setColumnStretch(3, 1) # Offset Label/Arrows/Font Size Spin
        padding_layout.setColumnStretch(4, 2) # Slider/Remove/Grid Check
        padding_layout.setColumnStretch(5, 1) # Slider/Reset/Grid Size
        padding_layout.setColumnStretch(6, 1) # Slider/Modify/Grid Size
        padding_layout.setColumnStretch(7, 1) # Duplicate/Color/Grid Size

        layout.addWidget(padding_params_group)
        layout.addStretch()

        return tab

    def open_modify_markers_dialog(self):
        """Opens the dialog to modify custom markers AND shapes."""

        # Ensure the lists exist and are lists, initialize if needed
        if not hasattr(self, "custom_markers") or not isinstance(self.custom_markers, list):
            self.custom_markers = []
        if not hasattr(self, "custom_shapes") or not isinstance(self.custom_shapes, list):
            self.custom_shapes = []

        # Check if there are *any* items to modify
        if not self.custom_markers and not self.custom_shapes:
            QMessageBox.information(self, "No Items", "There are no custom markers or shapes to modify.")
            return

        # --- CORRECTED CALL: Pass self.custom_shapes as the second argument ---
        dialog = ModifyMarkersDialog(
            list(self.custom_markers),  # Pass a copy of markers list
            list(self.custom_shapes),   # Pass a copy of shapes list <<-- FIX HERE
            self                       # Parent is still self
        )
        # --- End Correction ---

        if dialog.exec_() == QDialog.Accepted:
            # Retrieve both modified lists
            modified_markers, modified_shapes = dialog.get_modified_markers_and_shapes()

            # Check if either list has changed before saving state and updating
            markers_changed = (modified_markers != [tuple(m) for m in self.custom_markers]) # Compare with tuples if original was tuples
            shapes_changed = (modified_shapes != self.custom_shapes) # Direct dict list comparison should work

            if markers_changed or shapes_changed:
                 self.save_state() # Save previous state if changes were made
                 # Update the main application's lists
                 self.custom_markers = [list(m) for m in modified_markers] # Convert back to list of lists if needed
                 self.custom_shapes = modified_shapes
                 self.is_modified = True # Mark as modified
                 self.update_live_view() # Refresh display
    
    def add_column(self):
        """Add a new column to the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        if current_text.strip():
            self.top_marker_input.append("")  # Add a new line for a new column
        else:
            self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
    
    def remove_column(self):
        """Remove the last column from the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        lines = current_text.split("\n")
        if len(lines) > 1:
            lines.pop()  # Remove the last line
            self.top_marker_input.setPlainText("\n".join(lines))
        else:
            self.top_marker_input.clear()  # Clear the text if only one line exists

    
    def flip_vertical(self):
        self.save_state()
        """Flip the image vertically."""
        if self.image:
            self.image = self.image.mirrored(vertical=True, horizontal=False)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def flip_horizontal(self):
        self.save_state()
        """Flip the image horizontally."""
        if self.image:
            self.image = self.image.mirrored(vertical=False, horizontal=True)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def convert_to_black_and_white(self):
        """
        Converts the current self.image to grayscale, preserving alpha channel
        if present by creating an ARGB image where R=G=B=GrayValue.
        Aims for 16-bit grayscale precision when converting from color.
        """
        self.save_state()
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Error", "No image loaded.")
            return

        original_format = self.image.format()
        original_has_alpha = self.image.hasAlphaChannel() # Check QImage's report

        # Check if already effectively grayscale (but maybe not the specific format)
        if original_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16, QImage.Format_Mono]:
            # It's already a standard grayscale format (without explicit alpha in format enum)
            QMessageBox.information(self, "Info", f"Image is already grayscale (Format: {original_format}).")
            return

        converted_image = None
        try:
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")

            print(f"Original NumPy shape: {np_img.shape}, dtype: {np_img.dtype}") # Debug

            target_dtype = np.uint16 # Prefer 16-bit precision for grayscale conversion
            target_max_val = 65535.0

            if np_img.ndim == 3 and np_img.shape[2] == 4: # Color with Alpha (BGRA or RGBA)
                print("Processing image with Alpha channel.")
                # Assume input order from qimage_to_numpy is BGRA for ARGB32/RGB32
                # or RGBA for RGBA8888. cvtColor expects BGR.
                # Let's explicitly convert the color part assuming BGR input slice
                color_part = np_img[..., :3] # Slice BGR or RGB
                alpha_part = np_img[..., 3]  # Extract alpha channel

                # Convert color to grayscale using cv2 (expects BGR)
                # If the input was RGBA, this might be slightly off, but often acceptable.
                # A more robust way might check the original_format if needed.
                gray_np_8bit = cv2.cvtColor(color_part, cv2.COLOR_BGR2GRAY)

                # Scale grayscale to target bit depth (16-bit)
                gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)

                # Create a new 4-channel array (BGRA format for QImage.Format_ARGB32)
                # Replicate grayscale channel for B, G, R
                bgra_target = np.zeros((np_img.shape[0], np_img.shape[1], 4), dtype=target_dtype)
                bgra_target[..., 0] = gray_np_target # Blue = Gray
                bgra_target[..., 1] = gray_np_target # Green = Gray
                bgra_target[..., 2] = gray_np_target # Red = Gray

                # Handle Alpha channel: ensure it's the correct dtype for combining
                if alpha_part.dtype != target_dtype:
                     # Scale alpha if necessary (e.g., if source was 16-bit RGBA)
                     if alpha_part.dtype == np.uint16:
                         alpha_scaled = (alpha_part / 65535.0 * target_max_val).astype(target_dtype)
                     elif alpha_part.dtype == np.uint8:
                         alpha_scaled = (alpha_part / 255.0 * target_max_val).astype(target_dtype)
                     else: # Fallback: assume it doesn't need scaling or use 8-bit alpha
                          print(f"Warning: Unexpected alpha dtype {alpha_part.dtype}. Attempting direct use or scaling.")
                          try: # Try scaling assuming it's 0-max range
                              alpha_scaled = (alpha_part / np.max(alpha_part) * target_max_val).astype(target_dtype) if np.max(alpha_part) > 0 else np.zeros_like(alpha_part, dtype=target_dtype)
                          except: # Final fallback
                               alpha_scaled = (alpha_part.astype(np.float64) / 255.0 * target_max_val).astype(target_dtype) if np.max(alpha_part)>0 else np.zeros_like(alpha_part,dtype=target_dtype) #Assume 8 bit if failsafe
                     bgra_target[..., 3] = alpha_scaled
                else:
                    bgra_target[..., 3] = alpha_part # Alpha dtype already matches

                # Convert back to QImage - should become Format_ARGB32 due to 4 channels
                # Note: numpy_to_qimage needs to handle 16-bit 4-channel input if we want ARGB64
                # Currently, it scales down 16-bit color to 8-bit ARGB32. Let's adapt that.
                # --- MODIFICATION NEEDED in numpy_to_qimage for 16-bit ARGB ---
                # For now, let's convert bgra_target to uint8 for Format_ARGB32 output
                bgra_target_uint8 = (bgra_target / target_max_val * 255.0).astype(np.uint8)
                converted_image = self.numpy_to_qimage(bgra_target_uint8)
                if converted_image.isNull(): raise ValueError("Conversion of BGRA Grayscale to QImage failed.")
                print(f"Converted to Grayscale + Alpha. QImage format: {converted_image.format()}") # Debug


            elif np_img.ndim == 3 and np_img.shape[2] == 3: # Color without Alpha (BGR or RGB)
                print("Processing image without Alpha channel.")
                # Convert color to grayscale (standard procedure)
                gray_np_8bit = cv2.cvtColor(np_img[..., :3], cv2.COLOR_BGR2GRAY)
                 # Scale grayscale to target bit depth (16-bit)
                gray_np_target = (gray_np_8bit / 255.0 * target_max_val).astype(target_dtype)
                # Convert back to standard grayscale QImage
                converted_image = self.numpy_to_qimage(gray_np_target) # Should yield Format_Grayscale16
                if converted_image.isNull(): raise ValueError("Conversion of BGR Grayscale to QImage failed.")
                print(f"Converted to Grayscale. QImage format: {converted_image.format()}") # Debug


            elif np_img.ndim == 2: # Already grayscale (but QImage didn't report standard format)
                 # This case might occur for unusual grayscale formats QImage loads but doesn't map to standard enums.
                 QMessageBox.information(self, "Info", "Image is already grayscale (based on channel analysis).")
                 # Optionally convert to a standard format like Grayscale16
                 if np_img.dtype != target_dtype:
                     gray_np_target = (np_img.astype(np.float64) / np.max(np_img) * target_max_val).astype(target_dtype) if np.max(np_img)>0 else np.zeros_like(np_img,dtype=target_dtype)
                     converted_image = self.numpy_to_qimage(gray_np_target)
                 else:
                     converted_image = self.image.copy() # No conversion needed
            else:
                raise ValueError(f"Unsupported NumPy array dimension: {np_img.ndim}")


        except Exception as e:
             QMessageBox.critical(self, "Conversion Error", f"Could not convert image to grayscale: {e}")
             traceback.print_exc()
             return # Keep original image

        if converted_image and not converted_image.isNull():
             self.image = converted_image
             # Update backups consistently
             self.image_before_contrast = self.image.copy()
             self.image_contrasted = self.image.copy()
             # Reset padding state if format changes implicitly? Let's assume padding needs re-application.
             if self.image_padded:
                 self.image_before_padding = None # Invalidate padding backup
                 self.image_padded = False
             else:
                 self.image_before_padding = self.image.copy() if self.image else None # Ensure image exists before copying

             # Reset contrast/gamma sliders as appearance changed significantly
             self.reset_gamma_contrast() # Resets sliders and updates view
             self._update_status_bar()
             self.update_live_view() # Ensure view updates even if reset_gamma_contrast fails
        else:
             # This case should be less likely now with better error handling
             QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


    def invert_image(self):
        self.save_state()
        if self.image:
            inverted_image = self.image.copy()
            inverted_image.invertPixels()
            self.image = inverted_image
            self.update_live_view()
        self.image_before_contrast=self.image.copy()
        self.image_before_padding=self.image.copy()
        self.image_contrasted=self.image.copy()

    def keyPressEvent(self, event):
        # --- Escape Key Handling ---
        if event.key() == Qt.Key_Escape:
            # 1. PRIORITIZE CANCELLING CROP MODE if active
            if self.crop_rectangle_mode:
                # print("Debug: Escape pressed, cancelling crop mode.") # Optional Debug
                self.cancel_rectangle_crop_mode()
                # Optionally, clear the visual preview immediately on Escape
                self.live_view_label.clear_crop_preview()
                self.update_live_view() # Refresh view after clearing preview
                return # Consume the event, don't process further Escape logic

            # 2. Cancel Line/Rectangle Drawing Mode
            if self.drawing_mode in ['line', 'rectangle']:
                # print("Debug: Escape pressed, cancelling shape drawing.") # Optional Debug
                self.cancel_drawing_mode()
                self.update_live_view()
                return # Consume the event

            # 3. Cancel Marker Preview/Placement Modes
            if self.live_view_label.preview_marker_enabled:
                # print("Debug: Escape pressed, cancelling marker preview.") # Optional Debug
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                self.update_live_view() # Update to clear preview
                return # Consume the event

            if self.marker_mode is not None: # If any marker placement mode was active
                # print(f"Debug: Escape pressed, cancelling marker mode: {self.marker_mode}") # Optional Debug
                self.marker_mode = None
                self.live_view_label.mousePressEvent = None # Reset handler
                self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
                self.update_live_view()
                return # Consume the event

            # 4. Cancel Analysis Quadrilateral/Rectangle Definition/Move
            if self.live_view_label.measure_quantity_mode or self.live_view_label.mode in ["quad", "rectangle", "move"]:
                # print("Debug: Escape pressed, cancelling analysis area mode.") # Optional Debug
                self.live_view_label.measure_quantity_mode = False
                self.live_view_label.bounding_box_complete = False
                self.live_view_label.counter = 0
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_points = []
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self.live_view_label.selected_point = -1
                self.live_view_label.mousePressEvent = None
                self.live_view_label.mouseMoveEvent = None
                self.live_view_label.mouseReleaseEvent = None
                self.live_view_label.mode = None
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.update_live_view()
                return # Consume the event

            # If none of the above modes were active, let the base class handle it (if necessary)
            # super().keyPressEvent(event) # Or just pass if no specific base class action needed

        # --- Panning with Arrow Keys (Keep as is) ---
        if self.live_view_label.zoom_level != 1.0:
            step = 20
            offset_changed = False # Flag to update view only if changed
            current_x = self.live_view_label.pan_offset.x()
            current_y = self.live_view_label.pan_offset.y()

            if event.key() == Qt.Key_Left:
                self.live_view_label.pan_offset.setX(current_x - step)
                offset_changed = True
            elif event.key() == Qt.Key_Right:
                self.live_view_label.pan_offset.setX(current_x + step)
                offset_changed = True
            elif event.key() == Qt.Key_Up:
                self.live_view_label.pan_offset.setY(current_y - step)
                offset_changed = True
            elif event.key() == Qt.Key_Down:
                self.live_view_label.pan_offset.setY(current_y + step)
                offset_changed = True

            if offset_changed:
                self.update_live_view()
                return # Consume arrow keys if panning occurred

        # Allow other key presses to be handled by the base class or shortcuts
        super().keyPressEvent(event)
    
    def update_marker_text_font(self, font: QFont):
        """
        Updates the font of self.custom_marker_text_entry based on the selected font
        from self.custom_font_type_dropdown.
    
        :param font: QFont object representing the selected font from the combobox.
        """
        # Get the name of the selected font
        selected_font_name = font.family()
        
        # Set the font of the QLineEdit
        self.custom_marker_text_entry.setFont(QFont(selected_font_name))
    
    def arrow_marker(self, text: str):
        self.custom_marker_text_entry.clear()
        self.custom_marker_text_entry.setText(text)
    
    def enable_custom_marker_mode(self):
        """Enable the custom marker mode and set the mouse event."""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        custom_text = self.custom_marker_text_entry.text().strip()
    
        self.live_view_label.preview_marker_enabled = True
        self.live_view_label.preview_marker_text = custom_text
        self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
        self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
        self.live_view_label.marker_color=self.custom_marker_color
        
        self.live_view_label.setFocus()
        self.live_view_label.update()
        
        self.marker_mode = "custom"  # Indicate custom marker mode
        self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
        
    def remove_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers.pop()  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    def remove_last_custom_shape(self):
        """Removes the last added custom SHAPE (line or rectangle)."""
        if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
            self.save_state() # Save state before removal for undo
            removed_shape = self.custom_shapes.pop()  # Remove the last shape
            print(f"Removed last custom shape: {removed_shape}") # Optional debug
            self.is_modified = True
            self.update_live_view()  # Update the display
            # Optionally, provide user feedback via status bar or QMessageBox
            # self.statusBar().showMessage("Last custom shape removed.", 2000)
        else:
            print("No custom shapes to remove.") # Info message
            # Optionally, provide user feedback
            # self.statusBar().showMessage("No custom shapes to remove.", 2000)
        
    def reset_custom_marker_mode(self):
        """Resets (clears) all custom MARKERS and custom SHAPES."""
        markers_cleared = False
        shapes_cleared = False
        needs_save = False # Flag to check if state needs saving

        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            if not needs_save: # Save state only once before the first clear operation
                 self.save_state()
                 needs_save = True
            self.custom_markers.clear()
            markers_cleared = True
            print("Cleared all custom markers.")

        if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
            if not needs_save: # Save state only once if markers weren't cleared but shapes are
                 self.save_state()
                 needs_save = True
            self.custom_shapes.clear()
            shapes_cleared = True
            print("Cleared all custom shapes.")

        if markers_cleared or shapes_cleared:
            self.is_modified = True
            self.update_live_view()
        else:
            print("No custom markers or shapes to reset.")
        
    
    def place_custom_marker(self, event, custom_text):
        """Place a custom marker at the cursor location."""
        self.save_state()
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            
        # Adjust for snapping to grid
        # if self.show_grid_checkbox.isChecked():
        #     grid_size = self.grid_size_input.value()
        #     cursor_x = round(cursor_x / grid_size) * grid_size
        #     cursor_y = round(cursor_y / grid_size) * grid_size
        grid_size = self.grid_size_input.value() # Get grid size once

        if self.show_grid_checkbox_x.isChecked():
            if grid_size > 0: # Avoid division by zero
                 cursor_x = round(cursor_x / grid_size) * grid_size

        if self.show_grid_checkbox_y.isChecked():
            if grid_size > 0: # Avoid division by zero
                 cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        image_x = (cursor_x - x_offset) / scale
        image_y = (cursor_y - y_offset) / scale
        
            
        # Store the custom marker's position and text
        self.custom_markers = getattr(self, "custom_markers", [])
        self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value(),False,False))
        self.update_live_view()
        
    def select_custom_marker_color(self):
        """Open a color picker dialog to select the color for custom markers."""
        color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
        if color.isValid():
            self.custom_marker_color = color  # Update the custom marker color
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
    
    def update_all_labels(self):
        """Update all labels from the QTextEdit, supporting multiple columns."""
        self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
        input_text = self.top_marker_input.toPlainText()
        
        # Split the text into a list by commas and strip whitespace
        self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
        

        # if len(self.top_label) < len(self.top_markers):
        #     self.top_markers = self.top_markers[:len(self.top_label)]
        for i in range(0, len(self.top_markers)):
            try:
                self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
            except:
                self.top_markers[i] = (self.top_markers[i][0], str(""))

        # if len(self.marker_values) < len(self.left_markers):
        #     self.left_markers = self.left_markers[:len(self.marker_values)]
        for i in range(0, len(self.left_markers)):
            try:
                self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
            except:
                self.left_markers[i] = (self.left_markers[i][0], str(""))
                

        # if len(self.marker_values) < len(self.right_markers):
        #     self.right_markers = self.right_markers[:len(self.marker_values)]
        for i in range(0, len(self.right_markers)):
            try:
                self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
            except:
                self.right_markers[i] = (self.right_markers[i][0], str(""))
                

        
        # Trigger a refresh of the live view
        self.update_live_view()
        
    def reset_marker(self, marker_type, param):    
        self.save_state()
        if marker_type == 'left':
            if param == 'remove' and len(self.left_markers)!=0:
                self.left_markers.pop()  
                if self.current_left_marker_index > 0:
                    self.current_left_marker_index -= 1
            elif param == 'reset':
                self.left_markers.clear()
                self.current_left_marker_index = 0  

             
        elif marker_type == 'right' and len(self.right_markers)!=0:
            if param == 'remove':
                self.right_markers.pop()  
                if self.current_right_marker_index > 0:
                    self.current_right_marker_index -= 1
            elif param == 'reset':
                self.right_markers.clear()
                self.current_right_marker_index = 0

        elif marker_type == 'top' and len(self.top_markers)!=0:
            if param == 'remove':
                self.top_markers.pop() 
                if self.current_top_label_index > 0:
                    self.current_top_label_index -= 1
            elif param == 'reset':
                self.top_markers.clear()
                self.current_top_label_index = 0
    
        # Call update live view after resetting markers
        self.update_live_view()
        
    def duplicate_marker(self, marker_type):
        self.save_state() # Save state before making changes

        if marker_type == 'left' and self.right_markers:
            # Copy Right Markers TO Left
            self.left_markers = self.right_markers.copy()
            # Copy Right Offset value TO Left Offset variable
            self.left_marker_shift_added = self.right_marker_shift_added

            # Update Left Slider position to match the copied offset
            if hasattr(self, 'left_padding_slider'):
                min_val, max_val = self.left_padding_slider.minimum(), self.left_padding_slider.maximum()
                # Clamp the value to the slider's current range
                clamped_value = max(min_val, min(self.left_marker_shift_added, max_val))
                # Update slider visually, preventing signal emission temporarily
                self.left_padding_slider.blockSignals(True)
                self.left_padding_slider.setValue(clamped_value)
                self.left_padding_slider.blockSignals(False)
                # Ensure the internal variable matches the potentially clamped value
                self.left_marker_shift_added = clamped_value

        elif marker_type == 'right' and self.left_markers:
            # Copy Left Markers TO Right
            self.right_markers = self.left_markers.copy()
            # Copy Left Offset value TO Right Offset variable
            self.right_marker_shift_added = self.left_marker_shift_added

            # Update Right Slider position to match the copied offset
            if hasattr(self, 'right_padding_slider'):
                min_val, max_val = self.right_padding_slider.minimum(), self.right_padding_slider.maximum()
                # Clamp the value to the slider's current range
                clamped_value = max(min_val, min(self.right_marker_shift_added, max_val))
                # Update slider visually, preventing signal emission temporarily
                self.right_padding_slider.blockSignals(True)
                self.right_padding_slider.setValue(clamped_value)
                self.right_padding_slider.blockSignals(False)
                # Ensure the internal variable matches the potentially clamped value
                self.right_marker_shift_added = clamped_value


        # Call update live view after duplicating markers and updating offsets/sliders
        self.update_live_view()
        
    def on_combobox_changed(self):
        
        text=self.combo_box.currentText()
        """Handle the ComboBox change event to update marker values."""
        if text == "Custom":
            # Enable the textbox for custom values when "Custom" is selected
            self.marker_values_textbox.setEnabled(True)
            self.rename_input.setEnabled(True)
            
            # self.marker_values_textbox.setText()
        else:
            # Update marker values based on the preset option
            self.marker_values = self.marker_values_dict.get(text, [])
            self.marker_values_textbox.setEnabled(False)  # Disable the textbox
            self.rename_input.setEnabled(False)
            self.top_label = self.top_label_dict.get(text, [])
            self.top_label = [str(item) if not isinstance(item, str) else item for item in self.top_label]
            self.top_marker_input.setText(", ".join(self.top_label))
            try:
                
                # Ensure that the top_markers list only updates the top_label values serially
                for i in range(0, len(self.top_markers)):
                    self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                
                # # If self.top_label has more entries than current top_markers, add them
                # if len(self.top_label) > len(self.top_markers):
                #     additional_markers = [(self.top_markers[-1][0] + 50 * (i + 1), label) 
                #                           for i, label in enumerate(self.top_label[len(self.top_markers):])]
                #     self.top_markers.extend(additional_markers)
                
                # If self.top_label has fewer entries, truncate the list
                if len(self.top_label) < len(self.top_markers):
                    self.top_markers = self.top_markers[:len(self.top_label)]
                    
                
                
            except:
                pass
            try:
                self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
            except:
                pass

    # Functions for updating contrast and gamma
    
    def reset_gamma_contrast(self):
        try:
            if self.image_before_contrast==None:
                self.image_before_contrast=self.image_master.copy()
            self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
            self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
            self.high_slider.setValue(100)  # Reset contrast to default
            self.low_slider.setValue(100)  # Reset contrast to default
            self.gamma_slider.setValue(100)  # Reset gamma to default
            self.update_live_view()
        except:
            pass

    
    def update_image_contrast(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
            
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                self.update_live_view()
        except:
            pass
    
    def update_image_gamma(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
                
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                self.update_live_view()
        except:
            pass
    
    def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
        """
        Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
        preserving the original format (including color and bit depth) where possible.
        Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
        """
        if not qimage or qimage.isNull():
            return qimage

        original_format = qimage.format()
        try:
            img_array = self.qimage_to_numpy(qimage)
            if img_array is None: raise ValueError("NumPy conversion failed.")

            # Work with float64 for calculations to avoid precision issues
            img_array_float = img_array.astype(np.float64)

            # Determine max value based on original data type
            if img_array.dtype == np.uint16:
                max_val = 65535.0
            elif img_array.dtype == np.uint8:
                max_val = 255.0
            else: # Default for unexpected types (e.g., float input?)
                 max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0

            # --- Apply adjustments ---
            if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                num_channels = img_array.shape[2]
                adjusted_channels = []

                # Process only the color channels (first 3 usually)
                channels_to_process = min(num_channels, 3)
                for i in range(channels_to_process):
                    channel = img_array_float[:, :, i]
                    # Normalize to 0-1 range
                    channel_norm = channel / max_val

                    # Apply brightness (high_factor): Multiply
                    channel_norm = channel_norm * high_factor

                    # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                    mid_grey = 0.5
                    contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                    channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                    # Clip to 0-1 range after contrast/brightness
                    channel_norm = np.clip(channel_norm, 0.0, 1.0)

                    # Apply gamma correction
                    safe_gamma = max(0.01, gamma)
                    channel_norm = np.power(channel_norm, safe_gamma)

                    # Clip again after gamma
                    channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                    # Scale back to original range
                    adjusted_channels.append(channel_norm_clipped * max_val)

                # Reconstruct the image array
                img_array_final_float = np.stack(adjusted_channels, axis=2)

                # Keep the alpha channel (if present) untouched
                if num_channels == 4:
                    alpha_channel = img_array_float[:, :, 3] # Get original alpha
                    img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

            elif img_array.ndim == 2: # Grayscale Image
                # Normalize to 0-1 range
                img_array_norm = img_array_float / max_val

                # Apply brightness
                img_array_norm = img_array_norm * high_factor

                # Apply contrast
                mid_grey = 0.5
                contrast_factor = max(0.01, low_factor)
                img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                # Clip
                img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                # Apply gamma
                safe_gamma = max(0.01, gamma)
                img_array_norm = np.power(img_array_norm, safe_gamma)

                # Clip again
                img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                # Scale back
                img_array_final_float = img_array_norm_clipped * max_val
            else:
                return qimage # Return original if unsupported dimensions

            # Convert back to original data type
            img_array_final = img_array_final_float.astype(img_array.dtype)

            # Convert back to QImage using the helper function
            result_qimage = self.numpy_to_qimage(img_array_final)
            if result_qimage.isNull():
                raise ValueError("Conversion back to QImage failed.")

            # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
            return result_qimage

        except Exception as e:
            traceback.print_exc() # Print detailed traceback
            return qimage # Return original QImage on error
    

    def save_contrast_options(self):
        if self.image:
            self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
            self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
        else:
            QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

    def remove_config(self):
        try:
            # Get the currently selected marker label
            selected_marker = self.combo_box.currentText()
    
            # Ensure the selected marker is not "Custom" before deleting
            if selected_marker == "Custom":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' marker.")
                return
            
            elif selected_marker == "Precision Plus All Blue/Unstained":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
            
            elif selected_marker == "1 kB Plus":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
    
            # Remove the marker label and top_label if they exist
            if selected_marker in self.marker_values_dict:
                del self.marker_values_dict[selected_marker]
            if selected_marker in self.top_label_dict:
                del self.top_label_dict[selected_marker]
    
            # Save the updated configuration
            with open("Imaging_assistant_config.txt", "w") as f:
                config = {
                    "marker_values": self.marker_values_dict,
                    "top_label": self.top_label_dict
                }
                json.dump(config, f)
    
            # Remove from the ComboBox and reset UI
            self.combo_box.removeItem(self.combo_box.currentIndex())
            self.top_marker_input.clear()
    
            QMessageBox.information(self, "Success", f"Configuration '{selected_marker}' removed.")
    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error removing config: {e}")

    def save_config(self):
        """Rename the 'Custom' marker values option and save the configuration."""
        new_name = self.rename_input.text().strip()
        
        # Ensure that the correct dictionary (self.marker_values_dict) is being used
        if self.rename_input.text() != "Enter new name for Custom" and self.rename_input.text() != "":  # Correct condition
            # Save marker values list
            self.marker_values_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
            
            # Save top_label list under the new_name key
            self.top_label_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.top_marker_input.toPlainText().strip("[]").split(",")]

            try:
                # Save both the marker values and top label (under new_name) to the config file
                with open("Imaging_assistant_config.txt", "w") as f:
                    config = {
                        "marker_values": self.marker_values_dict,
                        "top_label": self.top_label_dict  # Save top_label_dict as a dictionary with new_name as key
                    }
                    json.dump(config, f)  # Save both marker_values and top_label_dict
            except Exception as e:
                pass

        self.combo_box.setCurrentText(new_name)
        self.load_config()  # Reload the configuration after saving
    
    
    def load_config(self):
        """
        Load configuration from file. If the file doesn't exist,
        create it with default popular marker standards.
        """
        # Define the path to the config file (e.g., in the same directory as the script)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_filepath = os.path.join(script_dir, "Imaging_assistant_config.txt")
        config_loaded_successfully = False

        # --- Define Default Marker Data ---
        # Use kDa for proteins, bp for DNA for clarity in keys, but values are just numbers
        default_marker_values = {
            # Proteins (kDa)
            "Precision Plus Protein All Blue Prestained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "Precision Plus Protein Unstained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "Precision Plus Protein Dual Color (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "PageRuler Prestained Protein Ladder (Thermo)": [250, 130, 100, 70, 55, 35, 25, 15, 10],
            "PageRuler Unstained Protein Ladder (Thermo)": [200, 150, 100, 70, 50, 40, 30, 20, 10],
            "Spectra Multicolor Broad Range Protein Ladder (Thermo)": [260, 140, 100, 75, 60, 45, 35, 25, 15, 10],
            # DNA (bp)
            "1 kb DNA Ladder (NEB N3232)": [10000, 8000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 500],
            "1 kb Plus DNA Ladder (Thermo 10787018)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 75],
            "TrackIt 1 Kb Plus DNA Ladder (Invitrogen 10488085)": [12000, 11000, 10000, 9000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100], # Often 100 bp, not 75
            "Lambda DNA/HindIII Marker (NEB N3012)": [23130, 9416, 6557, 4361, 2322, 2027, 564], # Approx values
        }
        # Generic default top labels
        default_top_labels_generic = ["MWM", "L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L13", "L14", "L15"]
        default_top_label_dict = {name: default_top_labels_generic[:] for name in default_marker_values.keys()} # Assign generic to all defaults

        default_config_data = {
            "marker_values": default_marker_values,
            "top_label": default_top_label_dict
        }
        # --- End Default Marker Data ---

        # --- Check if config file exists ---
        if os.path.exists(config_filepath):
            # --- Try to load existing file ---
            try:
                with open(config_filepath, "r") as f:
                    config = json.load(f)

                # Load marker values and top label from the file
                # Use .get() for safety, falling back to DEFAULTS if key is missing in file
                self.marker_values_dict = config.get("marker_values", default_marker_values.copy())
                self.top_label_dict = config.get("top_label", default_top_label_dict.copy())
                config_loaded_successfully = True

            except (json.JSONDecodeError, IOError) as e:
                print(f"Error loading config file '{config_filepath}': {e}. Using default values.")
                # Fallback to defaults if loading fails
                self.marker_values_dict = default_marker_values.copy()
                self.top_label_dict = default_top_label_dict.copy()
            except Exception as e:
                print(f"An unexpected error occurred loading config: {e}. Using default values.")
                traceback.print_exc() # Print full traceback for unexpected errors
                self.marker_values_dict = default_marker_values.copy()
                self.top_label_dict = default_top_label_dict.copy()

        else:
            # --- Config file NOT found, create it with defaults ---
            print(f"Config file not found at '{config_filepath}'. Creating with defaults.")
            try:
                with open(config_filepath, "w") as f:
                    json.dump(default_config_data, f, indent=4)

                # Use the defaults we just wrote
                self.marker_values_dict = default_marker_values.copy()
                self.top_label_dict = default_top_label_dict.copy()
                config_loaded_successfully = True # Technically created, not loaded, but state is set

            except IOError as e:
                print(f"Error creating default config file '{config_filepath}': {e}. Using in-memory defaults.")
                # Fallback to in-memory defaults if creation fails
                self.marker_values_dict = default_marker_values.copy()
                self.top_label_dict = default_top_label_dict.copy()
            except Exception as e:
                 print(f"An unexpected error occurred creating config: {e}. Using in-memory defaults.")
                 traceback.print_exc()
                 self.marker_values_dict = default_marker_values.copy()
                 self.top_label_dict = default_top_label_dict.copy()


        # --- Update UI based on the loaded/default data ---
        try:
            # Ensure combo_box exists before manipulating it
            if hasattr(self, 'combo_box'):
                self.combo_box.clear()
                # Add marker names sorted alphabetically for better usability
                sorted_marker_names = sorted(self.marker_values_dict.keys())
                self.combo_box.addItems(sorted_marker_names)
                self.combo_box.addItem("Custom") # Add Custom option last
                # Optionally set a default selection
                if "Precision Plus Protein All Blue Prestained (Bio-Rad)" in self.marker_values_dict:
                    self.combo_box.setCurrentText("Precision Plus Protein All Blue Prestained (Bio-Rad)")
                elif sorted_marker_names: # Select first alphabetically if default not present
                    self.combo_box.setCurrentText(sorted_marker_names[0])

                # Trigger initial population of text boxes based on current selection
                self.on_combobox_changed()
            else:
                 print("Warning: combo_box UI element not found during config load.")

            # Set the initial top_label based on the combobox selection, if possible
            if hasattr(self, 'top_marker_input'):
                current_selection = self.combo_box.currentText() if hasattr(self, 'combo_box') else None
                if current_selection and current_selection != "Custom" and current_selection in self.top_label_dict:
                    self.top_label = self.top_label_dict[current_selection][:] # Use a copy
                else:
                    # Use a generic default if selection is Custom or not found
                    self.top_label = default_top_labels_generic[:]
                # Ensure all items are strings before joining
                self.top_label = [str(item) for item in self.top_label]
                self.top_marker_input.setText(", ".join(self.top_label))
            else:
                 print("Warning: top_marker_input UI element not found during config load.")


        except Exception as e:
             print(f"Error updating UI after config load/create: {e}")
             traceback.print_exc()
    
    def paste_image(self):
        """Handle pasting image from clipboard."""
        self.is_modified = True
        self.reset_image() # Clear previous state first
        self.load_image_from_clipboard()
        # UI updates (label size, sliders) should happen within load_image_from_clipboard
        self.update_live_view()
        self.save_state() # Save state after pasting
    
    def load_image_from_clipboard(self):
        """
        Load an image from the clipboard into self.image, preserving format.
        Prioritizes file paths over raw image data to handle macOS file copying correctly.
        If loaded from a file path, attempts to load an associated config file.
        """
        # Check for unsaved changes before proceeding
        # if not self.prompt_save_if_needed(): # Optional: uncomment if needed
        #     return # Abort paste if user cancels save

        self.reset_image() # Clear previous state first

        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
        loaded_image = None
        source_info = "Clipboard" # Default source
        config_loaded_from_paste = False # Flag to track if config was loaded

        # --- PRIORITY 1: Check for File URLs ---
        if mime_data.hasUrls():
            urls = mime_data.urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if os.path.isfile(file_path) and file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                    loaded_image = QImage(file_path)
                    source_info = file_path

                    if loaded_image.isNull():
                        try:
                            pil_image = Image.open(file_path)
                            loaded_image = ImageQt.ImageQt(pil_image)
                            if loaded_image.isNull():
                                loaded_image = None
                            else:
                                source_info = f"{file_path} (Pillow)"
                        except Exception as e:
                            QMessageBox.warning(self, "File Load Error", f"Could not load image from file '{os.path.basename(file_path)}':\n{e}")
                            loaded_image = None

                    # --- *** ADDED: CONFIG FILE LOADING FOR PASTED FILE *** ---
                    if loaded_image and not loaded_image.isNull():
                        self.image_path = file_path # Store the path early for config lookup
                        base_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                        # Construct potential config file name (remove _original/_modified if present)
                        config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                        config_path = os.path.join(os.path.dirname(file_path), f"{config_base}_config.txt")

                        if os.path.exists(config_path):
                            try:
                                with open(config_path, "r") as config_file:
                                    config_data = json.load(config_file)
                                self.apply_config(config_data) # Apply loaded settings
                                config_loaded_from_paste = True # Set flag
                            except Exception as e:
                                QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply associated config file '{os.path.basename(config_path)}': {e}")

                    # --- *** END OF ADDED CONFIG LOADING *** ---


        # --- PRIORITY 2: Check for Raw Image Data (if no valid file was loaded) ---
        if loaded_image is None and mime_data.hasImage():
            loaded_image = clipboard.image()
            if not loaded_image.isNull():
                source_info = "Clipboard Image Data"
            else:
                try:
                    pil_image = ImageGrab.grabclipboard()
                    if isinstance(pil_image, Image.Image):
                        loaded_image = ImageQt.ImageQt(pil_image)
                        if loaded_image.isNull():
                             loaded_image = None
                        else:
                            source_info = "Clipboard Image Data (Pillow Grab)"
                    else:
                        loaded_image = None
                except Exception: # Keep quiet on Pillow grab errors
                    loaded_image = None

        # --- Process the successfully loaded image ---
        if loaded_image and not loaded_image.isNull():
            self.is_modified = True
            self.image = loaded_image

            # Initialize backups
            self.original_image = self.image.copy()
            self.image_master = self.image.copy()
            self.image_before_padding = None
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_padded = False
            # self.image_path is set earlier if loaded from file

            # --- Update UI Elements ---
            try:
                w = self.image.width()
                h = self.image.height()
                if w <= 0 or h <= 0: raise ValueError("Invalid image dimensions after load.")

                # Update preview label size
                ratio=w/h if h > 0 else 1
                target_width = int(self.label_size) # Use current label_size setting
                target_height = int(target_width / ratio)
                # Optional: Apply max height constraint
                # if hasattr(self, 'preview_label_max_height_setting') and target_height > self.preview_label_max_height_setting:
                #     target_height = self.preview_label_max_height_setting
                #     target_width = int(target_height * ratio)
                self._update_preview_label_size()

                # Update slider ranges based on *new* label size
                render_scale = 3
                render_width = self.live_view_label.width() * render_scale
                render_height = self.live_view_label.height() * render_scale
                self._update_marker_slider_ranges()

                # Set recommended padding only if config wasn't loaded
                if not config_loaded_from_paste:
                    self.left_padding_input.setText(str(int(w * 0.1)))
                    self.right_padding_input.setText(str(int(w * 0.1)))
                    self.top_padding_input.setText(str(int(h * 0.15)))
                    self.bottom_padding_input.setText("0")

                # Update window title and status bar
                display_source = os.path.basename(source_info.split(" (")[0]) # Show filename or simplified source
                title_suffix = " (+ Config)" if config_loaded_from_paste else ""
                self.setWindowTitle(f"{self.window_title}::{display_source}{title_suffix}")
                self._update_status_bar()

            except Exception as e:
                QMessageBox.warning(self, "UI Update Error", f"Could not update UI elements after pasting: {e}")
                self._update_preview_label_size()
                
            enable_pan = self.live_view_label.zoom_level > 1.0
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
            self.update_live_view()
            self.save_state()

        else:
            # If no image was successfully loaded
            QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
            # Clean up state
            self.image = None
            self.original_image = None
            self.image_master = None
            self.image_path = None
            self.live_view_label.clear()
            self._update_preview_label_size()
            self.setWindowTitle(self.window_title)
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
            self._update_status_bar()
            self.update_live_view()
        
    def update_font(self):
        self.save_state()
        """Update the font settings based on UI inputs"""
        # Update font family from the combo box
        self.font_family = self.font_combo_box.currentFont().family()
        
        # Update font size from the spin box
        self.font_size = self.font_size_spinner.value()
        
        self.font_rotation = int(self.font_rotation_input.value())
        
    
        # Once font settings are updated, update the live view immediately
        self.update_live_view()
    
    def select_font_color(self):
        self.save_state()
        color = QColorDialog.getColor()
        if color.isValid():
            self.font_color = color  # Store the selected color   
            self.update_font()
        self._update_color_button_style(self.font_color_button, self.font_color)
            
    def load_image(self):
        self.prompt_save_if_needed()
        self.is_modified = True # Mark as modified when loading new image
        # self.undo_stack = []
        # self.redo_stack = []
        self.reset_image() # Clear previous state

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
        )
        if file_path:
            self.image_path = file_path
            loaded_image = QImage(self.image_path)

            if loaded_image.isNull():
                # Try loading with Pillow as fallback
                try:
                    pil_image = Image.open(self.image_path)
                    # Use ImageQt for reliable conversion, preserves most formats
                    loaded_image = ImageQt.ImageQt(pil_image)
                    if loaded_image.isNull():
                        raise ValueError("Pillow could not convert to QImage.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}': {e}")
                    self.image_path = None
                    return

            # --- Keep the loaded image format ---
            self.image = loaded_image

            # --- Initialize backups with the loaded format ---
            if not self.image.isNull():
                self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                self.image_master = self.image.copy()   # Master copy for resets
                self.image_before_padding = None        # Reset padding state
                self.image_contrasted = self.image.copy() # Backup for contrast
                self.image_before_contrast = self.image.copy() # Backup for contrast
                self.image_padded = False               # Reset flag

                self.setWindowTitle(f"{self.window_title}::{self.image_path}")

                # --- Load Associated Config File (Logic remains the same) ---
                self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                config_name = ""
                if self.base_name.endswith("_original"):
                    config_name = self.base_name.replace("_original", "_config.txt")
                else:
                    config_name = self.base_name + "_config.txt"

                config_path = os.path.join(os.path.dirname(file_path), config_name)

                if os.path.exists(config_path):
                    try:
                        with open(config_path, "r") as config_file:
                            config_data = json.load(config_file)
                        self.apply_config(config_data) # Apply loaded settings
                    except Exception as e:
                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                # --- End Config File Loading ---

            else:
                 QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                 return

            # --- Update UI Elements (Label size, sliders) ---
            if self.image and not self.image.isNull():
                try:
                    # (UI update logic remains the same as before)
                    self._update_preview_label_size()

                    render_scale = 3
                    render_width = self.live_view_label.width() * render_scale
                    render_height = self.live_view_label.height() * render_scale
                    self._update_marker_slider_ranges()


                    if not os.path.exists(config_path):
                        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                        self.bottom_padding_input.setText("0")

                except Exception as e:
                    pass
            # --- End UI Element Update ---
            
            
            self._update_status_bar() # <--- Add this
            enable_pan = self.live_view_label.zoom_level > 1.0
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
            self.update_live_view() # Render the loaded image
            self.save_state() # Save initial loaded state
    
    def apply_config(self, config_data):
        self.left_padding_input.setText(config_data["adding_white_space"]["left"])
        self.right_padding_input.setText(config_data["adding_white_space"]["right"])
        self.top_padding_input.setText(config_data["adding_white_space"]["top"])
        
        try:
            self.transparency=int(config_data["adding_white_space"]["transparency"])
        except:
            pass
        
            
        try:
            self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
        except:
            pass
        
    
        try:
            self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
            self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
            self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
        except (KeyError, ValueError) as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
            pass
        try:
            self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
            self.top_marker_input.setText(", ".join(self.top_label))
        except KeyError as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
            pass
    
        self.font_family = config_data["font_options"]["font_family"]
        self.font_size = config_data["font_options"]["font_size"]
        self.font_rotation = config_data["font_options"]["font_rotation"]
        self.font_color = QColor(config_data["font_options"]["font_color"])
    
        try:
            top_padding_val = int(config_data["marker_padding"]["top"])
            left_padding_val = int(config_data["marker_padding"]["left"])
            right_padding_val = int(config_data["marker_padding"]["right"])
        except (KeyError, ValueError, TypeError) as e:
            print(f"Warning: Invalid 'marker_padding' data in config: {e}. Using defaults.")
            top_padding_val, left_padding_val, right_padding_val = 0, 0, 0

        # Load added shifts (internal variables used for rendering)
        try:
            self.left_marker_shift_added = int(config_data["added_shift"]["left"])
            self.right_marker_shift_added = int(config_data["added_shift"]["right"])
            self.top_marker_shift_added = int(config_data["added_shift"]["top"])
            # Validation: Ensure loaded shift matches padding value if both exist and are valid
            if self.left_marker_shift_added != left_padding_val or \
               self.right_marker_shift_added != right_padding_val or \
               self.top_marker_shift_added != top_padding_val:
                 print("Warning: Config 'added_shift' differs from 'marker_padding'. Prioritizing 'added_shift'.")
                 # Optionally update padding_val to match shift for consistency if desired:
                 # left_padding_val = self.left_marker_shift_added
                 # right_padding_val = self.right_marker_shift_added
                 # top_padding_val = self.top_marker_shift_added

        except KeyError:
            # If added_shift is missing (older config?), use marker_padding as the source
            print("Info: 'added_shift' not found in config, using 'marker_padding'.")
            self.left_marker_shift_added = left_padding_val
            self.right_marker_shift_added = right_padding_val
            self.top_marker_shift_added = top_padding_val
        except (ValueError, TypeError) as e:
             # Handle potential conversion errors if data is invalid
             print(f"Warning: Invalid 'added_shift' data in config: {e}. Resetting shifts.")
             self.left_marker_shift_added, left_padding_val = 0, 0
             self.right_marker_shift_added, right_padding_val = 0, 0
             self.top_marker_shift_added, top_padding_val = 0, 0

        # Load slider ranges (set range BEFORE value for Qt)
        try:
            lr_min, lr_max = map(int, config_data["slider_ranges"]["left"])
            self.left_padding_slider.setRange(lr_min, lr_max)
            self.left_slider_range = [lr_min, lr_max] # Store internally

            rr_min, rr_max = map(int, config_data["slider_ranges"]["right"])
            self.right_padding_slider.setRange(rr_min, rr_max)
            self.right_slider_range = [rr_min, rr_max] # Store internally

            tr_min, tr_max = map(int, config_data["slider_ranges"]["top"])
            self.top_padding_slider.setRange(tr_min, tr_max)
            self.top_slider_range = [tr_min, tr_max] # Store internally

        except (KeyError, ValueError, TypeError) as e:
            print(f"Warning: Slider ranges missing or invalid in config: {e}. Ranges will be updated.")
            # Don't set default ranges here, let _update_marker_slider_ranges handle it later
            pass

        # Now set the slider values using the determined padding_val
        # Qt might clamp these values if they are outside the range just set.
        self.top_padding_slider.setValue(top_padding_val)
        self.left_padding_slider.setValue(left_padding_val)
        self.right_padding_slider.setValue(right_padding_val)

        # --- CRITICAL RE-SYNC ---
        # After setting slider values (which might have been clamped by setRange),
        # ensure the internal _shift_added variables MATCH the final slider values.
        # This makes the internal state consistent with the UI state *after* loading.
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.right_marker_shift_added = self.right_padding_slider.value()
        self.top_marker_shift_added = self.top_padding_slider.value()
        
        try:
            try:
                x = list(map(int, config_data["marker_labels"]["left"]))
                self.marker_values_textbox.setText(str(x))
            except:
                x = list(map(int, config_data["marker_labels"]["right"]))
                self.marker_values_textbox.setText(str(x))
            if self.marker_values_textbox.text!="":
                self.combo_box.setCurrentText("Custom")
                self.marker_values_textbox.setEnabled(True)                
            else:
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.marker_values_textbox.setEnabled(False)
                
        except:
            pass
    
        try:
            loaded_custom_markers = []
            for marker_dict in config_data.get("custom_markers", []):
                try:
                    x = float(marker_dict["x"])
                    y = float(marker_dict["y"])
                    text = str(marker_dict["text"])
                    color = QColor(marker_dict["color"])
                    if not color.isValid(): color = QColor(0,0,0) # Fallback if invalid string
                    loaded_custom_markers.append((
                        float(marker_dict["x"]), float(marker_dict["y"]), str(marker_dict["text"]),
                        color, # Store the QColor object
                        str(marker_dict["font"]), int(marker_dict["font_size"]),
                        bool(marker_dict.get("bold", False)), bool(marker_dict.get("italic", False))
                    ))
                except (KeyError, ValueError, TypeError) as e:
                    print(f"Warning: Skipping invalid custom marker in config: {marker_dict}, Error: {e}")
            self.custom_markers = [list(m) for m in loaded_custom_markers] 
                
                
                
        except:
            pass
        
        try:
            loaded_custom_shapes = []
            for shape_dict in config_data.get("custom_shapes", []):
                try:
                    shape_type = shape_dict.get('type')
                    color_str = shape_dict.get('color')
                    thickness = int(shape_dict.get('thickness'))
    
                    # Validate basic fields
                    if shape_type not in ['line', 'rectangle'] or not color_str or thickness < 1:
                        raise ValueError("Missing or invalid basic shape data")
    
                    # Validate and convert coordinate data
                    parsed_shape = {'type': shape_type, 'color': color_str, 'thickness': thickness}
                    if shape_type == 'line':
                        start = tuple(map(float, shape_dict.get('start')))
                        end = tuple(map(float, shape_dict.get('end')))
                        if len(start) != 2 or len(end) != 2: raise ValueError("Invalid line coordinates")
                        parsed_shape['start'] = start
                        parsed_shape['end'] = end
                    elif shape_type == 'rectangle':
                        rect = tuple(map(float, shape_dict.get('rect'))) # x,y,w,h
                        if len(rect) != 4 or rect[2] < 0 or rect[3] < 0: raise ValueError("Invalid rectangle coordinates/size")
                        parsed_shape['rect'] = rect
                    else:
                        continue # Skip unknown types
    
                    loaded_custom_shapes.append(parsed_shape) # Add the validated shape dict
    
                except (KeyError, ValueError, TypeError, AttributeError) as e:
                     print(f"Warning: Skipping invalid custom shape in config: {shape_dict}, Error: {e}")
            self.custom_shapes = loaded_custom_shapes # Assign the loaded list
        except:
            pass
        # Set slider ranges from config_data
        try:
            self.left_padding_slider.setRange(
                int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
            )
            self.right_padding_slider.setRange(
                int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
            )
            self.top_padding_slider.setRange(
                int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
            )
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Slider ranges not found in config_data.")
            pass
        
        try:
            self.left_marker_shift_added=int(config_data["added_shift"]["left"])
            self.right_marker_shift_added=int(config_data["added_shift"]["right"])
            self.top_marker_shift_added=int(config_data["added_shift"]["top"])
            
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Added Shift not found in config_data.");
            pass
            
        # Apply font settings

        
        #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
        
        font_size_new=self.font_size
        font_rotation_new=self.font_rotation
        
        self.font_combo_box.setCurrentFont(QFont(self.font_family))
        self.font_size_spinner.setValue(font_size_new)
        self.font_rotation_input.setValue(font_rotation_new)

        self.update_live_view()
        
    def get_current_config(self):
        """Gathers the current application state into a dictionary for saving."""
        config = {
            "adding_white_space": {
                "left": self.left_padding_input.text(),
                "right": self.right_padding_input.text(),
                "top": self.top_padding_input.text(),
                "bottom": self.bottom_padding_input.text(),
                "transparency": self.transparency, # Assuming self.transparency exists
            },
            "marker_positions": {
                # Store positions only for standard markers
                "left": [(pos, label) for pos, label in getattr(self, 'left_markers', [])],
                "right": [(pos, label) for pos, label in getattr(self, 'right_markers', [])],
                "top": [(pos, label) for pos, label in getattr(self, 'top_markers', [])],
            },
            "marker_labels": { # Storing labels separately might be redundant if already in positions
                "top": getattr(self, 'top_label', []),
                "left": [marker[1] for marker in getattr(self, 'left_markers', [])],
                "right": [marker[1] for marker in getattr(self, 'right_markers', [])],
            },
            "marker_padding": { # Current slider values
                "top": self.top_padding_slider.value() if hasattr(self, 'top_padding_slider') else 0,
                "left": self.left_padding_slider.value() if hasattr(self, 'left_padding_slider') else 0,
                "right": self.right_padding_slider.value() if hasattr(self, 'right_padding_slider') else 0,
            },
            "font_options": { # Default font options for standard markers
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_rotation": self.font_rotation,
                "font_color": self.font_color.name(),
            },
            "slider_ranges": { # Current slider ranges
                 "left": getattr(self, 'left_slider_range', [-100, 1000]),
                 "right": getattr(self, 'right_slider_range', [-100, 1000]),
                 "top": getattr(self, 'top_slider_range', [-100, 1000]),
             },
            "added_shift": { # Store current added shifts
                 "left": getattr(self, 'left_marker_shift_added', 0),
                 "right": getattr(self, 'right_marker_shift_added', 0),
                 "top": getattr(self, 'top_marker_shift_added', 0),
            }
        }

        # --- Updated Custom Markers Section ---
        custom_markers_data = []
        # Use getattr for safety, default to empty list
        for marker_tuple in getattr(self, "custom_markers", []):
            try:
                # Unpack 8 elements
                x, y, text, color, font, font_size, is_bold, is_italic = marker_tuple
                custom_markers_data.append({
                    "x": x,
                    "y": y,
                    "text": text,
                    "color": color.name(), # Save color name/hex
                    "font": font,
                    "font_size": font_size,
                    "bold": is_bold,       # Save bold flag
                    "italic": is_italic    # Save italic flag
                })
            except (ValueError, TypeError, IndexError) as e:
                pass
        config["custom_markers"] = custom_markers_data
        config["custom_shapes"] = [dict(s) for s in getattr(self, "custom_shapes", [])]
        # --- End Updated Custom Markers ---

        return config
    
    def add_band(self, event):
        # --- Ensure internal lists match UI input BEFORE adding band ---
        # This call ensures self.marker_values and self.top_label are
        # updated based on the current text in the UI boxes.
        self.update_all_labels()
        # -------------------------------------------------------------

        self.save_state()
        # Ensure there's an image loaded and marker mode is active
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        if not self.image or not self.marker_mode:
            return

        # --- Get Coordinates and Scaling (Same as previous version) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        # if self.show_grid_checkbox.isChecked():
        #     grid_size = self.grid_size_input.value()
        #     cursor_x = round(cursor_x / grid_size) * grid_size
        #     cursor_y = round(cursor_y / grid_size) * grid_size
        grid_size = self.grid_size_input.value() # Get grid size once

        if self.show_grid_checkbox_x.isChecked():
            if grid_size > 0: # Avoid division by zero
                 cursor_x = round(cursor_x / grid_size) * grid_size

        if self.show_grid_checkbox_y.isChecked():
            if grid_size > 0: # Avoid division by zero
                 cursor_y = round(cursor_y / grid_size) * grid_size

        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image.width() > 0 else 1
        image_height = self.image.height() if self.image.height() > 0 else 1
        uniform_scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1
        offset_x = (displayed_width - image_width * uniform_scale) / 2
        offset_y = (displayed_height - image_height * uniform_scale) / 2
        image_x = (cursor_x - offset_x) / uniform_scale if uniform_scale != 0 else 0
        image_y = (cursor_y - offset_y) / uniform_scale if uniform_scale != 0 else 0
        
        x_start_percent = 0
        x_end_percent = 100
        y_start_percent = 0
        y_end_percent = 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)

        # --- Get Render Info for Slider Positioning (Same as previous version) ---
        render_scale = 3
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale

        # --- Validate Coordinates (Same as previous version) ---
        current_image_width = self.image.width()
        current_image_height = self.image.height()
        if not (0 <= image_y <= current_image_height) and self.marker_mode in ["left", "right"]:
             return
        if not (0 <= image_x <= current_image_width) and self.marker_mode == "top":
             return
        # --- End Coordinate/Scaling/Validation ---

        try:
            # --- Left Marker Logic ---
            if self.marker_mode == "left":
                # Determine label based on current count and *now updated* self.marker_values
                current_marker_count = len(self.left_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        pass

                    
                self.left_markers.append((image_y, marker_value_to_add))
                self.current_left_marker_index += 1 # Still increment conceptual index

                # Set slider position only for the *very first* marker placed
                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.left_padding_slider.setValue(0)
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.left_padding_slider.setValue(padding_value)
                    self.left_marker_shift_added = padding_value 

            # --- Right Marker Logic ---
            elif self.marker_mode == "right":
                current_marker_count = len(self.right_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        pass

                self.right_markers.append((image_y, marker_value_to_add))
                self.current_right_marker_index += 1

                if is_first_marker:
                    padding_value=int((image_x * (render_width / self.image.width())))
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.right_padding_slider.setValue(padding_value)
                    self.right_marker_shift_added = padding_value

            # --- Top Marker Logic ---
            elif self.marker_mode == "top":
                current_marker_count = len(self.top_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.top_label list, which was just updated
                if current_marker_count < len(self.top_label):
                    label_to_add = self.top_label[current_marker_count]
                else:
                    label_to_add = ""
                    if current_marker_count == len(self.top_label): # Print warning only once
                        pass     

                self.top_markers.append((image_x, label_to_add))
                self.current_top_label_index += 1

                if is_first_marker:
                    padding_value=int((image_y - y_start) * (render_height / self.image.height()))
                    self.top_padding_slider.setValue(0)
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
                    self.top_padding_slider.setValue(padding_value)
                    self.top_marker_shift_added = padding_value

        except Exception as e:
             # Catch other potential errors
             import traceback
             traceback.print_exc()
             QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

        # Update the live view to render the newly added marker
        self.update_live_view()
        

        
    def enable_left_marker_mode(self):
        self.marker_mode = "left"
        self.current_left_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)

    def enable_right_marker_mode(self):
        self.marker_mode = "right"
        self.current_right_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_top_marker_mode(self):
        self.marker_mode = "top"
        self.current_top_label_index
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
        
    # def remove_padding(self):
    #     if self.image_before_padding!=None:
    #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
    #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
    #     self.image_padded = False  # Reset the padding state
    #     w=self.image.width()
    #     h=self.image.height()
    #     # Preview window
    #     ratio=w/h
    #     self.label_width = 540
    #     label_height=int(self.label_width/ratio)
    #     if label_height>self.label_width:
    #         label_height=540
    #         self.label_width=ratio*label_height
    #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    #     self.update_live_view()
        
    def finalize_image(self): # Padding
        self.save_state()
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
            return

        try:
            # Ensure padding values are non-negative
            padding_left = max(0, int(self.left_padding_input.text()))
            padding_right = max(0, int(self.right_padding_input.text()))
            padding_top = max(0, int(self.top_padding_input.text()))
            padding_bottom = max(0, int(self.bottom_padding_input.text()))
            # Update input fields in case negative values were entered
            self.left_padding_input.setText(str(padding_left))
            self.right_padding_input.setText(str(padding_right))
            self.top_padding_input.setText(str(padding_top))
            self.bottom_padding_input.setText(str(padding_bottom))

        except ValueError:
            QMessageBox.warning(self, "Error", "Please enter valid non-negative integers for padding.")
            return

        # --- Check if padding is actually needed ---
        if padding_left == 0 and padding_right == 0 and padding_top == 0 and padding_bottom == 0:
            QMessageBox.information(self, "Info", "No padding specified. Image remains unchanged.")
            return # Nothing to do

        try:
            # --- Determine if the source image has transparency ---
            
            source_has_alpha = self.image.hasAlphaChannel()

            # --- Convert source QImage to NumPy array ---
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")
            
            fill_value = None
            # Check if NumPy array indicates alpha (4 channels) even if QImage didn't report it explicitly
            numpy_indicates_alpha = (np_img.ndim == 3 and np_img.shape[2] == 4)
            has_alpha = source_has_alpha or numpy_indicates_alpha # Combine checks

            if has_alpha:
                # Pad images WITH existing alpha using TRANSPARENT black
                # OpenCV expects (B, G, R, Alpha) or just Alpha if grayscale+alpha
                # Provide 4 values for safety, assuming color/grayscale + alpha format from qimage_to_numpy
                fill_value = (0, 0, 0, 0) # Transparent Black for NumPy/OpenCV
            elif np_img.ndim == 3: # Opaque Color (BGR)
                # Pad opaque color images with WHITE
                fill_value = (255, 255, 255) # White for BGR format expected by OpenCV
            elif np_img.ndim == 2: # Opaque Grayscale
                # Pad opaque grayscale images with WHITE
                # Determine white value based on dtype
                if np_img.dtype == np.uint16:
                    fill_value = 65535 # White for 16-bit grayscale
                else: # Assume uint8 or other standard grayscale
                    fill_value = 255 # White for 8-bit grayscale
            else:
                 raise ValueError(f"Unsupported image dimensions for padding: {np_img.ndim}")

            # --- Adjust markers BEFORE creating the new padded image ---
            self.adjust_markers_for_padding(padding_left, padding_right, padding_top, padding_bottom)

            # --- Pad using OpenCV's copyMakeBorder ---
            padded_np = cv2.copyMakeBorder(np_img, padding_top, padding_bottom,
                                           padding_left, padding_right,
                                           cv2.BORDER_CONSTANT, value=fill_value)

            # --- Convert padded NumPy array back to QImage ---
            # numpy_to_qimage should automatically select a format supporting alpha if padded_np has 4 channels
            padded_image = self.numpy_to_qimage(padded_np)
            if padded_image.isNull():
                 raise ValueError("Conversion back to QImage failed after padding.")


            # --- Update main image and backups ---
            # Store the image *before* this padding operation, unless padding was already applied
            if not self.image_padded:
                self.image_before_padding = self.image.copy()
            # Else: keep the existing image_before_padding from the *previous* unpadded state

            self.image = padded_image
            self.image_padded = True # Indicate padding has been applied

            # Update backups consistently to reflect the new padded state
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy() # Contrast baseline resets after padding

            # --- Update UI (label size, slider ranges, live view) ---
            # (Keep the existing UI update logic here)
 
            self._update_preview_label_size()

            render_scale = 3
            render_width = self.live_view_label.width() * render_scale
            render_height = self.live_view_label.height() * render_scale
            self._update_marker_slider_ranges()
            # --- End UI Update ---

            self.update_live_view() # Refresh display
            self._update_status_bar() # Update size/depth info

        except Exception as e:
            QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
            traceback.print_exc()
    # --- END: Modified Image Operations ---
    
    def adjust_markers_for_padding(self, padding_left, padding_right, padding_top, padding_bottom):
        """Adjust marker positions based on padding."""
        # Adjust left markers
        self.left_markers = [(y + padding_top, label) for y, label in self.left_markers]
        # Adjust right markers
        self.right_markers = [(y + padding_top, label) for y, label in self.right_markers]
        # Adjust top markers
        self.top_markers = [(x + padding_left, label) for x, label in self.top_markers]        
        
    
    def update_left_padding(self):
        # Update left padding when slider value changes
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.update_live_view()

    def update_right_padding(self):
        # Update right padding when slider value changes
        new_value = self.right_padding_slider.value()
        # --- Add check to prevent redundant updates ---
        if new_value != self.right_marker_shift_added:
            self.right_marker_shift_added = new_value
            self.update_live_view()
        
    def update_top_padding(self):
        # Update top padding when slider value changes
        self.top_marker_shift_added = self.top_padding_slider.value()
        self.update_live_view()

    def update_live_view(self):
        """Updates the preview label with the current image state, including transformations and markers."""
        if not self.image or self.image.isNull(): # Check if image is valid
            # If no valid image, clear the label and disable relevant UI
            if hasattr(self, 'live_view_label'):
                self.live_view_label.clear()
                self.live_view_label.update() # Ensure paintEvent clears overlays if needed

            # Disable buttons that require an image
            if hasattr(self, 'predict_button'): self.predict_button.setEnabled(False)
            if hasattr(self, 'save_action'): self.save_action.setEnabled(False)
            if hasattr(self, 'save_svg_action'): self.save_svg_action.setEnabled(False)
            if hasattr(self, 'copy_action'): self.copy_action.setEnabled(False)
            # ... disable others in different tabs if necessary ...
            return
        else:
            # If image is valid, enable relevant buttons
            if hasattr(self, 'save_action'): self.save_action.setEnabled(True)
            if hasattr(self, 'save_svg_action'): self.save_svg_action.setEnabled(True)
            if hasattr(self, 'copy_action'): self.copy_action.setEnabled(True)


        # Enable the "Predict Molecular Weight" button if markers are present
        if hasattr(self, 'predict_button'):
            # Use getattr for safety, default to empty list
            left_m = getattr(self, 'left_markers', [])
            right_m = getattr(self, 'right_markers', [])
            if left_m or right_m:
                self.predict_button.setEnabled(True)
            else:
                self.predict_button.setEnabled(False)

        # --- Define rendering parameters ---
        render_scale = 3
        try: # Protect against label size errors during initialization
            view_width = self.live_view_label.width()
            view_height = self.live_view_label.height()
            if view_width <= 0: view_width = 600 # Fallback width
            if view_height <= 0: view_height = 400 # Fallback height
            render_width = view_width * render_scale
            render_height = view_height * render_scale
        except Exception:
             # Use safe defaults if label size isn't ready or errors occur
             render_width = 1800
             render_height = 1200


        # --- Prepare image for transformations (Start with current self.image) ---
        # Ensure self.image is still valid before copying
        if not self.image or self.image.isNull():
            print("Error: self.image became invalid before transformation.")
            self.live_view_label.clear()
            return

        image_to_transform = self.image.copy() # Work on a copy

        # --- Apply Rotation ---
        orientation = 0.0
        # Check if the orientation slider UI element exists and is accessible
        if hasattr(self, 'orientation_slider') and self.orientation_slider:
            orientation = float(self.orientation_slider.value() / 20)
            # Update the label text only if the label exists
            if hasattr(self, 'orientation_label') and self.orientation_label:
                self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")

            if abs(orientation) > 0.01: # Apply only if rotation is significant
                 # Ensure image_to_transform is valid before rotating
                 if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
                      print("Warning: Skipping rotation due to invalid source image dimensions.")
                 else:
                     transform_rotate = QTransform()
                     w_rot, h_rot = image_to_transform.width(), image_to_transform.height()
                     transform_rotate.translate(w_rot / 2, h_rot / 2)
                     transform_rotate.rotate(orientation)
                     transform_rotate.translate(-w_rot / 2, -h_rot / 2)
                     # Store result temporarily to check validity
                     temp_rotated = image_to_transform.transformed(transform_rotate, Qt.SmoothTransformation)
                     if not temp_rotated.isNull():
                         image_to_transform = temp_rotated
                     else:
                         print("Warning: Rotation transform resulted in an invalid image. Skipping rotation.")
                         # Keep image_to_transform as it was before rotation attempt


        # --- Apply Skew / Taper ---
        taper_value = 0.0
        # Check if the skew slider UI element exists
        if hasattr(self, 'taper_skew_slider') and self.taper_skew_slider:
            taper_value = self.taper_skew_slider.value() / 100.0
            # Update the label text only if the label exists
            if hasattr(self, 'taper_skew_label') and self.taper_skew_label:
                self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

            if abs(taper_value) > 0.01: # Apply only if skew is significant
                # Ensure image_to_transform is valid before skewing
                if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
                     print("Warning: Skipping skew due to invalid source image dimensions.")
                else:
                    width = image_to_transform.width()
                    height = image_to_transform.height()
                    source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
                    destination_corners = QPolygonF(source_corners)
                    if taper_value > 0:
                        destination_corners[0].setX(width * taper_value / 2)
                        destination_corners[1].setX(width * (1 - taper_value / 2))
                    elif taper_value < 0:
                        destination_corners[2].setX(width * (-taper_value / 2))
                        destination_corners[3].setX(width * (1 + taper_value / 2))

                    transform_skew = QTransform()
                    # Check if quadToQuad succeeds before applying
                    if QTransform.quadToQuad(source_corners, destination_corners, transform_skew):
                        temp_skewed = image_to_transform.transformed(transform_skew, Qt.SmoothTransformation)
                        if not temp_skewed.isNull():
                            image_to_transform = temp_skewed
                        else:
                            print("Warning: Skew transform resulted in an invalid image. Skipping skew.")
                            # Keep image_to_transform as it was before skew attempt
                    else:
                         print("Warning: Failed to calculate skew transformation matrix.")


        # --- Scale the final transformed image for rendering ---
        # Check if image_to_transform is still valid after potential transforms
        if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
            print("Error: Image became invalid before final scaling for rendering.")
            self.live_view_label.clear()
            return

        scaled_image_for_render = image_to_transform.scaled(
            render_width,
            render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        if scaled_image_for_render.isNull():
            print("Error: Final scaling for rendering failed.")
            self.live_view_label.clear()
            return


        # --- Render on high-resolution canvas ---
        # Determine canvas format based on source image (after transforms)
        canvas_format = QImage.Format_ARGB32_Premultiplied if image_to_transform.hasAlphaChannel() else QImage.Format_RGB888
        canvas = QImage(render_width, render_height, canvas_format)
        # Check if canvas creation was successful
        if canvas.isNull():
             print("Error: Failed to create render canvas QImage.")
             self.live_view_label.clear()
             return

        # Fill canvas appropriately
        canvas.fill(Qt.white if canvas_format == QImage.Format_RGB888 else Qt.transparent)


        # ---> Pass the stored CROP offsets to the rendering function <---
        # These offsets tell the renderer where the top-left of the current
        # 'image_to_transform' is relative to the original, uncropped image.
        # Ensure crop_offset attributes exist (should be initialized in __init__)
        current_crop_offset_x = getattr(self, 'crop_offset_x', 0)
        current_crop_offset_y = getattr(self, 'crop_offset_y', 0)

        self.render_image_on_canvas(canvas, scaled_image_for_render,
                                    x_start=current_crop_offset_x,
                                    y_start=current_crop_offset_y,
                                    render_scale=render_scale)

        # --- Scale canvas down for display ---
        # Ensure canvas is valid before creating pixmap
        if canvas.isNull():
            print("Error: Render canvas became invalid after rendering.")
            self.live_view_label.clear()
            return

        pixmap = QPixmap.fromImage(canvas)
        if pixmap.isNull():
             print("Error: Failed to create QPixmap from render canvas.")
             self.live_view_label.clear()
             return

        # Scale pixmap to fit the actual live_view_label size
        scaled_pixmap_for_display = pixmap.scaled(
            self.live_view_label.size(), # Use current size of the label widget
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        if scaled_pixmap_for_display.isNull():
             print("Error: Failed to scale final pixmap for display.")
             self.live_view_label.clear()
             return


        # --- Apply Zoom/Pan Transformation for Display ---
        final_display_pixmap = scaled_pixmap_for_display
        if self.live_view_label.zoom_level != 1.0:
            # Ensure the base pixmap is valid before trying to zoom
            if final_display_pixmap.isNull():
                 print("Error: Cannot apply zoom, base pixmap is invalid.")
            else:
                # Use the size of the pixmap *before* zoom for the temporary canvas
                zoomed_pixmap = QPixmap(final_display_pixmap.size())
                zoomed_pixmap.fill(Qt.transparent) # Use transparent background for zoomed canvas

                zoom_painter = QPainter(zoomed_pixmap)
                # Check if painter is active before drawing
                if not zoom_painter.isActive():
                    print("Error: Failed to create QPainter for zooming.")
                else:
                    # Apply pan offset FIRST, then scale around the top-left (0,0)
                    zoom_painter.translate(self.live_view_label.pan_offset)
                    zoom_painter.scale(self.live_view_label.zoom_level, self.live_view_label.zoom_level)
                    zoom_painter.drawPixmap(0, 0, final_display_pixmap) # Draw the *scaled down* pixmap onto the zoomed canvas
                    zoom_painter.end() # End the painter

                    # Check if zoomed_pixmap is still valid after drawing
                    if zoomed_pixmap.isNull():
                        print("Error: Zoomed pixmap became invalid after drawing.")
                    else:
                         final_display_pixmap = zoomed_pixmap # Use the zoomed result


        # --- Set the final pixmap ---
        # Final check before setting
        if final_display_pixmap.isNull():
            print("Error: Final pixmap for display is invalid. Clearing label.")
            self.live_view_label.clear()
        else:
             self.live_view_label.setPixmap(final_display_pixmap)

        # Trigger the paintEvent of LiveViewLabel manually AFTER setting the pixmap
        # This is needed to draw overlays like the crop rectangle preview, markers, etc.
        self.live_view_label.update()
    
    def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
        painter = QPainter(canvas)
        x_offset = (canvas.width() - scaled_image.width()) // 2
        y_offset = (canvas.height() - scaled_image.height()) // 2
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
    
        # Draw the base image
        painter.drawImage(x_offset, y_offset, scaled_image)
        
        if hasattr(self, 'custom_shapes'):
            for shape_data in self.custom_shapes:
                try:
                    shape_type = shape_data.get('type')
                    color = QColor(shape_data.get('color', '#000000')) # Default black
                    thickness = int(shape_data.get('thickness', 1) / render_scale) # Scale thickness
                    if thickness < 1: thickness = 1

                    pen = QPen(color, thickness)
                    painter.setPen(pen)

                    # Calculate canvas coordinates based on stored image coordinates
                    def to_canvas_coords(img_coords):
                        # Inverse of the transformation in finalize_shape_draw
                        # Assumes x_start, y_start are relative to the self.image being rendered
                        # This needs careful checking against how render_image_on_canvas uses x_start/y_start
                        # Let's assume x_start/y_start are 0 for simplicity here if self.image is already cropped
                        canvas_x = x_offset + (img_coords[0] - x_start) * (scaled_image.width() / self.image.width())
                        canvas_y = y_offset + (img_coords[1] - y_start) * (scaled_image.height() / self.image.height())
                        return QPointF(canvas_x, canvas_y)

                    if shape_type == 'line':
                        start_img = shape_data.get('start')
                        end_img = shape_data.get('end')
                        if start_img and end_img:
                            start_canvas = to_canvas_coords(start_img)
                            end_canvas = to_canvas_coords(end_img)
                            painter.drawLine(start_canvas, end_canvas)
                    elif shape_type == 'rectangle':
                        rect_img = shape_data.get('rect') # (x, y, w, h) in image coords
                        if rect_img:
                            x_img, y_img, w_img, h_img = rect_img
                            # Convert top-left and calculate width/height on canvas
                            top_left_canvas = to_canvas_coords((x_img, y_img))
                            # Calculate width/height on canvas by scaling image dimensions
                            w_canvas = w_img * (scaled_image.width() / self.image.width())
                            h_canvas = h_img * (scaled_image.height() / self.image.height())
                            painter.drawRect(QRectF(top_left_canvas, QSizeF(w_canvas, h_canvas)))

                except Exception as e:
                    print(f"Error drawing shape {shape_data}: {e}") # Log error drawing specific shape
        # --- END: Draw Custom Shapes ---


        # --- START: Draw Shape Preview ---
        if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
             try:
                 start_pt = self.current_drawing_shape_preview['start'] # Already in view/label coords
                 end_pt = self.current_drawing_shape_preview['end']     # Already in view/label coords

                 # Get current style for preview
                 preview_color = self.custom_marker_color
                 preview_thickness = int(self.custom_font_size_spinbox.value() / render_scale) # Scale thickness
                 if preview_thickness < 1: preview_thickness = 1

                 preview_pen = QPen(preview_color, preview_thickness)
                 preview_pen.setStyle(Qt.DotLine) # Dashed line for preview
                 painter.setPen(preview_pen)

                 # --- Convert preview points to canvas coordinates ---
                 # (These points are relative to the view label, need scaling to canvas)
                 # This assumes the `transform_point` in LiveViewLabel already handled zoom/pan
                 # We need to map label coords -> canvas coords
                 # The label coords are potentially zoomed/panned, transform_point undoes that.
                 # We need to map the *unzoomed* label coords to the high-res canvas coords.

                 # Map label's top-left (0,0) and bottom-right to canvas space
                 canvas_w_eff = scaled_image.width() # Effective width of image on canvas
                 canvas_h_eff = scaled_image.height()
                 label_w = self.live_view_label.width() # Actual label size
                 label_h = self.live_view_label.height()

                 scale_label_to_canvas_x = canvas_w_eff / label_w if label_w > 0 else 1
                 scale_label_to_canvas_y = canvas_h_eff / label_h if label_h > 0 else 1

                 start_canvas_x = x_offset + start_pt.x() * scale_label_to_canvas_x
                 start_canvas_y = y_offset + start_pt.y() * scale_label_to_canvas_y
                 end_canvas_x = x_offset + end_pt.x() * scale_label_to_canvas_x
                 end_canvas_y = y_offset + end_pt.y() * scale_label_to_canvas_y
                 # --- End coordinate mapping ---


                 if self.drawing_mode == 'line':
                     painter.drawLine(QPointF(start_canvas_x, start_canvas_y), QPointF(end_canvas_x, end_canvas_y))
                 elif self.drawing_mode == 'rectangle':
                     painter.drawRect(QRectF(QPointF(start_canvas_x, start_canvas_y), QPointF(end_canvas_x, end_canvas_y)).normalized()) # Ensure positive width/height

             except Exception as e:
                 print(f"Error drawing shape preview: {e}")
    
        
           
        native_width = self.image.width() if self.image and self.image.width() > 0 else 1
        native_height = self.image.height() if self.image and self.image.height() > 0 else 1

        # Scale factor mapping native coordinates to the *scaled_image* dimensions
        scale_native_to_scaled_x = scaled_image.width() / native_width
        scale_native_to_scaled_y = scaled_image.height() / native_height

        # --- Draw Image 1 Overlay ---
        if hasattr(self, 'image1_original') and hasattr(self, 'image1_position') and not self.image1_original.isNull():
            try:
                scale_factor_overlay = self.image1_resize_slider.value() / 100.0
                target_overlay_w = int(self.image1_original.width() * scale_factor_overlay)
                target_overlay_h = int(self.image1_original.height() * scale_factor_overlay)
                self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())

                if target_overlay_w > 0 and target_overlay_h > 0:
                    # Resize the *original* overlay image
                    resized_overlay1 = self.image1_original.scaled(
                        target_overlay_w, target_overlay_h,
                        Qt.KeepAspectRatio, Qt.SmoothTransformation
                    )
                    if not resized_overlay1.isNull():
                        # Get NATIVE offsets stored
                        native_offset_x1 = self.image1_position[0]
                        native_offset_y1 = self.image1_position[1]

                        # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                        # Scale the NATIVE offset according to how the base image was scaled
                        # Then add the centering offset of the base image on the canvas
                        canvas_draw_x = x_offset + native_offset_x1 * scale_native_to_scaled_x
                        canvas_draw_y = y_offset + native_offset_y1 * scale_native_to_scaled_y

                        painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay1)
            except Exception as e:
                print(f"Error rendering overlay image 1: {e}")


        # --- Draw Image 2 Overlay ---
        if hasattr(self, 'image2_original') and hasattr(self, 'image2_position') and not self.image2_original.isNull():
            try:
                scale_factor_overlay = self.image2_resize_slider.value() / 100.0
                self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                target_overlay_w = int(self.image2_original.width() * scale_factor_overlay)
                target_overlay_h = int(self.image2_original.height() * scale_factor_overlay)

                if target_overlay_w > 0 and target_overlay_h > 0:
                    resized_overlay2 = self.image2_original.scaled(
                        target_overlay_w, target_overlay_h,
                        Qt.KeepAspectRatio, Qt.SmoothTransformation
                    )
                    if not resized_overlay2.isNull():
                        native_offset_x2 = self.image2_position[0]
                        native_offset_y2 = self.image2_position[1]

                        # <<< FIX HERE: Calculate position on PREVIEW canvas >>>
                        canvas_draw_x = x_offset + native_offset_x2 * scale_native_to_scaled_x
                        canvas_draw_y = y_offset + native_offset_y2 * scale_native_to_scaled_y

                        painter.drawImage(int(canvas_draw_x), int(canvas_draw_y), resized_overlay2)
            except Exception as e:
                print(f"Error rendering overlay image 2: {e}")
    
        # Get the selected font settings
        font = QFont(self.font_combo_box.currentFont().family(), self.font_size_spinner.value() * render_scale)
        font_color = self.font_color if hasattr(self, 'font_color') else QColor(0, 0, 0)  # Default to black if no color selected
    
        painter.setFont(font)
        painter.setPen(font_color)  # Set the font color
    
        # Measure text height for vertical alignment
        font_metrics = painter.fontMetrics()
        text_height = font_metrics.height()
        line_padding = 5 * render_scale  # Space between text and line
        
        y_offset_global = text_height / 4
    
        # Draw the left markers (aligned right)
        for y_pos, marker_value in self.left_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f"{marker_value} ⎯ "  ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
        
        
        # Draw the right markers (aligned left)
        for y_pos, marker_value in self.right_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f" ⎯ {marker_value}" ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.right_marker_shift_added),# + line_padding),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
                
    
        # Draw the top markers (if needed)
        for x_pos, top_label in self.top_markers:
            x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width())
            if 0 <= x_pos_cropped <= scaled_image.width():
                text = f"{top_label}"
                painter.save()
                text_width = font_metrics.horizontalAdvance(text)
                text_height= font_metrics.height()
                label_x = x_offset + x_pos_cropped 
                label_y = y_offset + self.top_marker_shift + self.top_marker_shift_added 
                painter.translate(int(label_x), int(label_y))
                painter.rotate(self.font_rotation)
                painter.drawText(0,int(y_offset_global), f"{top_label}")
                painter.restore()
    
        # Draw guide lines
        if draw_guides and self.show_guides_checkbox.isChecked():
            pen = QPen(Qt.red, 2 * render_scale)
            painter.setPen(pen)
            center_x = canvas.width() // 2
            center_y = canvas.height() // 2
            painter.drawLine(center_x, 0, center_x, canvas.height())  # Vertical line
            painter.drawLine(0, center_y, canvas.width(), center_y)  # Horizontal line
            
        
        
        
        # Draw the protein location marker (*)
        if hasattr(self, "protein_location") and self.run_predict_MW == False:
            x, y = self.protein_location
            text = "⎯⎯"
            text_width = font_metrics.horizontalAdvance(text)
            text_height= font_metrics.height()
            painter.drawText(
                int(x * render_scale - text_width / 2),  #Currently left edge # FOR Center horizontally use int(x * render_scale - text_width / 2)
                int(y * render_scale + text_height / 4),  # Center vertically
                text
            )
        for marker_tuple in getattr(self, "custom_markers", []):
            try:
                # -->> Unpack all 8 elements using the correct order <<--
                # The tuple structure is (x, y, text, color, font_family, font_size, is_bold, is_italic)
                x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple

                # Calculate position on canvas, scaling relative to the CROPPED image extent
                # Ensure original image dimensions are valid for scaling calculation
                orig_img_w = self.image.width() if self.image and self.image.width() > 0 else 1
                orig_img_h = self.image.height() if self.image and self.image.height() > 0 else 1
                # Dimensions of the image drawn on the canvas (after potential cropping/scaling)
                img_w_on_canvas = scaled_image.width()
                img_h_on_canvas = scaled_image.height()

                # Calculate the marker's position relative to the canvas origin (x_offset, y_offset)
                # Scale based on how the original coordinates map to the scaled_image dimensions
                x_pos_on_canvas = x_offset + (x_pos - x_start) * (img_w_on_canvas / orig_img_w)
                y_pos_on_canvas = y_offset + (y_pos - y_start) * (img_h_on_canvas / orig_img_h)

                # Only draw if within the visible scaled image bounds on the canvas
                # Check against the area occupied by scaled_image on the canvas
                if (x_offset <= x_pos_on_canvas <= x_offset + img_w_on_canvas and
                    y_offset <= y_pos_on_canvas <= y_offset + img_h_on_canvas):

                    # -->> Create and configure font using unpacked values <<--
                    # Ensure font_size is a valid integer before scaling
                    try:
                        # Use int() directly on font_size which should already be an int/float
                        scaled_font_size = int(int(font_size) * render_scale)
                        if scaled_font_size < 1: scaled_font_size = 1 # Ensure minimum size 1
                    except (ValueError, TypeError):
                        scaled_font_size = int(12 * render_scale) # Default scaled size (e.g., 12pt scaled)

                    # Ensure font_family is a string
                    marker_font = QFont(str(font_family), scaled_font_size)
                    # Ensure bold/italic are booleans and apply them
                    marker_font.setBold(bool(is_bold))
                    marker_font.setItalic(bool(is_italic))

                    # Set the configured font on the painter
                    painter.setFont(marker_font)

                    # -->> Set the color, ensuring it's a QColor <<--
                    if isinstance(color, str):
                        current_color = QColor(color) # Create QColor from name/hex string
                    elif isinstance(color, QColor):
                        current_color = color # It's already a QColor
                    else:
                         current_color = Qt.black # Fallback color
                    painter.setPen(current_color)

                    # Calculate text dimensions for alignment using the specific marker's font
                    font_metrics_custom = painter.fontMetrics()
                    # Ensure marker_text is a string
                    text_to_draw = str(marker_text)
                    text_width = font_metrics_custom.horizontalAdvance(text_to_draw)
                    text_height = font_metrics_custom.height()
                    # Approximate vertical offset to center text around the point's Y coordinate
                    y_offset_global_custom = text_height / 4

                    # Draw text centered horizontally at its calculated canvas position
                    painter.drawText(
                        int(x_pos_on_canvas - text_width / 2),          # Center horizontally
                        int(y_pos_on_canvas + y_offset_global_custom),  # Center vertically around baseline
                        text_to_draw                                    # Draw the text
                    )
            except (ValueError, TypeError, IndexError) as e:
                # Catch errors during unpacking or processing of a single marker tuple
                import traceback
                traceback.print_exc() # Print stack trace for debugging tuple issues
            
            
    

        # Draw the grid (if enabled)
        grid_size = self.grid_size_input.value() * render_scale
        if grid_size > 0: # Only draw if grid size is valid
            pen = QPen(Qt.red)
            pen.setStyle(Qt.DashLine)
            # You might want a slightly less prominent color/width for the grid
            # pen.setColor(QColor(200, 0, 0, 128)) # Semi-transparent red example
            # pen.setWidth(1) # Thinner line example
            painter.setPen(pen)

            if self.show_grid_checkbox_x.isChecked():
                # Draw vertical grid lines (for X snapping)
                for x in range(0, canvas.width(), grid_size):
                    painter.drawLine(x, 0, x, canvas.height())

            if self.show_grid_checkbox_y.isChecked():
                # Draw horizontal grid lines (for Y snapping)
                for y in range(0, canvas.height(), grid_size):
                    painter.drawLine(0, y, canvas.width(), y)
                
        painter.end()


     
    def crop_image(self):
        """Function to crop the current image."""
        if not self.image:
            return None
    
        # Get crop percentage from sliders
        x_start_percent = 0
        x_end_percent = 100
        y_start_percent = 0
        y_end_percent = 100
    
        # Calculate crop boundaries
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure cropping is valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            return None
    
        # Crop the image
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
        return cropped_image
    
    # Modify align_image and update_crop to preserve settings
    def reset_align_image(self):
        self.orientation_slider.setValue(0)
        try:
            if self.image: # Check if image exists after cropping
                self._update_preview_label_size()
        except Exception as e:
            # Fallback size?
            self._update_preview_label_size()


        self.update_live_view() # Final update with corrected markers and image
        
        
    def align_image(self):
        """
        Aligns (rotates) the main self.image based on the orientation slider value.
        This modifies the high-resolution self.image directly.
        """
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Rotation Error", "No image loaded to rotate.")
            return

        # Get the orientation value from the slider
        angle = float(self.orientation_slider.value() / 20.0) # Use float division

        if abs(angle) < 0.01: # Threshold to avoid unnecessary processing for tiny angles
            # print("Debug: Rotation angle negligible, skipping.") # Optional debug
            # Reset slider if needed, but don't modify image
            self.orientation_slider.setValue(0)
            return

        # --- Save state BEFORE modifying the image ---
        self.save_state()

        # Disable temporary guides before applying permanent change
        self.draw_guides = False
        if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)
        # Keep grid settings as they are independent of rotation


        try:
            # Perform rotation calculation
            transform = QTransform()
            # Rotate around the center of the *current* image
            current_width = self.image.width()
            current_height = self.image.height()
            transform.translate(current_width / 2.0, current_height / 2.0)
            transform.rotate(angle)
            transform.translate(-current_width / 2.0, -current_height / 2.0)

            # Apply the transformation to get the high-resolution rotated image
            # Qt.SmoothTransformation provides better quality
            # The resulting image might be larger to accommodate rotated corners
            rotated_image_high_res = self.image.transformed(transform, Qt.SmoothTransformation)

            if rotated_image_high_res.isNull():
                raise ValueError("Image transformation resulted in an invalid (null) image.")

            # --- Directly update the main image object ---
            self.image = rotated_image_high_res
            self.is_modified = True # Mark as modified

            # --- Update backup images to reflect the new state ---
            # Rotation invalidates previous padding
            self.image_before_padding = None # Or self.image.copy() if padding should persist conceptually
            self.image_padded = False
            # Contrast backups should also reflect the rotated image
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

            # --- Reset the orientation slider ---
            self.orientation_slider.setValue(0)

            # --- Update UI elements based on the new image dimensions ---
            # This will adjust the preview label size and marker slider ranges
            self._update_preview_label_size()
            self._update_marker_slider_ranges() # Important after potential size change
            self._update_status_bar()

            # --- Let update_live_view handle the rendering ---
            # This will correctly scale the new high-res self.image for display
            # and draw markers/overlays relative to the new image.
            self.update_live_view()

        except Exception as e:
            QMessageBox.critical(self, "Rotation Error", f"Failed to rotate image: {e}")
            traceback.print_exc()
            # Attempt to revert state if possible (though save_state was called)
            # self.undo_action_m() # Reverting might be complex if backups were already updated
        
        self.crop_x_start_slider.setEnabled(False) 
        self.crop_x_end_slider.setEnabled(False) 
        self.crop_y_start_slider.setEnabled(False) 
        self.crop_y_end_slider.setEnabled(False) 
    
    def _update_marker_slider_ranges(self):
        """Recalculates and updates ONLY the ranges for the marker offset sliders."""
        # --- Calculate new render dimensions and target ranges ---
        if not self.image or self.image.isNull() or not self.live_view_label:
            min_range, max_range_w, max_range_h = -100, 1000, 800 # Defaults
        else:
            try:
                render_scale = 3
                view_width = self.live_view_label.width(); view_height = self.live_view_label.height()
                if view_width <= 0: view_width = 600 # Fallback
                if view_height <= 0: view_height = 400 # Fallback
                render_width = view_width * render_scale
                render_height = view_height * render_scale
                margin = 100
                min_range = -margin
                max_range_w = max(margin, int(render_width) + margin)
                max_range_h = max(margin, int(render_height) + margin)
            except Exception as e:
                print(f"Warning: Error calculating slider ranges: {e}. Using defaults.")
                min_range, max_range_w, max_range_h = -100, 1000, 800

        # Update internal range storage (good practice)
        self.left_slider_range = [min_range, max_range_w]
        self.right_slider_range = [min_range, max_range_w]
        self.top_slider_range = [min_range, max_range_h]

        # --- Update Left Slider Range (Value is untouched here) ---
        if hasattr(self, 'left_padding_slider'):
            current_val_left = self.left_padding_slider.value() # Store current value
            self.left_padding_slider.blockSignals(True)
            self.left_padding_slider.setRange(min_range, max_range_w)
            self.left_padding_slider.setValue(current_val_left) # Re-apply value (Qt clamps if needed)
            self.left_padding_slider.blockSignals(False)
            # NO update to self.left_marker_shift_added

        # --- Update Right Slider Range (Value is untouched here) ---
        if hasattr(self, 'right_padding_slider'):
            current_val_right = self.right_padding_slider.value() # Store current value
            self.right_padding_slider.blockSignals(True)
            self.right_padding_slider.setRange(min_range, max_range_w)
            self.right_padding_slider.setValue(current_val_right) # Re-apply value (Qt clamps if needed)
            self.right_padding_slider.blockSignals(False)
            # NO update to self.right_marker_shift_added

        # --- Update Top Slider Range (Value is untouched here) ---
        if hasattr(self, 'top_padding_slider'):
            current_val_top = self.top_padding_slider.value() # Store current value
            self.top_padding_slider.blockSignals(True)
            self.top_padding_slider.setRange(min_range, max_range_h)
            self.top_padding_slider.setValue(current_val_top) # Re-apply value (Qt clamps if needed)
            self.top_padding_slider.blockSignals(False)
            # NO update to self.top_marker_shift_added
            
        
            
    def update_crop(self):
        """
        Update the image based on the defined crop rectangle.
        Adjusts marker and shape coordinates relative to the new image origin.
        """
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Crop Error", "No image loaded to crop.")
            return

        # Check if a crop rectangle has been defined
        if not self.crop_rectangle_coords:
            if self.crop_rectangle_mode:
                QMessageBox.information(self, "Crop", "Please finalize drawing the crop rectangle before applying.")
            else:
                QMessageBox.information(self, "Crop", "Please draw a crop rectangle first using the 'Draw Crop Rectangle' button.")
            return

        try:
            img_x, img_y, img_w, img_h = self.crop_rectangle_coords

            # Clamp coordinates to be within the current image bounds
            current_width = self.image.width()
            current_height = self.image.height()

            # Calculate the actual crop origin and dimensions, ensuring they are valid
            crop_x_start = max(0, img_x)
            crop_y_start = max(0, img_y)
            # Calculate potential end points based on original width/height
            crop_x_end_potential = crop_x_start + img_w
            crop_y_end_potential = crop_y_start + img_h
            # Clamp end points to image boundaries
            crop_x_end_clamped = min(current_width, crop_x_end_potential)
            crop_y_end_clamped = min(current_height, crop_y_end_potential)
            # Calculate final width and height based on clamped start/end
            crop_width = crop_x_end_clamped - crop_x_start
            crop_height = crop_y_end_clamped - crop_y_start

            if crop_width <= 0 or crop_height <= 0:
                 QMessageBox.warning(self, "Crop Error", "Calculated crop area has zero width or height. Aborting.")
                 # Clear invalid coords and preview
                 self.crop_rectangle_coords = None
                 self.live_view_label.clear_crop_preview()
                 self.cancel_rectangle_crop_mode() # Exit the mode
                 return

            print(f"Applying Crop: x={crop_x_start}, y={crop_y_start}, w={crop_width}, h={crop_height}") # Debug

            # --- Save state BEFORE modifying image and markers ---
            self.save_state()

            # --- Perform the actual crop ---
            cropped_image = self.image.copy(crop_x_start, crop_y_start, crop_width, crop_height)

            if cropped_image.isNull():
                raise ValueError("QImage.copy failed for cropping.")

            # --- Adjust marker positions relative to the new crop origin ---
            # The new origin (0,0) of cropped_image corresponds to
            # (crop_x_start, crop_y_start) in the original image's coordinates.
            # Any original coordinate (x_old, y_old) becomes (x_old - crop_x_start, y_old - crop_y_start)

            new_left_markers = []
            for y_old, label in getattr(self, 'left_markers', []):
                # Check if the original Y was within the vertical crop region
                if crop_y_start <= y_old < crop_y_start + crop_height:
                    y_new = y_old - crop_y_start
                    new_left_markers.append((y_new, label))
            self.left_markers = new_left_markers

            new_right_markers = []
            for y_old, label in getattr(self, 'right_markers', []):
                if crop_y_start <= y_old < crop_y_start + crop_height:
                    y_new = y_old - crop_y_start
                    new_right_markers.append((y_new, label))
            self.right_markers = new_right_markers

            new_top_markers = []
            for x_old, label in getattr(self, 'top_markers', []):
                # Check if the original X was within the horizontal crop region
                if crop_x_start <= x_old < crop_x_start + crop_width:
                    x_new = x_old - crop_x_start
                    new_top_markers.append((x_new, label))
            self.top_markers = new_top_markers

            new_custom_markers = []
            if hasattr(self, "custom_markers"):
                for marker_data in self.custom_markers:
                    try:
                        x_old, y_old = marker_data[0], marker_data[1]
                        # Check if the original point was within the crop rectangle
                        if crop_x_start <= x_old < crop_x_start + crop_width and \
                           crop_y_start <= y_old < crop_y_start + crop_height:
                            x_new = x_old - crop_x_start
                            y_new = y_old - crop_y_start
                            # Create new list/tuple keeping other data
                            updated_marker = [x_new, y_new] + list(marker_data[2:])
                            new_custom_markers.append(updated_marker)
                    except (IndexError, TypeError):
                        print(f"Warning: Skipping malformed custom marker during crop adjustment: {marker_data}")
                        continue # Skip malformed marker data
            self.custom_markers = new_custom_markers

            # --- Adjust custom shapes (Lines and Rectangles) ---
            new_custom_shapes = []
            if hasattr(self, "custom_shapes"):
                for shape_data in self.custom_shapes:
                    try:
                        stype = shape_data.get('type')
                        adjusted_shape = shape_data.copy() # Start with a copy
                        valid_after_crop = False

                        if stype == 'line':
                            sx_old, sy_old = shape_data['start']
                            ex_old, ey_old = shape_data['end']
                            # Simple check: Keep line only if *both* original points were inside crop bounds
                            # A more complex clipping algorithm could be implemented if needed
                            if (crop_x_start <= sx_old < crop_x_start + crop_width and
                                crop_y_start <= sy_old < crop_y_start + crop_height and
                                crop_x_start <= ex_old < crop_x_start + crop_width and
                                crop_y_start <= ey_old < crop_y_start + crop_height):
                                adjusted_shape['start'] = (sx_old - crop_x_start, sy_old - crop_y_start)
                                adjusted_shape['end'] = (ex_old - crop_x_start, ey_old - crop_y_start)
                                valid_after_crop = True

                        elif stype == 'rectangle':
                            rx_old, ry_old, rw_old, rh_old = shape_data['rect']
                            # Calculate intersection of original rectangle with crop area
                            overlap_x_start = max(crop_x_start, rx_old)
                            overlap_y_start = max(crop_y_start, ry_old)
                            overlap_x_end = min(crop_x_start + crop_width, rx_old + rw_old)
                            overlap_y_end = min(crop_y_start + crop_height, ry_old + rh_old)

                            # Check if there is a valid intersection
                            if overlap_x_end > overlap_x_start and overlap_y_end > overlap_y_start:
                                # Calculate new rect coords relative to the crop origin
                                rx_new = overlap_x_start - crop_x_start
                                ry_new = overlap_y_start - crop_y_start
                                rw_new = overlap_x_end - overlap_x_start
                                rh_new = overlap_y_end - overlap_y_start
                                adjusted_shape['rect'] = (rx_new, ry_new, rw_new, rh_new)
                                valid_after_crop = True

                        if valid_after_crop:
                            new_custom_shapes.append(adjusted_shape)

                    except (KeyError, IndexError, TypeError, ValueError):
                        print(f"Warning: Skipping malformed custom shape during crop adjustment: {shape_data}")
                        continue # Skip malformed shape data
            self.custom_shapes = new_custom_shapes
            # --- End Shape Adjustment ---

            # --- Update the main image and backups ---
            self.image = cropped_image
            self.is_modified = True
            # Backups should reflect the newly cropped state
            self.image_before_padding = self.image.copy() # Padding needs reapplication
            self.image_contrasted = self.image.copy()      # Contrast baseline resets
            self.image_before_contrast = self.image.copy() # Contrast baseline resets
            self.image_padded = False # Reset padding flag as crop removed it

            # --- Reset Crop UI elements ---
            self.crop_rectangle_coords = None
            self.live_view_label.clear_crop_preview()
            self.cancel_rectangle_crop_mode() # Exit the mode fully

            # --- Update UI ---
            self._update_preview_label_size() # Adjust label size for new image dimensions
            self._update_status_bar()         # Update status bar info
            self._update_marker_slider_ranges()
            self.update_live_view()           # Refresh the display

        except Exception as e:
            QMessageBox.critical(self, "Crop Error", f"An error occurred during cropping: {e}")
            traceback.print_exc()
            # Clear potentially invalid state
            self.crop_rectangle_coords = None
            self.live_view_label.clear_crop_preview()
            self.cancel_rectangle_crop_mode()
        
    def update_skew(self):
        self.save_state()
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            return
    
        # Apply the transformation
        # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

 
        self.image = self.image.transformed(transform, Qt.SmoothTransformation)
        self.taper_skew_slider.setValue(0)

    
        
    def save_image(self):
        self.draw_guides = False
        self.show_guides_checkbox.setChecked(False)
        self.show_grid_checkbox_x.setChecked(False)
        self.show_grid_checkbox_y.setChecked(False)
        self.update_live_view()
        """Saves the original image, the modified image (rendered view), and configuration."""
        if not self.image_master: # Check if an initial image was ever loaded/pasted
             QMessageBox.warning(self, "Error", "No image data to save.")
             return False # Indicate save failed or was aborted

        self.is_modified = False # Mark as saved

        options = QFileDialog.Options()
        # Suggest a filename based on the loaded image or a default
        suggested_name = ""
        if self.image_path:
            base = os.path.splitext(os.path.basename(self.image_path))[0]
            # Remove common suffixes if they exist from previous saves
            base = base.replace("_original", "").replace("_modified", "")
            suggested_name = f"{base}" # Suggest PNG for modified view
        elif hasattr(self, 'base_name') and self.base_name:
            suggested_name = f"{base}"
        else:
            suggested_name = "untitled_image"

        save_dir = os.path.dirname(self.image_path) if self.image_path else "" # Suggest save in original dir

        # --- Get the base save path from the user ---
        base_save_path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save Image Base Name", os.path.join(save_dir, suggested_name),
            "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
            options=options
        )
        # --- End getting base save path ---

        if not base_save_path: # User cancelled
            self.is_modified = True # Revert saved status if cancelled
            return False # Indicate save cancelled

        # Determine paths based on the base path
        base_name_nosuffix = (os.path.splitext(base_save_path)[0]).replace("_original", "").replace("_modified", "")
        # Determine the desired suffix based on the selected filter or default to .png
        if "png" in selected_filter.lower():
             suffix = ".png"
        elif "tif" in selected_filter.lower():
             suffix = ".tif"
        elif "jpg" in selected_filter.lower() or "jpeg" in selected_filter.lower():
             suffix = ".jpg"
        elif "bmp" in selected_filter.lower():
             suffix = ".bmp"
        else: # Default or All Files
             suffix = os.path.splitext(base_save_path)[1] # Keep user's suffix if provided
             if not suffix: suffix = ".png" # Default to png if no suffix given

        original_save_path = f"{base_name_nosuffix}_original{suffix}"
        modified_save_path = f"{base_name_nosuffix}_modified{suffix}"
        config_save_path = f"{base_name_nosuffix}_config.txt"

        # --- Save original image (using self.image_master) ---
        img_to_save_orig = self.image 
        if img_to_save_orig and not img_to_save_orig.isNull():
            save_format_orig = suffix.replace(".", "").upper() # Determine format from suffix
            if save_format_orig == "TIF": save_format_orig = "TIFF" # Use standard TIFF identifier
            elif save_format_orig == "JPEG": save_format_orig = "JPG"

            # QImage.save attempts to match format, quality is for lossy formats
            quality = 95 if save_format_orig in ["JPG", "JPEG"] else -1 # Default quality (-1) for lossless

            if not img_to_save_orig.save(original_save_path, format=save_format_orig if save_format_orig else None, quality=quality):
                QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "Original master image data is missing.")

        # --- Save modified image (Rendered view - likely RGB) ---
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        # Use ARGB32 for rendering to support potential transparency (good for PNG)
        # Use RGB888 if saving as JPG/BMP which don't support alpha well.
        save_format_mod = suffix.replace(".", "").upper()
        if save_format_mod in ["JPG", "JPEG", "BMP"]:
            canvas_format = QImage.Format_ARGB32_Premultiplied
            fill_color = Qt.transparent # Use white background for opaque formats
        else: # PNG, TIFF
            canvas_format = QImage.Format_ARGB32_Premultiplied # Good for rendering quality with alpha
            fill_color = Qt.transparent # Use transparent background

        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, canvas_format
        )
        high_res_canvas.fill(fill_color)

        # Use current self.image for rendering the modified view
        if self.image and not self.image.isNull():
             scaled_image_mod = self.image.scaled(
                 high_res_canvas_width, high_res_canvas_height,
                 Qt.KeepAspectRatio, Qt.SmoothTransformation)

             # Render onto the high-res canvas (render_image_on_canvas draws markers etc.)
             # Pass draw_guides=False to avoid saving them
             self.render_image_on_canvas(
                 high_res_canvas, scaled_image_mod,
                 x_start=0, y_start=0, # Rendering is relative to current self.image extent
                 render_scale=render_scale, draw_guides=False
             )

             # Save the rendered canvas
             quality_mod = 95 if save_format_mod in ["JPG", "JPEG"] else -1
             if not high_res_canvas.save(modified_save_path, format=save_format_mod if save_format_mod else None, quality=quality_mod):
                 QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "No current image to render for saving modified view.")


        # --- Save configuration ---
        config_data = self.get_current_config()
        try:
            with open(config_save_path, "w") as config_file:
                json.dump(config_data, config_file, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
            return False # Indicate save failure


        QMessageBox.information(self, "Saved", f"Files saved successfully:\n- {os.path.basename(original_save_path)}\n- {os.path.basename(modified_save_path)}\n- {os.path.basename(config_save_path)}")

        # Update window title (optional, remove base_name if confusing)
        self.setWindowTitle(f"{self.window_title}::{base_name_nosuffix}")
        return True # Indicate successful save
            
    def save_image_svg(self):
        """Save the processed image along with markers, labels, and custom shapes in SVG format."""
        if not self.image or self.image.isNull(): # Check current image validity
            QMessageBox.warning(self, "Warning", "No image to save.")
            return

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image as SVG for MS Word/Vector Editing", "", "SVG Files (*.svg)", options=options
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".svg"): # Ensure .svg extension
            file_path += ".svg"

        # --- Determine Render/Canvas Dimensions ---
        # Use the current image dimensions directly as the base SVG size
        # This avoids scaling issues if the view label size is different.
        # Markers and shapes will be positioned relative to this size.
        svg_width = self.image.width()
        svg_height = self.image.height()

        if svg_width <= 0 or svg_height <= 0:
             QMessageBox.warning(self, "Warning", "Invalid image dimensions for SVG.")
             return

        # --- Create SVG Drawing Object ---
        dwg = svgwrite.Drawing(file_path, profile='tiny', size=(f"{svg_width}px", f"{svg_height}px"))
        # Set viewbox to match image dimensions for correct coordinate system
        dwg.viewbox(0, 0, svg_width, svg_height)

        # --- Embed the Base Image ---
        try:
            # Convert the QImage to a base64-encoded PNG for embedding
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            # Save the *current* self.image (which might be cropped/transformed)
            if not self.image.save(buffer, "PNG"):
                raise IOError("Failed to save image to PNG buffer for SVG.")
            image_data = base64.b64encode(buffer.data()).decode('utf-8')
            buffer.close()

            # Embed the image at position (0, 0) with original dimensions
            dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}",
                              insert=(0, 0),
                              size=(f"{svg_width}px", f"{svg_height}px")))
        except Exception as e:
            QMessageBox.critical(self, "SVG Error", f"Failed to embed base image: {e}")
            return # Stop if base image fails

        # --- Define marker/label font style for SVG ---
        # Use the standard marker font settings
        svg_font_family = self.font_family
        svg_font_size_px = f"{self.font_size}px" # Use pixels for SVG consistency
        svg_font_color = self.font_color.name() if self.font_color else "#000000"

        # Calculate horizontal offset for left/right markers based on font size
        # Use QFontMetrics for accurate width calculation
        try:
             qfont_for_metrics = QFont(svg_font_family, self.font_size)
             font_metrics = QFontMetrics(qfont_for_metrics)
             # Use a representative character like 'm' or average width if needed
             avg_char_width = font_metrics.averageCharWidth()
             horizontal_offset = avg_char_width * 0.5 # Small offset from edge
             vertical_offset_adjust = font_metrics.ascent() * 0.75 # Adjustment for vertical alignment
        except Exception:
             horizontal_offset = 5 # Fallback offset
             vertical_offset_adjust = self.font_size * 0.75 # Fallback adjustment


        # --- Add Left Markers ---
        left_marker_x_pos = self.left_marker_shift_added # Use the absolute offset
        for y_pos, text in getattr(self, "left_markers", []):
            final_text = f"{text}" # Remove the line "⎯" for cleaner SVG text
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(left_marker_x_pos - horizontal_offset, y_pos + vertical_offset_adjust),
                    fill=svg_font_color,
                    font_family=svg_font_family,
                    font_size=svg_font_size_px,
                    text_anchor="end" # Align text right, ending at the x position
                )
            )

        # --- Add Right Markers ---
        right_marker_x_pos = self.right_marker_shift_added # Use the absolute offset
        for y_pos, text in getattr(self, "right_markers", []):
            final_text = f"{text}" # Remove the line "⎯"
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(right_marker_x_pos + horizontal_offset, y_pos + vertical_offset_adjust),
                    fill=svg_font_color,
                    font_family=svg_font_family,
                    font_size=svg_font_size_px,
                    text_anchor="start" # Align text left, starting at the x position
                )
            )

        # --- Add Top Markers ---
        top_marker_y_pos = self.top_marker_shift_added # Use the absolute offset
        for x_pos, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x_pos, top_marker_y_pos + vertical_offset_adjust), # Apply vertical adjust
                    fill=svg_font_color,
                    font_family=svg_font_family,
                    font_size=svg_font_size_px,
                    text_anchor="middle", # Center text horizontally over the x position
                    # Apply rotation around the insertion point
                    transform=f"rotate({self.font_rotation}, {x_pos}, {top_marker_y_pos + vertical_offset_adjust})"
                )
            )

        # --- Add Custom Markers ---
        for marker_tuple in getattr(self, "custom_markers", []):
            try:
                # Default values for optional elements
                is_bold = False
                is_italic = False

                # Unpack based on length for backward compatibility
                if len(marker_tuple) == 8:
                    x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple
                elif len(marker_tuple) == 6:
                    x_pos, y_pos, marker_text, color, font_family, font_size = marker_tuple
                else:
                    continue # Skip invalid marker data

                # Prepare SVG attributes
                text_content = str(marker_text)
                fill_color = QColor(color).name() if isinstance(color, QColor) else str(color) # Ensure hex/name
                font_family_svg = str(font_family)
                font_size_svg = f"{int(font_size)}px" if isinstance(font_size, (int, float)) else "12px"
                font_weight_svg = "bold" if bool(is_bold) else "normal"
                font_style_svg = "italic" if bool(is_italic) else "normal"

                # Adjust vertical position slightly for better alignment if needed
                # This might require font metrics specific to the marker's font
                # For simplicity, using the standard marker offset for now
                y_pos_adjusted = y_pos + vertical_offset_adjust

                # Add SVG text element, centered at the marker's coordinates
                dwg.add(
                    dwg.text(
                        text_content,
                        insert=(x_pos, y_pos_adjusted), # Position text anchor at the coordinate
                        fill=fill_color,
                        font_family=font_family_svg,
                        font_size=font_size_svg,
                        font_weight=font_weight_svg,
                        font_style=font_style_svg,
                        text_anchor="middle", # Center horizontally
                        dominant_baseline="central" # Attempt vertical centering (support varies)
                    )
                )
            except Exception as e:
                 print(f"Warning: Skipping invalid custom marker during SVG export: {e}")
                 # import traceback; traceback.print_exc() # Uncomment for detailed debug

        # --- START: Add Custom Shapes (Lines and Rectangles) ---
        for shape_data in getattr(self, "custom_shapes", []):
             try:
                 shape_type = shape_data.get('type')
                 color_str = shape_data.get('color', '#000000') # Default black
                 thickness = int(shape_data.get('thickness', 1))
                 if thickness < 1: thickness = 1

                 if shape_type == 'line':
                     start_coords = shape_data.get('start')
                     end_coords = shape_data.get('end')
                     if start_coords and end_coords:
                         dwg.add(dwg.line(start=start_coords,
                                          end=end_coords,
                                          stroke=color_str,
                                          stroke_width=thickness))
                 elif shape_type == 'rectangle':
                     rect_coords = shape_data.get('rect') # (x, y, w, h)
                     if rect_coords:
                         x, y, w, h = rect_coords
                         dwg.add(dwg.rect(insert=(x, y),
                                          size=(f"{w}px", f"{h}px"),
                                          stroke=color_str,
                                          stroke_width=thickness,
                                          fill="none")) # No fill for outline rectangle
             except Exception as e:
                 print(f"Warning: Skipping invalid custom shape during SVG export: {e}")
        # --- END: Add Custom Shapes ---

        # --- Save the SVG file ---
        try:
            dwg.save()
            QMessageBox.information(self, "Success", f"Image and annotations saved as SVG at\n{file_path}")
            self.is_modified = False # Mark as saved if successful
        except Exception as e:
            QMessageBox.critical(self, "SVG Save Error", f"Failed to save SVG file:\n{e}")
    
    
    def copy_to_clipboard(self):
        """Copy the image via a temporary file, cleaning up on exit."""
        # --- Initial Check ---
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Warning", "No image to copy.")
            return
    
        # --- Create the high-resolution rendered canvas ---
        # (Canvas creation, rendering, straight alpha conversion - KEEP AS IS)
        render_scale = 3
        try:
            label_width = self.live_view_label.width(); label_height = self.live_view_label.height()
            if label_width <= 0: label_width = 500 # Fallback
            if label_height <= 0: label_height = 500 # Fallback
            high_res_canvas_width = label_width * render_scale
            high_res_canvas_height = label_height * render_scale
            if high_res_canvas_width <= 0 or high_res_canvas_height <= 0: raise ValueError("Invalid canvas size")
        except Exception as e:
            QMessageBox.critical(self, "Internal Error", f"Could not calculate render dimensions: {e}")
            return
    
        render_canvas = QImage(high_res_canvas_width, high_res_canvas_height, QImage.Format_ARGB32_Premultiplied)
        render_canvas.fill(Qt.transparent)
    
        if self.image and not self.image.isNull():
            x_start, y_start = 0, 0 # Assuming self.image is pre-cropped
            # *** Robust Check ***
            if not self.image or self.image.isNull():
                 QMessageBox.critical(self, "Internal Error", "Image became invalid before scaling.")
                 self.cleanup_temp_clipboard_file() # Attempt cleanup
                 return
            try:
                scaled_image = self.image.scaled(high_res_canvas_width, high_res_canvas_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                if scaled_image.isNull(): raise ValueError("Scaling failed.")
            except Exception as e:
                QMessageBox.critical(self, "Copy Error", f"Failed to scale image: {e}")
                self.cleanup_temp_clipboard_file()
                return
    
            self.render_image_on_canvas(render_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False)
    
            final_canvas_for_clipboard = render_canvas
            if render_canvas.hasAlphaChannel():
                straight_alpha_canvas = render_canvas.convertToFormat(QImage.Format_ARGB32)
                if not straight_alpha_canvas.isNull():
                    final_canvas_for_clipboard = straight_alpha_canvas
            # --- End Rendering ---
    
            # --- Save to Temporary PNG File ---
            try:
                self.cleanup_temp_clipboard_file() # Clean up previous one first
    
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False, mode='wb') as temp_file:
                    self.temp_clipboard_file_path = temp_file.name
                    png_save_buffer = QBuffer()
                    png_save_buffer.open(QBuffer.WriteOnly)
                    if not final_canvas_for_clipboard.save(png_save_buffer, "PNG"):
                         raise IOError("Failed to save canvas to PNG buffer.")
                    temp_file.write(png_save_buffer.data())
                    png_save_buffer.close()
    
                if not os.path.exists(self.temp_clipboard_file_path):
                    raise IOError("Temporary file not created.")
    
            except Exception as e:
                QMessageBox.critical(self, "Copy Error", f"Failed to create/save temporary file: {e}")
                self.temp_clipboard_file_path = None
                return
            # --- End Save to Temp File ---
    
    
            # --- Prepare QMimeData ---
            clipboard = QApplication.clipboard()
            mime_data = QMimeData()
    
            # 1. Set File URL
            try:
                file_url = QUrl.fromLocalFile(self.temp_clipboard_file_path)
                if not file_url.isValid() or not file_url.isLocalFile(): raise ValueError("Invalid URL")
                mime_data.setUrls([file_url])
            except Exception as e:
                QMessageBox.warning(self, "Copy Warning", f"Could not set file URL: {e}")
    
            # 2. Set PNG Data fallback
            png_buffer_clip = QBuffer()
            png_buffer_clip.open(QBuffer.WriteOnly)
            if final_canvas_for_clipboard.save(png_buffer_clip, "PNG"):
                mime_data.setData("image/png", png_buffer_clip.data())
            png_buffer_clip.close()
    
            # 3. Set Standard Image Data fallback
            mime_data.setImageData(final_canvas_for_clipboard)
    
            # --- Set clipboard ---
            clipboard.setMimeData(mime_data)
    
            # --- REMOVED QTimer.singleShot ---
    
        else:
            QMessageBox.warning(self, "Warning", "No valid image data to render for copying.")
        
    def cleanup_temp_clipboard_file(self):
        """Deletes the temporary clipboard file if it exists."""
        path_to_delete = getattr(self, 'temp_clipboard_file_path', None)
        if path_to_delete:
            # print(f"Attempting to clean up temp file on exit: {path_to_delete}") # Debug
            # No need to clear self.temp_clipboard_file_path immediately here,
            # as the object is likely being destroyed soon anyway.
            if os.path.exists(path_to_delete):
                try:
                    os.remove(path_to_delete)
                    # print(f"Cleaned up temp file: {path_to_delete}") # Debug
                # Don't retry with QTimer here, the app is quitting.
                except OSError as e:
                    # Log it, but can't do much else at this point.
                    print(f"Warning: Could not delete temp clipboard file {path_to_delete} on exit: {e}")
        # Clear the attribute *after* attempting deletion (optional on exit)
        self.temp_clipboard_file_path = None
        
    def copy_to_clipboard_SVG(self):
        """Create a temporary SVG file with EMF data, copy it to clipboard."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
    
        # Get scaling factors to match the live view window
        view_width = self.live_view_label.width()
        view_height = self.live_view_label.height()
        scale_x = self.image.width() / view_width
        scale_y = self.image.height() / view_height
    
        # Create a temporary file to store the SVG
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_file:
            temp_file_path = temp_file.name
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(temp_file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x * scale_x, y * scale_y),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(self.left_marker_shift_added / scale_x, y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=((self.image.width() / scale_x + self.right_marker_shift_added / scale_x), y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x, self.top_marker_shift_added / scale_y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x}, {self.top_marker_shift_added / scale_y})"
                )
            )
    
        # Save the SVG to the temporary file
        dwg.save()
    
        # Read the SVG content
        with open(temp_file_path, "r", encoding="utf-8") as temp_file:
            svg_content = temp_file.read()
    
        # Copy SVG content to clipboard (macOS-compatible approach)
        clipboard = QApplication.clipboard()
        clipboard.setText(svg_content, mode=clipboard.Clipboard)
    
        QMessageBox.information(self, "Success", "SVG content copied to clipboard.")

        
    def clear_predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        if hasattr(self, "protein_location"):
            del self.protein_location  # Clear the protein location marker
        self.predict_size=False
        self.bounding_boxes=[]
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.live_view_label.bounding_box_preview = None
        self.quantities=[]
        self.peak_area_list=[]
        self.protein_quantities=[]
        self.standard_protein_values.setText("")
        self.standard_protein_areas=[]
        self.standard_protein_areas_text.setText("")
        self.live_view_label.quad_points=[]
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.rectangle_points = []
        self.latest_calculated_quantities = []
        self.quantities_peak_area_dict={}
        
        self.update_live_view()  # Update the display
        
    def predict_molecular_weight(self):
        """
        Initiates the molecular weight prediction process.
        Determines the available markers, sorts them by position,
        and sets up the mouse click event to get the target protein location.
        """
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.run_predict_MW = False # Reset prediction flag

        # Determine which markers to use (left or right)
        markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
        if not markers_raw_tuples:
            QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
            self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            return

        # --- Crucial Step: Sort markers by Y-position (migration distance) ---
        # This ensures the order reflects the actual separation on the gel/blot
        try:
            # Ensure marker values are numeric for sorting and processing
            numeric_markers = []
            for pos, val_str in markers_raw_tuples:
                try:
                    numeric_markers.append((float(pos), float(val_str)))
                except (ValueError, TypeError):
                    # Skip markers with non-numeric values for prediction
                    continue

            if len(numeric_markers) < 2:
                 QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 return

            sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
            # Separate sorted positions and values
            sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
            sorted_marker_values = np.array([val for pos, val in sorted_markers])

        except Exception as e:
            QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Ensure there are still at least two markers after potential filtering
        if len(sorted_marker_positions) < 2:
            QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Prompt user and set up the click handler
        QMessageBox.information(self, "Instruction",
                                "Click on the target protein location in the preview window.\n"
                                "The closest standard set (Gel or WB) will be used for calculation.")
        # Pass the *sorted* full marker data to the click handler
        self.live_view_label.mousePressEvent = lambda event: self.get_protein_location(
            event, sorted_marker_positions, sorted_marker_values
        )

    def get_protein_location(self, event, all_marker_positions, all_marker_values):
        """
        Handles the mouse click event for protein selection.
        Determines the relevant standard set based on click proximity,
        performs regression on that set, predicts MW, and plots the results.
        """
        # --- 1. Get Protein Click Position (Image Coordinates) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()

        # Account for zoom and pan
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        # Transform cursor position from view coordinates to image coordinates
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image and self.image.width() > 0 else 1
        image_height = self.image.height() if self.image and self.image.height() > 0 else 1

        scale = min(displayed_width / image_width, displayed_height / image_height)
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2

        protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
        # protein_x_image = (cursor_x - x_offset) / scale if scale != 0 else 0 # Keep X if needed later

        # Store the clicked location in *view* coordinates for drawing the marker later
        self.protein_location = (cursor_x, cursor_y) # Use view coordinates for the marker

        # --- 2. Identify Potential Standard Sets (Partitioning) ---
        transition_index = -1
        # Find the first index where the molecular weight INCREASES after initially decreasing
        # (indicates a likely switch from Gel high->low MW to WB high->low MW)
        initial_decrease = False
        for k in range(1, len(all_marker_values)):
             if all_marker_values[k] < all_marker_values[k-1]:
                 initial_decrease = True # Confirm we've started migrating down
             # Check for increase *after* we've seen a decrease
             if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                 transition_index = k
                 break # Found the likely transition

        # --- 3. Select the Active Standard Set based on Click Proximity ---
        active_marker_positions = None
        active_marker_values = None
        set_name = "Full Set" # Default name

        if transition_index != -1:
            # We have two potential sets
            set1_positions = all_marker_positions[:transition_index]
            set1_values = all_marker_values[:transition_index]
            set2_positions = all_marker_positions[transition_index:]
            set2_values = all_marker_values[transition_index:]

            # Check if both sets are valid (at least 2 points)
            valid_set1 = len(set1_positions) >= 2
            valid_set2 = len(set2_positions) >= 2

            if valid_set1 and valid_set2:
                # Calculate the mean Y position for each set
                mean_y_set1 = np.mean(set1_positions)
                mean_y_set2 = np.mean(set2_positions)

                # Assign the click to the set whose mean Y is closer
                if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                    active_marker_positions = set1_positions
                    active_marker_values = set1_values
                    set_name = "Set 1 (Gel?)" # Tentative name
                else:
                    active_marker_positions = set2_positions
                    active_marker_values = set2_values
                    set_name = "Set 2 (WB?)" # Tentative name
            elif valid_set1: # Only set 1 is valid
                 active_marker_positions = set1_positions
                 active_marker_values = set1_values
                 set_name = "Set 1 (Gel?)"
            elif valid_set2: # Only set 2 is valid
                 active_marker_positions = set2_positions
                 active_marker_values = set2_values
                 set_name = "Set 2 (WB?)"
            else: # Neither set is valid after splitting
                 QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return
        else:
            # Only one set detected, use all markers
            if len(all_marker_positions) >= 2:
                active_marker_positions = all_marker_positions
                active_marker_values = all_marker_values
                set_name = "Single Set"
            else: # Should have been caught earlier, but double-check
                 QMessageBox.warning(self, "Error", "Not enough markers in the single set.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return

        # --- 4. Perform Regression on the Active Set ---
        # Normalize distances *within the active set*
        min_pos_active = np.min(active_marker_positions)
        max_pos_active = np.max(active_marker_positions)
        if max_pos_active == min_pos_active: # Avoid division by zero if all points are the same
            QMessageBox.warning(self, "Error", "All markers in the selected set have the same position.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if hasattr(self, "protein_location"): del self.protein_location
            return

        normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)

        # Log transform marker values
        try:
            log_marker_values_active = np.log10(active_marker_values)
        except Exception as e:
             QMessageBox.warning(self, "Error", f"Could not log-transform marker values (are they all positive?): {e}")
             self.live_view_label.setCursor(Qt.ArrowCursor)
             if hasattr(self, "protein_location"): del self.protein_location
             return

        # Perform linear regression (log-linear fit)
        coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, 1)

        # Calculate R-squared for the fit on the active set
        residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1 # Handle case of perfect fit or single point mean

        # --- 5. Predict MW for the Clicked Protein ---
        # Normalize the protein's Y position *using the active set's min/max*
        normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)

        # Predict using the derived coefficients
        predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
        predicted_weight = 10 ** predicted_log10_weight

        # --- 6. Update View and Plot ---
        self.update_live_view() # Draw the '*' marker at self.protein_location

        # Plot the results, passing both active and all data for context
        self.plot_molecular_weight_graph(
            # Active set data for fitting and line plot:
            normalized_distances_active,
            active_marker_values,
            10 ** np.polyval(coefficients, normalized_distances_active), # Fitted values for active set
            # Full set data for context (plotting all points):
            all_marker_positions, # Pass original positions
            all_marker_values,
            min_pos_active,       # Pass min/max of *active* set for normalization context
            max_pos_active,
            # Prediction results:
            normalized_protein_position, # Position relative to active set scale
            predicted_weight,
            r_squared,
            set_name # Name of the set used
        )

        # Reset mouse event handler after prediction
        self.live_view_label.mousePressEvent = None
        self.live_view_label.setCursor(Qt.ArrowCursor)
        self.run_predict_MW = True # Indicate prediction was attempted/completed

        
        
    def plot_molecular_weight_graph(
        self,
        # Data for the active set (used for the fit line and highlighted points)
        active_norm_distances,  # Normalized distances for the active set points
        active_marker_values,   # Original MW values of the active set points
        active_fitted_values,   # Fitted MW values corresponding to active_norm_distances

        # Data for *all* markers (for plotting all points for context)
        all_marker_positions,   # *Original* Y positions of all markers
        all_marker_values,      # *Original* MW values of all markers
        active_min_pos,         # Min Y position of the *active* set (for normalizing all points)
        active_max_pos,         # Max Y position of the *active* set (for normalizing all points)

        # Prediction results
        predicted_norm_position,# Predicted protein position normalized relative to active set scale
        predicted_weight,       # Predicted MW value

        # Fit quality and set info
        r_squared,              # R-squared of the fit on the active set
        set_name                # Name of the set used (e.g., "Set 1", "Set 2", "Single Set")
    ):
        """
        Plots the molecular weight prediction graph and displays it in a custom dialog.
        """
        # --- Normalize *all* marker positions using the *active* set's scale ---
        if active_max_pos == active_min_pos: # Avoid division by zero
             all_norm_distances_for_plot = np.zeros_like(all_marker_positions)
        else:
            all_norm_distances_for_plot = (all_marker_positions - active_min_pos) / (active_max_pos - active_min_pos)

        # --- Create Plot ---
        # Adjust figsize and DPI for a more compact plot suitable for a dialog
        fig, ax = plt.subplots(figsize=(4.5, 3.5)) # Smaller figure size

        # 1. Plot *all* marker points lightly for context
        ax.scatter(all_norm_distances_for_plot, all_marker_values, color="grey", alpha=0.5, label="All Markers", s=25) # Slightly smaller

        # 2. Plot the *active* marker points prominently
        ax.scatter(active_norm_distances, active_marker_values, color="red", label=f"Active Set ({set_name})", s=40, marker='o') # Slightly smaller

        # 3. Plot the fitted line for the *active* set
        sort_indices = np.argsort(active_norm_distances)
        # Move R^2 out of the legend
        ax.plot(active_norm_distances[sort_indices], active_fitted_values[sort_indices],
                 color="blue", label="Fit Line", linewidth=1.5)

        # 4. Plot the predicted protein position
        # Move predicted value out of the legend
        ax.axvline(predicted_norm_position, color="green", linestyle="--",
                    label="Target Protein", linewidth=1.5)

        # --- Configure Plot ---
        ax.set_xlabel(f"Normalized Distance (Relative to {set_name})", fontsize=9)
        ax.set_ylabel("Molecular Weight (units)", fontsize=9)
        ax.set_yscale("log")
        # Smaller legend font size
        ax.legend(fontsize='x-small', loc='best') # Use 'best' location
        ax.set_title(f"Molecular Weight Prediction", fontsize=10) # Simpler title
        ax.grid(True, which='both', linestyle=':', linewidth=0.5)
        ax.tick_params(axis='both', which='major', labelsize=8) # Smaller tick labels

        plt.tight_layout(pad=0.5) # Adjust layout to prevent labels overlapping

        # --- Save Plot to Buffer ---
        pixmap = None
        try:
            buffer = BytesIO()
            # Use a moderate DPI suitable for screen display in a dialog
            plt.savefig(buffer, format="png", bbox_inches="tight", dpi=100)
            buffer.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buffer.read())
            buffer.close()
        except Exception as plot_err:
            QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
            # pixmap remains None
        finally:
             plt.close(fig) # Ensure figure is always closed using the fig object

        # --- Display Results in a Custom Dialog ---
        if pixmap:
            # Create a custom dialog instead of modifying QMessageBox layout
            dialog = QDialog(self)
            dialog.setWindowTitle("Prediction Result")

            # Layout for the dialog
            layout = QVBoxLayout(dialog)

            # Text label for results
            results_text = (
                f"<b>Prediction using {set_name}:</b><br>"
                f"Predicted MW: <b>{predicted_weight:.2f}</b> units<br>"
                f"Fit R²: {r_squared:.3f}"
            )
            info_label = QLabel(results_text)
            info_label.setTextFormat(Qt.RichText) # Allow basic HTML like <b>
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            # Label to display the plot
            plot_label = QLabel()
            plot_label.setPixmap(pixmap)
            plot_label.setAlignment(Qt.AlignCenter) # Center the plot
            layout.addWidget(plot_label)

            # OK button
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)

            dialog.setLayout(layout)
            dialog.exec_() # Show the custom dialog modally
        else:
             # If plot generation failed, show a simpler message box
             QMessageBox.information(self, "Prediction Result (No Plot)",
                f"Used {set_name} for prediction.\n"
                f"Predicted MW: {predicted_weight:.2f} units\n"
                f"Fit R²: {r_squared:.3f}"
             )

  
        
    def reset_image(self):
        # ... (Keep existing reset logic for image, markers, etc.) ...

        # --- Reset Crop Mode ---
        self.cancel_rectangle_crop_mode()
        self.crop_rectangle_coords = None
        self.live_view_label.clear_crop_preview()

        # --- Reset and Disable Crop Sliders ---
        slider_info = [
            (getattr(self, 'crop_x_start_slider', None), self.crop_slider_min),
            (getattr(self, 'crop_x_end_slider', None), self.crop_slider_max),
            (getattr(self, 'crop_y_start_slider', None), self.crop_slider_min),
            (getattr(self, 'crop_y_end_slider', None), self.crop_slider_max)
        ]
        for slider, default_value in slider_info:
            if slider:
                slider.blockSignals(True)
                slider.setValue(default_value)
                slider.setEnabled(False) # Disable sliders on reset
                slider.blockSignals(False)
        # --- End Reset Crop ---

        # ... (rest of reset_image method: clearing image, other sliders, markers, etc.) ...
        if hasattr(self, 'image_master') and self.image_master and not self.image_master.isNull():
            self.image = self.image_master.copy()
            self.image_before_padding = None
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_padded = False
        else:
            self.image = None
            self.original_image = None
            self.image_master = None
            self.image_before_padding = None
            self.image_contrasted = None
            self.image_before_contrast = None
            self.image_padded = False

        if hasattr(self, 'show_grid_checkbox_x'): self.show_grid_checkbox_x.setChecked(False)
        if hasattr(self, 'show_grid_checkbox_y'): self.show_grid_checkbox_y.setChecked(False)
        if hasattr(self, 'grid_size_input'): self.grid_size_input.setValue(20)

        self.warped_image=None
        if hasattr(self, 'left_markers'): self.left_markers.clear()
        if hasattr(self, 'right_markers'): self.right_markers.clear()
        if hasattr(self, 'top_markers'): self.top_markers.clear()
        if hasattr(self, 'custom_markers'): self.custom_markers.clear()
        if hasattr(self, 'custom_shapes'): self.custom_shapes.clear()
        self.cancel_drawing_mode()
        self.clear_predict_molecular_weight()

        if hasattr(self, 'orientation_slider'): self.orientation_slider.setValue(0)
        if hasattr(self, 'taper_skew_slider'): self.taper_skew_slider.setValue(0)
        if hasattr(self, 'high_slider'): self.high_slider.setValue(100)
        if hasattr(self, 'low_slider'): self.low_slider.setValue(100)
        if hasattr(self, 'gamma_slider'): self.gamma_slider.setValue(100)
        if hasattr(self, 'left_padding_slider'): self.left_padding_slider.setValue(0)
        if hasattr(self, 'right_padding_slider'): self.right_padding_slider.setValue(0)
        if hasattr(self, 'top_padding_slider'): self.top_padding_slider.setValue(0)
        if hasattr(self, 'left_padding_input'): self.left_padding_input.setText("0")
        if hasattr(self, 'right_padding_input'): self.right_padding_input.setText("0")
        if hasattr(self, 'top_padding_input'): self.top_padding_input.setText("0")
        if hasattr(self, 'bottom_padding_input'): self.bottom_padding_input.setText("0")

        self.marker_mode = None
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.left_marker_shift_added = 0
        self.right_marker_shift_added = 0
        self.top_marker_shift_added = 0
        self.live_view_label.mode = None
        self.live_view_label.quad_points = []
        self.live_view_label.setCursor(Qt.ArrowCursor)

        try:
            if hasattr(self, 'combo_box'):
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.on_combobox_changed()
        except Exception as e: pass

        if self.image and not self.image.isNull():
            try:
                self._update_preview_label_size()
                if hasattr(self, 'left_padding_input'): self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                if hasattr(self, 'right_padding_input'): self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                if hasattr(self, 'top_padding_input'): self.top_padding_input.setText(str(int(self.image.height()*0.15)))
            except Exception as e: pass
        else:
            self.live_view_label.clear()
            self._update_preview_label_size()

        self.live_view_label.zoom_level = 1.0
        self.live_view_label.pan_offset = QPointF(0, 0)
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)

        self._update_marker_slider_ranges()
        self.update_live_view()
        self._update_status_bar()


# if __name__ == "__main__":
#     try:
#         app = QApplication([])
#         app.setStyle("Fusion")
        
#         app.setStyleSheet("""
#         QSlider::handle:horizontal {
#             width: 100px;
#             height: 100px;
#             margin: -5px 0;
#             background: #FFFFFF;
#             border: 2px solid #555;
#             border-radius: 30px;
#         }
#         QStatusBar QLabel {
#             margin-left: 2px;
#             margin-right: 5px;
#             padding: 0px 0px;
#             border: none;
#         }
#         QPushButton:checked {
#             background-color: #a0d0a0; /* Light green background */
#             border: 1px solid #50a050; /* Darker green border */
#         }
#     """)
#         window = CombinedSDSApp()
#         window.show()
#         app.aboutToQuit.connect(window.cleanup_temp_clipboard_file)
#         app.exec_()
#     except Exception as e:
#         # This top-level except might catch errors *before* the app runs,
#         # so log_exception might not be called. Log directly here as a fallback.
#         print(f"FATAL: Application failed to start: {e}")
#         try:
#             # Try logging one last time if basicConfig worked at all
#             logging.critical("Application failed to start", exc_info=True)
#         except:
#             pass # Ignore errors during this final logging attemp
            
if __name__ == "__main__":
    # Minimal imports needed to START the app and show loading screen
    import sys
    import traceback # Import traceback for better error reporting
    from PyQt5.QtWidgets import QApplication
    # The LoadingDialog class defined above must be accessible here

    main_window = None # Initialize variable to hold the main window instance
    app = QApplication(sys.argv) # Create application first

    # --- Create and Show Loading Screen ---
    loading_dialog = LoadingDialog()
    try:
        loading_dialog.show()
        app.processEvents() # IMPORTANT: Allow the event loop to process and show the dialog
    except Exception as e:
        print(f"ERROR: Could not show loading dialog: {e}")
        # Continue without loading screen if it fails

    # --- Now perform heavy imports and main window initialization ---
    try:
        # Update loading message (optional)
        loading_dialog.set_message("Loading Libraries...")

        # <<< Move your main application class import HERE >>>
        # This delays the import until after the loading screen is shown
        # from Imaging_assistant_V7 import CombinedSDSApp # Assuming your file is named this
        # <<< You might also move other *very heavy* library imports here >>>
        # (e.g., if numpy, scipy, matplotlib import takes significant time)
        # import numpy as np
        # import matplotlib.pyplot as plt
        # ... etc ...
        # However, imports inside CombinedSDSApp.__init__ are usually the main delay.


        # Update loading message (optional)
        loading_dialog.set_message("Creating Main Window...")

        # Set Style *before* creating the main window if desired
        app.setStyle("Fusion")
        app.setStyleSheet("""
            QSlider::handle:horizontal {
                width: 100px;
                height: 100px;
                margin: -5px 0;
                background: #FFFFFF;
                border: 2px solid #555;
                border-radius: 30px;
            }
            QStatusBar QLabel {
                margin-left: 2px;
                margin-right: 5px;
                padding: 0px 0px;
                border: none;
            }
            QPushButton:checked {
                background-color: #a0d0a0; /* Light green background */
                border: 1px solid #50a050; /* Darker green border */
            }
        """)

        # --- Create the Main Window Instance ---
        # This is likely the most time-consuming part after imports
        main_window = CombinedSDSApp()

        # --- Initialization finished ---
        loading_dialog.close() # Close the loading screen

        main_window.show()
        app.aboutToQuit.connect(main_window.cleanup_temp_clipboard_file)
        sys.exit(app.exec_()) # Start the main event loop

    except ImportError as e:
        loading_dialog.close() # Ensure loading screen closes on import error
        print(f"FATAL ERROR: Missing required library: {e}")
        QMessageBox.critical(None, "Import Error", f"A required library is missing: {e}\nPlease install it and restart.")
        sys.exit(1)
    except Exception as e:
        loading_dialog.close() # Ensure loading screen closes on any error
        # Log the exception *before* showing the critical message
        log_exception(type(e), e, e.__traceback__) # Use your existing logger
        # Show a critical error message (log_exception might already do this)
        # QMessageBox.critical(None, "Application Error", f"An error occurred during startup:\n{e}\n\nCheck error_log.txt.")
        print(f"FATAL ERROR during application startup: {e}")
        traceback.print_exc()
        sys.exit(1)