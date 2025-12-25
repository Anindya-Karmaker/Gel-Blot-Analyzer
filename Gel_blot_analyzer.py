import sys
import re
import os
from PySide6.QtWidgets import (QApplication, QDialog, QLabel, QVBoxLayout,
                             QMessageBox)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont, QScreen, QGuiApplication
import logging
import traceback

class MinimalLoadingDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Loading...")
        self.setWindowFlags(Qt.SplashScreen | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground, False)

        self.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 1px solid #AAAAAA; /* Light gray border */
                border-radius: 8px;
            }
            QLabel {
                color: #333333;
                padding: 25px;
                background-color: transparent;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.label = QLabel("Gel Blot Analyzer v4.0\nDeveloped by Anindya Karmaker\nLoading software, please wait...")
        font = QFont("Arial", 11)
        font.setBold(True)
        self.label.setFont(font)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.setLayout(layout)

        self.setFixedSize(320, 120)
        self.center_on_screen()

    def center_on_screen(self):
        try:
            primary_screen = QGuiApplication.primaryScreen()
            if not primary_screen:
                screens = QGuiApplication.screens()
                if not screens:
                    print("Warning: No screens found to center loading dialog.")
                    self.move(100, 100)
                    return
                primary_screen = screens[0]

            screen_geo = primary_screen.availableGeometry()
            dialog_geo = self.frameGeometry()
            center_point = screen_geo.center()
            dialog_geo.moveCenter(center_point)
            self.move(dialog_geo.topLeft())
        except Exception as e:
            print(f"Warning: Could not center loading dialog: {e}")
            self.move(100, 100)

script_dir = os.path.dirname(os.path.abspath(__file__))
log_file_path = os.path.join(script_dir, "error_log.txt")


logging.basicConfig(
    filename=log_file_path,
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

try:
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not write initial log message: {e}")

logger = logging.getLogger()
logger.setLevel(logging.ERROR)

try:
    handler = logging.FileHandler(log_file_path, 'a', 'utf-8')
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not configure logging or write initial message: {e}")


def log_exception(exc_type, exc_value, exc_traceback):
    print("!!! log_exception called !!!")
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    try:
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
        
        for handler in logging.getLogger().handlers:
            handler.flush()

    except Exception as log_err:
        print(f"ERROR: Failed to log exception to file: {log_err}")

    try:
        error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}\n\n(Check error_log.txt for details)"
        QMessageBox.critical(
            None,
            "Unexpected Error",
            error_message,
            QMessageBox.Ok
        )
    except Exception as q_err:
         print(f"ERROR: Failed to show QMessageBox: {q_err}")


sys.excepthook = log_exception
	
if __name__ == "__main__":
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = None
    loading_dialog = None
    main_window = None

    try:
        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv if hasattr(sys, 'argv') and len(sys.argv) > 0 else [])
        else:
            print("INFO: Using existing QApplication instance.")

        
        app.setFont(QFont("Segoe UI", 10))

        try:
            loading_dialog = MinimalLoadingDialog()
            loading_dialog.show()
            if app: app.processEvents()
        except Exception as e_load_dialog:
            print(f"ERROR: Could not create/show minimal loading dialog: {e_load_dialog}")
            loading_dialog = None

        import sys
        import tempfile
        from tempfile import NamedTemporaryFile
        import base64
        from PIL import ImageDraw, ImageFont, ImageGrab, Image, ImageQt, ImageOps
        from io import BytesIO
        import io
        from PySide6.QtWidgets import (
            QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem,QToolBar,QStyle, QRadioButton, QButtonGroup,
            QScrollArea, QInputDialog, QFrame, QApplication, QSizePolicy,
            QMainWindow, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit,
            QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog,
            QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox, QDoubleSpinBox,
            QDialog, QHeaderView, QAbstractItemView, QMenu, QMenuBar, QFontDialog, QListWidget, 
        )
        from PySide6.QtGui import (
            QPixmap, QIcon, QPalette,QKeySequence, QImage, QPolygonF,QPainter, QBrush, QColor, QFont, QClipboard, QFontMetricsF,
            QPen, QTransform,QFontMetrics,QDesktopServices, QAction, QShortcut, QIntValidator, QFocusEvent, QDoubleValidator, QActionGroup,
        )
        from PySide6.QtCore import (
            Qt, QBuffer, QPoint, QPointF, QRect, QRectF, QUrl, QSize, QSizeF, QMimeData, Signal, QTimer, QEventLoop
        )
        import json
        import os
        import numpy as np
        import matplotlib.pyplot as plt
        import matplotlib.lines as mlines
        import matplotlib.patches as patches
        import platform
        import openpyxl
        from openpyxl.styles import Font
        from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
        from matplotlib.gridspec import GridSpec
        from skimage.restoration import rolling_ball 
        from scipy.signal import find_peaks
        from scipy.ndimage import gaussian_filter1d
        from scipy.ndimage import grey_opening, grey_erosion, grey_dilation
        from scipy.interpolate import interp1d
        SCIPY_AVAILABLE = False
        try:
            from scipy.optimize import curve_fit
            SCIPY_AVAILABLE = True
        except ImportError:
            print("WARNING: SciPy library not found. Advanced regression models (4-PL) will be disabled.")
            print("Install it with: pip install scipy")
        import cv2
        import datetime

        AMINO_ACID_RESIDUE_WEIGHTS = {
            'A': 71.0788, 'R': 156.1875, 'N': 114.1038, 'D': 115.0886,
            'C': 103.1388, 'E': 129.1155, 'Q': 128.1307, 'G': 57.0519,
            'H': 137.1411, 'I': 113.1594, 'L': 113.1594, 'K': 128.1741,
            'M': 131.1926, 'F': 147.1766, 'P': 97.1167, 'S': 87.0782,
            'T': 101.1051, 'W': 186.2132, 'Y': 163.1760, 'V': 99.1326
        }

        EXTINCTION_COEFFICIENTS = {
            'W': 5500,
            'Y': 1490,
            'C': 125
        }

        

        GLYCAN_MASSES_KDA = {
            "--- Select Glycan Type ---": 0.0,
            
            # N-linked High Mannose
            "N-linked High-Mannose (Man5)": 1.2,       # ~1235 Da
            "N-linked High-Mannose (Man9)": 1.9,       # ~1883 Da
            
            # N-linked Complex (Sialylated)
            "N-linked Complex (Bi-antennary A2G2S2)": 2.2,    # ~2224 Da (Standard "Complex")
            "N-linked Complex (Tri-antennary A3G3S3)": 2.9,   # ~2880 Da
            "N-linked Complex (Tetra-antennary A4G4S4)": 3.7, # ~3665 Da
            
            "Custom...": -1.0
        }

        def four_param_logistic(x, a, b, c, d):
            """ 4-Parameter Logistic Regression model (sigmoidal). y = d + (a - d) / (1 + (x / c)**b) """
            return d + (a - d) / (1 + (x / c)**b)

        

        def create_text_icon(font_type: QFont, icon_size: QSize, color: QColor, symbol: str) -> QIcon:
            """Creates a QIcon by drawing text/symbol onto a pixmap."""
            pixmap = QPixmap(icon_size)
            pixmap.fill(Qt.transparent)
            painter = QPainter(pixmap)
            painter.setRenderHint(QPainter.Antialiasing, True)
            painter.setRenderHint(QPainter.TextAntialiasing, True) # Good for text

            # Font settings (adjust as needed)
            font = QFont(font_type)
            # Make arrow slightly smaller than +/-, maybe not bold? Experiment.
            font.setPointSize(min(14, int(icon_size.height()*0.75)))
            # font.setBold(True) # Optional: Make arrows bold or not
            painter.setFont(font)
            painter.setPen(color)

            # Draw the symbol centered
            painter.drawText(pixmap.rect(), Qt.AlignCenter, symbol)
            painter.end()
            return QIcon(pixmap)

        # Set Style (can be done after app exists)


        class PredictionResultDialog(QDialog):
            """
            A dialog to display the molecular weight prediction, allowing for
            interactive changing of the regression model and internal standard calibration.
            """
            model_changed_in_dialog = Signal(str)

            def __init__(self, parent_app, all_marker_positions, all_marker_values, 
                        active_marker_positions, active_marker_values, protein_y_image, active_set_name,
                        initial_calibration=None):
                super().__init__(parent_app)
                self.setWindowTitle("Prediction & Calibration")
                self.setMinimumSize(650, 700)

                self.parent_app = parent_app
                self.all_marker_positions = all_marker_positions
                self.all_marker_values = all_marker_values
                self.active_marker_positions = active_marker_positions
                self.active_marker_values = active_marker_values
                self.protein_y_image = protein_y_image
                self.active_set_name = active_set_name
                
                self.final_model_type = "poly"
                self.final_coeffs = None
                self.final_min_max_pos = None
                self.final_predicted_mw = 0.0
                
                # Calibration State
                self.calibration_active = False
                self.calib_slope = 1.0
                self.calib_offset = 0.0

                main_layout = QVBoxLayout(self)
                
                # --- Model Selection ---
                controls_layout = QHBoxLayout()
                controls_layout.addWidget(QLabel("Regression Model:"))
                self.model_combo_dialog = QComboBox()
                
                model_items = ["Log-Linear (Degree 1)", "Log-Polynomial (Degree 2)", "Log-Polynomial (Degree 3)"]
                if SCIPY_AVAILABLE:
                    model_items.append("Log 4-PL")
                self.model_combo_dialog.addItems(model_items)

                self.model_combo_dialog.setCurrentText(self.parent_app.mw_regression_model_combo.currentText())
                self.model_combo_dialog.currentTextChanged.connect(self._recalculate_and_redraw)
                controls_layout.addWidget(self.model_combo_dialog, 1)
                main_layout.addLayout(controls_layout)

                # --- Plot ---
                self.fig, self.ax = plt.subplots(figsize=(5, 4))
                self.canvas = FigureCanvas(self.fig)
                main_layout.addWidget(self.canvas)

                # --- Results ---
                self.mw_label = QLabel("Predicted MW: -")
                self.r2_label = QLabel("Fit R²: -")
                font = self.mw_label.font(); font.setBold(True); font.setPointSize(12); self.mw_label.setFont(font)
                main_layout.addWidget(self.mw_label)
                main_layout.addWidget(self.r2_label)

                # --- INTERNAL CALIBRATION GROUP ---
                calib_group = QGroupBox("Internal Standard Calibration (1-Point or 2-Point)")
                calib_layout = QGridLayout(calib_group)
                calib_layout.addWidget(QLabel("Use this if you know the exact MW of bands in this lane (e.g. MALDI/LC-MS)."), 0, 0, 1, 4)

                # Point 1 Controls
                self.chk_calib1 = QCheckBox("Point 1")
                self.chk_calib1.stateChanged.connect(self._recalculate_and_redraw)
                
                self.spin_mw1 = QDoubleSpinBox()
                self.spin_mw1.setRange(0, 1000000)
                self.spin_mw1.setPrefix("Known MW: ")
                self.spin_mw1.setValue(0)
                self.spin_mw1.valueChanged.connect(self._recalculate_and_redraw)
                
                self.spin_y1 = QDoubleSpinBox()
                self.spin_y1.setRange(0, 100000)
                self.spin_y1.setPrefix("Y-Pos: ")
                self.spin_y1.setDecimals(2)
                self.spin_y1.setValue(self.protein_y_image) # Default to initial click
                self.spin_y1.valueChanged.connect(self._recalculate_and_redraw)
                
                self.btn_pick1 = QPushButton("Pick from Image")
                # Return special code 101 to indicate "Pick Point 1"
                self.btn_pick1.clicked.connect(lambda: self.done(101))
                
                calib_layout.addWidget(self.chk_calib1, 1, 0)
                calib_layout.addWidget(self.spin_mw1, 1, 1)
                calib_layout.addWidget(self.spin_y1, 1, 2)
                calib_layout.addWidget(self.btn_pick1, 1, 3)

                # Point 2 Controls
                self.chk_calib2 = QCheckBox("Point 2")
                self.chk_calib2.stateChanged.connect(self._recalculate_and_redraw)
                
                self.spin_mw2 = QDoubleSpinBox()
                self.spin_mw2.setRange(0, 1000000)
                self.spin_mw2.setPrefix("Known MW: ")
                self.spin_mw2.setValue(0)
                self.spin_mw2.valueChanged.connect(self._recalculate_and_redraw)
                
                self.spin_y2 = QDoubleSpinBox()
                self.spin_y2.setRange(0, 100000)
                self.spin_y2.setPrefix("Y-Pos: ")
                self.spin_y2.setDecimals(2)
                self.spin_y2.valueChanged.connect(self._recalculate_and_redraw)
                
                self.btn_pick2 = QPushButton("Pick from Image")
                # Return special code 102 to indicate "Pick Point 2"
                self.btn_pick2.clicked.connect(lambda: self.done(102))
                
                calib_layout.addWidget(self.chk_calib2, 2, 0)
                calib_layout.addWidget(self.spin_mw2, 2, 1)
                calib_layout.addWidget(self.spin_y2, 2, 2)
                calib_layout.addWidget(self.btn_pick2, 2, 3)
                
                main_layout.addWidget(calib_group)

                # --- Standard Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok)
                button_box.accepted.connect(self.accept)
                main_layout.addWidget(button_box)

                # --- RESTORE PREVIOUS STATE IF AVAILABLE ---
                if initial_calibration:
                    p1_data = initial_calibration.get("point1", {})
                    p2_data = initial_calibration.get("point2", {})
                    
                    if p1_data.get("active", False):
                        self.chk_calib1.setChecked(True)
                        self.spin_mw1.setValue(p1_data.get("mw", 0))
                        if "y" in p1_data and p1_data["y"] > 0:
                            self.spin_y1.setValue(p1_data["y"])
                    
                    if p2_data.get("active", False):
                        self.chk_calib2.setChecked(True)
                        self.spin_mw2.setValue(p2_data.get("mw", 0))
                        if "y" in p2_data and p2_data["y"] > 0:
                            self.spin_y2.setValue(p2_data["y"])

                self._recalculate_and_redraw()

            def _recalculate_and_redraw(self):
                selected_model_text = self.model_combo_dialog.currentText()
                min_pos_active = np.min(self.active_marker_positions)
                max_pos_active = np.max(self.active_marker_positions)
                
                # Helper for normalization
                def normalize(y_pixels):
                    return (y_pixels - min_pos_active) / (max_pos_active - min_pos_active)

                normalized_distances = normalize(self.active_marker_positions)
                log_marker_values = np.log10(self.active_marker_values)
                
                coefficients = None
                r_squared = 0.0
                predicted_log10_weight = 0.0

                self.ax.clear()

                # --- FIT THE STANDARD CURVE ---
                if "4-PL" in selected_model_text:
                    self.final_model_type = "4-PL"
                    if not SCIPY_AVAILABLE:
                        self.ax.text(0.5, 0.5, "SciPy required for 4-PL.", ha='center')
                        self.canvas.draw(); return
                    if len(normalized_distances) < 4:
                        self.ax.text(0.5, 0.5, f"Need 4+ points for 4-PL.", ha='center')
                        self.canvas.draw(); return
                    
                    try:
                        p0 = [np.max(log_marker_values), 1.0, np.median(normalized_distances), np.min(log_marker_values)]
                        coefficients, _ = curve_fit(four_param_logistic, normalized_distances, log_marker_values, p0=p0, maxfev=10000)
                        
                        residuals = log_marker_values - four_param_logistic(normalized_distances, *coefficients)
                        ss_res = np.sum(residuals**2)
                        ss_tot = np.sum((log_marker_values - np.mean(log_marker_values))**2)
                        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0
                    except RuntimeError:
                        self.ax.text(0.5, 0.5, "4-PL failed to converge.", ha='center')
                        self.canvas.draw(); return

                else: # Polynomial models
                    self.final_model_type = "poly"
                    poly_degree = 1
                    if "Degree 2" in selected_model_text: poly_degree = 2
                    elif "Degree 3" in selected_model_text: poly_degree = 3

                    if len(normalized_distances) <= poly_degree:
                        self.ax.text(0.5, 0.5, f"Not enough points.", ha='center')
                        self.canvas.draw(); return

                    coefficients = np.polyfit(normalized_distances, log_marker_values, poly_degree)
                    residuals = log_marker_values - np.polyval(coefficients, normalized_distances)
                    ss_res = np.sum(residuals**2)
                    ss_tot = np.sum((log_marker_values - np.mean(log_marker_values))**2)
                    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0

                # --- Helper to get Raw Prediction from Curve ---
                def get_raw_log_mw(y_pos_px):
                    norm_x = normalize(y_pos_px)
                    if coefficients is None: return 0
                    if self.final_model_type == "4-PL": return four_param_logistic(norm_x, *coefficients)
                    else: return np.polyval(coefficients, norm_x)

                # --- CALIBRATION CALCULATION ---
                self.calibration_active = False
                self.calib_slope = 1.0
                self.calib_offset = 0.0

                use_p1 = self.chk_calib1.isChecked() and self.spin_mw1.value() > 0
                use_p2 = self.chk_calib2.isChecked() and self.spin_mw2.value() > 0 and use_p1

                if use_p1 and coefficients is not None:
                    self.calibration_active = True
                    
                    # Point 1
                    obs_y1 = self.spin_y1.value()
                    known_log_mw1 = np.log10(self.spin_mw1.value())
                    pred_log_mw1 = get_raw_log_mw(obs_y1)

                    if use_p2:
                        # 2-Point Calibration (Slope + Offset)
                        obs_y2 = self.spin_y2.value()
                        known_log_mw2 = np.log10(self.spin_mw2.value())
                        pred_log_mw2 = get_raw_log_mw(obs_y2)
                        
                        delta_known = known_log_mw2 - known_log_mw1
                        delta_pred = pred_log_mw2 - pred_log_mw1
                        
                        if abs(delta_pred) > 1e-5:
                            self.calib_slope = delta_known / delta_pred
                        else:
                            self.calib_slope = 1.0 
                        
                        # Offset = Known - (Slope * Pred)
                        self.calib_offset = known_log_mw1 - (self.calib_slope * pred_log_mw1)
                        
                        # Visuals
                        self.ax.plot(normalize(obs_y1), known_log_mw1, 'g^', markersize=10, label="Calib Pt 1")
                        self.ax.plot(normalize(obs_y2), known_log_mw2, 'gv', markersize=10, label="Calib Pt 2")

                    else:
                        # 1-Point Calibration (Offset Only)
                        self.calib_offset = known_log_mw1 - pred_log_mw1
                        self.ax.plot(normalize(obs_y1), known_log_mw1, 'g^', markersize=10, label="Calib Pt 1")

                # --- Final Calculation ---
                raw_pred_log = get_raw_log_mw(self.protein_y_image)
                # Apply Calibration: Final = Slope * Raw + Offset
                final_log_mw = (self.calib_slope * raw_pred_log) + self.calib_offset
                predicted_weight = 10 ** final_log_mw
                
                self.final_coeffs = coefficients
                self.final_min_max_pos = (min_pos_active, max_pos_active)
                self.final_predicted_mw = predicted_weight

                # Update Labels
                calib_str = ""
                if self.calibration_active:
                    calib_str = f" [Calibrated: 1-Pt]" if not use_p2 else f" [Calibrated: 2-Pt]"
                
                self.mw_label.setText(f"Predicted MW: <b>{predicted_weight:.2f}</b> units {calib_str}")
                self.r2_label.setText(f"Std Curve R²: {r_squared:.4f}")
                
                # --- Plotting ---
                fit_line_x_dense_norm = np.linspace(0, 1, 200)
                if coefficients is not None:
                    if self.final_model_type == "4-PL": fit_y_log = four_param_logistic(fit_line_x_dense_norm, *coefficients)
                    else: fit_y_log = np.polyval(coefficients, fit_line_x_dense_norm)
                    
                    # Plot Standard Curve
                    self.ax.plot(fit_line_x_dense_norm, fit_y_log, color="blue", label="Std Curve", alpha=0.5, linestyle="--")
                    
                    # Plot Calibrated Curve
                    if self.calibration_active:
                        calib_y_log = (self.calib_slope * fit_y_log) + self.calib_offset
                        self.ax.plot(fit_line_x_dense_norm, calib_y_log, color="green", label="Calibrated Curve", linewidth=1.5)

                self.ax.scatter(normalized_distances, log_marker_values, color="red", label="Std Markers", s=30)
                
                norm_protein_pos = normalize(self.protein_y_image)
                self.ax.axvline(norm_protein_pos, color="orange", linestyle="-", label="Target")
                self.ax.plot(norm_protein_pos, final_log_mw, 'o', color="orange", markersize=8)

                self.ax.set_ylabel("Log(MW)")
                self.ax.set_xlabel("Normalized Distance")
                self.ax.legend(fontsize='x-small', loc='best')
                self.ax.grid(True, linestyle=':', linewidth=0.5)
                self.fig.tight_layout(pad=0.5)
                self.canvas.draw()
                self.model_changed_in_dialog.emit(selected_model_text)

            def get_final_prediction_model(self):
                # Convert numpy coeffs to list for JSON serialization
                coeffs_safe = None
                if self.final_coeffs is not None:
                    coeffs_safe = self.final_coeffs.tolist() if isinstance(self.final_coeffs, np.ndarray) else list(self.final_coeffs)

                return {
                    "model": self.final_model_type, 
                    "coeffs": coeffs_safe, 
                    "min_max_pos": self.final_min_max_pos,
                    "calibration": {
                        "active": self.calibration_active,
                        "slope": self.calib_slope,
                        "offset": self.calib_offset,
                        "point1": {
                            "active": self.chk_calib1.isChecked(),
                            "mw": self.spin_mw1.value(),
                            "y": self.spin_y1.value()
                        },
                        "point2": {
                            "active": self.chk_calib2.isChecked(),
                            "mw": self.spin_mw2.value(),
                            "y": self.spin_y2.value()
                        }
                    }
                }
            
            def get_final_predicted_mw(self):
                return self.final_predicted_mw

        class GlycosylationMapperDialog(QDialog):
            """
            A dialog for analyzing a protein sequence for N-glycosylation sites
            and calculating potential molecular weights of glycosylated and oligomeric fragments.
            """
            def __init__(self, sequence, base_mw, glycan_mass, num_oligomers, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Protein Size Analysis")
                self.setMinimumSize(900, 700) 

                if parent and hasattr(parent, 'styleSheet'):
                    self.setStyleSheet(parent.styleSheet())

                # Store initial values
                self.sequence = sequence
                self.base_mw = base_mw
                self.glycan_mass = glycan_mass
                self.num_glycosylation_sites = 0
                self.num_oligomers_to_model = num_oligomers

                # --- UI Setup --- (Layout remains the same as the improved two-column version)
                main_dialog_layout = QVBoxLayout(self)
                main_columns_layout = QHBoxLayout()
                left_column_layout = QVBoxLayout()
                right_column_layout = QVBoxLayout()
                readable_font = QFont("Courier New", 10)
                input_group = QGroupBox("Protein Sequence Input")
                input_layout = QVBoxLayout(input_group)
                self.sequence_entry = QTextEdit(self.sequence)
                self.sequence_entry.setPlaceholderText("Paste your protein sequence here (e.g., MNAEFGT...).")
                self.sequence_entry.setMinimumHeight(80)
                self.sequence_entry.setFont(readable_font)
                self.analyze_sequence_button = QPushButton("Analyze Sequence")
                self.analyze_sequence_button.setToolTip("Calculates all physicochemical properties and potential modifications from the sequence.")
                self.analyze_sequence_button.clicked.connect(self._analyze_sequence)
                input_layout.addWidget(self.sequence_entry)
                input_layout.addWidget(self.analyze_sequence_button, 0, Qt.AlignRight)
                left_column_layout.addWidget(input_group)
                analysis_group = QGroupBox("Sequence Analysis Results")
                analysis_layout = QVBoxLayout(analysis_group)
                self.sequence_analysis_text = QTextEdit()
                self.sequence_analysis_text.setReadOnly(True)
                self.sequence_analysis_text.setFont(readable_font)
                self.sequence_analysis_text.setLineWrapMode(QTextEdit.NoWrap)
                analysis_layout.addWidget(self.sequence_analysis_text)
                left_column_layout.addWidget(analysis_group, 1)
                output_group = QGroupBox("Calculated Potential Fragments")
                output_layout = QVBoxLayout(output_group)
                self.potential_fragments_text = QTextEdit(); self.potential_fragments_text.setReadOnly(True)
                self.potential_fragments_text.setFont(readable_font)
                output_layout.addWidget(self.potential_fragments_text)
                left_column_layout.addWidget(output_group, 1)
                physicochem_group = QGroupBox("Physicochemical Properties")
                props_layout = QGridLayout(physicochem_group)
                props_layout.addWidget(QLabel("<b>Isoelectric Point (pI):</b>"), 0, 0)
                self.pi_display = QLineEdit(); self.pi_display.setReadOnly(True)
                self.pi_display.setToolTip("Calculated isoelectric point based on provided pK values.")
                props_layout.addWidget(self.pi_display, 0, 1)
                props_layout.addWidget(QLabel("<b>Sequence Length:</b>"), 1, 0)
                self.sequence_length_display = QLineEdit(); self.sequence_length_display.setReadOnly(True)
                props_layout.addWidget(self.sequence_length_display, 1, 1)
                props_layout.addWidget(QLabel("<b>Ext. Coeff. (Reduced):</b>"), 2, 0)
                self.ext_coeff_reduced_display = QLineEdit(); self.ext_coeff_reduced_display.setReadOnly(True)
                self.ext_coeff_reduced_display.setToolTip("Assumes all Cysteine residues are reduced (free -SH groups).")
                props_layout.addWidget(self.ext_coeff_reduced_display, 2, 1)
                props_layout.addWidget(QLabel("<b>Absorbance (Reduced, 0.1%):</b>"), 3, 0)
                self.absorbance_reduced_display = QLineEdit(); self.absorbance_reduced_display.setReadOnly(True)
                self.absorbance_reduced_display.setToolTip("Calculated as (Ext. Coeff. Reduced) / (Molecular Weight).")
                props_layout.addWidget(self.absorbance_reduced_display, 3, 1)
                props_layout.addWidget(QLabel("<b>Ext. Coeff. (Oxidized):</b>"), 4, 0)
                self.ext_coeff_oxidized_display = QLineEdit(); self.ext_coeff_oxidized_display.setReadOnly(True)
                self.ext_coeff_oxidized_display.setToolTip("Assumes all Cysteine pairs form disulfide bonds.")
                props_layout.addWidget(self.ext_coeff_oxidized_display, 4, 1)
                props_layout.addWidget(QLabel("<b>Absorbance (Oxidized, 0.1%):</b>"), 5, 0)
                self.absorbance_oxidized_display = QLineEdit(); self.absorbance_oxidized_display.setReadOnly(True)
                self.absorbance_oxidized_display.setToolTip("Calculated as (Ext. Coeff. Oxidized) / (Molecular Weight).")
                props_layout.addWidget(self.absorbance_oxidized_display, 5, 1)
                props_layout.addWidget(QLabel("<b>Net Charge at pH:</b>"), 6, 0)
                charge_layout = QHBoxLayout()
                self.ph_input_spinbox = QDoubleSpinBox()
                self.ph_input_spinbox.setRange(0.0, 14.0)
                self.ph_input_spinbox.setDecimals(2)
                self.ph_input_spinbox.setSingleStep(0.25)
                self.ph_input_spinbox.setValue(7.0) # Default to physiological pH
                self.ph_input_spinbox.valueChanged.connect(self._update_charge_display)
                self.charge_display = QLineEdit(); self.charge_display.setReadOnly(True)
                charge_layout.addWidget(self.ph_input_spinbox)
                charge_layout.addWidget(self.charge_display)
                props_layout.addLayout(charge_layout, 6, 1)
                right_column_layout.addWidget(physicochem_group)
                params_group = QGroupBox("Fragment Modeling Parameters")
                params_layout = QGridLayout(params_group)
                params_layout.addWidget(QLabel("Base Protein MW (Da):"), 0, 0)
                self.base_protein_mw_input_da = QLineEdit()
                self.base_protein_mw_input_da.setValidator(QDoubleValidator(0, 2000000, 2, self))
                self.base_protein_mw_input_da.setPlaceholderText("e.g., 44324.55 (Editable)")
                self.base_protein_mw_input_da.textChanged.connect(self._on_mw_da_changed)
                params_layout.addWidget(self.base_protein_mw_input_da, 0, 1, 1, 3)
                params_layout.addWidget(QLabel("Avg. Glycan Mass (kDa):"), 1, 0)
                
                # --- THIS IS THE FIX ---
                glycan_layout = QHBoxLayout()
                self.glycan_type_combo = QComboBox()
                self.glycan_type_combo.addItems(GLYCAN_MASSES_KDA.keys())
                self.glycan_type_combo.currentTextChanged.connect(self._on_glycan_type_selected)
                
                self.glycan_mass_input = QLineEdit(str(self.glycan_mass) if self.glycan_mass > 0 else "")
                self.glycan_mass_input.setValidator(QDoubleValidator(0, 100, 2, self))
                self.glycan_mass_input.setPlaceholderText("e.g., 2.5")
                self.glycan_mass_input.setFixedWidth(80) # Give the input a fixed width
                self.glycan_mass_input.textChanged.connect(self.update_potential_fragments)
                self.glycan_mass_input.textChanged.connect(self._on_manual_glycan_mass_edit)
                
                glycan_layout.addWidget(self.glycan_type_combo, 1) # Let the combo box take most of the space
                glycan_layout.addWidget(self.glycan_mass_input, 0) # Give the input field no stretch
                # --- END OF FIX ---

                params_layout.addLayout(glycan_layout, 1, 1, 1, 3)
                params_layout.addWidget(QLabel("Number of Glycans:"), 2, 0)
                self.num_glycans_spinbox = QSpinBox(); self.num_glycans_spinbox.setRange(0, 50)
                self.num_glycans_spinbox.valueChanged.connect(self.update_potential_fragments)
                params_layout.addWidget(self.num_glycans_spinbox, 2, 1)
                params_layout.addWidget(QLabel("Number of Oligomers:"), 2, 2)
                self.num_oligomers_spinbox = QSpinBox(); self.num_oligomers_spinbox.setRange(1, 10); self.num_oligomers_spinbox.setValue(self.num_oligomers_to_model)
                self.num_oligomers_spinbox.valueChanged.connect(self.update_potential_fragments)
                params_layout.addWidget(self.num_oligomers_spinbox, 2, 3)
                right_column_layout.addWidget(params_group)
                right_column_layout.addStretch(1)
                main_columns_layout.addLayout(left_column_layout, 1)
                main_columns_layout.addLayout(right_column_layout, 1)
                main_dialog_layout.addLayout(main_columns_layout)
                io_layout = QHBoxLayout()
                self.export_button = QPushButton("Export Analysis"); self.export_button.clicked.connect(self._export_data)
                self.load_button = QPushButton("Load Analysis"); self.load_button.clicked.connect(self._load_data)
                io_layout.addWidget(self.load_button); io_layout.addWidget(self.export_button); io_layout.addStretch()
                main_dialog_layout.addLayout(io_layout)
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept); button_box.rejected.connect(self.reject)
                main_dialog_layout.addWidget(button_box)

                if self.sequence: self._analyze_sequence()
                if self.base_mw > 0: self.base_protein_mw_input_da.setText(f"{(self.base_mw * 1000):.2f}")

            # --- START MODIFICATION: pI Calculation using user-provided table ---
            def _calculate_isoelectric_point(self, sequence):
                """
                Calculates the isoelectric point (pI) of a protein sequence using an iterative
                method and pKa values from Bjellqvist et al. (1993, 1994).
                """
                if not sequence:
                    return None

                # pKa values from Bjellqvist et al., Electrophoresis 1993, 14, 1023-1031
                # and Bjellqvist et al., Electrophoresis 1994, 15, 529-539.
                # These values match the ExPASy pI/Mw calculator
                PK_DATA = {
                    'COOH': {
                        'A': 3.55, 'R': 3.55, 'N': 3.55, 'D': 4.55, 'C': 3.55, 
                        'E': 4.75, 'Q': 3.55, 'G': 3.55, 'H': 3.55, 'I': 3.55, 
                        'L': 3.55, 'K': 3.55, 'M': 3.55, 'F': 3.55, 'P': 3.55, 
                        'S': 3.55, 'T': 3.55, 'W': 3.55, 'Y': 3.55, 'V': 3.55
                    },
                    'NH3': {
                        'A': 7.59, 'R': 7.50, 'N': 7.22, 'D': 7.50, 'C': 7.50, 
                        'E': 7.70, 'Q': 7.50, 'G': 7.50, 'H': 7.50, 'I': 7.50, 
                        'L': 7.50, 'K': 7.50, 'M': 7.00, 'F': 7.50, 'P': 8.36, 
                        'S': 6.93, 'T': 6.82, 'W': 7.50, 'Y': 7.50, 'V': 7.44
                    },
                    'Side': {
                        'R': 12.0, 'D': 4.05, 'C': 9.0, 'E': 4.45, 
                        'H': 5.98, 'K': 10.0, 'Y': 10.0
                    }
                }
                
                n_term_aa = sequence[0]
                c_term_aa = sequence[-1]

                pk_n_term = PK_DATA['NH3'].get(n_term_aa, 7.50)
                pk_c_term = PK_DATA['COOH'].get(c_term_aa, 3.55)

                aa_counts = {aa: sequence.count(aa) for aa in PK_DATA['Side']}

                min_ph, max_ph = 0.0, 14.0
                current_ph = 7.0
                
                for _ in range(100):
                    # Charge of N-terminus (positive)
                    charge_n = 1.0 / (1.0 + 10**(current_ph - pk_n_term))
                    # Charge of C-terminus (negative)
                    charge_c = -1.0 / (1.0 + 10**(pk_c_term - current_ph))
                    
                    # Charges of positive side chains
                    charge_k = aa_counts.get('K', 0) / (1.0 + 10**(current_ph - PK_DATA['Side']['K']))
                    charge_r = aa_counts.get('R', 0) / (1.0 + 10**(current_ph - PK_DATA['Side']['R']))
                    charge_h = aa_counts.get('H', 0) / (1.0 + 10**(current_ph - PK_DATA['Side']['H']))
                    
                    # Charges of negative side chains
                    charge_d = -aa_counts.get('D', 0) / (1.0 + 10**(PK_DATA['Side']['D'] - current_ph))
                    charge_e = -aa_counts.get('E', 0) / (1.0 + 10**(PK_DATA['Side']['E'] - current_ph))
                    charge_cys = -aa_counts.get('C', 0) / (1.0 + 10**(PK_DATA['Side']['C'] - current_ph))
                    charge_tyr = -aa_counts.get('Y', 0) / (1.0 + 10**(PK_DATA['Side']['Y'] - current_ph))
                    
                    # Sum all charges
                    net_charge = (charge_n + charge_c + 
                                charge_k + charge_r + charge_h + 
                                charge_d + charge_e + charge_cys + charge_tyr)

                    # Bisection method to find pH where net_charge is zero
                    if abs(net_charge) < 1e-4:
                        break

                    if net_charge > 0:
                        min_ph = current_ph
                    else:
                        max_ph = current_ph
                    
                    current_ph = (min_ph + max_ph) / 2.0

                return current_ph
            
            def _calculate_net_charge_at_ph(self, sequence, ph):
                """Calculates the net charge of a protein sequence at a given pH value
                using pKa values from Bjellqvist et al. (1993, 1994).
                
                Args:
                    sequence: Protein amino acid sequence (string)
                    ph: pH value at which to calculate charge (float)
                
                Returns:
                    Net charge of the protein at the given pH (float)
                """
                if not sequence:
                    return None
                
                if ph < 0 or ph > 14:
                    raise ValueError("pH must be between 0 and 14")

                # pKa values from Bjellqvist et al., matching ExPASy pI/Mw calculator
                PK_DATA = {
                    'COOH': {
                        'A': 3.55, 'R': 3.55, 'N': 3.55, 'D': 4.55, 'C': 3.55, 
                        'E': 4.75, 'Q': 3.55, 'G': 3.55, 'H': 3.55, 'I': 3.55, 
                        'L': 3.55, 'K': 3.55, 'M': 3.55, 'F': 3.55, 'P': 3.55, 
                        'S': 3.55, 'T': 3.55, 'W': 3.55, 'Y': 3.55, 'V': 3.55
                    },
                    'NH3': {
                        'A': 7.59, 'R': 7.50, 'N': 7.22, 'D': 7.50, 'C': 7.50, 
                        'E': 7.70, 'Q': 7.50, 'G': 7.50, 'H': 7.50, 'I': 7.50, 
                        'L': 7.50, 'K': 7.50, 'M': 7.00, 'F': 7.50, 'P': 8.36, 
                        'S': 6.93, 'T': 6.82, 'W': 7.50, 'Y': 7.50, 'V': 7.44
                    },
                    'Side': {
                        'R': 12.0, 'D': 4.05, 'C': 9.0, 'E': 4.45, 
                        'H': 5.98, 'K': 10.0, 'Y': 10.0
                    }
                }
                
                n_term_aa = sequence[0]
                c_term_aa = sequence[-1]

                pk_n_term = PK_DATA['NH3'].get(n_term_aa, 7.50)
                pk_c_term = PK_DATA['COOH'].get(c_term_aa, 3.55)

                aa_counts = {aa: sequence.count(aa) for aa in PK_DATA['Side']}

                # Charge of N-terminus (positive)
                charge_n = 1.0 / (1.0 + 10**(ph - pk_n_term))
                # Charge of C-terminus (negative)
                charge_c = -1.0 / (1.0 + 10**(pk_c_term - ph))
                
                # Charges of positive side chains
                charge_k = aa_counts.get('K', 0) / (1.0 + 10**(ph - PK_DATA['Side']['K']))
                charge_r = aa_counts.get('R', 0) / (1.0 + 10**(ph - PK_DATA['Side']['R']))
                charge_h = aa_counts.get('H', 0) / (1.0 + 10**(ph - PK_DATA['Side']['H']))
                
                # Charges of negative side chains
                charge_d = -aa_counts.get('D', 0) / (1.0 + 10**(PK_DATA['Side']['D'] - ph))
                charge_e = -aa_counts.get('E', 0) / (1.0 + 10**(PK_DATA['Side']['E'] - ph))
                charge_cys = -aa_counts.get('C', 0) / (1.0 + 10**(PK_DATA['Side']['C'] - ph))
                charge_tyr = -aa_counts.get('Y', 0) / (1.0 + 10**(PK_DATA['Side']['Y'] - ph))
                
                # Sum all charges
                net_charge = (charge_n + charge_c + 
                            charge_k + charge_r + charge_h + 
                            charge_d + charge_e + charge_cys + charge_tyr)

                return net_charge

            def _on_mw_da_changed(self, text):
                if self.base_protein_mw_input_da.signalsBlocked(): return
                self.update_potential_fragments()

            def _format_sequence_with_numbers(self, sequence, line_length=10):
                output = []
                # Iterate through the sequence in steps of line_length (e.g., 10)
                for i in range(0, len(sequence), line_length):
                    # Get the chunk of the sequence for the current line
                    chunk = sequence[i : i + line_length]
                    
                    # Define start and end residue numbers for this line
                    start_num = i + 1
                    end_num = i + len(chunk)
                    
                    # Format the line with padding for alignment
                    # Format: [6 spaces for start #] [2 spaces] [sequence chunk] [2 spaces] [end #]
                    line = f"{str(start_num):>6}  {chunk}  {end_num}"
                    output.append(line)
                
                return "\n".join(output)
            
            def _update_charge_display(self):
                """Calculates and displays the net charge at the user-specified pH."""
                sequence = self.sequence_entry.toPlainText().strip().upper()
                if not sequence:
                    self.charge_display.clear()
                    return
                
                try:
                    ph = self.ph_input_spinbox.value()
                    charge = self._calculate_net_charge_at_ph(sequence, ph)
                    if charge is not None:
                        self.charge_display.setText(f"{charge:+.2f}")
                    else:
                        self.charge_display.clear()
                except Exception as e:
                    self.charge_display.setText("Error")
                    print(f"Error calculating charge: {e}")

            def _analyze_sequence(self):
                self.sequence = self.sequence_entry.toPlainText().strip().upper()
                if not self.sequence:
                    for widget in [self.sequence_analysis_text, self.base_protein_mw_input_da, self.ext_coeff_reduced_display,
                                   self.ext_coeff_oxidized_display, self.sequence_length_display, self.absorbance_reduced_display,
                                   self.absorbance_oxidized_display, self.pi_display]:
                        widget.clear()
                    self.num_glycosylation_sites = 0; self.num_glycans_spinbox.setValue(0); self.update_potential_fragments()
                    return

                calculated_pi = self._calculate_isoelectric_point(self.sequence)
                if calculated_pi is not None:
                    self.pi_display.setText(f"{calculated_pi:.2f}")
                else:
                    self.pi_display.clear()

                total_mass = 18.01528
                num_W = self.sequence.count('W'); num_Y = self.sequence.count('Y'); num_C = self.sequence.count('C')
                for aa in self.sequence: total_mass += AMINO_ACID_RESIDUE_WEIGHTS.get(aa, 0)
                if self.sequence: self.sequence_length_display.setText(f"{len(self.sequence)} residues")
                self.base_protein_mw_input_da.blockSignals(True)
                self.base_protein_mw_input_da.setText(f"{total_mass:.2f}" if total_mass > 18.1 else "")
                self.base_protein_mw_input_da.blockSignals(False)
                ext_coeff_reduced = (num_W * EXTINCTION_COEFFICIENTS['W']) + (num_Y * EXTINCTION_COEFFICIENTS['Y'])
                ext_coeff_oxidized = ext_coeff_reduced + ((num_C // 2) * EXTINCTION_COEFFICIENTS['C'])
                self.ext_coeff_reduced_display.setText(f"{ext_coeff_reduced:,} M⁻¹cm⁻¹")
                self.ext_coeff_oxidized_display.setText(f"{ext_coeff_oxidized:,} M⁻¹cm⁻¹")
                if total_mass > 18.1:
                    if ext_coeff_reduced > 0: self.absorbance_reduced_display.setText(f"{(ext_coeff_reduced / total_mass):.3f}")
                    else: self.absorbance_reduced_display.clear()
                    if ext_coeff_oxidized > 0: self.absorbance_oxidized_display.setText(f"{(ext_coeff_oxidized / total_mass):.3f}")
                    else: self.absorbance_oxidized_display.clear()
                else:
                    self.absorbance_reduced_display.clear(); self.absorbance_oxidized_display.clear()
                try:
                    pattern = re.compile(r'N[^P][ST]')
                    matches = pattern.finditer(self.sequence)
                    locations = [(match.start(), match.group()) for match in matches]
                    self.num_glycosylation_sites = len(locations)
                    self.num_glycans_spinbox.blockSignals(True); self.num_glycans_spinbox.setValue(self.num_glycosylation_sites); self.num_glycans_spinbox.blockSignals(False)
                    display_html_parts = ["<pre>", self._format_sequence_with_numbers(self.sequence), "</pre>", "<b>Potential N-Glycosylation Sites:</b>"]
                    if self.num_glycosylation_sites > 0:
                        site_lines = [f" - Site at position {loc + 1} (Sequence: {seq})" for loc, seq in locations]
                        display_html_parts.append("<ul><li>" + "</li><li>".join(site_lines) + "</li></ul>")
                        display_html_parts.append(f"<b>Total Potential Sites Found: {self.num_glycosylation_sites}</b>")
                    else: 
                        display_html_parts.append("<p> - None found.</p>")
                    self.sequence_analysis_text.setHtml("".join(display_html_parts))
                except Exception as e: self.sequence_analysis_text.setPlainText(f"An error occurred: {e}")
                self.update_potential_fragments()
                self._update_charge_display()

            def _on_glycan_type_selected(self, text):
                mass = GLYCAN_MASSES_KDA.get(text, 0.0)
                self.glycan_mass_input.blockSignals(True)
                if mass == -1.0: self.glycan_mass_input.clear(); self.glycan_mass_input.setFocus()
                elif mass >= 0.0: self.glycan_mass_input.setText(str(mass))
                self.glycan_mass_input.blockSignals(False); self.update_potential_fragments()

            def _on_manual_glycan_mass_edit(self):
                if not self.glycan_mass_input.signalsBlocked():
                    self.glycan_type_combo.blockSignals(True); self.glycan_type_combo.setCurrentText("Custom..."); self.glycan_type_combo.blockSignals(False)
            
            def process_sequence(self): self._analyze_sequence()

            def update_potential_fragments(self):
                try:
                    base_mw_da = float(self.base_protein_mw_input_da.text())
                    base_mw_kda = base_mw_da / 1000.0
                    glycan_mass = float(self.glycan_mass_input.text())
                    num_glycans_to_calc = self.num_glycans_spinbox.value()
                    num_oligomers_to_calc = self.num_oligomers_spinbox.value()
                    if base_mw_kda <= 0: raise ValueError
                    fragments_text_lines = []
                    oligomer_names = ["Monomer", "Dimer", "Trimer", "Tetramer", "Pentamer", "Hexamer", "Heptamer", "Octamer", "Nonamer", "Decamer"]
                    for j in range(1, num_oligomers_to_calc + 1):
                        oligomer_base_mw = base_mw_kda * j
                        name = oligomer_names[j-1] if j <= len(oligomer_names) else f"{j}-mer"
                        fragments_text_lines.append(f"--- {name} (Base MW: {oligomer_base_mw:.2f} kDa) ---")
                        fragments_text_lines.append(f" + 0 glycans: {oligomer_base_mw:.2f} kDa")
                        if glycan_mass > 0:
                            for i in range(1, num_glycans_to_calc + 1):
                                potential_mw = oligomer_base_mw + (i * glycan_mass)
                                fragments_text_lines.append(f" + {i} glycan(s): {potential_mw:.2f} kDa")
                        fragments_text_lines.append("")
                    self.potential_fragments_text.setPlainText("\n".join(fragments_text_lines))
                except (ValueError, TypeError): self.potential_fragments_text.clear()

            def _export_data(self):
                default_filename = f"Protein_Analysis_{datetime.datetime.now():%Y%m%d_%H%M%S}.txt"
                file_path, _ = QFileDialog.getSaveFileName(self, "Export Analysis Data", default_filename, "Text Files (*.txt)")
                if not file_path: return
                try:
                    content = [f"PROTEIN ANALYSIS REPORT", f"Generated: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}\n"]
                    content.extend(["--- INPUT SEQUENCE ---", self.sequence_entry.toPlainText() + "\n"])
                    content.extend(["--- SEQUENCE ANALYSIS ---", self.sequence_analysis_text.toPlainText() + "\n"])
                    content.extend(["--- PHYSICOCHEMICAL PROPERTIES ---",
                                    f"Sequence Length: {self.sequence_length_display.text()}",
                                    f"Isoelectric Point (pI): {self.pi_display.text()}",
                                    f"Molecular Weight (Da): {self.base_protein_mw_input_da.text()}",
                                    f"Ext. Coeff. (Reduced): {self.ext_coeff_reduced_display.text()}",
                                    f"Ext. Coeff. (Oxidized): {self.ext_coeff_oxidized_display.text()}",
                                    f"Absorbance (Reduced, 0.1%): {self.absorbance_reduced_display.text()}",
                                    f"Absorbance (Oxidized, 0.1%): {self.absorbance_oxidized_display.text()}\n"])
                    content.extend(["--- FRAGMENT MODELING PARAMETERS ---",
                                    f"Selected Glycan Type: {self.glycan_type_combo.currentText()}",
                                    f"Avg. Glycan Mass (kDa): {self.glycan_mass_input.text()}",
                                    f"Number of Glycans to Model: {self.num_glycans_spinbox.value()}",
                                    f"Number of Oligomers to Model: {self.num_oligomers_spinbox.value()}\n"])
                    content.extend(["--- CALCULATED POTENTIAL FRAGMENTS ---", self.potential_fragments_text.toPlainText()])
                    with open(file_path, 'w', encoding='utf-8') as f: f.write("\n".join(content))
                    QMessageBox.information(self, "Success", f"Analysis data exported successfully to:\n{file_path}")
                except Exception as e: QMessageBox.critical(self, "Export Error", f"Failed to export data: {e}")

            def _load_data(self):
                file_path, _ = QFileDialog.getOpenFileName(self, "Load Analysis Data", "", "Text Files (*.txt)")
                if not file_path: return
                try:
                    with open(file_path, 'r', encoding='utf-8') as f: content = f.read()
                    def get_section(header, text):
                        match = re.search(f"--- {header} ---\n(.*?)\n---", text, re.DOTALL)
                        return match.group(1).strip() if match else ""
                    def get_line_value(key, text):
                        match = re.search(f"^{re.escape(key)}: (.*)", text, re.MULTILINE)
                        return match.group(1).strip() if match else ""
                    sequence = get_section("INPUT SEQUENCE", content)
                    params_section = get_section("FRAGMENT MODELING PARAMETERS", content)
                    if not params_section: params_section = get_section("GLYCOSYLATION & OLIGOMERIZATION PARAMETERS", content) # Backward compatibility
                    props_section = get_section("PHYSICOCHEMICAL PROPERTIES", content)
                    
                    self.sequence_entry.setText(sequence)
                    base_mw = get_line_value("Molecular Weight (Da)", props_section)
                    self.base_protein_mw_input_da.setText(base_mw)
                    glycan_type = get_line_value("Selected Glycan Type", params_section)
                    glycan_mass = get_line_value("Avg. Glycan Mass (kDa)", params_section)
                    num_glycans = get_line_value("Number of Glycans to Model", params_section)
                    num_oligomers = get_line_value("Number of Oligomers to Model", params_section)
                    
                    self.glycan_type_combo.setCurrentText(glycan_type)
                    self.glycan_mass_input.setText(glycan_mass) 
                    self.num_glycans_spinbox.setValue(int(num_glycans) if num_glycans.isdigit() else 0)
                    self.num_oligomers_spinbox.setValue(int(num_oligomers) if num_oligomers.isdigit() else 1)
                    self._analyze_sequence()
                    QMessageBox.information(self, "Success", "Analysis data loaded successfully.")
                except Exception as e: QMessageBox.critical(self, "Load Error", f"Failed to parse analysis file: {e}")

            def get_results(self):
                try: base_mw_da = float(self.base_protein_mw_input_da.text())
                except ValueError: base_mw_da = 0.0
                try: glycan_mass = float(self.glycan_mass_input.text())
                except ValueError: glycan_mass = 0.0
                return {
                    "sequence": self.sequence_entry.toPlainText().strip(),
                    "base_mw": base_mw_da / 1000.0,
                    "glycan_mass": glycan_mass, "sites": self.num_glycosylation_sites,
                    "glycans_to_use": self.num_glycans_spinbox.value(),
                    "oligomers_to_use": self.num_oligomers_spinbox.value()
                }
            
        class ScaleDialog(QDialog):
            """A simple dialog to get a known length and its unit for image calibration."""
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Set Image Scale")
                layout = QGridLayout(self)
                

                layout.addWidget(QLabel("Known Length:"), 0, 0)
                self.length_spinbox = QDoubleSpinBox()
                self.length_spinbox.setRange(0.0001, 1000000.0)
                self.length_spinbox.setValue(1.0)
                self.length_spinbox.setDecimals(4)
                layout.addWidget(self.length_spinbox, 0, 1)

                layout.addWidget(QLabel("Units:"), 1, 0)
                self.unit_combo = QComboBox()
                # --- START MODIFICATION: Added more units ---
                self.unit_combo.addItems([
                    "pixels",
                    "µm (micrometers)",
                    "mm (millimeters)",
                    "cm (centimeters)",
                    "m (meters)",
                    "inches",
                    "feet"
                ])
                # Set a common default
                self.unit_combo.setCurrentText("mm (millimeters)")
                # --- END MODIFICATION ---
                layout.addWidget(self.unit_combo, 1, 1)

                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept)
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box, 2, 0, 1, 2)

            def get_values(self):
                """Returns the entered length and selected unit (short form)."""
                # --- START MODIFICATION: Return the short form of the unit ---
                full_unit_text = self.unit_combo.currentText()
                # Extract the short form (e.g., "mm" from "mm (millimeters)")
                short_unit = full_unit_text.split(" ")[0]
                return self.length_spinbox.value(), short_unit
                # --- END MODIFICATION ---
            
        class MeasurementToolWindow(QDialog):
            """A non-modal window to house all measurement tools and display results."""
            # Signal emitted when a tool button is toggled. The string is the mode name or None.
            tool_selected = Signal(str)
            clear_requested = Signal()

            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Measurement Tools")
                # Non-modal: setWindowFlags to not block the main window
                self.setWindowFlags(
                    Qt.Window |
                    Qt.CustomizeWindowHint |
                    Qt.WindowMinimizeButtonHint |
                    Qt.WindowCloseButtonHint
                )
                self.setMinimumWidth(350)
                
                layout = QGridLayout(self)
                layout.setSpacing(10)

                # --- Tool Buttons ---
                self.btn_set_scale = QPushButton("1. Set Scale")
                self.btn_set_scale.setToolTip("Calibrate by drawing a line on an object of known length.")
                self.btn_set_scale.setCheckable(True)

                self.btn_measure_distance = QPushButton("2. Measure Distance")
                self.btn_measure_distance.setToolTip("Click two points to measure the distance between them.")
                self.btn_measure_distance.setCheckable(True)
                
                self.btn_measure_area = QPushButton("3. Measure Area")
                self.btn_measure_area.setToolTip("Click and drag to draw a freehand shape and calculate its area.")
                self.btn_measure_area.setCheckable(True)

                # Use a button group to ensure only one tool is active at a time
                self.tool_button_group = QButtonGroup(self)
                self.tool_button_group.setExclusive(True)
                self.tool_button_group.addButton(self.btn_set_scale)
                self.tool_button_group.addButton(self.btn_measure_distance)
                self.tool_button_group.addButton(self.btn_measure_area)
                self.tool_button_group.buttonClicked.connect(self._on_tool_button_clicked)

                self.btn_clear_measurements = QPushButton("Clear All")
                self.btn_clear_measurements.setToolTip("Clear the current scale and any measurement overlays.")
                self.btn_clear_measurements.clicked.connect(self.clear_requested.emit)

                layout.addWidget(self.btn_set_scale, 0, 0)
                layout.addWidget(self.btn_measure_distance, 0, 1)
                layout.addWidget(self.btn_measure_area, 1, 0)
                layout.addWidget(self.btn_clear_measurements, 1, 1)

                # --- Display Labels ---
                self.scale_display_label = QLabel("<b>Scale:</b> Not Set")
                self.measurement_result_label = QLabel("<b>Result:</b> N/A")
                
                layout.addWidget(self.scale_display_label, 2, 0, 1, 2)
                layout.addWidget(self.measurement_result_label, 3, 0, 1, 2)
            
            def _on_tool_button_clicked(self, button):
                """Internal slot to translate button clicks into signals."""
                if not button.isChecked():
                    # If user clicks an already active button, it un-checks it.
                    self.tool_selected.emit(None) # Signal that no tool is active
                    return

                if button is self.btn_set_scale:
                    self.tool_selected.emit('set_scale')
                elif button is self.btn_measure_distance:
                    self.tool_selected.emit('measure_distance')
                elif button is self.btn_measure_area:
                    self.tool_selected.emit('measure_area')

            def uncheck_all_tools(self):
                """Method to programmatically uncheck all tool buttons."""
                # Temporarily disconnect the signal to prevent feedback loops
                self.tool_button_group.buttonClicked.disconnect(self._on_tool_button_clicked)
                checked_button = self.tool_button_group.checkedButton()
                if checked_button:
                    # Set exclusive to False, uncheck, then set back to True
                    self.tool_button_group.setExclusive(False)
                    checked_button.setChecked(False)
                    self.tool_button_group.setExclusive(True)
                # Reconnect the signal
                self.tool_button_group.buttonClicked.connect(self._on_tool_button_clicked)

            def update_scale_display(self, text):
                self.scale_display_label.setText(f"<b>Scale:</b> {text}")

            def update_result_display(self, text):
                self.measurement_result_label.setText(f"<b>Result:</b> {text}")

            def closeEvent(self, event):
                """When the window is closed, signal that no tool is active."""
                self.tool_selected.emit(None)
                super().closeEvent(event)
    
        class AutoLaneTuneDialog(QDialog):
            """
            Updated dialog with Center-of-Mass refinement for sub-pixel accuracy
            and robust 16-bit noise handling.
            """
            def __init__(self, pil_image_data, initial_settings, parent=None, is_from_quad_warp=False):
                super().__init__(parent)

                if is_from_quad_warp:
                    self.setWindowTitle("Tune Peaks (Warped Region)")
                else:
                    self.setWindowTitle("Tune Automatic Peak Detection")

                # Dynamic Screen-Relative Sizing
                screen_geo = QGuiApplication.primaryScreen().availableGeometry()
                target_w = min(1000, int(screen_geo.width() * 0.75))
                target_h = min(900, int(screen_geo.height() * 0.85))
                self.resize(target_w, target_h)

                # --- 16-BIT DATA SANITIZATION ---
                if not isinstance(pil_image_data, Image.Image):
                    raise TypeError("Input 'pil_image_data' must be a PIL Image object")

                self.pil_image_for_display = pil_image_data
                pil_mode = self.pil_image_for_display.mode
                self.original_max_value = 255.0

                try:
                    # Convert PIL to Numpy with handling for endianness
                    if pil_mode == 'I;16B': 
                        arr = np.array(self.pil_image_for_display, dtype='>u2')
                        self.intensity_array_original_range = arr.astype(np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode in ['I', 'I;16', 'I;16L', 'I;16N']:
                        arr = np.array(self.pil_image_for_display, dtype=np.uint16)
                        self.intensity_array_original_range = arr.astype(np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'F':
                        self.intensity_array_original_range = np.array(self.pil_image_for_display, dtype=np.float64)
                        max_val = np.max(self.intensity_array_original_range)
                        self.original_max_value = max(1.0, max_val)
                    else:
                        # 8-bit fallback
                        if pil_mode != 'L':
                            self.pil_image_for_display = self.pil_image_for_display.convert('L')
                        self.intensity_array_original_range = np.array(self.pil_image_for_display, dtype=np.uint8).astype(np.float64)
                        self.original_max_value = 255.0

                    # Handle multi-channel noise (if 16-bit RGBA slipped through)
                    if self.intensity_array_original_range.ndim == 3:
                        self.intensity_array_original_range = np.mean(self.intensity_array_original_range[:, :, :3], axis=2)

                    # Generate initial inverted profile (Dark bands = High values)
                    inverted_array = self.original_max_value - self.intensity_array_original_range
                    self.profile_original_inverted = np.sum(inverted_array, axis=1)

                except Exception as e:
                    print(f"Error processing image data in AutoLaneTuneDialog: {e}")
                    self.profile_original_inverted = np.zeros(100)
                    self.intensity_array_original_range = np.zeros((100, 10))

                self.selected_peak_index = -1 
                self.deleted_peak_indices = set()
                self._all_initial_peaks = np.array([])
                self.detected_peaks = np.array([]) 
                self.add_peak_mode_active = False
                self.is_inverted = False 

                # --- ROBUST DEFAULTS ---
                # Higher default smoothing for 16-bit images to kill static noise
                default_sigma = 3.0 if self.original_max_value > 255 else 1.0
                default_prominence = 0.05 

                self.smoothing_sigma = initial_settings.get('smoothing_sigma', default_sigma)
                self.peak_height_factor = initial_settings.get('peak_height_factor', 0.1)
                self.peak_distance = initial_settings.get('peak_distance', 10)
                self.peak_prominence_factor = initial_settings.get('peak_prominence_factor', default_prominence)
                self._final_settings = initial_settings.copy()

                self._setup_ui()
                self.run_peak_detection_and_plot()

            def _setup_ui(self):
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(10)

                plot_widget = QWidget()
                plot_layout = QVBoxLayout(plot_widget)
                plot_layout.setContentsMargins(0, 0, 0, 0)

                self.fig = plt.figure(figsize=(7, 5))
                gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig)
                self.fig.tight_layout(pad=0.5)
                self.ax_profile = self.fig.add_subplot(gs[0])
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax_profile)

                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.canvas.mpl_connect('button_press_event', self.on_canvas_click)
                plot_layout.addWidget(self.canvas)
                main_layout.addWidget(plot_widget, stretch=1)

                controls_group = QGroupBox("Peak Detection Parameters")
                controls_layout = QGridLayout(controls_group)
                controls_layout.setSpacing(8)

                controls_layout.addWidget(QLabel("Profile Method:"), 0, 0)
                profile_method_label = QLabel("Sum of Inverted Intensities")
                font = profile_method_label.font(); font.setBold(True); profile_method_label.setFont(font)
                controls_layout.addWidget(profile_method_label, 0, 1, 1, 2)

                self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
                self.smoothing_slider = QSlider(Qt.Horizontal)
                self.smoothing_slider.setRange(0, 100)
                self.smoothing_slider.setValue(int(self.smoothing_sigma * 10))
                self.smoothing_slider.valueChanged.connect(lambda val: self.smoothing_label.setText(f"Smoothing Sigma ({val/10.0:.1f})"))
                self.smoothing_slider.valueChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.smoothing_label, 1, 0)
                controls_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2)

                self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
                self.peak_prominence_slider = QSlider(Qt.Horizontal)
                self.peak_prominence_slider.setRange(0, 100)
                self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
                self.peak_prominence_slider.valueChanged.connect(lambda val: self.peak_prominence_slider_label.setText(f"Min Prominence ({val/100.0:.2f})"))
                self.peak_prominence_slider.valueChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
                controls_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)

                self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})")
                self.peak_height_slider = QSlider(Qt.Horizontal)
                self.peak_height_slider.setRange(0, 100)
                self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
                self.peak_height_slider.valueChanged.connect(lambda val: self.peak_height_slider_label.setText(f"Min Height ({val/100.0:.2f})"))
                self.peak_height_slider.valueChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.peak_height_slider_label, 3, 0)
                controls_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)

                self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
                self.peak_distance_slider = QSlider(Qt.Horizontal)
                self.peak_distance_slider.setRange(1, 200)
                self.peak_distance_slider.setValue(self.peak_distance)
                self.peak_distance_slider.valueChanged.connect(lambda val: self.peak_distance_slider_label.setText(f"Min Distance ({val}) px"))
                self.peak_distance_slider.valueChanged.connect(self.run_peak_detection_and_plot)
                controls_layout.addWidget(self.peak_distance_slider_label, 4, 0)
                controls_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)

                self.delete_peak_button = QPushButton("Delete Selected Peak")
                self.delete_peak_button.setEnabled(False)
                self.delete_peak_button.clicked.connect(self.delete_selected_peak)

                self.add_peak_button = QPushButton("Add Peak at Click")
                self.add_peak_button.setCheckable(True)
                self.add_peak_button.clicked.connect(self.toggle_add_peak_mode)

                self.invert_display_button = QPushButton("Invert Image")
                self.invert_display_button.setCheckable(True)
                self.invert_display_button.toggled.connect(self._toggle_inversion_display)
                self.invert_display_button.setToolTip("Toggle if your bands are Light-on-Dark vs Dark-on-Light.")
                
                button_hbox = QHBoxLayout()
                button_hbox.addWidget(self.delete_peak_button)
                button_hbox.addWidget(self.add_peak_button)
                button_hbox.addWidget(self.invert_display_button)
                
                controls_layout.addLayout(button_hbox, 5, 0, 1, 3) 
                controls_layout.setColumnStretch(1, 1)
                main_layout.addWidget(controls_group)

                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept_and_return_peaks)
                button_box.rejected.connect(self.reject)
                main_layout.addWidget(button_box)

            def _toggle_inversion_display(self, checked):
                self.is_inverted = checked
                self.run_peak_detection_and_plot()
            
            def toggle_add_peak_mode(self, checked):
                self.add_peak_mode_active = checked
                if checked:
                    self.canvas.setCursor(Qt.CrossCursor)
                    self.selected_peak_index = -1 
                    self.delete_peak_button.setEnabled(False)
                    self.update_plot_highlights()
                    QMessageBox.information(self, "Add Peak", "Click on the profile plot to add a peak.")
                else:
                    self.canvas.setCursor(Qt.ArrowCursor)

            def on_canvas_click(self, event):
                if event.inaxes != self.ax_profile or self.profile_original_inverted is None or event.button != 1:
                    return
                clicked_x = event.xdata # Keep float for accuracy
                
                if self.add_peak_mode_active:
                    clicked_int = int(round(clicked_x))
                    if 0 <= clicked_int < len(self.profile_original_inverted):
                        self.add_manual_peak(clicked_int)
                else:
                    if len(self.detected_peaks) > 0:
                        distances = np.abs(self.detected_peaks - clicked_x)
                        min_dist_idx = np.argmin(distances)
                        click_tolerance_x = max(5, self.peak_distance / 4.0)
                        if distances[min_dist_idx] <= click_tolerance_x:
                            self.selected_peak_index = self.detected_peaks[min_dist_idx]
                            self.delete_peak_button.setEnabled(True)
                        else:
                            self.selected_peak_index = -1
                            self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()
                    else:
                        self.selected_peak_index = -1
                        self.delete_peak_button.setEnabled(False)
                        self.update_plot_highlights()

            def add_manual_peak(self, x_coord):
                # Check for duplicates allowing for small float differences
                if any(abs(p - x_coord) < 0.1 for p in self.detected_peaks): return
                
                self.detected_peaks = np.sort(np.append(self.detected_peaks, float(x_coord)))
                self.update_plot_highlights()

            def delete_selected_peak(self):
                if self.selected_peak_index != -1:
                    # Remove peaks that match the selected one closely (float comparison)
                    self.detected_peaks = np.array([p for p in self.detected_peaks if abs(p - self.selected_peak_index) > 0.01])
                    self.selected_peak_index = -1; self.delete_peak_button.setEnabled(False)
                    self.update_plot_highlights()

            def _refine_peak_positions_center_of_mass(self, profile, peak_indices, window_radius=3):
                """
                Calculates the Center of Mass (Centroid) for each peak to achieve sub-pixel accuracy.
                """
                refined_peaks = []
                profile_len = len(profile)
                
                for peak_idx in peak_indices:
                    # Define a small window around the integer peak index
                    start = max(0, peak_idx - window_radius)
                    end = min(profile_len, peak_idx + window_radius + 1)
                    
                    window_vals = profile[start:end]
                    window_indices = np.arange(start, end)
                    
                    # Subtract local baseline to focus on the peak tip (optional but helpful)
                    local_min = np.min(window_vals)
                    weights = window_vals - local_min
                    
                    if np.sum(weights) > 0:
                        centroid = np.sum(window_indices * weights) / np.sum(weights)
                        refined_peaks.append(centroid)
                    else:
                        refined_peaks.append(float(peak_idx))
                        
                return np.array(refined_peaks)

            def run_peak_detection_and_plot(self):
                self.selected_peak_index = -1
                self.delete_peak_button.setEnabled(False)

                if self.profile_original_inverted is None: return
                
                self.smoothing_sigma = self.smoothing_slider.value() / 10.0
                self.peak_height_factor = self.peak_height_slider.value() / 100.0
                self.peak_distance = self.peak_distance_slider.value()
                self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
                
                smoothed_profile_base = self.profile_original_inverted.copy()
                
                # --- FIX: ALWAYS apply smoothing for detection if 16-bit to kill noise ---
                try:
                    current_sigma = self.smoothing_sigma
                    # Force a minimum smoothing of 1.0 for 16-bit unless user sets to 0
                    if self.original_max_value > 255 and current_sigma < 0.1 and self.smoothing_slider.value() > 0:
                         current_sigma = 1.0
                         
                    if current_sigma > 0.05:
                        smoothed_profile_base = gaussian_filter1d(smoothed_profile_base, sigma=current_sigma)
                except Exception: pass
                
                prof_min_base, prof_max_base = np.min(smoothed_profile_base), np.max(smoothed_profile_base)
                
                if self.is_inverted:
                    profile_for_detection = (prof_max_base + prof_min_base) - smoothed_profile_base
                else:
                    profile_for_detection = smoothed_profile_base
                
                # Normalize using percentiles to ignore outliers (hot pixels)
                p1, p99 = np.percentile(profile_for_detection, (1, 99))
                if p99 > p1 + 1e-6:
                    self.profile = np.clip((profile_for_detection - p1) / (p99 - p1), 0.0, 1.0) * 255.0
                else:
                    self.profile = np.zeros_like(profile_for_detection)

                profile_range_detect = np.ptp(self.profile)
                if profile_range_detect < 1e-6 : profile_range_detect = 1.0
                
                min_height_abs = np.min(self.profile) + profile_range_detect * self.peak_height_factor
                min_prominence_abs = profile_range_detect * self.peak_prominence_factor

                try:
                    # 1. Find raw integer peaks
                    peaks_indices_int, _ = find_peaks(self.profile, height=min_height_abs, prominence=min_prominence_abs, distance=self.peak_distance, width=1)
                    
                    # 2. Refine to Sub-Pixel Accuracy (Center of Mass)
                    # We use the smoothed profile for calculation to avoid noise affecting the centroid
                    peaks_refined = self._refine_peak_positions_center_of_mass(self.profile, peaks_indices_int)
                    
                    self._all_initial_peaks = np.sort(peaks_refined)
                    
                    # Filter deleted peaks (check proximity)
                    final_peaks = []
                    for p in self._all_initial_peaks:
                        # Check if this peak is close to any deleted peak index
                        is_deleted = False
                        for d in self.deleted_peak_indices:
                            if abs(p - d) < 1.0: # Tolerance
                                is_deleted = True
                                break
                        if not is_deleted:
                            final_peaks.append(p)
                            
                    self.detected_peaks = np.array(final_peaks)
                    
                except Exception as e:
                    print(f"Peak detection error: {e}")
                    self._all_initial_peaks = np.array([]); self.detected_peaks = np.array([])

                # 5. Plotting
                is_dark_theme = self.parent() and hasattr(self.parent(), 'current_theme') and self.parent().current_theme == "dark"
                if is_dark_theme:
                    bg_color, ax_bg_color, text_color, spine_color, grid_color = '#2D2D30', '#38383C', '#F1F1F1', '#707070', '#5A5A60'
                    profile_color, peak_marker_color, selected_peak_color = '#4DB6AC', '#FF8A65', '#42A5F5'
                else:
                    bg_color, ax_bg_color, text_color, spine_color, grid_color = 'white', 'white', 'black', '#555555', '#DDDDDD'
                    profile_color, peak_marker_color, selected_peak_color = 'black', 'red', 'blue'

                self.ax_profile.clear(); self.ax_image.clear()
                self.fig.patch.set_facecolor(bg_color)
                for axis in [self.ax_profile, self.ax_image]:
                    axis.patch.set_facecolor(ax_bg_color); [spine.set_color(spine_color) for spine in axis.spines.values()]
                    axis.tick_params(axis='x', colors=text_color); axis.tick_params(axis='y', colors=text_color)
                    axis.yaxis.label.set_color(text_color); axis.xaxis.label.set_color(text_color); axis.title.set_color(text_color)

                if profile_for_detection is not None and len(profile_for_detection) > 0:
                    self.ax_profile.plot(self.profile, label=f"Normalized Profile (σ={self.smoothing_sigma:.1f})", color=profile_color, lw=1.0)

                    if len(self.detected_peaks) > 0:
                        # Interpolate Y values for float X positions for plotting
                        peak_y_values = np.interp(self.detected_peaks, np.arange(len(self.profile)), self.profile)
                        self.peak_plot_artist, = self.ax_profile.plot(self.detected_peaks, peak_y_values, "x", color=peak_marker_color, markersize=8, label=f"Peaks")

                        if self.selected_peak_index != -1:
                            # Highlight selected
                            try:
                                # Find nearest peak to selected index
                                idx_in_valid = np.argmin(np.abs(self.detected_peaks - self.selected_peak_index))
                                if abs(self.detected_peaks[idx_in_valid] - self.selected_peak_index) < 0.1:
                                     self.ax_profile.plot(self.detected_peaks[idx_in_valid], peak_y_values[idx_in_valid], 'o', markersize=12, markeredgecolor=selected_peak_color, markerfacecolor='none')
                            except: pass

                    self.ax_profile.set_title("Intensity Profile (Normalized)", fontsize=10)
                    self.ax_profile.grid(True, linestyle=':', alpha=0.6, color=grid_color)
                    
                    try:
                        arr = self.intensity_array_original_range
                        if arr.ndim == 2:
                            arr_rot = np.rot90(arr, k=1) 
                            d_min, d_max = np.percentile(arr_rot, (1, 99))
                            cmap_val = 'gray' 
                            
                            profile_length = len(profile_for_detection)
                            extent = [0, profile_length - 1, 0, arr_rot.shape[0]]
                            
                            self.ax_image.imshow(arr_rot, cmap=cmap_val, aspect='auto', extent=extent, vmin=d_min, vmax=d_max)
                            self.ax_image.set_xlabel("Pixel Index", fontsize=9); self.ax_image.set_yticks([])
                            
                            # Draw lines on image for sub-pixel peaks
                            for p in self.detected_peaks:
                                self.ax_image.axvline(p, color='red', alpha=0.5, linewidth=1)
                                
                    except Exception as e: 
                        print(f"Preview error: {e}")

                self.canvas.draw_idle()

            def update_plot_highlights(self):
                if not hasattr(self, 'ax_profile'): return
                self.run_peak_detection_and_plot()

            def accept_and_return_peaks(self):
                self._final_settings = {
                    'smoothing_sigma': self.smoothing_slider.value() / 10.0,
                    'peak_height_factor': self.peak_height_slider.value() / 100.0,
                    'peak_distance': self.peak_distance_slider.value(),
                    'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
                    'band_estimation_method': "Sum",
                    'rolling_ball_radius': self._final_settings.get('rolling_ball_radius', 50),
                    'area_subtraction_method': self._final_settings.get('area_subtraction_method', "Rolling-valley"),
                }
                self.accept()

            def get_detected_peaks(self): return self.detected_peaks
            def get_final_settings(self): return self._final_settings
            
        
        
        class ModifyMarkersDialog(QDialog):
            """
            A dialog for modifying custom markers and shapes with global adjustments and
            individual item editing. This implementation uses a robust "repopulate-on-change"
            pattern to keep the UI view consistent with the data model, even after sorting.
            """
            shapes_adjusted_preview = Signal(list)
            global_markers_adjusted = Signal(list)
        
            def __init__(self, markers_list, shapes_list, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Modify Custom Markers and Shapes")
                self.setMinimumSize(950, 650)
        
                # The data model: two lists for markers to handle scaling correctly, one for shapes.
                self._original_markers_data = [tuple(m) for m in markers_list] # Pristine backup for scaling
                self.markers = [list(m) for m in markers_list]                 # Working copy for display
                self.shapes = [dict(s) for s in shapes_list]                   # Working copy
        
                self._block_signals = False
                self._current_image_width = parent.image.width() if parent and parent.image and not parent.image.isNull() else 1
                self._current_image_height = parent.image.height() if parent and parent.image and not parent.image.isNull() else 1
        
                self._setup_ui()
                self.populate_table()
        
            def _setup_ui(self):
                """Creates and arranges all UI elements for the dialog."""
                layout = QVBoxLayout(self)
        
                # --- Global Adjustment Controls ---
                global_adjust_group = QGroupBox("Global Adjustments for Markers")
                global_adjust_layout = QGridLayout(global_adjust_group)
                self.percent_precision_factor = 100.0
                self.scale_precision_factor = 10.0
        
                # ... (Slider creation logic remains unchanged) ...
                global_adjust_layout.addWidget(QLabel("Shift X (% Img W):"), 0, 0)
                self.abs_x_shift_slider = QSlider(Qt.Horizontal)
                self.abs_x_shift_slider.setRange(int(-100 * self.percent_precision_factor), int(100 * self.percent_precision_factor)); self.abs_x_shift_slider.setValue(0)
                self.abs_x_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_x_shift_slider.valueChanged.connect(lambda: self.abs_x_shift_slider.setFocus())
                self.abs_x_shift_label = QLabel("0.00%"); self.abs_x_shift_label.setFixedSize(80, 20)
                self.abs_x_shift_slider.valueChanged.connect(lambda val: self.abs_x_shift_label.setText(f"{val / self.percent_precision_factor:.2f}%"))
                global_adjust_layout.addWidget(self.abs_x_shift_slider, 0, 1); global_adjust_layout.addWidget(self.abs_x_shift_label, 0, 2)
                global_adjust_layout.addWidget(QLabel("Shift Y (% Img H):"), 1, 0)
                self.abs_y_shift_slider = QSlider(Qt.Horizontal)
                self.abs_y_shift_slider.setRange(int(-100 * self.percent_precision_factor), int(100 * self.percent_precision_factor)); self.abs_y_shift_slider.setValue(0)
                self.abs_y_shift_slider.valueChanged.connect(self._update_global_adjustments)
                self.abs_y_shift_slider.valueChanged.connect(lambda: self.abs_y_shift_slider.setFocus())
                self.abs_y_shift_label = QLabel("0.00%"); self.abs_y_shift_label.setFixedSize(80, 20)
                self.abs_y_shift_slider.valueChanged.connect(lambda val: self.abs_y_shift_label.setText(f"{val / self.percent_precision_factor:.2f}%"))
                global_adjust_layout.addWidget(self.abs_y_shift_slider, 1, 1); global_adjust_layout.addWidget(self.abs_y_shift_label, 1, 2)
                global_adjust_layout.addWidget(QLabel("Scale X (%):"), 2, 0)
                self.rel_x_scale_slider = QSlider(Qt.Horizontal)
                self.rel_x_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor)); self.rel_x_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.rel_x_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_x_scale_slider.valueChanged.connect(lambda: self.rel_x_scale_slider.setFocus())
                self.rel_x_scale_label = QLabel("100.0%"); self.rel_x_scale_label.setFixedSize(80, 20)
                self.rel_x_scale_slider.valueChanged.connect(lambda val: self.rel_x_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.rel_x_scale_slider, 2, 1); global_adjust_layout.addWidget(self.rel_x_scale_label, 2, 2)
                global_adjust_layout.addWidget(QLabel("Scale Y (%):"), 3, 0)
                self.rel_y_scale_slider = QSlider(Qt.Horizontal)
                self.rel_y_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor)); self.rel_y_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.rel_y_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.rel_y_scale_slider.valueChanged.connect(lambda: self.rel_y_scale_slider.setFocus())
                self.rel_y_scale_label = QLabel("100.0%"); self.rel_y_scale_label.setFixedSize(80, 20)
                self.rel_y_scale_slider.valueChanged.connect(lambda val: self.rel_y_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.rel_y_scale_slider, 3, 1); global_adjust_layout.addWidget(self.rel_y_scale_label, 3, 2)
                global_adjust_layout.addWidget(QLabel("Font Scale (%):"), 4, 0)
                self.font_scale_slider = QSlider(Qt.Horizontal)
                self.font_scale_slider.setRange(int(10 * self.scale_precision_factor), int(300 * self.scale_precision_factor)); self.font_scale_slider.setValue(int(100 * self.scale_precision_factor))
                self.font_scale_slider.valueChanged.connect(self._update_global_adjustments)
                self.font_scale_slider.valueChanged.connect(lambda: self.font_scale_slider.setFocus())
                self.font_scale_label = QLabel("100.0%"); self.font_scale_label.setFixedSize(80, 20)
                self.font_scale_slider.valueChanged.connect(lambda val: self.font_scale_label.setText(f"{val / self.scale_precision_factor:.1f}%"))
                global_adjust_layout.addWidget(self.font_scale_slider, 4, 1); global_adjust_layout.addWidget(self.font_scale_label, 4, 2)
        
                global_adjust_layout.setColumnStretch(1, 1)
                layout.addWidget(global_adjust_group)
        
                # --- Table Widget ---
                self.table_widget = QTableWidget()
                # --- FIX: Increase column count and update headers ---
                self.table_widget.setColumnCount(9)
                self.table_widget.setHorizontalHeaderLabels(["Type", "Text/Label", "Coordinates", "Font", "Size", "Bold", "Italic", "Color", "Actions"])
                # --- END FIX ---
                self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
                self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
                self.table_widget.setEditTriggers(QAbstractItemView.AnyKeyPressed | QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
                self.table_widget.setSortingEnabled(True)
                self.table_widget.itemChanged.connect(self.handle_item_changed)
                self.table_widget.cellDoubleClicked.connect(self.handle_cell_double_clicked)
                layout.addWidget(self.table_widget)
                self.table_widget.resizeColumnsToContents()
                self.table_widget.horizontalHeader().setStretchLastSection(True)
        
                # --- Dialog Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept)
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box)
                self.setLayout(layout)
        
            def _update_global_adjustments(self):
                """Recalculates all marker positions and font sizes based on sliders."""
                if self._block_signals: return
        
                # Get values from all sliders
                abs_x_shift_percent = self.abs_x_shift_slider.value() / self.percent_precision_factor
                abs_y_shift_percent = self.abs_y_shift_slider.value() / self.percent_precision_factor
                rel_x_scale_factor = self.rel_x_scale_slider.value() / self.scale_precision_factor / 100.0
                rel_y_scale_factor = self.rel_y_scale_slider.value() / self.scale_precision_factor / 100.0
                font_size_scale_factor = self.font_scale_slider.value() / self.scale_precision_factor / 100.0
        
                abs_x_shift_pixels = (abs_x_shift_percent / 100.0) * self._current_image_width
                abs_y_shift_pixels = (abs_y_shift_percent / 100.0) * self._current_image_height
        
                # Rebuild the working `self.markers` list from the pristine `_original_markers_data`
                for i, original_marker_tuple in enumerate(self._original_markers_data):
                    orig_x, orig_y, text, qcolor, font_family, orig_font_size, is_bold, is_italic = original_marker_tuple
                    
                    final_x = (orig_x * rel_x_scale_factor) + abs_x_shift_pixels
                    final_y = (orig_y * rel_y_scale_factor) + abs_y_shift_pixels
                    final_font_size = max(1, int(round(orig_font_size * font_size_scale_factor)))
                    
                    # Rebuild the entire marker entry to ensure all properties are preserved
                    self.markers[i] = [final_x, final_y, text, qcolor, font_family, final_font_size, is_bold, is_italic]
        
                self.populate_table()
                self.global_markers_adjusted.emit(list(self.markers))
        
            def populate_table(self):
                """Rebuilds the entire table from the current state of the data model."""
                self._block_signals = True
                self.table_widget.setSortingEnabled(False)
                self.table_widget.clearContents()
                self.table_widget.setRowCount(len(self.markers) + len(self.shapes))
                standard_table_font = QFont("Arial", 12)

                # --- Populate Markers (Unchanged) ---
                for i, marker_data in enumerate(self.markers):
                    x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_data
                    
                    type_item = QTableWidgetItem("Marker")
                    type_item.setData(Qt.UserRole, {'type': 'marker', 'original_index': i})
                    text_item = QTableWidgetItem(text)
                    coord_item = QTableWidgetItem(f"{x:.1f},{y:.1f}")
                    size_item = QTableWidgetItem(str(font_size))

                    items_to_set = [type_item, text_item, coord_item, size_item]
                    for col, item in enumerate(items_to_set):
                        item.setFont(standard_table_font)
                        if col in [1, 2, 3]: # Text, Coords, Size are editable
                            item.setFlags(item.flags() | Qt.ItemIsEditable)
                        
                        target_col = col if col < 3 else 4 # Skip col 3 (Font widget), go to col 4 (Size)
                        self.table_widget.setItem(i, target_col, item)

                    font_combo = QFontComboBox()
                    font_combo.setCurrentFont(QFont(font_family))
                    font_combo.currentFontChanged.connect(self.handle_marker_style_changed_from_widget)
                    self.table_widget.setCellWidget(i, 3, font_combo)

                    bold_checkbox = QCheckBox(); bold_checkbox.setChecked(is_bold)
                    bold_checkbox.stateChanged.connect(self.handle_marker_style_changed_from_widget)
                    cell_widget_bold = QWidget(); layout_bold = QHBoxLayout(cell_widget_bold); layout_bold.addWidget(bold_checkbox); layout_bold.setAlignment(Qt.AlignCenter); layout_bold.setContentsMargins(0,0,0,0)
                    self.table_widget.setCellWidget(i, 5, cell_widget_bold)

                    italic_checkbox = QCheckBox(); italic_checkbox.setChecked(is_italic)
                    italic_checkbox.stateChanged.connect(self.handle_marker_style_changed_from_widget)
                    cell_widget_italic = QWidget(); layout_italic = QHBoxLayout(cell_widget_italic); layout_italic.addWidget(italic_checkbox); layout_italic.setAlignment(Qt.AlignCenter); layout_italic.setContentsMargins(0,0,0,0)
                    self.table_widget.setCellWidget(i, 6, cell_widget_italic)

                    color_item = QTableWidgetItem(qcolor.name())
                    color_item.setBackground(QBrush(qcolor)); color_item.setForeground(QBrush(Qt.white if qcolor.lightness() < 128 else Qt.black))
                    color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable); color_item.setFont(standard_table_font)
                    self.table_widget.setItem(i, 7, color_item)
                    
                    delete_button = QPushButton("Delete"); delete_button.clicked.connect(self.delete_item)
                    self.table_widget.setCellWidget(i, 8, delete_button)

                # --- Populate Shapes (Modified) ---
                marker_count = len(self.markers)
                for i, shape_data in enumerate(self.shapes):
                    row_idx = marker_count + i
                    shape_type = shape_data.get('type', 'Unknown').capitalize()
                    type_item = QTableWidgetItem(shape_type); type_item.setData(Qt.UserRole, {'type': 'shape', 'original_index': i})
                    
                    # Text column unused for shapes
                    text_item = QTableWidgetItem(""); text_item.setFlags(text_item.flags() & ~Qt.ItemIsEditable)
                    
                    # Coordinates
                    details_str, tooltip_str = "", ""
                    if shape_type == 'Line':
                        start, end = shape_data.get('start', (0,0)), shape_data.get('end', (0,0))
                        details_str = f"{start[0]:.1f},{start[1]:.1f},{end[0]:.1f},{end[1]:.1f}"; tooltip_str = "Edit format: X1,Y1,X2,Y2"
                    elif shape_type == 'Rectangle':
                        x, y, w, h = shape_data.get('rect', (0,0,0,0))
                        details_str = f"{x:.1f},{y:.1f},{w:.1f},{h:.1f}"; tooltip_str = "Edit format: X,Y,Width,Height"
                    coord_item = QTableWidgetItem(details_str); coord_item.setToolTip(tooltip_str)
                    coord_item.setFlags(coord_item.flags() | Qt.ItemIsEditable)

                    # Col 3 (Font) - Placeholder for shapes
                    font_item = QTableWidgetItem("Shape"); font_item.setFlags(font_item.flags() & ~Qt.ItemIsEditable)

                    # --- FIX: Col 4 (Size) - Thickness, editable ---
                    thickness = int(shape_data.get('thickness', 1))
                    size_item = QTableWidgetItem(str(thickness))
                    size_item.setFlags(size_item.flags() | Qt.ItemIsEditable)
                    # --- END FIX ---

                    # Color
                    qcolor = QColor(shape_data.get('color', '#000000'))
                    color_item = QTableWidgetItem(qcolor.name()); color_item.setBackground(QBrush(qcolor))
                    color_item.setForeground(QBrush(Qt.white if qcolor.lightness() < 128 else Qt.black))
                    color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
                    
                    # Set items and fonts
                    items_to_set = [(0, type_item), (1, text_item), (2, coord_item), (3, font_item), (4, size_item), (7, color_item)]
                    for c_idx, item in items_to_set:
                        item.setFont(standard_table_font)
                        self.table_widget.setItem(row_idx, c_idx, item)
                    
                    delete_button = QPushButton("Delete"); delete_button.clicked.connect(self.delete_item)
                    self.table_widget.setCellWidget(row_idx, 8, delete_button)

                self.table_widget.resizeColumnsToContents()
                self.table_widget.setSortingEnabled(True)
                self._block_signals = False
        
            def handle_marker_style_changed_from_widget(self):
                """Unified handler for font combobox and style checkboxes."""
                if self._block_signals: return
        
                sender_widget = self.sender()
                if not sender_widget: return
        
                for row in range(self.table_widget.rowCount()):
                    # Check if the sender is a widget in the current row
                    font_widget = self.table_widget.cellWidget(row, 3)
                    bold_widget = self.table_widget.cellWidget(row, 5)
                    italic_widget = self.table_widget.cellWidget(row, 6)
                    
                    is_font_combo = (isinstance(font_widget, QFontComboBox) and sender_widget is font_widget)
                    is_bold_checkbox = (bold_widget and sender_widget is bold_widget.findChild(QCheckBox))
                    is_italic_checkbox = (italic_widget and sender_widget is italic_widget.findChild(QCheckBox))

                    if is_font_combo or is_bold_checkbox or is_italic_checkbox:
                        type_item = self.table_widget.item(row, 0)
                        if not type_item: return
        
                        item_data = type_item.data(Qt.UserRole)
                        if not item_data or item_data.get('type') != 'marker': return
                        
                        original_marker_index = item_data.get('original_index')
                        if original_marker_index is None or not (0 <= original_marker_index < len(self.markers)): return
                        
                        # --- Update the data model based on which widget sent the signal ---
                        if is_font_combo:
                            new_font_family = sender_widget.currentFont().family()
                            self.markers[original_marker_index][4] = new_font_family
                            temp_list = list(self._original_markers_data[original_marker_index]); temp_list[4] = new_font_family
                            self._original_markers_data[original_marker_index] = tuple(temp_list)
                        elif is_bold_checkbox:
                            is_checked = sender_widget.isChecked()
                            self.markers[original_marker_index][6] = is_checked
                            temp_list = list(self._original_markers_data[original_marker_index]); temp_list[6] = is_checked
                            self._original_markers_data[original_marker_index] = tuple(temp_list)
                        elif is_italic_checkbox:
                            is_checked = sender_widget.isChecked()
                            self.markers[original_marker_index][7] = is_checked
                            temp_list = list(self._original_markers_data[original_marker_index]); temp_list[7] = is_checked
                            self._original_markers_data[original_marker_index] = tuple(temp_list)
                        
                        # --- FIX: Repopulate the table to ensure UI consistency ---
                        self.populate_table()
                        self.global_markers_adjusted.emit(list(self.markers))
                        return
        
            def handle_item_changed(self, item):
                """Updates the data model when an editable cell's text is changed."""
                if self._block_signals: return
        
                row, col = item.row(), item.column()
                type_item = self.table_widget.item(row, 0)
                if not type_item: return
                item_data = type_item.data(Qt.UserRole)
                if not item_data or item_data['type'] == 'error': return
                
                item_type, original_index = item_data['type'], item_data['original_index']
                new_value = item.text()
                
                if item_type == 'marker':
                    # ... (Marker editing logic remains unchanged) ...
                    if col == 1: # Text/Label
                        self.markers[original_index][2] = new_value; temp_list = list(self._original_markers_data[original_index]); temp_list[2] = new_value; self._original_markers_data[original_index] = tuple(temp_list)
                    elif col == 2: # Coordinates
                        try:
                            x_str, y_str = new_value.split(','); new_x, new_y = float(x_str), float(y_str)
                            self.markers[original_index][0] = new_x; self.markers[original_index][1] = new_y
                            abs_x_shift = (self.abs_x_shift_slider.value() / self.percent_precision_factor / 100.0) * self._current_image_width
                            abs_y_shift = (self.abs_y_shift_slider.value() / self.percent_precision_factor / 100.0) * self._current_image_height
                            rel_x_scale = self.rel_x_scale_slider.value() / self.scale_precision_factor / 100.0; rel_y_scale = self.rel_y_scale_slider.value() / self.scale_precision_factor / 100.0
                            base_x = (new_x - abs_x_shift) / rel_x_scale if rel_x_scale != 0 else new_x; base_y = (new_y - abs_y_shift) / rel_y_scale if rel_y_scale != 0 else new_y
                            temp_list = list(self._original_markers_data[original_index]); temp_list[0], temp_list[1] = base_x, base_y; self._original_markers_data[original_index] = tuple(temp_list)
                        except ValueError:
                            self._block_signals = True; prev_x, prev_y = self.markers[original_index][0:2]; item.setText(f"{prev_x:.1f},{prev_y:.1f}"); self._block_signals = False
                            QMessageBox.warning(self, "Invalid Input", "Coordinates must be in 'X,Y' format.")
                    elif col == 4: # Size
                        try:
                            new_size = int(new_value)
                            if new_size < 1: new_size = 1
                            self.markers[original_index][5] = new_size
                            temp_list = list(self._original_markers_data[original_index]); temp_list[5] = new_size; self._original_markers_data[original_index] = tuple(temp_list)
                        except ValueError:
                            self._block_signals = True; item.setText(str(self.markers[original_index][5])); self._block_signals = False
                            QMessageBox.warning(self, "Invalid Input", "Font size must be an integer.")
                
                elif item_type == 'shape':
                    shape_data = self.shapes[original_index]
                    
                    if col == 2: # Coordinates
                        shape_type_internal = shape_data.get('type')
                        try:
                            coords = [float(c.strip()) for c in new_value.split(',')]
                            if shape_type_internal == 'line' and len(coords) == 4: shape_data['start'], shape_data['end'] = (coords[0], coords[1]), (coords[2], coords[3])
                            elif shape_type_internal == 'rectangle' and len(coords) == 4:
                                if coords[2] < 0 or coords[3] < 0: raise ValueError("Width/Height must be non-negative.")
                                shape_data['rect'] = (coords[0], coords[1], coords[2], coords[3])
                            else: raise ValueError("Incorrect number of coordinates.")
                        except ValueError as e:
                            self._block_signals = True; self.populate_table(); self._block_signals = False
                            QMessageBox.warning(self, "Invalid Input", f"Could not parse shape coordinates.\nError: {e}")
                    
                    # --- FIX: Handle Thickness in Column 4 ---
                    elif col == 4: # Size/Thickness
                        try:
                            new_thickness = int(new_value)
                            if new_thickness < 1: new_thickness = 1
                            shape_data['thickness'] = new_thickness
                        except ValueError:
                            self._block_signals = True
                            current_thickness = shape_data.get('thickness', 1)
                            item.setText(str(current_thickness))
                            self._block_signals = False
                            QMessageBox.warning(self, "Invalid Input", "Thickness must be an integer.")
                    # --- END FIX ---
        
                self.shapes_adjusted_preview.emit(list(self.shapes))
                self.global_markers_adjusted.emit(list(self.markers))
        
            def handle_cell_double_clicked(self, row, column):
                """Handles dialogs for non-text edits (color)."""
                # Let itemChanged handle editable text columns (1, 2, 4), and widgets handle 3, 5, 6
                if column != 7: return # Only color (col 7) is handled here now
        
                item_data = self.table_widget.item(row, 0).data(Qt.UserRole)
                if not item_data or item_data['type'] == 'error': return
                item_type, original_index = item_data['type'], item_data['original_index']
        
                if column == 7: # Color
                    current_color = QColor(self.markers[original_index][3] if item_type == 'marker' else self.shapes[original_index]['color'])
                    new_color = QColorDialog.getColor(current_color, self, "Select Color")
                    if new_color.isValid():
                        if item_type == 'marker':
                            self.markers[original_index][3] = new_color
                            temp_list = list(self._original_markers_data[original_index]); temp_list[3] = new_color; self._original_markers_data[original_index] = tuple(temp_list)
                        else: self.shapes[original_index]['color'] = new_color.name()
                        self.populate_table()
                
                self.shapes_adjusted_preview.emit(list(self.shapes))
                self.global_markers_adjusted.emit(list(self.markers))
        
            def delete_item(self):
                """Deletes an item from the data model and repopulates the table."""
                clicked_button = self.sender()
                if not clicked_button: return
        
                # Find the visual row of the button that was clicked
                for row in range(self.table_widget.rowCount()):
                    if self.table_widget.cellWidget(row, 7) is clicked_button:
                        item_data = self.table_widget.item(row, 0).data(Qt.UserRole)
                        if not item_data: return
                        
                        item_type, original_index = item_data['type'], item_data['original_index']
                        
                        if item_type == 'marker' and 0 <= original_index < len(self.markers):
                            del self.markers[original_index]
                            del self._original_markers_data[original_index]
                        elif item_type == 'shape' and 0 <= original_index < len(self.shapes):
                            del self.shapes[original_index]
        
                        # Repopulate the table to reflect the deletion
                        self.populate_table()
                        self.global_markers_adjusted.emit(list(self.markers))
                        self.shapes_adjusted_preview.emit(list(self.shapes))
                        return
        
            def get_modified_markers_and_shapes(self):
                """Returns the final state of the data model."""
                return [tuple(m) for m in self.markers], self.shapes

        class TableWindow(QDialog):
            HISTORY_FILE_NAME = "analysis_history.json"
            current_lane_pil_images = {} 

            def __init__(self, current_peak_areas_data, current_standard_dictionary,
                         current_is_standard_mode, current_calculated_quantities_data,
                         parent_app_instance=None, peak_details_data=None):
                super().__init__(parent_app_instance)
                self.setWindowTitle("Analysis Results and History")
                
                # --- START FIX: Dynamic Sizing ---
                screen_geometry = QGuiApplication.primaryScreen().availableGeometry()
                # Target 60% width, 80% height, constrained to available screen
                w = min(int(screen_geometry.width() * 0.5), 1200)
                h = min(int(screen_geometry.height() * 0.75), screen_geometry.height())
                self.resize(max(600, w), max(500, h))
                self.temp_clipboard_file_path = None
                self.parent_app = parent_app_instance
                self.current_results_data = {} 
                self.is_current_data_multi_lane = isinstance(current_peak_areas_data, dict)
                self.current_lane_pil_images = {}
                self.current_peak_details_data = peak_details_data if peak_details_data else {}
                
                self.current_model_name = "Linear"
                self.current_fit_params = None

                if self.is_current_data_multi_lane:
                    for lane_id_key in current_peak_areas_data.keys():
                        self.current_results_data[lane_id_key] = {
                            'areas': current_peak_areas_data.get(lane_id_key, []),
                            'quantities': current_calculated_quantities_data.get(lane_id_key, []) if current_calculated_quantities_data else [],
                            'details': self.current_peak_details_data.get(lane_id_key, []) if peak_details_data else []
                        }
                        if self.parent_app and hasattr(self.parent_app, 'multi_lane_definitions'):
                            for lane_def in self.parent_app.multi_lane_definitions:
                                if lane_def['id'] == lane_id_key: 
                                    q_img_region = None
                                    if lane_def['type'] == 'quad': q_img_region = self.parent_app.quadrilateral_to_rect(self.parent_app.image, lane_def['points_label'])
                                    elif lane_def['type'] == 'rectangle':
                                        img_coords = self.parent_app._map_label_rect_to_image_rect(lane_def['points_label'][0])
                                        if img_coords and self.parent_app.image: q_img_region = self.parent_app.image.copy(*img_coords)
                                    if q_img_region and not q_img_region.isNull():
                                        pil_for_lane = self.parent_app.convert_qimage_to_grayscale_pil(q_img_region)
                                        if pil_for_lane: self.current_lane_pil_images[lane_id_key] = pil_for_lane 
                                    break
                elif current_peak_areas_data is not None: 
                     self.current_results_data[1] = {
                            'areas': current_peak_areas_data,
                            'quantities': current_calculated_quantities_data if current_calculated_quantities_data is not None else [],
                            'details': self.current_peak_details_data.get(1, []) if peak_details_data else [] 
                     }
                     if self.parent_app and self.parent_app.image: 
                        extracted_qimage_single = None
                        if self.parent_app.live_view_label.quad_points and len(self.parent_app.live_view_label.quad_points) == 4: extracted_qimage_single = self.parent_app.quadrilateral_to_rect(self.parent_app.image, self.parent_app.live_view_label.quad_points)
                        elif self.parent_app.live_view_label.bounding_box_preview:
                            img_coords_s = self.parent_app._map_label_rect_to_image_rect(QRectF(QPointF(self.parent_app.live_view_label.bounding_box_preview[0],self.parent_app.live_view_label.bounding_box_preview[1]),QPointF(self.parent_app.live_view_label.bounding_box_preview[2],self.parent_app.live_view_label.bounding_box_preview[3])).normalized())
                            if img_coords_s: extracted_qimage_single = self.parent_app.image.copy(*img_coords_s)
                        if extracted_qimage_single and not extracted_qimage_single.isNull():
                            pil_for_lane = self.parent_app.convert_qimage_to_grayscale_pil(extracted_qimage)
                            if pil_for_lane: self.current_lane_pil_images[1] = pil_for_lane

                self.current_standard_dictionary = current_standard_dictionary if current_standard_dictionary is not None else {}
                self.current_is_standard_mode = current_is_standard_mode
                self.source_image_name_current = "Unknown"
                if self.parent_app and hasattr(self.parent_app, 'image_path') and self.parent_app.image_path:
                    if isinstance(self.parent_app.image_path, str): self.source_image_name_current = os.path.basename(self.parent_app.image_path)
                self.current_analysis_custom_name = self.source_image_name_current
                self.analysis_name_input_widget = None
                self.analysis_history = [] 
                self.delete_entry_button = None; self.export_previous_button = None; self.previous_sessions_listwidget = None
                self.previous_results_table = None; self.previous_plot_placeholder_label = None
                self.previous_plot_groupbox_layout = None; self.previous_plot_canvas_widget = None
                self.previous_lane_image_preview_label = None 
                self.previous_lane_tab_widget = None 
                self._load_history() 
                main_layout = QVBoxLayout(self); self.tab_widget = QTabWidget(); main_layout.addWidget(self.tab_widget)
                self._create_current_results_tab(); self._create_previous_results_tab() 
                self.dialog_button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                self.dialog_button_box.accepted.connect(self._accept_dialog_and_save_current); self.dialog_button_box.rejected.connect(self.reject)
                main_layout.addWidget(self.dialog_button_box); self.setLayout(main_layout)
                if self.current_results_data: self.tab_widget.setCurrentIndex(0)
                elif self.analysis_history: self.tab_widget.setCurrentIndex(1)
                else: self.tab_widget.setCurrentIndex(0)
                
                # Initial update for current tab if data exists
                if self.current_results_data:
                    self._update_analysis_display_for_tab(is_for_history=False)

            def _get_available_models(self):
                models = ["Linear", "Polynomial (Deg 2)", "Polynomial (Deg 3)"]
                if SCIPY_AVAILABLE:
                    models.extend(["4-PL Sigmoidal"])
                return models

            def _accept_dialog_and_save_current(self):
                if self.current_results_data: 
                    peak_dialog_settings_current = {}
                    if self.parent_app and hasattr(self.parent_app, 'peak_dialog_settings'):
                        peak_dialog_settings_current = self.parent_app.peak_dialog_settings.copy()
                    user_defined_analysis_name = self.analysis_name_input_widget.text().strip() if self.analysis_name_input_widget else ""
                    display_name_for_history = user_defined_analysis_name if user_defined_analysis_name else self.source_image_name_current
                    
                    new_entry = {
                        "timestamp": datetime.datetime.now().isoformat(),
                        "user_defined_name": display_name_for_history,
                        "source_image_name": self.source_image_name_current,
                        "is_multi_lane": self.is_current_data_multi_lane,
                        "results_data": self.current_results_data,
                        "standard_dictionary": self.current_standard_dictionary,
                        "analysis_settings": peak_dialog_settings_current,
                        "quantification_model": self.current_model_name # --- NEW: Save the selected model ---
                    }
                    self.analysis_history.insert(0, new_entry)
                    self._save_history()
                    if self.previous_sessions_listwidget: self._populate_previous_sessions_list()
                self.accept()

            def _create_current_results_tab(self):
                current_tab_widget = QWidget()
                current_main_layout = QVBoxLayout(current_tab_widget)
                current_main_layout.setSpacing(10)
                name_group = QGroupBox("Analysis Identification")
                name_layout = QHBoxLayout(name_group)
                name_layout.addWidget(QLabel("Analysis Name:"))
                self.analysis_name_input_widget = QLineEdit(self.current_analysis_custom_name)
                name_layout.addWidget(self.analysis_name_input_widget)
                current_main_layout.addWidget(name_group)
                
                # --- MODIFIED: Create plot group with model selector ---
                self.plot_group_current = QGroupBox("Standard Curve (Current Analysis)")
                plot_layout_current = QVBoxLayout(self.plot_group_current)
                model_layout_current = QHBoxLayout()
                model_layout_current.addWidget(QLabel("Regression Model:"))
                self.model_combo_current = QComboBox()
                self.model_combo_current.addItems(self._get_available_models())
                self.model_combo_current.currentTextChanged.connect(lambda: self._update_analysis_display_for_tab(is_for_history=False))
                model_layout_current.addWidget(self.model_combo_current, 1)
                plot_layout_current.addLayout(model_layout_current)
                self.plot_placeholder_current = QWidget() # Placeholder for the canvas
                plot_layout_current.addWidget(self.plot_placeholder_current, 1)
                self.plot_group_current.setMinimumHeight(280)
                current_main_layout.addWidget(self.plot_group_current)
                
                results_group = QGroupBox("Band Analysis")
                results_layout = QVBoxLayout(results_group)
                results_display_area = QWidget()
                self.results_display_layout_current = QVBoxLayout(results_display_area)
                self.results_display_layout_current.setContentsMargins(0,0,0,0)
                results_layout.addWidget(results_display_area)
                current_main_layout.addWidget(results_group, 1)
                current_buttons_layout = QHBoxLayout()
                copy_current_button = QPushButton("Copy Active Lane Table")
                copy_current_button.clicked.connect(self._copy_active_lane_table_data)
                export_current_button = QPushButton("Export All Lanes to Excel")
                export_current_button.clicked.connect(lambda: self._export_to_excel_generic(self.current_results_data, self.analysis_name_input_widget.text() or self.source_image_name_current, self.current_standard_dictionary, is_multi_lane_data=self.is_current_data_multi_lane))
                current_buttons_layout.addWidget(copy_current_button); current_buttons_layout.addStretch(); current_buttons_layout.addWidget(export_current_button)
                current_main_layout.addLayout(current_buttons_layout)
                self.tab_widget.addTab(current_tab_widget, "Current Analysis")

            def _create_previous_results_tab(self):
                previous_tab_widget = QWidget()
                previous_main_layout = QHBoxLayout(previous_tab_widget)
                left_pane_widget = QWidget(); left_layout = QVBoxLayout(left_pane_widget)
                left_layout.addWidget(QLabel("Saved Analyses:"))
                self.previous_sessions_listwidget = QListWidget(); self.previous_sessions_listwidget.itemSelectionChanged.connect(self._on_history_session_selected)
                left_layout.addWidget(self.previous_sessions_listwidget)
                history_buttons_layout = QHBoxLayout()
                self.delete_entry_button = QPushButton("Delete Selected"); self.delete_entry_button.clicked.connect(self._delete_selected_history_entry); self.delete_entry_button.setEnabled(False)
                self.clear_history_button = QPushButton("Clear All History"); self.clear_history_button.clicked.connect(self._clear_all_history)
                history_buttons_layout.addWidget(self.delete_entry_button); history_buttons_layout.addStretch(); history_buttons_layout.addWidget(self.clear_history_button)
                left_layout.addLayout(history_buttons_layout); previous_main_layout.addWidget(left_pane_widget, 1)
                
                right_pane_history_widget = QWidget(); right_layout = QVBoxLayout(right_pane_history_widget)
                
                # --- MODIFIED: Create plot group with model selector for history ---
                self.previous_plot_groupbox = QGroupBox("Standard Curve (Selected History)")
                self.previous_plot_groupbox_layout = QVBoxLayout(self.previous_plot_groupbox)
                model_layout_history = QHBoxLayout()
                model_layout_history.addWidget(QLabel("Regression Model:"))
                self.model_combo_history = QComboBox()
                self.model_combo_history.addItems(self._get_available_models())
                self.model_combo_history.currentTextChanged.connect(lambda: self._update_analysis_display_for_tab(is_for_history=True))
                model_layout_history.addWidget(self.model_combo_history, 1)
                self.previous_plot_groupbox_layout.addLayout(model_layout_history)
                self.plot_placeholder_history = QWidget() # Placeholder for the canvas
                self.previous_plot_placeholder_label = QLabel("Select an analysis from the list to view details."); self.previous_plot_placeholder_label.setAlignment(Qt.AlignCenter)
                self.plot_placeholder_history.setLayout(QVBoxLayout()); self.plot_placeholder_history.layout().addWidget(self.previous_plot_placeholder_label)
                self.previous_plot_groupbox_layout.addWidget(self.plot_placeholder_history, 1)
                self.previous_plot_groupbox.setMinimumHeight(280)
                right_layout.addWidget(self.previous_plot_groupbox)

                self.history_results_display_container = QWidget()
                self.history_results_display_layout = QVBoxLayout(self.history_results_display_container); self.history_results_display_layout.setContentsMargins(0,0,0,0)
                initial_hist_table_placeholder = QLabel("Lane data will appear here."); initial_hist_table_placeholder.setAlignment(Qt.AlignCenter)
                self.history_results_display_layout.addWidget(initial_hist_table_placeholder)
                right_layout.addWidget(self.history_results_display_container, 1)
                previous_table_buttons_layout = QHBoxLayout()
                copy_previous_button = QPushButton("Copy Active History Lane Table"); copy_previous_button.clicked.connect(self._copy_active_history_lane_table_data)
                self.export_previous_button = QPushButton("Export Selected History to Excel"); self.export_previous_button.clicked.connect(self._export_selected_history_to_excel); self.export_previous_button.setEnabled(False)
                previous_table_buttons_layout.addWidget(copy_previous_button); previous_table_buttons_layout.addStretch(); previous_table_buttons_layout.addWidget(self.export_previous_button)
                right_layout.addLayout(previous_table_buttons_layout); previous_main_layout.addWidget(right_pane_history_widget, 2)
                self.tab_widget.addTab(previous_tab_widget, "Analysis History")
                self._populate_previous_sessions_list()
            
            def _update_analysis_display_for_tab(self, is_for_history):
                """Central function to update plot, quantities, and tables for a given tab."""
                if is_for_history:
                    selected_items = self.previous_sessions_listwidget.selectedItems()
                    if not selected_items: return
                    selected_row_index = self.previous_sessions_listwidget.currentRow()
                    if not (0 <= selected_row_index < len(self.analysis_history)): return
                    entry = self.analysis_history[selected_row_index]
                    
                    std_dict = entry.get("standard_dictionary", {})
                    results_data = entry.get("results_data", {})
                    is_multi_lane = entry.get("is_multi_lane", False)
                    model_name = self.model_combo_history.currentText()
                    
                    # Update quantities in the results_data dictionary
                    if std_dict and self.parent_app:
                        std_qtys = list(std_dict.keys())
                        std_areas = list(std_dict.values())
                        for lane_id, lane_content in results_data.items():
                            sample_areas = lane_content.get('areas', [])
                            if sample_areas:
                                new_qtys, _ = self.parent_app._perform_quantification(model_name, std_qtys, std_areas, sample_areas)
                                results_data[lane_id]['quantities'] = new_qtys
                    
                    # Update plot
                    new_plot = self._create_standard_curve_plot_generic(std_dict, model_name)
                    old_widget = self.plot_placeholder_history.layout().takeAt(0).widget()
                    if old_widget: old_widget.deleteLater()
                    self.plot_placeholder_history.layout().addWidget(new_plot)

                    # Update tables
                    old_content = self.history_results_display_layout.takeAt(0).widget()
                    if old_content: old_content.deleteLater()
                    new_content = self._create_results_display_widget(results_data, std_dict, is_multi_lane, is_for_history=True)
                    self.history_results_display_layout.addWidget(new_content)

                else: # For current tab
                    model_name = self.model_combo_current.currentText()
                    self.current_model_name = model_name # Store for saving history

                    if self.current_standard_dictionary and self.parent_app:
                        std_qtys = list(self.current_standard_dictionary.keys())
                        std_areas = list(self.current_standard_dictionary.values())
                        for lane_id, lane_content in self.current_results_data.items():
                            sample_areas = lane_content.get('areas', [])
                            if sample_areas:
                                new_qtys, params = self.parent_app._perform_quantification(model_name, std_qtys, std_areas, sample_areas)
                                self.current_results_data[lane_id]['quantities'] = new_qtys
                                self.current_fit_params = params
                    
                    new_plot = self._create_standard_curve_plot_generic(self.current_standard_dictionary, model_name)
                    old_widget = self.plot_placeholder_current.layout().takeAt(0).widget() if self.plot_placeholder_current.layout() else None
                    if old_widget: old_widget.deleteLater()
                    if not self.plot_placeholder_current.layout(): self.plot_placeholder_current.setLayout(QVBoxLayout())
                    self.plot_placeholder_current.layout().addWidget(new_plot)
                    
                    old_content = self.results_display_layout_current.takeAt(0).widget() if self.results_display_layout_current.count() > 0 else None
                    if old_content: old_content.deleteLater()
                    new_content = self._create_results_display_widget(self.current_results_data, self.current_standard_dictionary, self.is_current_data_multi_lane, is_for_history=False)
                    self.results_display_layout_current.addWidget(new_content)
            
            def _create_results_display_widget(self, results_data, std_dict, is_multi_lane, is_for_history):
                container = QWidget()
                layout = QVBoxLayout(container); layout.setContentsMargins(0,0,0,0)
                is_std_mode = bool(std_dict)

                if is_multi_lane and len(results_data) > 0:
                    tabs = QTabWidget()
                    for lane_id in sorted(results_data.keys()):
                        lane_data = results_data[lane_id]
                        pil_img = self.current_lane_pil_images.get(lane_id) if not is_for_history else None
                        details = lane_data.get('details', [])
                        widget = self._create_lane_data_display_widget(lane_id, lane_data.get('areas', []), lane_data.get('quantities', []), is_std_mode, pil_img, details, is_for_history)
                        tabs.addTab(widget, f"Lane {lane_id}")
                    layout.addWidget(tabs)
                elif 1 in results_data:
                    lane_data = results_data[1]
                    pil_img = self.current_lane_pil_images.get(1) if not is_for_history else None
                    details = lane_data.get('details', [])
                    widget = self._create_lane_data_display_widget(1, lane_data.get('areas',[]), lane_data.get('quantities',[]), is_std_mode, pil_img, details, is_for_history)
                    layout.addWidget(widget)
                else:
                    layout.addWidget(QLabel("No analysis data to display.", alignment=Qt.AlignCenter))
                return container

            def _on_history_session_selected(self):
                if not self.previous_sessions_listwidget: return
                selected_items = self.previous_sessions_listwidget.selectedItems()
                if not selected_items or "No history available." in selected_items[0].text():
                    self._clear_previous_details_view(); return

                selected_row_index = self.previous_sessions_listwidget.currentRow()
                if not (0 <= selected_row_index < len(self.analysis_history)):
                    self._clear_previous_details_view(); return
                
                entry = self.analysis_history[selected_row_index]
                self.delete_entry_button.setEnabled(True); self.export_previous_button.setEnabled(True)

                # --- NEW: Set the model combo box from history ---
                saved_model = entry.get("quantification_model", "Linear") # Default to Linear if not saved
                self.model_combo_history.blockSignals(True)
                self.model_combo_history.setCurrentText(saved_model)
                self.model_combo_history.blockSignals(False)

                self._update_analysis_display_for_tab(is_for_history=True)
            
            def _create_standard_curve_plot_generic(self, standard_dictionary, model_name):
                is_std_mode = bool(standard_dictionary)
                if not is_std_mode or len(standard_dictionary) < 2:
                    return QLabel("Standard curve requires at least 2 standard points.", alignment=Qt.AlignCenter)
                try:
                    quantities = np.array(list(standard_dictionary.keys()), dtype=float)
                    areas = np.array(list(standard_dictionary.values()), dtype=float)
                    
                    is_dark_theme = self.parent_app and self.parent_app.current_theme == "dark"
                    if is_dark_theme:
                        bg_color, ax_bg_color, text_color, spine_color, grid_color, scatter_color, line_color = '#2D2D30', '#38383C', '#F1F1F1', '#707070', '#5A5A60', '#FF8A65', '#42A5F5'
                    else:
                        bg_color, ax_bg_color, text_color, spine_color, grid_color, scatter_color, line_color = 'white', 'white', 'black', '#555555', '#DDDDDD', 'red', 'blue'

                    fig, ax = plt.subplots(figsize=(4.5, 3.2)); fig.set_dpi(90)
                    fig.patch.set_facecolor(bg_color); ax.patch.set_facecolor(ax_bg_color)
                    for spine in ax.spines.values(): spine.set_color(spine_color)
                    ax.tick_params(axis='x', colors=text_color); ax.tick_params(axis='y', colors=text_color)
                    ax.yaxis.label.set_color(text_color); ax.xaxis.label.set_color(text_color); ax.title.set_color(text_color)

                    ax.scatter(quantities, areas, label='Standard Points', color=scatter_color, zorder=5, s=30)
                    
                    # --- NEW: Model fitting and plotting ---
                    q_min, q_max = np.min(quantities), np.max(quantities)
                    x_line = np.linspace(q_min, q_max, 200)
                    y_line = None
                    fit_label = f'{model_name}: Fit Failed'
                    
                    try:
                        if "Linear" in model_name or "Polynomial" in model_name:
                            degree = 1
                            if "Deg 2" in model_name: degree = 2
                            elif "Deg 3" in model_name: degree = 3
                            if len(quantities) > degree:
                                params = np.polyfit(quantities, areas, degree)
                                y_line = np.polyval(params, x_line)
                                residuals = areas - np.polyval(params, quantities)
                                ss_res = np.sum(residuals**2); ss_tot = np.sum((areas - np.mean(areas))**2)
                                r_squared = 1 - (ss_res / ss_tot) if ss_tot > 1e-9 else 1.0
                                fit_label = f'{model_name} (R² = {r_squared:.4f})'
                        
                        elif SCIPY_AVAILABLE and "4-PL" in model_name and len(quantities) >= 4:
                            p0 = [min(areas), 1.0, np.median(quantities), max(areas)]
                            params, _ = curve_fit(four_param_logistic, quantities, areas, p0=p0, maxfev=10000)
                            y_line = four_param_logistic(x_line, *params)
                            fit_label = '4-PL Sigmoidal Fit'


                        if y_line is not None:
                            ax.plot(x_line, y_line, label=fit_label, color=line_color, linewidth=1.2)
                    except (RuntimeError, ValueError) as fit_error:
                         print(f"Fit Error: {fit_error}")
                         ax.text(0.5, 0.5, f"Could not fit\n{model_name}", ha='center', va='center', color='red', transform=ax.transAxes)

                    ax.set_xlabel('Known Quantity', fontsize=8); ax.set_ylabel('Measured Peak Area', fontsize=8)
                    ax.set_title('Standard Curve', fontsize=9, fontweight='bold')
                    leg = ax.legend(fontsize='xx-small', loc='best'); leg.get_frame().set_facecolor(ax.get_facecolor())
                    for text in leg.get_texts(): text.set_color(text_color)
                    ax.grid(True, linestyle=':', alpha=0.7, linewidth=0.5, color=grid_color)
                    ax.tick_params(axis='both', which='major', labelsize=7)
                    if np.any(areas > 1e4) or (np.any(areas < 1e-2) and np.any(areas != 0)): ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                    if np.any(quantities > 1e4) or (np.any(quantities < 1e-2) and np.any(quantities != 0)): ax.ticklabel_format(style='sci', axis='x', scilimits=(0,0), useMathText=True)
                    try: fig.set_constrained_layout(True)
                    except AttributeError: plt.tight_layout(pad=0.3)
                    canvas = FigureCanvas(fig); canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding); canvas.updateGeometry()
                    plt.close(fig)
                    return canvas
                except Exception as e:
                    traceback.print_exc()
                    return QLabel(f"Error generating plot:\n{str(e)[:100]}...", alignment=Qt.AlignCenter, styleSheet="color: red;")
            
            # (All other TableWindow methods like _get_config_dir, _load_history, _save_history, etc., remain unchanged)
            # ...
            # The remaining methods from the previous version of TableWindow should be pasted here.
            # I am omitting them for brevity as they are not changed by this request.
            # The key is to replace the ENTIRE class with this new structure.
            # ...
            # MAKE SURE TO PASTE THE REST OF THE UNCHANGED TableWindow METHODS HERE
            # ...
            def _get_config_dir(self):
                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()
                return application_path

            def _load_history(self):
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                if os.path.exists(history_file_path):
                    try:
                        with open(history_file_path, "r", encoding='utf-8') as f:
                            loaded_data = json.load(f)
                        if isinstance(loaded_data, list):
                            self.analysis_history = [entry for entry in loaded_data if isinstance(entry, dict)]
                        else:
                            self.analysis_history = []
                    except (json.JSONDecodeError, IOError) as e:
                        self.analysis_history = []
                else:
                    self.analysis_history = []

            def _save_history(self):
                history_file_path = os.path.join(self._get_config_dir(), self.HISTORY_FILE_NAME)
                try:
                    with open(history_file_path, "w", encoding='utf-8') as f:
                        json.dump(self.analysis_history, f, indent=4)
                except IOError as e:
                    QMessageBox.critical(self, "Save History Error", f"Could not save analysis history to {history_file_path}: {e}")
            def _create_lane_data_display_widget(self, lane_id, peak_areas, calculated_quantities, is_std_mode,
                                                 pil_lane_image=None, peak_details_for_lane=None, is_for_history=False):
                lane_widget = QWidget()
                lane_layout = QHBoxLayout(lane_widget)

                table_scroll_area = QScrollArea(); table_scroll_area.setWidgetResizable(True)
                table_for_lane = QTableWidget(); table_for_lane.setColumnCount(4)
                table_for_lane.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
                table_for_lane.setEditTriggers(QTableWidget.NoEditTriggers); table_for_lane.setSelectionBehavior(QTableWidget.SelectRows)
                single_lane_data_for_populator = {1: {'areas': peak_areas, 'quantities': calculated_quantities}}
                self._populate_table_generic(table_for_lane, single_lane_data_for_populator, is_std_mode, is_multi_lane_data=False)
                table_scroll_area.setWidget(table_for_lane)
                lane_layout.addWidget(table_scroll_area, 3)

                if pil_lane_image and not is_for_history:
                    # Create the label directly, no QGroupBox
                    lane_image_label = QLabel(f"Lane {lane_id} Preview")
                    lane_image_label.setMinimumSize(150, 200)
                    lane_image_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding) 
                    lane_image_label.setAlignment(Qt.AlignCenter)
                    lane_image_label.setStyleSheet("border: 1px solid #C0C5CB; background-color: #333; border-radius: 4px;")
                    lane_image_label.setScaledContents(True)

                    try:
                        display_pil_image_oriented = pil_lane_image.copy()
                        
                        display_pil_image_rgba = None
                        if display_pil_image_oriented.mode in ['I', 'I;16'] or display_pil_image_oriented.mode.startswith("I;16"):
                            img_array = np.array(display_pil_image_oriented, dtype=np.float32); 
                            min_val, max_val = np.percentile(img_array, 1), np.percentile(img_array, 99)
                            if max_val <= min_val: min_val, max_val = np.min(img_array), np.max(img_array)
                            normalized_array = (img_array - min_val) / (max_val - min_val + 1e-9) 
                            normalized_array = np.clip(normalized_array, 0.0, 1.0)
                            img_8bit_gray = (normalized_array * 255).astype(np.uint8)
                            display_pil_image_rgba = Image.fromarray(img_8bit_gray).convert("RGBA")
                        elif display_pil_image_oriented.mode == 'L':
                            display_pil_image_rgba = ImageOps.autocontrast(display_pil_image_oriented, cutoff=1).convert("RGBA")
                        else:
                            display_pil_image_rgba = display_pil_image_oriented.convert("RGBA")

                        if display_pil_image_rgba:
                            draw = ImageDraw.Draw(display_pil_image_rgba)
                            
                            # --- FIX: Use a fixed, large pixel font size for cross-platform consistency ---
                            band_number_pil_font_size = 28 # Fixed pixel size
                            try:
                                font_pil = ImageFont.truetype("arialbd.ttf", band_number_pil_font_size)
                            except IOError: 
                                try: font_pil = ImageFont.truetype("arial.ttf", band_number_pil_font_size)
                                except: font_pil = ImageFont.load_default()
                            
                            text_color = (255, 0, 0, 255) # Bright Red

                            if peak_details_for_lane:
                                for i, peak_info in enumerate(peak_details_for_lane):
                                    y_pixel_on_oriented_image = int(peak_info['y_coord_in_lane_image'])
                                    y_pixel_on_oriented_image = max(0, min(y_pixel_on_oriented_image, display_pil_image_rgba.height -1))
                                    band_num_str = str(i + 1)
                                    text_bbox = draw.textbbox((0, 0), band_num_str, font=font_pil)
                                    text_width = text_bbox[2] - text_bbox[0]
                                    text_height = text_bbox[3] - text_bbox[1]
                                    
                                    x_text_pos = 5 # Small padding from the left edge
                                    actual_y_draw_pos = y_pixel_on_oriented_image - (text_height // 2) - text_bbox[1] 
                                    draw.text((x_text_pos, actual_y_draw_pos), band_num_str, fill=text_color, font=font_pil)
                                    
                                    line_start_x = x_text_pos + text_width + 4
                                    line_end_x = x_text_pos + text_width + 12
                                    line_y = y_pixel_on_oriented_image 
                                    draw.line([(line_start_x, line_y), (line_end_x, line_y)], fill=(0,255,0,180), width=2)

                            q_image_lane = ImageQt.ImageQt(display_pil_image_rgba)
                            if not q_image_lane.isNull():
                                pixmap_lane = QPixmap.fromImage(q_image_lane)
                                lane_image_label.setPixmap(pixmap_lane)
                        else:
                             lane_image_label.setText(f"Cannot display Lane {lane_id}\n(format error)")
                    except Exception as e_img: 
                        print(f"Error creating lane preview for lane {lane_id}: {e_img}"); traceback.print_exc()
                        lane_image_label.setText(f"Error displaying\nlane {lane_id} preview")
                
                    lane_layout.addWidget(lane_image_label, 1) # Image takes less horizontal stretch
                return lane_widget
            def _copy_active_lane_table_data(self):
                table_to_copy = None
                
                # Get the main container widget for the current analysis results
                container_widget = self.results_display_layout_current.itemAt(0).widget() if self.results_display_layout_current.count() > 0 else None
                
                if not container_widget:
                    QMessageBox.information(self, "Copy Error", "No analysis results are currently displayed.")
                    return

                # Find the QTabWidget within the container, if it exists (for multi-lane view)
                tab_widget = container_widget.findChild(QTabWidget)
                
                active_lane_widget = None
                if tab_widget:
                    # Multi-lane view: get the currently visible tab page
                    active_lane_widget = tab_widget.currentWidget()
                else:
                    # Single-lane view: the container itself holds the lane data
                    active_lane_widget = container_widget
                
                # Now find the QTableWidget within the active lane's widget
                if active_lane_widget:
                    table_widgets_found = active_lane_widget.findChildren(QTableWidget)
                    if table_widgets_found:
                        table_to_copy = table_widgets_found[0]
                
                if table_to_copy:
                    self._copy_table_data_generic(table_to_copy)
                else:
                    QMessageBox.information(self, "Copy Error", "Could not find the active lane's table to copy.")
            def _copy_active_history_lane_table_data(self):
                table_to_copy = None
                
                # Get the main container widget for the history results
                container_widget = self.history_results_display_layout.itemAt(0).widget() if self.history_results_display_layout.count() > 0 else None
                
                if not container_widget:
                    QMessageBox.information(self, "Copy Error", "No history results are currently displayed.")
                    return

                # Find the QTabWidget within the container, if it exists
                tab_widget = container_widget.findChild(QTabWidget)
                
                active_lane_widget = None
                if tab_widget:
                    # Multi-lane history view
                    active_lane_widget = tab_widget.currentWidget()
                else:
                    # Single-lane history view
                    active_lane_widget = container_widget
                
                # Now find the QTableWidget within the active lane's widget
                if active_lane_widget:
                    table_widgets_found = active_lane_widget.findChildren(QTableWidget)
                    if table_widgets_found:
                        table_to_copy = table_widgets_found[0]
                
                if table_to_copy:
                    self._copy_table_data_generic(table_to_copy)
                else:
                    QMessageBox.information(self, "Copy Error", "Could not find the active history lane's table to copy.")
            def _populate_previous_sessions_list(self):
                self.previous_sessions_listwidget.clear()
                if not self.analysis_history:
                    self.previous_sessions_listwidget.addItem("No history available.")
                    if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                    if self.export_previous_button: self.export_previous_button.setEnabled(False)
                    return

                for i, entry in enumerate(self.analysis_history):
                    ts_str = entry.get("timestamp", f"Entry {len(self.analysis_history) - i}")
                    entry_display_name = entry.get("user_defined_name", "").strip()
                    source_img_name = entry.get("source_image_name", "Unknown Image")
                    
                    if not entry_display_name:
                        entry_display_name = source_img_name
                    try:
                        dt_obj = datetime.datetime.fromisoformat(ts_str.split('.')[0])
                        final_display_string = f"{dt_obj.strftime('%Y-%m-%d %H:%M:%S')} ({entry_display_name})"
                    except ValueError:
                        final_display_string = f"{ts_str} ({entry_display_name})"
                    self.previous_sessions_listwidget.addItem(final_display_string)
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)
            def _clear_previous_details_view(self):
                if hasattr(self, 'history_results_display_layout'):
                    while self.history_results_display_layout.count() > 0:
                        item = self.history_results_display_layout.takeAt(0)
                        widget = item.widget()
                        if widget: widget.deleteLater()
                    initial_hist_table_placeholder = QLabel("Select an analysis to view details.")
                    initial_hist_table_placeholder.setAlignment(Qt.AlignCenter)
                    self.history_results_display_layout.addWidget(initial_hist_table_placeholder)

                # --- START OF FIX: Use a robust while loop to clear the layout ---
                if self.plot_placeholder_history and self.plot_placeholder_history.layout():
                    while self.plot_placeholder_history.layout().count() > 0:
                        item = self.plot_placeholder_history.layout().takeAt(0)
                        if item:
                            widget = item.widget()
                            if widget:
                                widget.deleteLater()
                # --- END OF FIX ---
                
                self.plot_placeholder_history.layout().addWidget(self.previous_plot_placeholder_label)
                self.previous_plot_placeholder_label.setText("Select an analysis from the list to view details.")
                self.previous_plot_placeholder_label.show()
                
                if self.delete_entry_button: self.delete_entry_button.setEnabled(False)
                if self.export_previous_button: self.export_previous_button.setEnabled(False)
            def _delete_selected_history_entry(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    reply = QMessageBox.question(self, "Confirm Delete",
                                                 "Are you sure you want to delete this history entry?",
                                                 QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        del self.analysis_history[current_row]
                        self._save_history()
                        self._populate_previous_sessions_list()
                        self._clear_previous_details_view()
                else:
                    QMessageBox.information(self, "No Selection", "Please select an entry to delete.")
            def _clear_all_history(self):
                reply = QMessageBox.question(self, "Confirm Clear All History",
                                             "Are you sure you want to delete ALL saved analysis entries?\nThis action cannot be undone.",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    self.analysis_history = []
                    self._save_history()
                    self._populate_previous_sessions_list()
                    self._clear_previous_details_view()
            def _export_selected_history_to_excel(self):
                if not self.previous_sessions_listwidget: return
                current_row = self.previous_sessions_listwidget.currentRow()
                if current_row >= 0 and current_row < len(self.analysis_history):
                    entry = self.analysis_history[current_row]
                    ts_str_raw = entry.get("timestamp", f"History_Entry_{current_row+1}")
                    try: dt_obj_export = datetime.datetime.fromisoformat(ts_str_raw.split('.')[0]); timestamp_str_for_file = dt_obj_export.strftime("%Y%m%d_%H%M%S")
                    except ValueError: timestamp_str_for_file = ts_str_raw.replace(":", "-").replace("T", "_").split('.')[0]
                    analysis_name_part = entry.get("user_defined_name", "").strip()
                    if not analysis_name_part: analysis_name_part = entry.get("source_image_name", "UnknownAnalysis")
                    analysis_name_part_sanitized = analysis_name_part.replace('.', '_').replace(' ', '_').replace(':', '-')
                    max_name_len = 50
                    if len(analysis_name_part_sanitized) > max_name_len: analysis_name_part_sanitized = analysis_name_part_sanitized[:max_name_len]
                    default_filename_base = f"Analysis_{timestamp_str_for_file}_{analysis_name_part_sanitized}"

                    hist_is_multi_lane = entry.get("is_multi_lane", False)
                    hist_results_data = entry.get("results_data", {})
                    
                    self._export_to_excel_generic(
                        hist_results_data,
                        default_filename_base,
                        entry.get("standard_dictionary", {}),
                        is_multi_lane_data=hist_is_multi_lane
                    )
                else:
                    QMessageBox.information(self, "No Selection", "Please select a history entry to export.")
            def _populate_table_generic(self, table_widget, results_data, is_standard_mode, is_multi_lane_data):
                table_widget.clearContents()
                table_widget.setRowCount(0)
                if not results_data:
                    placeholder_text = "No data to display in table."
                    row_span = 1
                    col_span = table_widget.columnCount()
                    table_widget.setRowCount(1)
                    item = QTableWidgetItem(placeholder_text)
                    item.setTextAlignment(Qt.AlignCenter)
                    table_widget.setItem(0, 0, item)
                    if col_span > 0 : table_widget.setSpan(0, 0, row_span, col_span)
                    table_widget.resizeColumnsToContents()
                    return

                current_row_idx = 0
                for lane_id_sorted in sorted(results_data.keys()): 
                    lane_data = results_data[lane_id_sorted]
                    peak_areas = lane_data.get('areas', [])
                    calculated_quantities = lane_data.get('quantities', [])
                    if not peak_areas: continue
                    total_area_this_lane = sum(peak_areas) if peak_areas else 0.0
                    table_widget.setRowCount(current_row_idx + len(peak_areas))
                    for band_idx, area in enumerate(peak_areas):
                        col_offset = 0
                        if is_multi_lane_data:
                            table_widget.setItem(current_row_idx, 0, QTableWidgetItem(str(lane_id_sorted)))
                            col_offset = 1
                        band_label = f"Band {band_idx + 1}"
                        table_widget.setItem(current_row_idx, col_offset + 0, QTableWidgetItem(band_label))
                        table_widget.setItem(current_row_idx, col_offset + 1, QTableWidgetItem(f"{area:.3f}"))
                        percentage_str = f"{(area / total_area_this_lane * 100):.2f}%" if total_area_this_lane != 0 else "0.00%"
                        table_widget.setItem(current_row_idx, col_offset + 2, QTableWidgetItem(percentage_str))
                        quantity_str = ""
                        if is_standard_mode and calculated_quantities and band_idx < len(calculated_quantities):
                            quantity_str = f"{calculated_quantities[band_idx]:.3f}" # More precision
                        table_widget.setItem(current_row_idx, col_offset + 3, QTableWidgetItem(quantity_str))
                        current_row_idx += 1
                if current_row_idx == 0:
                    self._populate_table_generic(table_widget, None, False, False)
                else:
                    table_widget.resizeColumnsToContents()
            def _copy_table_data_generic(self, table_widget_source):
                if not table_widget_source: return
                
                clipboard_string = ""
                column_count = table_widget_source.columnCount()
                row_count = table_widget_source.rowCount()

                # --- START OF FIX: Copy all headers and all rows, ignoring selection ---
                # Add headers to clipboard string
                header_data = [table_widget_source.horizontalHeaderItem(c).text() for c in range(column_count)]
                clipboard_string += "\t".join(header_data) + "\n"

                # Add all data rows to clipboard string
                for r in range(row_count):
                    row_data = []
                    # Check for placeholder text spanning the whole row
                    first_item = table_widget_source.item(r, 0)
                    if first_item and table_widget_source.columnSpan(r, 0) > 1:
                        if "No data" in first_item.text() or "Select an analysis" in first_item.text():
                            continue # Skip placeholder rows
                    
                    for c in range(column_count):
                        item = table_widget_source.item(r, c)
                        row_data.append(item.text() if item else "")
                    
                    if any(cell_text for cell_text in row_data):
                        clipboard_string += "\t".join(row_data) + "\n"
                # --- END OF FIX ---
                
                QApplication.clipboard().setText(clipboard_string.strip())
                QMessageBox.information(self, "Copied", "The active table's content has been copied to the clipboard.")
            def _export_to_excel_generic(self, results_data_for_export, analysis_name_for_filename_base="Analysis_Results", 
                                         standard_dict_for_export=None, is_multi_lane_data=False):
                safe_analysis_name = str(analysis_name_for_filename_base).replace('.', '_').replace(' ', '_').replace(':', '-')
                current_timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                final_default_filename = ""
                if "Analysis_" in safe_analysis_name and any(char.isdigit() for char in safe_analysis_name): 
                    final_default_filename = safe_analysis_name
                else: 
                    final_default_filename = f"Analysis_{current_timestamp_str}_{safe_analysis_name}"
                options = QFileDialog.Options(); file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Excel File", f"{final_default_filename}.xlsx", "Excel Files (*.xlsx)", options=options)
                if not file_path: return
                
                workbook = openpyxl.Workbook()
                if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
                if is_multi_lane_data and results_data_for_export:
                    worksheet_data = workbook.create_sheet("Multi-Lane Analysis")
                    headers = ["Lane ID", "Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"]
                    for col, header in enumerate(headers, start=1):
                        cell = worksheet_data.cell(row=1, column=col, value=header); cell.font = Font(bold=True)
                    current_excel_row = 2
                    for lane_id_sorted in sorted(results_data_for_export.keys()):
                        lane_data = results_data_for_export[lane_id_sorted]
                        peak_areas = lane_data.get('areas', [])
                        calculated_quantities = lane_data.get('quantities', [])
                        if not peak_areas: continue
                        total_area_this_lane = sum(peak_areas)

                        for band_idx, area_val in enumerate(peak_areas):
                            worksheet_data.cell(row=current_excel_row, column=1, value=lane_id_sorted)
                            worksheet_data.cell(row=current_excel_row, column=2, value=f"Band {band_idx + 1}")
                            worksheet_data.cell(row=current_excel_row, column=3, value=float(f"{area_val:.3f}"))
                            perc_val_str = f"{(area_val / total_area_this_lane * 100):.2f}%" if total_area_this_lane != 0 else "0.00%"
                            try:
                                perc_num = float(perc_val_str.replace('%','')) / 100.0
                                cell_perc = worksheet_data.cell(row=current_excel_row, column=4, value=perc_num)
                                cell_perc.number_format = '0.00%'
                            except ValueError: worksheet_data.cell(row=current_excel_row, column=4, value=perc_val_str)
                            qty_str_val = ""
                            if self.current_is_standard_mode and calculated_quantities and band_idx < len(calculated_quantities):
                                qty_str_val = f"{calculated_quantities[band_idx]:.3f}" # More precision
                                try: worksheet_data.cell(row=current_excel_row, column=5, value=float(qty_str_val))
                                except ValueError: worksheet_data.cell(row=current_excel_row, column=5, value=qty_str_val)
                            else: worksheet_data.cell(row=current_excel_row, column=5, value=qty_str_val)
                            current_excel_row += 1
                    for col_idx_letter in range(1, worksheet_data.max_column + 1):
                        column_letter = openpyxl.utils.get_column_letter(col_idx_letter)
                        max_length = 0; header_len = len(headers[col_idx_letter-1])
                        for cell in worksheet_data[column_letter]:
                            try: 
                                if cell.value: max_length = max(max_length, len(str(cell.value)) + (1 if cell.number_format == '0.00%' else 0) )
                            except: pass
                        adjusted_width = (max(max_length, header_len) + 2) * 1.1 
                        worksheet_data.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)

                elif not is_multi_lane_data and results_data_for_export and 1 in results_data_for_export:
                    worksheet_data_single = workbook.create_sheet("Single Lane Analysis")
                    headers_single = ["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"]
                    for col, header in enumerate(headers_single, start=1):
                        cell = worksheet_data_single.cell(row=1, column=col, value=header); cell.font = Font(bold=True)
                    single_lane_actual_data = results_data_for_export[1]
                    peak_areas_single = single_lane_actual_data.get('areas', [])
                    calculated_quantities_single = single_lane_actual_data.get('quantities', [])
                    total_area_single = sum(peak_areas_single) if peak_areas_single else 0.0

                    for r_idx, area_val in enumerate(peak_areas_single):
                        excel_r = r_idx + 2
                        worksheet_data_single.cell(row=excel_r, column=1, value=f"Band {r_idx+1}")
                        worksheet_data_single.cell(row=excel_r, column=2, value=float(f"{area_val:.3f}"))
                        perc_val_str = f"{(area_val / total_area_single * 100):.2f}%" if total_area_single != 0 else "0.00%"
                        try: 
                            perc_num = float(perc_val_str.replace('%','')) / 100.0
                            cell_perc = worksheet_data_single.cell(row=excel_r, column=3, value=perc_num)
                            cell_perc.number_format = '0.00%'
                        except ValueError: worksheet_data_single.cell(row=excel_r, column=3, value=perc_val_str)
                        qty_str_val = ""
                        if self.current_is_standard_mode and calculated_quantities_single and r_idx < len(calculated_quantities_single):
                            qty_str_val = f"{calculated_quantities_single[r_idx]:.3f}" # More precision
                            try: worksheet_data_single.cell(row=excel_r, column=4, value=float(qty_str_val))
                            except ValueError: worksheet_data_single.cell(row=excel_r, column=4, value=qty_str_val)
                        else: worksheet_data_single.cell(row=excel_r, column=4, value=qty_str_val)

                    for col_idx_letter in range(1, worksheet_data_single.max_column + 1):
                        column_letter = openpyxl.utils.get_column_letter(col_idx_letter)
                        max_length = 0; header_len = len(headers_single[col_idx_letter-1])
                        for cell in worksheet_data_single[column_letter]:
                            try: 
                                if cell.value: max_length = max(max_length, len(str(cell.value)) + (1 if cell.number_format == '0.00%' else 0))
                            except: pass
                        adjusted_width = (max(max_length, header_len) + 2) * 1.1
                        worksheet_data_single.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 50)
                else:
                    no_data_sheet = workbook.create_sheet("No Analysis Data")
                    no_data_sheet.cell(row=1, column=1, value="No analysis data was available for export.")

                if standard_dict_for_export and len(standard_dict_for_export) >= 2:
                    worksheet_std = workbook.create_sheet("Standard Curve Data")
                    worksheet_std.cell(row=1, column=1, value="Known Quantity").font = Font(bold=True)
                    worksheet_std.cell(row=1, column=2, value="Total Peak Area").font = Font(bold=True)
                    current_row_std = 2
                    for qty_key, area_val in sorted(standard_dict_for_export.items()):
                        try:
                            worksheet_std.cell(row=current_row_std, column=1, value=float(qty_key))
                            worksheet_std.cell(row=current_row_std, column=2, value=float(area_val))
                            current_row_std += 1
                        except ValueError: print(f"Warning: Skipping invalid standard data for Excel: Qty={qty_key}, Area={area_val}")
                    for col_dim_std_letter in range(1, worksheet_std.max_column + 1): 
                         column_letter_std = openpyxl.utils.get_column_letter(col_dim_std_letter)
                         worksheet_std.column_dimensions[column_letter_std].auto_size = True
                try:
                    workbook.save(file_path)
                    QMessageBox.information(self, "Success", f"Table data exported to\n{file_path}")
                except Exception as e:
                     QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                        

        class PeakAreaDialog(QDialog):
            """
            Interactive dialog to adjust peak regions and calculate peak areas.
            Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
            Peak region boundaries are now manipulated directly on the image strip plot.
            Optimized for fast, interactive boundary adjustments using blitting.
            """
            HANDLE_SIZE = 2 # Pixel size for draggable handles on ax_image

            def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
                super().__init__(parent)
                self.parent_app = parent
                self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
                
                # --- START FIX: Dynamic Sizing ---
                screen_geometry = QGuiApplication.primaryScreen().availableGeometry()
                screen_width = screen_geometry.width()
                screen_height = screen_geometry.height()
                
                # Set size to 80% of screen width/height, but enforce minimums
                dialog_width = max(600, int(screen_width * 0.6))
                dialog_height = max(900, int(screen_height * 0.9))
                
                # Ensure it fits on the screen
                dialog_width = min(dialog_width, screen_width)
                dialog_height = min(dialog_height, screen_height)
                
                self.resize(dialog_width, dialog_height)

                if not isinstance(cropped_data, Image.Image):
                    raise TypeError("Input 'cropped_data' must be a PIL Image object")

                self.original_pil_cropped_data = cropped_data
                self.enhanced_cropped_image_for_display = None

                self.original_max_value = 255.0
                pil_mode = self.original_pil_cropped_data.mode
                try:
                    if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        self.original_max_value = 65535.0
                    elif pil_mode == 'L':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        self.original_max_value = 255.0
                    elif pil_mode == 'F':
                        self.intensity_array_original_range = np.array(self.original_pil_cropped_data, dtype=np.float64)
                        max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
                        self.original_max_value = max(1.0, max_in_float)
                    else:
                        gray_img = self.original_pil_cropped_data.convert("L")
                        self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                        self.original_max_value = 255.0
                except Exception as e:
                    raise TypeError(f"Could not process input image mode '{pil_mode}': {e}")

                if self.intensity_array_original_range.ndim != 2:
                    raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

                # --- START MODIFICATION: Add inversion state ---
                self.is_inverted = current_settings.get('is_inverted', False)
                # --- END MODIFICATION ---

                self.profile_original_inverted = None; self.profile = None; self.background = None
                self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
                self.smoothing_sigma = current_settings.get('smoothing_sigma', 0.0)
                self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
                self.peak_distance = current_settings.get('peak_distance', 10)
                self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.00)
                self.area_subtraction_method = current_settings.get('area_subtraction_method', "Rolling-valley")
                self.auto_adjust_rb_radius = current_settings.get('auto_adjust_rb_radius', False)
                self.peaks = np.array([]); self.initial_valley_regions = []; self.peak_regions = []
                self.denoise_sigma = current_settings.get('denoise_sigma', 0.0)
                self.peak_areas_rolling_ball = []; self.peak_areas_straight_line = []; self.peak_areas_valley = []
                self._final_settings = {}; self._persist_enabled_on_exit = persist_checked
                
                self.manual_select_mode_active = False
                self.selected_peak_for_ui_focus = -1

                self.add_peak_mode_active = False; self.selected_peak_index_for_delete = -1

                self.background_blit = None
                self.dragging_handle_info = None
                self.interactive_artists = []
                
                if rolling_ball is None or find_peaks is None or gaussian_filter1d is None or interp1d is None or cv2 is None or ImageOps is None:
                    QMessageBox.critical(self, "Dependency Error","Missing required libraries...")
                
                self._setup_ui(persist_checked)
                self.regenerate_profile_and_detect()

            def _setup_ui(self, persist_checked_initial):
                # ... (code before peak_detect_group is the same) ...
                main_layout = QVBoxLayout(self)
                main_layout.setSpacing(10)
                self.fig = plt.figure(figsize=(10, 7.5))
                gs = GridSpec(2, 1, height_ratios=[3, 1], hspace=0.1, figure=self.fig)
                self.ax = self.fig.add_subplot(gs[0])
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)
                self.canvas = FigureCanvas(self.fig)
                self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.canvas.mpl_connect('draw_event', self.on_draw)
                self.canvas.mpl_connect('button_press_event', self.on_canvas_press)
                self.canvas.mpl_connect('motion_notify_event', self.on_canvas_motion)
                self.canvas.mpl_connect('button_release_event', self.on_canvas_release)
                main_layout.addWidget(self.canvas, stretch=1)
                controls_main_hbox = QHBoxLayout()
                left_controls_vbox = QVBoxLayout()
                global_settings_group = QGroupBox("Global Settings & Area Method")
                global_settings_layout = QGridLayout(global_settings_group)
                global_settings_layout.addWidget(QLabel("Profile Method:"), 0, 0)
                profile_method_label = QLabel("Sum of Pixel Intensities")
                font = profile_method_label.font(); font.setBold(True); profile_method_label.setFont(font)
                global_settings_layout.addWidget(profile_method_label, 0, 1, 1, 2)
                global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
                self.method_combobox = QComboBox()
                self.method_combobox.addItems(["Rolling-valley", "Rolling Ball", "Straight Line"])
                self.method_combobox.setCurrentText(self.area_subtraction_method)
                self.method_combobox.currentIndexChanged.connect(self._on_method_changed)
                global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2)
                self.rolling_ball_label = QLabel(f"Rolling Ball Radius ({int(self.rolling_ball_radius)})")
                self.rolling_ball_label.setMinimumWidth(160)
                self.rolling_ball_slider = QSlider(Qt.Horizontal)
                self.rolling_ball_slider.setRange(1, 500)
                self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
                self.rolling_ball_slider.valueChanged.connect(self._on_rb_slider_changed)
                self.rolling_ball_slider.valueChanged.connect(lambda: self.rolling_ball_slider.setFocus())
                self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_label: lbl.setText(f"Rolling Ball Radius ({val})"))
                self.auto_adjust_checkbox = QCheckBox("Auto")
                self.auto_adjust_checkbox.setToolTip("Automatically calculate the optimal rolling ball radius based on detected peak widths.")
                self.auto_adjust_checkbox.setChecked(self.auto_adjust_rb_radius)
                self.auto_adjust_checkbox.stateChanged.connect(self.toggle_auto_adjust_rb)
                global_settings_layout.addWidget(self.rolling_ball_label, 2, 0)
                global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1)
                global_settings_layout.addWidget(self.auto_adjust_checkbox, 2, 2)
                left_controls_vbox.addWidget(global_settings_group)
                manual_actions_group = QGroupBox("Manual Peak Adjustment")
                manual_actions_layout = QHBoxLayout(manual_actions_group)
                self.add_peak_manually_button = QPushButton("Add Peak")
                self.add_peak_manually_button.setCheckable(True)
                self.add_peak_manually_button.clicked.connect(self.toggle_add_peak_mode)
                self.delete_selected_peak_button = QPushButton("Delete Peak")
                self.delete_selected_peak_button.setEnabled(False)
                self.delete_selected_peak_button.clicked.connect(self.delete_selected_peak_action)
                self.identify_peak_button = QPushButton("Focus Peak")
                self.identify_peak_button.setCheckable(True)
                self.identify_peak_button.clicked.connect(self.toggle_manual_select_mode)
                self.invert_display_button = QPushButton("Invert Profile & Image")
                self.invert_display_button.setCheckable(True)
                self.invert_display_button.toggled.connect(self._toggle_inversion_display)
                self.invert_display_button.setToolTip("Invert the profile for peak detection and area calculation.\nWhen checked, light bands will be treated as peaks.")
                self.invert_display_button.setChecked(self.is_inverted)
                manual_actions_layout.addWidget(self.add_peak_manually_button)
                manual_actions_layout.addWidget(self.delete_selected_peak_button)
                manual_actions_layout.addWidget(self.identify_peak_button)
                manual_actions_layout.addWidget(self.invert_display_button)
                left_controls_vbox.addWidget(manual_actions_group)

                peak_detect_group = QGroupBox("Peak Detection Settings")
                peak_detect_layout = QGridLayout(peak_detect_group)
                peak_detect_layout.addWidget(QLabel("Detected Peaks:"), 0, 0); self.peak_number_input = QLineEdit(); self.peak_number_input.setPlaceholderText("#"); self.peak_number_input.setMaximumWidth(60); self.update_peak_number_button = QPushButton("Set"); self.update_peak_number_button.clicked.connect(self.manual_peak_number_update); peak_detect_layout.addWidget(self.peak_number_input, 0, 1); peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)
                
                # --- START OF FIX: Update signal connections ---
                self.denoise_sigma_label = QLabel(f"Denoise Sigma ({self.denoise_sigma:.1f})"); self.denoise_sigma_slider = QSlider(Qt.Horizontal); self.denoise_sigma_slider.setRange(0,50); self.denoise_sigma_slider.setValue(int(self.denoise_sigma*10))
                self.denoise_sigma_slider.valueChanged.connect(lambda val, lbl=self.denoise_sigma_label: (lbl.setText(f"Denoise Sigma ({val/10.0:.1f})"), setattr(self, 'denoise_sigma', val/10.0)))
                self.denoise_sigma_slider.sliderReleased.connect(self.regenerate_profile_and_detect)
                peak_detect_layout.addWidget(self.denoise_sigma_label,1,0); peak_detect_layout.addWidget(self.denoise_sigma_slider,1,1,1,2)
                
                self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})"); self.smoothing_slider = QSlider(Qt.Horizontal); self.smoothing_slider.setRange(0,100); self.smoothing_slider.setValue(int(self.smoothing_sigma*10))
                self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: (lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})"), setattr(self, 'smoothing_sigma', val/10.0)))
                self.smoothing_slider.sliderReleased.connect(self.regenerate_profile_and_detect)
                peak_detect_layout.addWidget(self.smoothing_label,2,0); peak_detect_layout.addWidget(self.smoothing_slider,2,1,1,2)
                
                self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.3f})")
                self.peak_prominence_slider = QSlider(Qt.Horizontal); self.peak_prominence_slider.setRange(0, 1000); self.peak_prominence_slider.setValue(int(self.peak_prominence_factor*1000))
                self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: (lbl.setText(f"Min Prominence ({val/1000.0:.3f})"), setattr(self, 'peak_prominence_factor', val/1000.0)))
                self.peak_prominence_slider.sliderReleased.connect(self.detect_peaks)
                peak_detect_layout.addWidget(self.peak_prominence_slider_label,3,0); peak_detect_layout.addWidget(self.peak_prominence_slider,3,1,1,2)
                
                self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f})"); self.peak_height_slider = QSlider(Qt.Horizontal); self.peak_height_slider.setRange(0,100); self.peak_height_slider.setValue(int(self.peak_height_factor*100))
                self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: (lbl.setText(f"Min Height ({val/100.0:.2f})"), setattr(self, 'peak_height_factor', val/100.0)))
                self.peak_height_slider.sliderReleased.connect(self.detect_peaks)
                peak_detect_layout.addWidget(self.peak_height_slider_label,4,0); peak_detect_layout.addWidget(self.peak_height_slider,4,1,1,2)
                
                self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px"); self.peak_distance_slider = QSlider(Qt.Horizontal); self.peak_distance_slider.setRange(1,200); self.peak_distance_slider.setValue(self.peak_distance)
                self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: (lbl.setText(f"Min Distance ({val}) px"), setattr(self, 'peak_distance', val)))
                self.peak_distance_slider.sliderReleased.connect(self.detect_peaks)
                peak_detect_layout.addWidget(self.peak_distance_slider_label,5,0); peak_detect_layout.addWidget(self.peak_distance_slider,5,1,1,2)
                # --- END OF FIX ---
                
                self.copy_regions_button = QPushButton("Copy Regions"); self.copy_regions_button.clicked.connect(self.copy_peak_regions_to_app)
                self.paste_regions_button = QPushButton("Paste Regions"); self.paste_regions_button.clicked.connect(self.paste_peak_regions_from_app);
                if not (self.parent_app and self.parent_app.copied_peak_regions_data.get("regions")): self.paste_regions_button.setEnabled(False)
                
                copy_paste_layout = QHBoxLayout()
                copy_paste_layout.addWidget(self.copy_regions_button)
                copy_paste_layout.addWidget(self.paste_regions_button)
                peak_detect_layout.addLayout(copy_paste_layout, 6, 0, 1, 3)

                left_controls_vbox.addWidget(peak_detect_group); left_controls_vbox.addStretch(1); controls_main_hbox.addLayout(left_controls_vbox, stretch=1); main_layout.addLayout(controls_main_hbox); bottom_button_layout = QHBoxLayout(); self.persist_settings_checkbox = QCheckBox("Persist Settings"); self.persist_settings_checkbox.setChecked(persist_checked_initial); bottom_button_layout.addWidget(self.persist_settings_checkbox); bottom_button_layout.addStretch(1); self.ok_button = QPushButton("OK"); self.ok_button.setDefault(True); self.ok_button.clicked.connect(self.accept_and_close); self.cancel_button = QPushButton("Cancel"); self.cancel_button.clicked.connect(self.reject); bottom_button_layout.addWidget(self.ok_button); bottom_button_layout.addWidget(self.cancel_button); main_layout.addLayout(bottom_button_layout); self.setLayout(main_layout)
                self.update_rb_controls_enabled_state()

            # --- START MODIFICATION: New handler method ---
            def _toggle_inversion_display(self, checked):
                """Toggles the inversion state and re-runs the entire analysis pipeline."""
                self.is_inverted = checked
                self.regenerate_profile_and_detect()
            # --- END MODIFICATION ---

            def _on_method_changed(self):
                """Handles area method changes by recalculating regions and then updating the plot."""
                self.update_rb_controls_enabled_state()
                self._recalculate_all_regions()
                self.update_plot()
            
            def _on_rb_slider_changed(self, value):
                """Handles real-time updates when the rolling ball slider is moved."""
                if not self.auto_adjust_checkbox.isChecked() and self.method_combobox.currentText() in ["Rolling Ball", "Rolling-valley"]:
                    self.rolling_ball_radius = value
                    self._recalculate_all_regions()
                    self.update_plot()

            def update_rb_controls_enabled_state(self):
                """Enables/disables controls based on the selected area method."""
                is_rb_dependent_method = self.method_combobox.currentText() in ["Rolling Ball", "Rolling-valley"]
                
                self.rolling_ball_slider.setEnabled(is_rb_dependent_method and not self.auto_adjust_checkbox.isChecked())
                self.auto_adjust_checkbox.setEnabled(is_rb_dependent_method)
                self.rolling_ball_label.setEnabled(is_rb_dependent_method)

            def toggle_auto_adjust_rb(self, state):
                """Handles the 'Auto' checkbox state change for rolling ball radius."""
                self.auto_adjust_rb_radius = bool(state)
                is_rb_dependent_method = self.method_combobox.currentText() in ["Rolling Ball", "Rolling-valley"]
                self.rolling_ball_slider.setEnabled(is_rb_dependent_method and not self.auto_adjust_rb_radius)
                self.detect_peaks()

            def _calculate_optimal_rb_radius(self):
                if self.profile is None or len(self.peaks) == 0:
                    return 50 
                try:
                    profile_range = np.ptp(self.profile)
                    if profile_range < 1e-6: profile_range = 1.0
                    min_height_abs = np.min(self.profile) + profile_range * self.peak_height_factor
                    min_prominence_abs = profile_range * self.peak_prominence_factor
                    _peaks_for_width, properties = find_peaks(
                        self.profile, height=min_height_abs, prominence=min_prominence_abs,
                        distance=self.peak_distance, width=1
                    )
                    if 'widths' in properties and len(properties['widths']) > 0:
                        avg_width = np.mean(properties['widths'])
                        optimal_radius = int(np.clip(2.5 * avg_width, 10, 500))
                        print(f"Auto-calculated optimal rolling ball radius: {optimal_radius} (from avg peak width: {avg_width:.2f})")
                        return optimal_radius
                    else:
                        return 50
                except Exception as e:
                    print(f"Error calculating optimal rolling ball radius: {e}")
                    return 50

            def _recalculate_all_regions(self):
                """The single source of truth for calculating peak_regions based on the current method."""
                self.method = self.method_combobox.currentText()

                # Calculate rolling ball background if needed for the current method.
                if self.method in ["Rolling Ball", "Rolling-valley"]:
                    if rolling_ball and self.profile_original_inverted is not None:
                        try:
                            profile_float = self.profile_original_inverted.astype(np.float64)
                            safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
                            self.background = self._custom_rolling_ball(profile_float, safe_radius) if len(profile_float) > 1 else profile_float.copy()
                            self.background = np.maximum(self.background, 0)
                        except Exception:
                            self.background = np.zeros_like(self.profile_original_inverted)
                    else:
                        self.background = np.zeros_like(self.profile_original_inverted) if self.profile_original_inverted is not None else np.array([])
                
                # Define peak region boundaries based on method.
                if self.method == "Rolling Ball":
                    # For Rolling Ball, boundaries are at the intersections with the background curve.
                    self._redefine_regions_from_background(self.background)
                else: # "Rolling-valley" and "Straight Line"
                    # For both Valley methods, boundaries are defined by troughs (and outer background intersections).
                    self._redefine_all_valley_regions()

            def _redefine_regions_from_background(self, background):
                """Helper to define peak regions based on intersections with a given background."""
                self.peak_regions = []
                profile_to_analyze = self.profile_original_inverted
                if profile_to_analyze is None or len(profile_to_analyze) <= 1 or len(self.peaks) == 0:
                    return

                midpoints = (self.peaks[:-1] + self.peaks[1:]) // 2 if len(self.peaks) > 1 else []
                search_boundaries_left = np.concatenate(([0], midpoints))
                search_boundaries_right = np.concatenate((midpoints, [len(profile_to_analyze) - 1]))

                for i, peak_idx in enumerate(self.peaks):
                    left_bound = int(search_boundaries_left[i])
                    right_bound = int(search_boundaries_right[i])
                    start, end = self._find_intersection_boundaries(
                        profile_to_analyze, background, peak_idx, left_bound, right_bound
                    )
                    self.peak_regions.append((start, end))
            
            def detect_peaks(self):
                if self.profile is None or len(self.profile) == 0:
                    self.peaks = np.array([])
                else:
                    # --- START OF FIX ---
                    # REMOVED: self.peak_height_factor = self.peak_height_slider.value()/100.0
                    # REMOVED: self.peak_distance = self.peak_distance_slider.value()
                    # REMOVED: self.peak_prominence_factor = self.peak_prominence_slider.value()/1000.0
                    # The method now uses the instance attributes directly.
                    # --- END OF FIX ---
                    profile_range = np.ptp(self.profile)
                    min_height_abs = np.min(self.profile) + profile_range * self.peak_height_factor
                    min_prominence_abs = profile_range * self.peak_prominence_factor
                    try:
                        self.peaks, _ = find_peaks(self.profile, height=min_height_abs, prominence=min_prominence_abs, distance=self.peak_distance, width=1)
                    except Exception:
                        self.peaks = np.array([])
                
                if self.auto_adjust_rb_radius:
                    self.rolling_ball_radius = self._calculate_optimal_rb_radius()
                    if hasattr(self, 'rolling_ball_slider'):
                        self.rolling_ball_slider.blockSignals(True)
                        self.rolling_ball_slider.setValue(self.rolling_ball_radius)
                        self.rolling_ball_slider.blockSignals(False)
                    if hasattr(self, 'rolling_ball_label'):
                        self.rolling_ball_label.setText(f"Rolling Ball Radius ({self.rolling_ball_radius})")

                if hasattr(self, 'peak_number_input') and not self.peak_number_input.hasFocus():
                    self.peak_number_input.setText(str(len(self.peaks)))
                
                self._recalculate_all_regions()
                self.update_plot()

            def accept_and_close(self):
                # --- START OF FIX ---
                # Read from instance attributes, not UI elements
                self._final_settings = {
                    'rolling_ball_radius': self.rolling_ball_radius,
                    'denoise_sigma': self.denoise_sigma,
                    'peak_height_factor': self.peak_height_factor,
                    'peak_distance': self.peak_distance,
                    'peak_prominence_factor': self.peak_prominence_factor,
                    'band_estimation_method': "Sum",
                    'area_subtraction_method': self.method_combobox.currentText(),
                    'smoothing_sigma': self.smoothing_sigma,
                    'auto_adjust_rb_radius': self.auto_adjust_checkbox.isChecked(),
                    'is_inverted': self.is_inverted,
                }
                # --- END OF FIX ---
                self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
                self.accept()
                
            def update_plot(self):
                if self.canvas is None: return
                profile_to_plot_and_calc = self.profile_original_inverted
                
                # --- Theme Colors (Unchanged) ---
                is_dark_theme = self.parent_app and self.parent_app.current_theme == "dark"
                if is_dark_theme:
                    bg_color, ax_bg_color = '#2D2D30', '#38383C'
                    text_color, spine_color, grid_color = '#F1F1F1', '#707070', '#5A5A60'
                    profile_color, peak_marker_color, focused_peak_color, selected_peak_color = '#4DB6AC', '#FF8A65', '#FFCA28', '#42A5F5'
                    bg_line_color, sl_line_color, rv_line_color = '#7E57C2', '#5C6BC0', '#42A5F5'
                    fill_color_rv = 'yellow'; fill_alpha_rv = 0.5
                else:
                    bg_color, ax_bg_color = 'white', 'white'
                    text_color, spine_color, grid_color = 'black', 'black', '#DDDDDD'
                    profile_color, peak_marker_color, focused_peak_color, selected_peak_color = 'black', 'red', 'orange', 'blue'
                    bg_line_color, sl_line_color, rv_line_color = 'purple', 'magenta', 'blue'
                    fill_color_rv = 'yellow'; fill_alpha_rv = 0.7

                self.fig.clf()
                
                # --- FIX 1: Adjusted height_ratios (2:1) and increased hspace (0.5) ---
                # This gives the bottom image more height and adds a larger gap between plots.
                gs = GridSpec(2, 1, height_ratios=[2, 1], hspace=0.15, figure=self.fig)
                
                self.ax = self.fig.add_subplot(gs[0])
                self.ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)
                self.fig.patch.set_facecolor(bg_color)
                self.interactive_artists.clear()
                
                for axis in [self.ax, self.ax_image]:
                    axis.patch.set_facecolor(ax_bg_color)
                    for spine in axis.spines.values(): spine.set_color(spine_color)
                    axis.tick_params(axis='x', colors=text_color, labelsize=9)
                    axis.tick_params(axis='y', colors=text_color, labelsize=9)
                    axis.yaxis.label.set_color(text_color); axis.xaxis.label.set_color(text_color); axis.title.set_color(text_color)
                
                if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
                     self.ax_image.set_xlabel("Pixel Index", color=text_color, fontsize=7)
                     self.ax.tick_params(axis='x', labelbottom=False)
                     self.ax_image.text(0.5, 0.5, 'No Profile Data', ha='center', va='center', color=text_color, transform=self.ax_image.transAxes)
                     self.canvas.draw_idle(); return
                
                profile_for_display = profile_to_plot_and_calc

                if not hasattr(self, 'background') or self.background is None or self.background.shape != profile_to_plot_and_calc.shape:
                    self.background = np.zeros_like(profile_to_plot_and_calc)

                self.ax.plot(profile_for_display, label="Profile", color=profile_color, lw=1.2, zorder=10)
                
                if len(self.peaks) > 0:
                     valid_peaks_indices = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
                     if len(valid_peaks_indices) > 0:
                         peak_y_on_displayed = profile_for_display[valid_peaks_indices]
                         self.ax.scatter(valid_peaks_indices, peak_y_on_displayed, color=peak_marker_color, marker='x', s=40, label="Peaks", zorder=15) 
                         if self.selected_peak_for_ui_focus != -1 and 0 <= self.selected_peak_for_ui_focus < len(self.peaks):
                             focused_peak_x_val = self.peaks[self.selected_peak_for_ui_focus]
                             self.ax.plot(focused_peak_x_val, profile_for_display[focused_peak_x_val], 'o', markersize=12, markeredgecolor=focused_peak_color, markerfacecolor='none', label='Focused', zorder=16)
                         if self.selected_peak_index_for_delete != -1:
                             self.ax.plot(self.selected_peak_index_for_delete, profile_for_display[self.selected_peak_index_for_delete], 's', markersize=14, markeredgecolor=selected_peak_color, markerfacecolor='none', label='Selected for Delete', zorder=17)
                
                self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
                
                profile_range_plot = np.ptp(profile_for_display) if np.ptp(profile_for_display) > 0 else 1.0
                max_y_for_plot_limit = np.max(profile_for_display) if len(profile_for_display) > 0 else 1
                text_positions = []

                # --- First pass: Calculate all areas ---
                for i in range(len(self.peak_regions)):
                    start_handle, end_handle = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                    if start_handle >= end_handle or i >= len(self.peaks): 
                        self.peak_areas_valley.append(0); self.peak_areas_straight_line.append(0); self.peak_areas_rolling_ball.append(0)
                        continue
                    
                    baseline_rb = self.background
                    area_rb = np.trapezoid(profile_to_plot_and_calc[start_handle:end_handle+1] - baseline_rb[start_handle:end_handle+1])
                    self.peak_areas_rolling_ball.append(max(0, area_rb))
                    
                    global_sl_baseline = None
                    if self.peak_regions:
                        trough_x = [self.peak_regions[0][0]] + [r[1] for r in self.peak_regions]
                        trough_y = [profile_to_plot_and_calc[x] for x in trough_x]
                        if len(trough_x) >= 2:
                            x_all = np.arange(len(profile_to_plot_and_calc))
                            global_sl_baseline = np.interp(x_all, trough_x, trough_y)
                    area_sl = 0.0
                    if global_sl_baseline is not None:
                        difference_sl = profile_to_plot_and_calc[start_handle:end_handle+1] - global_sl_baseline[start_handle:end_handle+1]
                        area_sl = np.trapezoid(np.maximum(0, difference_sl))
                    self.peak_areas_straight_line.append(max(0, area_sl))

                    y_baseline_rv_points = np.interp([start_handle, end_handle], [start_handle, end_handle], [profile_to_plot_and_calc[start_handle], profile_to_plot_and_calc[end_handle]])
                    baseline_rv_local_straight = np.interp(np.arange(start_handle, end_handle + 1), [start_handle, end_handle], y_baseline_rv_points)
                    final_rv_baseline = np.maximum(baseline_rv_local_straight, self.background[start_handle:end_handle+1])
                    area_valley = np.trapezoid(np.maximum(0, profile_to_plot_and_calc[start_handle:end_handle+1] - final_rv_baseline))
                    self.peak_areas_valley.append(max(0, area_valley))

                total_area = 0
                if self.method == "Rolling-valley": total_area = sum(self.peak_areas_valley)
                elif self.method == "Rolling Ball": total_area = sum(self.peak_areas_rolling_ball)
                elif self.method == "Straight Line": total_area = sum(self.peak_areas_straight_line)
                if total_area < 1e-9: total_area = 0.0

                # --- Second pass: Draw fills and labels ---
                for i in range(len(self.peak_regions)):
                    start_handle, end_handle = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                    if start_handle >= end_handle or i >= len(self.peaks): continue
                    peak_x = self.peaks[i]
                    
                    area_to_display = 0.0
                    if self.method == "Rolling-valley":
                        x_region_rv = np.arange(start_handle, end_handle + 1)
                        y_baseline_rv_points = np.interp([start_handle, end_handle], [start_handle, end_handle], [profile_to_plot_and_calc[start_handle], profile_to_plot_and_calc[end_handle]])
                        baseline_rv_local_straight = np.interp(x_region_rv, [start_handle, end_handle], y_baseline_rv_points)
                        final_rv_baseline = np.maximum(baseline_rv_local_straight, self.background[start_handle:end_handle+1])
                        
                        self.ax.fill_between(x_region_rv, final_rv_baseline, profile_for_display[x_region_rv], where=(profile_for_display[x_region_rv] >= final_rv_baseline), color=fill_color_rv, alpha=fill_alpha_rv, interpolate=True, zorder=1)
                        self.ax.plot(x_region_rv, final_rv_baseline, color=rv_line_color, lw=1.5, zorder=4)

                        if i == 0:
                            self.ax.get_lines()[-1].set_label("Valley BG")
                            self.ax.plot(np.arange(len(self.background)), self.background, color='magenta', ls=":", lw=1.0, label="RV Guide BG", zorder=3)
                        
                        area_to_display = self.peak_areas_valley[i]
                    
                    elif self.method == "Rolling Ball":
                         x_region = np.arange(start_handle, end_handle + 1)
                         self.ax.fill_between(x_region, self.background[x_region], profile_for_display[x_region], where=(profile_for_display[x_region] >= self.background[x_region]), color="yellow", alpha=0.4, interpolate=True, zorder=1)
                         if i == 0: self.ax.plot(np.arange(len(self.background)), self.background, color=bg_line_color, ls="--", lw=1, label="Rolling Ball BG", zorder=2)
                         area_to_display = self.peak_areas_rolling_ball[i]

                    elif self.method == "Straight Line":
                        if global_sl_baseline is not None:
                            x_region = np.arange(start_handle, end_handle + 1)
                            self.ax.fill_between(x_region, global_sl_baseline[x_region], profile_for_display[x_region], where=(profile_for_display[x_region] >= global_sl_baseline[x_region]), color="cyan", alpha=0.4, interpolate=True, zorder=1)
                            if i == 0: self.ax.plot(np.arange(len(global_sl_baseline)), global_sl_baseline, color=sl_line_color, ls="--", lw=1.2, label="SL BG", zorder=2)
                        area_to_display = self.peak_areas_straight_line[i]
                    
                    text_y_pos = profile_for_display[peak_x] + profile_range_plot * 0.03
                    if total_area > 0:
                        text_str = f"{(area_to_display / total_area * 100):.1f}%"
                    else:
                        text_str = f"{area_to_display:.0f}"
                    text_positions.append((peak_x, text_y_pos, text_str))
                    max_y_for_plot_limit = max(max_y_for_plot_limit, text_y_pos)

                last_text_end_x = -np.inf
                text_spacing_pixels = 5
                text_positions.sort(key=lambda item: item[0])
                for x, y, text_str in text_positions:
                    text_artist = self.ax.text(x, y, text_str, ha="center", va="bottom", fontsize=7, color=text_color, zorder=20,
                                               bbox=dict(boxstyle="round,pad=0.2", fc=ax_bg_color, ec=spine_color, alpha=0.8))
                    renderer = self.canvas.get_renderer()
                    bbox = text_artist.get_window_extent(renderer=renderer)
                    bbox_data = self.ax.transData.inverted().transform(bbox)
                    if bbox_data[0,0] < last_text_end_x + text_spacing_pixels:
                        new_y = y + profile_range_plot * 0.08; text_artist.set_y(new_y)
                        max_y_for_plot_limit = max(max_y_for_plot_limit, new_y)
                        bbox = text_artist.get_window_extent(renderer=renderer)
                        bbox_data = self.ax.transData.inverted().transform(bbox)
                    last_text_end_x = bbox_data[1,0]
                
                handles, labels = self.ax.get_legend_handles_labels()
                if handles:
                    # Legend placed in the extra bottom space
                    leg = self.fig.legend(handles, labels, loc='lower center', ncol=len(handles), bbox_to_anchor=(0.5, 0.01), fontsize='small', facecolor=bg_color, edgecolor=spine_color)
                    for text in leg.get_texts(): text.set_color(text_color)
                
                self.ax.set_ylabel("Intensity", fontsize=9)
                self.ax.set_title(f"Profile & Peak Regions ({self.method})", fontsize=9, weight='bold')
                self.ax.tick_params(axis='x', which='both', bottom=False, labelbottom=False)
                
                # --- FIX 2: Set X/Y limits with extra top padding for labels ---
                if len(profile_to_plot_and_calc) > 1:
                    self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1)
                    # Increased multiplier from 1.1 to 1.25 to prevent label cutoff
                    self.ax.set_ylim(bottom=min(0, np.min(profile_for_display)), top=max_y_for_plot_limit * 1.25)
                
                if np.max(profile_to_plot_and_calc) > 10000:
                    self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0), useMathText=True)
                
                self.ax_image.clear() 
                if hasattr(self, 'enhanced_cropped_image_for_display') and self.enhanced_cropped_image_for_display:
                    rotated_pil_image_display = self.enhanced_cropped_image_for_display.rotate(90, expand=True)
                    image_extent = [0, len(profile_to_plot_and_calc) - 1, 0, rotated_pil_image_display.height]
                    
                    cmap_val = 'gray_r' if self.is_inverted else 'gray'
                    self.ax_image.imshow(np.array(rotated_pil_image_display), cmap=cmap_val, aspect='auto', extent=image_extent)

                    self.ax_image.set_yticks([]); self.ax_image.set_ylabel("Lane Width", fontsize=9)
                    self.ax_image.set_xlabel("Pixel Index", fontsize=9)
                    
                    # Draw handles on image
                    for peak_idx, (start_px, end_px) in enumerate(self.peak_regions):
                        is_focused = peak_idx == self.selected_peak_for_ui_focus
                        line_color, zorder_val, lw = (focused_peak_color, 11, 2.0) if is_focused else ('blue', 10, 1.5)
                        start_line = mlines.Line2D([start_px, start_px], [0, rotated_pil_image_display.height], color=line_color, lw=lw, picker=self.HANDLE_SIZE, zorder=zorder_val); self.ax_image.add_line(start_line); self.interactive_artists.append((peak_idx, 'start_line', start_line))
                        end_line = mlines.Line2D([end_px, end_px], [0, rotated_pil_image_display.height], color=line_color, lw=lw, picker=self.HANDLE_SIZE, zorder=zorder_val); self.ax_image.add_line(end_line); self.interactive_artists.append((peak_idx, 'end_line', end_line))
                
                # --- FIX 3: Adjusted Margins for Title, Axis Labels, and Legend ---
                # top=0.85 (more room for title), bottom=0.22 (room for legend), left=0.15 (room for Y-label)
                self.fig.subplots_adjust(left=0.1, right=0.9, top=0.92, bottom=0.25)
                
                self.canvas.draw_idle()
                plt.close(self.fig)

            def on_draw(self, event): self.background_blit = self.canvas.copy_from_bbox(self.fig.bbox)
            def on_canvas_press(self, event):
                if event.inaxes != self.ax_image or event.button != 1:
                    if event.inaxes == self.ax: self.handle_profile_plot_click(event)
                    return
                for i, (peak_idx, handle_type, artist) in enumerate(self.interactive_artists):
                    contains, _ = artist.contains(event)
                    if contains:
                        self.dragging_handle_info = {'peak_index': peak_idx, 'type': handle_type, 'artist': artist}
                        self.selected_peak_for_ui_focus = peak_idx
                        for a in self.interactive_artists: a[2].set_animated(False)
                        artist.set_animated(True)
                        self.canvas.draw()
                        return
                self.dragging_handle_info = None
            def on_canvas_motion(self, event):
                if not self.dragging_handle_info or event.inaxes != self.ax_image: return
                self.canvas.restore_region(self.background_blit)
                artist = self.dragging_handle_info['artist']
                new_x = int(round(event.xdata)) if event.xdata is not None else 0
                artist.set_xdata([new_x, new_x])
                self.ax_image.draw_artist(artist)
                self.canvas.blit(self.ax_image.bbox)
            def on_canvas_release(self, event):
                if not self.dragging_handle_info: return
                artist = self.dragging_handle_info['artist']
                artist.set_animated(False)
                self.background_blit = None
                peak_idx = self.dragging_handle_info['peak_index']
                handle_type = self.dragging_handle_info['type']
                new_x = int(round(event.xdata)) if event.xdata is not None else 0
                profile_len = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                new_x = max(0, min(new_x, profile_len - 1))
                current_start, current_end = self.peak_regions[peak_idx]
                if handle_type == 'start_line':
                    new_start = min(new_x, current_end - 1 if current_end > 0 else 0)
                    self.peak_regions[peak_idx] = (new_start, current_end)
                elif handle_type == 'end_line':
                    new_end = max(new_x, current_start + 1 if current_start < profile_len - 1 else profile_len - 1)
                    self.peak_regions[peak_idx] = (current_start, new_end)
                self.dragging_handle_info = None
                self.update_plot()
                self.canvas.draw_idle()
            def handle_profile_plot_click(self, event):
                if self.add_peak_mode_active:
                    if event.button == 1 and event.xdata is not None:
                        clicked_x = int(round(event.xdata))
                        if self.profile_original_inverted is not None and 0 <= clicked_x < len(self.profile_original_inverted): self.add_manual_peak(clicked_x)
                elif self.manual_select_mode_active:
                    if event.button == 1 and event.xdata is not None and self.peaks.any():
                        clicked_x = int(round(event.xdata)); distances = np.abs(self.peaks - clicked_x); closest_peak_idx = np.argmin(distances)
                        click_tolerance = self.peak_distance / 2.0 if self.peak_distance > 0 else 10.0
                        if distances[closest_peak_idx] <= click_tolerance: self.selected_peak_for_ui_focus = closest_peak_idx
                        else: self.selected_peak_for_ui_focus = -1
                        self.update_plot()
                else:
                    if event.button == 1 and event.xdata is not None and self.peaks.any():
                        clicked_x = int(round(event.xdata)); distances = np.abs(self.peaks - clicked_x); closest_peak_idx = np.argmin(distances)
                        click_tolerance_delete = max(5, self.peak_distance / 4.0)
                        if distances[closest_peak_idx] <= click_tolerance_delete:
                            self.selected_peak_index_for_delete = self.peaks[closest_peak_idx]
                            self.delete_selected_peak_button.setEnabled(True)
                        else:
                            self.selected_peak_index_for_delete = -1
                            self.delete_selected_peak_button.setEnabled(False)
                        self.update_plot()
            def _find_intersection_boundaries(self, profile, baseline, peak_x, search_start, search_end):
                diff = profile - baseline
                if not np.any(diff): return search_start, search_end
                above_indices = np.where(diff > 0)[0]
                if len(above_indices) == 0: return peak_x, peak_x
                sign_changes = np.where(np.diff(np.sign(diff)))[0]
                left_intersections = sign_changes[sign_changes < peak_x]
                start_calc = np.max(left_intersections) + 1 if left_intersections.size > 0 else np.min(above_indices)
                right_intersections = sign_changes[sign_changes > peak_x]
                end_calc = np.min(right_intersections) if right_intersections.size > 0 else np.max(above_indices)
                start_calc = max(search_start, start_calc)
                end_calc = min(search_end, end_calc)
                if start_calc >= end_calc: return search_start, search_end
                return int(start_calc), int(end_calc)

            def get_final_peak_info(self):
                peak_info_list = []
                num_valid_peaks = len(self.peak_regions)
                current_area_list = []
                if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
                elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
                elif self.method == "Rolling-valley": current_area_list = self.peak_areas_valley
                num_peaks_to_process = min(num_valid_peaks, len(self.peaks), len(current_area_list))
                for i in range(num_peaks_to_process):
                    try:
                        original_peak_x_in_profile = int(self.peaks[i])
                        peak_info_list.append({'area': current_area_list[i],'y_coord_in_lane_image': original_peak_x_in_profile,'original_peak_index': original_peak_x_in_profile})
                    except IndexError: peak_info_list.append({'area': 0.0, 'y_coord_in_lane_image': 0, 'original_peak_index': -1})
                return peak_info_list
            def toggle_manual_select_mode(self, checked):
                self.manual_select_mode_active = checked
                if checked:
                    if self.add_peak_mode_active: self.add_peak_manually_button.setChecked(False); self.toggle_add_peak_mode(False) 
                    QMessageBox.information(self, "Identify Peak", "Click a peak in the profile plot to focus its handles.")
                else: self.selected_peak_for_ui_focus = -1; self.update_plot()
            def toggle_add_peak_mode(self, checked):
                self.add_peak_mode_active = checked
                if checked:
                    if self.manual_select_mode_active: self.identify_peak_button.setChecked(False); self.toggle_manual_select_mode(False)
                    self.selected_peak_index_for_delete = -1; self.delete_selected_peak_button.setEnabled(False)
                    self.update_plot(); QMessageBox.information(self, "Add Peak", "Click on the profile plot to add a peak.")
            def add_manual_peak(self, x_coord):
                if self.profile_original_inverted is None or x_coord in self.peaks: return
                self.peaks = np.array(sorted(self.peaks.tolist() + [x_coord]))
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self._recalculate_all_regions()
                self.update_plot()
            def _redefine_all_valley_regions(self):
                """
                Defines the start/end region for each peak based on its adjacent troughs (valleys).
                This is the definitive fix to prevent baselines from incorrectly connecting distant points.
                """
                self.peak_regions = []
                profile = self.profile_original_inverted
                if profile is None or len(profile) == 0 or len(self.peaks) == 0:
                    return

                # 1. Find the troughs (local minima) BETWEEN each pair of adjacent peaks.
                troughs = []
                for i in range(len(self.peaks) - 1):
                    start_search = self.peaks[i]
                    end_search = self.peaks[i+1]
                    if start_search >= end_search: continue
                    valley_region = profile[start_search:end_search]
                    if valley_region.size > 0:
                        local_min_idx = np.argmin(valley_region)
                        troughs.append(start_search + local_min_idx)

                # 2. Find the absolute outer boundaries of the entire peak cluster using the rolling ball background.
                left_outer_bound, _ = self._find_intersection_boundaries(profile, self.background, self.peaks[0], 0, self.peaks[0])
                _, right_outer_bound = self._find_intersection_boundaries(profile, self.background, self.peaks[-1], self.peaks[-1], len(profile) - 1)

                # 3. Create a complete list of all possible boundary points.
                all_boundary_points = sorted(list(set([left_outer_bound] + troughs + [right_outer_bound])))
                
                # 4. For each peak, find its correct start and end from the boundary points list.
                for i, peak_x in enumerate(self.peaks):
                    # Find the closest boundary point to the left.
                    left_boundaries = [b for b in all_boundary_points if b <= peak_x]
                    start_handle = max(left_boundaries) if left_boundaries else 0

                    # Find the closest boundary point to the right.
                    right_boundaries = [b for b in all_boundary_points if b >= peak_x]
                    end_handle = min(right_boundaries) if right_boundaries else len(profile) - 1
                    
                    if start_handle < end_handle:
                        self.peak_regions.append((start_handle, end_handle))
                    # Fallback for a peak that might not have a proper region found
                    elif len(self.peak_regions) < len(self.peaks):
                        self.peak_regions.append((max(0, peak_x - 5), min(len(profile)-1, peak_x + 5)))

            def delete_selected_peak_action(self):
                if self.selected_peak_index_for_delete == -1 or self.selected_peak_index_for_delete not in self.peaks: return
                self.peaks = np.array([p for p in self.peaks if p != self.selected_peak_index_for_delete])
                self.selected_peak_index_for_delete = -1; self.delete_selected_peak_button.setEnabled(False)
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self._recalculate_all_regions()
                self.update_plot()
            def copy_peak_regions_to_app(self):
                if not self.parent_app: return
                if not self.peak_regions: QMessageBox.information(self, "No Regions", "No peak regions to copy."); return
                self.parent_app.copied_peak_regions_data["regions"] = [tuple(r) for r in self.peak_regions]
                self.parent_app.copied_peak_regions_data["profile_length"] = len(self.profile_original_inverted) if self.profile_original_inverted is not None else 0
                self.parent_app.copied_peak_regions_data["peaks"] = self.peaks.tolist()
                QMessageBox.information(self, "Regions Copied", f"{len(self.peak_regions)} regions copied.")
                if hasattr(self, 'paste_regions_button'): self.paste_regions_button.setEnabled(True)
            def paste_peak_regions_from_app(self):
                if not self.parent_app or not self.parent_app.copied_peak_regions_data.get("regions"): QMessageBox.information(self, "No Regions to Paste", "No peak regions have been copied yet."); return
                if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0: QMessageBox.warning(self, "Error", "Current profile not available for pasting regions."); return
                copied_data = self.parent_app.copied_peak_regions_data; regions_to_paste = copied_data["regions"]; original_profile_len = copied_data["profile_length"]; copied_peaks_indices = np.array(copied_data.get("peaks", []))
                current_profile_len = len(self.profile_original_inverted); self.peak_regions = []
                scale_factor = 1.0
                if original_profile_len > 0 and current_profile_len > 0 and original_profile_len != current_profile_len: scale_factor = float(current_profile_len) / original_profile_len
                temp_derived_peaks = []
                if len(copied_peaks_indices) > 0 and len(copied_peaks_indices) == len(regions_to_paste): self.peaks = np.clip((copied_peaks_indices * scale_factor).round().astype(int), 0, current_profile_len - 1)
                for i, (start_orig, end_orig) in enumerate(regions_to_paste):
                    start_scaled = int(round(start_orig * scale_factor)); end_scaled = int(round(end_orig * scale_factor))
                    start_clamped = max(0, min(start_scaled, current_profile_len - 1)); end_clamped = max(0, min(end_scaled, current_profile_len - 1))
                    if start_clamped >= end_clamped: mid = (start_clamped + end_clamped) // 2; start_clamped, end_clamped = max(0, mid-1), min(current_profile_len-1, mid+1)
                    self.peak_regions.append((start_clamped, end_clamped))
                    if not (len(copied_peaks_indices) > 0 and len(copied_peaks_indices) == len(regions_to_paste)): temp_derived_peaks.append((start_clamped + end_clamped) // 2)
                if temp_derived_peaks: self.peaks = np.array(sorted(temp_derived_peaks))
                if hasattr(self, 'peak_number_input'): self.peak_number_input.setText(str(len(self.peaks)))
                self.update_plot(); QMessageBox.information(self, "Regions Pasted", f"{len(self.peak_regions)} regions applied.")
            def get_current_settings(self): return self._final_settings
            def should_persist_settings(self): return self._persist_enabled_on_exit
            def get_final_peak_area(self): return [info['area'] for info in self.get_final_peak_info()]
            
            def regenerate_profile_and_detect(self):
                if gaussian_filter1d is None or cv2 is None or ImageOps is None: return
                
                # --- START OF FIX ---
                # REMOVED: self.smoothing_sigma = self.smoothing_slider.value() / 10.0
                # REMOVED: self.denoise_sigma = self.denoise_sigma_slider.value() / 10.0
                # The method now uses self.smoothing_sigma and self.denoise_sigma directly,
                # which are updated by the slider signals.
                # --- END OF FIX ---

                base_img = self.original_pil_cropped_data.copy()
                if self.denoise_sigma > 0.01:
                    try:
                        base_img = Image.fromarray(cv2.GaussianBlur(np.array(base_img), (0,0), self.denoise_sigma).astype(np.array(base_img).dtype))
                    except: pass
                
                if base_img.mode.startswith('I') or base_img.mode == 'F': self.enhanced_cropped_image_for_display = Image.fromarray((np.clip((np.array(base_img, dtype=np.float32) - np.percentile(base_img, 2)) / (np.percentile(base_img, 98) - np.percentile(base_img, 2) + 1e-9), 0.0, 1.0) * 255).astype(np.uint8), mode='L')
                else: self.enhanced_cropped_image_for_display = ImageOps.autocontrast(base_img.convert('L'))
                
                base_array = self.intensity_array_original_range.astype(np.float64)
                if self.is_inverted:
                    array_for_summing = base_array
                else:
                    array_for_summing = self.original_max_value - base_array
                
                profile_to_process = np.sum(array_for_summing, axis=1)

                if self.smoothing_sigma > 0.1: 
                    self.profile_original_inverted = gaussian_filter1d(profile_to_process, sigma=self.smoothing_sigma)
                else:
                    self.profile_original_inverted = profile_to_process
                
                prof_min, prof_max = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
                if prof_max > prof_min + 1e-6: self.profile = (self.profile_original_inverted - prof_min) / (prof_max - prof_min) * 255.0
                else: self.profile = np.zeros_like(self.profile_original_inverted)
                
                self.detect_peaks()

            def _find_outward_troughs(self, profile, peak_idx, left_bound, right_bound):
                profile_len = len(profile)
                if not (0 <= left_bound <= peak_idx <= right_bound < profile_len):
                    w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                    return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)
                valley_left_idx = peak_idx
                for idx in range(peak_idx - 1, left_bound - 1, -1):
                    if profile[idx] > profile[idx + 1]: valley_left_idx = idx + 1; break
                    valley_left_idx = idx
                else: valley_left_idx = left_bound
                valley_right_idx = peak_idx
                for idx in range(peak_idx + 1, right_bound + 1):
                    if profile[idx] > profile[idx - 1]: valley_right_idx = idx - 1; break
                    valley_right_idx = idx
                else: valley_right_idx = right_bound
                if valley_left_idx >= valley_right_idx:
                    w = max(1, self.peak_distance // 8 if hasattr(self, 'peak_distance') else 2)
                    return max(0, peak_idx - w), min(profile_len - 1, peak_idx + w)
                return valley_left_idx, valley_right_idx
            def manual_peak_number_update(self):
                if self.profile_original_inverted is None: return
                try:
                    num_peaks = int(self.peak_number_input.text())
                    if num_peaks == len(self.peaks): return
                    if num_peaks < len(self.peaks): self.peaks = self.peaks[:num_peaks]
                    else:
                        num_to_add = num_peaks - len(self.peaks)
                        new_peaks = np.linspace(0, len(self.profile)-1, num_peaks + 2)[1:-1].astype(int)
                        self.peaks = np.sort(np.unique(np.concatenate((self.peaks, new_peaks))))[:num_peaks]
                    self._recalculate_all_regions()
                    self.update_plot()
                except ValueError: self.peak_number_input.setText(str(len(self.peaks)))
            def _custom_rolling_ball(self, profile, radius):
                if grey_opening is None or profile is None or profile.ndim != 1 or profile.size == 0 or radius <= 0: return np.zeros_like(profile) if profile is not None else np.array([])
                structure_size = int(max(1, 2 * radius + 1));
                if structure_size > profile.shape[0]: structure_size = profile.shape[0]
                try: return grey_opening(profile, size=structure_size, mode='reflect')
                except Exception: return np.zeros_like(profile)
            

        class LiveViewLabel(QLabel):
            mouseMovedInLabel = Signal(QPointF, QPointF, bool)
            CORNER_HANDLE_BASE_RADIUS = 6.0
            def __init__(self, font_type, font_size, marker_color, app_instance, parent=None):
                super().__init__(parent)
                self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
                self.setScaledContents(True)
                self.app_instance = app_instance
                self.setMouseTracking(True)
                self.setFocusPolicy(Qt.ClickFocus)

                # --- Simplified State for Previews and Drawing ---
                self.preview_marker_enabled = False
                self.preview_marker_text = ""
                self.preview_marker_position = None
                self.marker_font_type = font_type
                self.marker_font_size = font_size
                self.marker_color = marker_color
                self.standard_marker_preview_enabled = False
                self.standard_marker_preview_position = None
                self.standard_marker_preview_text = ""
                self.standard_marker_preview_mode = None
                self.mw_predict_preview_enabled = False
                self.mw_predict_preview_position = None
                
                self.mode = None # Tracks current interaction: 'define_rect', 'define_quad', etc.
                self.current_preview_points = [] # Unified list for drawing in-progress shapes
                self.drag_preview_quad_points = None
                self.drag_preview_rect = None

                # --- Zoom and Pan State ---
                self.zoom_level = 1.0
                self.pan_offset = QPointF(0, 0)
                self.is_panning = False
                self.pan_start_view_coords = None
                self.pan_offset_at_drag_start = None

                # --- Cropping State ---
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None

                # --- Custom Handlers ---
                self._custom_left_click_handler_from_app = None
                self._custom_mouseMoveEvent_from_app = None
                self._custom_mouseReleaseEvent_from_app = None
                
            def clear_crop_preview(self):
                """Clears any existing crop rectangle preview."""
                self.drawing_crop_rect = False
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.crop_rect_final_view = None
                self.update() # Trigger repaint

            def start_crop_preview(self, start_point_view):
                """Initiates drawing the crop rectangle preview."""
                self.drawing_crop_rect = True
                self.crop_rect_start_view = start_point_view
                self.crop_rect_end_view = start_point_view # Start and end are same initially
                self.crop_rect_final_view = None # Clear any finalized rect
                self.update()

            def update_crop_preview(self, start_point_view, current_point_view):
                """Updates the crop rectangle preview while dragging."""
                if self.drawing_crop_rect:
                    # Keep start point fixed, update end point
                    self.crop_rect_start_view = start_point_view
                    self.crop_rect_end_view = current_point_view
                    self.update()
                    

            def finalize_crop_preview(self, start_point_view, end_point_view):
                """Stores the final rectangle dimensions for persistent preview."""
                self.drawing_crop_rect = False
                # Store as a normalized QRectF for easier drawing
                self.crop_rect_final_view = QRectF(start_point_view, end_point_view).normalized()
                # Clear the temporary points used during drawing
                self.crop_rect_start_view = None
                self.crop_rect_end_view = None
                self.update()
                
            def wheelEvent(self, event: 'QWheelEvent'): # Add type hint for clarity
                if not self.app_instance or not self.app_instance.image: # Ensure app and image exist
                    super().wheelEvent(event)
                    return
            
                # --- Determine Zoom Factor ---
                # event.angleDelta().y() is typically +/- 120 for one notch on a mouse wheel
                num_degrees = event.angleDelta().y() / 8
                num_steps = num_degrees / 15.0 # Standard step factor
            
                zoom_factor_increment = 1.1
                zoom_factor_decrement = 1 / 1.1
            
                # --- Calculate new zoom level ---
                old_zoom_level = self.zoom_level
                if num_steps > 0: # Zooming in
                    self.zoom_level *= (zoom_factor_increment ** abs(num_steps))
                elif num_steps < 0: # Zooming out
                    self.zoom_level *= (zoom_factor_decrement ** abs(num_steps))
            
                # --- Clamp Zoom Level (optional, but good practice) ---
                min_zoom = 0.1 # Example minimum zoom
                max_zoom = 20.0  # Example maximum zoom
                self.zoom_level = max(min_zoom, min(self.zoom_level, max_zoom))
            
                # --- Zoom Towards Mouse Cursor ---
                mouse_point_widget = event.position() 
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_point_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_point_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_point_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_point_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if self.zoom_level <= 1.0001: # Use a small tolerance for float comparison
                    self.zoom_level = 1.0
                    self.pan_offset = QPointF(0, 0)
            
                if not self.is_panning: # is_panning check is important here
                    if self.zoom_level > 1.001: # Use a small tolerance for float comparison
                        self.setCursor(Qt.OpenHandCursor)
                    else:
                        self.setCursor(Qt.ArrowCursor)
                
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()
                event.accept()
                
            def mouseMoveEvent(self, event: 'QMouseEvent'):
                current_mouse_pos_widget = event.position() 
                untransformed_label_pos = self.transform_point(current_mouse_pos_widget)
                image_coords_actual = None
                is_image_coord_valid = False

                if self.app_instance and self.app_instance.image and not self.app_instance.image.isNull():
                    img_w_orig = float(self.app_instance.image.width())
                    img_h_orig = float(self.app_instance.image.height())
                    label_w_widget = float(self.width())
                    label_h_widget = float(self.height())

                    if img_w_orig > 0 and img_h_orig > 0 and label_w_widget > 0 and label_h_widget > 0:
                        scale_factor_img_to_label = min(label_w_widget / img_w_orig, label_h_widget / img_h_orig)
                        if scale_factor_img_to_label > 1e-9:
                            displayed_img_w_in_label = img_w_orig * scale_factor_img_to_label
                            displayed_img_h_in_label = img_h_orig * scale_factor_img_to_label
                            offset_x_img_in_label = (label_w_widget - displayed_img_w_in_label) / 2.0
                            offset_y_img_in_label = (label_h_widget - displayed_img_h_in_label) / 2.0

                            if untransformed_label_pos.x() >= offset_x_img_in_label and \
                               untransformed_label_pos.x() <= offset_x_img_in_label + displayed_img_w_in_label and \
                               untransformed_label_pos.y() >= offset_y_img_in_label and \
                               untransformed_label_pos.y() <= offset_y_img_in_label + displayed_img_h_in_label:
                                
                                relative_x_in_display = untransformed_label_pos.x() - offset_x_img_in_label
                                relative_y_in_display = untransformed_label_pos.y() - offset_y_img_in_label
                                
                                img_x = relative_x_in_display / scale_factor_img_to_label
                                img_y = relative_y_in_display / scale_factor_img_to_label
                                image_coords_actual = QPointF(img_x, img_y)
                                is_image_coord_valid = True
                
                self.mouseMovedInLabel.emit(untransformed_label_pos, image_coords_actual if image_coords_actual else QPointF(), is_image_coord_valid)

                if self.is_panning and (event.buttons() & Qt.RightButton):
                    if self.pan_start_view_coords:
                        delta = event.position() - self.pan_start_view_coords
                        self.pan_offset = self.pan_offset_at_drag_start + delta
                        if self.app_instance: self.app_instance.update_live_view()
                    event.accept()
                    return
            
                if hasattr(self, '_custom_mouseMoveEvent_from_app') and self._custom_mouseMoveEvent_from_app:
                    self._custom_mouseMoveEvent_from_app(event)
                    if event.isAccepted():
                        return

                if self.preview_marker_enabled: 
                    snapped_label_pos = untransformed_label_pos
                    if self.app_instance:
                        snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                    self.preview_marker_position = snapped_label_pos
                    self.update()
                    event.accept()
                    return
                
                elif self.standard_marker_preview_enabled:
                    snapped_label_pos = untransformed_label_pos
                    if self.app_instance:
                        snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                    self.standard_marker_preview_position = snapped_label_pos
                    self.update()
                    event.accept()
                    return
                
                if self.mw_predict_preview_enabled: 
                    if self.app_instance and hasattr(self.app_instance, 'update_mw_predict_preview'):
                         self.app_instance.update_mw_predict_preview(event)
                    else:
                        snapped_label_pos = untransformed_label_pos
                        if self.app_instance: snapped_label_pos = self.app_instance.snap_point_to_grid(untransformed_label_pos)
                        self.mw_predict_preview_position = snapped_label_pos
                        self.update()
                    event.accept()
                    return
        
                # --- REMOVED OBSOLETE BLOCK THAT USED self.selected_point ---
                
                if not event.isAccepted():
                    super().mouseMoveEvent(event)
                
            def mousePressEvent(self, event: 'QMouseEvent'):
                if event.button() == Qt.RightButton and self.zoom_level > 1.001:
                    self.is_panning = True
                    self.pan_start_view_coords = event.position()
                    self.pan_offset_at_drag_start = QPointF(self.pan_offset)
                    self.setCursor(Qt.ClosedHandCursor)
                    event.accept() 
                    return
                
                if event.button() != Qt.RightButton and hasattr(self, '_custom_left_click_handler_from_app') and self._custom_left_click_handler_from_app:
                    self._custom_left_click_handler_from_app(event)
                    if event.isAccepted():
                        return
                
                if self.preview_marker_enabled and event.button() == Qt.LeftButton:
                    if self.app_instance and hasattr(self.app_instance, 'place_custom_marker'):
                        self.app_instance.place_custom_marker(event, self.preview_marker_text)
                    self.update()
                    event.accept()
                    return
        
                # --- REMOVED OBSOLETE BLOCK THAT USED self.selected_point ---

                if not event.isAccepted():
                    super().mousePressEvent(event)

            def mouseReleaseEvent(self, event: 'QMouseEvent'):
                if event.button() == Qt.RightButton and self.is_panning:
                    self.is_panning = False
                    self.pan_start_view_coords = None
                    self.pan_offset_at_drag_start = None
                    if self.zoom_level > 1.001: self.setCursor(Qt.OpenHandCursor)
                    else: self.setCursor(Qt.ArrowCursor)
                    event.accept()
                    return
            
                if event.button() != Qt.RightButton and hasattr(self, '_custom_mouseReleaseEvent_from_app') and self._custom_mouseReleaseEvent_from_app:
                    self._custom_mouseReleaseEvent_from_app(event)
                    if event.isAccepted():
                        return
                
                # --- REMOVED OBSOLETE BLOCK THAT USED self.selected_point ---
                
                if not event.isAccepted():
                    super().mouseReleaseEvent(event)

            def transform_point(self, point):
                """Transform a point from widget coordinates to image coordinates."""
                if self.zoom_level != 1.0:
                    return QPointF(
                        (point.x() - self.pan_offset.x()) / self.zoom_level,
                        (point.y() - self.pan_offset.y()) / self.zoom_level
                    )
                return QPointF(point)
            
            
            def zoom_in(self): # Zooms towards view center by default
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level *= 1.1
                self.zoom_level = min(self.zoom_level, 20.0) # Max zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if not self.is_panning: self.setCursor(Qt.ArrowCursor)
                
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()

            def zoom_out(self): # Zooms towards view center
                mouse_center_widget = QPoint(self.width() // 2, self.height() // 2) # Center of the label
                old_zoom_level = self.zoom_level
                self.zoom_level /= 1.1
                self.zoom_level = max(self.zoom_level, 0.1) # Min zoom
            
                point_before_zoom_unzoomed_label = QPointF(
                    (mouse_center_widget.x() - self.pan_offset.x()) / old_zoom_level,
                    (mouse_center_widget.y() - self.pan_offset.y()) / old_zoom_level
                )
                new_pan_x = mouse_center_widget.x() - (point_before_zoom_unzoomed_label.x() * self.zoom_level)
                new_pan_y = mouse_center_widget.y() - (point_before_zoom_unzoomed_label.y() * self.zoom_level)
                self.pan_offset = QPointF(new_pan_x, new_pan_y)
            
                if self.zoom_level <= 1.0:
                    self.zoom_level = 1.0
                    self.pan_offset = QPointF(0, 0)
                    if not self.is_panning: self.setCursor(Qt.ArrowCursor)
                else:
                    if not self.is_panning: self.setCursor(Qt.ArrowCursor)
            
                if self.app_instance: self.app_instance.update_live_view()
                else: self.update()
                

            def paintEvent(self, event):
                """
                The complete, refactored paint event.
                This method handles drawing all elements onto the live view, including the base image,
                all finalized markers and shapes, selection highlights, and live previews for
                actions currently in progress (like defining a new analysis region).
                """
                super().paintEvent(event)
                painter = QPainter(self)
                painter.setRenderHint(QPainter.TextAntialiasing, True)
                
                painter.save()
                if self.zoom_level != 1.0:
                    painter.translate(self.pan_offset)
                    painter.scale(self.zoom_level, self.zoom_level)

                # --- Coordinate mapping helpers ---
                _image_to_label_space_valid = False
                _scale_factor_img_to_label = 1.0
                _img_w_orig_from_app = 1.0
                _img_h_orig_from_app = 1.0
                _offset_x_img_in_label = 0.0
                _offset_y_img_in_label = 0.0
                _displayed_img_w_in_label = 0.0
                _displayed_img_h_in_label = 0.0
                if self.app_instance and self.app_instance.image and not self.app_instance.image.isNull():
                    current_app_image = self.app_instance.image
                    label_w_widget = float(self.width())
                    label_h_widget = float(self.height())
                    _img_w_orig_from_app = float(current_app_image.width())
                    _img_h_orig_from_app = float(current_app_image.height())
                    if _img_w_orig_from_app > 0 and _img_h_orig_from_app > 0 and label_w_widget > 0 and label_h_widget > 0:
                        _scale_factor_img_to_label = min(label_w_widget / _img_w_orig_from_app, label_h_widget / _img_h_orig_from_app)
                        if _scale_factor_img_to_label > 1e-9:
                            _displayed_img_w_in_label = _img_w_orig_from_app * _scale_factor_img_to_label
                            _displayed_img_h_in_label = _img_h_orig_from_app * _scale_factor_img_to_label
                            _offset_x_img_in_label = (label_w_widget - _displayed_img_w_in_label) / 2.0
                            _offset_y_img_in_label = (label_h_widget - _displayed_img_h_in_label) / 2.0
                            _image_to_label_space_valid = True
                
                if _image_to_label_space_valid:
                    painter.save()
                    border_pen = QPen(Qt.magenta)
                    border_pen_width = max(0.5, 2.0 / self.zoom_level if self.zoom_level > 0 else 2.0)
                    border_pen.setWidthF(border_pen_width)
                    painter.setPen(border_pen)
                    painter.setBrush(Qt.NoBrush)
                    image_rect_in_label = QRectF(_offset_x_img_in_label, _offset_y_img_in_label, _displayed_img_w_in_label, _displayed_img_h_in_label)
                    painter.drawRect(image_rect_in_label)
                    painter.restore()

                def _app_image_coords_to_unzoomed_label_space(img_coords_tuple_or_qpointf):
                    if not _image_to_label_space_valid:
                        if isinstance(img_coords_tuple_or_qpointf, QPointF): return img_coords_tuple_or_qpointf
                        return QPointF(img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])
                    img_x, img_y = (img_coords_tuple_or_qpointf.x(), img_coords_tuple_or_qpointf.y()) if isinstance(img_coords_tuple_or_qpointf, QPointF) else (img_coords_tuple_or_qpointf[0], img_coords_tuple_or_qpointf[1])
                    x_ls = _offset_x_img_in_label + img_x * _scale_factor_img_to_label; y_ls = _offset_y_img_in_label + img_y * _scale_factor_img_to_label
                    return QPointF(x_ls, y_ls)
                
                # --- Draw standard markers, custom markers, custom shapes (from app instance) ---
                if _image_to_label_space_valid and self.app_instance:
                    std_marker_font = QFont(self.app_instance.font_family); std_marker_font.setPixelSize(self.app_instance.font_size); std_marker_color = self.app_instance.font_color if hasattr(self.app_instance, 'font_color') else QColor(Qt.black)
                    painter.setFont(std_marker_font); painter.setPen(std_marker_color); font_metrics_std = QFontMetrics(std_marker_font)
                    text_height_std_label_space = font_metrics_std.height(); y_offset_text_baseline_std = text_height_std_label_space * 0.3
                    if hasattr(self.app_instance, 'left_markers'):
                        left_marker_offset_x_label_space = self.app_instance.left_marker_shift_added * _scale_factor_img_to_label
                        for y_pos_img, marker_text_val in self.app_instance.left_markers:
                            anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y(); text_to_draw = f"{marker_text_val} ⎯"; text_width_label_space = font_metrics_std.horizontalAdvance(text_to_draw)
                            draw_x_ls = _offset_x_img_in_label + left_marker_offset_x_label_space - text_width_label_space; draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                            painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    if hasattr(self.app_instance, 'right_markers'):
                            right_marker_start_x_label_space = _offset_x_img_in_label + (self.app_instance.right_marker_shift_added * _scale_factor_img_to_label)
                            for y_pos_img, marker_text_val in self.app_instance.right_markers:
                                anchor_y_label_space = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y(); text_to_draw = f"⎯ {marker_text_val}"
                                draw_x_ls = right_marker_start_x_label_space; draw_y_ls = anchor_y_label_space + y_offset_text_baseline_std
                                painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    if hasattr(self.app_instance, 'top_markers'):
                        top_marker_offset_y_label = self.app_instance.top_marker_shift_added * _scale_factor_img_to_label; rotation_angle = self.app_instance.font_rotation
                        for x_pos_img, marker_text_val in self.app_instance.top_markers:
                            anchor_x_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, 0)).x(); text_to_draw = str(marker_text_val); painter.save()
                            draw_baseline_y_ls = _offset_y_img_in_label + top_marker_offset_y_label + y_offset_text_baseline_std
                            painter.translate(anchor_x_label_space, draw_baseline_y_ls); painter.rotate(rotation_angle); painter.drawText(QPointF(0, 0), text_to_draw); painter.restore()
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_shapes'):
                    for shape_data in self.app_instance.custom_shapes:
                        try:
                            shape_type = shape_data.get('type'); color = QColor(shape_data.get('color', '#000000')); base_thickness = float(shape_data.get('thickness', 0.5)); thickness_on_label = base_thickness * _scale_factor_img_to_label
                            effective_pen_width = max(0.5, thickness_on_label / self.zoom_level if self.zoom_level > 0 else thickness_on_label); pen = QPen(color); pen.setWidthF(effective_pen_width); painter.setPen(pen)
                            if shape_type == 'line':
                                start_img = shape_data.get('start'); end_img = shape_data.get('end')
                                if start_img and end_img: start_label_space = _app_image_coords_to_unzoomed_label_space(start_img); end_label_space = _app_image_coords_to_unzoomed_label_space(end_img); painter.drawLine(start_label_space, end_label_space)
                            elif shape_type == 'rectangle':
                                rect_img = shape_data.get('rect')
                                if rect_img: x_img, y_img, w_img, h_img = rect_img; top_left_label_space = _app_image_coords_to_unzoomed_label_space((x_img, y_img)); w_label_space = w_img * _scale_factor_img_to_label; h_label_space = h_img * _scale_factor_img_to_label; painter.drawRect(QRectF(top_left_label_space, QSizeF(w_label_space, h_label_space)))
                        except Exception as e: print(f"Error drawing custom shape in LiveViewLabel.paintEvent: {shape_data}, {e}")
                if _image_to_label_space_valid and hasattr(self.app_instance, 'custom_markers'):
                    for marker_data_list in self.app_instance.custom_markers:
                        try:
                            x_pos_img, y_pos_img, marker_text_str, qcolor_obj, font_family_str, font_size_int, is_bold, is_italic = marker_data_list
                            anchor_label_space = _app_image_coords_to_unzoomed_label_space((x_pos_img, y_pos_img)); current_marker_font = QFont(font_family_str); current_marker_font.setPixelSize(font_size_int); current_marker_font.setBold(is_bold); current_marker_font.setItalic(is_italic); painter.setFont(current_marker_font)
                            if not isinstance(qcolor_obj, QColor): qcolor_obj = QColor(qcolor_obj)
                            if not qcolor_obj.isValid(): qcolor_obj = Qt.black
                            painter.setPen(qcolor_obj); font_metrics_marker = QFontMetrics(current_marker_font); text_bounding_rect_marker = font_metrics_marker.boundingRect(marker_text_str)
                            draw_x_marker = anchor_label_space.x() - (text_bounding_rect_marker.left() + text_bounding_rect_marker.width() / 2.0); draw_y_marker = anchor_label_space.y() - (text_bounding_rect_marker.top() + text_bounding_rect_marker.height() / 2.0)
                            painter.drawText(QPointF(draw_x_marker, draw_y_marker), marker_text_str)
                        except Exception as e: print(f"Error drawing app_instance custom marker: {marker_data_list}, {e}")
                
                # --- Draw various previews (standard marker, custom marker, MW line) ---
                if self.standard_marker_preview_enabled and self.standard_marker_preview_position:
                    # (This block remains unchanged)
                    painter.save()
                    painter.setOpacity(0.7)
                    std_marker_font = QFont(self.app_instance.font_family); std_marker_font.setPixelSize(self.app_instance.font_size)
                    painter.setFont(std_marker_font); painter.setPen(self.app_instance.font_color)
                    font_metrics_std = QFontMetrics(std_marker_font); y_offset_text_baseline_std = font_metrics_std.height() * 0.3
                    preview_text = self.standard_marker_preview_text; preview_pos = self.standard_marker_preview_position
                    if self.standard_marker_preview_mode == 'left':
                        text_to_draw = f"{preview_text} ⎯"; text_width = font_metrics_std.horizontalAdvance(text_to_draw)
                        is_first = len(self.app_instance.left_markers) == 0
                        anchor_x_ls = preview_pos.x() if is_first else (_offset_x_img_in_label + self.app_instance.left_marker_shift_added * _scale_factor_img_to_label)
                        draw_x_ls = anchor_x_ls - text_width; draw_y_ls = preview_pos.y() + y_offset_text_baseline_std
                        painter.drawText(QPointF(draw_x_ls, draw_y_ls), text_to_draw)
                    elif self.standard_marker_preview_mode == 'right':
                        text_to_draw = f"⎯ {preview_text}"
                        is_first = len(self.app_instance.right_markers) == 0
                        anchor_x_ls = preview_pos.x() if is_first else (_offset_x_img_in_label + self.app_instance.right_marker_shift_added * _scale_factor_img_to_label)
                        draw_y_ls = preview_pos.y() + y_offset_text_baseline_std
                        painter.drawText(QPointF(anchor_x_ls, draw_y_ls), text_to_draw)
                    elif self.standard_marker_preview_mode == 'top':
                        rotation_angle = self.app_instance.font_rotation
                        is_first = len(self.app_instance.top_markers) == 0
                        anchor_y_ls = preview_pos.y() if is_first else (_offset_y_img_in_label + self.app_instance.top_marker_shift_added * _scale_factor_img_to_label)
                        draw_baseline_y_ls = anchor_y_ls + y_offset_text_baseline_std
                        painter.save(); painter.translate(preview_pos.x(), draw_baseline_y_ls); painter.rotate(rotation_angle)
                        painter.drawText(QPointF(0, 0), preview_text); painter.restore()
                    painter.restore()

                if self.preview_marker_enabled and self.preview_marker_position:
                    # (This block remains unchanged)
                    painter.setOpacity(0.7)
                    marker_preview_font = QFont(self.marker_font_type); marker_preview_font.setPixelSize(self.marker_font_size)
                    painter.setFont(marker_preview_font); painter.setPen(self.marker_color)
                    font_metrics_preview = QFontMetrics(marker_preview_font); preview_text_rect = font_metrics_preview.boundingRect(self.preview_marker_text)
                    draw_x_preview = self.preview_marker_position.x() - (preview_text_rect.left() + preview_text_rect.width() / 2.0)
                    draw_y_preview = self.preview_marker_position.y() - (preview_text_rect.top() + preview_text_rect.height() / 2.0)
                    painter.drawText(QPointF(draw_x_preview, draw_y_preview), self.preview_marker_text)
                    painter.setOpacity(1.0)
                
                # --- Selection Highlights ---
                if self.app_instance.overlay_mode_active and self.app_instance.selected_overlay_index > 0:
                    # (This block remains unchanged)
                    rect_ls = self.app_instance._get_overlay_rect_in_label_space(self.app_instance.selected_overlay_index)
                    if rect_ls:
                        highlight_pen = QPen(QColor(0, 255, 255, 200)); pen_width = max(0.5, 1.0 / self.zoom_level if self.zoom_level > 0 else 1.0)
                        highlight_pen.setWidthF(pen_width); highlight_pen.setStyle(Qt.DashLine); painter.setPen(highlight_pen); painter.drawRect(rect_ls)
                        handle_pen = QPen(QColor(255, 0, 0)); handle_pen.setWidthF(pen_width); painter.setPen(handle_pen)
                        painter.setBrush(QColor(255, 0, 0, 150)); handle_radius = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS
                        corners = [rect_ls.topLeft(), rect_ls.topRight(), rect_ls.bottomRight(), rect_ls.bottomLeft()]
                        for i, corner in enumerate(corners):
                            radius = handle_radius * 1.5 if self.app_instance.resizing_overlay_corner_index == i else handle_radius
                            painter.drawEllipse(corner, radius, radius)
                if self.app_instance and self.app_instance.moving_custom_item_info:
                    # (This block remains unchanged)
                    info = self.app_instance.moving_custom_item_info; is_resizing = self.app_instance.current_selection_mode == "resizing_custom_item"
                    selection_pen = QPen(Qt.magenta, max(0.5, 2.0 / self.zoom_level), Qt.DotLine); handle_pen = QPen(Qt.red, max(0.5, 2.0 / self.zoom_level))
                    handle_brush = QBrush(Qt.red); handle_radius = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS
                    selection_shape = None
                    if info['type'] == 'marker': selection_shape = self.app_instance._get_marker_bounding_box_in_label_space(info['index'])
                    elif info['type'] in ['left_marker', 'right_marker', 'top_marker']: selection_shape = self.app_instance._get_standard_marker_bounding_box_in_label_space(info['type'], info['index'])
                    elif info['type'] == 'shape':
                        body_ls, handles_ls = self.app_instance._get_shape_bounding_box_and_handles_in_label_space(info['index'])
                        if body_ls and handles_ls:
                            painter.setPen(selection_pen); painter.drawRect(body_ls.adjusted(-2,-2,2,2)); painter.setPen(handle_pen); painter.setBrush(handle_brush)
                            for idx, handle_pt in enumerate(handles_ls):
                                if is_resizing and idx == self.app_instance.resizing_corner_index: painter.drawEllipse(handle_pt, handle_radius * 1.2, handle_radius * 1.2)
                                else: painter.drawEllipse(handle_pt, handle_radius, handle_radius)
                            painter.setBrush(Qt.NoBrush)
                        selection_shape = None
                    else: selection_shape = None
                    if selection_shape:
                        painter.setPen(selection_pen)
                        if isinstance(selection_shape, QPolygonF): painter.drawPolygon(selection_shape)
                        elif isinstance(selection_shape, QRectF): painter.drawRect(selection_shape)

                # --- Crop and Custom Shape Drawing Previews ---
                preview_pen_crop = QPen(Qt.red); effective_pen_width_crop = max(0.5, 1.0 / self.zoom_level if self.zoom_level > 0 else 0.5); preview_pen_crop.setWidthF(effective_pen_width_crop)
                if self.drawing_crop_rect and self.crop_rect_start_view and self.crop_rect_end_view:
                    # (This block remains unchanged)
                    preview_pen_crop.setStyle(Qt.DashLine); painter.setPen(preview_pen_crop); rect_to_draw = QRectF(self.crop_rect_start_view, self.crop_rect_end_view).normalized(); painter.drawRect(rect_to_draw)
                elif self.crop_rect_final_view:
                    # (This block remains unchanged)
                    preview_pen_crop.setStyle(Qt.SolidLine); painter.setPen(preview_pen_crop); painter.drawRect(self.crop_rect_final_view)
                if self.app_instance and self.app_instance.drawing_mode in ['line', 'rectangle'] and self.app_instance.current_drawing_shape_preview:
                    # (This block remains unchanged)
                    try:
                        start_pt_ls = self.app_instance.current_drawing_shape_preview['start']; end_pt_ls = self.app_instance.current_drawing_shape_preview['end']; preview_color = self.app_instance.custom_marker_color
                        base_preview_thickness = float(self.app_instance.custom_font_size_spinbox.value()); effective_preview_thickness = min(1.0, base_preview_thickness / self.zoom_level if self.zoom_level > 0 else base_preview_thickness)
                        preview_pen_shape = QPen(preview_color); preview_pen_shape.setWidthF(effective_preview_thickness); preview_pen_shape.setStyle(Qt.DotLine); painter.setPen(preview_pen_shape)
                        if self.app_instance.drawing_mode == 'line': painter.drawLine(start_pt_ls, end_pt_ls)
                        elif self.app_instance.drawing_mode == 'rectangle': painter.drawRect(QRectF(start_pt_ls, end_pt_ls).normalized())
                    except Exception as e: print(f"Error drawing live shape preview in paintEvent: {e}")

                # --- Draw MW Prediction Lines ---
                anchor_point_ls = None
                if self.mw_predict_preview_enabled and self.mw_predict_preview_position: anchor_point_ls = self.mw_predict_preview_position
                elif self.app_instance and hasattr(self.app_instance, "protein_location") and self.app_instance.protein_location: anchor_point_ls = self.app_instance.protein_location
                if anchor_point_ls:
                    # (This entire complex block for drawing MW lines remains unchanged)
                    line_symbol = "⎯⎯"; base_font_size = 25
                    predict_font = QFont(self.app_instance.custom_font_type_dropdown.currentText()); scaled_pixel_size = max(8, int(base_font_size / self.zoom_level))
                    predict_font.setPixelSize(scaled_pixel_size); predict_fm = QFontMetricsF(predict_font); predict_line_rect = predict_fm.boundingRect(line_symbol)
                    center_x_ls = anchor_point_ls.x(); predict_line_draw_start_x_ls = center_x_ls - (predict_line_rect.left() + predict_line_rect.width() / 2.0)
                    if self.mw_predict_preview_enabled and self.mw_predict_preview_position:
                        painter.save(); painter.setOpacity(0.7); painter.setFont(predict_font); painter.setPen(Qt.darkGreen)
                        loc_y_ls = self.mw_predict_preview_position.y(); draw_y_mw_ls = loc_y_ls - (predict_line_rect.top() + predict_line_rect.height() / 2.0)
                        painter.drawText(QPointF(predict_line_draw_start_x_ls, draw_y_mw_ls), line_symbol); painter.restore()
                    should_draw_single_predict_line = (self.app_instance and hasattr(self.app_instance, "protein_location") and self.app_instance.protein_location and not self.app_instance.run_predict_MW)
                    if should_draw_single_predict_line:
                        painter.setFont(predict_font); painter.setPen(Qt.green)
                        loc_y_ls = self.app_instance.protein_location.y(); draw_y_mw_ls = loc_y_ls - (predict_line_rect.top() + predict_line_rect.height() / 2.0)
                        painter.drawText(QPointF(predict_line_draw_start_x_ls, draw_y_mw_ls), line_symbol)
                    should_draw_oligomer_overlay = (self.app_instance and hasattr(self.app_instance, 'show_oligomer_glyco_overlay_checkbox') and self.app_instance.show_oligomer_glyco_overlay_checkbox.isChecked() and self.app_instance.oligomer_products and self.app_instance.last_mw_prediction_model is not None)
                    if should_draw_oligomer_overlay:
                        line_colors = [QColor(255, 140, 0, 220), QColor(0, 128, 0, 220), QColor(0, 0, 139, 220), QColor(139, 0, 139, 220)]
                        text_colors = [QColor("#b35900"), QColor("#004d00"), QColor("#000052"), QColor("#520052")]
                        
                        # --- UPDATED: Use Standard Marker Font Settings ---
                        # Uses app_instance.font_family/font_size instead of custom settings
                        base_font_size = self.app_instance.font_size
                        font_family = self.app_instance.font_family
                        scaled_pixel_size = max(4, int(base_font_size / self.zoom_level))

                        text_font = QFont(font_family)
                        text_font.setPixelSize(scaled_pixel_size)
                        text_font.setBold(True)
                        painter.setFont(text_font)
                        # --------------------------------------------------

                        fm_text = QFontMetricsF(text_font)
                        text_height = fm_text.height()
                        min_text_spacing = text_height * 1.2
                        bands_to_draw = []
                        for i, mw in enumerate(self.app_instance.oligomer_products):
                            y_pos_img = self.app_instance._get_y_pos_from_mw(mw, self.app_instance.last_mw_prediction_model, self.app_instance.last_mw_prediction_min_max_pos)
                            if y_pos_img is not None:
                                y_pos_ls = _app_image_coords_to_unzoomed_label_space((0, y_pos_img)).y()
                                bands_to_draw.append({'mw': mw, 'y_ls': y_pos_ls, 'color': line_colors[i % len(line_colors)], 'text_color': text_colors[i % len(text_colors)]})
                        if not bands_to_draw: painter.restore(); return
                        bands_to_draw.sort(key=lambda b: b['y_ls'])
                        
                        last_text_y = -float('inf')
                        for band in bands_to_draw:
                            ideal_text_y = band['y_ls']
                            if ideal_text_y < last_text_y + min_text_spacing: band['text_y_ls'] = last_text_y + min_text_spacing
                            else: band['text_y_ls'] = ideal_text_y
                            last_text_y = band['text_y_ls']
                        
                        arrow_line_width = max(0.8, 2.0 / self.zoom_level)
                        
                        # Determine Start X (Cursor location or Image Center)
                        start_x = 0
                        if self.mw_predict_preview_enabled and self.mw_predict_preview_position:
                             start_x = self.mw_predict_preview_position.x()
                        elif self.app_instance and hasattr(self.app_instance, "protein_location") and self.app_instance.protein_location:
                             img_pt = self.app_instance.protein_location
                             # Convert Image Space -> Label Space
                             start_x = _app_image_coords_to_unzoomed_label_space(img_pt).x()
                        else:
                             start_x = _offset_x_img_in_label + _displayed_img_w_in_label / 2

                        for band in bands_to_draw:
                            pen = QPen(band['color']); pen.setWidthF(arrow_line_width); painter.setPen(pen)
                            text = f"{band['mw']:.1f} kDa"; text_rect = fm_text.boundingRect(text)
                            
                            # Draw stepped lines
                            line_len_1 = 30 / self.zoom_level
                            elbow_x = start_x + line_len_1
                            text_start_x = elbow_x + (5 / self.zoom_level)
                            
                            p_start = QPointF(start_x, band['y_ls'])
                            p_elbow_1 = QPointF(elbow_x, band['y_ls'])
                            p_elbow_2 = QPointF(elbow_x, band['text_y_ls'])
                            p_text_anchor = QPointF(text_start_x, band['text_y_ls'])

                            # Use Polyline for clean joints
                            polyline_points = [p_start, p_elbow_1, p_elbow_2, p_text_anchor]
                            painter.drawPolyline(QPolygonF(polyline_points))
                            
                            text_draw_y = band['text_y_ls'] - (text_rect.top() + text_rect.height() / 2.0)
                            
                            painter.setFont(text_font)

                            glow_color = QColor(255, 255, 255, 90)
                            glow_pen = QPen(glow_color)
                            glow_pen.setWidth(2) 
                            painter.setPen(glow_pen)
                            
                            glow_offset = max(0.5, 1.0 / self.zoom_level)
                            offsets = [
                                (-glow_offset, -glow_offset), (glow_offset, -glow_offset),
                                (-glow_offset, glow_offset), (glow_offset, glow_offset),
                                (0, -glow_offset), (0, glow_offset),
                                (-glow_offset, 0), (glow_offset, 0)
                            ]
                            for dx, dy in offsets:
                                painter.drawText(QPointF(text_start_x + dx, text_draw_y + dy), text)

                            painter.setPen(band['text_color'])
                            painter.drawText(QPointF(text_start_x, text_draw_y), text)

                # --- START OF NEW STREAMLINED PREVIEW LOGIC ---
                # This single block replaces all the old, fragmented preview drawing logic for analysis regions.
                if self.mode in ['define_rect', 'define_quad', 'auto_lane_rect', 'auto_lane_quad']:
                    preview_pen_width = max(0.5, 1.5 / self.zoom_level if self.zoom_level > 0 else 1.5)
                    handle_radius_view = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS
                    points_to_draw = self.current_preview_points

                    if self.mode in ['define_rect', 'auto_lane_rect'] and len(points_to_draw) == 2:
                        painter.setPen(QPen(Qt.darkMagenta, preview_pen_width, Qt.DotLine))
                        painter.drawRect(QRectF(points_to_draw[0], points_to_draw[1]).normalized())

                    elif self.mode in ['define_quad', 'auto_lane_quad'] and points_to_draw:
                        painter.setPen(QPen(Qt.magenta, preview_pen_width * 1.5, Qt.SolidLine))
                        for p_ls in points_to_draw:
                            painter.drawEllipse(p_ls, handle_radius_view, handle_radius_view)
                        if 0 < len(points_to_draw) < 4:
                            painter.setPen(QPen(Qt.darkMagenta, preview_pen_width, Qt.DotLine))
                            painter.drawPolyline(QPolygonF(points_to_draw))
                # --- END OF NEW STREAMLINED PREVIEW LOGIC ---

                # --- Draw Finalized Multi-Lane Shapes ---
                if self.app_instance and hasattr(self.app_instance, 'multi_lane_definitions') and self.app_instance.multi_lane_definitions:
                    lane_font = QFont("Arial")
                    font_pixel_size = max(4, int(10 / self.zoom_level if self.zoom_level > 0 else 10))
                    lane_font.setPixelSize(font_pixel_size)
                    lane_font.setBold(True)
                    
                    for i, lane_def in enumerate(self.app_instance.multi_lane_definitions):
                        is_selected_this_lane = (self.app_instance.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner", "skewing_edge"] and self.app_instance.moving_multi_lane_index == i)
                        pen_color = Qt.magenta if is_selected_this_lane else Qt.darkYellow
                        pen_width_multilane = max(0.5, (2.0 if is_selected_this_lane else 1.0) / self.zoom_level)
                        pen_defined_lane = QPen(pen_color, pen_width_multilane, Qt.SolidLine); painter.setPen(pen_defined_lane)
                        center_point, corners = QPointF(), []
                        
                        if lane_def['type'] == 'quad':
                            corners = lane_def['points_label']
                            painter.drawPolygon(QPolygonF(corners))
                        elif lane_def['type'] == 'rectangle':
                            rect_ls = lane_def['points_label'][0]
                            corners = [rect_ls.topLeft(), rect_ls.topRight(), rect_ls.bottomRight(), rect_ls.bottomLeft()]
                            painter.drawRect(rect_ls)
                        
                        if len(corners) == 4:
                            center_point = (corners[0] + corners[1] + corners[2] + corners[3]) / 4.0

                        if self.app_instance.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner", "skewing_edge"]:
                            point_pen_color = Qt.red if is_selected_this_lane and self.app_instance.current_selection_mode != "select_for_move" else Qt.blue
                            handle_radius_view = self.CORNER_HANDLE_BASE_RADIUS / self.zoom_level if self.zoom_level > 0 else self.CORNER_HANDLE_BASE_RADIUS
                            painter.setPen(QPen(point_pen_color, pen_width_multilane * 1.5))
                            for idx, p_ls_multi in enumerate(corners):
                                is_resizing_this_corner = (is_selected_this_lane and self.app_instance.current_selection_mode == "resizing_corner" and self.app_instance.resizing_corner_index == idx)
                                if is_resizing_this_corner: 
                                    painter.setBrush(QBrush(Qt.red)); painter.drawEllipse(p_ls_multi, handle_radius_view * 1.2, handle_radius_view * 1.2); painter.setBrush(Qt.NoBrush)
                                else: painter.drawEllipse(p_ls_multi, handle_radius_view, handle_radius_view)
                            
                            if len(corners) == 4:
                                edge_pen_color = Qt.green if is_selected_this_lane and self.app_instance.current_selection_mode == "skewing_edge" else point_pen_color
                                painter.setPen(QPen(edge_pen_color, pen_width_multilane)); painter.setBrush(QBrush(edge_pen_color))
                                
                                top_mid = (corners[0] + corners[1]) / 2.0
                                bottom_mid = (corners[3] + corners[2]) / 2.0
                                left_mid = (corners[0] + corners[3]) / 2.0
                                right_mid = (corners[1] + corners[2]) / 2.0

                                handle_width = handle_radius_view * 2.5
                                handle_height = handle_radius_view * 0.8
                                
                                # Top/Bottom edge handles (horizontal rectangles)
                                painter.drawRect(QRectF(top_mid.x() - handle_width/2, top_mid.y() - handle_height/2, handle_width, handle_height))
                                painter.drawRect(QRectF(bottom_mid.x() - handle_width/2, bottom_mid.y() - handle_height/2, handle_width, handle_height))

                                # --- START: Draw Left/Right Edge Handles ---
                                # Use vertical rectangles for the side handles
                                painter.drawRect(QRectF(left_mid.x() - handle_height/2, left_mid.y() - handle_width/2, handle_height, handle_width))
                                painter.drawRect(QRectF(right_mid.x() - handle_height/2, right_mid.y() - handle_width/2, handle_height, handle_width))
                                # --- END: Draw Left/Right Edge Handles ---

                                painter.setBrush(Qt.NoBrush)

                        if not center_point.isNull():
                            lane_id_str = str(lane_def['id'])
                            painter.setFont(lane_font)
                            
                            fm_lane = QFontMetrics(lane_font)
                            text_rect = fm_lane.boundingRect(lane_id_str)
                            
                            # --- START OF ALIGNMENT FIX ---
                            # 1. Determine the diameter needed for a circle that fits the text.
                            padding = 3 # Increase padding for better visual balance
                            diameter = max(text_rect.width(), text_rect.height()) + padding * 2
                            
                            # 2. Create a perfect square QRectF for the background circle.
                            bg_rect = QRectF(0, 0, diameter, diameter)
                            bg_rect.moveCenter(center_point)
                            
                            # 3. Draw the circle and the perfectly centered text.
                            painter.save()
                            painter.setBrush(QColor(255, 255, 255, 180))
                            painter.setPen(Qt.NoPen)
                            painter.drawEllipse(bg_rect)
                            painter.restore()
                            
                            painter.setPen(Qt.red if is_selected_this_lane else Qt.black)
                            painter.drawText(bg_rect, Qt.AlignCenter, lane_id_str)
                
                # Drag/Duplicate Preview
                if self.drag_preview_rect or self.drag_preview_quad_points:
                    # (This block remains unchanged)
                    painter.save()
                    preview_pen_width = max(0.5, 1.5 / self.zoom_level if self.zoom_level > 0 else 1.5)
                    preview_brush = QBrush(QColor(70, 130, 180, 80)); preview_pen = QPen(QColor(70, 130, 180), preview_pen_width, Qt.DashLine)
                    painter.setPen(preview_pen); painter.setBrush(preview_brush)
                    if self.drag_preview_rect and not self.drag_preview_rect.isNull(): painter.drawRect(self.drag_preview_rect)
                    elif self.drag_preview_quad_points: painter.drawPolygon(QPolygonF(self.drag_preview_quad_points))
                    painter.restore()

                # --- Draw Grid Lines ---
                if self.app_instance and hasattr(self.app_instance, 'grid_size_input') and hasattr(self.app_instance, 'show_grid_checkbox_x') and hasattr(self.app_instance, 'show_grid_checkbox_y'):
                    # (This block remains unchanged)
                    grid_size_label_space = self.app_instance.grid_size_input.value()
                    if grid_size_label_space > 0:
                        pen_grid_paint = QPen(Qt.red); pen_grid_paint.setStyle(Qt.DashLine); effective_pen_width_grid = max(0.25, 0.5 / self.zoom_level if self.zoom_level > 0 else 0.5)
                        pen_grid_paint.setWidthF(effective_pen_width_grid); painter.setPen(pen_grid_paint)
                        label_width_unzoomed = self.width() / (self.zoom_level if self.zoom_level > 0 else 0.5); label_height_unzoomed = self.height() / (self.zoom_level if self.zoom_level > 0 else 0.5)
                        view_origin_x_unzoomed = -self.pan_offset.x() / (self.zoom_level if self.zoom_level > 0 else 0.5); view_origin_y_unzoomed = -self.pan_offset.y() / (self.zoom_level if self.zoom_level > 0 else 0.5)
                        if self.app_instance.show_grid_checkbox_x.isChecked():
                            start_x_grid = (int(view_origin_x_unzoomed / grid_size_label_space) -1) * grid_size_label_space
                            for x_grid_ls in range(start_x_grid, int(view_origin_x_unzoomed + label_width_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(x_grid_ls, view_origin_y_unzoomed), QPointF(x_grid_ls, view_origin_y_unzoomed + label_height_unzoomed))
                        if self.app_instance.show_grid_checkbox_y.isChecked():
                            start_y_grid = (int(view_origin_y_unzoomed / grid_size_label_space)-1) * grid_size_label_space
                            for y_grid_ls in range(start_y_grid, int(view_origin_y_unzoomed + label_height_unzoomed + grid_size_label_space), grid_size_label_space):
                                painter.drawLine(QPointF(view_origin_x_unzoomed, y_grid_ls), QPointF(view_origin_x_unzoomed + label_width_unzoomed, y_grid_ls))

                if self.app_instance:
                        # 1. Draw the finalized, persistent shape and its result label first
                        if self.app_instance.finalized_measurement_shape:
                            painter.save()
                            final_pen = QPen(QColor(255, 255, 0, 220), max(1.0, 2.5 / self.zoom_level), Qt.SolidLine)
                            painter.setPen(final_pen)
                            painter.setBrush(QColor(255, 255, 0, 60) if self.app_instance.finalized_measurement_shape.isClosed() else Qt.NoBrush)
                            painter.drawPolygon(self.app_instance.finalized_measurement_shape)
    
                            if self.app_instance.last_measurement_result_text:
                                label_font = QFont("Arial", max(8, int(12 / self.zoom_level))); label_font.setBold(True)
                                painter.setFont(label_font)
                                text_rect = self.app_instance.finalized_measurement_shape.boundingRect()
                                text_pos = text_rect.center()
                                fm = QFontMetrics(label_font)
                                text_bound = fm.boundingRect(self.app_instance.last_measurement_result_text).adjusted(-4,-2,4,2)
                                text_bound.moveCenter(text_pos.toPoint())
                                painter.setBrush(QColor(0, 0, 0, 150)); painter.setPen(Qt.NoPen)
                                painter.drawRoundedRect(text_bound, 3, 3)
                                painter.setPen(Qt.white)
                                painter.drawText(text_bound, Qt.AlignCenter, self.app_instance.last_measurement_result_text)
                            painter.restore()

                        # 2. Draw the live, in-progress shape (markers and rubber-band line)
                        if self.app_instance.measurement_mode and self.app_instance.measurement_points:
                            painter.save()
                            # Define pens and radius for previews
                            preview_pen = QPen(QColor(0, 255, 255, 200), max(0.8, 2.0 / self.zoom_level), Qt.DotLine)
                            marker_pen = QPen(QColor(0, 255, 255, 220), max(1.0, 2.0 / self.zoom_level), Qt.SolidLine)
                            marker_radius = 5 / self.zoom_level if self.zoom_level > 0 else 5
                            
                            # Draw markers at each clicked point
                            painter.setPen(marker_pen)
                            for p in self.app_instance.measurement_points:
                                painter.drawEllipse(p, marker_radius, marker_radius)
                            
                            # Draw the polyline connecting the clicked points
                            painter.setPen(preview_pen)
                            if len(self.app_instance.measurement_points) > 1:
                                painter.drawPolyline(QPolygonF(self.app_instance.measurement_points))

                            # Draw the "rubber-band" line from the last point to the current cursor position
                            # 'current_preview_points' should contain only the live cursor position from mouseMove
                            if self.app_instance.live_view_label.current_preview_points:
                                last_clicked_point = self.app_instance.measurement_points[-1]
                                cursor_point = self.app_instance.live_view_label.current_preview_points[0]
                                painter.drawLine(last_clicked_point, cursor_point)
                            
                            painter.restore()

                painter.restore()
            
            def leaveEvent(self, event):
                super().leaveEvent(event)
                # Emit signal with invalid label coordinates and False for image coordinate validity
                self.mouseMovedInLabel.emit(QPointF(-1,-1), QPointF(), False)

            def keyPressEvent(self, event):
                if event.key() == Qt.Key_Escape:
                    if self.app_instance and hasattr(self.app_instance, 'keyPressEvent'):
                        self.app_instance.keyPressEvent(event)
                        if event.isAccepted():
                            return
                
                super().keyPressEvent(event)

        class CombinedSDSApp(QMainWindow):
            CONFIG_PRESET_FILE_NAME = "Gel_Blot_Analyzer_preset_config.txt"
            MIME_TYPE_CUSTOM_ITEMS = "application/x-Gel-Blot-Analyzer.customitems+json"
            light_stylesheet = """
                /* 
                ================================================================================
                POLISHED COMPACT STYLESHEET for Gel Blot Analyzer (LIGHT THEME)
                ================================================================================
                */
                QMainWindow, QDialog { background-color: #F0F2F5; }
                QWidget { font-family: "Segoe UI", Arial, sans-serif; font-size: 12px; color: #333333; }
                QGroupBox { background-color: #FBFCFD; border: 1px solid #D0D5DB; border-radius: 6px; margin-top: 18px; padding: 8px 5px 5px 5px; }
                QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 2px 8px; left: 10px; color: #FFFFFF; background-color: #5D98D4; border-top-left-radius: 4px; border-top-right-radius: 4px; font-weight: bold; }
                QPushButton { background-color: #FFFFFF; border: 1px solid #C0C5CB; border-radius: 3px; padding: 4px 8px; min-height: 18px; }
                QPushButton:hover { background-color: #E6F0F9; border-color: #5D98D4; }
                QPushButton:pressed { background-color: #D0E0EF; }
                QPushButton:checked { background-color: #D4EDDA; border: 1px solid #74B882; }
                QPushButton:disabled { background-color: #F0F2F5; color: #AAAAAA; }
                QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QComboBox, QFontComboBox { background-color: #FFFFFF; border: 1px solid #D0D5DB; border-radius: 4px; padding: 4px 6px; min-height: 20px; selection-background-color: #5D98D4; selection-color: white; color: #333333; }
                QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus, QFontComboBox:focus { border: 1px solid #5D98D4; }
                QLineEdit:disabled, QTextEdit:disabled, QSpinBox:disabled, QDoubleSpinBox:disabled, QComboBox:disabled, QFontComboBox:disabled { background-color: #F0F2F5; color: #999999; }
                QComboBox::drop-down, QFontComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 22px; border-left-width: 1px; border-left-color: #D0D5DB; border-left-style: solid; border-top-right-radius: 3px; border-bottom-right-radius: 3px; }
                QComboBox::down-arrow, QFontComboBox::down-arrow { image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZHToPSIxNiIgdmlld0JveD0iMCAwIDE2IDE2Ij48cGF0aCBmaWxsPSIjNTU1NTU1IiBkPSJNMyA2bDUgNS4wMDFM MTQgNnoiLz48L3N2Zz4=); width: 12px; height: 12px; }
                QComboBox QAbstractItemView, QFontComboBox QAbstractItemView { background-color: #FFFFFF; border: 1px solid #C0C5CB; selection-background-color: #5D98D4; }
                QSlider::groove:horizontal { border: 1px solid #C0C5CB; background: #FFFFFF; height: 4px; border-radius: 2px; }
                QSlider::handle:horizontal { background: #5D98D4; border: 1px solid #4A78A9; width: 14px; height: 14px; margin: -5px 0; border-radius: 7px; }
                QTabWidget::pane { border-top: 1px solid #D0D5DB; }
                QTabBar::tab { background-color: #E4E7EB; border: 1px solid #D0D5DB; border-bottom: none; border-top-left-radius: 3px; border-top-right-radius: 3px; padding: 5px 10px; margin-right: 2px; }
                QTabBar::tab:selected { background: #FBFCFD; border-bottom-color: #FBFCFD; font-weight: bold; }
                QTabBar::tab:!selected:hover { background: #EFF2F5; }
                QCheckBox { spacing: 5px; }
                QCheckBox::indicator {
                    width: 13px;
                    height: 13px;
                    border-radius: 3px;
                    border: 1px solid #C0C5CB;
                    background-color: #FFFFFF;
                }
                QCheckBox::indicator:hover {
                    border-color: #5D98D4;
                }
                QCheckBox::indicator:checked {
                    background-color: #5D98D4;
                    border-color: #4A78A9;
                    image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxMiIgaGVpZ2h0PSIxMiIgdmlld0JveD0iMCAwIDI0IDI0Ij48cGF0aCBmaWxsPSJ3aGl0ZSIgZD0iTTkgMTYuMTdMNC44MyAxMmwtMS40MiAxLjQxTDkgMTlsNy03Ljg5TDIxIDdsLTEuNDEtMS40MUw5IDE2LjE3eiIvPjwvc3ZnPg==);
                }
                QCheckBox::indicator:disabled {
                    background-color: #E0E0E0;
                }
                QToolBar { background-color: #E4E7EB; border: none; padding: 1px; }
                QToolBar QToolButton {
                    border: 1px solid transparent;
                    border-radius: 3px;
                    padding: 2px;
                }
                QToolBar QToolButton:hover {
                    background-color: #E6F0F9;
                    border: 1px solid #5D98D4;
                }
                QToolBar QToolButton:pressed {
                    background-color: #D0E0EF;
                }
                QToolBar QToolButton:checked { background-color: #A0D0A0; }
                QStatusBar { background-color: #E4E7EB; }
                #LiveViewLabel { background-color: white; border: 1px solid #AAAAAA; }

                /* --- COMPLETE STYLES FOR TABLES, LISTS, AND SCROLLBARS --- */
                QTableView, QTableWidget, QListView, QListWidget {
                    border: 1px solid #D0D5DB;
                    gridline-color: #EAEAEA;
                    selection-color: white; /* Text color for selected items */
                }
                
                /* --- START OF FIX: Explicitly set item colors and selection highlight --- */
                QTableView::item, QTableWidget::item, QListView::item, QListWidget::item {
                    color: #333333; /* Dark gray for readability on white background */
                    background-color: #FFFFFF; /* Ensure item background is white */
                }
                QTableView::item:selected, QTableWidget::item:selected, QListView::item:selected, QListWidget::item:selected {
                    background-color: #5D98D4; /* Blue highlight color */
                    color: white; /* White text on blue highlight */
                }
                /* --- END OF FIX --- */
                
                QAbstractItemView {
                    background-color: #FFFFFF;
                }

                QHeaderView {
                    background-color: #F0F2F5;
                }

                QHeaderView::section {
                    background-color: #E4E7EB;
                    color: #333333;
                    padding: 4px;
                    border-top: 0px;
                    border-left: 0px;
                    border-right: 1px solid #D0D5DB;
                    border-bottom: 2px solid #C0C5CB;
                    font-weight: bold;
                }

                QTableCornerButton::section {
                    background-color: #E4E7EB;
                    border-right: 1px solid #D0D5DB;
                    border-bottom: 2px solid #C0C5CB;
                }

                /* --- Light Scrollbars --- */
                QScrollBar:vertical {
                    border: none;
                    background-color: #F0F2F5;
                    width: 12px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background-color: #C0C5CB;
                    min-height: 25px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical:hover {
                    background-color: #A8B0B6;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }

                QScrollBar:horizontal {
                    border: none;
                    background-color: #F0F2F5;
                    height: 12px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background-color: #C0C5CB;
                    min-width: 25px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal:hover {
                    background-color: #A8B0B6;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: none;
                }
            """
            
            dark_stylesheet = """
                /* 
                ================================================================================
                POLISHED COMPACT STYLESHEET for Gel Blot Analyzer (DARK THEME - FINAL)
                ================================================================================
                */
                QMainWindow, QDialog { background-color: #2D2D30; }
                QWidget { font-family: "Segoe UI", Arial, sans-serif; font-size: 12px; color: #F1F1F1; }
                
                /* --- FIX: Force Dark Backgrounds on Containers --- */
                QScrollArea { background-color: #2D2D30; border: none; }
                QScrollArea > QWidget { background-color: #2D2D30; }
                QScrollArea > QWidget > QWidget { background-color: #2D2D30; }
                
                QTabWidget { background-color: #2D2D30; }
                QTabWidget::pane { border-top: 1px solid #505055; background-color: #2D2D30; }
                /* ------------------------------------------------ */

                QGroupBox { background-color: #38383C; border: 1px solid #505055; border-radius: 6px; margin-top: 18px; padding: 8px 5px 5px 5px; }
                QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 2px 8px; left: 10px; color: #F1F1F1; background-color: #007ACC; border-top-left-radius: 4px; border-top-right-radius: 4px; font-weight: bold; }
                
                QPushButton { background-color: #4A4A4F; border: 1px solid #606065; border-radius: 3px; padding: 4px 8px; min-height: 18px; color: #F1F1F1; }
                QPushButton:hover { background-color: #5A5A60; border-color: #007ACC; }
                QPushButton:pressed { background-color: #6A6A70; }
                QPushButton:checked { background-color: #3D984E; border: 1px solid #5DBB6F; color: white; }
                QPushButton:disabled { background-color: #3A3A3D; color: #707070; border-color: #4A4A4F; }
                
                QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QComboBox, QFontComboBox { background-color: #3C3C3F; border: 1px solid #505055; border-radius: 4px; padding: 4px 6px; min-height: 20px; selection-background-color: #007ACC; selection-color: white; color: #F1F1F1; }
                QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus, QFontComboBox:focus { border: 1px solid #007ACC; }
                QLineEdit:disabled, QTextEdit:disabled, QSpinBox:disabled, QDoubleSpinBox:disabled, QComboBox:disabled, QFontComboBox:disabled { background-color: #3A3A3D; color: #707070; }
                
                QComboBox::drop-down, QFontComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 22px; border-left-width: 1px; border-left-color: #505055; border-left-style: solid; border-top-right-radius: 3px; border-bottom-right-radius: 3px; }
                QComboBox::down-arrow, QFontComboBox::down-arrow { image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxNiIgaGVpZ2h0PSIxNiIgdmlld0JveD0iMCAwIDE2IDE2Ij48cGF0aCBmaWxsPSIjRjFGNEYxIiBkPSJNMyA2bDUgNS4wMDFM MTQgNnoiLz48L3N2Zz4=); width: 12px; height: 12px; }
                QComboBox QAbstractItemView, QFontComboBox QAbstractItemView { background-color: #3C3C3F; border: 1px solid #505055; selection-background-color: #007ACC; }
                
                QSlider::groove:horizontal { border: 1px solid #505055; background: #3C3C3F; height: 4px; border-radius: 2px; }
                QSlider::handle:horizontal { background: #007ACC; border: 1px solid #009AFF; width: 14px; height: 14px; margin: -5px 0; border-radius: 7px; }
                
                QTabWidget::pane { border-top: 1px solid #505055; }
                QTabBar::tab { background-color: #38383C; border: 1px solid #505055; border-bottom: none; color: #A0A0A0; border-top-left-radius: 3px; border-top-right-radius: 3px; padding: 5px 10px; margin-right: 2px; }
                QTabBar::tab:selected { background: #2D2D30; border-bottom-color: #2D2D30; font-weight: bold; color: #F1F1F1; }
                QTabBar::tab:!selected:hover { background: #4A4A4F; color: #F1F1F1; }
                QCheckBox { spacing: 5px; }
                QCheckBox::indicator {
                    width: 13px;
                    height: 13px;
                    border-radius: 3px;
                    border: 1px solid #505055;
                    background-color: #3C3C3F;
                }
                QCheckBox::indicator:hover {
                    border-color: #007ACC;
                }
                QCheckBox::indicator:checked {
                    background-color: #007ACC;
                    border-color: #009AFF;
                    image: url(data:image/svg+xml;base64,PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIxMiIgaGVpZ2h0PSIxMiIgdmlld0JveD0iMCAwIDI0IDI0Ij48cGF0aCBmaWxsPSJ3aGl0ZSIgZD0iTTkgMTYuMTdMNC44MyAxMmwtMS40MiAxLjQxTDkgMTlsNy03Ljg5TDIxIDdsLTEuNDEtMS40MUw5IDE2LjE3eiIvPjwvc3ZnPg==);
                }
                QCheckBox::indicator:disabled {
                    background-color: #3A3A3D;
                    border-color: #4A4A4F;
                }
                QToolBar { background-color: #4F4F54; border: none; padding: 1px; }
                QToolBar QToolButton {
                    border: 1px solid transparent;
                    border-radius: 3px;
                    padding: 2px;
                }
                QToolBar QToolButton:hover {
                    background-color: #5A5A60;
                    border: 1px solid #009AFF;
                }
                QToolBar QToolButton:pressed {
                    background-color: #6A6A70;
                }
                QToolBar QToolButton:checked { 
                    background-color: #3D984E; 
                    border: 1px solid #5DBB6F;
                }
                QStatusBar { background-color: #007ACC; }
                QStatusBar QLabel { color: white; }
                
                #LiveViewLabel {
                    background-color: #5A5A60;
                    border: 1px solid #404040;
                }

                /* --- COMPLETE STYLES FOR TABLES, LISTS, AND SCROLLBARS --- */
                
                QAbstractItemView {
                    background-color: #3C3C3F;
                }

                QTableView, QTableWidget, QListView, QListWidget {
                    border: 1px solid #505055;
                    gridline-color: #505055;
                    color: #F1F1F1;
                    selection-background-color: #007ACC;
                    selection-color: white;
                }

                QHeaderView {
                    background-color: #38383C;
                }

                QHeaderView::section {
                    background-color: #4A4A4F;
                    color: #F1F1F1;
                    padding: 4px;
                    border-top: 0px;
                    border-left: 0px;
                    border-right: 1px solid #606065;
                    border-bottom: 2px solid #606065;
                    font-weight: bold;
                }
                
                QTableCornerButton::section {
                    background-color: #4A4A4F;
                    border-right: 1px solid #606065;
                    border-bottom: 2px solid #606065;
                }
                
                /* --- Dark Scrollbars --- */
                QScrollBar:vertical {
                    border: none;
                    background-color: #2D2D30;
                    width: 12px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background-color: #5A5A60;
                    min-height: 25px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical:hover {
                    background-color: #6A6A70;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }

                QScrollBar:horizontal {
                    border: none;
                    background-color: #2D2D30;
                    height: 12px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background-color: #5A5A60;
                    min-width: 25px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal:hover {
                    background-color: #6A6A70;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: none;
                }
            """

            def __init__(self):
                super().__init__()
                self.setAcceptDrops(True)
                # --- ADD THIS LINE AT THE START OF __init__ ---
                # --- (The rest of your __init__ method continues as before) ---
                self.current_theme='light'
                primary_screen_obj = QApplication.primaryScreen()
                if primary_screen_obj:
                    self.screen = primary_screen_obj.geometry() # Or .availableGeometry()
                else: # Fallback if no primary screen
                    screens = QApplication.screens()
                    if screens:
                        self.screen = screens[0].geometry()
                    else:
                        # Further fallback or error if no screens detected
                        print("Warning: No screens detected. Using default screen dimensions.")
                        self.screen = QRect(0, 0, 1024, 768) # Arbitrary default
                self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
                self.hist_fig, self.hist_ax, self.hist_canvas = None, None, None
                self.overlay_mode_active = False
                self.measurement_mode = None  # Tracks current tool: 'set_scale', 'measure_distance', 'measure_area'
                self.pixels_per_unit = None
                self.scale_unit = "pixels"
                self.measurement_points = []  # For drawing in-progress shapes
                self.finalized_measurement_shape = None  # Stores the final QPolygonF for persistent drawing
                self.last_measurement_result_text = ""
                self.adjustment_context = "Main Image" # 'Main Image', 'Overlay 1', or 'Overlay 2'
                self.image1_adjustments = {}
                self.image2_adjustments = {}
                self.image1_adjusted_preview = None
                self.image2_adjusted_preview = None
                self.is_in_dedicated_edit_mode = False
                self.main_view_widgets = {}
                self.image_before_padding = None
                self.image_contrasted=None
                self.image_before_contrast=None
                self.contrast_applied=False
                self.image_padded=False
                self.main_image_is_inverted = False
                self.selected_overlay_index = 0  # 0 for none, 1 for image1, 2 for image2
                self.overlay_interaction_mode = None  # None, 'moving', or 'resizing'
                self.resizing_overlay_corner_index = -1 # 0:TL, 1:TR, 2:BR, 3:BL
                self.drag_start_overlay_state = {}
                self._is_restoring_state = False
                # window_width = int(self.screen_width * 0.5)  # 60% of screen width
                # window_height = int(self.screen_height * 0.75)  # 95% of screen height
                self.preview_label_width_setting = 400  # A smaller base width for the minimum calculation
                self.preview_label_max_height_setting = 300 # A smaller base height for the minimum calculation
                self.label_size = self.preview_label_width_setting
                self.window_title="GEL BLOT ANALYZER v4.0"
                self.protein_sequence = ""
                self.base_protein_mw = 0.0
                self.avg_glycan_mass = 0.0
                self.last_predicted_mw = 0.0 
                self.num_oligomers_to_model = 1
                self.num_glycans_to_model = 0
                #self.num_glycosylation_sites = 0
                self.oligomer_products = [] # Combined list for all forms
                self.last_mw_prediction_model = None
                self.last_mw_prediction_min_max_pos = None
                # --- Initialize Status Bar Labels ---
                self.size_label = QLabel("Image Size: N/A")
                self.depth_label = QLabel("Bit Depth: N/A")
                self.location_label = QLabel("Source: N/A")
                self.mouse_coord_label = QLabel("X: --, Y: --")
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None  # 'quad' or 'rectangle'
                self.multi_lane_definitions = []  # List of dicts: {'type': 'quad'/'rect', 'points_label': [...], 'id': int}
                                                  # For rect: 'points_label' will be [QRectF_in_label_space]
                                                  # For quad: 'points_label' will be list of 4 QPointF_in_label_space
                self.current_multi_lane_points = []  # Temporary for defining current quad
                self.current_multi_lane_rect_start = None # For defining current rect
                self.latest_multi_lane_peak_areas = {} # Key: lane_id (int), Value: list of areas
                self.latest_multi_lane_peak_details = {} # Key: lane_id, Value: list of dicts with area & y_coord
                self.latest_multi_lane_calculated_quantities = {} # Key: lane_id (int), Value: list of quantities
                self.multi_lane_processing_finished = False # Flag
                self.moving_multi_lane_index = -1 
                self.moving_custom_item_info = None # Will be {'type': 'marker'/'shape', 'index': int}
                self.current_selection_mode = None # None, "select_for_move", "dragging_shape", "resizing_corner", "select_custom_item", ...
                self.resizing_corner_index = -1 # Index of the corner (0-3) being resized
                self.is_duplicating_shape = False 
                self.duplication_source_info = None 
                self.overlay_mode_active = False
                self.selected_overlay_index = 0  # 0 for none, 1 for image1, 2 for image2
                self.overlay_interaction_mode = None  # None, 'moving', or 'resizing'
                self.resizing_overlay_corner_index = -1 # 0:TL, 1:TR, 2:BR, 3:BL
                self.drag_start_overlay_state = {} # Stores rect, mouse pos, etc. at drag start
                

                # --- Add Labels to Status Bar ---
                statusbar = self.statusBar() # Get the status bar instance
                statusbar.addWidget(self.size_label)
                statusbar.addWidget(self.depth_label)
                statusbar.addWidget(self.mouse_coord_label) # Add new label here
                statusbar.addWidget(self.location_label, 1) # location_label remains stretched
                self.setWindowTitle(self.window_title)
                self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
                self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
                self.undo_stack = []
                self.redo_stack = []
                self.custom_shapes = []  # NEW: List to store lines/rectangles
                self.drawing_mode = None # NEW: None, 'line', 'rectangle'
                self.current_drawing_shape_preview = None # NEW: For live preview data
                self.quantities_peak_area_dict = {}
                self.latest_peak_details = [] # List of dicts: [{'area':val, 'y_coord_in_lane_image':val}, ...]
                self.latest_peak_areas = []
                self.latest_calculated_quantities = []
                self.image_path = None
                self.image = None
                self.image_master= None
                self.channel_mixer_data = {'r': 100, 'g': 100, 'b': 100, 'mono': False}
                self.unsharp_mask_data = {'amount': 0, 'radius': 1.0, 'threshold': 0}
                self.clahe_data = {'clip_limit': 1.0, 'tile_size': 8}
                self.predict_size=False
                self.warped_image=None
                self.transparency=1
                self.left_markers = []
                self.right_markers = []
                self.top_markers = []
                self.custom_markers=[]
                self.current_left_marker_index = 0
                self.current_right_marker_index = 0
                self.current_top_label_index = 0
                self.font_rotation=-45
                self.image_width=0
                self.image_height=0
                self.new_image_width=0
                self.new_image_height=0
                self.base_name="Image"
                self.image_path=""
                self.x_offset_s=0
                self.y_offset_s=0
                self.peak_area=[]
                self.auto_marker_side = None # To store 'left' or 'right' for auto mode
                self.auto_lane_quad_points = [] # Store points defined for auto lane
                self.is_modified=False
                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None # Temp storage for starting point in view coords
                self.crop_rectangle_coords = None # Stores final (x, y, w, h) in *image* coordinates
                self.crop_offset_x = 0
                self.crop_offset_y = 0
                
                self.copied_peak_regions_data = {
                    "regions": None,        # List of (start, end) tuples
                    "profile_length": None  # Length of the profile from which regions were copied
                }
                self.peak_dialog_settings = {
                    'rolling_ball_radius': 50,
                    'peak_height_factor': 0.1,
                    'peak_distance': 10,
                    'peak_prominence_factor': 0.00,
                    'valley_offset_pixels': 0,  # Added to persist settings
                    'band_estimation_method': "Mean",
                    'area_subtraction_method': "Rolling-valley",
                    'smoothing_sigma': 0.0, # Added to persist settings
                }
                
                self.persist_peak_settings_enabled = True # State of the checkbox
                
                
                # Variables to store bounding boxes and quantities
                self.bounding_boxes = []
                self.up_bounding_boxes = []
                self.standard_protein_areas=[]
                self.quantities = []
                self.protein_quantities = []
                self.measure_quantity_mode = False
                # Initialize self.marker_values to None initially
                self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"] # Default internal top_label
                self.marker_values = [] # Default internal marker_values (for L/R) - will be populated by preset

                self.custom_marker_name = "Custom"
                self.presets_data = {}
                self.custom_marker_name = "Custom"
                # Load the config file if exists
                
                self.marker_mode = None
                self.left_marker_shift = 0   # Additional shift for marker text
                self.right_marker_shift = 0   # Additional shift for marker tex
                self.top_marker_shift=0 
                self.left_marker_shift_added=0
                self.right_marker_shift_added=0
                self.top_marker_shift_added= 0
                self.left_slider_range=[-1000,1000]
                self.right_slider_range=[-1000,1000]
                self.top_slider_range=[-1000,1000]
                       
                
                self.top_padding = 0
                self.font_color = QColor(0, 0, 0)  # Default to black
                self.custom_marker_color = QColor(0, 0, 0)  # Default to black
                self.font_family = "Arial"  # Default font family
                self.font_size = 16  # Default font size
                self.image_array_backup= None
                self.run_predict_MW=False
                
                
                
                # Main container widget
                self.main_widget = QWidget()
                self.setCentralWidget(self.main_widget)
                self.table_window_instance = None
                self._deactivate_all_previews()

                # --- Create the two main widgets that will be rearranged ---
                # 1. The image viewer widget, containing the live view label
                self.live_view_label = LiveViewLabel(
                    font_type=QFont("Arial"),
                    font_size=int(24),
                    marker_color=QColor(0,0,0),
                    app_instance=self, # Pass the CombinedSDSApp instance
                    parent=self,
                )
                self.live_view_label.setObjectName("LiveViewLabel")
                self.live_view_label.setAlignment(Qt.AlignCenter) # This centers the image within the label
                
                self.live_view_label.mouseMovedInLabel.connect(self.update_mouse_coords_in_statusbar)

                self._create_actions()
                self.create_tool_bar()
                
                
                # 2. The tab widget for settings (created but not yet placed in a layout)
                self.tab_widget = QTabWidget()
                self.tab_widget.addTab(self.font_and_image_tab(), "Image and Contrast")
                self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
                self.tab_widget.addTab(self.create_markers_tab(), "Markers")
                self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
                self.tab_widget.addTab(self.analysis_tab(), "Analysis")

                
                #self.load_shortcut = QShortcut(QKeySequence.Open, self) # Ctrl+O / Cmd+O
                #self.load_shortcut.activated.connect(self.load_action.trigger) # Trigger the action

                #self.save_shortcut = QShortcut(QKeySequence.Save, self) # Ctrl+S / Cmd+S
                #self.save_shortcut.activated.connect(self.save_action.trigger) # Trigger the action

                #self.copy_shortcut = QShortcut(QKeySequence.Copy, self) # Ctrl+C / Cmd+C
                #self.copy_shortcut.activated.connect(self.copy_action.trigger) # Trigger the action

                #self.paste_shortcut = QShortcut(QKeySequence.Paste, self) # Ctrl+V / Cmd+V
                #self.paste_shortcut.activated.connect(self.paste_action.trigger) # Trigger the action

                #self.undo_shortcut = QShortcut(QKeySequence.Undo, self) # Ctrl+Z / Cmd+Z
                #self.undo_shortcut.activated.connect(self.undo_action_m) # Connect directly to method

                #self.redo_shortcut = QShortcut(QKeySequence.Redo, self) # Ctrl+Y / Cmd+Shift+Z
                #self.redo_shortcut.activated.connect(self.redo_action_m) # Connect directly to method

                #self.zoom_out_shortcut = QShortcut(QKeySequence.ZoomOut, self) # Ctrl+- / Cmd+-
                #self.zoom_out_shortcut.activated.connect(self.zoom_out_action.trigger) # Trigger the action

                # Zoom In - Keep the standard one
                #self.zoom_in_shortcut = QShortcut(QKeySequence.ZoomIn, self) # Attempts Ctrl++ / Cmd++
                #self.zoom_in_shortcut.activated.connect(self.zoom_in_action.trigger) # Trigger the action
                # ======================================================


                # === KEEP QShortcut definitions for NON-STANDARD actions ===
                # self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
                # self.save_svg_shortcut.activated.connect(self.save_svg_action.trigger)

                #self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
                #self.reset_shortcut.activated.connect(self.reset_action.trigger)

                self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
                self.predict_shortcut.activated.connect(self.predict_molecular_weight)

                self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
                self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)

                self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
                self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)

                self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
                self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)

                self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
                self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)

                self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
                self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
                self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
                self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
                self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
                self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

                self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
                self.grid_shortcut.activated.connect(
                    lambda: (
                        # Determine the target state (True if both are currently False, False otherwise)
                        target_state := not (self.show_grid_checkbox_x.isChecked() or self.show_grid_checkbox_y.isChecked()),
                        # Set both checkboxes to the target state
                        self.show_grid_checkbox_x.setChecked(target_state),
                        self.show_grid_checkbox_y.setChecked(target_state)
                    ) if hasattr(self, 'show_grid_checkbox_x') and hasattr(self, 'show_grid_checkbox_y') else None
                )
                
                self.grid_shortcut_x = QShortcut(QKeySequence("Ctrl+Shift+X"), self) # Use uppercase X for consistency
                self.grid_shortcut_x.activated.connect(
                    lambda: self.show_grid_checkbox_x.setChecked(not self.show_grid_checkbox_x.isChecked())
                    if hasattr(self, 'show_grid_checkbox_x') else None
                )

                # Shortcut for Y Grid Snapping
                self.grid_shortcut_y = QShortcut(QKeySequence("Ctrl+Shift+Y"), self) # Use uppercase Y
                self.grid_shortcut_y.activated.connect(
                    lambda: self.show_grid_checkbox_y.setChecked(not self.show_grid_checkbox_y.isChecked())
                    if hasattr(self, 'show_grid_checkbox_y') else None
                )

                self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
                self.guidelines_shortcut.activated.connect(
                    lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked())
                    if hasattr(self, 'show_guides_checkbox') else None
                )

                self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
                self.increase_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1)
                    if hasattr(self, 'grid_size_input') else None
                )
                self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
                self.decrease_grid_size_shortcut.activated.connect(
                    lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1)
                    if hasattr(self, 'grid_size_input') else None
                )

                self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
                self.invert_shortcut.activated.connect(self.invert_image)

                self.bw_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
                self.bw_shortcut.activated.connect(self.convert_to_black_and_white)

                # Tab movement shortcuts
                self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
                self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
                self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
                self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
                self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
                self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
                self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
                self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
                self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
                self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))

                
                self.viewer_position = "Top" # Default value before loading
                self.load_config()
                self._update_main_layout(self.viewer_position)

                # Update the checked state of the correct layout button in the toolbar
                if self.viewer_position == "Top": self.layout_top_action.setChecked(True)
                elif self.viewer_position == "Bottom": self.layout_bottom_action.setChecked(True)
                elif self.viewer_position == "Left": self.layout_left_action.setChecked(True)
                elif self.viewer_position == "Right": self.layout_right_action.setChecked(True)
                
                self._apply_initial_theme('dark')
                self._apply_initial_theme('light')
                self._update_toolbar_icons()
            
            def _apply_initial_theme(self, current_theme):
                """Applies the theme stylesheet based on the loaded preference."""
                app = QApplication.instance()
                if not app: return

                if current_theme == "dark":
                    app.setStyleSheet(self.dark_stylesheet)
                    if hasattr(self, 'theme_action'):
                        self.theme_action.setChecked(True)
                else:
                    app.setStyleSheet(self.light_stylesheet)
                    if hasattr(self, 'theme_action'):
                        self.theme_action.setChecked(False)

            def _on_table_window_closed(self):
                """Slot to clear the reference to the TableWindow when it closes."""
                # --- START OF THE FIX ---
                # This ensures the reference is cleared, preventing attempts to access a deleted object.
                if hasattr(self, 'table_window_instance'):
                    self.table_window_instance = None


            # --- ADD THIS NEW METHOD ---
            def _toggle_theme(self, checked):
                """Switches the application stylesheet between light and dark modes."""
                app = QApplication.instance()
                if not app: return

                if checked:
                    app.setStyleSheet(self.dark_stylesheet)
                    self.current_theme = "dark"
                else:
                    app.setStyleSheet(self.light_stylesheet)
                    self.current_theme = "light"
                
                # --- FIX: Call the helper method to regenerate icons with the new theme color ---
                self._update_toolbar_icons()
                
                self._update_levels_histogram()

            # --- ADD THIS NEW METHOD ---
            def save_app_settings(self):
                """Saves non-preset global settings like theme and layout to the config file."""
                config_filepath = os.path.join(
                    os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)),
                    self.CONFIG_PRESET_FILE_NAME
                )
                
                config_data = {}
                try:
                    if os.path.exists(config_filepath):
                        with open(config_filepath, "r", encoding='utf-8') as f:
                            config_data = json.load(f)
                    if not isinstance(config_data, dict): # Ensure it's a dict
                        config_data = {}
                except (json.JSONDecodeError, IOError):
                    config_data = {} # Start fresh on parsing error
                
                # Update the settings we want to save
                config_data["theme"] = self.current_theme
                config_data["viewer_position"] = self.viewer_position
                
                # Ensure presets data is not lost
                if "presets" not in config_data:
                    config_data["presets"] = self.presets_data

                try:
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump(config_data, f, indent=4)
                except Exception as e:
                    print(f"Warning: Could not save global app settings: {e}")

            def _update_histogram_markers_only(self):
                """A lightweight function that only updates the position of the level markers on the histogram canvas."""
                if not hasattr(self, 'hist_ax') or not self.hist_ax or not hasattr(self, 'hist_black_line') or not self.hist_black_line:
                    return # Do nothing if the histogram plot isn't ready

                try:
                    is_16bit = self.image_master.format() == QImage.Format_Grayscale16 if self.image_master else False
                    hist_range = (0, 65535) if is_16bit else (0, 255)
                    
                    black_point_slider_val = self.black_point_slider.value()
                    white_point_slider_val = self.white_point_slider.value()
                    slider_max = self.black_point_slider.maximum()

                    black_point_pos = (black_point_slider_val / slider_max) * hist_range[1]
                    white_point_pos = (white_point_slider_val / slider_max) * hist_range[1]
                    
                    # Update the x-position of the line and marker artists
                    self.hist_black_line.set_xdata([black_point_pos, black_point_pos])
                    self.hist_black_marker.set_xdata([black_point_pos])
                    self.hist_white_line.set_xdata([white_point_pos, white_point_pos])
                    self.hist_white_marker.set_xdata([white_point_pos])
                    
                    # Redraw just the canvas, which is very fast
                    self.hist_canvas.draw_idle()
                except (AttributeError, RuntimeError):
                    # Fail silently if artists have been deleted or are not ready
                    pass

            def _update_levels_histogram(self):
                """Draws or updates the intensity histogram for the currently selected adjustment context."""
                # Get the base image and settings for the current context
                source_image = None
                settings_dict = {}
            
                if self.adjustment_context == "Main Image":
                    source_image = self.image_master # Use the high-bit-depth master for accuracy
                    settings_dict = {
                        'is_inverted': self.main_image_is_inverted,
                        'channel_mixer': self.channel_mixer_data,
                        'unsharp_mask': self.unsharp_mask_data,
                        'clahe': self.clahe_data
                    }
                elif self.adjustment_context == "Overlay 1 (Base)":
                    source_image = getattr(self, 'image1_original', None)
                    settings_dict = self.image1_adjustments
                elif self.adjustment_context == "Overlay 2 (Overlay)":
                    source_image = getattr(self, 'image2_original', None)
                    settings_dict = self.image2_adjustments
            
                # Generate the pre-levels image for the histogram by replicating the adjustment pipeline
                source_image_for_hist = None
                if source_image and not source_image.isNull():
                    temp_image = source_image.copy()
                    if settings_dict.get('is_inverted', False):
                        temp_image.invertPixels()
            
                    np_img_full = self.qimage_to_numpy(temp_image)
                    if np_img_full is not None:
                        np_content = np_img_full
                        is_16bit = np_content.dtype == np.uint16
                        max_val = 65535.0 if is_16bit else 255.0
            
                        # Apply Channel Mixer
                        cm_settings = settings_dict.get('channel_mixer', self._get_default_adjustments()['channel_mixer'])
                        if np_content.ndim == 3:
                            is_mono = cm_settings.get('mono', False)
                            r, g, b = cm_settings.get('r',100)/100.0, cm_settings.get('g',100)/100.0, cm_settings.get('b',100)/100.0
                            np_float = np_content.astype(np.float32)
                            if is_mono:
                                gray = cv2.transform(np_float[...,:3], np.array([[b],[g],[r]]).T)
                                np_content = np.clip(gray, 0, max_val).astype(np_content.dtype)
                            else:
                                np_float[..., 0] *= b; np_float[..., 1] *= g; np_float[..., 2] *= r
                                np_content = np.clip(np_float, 0, max_val).astype(np_content.dtype)
            
                        # --- CLAHE FIX FOR 16-BIT COLOR ---
                        clahe_settings = settings_dict.get('clahe', self._get_default_adjustments()['clahe'])
                        if clahe_settings.get('clip_limit', 1.0) > 1.0:
                            tile_size = clahe_settings.get('tile_size', 8)
                            effective_clip = clahe_settings['clip_limit'] * 256.0 if is_16bit else clahe_settings['clip_limit']
                            clahe = cv2.createCLAHE(clipLimit=effective_clip, tileGridSize=(tile_size, tile_size))
                            
                            if np_content.ndim == 2:
                                np_content = clahe.apply(np_content)
                            elif np_content.ndim == 3:
                                bgr = np_content[...,:3]
                                alpha = np_content[..., 3] if np_content.shape[2] == 4 else None

                                if is_16bit:
                                    # Convert 16-bit Int -> 32-bit Float -> LAB -> CLAHE -> Back
                                    bgr_float = bgr.astype(np.float32) / 65535.0
                                    lab_float = cv2.cvtColor(bgr_float, cv2.COLOR_BGR2LAB)
                                    l_channel = lab_float[..., 0]
                                    l_uint16 = (l_channel / 100.0 * 65535.0).astype(np.uint16)
                                    l_clahe = clahe.apply(l_uint16)
                                    lab_float[..., 0] = (l_clahe.astype(np.float32) / 65535.0 * 100.0)
                                    bgr_float_out = cv2.cvtColor(lab_float, cv2.COLOR_LAB2BGR)
                                    bgr_out = np.clip(bgr_float_out * 65535.0, 0, 65535).astype(np.uint16)
                                    
                                    if alpha is not None: np_content = np.dstack((bgr_out, alpha))
                                    else: np_content = bgr_out
                                else:
                                    # 8-bit path
                                    lab = cv2.cvtColor(bgr, cv2.COLOR_BGR2LAB)
                                    lab[..., 0] = clahe.apply(lab[..., 0])
                                    bgr_out = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
                                    if alpha is not None: np_content = np.dstack((bgr_out, alpha))
                                    else: np_content = bgr_out
                        
                        # Apply Unsharp Mask
                        usm_settings = settings_dict.get('unsharp_mask', self._get_default_adjustments()['unsharp_mask'])
                        if usm_settings.get('amount', 0) > 0:
                            amount = usm_settings['amount'] / 100.0
                            sigma = max(0.1, usm_settings['radius'])
                            blurred = cv2.GaussianBlur(np_content, (0, 0), sigma)
                            np_content = cv2.addWeighted(np_content, 1.0 + amount, blurred, -amount, 0)
                        
                        source_image_for_hist = self.numpy_to_qimage(np_content)
                    else:
                        source_image_for_hist = temp_image
                
                # ... (The rest of the function for plotting the histogram remains unchanged) ...
                # Ensure the plot canvas exists and there's an image to analyze.
                if not self.hist_ax or not source_image_for_hist or source_image_for_hist.isNull():
                    if self.hist_ax:
                        self.hist_ax.clear()
                        # Set background color based on theme
                        bg_color = '#38383C' if self.current_theme == 'dark' else '#F0F2F5' # Match main window bg
                        self.hist_fig.patch.set_facecolor(bg_color)
                        self.hist_ax.patch.set_facecolor(bg_color)
                        text_color = '#A0A0A0' if self.current_theme == 'dark' else 'gray'
                        
                        self.hist_ax.text(0.5, 0.5, 'No Image Selected', transform=self.hist_ax.transAxes, ha='center', va='center', fontsize=9, color=text_color)
                        self.hist_ax.set_xticks([]); self.hist_ax.set_yticks([])
                        self.hist_fig.tight_layout(pad=0.2)
                        self.hist_canvas.draw_idle()
                    return

                try:
                    np_img = self.qimage_to_numpy(source_image_for_hist)
                    if np_img is None: return

                    if np_img.ndim == 3:
                        gray_np = cv2.cvtColor(np_img[:,:,:3], cv2.COLOR_BGR2GRAY) if np_img.shape[2] >= 3 else np_img[:,:,0]
                    else:
                        gray_np = np_img
                    
                    is_16bit = gray_np.dtype == np.uint16
                    bins = 255
                    hist_range = (0, 65535) if is_16bit else (0, 255)
                    
                    # Calculate Histogram
                    hist, bin_edges = np.histogram(gray_np.ravel(), bins=bins, range=hist_range)
                    # Add 1 to avoid log(0)
                    log_hist = np.log1p(hist)
                    bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

                    self.hist_ax.clear()

                    # Set colors based on the current theme
                    if self.current_theme == 'dark':
                        bg_color = '#38383C'
                        plot_bg_color = '#2D2D30'
                        line_color = '#64B5F6' # Soft Blue
                        fill_color = '#64B5F6'
                        fill_alpha = 0.3
                        text_color = '#E0E0E0'
                        grid_color = '#505050'
                        marker_color = '#FF5252' # Red Accent
                    else:
                        bg_color = '#F0F2F5'
                        plot_bg_color = '#FFFFFF'
                        line_color = '#1976D2' # Darker Blue
                        fill_color = '#2196F3'
                        fill_alpha = 0.3
                        text_color = '#333333'
                        grid_color = '#E0E0E0'
                        marker_color = '#D32F2F' # Dark Red

                    self.hist_fig.patch.set_facecolor(bg_color)
                    self.hist_ax.patch.set_facecolor(plot_bg_color)
                    
                    # Draw Data
                    self.hist_ax.plot(bin_centers, log_hist, color=line_color, linewidth=1.0)
                    self.hist_ax.fill_between(bin_centers, 0, log_hist, color=fill_color, alpha=fill_alpha)
                    
                    # --- AXIS FORMATTING ---
                    self.hist_ax.grid(True, linestyle=':', alpha=0.6, color=grid_color)
                    
                    # X-Axis Labels
                    x_ticks = np.linspace(hist_range[0], hist_range[1], 5)
                    self.hist_ax.set_xticks(x_ticks)
                    
                    if is_16bit:
                        # Format 16-bit ticks (e.g., 0, 16k, 32k, 48k, 64k)
                        labels = [f"{int(x/1000)}k" if x > 0 else "0" for x in x_ticks]
                        self.hist_ax.set_xticklabels(labels, fontsize=7, color=text_color)
                    else:
                        self.hist_ax.set_xticklabels([f"{int(x)}" for x in x_ticks], fontsize=7, color=text_color)

                    # Y-Axis (Log Density)
                    self.hist_ax.set_yticks([]) # Hide Y ticks as log values aren't intuitive for users here
                    self.hist_ax.set_ylabel("Log Density", fontsize=7, color=text_color)
                    
                    # Title
                    title_str = 'Intensity Histogram (16-bit)' if is_16bit else 'Intensity Histogram (8-bit)'
                    self.hist_ax.set_title(title_str, fontsize=8, pad=3, color=text_color, fontweight='bold')
                    
                    # Spines
                    for spine in self.hist_ax.spines.values():
                        spine.set_color(grid_color)
                        spine.set_linewidth(0.5)

                    self.hist_ax.set_xlim(hist_range[0], hist_range[1])
                    self.hist_ax.set_ylim(bottom=0)
                    
                    # --- LEVEL MARKERS ---
                    black_point_slider_val = self.black_point_slider.value()
                    white_point_slider_val = self.white_point_slider.value()
                    slider_max = self.black_point_slider.maximum()
                    
                    # Map slider (0-65535) to current hist range
                    scale = hist_range[1] / 65535.0
                    black_point_pos = black_point_slider_val * scale
                    white_point_pos = white_point_slider_val * scale
                    
                    # Draw Lines
                    self.hist_black_line = self.hist_ax.axvline(black_point_pos, color=text_color, lw=1.0, linestyle='--')
                    self.hist_white_line = self.hist_ax.axvline(white_point_pos, color=text_color, lw=1.0, linestyle='--')
                    
                    # Draw Triangles at bottom
                    self.hist_black_marker = self.hist_ax.plot(black_point_pos, 0, marker='^', color='black', markeredgecolor='white', markersize=8, clip_on=False, zorder=10)[0]
                    self.hist_white_marker = self.hist_ax.plot(white_point_pos, 0, marker='^', color='white', markeredgecolor='black', markersize=8, clip_on=False, zorder=10)[0]
                    
                    # Adjust margins to fit labels
                    self.hist_fig.subplots_adjust(left=0.12, right=0.95, top=0.88, bottom=0.25)
                    self.hist_canvas.draw_idle()
                except Exception as e:
                    print(f"Error updating levels histogram: {e}")
                    # traceback.print_exc()
                
            def update_mouse_coords_in_statusbar(self, label_pos: QPointF, image_pos: QPointF = None):
                if image_pos is not None:
                    # Display image coordinates if available (mouse is over the image)
                    self.mouse_coord_label.setText(f"Img X: {image_pos.x():.0f}, Y: {image_pos.y():.0f}")
                elif label_pos is not None:
                    # Fallback to label coordinates if not over image or no image loaded
                    self.mouse_coord_label.setText(f"View X: {label_pos.x():.0f}, Y: {label_pos.y():.0f}")
                else:
                    # Default if no valid position data (e.g., mouse exited label)
                    self.mouse_coord_label.setText("X: --, Y: --")
                    
            def _find_image_content_borders(self):
                """
                Scans the current self.image to find the left and right borders of the
                non-padding content. Assumes padding is transparent.
                Returns a tuple (left_border_x, right_border_x), or (None, None) on failure.
                """
                if not self.image or self.image.isNull():
                    return None, None

                # Use a copy to avoid modifying the original
                qimg_to_scan = self.image_master.copy()
                # Ensure the image has an alpha channel for reliable detection
                if not qimg_to_scan.hasAlphaChannel():
                    qimg_to_scan = qimg_to_scan.convertToFormat(QImage.Format_ARGB32)

                np_img = self.qimage_to_numpy(qimg_to_scan)
                if np_img is None or np_img.ndim < 3 or np_img.shape[2] < 4:
                    print("Warning: Could not get image with alpha channel for border detection.")
                    return None, None

                height, width, _ = np_img.shape
                # The alpha channel is the 4th channel (index 3)
                alpha_channel = np_img[:, :, 3]

                # Scan from left to right to find the first non-transparent column
                left_border_x = None
                for x in range(width):
                    # np.any is fast for checking if any value in the column is non-zero
                    if np.any(alpha_channel[:, x] > 0):
                        left_border_x = x
                        break

                # Scan from right to left to find the first non-transparent column
                right_border_x = None
                for x in range(width - 1, -1, -1):
                    if np.any(alpha_channel[:, x] > 0):
                        right_border_x = x
                        break
                
                if left_border_x is None or right_border_x is None:
                    # This case happens if the image is fully transparent
                    return 0, width-1 # Fallback to image edges

                return left_border_x, right_border_x
                    
            def _reset_live_view_label_custom_handlers(self):
                """Helper to reset all custom mouse interaction hooks on LiveViewLabel."""
                if hasattr(self.live_view_label, '_custom_left_click_handler_from_app'):
                    self.live_view_label._custom_left_click_handler_from_app = None
                if hasattr(self.live_view_label, '_custom_mouseMoveEvent_from_app'):
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                if hasattr(self.live_view_label, '_custom_mouseReleaseEvent_from_app'):
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                # Ensure mouse tracking is on if it was turned off by a mode
                self.live_view_label.setMouseTracking(True)
                self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            
            def _create_actions(self):
                """Create QAction objects for menus and toolbars."""
                # --- Actions are now created WITHOUT icons initially ---
                # --- File Actions ---
                self.load_action = QAction("&Load Image...", self)
                self.save_action = QAction("&Save with Config", self)
                self.reset_action = QAction("&Reset Image", self)
                self.exit_action = QAction("E&xit", self)

                # --- Edit Actions ---
                self.undo_action = QAction("&Undo", self)
                self.redo_action = QAction("&Redo", self)
                self.copy_action = QAction("&Copy Image", self)
                self.paste_action = QAction("&Paste Image", self)

                # --- View Actions ---
                self.zoom_in_action = QAction("Zoom &In", self)
                self.zoom_out_action = QAction("Zoom &Out", self)
                self.pan_left_action = QAction("Pan Left", self)
                self.pan_right_action = QAction("Pan Right", self)
                self.pan_up_action = QAction("Pan Up", self)
                self.pan_down_action = QAction("Pan Down", self)
                self.auto_lane_action = QAction("&Automatic Lane Markers", self)

                # --- Layout Actions ---
                self.layout_top_action = QAction("Viewer on Top", self)
                self.layout_bottom_action = QAction("Viewer on Bottom", self)
                self.layout_left_action = QAction("Viewer on Left", self)
                self.layout_right_action = QAction("Viewer on Right", self)
                layout_actions = [self.layout_top_action, self.layout_bottom_action, self.layout_left_action, self.layout_right_action]
                self.layout_action_group = QActionGroup(self)
                self.layout_action_group.setExclusive(True)
                for action in layout_actions:
                    action.setCheckable(True)
                    self.layout_action_group.addAction(action)
                self.layout_top_action.setChecked(True)

                # --- Theme Action ---
                self.theme_action = QAction("Toggle Dark/Light Mode", self)
                self.theme_action.setCheckable(True)
                self.theme_action.setShortcut(QKeySequence("Ctrl+Shift+D"))

                # --- Other Tool Actions ---
                self.info_action = QAction("&Info/GitHub", self)
                self.draw_bounding_box_action = QAction("Draw &Bounding Box", self)
                self.draw_line_action = QAction("Draw &a line", self)
                self.copy_custom_items_action = QAction("Copy Custom Markers/Shapes", self)
                self.paste_custom_items_action = QAction("Paste Custom Markers/Shapes", self)

                # --- FIX: Call the new helper method to set all icons ---
                self._update_toolbar_icons()

                # --- The rest of the method (tooltips, shortcuts, connections) is unchanged ---
                self.layout_top_action.setToolTip("Place the image viewer above the settings tabs.")
                self.layout_bottom_action.setToolTip("Place the image viewer below the settings tabs.")
                self.layout_left_action.setToolTip("Place the image viewer to the left of the settings tabs.")
                self.layout_right_action.setToolTip("Place the image viewer to the right of the settings tabs.")
                self.auto_lane_action.setToolTip("Automatically detect and place lane markers based on a defined region.")
                self.theme_action.setToolTip("Switch between light and dark themes (Ctrl+Shift+D or Cmd+Shift+D).")
                self.info_action.setToolTip("Open Project GitHub Page")
                self.load_action.setToolTip("Load an image file (Ctrl+O)")
                self.save_action.setToolTip("Save image and configuration (Ctrl+S)")
                self.reset_action.setToolTip("Reset image and all annotations (Ctrl+R)")
                self.exit_action.setToolTip("Exit the application")
                self.undo_action.setToolTip("Undo last action (Default Shortcut: OS dependent")
                self.redo_action.setToolTip("Redo last undone action (Default Shortcut: OS dependent)")
                self.copy_action.setToolTip("Copy rendered image to clipboard (Ctrl+C)")
                self.paste_action.setToolTip("Paste image from clipboard (Ctrl+V)")
                self.zoom_in_action.setToolTip("Increase zoom level (Ctrl+= or mouse scroll bar)")
                self.zoom_out_action.setToolTip("Decrease zoom level (Ctrl+- or mouse scroll bar)). Auto resets the zoom when reaches zero.")
                self.pan_left_action.setToolTip("Pan the view left (when zoomed) (Arrow key left or mouse right click)")
                self.pan_right_action.setToolTip("Pan the view right (when zoomed) (Arrow key right or mouse right click")
                self.pan_up_action.setToolTip("Pan the view up (when zoomed) (Arrow key up or mouse right click)")
                self.pan_down_action.setToolTip("Pan the view down (when zoomed) (Arrow key down or mouse right click")
                self.draw_bounding_box_action.setToolTip("Draw a bounding rectangle on the image. Use the marker tab, custom marker options for color and size")
                self.draw_line_action.setToolTip("Draw a line on the image. Use the marker tab, custom marker options for color and size")
                self.copy_custom_items_action.setToolTip("Copy all custom markers and shapes to the clipboard.")
                self.paste_custom_items_action.setToolTip("Paste custom markers and shapes from the clipboard.\nItems will be added to existing ones.")
                
                self.load_action.setShortcut(QKeySequence.Open)
                self.save_action.setShortcut(QKeySequence.Save)
                self.copy_action.setShortcut(QKeySequence.Copy)
                self.paste_action.setShortcut(QKeySequence.Paste)
                self.undo_action.setShortcut(QKeySequence.Undo)
                self.redo_action.setShortcut(QKeySequence.Redo)
                self.zoom_in_action.setShortcut(QKeySequence("Ctrl+="))
                self.zoom_out_action.setShortcut(QKeySequence.ZoomOut)
                self.reset_action.setShortcut(QKeySequence("Ctrl+R"))

                self.load_action.triggered.connect(self.load_image)
                self.save_action.triggered.connect(self.save_image)
                self.reset_action.triggered.connect(self.reset_image)
                self.exit_action.triggered.connect(self.close)
                self.undo_action.triggered.connect(self.undo_action_m)
                self.redo_action.triggered.connect(self.redo_action_m)
                self.copy_action.triggered.connect(self.copy_to_clipboard)
                self.paste_action.triggered.connect(self.paste_image)
                self.zoom_in_action.triggered.connect(self.zoom_in)
                self.zoom_out_action.triggered.connect(self.zoom_out)
                pan_step = 30
                self.pan_right_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_right_action.triggered.connect(self.update_live_view)
                self.pan_left_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_left_action.triggered.connect(self.update_live_view)
                self.pan_down_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_down_action.triggered.connect(self.update_live_view)
                self.pan_up_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
                self.pan_up_action.triggered.connect(self.update_live_view)
                self.copy_custom_items_action.triggered.connect(self.copy_custom_items)
                self.paste_custom_items_action.triggered.connect(self.paste_custom_items)
                self.layout_top_action.triggered.connect(lambda: self._transition_layout_change("Top"))
                self.layout_bottom_action.triggered.connect(lambda: self._transition_layout_change("Bottom"))
                self.layout_left_action.triggered.connect(lambda: self._transition_layout_change("Left"))
                self.layout_right_action.triggered.connect(lambda: self._transition_layout_change("Right"))
                self.auto_lane_action.triggered.connect(self.start_auto_lane_marker)
                self.draw_line_action.triggered.connect(self.enable_line_drawing_mode)
                self.draw_bounding_box_action.triggered.connect(self.enable_rectangle_drawing_mode)
                self.info_action.triggered.connect(self.open_github)
                self.theme_action.triggered.connect(self._toggle_theme)

                self.pan_left_action.setEnabled(False)
                self.pan_right_action.setEnabled(False)
                self.pan_up_action.setEnabled(False)
                self.pan_down_action.setEnabled(False)

            def _update_main_layout(self, position: str):
                """
                Rebuilds the main window's layout.
                - Viewer is FIXED size (550x350) to prevent issues.
                - Controls area dynamically sizes to fill the rest of the screen 
                  to minimize scrollbar usage.
                - Left/Right modes now maximize vertical height usage.
                """
                # --- START OF MODIFICATION ---
                # Fixed dimensions for the viewer as requested
                VIEWER_FIXED_WIDTH = 550
                VIEWER_FIXED_HEIGHT = 350
                
                # Get available screen geometry
                screen_geo = QGuiApplication.primaryScreen().availableGeometry()
                screen_w = screen_geo.width()
                screen_h = screen_geo.height()

                new_main_widget = QWidget()
                new_layout = None

                # Create a ScrollArea for the Tab Widget (Controls) if needed
                if not hasattr(self, 'controls_scroll_area'):
                    self.controls_scroll_area = QScrollArea()
                    
                self.controls_scroll_area.setWidget(self.tab_widget)
                self.controls_scroll_area.setWidgetResizable(True)
                self.controls_scroll_area.setFrameShape(QFrame.NoFrame)

                # --- 1. Enforce Fixed Viewer Size ---
                self.live_view_label.setFixedSize(VIEWER_FIXED_WIDTH, VIEWER_FIXED_HEIGHT)
                self.live_view_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

                # --- 2. Unconstrain Tab Widget ---
                self.tab_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
                self.tab_widget.setMinimumSize(0, 0)
                self.tab_widget.setMaximumSize(16777215, 16777215)

                if position in ["Top", "Bottom"]:
                    new_layout = QVBoxLayout(new_main_widget)
                    
                    # Vertical Layout: Calculate remaining height for controls
                    # Screen Height - Fixed Viewer Height - Window Margins (~120px)
                    available_height = max(400, screen_h - VIEWER_FIXED_HEIGHT - 150)
                    
                    self.controls_scroll_area.setMinimumHeight(available_height)
                    self.controls_scroll_area.setMinimumWidth(0)
                    
                    if position == "Top":
                        new_layout.addWidget(self.live_view_label, 0, Qt.AlignCenter)
                        new_layout.addWidget(self.create_separator())
                        new_layout.addWidget(self.controls_scroll_area, 1)
                    else:  # Bottom
                        new_layout.addWidget(self.controls_scroll_area, 1)
                        new_layout.addWidget(self.create_separator())
                        new_layout.addWidget(self.live_view_label, 0, Qt.AlignCenter)
                
                elif position in ["Left", "Right"]:
                    separator = QFrame()
                    separator.setFrameShape(QFrame.VLine)
                    separator.setFrameShadow(QFrame.Sunken)
                    new_layout = QHBoxLayout(new_main_widget)
                    
                    # Horizontal Layout Width: Screen Width - Fixed Viewer Width - Margins (~60px)
                    available_width = max(450, screen_w - VIEWER_FIXED_WIDTH - 60)
                    
                    # Horizontal Layout Height: Use almost full screen height to minimize vertical scrolling
                    # Screen Height - Window Titlebar/Taskbar buffer (~100px)
                    available_height = max(600, screen_h - VIEWER_FIXED_HEIGHT - 150)
                    
                    self.controls_scroll_area.setMinimumWidth(available_width)
                    self.controls_scroll_area.setMinimumHeight(available_height) # Maximize height usage
                    
                    if position == "Left":
                        new_layout.addWidget(self.live_view_label, 0, Qt.AlignCenter)
                        new_layout.addWidget(separator)
                        new_layout.addWidget(self.controls_scroll_area, 1) # Controls take remaining width
                    else:  # Right
                        new_layout.addWidget(self.controls_scroll_area, 1) # Controls take remaining width
                        new_layout.addWidget(separator)
                        new_layout.addWidget(self.live_view_label, 0, Qt.AlignCenter)
                # --- END OF MODIFICATION ---

                if new_layout:
                    old_central_widget = self.centralWidget()
                    self.setCentralWidget(new_main_widget)
                    self.main_widget = new_main_widget
                    if old_central_widget:
                        old_central_widget.deleteLater()

                
            def _transition_layout_change(self, position: str):
                """
                Rebuilds the main window's layout, handling the case where we might be in editor mode.
                """
                self.viewer_position = position
                
                # --- Save the new layout preference to the config file ---
                if getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                else:
                    try:
                        application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError:
                        application_path = os.getcwd()
                config_filepath = os.path.join(application_path, self.CONFIG_PRESET_FILE_NAME)
                
                config_data = {}
                try: 
                    if os.path.exists(config_filepath):
                        with open(config_filepath, "r", encoding='utf-8') as f:
                            config_data = json.load(f)
                    if not isinstance(config_data, dict):
                        config_data = {}
                except (json.JSONDecodeError, IOError):
                    config_data = {}
                
                config_data["viewer_position"] = self.viewer_position
                if "presets" not in config_data:
                    config_data["presets"] = self.presets_data
                
                try: 
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump(config_data, f, indent=4)
                except Exception as e:
                    print(f"Warning: Could not save global config: {e}")

                # --- FIX: Only change the layout if NOT in dedicated editor mode ---
                if not self.is_in_dedicated_edit_mode:
                    # Use a "blink" transition to avoid visual artifacts
                    self.live_view_label.hide()
                    self.tab_widget.hide()
                    
                    # Rebuild the main layout
                    self._update_main_layout(position)
                    
                    # Process events to apply the layout change
                    QApplication.processEvents()
                    
                    # Adjust window size to fit the new layout
                    self.adjustSize()
                    
                    # Show the widgets again in their new positions
                    self.live_view_label.show()
                    self.tab_widget.show()
                    
                    # Redraw the content
                    self.update_live_view()

                

            def _constrain_point_orthogonally(self, start_point: QPointF, current_point: QPointF) -> QPointF:
                """Constrains the current point to be on a straight horizontal or vertical line from the start point."""
                delta_x = abs(current_point.x() - start_point.x())
                delta_y = abs(current_point.y() - start_point.y())

                if delta_x > delta_y:
                    # Horizontal movement is dominant, lock the Y-coordinate
                    return QPointF(current_point.x(), start_point.y())
                else:
                    # Vertical movement is dominant, lock the X-coordinate
                    return QPointF(start_point.x(), current_point.y()) 
                
            def copy_custom_items(self):
                """Copies custom markers and shapes to the clipboard as JSON."""
                if not hasattr(self, "custom_markers"): self.custom_markers = []
                if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                if not self.custom_markers and not self.custom_shapes:
                    QMessageBox.information(self, "Nothing to Copy", "There are no custom markers or shapes to copy.")
                    return

                # Serialize markers and shapes
                # custom_markers are stored as lists, custom_shapes as dicts
                # For markers, we need to convert QColor to string for JSON
                serializable_markers = []
                for marker_data in self.custom_markers:
                    try:
                        # marker_data is a list: [x, y, text, qcolor, font_family, font_size, is_bold, is_italic]
                        m_copy = list(marker_data)
                        if isinstance(m_copy[3], QColor):
                            m_copy[3] = m_copy[3].name() # Convert QColor to hex string
                        serializable_markers.append(m_copy)
                    except IndexError:
                        print(f"Warning: Skipping marker during copy due to unexpected format: {marker_data}")
                
                items_to_copy = {
                    "custom_markers": serializable_markers,
                    "custom_shapes": self.custom_shapes # Already list of dicts
                }

                try:
                    json_data = json.dumps(items_to_copy)
                    mime_data = QMimeData()
                    mime_data.setData(self.MIME_TYPE_CUSTOM_ITEMS, json_data.encode('utf-8'))
                    
                    # Also set text for basic paste compatibility (e.g., into a text editor)
                    mime_data.setText(f"Gel Blot Analyzer Custom Items (JSON):\n{json_data}")
                    
                    QApplication.clipboard().setMimeData(mime_data)
                    QMessageBox.information(self, "Items Copied", 
                                            f"{len(self.custom_markers)} markers and {len(self.custom_shapes)} shapes copied to clipboard.")
                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Could not copy custom items: {e}")
                    traceback.print_exc()

            def paste_custom_items(self):
                """Pastes custom markers and shapes from the clipboard."""
                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()

                if not mime_data.hasFormat(self.MIME_TYPE_CUSTOM_ITEMS):
                    QMessageBox.information(self, "Paste Error", 
                                            "Clipboard does not contain Gel Blot Analyzer custom items in the expected format.")
                    return

                try:
                    json_bytes = mime_data.data(self.MIME_TYPE_CUSTOM_ITEMS)
                    json_data_str = json_bytes.data().decode('utf-8') # Convert QByteArray to bytes then to string
                    pasted_items = json.loads(json_data_str)

                    pasted_markers_data = pasted_items.get("custom_markers", [])
                    pasted_shapes_data = pasted_items.get("custom_shapes", [])

                    if not pasted_markers_data and not pasted_shapes_data:
                        QMessageBox.information(self, "Paste", "No custom markers or shapes found in clipboard data.")
                        return

                    self.save_state() # Save current state before adding new items

                    if not hasattr(self, "custom_markers"): self.custom_markers = []
                    if not hasattr(self, "custom_shapes"): self.custom_shapes = []

                    num_pasted_markers = 0
                    for marker_data_serial in pasted_markers_data:
                        try:
                            # marker_data_serial is a list: [x, y, text, color_str, font_family, font_size, is_bold, is_italic]
                            m_copy = list(marker_data_serial)
                            m_copy[3] = QColor(m_copy[3]) # Convert color string back to QColor
                            if not m_copy[3].isValid(): m_copy[3] = QColor(Qt.black) # Fallback
                            self.custom_markers.append(m_copy)
                            num_pasted_markers += 1
                        except (IndexError, TypeError, ValueError) as e:
                            print(f"Warning: Skipping pasted marker due to format error: {marker_data_serial}, {e}")
                    
                    num_pasted_shapes = 0
                    for shape_data in pasted_shapes_data:
                        # Basic validation (can be extended)
                        if isinstance(shape_data, dict) and 'type' in shape_data:
                            self.custom_shapes.append(shape_data)
                            num_pasted_shapes +=1
                        else:
                             print(f"Warning: Skipping pasted shape due to format error: {shape_data}")


                    self.is_modified = True
                    self.update_live_view()
                    QMessageBox.information(self, "Items Pasted",
                                            f"{num_pasted_markers} markers and {num_pasted_shapes} shapes pasted.")

                except Exception as e:
                    QMessageBox.critical(self, "Paste Error", f"Could not paste custom items: {e}")
                    traceback.print_exc()
                
            def _update_preview_label_size(self):
                """
                Updates the fixed size of the live_view_label, respecting max width/height
                and maintaining aspect ratio. Prioritizes fixing height, but adjusts if
                width constraint is violated.
                """
                # --- Define Defaults and Minimums ---
                default_max_width = 600  # Fallback if width setting is missing/invalid
                default_max_height = 500 # Fallback if height setting is missing/invalid
                min_dim = 50             # Minimum allowed dimension for the label

                # --- Determine Maximum Constraints (with validation) ---
                max_w = default_max_width
                if hasattr(self, 'preview_label_width_setting'):
                    try:
                        setting_width = int(self.preview_label_width_setting)
                        if setting_width > 0:
                            max_w = setting_width
                        else:
                            print("Warning: preview_label_width_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_width_setting is invalid, using default.")
                else:
                    print("Warning: preview_label_width_setting attribute not found, using default.")
                max_w = max(min_dim, max_w) # Apply minimum constraint

                max_h = default_max_height
                if hasattr(self, 'preview_label_max_height_setting'):
                    try:
                        setting_height = int(self.preview_label_max_height_setting)
                        if setting_height > 0:
                            max_h = setting_height
                        else:
                            print("Warning: preview_label_max_height_setting is not positive, using default.")
                    except (TypeError, ValueError):
                        print("Warning: preview_label_max_height_setting is invalid, using default.")
                max_h = max(min_dim, max_h) # Apply minimum constraint


                # --- Initialize Final Dimensions ---
                final_w = max_w
                final_h = max_h

                # --- Calculate Dimensions Based on Image ---
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()

                    if w > 0 and h > 0:
                        img_ratio = w / h

                        # Attempt 1: Calculate width based on fixed max height
                        calc_w_based_on_h = max_h * img_ratio

                        if calc_w_based_on_h <= max_w:
                            # Width is within limits, height is fixed
                            final_w = calc_w_based_on_h
                            final_h = max_h
                        else:
                            # Calculated width exceeds max width, so fix width and calculate height
                            final_w = max_w
                            final_h = max_w / img_ratio

                        # Ensure calculated dimensions are not below minimum
                        final_w = max(min_dim, int(final_w))
                        final_h = max(min_dim, int(final_h))

                    else:
                        # Handle invalid image dimensions (0 width or height)
                        final_w = max_w # Already set above
                        final_h = max_h # Already set above

                else:
                    # No image loaded - Use max constraints as default size
                    final_w = max_w # Already set above
                    final_h = max_h # Already set above
                    # self.live_view_label.clear() # Optionally clear the label

                # --- Set the Calculated Fixed Size ---
                self.live_view_label.setMinimumSize(final_w, final_h)
                
            def _update_status_bar(self):
                """Updates the status bar labels with current image information."""
                if self.image and not self.image.isNull():
                    w = self.image.width()
                    h = self.image.height()
                    self.size_label.setText(f"Image Size: {w}x{h}")

                    # Determine bit depth/format string
                    img_format = self.image.format()
                    depth_str = "Unknown"
                    if img_format == QImage.Format_Grayscale8:
                        depth_str = "8-bit Grayscale"
                    elif img_format == QImage.Format_Grayscale16:
                        depth_str = "16-bit Grayscale"
                    elif img_format == QImage.Format_RGB888:
                         depth_str = "24-bit RGB"
                    elif img_format in (QImage.Format_RGB32, QImage.Format_ARGB32,
                                        QImage.Format_ARGB32_Premultiplied, QImage.Format_RGBA8888):
                         depth_str = "32-bit (A)RGB"
                    elif img_format == QImage.Format_Indexed8:
                         depth_str = "8-bit Indexed"
                    elif img_format == QImage.Format_Mono:
                         depth_str = "1-bit Mono"
                    else:
                        # Fallback to depth() if format is unusual
                        depth_val = self.image.depth()
                        depth_str = f"{depth_val}-bit Depth"
                    self.depth_label.setText(f"Bit Depth: {depth_str}")

                    # Update location
                    location = "N/A"
                    if hasattr(self, 'image_path') and self.image_path:
                        # Show only filename if it's a path, otherwise show the source info
                        if os.path.exists(self.image_path):
                            location = os.path.basename(self.image_path)
                        else: # Likely "Clipboard..." or similar
                            location = self.image_path
                    else:
                        location = "N/A"
                    self.location_label.setText(f"Source: {location}")

                else:
                    # No image loaded
                    self.size_label.setText("Image Size: N/A")
                    self.depth_label.setText("Bit Depth: N/A")
                    self.location_label.setText("Source: N/A")
                
            def prompt_save_if_needed(self):
                if not self.is_modified:
                    return True # No changes, proceed
                """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
                reply = QMessageBox.question(self, 'Unsaved Changes',
                                             "Do you want to save the file?",
                                             QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                             QMessageBox.Save) # Default button is Save

                if reply == QMessageBox.Save:
                    return self.save_image() # Returns True if saved, False if save cancelled
                elif reply == QMessageBox.Discard:
                    return True # Proceed without saving
                else: # Cancelled
                    return False # Abort the current action (close/load)

            def closeEvent(self, event):
                """Overrides the default close event to prompt for saving and save app settings."""
                if self.prompt_save_if_needed():
                    self.save_app_settings() # Save theme and layout choice on exit
                    event.accept() # Proceed with closing
                else:
                    event.ignore() # Abort closing
                    
            def enable_line_drawing_mode(self):
                self.save_state()
                self.drawing_mode = 'line'
                self.live_view_label.mode = 'draw_shape' 
                self.current_drawing_shape_preview = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                
                # Use custom hooks instead of direct overwrite
                self.live_view_label._custom_left_click_handler_from_app = self.start_shape_draw
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_shape_draw
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_shape_draw

            def enable_rectangle_drawing_mode(self):
                self.save_state()
                self.drawing_mode = 'rectangle'
                self.live_view_label.mode = 'draw_shape'
                self.current_drawing_shape_preview = None
                self.live_view_label.setCursor(Qt.CrossCursor)

                self.live_view_label._custom_left_click_handler_from_app = self.start_shape_draw
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_shape_draw
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_shape_draw

            def cancel_drawing_mode(self):
                """Resets drawing mode and cursor."""
                self.drawing_mode = None
                self.live_view_label.mode = None # App-level mode for LiveViewLabel's paintEvent logic
                self.current_drawing_shape_preview = None
                self._reset_live_view_label_custom_handlers() # Use helper
                self.update_live_view()
                
            def start_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle']:
                    start_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    self.current_drawing_shape_preview = {'start': snapped_start_point, 'end': snapped_start_point}
                    self.update_live_view()

            def update_shape_draw(self, event):
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
                    start_point = self.current_drawing_shape_preview['start']

                    # --- MODIFIED: Constrain movement if Shift is pressed ---
                    if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                        if self.drawing_mode == 'line':
                            # --- Angle snapping logic for lines (0, 45, 90 degrees) ---
                            delta = snapped_end_point - start_point
                            angle_rad = np.arctan2(delta.y(), delta.x())
                            angle_deg = np.degrees(angle_rad)
                            
                            # Snap angle to the nearest 45 degrees
                            snapped_angle_deg = round(angle_deg / 45.0) * 45.0
                            snapped_angle_rad = np.radians(snapped_angle_deg)
                            
                            # Recalculate the end point based on the original length and snapped angle
                            length = np.sqrt(delta.x()**2 + delta.y()**2)
                            snapped_end_point = QPointF(
                                start_point.x() + length * np.cos(snapped_angle_rad),
                                start_point.y() + length * np.sin(snapped_angle_rad)
                            )
                        elif self.drawing_mode == 'rectangle':
                            # --- Square constraint logic for rectangles ---
                            delta_x = snapped_end_point.x() - start_point.x()
                            delta_y = snapped_end_point.y() - start_point.y()
                            max_delta = max(abs(delta_x), abs(delta_y))
                            
                            new_x = start_point.x() + (max_delta if delta_x > 0 else -max_delta)
                            new_y = start_point.y() + (max_delta if delta_y > 0 else -max_delta)
                            snapped_end_point = QPointF(new_x, new_y)
                    
                    self.current_drawing_shape_preview['end'] = snapped_end_point
                    self.update_live_view()

            def finalize_shape_draw(self, event):
                """Finalizes the shape and adds it to the custom_shapes list."""
                if self.drawing_mode in ['line', 'rectangle'] and self.current_drawing_shape_preview:
                    # start_point_label_space is already snapped from start_shape_draw
                    start_point_label_space = self.current_drawing_shape_preview['start']
                    
                    # --- THE FIX ---
                    # Use the final, constrained end point from the preview data, not the raw mouse event.
                    # The preview has already handled grid snapping and Shift key constraints.
                    end_point_snapped_label_space = self.current_drawing_shape_preview['end']
                    # --- END FIX ---
    
                    # ... (Get current style settings: color, thickness) ...
                    color = self.custom_marker_color
                    thickness = self.custom_font_size_spinbox.value()


                    displayed_width = float(self.live_view_label.width())
                    displayed_height = float(self.live_view_label.height())
                    current_image_width = float(self.image.width()) if self.image and self.image.width() > 0 else 1.0
                    current_image_height = float(self.image.height()) if self.image and self.image.height() > 0 else 1.0
                    if current_image_width <= 0 or current_image_height <= 0: # Safety check
                        scale_img_to_label = 1.0
                    else:
                        scale_x_fit = displayed_width / current_image_width
                        scale_y_fit = displayed_height / current_image_height
                        scale_img_to_label = min(scale_x_fit, scale_y_fit)
                    display_img_w = current_image_width * scale_img_to_label
                    display_img_h = current_image_height * scale_img_to_label
                    label_offset_x = (displayed_width - display_img_w) / 2.0
                    label_offset_y = (displayed_height - display_img_h) / 2.0
                    def label_to_image_coords(label_point):
                        if scale_img_to_label <= 1e-9: return (0.0, 0.0)
                        relative_x_in_display = label_point.x() - label_offset_x
                        relative_y_in_display = label_point.y() - label_offset_y
                        img_x = relative_x_in_display / scale_img_to_label
                        img_y = relative_y_in_display / scale_img_to_label
                        return (img_x, img_y)
                    start_img_coords = label_to_image_coords(start_point_label_space) # Use snapped start
                    end_img_coords = label_to_image_coords(end_point_snapped_label_space) # Use snapped end
    
                    shape_data = {
                        'type': self.drawing_mode,
                        'color': color.name(),
                        'thickness': thickness
                    }
                    valid_shape = False
                    if self.drawing_mode == 'line':
                        if abs(start_img_coords[0] - end_img_coords[0]) > 0.5 or abs(start_img_coords[1] - end_img_coords[1]) > 0.5:
                            shape_data['start'] = start_img_coords
                            shape_data['end'] = end_img_coords
                            valid_shape = True
                    elif self.drawing_mode == 'rectangle':
                        x_img = min(start_img_coords[0], end_img_coords[0])
                        y_img = min(start_img_coords[1], end_img_coords[1])
                        w_img = abs(end_img_coords[0] - start_img_coords[0])
                        h_img = abs(end_img_coords[1] - start_img_coords[1])
                        if w_img > 0.5 and h_img > 0.5:
                            shape_data['rect'] = (x_img, y_img, w_img, h_img)
                            valid_shape = True
                    if valid_shape:
                        self.custom_shapes.append(shape_data)
                        self.save_state()
                        self.is_modified = True
                    self.cancel_drawing_mode()
                else:
                    self.cancel_drawing_mode()
                    
            def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
                """Converts QImage to NumPy array, preserving format and handling row padding."""
                if qimage.isNull(): return None

                img_format = qimage.format()
                height = qimage.height()
                width = qimage.width()
                
                # PySide6/Qt6 uses sizeInBytes()
                try: expected_total_bytes = qimage.sizeInBytes()
                except AttributeError: expected_total_bytes = qimage.byteCount()

                ptr = qimage.constBits()
                if not ptr: return None
                
                try: ptr.setsize(expected_total_bytes)
                except AttributeError: pass
                
                buffer_data = bytes(ptr)

                # --- 16-bit Grayscale ---
                if img_format == QImage.Format_Grayscale16:
                    bytes_per_line = qimage.bytesPerLine()
                    expected_bytes_per_line = width * 2
                    if bytes_per_line == expected_bytes_per_line:
                        arr = np.frombuffer(buffer_data, dtype=np.uint16).reshape(height, width)
                        return arr.copy()
                    else:
                        arr = np.zeros((height, width), dtype=np.uint16)
                        for y in range(height):
                            start = y * bytes_per_line
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            arr[y] = np.frombuffer(row_data, dtype=np.uint16)
                        return arr

                # --- FIX: 16-bit RGBA (RGBA64) Handling ---
                elif img_format in [QImage.Format_RGBA64, QImage.Format_RGBX64]:
                    bytes_per_pixel = 8 # 4 channels * 2 bytes
                    bytes_per_line = qimage.bytesPerLine()
                    expected_bytes_per_line = width * bytes_per_pixel
                    
                    if bytes_per_line == expected_bytes_per_line:
                        # Direct reshape if no padding
                        arr = np.frombuffer(buffer_data, dtype=np.uint16).reshape(height, width, 4)
                        return arr.copy()
                    else:
                        # Handle row padding
                        arr = np.zeros((height, width, 4), dtype=np.uint16)
                        for y in range(height):
                            start = y * bytes_per_line
                            # Read exactly the bytes for the width, ignore padding
                            row_data = buffer_data[start : start + expected_bytes_per_line]
                            arr[y] = np.frombuffer(row_data, dtype=np.uint16).reshape(width, 4)
                        return arr

                # --- 8-bit Grayscale ---
                elif img_format == QImage.Format_Grayscale8:
                    bytes_per_pixel = 1
                    bytes_per_line = qimage.bytesPerLine()
                    if bytes_per_line == width:
                        return np.frombuffer(buffer_data, dtype=np.uint8).reshape(height, width).copy()
                    else:
                        arr = np.zeros((height, width), dtype=np.uint8)
                        for y in range(height):
                            s = y * bytes_per_line
                            arr[y] = np.frombuffer(buffer_data[s : s + width], dtype=np.uint8)
                        return arr

                # --- 8-bit Color (ARGB32/RGB32/RGB888) ---
                elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied, QImage.Format_RGB32):
                    bytes_per_pixel = 4
                    bytes_per_line = qimage.bytesPerLine()
                    expected_bytes_per_line = width * 4
                    if bytes_per_line == expected_bytes_per_line:
                        return np.frombuffer(buffer_data, dtype=np.uint8).reshape(height, width, 4).copy()
                    else:
                        arr = np.zeros((height, width, 4), dtype=np.uint8)
                        for y in range(height):
                            s = y * bytes_per_line
                            arr[y] = np.frombuffer(buffer_data[s : s + expected_bytes_per_line], dtype=np.uint8).reshape(width, 4)
                        return arr
                
                elif img_format == QImage.Format_RGB888:
                    bytes_per_pixel = 3
                    bytes_per_line = qimage.bytesPerLine()
                    expected_bytes_per_line = width * 3
                    if bytes_per_line == expected_bytes_per_line:
                        return np.frombuffer(buffer_data, dtype=np.uint8).reshape(height, width, 3).copy()
                    else:
                        arr = np.zeros((height, width, 3), dtype=np.uint8)
                        for y in range(height):
                            s = y * bytes_per_line
                            arr[y] = np.frombuffer(buffer_data[s : s + expected_bytes_per_line], dtype=np.uint8).reshape(width, 3)
                        return arr

                # Fallback conversion
                else:
                    try:
                        qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                        return self.qimage_to_numpy(qimage_conv)
                    except: return None

            def numpy_to_qimage(self, array: np.ndarray) -> QImage:
                """Converts NumPy array to QImage, selecting appropriate format."""
                if array is None: return QImage()
                if not isinstance(array, np.ndarray): return QImage()
            
                try:
                    if array.ndim == 2: # Grayscale
                        height, width = array.shape
                        if array.dtype == np.uint16:
                            bytes_per_line = width * 2
                            contiguous_array = np.ascontiguousarray(array)
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                            return qimg.copy()
                        elif array.dtype == np.uint8:
                            bytes_per_line = width * 1
                            contiguous_array = np.ascontiguousarray(array)
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                            return qimg.copy()
                        elif np.issubdtype(array.dtype, np.floating):
                             img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                             bytes_per_line = width * 1
                             contiguous_array = np.ascontiguousarray(img_norm)
                             qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                             return qimg.copy()
            
                    elif array.ndim == 3: # Color
                        height, width, channels = array.shape
                        
                        if channels == 3 and array.dtype == np.uint16:
                             # FIX: Support 16-bit Color via RGBA64
                             # QImage doesn't have a simple RGB48 format, so we use RGBA64.
                             # We must add an alpha channel (opaque 65535).
                             alpha = np.full((height, width, 1), 65535, dtype=np.uint16)
                             rgba_16 = np.concatenate((array, alpha), axis=2)
                             contiguous_array = np.ascontiguousarray(rgba_16)
                             bytes_per_line = width * 8 # 4 channels * 2 bytes
                             qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_RGBA64)
                             return qimg.copy()

                        elif channels == 4 and array.dtype == np.uint16:
                             # FIX: Support 16-bit RGBA directly
                             contiguous_array = np.ascontiguousarray(array)
                             bytes_per_line = width * 8 # 4 channels * 2 bytes
                             qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_RGBA64)
                             return qimg.copy()

                        elif channels == 3 and array.dtype == np.uint8:
                            contiguous_array = np.ascontiguousarray(array)
                            bytes_per_line = width * 3
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_RGB888)
                            return qimg.copy()

                        elif channels == 4 and array.dtype == np.uint8:
                            contiguous_array = np.ascontiguousarray(array)
                            bytes_per_line = width * 4
                            qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                            return qimg.copy()
                    
                    return QImage()
            
                except Exception as e:
                    print(f"Error in numpy_to_qimage: {e}")
                    traceback.print_exc()
                    return QImage()

            def get_image_format(self, image=None):
                """Helper to safely get the format of self.image or a provided image."""
                img_to_check = image if image is not None else self.image
                if img_to_check and not img_to_check.isNull():
                    return img_to_check.format()
                return None

            def get_compatible_grayscale_format(self, image=None):
                """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
                current_format = self.get_image_format(image)
                if current_format == QImage.Format_Grayscale16:
                    return QImage.Format_Grayscale16
                else: # Treat 8-bit grayscale or color as needing 8-bit target
                    return QImage.Format_Grayscale8
            # --- END: Helper Functions ---
                
            def quadrilateral_to_rect(self, image, quad_points):
                """
                Warps the quadrilateral region defined by quad_points in the input image
                to a rectangular image using perspective transformation.
                Crucially, preserves the original bit depth (e.g., uint16) of the input image.
                """
                if not image or image.isNull():
                    QMessageBox.warning(self, "Warp Error", "Invalid input image provided for warping.")
                    return None
                if len(quad_points) != 4:
                     QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
                     return None
                if cv2 is None:
                     QMessageBox.critical(self, "Dependency Error", "OpenCV (cv2) is required for quadrilateral warping but is not installed.")
                     return None
             
                try:
                    img_array = self.qimage_to_numpy(image)
                    if img_array is None: raise ValueError("NumPy Conversion returned None")
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
                    return None
             
                # --- START FIX: Correctly transform points from Label space to Image space ---
                # The incoming quad_points are already in the un-zoomed, un-panned "label space".
                # We need to map them to the native image pixel coordinates.
                src_points_img = self._map_label_points_to_image_points(quad_points)
                if src_points_img is None:
                    QMessageBox.warning(self, "Warp Error", "Failed to map label coordinates to image coordinates.")
                    return None
                
                src_np = np.array([[p.x(), p.y()] for p in src_points_img], dtype=np.float32)
                # --- END FIX ---
             
                width_a = np.linalg.norm(src_np[0] - src_np[1])
                width_b = np.linalg.norm(src_np[2] - src_np[3])
                height_a = np.linalg.norm(src_np[0] - src_np[3])
                height_b = np.linalg.norm(src_np[1] - src_np[2])
                max_width = int(max(width_a, width_b))
                max_height = int(max(height_a, height_b))
             
                if max_width <= 0 or max_height <= 0:
                     QMessageBox.warning(self, "Warp Error", f"Invalid destination rectangle size ({max_width}x{max_height}).")
                     return None
             
                dst_np = np.array([
                    [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
                ], dtype=np.float32)
             
                try:
                    matrix = cv2.getPerspectiveTransform(src_np, dst_np)
                    if img_array.ndim == 3:
                         border_val = (0, 0, 0, 0) if img_array.shape[2] == 4 else (0, 0, 0)
                    else:
                         border_val = 0
             
                    warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                                       flags=cv2.INTER_LINEAR,
                                                       borderMode=cv2.BORDER_CONSTANT,
                                                       borderValue=border_val)
                except Exception as e:
                     QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
                     traceback.print_exc()
                     return None
             
                try:
                    warped_qimage = self.numpy_to_qimage(warped_array)
                    if warped_qimage.isNull(): raise ValueError("numpy_to_qimage conversion failed.")
                    return warped_qimage
                except Exception as e:
                    QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array back to QImage: {e}")
                    return None
            
                
            def create_menu_bar(self):
                menubar = self.menuBar()

                # --- File Menu ---
                file_menu = menubar.addMenu("&File")
                file_menu.addAction(self.load_action)
                file_menu.addAction(self.save_action)
                # file_menu.addAction(self.save_svg_action)
                file_menu.addSeparator()
                file_menu.addAction(self.reset_action)
                file_menu.addSeparator()
                file_menu.addAction(self.exit_action)

                # --- Edit Menu ---
                edit_menu = menubar.addMenu("&Edit")
                edit_menu.addAction(self.undo_action)
                edit_menu.addAction(self.redo_action)
                edit_menu.addSeparator()
                edit_menu.addAction(self.copy_action)
                edit_menu.addAction(self.paste_action)

                # --- View Menu ---
                view_menu = menubar.addMenu("&View")
                view_menu.addAction(self.zoom_in_action)
                view_menu.addAction(self.zoom_out_action)
                
                tools_menu = menubar.addMenu("&Tools")
                tools_menu.addAction(self.auto_lane_action)

                # --- About Menu ---
                about_menu = menubar.addMenu("&About")
                # Add "GitHub" action directly here or create it in _create_actions
                github_action = QAction("&GitHub", self)
                github_action.setToolTip("Open the project's GitHub page")
                github_action.triggered.connect(self.open_github)
                about_menu.addAction(github_action)
                
                # self.statusBar().showMessage("Ready")
            def open_github(self):
                # Open the GitHub link in the default web browser
                QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Gel-Blot-Analyzer"))
                
            def _update_toolbar_icons(self):
                """Regenerates all toolbar icons based on the current theme's palette."""
                icon_size = QSize(30, 30) 
                # --- FIX: Get the text color from the CURRENT palette ---
                text_color = self.palette().color(QPalette.Text) # Use QPalette.Text for better contrast guarantee

                # Create and set icons for each action
                self.load_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "1"))
                self.save_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "="))
                self.reset_action.setIcon(create_text_icon("Wingdings 3", icon_size, text_color, "Q"))
                self.exit_action.setIcon(create_text_icon("Wingdings 2", icon_size, text_color, "V"))
                self.undo_action.setIcon(create_text_icon("Wingdings 3", icon_size, text_color, "O"))
                self.redo_action.setIcon(create_text_icon("Wingdings 3", icon_size, text_color, "N"))
                self.copy_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "4"))
                self.paste_action.setIcon(create_text_icon("Wingdings 2", icon_size, text_color, "2"))
                self.zoom_in_action.setIcon(create_text_icon("Arial", icon_size, text_color, "+"))
                self.zoom_out_action.setIcon(create_text_icon("Arial", icon_size, text_color, "-"))
                self.pan_left_action.setIcon(create_text_icon("Arial", icon_size, text_color, "←"))
                self.pan_right_action.setIcon(create_text_icon("Arial", icon_size, text_color, "→"))
                self.pan_up_action.setIcon(create_text_icon("Arial", icon_size, text_color, "↑"))
                self.pan_down_action.setIcon(create_text_icon("Arial", icon_size, text_color, "↓"))
                self.auto_lane_action.setIcon(create_text_icon("Arial", icon_size, text_color, "A"))
                self.layout_top_action.setIcon(create_text_icon("Webdings", icon_size, text_color, "5"))
                self.layout_bottom_action.setIcon(create_text_icon("Webdings", icon_size, text_color, "6"))
                self.layout_left_action.setIcon(create_text_icon("Webdings", icon_size, text_color, "7"))
                self.layout_right_action.setIcon(create_text_icon("Webdings", icon_size, text_color, "8"))
                self.theme_action.setIcon(create_text_icon("Webdings", icon_size, text_color, "N"))
                self.info_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "'"))
                self.draw_bounding_box_action.setIcon(create_text_icon("Wingdings 2", icon_size, text_color, "0"))
                self.draw_line_action.setIcon(create_text_icon("Arial", icon_size, text_color, "__"))
                self.copy_custom_items_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "B"))
                self.paste_custom_items_action.setIcon(create_text_icon("Wingdings", icon_size, text_color, "A"))
            
            def create_tool_bar(self):
                """Create the main application toolbar."""
                self.tool_bar = QToolBar("Main Toolbar")
                self.tool_bar.setIconSize(QSize(25, 25))
                self.tool_bar.setMovable(False)
                self.tool_bar.setToolButtonStyle(Qt.ToolButtonIconOnly)

                # Add actions to the toolbar
                self.tool_bar.addAction(self.load_action)
                self.tool_bar.addAction(self.paste_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.save_action)
                self.tool_bar.addAction(self.copy_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.undo_action)
                self.tool_bar.addAction(self.redo_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.zoom_in_action)
                self.tool_bar.addAction(self.zoom_out_action)
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.pan_left_action)
                self.tool_bar.addAction(self.pan_up_action)
                self.tool_bar.addAction(self.pan_down_action)
                self.tool_bar.addAction(self.pan_right_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.auto_lane_action)
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.draw_line_action) 
                self.tool_bar.addAction(self.draw_bounding_box_action) 
                self.tool_bar.addSeparator() 
                self.tool_bar.addAction(self.copy_custom_items_action)
                self.tool_bar.addAction(self.paste_custom_items_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.reset_action)
                self.tool_bar.addSeparator()
                self.tool_bar.addAction(self.layout_top_action)
                self.tool_bar.addAction(self.layout_bottom_action)
                self.tool_bar.addAction(self.layout_left_action)
                self.tool_bar.addAction(self.layout_right_action)
                self.tool_bar.addSeparator()
                # --- ADD THIS ACTION TO THE TOOLBAR ---
                self.tool_bar.addAction(self.theme_action)
                # --- END ADD ---
                self.tool_bar.addAction(self.info_action)

                self.addToolBar(Qt.TopToolBarArea, self.tool_bar)
                self.tool_bar.setContextMenuPolicy(Qt.PreventContextMenu)

                
            def start_auto_lane_marker(self):
                self._reset_live_view_label_custom_handlers()
                """Initiates the automatic lane marker placement process using a single, dynamically created dialog."""
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    return

                # --- START: Create a simple dialog on-the-fly without a new class ---
                dialog = QDialog(self)
                dialog.setWindowTitle("Automatic Lane Setup")
                dialog.setMinimumWidth(350)
                layout = QVBoxLayout(dialog)

                # --- Side Selection ---
                side_group = QGroupBox("Select Marker Side")
                side_layout = QHBoxLayout(side_group)
                radio_left = QRadioButton("Left")
                radio_right = QRadioButton("Right")
                radio_left.setChecked(True)
                side_layout.addWidget(radio_left)
                side_layout.addWidget(radio_right)
                side_layout.addStretch()
                layout.addWidget(side_group)

                # --- Region Type Selection ---
                region_group = QGroupBox("Select Lane Region Type")
                region_layout = QVBoxLayout(region_group)
                radio_rect = QRadioButton("Rectangle (for straight lanes)")
                radio_quad = QRadioButton("Quadrilateral (for skewed lanes)")
                radio_rect.setChecked(True)
                region_layout.addWidget(radio_rect)
                region_layout.addWidget(radio_quad)
                layout.addWidget(region_group)

                # --- OK/Cancel Buttons ---
                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(dialog.accept)
                button_box.rejected.connect(dialog.reject)
                layout.addWidget(button_box)

                # --- Execute the dialog and get results ---
                if dialog.exec() == QDialog.Accepted:
                    side = "left" if radio_left.isChecked() else "right"
                    region_type = "rectangle" if radio_rect.isChecked() else "quadrilateral"
                    self.auto_marker_side = side
                else:
                    return # User cancelled
                # --- END: On-the-fly dialog logic ---

                # Clear all preview states (this part remains the same)
                self.live_view_label.quad_points = []
                self.live_view_label.current_preview_points = []
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)

                if region_type == "rectangle":
                    self.live_view_label.mode = 'auto_lane_rect'
                    self.live_view_label._custom_left_click_handler_from_app = self.start_rectangle
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.update_rectangle_preview
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_rectangle_for_auto_lane
                elif region_type == "quadrilateral":
                    self.live_view_label.mode = 'auto_lane_quad'
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_auto_lane_quad_click

                self.update_live_view()

            def handle_auto_lane_quad_click(self, event):
                if self.live_view_label.mode != 'auto_lane_quad':
                    return
    
                if event.button() == Qt.LeftButton:
                    point_label_space_transformed = self.live_view_label.transform_point(event.position())
                    snapped_point_label_space = self.snap_point_to_grid(point_label_space_transformed) # Snap it
                    
                    # --- FIX: Append to the correct preview list ---
                    self.live_view_label.current_preview_points.append(snapped_point_label_space)
                    # --- END FIX ---
                    self.update_live_view()
    
                    if len(self.live_view_label.current_preview_points) == 4:
                        self.finalize_quad_for_auto_lane(self.live_view_label.current_preview_points)

            def finalize_quad_for_auto_lane(self, quad_points_label_space):
                """
                Finalizes the quadrilateral using points from "label space",
                warps the corresponding region from self.image, and proceeds to processing.
                quad_points_label_space: List of 4 QPointF in "label space".
                """
                if len(quad_points_label_space) != 4:
                    QMessageBox.warning(self, "Error", "Quadrilateral definition incomplete for auto lane.")
                    self.live_view_label.mode = None # Reset mode
                    self.live_view_label.quad_points = []
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self._reset_live_view_label_custom_handlers() # Reset after finalization
                    self.live_view_label.mode = None # Reset LiveViewLabel's internal mode flag
                    self.update_live_view()
                    return
                
                # --- START FIX ---
                # Use the fully adjusted master image as the source for warping.
                adjusted_master = self._get_fully_adjusted_image_for_analysis()
                if not adjusted_master or adjusted_master.isNull():
                    QMessageBox.warning(self, "Error", "Could not get adjusted image for analysis.")
                    return
                warped_qimage_region = self.quadrilateral_to_rect(adjusted_master, quad_points_label_space)
                # --- END FIX ---

                if warped_qimage_region and not warped_qimage_region.isNull():
                    quad_points_image_space_for_marker_placement = []
                    try:
                        # Use self.image for dimension mapping as it's the basis for the view
                        native_img_w = float(self.image.width())
                        native_img_h = float(self.image.height())
                        label_w_widget = float(self.live_view_label.width())
                        label_h_widget = float(self.live_view_label.height())
                        if native_img_w <= 0 or native_img_h <= 0 or label_w_widget <= 0 or label_h_widget <= 0:
                            raise ValueError("Invalid image_to_warp or label dimensions.")
                        scale_native_to_label = min(label_w_widget / native_img_w, label_h_widget / native_img_h)
                        displayed_w_in_label = native_img_w * scale_native_to_label
                        displayed_h_in_label = native_img_h * scale_native_to_label
                        offset_x_centering = (label_w_widget - displayed_w_in_label) / 2.0
                        offset_y_centering = (label_h_widget - displayed_h_in_label) / 2.0
                        for p_label in quad_points_label_space:
                            x_rel_display = p_label.x() - offset_x_centering
                            y_rel_display = p_label.y() - offset_y_centering
                            if scale_native_to_label < 1e-9: raise ValueError("Scale factor too small.")
                            img_x = x_rel_display / scale_native_to_label
                            img_y = y_rel_display / scale_native_to_label
                            quad_points_image_space_for_marker_placement.append(QPointF(img_x, img_y))
                        
                        # print(f"DEBUG AutoLaneQuad (finalize): Mapped image space points for marker placement: {[(p.x(), p.y()) for p in quad_points_image_space_for_marker_placement]}")

                        self.process_auto_lane_region(warped_qimage_region, 
                                                      quad_points_image_space_for_marker_placement, # Pass image-space points
                                                      is_quad_warp=True)
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to map quad points for marker placement: {e}")
                        traceback.print_exc()
                else:
                    QMessageBox.warning(self, "Error", "Failed to warp quadrilateral region for auto lane.")

                # Reset state after processing or error
                self.live_view_label.mode = None
                self.live_view_label.quad_points = [] # Clear points from label
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self._reset_live_view_label_custom_handlers() # Important to reset
                self.update_live_view()


            def finalize_rectangle_for_auto_lane(self, event):
                # --- START OF THE FIX ---
                # Check the new preview list, not the old 'rectangle_start' variable.
                if self.live_view_label.mode != 'auto_lane_rect' or not self.live_view_label.current_preview_points:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event)
                    return
                # --- END OF THE FIX ---
    
                if event.button() == Qt.LeftButton:
                    try:
                        # --- START OF THE FIX ---
                        # Get the final start and end points from the correct preview list.
                        preview_points = self.live_view_label.current_preview_points
                        if len(preview_points) != 2:
                            raise ValueError("Rectangle definition incomplete.")
                        start_point_label = preview_points[0]
                        end_point_label = preview_points[1] # The final end point is already in the list
                        
                        rect_in_label_space = QRectF(start_point_label, end_point_label).normalized()
                        # --- END OF THE FIX ---

                        rect_coords_img = self._map_label_rect_to_image_rect(rect_in_label_space)
                        if rect_coords_img is None:
                            raise ValueError("Coordinate mapping from label to image failed.")
                        
                        # ... (rest of the function is correct) ...
    
                        adjusted_master = self._get_fully_adjusted_image_for_analysis()
                        if not adjusted_master or adjusted_master.isNull():
                            raise ValueError("Could not get adjusted image for analysis.")
                        
                        extracted_qimage_region = adjusted_master.copy(*rect_coords_img) 
                        
                        if extracted_qimage_region.isNull():
                            raise ValueError("QImage copy failed for rectangle.")
    
                        self.process_auto_lane_region(extracted_qimage_region, rect_coords_img, is_quad_warp=False)
    
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to finalize rectangle for auto lane: {e}")
                        traceback.print_exc()
                    finally:
                        self._reset_live_view_label_custom_handlers()
                        self.live_view_label.mode = None
                        self.live_view_label.setCursor(Qt.ArrowCursor)
                        # --- FIX: Clear all preview-related variables ---
                        self.live_view_label.bounding_box_preview = None
                        self.live_view_label.rectangle_start = None
                        self.live_view_label.current_preview_points = []
                        # --- END FIX ---
                        self.update_live_view()

            def process_auto_lane_region(self, qimage_region, original_region_definition, is_quad_warp):
                """
                Processes the extracted/warped region, opens tuning dialog, and places markers.
                qimage_region: The QImage (rectangular) to be analyzed.
                original_region_definition: For rect, tuple (x,y,w,h) in original image space.
                                            For quad, list of 4 QPointF in original image space.
                is_quad_warp: Boolean indicating if qimage_region came from a warped quadrilateral.
                """
                if qimage_region.isNull():
                    QMessageBox.warning(self, "Error", "No valid image region provided for auto lane processing.")
                    return

                dialog_image_pil = self.convert_qimage_to_grayscale_pil(qimage_region)
                if not dialog_image_pil:
                    QMessageBox.warning(self, "Error", "Failed to convert extracted lane to grayscale for analysis.")
                    return

                # Store height AND WIDTH of the image passed to the dialog for later mapping
                dialog_image_height = dialog_image_pil.height
                dialog_image_width = dialog_image_pil.width # <<< --- ADDED

                dialog = AutoLaneTuneDialog(
                    pil_image_data=dialog_image_pil,
                    initial_settings=self.peak_dialog_settings,
                    parent=self,
                    is_from_quad_warp=is_quad_warp
                )

                if dialog.exec() == QDialog.Accepted:
                    detected_peak_coords_dialog = dialog.get_detected_peaks()
                    final_settings = dialog.get_final_settings()

                    if self.persist_peak_settings_enabled:
                        self.peak_dialog_settings.update(final_settings)

                    if detected_peak_coords_dialog is not None and len(detected_peak_coords_dialog) > 0:
                        self.place_markers_from_dialog(
                            original_region_definition,
                            detected_peak_coords_dialog,
                            self.auto_marker_side,
                            dialog_image_height,
                            dialog_image_width, # <<< --- PASS WIDTH
                            is_quad_warp
                        )
                    else:
                        QMessageBox.information(self, "Auto Lane", "No peaks were detected with the current settings.")
                else:
                    print("Automatic lane marker tuning cancelled.")

                self.live_view_label.bounding_box_preview = None
                self.live_view_label.quad_points = []
                self.update_live_view()


            def place_markers_from_dialog(self, original_region_definition, peak_coords_in_dialog, side,
                                          dialog_image_height, dialog_image_width, 
                                          is_from_quad_warp):
                if not peak_coords_in_dialog.any() or dialog_image_height <= 0 or dialog_image_width <= 0:
                    print("No peak coordinates or invalid dialog image dimensions for auto lane marker placement.")
                    return

                self.save_state()
                target_marker_list = None
                if side == 'left':
                    self.left_markers.clear(); self.current_left_marker_index = 0
                    target_marker_list = self.left_markers
                elif side == 'right':
                    self.right_markers.clear(); self.current_right_marker_index = 0
                    target_marker_list = self.right_markers
                else:
                    print(f"Invalid side '{side}' for auto lane marker placement."); return

                self.on_combobox_changed() 
                current_marker_values_for_labels = self.marker_values 

                if not current_marker_values_for_labels: 
                    current_marker_values_for_labels = [""] * len(peak_coords_in_dialog)
                elif len(current_marker_values_for_labels) < len(peak_coords_in_dialog):
                    current_marker_values_for_labels.extend([""] * (len(peak_coords_in_dialog) - len(current_marker_values_for_labels)))

                sorted_peaks_dialog = np.sort(peak_coords_in_dialog.astype(float)) 
                
                inv_perspective_matrix = None
                src_points_for_transform_np = None

                if is_from_quad_warp:
                    if len(original_region_definition) != 4:
                        print("Error: Quadrilateral definition requires 4 points for inverse mapping in auto lane.")
                        return
                    src_points_for_transform_list = []
                    for p in original_region_definition:
                        try: src_points_for_transform_list.append([float(p.x()), float(p.y())])
                        except AttributeError: src_points_for_transform_list.append([float(p[0]), float(p[1])])
                    src_points_for_transform_np = np.array(src_points_for_transform_list, dtype=np.float32)
                    
                    dst_points_for_dialog_transform_np = np.array([
                        [0.0, 0.0],
                        [float(dialog_image_width) - 1.0, 0.0],
                        [float(dialog_image_width) - 1.0, float(dialog_image_height) - 1.0],
                        [0.0, float(dialog_image_height) - 1.0]
                    ], dtype=np.float32)
                    try:
                        inv_perspective_matrix = cv2.getPerspectiveTransform(dst_points_for_dialog_transform_np, src_points_for_transform_np)
                    except Exception as e:
                        print(f"Error calculating inverse perspective matrix for auto lane: {e}")
                        return

                for i, peak_y_dialog in enumerate(sorted_peaks_dialog):
                    label = str(current_marker_values_for_labels[i]) if i < len(current_marker_values_for_labels) else ""
                    y_final_img_calc = 0.0
                    
                    if not is_from_quad_warp:
                        rect_x_img, rect_y_img, rect_w_img, rect_h_img = map(float, original_region_definition)
                        t = peak_y_dialog / (float(dialog_image_height) - 1.0) if dialog_image_height > 1 else 0.5
                        t = max(0.0, min(1.0, t))
                        y_final_img_calc = rect_y_img + t * rect_h_img
                    else:
                        if inv_perspective_matrix is None: continue
                        x_in_dialog_space = float(dialog_image_width) * 0.05 if side == 'left' else float(dialog_image_width) * 0.95
                        point_in_dialog_np = np.array([[[x_in_dialog_space, peak_y_dialog]]], dtype=np.float32)
                        try:
                            original_image_point_np = cv2.perspectiveTransform(point_in_dialog_np, inv_perspective_matrix)
                            y_final_img_calc = float(original_image_point_np[0,0,1])
                        except Exception as e_transform:
                            print(f"Error transforming point back for auto lane: {e_transform}"); continue
                    
                    final_y_to_store = float(y_final_img_calc)
                    target_marker_list.append((final_y_to_store, label))
                
                # --- MODIFICATION START (with pixel-perfect refinement) ---
                # Override X position logic to use detected content borders instead of region geometry.
                left_border, right_border = self._find_image_content_borders()
                
                target_x_in_image_space_for_slider = 0.0

                if side == 'left':
                    if left_border is not None:
                        # For the left marker, the text "value ⎯" is drawn to the left of the anchor.
                        # Setting the anchor exactly at the border ensures the line sits right on it.
                        target_x_in_image_space_for_slider = left_border
                        print(f"INFO: Auto lane marker X position set to detected left content border: {left_border}")
                    else:
                        print("WARNING: Could not detect image content border. Falling back to edge.")
                        target_x_in_image_space_for_slider = 0 
                elif side == 'right':
                    if right_border is not None:
                        # For the right marker, the text "⎯ value" is drawn starting at the anchor.
                        # Setting the anchor to the border + 1 ensures the line starts just outside the content.
                        target_x_in_image_space_for_slider = right_border + 1
                        print(f"INFO: Auto lane marker X position set to detected right content border + 1: {right_border + 1}")
                    else:
                        print("WARNING: Could not detect image content border. Falling back to edge.")
                        target_x_in_image_space_for_slider = self.image.width() - 1 if self.image else 0
                # --- MODIFICATION END ---
                
                slider_target_value_native_pixels = int(round(target_x_in_image_space_for_slider))
                self._update_marker_slider_ranges()

                if side == 'left':
                    if hasattr(self, 'left_padding_slider'):
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(
                            max(self.left_slider_range[0], min(slider_target_value_native_pixels, self.left_slider_range[1]))
                        )
                        self.left_padding_slider.blockSignals(False)
                        self.left_marker_shift_added = self.left_padding_slider.value()
                elif side == 'right':
                     if hasattr(self, 'right_padding_slider'):
                         self.right_padding_slider.blockSignals(True)
                         self.right_padding_slider.setValue(
                             max(self.right_slider_range[0], min(slider_target_value_native_pixels, self.right_slider_range[1]))
                         )
                         self.right_padding_slider.blockSignals(False)
                         self.right_marker_shift_added = self.right_padding_slider.value()

                self.is_modified = True
                self.update_live_view()
                
            def zoom_in(self):
                self.live_view_label.zoom_in()

            def zoom_out(self):
                self.live_view_label.zoom_out()
                
            
            def enable_standard_protein_mode(self):
                """"Enable mode to define standard protein amounts for creating a standard curve."""
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                # Assign mouse event handlers for bounding box creation
                self._reset_live_view_label_custom_handlers() # Good practice to reset first
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.start_bounding_box(event)
                self.live_view_label._custom_mouseReleaseEvent_from_app = lambda event: self.end_standard_bounding_box(event)
            
            def enable_measure_protein_mode(self):
                """Enable mode to measure protein quantity using the standard curve."""
                if len(self.quantities) < 2:
                    QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
                self.measure_quantity_mode = True
                self.live_view_label.measure_quantity_mode = True
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
                self.setMouseTracking(True)  # Ensure parent also tracks mouse
                self._reset_live_view_label_custom_handlers() # Good practice
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.start_bounding_box(event)
                self.live_view_label._custom_mouseReleaseEvent_from_app = lambda event: self.end_measure_bounding_box(event)

            def call_live_view(self):
                self.update_live_view()      

            def analyze_bounding_box(self, pil_image_for_dialog, standard):
                peak_info_result = None # Will hold list of dicts
                self.latest_peak_areas = [] # Clear for single lane mode
                self.latest_peak_details = [] # Clear for single lane mode
                self.latest_calculated_quantities = []

                if standard:
                    quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
                    if ok and quantity:
                        try:
                            quantity_value = float(quantity.split()[0])
                            peak_info_result = self.calculate_peak_area(pil_image_for_dialog)
                            
                            if peak_info_result and len(peak_info_result) > 0:
                                areas_for_standard = [info['area'] for info in peak_info_result]
                                total_area = sum(areas_for_standard)
                                self.quantities_peak_area_dict[quantity_value] = total_area
                                
                                # FIX: Format numbers for clean display
                                formatted_quantities = [f"{qty:.2f}" for qty in self.quantities_peak_area_dict.keys()]
                                formatted_areas = [f"{area:.3f}" for area in self.quantities_peak_area_dict.values()]
                                self.standard_protein_values.setText(", ".join(formatted_quantities))
                                self.standard_protein_areas_text.setText(", ".join(formatted_areas))

                                print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                                self.latest_peak_areas = [round(a, 3) for a in areas_for_standard]
                                self.latest_peak_details = peak_info_result # Store full details
                            else:
                                 print("Peak area calculation cancelled or failed for standard.")
                        except (ValueError, IndexError) as e:
                            QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                        except Exception as e:
                             QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                    else:
                        print("Standard quantity input cancelled.")
                else: # Analyze sample
                    peak_info_result = self.calculate_peak_area(pil_image_for_dialog)
                    if peak_info_result and len(peak_info_result) > 0:
                        self.latest_peak_areas = [round(info['area'], 3) for info in peak_info_result]
                        self.latest_peak_details = peak_info_result # Store full details
                        print(f"Sample Analysis: Calculated Peak Info = {self.latest_peak_details}")

                        if len(self.quantities_peak_area_dict) >= 2:
                            # --- NEW: Use the central quantification engine ---
                            # The model name is taken from the TableWindow's current tab, if open.
                            model_to_use = "Linear" # Default
                            if self.table_window_instance and not self.table_window_instance.isHidden():
                                model_to_use = self.table_window_instance.model_combo_current.currentText()
                            
                            quantities, _ = self._perform_quantification(
                                model_to_use,
                                list(self.quantities_peak_area_dict.keys()),
                                list(self.quantities_peak_area_dict.values()),
                                self.latest_peak_areas
                            )
                            self.latest_calculated_quantities = quantities
                        else: self.latest_calculated_quantities = []
                        try: 
                            # FIX: Format numbers for clean display
                            formatted_areas = [f"{area:.3f}" for area in self.latest_peak_areas]
                            self.target_protein_areas_text.setText(", ".join(formatted_areas))
                        except Exception as e: 
                            print(f"Error displaying sample areas: {e}"); self.target_protein_areas_text.setText("Error")
                    else:
                         print("Peak area calculation cancelled or failed for sample.")
                         self.target_protein_areas_text.setText("N/A")
                self.update_live_view()
            
            def calculate_peak_area(self, pil_image_for_dialog):
                if pil_image_for_dialog is None: # ... (existing validation) ...
                    print("Error: No PIL Image data provided to calculate_peak_area.")
                    return None # Return None if no areas/info
                if not isinstance(pil_image_for_dialog, Image.Image):
                     QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
                     return None

                dialog = PeakAreaDialog(
                    cropped_data=pil_image_for_dialog,
                    current_settings=self.peak_dialog_settings,
                    persist_checked=self.persist_peak_settings_enabled,
                    parent=self
                )

                peak_info_list = None # Will be list of dicts
                if dialog.exec() == QDialog.Accepted:
                    peak_info_list = dialog.get_final_peak_info() # Get the detailed info
                    if dialog.should_persist_settings():
                        self.peak_dialog_settings = dialog.get_current_settings()
                        self.persist_peak_settings_enabled = True
                    else:
                        self.persist_peak_settings_enabled = False
                else:
                    print("PeakAreaDialog cancelled.")
                
                return peak_info_list if peak_info_list is not None else [] # Return list of dicts or empty list
            
            
            def _perform_quantification(self, model_name, standard_quantities, standard_areas, sample_areas):
                """
                Central quantification engine. Fits a standard curve using the specified model
                and calculates unknown quantities for sample areas.
                """
                if len(standard_quantities) < 2 or not sample_areas:
                    return ([0.0] * len(sample_areas), None)

                std_qty_np = np.array(standard_quantities, dtype=float)
                std_area_np = np.array(standard_areas, dtype=float)
                sample_areas_np = np.array(sample_areas, dtype=float)
                
                fit_params = None
                calculated_quantities = []

                try:
                    # --- START OF THE FIX ---
                    # For polynomial models, we fit Quantity as a function of Area.
                    # This makes calculating the unknown quantity a simple evaluation, not a complex root-finding problem.
                    if "Linear" in model_name or "Polynomial" in model_name:
                        degree = 1
                        if "Deg 2" in model_name: degree = 2
                        elif "Deg 3" in model_name: degree = 3
                        
                        # Ensure enough points for the chosen degree
                        if len(std_qty_np) <= degree: return ([0.0] * len(sample_areas), None)

                        # Fit Quantity = f(Area)
                        fit_params = np.polyfit(std_area_np, std_qty_np, degree)
                        
                        # Calculate quantities by simply evaluating the polynomial with the sample areas
                        calculated_quantities = np.polyval(fit_params, sample_areas_np).tolist()

                    # --- END OF THE FIX ---

                    elif SCIPY_AVAILABLE and "4-PL" in model_name:
                        # For sigmoidal models, we fit Area = f(Quantity) and then use the mathematical inverse.
                        if len(std_qty_np) < 4: return ([0.0] * len(sample_areas), None)
                        
                        p0 = [min(std_area_np), 1.0, np.median(std_qty_np), max(std_area_np)]
                        fit_params, _ = curve_fit(four_param_logistic, std_qty_np, std_area_np, p0=p0, maxfev=10000)
                        a, b, c, d = fit_params

                        # Calculate quantity from area using the inverse of the 4-PL function
                        term = ((a - d) / (sample_areas_np - d)) - 1
                        quantities = np.full_like(sample_areas_np, 0.0)
                        valid_mask = term > 0
                        quantities[valid_mask] = c * (term[valid_mask]**(1/b))
                        calculated_quantities = quantities.tolist()
                
                except (RuntimeError, ValueError, np.linalg.LinAlgError) as e:
                    print(f"ERROR: Could not fit model '{model_name}': {e}")
                    return ([0.0] * len(sample_areas), None)

                return (calculated_quantities, fit_params)

                
            def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
                """Draw quantity text at the correct position."""
                text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
                painter.drawText(text_position, str(quantity))
            
            def update_standard_protein_quantities(self):
                self.standard_protein_values.text()
            
            def move_tab(self,tab):
                self.tab_widget.setCurrentIndex(tab)
                
            def save_state(self):
                """
                Saves a complete snapshot of the application's visual and logical state to the undo stack.
                This is the new, robust state management core.
                """
                if not self.image_master or self.image_master.isNull():
                    return

                UNDO_LIMIT = 20
                while len(self.undo_stack) >= UNDO_LIMIT:
                    self.undo_stack.pop(0)

                state = {
                    "selected_preset": self.combo_box.currentText() if hasattr(self, 'combo_box') else "Custom",
                    # --- Core Image ---
                    "image_master": self.image_master.copy(),

                    # --- Annotations ---
                    "left_markers": [list(m) for m in self.left_markers],
                    "right_markers": [list(m) for m in self.right_markers],
                    "top_markers": [list(m) for m in self.top_markers],
                    "custom_markers": [list(m) for m in self.custom_markers],
                    "custom_shapes": [dict(s) for s in self.custom_shapes],

                    # --- Marker Alignment ---
                    "left_marker_shift_added": self.left_marker_shift_added,
                    "right_marker_shift_added": self.right_marker_shift_added,
                    "top_marker_shift_added": self.top_marker_shift_added,

                    # --- Font Settings (Standard Markers) ---
                    "font_family": self.font_family,
                    "font_size": self.font_size,
                    "font_color": self.font_color.name(),
                    "font_rotation": self.font_rotation,

                    # --- Image Adjustments ---
                    "main_image_is_inverted": self.main_image_is_inverted,
                    "levels_gamma": {
                        'black_point': self.black_point_slider.value(),
                        'white_point': self.white_point_slider.value(),
                        'gamma': self.gamma_slider.value()
                    },
                    "channel_mixer_data": self.channel_mixer_data.copy(),
                    "unsharp_mask_data": self.unsharp_mask_data.copy(),
                    "clahe_data": self.clahe_data.copy(),
                    
                    # --- Padding State ---
                    "image_padded": self.image_padded,

                    # --- ANALYSIS REGIONS (THE FIX) ---
                    "multi_lane_definitions": [
                        {
                            'type': d['type'], 'id': d['id'],
                            'points_label': (
                                [(p.x(), p.y()) for p in d['points_label']] if d['type'] == 'quad' else
                                [(d['points_label'][0].x(), d['points_label'][0].y(), d['points_label'][0].width(), d['points_label'][0].height())]
                            )
                        } for d in self.multi_lane_definitions
                    ],
                    "single_quad_points": [(p.x(), p.y()) for p in self.live_view_label.quad_points],
                    "single_bounding_box": self.live_view_label.bounding_box_preview,
                    # --- END OF FIX ---
                }
                
                self.undo_stack.append(state)
                self.redo_stack.clear()

            def _restore_state_from_dict(self, state_dict):
                self._is_restoring_state = True # LOCK state saving

                try:
                    # 1. Restore Data Variables
                    self.image_master = state_dict['image_master'].copy()
                    self.left_markers = state_dict['left_markers']
                    self.right_markers = state_dict['right_markers']
                    self.top_markers = state_dict['top_markers']
                    self.custom_markers = state_dict['custom_markers']
                    self.custom_shapes = state_dict['custom_shapes']
                    
                    self.left_marker_shift_added = state_dict['left_marker_shift_added']
                    self.right_marker_shift_added = state_dict['right_marker_shift_added']
                    self.top_marker_shift_added = state_dict['top_marker_shift_added']
                    
                    self.font_family = state_dict['font_family']
                    self.font_size = state_dict['font_size']
                    self.font_color = QColor(state_dict['font_color'])
                    self.font_rotation = state_dict['font_rotation']

                    self.main_image_is_inverted = state_dict['main_image_is_inverted']
                    self.channel_mixer_data = state_dict['channel_mixer_data']
                    self.unsharp_mask_data = state_dict['unsharp_mask_data']
                    self.clahe_data = state_dict['clahe_data']
                    self.image_padded = state_dict['image_padded']

                    # Restore Analysis Regions
                    serialized_defs = state_dict.get("multi_lane_definitions", [])
                    self.multi_lane_definitions = []
                    for d in serialized_defs:
                        restored_def = {'type': d['type'], 'id': d['id']}
                        if d['type'] == 'quad':
                            restored_def['points_label'] = [QPointF(x, y) for x, y in d['points_label']]
                        elif d['type'] == 'rectangle':
                            x, y, w, h = d['points_label'][0]
                            restored_def['points_label'] = [QRectF(x, y, w, h)]
                        self.multi_lane_definitions.append(restored_def)

                    serialized_quad = state_dict.get("single_quad_points", [])
                    self.live_view_label.quad_points = [QPointF(x, y) for x, y in serialized_quad]
                    self.live_view_label.bounding_box_preview = state_dict.get("single_bounding_box", None)

                    # 2. Restore UI - Adjustment Sliders (Safe to do early)
                    self._load_adjustments_to_ui("Main Image")
                    lg_settings = state_dict['levels_gamma']
                    self.black_point_slider.setValue(lg_settings.get('black_point', 0))
                    self.white_point_slider.setValue(lg_settings.get('white_point', 65535))
                    self.gamma_slider.setValue(lg_settings.get('gamma', 100))

                    # 3. CRITICAL: Generate the full image NOW so we have correct dimensions
                    self.apply_all_adjustments() 

                    # 4. CRITICAL: Update Slider Ranges based on the restored image size
                    self._update_marker_slider_ranges()

                    # 5. CRITICAL: Restore Marker Sliders WITHOUT triggering signals
                    # (Triggering signals now would overwrite the variables we just restored)
                    if hasattr(self, 'left_padding_slider'):
                        self.left_padding_slider.setEnabled(True)
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(self.left_marker_shift_added)
                        self.left_padding_slider.blockSignals(False)
                    
                    if hasattr(self, 'right_padding_slider'):
                        self.right_padding_slider.setEnabled(True)
                        self.right_padding_slider.blockSignals(True)
                        self.right_padding_slider.setValue(self.right_marker_shift_added)
                        self.right_padding_slider.blockSignals(False)
                        
                    if hasattr(self, 'top_padding_slider'):
                        self.top_padding_slider.setEnabled(True)
                        self.top_padding_slider.blockSignals(True)
                        self.top_padding_slider.setValue(self.top_marker_shift_added)
                        self.top_padding_slider.blockSignals(False)

                    # 6. Restore Font UI
                    self.font_combo_box.blockSignals(True)
                    self.font_combo_box.setCurrentFont(QFont(self.font_family))
                    self.font_combo_box.blockSignals(False)
                    
                    self.font_size_spinner.blockSignals(True)
                    self.font_size_spinner.setValue(self.font_size)
                    self.font_size_spinner.blockSignals(False)
                    
                    self.font_rotation_input.blockSignals(True)
                    self.font_rotation_input.setValue(self.font_rotation)
                    self.font_rotation_input.blockSignals(False)
                    
                    self._update_color_button_style(self.font_color_button, self.font_color)

                    # 7. NEW: Restore Preset Selection (Prevent "Bio-Rad" from overwriting "Custom")
                    saved_preset = state_dict.get("selected_preset", "Custom")
                    if hasattr(self, 'combo_box'):
                        self.combo_box.blockSignals(True)
                        index = self.combo_box.findText(saved_preset)
                        if index != -1:
                            self.combo_box.setCurrentIndex(index)
                        else:
                            # If saved preset not found, default to Custom to preserve markers
                            custom_idx = self.combo_box.findText("Custom")
                            if custom_idx != -1: self.combo_box.setCurrentIndex(custom_idx)
                        self.combo_box.blockSignals(False)
                        
                        # Manually trigger UI updates for preset text boxes without saving state
                        if saved_preset == "Custom":
                            self.marker_values_textbox.setEnabled(True)
                            self.rename_input.setEnabled(True)
                        else:
                            self.marker_values_textbox.setEnabled(False)
                            self.rename_input.setEnabled(False)

                    self._update_status_bar()

                except KeyError as e:
                    print(f"Undo Error: Missing key {e}")
                except Exception as e:
                    print(f"Undo Error: {e}")
                    import traceback
                    traceback.print_exc()
                finally:
                    self._is_restoring_state = False
                    self.update_live_view()
            
            def undo_action_m(self):
                """Undo the last action by restoring the previous state."""
                if len(self.undo_stack) > 1: # Need at least one state to revert TO
                    current_state = self.undo_stack.pop()
                    self.redo_stack.append(current_state)
                    
                    state_to_restore = self.undo_stack[-1] # Peek at the new top of the undo stack
                    self._restore_state_from_dict(state_to_restore)
                    self.is_modified = True

            def redo_action_m(self):
                """Redo the last undone action by restoring the next state."""
                if self.redo_stack:
                    state_to_restore = self.redo_stack.pop()
                    self.undo_stack.append(state_to_restore) # Push it back onto the undo stack
                    
                    self._restore_state_from_dict(state_to_restore)
                    self.is_modified = True

            def get_current_config_for_state(self):
                # Helper to gather current state for undo/redo stack
                return {
                    "image": self.image.copy() if self.image else None,
                    # --- START FIX ---
                    # This was the missing line. Without it, the redo stack receives
                    # incomplete state objects, causing the "no valid master image" error.
                    "image_master": self.image_master.copy() if self.image_master else None,
                    # --- END FIX ---
                    "left_markers": self.left_markers.copy(), "right_markers": self.right_markers.copy(), "top_markers": self.top_markers.copy(),
                    "custom_markers": [list(m) for m in getattr(self, "custom_markers", [])], "custom_shapes": [dict(s) for s in getattr(self, "custom_shapes", [])],
                    "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                    "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                    "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                    "font_family": self.font_family, "font_size": self.font_size, "font_color": self.font_color, "font_rotation": self.font_rotation,
                    "left_marker_shift_added": self.left_marker_shift_added, "right_marker_shift_added": self.right_marker_shift_added, "top_marker_shift_added": self.top_marker_shift_added,
                    "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(), "quantities": self.quantities.copy(), "protein_quantities": self.protein_quantities.copy(), "standard_protein_areas": self.standard_protein_areas.copy(),
                    "custom_marker_color": self.custom_marker_color, "custom_font_family": self.custom_font_type_dropdown.currentText(), "custom_font_size": self.custom_font_size_spinbox.value(),
                    "channel_mixer_data": self.channel_mixer_data.copy(), "unsharp_mask_data": self.unsharp_mask_data.copy(), "clahe_data": self.clahe_data.copy(),
                    "black_point": self.black_point_slider.value(), "white_point": self.white_point_slider.value(), "gamma": self.gamma_slider.value()
                }
                    
            def analysis_tab(self):
                tab = QWidget()
                main_layout = QVBoxLayout(tab)
                main_layout.setSpacing(10)

                # --- Group 1: Molecular Weight & Measurement Tools (Combined) ---
                mw_group = QGroupBox("Molecular Weight and Measurements")
                mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                mw_layout = QGridLayout(mw_group)
                
                # --- Row 1: MW Prediction ---
                mw_layout.addWidget(QLabel("Regression Model:"), 0, 0)
                self.mw_regression_model_combo = QComboBox()
                self.mw_regression_model_combo.addItems(["Log-Linear (Degree 1)", "Log-Polynomial (Degree 2)", "Log-Polynomial (Degree 3)","Log 4-PL"])
                self.mw_regression_model_combo.setCurrentText("Log-Polynomial (Degree 2)")
                mw_layout.addWidget(self.mw_regression_model_combo, 0, 1)
                
                self.predict_button = QPushButton("Predict Molecular Weight")
                self.predict_button.setToolTip("Click a point on a lane to predict its molecular weight/size based on the active standard markers.\nShortcut: Ctrl+P")
                self.predict_button.setEnabled(False)
                self.predict_button.clicked.connect(self.predict_molecular_weight)
                mw_layout.addWidget(self.predict_button, 0, 2)

                # --- Row 2: Protein Analysis & Measurement Button ---
                self.open_glyco_mapper_button = QPushButton("Protein Analysis")
                self.open_glyco_mapper_button.setToolTip("Analyze a protein sequence for modifications and oligomerization.")
                self.open_glyco_mapper_button.clicked.connect(self.open_protein_analyzer)
                mw_layout.addWidget(self.open_glyco_mapper_button, 0, 3)

                self.btn_open_measure_window = QPushButton("Measurement Tools")
                self.btn_open_measure_window.setToolTip("Opens a separate window for image calibration, distance, and area measurements.")
                self.btn_open_measure_window.clicked.connect(self.open_measurement_window)
                mw_layout.addWidget(self.btn_open_measure_window, 0, 4) # Moved here

                mw_layout.setColumnStretch(1, 1)
                self.show_oligomer_glyco_overlay_checkbox = QCheckBox("Show Oligomer/Glyco Overlay")
                self.show_oligomer_glyco_overlay_checkbox.setToolTip("After predicting an MW, overlay potential oligomeric and glycosylated bands on the image.")
                self.show_oligomer_glyco_overlay_checkbox.stateChanged.connect(self.update_live_view)
                mw_layout.addWidget(self.show_oligomer_glyco_overlay_checkbox, 1, 0, 1, 4)
                
                main_layout.addWidget(mw_group)

                # --- Group 2: Lane Quantification Workflow ---
                quant_workflow_group = QGroupBox("Lane Quantification Workflow")
                # ... (The rest of your analysis_tab code for the quantification workflow is unchanged) ...
                quant_workflow_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                quant_workflow_layout = QVBoxLayout(quant_workflow_group)
                quant_workflow_layout.setSpacing(8)
                step1_group = QGroupBox("Define and Manage Analysis Regions")
                step1_layout = QGridLayout(step1_group)
                self.btn_define_quad = QPushButton("Quadrilateral Region")
                self.btn_define_quad.setToolTip("Start a session to define one or more skewed lane regions by clicking 4 corner points for each.\nPress ESC to finish the session.")
                self.btn_define_quad.clicked.connect(lambda: self.start_region_definition_session('quad'))
                self.btn_define_quad.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.btn_define_rec = QPushButton("Rectangular Region")
                self.btn_define_rec.setToolTip("Start a session to define one or more straight lane regions by clicking and dragging for each.\nPress ESC to finish the session.")
                self.btn_define_rec.clicked.connect(lambda: self.start_region_definition_session('rectangle'))
                self.btn_define_rec.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.btn_sel_rec = QPushButton("Move/Resize/Delete Area")
                self.btn_sel_rec.setToolTip(
                    "Click a lane to select it. Then:\n"
                    "- Drag Body: Move shape freely.\n"
                    "- Shift+Drag Body: Move horizontally or vertically.\n"
                    "- Cmd/Ctrl+Shift+Drag Body: Duplicate the shape.\n"
                    "- Drag Corner: Resize shape.\n"
                    "- Shift+Drag Corner: Resize horizontally or vertically.\n"
                    "- Drag Edge: Move edge perpendicularly (e.g., top edge moves up/down).\n"
                    "- Shift+Drag Edge: Slide edge in parallel (e.g., top edge moves left/right).\n"
                    "- Cmd/Ctrl+Drag Edge: Skew edge to create a trapezoid."
                )
                self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
                self.btn_sel_rec.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                step1_layout.addWidget(self.btn_define_quad, 0, 0)
                step1_layout.addWidget(self.btn_define_rec, 0, 1)
                step1_layout.addWidget(self.btn_sel_rec, 0, 2)
                quant_workflow_layout.addWidget(step1_group,0)
                step2_group = QGroupBox("Process Regions and View Results")
                step2_layout = QGridLayout(step2_group)
                process_buttons_layout = QHBoxLayout()
                self.btn_process_std = QPushButton("Process Selected as Standard"); self.btn_process_std.setToolTip("Analyze a pre-defined and selected region as a single standard lane.")
                self.btn_process_std.clicked.connect(self.process_standard)
                self.btn_analyze_sample = QPushButton("Analyze as Sample(s)"); self.btn_analyze_sample.setToolTip("Analyze the defined region(s) as sample lane(s), using the standard curve if available.")
                self.btn_analyze_sample.clicked.connect(self.process_sample)
                self.btn_process_std.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.btn_analyze_sample.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                process_buttons_layout.addWidget(self.btn_process_std)
                process_buttons_layout.addWidget(self.btn_analyze_sample)
                step2_layout.addLayout(process_buttons_layout, 0, 0, 1, 3)
                step2_layout.addWidget(self.create_separator(), 1, 0, 1, 3)
                step2_layout.addWidget(QLabel("Std. Quantities:"), 2, 0); self.standard_protein_values = QLineEdit(); self.standard_protein_values.setPlaceholderText("Known quantities (comma-separated)")
                step2_layout.addWidget(self.standard_protein_values, 2, 1)
                self.update_standards_button = QPushButton("Update Curve"); self.update_standards_button.setToolTip("Apply manual changes to the Quantities and Areas fields to update the standard curve.")
                self.update_standards_button.clicked.connect(self.update_standards_from_text_fields)
                step2_layout.addWidget(self.update_standards_button, 2, 2)
                step2_layout.addWidget(QLabel("Std. Areas:"), 3, 0); self.standard_protein_areas_text = QLineEdit(); self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (comma-separated)")
                step2_layout.addWidget(self.standard_protein_areas_text, 3, 1, 1, 2)
                step2_layout.addWidget(QLabel("Sample Results:"), 4, 0, Qt.AlignTop); self.target_protein_areas_text = QTextEdit(); self.target_protein_areas_text.setPlaceholderText("Calculated peak areas and quantities will appear here for each lane.")
                self.target_protein_areas_text.setReadOnly(True); self.target_protein_areas_text.setFixedHeight(60)
                step2_layout.addWidget(self.target_protein_areas_text, 4, 1, 1, 2)
                bottom_buttons_layout = QHBoxLayout()
                self.table_export_button = QPushButton("View Full Results and History")
                self.table_export_button.setToolTip("Open a new window to view, export, and manage current and past analysis results.")
                self.table_export_button.clicked.connect(self.open_table_window)
                self.save_analysis_button = QPushButton("Save Analysis")
                self.save_analysis_button.setToolTip("Saves all defined regions, standard curve data, and results to a config file associated with the current image.")
                self.save_analysis_button.clicked.connect(self.save_analysis_to_config)
                self.load_analysis_button = QPushButton("Load Analysis")
                self.load_analysis_button.setToolTip("Loads analysis regions and standard curve data from a config file.")
                self.load_analysis_button.clicked.connect(self.load_analysis_from_config)
                self.clear_predict_button = QPushButton("Reset Analysis")
                self.clear_predict_button.setToolTip("Clears MW prediction line, all analysis regions, and standard curve data.\nShortcut: Ctrl+Shift+P")
                self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
                self.table_export_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.save_analysis_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.load_analysis_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.clear_predict_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                bottom_buttons_layout.addWidget(self.table_export_button)
                bottom_buttons_layout.addWidget(self.save_analysis_button)
                bottom_buttons_layout.addWidget(self.load_analysis_button)
                bottom_buttons_layout.addWidget(self.clear_predict_button)
                step2_layout.addLayout(bottom_buttons_layout, 5, 0, 1, 3)
                quant_workflow_layout.addWidget(step2_group)
                main_layout.addWidget(quant_workflow_group)

                main_layout.addStretch(1)
                
                return tab

            
            def open_measurement_window(self):
                """Creates (if necessary) and shows the non-modal measurement tool window."""
                if not hasattr(self, 'measurement_tool_window') or not self.measurement_tool_window:
                    self.measurement_tool_window = MeasurementToolWindow(self)
                    # Connect signals from the tool window to methods in the main app
                    self.measurement_tool_window.tool_selected.connect(self.activate_measurement_tool)
                    self.measurement_tool_window.clear_requested.connect(self.clear_measurement_mode)
                    # When the dialog is closed, clean up the reference to it
                    self.measurement_tool_window.finished.connect(lambda: setattr(self, 'measurement_tool_window', None))
                
                self.measurement_tool_window.show()
                self.measurement_tool_window.raise_()
                self.measurement_tool_window.activateWindow()

            def activate_measurement_tool(self, mode):
                """Central controller to activate a specific measurement tool."""
                if mode == self.measurement_mode: # User clicked the same button again
                    self._exit_current_tool_mode()
                    return

                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    if self.measurement_tool_window:
                        self.measurement_tool_window.uncheck_all_tools()
                    return

                # Clear previous interaction state, but keep results on screen
                self._exit_current_tool_mode()
                # If a new measurement is started, then we clear the old finalized shape
                self.finalized_measurement_shape = None
                self.last_measurement_result_text = ""

                self.measurement_mode = mode
                
                # Setup mouse handlers and show instructions
                self.live_view_label.setCursor(Qt.CrossCursor)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.handle_measurement_mouse_press
                self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_measurement_mouse_move
                
                if mode == 'set_scale':
                    QMessageBox.information(self, "Set Scale", "Click two points to define the calibration line.")
                elif mode == 'measure_distance':
                    QMessageBox.information(self, "Measure Distance", "Click two points to measure the distance.")
                elif mode == 'measure_area':
                    QMessageBox.information(self, "Measure Area", "Click points to draw a polygon.\nClick the first point again to close the shape and calculate the area.")

            def _exit_current_tool_mode(self):
                """Helper to reset the UI interaction state of any measurement tool."""
                self.measurement_mode = None
                self.measurement_points = []  # Clear temporary drawing points
                self._reset_live_view_label_custom_handlers()
                # If the tool window is open, tell it to uncheck its buttons
                if hasattr(self, 'measurement_tool_window') and self.measurement_tool_window:
                    self.measurement_tool_window.uncheck_all_tools()
                self.update_live_view()

            def clear_measurement_mode(self, clear_scale=True):
                """Clears all measurement results, overlays, and exits any active tool mode."""
                if clear_scale:
                    self.pixels_per_unit = None
                    self.scale_unit = "pixels"
                    if hasattr(self, 'measurement_tool_window') and self.measurement_tool_window:
                        self.measurement_tool_window.update_scale_display("Not Set")

                # Clear the finalized results that are being displayed
                self.finalized_measurement_shape = None
                self.last_measurement_result_text = ""
                if hasattr(self, 'measurement_tool_window') and self.measurement_tool_window:
                    self.measurement_tool_window.update_result_display("N/A")

                # Exit the active tool (resets cursor, mouse handlers, etc.)
                self._exit_current_tool_mode()

            def handle_measurement_mouse_press(self, event):
                if not self.measurement_mode or event.button() != Qt.LeftButton:
                    return

                # --- START OF THE FIX ---
                # Instead of recalculating the point from the event, we use the PREVIEW point
                # that was calculated during the last mouse move. This is the point the user actually sees and intends to click.
                if self.live_view_label.current_preview_points:
                    # The preview point is already transformed, snapped, and constrained.
                    final_click_point_ls = self.live_view_label.current_preview_points[0]
                else:
                    # This is a fallback for safety, in case a click happens without a preceding move event.
                    point_ls = self.live_view_label.transform_point(event.position())
                    final_click_point_ls = self.snap_point_to_grid(point_ls)
                # --- END OF THE FIX ---

                if self.measurement_mode == 'measure_area':
                    if len(self.measurement_points) > 1:
                        first_point = self.measurement_points[0]
                        click_radius_threshold = 10 / self.live_view_label.zoom_level
                        if (final_click_point_ls - first_point).manhattanLength() < click_radius_threshold:
                            # If closing the shape, add the first point to ensure it's perfectly closed.
                            self.measurement_points.append(first_point)
                            self.finalize_measurement()
                            return
                    # Add the confirmed, constrained point to the list.
                    self.measurement_points.append(final_click_point_ls)

                else: # 'set_scale' or 'measure_distance' modes
                    # Add the confirmed, constrained point to the list.
                    self.measurement_points.append(final_click_point_ls)
                    if len(self.measurement_points) == 2:
                        self.finalize_measurement()
                
                self.update_live_view()

            def handle_measurement_mouse_move(self, event):
                """Handles mouse move for measurement previews, including Shift-key constraint for all modes."""
                if not self.measurement_mode or not self.measurement_points:
                    self.live_view_label.current_preview_points = []
                    return
                
                current_point_ls = self.live_view_label.transform_point(event.position())
                snapped_current_point = self.snap_point_to_grid(current_point_ls)
                
                # --- Unified Shift-key constraint logic ---
                if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                    start_point = None
                    if self.measurement_mode in ['set_scale', 'measure_distance']:
                        start_point = self.measurement_points[0]
                    elif self.measurement_mode == 'measure_area':
                        start_point = self.measurement_points[-1]
                    
                    if start_point:
                        delta_x = abs(snapped_current_point.x() - start_point.x())
                        delta_y = abs(snapped_current_point.y() - start_point.y())

                        if delta_x > delta_y:
                            snapped_current_point.setY(start_point.y())
                        else:
                            snapped_current_point.setX(start_point.x())

                # This list now ONLY contains the live, constrained cursor position for the rubber-band line.
                self.live_view_label.current_preview_points = [snapped_current_point]
                
                self.update_live_view()

            def finalize_measurement(self):
                """Performs calculations, sets the final shape for display, and exits the tool's interaction mode."""
                # This method's logic remains largely the same, but it's now called at the right time.
                if not self.measurement_mode or not self.measurement_points:
                    self._exit_current_tool_mode()
                    return

                self.finalized_measurement_shape = QPolygonF(self.measurement_points)
                points_img = self._map_label_points_to_image_points(self.measurement_points)
                if not points_img:
                    self.clear_measurement_mode()
                    return

                if self.measurement_mode == 'set_scale':
                    dialog = ScaleDialog(self)
                    if dialog.exec() == QDialog.Accepted:
                        known_length, unit = dialog.get_values()
                        pixel_dist = np.linalg.norm(np.array(points_img[0].toTuple()) - np.array(points_img[-1].toTuple()))
                        if pixel_dist > 1e-6 and known_length > 0:
                            self.pixels_per_unit = pixel_dist / known_length
                            self.scale_unit = unit
                            if self.measurement_tool_window:
                                self.measurement_tool_window.update_scale_display(f"{self.pixels_per_unit:.2f} px/{self.scale_unit}")
                    self.last_measurement_result_text = ""
                
                elif self.measurement_mode == 'measure_distance':
                    if self.pixels_per_unit and self.pixels_per_unit > 0:
                        pixel_dist = np.linalg.norm(np.array(points_img[0].toTuple()) - np.array(points_img[-1].toTuple()))
                        real_dist = pixel_dist / self.pixels_per_unit
                        self.last_measurement_result_text = f"{real_dist:.3f} {self.scale_unit}"
                        if self.measurement_tool_window:
                            self.measurement_tool_window.update_result_display(self.last_measurement_result_text)
                    else:
                        self.last_measurement_result_text = "No Scale Set"
                        if self.measurement_tool_window:
                            self.measurement_tool_window.update_result_display("Set scale first")
                
                elif self.measurement_mode == 'measure_area':
                    if self.pixels_per_unit and self.pixels_per_unit > 0:
                        points_img_np = np.array([p.toTuple() for p in points_img], dtype=np.int32)
                        area_px2 = cv2.contourArea(points_img_np)
                        area_real2 = area_px2 / (self.pixels_per_unit ** 2)
                        self.last_measurement_result_text = f"{area_real2:.3f} {self.scale_unit}\u00b2"
                        if self.measurement_tool_window:
                            self.measurement_tool_window.update_result_display(self.last_measurement_result_text)
                    else:
                        self.last_measurement_result_text = "No Scale Set"
                        if self.measurement_tool_window:
                            self.measurement_tool_window.update_result_display("Set scale first")
                
                self._exit_current_tool_mode()
            
            def save_analysis_to_config(self):
                """
                Saves the complete current configuration (including analysis data, regions,
                adjustments, and markers) to a text file. This will overwrite any
                existing config file for the current image.
                """
                if not self.image_path:
                    QMessageBox.warning(self, "Save Error", "Please save the main image first to associate it with a file name before saving the analysis.")
                    return

                # Construct the config file path based on the main image path
                base_name_no_ext = os.path.splitext(os.path.basename(self.image_path))[0]
                config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                config_save_path = os.path.join(os.path.dirname(self.image_path), f"{config_base}_config.txt")

                # --- START OF THE FIX: Simplify to always save the full current state ---

                # 1. Get the complete current configuration dictionary.
                #    This now correctly includes all boundary boxes and protein analysis data.
                config_data_to_save = self.get_current_config()

                # --- END OF THE FIX ---

                try:
                    with open(config_save_path, "w", encoding='utf-8') as config_file:
                        json.dump(config_data_to_save, config_file, indent=4)
                    
                    self.is_modified = False # Mark as saved
                    self.setWindowTitle(f"{self.window_title}::{config_base}")
                    self._update_status_bar()
                    QMessageBox.information(self, "Analysis Saved", f"Analysis configuration saved successfully to:\n{os.path.basename(config_save_path)}")
                
                except Exception as e:
                    QMessageBox.critical(self, "Save Analysis Error", f"Could not save the analysis config file: {e}")
                    traceback.print_exc()

            def load_analysis_from_config(self):
                """
                Loads analysis configuration from the config file associated with the current image,
                without opening a file dialog.
                """
                if not self.image_path:
                    QMessageBox.warning(self, "Load Error", "An image must be loaded and saved to have an associated analysis file.")
                    return

                # --- START OF MODIFICATION: Auto-find the config file ---
                base_name_no_ext = os.path.splitext(os.path.basename(self.image_path))[0]
                config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                config_path = os.path.join(os.path.dirname(self.image_path), f"{config_base}_config.txt")

                if not os.path.exists(config_path):
                    QMessageBox.information(self, "Load Analysis", f"No analysis file found for this image.\n(Looked for: {os.path.basename(config_path)})")
                    return
                # --- END OF MODIFICATION ---

                try:
                    with open(config_path, "r", encoding='utf-8') as config_file:
                        config_data = json.load(config_file)
                    
                    self.save_state()

                    # --- MODIFICATION: Call apply_config to load ALL data ---
                    self.apply_config(config_data,load_analysis=True)

                    # Specifically load analysis-related data that might be separate
                    raw_qpa_dict = config_data.get("quantities_peak_area_dict", {})
                    self.quantities_peak_area_dict = {float(k): float(v) for k, v in raw_qpa_dict.items()}
                    self.multi_lane_processing_finished = config_data.get("multi_lane_processing_finished", False)
                    
                    # Update UI text fields for standards
                    formatted_quantities = [f"{qty:.2f}" for qty in self.quantities_peak_area_dict.keys()]
                    formatted_areas = [f"{area:.3f}" for area in self.quantities_peak_area_dict.values()]
                    self.standard_protein_values.setText(", ".join(formatted_quantities))
                    self.standard_protein_areas_text.setText(", ".join(formatted_areas))

                    # Load Protein Analysis data
                    protein_analysis_data = config_data.get("protein_analysis_data", {})
                    self.protein_sequence = protein_analysis_data.get("protein_sequence", "")
                    self.base_protein_mw = float(protein_analysis_data.get("base_protein_mw", 0.0))
                    self.avg_glycan_mass = float(protein_analysis_data.get("avg_glycan_mass", 0.0))
                    self.num_oligomers_to_model = int(protein_analysis_data.get("num_oligomers_to_model", 1))
                    self.num_glycans_to_model = int(protein_analysis_data.get("num_glycans_to_model", 0))
                    self.last_predicted_mw = float(protein_analysis_data.get("last_predicted_mw", 0.0))
                    self.last_mw_prediction_model = protein_analysis_data.get("last_mw_prediction_model", None)
                    if self.last_mw_prediction_model and "coeffs" in self.last_mw_prediction_model:
                        c = self.last_mw_prediction_model["coeffs"]
                        if isinstance(c, list):
                            self.last_mw_prediction_model["coeffs"] = np.array(c)
                    
                    self.is_modified = True
                    self.update_live_view()
                    QMessageBox.information(self, "Analysis Loaded", f"Analysis configuration loaded successfully from:\n{os.path.basename(config_path)}")

                except Exception as e:
                    QMessageBox.critical(self, "Load Analysis Error", f"Failed to load or apply the configuration file: {e}")
                    traceback.print_exc()
            
            def update_standards_from_text_fields(self):
                """
                Reads the editable standard fields, validates them, updates the internal dictionary,
                and automatically recalculates quantities for the last analyzed samples using the
                currently selected regression model from the TableWindow.
                """
                try:
                    # Read and parse quantities and areas
                    qty_text = self.standard_protein_values.text()
                    quantities = [float(q.strip()) for q in qty_text.split(',') if q.strip()]
                    area_text = self.standard_protein_areas_text.text()
                    areas = [float(a.strip()) for a in area_text.split(',') if a.strip()]

                    if len(quantities) != len(areas):
                        QMessageBox.warning(self, "Mismatch Error", "The number of quantities must match the number of areas.")
                        return

                    # Rebuild the standard dictionary
                    self.quantities_peak_area_dict = dict(zip(quantities, areas))
                    
                    # Re-format and display the cleaned data back in the text boxes
                    formatted_quantities = [f"{qty:.2f}" for qty in self.quantities_peak_area_dict.keys()]
                    formatted_areas = [f"{area:.3f}" for area in self.quantities_peak_area_dict.values()]
                    self.standard_protein_values.setText(", ".join(formatted_quantities))
                    self.standard_protein_areas_text.setText(", ".join(formatted_areas))

                    QMessageBox.information(self, "Success", "Standard curve data has been updated.")
                    
                    # --- START OF MODIFICATION ---
                    # AUTOMATIC RECALCULATION LOGIC using the new quantification engine

                    if len(self.quantities_peak_area_dict) < 2:
                        # If standards are no longer valid, clear quantities
                        self.latest_calculated_quantities = []
                        self.latest_multi_lane_calculated_quantities = {}
                        self.target_protein_areas_text.clear() # Clear the entire results box
                        QMessageBox.information(self, "Info", "Standard curve is no longer valid (less than 2 points). Sample quantities cleared.")
                        return

                    # Determine which regression model to use from the (potentially open) TableWindow
                    model_to_use = "Linear" # Default if window isn't open or accessible
                    if self.table_window_instance and not self.table_window_instance.isHidden():
                        model_to_use = self.table_window_instance.model_combo_current.currentText()
                    
                    all_lanes_text_results = []
                    std_qtys = list(self.quantities_peak_area_dict.keys())
                    std_areas = list(self.quantities_peak_area_dict.values())

                    # Check if the last analysis was multi-lane
                    if self.latest_multi_lane_peak_areas:
                        for lane_id, lane_areas in self.latest_multi_lane_peak_areas.items():
                            if lane_areas:
                                new_quantities, _ = self._perform_quantification(
                                    model_to_use, std_qtys, std_areas, lane_areas
                                )
                                self.latest_multi_lane_calculated_quantities[lane_id] = new_quantities
                                
                                # Rebuild the display text for this lane
                                formatted_areas = ', '.join([f"{a:.3f}" for a in lane_areas])
                                formatted_quantities = ', '.join([f"{q:.2f}" for q in new_quantities])
                                all_lanes_text_results.append(f"Lane {lane_id}: Areas=[{formatted_areas}], Qty=[{formatted_quantities}]")
                            else:
                                all_lanes_text_results.append(f"Lane {lane_id}: No areas analyzed.")

                    # Check if the last analysis was single-lane
                    elif self.latest_peak_areas:
                        new_quantities, _ = self._perform_quantification(
                            model_to_use, std_qtys, std_areas, self.latest_peak_areas
                        )
                        self.latest_calculated_quantities = new_quantities
                        
                        # Rebuild the display text for the single lane
                        formatted_areas = ', '.join([f"{a:.3f}" for a in self.latest_peak_areas])
                        formatted_quantities = ', '.join([f"{q:.2f}" for q in new_quantities])
                        all_lanes_text_results.append(f"Areas=[{formatted_areas}], Qty=[{formatted_quantities}]")

                    # Update the UI text box with the newly calculated results
                    if all_lanes_text_results:
                        self.target_protein_areas_text.setText("\n".join(all_lanes_text_results))
                    
                    # If the table window is open, refresh it to show the new curve and quantities
                    if self.table_window_instance and not self.table_window_instance.isHidden():
                        self.table_window_instance._update_analysis_display_for_tab(is_for_history=False)

                    # --- END OF MODIFICATION ---

                except ValueError:
                    QMessageBox.critical(self, "Input Error", "Please ensure all values are comma-separated numbers.")
                except Exception as e:
                    QMessageBox.critical(self, "Update Error", f"An unexpected error occurred: {e}")
            
            def open_protein_analyzer(self):
                # Use the last predicted MW if available and positive, otherwise use the stored base_mw
                base_mw_for_dialog = self.base_protein_mw if self.base_protein_mw > 0 else self.last_predicted_mw
                # Default glycan mass to 0 if it hasn't been set to a positive value
                glycan_mass_for_dialog = self.avg_glycan_mass if self.avg_glycan_mass > 0 else 0.0

                dialog = GlycosylationMapperDialog(
                    self.protein_sequence,
                    base_mw_for_dialog,
                    glycan_mass_for_dialog,
                    self.num_oligomers_to_model,
                    self
                )
                if dialog.exec() == QDialog.Accepted:
                    results = dialog.get_results()
                    self.protein_sequence = results["sequence"]
                    self.base_protein_mw = results["base_mw"]
                    self.avg_glycan_mass = results["glycan_mass"]
                    self.num_glycosylation_sites = results["sites"]
                    self.num_glycans_to_model = results["glycans_to_use"]
                    self.num_oligomers_to_model = results["oligomers_to_use"]

                    # Recalculate and store all potential products
                    self.oligomer_products = []
                    if self.base_protein_mw > 0:
                        for j in range(1, self.num_oligomers_to_model + 1):
                            oligomer_base_mw = self.base_protein_mw * j
                            # Add the unglycosylated oligomer itself
                            self.oligomer_products.append(oligomer_base_mw)
                            # Add glycosylated forms if glycan mass is positive
                            if self.avg_glycan_mass > 0:
                                for i in range(1, self.num_glycans_to_model + 1):
                                    self.oligomer_products.append(oligomer_base_mw + (i * self.avg_glycan_mass))
                    
                    self.update_live_view()

            def _get_y_pos_from_mw(self, mw, model_data, min_max_pos):
                """Calculates the image Y-position for a given MW using the inverse regression model + calibration."""
                if mw <= 0 or model_data is None: return None
                if min_max_pos is None: min_max_pos = model_data.get("min_max_pos")
                if min_max_pos is None: return None

                min_pos, max_pos = min_max_pos
                log_mw = np.log10(mw)
                
                # --- APPLY CALIBRATION INVERSE ---
                calibration = model_data.get("calibration", {})
                if calibration.get("active", False):
                    slope = calibration.get("slope", 1.0)
                    offset = calibration.get("offset", 0.0)
                    if abs(slope) > 1e-9:
                        log_mw = (log_mw - offset) / slope
                
                model_type = model_data.get("model", "poly")
                coeffs = model_data.get("coeffs")
                if coeffs is None: return None

                try:
                    norm_dist = None
                    if model_type == "4-PL":
                        if not SCIPY_AVAILABLE: return None
                        a, b, c, d = coeffs
                        if (a - d) == 0 or (log_mw - d) == 0: return None
                        term = ((a - d) / (log_mw - d)) - 1
                        if term < 0: return None 
                        norm_dist = c * (term**(1/b))
                    else: # Polynomial
                        poly_coeffs = list(coeffs)
                        poly_coeffs[-1] -= log_mw
                        roots = np.roots(poly_coeffs)
                        valid_roots = [r.real for r in roots if np.isreal(r) and -0.5 <= r.real <= 1.5]
                        if not valid_roots: return None
                        norm_dist = min(valid_roots, key=lambda r: abs(r - 0.5))

                    if norm_dist is not None:
                        return min_pos + norm_dist * (max_pos - min_pos)
                    return None
                except (ValueError, ZeroDivisionError, RuntimeError):
                    return None
                
            def start_region_definition_session(self, region_type):
                """
                Unified entry point to start or continue a multi-lane definition session.
                If a session is not active, it starts a new one (clearing old regions).
                If a session is already active, it just switches the tool for the next shape.
                """
                self._reset_live_view_label_custom_handlers()
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "Please load an image first.")
                    return

                # If a multi-lane session is NOT already active, start a new one.
                if not self.multi_lane_mode_active:
                    self.save_state()
                    self.multi_lane_definitions = [] # Start a new session, clear old definitions
                    self.latest_multi_lane_peak_areas.clear()
                    self.latest_multi_lane_calculated_quantities.clear()
                    self.multi_lane_processing_finished = False
                    self.target_protein_areas_text.clear()
                    self.multi_lane_mode_active = True
                    print("INFO: Starting new multi-lane definition session.")

                # Set up the UI and handlers for the next region to be drawn.
                self._set_next_region_type(region_type)

            def _set_next_region_type(self, region_type):
                """Sets up the UI and handlers for defining the next region of a given type."""
                self.multi_lane_definition_type = region_type
                next_lane_id = len(self.multi_lane_definitions) + 1
                
                if region_type == 'quad':
                    self.live_view_label.mode = 'define_quad'
                    self.live_view_label.current_preview_points = []
                    QMessageBox.information(self, "Define Lane", f"Click 4 corners for Lane {next_lane_id}.\nPress ESC to finish the session.")
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_region_definition_click
                elif region_type == 'rectangle':
                    self.live_view_label.mode = 'define_rect'
                    self.live_view_label.current_preview_points = []
                    QMessageBox.information(self, "Define Lane", f"Draw Lane {next_lane_id}.\nPress ESC to finish the session.")
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_region_definition_start
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_region_definition_move
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_current_multi_lane_definition
                
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)
                self.update_live_view()

            def handle_region_definition_click(self, event):
                """Handles clicks for point-by-point shape definitions (e.g., quadrilaterals)."""
                if not self.multi_lane_mode_active or event.button() != Qt.LeftButton: return

                point_transformed = self.live_view_label.transform_point(event.position())
                snapped_point = self.snap_point_to_grid(point_transformed)
                self.live_view_label.current_preview_points.append(snapped_point)
                self.update_live_view()

                if self.multi_lane_definition_type == 'quad' and len(self.live_view_label.current_preview_points) == 4:
                    self.finalize_current_multi_lane_definition(event)

            def handle_region_definition_start(self, event):
                """Handles mouse press for drag-based shape definitions (e.g., rectangles)."""
                if not self.multi_lane_mode_active or event.button() != Qt.LeftButton: return
                start_point = self.snap_point_to_grid(self.live_view_label.transform_point(event.position()))
                self.live_view_label.current_preview_points = [start_point, start_point]
                self.update_live_view()

            def handle_region_definition_move(self, event):
                """Handles mouse move for drag-based shape definitions (e.g., rectangles)."""
                if not self.multi_lane_mode_active or not (event.buttons() & Qt.LeftButton):
                    return

                if self.multi_lane_definition_type == 'rectangle' and len(self.live_view_label.current_preview_points) == 2:
                    current_point = self.snap_point_to_grid(self.live_view_label.transform_point(event.position()))
                    # Update the second point (the end point) of the preview list
                    self.live_view_label.current_preview_points[1] = current_point
                    self.update_live_view()

            def handle_current_multi_lane_quad_click(self, event):
                if self.multi_lane_mode_active and self.multi_lane_definition_type == 'quad':
                    if event.button() == Qt.LeftButton:
                        point_transformed = self.live_view_label.transform_point(event.position())
                        snapped_point = self.snap_point_to_grid(point_transformed)
                        self.current_multi_lane_points.append(snapped_point)
                        # Update live_view_label's quad_points for drawing the current quad being defined
                        self.live_view_label.quad_points = self.current_multi_lane_points[:]
                        self.update_live_view()

                        if len(self.current_multi_lane_points) == 4:
                            self.finalize_current_multi_lane_definition(event) # Pass event for consistency if needed


            def finalize_current_multi_lane_definition(self, event):
                if not self.multi_lane_mode_active: return

                lane_id = len(self.multi_lane_definitions) + 1
                preview_points = self.live_view_label.current_preview_points
                definition_to_store = None

                if self.multi_lane_definition_type == 'rectangle':
                    if len(preview_points) == 2:
                        rect_ls = QRectF(preview_points[0], preview_points[1]).normalized()
                        if rect_ls.width() < 2 or rect_ls.height() < 2:
                            QMessageBox.warning(self, "Info", "Rectangle too small, please redraw.")
                            self.live_view_label.current_preview_points = [] # Clear preview
                            self.update_live_view()
                            return
                        definition_to_store = {'type': 'rectangle', 'points_label': [rect_ls], 'id': lane_id}
                
                elif self.multi_lane_definition_type == 'quad':
                    if len(preview_points) == 4:
                        definition_to_store = {'type': 'quad', 'points_label': preview_points, 'id': lane_id}

                if definition_to_store:
                    self.multi_lane_definitions.append(definition_to_store)
                    self.is_modified = True
                    # Immediately set up for the next shape of the same type
                    self._set_next_region_type(self.multi_lane_definition_type)
                else:
                    self.update_live_view()

        
            def cancel_multi_lane_mode(self):
                """
                Finishes the multi-lane definition session (triggered by ESC).
                It finalizes the mode but preserves the defined regions for processing.
                """
                if not self.multi_lane_mode_active:
                    return

                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None
                self.live_view_label.mode = None
                self.live_view_label.current_preview_points = []
                self._reset_live_view_label_custom_handlers()
                self.update_live_view()
                QMessageBox.information(self, "Finished Defining Lanes", 
                                        f"{len(self.multi_lane_definitions)} lane(s) defined. You can now process them as standards or samples.")

            
            def enable_move_selection_mode(self):
                """Enables mode to select an area for moving/resizing."""
                # --- START OF THE FIX ---
                # The single source of truth for all defined regions is now self.multi_lane_definitions.
                # The check is simplified to see if this list is empty.
                if not self.multi_lane_definitions:
                    QMessageBox.information(self, "Move/Resize Area", "No area is currently defined to select.")
                    self.cancel_selection_or_move_mode()
                    return
                # --- END OF THE FIX ---

                self.current_selection_mode = "select_for_move"
                self.live_view_label.mode = "select_for_move" 
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.handle_area_selection_click
                
                QMessageBox.information(self, "Select Area to Move/Resize", "Click on a defined lane to select it. Click near a corner to resize, or inside to move the whole shape.")
                self.moving_multi_lane_index = -1 
                self.resizing_corner_index = -1
                self.update_live_view()
                
            def handle_area_selection_click(self, event):
                if self.current_selection_mode != "select_for_move" or event.button() != Qt.LeftButton:
                    return

                clicked_point_ls = self.live_view_label.transform_point(event.position())
                click_radius_threshold = self.live_view_label.CORNER_HANDLE_BASE_RADIUS * 2.5 / self.live_view_label.zoom_level

                # Pass 1: Check for corner handles
                for i, lane_def in reversed(list(enumerate(self.multi_lane_definitions))):
                    corners = lane_def['points_label'] if lane_def['type'] == 'quad' else [lane_def['points_label'][0].topLeft(), lane_def['points_label'][0].topRight(), lane_def['points_label'][0].bottomRight(), lane_def['points_label'][0].bottomLeft()]
                    for corner_idx, corner_pt in enumerate(corners):
                        if (clicked_point_ls - corner_pt).manhattanLength() < click_radius_threshold:
                            # (This part for corner resizing is correct and remains the same)
                            self.current_selection_mode = "resizing_corner"; self.live_view_label.mode = "resizing_corner"
                            self.moving_multi_lane_index = i; self.resizing_corner_index = corner_idx
                            self.shape_points_at_drag_start_label = [QPointF(p) for p in corners]
                            self.initial_mouse_pos_for_shape_drag_label = clicked_point_ls
                            self.live_view_label.setCursor(Qt.CrossCursor)
                            self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_drag_operation
                            self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_drag_release
                            self.update_live_view(); return

                # Pass 2: Check for edge handles
                for i, lane_def in reversed(list(enumerate(self.multi_lane_definitions))):
                    corners = lane_def['points_label'] if lane_def['type'] == 'quad' else [lane_def['points_label'][0].topLeft(), lane_def['points_label'][0].topRight(), lane_def['points_label'][0].bottomRight(), lane_def['points_label'][0].bottomLeft()]
                    if len(corners) == 4:
                        edge_map = {0: (corners[0] + corners[1]) / 2.0, 1: (corners[3] + corners[2]) / 2.0, 2: (corners[0] + corners[3]) / 2.0, 3: (corners[1] + corners[2]) / 2.0}
                        for edge_idx, mid_point in edge_map.items():
                            if (clicked_point_ls - mid_point).manhattanLength() < click_radius_threshold:
                                # (This part for edge skewing is correct and remains the same)
                                self.current_selection_mode = "skewing_edge"; self.live_view_label.mode = "skewing_edge"
                                self.moving_multi_lane_index = i; self.skewing_edge_index = edge_idx
                                self.shape_points_at_drag_start_label = [QPointF(p) for p in corners]
                                self.initial_mouse_pos_for_shape_drag_label = clicked_point_ls
                                cursor = Qt.SizeVerCursor if edge_idx < 2 else Qt.SizeHorCursor
                                self.live_view_label.setCursor(cursor)
                                self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_drag_operation
                                self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_drag_release
                                self.update_live_view(); return

                # Pass 3: Check for body clicks (and duplication shortcut)
                for i, lane_def in reversed(list(enumerate(self.multi_lane_definitions))):
                    body_shape = QPolygonF(lane_def['points_label']) if lane_def['type'] == 'quad' else lane_def['points_label'][0]
                    if (isinstance(body_shape, QRectF) and body_shape.contains(clicked_point_ls)) or \
                       (isinstance(body_shape, QPolygonF) and body_shape.containsPoint(clicked_point_ls, Qt.OddEvenFill)):
                        corners = lane_def['points_label'] if lane_def['type'] == 'quad' else [body_shape.topLeft(), body_shape.topRight(), body_shape.bottomRight(), body_shape.bottomLeft()]
                        
                        # --- START OF THE FIX: Correctly check for duplication shortcut ---
                        modifiers = QApplication.keyboardModifiers()
                        is_duplicate = (modifiers == (Qt.ShiftModifier | Qt.ControlModifier)) or \
                                       (modifiers == (Qt.ShiftModifier | Qt.MetaModifier))
                        self.is_duplicating_shape = is_duplicate
                        # --- END OF THE FIX ---
                        
                        self.current_selection_mode = "dragging_shape"; self.live_view_label.mode = "dragging_shape"
                        self.moving_multi_lane_index = i; self.resizing_corner_index = -1
                        self.shape_points_at_drag_start_label = [QPointF(p) for p in corners]
                        self.initial_mouse_pos_for_shape_drag_label = clicked_point_ls
                        self.live_view_label.setCursor(Qt.CrossCursor if is_duplicate else Qt.SizeAllCursor)
                        self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_drag_operation
                        self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_drag_release
                        self.update_live_view(); return
                
                self.moving_multi_lane_index = -1; self.resizing_corner_index = -1; self.update_live_view()
            
            def handle_drag_operation(self, event):
                if not self.moving_multi_lane_index >= 0 or not (event.buttons() & Qt.LeftButton):
                    return

                current_mouse_pos_ls = self.live_view_label.transform_point(event.position())
                new_shape_points_label = list(self.shape_points_at_drag_start_label)
                lane_def = self.multi_lane_definitions[self.moving_multi_lane_index]
                
                if self.current_selection_mode == "dragging_shape":
                    delta = self.snap_point_to_grid(current_mouse_pos_ls) - self.initial_mouse_pos_for_shape_drag_label
                    if event.modifiers() & Qt.ShiftModifier:
                        if abs(delta.x()) > abs(delta.y()): delta.setY(0)
                        else: delta.setX(0)
                    new_shape_points_label = [p + delta for p in self.shape_points_at_drag_start_label]
                
                elif self.current_selection_mode == "resizing_corner":
                    snapped_mouse = self.snap_point_to_grid(current_mouse_pos_ls)
                    if event.modifiers() & Qt.ShiftModifier:
                        start_corner = self.shape_points_at_drag_start_label[self.resizing_corner_index]
                        delta = snapped_mouse - start_corner
                        if abs(delta.x()) > abs(delta.y()): snapped_mouse.setY(start_corner.y())
                        else: snapped_mouse.setX(start_corner.x())
                    if lane_def['type'] == 'quad':
                        new_shape_points_label[self.resizing_corner_index] = snapped_mouse
                    elif lane_def['type'] == 'rectangle':
                        fixed_corner = self.shape_points_at_drag_start_label[(self.resizing_corner_index + 2) % 4]
                        new_rect = QRectF(fixed_corner, snapped_mouse).normalized()
                        new_shape_points_label = [new_rect.topLeft(), new_rect.topRight(), new_rect.bottomRight(), new_rect.bottomLeft()]

                elif self.current_selection_mode == "skewing_edge":
                    delta = current_mouse_pos_ls - self.initial_mouse_pos_for_shape_drag_label
                    snapped_mouse = self.snap_point_to_grid(current_mouse_pos_ls)
                    modifiers = event.modifiers()
                    
                    if modifiers & (Qt.ControlModifier | Qt.MetaModifier): # CMD/CTRL for Trapezoid Skew
                        edge_points_indices = {0: (0, 1), 1: (3, 2), 2: (0, 3), 3: (1, 2)}[self.skewing_edge_index]
                        dist1 = (self.initial_mouse_pos_for_shape_drag_label - self.shape_points_at_drag_start_label[edge_points_indices[0]]).manhattanLength()
                        dist2 = (self.initial_mouse_pos_for_shape_drag_label - self.shape_points_at_drag_start_label[edge_points_indices[1]]).manhattanLength()
                        corner_to_move_idx = edge_points_indices[0] if dist1 < dist2 else edge_points_indices[1]
                        is_vertical_edge = self.skewing_edge_index >= 2
                        if is_vertical_edge: new_shape_points_label[corner_to_move_idx].setY(snapped_mouse.y())
                        else: new_shape_points_label[corner_to_move_idx].setX(snapped_mouse.x())
                    
                    else:
                        is_vertical_edge = self.skewing_edge_index >= 2
                        # --- START OF THE FIX: Correctly swap default and shift behavior ---
                        # SHIFT should be parallel slide (e.g., top edge moves left/right)
                        # DEFAULT should be perpendicular slide (e.g., top edge moves up/down)
                        if modifiers & Qt.ShiftModifier: # Parallel slide
                            if is_vertical_edge: delta.setX(0) # Left/right edges slide vertically
                            else: delta.setY(0) # Top/bottom edges slide horizontally
                        else: # Default: Perpendicular slide
                            if is_vertical_edge: delta.setY(0) # Left/right edges slide horizontally
                            else: delta.setX(0) # Top/bottom edges slide vertically
                        # --- END OF THE FIX ---
                        
                        indices_to_move = {0: (0, 1), 1: (2, 3), 2: (0, 3), 3: (1, 2)}[self.skewing_edge_index]
                        new_shape_points_label[indices_to_move[0]] = self.snap_point_to_grid(self.shape_points_at_drag_start_label[indices_to_move[0]] + delta)
                        new_shape_points_label[indices_to_move[1]] = self.snap_point_to_grid(self.shape_points_at_drag_start_label[indices_to_move[1]] + delta)

                if new_shape_points_label:
                    self.live_view_label.drag_preview_quad_points = new_shape_points_label
                    self.update_live_view()

            def handle_drag_release(self, event):
                if self.current_selection_mode in ["dragging_shape", "resizing_corner", "skewing_edge"] and event.button() == Qt.LeftButton:
                    final_points = self.live_view_label.drag_preview_quad_points
                    if not final_points: return

                    self.save_state()
                    
                    # --- START OF THE FIX: Correctly handle duplication on release ---
                    if self.is_duplicating_shape:
                        new_lane_id = len(self.multi_lane_definitions) + 1
                        new_def = {'type': 'quad', 'points_label': final_points, 'id': new_lane_id}
                        self.multi_lane_definitions.append(new_def)
                    # --- END OF THE FIX ---
                    
                    elif self.moving_multi_lane_index >= 0:
                        lane = self.multi_lane_definitions[self.moving_multi_lane_index]
                        if self.current_selection_mode == "skewing_edge":
                            lane['type'] = 'quad'
                            lane['points_label'] = final_points
                        elif lane['type'] == 'quad':
                            lane['points_label'] = final_points
                        elif lane['type'] == 'rectangle':
                            new_rect = QPolygonF(final_points).boundingRect()
                            lane['points_label'] = [new_rect]
                    
                    self.is_modified = True
                    self.live_view_label.drag_preview_quad_points = None
                    self.shape_points_at_drag_start_label = []; self.resizing_corner_index = -1
                    if hasattr(self, 'skewing_edge_index'): self.skewing_edge_index = -1
                    self.is_duplicating_shape = False # Reset duplication flag
                    
                    self.current_selection_mode = "select_for_move"; self.live_view_label.mode = "select_for_move"
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_area_selection_click
                    self.update_live_view()
            
            def cancel_selection_or_move_mode(self):
                self.current_selection_mode = None
                self.live_view_label.mode = None 
                self.moving_multi_lane_index = -1
                self.resizing_corner_index = -1
                self._reset_live_view_label_custom_handlers() # Use helper
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                self.live_view_label.draw_edges = True
                self.update_live_view()
                
                
                     
            def get_nearest_point(self, mouse_pos, points):
                """Get the nearest point to the mouse position."""
                min_distance = float('inf')
                nearest_point = None
                for point in points:
                    distance = (mouse_pos - point).manhattanLength()
                    if distance < min_distance:
                        min_distance = distance
                        nearest_point = point
                return nearest_point
            
            def open_table_window(self):
                # --- START OF THE FIX ---
                # Check if an instance exists and is still a valid Qt widget.
                # The 'isinstance' check prevents errors if the C++ part of the object was deleted.
                if hasattr(self, 'table_window_instance') and self.table_window_instance and isinstance(self.table_window_instance, QDialog):
                    # If it exists, bring it to the front instead of creating a new one.
                    self.table_window_instance.raise_()
                    self.table_window_instance.activateWindow()
                    return # Exit the function to prevent creating a duplicate window
                # --- END OF THE FIX ---
                
                # The rest of the logic to gather data and create the window is correct.
                is_multi_lane_results = bool(self.latest_multi_lane_peak_areas)

                peak_areas_data_for_table = None
                quantities_data_for_table = None
                peak_details_for_table = None

                if is_multi_lane_results:
                    peak_areas_data_for_table = self.latest_multi_lane_peak_areas 
                    quantities_data_for_table = self.latest_multi_lane_calculated_quantities
                    peak_details_for_table = self.latest_multi_lane_peak_details
                else: 
                    peak_areas_data_for_table = {1: self.latest_peak_areas} if self.latest_peak_areas else {}
                    quantities_data_for_table = {1: self.latest_calculated_quantities} if self.latest_calculated_quantities else {}
                    peak_details_for_table = {1: self.latest_peak_details} if self.latest_peak_details else {}
                
                standard_dict_to_show_current = self.quantities_peak_area_dict
                is_standard_mode_current = len(standard_dict_to_show_current) >= 2
            
                self.table_window_instance = TableWindow(
                    peak_areas_data_for_table, 
                    standard_dict_to_show_current,
                    is_standard_mode_current,
                    quantities_data_for_table,
                    self,
                    peak_details_data=peak_details_for_table
                )

                self.table_window_instance.finished.connect(self._on_table_window_closed)
                self.table_window_instance.show()
            
            def snap_point_to_grid(self, point: QPointF) -> QPointF:
                """Snaps a QPointF to the grid if enabled."""
                snapped_x = point.x()
                snapped_y = point.y()
    
                grid_size = 0
                snap_x_enabled = False
                snap_y_enabled = False
    
                if hasattr(self, 'grid_size_input'):
                    grid_size = self.grid_size_input.value()
                if hasattr(self, 'show_grid_checkbox_x'):
                    snap_x_enabled = self.show_grid_checkbox_x.isChecked()
                if hasattr(self, 'show_grid_checkbox_y'):
                    snap_y_enabled = self.show_grid_checkbox_y.isChecked()
    
                if grid_size > 0:
                    if snap_x_enabled:
                        snapped_x = round(point.x() / grid_size) * grid_size
                    if snap_y_enabled:
                        snapped_y = round(point.y() / grid_size) * grid_size
                
                return QPointF(snapped_x, snapped_y)
            
            def start_rectangle(self, event):
                """Record the start position of the rectangle (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"]:
                    start_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_start_point = self.snap_point_to_grid(start_point_transformed) # Snap it
                    
                    # --- START OF THE FIX ---
                    # Populate the modern preview data structure that paintEvent uses.
                    self.live_view_label.current_preview_points = [snapped_start_point, snapped_start_point]
                    # Clear the old variables to avoid confusion.
                    self.live_view_label.rectangle_start = None 
                    self.live_view_label.bounding_box_preview = None
                    # --- END OF THE FIX ---
            
            def update_rectangle_preview(self, event):
                """Update the rectangle preview as the mouse moves (Works for 'rectangle' and 'auto_lane_rect' modes)."""
                # --- START OF THE FIX ---
                # This entire function is replaced to use the new preview data structure.
                if self.live_view_label.mode in ["rectangle", "auto_lane_rect"] and self.live_view_label.current_preview_points:
                    if event.buttons() & Qt.LeftButton:
                        current_end_point_transformed = self.live_view_label.transform_point(event.position())
                        snapped_end_point = self.snap_point_to_grid(current_end_point_transformed) # Snap it
        
                        # Update the end point of the preview list.
                        if len(self.live_view_label.current_preview_points) == 2:
                            self.live_view_label.current_preview_points[1] = snapped_end_point
                        
                        self.update_live_view()
            
            def finalize_rectangle(self, event):
                """Finalize the rectangle when the mouse is released."""
                if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
                    end_point_transformed = self.live_view_label.transform_point(event.position())
                    snapped_end_point = self.snap_point_to_grid(end_point_transformed) # Snap it
    
                    self.live_view_label.rectangle_end = snapped_end_point
                    # self.live_view_label.rectangle_start is already snapped
                    self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, snapped_end_point]
                    
                    self.live_view_label.bounding_box_preview = (
                        self.live_view_label.rectangle_start.x(),
                        self.live_view_label.rectangle_start.y(),
                        snapped_end_point.x(), # Use snapped end point
                        snapped_end_point.y(), # Use snapped end point
                    )
                    
                    self.live_view_label.mode = None
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    self.update_live_view()
                
            def _reset_to_selection_mode(self):
                """
                Resets the interaction state back to 'select_for_move' without fully cancelling the mode.
                This is used after an action (like processing a standard) interrupts the mouse event flow.
                """
                self.current_selection_mode = "select_for_move"
                self.live_view_label.mode = "select_for_move"

                # Clear any temporary drag/resize state variables
                self.shape_points_at_drag_start_label = []
                self.initial_mouse_pos_for_shape_drag_label = QPointF()
                self.resizing_corner_index = -1
                self.is_duplicating_shape = False
                self.duplication_source_info = None

                # Reset the event handlers to allow for a new selection click
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.handle_area_selection_click
                self.live_view_label.setCursor(Qt.ArrowCursor)

                # Refresh the view to clear any selection-specific highlights
                self.update_live_view()

            def process_standard(self):
                extracted_qimage = None
                region_type_for_message = ""
                
                # --- Get the fully adjusted image for analysis ---
                adjusted_master = self._get_fully_adjusted_image_for_analysis()
                if not adjusted_master or adjusted_master.isNull():
                    QMessageBox.warning(self, "Error", "Could not get adjusted image for analysis.")
                    return

                # --- START: New, Robust Logic to Find the Target Region ---
                target_lane_def = None # This will hold the definition of the one region we will process.

                # Case 1: A specific multi-lane region is already selected.
                if hasattr(self, 'moving_multi_lane_index') and self.moving_multi_lane_index >= 0:
                    if self.moving_multi_lane_index < len(self.multi_lane_definitions):
                        target_lane_def = self.multi_lane_definitions[self.moving_multi_lane_index]
                        region_type_for_message = f"Selected Multi-Lane {target_lane_def['id']}"

                # Case 2: No selection, check for single defined regions.
                else:
                    is_single_quad = len(self.live_view_label.quad_points) == 4
                    is_single_rect = self.live_view_label.bounding_box_preview is not None
                    has_multi_lanes = bool(self.multi_lane_definitions)

                    if has_multi_lanes and len(self.multi_lane_definitions) > 1:
                        QMessageBox.warning(self, "Ambiguous Selection", "Multiple analysis regions are defined. Please use 'Move/Resize Area' to click on the single lane you want to process as a standard.")
                        return
                    elif has_multi_lanes and len(self.multi_lane_definitions) == 1:
                        target_lane_def = self.multi_lane_definitions[0]
                        region_type_for_message = f"Single Defined Multi-Lane {target_lane_def['id']}"
                    elif is_single_quad:
                        target_lane_def = {'type': 'quad', 'points_label': self.live_view_label.quad_points}
                        region_type_for_message = "Defined Single Quadrilateral"
                    elif is_single_rect:
                        rect_ls = QRectF(QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]),
                                         QPointF(self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3])).normalized()
                        target_lane_def = {'type': 'rectangle', 'points_label': [rect_ls]}
                        region_type_for_message = "Defined Single Rectangle"

                if not target_lane_def:
                    QMessageBox.warning(self, "Input Error", "Please define an analysis region first.")
                    return
                # --- END: New Logic ---

                # --- Extract the image data from the identified target region ---
                print(f"Processing Standard: {region_type_for_message}")
                if target_lane_def['type'] == 'quad':
                    extracted_qimage = self.quadrilateral_to_rect(adjusted_master, target_lane_def['points_label'])
                elif target_lane_def['type'] == 'rectangle':
                    img_coords_rect = self._map_label_rect_to_image_rect(target_lane_def['points_label'][0])
                    if img_coords_rect:
                        extracted_qimage = adjusted_master.copy(*img_coords_rect)

                if not extracted_qimage or extracted_qimage.isNull():
                    QMessageBox.warning(self, "Error", f"Could not extract or warp the defined region: {region_type_for_message}.")
                    return

                # --- Final analysis step (unchanged) ---
                processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                if processed_data_pil:
                    self.analyze_bounding_box(processed_data_pil, standard=True)
                else:
                    QMessageBox.warning(self, "Error", f"Could not convert {region_type_for_message} to grayscale for analysis.")
                
                self._reset_to_selection_mode()
            
            def process_sample(self):
                extracted_qimage = None
                
                adjusted_master = self._get_fully_adjusted_image_for_analysis()
                if not adjusted_master or adjusted_master.isNull():
                    QMessageBox.warning(self, "Error", "Could not get adjusted image for analysis.")
                    return
                
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                self.update_live_view()  
                
                self.latest_peak_areas = [] 
                self.latest_peak_details = []
                self.latest_calculated_quantities = []
                self.latest_multi_lane_peak_areas.clear()
                self.latest_multi_lane_peak_details.clear()
                self.latest_multi_lane_calculated_quantities.clear()

                extracted_regions_info = []

                # --- START OF THE FIX ---
                # New, more robust logic to decide which regions to process.
                # It no longer depends on the faulty 'multi_lane_processing_finished' flag after loading.

                # Priority 1: If multi-lane definitions exist, always use them.
                if self.multi_lane_definitions:
                    print(f"Processing {len(self.multi_lane_definitions)} multiple lanes as samples.")
                    for lane_def in self.multi_lane_definitions:
                        extracted_qimage = None
                        if lane_def['type'] == 'quad':
                            extracted_qimage = self.quadrilateral_to_rect(adjusted_master, lane_def['points_label'])
                        elif lane_def['type'] == 'rectangle':
                            img_coords_rect = self._map_label_rect_to_image_rect(lane_def['points_label'][0])
                            if img_coords_rect and adjusted_master:
                                extracted_qimage = adjusted_master.copy(*img_coords_rect)
                        
                        if extracted_qimage and not extracted_qimage.isNull():
                            pil_for_this_lane = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                            if pil_for_this_lane:
                                extracted_regions_info.append({'pil': pil_for_this_lane, 'id': lane_def['id']})
                            else:
                                QMessageBox.warning(self, "Error", f"Could not convert Lane {lane_def['id']} to PIL for analysis.")
                        else:
                             QMessageBox.warning(self, "Error", f"Could not extract/warp Lane {lane_def['id']}.")
                
                # Priority 2: If no multi-lanes, check for a single defined quad.
                elif len(self.live_view_label.quad_points) == 4: 
                    print("Processing Sample: Single Quadrilateral")
                    extracted_qimage = self.quadrilateral_to_rect(adjusted_master, self.live_view_label.quad_points)
                    if extracted_qimage and not extracted_qimage.isNull():
                        pil_img = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                        if pil_img:
                             extracted_regions_info.append({'pil': pil_img, 'id': 1})
                        else: QMessageBox.warning(self, "Error", "Could not convert quad region to PIL.")
                    else: QMessageBox.warning(self, "Error", "Quadrilateral warping failed for sample.")

                # Priority 3: If still nothing, check for a single defined rectangle.
                elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4: 
                    print("Processing Sample: Single Rectangle")
                    try:
                        img_coords_rect = self._map_label_rect_to_image_rect(QRectF(
                            QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]),
                            QPointF(self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3])
                        ).normalized())
                        if img_coords_rect and adjusted_master:
                            extracted_qimage = adjusted_master.copy(*img_coords_rect)
                            if extracted_qimage and not extracted_qimage.isNull():
                                pil_img = self.convert_qimage_to_grayscale_pil(extracted_qimage)
                                if pil_img:
                                     extracted_regions_info.append({'pil': pil_img, 'id': 1})
                                else: QMessageBox.warning(self, "Error", "Could not convert rect region to PIL.")
                            else: raise ValueError("QImage.copy failed for rectangle.")
                        else: raise ValueError("Failed to map rectangle to image coords or image is missing.")
                    except Exception as e:
                         print(f"Error processing rectangle region for sample: {e}"); QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                
                # Priority 4: If no regions of any kind are defined, show the error message.
                else:
                    QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first, or finish defining multiple lanes.")
                    return
                # --- END OF THE FIX ---

                all_lanes_text_results = []
                if not extracted_regions_info:
                    QMessageBox.warning(self, "Analysis Error", "No regions were successfully extracted for analysis.")
                    return

                model_to_use = "Linear"
                std_qtys = list(self.quantities_peak_area_dict.keys())
                std_areas = list(self.quantities_peak_area_dict.values())
                
                for region_info in extracted_regions_info:
                    lane_id = region_info['id']
                    pil_image_for_dialog = region_info['pil'] 
                    peak_info_for_lane = self.calculate_peak_area(pil_image_for_dialog) 

                    if peak_info_for_lane and len(peak_info_for_lane) > 0:
                        areas_for_this_lane = [float(round(info['area'], 3)) for info in peak_info_for_lane]
                        self.latest_multi_lane_peak_areas[lane_id] = areas_for_this_lane
                        self.latest_multi_lane_peak_details[lane_id] = peak_info_for_lane 

                        if len(self.quantities_peak_area_dict) >= 2:
                            quantities_for_lane, _ = self._perform_quantification(model_to_use, std_qtys, std_areas, areas_for_this_lane)
                            self.latest_multi_lane_calculated_quantities[lane_id] = quantities_for_lane
                            formatted_areas = ', '.join([f"{a:.3f}" for a in areas_for_this_lane])
                            formatted_quantities = ', '.join([f"{q:.2f}" for q in quantities_for_lane])
                            all_lanes_text_results.append(f"Lane {lane_id}: Areas=[{formatted_areas}], Qty=[{formatted_quantities}]")
                        else:
                            self.latest_multi_lane_calculated_quantities[lane_id] = []
                            formatted_areas = ', '.join([f"{a:.3f}" for a in areas_for_this_lane])
                            all_lanes_text_results.append(f"Lane {lane_id}: Areas=[{formatted_areas}] (No std curve for qty)")
                    else:
                        all_lanes_text_results.append(f"Lane {lane_id}: Analysis failed or no peaks.")
                        self.latest_multi_lane_peak_areas[lane_id] = []
                        self.latest_multi_lane_peak_details[lane_id] = [] 
                        self.latest_multi_lane_calculated_quantities[lane_id] = []
                
                if 1 in self.latest_multi_lane_peak_areas:
                     self.latest_peak_areas = self.latest_multi_lane_peak_areas[1]
                if 1 in self.latest_multi_lane_peak_details: 
                     self.latest_peak_details = self.latest_multi_lane_peak_details[1]
                if 1 in self.latest_multi_lane_calculated_quantities:
                     self.latest_calculated_quantities = self.latest_multi_lane_calculated_quantities[1]
                
                if all_lanes_text_results:
                    self.target_protein_areas_text.setText("\n".join(all_lanes_text_results))
                else:
                    self.target_protein_areas_text.setText("N/A or analysis failed for all lanes.")
                
                self._reset_to_selection_mode()
                self.update_live_view()
                
            def _map_label_rect_to_image_rect(self, rect_label_space: QRectF):
                if not self.image or self.image.isNull(): return None
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                if not (img_w > 0 and img_h > 0 and label_w > 0 and label_h > 0): return None

                scale_factor = min(label_w / img_w, label_h / img_h)
                display_offset_x = (label_w - img_w * scale_factor) / 2.0
                display_offset_y = (label_h - img_h * scale_factor) / 2.0

                start_x_img = (rect_label_space.left() - display_offset_x) / scale_factor
                start_y_img = (rect_label_space.top() - display_offset_y) / scale_factor
                end_x_img = (rect_label_space.right() - display_offset_x) / scale_factor
                end_y_img = (rect_label_space.bottom() - display_offset_y) / scale_factor
                
                x_img = int(min(start_x_img, end_x_img))
                y_img = int(min(start_y_img, end_y_img))
                w_img = int(abs(end_x_img - start_x_img))
                h_img = int(abs(end_y_img - start_y_img))

                x_clamped = max(0, x_img)
                y_clamped = max(0, y_img)
                w_clamped = max(1, min(w_img, img_w - x_clamped))
                h_clamped = max(1, min(h_img, img_h - y_clamped))
                return (x_clamped, y_clamped, w_clamped, h_clamped)

            # Helper: _map_label_points_to_image_points
            def _map_label_points_to_image_points(self, points_label_space: list): # list of QPointF
                if not self.image or self.image.isNull() or not points_label_space: return None
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                if not (img_w > 0 and img_h > 0 and label_w > 0 and label_h > 0): return None

                scale_factor = min(label_w / img_w, label_h / img_h)
                display_offset_x = (label_w - img_w * scale_factor) / 2.0
                display_offset_y = (label_h - img_h * scale_factor) / 2.0
                
                image_points = []
                for p_label in points_label_space:
                    x_img = (p_label.x() - display_offset_x) / scale_factor
                    y_img = (p_label.y() - display_offset_y) / scale_factor
                    image_points.append(QPointF(x_img, y_img))
                return image_points


            def convert_qimage_to_grayscale_pil(self, qimg):
                """
                Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
                for use with PeakAreaDialog. Handles 64-bit (16-bit RGBA) correctly by averaging channels.
                """
                if not qimg or qimg.isNull():
                    return None

                fmt = qimg.format()
                
                try:
                    # 1. Handle 64-bit / 16-bit RGBA explicitly
                    if fmt in [QImage.Format_RGBA64, QImage.Format_RGBX64]:
                        np_array = self.qimage_to_numpy(qimg) # Returns (H, W, 4) uint16
                        if np_array is not None and np_array.ndim == 3 and np_array.shape[2] == 4:
                            # Convert 16-bit RGBA to 16-bit Grayscale
                            # We take the max or mean of RGB channels, ignoring Alpha (index 3)
                            # Using max helps preserve faint bands in gels.
                            # Slice: [All rows, All cols, First 3 channels (R,G,B)]
                            rgb_data = np_array[..., :3]
                            grayscale_16 = np.max(rgb_data, axis=2).astype(np.uint16)
                            return Image.fromarray(grayscale_16, mode='I;16')

                    # 2. Already grayscale? Convert directly if possible.
                    if fmt == QImage.Format_Grayscale16:
                        np_array = self.qimage_to_numpy(qimg)
                        if np_array is not None and np_array.dtype == np.uint16:
                            return Image.fromarray(np_array, mode='I;16')
                    elif fmt == QImage.Format_Grayscale8:
                        np_array = self.qimage_to_numpy(qimg)
                        if np_array is not None and np_array.dtype == np.uint8:
                            return Image.fromarray(np_array, mode='L')
                    
                    # 3. Fallback: Color 8-bit or other formats
                    # Use NumPy for robust conversion
                    np_img = self.qimage_to_numpy(qimg)
                    if np_img is None: raise ValueError("NumPy conversion failed for color.")
                    
                    if np_img.ndim == 3:
                        # Convert using OpenCV (Standard BGR/RGB -> Gray)
                        # Note: qimage_to_numpy returns RGBA order usually for Qt6
                        # Simple average or cv2 conversion
                        gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) 
                        
                        # If we want to simulate 16-bit range for consistent processing downstream
                        gray_np_16bit = (gray_np.astype(np.float32) / 255.0 * 65535.0).astype(np.uint16)
                        return Image.fromarray(gray_np_16bit, mode='I;16')
                        
                    elif np_img.ndim == 2: 
                         # Assume uint8 or other, convert to L
                         return Image.fromarray(np_img).convert('L')

                    raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

                except Exception as e:
                    print(f"Error converting QImage to Grayscale PIL: {e}")
                    traceback.print_exc()
                    return None     
                
            def combine_image_tab(self):
                tab = QWidget()
                main_layout = QVBoxLayout(tab)
                main_layout.setSpacing(10)

                # --- Top Row: Side-by-Side Overlay Controls ---
                top_row_layout = QHBoxLayout()
                
                # --- Image 1 (Base) Controls ---
                image1_group = QGroupBox("Image 1 (Base)")
                image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image1_layout = QGridLayout(image1_group)

                copy_image1_button = QPushButton("Copy Current"); copy_image1_button.clicked.connect(self.save_image1)
                place_image1_button = QPushButton("Place"); place_image1_button.clicked.connect(self.place_image1)
                remove_image1_button = QPushButton("Remove"); remove_image1_button.clicked.connect(lambda: (self.remove_image1(), self.update_live_view()))
                reset_overlay1_button = QPushButton("Reset"); reset_overlay1_button.clicked.connect(self.reset_overlay1_transform)

                image1_layout.addWidget(copy_image1_button, 0, 0, 1, 1)
                image1_layout.addWidget(place_image1_button, 0, 1, 1, 1)
                image1_layout.addWidget(remove_image1_button, 1, 0, 1, 1)
                image1_layout.addWidget(reset_overlay1_button, 1, 1, 1, 1)

                image1_layout.addWidget(QLabel("X:"), 2, 0)
                self.image1_left_slider = QSlider(Qt.Horizontal); self.image1_left_slider.valueChanged.connect(lambda: self._update_overlay_position_from_sliders()); self.image1_left_slider.valueChanged.connect(lambda: self.image1_left_slider.setFocus())
                image1_layout.addWidget(self.image1_left_slider, 2, 1)
                self.image1_pos_x_label = QLabel("0"); self.image1_pos_x_label.setFixedWidth(40)
                self.image1_left_slider.valueChanged.connect(lambda val, lbl=self.image1_pos_x_label: lbl.setText(str(val)))
                image1_layout.addWidget(self.image1_pos_x_label, 2, 2)
                
                image1_layout.addWidget(QLabel("Y:"), 3, 0)
                self.image1_top_slider = QSlider(Qt.Horizontal); self.image1_top_slider.valueChanged.connect(lambda: self._update_overlay_position_from_sliders()); self.image1_top_slider.valueChanged.connect(lambda: self.image1_top_slider.setFocus())
                image1_layout.addWidget(self.image1_top_slider, 3, 1)
                self.image1_pos_y_label = QLabel("0"); self.image1_pos_y_label.setFixedWidth(40)
                self.image1_top_slider.valueChanged.connect(lambda val, lbl=self.image1_pos_y_label: lbl.setText(str(val)))
                image1_layout.addWidget(self.image1_pos_y_label, 3, 2)

                image1_layout.addWidget(QLabel("Resize:"), 4, 0)
                self.image1_resize_slider = QSlider(Qt.Horizontal); self.image1_resize_slider.setRange(10, 300); self.image1_resize_slider.setValue(100); self.image1_resize_slider.valueChanged.connect(lambda: self.update_live_view()); self.image1_resize_slider.valueChanged.connect(lambda:  self.image1_resize_slider.setFocus())
                image1_layout.addWidget(self.image1_resize_slider, 4, 1)
                self.image1_resize_label = QLabel("100%"); self.image1_resize_label.setFixedWidth(40)
                self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
                image1_layout.addWidget(self.image1_resize_label, 4, 2)

                image1_layout.addWidget(QLabel("Rotate:"), 5, 0)
                self.image1_rotation_slider = QSlider(Qt.Horizontal); self.image1_rotation_slider.setRange(-1800, 1800); self.image1_rotation_slider.setValue(0); self.image1_rotation_slider.valueChanged.connect(lambda: self.update_live_view()); self.image1_rotation_slider.valueChanged.connect(lambda:  self.image1_rotation_slider.setFocus())
                image1_layout.addWidget(self.image1_rotation_slider, 5, 1)
                self.image1_rotation_label = QLabel("0.0°"); self.image1_rotation_label.setFixedWidth(40)
                self.image1_rotation_slider.valueChanged.connect(lambda val, lbl=self.image1_rotation_label: lbl.setText(f"{val/10.0:.1f}°"))
                image1_layout.addWidget(self.image1_rotation_label, 5, 2)
                
                image1_layout.setColumnStretch(1, 1)
                top_row_layout.addWidget(image1_group)

                # --- Image 2 (Overlay) Controls ---
                image2_group = QGroupBox("Image 2 (Overlay)")
                image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                image2_layout = QGridLayout(image2_group)

                copy_image2_button = QPushButton("Copy Current"); copy_image2_button.clicked.connect(self.save_image2)
                place_image2_button = QPushButton("Place"); place_image2_button.clicked.connect(self.place_image2)
                remove_image2_button = QPushButton("Remove"); remove_image2_button.clicked.connect(lambda: (self.remove_image2(), self.update_live_view()))
                reset_overlay2_button = QPushButton("Reset"); reset_overlay2_button.clicked.connect(self.reset_overlay2_transform)

                image2_layout.addWidget(copy_image2_button, 0, 0, 1, 1)
                image2_layout.addWidget(place_image2_button, 0, 1, 1, 1)
                image2_layout.addWidget(remove_image2_button, 1, 0, 1, 1)
                image2_layout.addWidget(reset_overlay2_button, 1, 1, 1, 1)

                image2_layout.addWidget(QLabel("X:"), 2, 0)
                self.image2_left_slider = QSlider(Qt.Horizontal); self.image2_left_slider.valueChanged.connect(lambda: self._update_overlay_position_from_sliders()); self.image2_left_slider.valueChanged.connect(lambda: self.image2_left_slider.setFocus())
                image2_layout.addWidget(self.image2_left_slider, 2, 1)
                self.image2_pos_x_label = QLabel("0"); self.image2_pos_x_label.setFixedWidth(40)
                self.image2_left_slider.valueChanged.connect(lambda val, lbl=self.image2_pos_x_label: lbl.setText(str(val)))
                image2_layout.addWidget(self.image2_pos_x_label, 2, 2)

                image2_layout.addWidget(QLabel("Y:"), 3, 0)
                self.image2_top_slider = QSlider(Qt.Horizontal); self.image2_top_slider.valueChanged.connect(lambda: self._update_overlay_position_from_sliders()); self.image2_top_slider.valueChanged.connect(lambda: self.image2_top_slider.setFocus())
                image2_layout.addWidget(self.image2_top_slider, 3, 1)
                self.image2_pos_y_label = QLabel("0"); self.image2_pos_y_label.setFixedWidth(40)
                self.image2_top_slider.valueChanged.connect(lambda val, lbl=self.image2_pos_y_label: lbl.setText(str(val)))
                image2_layout.addWidget(self.image2_pos_y_label, 3, 2)
                
                image2_layout.addWidget(QLabel("Resize:"), 4, 0)
                self.image2_resize_slider = QSlider(Qt.Horizontal); self.image2_resize_slider.setRange(10, 300); self.image2_resize_slider.setValue(100); self.image2_resize_slider.valueChanged.connect(lambda: self.update_live_view()); self.image2_resize_slider.valueChanged.connect(lambda: self.image2_resize_slider.setFocus())
                image2_layout.addWidget(self.image2_resize_slider, 4, 1)
                self.image2_resize_label = QLabel("100%"); self.image2_resize_label.setFixedWidth(40)
                self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
                image2_layout.addWidget(self.image2_resize_label, 4, 2)

                image2_layout.addWidget(QLabel("Rotate:"), 5, 0)
                self.image2_rotation_slider = QSlider(Qt.Horizontal); self.image2_rotation_slider.setRange(-1800, 1800); self.image2_rotation_slider.setValue(0); self.image2_rotation_slider.valueChanged.connect(lambda: self.update_live_view());self.image2_rotation_slider.valueChanged.connect(lambda: self.image2_rotation_slider.setFocus())
                image2_layout.addWidget(self.image2_rotation_slider, 5, 1)
                self.image2_rotation_label = QLabel("0.0°"); self.image2_rotation_label.setFixedWidth(40)
                self.image2_rotation_slider.valueChanged.connect(lambda val, lbl=self.image2_rotation_label: lbl.setText(f"{val/10.0:.1f}°"))
                image2_layout.addWidget(self.image2_rotation_label, 5, 2)

                image2_layout.setColumnStretch(1, 1)
                top_row_layout.addWidget(image2_group)
                main_layout.addLayout(top_row_layout)

                # --- Middle Section: Global Controls ---
                global_controls_group = QGroupBox("Global Overlay Controls")
                global_controls_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                global_controls_layout = QGridLayout(global_controls_group)
                
                global_controls_layout.addWidget(QLabel("Set Base:"), 0, 0)
                set_base_button = QPushButton("Set Current as Base (Image 1)")
                set_base_button.setToolTip("Copies the main image to the 'Image 1' buffer and places it as the bottom layer.")
                set_base_button.clicked.connect(self.set_overlay_base)
                global_controls_layout.addWidget(set_base_button, 0, 1, 1, 2)

                self.load_overlay_button = QPushButton("Load Overlay (Image 2)")
                self.load_overlay_button.setToolTip("Loads a second image from a file into the 'Image 2' buffer and places it as the top layer.")
                self.load_overlay_button.clicked.connect(self.load_overlay_image)
                self.load_overlay_button.setEnabled(False)
                global_controls_layout.addWidget(self.load_overlay_button, 0, 3, 1, 2)
                
                # --- Mixing Controls ---
                # FIX: Updated Label text to be explicit about mixing
                global_controls_layout.addWidget(QLabel("Mixing % (Overlay):"), 1, 0)
                self.blend_slider = QSlider(Qt.Horizontal)
                self.blend_slider.setRange(0, 100); self.blend_slider.setValue(50)
                self.blend_slider.setToolTip("Sets the mixing percentage for Image 2 (Overlay).\n0% = Overlay invisible (100% Base)\n50% = 50% Overlay / 50% Base\n100% = Overlay opaque (0% Base)")
                self.blend_slider.valueChanged.connect(lambda: self.update_live_view()); self.blend_slider.valueChanged.connect(lambda: self.blend_slider.setFocus())
                global_controls_layout.addWidget(self.blend_slider, 1, 1, 1, 4)

                self.interactive_overlay_button = QPushButton("Activate Interactive Alignment"); self.interactive_overlay_button.setCheckable(True); self.interactive_overlay_button.clicked.connect(self.toggle_interactive_overlay_mode)
                finalize_button = QPushButton("Rasterize Image"); finalize_button.clicked.connect(self.finalize_combined_image)
                global_controls_layout.addWidget(self.interactive_overlay_button, 2, 1, 1, 2)
                global_controls_layout.addWidget(finalize_button, 2, 3, 1, 2)

                global_controls_layout.setColumnStretch(1, 1); global_controls_layout.setColumnStretch(3, 1)
                main_layout.addWidget(global_controls_group)

                main_layout.addStretch()
                self._update_overlay_slider_ranges()
                return tab
            
            def set_overlay_base(self):
                """Convenience method to copy current image and place it as the base (Image 1)."""
                if self.image_master and not self.image_master.isNull():
                    self.save_image1()
                    self.place_image1()
                    self.adjustment_context_combo.model().item(1).setEnabled(True)
                    self.adjustment_context_combo.setCurrentText("Overlay 1 (Base)")
                    if hasattr(self, 'load_overlay_button'):
                        self.load_overlay_button.setEnabled(True)
                    QMessageBox.information(self, "Success", "Current image set as the base (Image 1).")
                else:
                    QMessageBox.warning(self, "Error", "No image is loaded to set as base.")

            def load_overlay_image(self):
                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getOpenFileName(self, "Open Overlay Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options)
                if not file_path: return

                overlay_image = QImage(file_path)
                if overlay_image.isNull():
                    try:
                        pil_image = Image.open(file_path); np_array = np.array(pil_image)
                        overlay_image = self.numpy_to_qimage(np_array)
                        if overlay_image.isNull(): raise ValueError("Pillow/NumPy conversion failed.")
                    except Exception as e:
                        QMessageBox.warning(self, "Error", f"Failed to load overlay image: {e}")
                        return
                
                # --- Handle main image transparency setup (existing code) ---
                if hasattr(self, 'image1_original') and self.image1_original and not self.image1_original.isNull():
                    width = self.image_master.width()
                    height = self.image_master.height()
                    transparent_canvas = QImage(width, height, QImage.Format_ARGB32_Premultiplied)
                    transparent_canvas.fill(Qt.transparent)
                    
                    self.image = transparent_canvas
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()

                    default_settings = self._get_default_adjustments()
                    self.channel_mixer_data = default_settings['channel_mixer'].copy()
                    self.unsharp_mask_data = default_settings['unsharp_mask'].copy()
                    self.clahe_data = default_settings['clahe'].copy()
                    self.main_image_is_inverted = False

                    if self.adjustment_context == "Main Image":
                        self._load_adjustments_to_ui("Main Image")

                    print("INFO: Main image display replaced with a transparent canvas for overlay-only view.")

                self.image2_adjustments = self._get_default_adjustments()
                self.image2_original = overlay_image

                # --- ADDED: Auto-invert 16-bit overlay ---
                if self.image2_original.format() == QImage.Format_Grayscale16:
                    self.image2_adjustments['is_inverted'] = True
                # -----------------------------------------

                self._update_overlay_preview(2)
                self.place_image2()
                self.adjustment_context_combo.model().item(2).setEnabled(True)
                self.adjustment_context_combo.setCurrentText("Overlay 2 (Overlay)")
                QMessageBox.information(self, "Success", "Overlay image loaded (Image 2). Use 'Interactive Alignment' to position it.")
            
            def remove_image1(self):
                """Hides Image 1 and resets its sliders, without triggering a redraw."""
                if hasattr(self, 'image1_position'):
                    del self.image1_position
                if hasattr(self, 'image1_original'):
                    del self.image1_original

                self.image1_adjusted_preview = None
                self.image1_adjustments = {}

                if hasattr(self, 'load_overlay_button'):
                    self.load_overlay_button.setEnabled(False)
                    
                if hasattr(self, 'adjustment_context_combo'):
                    self.adjustment_context_combo.model().item(1).setEnabled(False)
                    if self.adjustment_context_combo.currentText() == "Overlay 1 (Base)":
                        self.adjustment_context_combo.setCurrentText("Main Image")

                self.reset_overlay1_transform()
            
            def remove_image2(self):
                """Hides Image 2 and resets its sliders, without triggering a redraw."""
                if hasattr(self, 'image2_position'):
                    del self.image2_position
                if hasattr(self, 'image2_original'):
                    del self.image2_original
                
                self.image2_adjusted_preview = None
                self.image2_adjustments = {}
                
                if hasattr(self, 'adjustment_context_combo'):
                    self.adjustment_context_combo.model().item(2).setEnabled(False)
                    if self.adjustment_context_combo.currentText() == "Overlay 2 (Overlay)":
                        self.adjustment_context_combo.setCurrentText("Main Image")

                self.reset_overlay2_transform()
            
            def save_image1(self):
                if self.image_master and not self.image_master.isNull():
                    if hasattr(self, 'image1_position'):
                        del self.image1_position
                    
                    # Capture current adjustments from main image
                    current_main_image_adjustments = {
                        'is_inverted': self.main_image_is_inverted,
                        'levels_gamma': {
                            'black_point': self.black_point_slider.value(),
                            'white_point': self.white_point_slider.value(),
                            'gamma': self.gamma_slider.value()
                        },
                        'channel_mixer': self.channel_mixer_data.copy(),
                        'unsharp_mask': self.unsharp_mask_data.copy(),
                        'clahe': self.clahe_data.copy()
                    }
                    self.image1_adjustments = current_main_image_adjustments
                    
                    self.image1_original = self.image_master.copy()
                    self._update_overlay_preview(1)
                    self.reset_overlay1_transform()
                    self.update_live_view() 
                    QMessageBox.information(self, "Success", "Image 1 copied to buffer.")
                else:
                    QMessageBox.warning(self, "Error", "No valid master image to copy.")

            def save_image2(self):
                if self.image_master and not self.image_master.isNull():
                    if hasattr(self, 'image2_position'):
                        del self.image2_position

                    current_main_image_adjustments = {
                        'is_inverted': self.main_image_is_inverted,
                        'levels_gamma': {
                            'black_point': self.black_point_slider.value(),
                            'white_point': self.white_point_slider.value(),
                            'gamma': self.gamma_slider.value()
                        },
                        'channel_mixer': self.channel_mixer_data.copy(),
                        'unsharp_mask': self.unsharp_mask_data.copy(),
                        'clahe': self.clahe_data.copy()
                    }
                    self.image2_adjustments = current_main_image_adjustments
                    
                    self.image2_original = self.image_master.copy()
                    self._update_overlay_preview(2)
                    self.reset_overlay2_transform()
                    self.update_live_view()
                    QMessageBox.information(self, "Success", "Image 2 copied to buffer.")
                else:
                    QMessageBox.warning(self, "Error", "No valid master image to copy.")
            
            def place_image1(self):
                if hasattr(self, 'image1_original') and self.image1_original and not self.image1_original.isNull():
                    self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
                    self.update_live_view()
                else:
                    QMessageBox.warning(self, "Info", "No image copied to Image 1 buffer yet.")
            
            def place_image2(self):
                if hasattr(self, 'image2_original') and self.image2_original and not self.image2_original.isNull():
                    self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                    self.update_live_view()
                else:
                    QMessageBox.warning(self, "Info", "No image copied or loaded to Image 2 buffer yet.")

            def reset_overlay1_transform(self):
                """Resets position, size, and rotation for Image 1 overlay."""
                if hasattr(self, 'image1_left_slider'):
                    self.image1_left_slider.setValue(0)
                if hasattr(self, 'image1_top_slider'):
                    self.image1_top_slider.setValue(0)
                if hasattr(self, 'image1_resize_slider'):
                    self.image1_resize_slider.setValue(100)
                if hasattr(self, 'image1_rotation_slider'):
                    self.image1_rotation_slider.setValue(0)
                self.update_live_view()

            def reset_overlay2_transform(self):
                """Resets position, size, and rotation for Image 2 overlay."""
                if hasattr(self, 'image2_left_slider'):
                    self.image2_left_slider.setValue(0)
                if hasattr(self, 'image2_top_slider'):
                    self.image2_top_slider.setValue(0)
                if hasattr(self, 'image2_resize_slider'):
                    self.image2_resize_slider.setValue(100)
                if hasattr(self, 'image2_rotation_slider'):
                    self.image2_rotation_slider.setValue(0)
                self.update_live_view()
            
            
            
            def finalize_combined_image(self):
                """ Rasterizes placed overlays onto a new canvas, baking in their individual adjustments. """
                has_img1 = hasattr(self, 'image1_original') and self.image1_original and not self.image1_original.isNull() and hasattr(self, 'image1_position')
                has_img2 = hasattr(self, 'image2_original') and self.image2_original and not self.image2_original.isNull() and hasattr(self, 'image2_position')
                
                if hasattr(self, 'interactive_overlay_button'):
                    self.interactive_overlay_button.setChecked(False)
                    
                if not (has_img1 or has_img2):
                    QMessageBox.information(self, "Info", "No overlays are currently placed to rasterize.")
                    return

                self.save_state()

                # Get the adjusted previews (which contain the visual state including inversions)
                adjusted_img1 = getattr(self, 'image1_adjusted_preview', None)
                adjusted_img2 = getattr(self, 'image2_adjusted_preview', None)

                # --- 1. Determine Output Bit Depth ---
                # "If one image is 16 bit and another is 8 bit then final should be 8 bit but if both 16 bit then 16 bit"
                
                is_img1_16 = False
                if has_img1 and adjusted_img1:
                    is_img1_16 = adjusted_img1.format() in [QImage.Format_Grayscale16, QImage.Format_RGBA64, QImage.Format_RGBX64]
                
                is_img2_16 = False
                if has_img2 and adjusted_img2:
                    is_img2_16 = adjusted_img2.format() in [QImage.Format_Grayscale16, QImage.Format_RGBA64, QImage.Format_RGBX64]

                # Determine if we go high depth (16-bit)
                # If only one image exists, respect its depth. If both exist, BOTH must be 16-bit.
                target_16_bit = False
                if has_img1 and has_img2:
                    target_16_bit = is_img1_16 and is_img2_16
                elif has_img1:
                    target_16_bit = is_img1_16
                elif has_img2:
                    target_16_bit = is_img2_16

                canvas_format = QImage.Format_RGBA64 if target_16_bit else QImage.Format_ARGB32_Premultiplied
                
                # --- 2. Setup Canvas ---
                # Use master size, or if master is hidden/replaced by overlay flow, use Img1 size
                canvas_w = self.image_master.width()
                canvas_h = self.image_master.height()
                
                final_canvas = QImage(canvas_w, canvas_h, canvas_format)
                # Fill with white (standard for gels) or transparent. 
                # Transparent allows checking if alignment leaves gaps.
                final_canvas.fill(Qt.transparent) 

                painter = QPainter(final_canvas)
                painter.setRenderHint(QPainter.SmoothPixmapTransform, False) # Preserves pixel data better
                painter.setRenderHint(QPainter.Antialiasing, True)

                blend_value = self.blend_slider.value()

                # --- 3. Draw Image 1 (Base) ---
                if has_img1 and adjusted_img1:
                    # Base image drawn at 100% opacity to ensure visibility
                    painter.setOpacity(1.0)
                    
                    rect1_native = QRectF(
                        QPointF(*self.image1_position), 
                        QSizeF(adjusted_img1.width() * (self.image1_resize_slider.value()/100.0), 
                               adjusted_img1.height() * (self.image1_resize_slider.value()/100.0))
                    )
                    rotation1 = self.image1_rotation_slider.value() / 10.0
                    if abs(rotation1) > 0.01:
                        center_point = rect1_native.center()
                        painter.save(); painter.translate(center_point); painter.rotate(rotation1); painter.translate(-center_point)
                        painter.drawImage(rect1_native, adjusted_img1)
                        painter.restore()
                    else:
                        painter.drawImage(rect1_native, adjusted_img1)
                
                # --- 4. Draw Image 2 (Overlay) ---
                if has_img2 and adjusted_img2:
                    # Overlay blended on top. 0% blend = Invisible, 100% blend = Opaque
                    painter.setOpacity(blend_value / 100.0)
                    
                    rect2_native = QRectF(
                        QPointF(*self.image2_position), 
                        QSizeF(adjusted_img2.width() * (self.image2_resize_slider.value()/100.0), 
                               adjusted_img2.height() * (self.image2_resize_slider.value()/100.0))
                    )
                    rotation2 = self.image2_rotation_slider.value() / 10.0
                    if abs(rotation2) > 0.01:
                        center_point = rect2_native.center()
                        painter.save(); painter.translate(center_point); painter.rotate(rotation2); painter.translate(-center_point)
                        painter.drawImage(rect2_native, adjusted_img2)
                        painter.restore()
                    else:
                        painter.drawImage(rect2_native, adjusted_img2)

                painter.end()
                
                # --- 5. Update App State ---
                self.image_master = final_canvas.copy()
                self.image_before_contrast = self.image_master.copy()
                self.image_contrasted = self.image_master.copy()
                self.image_before_padding = self.image_master.copy()
                self.image_padded = False
                self.is_modified = True
                
                self.image = final_canvas.copy()
                self.main_image_is_inverted = False 

                self.remove_image1(); self.remove_image2()
                self.adjustment_context_combo.model().item(1).setEnabled(False)
                self.adjustment_context_combo.model().item(2).setEnabled(False)
                self.adjustment_context_combo.setCurrentText("Main Image")

                self._update_status_bar()
                self._update_marker_slider_ranges()
                self._update_overlay_slider_ranges()
                
                self.reset_all_adjustments()
                self.update_live_view()
                
                QMessageBox.information(self, "Success", "The overlay(s) have been rasterized onto the image.")

            def _update_overlay_position_from_sliders(self):
                if hasattr(self, 'image1_position'):
                    self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
                
                if hasattr(self, 'image2_position'):
                    self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
                
                self.update_live_view()
            
            def toggle_interactive_overlay_mode(self, checked):
                """Toggles the interactive overlay manipulation mode on and off."""
                if checked:
                    # Cancel any other active modes
                    self.cancel_drawing_mode()
                    self.cancel_rectangle_crop_mode()
                    self.cancel_selection_or_move_mode()
                    self.cancel_custom_item_interaction_mode()

                    self.overlay_mode_active = True
                    self.selected_overlay_index = 0 # No overlay selected initially
                    self.live_view_label.setCursor(Qt.PointingHandCursor)
                    QMessageBox.information(self, "Interactive Overlay Mode",
                                            "Mode Activated:\n\n"
                                            " - Click an overlay to select it.\n"
                                            " - Drag the selected overlay to move it.\n"
                                            " - Use Mouse Wheel over it to resize.\n"
                                            " - Use Arrow Keys to nudge it.\n\n"
                                            "Uncheck the button or press ESC to exit.")
                    
                    self._reset_live_view_label_custom_handlers()
                    self.live_view_label._custom_left_click_handler_from_app = self._handle_overlay_mouse_press
                    self.live_view_label._custom_mouseMoveEvent_from_app = self._handle_overlay_mouse_move
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self._handle_overlay_mouse_release
                else:
                    self.cancel_interactive_overlay_mode()

            def cancel_interactive_overlay_mode(self):
                """Safely exits the interactive overlay mode."""
                self.overlay_mode_active = False
                self.selected_overlay_index = 0
                self.dragging_overlay = False
                self._reset_live_view_label_custom_handlers()
                if hasattr(self, 'interactive_overlay_button') and self.interactive_overlay_button.isChecked():
                    self.interactive_overlay_button.setChecked(False)
                self.update_live_view() # Redraw to remove selection highlight

            def _get_overlay_rect_in_label_space(self, overlay_index):
                """Calculates the QRectF of an overlay image in the main label's coordinate space."""
                if not (self.image and not self.image.isNull()): return None
                
                overlay_image_orig = getattr(self, f'image{overlay_index}_original', None)
                overlay_position_native = getattr(self, f'image{overlay_index}_position', (0,0))
                overlay_resize_slider = getattr(self, f'image{overlay_index}_resize_slider', None)
                
                if not overlay_image_orig or not overlay_resize_slider: return None

                # Calculate scaled overlay dimensions in native pixels
                scale_factor = overlay_resize_slider.value() / 100.0
                overlay_w_native = overlay_image_orig.width() * scale_factor
                overlay_h_native = overlay_image_orig.height() * scale_factor
                
                # Create a QRectF in native image coordinates
                rect_native = QRectF(QPointF(overlay_position_native[0], overlay_position_native[1]), QSizeF(overlay_w_native, overlay_h_native))
                
                # Map this native rect to label space
                native_w = self.image.width(); native_h = self.image.height()
                label_w = self.live_view_label.width(); label_h = self.live_view_label.height()
                if not (native_w > 0 and native_h > 0 and label_w > 0 and label_h > 0): return None
                
                scale_to_label = min(label_w / native_w, label_h / native_h)
                offset_x_ls = (label_w - native_w * scale_to_label) / 2.0
                offset_y_ls = (label_h - native_h * scale_to_label) / 2.0
                
                topLeft_ls = QPointF(rect_native.left() * scale_to_label + offset_x_ls, rect_native.top() * scale_to_label + offset_y_ls)
                bottomRight_ls = QPointF(rect_native.right() * scale_to_label + offset_x_ls, rect_native.bottom() * scale_to_label + offset_y_ls)
                
                return QRectF(topLeft_ls, bottomRight_ls)

            def _handle_overlay_mouse_press(self, event):
                if not self.overlay_mode_active or event.button() != Qt.LeftButton: return
                
                clicked_point_ls = self.live_view_label.transform_point(event.position())
                
                # Check for resize handle click on the ALREADY selected overlay first
                if self.selected_overlay_index > 0:
                    rect_ls = self._get_overlay_rect_in_label_space(self.selected_overlay_index)
                    if rect_ls:
                        corners = [rect_ls.topLeft(), rect_ls.topRight(), rect_ls.bottomRight(), rect_ls.bottomLeft()]
                        click_radius = self.live_view_label.CORNER_HANDLE_BASE_RADIUS * 1.5 / self.live_view_label.zoom_level
                        for i, corner in enumerate(corners):
                            if (clicked_point_ls - corner).manhattanLength() < click_radius:
                                self.overlay_interaction_mode = 'resizing'
                                self.resizing_overlay_corner_index = i
                                self.drag_start_overlay_state = {
                                    'rect_ls': rect_ls,
                                    'mouse_pos_ls': clicked_point_ls,
                                    'opposite_corner_ls': corners[(i + 2) % 4]
                                }
                                self.live_view_label.setCursor(Qt.CrossCursor)
                                self.update_live_view()
                                return

                # If not resizing, check for a click on ANY overlay body to select/move
                rect2_ls = self._get_overlay_rect_in_label_space(2)
                rect1_ls = self._get_overlay_rect_in_label_space(1)
                
                new_selection = 0
                if rect2_ls and rect2_ls.contains(clicked_point_ls):
                    new_selection = 2
                elif rect1_ls and rect1_ls.contains(clicked_point_ls):
                    new_selection = 1
                
                self.selected_overlay_index = new_selection
                
                if self.selected_overlay_index > 0: # Start a move operation
                    self.overlay_interaction_mode = 'moving'
                    self.drag_start_overlay_state = {
                        'mouse_pos_ls': clicked_point_ls,
                        'position_native': getattr(self, f'image{self.selected_overlay_index}_position', (0,0))
                    }
                    self.live_view_label.setCursor(Qt.ClosedHandCursor)
                
                self.update_live_view()

            def _handle_overlay_mouse_move(self, event):
                if not self.overlay_interaction_mode or not (event.buttons() & Qt.LeftButton): return
                
                current_pos_ls = self.live_view_label.transform_point(event.position())

                # --- Map label space to native image space ---
                native_w = self.image.width(); native_h = self.image.height()
                label_w = self.live_view_label.width(); label_h = self.live_view_label.height()
                if not (native_w > 0 and native_h > 0 and label_w > 0 and label_h > 0): return
                scale_to_label = min(label_w / native_w, label_h / native_h)

                if self.overlay_interaction_mode == 'moving':
                    delta_ls = current_pos_ls - self.drag_start_overlay_state['mouse_pos_ls']
                    delta_native_x = delta_ls.x() / scale_to_label
                    delta_native_y = delta_ls.y() / scale_to_label
                    
                    start_native = self.drag_start_overlay_state['position_native']
                    new_native_x = start_native[0] + delta_native_x
                    new_native_y = start_native[1] + delta_native_y
                    
                    getattr(self, f'image{self.selected_overlay_index}_left_slider').setValue(int(new_native_x))
                    getattr(self, f'image{self.selected_overlay_index}_top_slider').setValue(int(new_native_y))

                elif self.overlay_interaction_mode == 'resizing':
                    fixed_corner_ls = self.drag_start_overlay_state['opposite_corner_ls']
                    new_rect_ls = QRectF(fixed_corner_ls, current_pos_ls).normalized()
                    
                    # Enforce aspect ratio
                    overlay_img = getattr(self, f'image{self.selected_overlay_index}_original')
                    if not overlay_img: return
                    aspect_ratio = overlay_img.width() / overlay_img.height() if overlay_img.height() > 0 else 1
                    
                    if new_rect_ls.width() / aspect_ratio > new_rect_ls.height():
                        new_rect_ls.setHeight(new_rect_ls.width() / aspect_ratio)
                    else:
                        new_rect_ls.setWidth(new_rect_ls.height() * aspect_ratio)

                    # Determine new top-left in native coords
                    offset_x_ls = (label_w - native_w * scale_to_label) / 2.0
                    offset_y_ls = (label_h - native_h * scale_to_label) / 2.0
                    new_topleft_native_x = (new_rect_ls.left() - offset_x_ls) / scale_to_label
                    new_topleft_native_y = (new_rect_ls.top() - offset_y_ls) / scale_to_label

                    # Determine new size percentage
                    new_width_native = new_rect_ls.width() / scale_to_label
                    new_size_perc = (new_width_native / overlay_img.width()) * 100 if overlay_img.width() > 0 else 100
                    
                    # Update sliders
                    getattr(self, f'image{self.selected_overlay_index}_left_slider').setValue(int(new_topleft_native_x))
                    getattr(self, f'image{self.selected_overlay_index}_top_slider').setValue(int(new_topleft_native_y))
                    getattr(self, f'image{self.selected_overlay_index}_resize_slider').setValue(int(new_size_perc))

            def _handle_overlay_mouse_release(self, event):
                if event.button() == Qt.LeftButton and self.overlay_interaction_mode:
                    self.save_state()
                    self.overlay_interaction_mode = None
                    self.resizing_overlay_corner_index = -1
                    self.drag_start_overlay_state = {}
                    self.live_view_label.setCursor(Qt.PointingHandCursor if self.overlay_mode_active else Qt.ArrowCursor)  
            


            def font_and_image_tab(self):
                tab = QWidget()
                main_layout = QHBoxLayout(tab)
                main_layout.setSpacing(15)

                left_column_layout = QVBoxLayout()

                # --- NEW: Adjustment Context Switcher ---
                context_group = QGroupBox("Adjustment Target")
                context_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                context_layout = QHBoxLayout(context_group)
                context_layout.addWidget(QLabel("Adjusting:"))
                self.adjustment_context_combo = QComboBox()
                self.adjustment_context_combo.addItems(["Main Image", "Overlay 1 (Base)", "Overlay 2 (Overlay)"])
                self.adjustment_context_combo.model().item(1).setEnabled(False) # Disable Overlay 1
                self.adjustment_context_combo.model().item(2).setEnabled(False) # Disable Overlay 2
                self.adjustment_context_combo.currentTextChanged.connect(self._on_adjustment_context_changed)
                context_layout.addWidget(self.adjustment_context_combo, 1)
                left_column_layout.addWidget(context_group)
                # --- END NEW ---

                levels_group = QGroupBox("Levels and Gamma"); levels_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                levels_layout = QGridLayout(levels_group)
                from matplotlib.figure import Figure
                self.hist_fig = Figure(figsize=(5, 1.8), dpi=100)
                self.hist_fig.patch.set_facecolor('none')
                self.hist_ax = self.hist_fig.add_subplot(111)
                self.hist_ax.patch.set_facecolor('white')
                self.hist_canvas = FigureCanvas(self.hist_fig)
                self.hist_canvas.setFixedHeight(120) 
                self._update_levels_histogram() 
                levels_layout.addWidget(self.hist_canvas, 0, 0, 1, 3) 
                
                self.black_point_label = QLabel("Black Point:"); levels_layout.addWidget(self.black_point_label, 1, 0)
                self.black_point_slider = QSlider(Qt.Horizontal); self.black_point_slider.setRange(0, 65535); self.black_point_slider.setValue(0); levels_layout.addWidget(self.black_point_slider, 1, 1)
                self.black_point_value_label = QLabel("0"); self.black_point_value_label.setFixedWidth(50); levels_layout.addWidget(self.black_point_value_label, 1, 2)
                
                self.white_point_label = QLabel("White Point:"); levels_layout.addWidget(self.white_point_label, 2, 0)
                self.white_point_slider = QSlider(Qt.Horizontal); self.white_point_slider.setRange(0, 65535); self.white_point_slider.setValue(65535); levels_layout.addWidget(self.white_point_slider, 2, 1)
                self.white_point_value_label = QLabel("65535"); self.white_point_value_label.setFixedWidth(50); levels_layout.addWidget(self.white_point_value_label, 2, 2)
                
                gamma_label = QLabel("Gamma:"); levels_layout.addWidget(gamma_label, 3, 0)
                self.gamma_slider = QSlider(Qt.Horizontal); self.gamma_slider.setRange(10, 500); self.gamma_slider.setValue(100); levels_layout.addWidget(self.gamma_slider, 3, 1)
                self.gamma_value_label = QLabel("1.00"); self.gamma_value_label.setFixedWidth(50); levels_layout.addWidget(self.gamma_value_label, 3, 2)
                
                left_column_layout.addWidget(levels_group)

                actions_group = QGroupBox("General Image Actions"); actions_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                actions_layout = QHBoxLayout(actions_group)
                self.bw_button = QPushButton("Grayscale"); self.bw_button.clicked.connect(self.convert_to_black_and_white)
                invert_button = QPushButton("Invert"); invert_button.clicked.connect(self.invert_image)
                reset_button = QPushButton("Reset Current Adjustments"); reset_button.clicked.connect(self.reset_all_adjustments)
                actions_layout.addWidget(self.bw_button); actions_layout.addWidget(invert_button); actions_layout.addStretch(); actions_layout.addWidget(reset_button)
                left_column_layout.addWidget(actions_group)
                left_column_layout.addStretch(1); main_layout.addLayout(left_column_layout, 1)

                right_column_layout = QVBoxLayout()
                cm_group = QGroupBox("Channel Mixer (for Color Images)"); cm_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                cm_layout = QGridLayout(cm_group)
                self.cm_red_slider = QSlider(Qt.Horizontal); self.cm_red_slider.setRange(0, 200); self.cm_red_slider.setValue(100)
                self.cm_red_label = QLabel("100%"); self.cm_red_label.setFixedWidth(40); cm_layout.addWidget(QLabel("Red:"), 0, 0); cm_layout.addWidget(self.cm_red_slider, 0, 1); cm_layout.addWidget(self.cm_red_label, 0, 2)
                self.cm_green_slider = QSlider(Qt.Horizontal); self.cm_green_slider.setRange(0, 200); self.cm_green_slider.setValue(100)
                self.cm_green_label = QLabel("100%"); self.cm_green_label.setFixedWidth(40); cm_layout.addWidget(QLabel("Green:"), 1, 0); cm_layout.addWidget(self.cm_green_slider, 1, 1); cm_layout.addWidget(self.cm_green_label, 1, 2)
                self.cm_blue_slider = QSlider(Qt.Horizontal); self.cm_blue_slider.setRange(0, 200); self.cm_blue_slider.setValue(100)
                self.cm_blue_label = QLabel("100%"); self.cm_blue_label.setFixedWidth(40); cm_layout.addWidget(QLabel("Blue:"), 2, 0); cm_layout.addWidget(self.cm_blue_slider, 2, 1); cm_layout.addWidget(self.cm_blue_label, 2, 2)
                self.cm_mono_checkbox = QCheckBox("Monochrome"); cm_layout.addWidget(self.cm_mono_checkbox, 3, 1)
                
                right_column_layout.addWidget(cm_group)

                usm_group = QGroupBox("Sharpening (Unsharp Mask)"); usm_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                usm_layout = QGridLayout(usm_group)
                self.usm_amount_slider = QSlider(Qt.Horizontal); self.usm_amount_slider.setRange(0, 500); self.usm_amount_slider.setValue(0)
                self.usm_amount_label = QLabel("0%"); self.usm_amount_label.setFixedWidth(40); usm_layout.addWidget(QLabel("Amount:"), 0, 0); usm_layout.addWidget(self.usm_amount_slider, 0, 1); usm_layout.addWidget(self.usm_amount_label, 0, 2)
                self.usm_radius_slider = QSlider(Qt.Horizontal); self.usm_radius_slider.setRange(1, 250); self.usm_radius_slider.setValue(10)
                self.usm_radius_label = QLabel("1.0 px"); self.usm_radius_label.setFixedWidth(50); usm_layout.addWidget(QLabel("Radius:"), 1, 0); usm_layout.addWidget(self.usm_radius_slider, 1, 1); usm_layout.addWidget(self.usm_radius_label, 1, 2)
                self.usm_threshold_slider = QSlider(Qt.Horizontal); self.usm_threshold_slider.setRange(0, 255); self.usm_threshold_slider.setValue(0)
                self.usm_threshold_label = QLabel("0"); self.usm_threshold_label.setFixedWidth(40); usm_layout.addWidget(QLabel("Threshold:"), 2, 0); usm_layout.addWidget(self.usm_threshold_slider, 2, 1); usm_layout.addWidget(self.usm_threshold_label, 2, 2)

                right_column_layout.addWidget(usm_group)

                clahe_group = QGroupBox("Local Contrast (CLAHE)"); clahe_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                clahe_layout = QGridLayout(clahe_group)
                self.clahe_clip_slider = QSlider(Qt.Horizontal); self.clahe_clip_slider.setRange(10, 100); self.clahe_clip_slider.setValue(10)
                self.clahe_clip_label = QLabel("1.0"); self.clahe_clip_label.setFixedWidth(40); clahe_layout.addWidget(QLabel("Clip Limit:"), 0, 0); clahe_layout.addWidget(self.clahe_clip_slider, 0, 1); clahe_layout.addWidget(self.clahe_clip_label, 0, 2)
                self.clahe_tile_slider = QSlider(Qt.Horizontal); self.clahe_tile_slider.setRange(2, 32); self.clahe_tile_slider.setValue(8)
                self.clahe_tile_label = QLabel("8x8"); self.clahe_tile_label.setFixedWidth(40); clahe_layout.addWidget(QLabel("Tile Size:"), 1, 0); clahe_layout.addWidget(self.clahe_tile_slider, 1, 1); clahe_layout.addWidget(self.clahe_tile_label, 1, 2)

                right_column_layout.addWidget(clahe_group)
                right_column_layout.addStretch(); main_layout.addLayout(right_column_layout, 1)

                # --- UPDATED CONNECTION LOGIC ---
                
                # Helper lambdas
                # Live preview: Update image, don't save to undo stack
                preview_update = lambda: self.apply_all_adjustments(save_history=False)
                # Commit: Save to undo stack (and ensure image is updated)
                commit_update = lambda: self.apply_all_adjustments(save_history=True)

                # --- Levels and Gamma ---
                # valueChanged: Updates histogram markers AND live preview (fixes arrow keys)
                self.black_point_slider.valueChanged.connect(self._update_histogram_markers_only)
                self.black_point_slider.valueChanged.connect(preview_update)
                self.black_point_slider.valueChanged.connect(lambda: self.black_point_slider.setFocus())
                self.black_point_slider.sliderReleased.connect(commit_update)

                self.white_point_slider.valueChanged.connect(self._update_histogram_markers_only)
                self.white_point_slider.valueChanged.connect(preview_update)
                self.white_point_slider.valueChanged.connect(lambda: self.white_point_slider.setFocus())
                self.white_point_slider.sliderReleased.connect(commit_update)

                self.gamma_slider.valueChanged.connect(preview_update)
                self.gamma_slider.valueChanged.connect(lambda: self.gamma_slider.setFocus())
                self.gamma_slider.sliderReleased.connect(commit_update)

                # Labels (Live update)
                self.black_point_slider.valueChanged.connect(lambda val: self.black_point_value_label.setText(f"{val}"))
                self.white_point_slider.valueChanged.connect(lambda val: self.white_point_value_label.setText(f"{val}"))
                self.gamma_slider.valueChanged.connect(lambda val: self.gamma_value_label.setText(f"{val/100.0:.2f}"))

                # --- Channel Mixer ---
                self.cm_red_slider.valueChanged.connect(preview_update)
                self.cm_red_slider.sliderReleased.connect(commit_update)
                self.cm_red_slider.sliderReleased.connect(lambda: self.cm_red_slider.setFocus())

                self.cm_green_slider.valueChanged.connect(preview_update)
                self.cm_green_slider.sliderReleased.connect(commit_update)
                self.cm_green_slider.sliderReleased.connect(lambda: self.cm_green_slider.setFocus())

                self.cm_blue_slider.valueChanged.connect(preview_update)
                self.cm_blue_slider.sliderReleased.connect(commit_update)
                self.cm_blue_slider.sliderReleased.connect(lambda: self.cm_blue_slider.setFocus())
                
                self.cm_mono_checkbox.stateChanged.connect(commit_update) # Checkboxes are instant commits

                # Labels
                self.cm_red_slider.valueChanged.connect(lambda v: self.cm_red_label.setText(f"{v}%"))
                self.cm_green_slider.valueChanged.connect(lambda v: self.cm_green_label.setText(f"{v}%"))
                self.cm_blue_slider.valueChanged.connect(lambda v: self.cm_blue_label.setText(f"{v}%"))

                # --- Unsharp Mask ---
                self.usm_amount_slider.valueChanged.connect(preview_update)
                self.usm_amount_slider.sliderReleased.connect(commit_update)
                self.usm_amount_slider.sliderReleased.connect(lambda: self.usm_amount_slider.setFocus())

                self.usm_radius_slider.valueChanged.connect(preview_update)
                self.usm_radius_slider.sliderReleased.connect(commit_update)
                self.usm_radius_slider.sliderReleased.connect(lambda: self.usm_radius_slider.setFocus())

                self.usm_threshold_slider.valueChanged.connect(preview_update)
                self.usm_threshold_slider.sliderReleased.connect(commit_update)
                self.usm_threshold_slider.sliderReleased.connect(lambda: self.usm_threshold_slider.setFocus())

                # Labels
                self.usm_amount_slider.valueChanged.connect(lambda v: self.usm_amount_label.setText(f"{v}%"))
                self.usm_radius_slider.valueChanged.connect(lambda v: self.usm_radius_label.setText(f"{(v/10.0):.1f} px"))
                self.usm_threshold_slider.valueChanged.connect(lambda v: self.usm_threshold_label.setText(f"{v}"))

                # --- CLAHE ---
                self.clahe_clip_slider.valueChanged.connect(preview_update)
                self.clahe_clip_slider.sliderReleased.connect(commit_update)
                self.clahe_clip_slider.sliderReleased.connect(lambda: self.clahe_clip_slider.setFocus())

                self.clahe_tile_slider.valueChanged.connect(preview_update)
                self.clahe_tile_slider.sliderReleased.connect(commit_update)
                self.clahe_tile_slider.sliderReleased.connect(lambda: self.clahe_tile_slider.setFocus())

                # Labels
                self.clahe_clip_slider.valueChanged.connect(lambda v: self.clahe_clip_label.setText(f"{(v/10.0):.1f}"))
                self.clahe_tile_slider.valueChanged.connect(lambda v: self.clahe_tile_label.setText(f"{v}x{v}"))

                return tab
            
            def _get_default_adjustments(self):
                """Returns a dictionary with default adjustment settings."""
                return {
                    'is_inverted': False, # <-- ADD THIS LINE
                    'levels_gamma': {'black_point': 0, 'white_point': 65535, 'gamma': 100},
                    'channel_mixer': {'r': 100, 'g': 100, 'b': 100, 'mono': False},
                    'unsharp_mask': {'amount': 0, 'radius': 1.0, 'threshold': 0},
                    'clahe': {'clip_limit': 1.0, 'tile_size': 8}
                }

            def _save_current_ui_adjustments(self):
                """Saves the current state of the UI sliders into the appropriate settings dictionary."""
                target_dict = None
                if self.adjustment_context == "Main Image":
                    # For main image, we save to the top-level attributes
                    self.channel_mixer_data = {'r': self.cm_red_slider.value(), 'g': self.cm_green_slider.value(), 'b': self.cm_blue_slider.value(), 'mono': self.cm_mono_checkbox.isChecked()}
                    self.unsharp_mask_data = {'amount': self.usm_amount_slider.value(), 'radius': self.usm_radius_slider.value() / 10.0, 'threshold': self.usm_threshold_slider.value()}
                    self.clahe_data = {'clip_limit': self.clahe_clip_slider.value() / 10.0, 'tile_size': self.clahe_tile_slider.value()}
                    # The main image levels are not stored separately, they are part of the main undo stack
                    return
                elif self.adjustment_context == "Overlay 1 (Base)":
                    target_dict = self.image1_adjustments
                elif self.adjustment_context == "Overlay 2 (Overlay)":
                    target_dict = self.image2_adjustments

                if target_dict:
                    target_dict['levels_gamma'] = {'black_point': self.black_point_slider.value(), 'white_point': self.white_point_slider.value(), 'gamma': self.gamma_slider.value()}
                    target_dict['channel_mixer'] = {'r': self.cm_red_slider.value(), 'g': self.cm_green_slider.value(), 'b': self.cm_blue_slider.value(), 'mono': self.cm_mono_checkbox.isChecked()}
                    target_dict['unsharp_mask'] = {'amount': self.usm_amount_slider.value(), 'radius': self.usm_radius_slider.value() / 10.0, 'threshold': self.usm_threshold_slider.value()}
                    target_dict['clahe'] = {'clip_limit': self.clahe_clip_slider.value() / 10.0, 'tile_size': self.clahe_tile_slider.value()}

            def _load_adjustments_to_ui(self, context):
                """Loads settings from the appropriate dictionary and updates the UI sliders."""
                settings = {}
                if context == "Main Image":
                    settings['levels_gamma'] = {'black_point': self.black_point_slider.value(), 'white_point': self.white_point_slider.value(), 'gamma': self.gamma_slider.value()} # Use current values
                    settings['channel_mixer'] = self.channel_mixer_data
                    settings['unsharp_mask'] = self.unsharp_mask_data
                    settings['clahe'] = self.clahe_data
                elif context == "Overlay 1 (Base)" and self.image1_adjustments:
                    settings = self.image1_adjustments
                elif context == "Overlay 2 (Overlay)" and self.image2_adjustments:
                    settings = self.image2_adjustments
                else: # Fallback to defaults if dictionary is missing
                    settings = self._get_default_adjustments()

                # Block signals to prevent feedback loops
                for slider in [self.black_point_slider, self.white_point_slider, self.gamma_slider, self.cm_red_slider, self.cm_green_slider, self.cm_blue_slider, self.usm_amount_slider, self.usm_radius_slider, self.usm_threshold_slider, self.clahe_clip_slider, self.clahe_tile_slider, self.cm_mono_checkbox]:
                    slider.blockSignals(True)
                
                # Update Levels/Gamma
                lg_settings = settings.get('levels_gamma', self._get_default_adjustments()['levels_gamma'])
                self.black_point_slider.setValue(lg_settings.get('black_point', 0))
                self.white_point_slider.setValue(lg_settings.get('white_point', 65535))
                self.gamma_slider.setValue(lg_settings.get('gamma', 100))

                # Update Channel Mixer
                cm_settings = settings.get('channel_mixer', self._get_default_adjustments()['channel_mixer'])
                self.cm_red_slider.setValue(cm_settings.get('r', 100))
                self.cm_green_slider.setValue(cm_settings.get('g', 100))
                self.cm_blue_slider.setValue(cm_settings.get('b', 100))
                self.cm_mono_checkbox.setChecked(cm_settings.get('mono', False))

                # Update Unsharp Mask
                usm_settings = settings.get('unsharp_mask', self._get_default_adjustments()['unsharp_mask'])
                self.usm_amount_slider.setValue(usm_settings.get('amount', 0))
                self.usm_radius_slider.setValue(int(usm_settings.get('radius', 1.0) * 10))
                self.usm_threshold_slider.setValue(usm_settings.get('threshold', 0))

                # Update CLAHE
                clahe_settings = settings.get('clahe', self._get_default_adjustments()['clahe'])
                self.clahe_clip_slider.setValue(int(clahe_settings.get('clip_limit', 1.0) * 10))
                self.clahe_tile_slider.setValue(clahe_settings.get('tile_size', 8))

                # Unblock signals
                for slider in [self.black_point_slider, self.white_point_slider, self.gamma_slider, self.cm_red_slider, self.cm_green_slider, self.cm_blue_slider, self.usm_amount_slider, self.usm_radius_slider, self.usm_threshold_slider, self.clahe_clip_slider, self.clahe_tile_slider, self.cm_mono_checkbox]:
                    slider.blockSignals(False)

                # Manually trigger label updates
                self.black_point_value_label.setText(f"{self.black_point_slider.value()}")
                self.white_point_value_label.setText(f"{self.white_point_slider.value()}")
                self.gamma_value_label.setText(f"{self.gamma_slider.value() / 100.0:.2f}")
                self.cm_red_label.setText(f"{self.cm_red_slider.value()}%")
                self.cm_green_label.setText(f"{self.cm_green_slider.value()}%")
                self.cm_blue_label.setText(f"{self.cm_blue_slider.value()}%")
                self.usm_amount_label.setText(f"{self.usm_amount_slider.value()}%")
                self.usm_radius_label.setText(f"{self.usm_radius_slider.value() / 10.0:.1f} px")
                self.usm_threshold_label.setText(f"{self.usm_threshold_slider.value()}")
                self.clahe_clip_label.setText(f"{self.clahe_clip_slider.value() / 10.0:.1f}")
                self.clahe_tile_label.setText(f"{self.clahe_tile_slider.value()}x{self.clahe_tile_slider.value()}")

            def _on_adjustment_context_changed(self, new_context):
                # 1. Save UI state from the old context before switching
                self._save_current_ui_adjustments()
                
                # 2. Update the internal context tracker
                self.adjustment_context = new_context
                
                # 3. Set the flag that tells the rendering engine if we are in an isolated editor view
                if new_context in ["Overlay 1 (Base)", "Overlay 2 (Overlay)"]:
                    self.is_in_dedicated_edit_mode = True
                else: # "Main Image"
                    self.is_in_dedicated_edit_mode = False
                    
                # 4. Load the new context's settings into the UI sliders
                self._load_adjustments_to_ui(new_context)
                
                # 5. Update the histogram to reflect the newly selected image's base state
                self._update_levels_histogram()
                
                # 6. Refresh the entire view. The rendering functions will now use the new context.
                self.update_live_view()
            
            def _update_channel_mixer(self):
                if not self.image or self.image.isNull(): return
                self.channel_mixer_data = {
                    'r': self.cm_red_slider.value(), 'g': self.cm_green_slider.value(),
                    'b': self.cm_blue_slider.value(), 'mono': self.cm_mono_checkbox.isChecked()
                }
                self.apply_all_adjustments()

            def _update_unsharp_mask(self):
                if not self.image or self.image.isNull(): return
                self.unsharp_mask_data = {
                    'amount': self.usm_amount_slider.value(),
                    'radius': self.usm_radius_slider.value() / 10.0,
                    'threshold': self.usm_threshold_slider.value()
                }
                self.apply_all_adjustments()

            def _update_clahe(self):
                if not self.image or self.image.isNull(): return
                self.clahe_data = {
                    'clip_limit': self.clahe_clip_slider.value() / 10.0,
                    'tile_size': self.clahe_tile_slider.value()
                }
                self.apply_all_adjustments()

            def _update_overlay_preview(self, overlay_index):
                """
                Recalculates the adjusted 8-bit preview for a specific overlay and caches it.
                This is the performance-intensive operation that is now called only when needed.
                """
                if overlay_index not in [1, 2]:
                    return

                source_image = getattr(self, f'image{overlay_index}_original', None)
                adjustments = getattr(self, f'image{overlay_index}_adjustments', None)

                if source_image and not source_image.isNull() and adjustments:
                    # Perform the full adjustment pipeline on the high-fidelity original
                    adjusted_image = self._apply_all_adjustments_to_image(source_image, adjustments)
                    # Store the 8-bit result in the preview cache attribute
                    setattr(self, f'image{overlay_index}_adjusted_preview', adjusted_image)
                else:
                    # If the overlay doesn't exist, clear its cache
                    setattr(self, f'image{overlay_index}_adjusted_preview', None)

            def apply_all_adjustments(self, save_history=True):
                """A single function to apply all adjustments in order, respecting transparency."""
                # Save UI state to the current context's dictionary first
                self._save_current_ui_adjustments()
                
                # Only save to the undo stack if requested and not restoring
                if save_history and not self._is_restoring_state:
                    self.save_state()

                if self.adjustment_context == "Main Image":
                    if self.image_master and not self.image_master.isNull():
                        base_image = self.image_master.copy()
                        self.image = self._apply_all_adjustments_to_image(base_image, {
                            'is_inverted': self.main_image_is_inverted,
                            'levels_gamma': {'black_point': self.black_point_slider.value(), 'white_point': self.white_point_slider.value(), 'gamma': self.gamma_slider.value()},
                            'channel_mixer': self.channel_mixer_data,
                            'unsharp_mask': self.unsharp_mask_data,
                            'clahe': self.clahe_data
                        })
                elif self.adjustment_context == "Overlay 1 (Base)":
                    self._update_overlay_preview(1)
                elif self.adjustment_context == "Overlay 2 (Overlay)":
                    self._update_overlay_preview(2)

                self.update_live_view()
                self._update_levels_histogram()

            def _apply_all_adjustments_to_image(self, source_image, settings_dict):
                """Applies a full suite of adjustments from a settings dict to a source QImage."""
                if not source_image or source_image.isNull(): return source_image
                
                # 1. Apply inversion FIRST if the flag is set.
                temp_image = source_image.copy()
                is_inverted = settings_dict.get('is_inverted', False)
                if is_inverted:
                    temp_image.invertPixels()
                
                # 2. All subsequent operations now use the (potentially) inverted temp_image.
                np_img_full = self.qimage_to_numpy(temp_image)

                if np_img_full is None: return source_image

                # --- DETERMINE BIT DEPTH AND MAX VALUE ---
                is_16bit = np_img_full.dtype == np.uint16
                max_val = 65535.0 if is_16bit else 255.0
                
                content_rect = None
                has_alpha = np_img_full.ndim == 3 and np_img_full.shape[2] == 4
                if has_alpha:
                    alpha_channel = np_img_full[:, :, 3]
                    rows, cols = np.any(alpha_channel, axis=1), np.any(alpha_channel, axis=0)
                    if np.any(rows) and np.any(cols):
                        ymin, ymax = np.where(rows)[0][[0, -1]]
                        xmin, xmax = np.where(cols)[0][[0, -1]]
                        content_rect = (xmin, ymin, xmax - xmin + 1, ymax - ymin + 1)
                        np_content = np_img_full[ymin:ymax+1, xmin:xmax+1]
                    else: np_content = np_img_full
                else:
                    np_content = np_img_full

                # --- CHANNEL MIXER ---
                channel_mixer_settings = settings_dict.get('channel_mixer', self._get_default_adjustments()['channel_mixer'])
                
                # Check if mixer is active or we need to promote grayscale
                is_mono = channel_mixer_settings.get('mono', False)
                r_val = channel_mixer_settings.get('r', 100)
                g_val = channel_mixer_settings.get('g', 100)
                b_val = channel_mixer_settings.get('b', 100)
                
                # If image is grayscale (2D) and settings are non-default, promote to RGB
                if np_content.ndim == 2 and (not is_mono and (r_val != 100 or g_val != 100 or b_val != 100)):
                    np_content = np.stack((np_content,)*3, axis=-1)

                if np_content.ndim == 3:
                    r_scale = r_val / 100.0
                    g_scale = g_val / 100.0
                    b_scale = b_val / 100.0
                    np_float = np_content.astype(np.float32)
                    
                    if is_mono:
                        gray = cv2.transform(np_float[...,:3], np.array([[b_scale],[g_scale],[r_scale]]).T)
                        np_content = np.clip(gray, 0, max_val).astype(np_content.dtype)
                    else:
                        if np_content.shape[2] == 4: # BGRA
                            np_float[..., 0] *= b_scale; np_float[..., 1] *= g_scale; np_float[..., 2] *= r_scale
                        else: # RGB/BGR
                            np_float[..., 0] *= r_scale; np_float[..., 1] *= g_scale; np_float[..., 2] *= b_scale
                        np_content = np.clip(np_float, 0, max_val).astype(np_content.dtype)

                # --- CLAHE FIX FOR 16-BIT COLOR ---
                clahe_settings = settings_dict.get('clahe', self._get_default_adjustments()['clahe'])
                clip_limit = clahe_settings.get('clip_limit', 1.0)
                if clip_limit > 1.0:
                    tile_size = clahe_settings.get('tile_size', 8)
                    # Scale clip limit for 16-bit range if we apply it to 16-bit integers directly
                    # However, logic below converts L to 16-bit int, so we scale it.
                    effective_clip = clip_limit * 256.0 if is_16bit else clip_limit
                    clahe = cv2.createCLAHE(clipLimit=effective_clip, tileGridSize=(tile_size, tile_size))
                    
                    if np_content.ndim == 2: # Grayscale (Works fine for 16-bit Int)
                        np_content = clahe.apply(np_content)
                    
                    elif np_content.ndim == 3: # Color
                        # Separate Alpha if present
                        bgr = np_content[..., :3]
                        alpha = np_content[..., 3] if np_content.shape[2] == 4 else None

                        if is_16bit:
                            # --- CRITICAL FIX START ---
                            # OpenCV's cvtColor(BGR2LAB) DOES NOT support uint16. 
                            # We MUST convert to float32 (0.0 - 1.0) first.
                            
                            # 1. Convert 16-bit Int to 32-bit Float
                            bgr_float = bgr.astype(np.float32) / 65535.0
                            
                            # 2. Convert Float BGR to Float LAB
                            lab_float = cv2.cvtColor(bgr_float, cv2.COLOR_BGR2LAB)
                            
                            # 3. Extract L channel (Float range 0.0 - 100.0 in OpenCV LAB)
                            l_channel = lab_float[..., 0]
                            
                            # 4. Map Float L (0-100) to UInt16 (0-65535) for CLAHE
                            l_uint16 = (l_channel / 100.0 * 65535.0).astype(np.uint16)
                            
                            # 5. Apply CLAHE to the 16-bit L channel
                            l_clahe = clahe.apply(l_uint16)
                            
                            # 6. Map back to Float L (0-100)
                            lab_float[..., 0] = (l_clahe.astype(np.float32) / 65535.0 * 100.0)
                            
                            # 7. Convert Float LAB back to Float BGR
                            bgr_float_out = cv2.cvtColor(lab_float, cv2.COLOR_LAB2BGR)
                            
                            # 8. Convert Float BGR back to UInt16
                            bgr_out = np.clip(bgr_float_out * 65535.0, 0, 65535).astype(np.uint16)
                            # --- CRITICAL FIX END ---
                            
                            # Recombine with alpha
                            if alpha is not None:
                                np_content = np.dstack((bgr_out, alpha))
                            else:
                                np_content = bgr_out

                        else: # 8-bit Color (Standard path, works with uint8)
                            lab = cv2.cvtColor(bgr, cv2.COLOR_BGR2LAB)
                            lab[..., 0] = clahe.apply(lab[..., 0])
                            bgr_out = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
                            if alpha is not None:
                                np_content = np.dstack((bgr_out, alpha))
                            else:
                                np_content = bgr_out
                
                # --- RECOMBINE WITH PADDING IF NEEDED ---
                if content_rect:
                    np_img = np_img_full.copy()
                    xmin, ymin, w, h = content_rect
                    if np_content.ndim == 2 and np_img.ndim == 3:
                        if np_img.shape[2] == 4: np_img[ymin:ymin+h, xmin:xmin+w] = cv2.cvtColor(np_content, cv2.COLOR_GRAY2BGRA)
                        else: np_img[ymin:ymin+h, xmin:xmin+w] = cv2.cvtColor(np_content, cv2.COLOR_GRAY2BGR)
                    elif np_content.ndim == 3 and np_img.ndim == 2:
                        np_img = np.stack((np_img,)*3, axis=-1) if np_content.shape[2]==3 else cv2.cvtColor(np_img, cv2.COLOR_GRAY2BGRA)
                        np_img[ymin:ymin+h, xmin:xmin+w] = np_content
                    else:
                        np_img[ymin:ymin+h, xmin:xmin+w] = np_content
                else: np_img = np_content
                
                # --- UNSHARP MASK ---
                unsharp_mask_settings = settings_dict.get('unsharp_mask', self._get_default_adjustments()['unsharp_mask'])
                amount = unsharp_mask_settings.get('amount', 0) / 100.0
                if amount > 0:
                    radius = unsharp_mask_settings.get('radius', 1.0)
                    threshold = unsharp_mask_settings.get('threshold', 0)
                    if is_16bit: threshold = threshold * 256
                    sigma = max(0.1, radius)
                    img_float = np_img.astype(np.float32)
                    blurred = cv2.GaussianBlur(img_float, (0, 0), sigma)
                    mask = np.abs(img_float - blurred) > threshold
                    sharpened = np.zeros_like(img_float)
                    sharpened[~mask] = img_float[~mask]
                    sharpened[mask] = img_float[mask] + (img_float[mask] - blurred[mask]) * amount
                    np_img = np.clip(sharpened, 0, max_val).astype(np_img.dtype)

                temp_image_after_effects = self.numpy_to_qimage(np_img)

                # Finally, apply levels and gamma
                lg_settings = settings_dict.get('levels_gamma', self._get_default_adjustments()['levels_gamma'])
                self._update_levels_histogram()
                return self.apply_levels_gamma(temp_image_after_effects, lg_settings['black_point'], lg_settings['white_point'], lg_settings['gamma'] / 100.0)

            def reset_all_adjustments(self):
                # This now only resets the settings for the CURRENT context
                if not self._is_restoring_state:
                    self.save_state() # Save state before resetting
                
                default_settings = self._get_default_adjustments()

                if self.adjustment_context == "Main Image":
                    # For the main image, we reset the top-level attributes
                    self.image_before_contrast = self.image_master.copy() if self.image_master else None
                    self.image_contrasted = self.image_master.copy() if self.image_master else None
                    
                    self.channel_mixer_data = default_settings['channel_mixer'].copy()
                    self.unsharp_mask_data = default_settings['unsharp_mask'].copy()
                    self.clahe_data = default_settings['clahe'].copy()
                    
                    # --- START FIX ---
                    # Call the dedicated function to reset sliders to their default full range
                    self._update_level_slider_ranges_and_defaults() 
                    # --- END FIX ---

                    if hasattr(self, 'gamma_slider'): self.gamma_slider.setValue(100)
                    
                elif self.adjustment_context == "Overlay 1 (Base)":
                    self.image1_adjustments = self._get_default_adjustments()
                elif self.adjustment_context == "Overlay 2 (Overlay)":
                    self.image2_adjustments = self._get_default_adjustments()
                
                # Load these newly reset default settings into the UI sliders
                self._load_adjustments_to_ui(self.adjustment_context)
                
                # Apply the reset adjustments to generate the new view
                if not self._is_restoring_state:
                    self.apply_all_adjustments()
            
            def reset_levels_and_gamma(self):
                # This function is now just an alias for the more comprehensive reset
                self.reset_all_adjustments()
                
            def apply_levels_gamma(self, qimage_base, black_point_ui, white_point_ui, gamma_ui_factor):
                if not qimage_base or qimage_base.isNull():
                    return qimage_base

                try:
                    img_array = self.qimage_to_numpy(qimage_base)
                    if img_array is None: raise ValueError("NumPy conversion failed.")

                    img_array_float = img_array.astype(np.float64)
                    original_dtype = img_array.dtype

                    # Determine max pixel value of the image type
                    if original_dtype == np.uint16:
                        max_dtype_val = 65535.0
                    elif original_dtype == np.uint8:
                        max_dtype_val = 255.0
                    elif np.issubdtype(original_dtype, np.floating):
                        max_dtype_val = np.max(img_array_float) if np.any(img_array_float) else 1.0
                    else:
                        max_dtype_val = 255.0

                    if max_dtype_val == 0: max_dtype_val = 1.0 

                    # Scale slider logic (0-65535 or 0-255) to image range
                    slider_max = 65535.0 
                    if hasattr(self, 'black_point_slider'):
                        slider_max = float(self.black_point_slider.maximum())
                    if slider_max == 0: slider_max = 1.0
                    
                    scale_factor_slider_to_img = max_dtype_val / slider_max

                    current_black = float(black_point_ui) * scale_factor_slider_to_img
                    current_white = float(white_point_ui) * scale_factor_slider_to_img

                    if current_black >= current_white:
                        if current_black >= max_dtype_val - 1:
                            current_black = max_dtype_val - 2.0; current_white = max_dtype_val
                        else:
                            current_white = current_black + 1.0

                    denominator = current_white - current_black
                    if abs(denominator) < 1e-9: denominator = 1e-9

                    # --- Process Channels ---
                    if img_array_float.ndim == 3: # Color
                        processed_channels = []
                        num_channels = img_array_float.shape[2]
                        channels_to_adjust = min(num_channels, 3)

                        for i in range(channels_to_adjust):
                            channel_data = img_array_float[..., i]
                            channel_levels = (channel_data - current_black) / denominator
                            channel_levels = np.clip(channel_levels, 0.0, 1.0)
                            channel_gamma = np.power(channel_levels, max(0.01, gamma_ui_factor))
                            processed_channels.append(np.clip(channel_gamma, 0.0, 1.0) * max_dtype_val)

                        img_array_final_float = np.stack(processed_channels, axis=-1)
                        if num_channels == 4:
                            img_array_final_float = np.dstack((img_array_final_float, img_array_float[..., 3]))

                    elif img_array_float.ndim == 2: # Grayscale
                        img_levels = (img_array_float - current_black) / denominator
                        img_levels = np.clip(img_levels, 0.0, 1.0)
                        img_gamma = np.power(img_levels, max(0.01, gamma_ui_factor))
                        img_array_final_float = np.clip(img_gamma, 0.0, 1.0) * max_dtype_val
                    else:
                        return qimage_base

                    # --- FIX: Preserve Bit Depth ---
                    if original_dtype == np.uint16:
                        img_array_final = img_array_final_float.astype(np.uint16)
                    else:
                        img_array_final = img_array_final_float.astype(np.uint8)

                    result_qimage = self.numpy_to_qimage(img_array_final)
                    if result_qimage.isNull(): raise ValueError("Conversion back to QImage failed.")
                    return result_qimage

                except Exception as e:
                    print(f"Error in apply_levels_gamma: {e}")
                    traceback.print_exc()
                    return qimage_base
                
            def _get_fully_adjusted_image_for_analysis(self):
                """
                Creates a temporary, high-fidelity QImage with ALL 'Main Image' visual adjustments
                (inversion, channels, sharpen, CLAHE, and Levels/Gamma) applied.
                This is the definitive "What You See Is What You Analyze" image source.
                """
                if not self.image_master or self.image_master.isNull():
                    return None

                # Create a temporary high-quality copy to work on
                adjusted_master_copy = self.image_master.copy()

                # Apply inversion if needed
                if self.main_image_is_inverted:
                    adjusted_master_copy.invertPixels()
                
                # Use NumPy for the next series of transformations
                np_adjusted_hq = self.qimage_to_numpy(adjusted_master_copy)
                if np_adjusted_hq is None:
                    return adjusted_master_copy 

                is_16bit = np_adjusted_hq.dtype == np.uint16
                max_val = 65535.0 if is_16bit else 255.0

                # Channel Mixer
                cm_settings = self.channel_mixer_data
                if np_adjusted_hq.ndim == 3:
                    np_float = np_adjusted_hq.astype(np.float32)
                    if cm_settings.get('mono', False):
                        r, g, b = cm_settings.get('r',100)/100.0, cm_settings.get('g',100)/100.0, cm_settings.get('b',100)/100.0
                        gray = cv2.transform(np_float[...,:3], np.array([[b],[g],[r]]).T)
                        np_adjusted_hq = gray.astype(np_adjusted_hq.dtype)
                    else:
                        r, g, b = cm_settings.get('r',100)/100.0, cm_settings.get('g',100)/100.0, cm_settings.get('b',100)/100.0
                        np_float[..., 0] *= b; np_float[..., 1] *= g; np_float[..., 2] *= r
                        np_adjusted_hq = np.clip(np_float, 0, max_val).astype(np_adjusted_hq.dtype)

                # --- CLAHE FIX FOR 16-BIT COLOR ---
                clahe_settings = self.clahe_data
                if clahe_settings.get('clip_limit', 1.0) > 1.0:
                    tile_size = clahe_settings.get('tile_size', 8)
                    effective_clip = clahe_settings['clip_limit'] * 256.0 if is_16bit else clahe_settings['clip_limit']
                    clahe = cv2.createCLAHE(clipLimit=effective_clip, tileGridSize=(tile_size, tile_size))
                    
                    if np_adjusted_hq.ndim == 2:
                        np_adjusted_hq = clahe.apply(np_adjusted_hq)
                    elif np_adjusted_hq.ndim == 3:
                        bgr = np_adjusted_hq[...,:3]
                        alpha = np_adjusted_hq[..., 3] if np_adjusted_hq.shape[2] == 4 else None

                        if is_16bit:
                            # 16-bit Int -> 32-bit Float -> LAB -> CLAHE -> Back
                            bgr_float = bgr.astype(np.float32) / 65535.0
                            lab_float = cv2.cvtColor(bgr_float, cv2.COLOR_BGR2LAB)
                            l_channel = lab_float[..., 0]
                            l_uint16 = (l_channel / 100.0 * 65535.0).astype(np.uint16)
                            l_clahe = clahe.apply(l_uint16)
                            lab_float[..., 0] = (l_clahe.astype(np.float32) / 65535.0 * 100.0)
                            bgr_float_out = cv2.cvtColor(lab_float, cv2.COLOR_LAB2BGR)
                            bgr_out = np.clip(bgr_float_out * 65535.0, 0, 65535).astype(np.uint16)
                            
                            if alpha is not None: np_adjusted_hq = np.dstack((bgr_out, alpha))
                            else: np_adjusted_hq = bgr_out
                        else:
                            # 8-bit
                            lab = cv2.cvtColor(bgr, cv2.COLOR_BGR2LAB)
                            lab[..., 0] = clahe.apply(lab[..., 0])
                            bgr_out = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
                            if alpha is not None: np_adjusted_hq = np.dstack((bgr_out, alpha))
                            else: np_adjusted_hq = bgr_out
                
                # Unsharp Mask
                usm_settings = self.unsharp_mask_data
                if usm_settings.get('amount', 0) > 0:
                    amount = usm_settings['amount'] / 100.0; sigma = max(0.1, usm_settings['radius'])
                    blurred = cv2.GaussianBlur(np_adjusted_hq, (0, 0), sigma)
                    np_adjusted_hq = cv2.addWeighted(np_adjusted_hq, 1.0 + amount, blurred, -amount, 0)
                
                # Convert back to QImage before applying levels
                qimage_after_effects = self.numpy_to_qimage(np_adjusted_hq)
                
                # Finally, apply the Levels/Gamma adjustments to this temporary image
                black_point_val = self.black_point_slider.value()
                white_point_val = self.white_point_slider.value()
                gamma_factor = self.gamma_slider.value() / 100.0

                fully_adjusted_image = self.apply_levels_gamma(
                    qimage_after_effects, 
                    black_point_val, 
                    white_point_val, 
                    gamma_factor
                )
                
                return fully_adjusted_image
            
            def update_image_levels_and_gamma(self):
                """Applies levels and gamma adjustments based on current slider values."""
                if not self.image or self.image.isNull():
                    return
                if not hasattr(self, 'image_contrasted') or not self.image_contrasted or self.image_contrasted.isNull():
                     if self.image_before_contrast and not self.image_before_contrast.isNull():
                         self.image_contrasted = self.image_before_contrast.copy()
                     elif self.image_master and not self.image_master.isNull():
                         self.image_contrasted = self.image_master.copy()
                     else: # Should not happen if self.image is valid
                          self.image_contrasted = self.image.copy()


                if self.contrast_applied == False: # Ensure base image for contrast is set
                    self.image_before_contrast = self.image.copy() # self.image is the one before any adjustments
                    self.image_contrasted = self.image_before_contrast.copy()
                    self.contrast_applied = True

                try:
                    black_point_val = 0
                    white_point_val = 255
                    gamma_val = 100 # Slider value for gamma (1.0 factor)

                    if hasattr(self, 'black_point_slider'): black_point_val = self.black_point_slider.value()
                    if hasattr(self, 'white_point_slider'): white_point_val = self.white_point_slider.value()
                    if hasattr(self, 'gamma_slider'): gamma_val = self.gamma_slider.value()
                    
                    gamma_factor = gamma_val / 100.0

                    # Ensure black point is less than white point for UI logic
                    if black_point_val >= white_point_val:
                        # If black >= white, clamp white to be slightly above black
                        # This prevents issues in apply_levels_gamma and provides a very high contrast visual
                        if self.black_point_slider.value() >= self.white_point_slider.value(): # Check slider values directly
                            self.white_point_slider.blockSignals(True)
                            self.white_point_slider.setValue(min(self.black_point_slider.value() + 1, self.white_point_slider.maximum()))
                            self.white_point_slider.blockSignals(False)
                            white_point_val = self.white_point_slider.value() # Get corrected value
                            if hasattr(self, 'white_point_value_label'): self.white_point_value_label.setText(str(white_point_val))

                    # Use self.image_contrasted as the base for adjustment
                    # self.image_contrasted holds the image state *before* current levels/gamma
                    # OR it holds the image state after a "permanent" operation like grayscale/invert/padding.
                    base_image_for_adjustment = self.image_contrasted
                    if not base_image_for_adjustment or base_image_for_adjustment.isNull():
                        # Fallback if image_contrasted is somehow invalid
                        base_image_for_adjustment = self.image_master.copy() if self.image_master else self.image.copy()


                    self.image = self.apply_levels_gamma(base_image_for_adjustment, black_point_val, white_point_val, gamma_factor)
                    self.update_live_view()
                except Exception as e:
                    print(f"Error in update_image_levels_and_gamma: {e}")
                    traceback.print_exc()
            
            def _update_level_slider_ranges_and_defaults(self):
                """
                Sets the Black/White Point slider ranges and default values based on the master image's format.
                8-bit images get 0-255. 16-bit/Color images get 0-65535.
                """
                # Default for 16-bit or unknown
                new_max = 65535
                
                img_to_check = self.image_master
                if img_to_check and not img_to_check.isNull():
                    current_format = img_to_check.format()
                    # Check for 8-bit formats
                    if current_format in [QImage.Format_Grayscale8, QImage.Format_Indexed8]:
                        new_max = 255
                
                # Update Black Point Slider
                if hasattr(self, 'black_point_slider'):
                    self.black_point_slider.blockSignals(True)
                    self.black_point_slider.setRange(0, new_max)
                    self.black_point_slider.setValue(0)
                    self.black_point_slider.blockSignals(False)
                    if hasattr(self, 'black_point_value_label'): 
                        self.black_point_value_label.setText("0")

                # Update White Point Slider
                if hasattr(self, 'white_point_slider'):
                    self.white_point_slider.blockSignals(True)
                    self.white_point_slider.setRange(0, new_max)
                    self.white_point_slider.setValue(new_max)
                    self.white_point_slider.blockSignals(False)
                    if hasattr(self, 'white_point_value_label'): 
                        self.white_point_value_label.setText(str(new_max))

            
            def _update_color_button_style(self, button, color):
                """ Helper to update button background color preview """
                if color.isValid():
                    button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
                else:
                    button.setStyleSheet("") # Reset to default stylesheet
            
            def create_separator(self):
                """Creates a horizontal separator line."""
                line = QFrame()
                line.setFrameShape(QFrame.HLine)
                line.setFrameShadow(QFrame.Sunken)
                return line
            
            
            def create_cropping_tab(self):
                """
                Creates a polished and user-friendly tab for all image transformations:
                Alignment, Skew, Cropping, and Padding, arranged in a balanced two-column layout.
                """
                tab = QWidget()
                main_layout = QHBoxLayout(tab)
                main_layout.setSpacing(15)

                # ======================================================================
                # === LEFT COLUMN: Global Transformations (Alignment, Skew, Flip) ===
                # ======================================================================
                left_column_layout = QVBoxLayout()
                left_column_layout.setSpacing(10)

                # --- Rotation & Orientation Group ---
                rotation_group = QGroupBox("Rotation and Orientation")
                rotation_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                rotation_layout = QGridLayout(rotation_group)
                rotation_layout.setSpacing(8)
                self.orientation_label = QLabel("Rotation Angle (+0.00°)")
                self.orientation_label.setFixedWidth(150)
                self.orientation_slider = QSlider(Qt.Horizontal)
                self.orientation_slider.setRange(-3600, 3600)
                self.orientation_slider.setValue(0)
                self.orientation_slider.valueChanged.connect(self._update_rotation_label)
                self.orientation_slider.valueChanged.connect(lambda: self.update_live_view()); self.orientation_slider.valueChanged.connect(lambda: self.orientation_slider.setFocus())
                self.align_button = QPushButton("Apply"); self.align_button.clicked.connect(self.align_image)
                self.reset_align_button = QPushButton("Reset"); self.reset_align_button.clicked.connect(self.reset_align_image)
                rotation_layout.addWidget(self.orientation_label, 0, 0)
                rotation_layout.addWidget(self.orientation_slider, 0, 1)
                rotation_layout.addWidget(self.align_button, 0, 2)
                rotation_layout.addWidget(self.reset_align_button, 0, 3)
                guides_flip_layout = QHBoxLayout()
                guides_flip_layout.addWidget(QLabel("Show Guide Lines: "))
                self.show_guides_checkbox = QCheckBox(""); self.show_guides_checkbox.setChecked(False)
                self.show_guides_checkbox.setToolTip("Show a center line to align/rotate the image. Shortcut: CTRL+G")
                self.show_guides_checkbox.stateChanged.connect(self.update_live_view)
                guides_flip_layout.addWidget(self.show_guides_checkbox)
                guides_flip_layout.addStretch()
                self.flip_vertical_button = QPushButton("Flip Vertical"); self.flip_vertical_button.clicked.connect(self.flip_vertical)
                self.flip_horizontal_button = QPushButton("Flip Horizontal"); self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
                guides_flip_layout.addWidget(self.flip_vertical_button)
                guides_flip_layout.addWidget(self.flip_horizontal_button)
                rotation_layout.addLayout(guides_flip_layout, 1, 0, 1, 4)
                rotation_layout.setColumnStretch(1, 1)
                left_column_layout.addWidget(rotation_group)

                # --- Skew Correction Group ---
                skew_group = QGroupBox("Skew Correction")
                skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                skew_layout = QGridLayout(skew_group)
                skew_layout.setSpacing(8)
                self.taper_skew_label = QLabel("Skew (+0.00)")
                self.taper_skew_label.setFixedWidth(150)
                self.taper_skew_slider = QSlider(Qt.Horizontal)
                self.taper_skew_slider.setRange(-70, 70); self.taper_skew_slider.setValue(0)
                self.taper_skew_slider.valueChanged.connect(lambda value: self.taper_skew_label.setText(f"Skew ({value / 100.0:+0.2f})"))
                self.taper_skew_slider.valueChanged.connect(lambda: self.update_live_view()); self.taper_skew_slider.valueChanged.connect(lambda: self.taper_skew_slider.setFocus())
                self.skew_button = QPushButton("Apply"); self.skew_button.clicked.connect(self.update_skew)
                self.reset_skew_button = QPushButton("Reset"); self.reset_skew_button.clicked.connect(lambda: self.taper_skew_slider.setValue(0))
                skew_layout.addWidget(self.taper_skew_label, 0, 0)
                skew_layout.addWidget(self.taper_skew_slider, 0, 1)
                skew_layout.addWidget(self.skew_button, 0, 2)
                skew_layout.addWidget(self.reset_skew_button, 0, 3)
                skew_layout.setColumnStretch(1, 1)
                left_column_layout.addWidget(skew_group)
                left_column_layout.addStretch(1)
                
                main_layout.addLayout(left_column_layout,1) # Add the completed left column layout

                # ======================================================================
                # === RIGHT COLUMN: Destructive Edits (Crop & Pad) ===
                # ======================================================================
                right_column_layout = QVBoxLayout()
                right_column_layout.setSpacing(10)

                # --- Cropping Group ---
                cropping_group = QGroupBox("Crop Image")
                cropping_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                cropping_layout = QVBoxLayout(cropping_group)
                crop_actions_layout = QHBoxLayout()
                self.draw_crop_rect_button = QPushButton("Draw Crop Area"); self.draw_crop_rect_button.setCheckable(True)
                self.draw_crop_rect_button.clicked.connect(self.toggle_rectangle_crop_mode)
                self.apply_crop_button = QPushButton("Apply Crop"); self.apply_crop_button.clicked.connect(self.update_crop)
                crop_actions_layout.addWidget(self.draw_crop_rect_button, 1); crop_actions_layout.addStretch(); crop_actions_layout.addWidget(self.apply_crop_button)
                cropping_layout.addLayout(crop_actions_layout)
                cropping_layout.addWidget(self.create_separator())
                crop_slider_layout = QGridLayout()
                self.crop_slider_min, self.crop_slider_max, self.crop_slider_precision_factor = 0, 10000, 100.0
                def create_value_label(initial_value=0.0):
                    lbl = QLabel(f"{initial_value:.2f}%"); lbl.setMinimumWidth(55); lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    return lbl
                self.crop_x_start_slider = QSlider(Qt.Horizontal); self.crop_x_start_slider.setRange(self.crop_slider_min, self.crop_slider_max); self.crop_x_start_slider.setValue(self.crop_slider_min); self.crop_x_start_slider.setEnabled(False)
                self.crop_x_start_value_label = create_value_label(0.00); self.crop_x_start_slider.valueChanged.connect(lambda val, lbl=self.crop_x_start_value_label: lbl.setText(f"{val/self.crop_slider_precision_factor:.2f}%")); self.crop_x_start_slider.valueChanged.connect(lambda: self._update_crop_from_sliders); self.crop_x_start_slider.valueChanged.connect(lambda: self.crop_x_start_slider.setFocus())
                self.crop_x_end_slider = QSlider(Qt.Horizontal); self.crop_x_end_slider.setRange(self.crop_slider_min, self.crop_slider_max); self.crop_x_end_slider.setValue(self.crop_slider_max); self.crop_x_end_slider.setEnabled(False)
                self.crop_x_end_value_label = create_value_label(100.00); self.crop_x_end_slider.valueChanged.connect(lambda val, lbl=self.crop_x_end_value_label: lbl.setText(f"{val/self.crop_slider_precision_factor:.2f}%")); self.crop_x_end_slider.valueChanged.connect(lambda: self._update_crop_from_sliders); self.crop_x_end_slider.valueChanged.connect(lambda: self.crop_x_end_slider.setFocus())
                self.crop_y_start_slider = QSlider(Qt.Horizontal); self.crop_y_start_slider.setRange(self.crop_slider_min, self.crop_slider_max); self.crop_y_start_slider.setValue(self.crop_slider_min); self.crop_y_start_slider.setEnabled(False)
                self.crop_y_start_value_label = create_value_label(0.00); self.crop_y_start_slider.valueChanged.connect(lambda val, lbl=self.crop_y_start_value_label: lbl.setText(f"{val/self.crop_slider_precision_factor:.2f}%")); self.crop_y_start_slider.valueChanged.connect(lambda: self._update_crop_from_sliders); self.crop_y_start_slider.valueChanged.connect(lambda: self.crop_y_start_slider.setFocus())
                self.crop_y_end_slider = QSlider(Qt.Horizontal); self.crop_y_end_slider.setRange(self.crop_slider_min, self.crop_slider_max); self.crop_y_end_slider.setValue(self.crop_slider_max); self.crop_y_end_slider.setEnabled(False)
                self.crop_y_end_value_label = create_value_label(100.00); self.crop_y_end_slider.valueChanged.connect(lambda val, lbl=self.crop_y_end_value_label: lbl.setText(f"{val/self.crop_slider_precision_factor:.2f}%")); self.crop_y_end_slider.valueChanged.connect(lambda: self._update_crop_from_sliders); self.crop_y_end_slider.valueChanged.connect(lambda: self.crop_y_end_slider.setFocus())
                crop_slider_layout.addWidget(QLabel("Left:"), 0, 0); crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1); crop_slider_layout.addWidget(self.crop_x_start_value_label, 0, 2)
                crop_slider_layout.addWidget(QLabel("Right:"), 0, 3); crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 4); crop_slider_layout.addWidget(self.crop_x_end_value_label, 0, 5)
                crop_slider_layout.addWidget(QLabel("Top:"), 1, 0); crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1); crop_slider_layout.addWidget(self.crop_y_start_value_label, 1, 2)
                crop_slider_layout.addWidget(QLabel("Bottom:"), 1, 3); crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 4); crop_slider_layout.addWidget(self.crop_y_end_value_label, 1, 5)
                crop_slider_layout.setColumnStretch(1, 1); crop_slider_layout.setColumnStretch(4, 1)
                cropping_layout.addLayout(crop_slider_layout)
                right_column_layout.addWidget(cropping_group)

                # --- Padding Group ---
                padding_group = QGroupBox("Add White Space (Padding)")
                padding_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                padding_layout = QGridLayout(padding_group)
                int_validator = QIntValidator(0, 5000, self)
                self.left_padding_input = QLineEdit("100"); self.left_padding_input.setValidator(int_validator)
                self.right_padding_input = QLineEdit("100"); self.right_padding_input.setValidator(int_validator)
                self.top_padding_input = QLineEdit("100"); self.top_padding_input.setValidator(int_validator)
                self.bottom_padding_input = QLineEdit("0"); self.bottom_padding_input.setValidator(int_validator)
                padding_layout.addWidget(QLabel("Left (px):"), 0, 0); padding_layout.addWidget(self.left_padding_input, 0, 1)
                padding_layout.addWidget(QLabel("Right (px):"), 0, 2); padding_layout.addWidget(self.right_padding_input, 0, 3)
                padding_layout.addWidget(QLabel("Top (px):"), 1, 0); padding_layout.addWidget(self.top_padding_input, 1, 1)
                padding_layout.addWidget(QLabel("Bottom (px):"), 1, 2); padding_layout.addWidget(self.bottom_padding_input, 1, 3)
                padding_actions_layout = QHBoxLayout()
                self.recommend_button = QPushButton("Set Recommended Values"); self.recommend_button.clicked.connect(self.recommended_values)
                self.clear_padding_button = QPushButton("Clear"); self.clear_padding_button.clicked.connect(self.clear_padding_values)
                self.finalize_button = QPushButton("Apply Padding"); self.finalize_button.clicked.connect(self.finalize_image)
                padding_actions_layout.addWidget(self.recommend_button); padding_actions_layout.addWidget(self.clear_padding_button); padding_actions_layout.addStretch()
                padding_actions_layout.addWidget(self.finalize_button)
                padding_layout.addLayout(padding_actions_layout, 2, 0, 1, 4)
                right_column_layout.addWidget(padding_group)
                right_column_layout.addStretch(1)

                main_layout.addLayout(right_column_layout,1) # Add the completed right column layout

                return tab
            
            def _update_rotation_label(self, value):
                """Updates the rotation label text with consistent formatting."""
                if hasattr(self, 'orientation_label'):
                    # FIX: The divisor is 20.0 to match the slider range of -3600 to 3600, yielding -180.0 to +180.0
                    orientation = value / 20.0
                    # Use f-string formatting to always show the sign (+/-) and pad with zeros.
                    self.orientation_label.setText(f"Rotation Angle ({orientation:+06.2f}°)")

            # def _update_skew_label(self, value):
            #     """Updates the skew label text."""
            #     if hasattr(self, 'taper_skew_label'):
            #         taper_value = value / 100.0
            #         self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")

            def _update_crop_from_sliders(self):
                """Updates crop rectangle based on slider percentages and ensures labels update."""
                if not self.image or self.image.isNull():
                    return
            
                # Read integer slider values
                try:
                    x_start_val = self.crop_x_start_slider.value()
                    x_end_val = self.crop_x_end_slider.value()
                    y_start_val = self.crop_y_start_slider.value()
                    y_end_val = self.crop_y_end_slider.value()
                except AttributeError:
                    print("Error: Crop sliders not fully initialized.")
                    return
            
                # Convert to percentage
                x_start_perc = x_start_val / self.crop_slider_precision_factor
                x_end_perc = x_end_val / self.crop_slider_precision_factor
                y_start_perc = y_start_val / self.crop_slider_precision_factor
                y_end_perc = y_end_val / self.crop_slider_precision_factor
            
                # Validation: Ensure Start <= End
                x_start = min(x_start_perc, x_end_perc)
                x_end = max(x_start_perc, x_end_perc)
                y_start = min(y_start_perc, y_end_perc)
                y_end = max(y_start_perc, y_end_perc)
            
                # Convert corrected percentages back to integer slider values
                x_start_val_corr = int(round(x_start * self.crop_slider_precision_factor))
                x_end_val_corr = int(round(x_end * self.crop_slider_precision_factor))
                y_start_val_corr = int(round(y_start * self.crop_slider_precision_factor))
                y_end_val_corr = int(round(y_end * self.crop_slider_precision_factor))
            
                # Reflect potentially swapped values back in sliders (only if different)
                # Use a flag to track if signals were blocked to avoid redundant label updates later
                signals_were_blocked = False
                sliders_to_check = [
                    (self.crop_x_start_slider, x_start_val_corr), (self.crop_x_end_slider, x_end_val_corr),
                    (self.crop_y_start_slider, y_start_val_corr), (self.crop_y_end_slider, y_end_val_corr)
                ]
                for slider, correct_value in sliders_to_check:
                    if slider.value() != correct_value:
                        slider.blockSignals(True)
                        slider.setValue(correct_value)
                        slider.blockSignals(False)
                        signals_were_blocked = True # Record that we manually set a value
            
                # --- Explicitly Update Labels AFTER potential corrections ---
                # This ensures labels always reflect the validated slider state, even if
                # the correction logic blocked the initial valueChanged signal for the label lambda.
                # Use the corrected integer values derived above.
                try:
                    self.crop_x_start_value_label.setText(f"{x_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_x_end_value_label.setText(f"{x_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_start_value_label.setText(f"{y_start_val_corr / self.crop_slider_precision_factor:.2f}%")
                    self.crop_y_end_value_label.setText(f"{y_end_val_corr / self.crop_slider_precision_factor:.2f}%")
                except AttributeError:
                    print("Error: Crop value labels not fully initialized for explicit update.")
                    # Handle case where labels might not exist yet if called too early
                    pass
                # --- End Explicit Label Update ---
            
            
                # Deactivate drawing mode if sliders are used
                if self.crop_rectangle_mode:
                    self.cancel_rectangle_crop_mode()
            
                # Calculate Image Coordinates using corrected percentages
                img_w = float(self.image.width())
                img_h = float(self.image.height())
                if img_w <= 0 or img_h <= 0: return
            
                img_x = img_w * (x_start / 100.0)
                img_y = img_h * (y_start / 100.0)
                img_crop_w = img_w * ((x_end - x_start) / 100.0)
                img_crop_h = img_h * ((y_end - y_start) / 100.0)
            
                # Store the calculated *image* coordinates
                self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))
            
                # Calculate View Coordinates for Preview
                label_w = float(self.live_view_label.width())
                label_h = float(self.live_view_label.height())
                if label_w <= 0 or label_h <= 0: return
            
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                if scale_factor <= 1e-9: scale_factor = 1
            
                display_img_w = img_w * scale_factor
                display_img_h = img_h * scale_factor
                display_offset_x = (label_w - display_img_w) / 2.0
                display_offset_y = (label_h - display_img_h) / 2.0
            
                view_x_start = display_offset_x + img_x * scale_factor
                view_y_start = display_offset_y + img_y * scale_factor
                view_x_end = view_x_start + img_crop_w * scale_factor
                view_y_end = view_y_start + img_crop_h * scale_factor
            
                # Update the visual preview rectangle
                self.live_view_label.crop_rect_final_view = QRectF(
                    QPointF(view_x_start, view_y_start),
                    QPointF(view_x_end, view_y_end)
                ).normalized()
                self.live_view_label.drawing_crop_rect = False
            
                self.update_live_view()
            
            def toggle_rectangle_crop_mode(self, checked):
                if checked:
                    self.enable_rectangle_crop_mode()
                else:
                    self.cancel_rectangle_crop_mode()

            def enable_rectangle_crop_mode(self):
                if self.crop_rectangle_mode: return
                self.marker_mode = None # Deactivate other modes
                self.crop_rectangle_mode = True
                self.crop_rectangle_coords = None
                self.live_view_label.clear_crop_preview()
                self.draw_crop_rect_button.setChecked(True)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label._custom_left_click_handler_from_app = self.start_crop_rectangle
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_crop_rectangle_preview
                self.live_view_label._custom_mouseReleaseEvent_from_app = self.finalize_crop_rectangle

            def cancel_rectangle_crop_mode(self):
                """Deactivates the rectangle drawing mode."""
                if not self.crop_rectangle_mode: return

                self.crop_rectangle_mode = False
                self.crop_rect_start_view = None
                if hasattr(self, 'draw_crop_rect_button'):
                     self.draw_crop_rect_button.setChecked(False)
                self._reset_live_view_label_custom_handlers() # Use helper
                self.update_live_view()

            def start_crop_rectangle(self, event):
                """Handles mouse press to start drawing the crop rectangle."""
                if not self.crop_rectangle_mode:
                    # If not in crop mode, pass event to the base class or intended handler
                    super(LiveViewLabel, self.live_view_label.__class__).mousePressEvent(self.live_view_label, event)
                    return

                if event.button() == Qt.LeftButton:
                    # Get start point in *view* coordinates (unzoomed)
                    self.crop_rect_start_view = self.live_view_label.transform_point(event.position())
                    # Tell LiveViewLabel to start drawing the preview
                    self.live_view_label.start_crop_preview(self.crop_rect_start_view)

            def update_crop_rectangle_preview(self, event):
                """Handles mouse move to update the crop rectangle preview."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                     # If not in crop mode or not dragging, pass event along
                     super(LiveViewLabel, self.live_view_label.__class__).mouseMoveEvent(self.live_view_label, event)
                     return

                # Check if left button is held down (QApplication.mouseButtons() might be needed)
                if event.buttons() & Qt.LeftButton:
                    current_point_view = self.live_view_label.transform_point(event.position())
                    if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                        current_point_view = self._constrain_point_orthogonally(self.crop_rect_start_view, current_point_view)
                    # Tell LiveViewLabel to update the preview
                    self.live_view_label.update_crop_preview(self.crop_rect_start_view, current_point_view)

            def finalize_crop_rectangle(self, event):
                """Handles mouse release to finalize the crop rectangle and updates sliders/labels."""
                if not self.crop_rectangle_mode or not self.crop_rect_start_view:
                    # If not in crop mode or drag never started, pass event along
                    # Use super() with the correct class hierarchy if LiveViewLabel inherits directly
                    # If LiveViewLabel is just an instance variable, this super() call is incorrect.
                    # Assuming LiveViewLabel is a custom class inheriting QLabel:
                    # super(LiveViewLabel, self.live_view_label).mouseReleaseEvent(event)
                    # If LiveViewLabel is just a QLabel instance, default behaviour is likely sufficient:
                    if hasattr(self.live_view_label, 'mouseReleaseEvent') and callable(getattr(QLabel, 'mouseReleaseEvent', None)):
                         QLabel.mouseReleaseEvent(self.live_view_label, event) # Call base QLabel handler if needed
                    return

                if event.button() == Qt.LeftButton:
                    end_point_view = self.live_view_label.transform_point(event.position())
                    start_point_view = self.crop_rect_start_view

                    # Finalize the visual preview in LiveViewLabel
                    self.live_view_label.finalize_crop_preview(start_point_view, end_point_view)

                    try:
                        # --- Coordinate Conversion (same as before) ---
                        if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                        img_w = float(self.image.width())
                        img_h = float(self.image.height())
                        label_w = float(self.live_view_label.width())
                        label_h = float(self.live_view_label.height())

                        if img_w <= 0 or img_h <= 0 or label_w <= 0 or label_h <= 0:
                            raise ValueError("Invalid image or label dimensions.")

                        scale_factor = min(label_w / img_w, label_h / img_h)
                        if scale_factor <= 1e-9: raise ValueError("Scale factor too small.")

                        display_img_w = img_w * scale_factor
                        display_img_h = img_h * scale_factor
                        display_offset_x = (label_w - display_img_w) / 2.0
                        display_offset_y = (label_h - display_img_h) / 2.0

                        start_x_img = (start_point_view.x() - display_offset_x) / scale_factor
                        start_y_img = (start_point_view.y() - display_offset_y) / scale_factor
                        end_x_img = (end_point_view.x() - display_offset_x) / scale_factor
                        end_y_img = (end_point_view.y() - display_offset_y) / scale_factor

                        img_x = min(start_x_img, end_x_img)
                        img_y = min(start_y_img, end_y_img)
                        img_crop_w = abs(end_x_img - start_x_img)
                        img_crop_h = abs(end_y_img - start_y_img)
                        # --- End Coordinate Conversion ---

                        if img_crop_w < 1 or img_crop_h < 1:
                             # Handle case where drawn rectangle is too small
                             self.crop_rectangle_coords = None
                             self.live_view_label.clear_crop_preview()
                             # Keep sliders disabled
                             self.crop_x_start_slider.setEnabled(False)
                             self.crop_x_end_slider.setEnabled(False)
                             self.crop_y_start_slider.setEnabled(False)
                             self.crop_y_end_slider.setEnabled(False)
                        else:
                            # Store valid image coordinates
                            self.crop_rectangle_coords = (int(img_x), int(img_y), int(img_crop_w), int(img_crop_h))

                            # Calculate percentages from valid image coordinates
                            x_start_perc = (img_x / img_w) * 100.0 if img_w > 0 else 0
                            y_start_perc = (img_y / img_h) * 100.0 if img_h > 0 else 0
                            x_end_perc = ((img_x + img_crop_w) / img_w) * 100.0 if img_w > 0 else 100
                            y_end_perc = ((img_y + img_crop_h) / img_h) * 100.0 if img_h > 0 else 100

                            # Convert percentages to integer slider values
                            x_start_val = int(round(x_start_perc * self.crop_slider_precision_factor))
                            y_start_val = int(round(y_start_perc * self.crop_slider_precision_factor))
                            x_end_val = int(round(x_end_perc * self.crop_slider_precision_factor))
                            y_end_val = int(round(y_end_perc * self.crop_slider_precision_factor))

                            # --- Update Sliders (Block signals) ---
                            self.crop_x_start_slider.blockSignals(True); self.crop_x_start_slider.setValue(x_start_val); self.crop_x_start_slider.blockSignals(False)
                            self.crop_x_end_slider.blockSignals(True); self.crop_x_end_slider.setValue(x_end_val); self.crop_x_end_slider.blockSignals(False)
                            self.crop_y_start_slider.blockSignals(True); self.crop_y_start_slider.setValue(y_start_val); self.crop_y_start_slider.blockSignals(False)
                            self.crop_y_end_slider.blockSignals(True); self.crop_y_end_slider.setValue(y_end_val); self.crop_y_end_slider.blockSignals(False)

                            # --- Explicitly Update Labels AFTER setting sliders ---
                            self.crop_x_start_value_label.setText(f"{x_start_perc:.2f}%")
                            self.crop_x_end_value_label.setText(f"{x_end_perc:.2f}%")
                            self.crop_y_start_value_label.setText(f"{y_start_perc:.2f}%")
                            self.crop_y_end_value_label.setText(f"{y_end_perc:.2f}%")
                            # --- End Explicit Label Update ---

                            # --- Enable the sliders ---
                            self.crop_x_start_slider.setEnabled(True)
                            self.crop_x_end_slider.setEnabled(True)
                            self.crop_y_start_slider.setEnabled(True)
                            self.crop_y_end_slider.setEnabled(True)

                    except Exception as e:
                         # Handle errors during coordinate calculation or UI update
                         QMessageBox.warning(self, "Coordinate Error", f"Could not calculate/update crop: {e}")
                         traceback.print_exc()
                         self.crop_rectangle_coords = None
                         self.live_view_label.clear_crop_preview()
                         # Ensure sliders remain disabled on error
                         self.crop_x_start_slider.setEnabled(False)
                         self.crop_x_end_slider.setEnabled(False)
                         self.crop_y_start_slider.setEnabled(False)
                         self.crop_y_end_slider.setEnabled(False)

                    # Reset the temporary start point used for drawing
                    self.crop_rect_start_view = None
            
            def clear_padding_values(self):
                self.bottom_padding_input.setText("0")
                self.top_padding_input.setText("0")
                self.right_padding_input.setText("0")
                self.left_padding_input.setText("0")
                
            def recommended_values(self):
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to set recommended padding.")
                    return

                # This function should ONLY set the text in the QLineEdit fields for padding.
                # It should NOT directly change slider ranges or _marker_shift_added values.
                # The marker offsets are independent of padding until "Apply Padding" is clicked.

                try:
                    native_width = self.image.width()
                    native_height = self.image.height()

                    if native_width <= 0 or native_height <= 0:
                        QMessageBox.warning(self, "Error", "Current image has invalid dimensions.")
                        return

                    # Set recommended padding values in the QLineEdit fields
                    self.left_padding_input.setText(str(int(native_width * 0.15)))
                    self.right_padding_input.setText(str(int(native_width * 0.15)))
                    self.top_padding_input.setText(str(int(native_height * 0.15)))
                    self.bottom_padding_input.setText(str(int(0))) # Or another sensible default

                    # DO NOT update slider ranges or _marker_shift_added here.
                    # Slider ranges are updated by _update_marker_slider_ranges when self.image changes.
                    # _marker_shift_added are updated by sliders or by crop/pad operations.
                    
                    # No need to call self.update_live_view() here either, as only
                    # text fields for a future operation were changed.
                    # QMessageBox.information(self, "Recommended Values Set", 
                    #                         "Recommended padding values have been entered into the fields.\n"
                    #                         "Click 'Apply Padding' to use them.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not set recommended values: {e}")
                
                
                
            # In class CombinedSDSApp:

            def create_markers_tab(self):
                """Create the Markers tab with a more compact and organized layout."""
                tab = QWidget()
                main_layout = QVBoxLayout(tab)
                main_layout.setSpacing(10)

                # --- Group 1: Marker Data (Presets and Labels) ---
                presets_group = QGroupBox("Marker Presets and Labels")
                presets_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                presets_layout = QGridLayout(presets_group)
                presets_layout.addWidget(QLabel("Preset:"), 0, 0)
                self.combo_box = QComboBox(self)
                if hasattr(self, 'presets_data') and self.presets_data: self.combo_box.addItems(sorted(self.presets_data.keys()))
                self.combo_box.addItem("Custom"); self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
                self.combo_box.setFixedWidth(450)
                presets_layout.addWidget(self.combo_box, 0, 1)
                self.rename_input = QLineEdit(self)
                self.rename_input.setPlaceholderText("Enter new name to save preset..."); self.rename_input.setEnabled(False)
                presets_layout.addWidget(self.rename_input, 0, 2)
                self.save_button = QPushButton("Save"); self.save_button.setToolTip("Saves the current L/R, Top, Custom Markers/Shapes to the selected/new preset name.")
                self.save_button.setFixedWidth(100)
                self.save_button.clicked.connect(self.save_config)
                presets_layout.addWidget(self.save_button, 0, 3)
                self.remove_config_button = QPushButton("Remove"); self.remove_config_button.setFixedWidth(100); self.remove_config_button.clicked.connect(self.remove_config)
                presets_layout.addWidget(self.remove_config_button, 0, 4)
                self.load_custom_from_preset_checkbox = QCheckBox("Load Custom Markers/Shapes from Preset")
                self.load_custom_from_preset_checkbox.setChecked(True) # Default to checked
                self.load_custom_from_preset_checkbox.setToolTip("If checked, changing the preset will also load any custom markers/shapes saved with it, overwriting existing ones.")
                presets_layout.addWidget(self.load_custom_from_preset_checkbox, 1, 1, 1, 4)
                presets_layout.addWidget(QLabel("L/R Values:"), 2, 0)
                self.marker_values_textbox = QLineEdit(self)
                self.marker_values_textbox.setPlaceholderText("Custom L/R values (comma-separated)"); self.marker_values_textbox.setEnabled(False)
                presets_layout.addWidget(self.marker_values_textbox, 2, 1, 1, 4)
                presets_layout.addWidget(QLabel("Top Labels:"), 3, 0, Qt.AlignTop)
                self.top_marker_input = QTextEdit(self)
                self.top_marker_input.setText(", ".join(map(str, getattr(self, 'top_label', []))))
                self.top_marker_input.setFixedHeight(50); self.top_marker_input.setPlaceholderText("Top labels (comma-separated)")
                presets_layout.addWidget(self.top_marker_input, 3, 1, 1, 3)
                self.update_labels_button = QPushButton("Update Labels")
                self.update_labels_button.setToolTip("Apply values from the L/R and Top text boxes to any markers currently on the image.")
                self.update_labels_button.clicked.connect(self.update_all_labels)
                presets_layout.addWidget(self.update_labels_button, 3, 4)
                main_layout.addWidget(presets_group)

                # --- Group 2: Standard Marker Tools ---
                standard_group = QGroupBox("Standard Marker Tools")
                standard_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                standard_layout = QGridLayout(standard_group)
                standard_layout.setColumnStretch(1, 1)
                font_options_layout = QHBoxLayout()
                self.font_combo_box = QFontComboBox(); self.font_combo_box.setCurrentFont(QFont(self.font_family)); self.font_combo_box.currentFontChanged.connect(self.update_font)
                self.font_size_spinner = QSpinBox(); self.font_size_spinner.setRange(6, 72); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.valueChanged.connect(self.update_font)
                self.font_color_button = QPushButton("Color"); self.font_color_button.clicked.connect(self.select_font_color); self._update_color_button_style(self.font_color_button, self.font_color)
                self.font_rotation_input = QSpinBox(); self.font_rotation_input.setRange(-180, 180); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.setSuffix(" °"); self.font_rotation_input.valueChanged.connect(self.update_font)
                font_options_layout.addWidget(QLabel("Font:")); font_options_layout.addWidget(self.font_combo_box, 1); font_options_layout.addWidget(self.font_size_spinner)
                font_options_layout.addWidget(self.font_color_button); font_options_layout.addWidget(QLabel("Top Rotation:")); font_options_layout.addWidget(self.font_rotation_input)
                standard_layout.addLayout(font_options_layout, 0, 0, 1, 3)
                standard_layout.addWidget(self.create_separator(), 1, 0, 1, 3)
                left_buttons = QHBoxLayout(); left_marker_button = QPushButton("Place Left"); left_marker_button.clicked.connect(self.enable_left_marker_mode); remove_left_button = QPushButton("Remove Last"); remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove')); reset_left_button = QPushButton("Reset All"); reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
                left_buttons.addWidget(left_marker_button); left_buttons.addWidget(remove_left_button); left_buttons.addWidget(reset_left_button); standard_layout.addLayout(left_buttons, 2, 0)
                self.left_padding_slider = QSlider(Qt.Horizontal); self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1]); self.left_padding_slider.setValue(self.left_marker_shift_added); self.left_padding_slider.valueChanged.connect(lambda: self.update_left_padding()); self.left_padding_slider.valueChanged.connect(lambda: self.left_padding_slider.setFocus())
                standard_layout.addWidget(self.left_padding_slider, 2, 1)
                duplicate_left_button = QPushButton("Copy →"); duplicate_left_button.setToolTip("Copy Right Markers & Offset to Left"); duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left')); standard_layout.addWidget(duplicate_left_button, 2, 2)
                right_buttons = QHBoxLayout(); right_marker_button = QPushButton("Place Right"); right_marker_button.clicked.connect(self.enable_right_marker_mode); remove_right_button = QPushButton("Remove Last"); remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove')); reset_right_button = QPushButton("Reset All"); reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
                right_buttons.addWidget(right_marker_button); right_buttons.addWidget(remove_right_button); right_buttons.addWidget(reset_right_button); standard_layout.addLayout(right_buttons, 3, 0)
                self.right_padding_slider = QSlider(Qt.Horizontal); self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1]); self.right_padding_slider.setValue(self.right_marker_shift_added); self.right_padding_slider.valueChanged.connect(lambda: self.update_right_padding()); self.right_padding_slider.valueChanged.connect(lambda: self.right_padding_slider.setFocus())
                standard_layout.addWidget(self.right_padding_slider, 3, 1)
                duplicate_right_button = QPushButton("← Copy"); duplicate_right_button.setToolTip("Copy Left Markers & Offset to Right"); duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right')); standard_layout.addWidget(duplicate_right_button, 3, 2)
                top_buttons = QHBoxLayout(); top_marker_button = QPushButton("Place Top"); top_marker_button.clicked.connect(self.enable_top_marker_mode); remove_top_button = QPushButton("Remove Last"); remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove')); reset_top_button = QPushButton("Reset All"); reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
                top_buttons.addWidget(top_marker_button); top_buttons.addWidget(remove_top_button); top_buttons.addWidget(reset_top_button); standard_layout.addLayout(top_buttons, 4, 0)
                self.top_padding_slider = QSlider(Qt.Horizontal); self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1]); self.top_padding_slider.setValue(self.top_marker_shift_added); self.top_padding_slider.valueChanged.connect(lambda: self.update_top_padding()); self.top_padding_slider.valueChanged.connect(lambda: self.top_padding_slider.setFocus())
                standard_layout.addWidget(self.top_padding_slider, 4, 1)
                main_layout.addWidget(standard_group)

                # --- Group 3: Custom Markers, Shapes & Grid ---
                custom_group = QGroupBox("Custom Markers, Shapes and Grid")
                custom_group.setStyleSheet("QGroupBox { font-weight: bold; }")
                custom_layout = QVBoxLayout(custom_group); custom_layout.setSpacing(6)
                row1_layout = QHBoxLayout(); row1_layout.setSpacing(6)
                self.custom_marker_button = QPushButton("Place Custom", self); self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
                self.custom_marker_text_entry = QLineEdit(self); self.custom_marker_text_entry.setPlaceholderText("Custom text...")
                arrow_buttons_layout = QHBoxLayout(); arrow_buttons_layout.setContentsMargins(0, 0, 0, 0); arrow_buttons_layout.setSpacing(2)
                arrow_size = 35
                self.custom_marker_button_left_arrow = QPushButton("←"); self.custom_marker_button_left_arrow.setToolTip("Ctrl+Left")
                self.custom_marker_button_right_arrow = QPushButton("→"); self.custom_marker_button_right_arrow.setToolTip("Ctrl+Right")
                self.custom_marker_button_top_arrow = QPushButton("↑"); self.custom_marker_button_top_arrow.setToolTip("Ctrl+Up")
                self.custom_marker_button_bottom_arrow = QPushButton("↓"); self.custom_marker_button_bottom_arrow.setToolTip("Ctrl+Down")

                # --- START OF BUG FIX ---
                # 1. Enlarge the arrow symbols by applying a larger, bold font.
                arrow_font = QFont(); arrow_font.setPointSize(14); arrow_font.setBold(True)
                self.custom_marker_button_left_arrow.setFont(arrow_font)
                self.custom_marker_button_right_arrow.setFont(arrow_font)
                self.custom_marker_button_top_arrow.setFont(arrow_font)
                self.custom_marker_button_bottom_arrow.setFont(arrow_font)
                
                self.custom_marker_button_left_arrow.setFixedSize(arrow_size, arrow_size)
                self.custom_marker_button_right_arrow.setFixedSize(arrow_size, arrow_size)
                self.custom_marker_button_top_arrow.setFixedSize(arrow_size, arrow_size)
                self.custom_marker_button_bottom_arrow.setFixedSize(arrow_size, arrow_size)
                
                arrow_buttons_layout.addWidget(self.custom_marker_button_left_arrow); arrow_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
                arrow_buttons_layout.addWidget(self.custom_marker_button_top_arrow); arrow_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
                self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
                self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
                self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
                self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))
                self.custom_font_type_dropdown = QFontComboBox(); self.custom_font_type_dropdown.setCurrentFont(QFont("Arial")); self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
                self.custom_font_size_spinbox = QSpinBox(); self.custom_font_size_spinbox.setRange(1, 150); self.custom_font_size_spinbox.setValue(12); self.custom_font_size_spinbox.setPrefix("Size (px): ")
                self.custom_marker_color_button = QPushButton("Color"); self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
                if not hasattr(self, 'custom_marker_color'): self.custom_marker_color = QColor(0,0,0)
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                
                row1_layout.addWidget(self.custom_marker_button)
                # 2. Adjust stretch factors to make the text box longer.
                row1_layout.addWidget(self.custom_marker_text_entry, 2) # Give text entry a stretch factor of 2
                row1_layout.addLayout(arrow_buttons_layout)
                row1_layout.addWidget(self.custom_font_type_dropdown, 1) # Give font dropdown a stretch factor of 1
                row1_layout.addWidget(self.custom_font_size_spinbox)
                row1_layout.addWidget(self.custom_marker_color_button)
                # --- END OF BUG FIX ---
                
                custom_layout.addLayout(row1_layout)
                row2_layout = QHBoxLayout(); row2_layout.setSpacing(6)
                self.remove_custom_marker_button = QPushButton("Remove Last"); self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
                self.reset_custom_marker_button = QPushButton("Reset All"); self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
                shape_size = 25
                self.draw_line_button = QPushButton("L"); self.draw_line_button.setToolTip("Draw Line"); self.draw_line_button.setFixedSize(shape_size, shape_size); self.draw_line_button.clicked.connect(self.enable_line_drawing_mode)
                self.draw_rect_button = QPushButton("R"); self.draw_rect_button.setToolTip("Draw Rectangle"); self.draw_rect_button.setFixedSize(shape_size, shape_size); self.draw_rect_button.clicked.connect(self.enable_rectangle_drawing_mode)
                self.remove_shape_button = QPushButton("X"); self.remove_shape_button.setToolTip("Remove Last Shape"); self.remove_shape_button.setFixedSize(shape_size, shape_size); self.remove_shape_button.clicked.connect(self.remove_last_custom_shape)
                self.show_grid_checkbox_x = QCheckBox("Snap X"); self.show_grid_checkbox_x.setToolTip("Snap horizontally. Ctrl+Shift+X or CMD+Shift+X toggles X and Ctrl+Shift+G or CMD+Shift+G for both X and Y.")
                self.show_grid_checkbox_x.setFixedWidth(90)
                self.show_grid_checkbox_x.stateChanged.connect(self.update_live_view)
                self.show_grid_checkbox_y = QCheckBox("Snap Y"); self.show_grid_checkbox_y.setToolTip("Snap vertically. Ctrl+Shift+Y or CMD+Shift+Y  toggles Y and Ctrl+Shift+G or CMD+Shift+G for both X and Y.")
                self.show_grid_checkbox_y.setFixedWidth(90)
                self.show_grid_checkbox_y.stateChanged.connect(self.update_live_view)
                self.grid_size_input = QSpinBox(); self.grid_size_input.setRange(5, 100); self.grid_size_input.setValue(20); self.grid_size_input.setPrefix("Grid (px): ")
                self.grid_size_input.valueChanged.connect(self.update_live_view)
                self.grid_size_input.setToolTip("Can increase or decrease grid pixel size by CTRL+Shift+Up or CTRL+Shift+Down")
                self.move_resize_button = QPushButton("Move/Resize"); self.move_resize_button.setToolTip("Toggle mode to move/resize custom markers and shapes on the image."); self.move_resize_button.setCheckable(True); self.move_resize_button.clicked.connect(self.toggle_custom_item_interaction_mode)
                self.modify_custom_marker_button = QPushButton("Modify All"); self.modify_custom_marker_button.setToolTip("Modify/Delete Custom Markers & Shapes"); self.modify_custom_marker_button.clicked.connect(self.open_modify_markers_dialog)
                row2_layout.addWidget(self.remove_custom_marker_button); row2_layout.addWidget(self.reset_custom_marker_button); row2_layout.addSpacing(10)
                row2_layout.addWidget(QLabel("Shapes:")); row2_layout.addWidget(self.draw_line_button); row2_layout.addWidget(self.draw_rect_button); row2_layout.addWidget(self.remove_shape_button); row2_layout.addSpacing(10)
                row2_layout.addWidget(self.show_grid_checkbox_x); row2_layout.addWidget(self.show_grid_checkbox_y); row2_layout.addWidget(self.grid_size_input); row2_layout.addStretch(1)
                row2_layout.addWidget(self.move_resize_button); row2_layout.addWidget(self.modify_custom_marker_button)
                custom_layout.addLayout(row2_layout)
                main_layout.addWidget(custom_group)

                main_layout.addStretch()
                return tab

            def open_modify_markers_dialog(self):
                self._backup_custom_markers_before_modify_dialog = [list(m) for m in self.custom_markers]
                self._backup_custom_shapes_before_modify_dialog = [dict(s) for s in self.custom_shapes] # Backup shapes
                
                if not hasattr(self, "custom_markers") or not isinstance(self.custom_markers, list):
                    self.custom_markers = []
                if not hasattr(self, "custom_shapes") or not isinstance(self.custom_shapes, list):
                    self.custom_shapes = []
            
                if not self.custom_markers and not self.custom_shapes:
                    QMessageBox.information(self, "No Items", "There are no custom markers or shapes to modify.")
                    return
            
                dialog = ModifyMarkersDialog(
                    [list(m) for m in self.custom_markers],
                    [dict(s) for s in self.custom_shapes],
                    self
                )
            
                dialog.global_markers_adjusted.connect(self.handle_live_marker_adjustment_preview)
                dialog.shapes_adjusted_preview.connect(self.handle_live_shape_adjustment_preview)
            
                if dialog.exec() == QDialog.Accepted:
                    modified_markers_tuples, modified_shapes_dicts = dialog.get_modified_markers_and_shapes()
                    modified_markers_lists = [list(m) for m in modified_markers_tuples]
            
                    markers_changed = (modified_markers_lists != self._backup_custom_markers_before_modify_dialog)
                    shapes_changed = (modified_shapes_dicts != self._backup_custom_shapes_before_modify_dialog) # FIX: Compare to shape backup
            
                    if markers_changed or shapes_changed:
                        self.save_state()
                        self.custom_markers = modified_markers_lists
                        self.custom_shapes = modified_shapes_dicts
                        self.is_modified = True
                        self.update_live_view()
                    else:
                        # If OK was clicked but no changes were made (e.g., user undid their changes in dialog)
                        self.custom_markers = self._backup_custom_markers_before_modify_dialog
                        # FIX: Also revert shapes if no changes were accepted.
                        self.custom_shapes = self._backup_custom_shapes_before_modify_dialog
                        self.update_live_view()
            
                else: # Dialog was cancelled or closed
                    # Revert both markers and shapes to their state before the dialog was opened.
                    self.custom_markers = self._backup_custom_markers_before_modify_dialog
                    # FIX: Add the missing line to revert shapes on cancel.
                    self.custom_shapes = self._backup_custom_shapes_before_modify_dialog
                    self.update_live_view() # Refresh to show the original state
            
                # Clean up the backup attributes
                if hasattr(self, '_backup_custom_markers_before_modify_dialog'):
                    del self._backup_custom_markers_before_modify_dialog
                if hasattr(self, '_backup_custom_shapes_before_modify_dialog'):
                    del self._backup_custom_shapes_before_modify_dialog
                
                # Disconnect the signals
                try:
                    dialog.global_markers_adjusted.disconnect(self.handle_live_marker_adjustment_preview)
                    dialog.shapes_adjusted_preview.disconnect(self.handle_live_shape_adjustment_preview)
                except (TypeError, RuntimeError): 
                    pass
            
            
            def handle_live_marker_adjustment_preview(self, temporarily_adjusted_markers):
                """
                Temporarily updates self.custom_markers and refreshes the live_view_label
                to show the effect of global adjustments from ModifyMarkersDialog.
                This does NOT mark the image as permanently modified or save to undo stack.
                """
                # We are directly modifying self.custom_markers here for the preview.
                # The backup in open_modify_markers_dialog handles reverting on cancel.
                self.custom_markers = temporarily_adjusted_markers # Update with the live adjusted list
                self.update_live_view()
                
            def handle_live_shape_adjustment_preview(self, temporarily_adjusted_shapes):
                """
                Temporarily updates self.custom_shapes and refreshes the live_view_label
                to show the effect of shape edits from ModifyMarkersDialog.
                """
                self.custom_shapes = temporarily_adjusted_shapes # Update with the live adjusted list
                self.update_live_view()
                         
            def _serialize_custom_markers(self, custom_markers_list):
                """Converts a list of custom marker tuples (with QColor) to a list of serializable dicts."""
                serialized_list = []
                for marker_tuple in custom_markers_list:
                    try:
                        # Ensure the tuple has the expected 8 elements
                        if len(marker_tuple) == 8:
                            x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_tuple
                            serialized_list.append({
                                "x": x, "y": y, "text": text, "color": qcolor.name(), # Store color name
                                "font_family": font_family, "font_size": font_size,
                                "bold": is_bold, "italic": is_italic
                            })
                        else:
                            # Handle older format or malformed data gracefully if necessary
                            # For now, we'll skip markers not matching the 8-element structure
                            print(f"Warning: Skipping marker during serialization due to unexpected format: {marker_tuple}")
                    except (ValueError, TypeError, IndexError, AttributeError) as e:
                        print(f"Warning: Skipping marker during serialization: {marker_tuple}, Error: {e}")
                return serialized_list

            def _deserialize_custom_markers(self, custom_markers_config_list):
                """Converts a list of serializable dicts back to custom marker tuples (with QColor)."""
                deserialized_list = []
                if not isinstance(custom_markers_config_list, list):
                    print("Warning: custom_markers_config_list is not a list, cannot deserialize.")
                    return []
                    
                for marker_conf in custom_markers_config_list:
                    if not isinstance(marker_conf, dict):
                        print(f"Warning: Skipping non-dictionary item in custom_markers_config_list: {marker_conf}")
                        continue
                    try:
                        color_str = marker_conf.get("color", "#000000") # Default to black if missing
                        qcolor = QColor(color_str)
                        if not qcolor.isValid():
                            print(f"Warning: Invalid color string '{color_str}' for marker, using black.")
                            qcolor = QColor(0,0,0) # Fallback to black

                        deserialized_list.append([
                            float(marker_conf.get("x", 0.0)),
                            float(marker_conf.get("y", 0.0)),
                            str(marker_conf.get("text", "")),
                            qcolor,
                            str(marker_conf.get("font_family", "Arial")),
                            int(marker_conf.get("font_size", 16)),
                            bool(marker_conf.get("bold", False)),
                            bool(marker_conf.get("italic", False))
                        ])
                    except (ValueError, TypeError, KeyError) as e:
                        print(f"Warning: Skipping marker during deserialization: {marker_conf}, Error: {e}")
                return deserialized_list
            
            def add_column(self):
                """Add a new column to the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                if current_text.strip():
                    self.top_marker_input.append("")  # Add a new line for a new column
                else:
                    self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
            
            def remove_column(self):
                """Remove the last column from the top marker labels."""
                current_text = self.top_marker_input.toPlainText()
                lines = current_text.split("\n")
                if len(lines) > 1:
                    lines.pop()  # Remove the last line
                    self.top_marker_input.setPlainText("\n".join(lines))
                else:
                    self.top_marker_input.clear()  # Clear the text if only one line exists

            
            def flip_vertical(self):
                self.save_state()
                """Flip the image vertically."""
                if self.image and not self.image.isNull():
                    transform = QTransform()
                    # Scale by -1 in Y, then translate back because scaling is around (0,0)
                    transform.scale(1, -1)
                    transform.translate(0, -self.image.height())
                    self.image = self.image.transformed(transform) # Get a new transformed image

                    # Update backups after transformation
                    if not self.image.isNull(): # Check if transform was successful
                        self.image_master = self.image.copy()
                        self.image_before_contrast = self.image.copy()
                        self.image_before_padding = self.image.copy() # Or None if padding invalidates this
                        self.image_contrasted = self.image.copy()
                        self.update_live_view()
                    else:
                        print("Warning: Vertical flip resulted in a null image.")
            
            def flip_horizontal(self):
                self.save_state()
                """Flip the image horizontally."""
                if self.image and not self.image.isNull():
                    transform = QTransform()
                    # Scale by -1 in X, then translate back
                    transform.scale(-1, 1)
                    transform.translate(-self.image.width(), 0)
                    self.image = self.image.transformed(transform) # Get a new transformed image

                    # Update backups after transformation
                    if not self.image.isNull(): # Check if transform was 
                        self.image_master = self.image.copy()
                        self.image_before_contrast = self.image.copy()
                        self.image_before_padding = self.image.copy() # Or None
                        self.image_contrasted = self.image.copy()
                        self.update_live_view()
                    else:
                        print("Warning: Horizontal flip resulted in a null image.")
            
            def convert_to_black_and_white(self):
                """
                Converts the current self.image to grayscale, preserving alpha channel.
                Aims for 16-bit grayscale precision, fixing 64-bit RGBA handling.
                """
                self.save_state()
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded.")
                    return

                original_format = self.image.format()
                
                # Check if already effectively grayscale
                if original_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16, QImage.Format_Mono]:
                    QMessageBox.information(self, "Info", f"Image is already grayscale (Format: {original_format}).")
                    return

                converted_image = None
                try:
                    np_img = self.qimage_to_numpy(self.image)
                    if np_img is None: raise ValueError("NumPy conversion failed.")

                    target_dtype = np.uint16
                    target_max_val = 65535.0

                    if np_img.ndim == 3 and np_img.shape[2] == 4: # 4-Channel (Color + Alpha)
                        alpha_part = np_img[..., 3]
                        color_part = np_img[..., :3]
                        
                        # --- FIX: Handle 16-bit RGBA correctly ---
                        if np_img.dtype == np.uint16:
                            # 16-bit RGBA (Qt Format_RGBA64) is R,G,B,A in memory
                            # OpenCV accepts uint16 and returns uint16
                            gray_data = cv2.cvtColor(color_part, cv2.COLOR_RGB2GRAY)
                        else:
                            # 8-bit BGRA (Qt Format_ARGB32) is B,G,R,A in memory
                            # OpenCV accepts uint8 and returns uint8
                            gray_8 = cv2.cvtColor(color_part, cv2.COLOR_BGR2GRAY)
                            # Scale 8-bit result to 16-bit
                            gray_data = (gray_8.astype(np.float32) / 255.0 * target_max_val).astype(target_dtype)

                        # Handle Alpha Channel Scaling
                        if alpha_part.dtype == np.uint16:
                            alpha_16 = alpha_part
                        else:
                            alpha_16 = (alpha_part.astype(np.float32) / 255.0 * target_max_val).astype(target_dtype)

                        # Construct 16-bit RGBA (R=G=B=Gray)
                        # numpy_to_qimage for 16-bit 4-channel expects RGBA order
                        out_img = np.dstack((gray_data, gray_data, gray_data, alpha_16))
                        
                        # --- FIX: Pass 16-bit array directly (Do NOT downsample to uint8) ---
                        converted_image = self.numpy_to_qimage(out_img)

                    elif np_img.ndim == 3 and np_img.shape[2] == 3: # 3-Channel (Color, No Alpha)
                        if np_img.dtype == np.uint16:
                            # 16-bit RGB
                            gray_data = cv2.cvtColor(np_img, cv2.COLOR_RGB2GRAY)
                        else:
                            # 8-bit BGR
                            gray_8 = cv2.cvtColor(np_img, cv2.COLOR_BGR2GRAY)
                            gray_data = (gray_8.astype(np.float32) / 255.0 * target_max_val).astype(target_dtype)
                        
                        # Convert to standard grayscale QImage (Format_Grayscale16)
                        converted_image = self.numpy_to_qimage(gray_data)

                    elif np_img.ndim == 2:
                         # Already grayscale, just promote bit depth if needed
                         if np_img.dtype == np.uint8:
                             gray_data = (np_img.astype(np.float32) / 255.0 * target_max_val).astype(target_dtype)
                             converted_image = self.numpy_to_qimage(gray_data)
                         else:
                             converted_image = self.image.copy()

                except Exception as e:
                     QMessageBox.critical(self, "Conversion Error", f"Could not convert image to grayscale: {e}")
                     import traceback
                     traceback.print_exc()
                     return

                if converted_image and not converted_image.isNull():
                     self.image = converted_image
                     self.image_master = self.image.copy()
                     self.image_before_contrast = self.image.copy()
                     self.image_contrasted = self.image.copy()
                     
                     if self.image_padded:
                         self.image_before_padding = None 
                         self.image_padded = False
                     else:
                         self.image_before_padding = self.image.copy()

                     self.reset_gamma_contrast()
                     self._update_status_bar()
                     self.update_live_view()
                     self._update_levels_histogram()
                else:
                     QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


            def invert_image(self):
                """Toggles the inversion state for the current adjustment context and re-applies all adjustments."""
                self.save_state()
                context = self.adjustment_context
                
                if context == "Main Image":
                    if not self.image or self.image.isNull():
                        QMessageBox.warning(self, "Invert Error", "No main image loaded to invert.")
                        return
                    self.main_image_is_inverted = not self.main_image_is_inverted
                
                elif context == "Overlay 1 (Base)":
                    if not hasattr(self, 'image1_original') or not self.image1_original:
                        QMessageBox.warning(self, "Invert Error", "No image loaded in Overlay 1 buffer.")
                        return
                    self.image1_adjustments['is_inverted'] = not self.image1_adjustments.get('is_inverted', False)

                elif context == "Overlay 2 (Overlay)":
                    if not hasattr(self, 'image2_original') or not self.image2_original:
                        QMessageBox.warning(self, "Invert Error", "No image loaded in Overlay 2 buffer.")
                        return
                    self.image2_adjustments['is_inverted'] = not self.image2_adjustments.get('is_inverted', False)

                # After toggling the flag, re-run the entire adjustment pipeline.
                # This will now apply inversion first, then all other effects.
                self.apply_all_adjustments()
                self._update_levels_histogram()

            def keyPressEvent(self, event):
                key = event.key()

                # --- Handle Deletion of Selected Items (Analysis Regions OR Custom Items) ---
                if key in (Qt.Key_Delete, Qt.Key_Backspace):
                    item_deleted = False

                    # Priority 1: Check if an ANALYSIS REGION is selected for deletion
                    if self.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"]:
                        selected_index = self.moving_multi_lane_index

                        if selected_index >= 0 and selected_index < len(self.multi_lane_definitions):
                            self.save_state()
                            del self.multi_lane_definitions[selected_index]
                            for i in range(selected_index, len(self.multi_lane_definitions)):
                                self.multi_lane_definitions[i]['id'] = i + 1
                            item_deleted = True
                        
                        if item_deleted:
                            self.is_modified = True
                            self.moving_multi_lane_index = -1
                            self.resizing_corner_index = -1
                            self.update_live_view()
                            event.accept()
                            return

                    # Priority 2: If no analysis region was deleted, check if a CUSTOM ITEM is selected
                    if not item_deleted and self.moving_custom_item_info:
                        info = self.moving_custom_item_info; item_type = info['type']; item_index = info['index']
                        custom_item_deleted = False
                        
                        if item_type == 'marker' and 0 <= item_index < len(self.custom_markers):
                            self.save_state(); del self.custom_markers[item_index]; custom_item_deleted = True
                        elif item_type == 'shape' and 0 <= item_index < len(self.custom_shapes):
                            self.save_state(); del self.custom_shapes[item_index]; custom_item_deleted = True
                        elif item_type in ['left_marker', 'right_marker', 'top_marker']:
                            marker_list_to_modify = None
                            if item_type == 'left_marker': marker_list_to_modify = self.left_markers
                            elif item_type == 'right_marker': marker_list_to_modify = self.right_markers
                            elif item_type == 'top_marker': marker_list_to_modify = self.top_markers
                            
                            if marker_list_to_modify is not None and 0 <= item_index < len(marker_list_to_modify):
                                self.save_state(); del marker_list_to_modify[item_index]
                                self._relabel_standard_markers(item_type); custom_item_deleted = True

                        if custom_item_deleted:
                            self.is_modified = True; self.moving_custom_item_info = None
                            self.update_live_view(); event.accept(); return

                # --- Escape Key Handling ---
                if key == Qt.Key_Escape:
                    a_mode_was_cancelled_or_view_reset = False
                    self._deactivate_all_previews()

                    if self.measurement_mode:
                        self.clear_measurement_mode()
                        a_mode_was_cancelled_or_view_reset = True
                    elif self.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"] and self.moving_multi_lane_index != -1:
                        self.moving_multi_lane_index = -1; self._reset_to_selection_mode(); self.setFocus(); event.accept(); return 
                    elif self.current_selection_mode in ["select_custom_item", "dragging_custom_item", "resizing_custom_item"]:
                        self.cancel_custom_item_interaction_mode(); a_mode_was_cancelled_or_view_reset = True
                    elif self.overlay_mode_active:
                        self.cancel_interactive_overlay_mode(); a_mode_was_cancelled_or_view_reset = True
                    elif self.current_selection_mode in ["select_for_move", "dragging_shape", "resizing_corner"]:
                        if self.current_selection_mode in ["dragging_shape", "resizing_corner"] and self.shape_points_at_drag_start_label:
                            pass # Revert logic can be added here if needed
                        self.cancel_selection_or_move_mode(); a_mode_was_cancelled_or_view_reset = True
                    
                    # --- START OF THE FIX ---
                    # The multi_lane_mode_active check is now the single point of truth for cancelling region definition.
                    elif self.multi_lane_mode_active:
                        self.cancel_multi_lane_mode()
                        a_mode_was_cancelled_or_view_reset = True
                    # --- The obsolete block that caused the crash has been removed. ---
                    # --- END OF THE FIX ---

                    elif self.crop_rectangle_mode:
                        self.cancel_rectangle_crop_mode(); self.live_view_label.clear_crop_preview(); a_mode_was_cancelled_or_view_reset = True
                    elif self.drawing_mode in ['line', 'rectangle']:
                        self.cancel_drawing_mode(); a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.preview_marker_enabled:
                        self.live_view_label.preview_marker_enabled = False; self.live_view_label.preview_marker_position = None; a_mode_was_cancelled_or_view_reset = True
                    elif self.marker_mode is not None:
                        self.marker_mode = None; a_mode_was_cancelled_or_view_reset = True
                    elif self.live_view_label.mw_predict_preview_enabled:
                        self.live_view_label.mw_predict_preview_enabled = False; self.live_view_label.mw_predict_preview_position = None
                        self.live_view_label.setMouseTracking(False); a_mode_was_cancelled_or_view_reset = True

                    self._reset_live_view_label_custom_handlers()

                    if self.live_view_label.zoom_level != 1.0 or a_mode_was_cancelled_or_view_reset:
                        self.live_view_label.zoom_level = 1.0; self.live_view_label.pan_offset = QPointF(0, 0); a_mode_was_cancelled_or_view_reset = True

                    if a_mode_was_cancelled_or_view_reset:
                        self.update_live_view()
                    event.accept()
                    return

                # --- Panning with Arrow Keys (Unchanged) ---
                if self.live_view_label.zoom_level != 1.0:
                    step = 20; offset_changed = False; current_x = self.live_view_label.pan_offset.x(); current_y = self.live_view_label.pan_offset.y()
                    if key == Qt.Key_Left: self.live_view_label.pan_offset.setX(current_x - step); offset_changed = True
                    elif key == Qt.Key_Right: self.live_view_label.pan_offset.setX(current_x + step); offset_changed = True
                    elif key == Qt.Key_Up: self.live_view_label.pan_offset.setY(current_y - step); offset_changed = True
                    elif key == Qt.Key_Down: self.live_view_label.pan_offset.setY(current_y + step); offset_changed = True
                    if offset_changed: self.update_live_view(); event.accept(); return
                
                # --- Overlay Nudging (Unchanged) ---
                if self.overlay_mode_active and self.selected_overlay_index > 0:
                    pos_slider_x = getattr(self, f'image{self.selected_overlay_index}_left_slider'); pos_slider_y = getattr(self, f'image{self.selected_overlay_index}_top_slider')
                    if pos_slider_x and pos_slider_y:
                        if key == Qt.Key_Left: pos_slider_x.setValue(pos_slider_x.value() - 1); event.accept(); return
                        elif key == Qt.Key_Right: pos_slider_x.setValue(pos_slider_x.value() + 1); event.accept(); return
                        elif key == Qt.Key_Up: pos_slider_y.setValue(pos_slider_y.value() - 1); event.accept(); return
                        elif key == Qt.Key_Down: pos_slider_y.setValue(pos_slider_y.value() + 1); event.accept(); return

                super().keyPressEvent(event)
                
            def toggle_custom_item_interaction_mode(self, checked):
                if checked:
                    # Cancel any other conflicting modes
                    self.cancel_drawing_mode()
                    self.cancel_rectangle_crop_mode()
                    self.cancel_selection_or_move_mode() # Cancel analysis area selection
                    if hasattr(self, 'marker_mode'): self.marker_mode = None

                    self.current_selection_mode = "select_custom_item"
                    self.live_view_label.mode = "select_custom_item"
                    self.moving_custom_item_info = None
                    self.resizing_corner_index = -1

                    self._reset_live_view_label_custom_handlers()
                    self.live_view_label._custom_left_click_handler_from_app = self.handle_custom_item_selection_click

                    QMessageBox.information(self, "Move/Resize Custom Items",
                                            "Click on a marker or shape to select it.\n"
                                            "Drag its body to move, or a corner handle to resize (shapes only).\n"
                                            "Press ESC or un-check the button to exit this mode.")
                    self.update_live_view()
                else:
                    self.cancel_custom_item_interaction_mode()

            def cancel_custom_item_interaction_mode(self):
                if self.current_selection_mode not in ["select_custom_item", "dragging_custom_item", "resizing_custom_item"]:
                    return

                self.current_selection_mode = None
                self.live_view_label.mode = None
                self.moving_custom_item_info = None
                self.resizing_corner_index = -1
                self.shape_points_at_drag_start_label = []

                self._reset_live_view_label_custom_handlers()

                if hasattr(self, 'move_resize_button') and self.move_resize_button.isChecked():
                    self.move_resize_button.setChecked(False)

                self.update_live_view()
            
            def _get_standard_marker_bounding_box_in_label_space(self, marker_type, marker_index):
                """
                Calculates the bounding shape (QRectF or QPolygonF) of a standard marker 
                in the LiveViewLabel's coordinate space.
                """
                if not (self.image and not self.image.isNull()): return None
                
                marker_list = None
                if marker_type == 'left_marker': marker_list = self.left_markers
                elif marker_type == 'right_marker': marker_list = self.right_markers
                elif marker_type == 'top_marker': marker_list = self.top_markers
                else: return None

                if not (0 <= marker_index < len(marker_list)): return None

                pos, text = marker_list[marker_index]
                
                font = QFont(self.font_family, self.font_size)
                fm = QFontMetrics(font)
                padding = 4 # Click padding
                text_rect_unscaled = fm.boundingRect(str(text)).adjusted(-padding, -padding, padding, padding)

                label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                img_w = float(self.image.width()); img_h = float(self.image.height())
                if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return None
                scale = min(label_w / img_w, label_h / img_h)
                offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0
                
                y_offset_baseline = fm.height() * 0.3

                if marker_type in ['left_marker', 'right_marker']:
                    y_ls = pos * scale + offset_y
                    full_text = f"{text} ⎯" if marker_type == 'left_marker' else f"⎯ {text}"
                    text_width_ls = fm.horizontalAdvance(full_text) + (2 * padding)
                    
                    x_ls = (self.left_marker_shift_added if marker_type == 'left_marker' else self.right_marker_shift_added) * scale + offset_x
                    
                    rect_height = text_rect_unscaled.height()
                    top_y = y_ls # Center around the baseline

                    if marker_type == 'left_marker':
                        return QRectF(x_ls - text_width_ls, top_y - rect_height / 2, text_width_ls, rect_height)
                    else:
                        return QRectF(x_ls, top_y - rect_height / 2, text_width_ls, rect_height)

                elif marker_type == 'top_marker':
                    # --- START MODIFICATION: Calculate a rotated QPolygonF ---
                    anchor_x_ls = pos * scale + offset_x
                    anchor_y_ls = self.top_marker_shift_added * scale + offset_y + y_offset_baseline

                    # Get the corners of the unrotated bounding rect, relative to a (0,0) draw point
                    corners = [
                        text_rect_unscaled.topLeft(),
                        text_rect_unscaled.topRight(),
                        text_rect_unscaled.bottomRight(),
                        text_rect_unscaled.bottomLeft()
                    ]

                    # Create a transform that rotates around (0,0) and then translates to the anchor.
                    # This exactly mimics the painter's transformation.
                    transform = QTransform()
                    transform.translate(anchor_x_ls, anchor_y_ls)
                    transform.rotate(self.font_rotation)
                    
                    # Map the corners of the original bounding rect to their new rotated and translated positions
                    rotated_corners = [transform.map(p) for p in corners]
                    
                    return QPolygonF(rotated_corners)
                    # --- END MODIFICATION ---
                    
                return None

            def handle_custom_item_selection_click(self, event):
                if self.current_selection_mode != "select_custom_item" or event.button() != Qt.LeftButton:
                    return

                clicked_point_ls = self.live_view_label.transform_point(event.position())
                click_radius_threshold = self.live_view_label.CORNER_HANDLE_BASE_RADIUS * 1.5
                item_selected = False

                # Iterate in reverse to select topmost items first
                # Check shape handles first (higher priority than body)
                for i in range(len(self.custom_shapes) - 1, -1, -1):
                    _body, handles_ls = self._get_shape_bounding_box_and_handles_in_label_space(i)
                    if handles_ls:
                        for corner_idx, handle_pt in enumerate(handles_ls):
                            if (clicked_point_ls - handle_pt).manhattanLength() < click_radius_threshold:
                                self.moving_custom_item_info = {'type': 'shape', 'index': i}
                                self.resizing_corner_index = corner_idx
                                self.shape_points_at_drag_start_label = handles_ls
                                item_selected = True
                                break
                    if item_selected: break

                # Check shape bodies if no handle was clicked
                if not item_selected:
                    for i in range(len(self.custom_shapes) - 1, -1, -1):
                        body_ls, handles_ls = self._get_shape_bounding_box_and_handles_in_label_space(i)
                        if body_ls and body_ls.contains(clicked_point_ls):
                            self.moving_custom_item_info = {'type': 'shape', 'index': i}
                            self.resizing_corner_index = -1 # Body move
                            self.shape_points_at_drag_start_label = handles_ls # Store corners for move calculation
                            item_selected = True
                            break

                # Check marker bodies if no shape was clicked
                if not item_selected:
                    for i in range(len(self.custom_markers) - 1, -1, -1):
                        bbox_ls = self._get_marker_bounding_box_in_label_space(i)
                        if bbox_ls and bbox_ls.contains(clicked_point_ls):
                            self.moving_custom_item_info = {'type': 'marker', 'index': i}
                            self.resizing_corner_index = -1 # Markers can't be resized this way
                            self.shape_points_at_drag_start_label = [bbox_ls.center()]
                            item_selected = True
                            break
                
                # Check standard markers if no custom item was selected
                if not item_selected:
                    for marker_type_str, marker_list in [('left_marker', self.left_markers), ('right_marker', self.right_markers), ('top_marker', self.top_markers)]:
                        for i in range(len(marker_list) - 1, -1, -1):
                            selection_shape = self._get_standard_marker_bounding_box_in_label_space(marker_type_str, i)
                            is_inside = False
                            if isinstance(selection_shape, QPolygonF):
                                is_inside = selection_shape.containsPoint(clicked_point_ls, Qt.OddEvenFill)
                            elif isinstance(selection_shape, QRectF):
                                is_inside = selection_shape.contains(clicked_point_ls)

                            if selection_shape and is_inside:
                                self.moving_custom_item_info = {'type': marker_type_str, 'index': i}
                                self.resizing_corner_index = -1
                                
                                # --- START FIX for DRAG ANCHOR ---
                                if marker_type_str == 'top_marker':
                                    # For top markers (QPolygonF), use its bounding rect's center.
                                    self.shape_points_at_drag_start_label = [selection_shape.boundingRect().center()]
                                else:
                                    # For L/R markers (QRectF), the object itself is the rect. Use its center.
                                    self.shape_points_at_drag_start_label = [selection_shape.center()]
                                # --- END FIX ---

                                item_selected = True
                                break
                        if item_selected: break

                if item_selected:
                    self.initial_mouse_pos_for_shape_drag_label = clicked_point_ls
                    if self.resizing_corner_index != -1:
                        self.current_selection_mode = "resizing_custom_item"
                        self.live_view_label.mode = "resizing_custom_item"
                    else:
                        self.current_selection_mode = "dragging_custom_item"
                        self.live_view_label.mode = "dragging_custom_item"

                    self.live_view_label.setCursor(Qt.CrossCursor)
                    self.live_view_label._custom_mouseMoveEvent_from_app = self.handle_custom_item_drag
                    self.live_view_label._custom_mouseReleaseEvent_from_app = self.handle_custom_item_drag_release
                else:
                    self.moving_custom_item_info = None
                    self.resizing_corner_index = -1

                self.update_live_view()

            def _relabel_standard_markers(self, marker_type):
                """
                Re-applies labels to a standard marker list from its source value list.
                This is called after a marker has been deleted to "shift" the subsequent labels up.
                """
                marker_list_to_update = None
                label_source_list = None

                if marker_type == 'left_marker':
                    marker_list_to_update = self.left_markers
                    label_source_list = self.marker_values
                elif marker_type == 'right_marker':
                    marker_list_to_update = self.right_markers
                    label_source_list = self.marker_values
                elif marker_type == 'top_marker':
                    marker_list_to_update = self.top_markers
                    label_source_list = self.top_label
                else:
                    return # Not a standard marker type

                # Iterate through the remaining markers and assign new labels from the source list
                for i in range(len(marker_list_to_update)):
                    original_position = marker_list_to_update[i][0]
                    # Get the corresponding label from the source list
                    new_label = label_source_list[i] if i < len(label_source_list) else ""
                    # Update the tuple in the list
                    marker_list_to_update[i] = (original_position, new_label)

            def handle_custom_item_drag(self, event):
                if not self.moving_custom_item_info or not (event.buttons() & Qt.LeftButton):
                    return

                current_mouse_pos_ls = self.live_view_label.transform_point(event.position())
                info = self.moving_custom_item_info

                # Coordinate transformation helper
                label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                img_w = float(self.image.width()); img_h = float(self.image.height())
                if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return
                scale = min(label_w / img_w, label_h / img_h)
                offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0
                def label_to_image(p_ls):
                    if scale == 0: return QPointF(0,0)
                    return QPointF((p_ls.x() - offset_x) / scale, (p_ls.y() - offset_y) / scale)

                # --- DRAG LOGIC ---
                if self.current_selection_mode == "dragging_custom_item":
                    raw_delta_ls = current_mouse_pos_ls - self.initial_mouse_pos_for_shape_drag_label
                    
                    if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                        if abs(raw_delta_ls.x()) > abs(raw_delta_ls.y()): raw_delta_ls.setY(0)
                        else: raw_delta_ls.setX(0)
                    
                    target_ref_point_ls = self.shape_points_at_drag_start_label[0] + raw_delta_ls
                    snapped_target_ref_point_ls = self.snap_point_to_grid(target_ref_point_ls)
                    effective_delta_ls = snapped_target_ref_point_ls - self.shape_points_at_drag_start_label[0]

                    if info['type'] == 'marker':
                        new_center_ls = self.shape_points_at_drag_start_label[0] + effective_delta_ls
                        new_center_img = label_to_image(new_center_ls)
                        self.custom_markers[info['index']][0] = new_center_img.x()
                        self.custom_markers[info['index']][1] = new_center_img.y()
                    
                    # --- START MODIFICATION ---
                    elif info['type'] in ['left_marker', 'right_marker']:
                        # Constrain movement to Y-axis
                        new_center_ls = QPointF(self.shape_points_at_drag_start_label[0].x(), snapped_target_ref_point_ls.y())
                        new_pos_img = label_to_image(new_center_ls)
                        marker_list = self.left_markers if info['type'] == 'left_marker' else self.right_markers
                        # Update only the position (index 0) of the tuple
                        original_text = marker_list[info['index']][1]
                        marker_list[info['index']] = (new_pos_img.y(), original_text)
                    
                    elif info['type'] == 'top_marker':
                        # Constrain movement to X-axis
                        new_center_ls = QPointF(snapped_target_ref_point_ls.x(), self.shape_points_at_drag_start_label[0].y())
                        new_pos_img = label_to_image(new_center_ls)
                        # Update only the position (index 0) of the tuple
                        original_text = self.top_markers[info['index']][1]
                        self.top_markers[info['index']] = (new_pos_img.x(), original_text)
                    # --- END MODIFICATION ---

                    elif info['type'] == 'shape':
                        new_points_ls = [p + effective_delta_ls for p in self.shape_points_at_drag_start_label]
                        new_points_img = [label_to_image(p) for p in new_points_ls]
                        shape_data = self.custom_shapes[info['index']]
                        if shape_data['type'] == 'rectangle':
                            xs = [p.x() for p in new_points_img]; ys = [p.y() for p in new_points_img]
                            shape_data['rect'] = (min(xs), min(ys), max(xs) - min(xs), max(ys) - min(ys))
                        elif shape_data['type'] == 'line':
                            shape_data['start'] = (new_points_img[0].x(), new_points_img[0].y())
                            shape_data['end'] = (new_points_img[1].x(), new_points_img[1].y())

                # --- RESIZE LOGIC ---
                elif self.current_selection_mode == "resizing_custom_item":
                    snapped_mouse_ls = self.snap_point_to_grid(current_mouse_pos_ls)
                    shape_data = self.custom_shapes[info['index']]
                    if shape_data['type'] == 'rectangle':
                        fixed_corner_ls = self.shape_points_at_drag_start_label[(self.resizing_corner_index + 2) % 4]
                        
                        # Constrain to square if Shift is pressed
                        if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                            delta_x = snapped_mouse_ls.x() - fixed_corner_ls.x()
                            delta_y = snapped_mouse_ls.y() - fixed_corner_ls.y()
                            max_delta = max(abs(delta_x), abs(delta_y))
                            
                            new_x = fixed_corner_ls.x() + (max_delta if delta_x > 0 else -max_delta)
                            new_y = fixed_corner_ls.y() + (max_delta if delta_y > 0 else -max_delta)
                            snapped_mouse_ls = QPointF(new_x, new_y)

                        p1_img = label_to_image(fixed_corner_ls); p2_img = label_to_image(snapped_mouse_ls)
                        xs = sorted([p1_img.x(), p2_img.x()]); ys = sorted([p1_img.y(), p2_img.y()])
                        shape_data['rect'] = (xs[0], ys[0], xs[1] - xs[0], ys[1] - ys[0])
                    elif shape_data['type'] == 'line':
                        fixed_endpoint_ls = self.shape_points_at_drag_start_label[(self.resizing_corner_index + 1) % 2]
                        
                        # Snap to angles if Shift is pressed
                        if QApplication.keyboardModifiers() == Qt.ShiftModifier:
                            delta = snapped_mouse_ls - fixed_endpoint_ls
                            angle_rad = np.arctan2(delta.y(), delta.x())
                            angle_deg = np.degrees(angle_rad)
                            
                            # Snap angle to nearest 45 degrees
                            snapped_angle_deg = round(angle_deg / 45.0) * 45.0
                            snapped_angle_rad = np.radians(snapped_angle_deg)
                            
                            # Recalculate the snapped mouse position
                            length = np.sqrt(delta.x()**2 + delta.y()**2)
                            snapped_mouse_ls = QPointF(
                                fixed_endpoint_ls.x() + length * np.cos(snapped_angle_rad),
                                fixed_endpoint_ls.y() + length * np.sin(snapped_angle_rad)
                            )

                        p1_img = label_to_image(fixed_endpoint_ls)
                        p2_img = label_to_image(snapped_mouse_ls)
                        
                        if self.resizing_corner_index == 0:
                            shape_data['start'], shape_data['end'] = (p2_img.x(), p2_img.y()), (p1_img.x(), p1_img.y())
                        else:
                            shape_data['start'], shape_data['end'] = (p1_img.x(), p1_img.y()), (p2_img.x(), p2_img.y())
                
                self.update_live_view()

            def handle_custom_item_drag_release(self, event):
                if self.current_selection_mode in ["dragging_custom_item", "resizing_custom_item"] and event.button() == Qt.LeftButton:
                    self.save_state()
                    self.is_modified = True
                    self.shape_points_at_drag_start_label = []
                    self.initial_mouse_pos_for_shape_drag_label = QPointF()
                    self.resizing_corner_index = -1
                    self.current_selection_mode = "select_custom_item"
                    self.live_view_label.mode = "select_custom_item"
                    self.live_view_label._custom_mouseMoveEvent_from_app = None
                    self.live_view_label._custom_mouseReleaseEvent_from_app = None
                    self.update_live_view()
            
            def _get_marker_bounding_box_in_label_space(self, marker_index):
                if not (self.image and not self.image.isNull()): return None
                try:
                    marker_data = self.custom_markers[marker_index]
                    x_img, y_img, text, _color, font_family, font_size, is_bold, is_italic = marker_data
                    label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                    img_w = float(self.image.width()); img_h = float(self.image.height())
                    if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return None
                    scale = min(label_w / img_w, label_h / img_h)
                    offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0
                    anchor_x_ls = x_img * scale + offset_x; anchor_y_ls = y_img * scale + offset_y
                    
                    font = QFont(font_family, int(font_size)); font.setBold(is_bold); font.setItalic(is_italic)
                    fm = QFontMetrics(font)
                    # Add a small padding to the bounding box to make it easier to click
                    padding = 4
                    text_rect = fm.boundingRect(text).adjusted(-padding, -padding, padding, padding)
                    
                    return QRectF(
                        anchor_x_ls - text_rect.width() / 2.0,
                        anchor_y_ls - text_rect.height() / 2.0,
                        text_rect.width(),
                        text_rect.height()
                    )
                except (IndexError, Exception):
                    return None

            def _get_shape_bounding_box_and_handles_in_label_space(self, shape_index):
                if not (self.image and not self.image.isNull()): return None, None
                try:
                    shape_data = self.custom_shapes[shape_index]
                    label_w = float(self.live_view_label.width()); label_h = float(self.live_view_label.height())
                    img_w = float(self.image.width()); img_h = float(self.image.height())
                    if not (label_w > 0 and label_h > 0 and img_w > 0 and img_h > 0): return None, None
                    scale = min(label_w / img_w, label_h / img_h)
                    offset_x = (label_w - img_w * scale) / 2.0; offset_y = (label_h - img_h * scale) / 2.0

                    def img_to_label(p): return QPointF(p[0] * scale + offset_x, p[1] * scale + offset_y)
                        
                    shape_type = shape_data.get('type')
                    body = None; handles = []
                    if shape_type == 'rectangle':
                        x, y, w, h = shape_data['rect']
                        p1 = img_to_label((x, y)); p2 = img_to_label((x + w, y + h))
                        body = QRectF(p1, p2).normalized()
                        handles = [body.topLeft(), body.topRight(), body.bottomRight(), body.bottomLeft()]
                    elif shape_type == 'line':
                        p1 = img_to_label(shape_data['start']); p2 = img_to_label(shape_data['end'])
                        # The "body" for clicking a line needs tolerance
                        body = QRectF(p1, p2).normalized().adjusted(-5, -5, 5, 5) 
                        handles = [p1, p2]
                    return body, handles
                except (IndexError, KeyError, Exception):
                    return None, None
            
            def update_marker_text_font(self, font: QFont):
                """
                Updates the font of self.custom_marker_text_entry based on the selected font
                from self.custom_font_type_dropdown.
            
                :param font: QFont object representing the selected font from the combobox.
                """
                # Get the name of the selected font
                selected_font_name = font.family()
                
                # Set the font of the QLineEdit
                self.custom_marker_text_entry.setFont(QFont(selected_font_name))
            
            def arrow_marker(self, text: str):
                self.custom_marker_text_entry.clear()
                self.custom_marker_text_entry.setText(text)
            
            def enable_custom_marker_mode(self):
                """Enable the custom marker mode and set the mouse event."""
                self.live_view_label.setCursor(Qt.ArrowCursor) # Often Arrow for text placement
                custom_text = self.custom_marker_text_entry.text().strip()
                self._reset_live_view_label_custom_handlers()
            
                self.live_view_label.preview_marker_enabled = True
                self.live_view_label.preview_marker_text = custom_text
                self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
                self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
                self.live_view_label.marker_color=self.custom_marker_color
                
                self.live_view_label.setFocus()
                self.live_view_label.update()
                
                self.marker_mode = "custom"  # Indicate custom marker mode
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.place_custom_marker(event, custom_text)
                # self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
                
            def remove_custom_marker_mode(self):
                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    self.custom_markers.pop()  # Remove the last entry from the list           
                self.update_live_view()  # Update the display
                
            def remove_last_custom_shape(self):
                """Removes the last added custom SHAPE (line or rectangle)."""
                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    self.save_state() # Save state before removal for undo
                    removed_shape = self.custom_shapes.pop()  # Remove the last shape
                    print(f"Removed last custom shape: {removed_shape}") # Optional debug
                    self.is_modified = True
                    self.update_live_view()  # Update the display
                    # Optionally, provide user feedback via status bar or QMessageBox
                    # self.statusBar().showMessage("Last custom shape removed.", 2000)
                else:
                    print("No custom shapes to remove.") # Info message
                    # Optionally, provide user feedback
                    # self.statusBar().showMessage("No custom shapes to remove.", 2000)
                
            def reset_custom_marker_mode(self):
                """Resets (clears) all custom MARKERS and custom SHAPES."""
                markers_cleared = False
                shapes_cleared = False
                needs_save = False # Flag to check if state needs saving

                if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
                    if not needs_save: # Save state only once before the first clear operation
                         self.save_state()
                         needs_save = True
                    self.custom_markers.clear()
                    markers_cleared = True
                    print("Cleared all custom markers.")

                if hasattr(self, "custom_shapes") and isinstance(self.custom_shapes, list) and self.custom_shapes:
                    if not needs_save: # Save state only once if markers weren't cleared but shapes are
                         self.save_state()
                         needs_save = True
                    self.custom_shapes.clear()
                    shapes_cleared = True
                    print("Cleared all custom shapes.")

                if markers_cleared or shapes_cleared:
                    self.is_modified = True
                    self.update_live_view()
                else:
                    print("No custom markers or shapes to reset.")
                
            
            def place_custom_marker(self, event, custom_text):
                """Place a custom marker at the cursor location."""
                self.save_state()
                # Get cursor position from the event
                pos = event.position()
                cursor_x, cursor_y = pos.x(), pos.y()
                
                
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                grid_size = self.grid_size_input.value() # Get grid size once

                if self.show_grid_checkbox_x.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_x = round(cursor_x / grid_size) * grid_size

                if self.show_grid_checkbox_y.isChecked():
                    if grid_size > 0: # Avoid division by zero
                         cursor_y = round(cursor_y / grid_size) * grid_size
            
                # Dimensions of the displayed image
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
            
                # Dimensions of the actual image
                image_width = self.image.width()
                image_height = self.image.height()
            
                # Calculate scaling factors
                scale = min(displayed_width / image_width, displayed_height / image_height)
            
                # Calculate offsets
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2
            
                # Transform cursor position to image space
                image_x = (cursor_x - x_offset) / scale
                image_y = (cursor_y - y_offset) / scale
                
                    
                # Store the custom marker's position and text
                self.custom_markers = getattr(self, "custom_markers", [])
                self.custom_markers.append([image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value(),False,False])
                self.update_live_view()
                
            def select_custom_marker_color(self):
                """Open a color picker dialog to select the color for custom markers."""
                color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
                if color.isValid():
                    self.custom_marker_color = color  # Update the custom marker color
                self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
            
            def update_all_labels(self):
                """Update all labels from the QTextEdit, supporting multiple columns."""
                self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
                input_text = self.top_marker_input.toPlainText()
                
                # Split the text into a list by commas and strip whitespace
                self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
                

                # if len(self.top_label) < len(self.top_markers):
                #     self.top_markers = self.top_markers[:len(self.top_label)]
                for i in range(0, len(self.top_markers)):
                    try:
                        self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                    except:
                        self.top_markers[i] = (self.top_markers[i][0], str(""))

                # if len(self.marker_values) < len(self.left_markers):
                #     self.left_markers = self.left_markers[:len(self.marker_values)]
                for i in range(0, len(self.left_markers)):
                    try:
                        self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
                    except:
                        self.left_markers[i] = (self.left_markers[i][0], str(""))
                        

                # if len(self.marker_values) < len(self.right_markers):
                #     self.right_markers = self.right_markers[:len(self.marker_values)]
                for i in range(0, len(self.right_markers)):
                    try:
                        self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
                    except:
                        self.right_markers[i] = (self.right_markers[i][0], str(""))
                        

                
                # Trigger a refresh of the live view
                self.update_live_view()
                
            def reset_marker(self, marker_type, param):    
                self.save_state()
                if marker_type == 'left':
                    if param == 'remove' and len(self.left_markers)!=0:
                        self.left_markers.pop()  
                        if self.current_left_marker_index > 0:
                            self.current_left_marker_index -= 1
                    elif param == 'reset':
                        self.left_markers.clear()
                        self.current_left_marker_index = 0  

                     
                elif marker_type == 'right' and len(self.right_markers)!=0:
                    if param == 'remove':
                        self.right_markers.pop()  
                        if self.current_right_marker_index > 0:
                            self.current_right_marker_index -= 1
                    elif param == 'reset':
                        self.right_markers.clear()
                        self.current_right_marker_index = 0

                elif marker_type == 'top' and len(self.top_markers)!=0:
                    if param == 'remove':
                        self.top_markers.pop() 
                        if self.current_top_label_index > 0:
                            self.current_top_label_index -= 1
                    elif param == 'reset':
                        self.top_markers.clear()
                        self.current_top_label_index = 0
            
                # Call update live view after resetting markers
                self.update_live_view()
                
            def duplicate_marker(self, marker_type):
                self.save_state() # Save state before making changes

                # --- NEW: Detect the actual content borders ---
                left_border, right_border = self._find_image_content_borders()

                if left_border is None or right_border is None:
                    QMessageBox.warning(self, "Border Detection Failed", 
                                        "Could not detect image content borders. Please ensure image has padding or clear boundaries.")
                    # Revert state if we can't proceed
                    if self.undo_stack: self.undo_stack.pop()
                    return

                if marker_type == 'left' and self.right_markers:
                    # Copy Right Markers TO Left
                    self.left_markers = self.right_markers.copy()
                    
                    # --- FIX: Set offset to the detected LEFT border ---
                    self.left_marker_shift_added = left_border

                    # Update Left Slider position to match the detected border
                    if hasattr(self, 'left_padding_slider'):
                        min_val, max_val = self.left_padding_slider.minimum(), self.left_padding_slider.maximum()
                        clamped_value = max(min_val, min(self.left_marker_shift_added, max_val))
                        self.left_padding_slider.blockSignals(True)
                        self.left_padding_slider.setValue(clamped_value)
                        self.left_padding_slider.blockSignals(False)
                        # Re-sync internal variable with the actual slider value after potential clamping
                        self.left_marker_shift_added = self.left_padding_slider.value()

                elif marker_type == 'right' and self.left_markers:
                    # Copy Left Markers TO Right
                    self.right_markers = self.left_markers.copy()

                    # --- FIX: Set offset to the detected RIGHT border ---
                    self.right_marker_shift_added = right_border

                    # Update Right Slider position to match the detected border
                    if hasattr(self, 'right_padding_slider'):
                        min_val, max_val = self.right_padding_slider.minimum(), self.right_padding_slider.maximum()
                        clamped_value = max(min_val, min(self.right_marker_shift_added, max_val))
                        self.right_padding_slider.blockSignals(True)
                        self.right_padding_slider.setValue(clamped_value)
                        self.right_padding_slider.blockSignals(False)
                        # Re-sync internal variable
                        self.right_marker_shift_added = self.right_padding_slider.value()

                # Call update live view after duplicating markers and updating offsets/sliders
                self.update_live_view()
                
            def on_combobox_changed(self):
                preset_name = self.combo_box.currentText()

                if not self._is_restoring_state and (self.is_modified or preset_name != "Custom"):
                    self.save_state()
                
                if self.is_modified or preset_name != "Custom":
                    self.save_state()

                if preset_name == "Custom":
                    self.marker_values_textbox.setEnabled(True)
                    self.rename_input.setEnabled(True)
                    self.rename_input.clear()
                    self.marker_values_textbox.setPlaceholderText("Current L/R values, edit to save as new")
                    current_lr_display_values = [str(v) for v in getattr(self, 'marker_values', [])]
                    self.marker_values_textbox.setText(", ".join(current_lr_display_values))
                    current_top_display_labels = [str(v) for v in getattr(self, 'top_label', [])]
                    self.top_marker_input.setText(", ".join(current_top_display_labels))

                elif preset_name in self.presets_data:
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.rename_input.clear()

                    preset_config = self.presets_data[preset_name]
                    loaded_settings = preset_config.get("peak_dialog_settings", {})
                    # Update the main app's settings dictionary with the loaded values
                    self.peak_dialog_settings.update(loaded_settings)
                    
                    self.marker_values = list(preset_config.get("marker_values", []))
                    display_marker_values = [str(v) for v in self.marker_values]
                    self.marker_values_textbox.setText(", ".join(display_marker_values))

                    self.top_label = list(preset_config.get("top_labels", []))
                    self.top_marker_input.setText(", ".join(map(str, self.top_label)))

                    # --- START FIX: Make loading custom items conditional ---
                    if self.load_custom_from_preset_checkbox.isChecked():
                        # This code now only runs if the user wants to load custom items from the preset.
                        custom_markers_config_from_preset = preset_config.get("custom_markers_config", [])
                        if not isinstance(custom_markers_config_from_preset, list):
                            custom_markers_config_from_preset = []
                        deserialized_markers = self._deserialize_custom_markers(custom_markers_config_from_preset)
                        self.custom_markers = [list(m) for m in deserialized_markers]

                        custom_shapes_config_from_preset = preset_config.get("custom_shapes_config", [])
                        if not isinstance(custom_shapes_config_from_preset, list):
                            custom_shapes_config_from_preset = []
                        self.custom_shapes = [dict(s) for s in custom_shapes_config_from_preset if isinstance(s, dict)]
                    # --- END FIX ---
                    
                    self.update_all_labels() 
                    
                else: # Preset name not found in data
                    self.marker_values_textbox.setEnabled(False)
                    self.rename_input.setEnabled(False)
                    self.marker_values_textbox.clear()
                    self.top_marker_input.clear()
                    
                    # --- START FIX: Make clearing custom items conditional ---
                    if self.load_custom_from_preset_checkbox.isChecked():
                        self.custom_markers.clear()
                        self.custom_shapes.clear()
                    # --- END FIX ---

                    self.marker_values = []
                    self.top_label = []
                    self.update_live_view()

            
            def reset_gamma_contrast(self):
                try:
                    if self.image_before_contrast==None:
                        self.image_before_contrast=self.image_master.copy()
                    self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
                    self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
                    self.high_slider.setValue(100)  # Reset contrast to default
                    self.low_slider.setValue(100)  # Reset contrast to default
                    self.gamma_slider.setValue(100)  # Reset gamma to default
                    self.update_live_view()
                except:
                    pass

            
            def update_image_contrast(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                    
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                        self.update_live_view()
                except:
                    pass
            
            def update_image_gamma(self):
                try:
                    if self.contrast_applied==False:
                        self.image_before_contrast=self.image.copy()
                        self.contrast_applied=True
                        
                    if self.image:
                        high_contrast_factor = self.high_slider.value() / 100.0
                        low_contrast_factor = self.low_slider.value() / 100.0
                        gamma_factor = self.gamma_slider.value() / 100.0
                        self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                        self.update_live_view()
                except:
                    pass
            
            def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
                """
                Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
                preserving the original format (including color and bit depth) where possible.
                Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
                """
                if not qimage or qimage.isNull():
                    return qimage

                original_format = qimage.format()
                try:
                    img_array = self.qimage_to_numpy(qimage)
                    if img_array is None: raise ValueError("NumPy conversion failed.")

                    # Work with float64 for calculations to avoid precision issues
                    img_array_float = img_array.astype(np.float64)

                    # Determine max value based on original data type
                    if img_array.dtype == np.uint16:
                        max_val = 65535.0
                    elif img_array.dtype == np.uint8:
                        max_val = 255.0
                    else: # Default for unexpected types (e.g., float input?)
                         max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0

                    # --- Apply adjustments ---
                    if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                        num_channels = img_array.shape[2]
                        adjusted_channels = []

                        # Process only the color channels (first 3 usually)
                        channels_to_process = min(num_channels, 3)
                        for i in range(channels_to_process):
                            channel = img_array_float[:, :, i]
                            # Normalize to 0-1 range
                            channel_norm = channel / max_val

                            # Apply brightness (high_factor): Multiply
                            channel_norm = channel_norm * high_factor

                            # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                            mid_grey = 0.5
                            contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                            channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                            # Clip to 0-1 range after contrast/brightness
                            channel_norm = np.clip(channel_norm, 0.0, 1.0)

                            # Apply gamma correction
                            safe_gamma = max(0.01, gamma)
                            channel_norm = np.power(channel_norm, safe_gamma)

                            # Clip again after gamma
                            channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                            # Scale back to original range
                            adjusted_channels.append(channel_norm_clipped * max_val)

                        # Reconstruct the image array
                        img_array_final_float = np.stack(adjusted_channels, axis=2)

                        # Keep the alpha channel (if present) untouched
                        if num_channels == 4:
                            alpha_channel = img_array_float[:, :, 3] # Get original alpha
                            img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

                    elif img_array.ndim == 2: # Grayscale Image
                        # Normalize to 0-1 range
                        img_array_norm = img_array_float / max_val

                        # Apply brightness
                        img_array_norm = img_array_norm * high_factor

                        # Apply contrast
                        mid_grey = 0.5
                        contrast_factor = max(0.01, low_factor)
                        img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                        # Clip
                        img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                        # Apply gamma
                        safe_gamma = max(0.01, gamma)
                        img_array_norm = np.power(img_array_norm, safe_gamma)

                        # Clip again
                        img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                        # Scale back
                        img_array_final_float = img_array_norm_clipped * max_val
                    else:
                        return qimage # Return original if unsupported dimensions

                    # Convert back to original data type
                    img_array_final = img_array_final_float.astype(original_dtype)

                    # Convert back to QImage using the helper function
                    result_qimage = self.numpy_to_qimage(img_array_final)
                    if result_qimage.isNull():
                        raise ValueError("Conversion back to QImage failed.")

                    # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
                    return result_qimage

                except Exception as e:
                    traceback.print_exc() # Print detailed traceback
                    return qimage # Return original QImage on error
            

            def save_contrast_options(self):
                if self.image:
                    self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
                    self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
                else:
                    QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

            def remove_config(self): # This is for "Remove Preset" button
                selected_preset_name = self.combo_box.currentText()

                if selected_preset_name == "Custom":
                    QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' option.")
                    return
                
                # Prevent deletion of essential default presets if desired (example)
                # essential_defaults = ["Precision Plus Protein All Blue Prestained (Bio-Rad)", "1 kb Plus DNA Ladder (Thermo 10787018)"]
                # if selected_preset_name in essential_defaults:
                #     QMessageBox.warning(self, "Error", f"Cannot remove the built-in preset: '{selected_preset_name}'.")
                #     return

                if selected_preset_name not in self.presets_data:
                    QMessageBox.warning(self, "Error", f"Preset '{selected_preset_name}' not found in data.")
                    return

                reply = QMessageBox.question(self, "Confirm Remove",
                                             f"Are you sure you want to remove the preset '{selected_preset_name}'?",
                                             QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return

                try:
                    del self.presets_data[selected_preset_name]

                    # Save the updated configuration
                    config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)

                    # Update ComboBox
                    self.combo_box.blockSignals(True)
                    current_idx = self.combo_box.findText(selected_preset_name)
                    if current_idx != -1:
                        self.combo_box.removeItem(current_idx)
                    # Select "Custom" or the first available preset after removal
                    if self.combo_box.count() > 1: # More than just "Custom" left
                        custom_idx = self.combo_box.findText("Custom")
                        if self.combo_box.currentIndex() == custom_idx and custom_idx > 0: # If custom was selected and not first
                             self.combo_box.setCurrentIndex(custom_idx -1) # Select item before custom
                        elif self.combo_box.currentIndex() == custom_idx and custom_idx == 0 and self.combo_box.count() > 1: # custom is first, but others exist
                            self.combo_box.setCurrentIndex(1)
                        # else keep current selection if it's not the removed one
                    elif self.combo_box.count() == 1 and self.combo_box.itemText(0) == "Custom":
                        self.combo_box.setCurrentIndex(0)

                    self.combo_box.blockSignals(False)
                    self.on_combobox_changed() # Refresh UI based on new selection

                    QMessageBox.information(self, "Success", f"Preset '{selected_preset_name}' removed.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Error removing preset: {e}")
                    traceback.print_exc()

            def save_config(self): # This is for "Save Preset" button
                """Saves the current settings (L/R/Top values, custom markers/shapes) to the selected/new preset name."""
                
                current_combo_text = self.combo_box.currentText()
                new_name_from_input = self.rename_input.text().strip()
                
                name_to_save = ""
                is_new_preset_creation = False

                if new_name_from_input: # User entered a name in rename_input
                    name_to_save = new_name_from_input
                    is_new_preset_creation = True
                    if name_to_save == "Custom":
                        QMessageBox.warning(self, "Invalid Name", "Cannot save a preset with the name 'Custom'. Please choose another name or clear the rename field to update the selected preset.")
                        return
                elif current_combo_text != "Custom": # Updating an existing preset
                    name_to_save = current_combo_text
                else: # "Custom" is selected, and rename_input is empty
                    QMessageBox.information(self, "Save Preset", "Please enter a new name in the 'New name for Custom preset' field to save the current settings as a new preset.")
                    return

                # Confirmation for overwriting
                if name_to_save in self.presets_data and is_new_preset_creation : # Only ask if it's a new name that already exists
                     reply = QMessageBox.question(self, "Overwrite Preset?",
                                                      f"A preset named '{name_to_save}' already exists. Overwrite it?",
                                                      QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                     if reply == QMessageBox.No:
                         return
                
                # --- Gather data to save ---
                # L/R Marker Values: Always take from the textbox.
                try:
                    raw_markers_text = self.marker_values_textbox.text().strip("[]")
                    if raw_markers_text:
                        marker_values_to_save = [
                            int(num.strip()) if num.strip().isdigit() else float(num.strip()) if '.' in num.strip() else num.strip()
                            for num in raw_markers_text.split(",") if num.strip()
                        ]
                    else:
                        marker_values_to_save = []
                except ValueError:
                    QMessageBox.warning(self, "Input Error", "Invalid format for L/R marker values. Please use comma-separated numbers or text.")
                    return

                # Top Labels: Always take from the textedit.
                top_labels_to_save = [label.strip() for label in self.top_marker_input.toPlainText().split(",") if label.strip()]
                
                # Custom Markers: Serialize current self.custom_markers
                custom_markers_config_to_save = self._serialize_custom_markers(getattr(self, 'custom_markers', []))
                
                # Custom Shapes: Are already serializable
                custom_shapes_config_to_save = [dict(s) for s in getattr(self, 'custom_shapes', [])] # Ensure they are dicts

                # --- Update self.presets_data in memory ---
                self.presets_data[name_to_save] = {
                    "marker_values": marker_values_to_save,
                    "top_labels": top_labels_to_save,
                    "custom_markers_config": custom_markers_config_to_save,
                    "custom_shapes_config": custom_shapes_config_to_save,
                    "peak_dialog_settings": self.peak_dialog_settings.copy() # Save a copy of the current settings
                }

                # --- Save to file ---
                config_filepath = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), self.CONFIG_PRESET_FILE_NAME)
                try:
                    with open(config_filepath, "w", encoding='utf-8') as f:
                        json.dump({"presets": self.presets_data}, f, indent=4)
                    
                    # --- Update ComboBox if a new preset was created ---
                    if is_new_preset_creation and self.combo_box.findText(name_to_save) == -1:
                        self.combo_box.blockSignals(True)
                        custom_idx = self.combo_box.findText("Custom")
                        if custom_idx != -1:
                            self.combo_box.insertItem(custom_idx, name_to_save)
                        else:
                            self.combo_box.addItem(name_to_save) # Fallback
                        self.combo_box.setCurrentText(name_to_save) # Select the new/updated preset
                        self.combo_box.blockSignals(False)
                        # self.on_combobox_changed() # Triggered by setCurrentText if not blocked

                    self.rename_input.clear() # Clear rename field after successful save
                    self.marker_values_textbox.setEnabled(False) # Disable textbox after saving to a named preset
                    self.rename_input.setEnabled(False)

                    QMessageBox.information(self, "Preset Saved", f"Preset '{name_to_save}' saved successfully.")

                except Exception as e:
                    QMessageBox.critical(self, "Save Preset Error", f"Could not save preset configuration:\n{e}")
                    traceback.print_exc()
            
            
            def load_config(self):
                """
                Load preset configuration from file. If the file doesn't exist,
                create it with default popular marker standards and empty custom templates.
                """
                config_loaded_successfully = False
                self.presets_data.clear() # Start with an empty dictionary

                if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    application_path = os.path.dirname(sys.executable)
                elif getattr(sys, 'frozen', False):
                     application_path = os.path.dirname(sys.executable)
                else:
                    try: application_path = os.path.dirname(os.path.abspath(__file__))
                    except NameError: application_path = os.getcwd()

                config_filepath = os.path.join(application_path, self.CONFIG_PRESET_FILE_NAME)
                print(f"INFO: Attempting to load/create preset config at: {config_filepath}")

                default_marker_standards = {
                    # Proteins (kDa)
                    "Precision Plus Protein All Blue Prestained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Unstained (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "Precision Plus Protein Dual Color (Bio-Rad)": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                    "PageRuler Prestained Protein Ladder (Thermo)": [250, 130, 100, 70, 55, 35, 25, 15, 10],
                    "PageRuler Unstained Protein Ladder (Thermo)": [200, 150, 100, 70, 50, 40, 30, 20, 10],
                    "Spectra Multicolor Broad Range Protein Ladder (Thermo)": [260, 140, 100, 75, 60, 45, 35, 25, 15, 10],
                    # DNA (bp)
                    "1 kb DNA Ladder (NEB N3232)": [10000, 8000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 500],
                    "1 kb Plus DNA Ladder (Invitrogen 10787018)": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
                    "GeneRuler 1 kb Plus DNA Ladder (Thermo SM1331)": [20000, 10000, 7000, 5000, 4000, 3000, 2000, 1500, 1000, 700, 500, 400, 300, 200, 75],
                    "Lambda DNA/HindIII Marker (NEB N3012)": [23130, 9416, 6557, 4361, 2322, 2027, 564, 125],
                }
                # Generic default top labels
                default_top_labels_generic = ["MWM", "L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9", "L10", "L11", "L12", "L13", "L14", "L15"]
                # default_top_label_dict = {name: default_top_labels_generic[:] for name in default_marker_values.keys()} # Assign generic to all defaults
                default_presets_init = {}
                for name, markers in default_marker_standards.items():
                    default_presets_init[name] = {
                        "marker_values": list(markers), # Ensure it's a list
                        "top_labels": list(default_top_labels_generic), # Ensure it's a list
                        "custom_markers_config": [], # Empty by default
                        "custom_shapes_config": []   # Empty by default
                    }
                # --- End Default Marker Data ---

                # --- Initialize internal dictionaries BEFORE trying to load/create ---
                # Start with defaults, then overwrite if load is successful.
                if os.path.exists(config_filepath):
                    try:
                        with open(config_filepath, "r", encoding='utf-8') as f:
                            loaded_json = json.load(f)

                        self.viewer_position = loaded_json.get("viewer_position", "Top")
                        
                        # Check for new "presets" key
                        if "presets" in loaded_json and isinstance(loaded_json["presets"], dict):
                            self.presets_data = loaded_json["presets"]
                            # Ensure all default fields exist for loaded presets
                            for preset_name, data in self.presets_data.items():
                                if "marker_values" not in data: data["marker_values"] = []
                                if "top_labels" not in data: data["top_labels"] = []
                                if "custom_markers_config" not in data: data["custom_markers_config"] = []
                                if "custom_shapes_config" not in data: data["custom_shapes_config"] = []
                            config_loaded_successfully = True
                        else: # Try to migrate old format
                            print("INFO: Old config format detected. Attempting migration...")
                            migrated_data = {}
                            old_marker_values = loaded_json.get("marker_values", {})
                            old_top_labels = loaded_json.get("top_label", {})
                            all_preset_names = set(old_marker_values.keys()) | set(old_top_labels.keys())

                            for name in all_preset_names:
                                migrated_data[name] = {
                                    "marker_values": list(old_marker_values.get(name, [])),
                                    "top_labels": list(old_top_labels.get(name, default_top_labels_generic)),
                                    "custom_markers_config": [], # No custom markers in old format
                                    "custom_shapes_config": []   # No custom shapes in old format
                                }
                            self.presets_data = migrated_data
                            # Save immediately in new format after migration
                            with open(config_filepath, "w", encoding='utf-8') as f_new:
                                json.dump({"viewer_position": self.viewer_position, "presets": self.presets_data}, f_new, indent=4)
                            print("INFO: Config migrated to new format and saved.")
                            config_loaded_successfully = True

                    except (json.JSONDecodeError, IOError, TypeError) as e:
                        QMessageBox.warning(self, "Preset Config Load Error", f"Could not load '{self.CONFIG_PRESET_FILE_NAME}':\n{e}\n\nUsing defaults.")
                        self.presets_data = default_presets_init.copy()
                        self.viewer_position = "Top"
                        
                    except Exception as e:
                        traceback.print_exc()
                        QMessageBox.warning(self, "Preset Config Load Error",
                                            f"Unexpected error loading '{self.CONFIG_PRESET_FILE_NAME}'.\n\nUsing default presets.")
                        self.presets_data = default_presets_init.copy() # Fallback
                else:
                    self.presets_data = default_presets_init.copy()
                    self.viewer_position = "Top" 
                    try:
                        with open(config_filepath, "w", encoding='utf-8') as f:
                            json.dump({"viewer_position": self.viewer_position, "presets": self.presets_data}, f, indent=4)
                        config_loaded_successfully = True # Created successfully
                        print(f"INFO: Default preset config created at: {config_filepath}")
                    except Exception as e:
                        QMessageBox.critical(self, "Preset Config Creation Error",
                                             f"Could not create default preset config file:\n{e}\n\nDefault presets will be used for this session only.")
                        # self.presets_data remains default_presets_init

                # --- Update UI (ComboBox) ---
                try:
                    if hasattr(self, 'combo_box'):
                        self.combo_box.blockSignals(True)
                        self.combo_box.clear()
                        sorted_preset_names = sorted(self.presets_data.keys())
                        self.combo_box.addItems(sorted_preset_names)
                        self.combo_box.addItem("Custom")
                        default_biorad = "Precision Plus Protein All Blue Prestained (Bio-Rad)"
                        idx_to_select = self.combo_box.findText(default_biorad)
                        if idx_to_select == -1 and sorted_preset_names:
                            idx_to_select = 0 # Select first actual preset if BioRad not found
                        elif idx_to_select == -1: # No actual presets, only "Custom"
                            idx_to_select = self.combo_box.findText("Custom")
                        
                        if idx_to_select != -1:
                             self.combo_box.setCurrentIndex(idx_to_select)
                        
                        self.combo_box.blockSignals(False)
                        self.on_combobox_changed() # Trigger update based on selection
                    else:
                        print("Warning: combo_box UI element not found during config load.")
                except Exception as e_ui:
                    traceback.print_exc()
                    QMessageBox.warning(self, "UI Error", f"Error populating preset combobox: {e_ui}")
            
            def paste_image(self):
                """Handle pasting image from clipboard."""
                self.is_modified = True
                self.reset_image() # Clear previous state first
                self.load_image_from_clipboard()
                self._update_overlay_slider_ranges()
                # UI updates (label size, sliders) should happen within load_image_from_clipboard
                self.update_live_view()
                self.save_state() # Save state after pasting
            
            def load_image_from_clipboard(self):
                """
                Load an image from the clipboard into self.image, preserving format.
                Prioritizes file paths over raw image data to handle macOS file copying correctly.
                If loaded from a file path, attempts to load an associated config file.
                """
                # Check for unsaved changes before proceeding
                # if not self.prompt_save_if_needed(): # Optional: uncomment if needed
                #     return # Abort paste if user cancels save

                self.reset_image() # Clear previous state first

                clipboard = QApplication.clipboard()
                mime_data = clipboard.mimeData()
                loaded_image = None
                source_info = "Clipboard" # Default source
                config_loaded_from_paste = False # Flag to track if config was loaded
                

                # --- PRIORITY 1: Check for File URLs ---
                if mime_data.hasUrls():
                    urls = mime_data.urls()
                    if urls:
                        file_path = urls[0].toLocalFile()
                        if os.path.isfile(file_path) and file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                            loaded_image = QImage(file_path)
                            source_info = file_path

                            if loaded_image.isNull():
                                try:
                                    pil_image = Image.open(file_path)
                                    # Use our numpy converter which correctly handles 16-bit arrays
                                    np_array = np.array(pil_image)
                                    loaded_image = self.numpy_to_qimage(np_array)
                                    if loaded_image.isNull():
                                        loaded_image = None # Ensure it's None if conversion failed
                                    else:
                                        source_info = f"{file_path} (Pillow/NumPy)" # Update source info for clarity
                                except Exception as e:
                                    QMessageBox.warning(self, "File Load Error", f"Could not load image from file '{os.path.basename(file_path)}':\n{e}")
                                    loaded_image = None

                            # --- CONFIG FILE LOADING FOR PASTED FILE ---
                            if loaded_image and not loaded_image.isNull():
                                self.image_path = file_path # Store the path early for config lookup
                                base_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                                # Construct potential config file name (remove _original/_modified if present)
                                config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                                config_path = os.path.join(os.path.dirname(file_path), f"{config_base}_config.txt")

                                if os.path.exists(config_path):
                                    try:
                                        with open(config_path, "r") as config_file:
                                            config_data = json.load(config_file)
                                        self.apply_config(config_data,load_analysis=False) # Apply loaded settings
                                        config_loaded_from_paste = True # Set flag
                                    except Exception as e:
                                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply associated config file '{os.path.basename(config_path)}': {e}")


                # --- PRIORITY 2: Check for Raw Image Data (if no valid file was loaded) ---
                if loaded_image is None and mime_data.hasImage():
                    loaded_image = clipboard.image()
                    if not loaded_image.isNull():
                        source_info = "Clipboard Image Data"
                    else:
                        try:
                            pil_image = ImageGrab.grabclipboard()
                            if isinstance(pil_image, Image.Image):
                                loaded_image = ImageQt.ImageQt(pil_image)
                                if loaded_image.isNull():
                                     loaded_image = None
                                else:
                                    source_info = "Clipboard Image Data (Pillow Grab)"
                            else:
                                loaded_image = None
                        except Exception: # Keep quiet on Pillow grab errors
                            loaded_image = None

                # --- Process the successfully loaded image ---
                if loaded_image and not loaded_image.isNull():
                    self.is_modified = True
                    self.image = loaded_image

                    # --- ADDED: Auto-invert 16-bit (Only if no config loaded) ---
                    if not config_loaded_from_paste:
                        if self.image.format() == QImage.Format_Grayscale16:
                            self.main_image_is_inverted = True
                        else:
                            self.main_image_is_inverted = False
                    # -----------------------------------------------------------

                    # Initialize backups
                    self.image_master = self.image.copy()
                    self.original_image = self.image.copy()
                    self.image_before_padding = None
                    self.image_contrasted = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False
                    # self.image_path is set earlier if loaded from file

                    # --- Update UI Elements ---
                    try:
                        w = self.image.width()
                        h = self.image.height()
                        if w <= 0 or h <= 0: raise ValueError("Invalid image dimensions after load.")

                        # Update preview label size
                        ratio=w/h if h > 0 else 1
                        target_width = int(self.label_size) # Use current label_size setting
                        target_height = int(target_width / ratio)

                        # Update slider ranges based on *new* label size
                        render_scale = 3
                        render_width = self.live_view_label.width() * render_scale
                        render_height = self.live_view_label.height() * render_scale
                        self._update_marker_slider_ranges()

                        # Set recommended padding only if config wasn't loaded
                        if not config_loaded_from_paste:
                            self.left_padding_input.setText(str(int(w * 0.1)))
                            self.right_padding_input.setText(str(int(w * 0.1)))
                            self.top_padding_input.setText(str(int(h * 0.15)))
                            self.bottom_padding_input.setText("0")

                        # Update window title and status bar
                        display_source = os.path.basename(source_info.split(" (")[0]) # Show filename or simplified source
                        title_suffix = " (+ Config)" if config_loaded_from_paste else ""
                        self.setWindowTitle(f"{self.window_title}::{display_source}{title_suffix}")
                        self._update_status_bar()

                    except Exception as e:
                        QMessageBox.warning(self, "UI Update Error", f"Could not update UI elements after pasting: {e}")
                    
                    # --- ADDED: Apply adjustments ---
                    self.apply_all_adjustments()
                    # ------------------------------

                    enable_pan = self.live_view_label.zoom_level > 1.0
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                    QTimer.singleShot(0, self.update_live_view) # Render the loaded image
                    self._update_levels_histogram() # Update histogram for new image
                    self.save_state()

                else:
                    # If no image was successfully loaded
                    QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
                    # Clean up state
                    self.image = None
                    self.original_image = None
                    self.image_master = None
                    self.image_path = None
                    self.live_view_label.clear()
                    self.setWindowTitle(self.window_title)
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                    self._update_status_bar()
                    self.adjustSize()
                    QTimer.singleShot(0, self.update_live_view) # Render the loaded image
                    self._update_levels_histogram() # Update histogram for new image
                    
                
            def update_font(self):
                self.save_state()
                """Update the font settings based on UI inputs"""
                # Update font family from the combo box
                self.font_family = self.font_combo_box.currentFont().family()
                
                # Update font size from the spin box
                self.font_size = self.font_size_spinner.value()
                
                self.font_rotation = int(self.font_rotation_input.value())
                
            
                # Once font settings are updated, update the live view immediately
                self.update_live_view()
            
            def select_font_color(self):
                self.save_state()
                color = QColorDialog.getColor()
                if color.isValid():
                    self.font_color = color  # Store the selected color   
                    self.update_font()
                self._update_color_button_style(self.font_color_button, self.font_color)

            def dragEnterEvent(self, event):
                """Handle file dragging into the window."""
                if event.mimeData().hasUrls():
                    urls = event.mimeData().urls()
                    if urls and urls[0].isLocalFile():
                        file_path = urls[0].toLocalFile().lower()
                        # Check if the file has a valid image extension
                        if file_path.endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                            event.acceptProposedAction()
                            return
                event.ignore()

            def dropEvent(self, event):
                """Handle dropping the file."""
                urls = event.mimeData().urls()
                if urls and urls[0].isLocalFile():
                    file_path = urls[0].toLocalFile()
                    
                    # Check for unsaved changes before loading the new file
                    if not self.prompt_save_if_needed():
                        return 
                        
                    self.open_image_from_path(file_path)
                    
            def load_image(self):
                """Opens file dialog to select an image."""
                # Check for unsaved changes
                if not self.prompt_save_if_needed():
                    return

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getOpenFileName(
                    self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
                )
                
                if file_path:
                    self.open_image_from_path(file_path)

            def open_image_from_path(self, file_path):
                """Loads the image from a specific path (Used by Load Action and Drag & Drop)."""
                self.reset_image() # Clear previous state

                self.image_path = file_path
                loaded_image = QImage(self.image_path)                    

                if loaded_image.isNull():
                    # Try loading with Pillow as fallback
                    try:
                        pil_image = Image.open(self.image_path)
                        # Use our numpy converter which correctly handles 16-bit arrays
                        np_array = np.array(pil_image)
                        loaded_image = self.numpy_to_qimage(np_array)
                        if loaded_image.isNull():
                            raise ValueError("Pillow/NumPy could not convert to QImage.")
                    except Exception as e_pil:
                        print(f"Pillow/NumPy load failed: {e_pil}. Falling back to Qt loader.")
                        # Fallback to Qt's loader if Pillow fails
                        loaded_image = QImage(self.image_path)
                        if loaded_image.isNull():
                            QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}' with both Pillow and Qt.")
                            self.image_path = None
                            return

                # --- Keep the loaded image format ---
                self.image = loaded_image
                self._update_overlay_slider_ranges()

                # --- Initialize backups with the loaded format ---
                if not self.image.isNull():
                    # --- ADDED: Auto-invert 16-bit grayscale ---
                    # 16-bit gel images often have inverted intensities (0=white, 65535=black data).
                    if self.image.format() == QImage.Format_Grayscale16:
                        self.main_image_is_inverted = True
                    else:
                        self.main_image_is_inverted = False
                    # -------------------------------------------

                    self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                    self.image_master = self.image.copy()   # Master copy for resets
                    self.image_before_padding = None        # Reset padding state
                    self.image_contrasted = self.image.copy() # Backup for contrast
                    self.image_before_contrast = self.image.copy()
                    self.image_padded = False               # Reset flag

                    self.setWindowTitle(f"{self.window_title}::{self.image_path}")

                    # --- Load Associated Config File ---
                    self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                    config_name = ""
                    if self.base_name.endswith("_original"):
                        config_name = self.base_name.replace("_original", "_config.txt")
                    else:
                        config_name = self.base_name + "_config.txt"

                    config_path = os.path.join(os.path.dirname(file_path), config_name)

                    if os.path.exists(config_path):
                        try:
                            with open(config_path, "r") as config_file:
                                config_data = json.load(config_file)
                            # Apply loaded settings (this will overwrite main_image_is_inverted if defined in config)
                            self.apply_config(config_data, load_analysis=False) 
                        except Exception as e:
                            QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                    # --- End Config File Loading ---
                    
                    self.is_modified = True # Mark as modified when loading new image
                    
                    # --- ADDED: Apply adjustments ---
                    # Ensure the display reflects the inversion (or config settings) immediately
                    self.apply_all_adjustments()
                    # ------------------------------

                else:
                        QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                        return

                # --- Update UI Elements (Label size, sliders) ---
                if self.image and not self.image.isNull():
                    try:
                        render_scale = 3
                        render_width = self.live_view_label.width() * render_scale
                        render_height = self.live_view_label.height() * render_scale
                        self._update_marker_slider_ranges()

                        if not os.path.exists(config_path):
                            self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                            self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                            self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                            self.bottom_padding_input.setText("0")

                    except Exception as e:
                        pass
                # --- End UI Element Update ---
                
                self._update_status_bar()
                enable_pan = self.live_view_label.zoom_level > 1.0
                if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
                if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
                QTimer.singleShot(0, self.update_live_view)
                self._update_levels_histogram() 
                self.save_state()
            
            def apply_config(self, config_data,load_analysis=False):
                # --- 1. Load data from config_data into self attributes ---
                if "peak_dialog_settings" in config_data:
                    self.peak_dialog_settings.update(config_data["peak_dialog_settings"])

                # Padding text inputs
                adding_white_space_data = config_data.get("adding_white_space", {})
                self.left_padding_input.setText(str(adding_white_space_data.get("left", "0")))
                self.right_padding_input.setText(str(adding_white_space_data.get("right", "0")))
                self.top_padding_input.setText(str(adding_white_space_data.get("top", "0")))
                self.bottom_padding_input.setText(str(adding_white_space_data.get("bottom", "0")))
                try:
                    self.transparency = int(adding_white_space_data.get("transparency", 1))
                except (ValueError, TypeError):
                    self.transparency = 1

                # Standard marker positions
                marker_positions_data = config_data.get("marker_positions", {})
                self.left_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("left", [])]
                self.right_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("right", [])]
                self.top_markers = [(float(pos), str(label)) for pos, label in marker_positions_data.get("top", [])]

                # Top labels (internal list)
                marker_labels_data = config_data.get("marker_labels", {})
                self.top_label = [str(label).strip() for label in marker_labels_data.get("top", []) if str(label).strip()]

                # L/R marker values (internal list) - Determine if they are "custom" from this image config
                custom_lr_marker_values_from_config = []
                config_suggests_custom_lr_for_combobox = False
                lr_source_key = None
                # Check "left" first, then "right" as the source for L/R values in an image-specific config
                if "left" in marker_labels_data and isinstance(marker_labels_data["left"], list):
                    lr_source_key = "left"
                elif "right" in marker_labels_data and isinstance(marker_labels_data["right"], list):
                    lr_source_key = "right"
                
                if lr_source_key:
                    raw_values = marker_labels_data[lr_source_key]
                    for v_str in raw_values:
                        s_val = str(v_str).strip()
                        if not s_val: continue # Skip empty strings
                        try: custom_lr_marker_values_from_config.append(int(s_val))
                        except ValueError:
                            try: custom_lr_marker_values_from_config.append(float(s_val))
                            except ValueError: custom_lr_marker_values_from_config.append(s_val) # Keep as string if not number
                    if custom_lr_marker_values_from_config:
                         config_suggests_custom_lr_for_combobox = True
                
                if config_suggests_custom_lr_for_combobox:
                    self.marker_values = custom_lr_marker_values_from_config # Set internal list
                else:
                    # If no L/R values from this specific image config,
                    # self.marker_values will be populated by on_combobox_changed based on the selected preset.
                    self.marker_values = []

                if load_analysis:
                    serialized_defs = config_data.get("multi_lane_definitions", [])
                    self.multi_lane_definitions = []
                    for d in serialized_defs:
                        restored_def = {'type': d['type'], 'id': d['id']}
                        if d['type'] == 'quad':
                            restored_def['points_label'] = [QPointF(x, y) for x, y in d['points_label']]
                        elif d['type'] == 'rectangle':
                            x, y, w, h = d['points_label'][0]
                            restored_def['points_label'] = [QRectF(x, y, w, h)]
                        self.multi_lane_definitions.append(restored_def)

                    serialized_quad = config_data.get("single_quad_points", [])
                    self.live_view_label.quad_points = [QPointF(x, y) for x, y in serialized_quad]

                    self.live_view_label.bounding_box_preview = config_data.get("single_bounding_box", None)

                adj_settings = config_data.get("image_adjustments", {})
                if adj_settings:
                    # Load the values into the application's state variables
                    self.main_image_is_inverted = adj_settings.get("is_inverted", False)
                    self.channel_mixer_data = adj_settings.get("channel_mixer", self._get_default_adjustments()['channel_mixer'])
                    self.unsharp_mask_data = adj_settings.get("unsharp_mask", self._get_default_adjustments()['unsharp_mask'])
                    self.clahe_data = adj_settings.get("clahe", self._get_default_adjustments()['clahe'])
                    
                    # Update the UI sliders to match the loaded values
                    self._load_adjustments_to_ui("Main Image") # This helper will sync all sliders
                    
                    # Explicitly set the levels/gamma sliders from the loaded config
                    lg_settings = adj_settings.get('levels_gamma', {})
                    self.black_point_slider.setValue(lg_settings.get('black_point', 0))
                    self.white_point_slider.setValue(lg_settings.get('white_point', 65535))
                    self.gamma_slider.setValue(lg_settings.get('gamma', 100))

                self.apply_all_adjustments()

                # Font options for standard markers
                font_options_data = config_data.get("font_options", {})
                self.font_family = font_options_data.get("font_family", "Arial")
                self.font_size = int(font_options_data.get("font_size", 12))
                self.font_rotation = int(font_options_data.get("font_rotation", -45))
                font_color_str = font_options_data.get("font_color", "#000000")
                self.font_color = QColor(font_color_str)
                if not self.font_color.isValid(): self.font_color = QColor(0,0,0) # Fallback


                # --- Marker Shifts and Slider Value Logic ---
                marker_padding_data = config_data.get("marker_padding", {})
                added_shift_data = config_data.get("added_shift", {})

                # 1. Determine the values from "marker_padding" (these will directly set the sliders)
                try: padding_val_left = int(marker_padding_data.get("left", 0))
                except (ValueError, TypeError): padding_val_left = 0
                try: padding_val_right = int(marker_padding_data.get("right", 0))
                except (ValueError, TypeError): padding_val_right = 0
                try: padding_val_top = int(marker_padding_data.get("top", 0))
                except (ValueError, TypeError): padding_val_top = 0

                # 2. Load "added_shift" values from config. These are the true internal state.
                #    If "added_shift" is missing, fall back to "marker_padding" values for the internal state.
                self.left_marker_shift_added = int(added_shift_data.get("left", padding_val_left))
                self.right_marker_shift_added = int(added_shift_data.get("right", padding_val_right))
                self.top_marker_shift_added = int(added_shift_data.get("top", padding_val_top))

                # 3. Load and set slider RANGES.
                slider_ranges_data = config_data.get("slider_ranges", {})
                lr_range_conf = slider_ranges_data.get("left", [])
                rr_range_conf = slider_ranges_data.get("right", [])
                tr_range_conf = slider_ranges_data.get("top", [])
                default_min_range, default_max_range_w, default_max_range_h = -1000, 2000, 2000 # Adjusted defaults

                if hasattr(self, 'left_padding_slider'):
                    try:
                        lr_min, lr_max = (int(lr_range_conf[0]), int(lr_range_conf[1])) if len(lr_range_conf) == 2 else (default_min_range, default_max_range_w)
                        self.left_padding_slider.setRange(lr_min, lr_max); self.left_slider_range = [lr_min, lr_max]
                    except (IndexError, TypeError, ValueError): self.left_padding_slider.setRange(default_min_range, default_max_range_w)
                if hasattr(self, 'right_padding_slider'):
                    try:
                        rr_min, rr_max = (int(rr_range_conf[0]), int(rr_range_conf[1])) if len(rr_range_conf) == 2 else (default_min_range, default_max_range_w)
                        self.right_padding_slider.setRange(rr_min, rr_max); self.right_slider_range = [rr_min, rr_max]
                    except (IndexError, TypeError, ValueError): self.right_padding_slider.setRange(default_min_range, default_max_range_w)
                if hasattr(self, 'top_padding_slider'):
                    try:
                        tr_min, tr_max = (int(tr_range_conf[0]), int(tr_range_conf[1])) if len(tr_range_conf) == 2 else (default_min_range, default_max_range_h)
                        self.top_padding_slider.setRange(tr_min, tr_max); self.top_slider_range = [tr_min, tr_max]
                    except (IndexError, TypeError, ValueError): self.top_padding_slider.setRange(default_min_range, default_max_range_h)

                # 4. Set slider VALUES using self.xxx_marker_shift_added (which now holds the true state)
                #    Signals are blocked. Qt will clamp these if they are outside the range just set.
                if hasattr(self, 'left_padding_slider'):
                    self.left_padding_slider.blockSignals(True); self.left_padding_slider.setValue(self.left_marker_shift_added); self.left_padding_slider.blockSignals(False)
                if hasattr(self, 'right_padding_slider'):
                    self.right_padding_slider.blockSignals(True); self.right_padding_slider.setValue(self.right_marker_shift_added); self.right_padding_slider.blockSignals(False)
                if hasattr(self, 'top_padding_slider'):
                    self.top_padding_slider.blockSignals(True); self.top_padding_slider.setValue(self.top_marker_shift_added); self.top_padding_slider.blockSignals(False)

                # 5. CRITICAL RE-SYNC: Ensure internal `_marker_shift_added` variables
                #    MATCH the SLIDER'S ACTUAL VALUE after potential clamping by setValue.
                if hasattr(self, 'left_padding_slider'): self.left_marker_shift_added = self.left_padding_slider.value()
                if hasattr(self, 'right_padding_slider'): self.right_marker_shift_added = self.right_padding_slider.value()
                if hasattr(self, 'top_padding_slider'): self.top_marker_shift_added = self.top_padding_slider.value()
                # --- End Marker Shifts and Slider Logic ---


                # --- ComboBox and UI Update for L/R values and Top Labels ---
                if hasattr(self, 'combo_box'):
                    self.combo_box.blockSignals(True)
                    if config_suggests_custom_lr_for_combobox: # Implies self.marker_values is set from image config
                        custom_idx = self.combo_box.findText("Custom")
                        if custom_idx != -1: self.combo_box.setCurrentIndex(custom_idx)
                    else:
                        # If no L/R in image config, try to load saved preset name from config_data itself.
                        # This 'selected_preset_name' would have been saved if the image was last saved with a specific preset selected.
                        selected_preset_name_from_config = config_data.get("selected_preset_name", None)
                        if selected_preset_name_from_config and self.combo_box.findText(selected_preset_name_from_config) != -1:
                            self.combo_box.setCurrentText(selected_preset_name_from_config)
                        else: # Fallback to a default preset if nothing specific is found.
                            default_biorad = "Precision Plus Protein All Blue Prestained (Bio-Rad)"
                            idx_biorad = self.combo_box.findText(default_biorad)
                            if idx_biorad != -1: self.combo_box.setCurrentIndex(idx_biorad)
                            elif self.combo_box.count() > 0: # If BioRad not found, select first available or "Custom"
                                custom_idx_fallback = self.combo_box.findText("Custom")
                                if custom_idx_fallback == 0 and self.combo_box.count() > 1: self.combo_box.setCurrentIndex(1)
                                elif custom_idx_fallback != -1: self.combo_box.setCurrentIndex(custom_idx_fallback)
                                else: self.combo_box.setCurrentIndex(0)
                    self.combo_box.blockSignals(False)
                    self.on_combobox_changed() # Updates text boxes based on current combobox selection and internal lists

                # Load custom markers and shapes
                custom_markers_data_from_image_config = config_data.get("custom_markers_config", config_data.get("custom_markers", [])) # Check both keys
                self.custom_markers = [list(m) for m in self._deserialize_custom_markers(custom_markers_data_from_image_config)]
                
                self.custom_shapes = []
                loaded_shapes_from_config = config_data.get("custom_shapes_config", config_data.get("custom_shapes", [])) # Check both keys
                if isinstance(loaded_shapes_from_config, list):
                    for shape_item in loaded_shapes_from_config:
                        if isinstance(shape_item, dict):
                            self.custom_shapes.append(dict(shape_item))

                # Update standard font UI elements
                if hasattr(self, 'font_combo_box'):
                    self.font_combo_box.blockSignals(True); self.font_combo_box.setCurrentFont(QFont(self.font_family)); self.font_combo_box.blockSignals(False)
                if hasattr(self, 'font_size_spinner'):
                    self.font_size_spinner.blockSignals(True); self.font_size_spinner.setValue(self.font_size); self.font_size_spinner.blockSignals(False)
                if hasattr(self, 'font_rotation_input'):
                    self.font_rotation_input.blockSignals(True); self.font_rotation_input.setValue(self.font_rotation); self.font_rotation_input.blockSignals(False)
                if hasattr(self, 'font_color_button'):
                    self._update_color_button_style(self.font_color_button, self.font_color)
                
                # Apply UI state for custom marker creation tools (if saved in config, e.g. from a full preset)
                ui_custom_settings = config_data.get("ui_custom_marker_settings", {})
                
                custom_marker_color_str = ui_custom_settings.get("color", self.custom_marker_color.name() if hasattr(self, 'custom_marker_color') else "#000000")
                temp_custom_color = QColor(custom_marker_color_str)
                self.custom_marker_color = temp_custom_color if temp_custom_color.isValid() else QColor(0,0,0)

                if hasattr(self, 'custom_font_type_dropdown'):
                    default_custom_family = self.custom_font_type_dropdown.currentFont().family() if self.custom_font_type_dropdown.count() > 0 else "Arial"
                    custom_font_family_str = ui_custom_settings.get("font_family", default_custom_family)
                    self.custom_font_type_dropdown.blockSignals(True)
                    self.custom_font_type_dropdown.setCurrentFont(QFont(custom_font_family_str))
                    self.custom_font_type_dropdown.blockSignals(False)
                    self.update_marker_text_font(self.custom_font_type_dropdown.currentFont())

                if hasattr(self, 'custom_font_size_spinbox'):
                    default_custom_size = self.custom_font_size_spinbox.value() if hasattr(self, 'custom_font_size_spinbox') else 12
                    custom_font_size_val = int(ui_custom_settings.get("font_size", default_custom_size))
                    self.custom_font_size_spinbox.blockSignals(True)
                    self.custom_font_size_spinbox.setValue(custom_font_size_val)
                    self.custom_font_size_spinbox.blockSignals(False)

                if hasattr(self, 'custom_marker_color_button'):
                    self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
                
            def get_current_config(self):
                """Gathers the current application state into a dictionary for saving."""
                def make_json_serializable(value):
                    if isinstance(value, (np.float16, np.float32, np.float64, np.floating)):
                        return float(value)
                    if isinstance(value, (np.int8, np.int16, np.int32, np.int64, np.integer)):
                        return int(value)
                    if isinstance(value, np.bool_):
                        return bool(value)
                    if isinstance(value, list):
                        return [make_json_serializable(v) for v in value]
                    if isinstance(value, tuple):
                        if len(value) == 2 and isinstance(value[1], str):
                            return [make_json_serializable(value[0]), str(value[1])]
                        return [make_json_serializable(v) for v in value]
                    if isinstance(value, dict):
                        return {make_json_serializable(k): make_json_serializable(v) for k, v in value.items()}
                    return value
                
                config = {
                    "adding_white_space": {
                        "left": self.left_padding_input.text(),
                        "right": self.right_padding_input.text(),
                        "top": self.top_padding_input.text(),
                        "bottom": self.bottom_padding_input.text(),
                        "transparency": self.transparency,
                    },
                    "marker_positions": {
                        "left": [(pos, label) for pos, label in getattr(self, 'left_markers', [])],
                        "right": [(pos, label) for pos, label in getattr(self, 'right_markers', [])],
                        "top": [(pos, label) for pos, label in getattr(self, 'top_markers', [])],
                    },
                    "marker_labels": {
                        "top": getattr(self, 'top_label', []),
                        "left": [marker[1] for marker in getattr(self, 'left_markers', [])],
                        "right": [marker[1] for marker in getattr(self, 'right_markers', [])],
                    },
                    "marker_padding": {
                        "top": self.top_padding_slider.value() if hasattr(self, 'top_padding_slider') else 0,
                        "left": self.left_padding_slider.value() if hasattr(self, 'left_padding_slider') else 0,
                        "right": self.right_padding_slider.value() if hasattr(self, 'right_padding_slider') else 0,
                    },
                    "font_options": {
                        "font_family": self.font_family,
                        "font_size": self.font_size,
                        "font_rotation": self.font_rotation,
                        "font_color": self.font_color.name(),
                    },
                    "slider_ranges": {
                         "left": getattr(self, 'left_slider_range', [-100, 1000]),
                         "right": getattr(self, 'right_slider_range', [-100, 1000]),
                         "top": getattr(self, 'top_slider_range', [-100, 1000]),
                     },
                    "added_shift": {
                         "left": getattr(self, 'left_marker_shift_added', 0),
                         "right": getattr(self, 'right_marker_shift_added', 0),
                         "top": getattr(self, 'top_marker_shift_added', 0),
                    },
                    "image_adjustments": {
                        "is_inverted": self.main_image_is_inverted,
                        "levels_gamma": {
                            "black_point": self.black_point_slider.value(),
                            "white_point": self.white_point_slider.value(),
                            "gamma": self.gamma_slider.value()
                        },
                        "channel_mixer": self.channel_mixer_data.copy(),
                        "unsharp_mask": self.unsharp_mask_data.copy(),
                        "clahe": self.clahe_data.copy()
                    }
                }

                # --- START OF THE FIX ---
                # Explicitly add all analysis region definitions (single and multi-lane)
                
                # 1. Multi-lane definitions (already in the save_state function, let's add it here too for consistency)
                config["multi_lane_definitions"] = [
                    {
                        'type': d['type'], 'id': d['id'],
                        'points_label': (
                            [(p.x(), p.y()) for p in d['points_label']] if d['type'] == 'quad' else
                            [(d['points_label'][0].x(), d['points_label'][0].y(), d['points_label'][0].width(), d['points_label'][0].height())]
                        )
                    } for d in self.multi_lane_definitions
                ]

                # 2. Single quadrilateral points
                config["single_quad_points"] = [(p.x(), p.y()) for p in self.live_view_label.quad_points]

                # 3. Single rectangle points
                config["single_bounding_box"] = self.live_view_label.bounding_box_preview
                # --- END OF BOUNDARY BOX FIX ---

                custom_markers_data = []
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        x, y, text, color, font, font_size, is_bold, is_italic = marker_tuple
                        custom_markers_data.append({
                            "x": x, "y": y, "text": text, "color": color.name(),
                            "font": font, "font_size": font_size, "bold": is_bold, "italic": is_italic
                        })
                    except (ValueError, TypeError, IndexError): pass
                config["custom_markers"] = custom_markers_data
                config["custom_shapes"] = [dict(s) for s in getattr(self, "custom_shapes", [])]

                config["quantities_peak_area_dict"] = self.quantities_peak_area_dict

                # --- START OF THE PROTEIN SEQUENCE FIX ---
                config["protein_analysis_data"] = {
                    "protein_sequence": self.protein_sequence,
                    "base_protein_mw": self.base_protein_mw,
                    "avg_glycan_mass": self.avg_glycan_mass,
                    "num_oligomers_to_model": self.num_oligomers_to_model,
                    "num_glycans_to_model": self.num_glycans_to_model,
                    "last_predicted_mw": self.last_predicted_mw,
                    "last_mw_prediction_model": self.last_mw_prediction_model,
                }

                config["peak_dialog_settings"] = self.peak_dialog_settings
                # --- END OF PROTEIN SEQUENCE FIX ---
                
                return config
            
            def add_band(self, event):
                self.update_all_labels() 
                self.save_state()
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                
                if not self.image or self.image.isNull() or not self.marker_mode:
                    return

                # --- 1. Get Click Position in Unzoomed Label Space (handles zoom/pan) ---
                click_pos_unzoomed_label_space = self.live_view_label.transform_point(event.position())
                
                # --- 2. Apply Grid Snapping in Unzoomed Label Space ---
                snapped_click_pos_label_space = self.snap_point_to_grid(click_pos_unzoomed_label_space)
                cursor_x_ls = snapped_click_pos_label_space.x() 
                cursor_y_ls = snapped_click_pos_label_space.y()

                # --- 3. Transform Snapped Label Space Coords to Native Image Coords ---
                # (This logic is unchanged)
                current_app_image = self.image
                label_w_widget = float(self.live_view_label.width())
                label_h_widget = float(self.live_view_label.height())
                img_w_native = float(current_app_image.width())
                img_h_native = float(current_app_image.height())
                if img_w_native <= 0 or img_h_native <= 0 or label_w_widget <= 0 or label_h_widget <= 0: return
                scale_native_to_label = min(label_w_widget / img_w_native, label_h_widget / img_h_native)
                if scale_native_to_label <= 1e-9: return
                displayed_img_w_in_label = img_w_native * scale_native_to_label
                displayed_img_h_in_label = img_h_native * scale_native_to_label
                offset_x_img_in_label = (label_w_widget - displayed_img_w_in_label) / 2.0
                offset_y_img_in_label = (label_h_widget - displayed_img_h_in_label) / 2.0
                image_x = (cursor_x_ls - offset_x_img_in_label) / scale_native_to_label
                image_y = (cursor_y_ls - offset_y_img_in_label) / scale_native_to_label
                image_x = max(0.0, min(image_x, img_w_native))
                image_y = max(0.0, min(image_y, img_h_native))
                
                try:
                    target_list = None
                    label_source = None
                    new_position = 0.0
                    is_first_marker = False

                    if self.marker_mode == "left":
                        target_list = self.left_markers
                        label_source = self.marker_values
                        new_position = image_y
                        is_first_marker = (len(self.left_markers) == 0)
                    elif self.marker_mode == "right":
                        target_list = self.right_markers
                        label_source = self.marker_values
                        new_position = image_y
                        is_first_marker = (len(self.right_markers) == 0)
                    elif self.marker_mode == "top":
                        target_list = self.top_markers
                        label_source = self.top_label
                        new_position = image_x
                        is_first_marker = (len(self.top_markers) == 0)
                    
                    if target_list is not None:
                        # 1. Add the new marker with a temporary blank label
                        target_list.append((new_position, ""))

                        # 2. Sort the entire list by position (the first element of the tuple)
                        target_list.sort(key=lambda m: m[0])

                        # 3. Re-label the entire sorted list from scratch
                        for i in range(len(target_list)):
                            pos = target_list[i][0]
                            new_label = str(label_source[i]) if i < len(label_source) else ""
                            target_list[i] = (pos, new_label)

                        # 4. Handle the special case of setting the offset slider on the very first marker
                        if is_first_marker:
                            slider_to_update = None
                            range_to_use = []
                            shift_attr_to_set = ""
                            pos_to_use_for_slider = 0

                            if self.marker_mode == "left":
                                slider_to_update = self.left_padding_slider
                                range_to_use = self.left_slider_range
                                shift_attr_to_set = "left_marker_shift_added"
                                pos_to_use_for_slider = image_x
                            elif self.marker_mode == "right":
                                slider_to_update = self.right_padding_slider
                                range_to_use = self.right_slider_range
                                shift_attr_to_set = "right_marker_shift_added"
                                pos_to_use_for_slider = image_x
                            elif self.marker_mode == "top":
                                slider_to_update = self.top_padding_slider
                                range_to_use = self.top_slider_range
                                shift_attr_to_set = "top_marker_shift_added"
                                pos_to_use_for_slider = image_y

                            if slider_to_update:
                                slider_target_value = int(round(pos_to_use_for_slider))
                                self._update_marker_slider_ranges() # Ensure range is up-to-date
                                slider_to_update.blockSignals(True)
                                slider_to_update.setValue(max(range_to_use[0], min(slider_target_value, range_to_use[1])))
                                slider_to_update.blockSignals(False)
                                setattr(self, shift_attr_to_set, slider_to_update.value())

                        # 5. Update the preview text for the *next* click
                        next_index = len(target_list)
                        next_label_for_preview = str(label_source[next_index]) if next_index < len(label_source) else ""
                        self.live_view_label.standard_marker_preview_text = next_label_for_preview

                except Exception as e:
                     import traceback
                     traceback.print_exc()
                     QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

                self.update_live_view()
                
            def _deactivate_all_previews(self):
                """Helper to turn off all live preview modes on the label."""
                if hasattr(self, 'live_view_label'):
                    self.live_view_label.preview_marker_enabled = False
                    self.live_view_label.mw_predict_preview_enabled = False
                    self.live_view_label.standard_marker_preview_enabled = False
                    self.live_view_label.standard_marker_preview_position = None
                    self.live_view_label.preview_marker_position = None
                    self.live_view_label.mw_predict_preview_position = None
                
            def enable_left_marker_mode(self):
                self._deactivate_all_previews()
                self.marker_mode = "left"
                self.current_left_marker_index = len(self.left_markers) # Start from the next available
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)

                # --- NEW: Enable preview ---
                self.live_view_label.standard_marker_preview_enabled = True
                next_index = len(self.left_markers)
                next_label = str(self.marker_values[next_index]) if next_index < len(self.marker_values) else ""
                self.live_view_label.standard_marker_preview_text = next_label
                self.live_view_label.standard_marker_preview_mode = "left"

            def enable_right_marker_mode(self):
                self._deactivate_all_previews()
                self.marker_mode = "right"
                self.current_right_marker_index = len(self.right_markers)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)

                # --- NEW: Enable preview ---
                self.live_view_label.standard_marker_preview_enabled = True
                next_index = len(self.right_markers)
                next_label = str(self.marker_values[next_index]) if next_index < len(self.marker_values) else ""
                self.live_view_label.standard_marker_preview_text = next_label
                self.live_view_label.standard_marker_preview_mode = "right"
            
            def enable_top_marker_mode(self):
                self._deactivate_all_previews()
                self.marker_mode = "top"
                self.current_top_label_index = len(self.top_markers)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = self.add_band
                self.live_view_label.setCursor(Qt.CrossCursor)

                # --- NEW: Enable preview ---
                self.live_view_label.standard_marker_preview_enabled = True
                next_index = len(self.top_markers)
                next_label = str(self.top_label[next_index]) if next_index < len(self.top_label) else ""
                self.live_view_label.standard_marker_preview_text = next_label
                self.live_view_label.standard_marker_preview_mode = "top"
                
                
            def finalize_image(self, *args, **kwargs): # Padding
                if not self.image or self.image.isNull():
                    QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
                    return
                try:
                    padding_left = max(0, int(self.left_padding_input.text()))
                    padding_right = max(0, int(self.right_padding_input.text()))
                    padding_top = max(0, int(self.top_padding_input.text()))
                    padding_bottom = max(0, int(self.bottom_padding_input.text()))
                    # Normalize text fields
                    self.left_padding_input.setText(str(padding_left))
                    self.right_padding_input.setText(str(padding_right))
                    self.top_padding_input.setText(str(padding_top))
                    self.bottom_padding_input.setText(str(padding_bottom))
                except ValueError:
                    QMessageBox.warning(self, "Error", "Please enter valid non-negative integers for padding.")
                    return

                if padding_left == 0 and padding_right == 0 and padding_top == 0 and padding_bottom == 0:
                    QMessageBox.information(self, "Info", "No padding specified. Image remains unchanged.")
                    return

                self.save_state()

                try:
                    # 1. Update internal marker coordinates first
                    self.adjust_elements_for_padding(padding_left, padding_top)

                    # 2. Perform Padding on Image Data
                    np_img = self.qimage_to_numpy(self.image_master)
                    if np_img is None:
                        raise ValueError("Failed to convert source image to NumPy array for padding.")

                    original_height, original_width = np_img.shape[:2]
                    new_width = original_width + padding_left + padding_right
                    new_height = original_height + padding_top + padding_bottom
                    target_dtype = np_img.dtype

                    padded_shape = (new_height, new_width, 4) if np_img.ndim == 3 else (new_height, new_width)
                    padded_np = np.zeros(padded_shape, dtype=target_dtype)

                    target_slice = padded_np[padding_top:padding_top + original_height, padding_left:padding_left + original_width]

                    if np_img.ndim == 2:
                        target_slice[:, :] = np_img
                    elif np_img.ndim == 3:
                        target_slice[:, :, :] = np_img
                    
                    padded_image = self.numpy_to_qimage(padded_np)
                    if padded_image.isNull():
                        raise ValueError("Conversion back to QImage failed after padding.")

                    # 3. Update Master Image State
                    self.image_master = padded_image.copy()
                    self.image_before_contrast = self.image_master.copy()
                    self.image_contrasted = self.image_master.copy()
                    self.image_before_padding = None
                    self.image_padded = True
                    self.is_modified = True
                    
                    # 4. CRITICAL FIX: Refresh self.image (display image) so it has NEW DIMENSIONS
                    # This must happen BEFORE updating slider ranges.
                    self.apply_all_adjustments()

                    # 5. Now update ranges (uses new self.image dimensions) and sync slider values
                    self._update_status_bar()
                    self._update_marker_slider_ranges()
                    self._update_overlay_slider_ranges()

                    # Force sliders to the new adjusted values (which now fit in the new ranges)
                    for slider, value_to_set in [
                        (self.left_padding_slider, self.left_marker_shift_added), 
                        (self.right_padding_slider, self.right_marker_shift_added), 
                        (self.top_padding_slider, self.top_marker_shift_added)
                    ]:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(value_to_set)
                            slider.blockSignals(False)

                except Exception as e:
                    QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
                    traceback.print_exc()
                    if self.undo_stack:
                        try: self.undo_action_m()
                        except: pass
                    

            def adjust_elements_for_padding(self, padding_left, padding_top):
                """
                Adjusts coordinates of all markers and custom shapes for added padding.
                Also adjusts the _marker_shift_added values.
                This function is called BEFORE the image is actually padded.
                """

                # Adjust standard markers (Y for L/R, X for Top)
                self.left_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'left_markers', [])]
                self.right_markers = [(y_pos_img + padding_top, label) for y_pos_img, label in getattr(self, 'right_markers', [])]
                self.top_markers = [(x_pos_img + padding_left, label) for x_pos_img, label in getattr(self, 'top_markers', [])]

                # Adjust custom markers (both X and Y)
                new_custom_markers = []
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        m_list = list(marker_data) # Ensure mutable
                        m_list[0] = float(m_list[0]) + padding_left  # Adjust X
                        m_list[1] = float(m_list[1]) + padding_top   # Adjust Y
                        new_custom_markers.append(m_list)
                    except (IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom marker during padding adjustment: {marker_data}")
                         new_custom_markers.append(marker_data) # Keep original if error
                self.custom_markers = new_custom_markers

                # Adjust Custom Shapes
                new_custom_shapes = []
                for shape_data_orig in getattr(self, "custom_shapes", []):
                    shape_data = dict(shape_data_orig) # Work on a copy
                    try:
                        shape_type = shape_data.get('type')
                        if shape_type == 'line':
                            sx, sy = shape_data['start']
                            ex, ey = shape_data['end']
                            shape_data['start'] = (float(sx) + padding_left, float(sy) + padding_top)
                            shape_data['end'] = (float(ex) + padding_left, float(ey) + padding_top)
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            shape_data['rect'] = (float(x) + padding_left, float(y) + padding_top, w, h)
                        # Add other shape types here if needed
                        new_custom_shapes.append(shape_data)
                    except (KeyError, IndexError, TypeError, ValueError):
                         print(f"Warning: Skipping malformed custom shape during padding adjustment: {shape_data_orig}")
                         new_custom_shapes.append(shape_data_orig)
                self.custom_shapes = new_custom_shapes

                # Adjust the absolute marker shift variables (these are in native image pixels)
                self.left_marker_shift_added += padding_left
                self.right_marker_shift_added += padding_left
                self.top_marker_shift_added += padding_top      
                self._update_overlay_slider_ranges
                
            
            def update_left_padding(self):
                # Update left padding when slider value changes
                self.left_marker_shift_added = self.left_padding_slider.value()
                self.update_live_view()

            def update_right_padding(self):
                # Update right padding when slider value changes
                new_value = self.right_padding_slider.value()
                # --- Add check to prevent redundant updates ---
                if new_value != self.right_marker_shift_added:
                    self.right_marker_shift_added = new_value
                    self.update_live_view()
                
            def update_top_padding(self):
                # Update top padding when slider value changes
                self.top_marker_shift_added = self.top_padding_slider.value()
                self.update_live_view()

            def update_live_view(self):
                # --- START FIX: Context-aware rendering ---
                image_for_view = None
                
                if self.is_in_dedicated_edit_mode:
                    # In an isolated editor view, get the specific pre-calculated overlay preview
                    if self.adjustment_context == "Overlay 1 (Base)":
                        image_for_view = getattr(self, 'image1_adjusted_preview', None)
                    elif self.adjustment_context == "Overlay 2 (Overlay)":
                        image_for_view = getattr(self, 'image2_adjusted_preview', None)
                    else: # Should not happen
                        image_for_view = self.image

                    # If the overlay isn't loaded or cache is empty, show a blank placeholder
                    if not image_for_view or image_for_view.isNull():
                         image_for_view = QImage(600, 400, QImage.Format_RGB32)
                         image_for_view.fill(Qt.darkGray)

                else:
                    # In the main view mode, use the main self.image (which is the 8-bit display cache for the base image)
                    image_for_view = self.image
                # --- END FIX ---

                if hasattr(self, 'live_view_label') and hasattr(self, 'pan_left_action'):
                    enable_pan_actions = self.live_view_label.zoom_level > 1.0
                    self.pan_left_action.setEnabled(enable_pan_actions)
                    self.pan_right_action.setEnabled(enable_pan_actions)
                    self.pan_up_action.setEnabled(enable_pan_actions)
                    self.pan_down_action.setEnabled(enable_pan_actions)
                
                if not image_for_view or image_for_view.isNull(): 
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear(); self.live_view_label.update()
                    # ... (rest of the null-image handling)
                    if hasattr(self, 'predict_button'): self.predict_button.setEnabled(False)
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(False)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(False)
                    return
                else:
                    if hasattr(self, 'save_action'): self.save_action.setEnabled(True)
                    if hasattr(self, 'copy_action'): self.copy_action.setEnabled(True)
                    if hasattr(self, 'predict_button'):
                        left_m = getattr(self, 'left_markers', []); right_m = getattr(self, 'right_markers', [])
                        self.predict_button.setEnabled(bool(left_m or right_m))

                # ... (The rest of the update_live_view method from here on is UNCHANGED) ...
                # ... It will now correctly use `image_for_view` for all subsequent rendering steps ...

                current_zoom_level = 1.0
                current_pan_offset = QPointF(0, 0)
                if hasattr(self, 'live_view_label'):
                    current_zoom_level = self.live_view_label.zoom_level
                    current_pan_offset = QPointF(self.live_view_label.pan_offset)

                render_scale = 3
                try: 
                    view_width = self.live_view_label.width(); view_height = self.live_view_label.height()
                    if view_width <= 0: view_width = 600
                    if view_height <= 0: view_height = 400
                    render_width = view_width * render_scale; render_height = view_height * render_scale
                except AttributeError:
                     render_width = 1800; render_height = 1200
                except Exception:
                     render_width = 1800; render_height = 1200
                
                if not image_for_view or image_for_view.isNull():
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                
                image_to_transform = image_for_view.copy()

                orientation = 0.0
                if hasattr(self, 'orientation_slider') and self.orientation_slider:
                    orientation = float(self.orientation_slider.value() / 20)
                    if hasattr(self, 'orientation_label') and self.orientation_label:
                        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")
                    if abs(orientation) > 0.01: 
                         if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                             transform_rotate = QTransform()
                             w_rot, h_rot = image_to_transform.width(), image_to_transform.height()
                             transform_rotate.translate(w_rot / 2.0, h_rot / 2.0)
                             transform_rotate.rotate(orientation)
                             transform_rotate.translate(-w_rot / 2.0, -h_rot / 2.0)
                             temp_rotated = image_to_transform.transformed(transform_rotate, Qt.SmoothTransformation)
                             if not temp_rotated.isNull(): image_to_transform = temp_rotated
                
                taper_value = 0.0
                if hasattr(self, 'taper_skew_slider') and self.taper_skew_slider:
                    taper_value = self.taper_skew_slider.value() / 100.0
                    if hasattr(self, 'taper_skew_label') and self.taper_skew_label:
                        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")
                    if abs(taper_value) > 0.01: 
                        if not image_to_transform.isNull() and image_to_transform.width() > 0 and image_to_transform.height() > 0:
                            try:
                                np_image = self.qimage_to_numpy(image_to_transform)
                                if np_image is None: raise ValueError("Failed to convert QImage to NumPy array.")
                                height, width = np_image.shape[:2]
                                source_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                                destination_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                                if taper_value > 0:
                                    destination_np[0][0] = width * taper_value / 2.0
                                    destination_np[1][0] = width * (1 - taper_value / 2.0)
                                elif taper_value < 0:
                                    destination_np[3][0] = width * (-taper_value / 2.0)
                                    destination_np[2][0] = width * (1 + taper_value / 2.0)
                                matrix = cv2.getPerspectiveTransform(source_np, destination_np)
                                skewed_np_image = cv2.warpPerspective(np_image, matrix, (width, height))
                                temp_skewed_qimage = self.numpy_to_qimage(skewed_np_image)
                                if not temp_skewed_qimage.isNull(): image_to_transform = temp_skewed_qimage
                            except Exception as e: print(f"Error during OpenCV skew preview: {e}")

                if image_to_transform.isNull() or image_to_transform.width() <= 0 or image_to_transform.height() <= 0:
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return

                scaled_image_for_render_canvas = image_to_transform.scaled(
                    render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                if scaled_image_for_render_canvas.isNull():
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return

                canvas_format = QImage.Format_ARGB32_Premultiplied if image_to_transform.hasAlphaChannel() else QImage.Format_RGB888
                render_canvas = QImage(render_width, render_height, canvas_format)
                if render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     return
                render_canvas.fill(Qt.white if canvas_format == QImage.Format_RGB888 else Qt.transparent)

                current_crop_offset_x = getattr(self, 'crop_offset_x', 0)
                current_crop_offset_y = getattr(self, 'crop_offset_y', 0)

                self.render_image_on_canvas(render_canvas, scaled_image_for_render_canvas,
                                            x_start=current_crop_offset_x, 
                                            y_start=current_crop_offset_y, 
                                            render_scale=render_scale,
                                            draw_guides=True)

                if render_canvas.isNull():
                    if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                    return
                
                # --- THIS IS THE HIGH-RESOLUTION RENDER ---
                pixmap_from_render_canvas = QPixmap.fromImage(render_canvas)
                if pixmap_from_render_canvas.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     return

                # This is the pixmap representing the 100% zoom, unpanned view, scaled to fit the label
                scaled_pixmap_for_label_fit = pixmap_from_render_canvas.scaled(
                    self.live_view_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                if scaled_pixmap_for_label_fit.isNull():
                     if hasattr(self, 'live_view_label'): self.live_view_label.clear()
                     return

                final_pixmap_to_set_on_label = scaled_pixmap_for_label_fit

                # --- START OF THE FIX ---
                if current_zoom_level != 1.0:
                    zoomed_display_pixmap = QPixmap(self.live_view_label.size()) 
                    if zoomed_display_pixmap.isNull():
                        print("Error: Failed to create zoomed_display_pixmap.")
                    else:
                        zoomed_display_pixmap.fill(Qt.transparent) 
                        zoom_painter = QPainter(zoomed_display_pixmap)
                        if not zoom_painter.isActive():
                            print("Error: Failed to create QPainter for zooming in update_live_view.")
                        else:
                            # Enable high-quality transformations for the painter
                            zoom_painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
                            
                            zoom_painter.translate(current_pan_offset)
                            zoom_painter.scale(current_zoom_level, current_zoom_level)
                            
                            # THE FIX: Draw the HIGH-RESOLUTION render canvas into the
                            # target rectangle of the 100% view. The painter will handle
                            # scaling it correctly, providing a sharp zoom.
                            target_rect = scaled_pixmap_for_label_fit.rect()
                            zoom_painter.drawPixmap(target_rect, pixmap_from_render_canvas)
                            
                            zoom_painter.end()

                            if not zoomed_display_pixmap.isNull():
                                 final_pixmap_to_set_on_label = zoomed_display_pixmap
                # --- END OF THE FIX ---
                
                if final_pixmap_to_set_on_label.isNull():
                    self.live_view_label.clear()
                else:
                     self.live_view_label.setPixmap(final_pixmap_to_set_on_label)

                self.live_view_label.update()
                
            def _update_overlay_slider_ranges(self):
                """
                Updates the ranges of overlay position sliders.
                The ranges are based on the NATIVE dimensions of the current self.image.
                Slider values will represent native pixel offsets.
                """
                native_width = 1000  # Fallback
                native_height = 1000 # Fallback
        
                if self.image and not self.image.isNull() and self.image.width() > 0 and self.image.height() > 0:
                    native_width = self.image.width()
                    native_height = self.image.height()
        
                # Allow positioning overlay's top-left from -1*dimension to +2*dimension
                # relative to the main image's top-left in NATIVE pixels.
                x_range_min = -native_width
                x_range_max = native_width * 2
                y_range_min = -native_height
                y_range_max = native_height * 2
                
                
                sliders_to_update = [
                    (getattr(self, 'image1_left_slider', None), x_range_min, x_range_max, getattr(self, 'image1_position', (0,0))[0]),
                    (getattr(self, 'image1_top_slider', None), y_range_min, y_range_max, getattr(self, 'image1_position', (0,0))[1]),
                    (getattr(self, 'image2_left_slider', None), x_range_min, x_range_max, getattr(self, 'image2_position', (0,0))[0]),
                    (getattr(self, 'image2_top_slider', None), y_range_min, y_range_max, getattr(self, 'image2_position', (0,0))[1]),
                ]
        
                for slider, min_val, max_val, current_pos_val in sliders_to_update:
                    if slider:
                        slider.blockSignals(True)
                        slider.setRange(min_val, max_val)
                        # Try to set the slider to the current stored native position, clamped by new range
                        clamped_val = max(min_val, min(current_pos_val, max_val))
                        slider.setValue(clamped_val)
                        slider.blockSignals(False)
            
            def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
                painter = QPainter(canvas)
                painter.setRenderHint(QPainter.Antialiasing, True)
                painter.setRenderHint(QPainter.TextAntialiasing, True)

                # --- 1. Logic for Isolated Edit Mode ---
                # (Viewing just one overlay while adjusting its settings)
                if self.is_in_dedicated_edit_mode:
                    image_to_draw = None
                    if self.adjustment_context == "Overlay 1 (Base)":
                        image_to_draw = getattr(self, 'image1_adjusted_preview', None)
                    elif self.adjustment_context == "Overlay 2 (Overlay)":
                        image_to_draw = getattr(self, 'image2_adjusted_preview', None)
                    else:
                        image_to_draw = scaled_image

                    if not image_to_draw or image_to_draw.isNull():
                         painter.end(); return

                    # Draw the single image centered and scaled to fit the view
                    scaled_overlay_for_view = image_to_draw.scaled(canvas.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    x_offset = (canvas.width() - scaled_overlay_for_view.width()) // 2
                    y_offset = (canvas.height() - scaled_overlay_for_view.height()) // 2
                    self.x_offset_s = x_offset
                    self.y_offset_s = y_offset
                    painter.drawImage(x_offset, y_offset, scaled_overlay_for_view)
                
                # --- 2. Logic for Main View (Combined) ---
                else:
                    # Draw the base container image centered
                    x_offset = (canvas.width() - scaled_image.width()) // 2
                    y_offset = (canvas.height() - scaled_image.height()) // 2
                    self.x_offset_s = x_offset
                    self.y_offset_s = y_offset
                    painter.drawImage(x_offset, y_offset, scaled_image)

                    # Check which overlays are active
                    has_img1 = hasattr(self, 'image1_adjusted_preview') and self.image1_adjusted_preview and hasattr(self, 'image1_position')
                    has_img2 = hasattr(self, 'image2_adjusted_preview') and self.image2_adjusted_preview and hasattr(self, 'image2_position')
                    
                    # Get mixing value (0.0 to 1.0)
                    # 0% = Overlay invisible, 100% = Overlay opaque
                    blend_value = (self.blend_slider.value() / 100.0) if hasattr(self, 'blend_slider') else 0.5

                    # --- Draw Image 1 (Base) ---
                    # Always drawn at 100% opacity to act as the "A" in the mixing equation
                    if has_img1:
                        adjusted_img1 = getattr(self, 'image1_adjusted_preview', None)
                        if adjusted_img1:
                            painter.setOpacity(1.0) # Base is opaque
                            
                            rect_in_label_space = self._get_overlay_rect_in_label_space(1)
                            if rect_in_label_space:
                                # Scale rect to high-res canvas coordinates
                                rect_in_canvas_space = QRectF(
                                    rect_in_label_space.left() * render_scale, 
                                    rect_in_label_space.top() * render_scale, 
                                    rect_in_label_space.width() * render_scale, 
                                    rect_in_label_space.height() * render_scale
                                )
                                rotation1 = self.image1_rotation_slider.value() / 10.0 if hasattr(self, 'image1_rotation_slider') else 0.0

                                if abs(rotation1) > 0.01:
                                    center_point_canvas = rect_in_canvas_space.center()
                                    painter.save()
                                    painter.translate(center_point_canvas)
                                    painter.rotate(rotation1)
                                    painter.translate(-center_point_canvas)
                                    painter.drawImage(rect_in_canvas_space, adjusted_img1)
                                    painter.restore()
                                else:
                                    painter.drawImage(rect_in_canvas_space, adjusted_img1)

                    # --- Draw Image 2 (Overlay) ---
                    # Drawn with opacity = blend_value to act as the "B" in the mixing equation
                    # Result = A * (1-alpha) + B * alpha (Standard painter composition)
                    if has_img2:
                        adjusted_img2 = getattr(self, 'image2_adjusted_preview', None)
                        if adjusted_img2:
                            # Apply mixing percentage
                            painter.setOpacity(blend_value)

                            rect_in_label_space = self._get_overlay_rect_in_label_space(2)
                            if rect_in_label_space:
                                rect_in_canvas_space = QRectF(
                                    rect_in_label_space.left() * render_scale, 
                                    rect_in_label_space.top() * render_scale, 
                                    rect_in_label_space.width() * render_scale, 
                                    rect_in_label_space.height() * render_scale
                                )
                                rotation2 = self.image2_rotation_slider.value() / 10.0 if hasattr(self, 'image2_rotation_slider') else 0.0

                                if abs(rotation2) > 0.01:
                                    center_point_canvas = rect_in_canvas_space.center()
                                    painter.save()
                                    painter.translate(center_point_canvas)
                                    painter.rotate(rotation2)
                                    painter.translate(-center_point_canvas)
                                    painter.drawImage(rect_in_canvas_space, adjusted_img2)
                                    painter.restore()
                                else:
                                    painter.drawImage(rect_in_canvas_space, adjusted_img2)
                            
                            # Reset opacity for subsequent drawing operations
                            painter.setOpacity(1.0)
                
                # --- 3. Draw Guide Lines ---
                if draw_guides and hasattr(self, 'show_guides_checkbox') and self.show_guides_checkbox.isChecked():
                    pen_guides = QPen(Qt.red, 2 * render_scale)
                    painter.setPen(pen_guides)
                    center_x_canvas = canvas.width() // 2
                    center_y_canvas = canvas.height() // 2
                    painter.drawLine(center_x_canvas, 0, center_x_canvas, canvas.height())
                    painter.drawLine(0, center_y_canvas, canvas.width(), center_y_canvas)
            
                painter.end()

            def crop_image(self):
                """Function to crop the current image."""
                if not self.image:
                    return None
            
                # Get crop percentage from sliders
                x_start_percent = 0
                x_end_percent = 100
                y_start_percent = 0
                y_end_percent = 100
            
                # Calculate crop boundaries
                x_start = int(self.image.width() * x_start_percent)
                x_end = int(self.image.width() * x_end_percent)
                y_start = int(self.image.height() * y_start_percent)
                y_end = int(self.image.height() * y_end_percent)
            
                # Ensure cropping is valid
                if x_start >= x_end or y_start >= y_end:
                    QMessageBox.warning(self, "Warning", "Invalid cropping values.")
                    return None
            
                # Crop the image
                cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
                return cropped_image
            
            # Modify align_image and update_crop to preserve settings
            def reset_align_image(self):
                self.orientation_slider.setValue(0)
                self.update_live_view() # Final update with corrected markers and image

            def _finalize_permanent_transformation(self, new_master_image):
                """
                A centralized function to correctly update all state after a permanent
                image transformation like crop, rotate, skew, or rasterize.
                """
                if not new_master_image or new_master_image.isNull():
                    QMessageBox.critical(self, "Error", "A permanent transformation resulted in an invalid image.")
                    if self.undo_stack: self.undo_action_m()
                    return

                self.image_master = new_master_image.copy()
                self.image_before_contrast = self.image_master.copy()
                self.image_contrasted = self.image_master.copy()
                self.image_before_padding = None
                self.image_padded = False
                self.main_image_is_inverted = False
                self.is_modified = True

                self.reset_all_adjustments()

                self._update_status_bar()
                self._update_marker_slider_ranges()
                self._update_overlay_slider_ranges()

                self.update_live_view()
                
                
            def align_image(self):
                # ... (omitted checks)
                angle = float(self.orientation_slider.value() / 20.0)
                if abs(angle) < 0.01:
                    self.orientation_slider.setValue(0); return

                self.save_state() # State saved before rotation

                try:
                    # --- NEW MARKER TRANSFORMATION CODE ---
                    # Note: We must transform the markers based on the original image dimensions,
                    # but the actual rotation is applied to the image_master itself.
                    
                    # 1. Calculate transformation matrix
                    transform_marker = QTransform()
                    current_width = self.image_master.width()
                    current_height = self.image_master.height()
                    transform_marker.translate(current_width / 2.0, current_height / 2.0)
                    transform_marker.rotate(angle)
                    transform_marker.translate(-current_width / 2.0, -current_height / 2.0)
                    
                    # 2. Apply transformation to ALL markers and shapes
                    
                    # Apply to Left/Right Markers (only Y coordinate needs transformation)
                    # However, because rotation changes the entire coordinate system, L/R markers must be treated as custom points (X, Y)
                    # where X is the marker shift added. This is complex and usually requires a coordinate transformation step.
                    # Since standard markers are typically constrained to X=shift_added and only Y changes, they become invalid after rotation.
                    
                    # SOLUTION: Clear L/R/Top markers and warn user if they exist before rotation.
                    if self.left_markers or self.right_markers or self.top_markers:
                        QMessageBox.warning(self, "Warning", "Standard markers (L/R/Top) were cleared due to rotation, as they rely on a rectangular coordinate system alignment.")
                        self.left_markers.clear(); self.right_markers.clear(); self.top_markers.clear()
                        self.left_marker_shift_added = 0; self.right_marker_shift_added = 0; self.top_marker_shift_added = 0

                    # Apply to Custom Markers
                    new_custom_markers = []
                    for marker_data in getattr(self, "custom_markers", []):
                        try:
                            m_list = list(marker_data) 
                            x_old, y_old = float(m_list[0]), float(m_list[1])
                            # Convert image point to QPointF, apply transformation, convert back
                            p_img = QPointF(x_old, y_old)
                            p_new = transform_marker.map(p_img)
                            
                            m_list[0], m_list[1] = p_new.x(), p_new.y()
                            new_custom_markers.append(m_list)
                        except Exception as e:
                            print(f"Error rotating custom marker: {e}")
                            new_custom_markers.append(marker_data) # Keep original if error
                    self.custom_markers = new_custom_markers
                    
                    # Apply to Custom Shapes
                    new_custom_shapes = []
                    for shape_data_orig in getattr(self, "custom_shapes", []):
                        shape_data = dict(shape_data_orig)
                        shape_type = shape_data.get('type')
                        if shape_type == 'line':
                            sx, sy = shape_data['start']; ex, ey = shape_data['end']
                            p_s = transform_marker.map(QPointF(sx, sy)); p_e = transform_marker.map(QPointF(ex, ey))
                            shape_data['start'] = (p_s.x(), p_s.y()); shape_data['end'] = (p_e.x(), p_e.y())
                        elif shape_type == 'rectangle':
                            # Rotating a rect means it is no longer axis-aligned. It must be converted to 4 points.
                            # Simplification: Only store the bounding box of the rotated rectangle.
                            x, y, w, h = shape_data['rect']
                            
                            # Create a polygon from the 4 corners
                            corners_img = [
                                QPointF(x, y), QPointF(x + w, y), 
                                QPointF(x + w, y + h), QPointF(x, y + h)
                            ]
                            rotated_corners = [transform_marker.map(p) for p in corners_img]
                            
                            # Calculate the new axis-aligned bounding box (AABB)
                            new_xs = [p.x() for p in rotated_corners]; new_ys = [p.y() for p in rotated_corners]
                            new_x_min, new_x_max = min(new_xs), max(new_xs)
                            new_y_min, new_y_max = min(new_ys), max(new_ys)
                            
                            # Update shape data with the AABB
                            shape_data['rect'] = (new_x_min, new_y_min, new_x_max - new_x_min, new_y_max - new_y_min)
                        
                        new_custom_shapes.append(shape_data)

                    self.custom_shapes = new_custom_shapes
                    # --- END NEW MARKER TRANSFORMATION CODE ---
                    
                    rotated_image_high_res = self.image_master.transformed(transform_marker, Qt.SmoothTransformation) # Use the same transform
                    
                    self.orientation_slider.setValue(0)
                    
                    # --- BUG FIX: Replace call to _finalize_permanent_transformation ---
                    if rotated_image_high_res.isNull():
                        raise ValueError("Rotation resulted in an invalid image.")

                    self.image_master = rotated_image_high_res.copy()
                    self.image_before_contrast = self.image_master.copy()
                    self.image_contrasted = self.image_master.copy()
                    self.image_before_padding = None
                    self.image_padded = False
                    self.is_modified = True

                    self._update_status_bar()
                    self._update_marker_slider_ranges()
                    self._update_overlay_slider_ranges()
                    self.apply_all_adjustments() # Re-apply visual adjustments
                    # --- END BUG FIX ---

                except Exception as e:
                    QMessageBox.critical(self, "Rotation Error", f"Failed to rotate image: {e}")
                    traceback.print_exc()
            
            def _update_marker_slider_ranges(self):
                """
                Updates slider ranges based on image size AND current marker positions.
                Ensures the range always encompasses the current marker shift values.
                """
                if not self.image or self.image.isNull() or not self.live_view_label:
                    min_abs, max_abs_w, max_abs_h = -100, 1000, 800
                else:
                    try:
                        native_img_width = self.image.width()
                        native_img_height = self.image.height()
                        margin = 100 

                        # Base ranges on image dimensions
                        min_abs_x = -margin
                        max_abs_x = native_img_width + margin
                        min_abs_y = -margin
                        max_abs_y = native_img_height + margin

                        # Expand ranges if current shift values are outside bounds
                        # (e.g. if a crop moved the origin significantly)
                        current_left = getattr(self, 'left_marker_shift_added', 0)
                        current_right = getattr(self, 'right_marker_shift_added', 0)
                        current_top = getattr(self, 'top_marker_shift_added', 0)

                        min_abs_x = min(min_abs_x, current_left - margin, current_right - margin)
                        max_abs_x = max(max_abs_x, current_left + margin, current_right + margin)
                        min_abs_y = min(min_abs_y, current_top - margin)
                        max_abs_y = max(max_abs_y, current_top + margin)

                        max_abs_w = max_abs_x
                        max_abs_h = max_abs_y
                        min_abs = min(min_abs_x, min_abs_y)

                    except Exception as e:
                        print(f"Warning: Error calculating slider ranges: {e}. Using defaults.")
                        min_abs, max_abs_w, max_abs_h = -100, 1000, 800

                self.left_slider_range = [min_abs, max_abs_w]
                self.right_slider_range = [min_abs, max_abs_w]
                self.top_slider_range = [min_abs, max_abs_h]

                sliders_and_ranges = [
                    (getattr(self, 'left_padding_slider', None), self.left_slider_range),
                    (getattr(self, 'right_padding_slider', None), self.right_slider_range),
                    (getattr(self, 'top_padding_slider', None), self.top_slider_range)
                ]

                for slider, new_range in sliders_and_ranges:
                    if slider:
                        current_val = slider.value()
                        slider.blockSignals(True)
                        slider.setRange(new_range[0], new_range[1])
                        # Only re-apply if we aren't about to overwrite it externally
                        # (This preserves value if range shrinks/grows around it)
                        slider.setValue(current_val) 
                        slider.blockSignals(False)
                    
                
                    
            def update_crop(self):
                if not self.image_master or self.image_master.isNull() or not self.crop_rectangle_coords:
                    QMessageBox.warning(self, "Crop Error", "No image loaded or crop area defined.")
                    return

                try:
                    img_x_intent, img_y_intent, img_w_intent, img_h_intent = self.crop_rectangle_coords
                    original_image_width_before_crop = self.image_master.width()
                    original_image_height_before_crop = self.image_master.height()
                    
                    crop_x_start = max(0, int(round(img_x_intent)))
                    crop_y_start = max(0, int(round(img_y_intent)))
                    crop_width = max(1, min(int(round(img_w_intent)), original_image_width_before_crop - crop_x_start))
                    crop_height = max(1, min(int(round(img_h_intent)), original_image_height_before_crop - crop_y_start))
                    
                    if crop_width <= 0 or crop_height <= 0: return

                    self.save_state()

                    # --- Adjust Standard Markers (Lists) ---
                    new_left_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'left_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.left_markers = new_left_markers
                    
                    new_right_markers = [(y_old - crop_y_start, label) for y_old, label in getattr(self, 'right_markers', []) if crop_y_start <= y_old < crop_y_start + crop_height]
                    self.right_markers = new_right_markers
                    
                    new_top_markers = [(x_old - crop_x_start, label) for x_old, label in getattr(self, 'top_markers', []) if crop_x_start <= x_old < crop_x_start + crop_width]
                    self.top_markers = new_top_markers

                    # --- Adjust Custom Markers ---
                    new_custom_markers = [] 
                    if hasattr(self, "custom_markers"):
                        for marker_data in self.custom_markers:
                            try:
                                m_list = list(marker_data); x_old, y_old = float(m_list[0]), float(m_list[1])
                                x_new, y_new = x_old - crop_x_start, y_old - crop_y_start
                                if 0 <= x_new < crop_width and 0 <= y_new < crop_height:
                                    m_list[0] = x_new; m_list[1] = y_new
                                    new_custom_markers.append(m_list)
                            except: pass
                    self.custom_markers = new_custom_markers

                    # --- Adjust Custom Shapes ---
                    new_custom_shapes = []
                    if hasattr(self, "custom_shapes"):
                        for shape_data_orig in self.custom_shapes:
                            shape_data = dict(shape_data_orig); adjusted_shape_data = None
                            try:
                                stype = shape_data.get('type')
                                if stype == 'line':
                                    sx_old, sy_old = map(float, shape_data['start']); ex_old, ey_old = map(float, shape_data['end'])
                                    adjusted_shape_data = shape_data.copy()
                                    adjusted_shape_data['start'] = (sx_old - crop_x_start, sy_old - crop_y_start)
                                    adjusted_shape_data['end'] = (ex_old - crop_x_start, ey_old - crop_y_start)
                                elif stype == 'rectangle':
                                    rx_old, ry_old, rw_old, rh_old = map(float, shape_data['rect'])
                                    overlap_x1 = max(crop_x_start, rx_old); overlap_y1 = max(crop_y_start, ry_old)
                                    overlap_x2 = min(crop_x_start + crop_width, rx_old + rw_old)
                                    overlap_y2 = min(crop_y_start + crop_height, ry_old + rh_old)
                                    if overlap_x2 > overlap_x1 and overlap_y2 > overlap_y1:
                                        adjusted_shape_data = shape_data.copy()
                                        adjusted_shape_data['rect'] = (overlap_x1 - crop_x_start, overlap_y1 - crop_y_start, overlap_x2 - overlap_x1, overlap_y2 - overlap_y1)
                                if adjusted_shape_data: new_custom_shapes.append(adjusted_shape_data)
                            except: pass
                    self.custom_shapes = new_custom_shapes

                    # --- Update Internal Shifts (The Correct New Values) ---
                    self.left_marker_shift_added -= crop_x_start
                    self.right_marker_shift_added -= crop_x_start
                    self.top_marker_shift_added -= crop_y_start

                    # --- Perform Image Crop ---
                    cropped_qimage = self.image_master.copy(crop_x_start, crop_y_start, crop_width, crop_height)
                    if cropped_qimage.isNull(): raise ValueError("Cropping resulted in an invalid image.")

                    # --- Clear Crop UI State ---
                    self.crop_rectangle_coords = None 
                    self.live_view_label.clear_crop_preview() 
                    self.cancel_rectangle_crop_mode() 

                    # --- Update Master Image ---
                    self.image_master = cropped_qimage.copy()
                    self.image_before_contrast = self.image_master.copy()
                    self.image_contrasted = self.image_master.copy()
                    self.image_before_padding = None
                    self.image_padded = False
                    self.is_modified = True

                    # --- Update Display Image (Dimensions Change Here) ---
                    self.apply_all_adjustments()

                    # --- CRITICAL FIX: Sync Sliders with New Dimensions AND New Shift Values ---
                    self._update_status_bar()
                    
                    # 1. Update ranges first (will use new image size + current shift vars)
                    self._update_marker_slider_ranges()
                    self._update_overlay_slider_ranges()

                    # 2. Force sliders to the calculated shift values
                    # This must be done AFTER setRange to ensure the value isn't clamped to old bounds
                    for slider, value_to_set in [
                        (self.left_padding_slider, self.left_marker_shift_added),
                        (self.right_padding_slider, self.right_marker_shift_added),
                        (self.top_padding_slider, self.top_marker_shift_added)
                    ]:
                        if slider:
                            slider.blockSignals(True)
                            slider.setValue(int(value_to_set)) # Ensure integer
                            slider.blockSignals(False)

                except Exception as e:
                    QMessageBox.critical(self, "Crop Error", f"An error occurred during cropping: {e}")
                    traceback.print_exc()
                    if self.undo_stack: self.undo_action_m()
                
            def update_skew(self):
                if not self.image_master or self.image_master.isNull():
                    QMessageBox.warning(self, "Skew Error", "No master image loaded to apply skew.")
                    return

                taper_value = self.taper_skew_slider.value() / 100.0
                if abs(taper_value) < 0.01: # No significant skew to apply
                    self.taper_skew_slider.setValue(0)
                    return

                self.save_state() # Save state before applying the permanent skew

                try:
                    # --- START: New Marker/Shape Transformation Logic ---

                    # 1. Warn user and clear standard markers as skew breaks their alignment logic
                    if self.left_markers or self.right_markers or self.top_markers:
                        QMessageBox.warning(self, "Warning", "Standard markers (L/R/Top) were cleared due to skewing, as they rely on a rectangular coordinate system alignment.")
                        self.left_markers.clear(); self.right_markers.clear(); self.top_markers.clear()
                        self.left_marker_shift_added = 0; self.right_marker_shift_added = 0; self.top_marker_shift_added = 0

                    # 2. Calculate the perspective transformation matrix
                    source_image = self.image_master.copy()
                    np_image_for_dims = self.qimage_to_numpy(source_image)
                    if np_image_for_dims is None: raise ValueError("Failed to get image dimensions.")
                    height, width = np_image_for_dims.shape[:2]

                    source_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])
                    destination_np = np.float32([[0, 0], [width, 0], [width, height], [0, height]])

                    if taper_value > 0:
                        destination_np[0][0] = width * taper_value / 2.0
                        destination_np[1][0] = width * (1 - taper_value / 2.0)
                    elif taper_value < 0:
                        destination_np[3][0] = width * (-taper_value / 2.0)
                        destination_np[2][0] = width * (1 + taper_value / 2.0)

                    matrix = cv2.getPerspectiveTransform(source_np, destination_np)

                    # 3. Apply transformation to custom markers
                    new_custom_markers = []
                    if hasattr(self, "custom_markers") and self.custom_markers:
                        points_to_transform = np.array([[[m[0], m[1]]] for m in self.custom_markers], dtype=np.float32)
                        transformed_points = cv2.perspectiveTransform(points_to_transform, matrix)
                        for i, marker_data in enumerate(self.custom_markers):
                            m_list = list(marker_data)
                            m_list[0] = float(transformed_points[i, 0, 0])
                            m_list[1] = float(transformed_points[i, 0, 1])
                            new_custom_markers.append(m_list)
                    self.custom_markers = new_custom_markers

                    # 4. Apply transformation to custom shapes
                    new_custom_shapes = []
                    if hasattr(self, "custom_shapes") and self.custom_shapes:
                        for shape_data in self.custom_shapes:
                            shape_copy = dict(shape_data)
                            shape_type = shape_copy.get('type')
                            if shape_type == 'line':
                                points = np.array([[shape_copy['start'], shape_copy['end']]], dtype=np.float32)
                                transformed = cv2.perspectiveTransform(points, matrix)
                                shape_copy['start'] = (float(transformed[0,0,0]), float(transformed[0,0,1]))
                                shape_copy['end'] = (float(transformed[0,1,0]), float(transformed[0,1,1]))
                            elif shape_type == 'rectangle':
                                x, y, w, h = shape_copy['rect']
                                corners = np.array([[[x, y], [x + w, y], [x + w, y + h], [x, y + h]]], dtype=np.float32)
                                transformed_corners = cv2.perspectiveTransform(corners, matrix)[0]
                                # Store the new axis-aligned bounding box of the skewed rectangle
                                new_x_min, new_y_min = np.min(transformed_corners, axis=0)
                                new_x_max, new_y_max = np.max(transformed_corners, axis=0)
                                shape_copy['rect'] = (float(new_x_min), float(new_y_min), float(new_x_max - new_x_min), float(new_y_max - new_y_min))
                            new_custom_shapes.append(shape_copy)
                    self.custom_shapes = new_custom_shapes

                    # --- END: New Marker/Shape Transformation Logic ---

                    # 5. Apply skew to the actual image
                    np_image_to_skew = self.qimage_to_numpy(self.image_master)
                    if np_image_to_skew is None: raise ValueError("Failed to convert master image for skewing.")
                    skewed_np_image = cv2.warpPerspective(np_image_to_skew, matrix, (width, height))
                    skewed_image = self.numpy_to_qimage(skewed_np_image)
                    if skewed_image.isNull(): raise ValueError("Skew resulted in an invalid image.")

                    self.taper_skew_slider.setValue(0)

                    # --- BUG FIX: Replace call to _finalize_permanent_transformation ---
                    self.image_master = skewed_image.copy()
                    self.image_before_contrast = self.image_master.copy()
                    self.image_contrasted = self.image_master.copy()
                    self.image_before_padding = None
                    self.image_padded = False
                    self.is_modified = True

                    self._update_status_bar()
                    self._update_marker_slider_ranges()
                    self._update_overlay_slider_ranges()
                    self.apply_all_adjustments() # Re-apply visual adjustments
                    # --- END BUG FIX ---

                except Exception as e:
                    QMessageBox.critical(self, "Skew Error", f"Failed to apply skew: {e}")
                    traceback.print_exc()
                    if self.undo_stack:
                        self.undo_action_m()

            def _draw_oligomer_overlay_on_canvas(self, painter, render_scale, font_scale_factor):
                """Helper to draw the oligomer overlay on a high-res canvas (Save/Copy)."""
                if not (hasattr(self, 'show_oligomer_glyco_overlay_checkbox') and 
                        self.show_oligomer_glyco_overlay_checkbox.isChecked() and 
                        self.oligomer_products and 
                        self.last_mw_prediction_model is not None):
                    return

                line_colors = [QColor(255, 140, 0, 220), QColor(0, 128, 0, 220), QColor(0, 0, 139, 220), QColor(139, 0, 139, 220)]
                text_colors = [QColor("#b35900"), QColor("#004d00"), QColor("#000052"), QColor("#520052")]
                
                # --- UPDATE: Use Standard Marker Font Settings ---
                # This ensures the size matches the Left/Right markers exactly
                base_font_size = self.font_size 
                scaled_font_size = int(base_font_size * font_scale_factor)
                
                text_font = QFont(self.font_family)
                text_font.setPixelSize(max(4, scaled_font_size))
                text_font.setBold(True) # Keep bold to distinguish specific bands, or remove for exact match
                painter.setFont(text_font)
                # -----------------------------------------------
                
                fm_text = QFontMetricsF(text_font)
                text_height = fm_text.height()
                min_text_spacing = text_height * 1.2

                bands_to_draw = []
                for i, mw in enumerate(self.oligomer_products):
                    y_pos_img = self._get_y_pos_from_mw(mw, self.last_mw_prediction_model, self.last_mw_prediction_min_max_pos)
                    if y_pos_img is not None:
                        y_pos_canvas = y_pos_img * render_scale
                        bands_to_draw.append({
                            'mw': mw, 
                            'y_canvas': y_pos_canvas, 
                            'color': line_colors[i % len(line_colors)], 
                            'text_color': text_colors[i % len(text_colors)]
                        })
                
                if not bands_to_draw: return

                bands_to_draw.sort(key=lambda b: b['y_canvas'])
                last_text_y = -float('inf')
                
                for band in bands_to_draw:
                    ideal_text_y = band['y_canvas']
                    if ideal_text_y < last_text_y + min_text_spacing:
                        band['text_y_canvas'] = last_text_y + min_text_spacing
                    else:
                        band['text_y_canvas'] = ideal_text_y
                    last_text_y = band['text_y_canvas']

                arrow_line_width = max(1.0, 2.0 * font_scale_factor)
                
                # Determine Start X (Cursor location or Image Center)
                if hasattr(self, "protein_location") and self.protein_location:
                    start_x = self.protein_location.x() * render_scale
                else:
                    start_x = (self.image_master.width() * render_scale) / 2.0

                # Visual offsets scaled by font_scale_factor to match zoom/view ratio
                line_len_1 = 30 * font_scale_factor
                text_offset = 5 * font_scale_factor
                elbow_x = start_x + line_len_1
                text_start_x = elbow_x + text_offset

                for band in bands_to_draw:
                    pen = QPen(band['color'])
                    pen.setWidthF(arrow_line_width)
                    painter.setPen(pen)
                    
                    text = f"{band['mw']:.1f} kDa"
                    text_rect = fm_text.boundingRect(text)
                    
                    p_start = QPointF(start_x, band['y_canvas'])
                    p_elbow_1 = QPointF(elbow_x, band['y_canvas'])
                    p_elbow_2 = QPointF(elbow_x, band['text_y_canvas'])
                    p_text_anchor = QPointF(text_start_x, band['text_y_canvas'])
                    
                    polyline_points = [p_start, p_elbow_1, p_elbow_2, p_text_anchor]
                    painter.drawPolyline(QPolygonF(polyline_points))
                    
                    text_draw_y = band['text_y_canvas'] - (text_rect.top() + text_rect.height() / 2.0)
                    
                    # Glow Effect
                    glow_color = QColor(255, 255, 255, 90)
                    glow_pen = QPen(glow_color)
                    glow_pen.setWidth(int(3 * font_scale_factor)) 
                    painter.setPen(glow_pen)
                    
                    glow_offset = max(1.0, 1.0 * font_scale_factor)
                    offsets = [(-glow_offset, -glow_offset), (glow_offset, -glow_offset), (-glow_offset, glow_offset), (glow_offset, glow_offset)]
                    for dx, dy in offsets:
                        painter.drawText(QPointF(text_start_x + dx, text_draw_y + dy), text)
                        
                    painter.setPen(band['text_color'])
                    painter.drawText(QPointF(text_start_x, text_draw_y), text)
                
            def save_image(self):
                self.draw_guides = False
                if hasattr(self, 'show_guides_checkbox'): self.show_guides_checkbox.setChecked(False)

                if not self.image or self.image.isNull():
                     QMessageBox.warning(self, "Error", "No image data to save.")
                     return False

                # --- 1. Generate name and Open Dialog (Unchanged) ---
                suggested_name = os.path.splitext(os.path.basename(self.image_path))[0] if self.image_path else "untitled_image"
                base_name_clean_for_dialog = suggested_name.replace("_original", "").replace("_modified", "")
                save_dir = os.path.dirname(self.image_path) if self.image_path else ""
                
                options = QFileDialog.Options()
                base_save_path, selected_filter = QFileDialog.getSaveFileName(
                    self, "Save Image Base Name", os.path.join(save_dir, base_name_clean_for_dialog),
                    "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
                    options=options
                )
                if not base_save_path: return False

                user_selected_base_name = os.path.splitext(base_save_path)[0]
                final_clean_base = user_selected_base_name.replace("_original", "").replace("_modified", "")
                
                suffix = os.path.splitext(base_save_path)[1].lower() or ".png"
                if "tif" in selected_filter.lower(): suffix = ".tif"
                elif "jpg" in selected_filter.lower(): suffix = ".jpg"
                elif "bmp" in selected_filter.lower(): suffix = ".bmp"

                original_save_path = f"{final_clean_base}_original{suffix}"
                modified_save_path = f"{final_clean_base}_modified{suffix}"
                config_save_path = f"{final_clean_base}_config.txt"

                # --- Save _original image (Unchanged) ---
                if self.image_master and not self.image_master.isNull():
                    save_format = suffix.replace(".", "").upper()
                    if save_format == "TIF": save_format = "TIFF"
                    quality = 95 if save_format in ["JPG", "JPEG"] else -1
                    if not self.image_master.save(original_save_path, format=save_format if save_format else None, quality=quality):
                        QMessageBox.warning(self, "Error", f"Failed to save original image.")

                # --- Create and save _modified image ---
                render_scale = 3
                native_width = self.image_master.width(); native_height = self.image_master.height()
                if native_width <= 0 or native_height <= 0: return False

                current_adjustment_settings = {
                    'is_inverted': self.main_image_is_inverted,
                    'levels_gamma': {
                        'black_point': self.black_point_slider.value(),
                        'white_point': self.white_point_slider.value(),
                        'gamma': self.gamma_slider.value()
                    },
                    'channel_mixer': self.channel_mixer_data.copy(),
                    'unsharp_mask': self.unsharp_mask_data.copy(),
                    'clahe': self.clahe_data.copy()
                }

                # Apply settings to master image (this now returns 16-bit if appropriate)
                fully_adjusted_master = self._apply_all_adjustments_to_image(self.image_master, current_adjustment_settings)

                # --- FIX: Determine Canvas Format based on Image Depth ---
                canvas_width = native_width * render_scale
                canvas_height = native_height * render_scale
                
                # Check if result is high depth (Grayscale16 or RGBA64/RGBX64)
                fmt = fully_adjusted_master.format()
                is_high_depth = fmt in [QImage.Format_Grayscale16, QImage.Format_RGBA64, QImage.Format_RGBX64]
                
                if is_high_depth:
                    # Use 64-bit RGBA for the canvas to preserve depth while allowing colored annotations
                    canvas_format = QImage.Format_RGBA64
                else:
                    canvas_format = QImage.Format_ARGB32_Premultiplied

                modified_canvas = QImage(canvas_width, canvas_height, canvas_format)
                modified_canvas.fill(Qt.transparent)
                
                painter = QPainter(modified_canvas)
                painter.setRenderHint(QPainter.SmoothPixmapTransform, False)
                
                # Draw the image
                painter.drawImage(QRectF(0.0, 0.0, float(canvas_width), float(canvas_height)), fully_adjusted_master, QRectF(fully_adjusted_master.rect()))

                # ... (Draw Annotations - Code remains exactly the same as previous) ...
                label_width = float(self.live_view_label.width()); label_height = float(self.live_view_label.height())
                scale_native_to_view = min(label_width / native_width, label_height / native_height) if label_width > 0 and label_height > 0 else 1.0
                font_scale_factor = render_scale / scale_native_to_view if scale_native_to_view > 1e-6 else render_scale
                
                painter.setRenderHint(QPainter.Antialiasing, True); painter.setRenderHint(QPainter.TextAntialiasing, True)
                
                def map_img_coords_to_canvas(img_x, img_y):
                    return QPointF(img_x * render_scale, img_y * render_scale)

                std_font = QFont(self.font_family); std_font.setPixelSize(int(self.font_size * font_scale_factor))
                painter.setFont(std_font); painter.setPen(self.font_color)
                fm_std = QFontMetrics(std_font); y_offset_baseline = fm_std.height() * 0.3
                for y_img, text in self.left_markers:
                    anchor_x = self.left_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    full_text = f"{text} ⎯"; painter.drawText(QPointF(anchor_x - fm_std.horizontalAdvance(full_text), anchor_y + y_offset_baseline), full_text)
                for y_img, text in self.right_markers:
                    anchor_x = self.right_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    painter.drawText(QPointF(anchor_x, anchor_y + y_offset_baseline), f"⎯ {text}")
                for x_img, text in self.top_markers:
                    painter.save(); anchor_x = x_img * render_scale; anchor_y = self.top_marker_shift_added * render_scale
                    painter.translate(anchor_x, anchor_y + y_offset_baseline); painter.rotate(self.font_rotation)
                    painter.drawText(QPointF(0, 0), str(text)); painter.restore()
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        x, y, text, color, font, size, is_bold, is_italic = marker_data
                        custom_font = QFont(font); custom_font.setPixelSize(int(size * font_scale_factor)); custom_font.setBold(is_bold); custom_font.setItalic(is_italic)
                        painter.setFont(custom_font); painter.setPen(QColor(color))
                        fm = QFontMetrics(custom_font); rect = fm.boundingRect(text)
                        draw_pos = QPointF(x * render_scale - rect.center().x(), y * render_scale - rect.center().y())
                        painter.drawText(draw_pos, text)
                    except Exception: pass
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type, color_str, thickness = shape_data.get('type'), shape_data.get('color'), shape_data.get('thickness')
                        pen = QPen(QColor(color_str), max(1.0, thickness * render_scale)); painter.setPen(pen)
                        if shape_type == 'line':
                            painter.drawLine(map_img_coords_to_canvas(*shape_data['start']), map_img_coords_to_canvas(*shape_data['end']))
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            painter.drawRect(QRectF(map_img_coords_to_canvas(x, y), QSizeF(w * render_scale, h * render_scale)))
                    except Exception: pass
                
                self._draw_oligomer_overlay_on_canvas(painter, render_scale, font_scale_factor)
                painter.end()

                save_format_mod = suffix.replace(".", "").upper()
                if save_format_mod == "TIF": save_format_mod = "TIFF"
                if not modified_canvas.save(modified_save_path, format=save_format_mod if save_format_mod else None, quality=-1):
                    QMessageBox.warning(self, "Error", f"Failed to save modified image.")
                    return False
                
                # --- Save Config File ---
                config_data = self.get_current_config()
                try:
                    with open(config_save_path, "w", encoding='utf-8') as config_file:
                        json.dump(config_data, config_file, indent=4)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
                    return False

                self.is_modified = False
                self.image_path = original_save_path
                
                QMessageBox.information(self, "Saved", f"Files saved successfully to '{os.path.dirname(base_save_path)}'")
                self.setWindowTitle(f"{self.window_title}::{final_clean_base}")
                self._update_status_bar()

                return True
                    
            def save_image_svg(self):
                """Save the processed image along with markers, labels, and custom shapes in SVG format."""
                if not self.image or self.image.isNull(): # Check current image validity
                    QMessageBox.warning(self, "Warning", "No image to save.")
                    return

                options = QFileDialog.Options()
                file_path, _ = QFileDialog.getSaveFileName(
                    self, "Save Image as SVG for MS Word/Vector Editing", "", "SVG Files (*.svg)", options=options
                )

                if not file_path:
                    return

                if not file_path.lower().endswith(".svg"): # Ensure .svg extension
                    file_path += ".svg"

                # --- Determine Render/Canvas Dimensions ---
                # Use the current image dimensions directly as the base SVG size
                # This avoids scaling issues if the view label size is different.
                # Markers and shapes will be positioned relative to this size.
                svg_width = self.image.width()
                svg_height = self.image.height()

                if svg_width <= 0 or svg_height <= 0:
                     QMessageBox.warning(self, "Warning", "Invalid image dimensions for SVG.")
                     return

                # --- Create SVG Drawing Object ---
                dwg = svgwrite.Drawing(file_path, profile='tiny', size=(f"{svg_width}px", f"{svg_height}px"))
                # Set viewbox to match image dimensions for correct coordinate system
                dwg.viewbox(0, 0, svg_width, svg_height)

                # --- Embed the Base Image ---
                try:
                    # Convert the QImage to a base64-encoded PNG for embedding
                    buffer = QBuffer()
                    buffer.open(QBuffer.ReadWrite)
                    # Save the *current* self.image (which might be cropped/transformed)
                    if not self.image.save(buffer, "PNG"):
                        raise IOError("Failed to save image to PNG buffer for SVG.")
                    image_data = base64.b64encode(buffer.data()).decode('utf-8')
                    buffer.close()

                    # Embed the image at position (0, 0) with original dimensions
                    dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}",
                                      insert=(0, 0),
                                      size=(f"{svg_width}px", f"{svg_height}px")))
                except Exception as e:
                    QMessageBox.critical(self, "SVG Error", f"Failed to embed base image: {e}")
                    return # Stop if base image fails

                # --- Define marker/label font style for SVG ---
                # Use the standard marker font settings
                svg_font_family = self.font_family
                svg_font_size_px = f"{self.font_size}px" # Use pixels for SVG consistency
                svg_font_color = self.font_color.name() if self.font_color else "#000000"

                # Calculate horizontal offset for left/right markers based on font size
                # Use QFontMetrics for accurate width calculation
                try:
                     qfont_for_metrics = QFont(svg_font_family, self.font_size)
                     font_metrics = QFontMetrics(qfont_for_metrics)
                     # Use a representative character like 'm' or average width if needed
                     avg_char_width = font_metrics.averageCharWidth()
                     horizontal_offset = avg_char_width * 0.5 # Small offset from edge
                     vertical_offset_adjust = font_metrics.ascent() * 0.75 # Adjustment for vertical alignment
                except Exception:
                     horizontal_offset = 5 # Fallback offset
                     vertical_offset_adjust = self.font_size * 0.75 # Fallback adjustment


                # --- Add Left Markers ---
                left_marker_x_pos = self.left_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "left_markers", []):
                    final_text = f"{text}" # Remove the line "⎯" for cleaner SVG text
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(left_marker_x_pos - horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="end" # Align text right, ending at the x position
                        )
                    )

                # --- Add Right Markers ---
                right_marker_x_pos = self.right_marker_shift_added # Use the absolute offset
                for y_pos, text in getattr(self, "right_markers", []):
                    final_text = f"{text}" # Remove the line "⎯"
                    dwg.add(
                        dwg.text(
                            final_text,
                            insert=(right_marker_x_pos + horizontal_offset, y_pos + vertical_offset_adjust),
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="start" # Align text left, starting at the x position
                        )
                    )

                # --- Add Top Markers ---
                top_marker_y_pos = self.top_marker_shift_added # Use the absolute offset
                for x_pos, text in getattr(self, "top_markers", []):
                    dwg.add(
                        dwg.text(
                            text,
                            insert=(x_pos, top_marker_y_pos + vertical_offset_adjust), # Apply vertical adjust
                            fill=svg_font_color,
                            font_family=svg_font_family,
                            font_size=svg_font_size_px,
                            text_anchor="middle", # Center text horizontally over the x position
                            # Apply rotation around the insertion point
                            transform=f"rotate({self.font_rotation}, {x_pos}, {top_marker_y_pos + vertical_offset_adjust})"
                        )
                    )

                # --- Add Custom Markers ---
                for marker_tuple in getattr(self, "custom_markers", []):
                    try:
                        # Default values for optional elements
                        is_bold = False
                        is_italic = False

                        # Unpack based on length for backward compatibility
                        if len(marker_tuple) == 8:
                            x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple
                        elif len(marker_tuple) == 6:
                            x_pos, y_pos, marker_text, color, font_family, font_size = marker_tuple
                        else:
                            continue # Skip invalid marker data

                        # Prepare SVG attributes
                        text_content = str(marker_text)
                        fill_color = QColor(color).name() if isinstance(color, QColor) else str(color) # Ensure hex/name
                        font_family_svg = str(font_family)
                        font_size_svg = f"{int(font_size)}px" if isinstance(font_size, (int, float)) else "12px"
                        font_weight_svg = "bold" if bool(is_bold) else "normal"
                        font_style_svg = "italic" if bool(is_italic) else "normal"

                        # Adjust vertical position slightly for better alignment if needed
                        # This might require font metrics specific to the marker's font
                        # For simplicity, using the standard marker offset for now
                        y_pos_adjusted = y_pos + vertical_offset_adjust

                        # Add SVG text element, centered at the marker's coordinates
                        dwg.add(
                            dwg.text(
                                text_content,
                                insert=(x_pos, y_pos_adjusted), # Position text anchor at the coordinate
                                fill=fill_color,
                                font_family=font_family_svg,
                                font_size=font_size_svg,
                                font_weight=font_weight_svg,
                                font_style=font_style_svg,
                                text_anchor="middle", # Center horizontally
                                dominant_baseline="central" # Attempt vertical centering (support varies)
                            )
                        )
                    except Exception as e:
                         print(f"Warning: Skipping invalid custom marker during SVG export: {e}")
                         # import traceback; traceback.print_exc() # Uncomment for detailed debug

                # --- START: Add Custom Shapes (Lines and Rectangles) ---
                for shape_data in getattr(self, "custom_shapes", []):
                     try:
                         shape_type = shape_data.get('type')
                         color_str = shape_data.get('color', '#000000') # Default black
                         thickness = int(shape_data.get('thickness', 1))
                         if thickness < 1: thickness = 1

                         if shape_type == 'line':
                             start_coords = shape_data.get('start')
                             end_coords = shape_data.get('end')
                             if start_coords and end_coords:
                                 dwg.add(dwg.line(start=start_coords,
                                                  end=end_coords,
                                                  stroke=color_str,
                                                  stroke_width=thickness))
                         elif shape_type == 'rectangle':
                             rect_coords = shape_data.get('rect') # (x, y, w, h)
                             if rect_coords:
                                 x, y, w, h = rect_coords
                                 dwg.add(dwg.rect(insert=(x, y),
                                                  size=(f"{w}px", f"{h}px"),
                                                  stroke=color_str,
                                                  stroke_width=thickness,
                                                  fill="none")) # No fill for outline rectangle
                     except Exception as e:
                         print(f"Warning: Skipping invalid custom shape during SVG export: {e}")
                # --- END: Add Custom Shapes ---

                # --- Save the SVG file ---
                try:
                    dwg.save()
                    QMessageBox.information(self, "Success", f"Image and annotations saved as SVG at\n{file_path}")
                    self.is_modified = False # Mark as saved if successful
                except Exception as e:
                    QMessageBox.critical(self, "SVG Save Error", f"Failed to save SVG file:\n{e}")
            
            
            def copy_to_clipboard(self):
                if not self.image_master or self.image_master.isNull(): # Check against master image
                    QMessageBox.warning(self, "Warning", "No image to copy.")
                    return

                # Clean up any previous temp file from this session
                self.cleanup_temp_clipboard_file()

                # --- Render high-resolution canvas ---
                render_scale = 3
                native_width = self.image_master.width()
                native_height = self.image_master.height()
                if native_width <= 0 or native_height <= 0: return

                # --- START FIX: Create a fully adjusted, high-quality base image ---
                # 1. Gather all current adjustment settings
                current_adjustment_settings = {
                    'is_inverted': self.main_image_is_inverted,
                    'levels_gamma': {
                        'black_point': self.black_point_slider.value(),
                        'white_point': self.white_point_slider.value(),
                        'gamma': self.gamma_slider.value()
                    },
                    'channel_mixer': self.channel_mixer_data.copy(),
                    'unsharp_mask': self.unsharp_mask_data.copy(),
                    'clahe': self.clahe_data.copy()
                }

                # 2. Use the single, reliable helper function to apply all adjustments
                #    to the master image, just like save_image() does.
                base_image_for_copy = self._apply_all_adjustments_to_image(self.image_master, current_adjustment_settings)
                # --- END FIX ---


                canvas_width = native_width * render_scale
                canvas_height = native_height * render_scale
                # Use a high-quality format for the canvas
                render_canvas = QImage(canvas_width, canvas_height, QImage.Format_ARGB32_Premultiplied)
                render_canvas.fill(Qt.transparent)
                
                painter = QPainter(render_canvas)
                if not painter.isActive(): return
                
                painter.setRenderHint(QPainter.SmoothPixmapTransform, False)
                # Draw the fully adjusted high-quality image as the base
                painter.drawImage(QRectF(0.0, 0.0, float(canvas_width), float(canvas_height)), base_image_for_copy, QRectF(base_image_for_copy.rect()))
                
                # --- The rest of the drawing logic for annotations is correct and remains unchanged ---
                label_width = float(self.live_view_label.width()); label_height = float(self.live_view_label.height())
                scale_native_to_view = min(label_width / native_width, label_height / native_height) if label_width > 0 and label_height > 0 else 1.0
                font_scale_factor = render_scale / scale_native_to_view if scale_native_to_view > 1e-6 else render_scale

                painter.setRenderHint(QPainter.Antialiasing, True); painter.setRenderHint(QPainter.TextAntialiasing, True)
                def map_img_coords_to_canvas(img_x, img_y): return QPointF(img_x * render_scale, img_y * render_scale)
                std_font = QFont(self.font_family)
                std_font.setPixelSize(int(self.font_size * font_scale_factor))
                painter.setFont(std_font); painter.setPen(self.font_color)
                fm_std = QFontMetrics(std_font); y_offset_baseline = fm_std.height() * 0.3
                for y_img, text in self.left_markers:
                    anchor_x = self.left_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    full_text = f"{text} ⎯"; painter.drawText(QPointF(anchor_x - fm_std.horizontalAdvance(full_text), anchor_y + y_offset_baseline), full_text)
                for y_img, text in self.right_markers:
                    anchor_x = self.right_marker_shift_added * render_scale; anchor_y = y_img * render_scale
                    painter.drawText(QPointF(anchor_x, anchor_y + y_offset_baseline), f"⎯ {text}")
                for x_img, text in self.top_markers:
                    painter.save(); anchor_x = x_img * render_scale; anchor_y = self.top_marker_shift_added * render_scale
                    painter.translate(anchor_x, anchor_y + y_offset_baseline); painter.rotate(self.font_rotation)
                    painter.drawText(QPointF(0, 0), str(text)); painter.restore()
                for marker_data in getattr(self, "custom_markers", []):
                    try:
                        x, y, text, color, font, size, is_bold, is_italic = marker_data
                        custom_font = QFont(font); custom_font.setPixelSize(int(size * font_scale_factor)); custom_font.setBold(is_bold); custom_font.setItalic(is_italic)
                        painter.setFont(custom_font); painter.setPen(QColor(color))
                        fm = QFontMetrics(custom_font); rect = fm.boundingRect(text)
                        draw_pos = QPointF(x * render_scale - rect.center().x(), y * render_scale - rect.center().y())
                        painter.drawText(draw_pos, text)
                    except Exception: pass
                for shape_data in getattr(self, "custom_shapes", []):
                    try:
                        shape_type, color_str, thickness = shape_data.get('type'), shape_data.get('color'), shape_data.get('thickness')
                        pen = QPen(QColor(color_str), max(1.0, thickness * render_scale)); painter.setPen(pen)
                        if shape_type == 'line': painter.drawLine(map_img_coords_to_canvas(*shape_data['start']), map_img_coords_to_canvas(*shape_data['end']))
                        elif shape_type == 'rectangle':
                            x, y, w, h = shape_data['rect']
                            painter.drawRect(QRectF(map_img_coords_to_canvas(x, y), QSizeF(w * render_scale, h * render_scale)))
                    except Exception: pass
                
                self._draw_oligomer_overlay_on_canvas(painter, render_scale, font_scale_factor)
                painter.end()
                
                # --- The rest of the method, which saves to a temp file, remains the same ---
                try:
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_file:
                        temp_file_path = temp_file.name
                    
                    self.temp_clipboard_file_path = temp_file_path

                    if not render_canvas.save(temp_file_path, "PNG"):
                        raise IOError(f"Failed to save temporary file to {temp_file_path}")

                    mime_data = QMimeData()
                    mime_data.setUrls([QUrl.fromLocalFile(temp_file_path)])
                    mime_data.setImageData(render_canvas) # Fallback for some apps
                    
                    QApplication.clipboard().setMimeData(mime_data)
                    QMessageBox.information(self, "Copied", "High-resolution image with transparency copied to clipboard.")

                except Exception as e:
                    QMessageBox.critical(self, "Copy Error", f"Could not copy image to clipboard via temporary file:\n{e}")
                    self.cleanup_temp_clipboard_file()


            def cleanup_temp_clipboard_file(self):
                """Deletes the temporary clipboard file if it exists."""
                path_to_delete = getattr(self, 'temp_clipboard_file_path', None)
                if path_to_delete:
                    if os.path.exists(path_to_delete):
                        try:
                            os.remove(path_to_delete)
                        except OSError as e:
                            print(f"WARNING: Could not delete temp clipboard file {path_to_delete}: {e}")
                    self.temp_clipboard_file_path = None
                
                
            def clear_predict_molecular_weight(self):
                # ... (existing clear logic for MW prediction, single bounding boxes, etc.) ...
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.preview_marker_text = ""
                self.live_view_label.setCursor(Qt.ArrowCursor)
                if hasattr(self, "protein_location"):
                    del self.protein_location 
                self.predict_size=False
                self.bounding_boxes=[]
                self.clear_measurement_mode()
                self.bounding_box_start = None
                # self.live_view_label.bounding_box_start = None # This is not an attribute of LiveViewLabel
                # self.live_view_label.bounding_box_preview = None # Will be cleared below
                self.quantities=[]
                self.peak_area_list=[]
                self.protein_quantities=[]
                self.standard_protein_values.setText("")
                self.standard_protein_areas=[]
                self.standard_protein_areas_text.setText("")
                # self.live_view_label.quad_points=[] # Will be cleared below
                # self.live_view_label.bounding_box_preview = None # Will be cleared below
                # self.live_view_label.rectangle_points = [] # This is likely for single rect, also clear
                self.latest_calculated_quantities = []
                self.quantities_peak_area_dict={}
                self.last_mw_prediction_model = None
                self.last_mw_prediction_min_max_pos = None
                self.oligomer_products = []
                self.last_mw_prediction_marker_x_ls = None
                self.last_mw_prediction_marker_width_ls = None
                if hasattr(self, 'show_oligomer_glyco_overlay_checkbox'):
                    self.show_oligomer_glyco_overlay_checkbox.setChecked(False)
                
                # --- MODIFIED/ADDED: Clear multi-lane and LiveViewLabel preview states ---
                self.multi_lane_mode_active = False
                self.multi_lane_definition_type = None
                self.multi_lane_definitions = []
                self.current_multi_lane_points = []      # Clears app's buffer for current multi-lane quad
                self.current_multi_lane_rect_start = None # Clears app's buffer for current multi-lane rect
                self.latest_multi_lane_peak_areas = {}
                self.latest_multi_lane_calculated_quantities = {}
                self.latest_multi_lane_peak_details = {} # Also clear this
                self.target_protein_areas_text.clear()
                # REMOVED the check for the non-existent button here
                self.multi_lane_processing_finished = False

                # Explicitly clear LiveViewLabel's general-purpose preview buffers
                self.live_view_label.quad_points = []
                self.live_view_label.bounding_box_preview = None
                self.live_view_label.rectangle_points = [] # Also clear this, it's used for single rect
                self.live_view_label.rectangle_start = None
                self.live_view_label.rectangle_end = None
                self._reset_live_view_label_custom_handlers()
                self.update_live_view()
                
            def predict_molecular_weight(self):
                # --- Step 1: Check for available valid marker sets ---
                has_left = hasattr(self, 'left_markers') and len(self.left_markers) >= 2
                has_right = hasattr(self, 'right_markers') and len(self.right_markers) >= 2
        
                choices = []
                if has_left:
                    choices.append("Left Markers")
                if has_right:
                    choices.append("Right Markers")
                if has_left and has_right:
                    choices.append("Average of Left & Right")
        
                if not choices:
                    QMessageBox.warning(self, "Error", "At least one set of two or more markers (Left or Right) is required for prediction.")
                    return
        
                # --- Step 2: Prompt user to select the marker source ---
                source = choices[0] # Default to the first available option
                if len(choices) > 1:
                    selected_source, ok = QInputDialog.getItem(self, "Select Marker Source",
                                                               "Which markers should be used for the standard curve?",
                                                               choices, 2, False)
                    if not ok or not selected_source:
                        return # User cancelled
                    source = selected_source
        
                # --- Step 3: Prepare the marker data based on user's choice ---
                markers_raw_tuples = []
                if source == "Left Markers":
                    markers_raw_tuples = self.left_markers
                elif source == "Right Markers":
                    markers_raw_tuples = self.right_markers
                elif source == "Average of Left & Right":
                    
                    # --- Self-contained logic for partitioning and averaging ---
                    def _partition(marker_list):
                        """Nested helper to partition a marker list into one or two sets."""
                        if not marker_list or len(marker_list) < 2: return (marker_list, [])
                        
                        numeric_markers = sorted(
                            [(float(pos), float(val)) for pos, val in marker_list if str(val).replace('.', '', 1).isdigit()],
                            key=lambda item: item[0]
                        )
                        if len(numeric_markers) < 2: return (numeric_markers, [])
                        
                        values = [val for pos, val in numeric_markers]
                        transition_index = -1
                        initial_decrease = any(values[k] < values[k-1] for k in range(1, len(values)))

                        if initial_decrease:
                            for k in range(1, len(values)):
                                if values[k] > values[k-1]:
                                    transition_index = k
                                    break
                        
                        return (numeric_markers[:transition_index], numeric_markers[transition_index:]) if transition_index != -1 else (numeric_markers, [])

                    def _average(set_a, set_b):
                        """Nested helper to average two corresponding sets."""
                        if not set_a or not set_b: return []
                        map_a = {val: pos for pos, val in set_a}
                        map_b = {val: pos for pos, val in set_b}
                        common_values = set(map_a.keys()) & set(map_b.keys())
                        return [( (map_a[val] + map_b[val]) / 2.0, val ) for val in sorted(list(common_values))]

                    left_set1, left_set2 = _partition(self.left_markers)
                    right_set1, right_set2 = _partition(self.right_markers)

                    avg_set1 = _average(left_set1, right_set1)
                    avg_set2 = _average(left_set2, right_set2)
                    
                    markers_raw_tuples = avg_set1 + avg_set2
                    
                    if len(markers_raw_tuples) < 2:
                        QMessageBox.warning(self, "Averaging Error", 
                                            "Fewer than two common marker values were found between the Left and Right sets after partitioning.\n"
                                            "Cannot create an averaged standard curve.")
                        return
                    # --- End of self-contained logic ---

                # --- Step 4: Process the selected markers (validation, sorting) ---
                self.live_view_label.preview_marker_enabled = False
                self.live_view_label.mw_predict_preview_enabled = True
                self.live_view_label.mw_predict_preview_position = None
                self.live_view_label.setCursor(Qt.CrossCursor)
                self.live_view_label.setMouseTracking(True)
        
                try:
                    numeric_markers = []
                    for pos, val in markers_raw_tuples:
                        try:
                            # If from "Average", value is already float. Otherwise, it could be a string.
                            numeric_markers.append((float(pos), float(val)))
                        except (ValueError, TypeError):
                            continue # Skip non-numeric markers

                    if len(numeric_markers) < 2:
                         QMessageBox.warning(self, "Error", f"The selected set ('{source}') has fewer than two valid numeric markers.")
                         self.live_view_label.setCursor(Qt.ArrowCursor)
                         return

                    # Sort markers by Y-position (migration distance)
                    sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
                    sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
                    sorted_marker_values = np.array([val for pos, val in sorted_markers])

                except Exception as e:
                    QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
                    self.live_view_label.setCursor(Qt.ArrowCursor)
                    return
        
                # --- Step 5: Set up the click handler with the processed marker data ---
                QMessageBox.information(self, "Instruction",
                                        f"Using '{source}' for calculation.\n\n"
                                        "Click on the target protein location in the preview window.")
                
                self._reset_live_view_label_custom_handlers()
                self.live_view_label._custom_left_click_handler_from_app = lambda event: self.get_protein_location_and_clear_preview(
                    event, sorted_marker_positions, sorted_marker_values
                )
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_mw_predict_preview

            def _update_main_model_from_dialog(self, model_name: str):
                """Slot to update the main window's regression combo box from the dialog."""
                if hasattr(self, 'mw_regression_model_combo'):
                    self.mw_regression_model_combo.blockSignals(True)
                    self.mw_regression_model_combo.setCurrentText(model_name)
                    self.mw_regression_model_combo.blockSignals(False)

            def _calculate_refined_y_from_event(self, event):
                pos = event.position()
                cursor_x, cursor_y = pos.x(), pos.y()

                # Coordinate Transforms
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
                image_width = self.image.width() if self.image and self.image.width() > 0 else 1
                image_height = self.image.height() if self.image and self.image.height() > 0 else 1

                scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1.0
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2

                raw_img_x = (cursor_x - x_offset) / scale
                raw_img_y = (cursor_y - y_offset) / scale
                refined_y = raw_img_y

                # Centroid Refinement
                if self.image and not self.image.isNull():
                    try:
                        np_img = self.qimage_to_numpy(self.image)
                        if np_img is not None:
                            h, w = np_img.shape[:2]
                            search_radius = 15
                            center_x_int = int(round(raw_img_x))
                            center_y_int = int(round(raw_img_y))
                            
                            y_start = max(0, center_y_int - search_radius)
                            y_end = min(h, center_y_int + search_radius + 1)
                            x_col = max(0, min(w - 1, center_x_int))

                            if np_img.ndim == 3: slice_data = np.mean(np_img[y_start:y_end, x_col, :3], axis=1)
                            else: slice_data = np_img[y_start:y_end, x_col]

                            if not self.main_image_is_inverted:
                                max_val = 65535.0 if slice_data.dtype == np.uint16 else 255.0
                                slice_data = max_val - slice_data

                            slice_data = slice_data.astype(float) - np.min(slice_data)
                            total_mass = np.sum(slice_data)
                            if total_mass > 0:
                                relative_centroid = np.sum(np.arange(len(slice_data)) * slice_data) / total_mass
                                refined_y = y_start + relative_centroid
                    except Exception: pass
                
                return refined_y

            # --- HELPER: Picking Loop for Calibration ---
            def _pick_calibration_point_loop(self):
                """Enters a local event loop to let user pick a point, returning refined Y."""
                
                # 1. Setup UI for picking
                self.live_view_label.setCursor(Qt.CrossCursor)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.mw_predict_preview_enabled = True # Enable preview line
                self.live_view_label.setMouseTracking(True)
                
                # 2. Local Event Loop
                loop = QEventLoop()
                picked_data = {"y": None}

                def pick_handler(event):
                    refined_y = self._calculate_refined_y_from_event(event)
                    picked_data["y"] = refined_y
                    loop.quit()

                self.live_view_label._custom_left_click_handler_from_app = pick_handler
                self.live_view_label._custom_mouseMoveEvent_from_app = self.update_mw_predict_preview
                
                loop.exec() # BLOCK here until clicked

                # 3. Cleanup
                self.live_view_label.mw_predict_preview_enabled = False
                self.live_view_label.setMouseTracking(False)
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.update()
                
                return picked_data["y"]

            def get_protein_location(self, event, all_marker_positions, all_marker_values):
                # Calculate initial position using the shared helper
                refined_y = self._calculate_refined_y_from_event(event)
                
                # Store simple point for line drawing compatibility
                pos = event.position()
                cursor_x = pos.x()
                if self.live_view_label.zoom_level != 1.0:
                    cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                
                # (Re-calculate X mapping locally)
                displayed_width = self.live_view_label.width(); displayed_height = self.live_view_label.height()
                image_width = self.image.width() if self.image else 1; image_height = self.image.height() if self.image else 1
                scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 else 1.0
                x_offset = (displayed_width - image_width * scale) / 2
                raw_img_x = (cursor_x - x_offset) / scale
                
                self.protein_location = QPointF(raw_img_x, refined_y)

                # Determine Marker Set
                transition_index = -1; initial_decrease = False
                for k in range(1, len(all_marker_values)):
                    if all_marker_values[k] < all_marker_values[k-1]: initial_decrease = True
                    if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                        transition_index = k; break

                active_marker_positions, active_marker_values, set_name = None, None, "Full Set"
                if transition_index != -1:
                    set1_pos = all_marker_positions[:transition_index]
                    set1_val = all_marker_values[:transition_index]
                    set2_pos = all_marker_positions[transition_index:]
                    set2_val = all_marker_values[transition_index:]
                    
                    mean1 = np.mean(set1_pos) if len(set1_pos)>0 else -9e9
                    mean2 = np.mean(set2_pos) if len(set2_pos)>0 else -9e9
                    
                    if abs(refined_y - mean1) <= abs(refined_y - mean2):
                        active_marker_positions, active_marker_values, set_name = set1_pos, set1_val, "Set 1"
                    else:
                        active_marker_positions, active_marker_values, set_name = set2_pos, set2_val, "Set 2"
                else:
                    active_marker_positions, active_marker_values = all_marker_positions, all_marker_values

                if active_marker_positions is None or len(active_marker_positions) < 2:
                    QMessageBox.warning(self, "Error", f"Insufficient points for prediction.")
                    if hasattr(self, "protein_location"): del self.protein_location
                    self._reset_live_view_label_custom_handlers(); self.live_view_label.setCursor(Qt.ArrowCursor); return

                # --- RETRIEVE SAVED CALIBRATION ---
                existing_calibration = None
                if self.last_mw_prediction_model and "calibration" in self.last_mw_prediction_model:
                    existing_calibration = self.last_mw_prediction_model["calibration"]

                # --- DIALOG LOOP FOR INTERACTIVE PICKING ---
                dialog = PredictionResultDialog(
                    self, all_marker_positions, all_marker_values, 
                    active_marker_positions, active_marker_values, 
                    refined_y, set_name,
                    initial_calibration=existing_calibration
                )
                dialog.model_changed_in_dialog.connect(self._update_main_model_from_dialog)
                
                final_result = QDialog.Rejected
                
                while True:
                    result = dialog.exec()
                    
                    if result == QDialog.Accepted:
                        final_result = QDialog.Accepted; break
                    elif result == QDialog.Rejected:
                        final_result = QDialog.Rejected; break
                    elif result == 101: # Pick Point 1
                        new_y = self._pick_calibration_point_loop()
                        if new_y is not None:
                            dialog.spin_y1.setValue(new_y)
                            dialog.chk_calib1.setChecked(True)
                    elif result == 102: # Pick Point 2
                        new_y = self._pick_calibration_point_loop()
                        if new_y is not None:
                            dialog.spin_y2.setValue(new_y)
                            dialog.chk_calib2.setChecked(True)

                if final_result == QDialog.Accepted:
                    model_data = dialog.get_final_prediction_model()
                    self.last_mw_prediction_model = model_data 
                    self.last_mw_prediction_min_max_pos = model_data.get("min_max_pos")
                    self.run_predict_MW = True

                    predicted_mw = dialog.get_final_predicted_mw()
                    if predicted_mw > 0: self.last_predicted_mw = predicted_mw
                    
                    if self.avg_glycan_mass <= 0: self.avg_glycan_mass = 0.0
                    self.oligomer_products = []
                    if self.base_protein_mw > 0:
                        for j in range(1, self.num_oligomers_to_model + 1):
                            oligomer_base_mw = self.base_protein_mw * j
                            self.oligomer_products.append(oligomer_base_mw)
                            if self.avg_glycan_mass > 0:
                                for i in range(1, self.num_glycans_to_model + 1):
                                    self.oligomer_products.append(oligomer_base_mw + (i * self.avg_glycan_mass))
                else:
                    self.last_mw_prediction_model = None
                    self.run_predict_MW = False
                    if hasattr(self, "protein_location"): del self.protein_location

                self._reset_live_view_label_custom_handlers()
                self.live_view_label.setCursor(Qt.ArrowCursor)
                self.update_live_view()

            

            def get_protein_location_and_clear_preview(self, event, all_marker_positions, all_marker_values):
                self.get_protein_location(event, all_marker_positions, all_marker_values)
                self.live_view_label.mw_predict_preview_enabled = False
                self.live_view_label.mw_predict_preview_position = None
                self.live_view_label.setMouseTracking(False)
                # Reset the custom handlers after the action is complete
                self._reset_live_view_label_custom_handlers()
                self.live_view_label.update()
                
            def update_mw_predict_preview(self, event):
                if self.live_view_label.mw_predict_preview_enabled:
                    untransformed_label_pos = self.live_view_label.transform_point(event.position())
                    # Snapping for MW preview is optional, but can be nice
                    snapped_label_pos = self.snap_point_to_grid(untransformed_label_pos) 
                    self.live_view_label.mw_predict_preview_position = snapped_label_pos
                    self.live_view_label.update()
                elif hasattr(self.live_view_label, '_original_mouseMoveEvent'): # Pass to original if exists
                     self.live_view_label._original_mouseMoveEvent(event)
                
            def reset_image(self):
                # 1. Save state exactly ONCE at the start
                self.save_state() 

                # 2. LOCK: Prevent sub-functions from saving state or clearing redo stack
                self._is_restoring_state = True 

                try:
                    self.cancel_rectangle_crop_mode()
                    self.crop_rectangle_coords = None
                    self.live_view_label.clear_crop_preview()

                    slider_info = [
                        (getattr(self, 'crop_x_start_slider', None), self.crop_slider_min),
                        (getattr(self, 'crop_x_end_slider', None), self.crop_slider_max),
                        (getattr(self, 'crop_y_start_slider', None), self.crop_slider_min),
                        (getattr(self, 'crop_y_end_slider', None), self.crop_slider_max)
                    ]
                    for slider, default_value in slider_info:
                        if slider:
                            slider.blockSignals(True); slider.setValue(default_value); slider.setEnabled(False); slider.blockSignals(False)

                    # Restore Master Image
                    if hasattr(self, 'original_image') and self.original_image and not self.original_image.isNull():
                        self.image_master = self.original_image.copy()
                        self.image = self.image_master.copy() 
                        self.image_before_padding = None
                        self.image_contrasted = self.image_master.copy()
                        self.image_before_contrast = self.image_master.copy()
                        self.image_padded = False
                        self.contrast_applied = False 
                    else:
                        self.image = None; self.image_master = None; self.original_image = None
                        self.image_before_padding = None; self.image_contrasted = None; self.image_before_contrast = None
                        self.image_padded = False; self.contrast_applied = False
                    
                    # Reset UI elements
                    if hasattr(self, 'blend_slider'): self.blend_slider.setValue(50)
                    self.cancel_interactive_overlay_mode()
                    
                    if hasattr(self, 'show_grid_checkbox_x'): self.show_grid_checkbox_x.setChecked(False)
                    if hasattr(self, 'show_grid_checkbox_y'): self.show_grid_checkbox_y.setChecked(False)
                    if hasattr(self, 'grid_size_input'): self.grid_size_input.setValue(20)

                    self.warped_image=None
                    if hasattr(self, 'left_markers'): self.left_markers.clear()
                    if hasattr(self, 'right_markers'): self.right_markers.clear()
                    if hasattr(self, 'top_markers'): self.top_markers.clear()
                    if hasattr(self, 'custom_markers'): self.custom_markers.clear()
                    if hasattr(self, 'custom_shapes'): self.custom_shapes.clear()
                    self.cancel_drawing_mode()
                    self.clear_predict_molecular_weight()

                    if hasattr(self, 'orientation_slider'): self.orientation_slider.setValue(0)
                    if hasattr(self, 'taper_skew_slider'): self.taper_skew_slider.setValue(0)

                    self.marker_mode = None
                    self.current_left_marker_index = 0; self.current_right_marker_index = 0; self.current_top_label_index = 0
                    self.left_marker_shift_added = 0; self.right_marker_shift_added = 0; self.top_marker_shift_added = 0
                    self.live_view_label.mode = None; self.live_view_label.quad_points = []; self.live_view_label.setCursor(Qt.ArrowCursor)

                    # Reset Preset to Bio-Rad (or default), blocking signals to prevent auto-save
                    try:
                        if hasattr(self, 'combo_box'):
                            biorad_index = self.combo_box.findText("Precision Plus Protein All Blue Prestained (Bio-Rad)")
                            if biorad_index != -1: self.combo_box.setCurrentIndex(biorad_index)
                            # We call this to update the text boxes, but _is_restoring_state=True prevents it from saving state
                            self.on_combobox_changed()
                    except Exception: pass

                    # Reset padding text fields
                    if self.image and not self.image.isNull():
                        if hasattr(self, 'left_padding_input'): self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'right_padding_input'): self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        if hasattr(self, 'top_padding_input'): self.top_padding_input.setText(str(int(self.image.height()*0.15)))

                    self.live_view_label.zoom_level = 1.0; self.live_view_label.pan_offset = QPointF(0, 0)
                    if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
                    if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
                    if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
                    if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
                    
                    # Reset adjustments (will not save state due to lock)
                    self.reset_all_adjustments() 
                    
                    self._update_overlay_slider_ranges()
                    self._update_marker_slider_ranges()
                    self.update_live_view()
                    self._update_levels_histogram()
                    self._update_status_bar()

                finally:
                    # 3. UNLOCK: Re-enable saving for future actions
                    self._is_restoring_state = False
                


        main_window = CombinedSDSApp()

        # --- Close Loading Screen and Show Main Window ---
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # YOU ALREADY HAVE THIS, WHICH IS GOOD!
            loading_dialog = None      # <<--- ADD THIS LINE HERE

        main_window.show()

        # Connect cleanup
        if main_window and app: # Check both exist
            app.aboutToQuit.connect(main_window.cleanup_temp_clipboard_file)

        # --- Start Main Event Loop ---
        if app:
            exit_code = app.exec()
            sys.exit(exit_code)
        else:
            print("FATAL ERROR: QApplication could not be initialized.")
            sys.exit(1)

    except ImportError as e_imp:
        print(f"FATAL ERROR: Missing required library: {e_imp}")
        traceback.print_exc()
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # Good
            loading_dialog = None

        # Use existing app instance for QMessageBox if possible, or create one carefully
        error_app_instance = QApplication.instance()
        temp_app_created = False
        if not error_app_instance:
            try:
                error_app_instance = QApplication([])
                temp_app_created = True
            except RuntimeError: # Catch "already exists" if some other part created it
                error_app_instance = QApplication.instance() # Try to grab it again

        if error_app_instance:
            QMessageBox.critical(None, "Import Error", f"A required library is missing: {e_imp}\nPlease install it and restart the application.")
            if temp_app_created:
                # If we created a temporary app just for the message box,
                # we don't call exec on it.
                pass
        else:
            print("Could not display import error message box: No QApplication.")
        sys.exit(1)

    except Exception as e_startup:
        print(f"FATAL ERROR during application startup: {e_startup}")
        traceback.print_exc()
        if loading_dialog:
            loading_dialog.close()
            loading_dialog.deleteLater() # Good
            loading_dialog = None

        # Log exception using your global handler if sys.excepthook is set
        if sys.excepthook is not sys.__excepthook__: # Check if it's your custom hook
            try:
                sys.excepthook(type(e_startup), e_startup, e_startup.__traceback__)
            except Exception as log_hook_err:
                print(f"ERROR calling custom excepthook: {log_hook_err}")
        else: # Fallback to direct logging if hook wasn't set or failed
            try:
                logging.error("Uncaught startup exception", exc_info=(type(e_startup), e_startup, e_startup.__traceback__))
            except Exception as log_err:
                print(f"Failed to log startup exception directly: {log_err}")

        # Show critical error message box using existing app or a temporary one
        error_app_instance = QApplication.instance()
        temp_app_created = False
        if not error_app_instance:
            try:
                error_app_instance = QApplication([])
                temp_app_created = True
            except RuntimeError:
                error_app_instance = QApplication.instance()

        if error_app_instance:
            QMessageBox.critical(None, "Application Startup Error", f"An unexpected error occurred during startup:\n{e_startup}\n\nCheck error_log.txt for details.")
            if temp_app_created:
                pass
        else:
            print("Could not display startup error message box: No QApplication.")
        sys.exit(1)

    finally:
        # Ensure the loading dialog is closed and cleaned up
        if loading_dialog is not None: # Check if the Python variable still holds an object
            try:
                if loading_dialog.isVisible():
                    loading_dialog.close()
                # Regardless of visibility, if it's not None, schedule for deletion
                # to handle cases where it might exist but wasn't properly closed.
                loading_dialog.deleteLater()
            except RuntimeError:
                # This catches the "Internal C++ object already deleted" specifically
                # print("INFO: loading_dialog was already deleted (RuntimeError in finally).")
                pass # Object is already gone, nothing more to do with it
            except Exception as e_finally_close:
                # Catch any other unexpected error during cleanup
                print(f"ERROR in finally block trying to clean up loading_dialog: {e_finally_close}")
            finally:
                loading_dialog = None # Ensure the Python variable is cleared