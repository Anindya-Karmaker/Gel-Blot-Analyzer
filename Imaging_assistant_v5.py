import logging
import traceback
import sys
import svgwrite
import tempfile
from tempfile import NamedTemporaryFile
import base64
from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
from io import BytesIO
import io
from PyQt5.QtWidgets import (
    QDesktopWidget, QSpacerItem, QTableWidget, QTableWidgetItem, QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy, QMainWindow, QApplication, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit, QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog, QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox
)
from PyQt5.QtGui import QPixmap, QKeySequence, QImage, QPolygonF,QPainter, QColor, QFont, QKeySequence, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF,QUrl
import json
import os
import numpy as np
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font
# import ctypes

from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QSlider,QMenuBar, QMenu, QAction
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.gridspec import GridSpec
from skimage.restoration import rolling_ball 
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import interp1d # Needed for interpolation

# --- Style Sheet Definition ---
STYLE_SHEET = """
QMainWindow {
    background-color: #f0f0f0; /* Light gray background */
}

QTabWidget::pane { /* The tab widget frame */
    border-top: 1px solid #C2C7CB;
    padding: 10px;
    background-color: #ffffff; /* White background for tab content */
}

QTabBar::tab {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
    border: 1px solid #C4C4C3;
    border-bottom-color: #C2C7CB; /* same as the pane border */
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    min-width: 8ex;
    padding: 5px 10px;
    margin-right: 2px; /* space between tabs */
}

QTabBar::tab:selected, QTabBar::tab:hover {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                stop: 0 #fafafa, stop: 0.4 #f4f4f4,
                                stop: 0.5 #e7e7e7, stop: 1.0 #fafafa);
}

QTabBar::tab:selected {
    border-color: #9B9B9B;
    border-bottom-color: #ffffff; /* same as pane background */
    margin-left: -2px; /* make selected tab look connected */
    margin-right: -2px;
}

QTabBar::tab:!selected {
    margin-top: 2px; /* make non-selected tabs look smaller */
}

QPushButton {
    background-color: #e0e0e0;
    border: 1px solid #c0c0c0;
    padding: 5px 10px;
    border-radius: 4px;
    min-height: 20px; /* Ensure minimum height */
}

QPushButton:hover {
    background-color: #d0d0d0;
    border: 1px solid #b0b0b0;
}

QPushButton:pressed {
    background-color: #c0c0c0;
}

QPushButton:disabled {
    background-color: #f5f5f5;
    color: #a0a0a0;
    border: 1px solid #d5d5d5;
}

QGroupBox {
    background-color: #fafafa; /* Slightly off-white */
    border: 1px solid #d0d0d0;
    border-radius: 5px;
    margin-top: 1ex; /* spacing above the title */
    padding: 10px; /* internal padding */
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left; /* position at the top left */
    padding: 0 3px;
    left: 10px; /* position title slightly indented */
    color: #333;
    font-weight: bold;
}

QLabel {
    color: #333; /* Darker text for labels */
    padding-bottom: 2px; /* Small spacing below labels */
}

QLineEdit, QTextEdit, QSpinBox, QComboBox, QFontComboBox {
    border: 1px solid #c0c0c0;
    border-radius: 3px;
    padding: 3px 5px;
    background-color: white;
    min-height: 20px;
}

QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QComboBox:focus, QFontComboBox:focus {
    border: 1px solid #88aaff; /* Highlight focus */
}

QSlider::groove:horizontal {
    border: 1px solid #bbb;
    background: white;
    height: 8px; /* Slider groove height */
    border-radius: 4px;
}

QSlider::handle:horizontal {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #eee, stop:1 #ccc);
    border: 1px solid #777;
    width: 13px; /* Handle width */
    margin: -2px 0; /* handle is placed vertically centered */
    border-radius: 4px;
}

QSlider::add-page:horizontal {
    background: #d0d0d0; /* Color for the part after the handle */
    border: 1px solid #bbb;
    border-radius: 4px;
}

QSlider::sub-page:horizontal {
    background: #88aaff; /* Color for the part before the handle */
    border: 1px solid #bbb;
    border-radius: 4px;
}

QTableWidget {
    border: 1px solid #c0c0c0;
    gridline-color: #d0d0d0; /* Lighter grid lines */
    background-color: white;
}

QHeaderView::section {
    background-color: #e8e8e8; /* Header background */
    padding: 4px;
    border: 1px solid #c0c0c0;
    font-weight: bold;
}

QTableWidgetItem {
    padding: 3px;
}

QScrollArea {
    border: none; /* Remove border from scroll area itself */
}
/* Make the LiveViewLabel border slightly softer */
#LiveViewLabel {
    border: 1px solid #c0c0c0;
}
"""

# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
logging.basicConfig(
    filename="error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    # Log the exception
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))

    # Display a QMessageBox with the error details
    error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}"
    QMessageBox.critical(
        None,  # No parent window
        "Unexpected Error",  # Title of the message box
        error_message,  # Error message to display
        QMessageBox.Ok  # Button to close the dialog
    )
    

# # Set the custom exception handler
sys.excepthook = log_exception

class TableWindow(QDialog):
    def __init__(self, peak_areas, standard_dictionary, standard, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Table Export")
        self.setGeometry(100, 100, 600, 400)

        # Create a scroll area
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)  # Allow the table to resize within the scroll area

        # Create a table widget
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)  # Enable row selection
        self.table.setSelectionMode(QTableWidget.ContiguousSelection)  # Allow copying multiple rows

        # Populate the table with peak areas and percentages
        self.populate_table(peak_areas, standard_dictionary, standard)

        # Set the table as the widget for the scroll area
        self.scroll_area.setWidget(self.table)

        # Add an "Export to Excel" button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)  # Add the scroll area to the layout
        layout.addWidget(self.export_button)  # Add the export button
        self.setLayout(layout)

    def populate_table(self, peak_areas, standard_dictionary, standard):
        # Calculate the total sum of peak areas
        total_area = sum(peak_areas)

        # Set the number of rows in the table
        self.table.setRowCount(len(peak_areas))

        # Populate the table with band labels, peak areas, and percentages
        for row, area in enumerate(peak_areas):
            # Bang label (e.g., "Band 1", "Band 2", etc.)
            band_label = f"Band {row + 1}"
            self.table.setItem(row, 0, QTableWidgetItem(band_label))

            # Peak area (rounded to 3 decimal places)
            peak_area_rounded = round(area, 3)
            self.table.setItem(row, 1, QTableWidgetItem(str(peak_area_rounded)))

            # Percentage (rounded to 2 decimal places)
            if total_area != 0:  # Avoid division by zero
                percentage = (area / total_area) * 100
                percentage_rounded = round(percentage, 2)
            else:
                percentage_rounded = 0.0
            self.table.setItem(row, 2, QTableWidgetItem(f"{percentage_rounded:.2f}%"))

            # Quantity (if standard is True)
            if standard:
                std_peak_area = list(standard_dictionary.values())
                known_quantities = list(standard_dictionary.keys())
                coefficients = np.polyfit(std_peak_area, known_quantities, 1)
                unknown_quantity = np.polyval(coefficients, area)
                self.table.setItem(row, 3, QTableWidgetItem(f"{unknown_quantity:.2f}"))

        # Enable copying data from the table
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)  # Make the table read-only
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)  # Enable right-click context menu
        self.table.customContextMenuRequested.connect(self.copy_table_data)

    def copy_table_data(self):
        """Copy selected table data to the clipboard."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # Get the selected rows and columns
        rows = set(item.row() for item in selected_items)
        cols = set(item.column() for item in selected_items)

        # Prepare the data for copying
        data = []
        for row in sorted(rows):
            row_data = []
            for col in sorted(cols):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append("\t".join(row_data))

        # Copy the data to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(data))

    def export_to_excel(self):
        """Export the table data to an Excel file."""
        # Prompt the user to select a save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
        )
        if not file_path:
            return

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Peak Areas"

        # Write the table headers to the Excel sheet
        headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
            worksheet.cell(row=1, column=col).font = Font(bold=True)  # Make headers bold

        # Write the table data to the Excel sheet
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                worksheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

        # Save the Excel file
        workbook.save(file_path)
        QMessageBox.information(self, "Success", f"Table data exported to {file_path}")
                

class PeakAreaDialog(QDialog):
    """
    Interactive dialog to adjust peak regions and calculate peak areas.
    Settings persistence (prominence, height, distance, spread, methods)
    is managed by the parent application via passed variables.
    """
    # No SETTINGS_FILE needed as persistence is handled by the parent

    def __init__(self, cropped_image, current_settings, persist_checked, parent=None):
        """
        Initializes the dialog.

        Args:
            cropped_image (PIL.Image): The cropped image (should be grayscale).
            current_settings (dict): Dictionary containing the last used settings
                                      from the parent application.
            persist_checked (bool): The last state of the "Persist Settings"
                                     checkbox from the parent.
            parent (QWidget, optional): The parent widget. Defaults to None.
        """
        super().__init__(parent)
        self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
        self.setGeometry(100, 100, 1100, 850) # Adjusted size for better layout

        # Ensure input is a PIL Image and store it
        if not isinstance(cropped_image, Image.Image):
             # Handle error or try conversion if applicable
             raise TypeError("cropped_image must be a PIL Image object")
        self.cropped_image = cropped_image
        # Convert to grayscale numpy array for processing
        self.intensity_array = np.array(self.cropped_image.convert("L"), dtype=np.float64)

        self.profile = None
        self.background = None # Stores the rolling ball background result

        # --- Apply passed settings or use defaults ---
        # Get settings from the dictionary passed by the parent
        self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
        self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
        self.peak_distance = current_settings.get('peak_distance', 30)
        self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.02)
        self.peak_spread_pixels = current_settings.get('peak_spread_pixels', 10)
        self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
        self.area_subtraction_method = current_settings.get('area_subtraction_method', "Valley-to-Valley")
        # --- End Apply passed settings ---

        # Internal state for peak data
        self.peaks = np.array([])           # Indices of detected peaks
        self.initial_peak_regions = []      # Regions calculated directly from detection (before spread)
        self.peak_regions = []              # Final regions after spread/slider adjustments
        self.peak_areas_rolling_ball = []   # Calculated areas (method 1)
        self.peak_areas_straight_line = []  # Calculated areas (method 2)
        self.peak_areas_valley = []         # Calculated areas (method 3)
        self.peak_sliders = []              # References to individual QSlider widgets

        # --- Stored state for parent access upon closing ---
        self._final_settings = {} # Will store the settings when OK is clicked
        self._persist_enabled_on_exit = persist_checked # Store initial state from parent

        # --- Build UI ---
        self._setup_ui(persist_checked) # Pass initial checkbox state to UI setup

        # --- Initial Setup ---
        self.regenerate_profile_and_detect() # Generate profile and detect peaks

    def _setup_ui(self, persist_checked_initial):
        """Creates and arranges the UI elements."""
        # --- Main Layout ---
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # --- Matplotlib Plot Canvas ---
        self.fig = plt.figure(figsize=(10, 5)) # Adjusted figure size
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.canvas, stretch=3) # Give canvas more vertical stretch

        # --- Controls Layout (Horizontal Box for side-by-side groups) ---
        controls_hbox = QHBoxLayout()
        controls_hbox.setSpacing(15)

        # --- Left Controls Column (Global & Detection) ---
        left_controls_vbox = QVBoxLayout()

        # Group 1: Global Settings
        global_settings_group = QGroupBox("Global Settings")
        global_settings_layout = QGridLayout(global_settings_group)
        global_settings_layout.setSpacing(8) # Spacing within the grid

        # Band Estimation Method Dropdown
        self.band_estimation_combobox = QComboBox()
        self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
        self.band_estimation_combobox.setCurrentText(self.band_estimation_method) # Initialize with current value
        self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
        global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
        global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1)

        # Area Subtraction Method Dropdown
        self.method_combobox = QComboBox()
        self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
        self.method_combobox.setCurrentText(self.area_subtraction_method) # Initialize with current value
        self.method_combobox.currentIndexChanged.connect(self.update_plot) # Update plot when method changes
        global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
        global_settings_layout.addWidget(self.method_combobox, 1, 1)

        # Rolling Ball Radius Slider
        self.rolling_ball_label = QLabel(f"Rolling Ball Radius ({self.rolling_ball_radius:.0f})")
        self.rolling_ball_slider = QSlider(Qt.Horizontal)
        self.rolling_ball_slider.setRange(1, 500) # Adjust range as needed
        self.rolling_ball_slider.setValue(int(self.rolling_ball_radius)) # Initialize with current value
        self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_label: lbl.setText(f"Rolling Ball Radius ({val})")) # Update label
        self.rolling_ball_slider.valueChanged.connect(self.update_plot) # Update plot when radius changes
        global_settings_layout.addWidget(self.rolling_ball_label, 2, 0)
        global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1)

        left_controls_vbox.addWidget(global_settings_group)

        # Group 2: Peak Detection Parameters
        peak_detect_group = QGroupBox("Peak Detection Parameters")
        peak_detect_layout = QGridLayout(peak_detect_group)
        peak_detect_layout.setSpacing(8)

        # Manual Peak Number Input & Update Button (Row 0)
        self.peak_number_label = QLabel("Detected Peaks:")
        self.peak_number_input = QLineEdit()
        self.peak_number_input.setPlaceholderText("#")
        self.peak_number_input.setMaximumWidth(60) # Limit width for aesthetics
        self.update_peak_number_button = QPushButton("Set")
        self.update_peak_number_button.setToolTip("Manually override the number of peaks detected.")
        self.update_peak_number_button.clicked.connect(self.manual_peak_number_update)
        peak_detect_layout.addWidget(self.peak_number_label, 0, 0)
        peak_detect_layout.addWidget(self.peak_number_input, 0, 1)
        peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)

        # Peak Prominence Slider (Row 1)
        self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        self.peak_prominence_slider = QSlider(Qt.Horizontal)
        self.peak_prominence_slider.setRange(0, 50) # Range 0.0 to 0.5
        self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100)) # Initialize
        self.peak_prominence_slider.valueChanged.connect(self.detect_peaks) # Trigger peak detection
        peak_detect_layout.addWidget(self.peak_prominence_slider_label, 1, 0)
        peak_detect_layout.addWidget(self.peak_prominence_slider, 1, 1, 1, 2) # Span 2 columns

        # Peak Height Slider (Row 2)
        self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f}) %")
        self.peak_height_slider = QSlider(Qt.Horizontal)
        self.peak_height_slider.setRange(0, 100) # Range 0% to 100%
        self.peak_height_slider.setValue(int(self.peak_height_factor * 100)) # Initialize
        self.peak_height_slider.valueChanged.connect(self.detect_peaks) # Trigger peak detection
        peak_detect_layout.addWidget(self.peak_height_slider_label, 2, 0)
        peak_detect_layout.addWidget(self.peak_height_slider, 2, 1, 1, 2) # Span 2 columns

        # Peak Distance Slider (Row 3)
        self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
        self.peak_distance_slider = QSlider(Qt.Horizontal)
        self.peak_distance_slider.setRange(1, 200) # Adjust range as needed
        self.peak_distance_slider.setValue(self.peak_distance) # Initialize
        self.peak_distance_slider.valueChanged.connect(self.detect_peaks) # Trigger peak detection
        peak_detect_layout.addWidget(self.peak_distance_slider_label, 3, 0)
        peak_detect_layout.addWidget(self.peak_distance_slider, 3, 1, 1, 2) # Span 2 columns

        left_controls_vbox.addWidget(peak_detect_group)
        left_controls_vbox.addStretch(1) # Push groups up in the left column

        controls_hbox.addLayout(left_controls_vbox, stretch=1) # Add left column to HBox

        # --- Right Controls Column (Peak Region Adjustments) ---
        right_controls_vbox = QVBoxLayout()

        # Group 3: Peak Region Spread (Global Adjustment)
        peak_spread_group = QGroupBox("Peak Region Adjustments")
        peak_spread_layout = QGridLayout(peak_spread_group)
        peak_spread_layout.setSpacing(8)

        # Peak Spread Slider
        self.peak_spread_label = QLabel(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        self.peak_spread_slider = QSlider(Qt.Horizontal)
        self.peak_spread_slider.setRange(0, 100) # Spread 0 to 100 pixels around center
        self.peak_spread_slider.setValue(self.peak_spread_pixels) # Initialize
        self.peak_spread_slider.setToolTip(
            "Adjusts the width of all detected peak regions simultaneously.\n"
            "Regions expand/contract around the initial detected peak center."
        )
        self.peak_spread_slider.valueChanged.connect(self.apply_peak_spread) # Connect to handler
        # Update label text when slider changes
        self.peak_spread_slider.valueChanged.connect(
            lambda value, lbl=self.peak_spread_label: lbl.setText(f"Peak Spread (+/- {value} px)")
        )
        peak_spread_layout.addWidget(self.peak_spread_label, 0, 0)
        peak_spread_layout.addWidget(self.peak_spread_slider, 0, 1)

        right_controls_vbox.addWidget(peak_spread_group)

        # Scroll Area for Individual Peak Sliders
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(250) # Ensure scroll area has reasonable initial height
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding) # Allow vertical expansion

        self.container = QWidget() # Container widget for the layout inside scroll area
        self.peak_sliders_layout = QVBoxLayout(self.container) # Layout for individual peak groups
        self.peak_sliders_layout.setSpacing(10) # Space between peak groups

        scroll_area.setWidget(self.container) # Put the container in the scroll area
        right_controls_vbox.addWidget(scroll_area, stretch=1) # Give scroll area vertical stretch

        controls_hbox.addLayout(right_controls_vbox, stretch=2) # Add right column, give more horizontal space

        main_layout.addLayout(controls_hbox) # Add the HBox containing controls to main layout

        # --- Bottom Button Layout ---
        bottom_button_layout = QHBoxLayout()

        # Persist Settings Checkbox
        self.persist_settings_checkbox = QCheckBox("Persist Settings")
        self.persist_settings_checkbox.setChecked(persist_checked_initial) # Use initial state from parent
        self.persist_settings_checkbox.setToolTip("Save current detection parameters for the next time this dialog opens during this session.")
        bottom_button_layout.addWidget(self.persist_settings_checkbox)

        bottom_button_layout.addStretch(1) # Spacer pushes OK button to the right
        self.ok_button = QPushButton("OK")
        self.ok_button.setMinimumWidth(100)
        self.ok_button.setDefault(True) # Make OK the default button (responds to Enter)
        self.ok_button.clicked.connect(self.accept_and_close) # Connect to custom accept method
        bottom_button_layout.addWidget(self.ok_button)

        main_layout.addLayout(bottom_button_layout) # Add bottom buttons

    # --- Methods for Parent Interaction ---

    def accept_and_close(self):
        """Stores the current settings and persist state before accepting."""
        # Capture the current state of all relevant controls into the dictionary
        self._final_settings = {
            'rolling_ball_radius': self.rolling_ball_slider.value(),
            'peak_height_factor': self.peak_height_slider.value() / 100.0,
            'peak_distance': self.peak_distance_slider.value(),
            'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
            'peak_spread_pixels': self.peak_spread_slider.value(),
            'band_estimation_method': self.band_estimation_combobox.currentText(),
            'area_subtraction_method': self.method_combobox.currentText()
        }
        # Capture the state of the checkbox when OK is clicked
        self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
        self.accept() # Standard Qt accept method to close the dialog

    def get_current_settings(self):
        """Returns the settings dictionary captured when OK was clicked."""
        return self._final_settings

    def should_persist_settings(self):
        """Returns the state of the 'Persist Settings' checkbox when OK was clicked."""
        return self._persist_enabled_on_exit

    # --- Core Logic Methods ---

    def regenerate_profile_and_detect(self):
        """Calculates the profile based on combobox selection and detects peaks."""
        # Update instance attributes from comboboxes *before* using them
        self.band_estimation_method = self.band_estimation_combobox.currentText()
        self.area_subtraction_method = self.method_combobox.currentText()

        # Calculate profile based on selected technique
        if self.band_estimation_method == "Mean":
            self.profile = np.mean(self.intensity_array, axis=1)
        elif self.band_estimation_method.startswith("Percentile"):
            try:
                percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                percent = max(0, min(100, percent))
                self.profile = np.percentile(self.intensity_array, percent, axis=1)
            except (ValueError, IndexError):
                self.profile = np.percentile(self.intensity_array, 5, axis=1) # Fallback
                print("Warning: Invalid percentile format, defaulting to 5%")
        else:
            self.profile = np.mean(self.intensity_array, axis=1) # Default to Mean

        # Check if profile calculation resulted in NaN or Inf values (can happen with empty images/regions)
        if not np.all(np.isfinite(self.profile)):
            print("Warning: Profile contains non-finite values. Setting to zero.")
            self.profile = np.zeros(self.intensity_array.shape[0]) # Create zero profile matching height


        # Normalize and Invert profile
        prof_min, prof_max = np.min(self.profile), np.max(self.profile)
        if prof_max > prof_min:
            # Added check for non-finite min/max
            if np.isfinite(prof_min) and np.isfinite(prof_max):
                 self.profile = (self.profile - prof_min) / (prof_max - prof_min) * 255.0
            else: # Handle case where min/max might be NaN/Inf if profile was bad
                 self.profile = np.zeros_like(self.profile)
        else:
             self.profile = np.zeros_like(self.profile) # Handle flat profile
        self.profile = 255.0 - self.profile # Invert (dark bands = high peaks)

        # Apply Gaussian smoothing
        try:
            # Ensure profile length is sufficient for smoothing
            if len(self.profile) > 6: # sigma*3*2 rule of thumb for Gaussian
                self.profile = gaussian_filter1d(self.profile, sigma=2)
            elif len(self.profile) > 0:
                print("Warning: Profile too short for significant Gaussian smoothing.")
                # Optionally apply minimal smoothing or skip
        except Exception as smooth_err:
            print(f"Error during Gaussian smoothing: {smooth_err}")
            # Continue without smoothing if it fails

        # Detect peaks using the new profile
        self.detect_peaks()

    def detect_peaks(self):
        """Detect peaks, set initial regions, and apply current spread."""
        if self.profile is None or len(self.profile) == 0:
            print("Profile not generated yet for peak detection.")
            self.peaks = np.array([])
            self.initial_peak_regions = []
            self.peak_regions = []
            self.peak_number_input.setText("0")
            self.update_sliders()
            self.update_plot()
            return

        # Update parameters from sliders INTO the instance attributes
        self.peak_height_factor = self.peak_height_slider.value() / 100.0
        self.peak_distance = self.peak_distance_slider.value()
        self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
        self.peak_spread_pixels = self.peak_spread_slider.value()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # Update UI Labels to reflect current instance attributes
        self.peak_height_slider_label.setText(f"Min Height ({self.peak_height_factor:.2f}) %")
        self.peak_distance_slider_label.setText(f"Min Distance ({self.peak_distance}) px")
        self.peak_prominence_slider_label.setText(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        self.peak_spread_label.setText(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        self.rolling_ball_label.setText(f"Rolling Ball Radius ({self.rolling_ball_radius:.0f})")

        # Calculate thresholds using the UPDATED instance attributes
        profile_range = np.ptp(self.profile)
        if profile_range == 0: profile_range = 1 # Avoid division by zero
        min_height = np.min(self.profile) + profile_range * self.peak_height_factor
        min_prominence = profile_range * self.peak_prominence_factor

        try:
            # Detect peaks using UPDATED instance attributes
            peaks_indices, properties = find_peaks(
                self.profile,
                height=min_height,
                prominence=max(1, min_prominence), # Ensure prominence is at least 1
                distance=self.peak_distance,
                width=1,       # Request width properties
                rel_height=0.5 # Measure width at half prominence
            )

            # Extract width properties
            left_ips = properties.get('left_ips', [])
            right_ips = properties.get('right_ips', [])

            self.peaks = peaks_indices
            self.initial_peak_regions = [] # Reset INITIAL regions

            # Set INITIAL peak regions based on width properties or fallback
            if len(left_ips) == len(self.peaks) and len(right_ips) == len(self.peaks):
                for i, peak_idx in enumerate(self.peaks):
                    start = int(np.floor(left_ips[i]))
                    end = int(np.ceil(right_ips[i]))
                    start = max(0, start)
                    end = min(len(self.profile) - 1, end)
                    if start >= end: # Handle zero/negative width from properties
                        fallback_width = max(1, self.peak_distance // 4) # Use instance distance
                        start = max(0, peak_idx - fallback_width)
                        end = min(len(self.profile) - 1, peak_idx + fallback_width)
                        if start >= end: end = min(len(self.profile) - 1, start + 1)
                    self.initial_peak_regions.append((start, end))
            else:
                # Fallback if width properties aren't available/reliable
                print("Warning: Width properties inconsistent, using distance-based fallback.")
                for i, peak_idx in enumerate(self.peaks):
                    width_estimate = self.peak_distance // 2 # Use instance distance
                    start = max(0, peak_idx - width_estimate)
                    end = min(len(self.profile) - 1, peak_idx + width_estimate)
                    if start >= end: # Minimal fallback
                         start = max(0, peak_idx - 2)
                         end = min(len(self.profile) - 1, peak_idx + 2)
                         if start >= end: end = min(len(self.profile) - 1, start + 1)
                    self.initial_peak_regions.append((start, end))

        except Exception as e:
            print(f"Error during peak detection: {e}")
            QMessageBox.warning(self, "Peak Detection Error", f"An error occurred during peak detection:\n{e}")
            self.peaks = np.array([])
            self.initial_peak_regions = []

        # Update the peak number input field only if not focused or empty
        if not self.peak_number_input.hasFocus() or self.peak_number_input.text() == "":
             self.peak_number_input.setText(str(len(self.peaks)))

        # Apply the current spread setting using the updated instance attribute
        # This call is crucial to sync regions after detection
        self.apply_peak_spread(self.peak_spread_pixels) # Calls update_sliders & update_plot

    def apply_peak_spread(self, spread_value):
        """Applies the spread value to the initial peak regions."""
        self.peak_spread_pixels = spread_value # Update internal state for persistence
        self.peak_regions = [] # Clear the final regions list to rebuild

        if self.profile is None or len(self.profile) == 0:
            self.update_sliders() # Ensure sliders are cleared if no profile
            self.update_plot()
            return

        profile_len = len(self.profile)

        # Use the initial regions list as the base for applying spread
        num_initial = min(len(self.peaks), len(self.initial_peak_regions))
        if len(self.peaks) != len(self.initial_peak_regions):
             print(f"Warning: Mismatch between peaks ({len(self.peaks)}) and initial regions "
                   f"({len(self.initial_peak_regions)}). Applying spread to {num_initial}.")

        for i in range(num_initial):
            # Use the detected peak index as the center for spread calculation
            # This is more robust if initial_peak_regions were just placeholders
            peak_idx = self.peaks[i]
            center = peak_idx

            # Calculate new boundaries based on peak center and spread
            new_start = center - self.peak_spread_pixels
            new_end = center + self.peak_spread_pixels

            # Clamp to profile boundaries and ensure integer values
            new_start = max(0, int(new_start))
            new_end = min(profile_len - 1, int(new_end))

            # Ensure start <= end, making it a single point if spread is excessive
            if new_start > new_end:
                new_start = new_end

            self.peak_regions.append((new_start, new_end))

        # Adjust final regions list if manual peak# change occurred BEFORE spread application
        if len(self.peak_regions) != len(self.peaks):
            print(f"Adjusting final regions list length from {len(self.peak_regions)} to {len(self.peaks)} after manual change.")
            # This assumes peaks list is the source of truth for length
            self.peak_regions = self.peak_regions[:len(self.peaks)]
            # If peaks were added manually, need corresponding placeholder regions added here?
            # Manual_peak_number_update should handle adding initial regions first.

        # Recreate sliders based on NEW self.peak_regions & redraw plot
        self.update_sliders()
        self.update_plot()

    def manual_peak_number_update(self):
        """Handles manual changes to the number of peaks."""
        if self.profile is None or len(self.profile) == 0:
            QMessageBox.warning(self, "Error", "Profile must be generated first.")
            return

        try:
            num_peaks_manual = int(self.peak_number_input.text())
            if num_peaks_manual < 0:
                self.peak_number_input.setText(str(len(self.peaks))) # Revert display
                QMessageBox.warning(self, "Input Error", "Number of peaks cannot be negative.")
                return

            current_num_peaks = len(self.peaks)
            if num_peaks_manual == current_num_peaks:
                 return # No change needed

            if num_peaks_manual == 0:
                self.peaks = np.array([])
                self.initial_peak_regions = []
                self.peak_regions = [] # Clear final regions too
            elif num_peaks_manual < current_num_peaks:
                # Truncate peaks and initial regions
                self.peaks = self.peaks[:num_peaks_manual]
                self.initial_peak_regions = self.initial_peak_regions[:num_peaks_manual]
                # Final regions will be adjusted by apply_peak_spread later
            elif num_peaks_manual > current_num_peaks:
                # Add dummy peaks and corresponding initial regions
                num_to_add = num_peaks_manual - current_num_peaks
                profile_center = len(self.profile) // 2
                peaks_list = self.peaks.tolist()
                initial_regions_list = list(self.initial_peak_regions) # Work with list copy

                for i in range(num_to_add):
                    # Add a dummy peak (e.g., near center)
                    new_peak_pos = profile_center + np.random.randint(-50, 50) # Simple placement
                    new_peak_pos = max(0, min(len(self.profile) - 1, new_peak_pos))
                    peaks_list.append(new_peak_pos)

                    # Add a placeholder *initial* region for the new peak
                    # The size will be determined by the spread slider later
                    # Use a minimal default width for the placeholder initial region
                    placeholder_width = 5 # e.g., +/- 5 pixels
                    initial_start = max(0, new_peak_pos - placeholder_width)
                    initial_end = min(len(self.profile) - 1, new_peak_pos + placeholder_width)
                    if initial_start >= initial_end: initial_end = min(len(self.profile)-1, initial_start + 1)
                    initial_regions_list.append((initial_start, initial_end))

                # Sort peaks and initial regions together after adding dummies
                if peaks_list:
                    # Ensure initial_regions_list has the same length as peaks_list
                    if len(initial_regions_list) != len(peaks_list):
                        # This case should ideally not happen if logic is correct, but handle defensively
                        print("Warning: Peak and initial region lists length mismatch during manual add.")
                        min_len = min(len(peaks_list), len(initial_regions_list))
                        peaks_list = peaks_list[:min_len]
                        initial_regions_list = initial_regions_list[:min_len]

                    combined = sorted(zip(peaks_list, initial_regions_list), key=lambda pair: pair[0])
                    sorted_peaks, sorted_initial_regions = zip(*combined)
                    self.peaks = np.array(sorted_peaks)
                    self.initial_peak_regions = list(sorted_initial_regions)
                else:
                    self.peaks = np.array([])
                    self.initial_peak_regions = []

            # Crucially, after modifying peaks and initial_regions,
            # re-apply the current spread setting to update self.peak_regions
            self.apply_peak_spread(self.peak_spread_slider.value()) # This calls update_sliders/plot

        except ValueError:
            self.peak_number_input.setText(str(len(self.peaks))) # Revert display
            QMessageBox.warning(self, "Input Error", "Please enter a valid integer for the number of peaks.")
        except Exception as e:
             print(f"Error in manual peak update: {e}")
             QMessageBox.critical(self, "Error", f"An error occurred during manual peak update:\n{e}")
             self.peak_number_input.setText(str(len(self.peaks))) # Revert on error

    def update_sliders(self):
        """Update the sliders based on the current self.peak_regions (after spread)."""
        # Clear existing sliders widgets first
        for i in reversed(range(self.peak_sliders_layout.count())):
            item = self.peak_sliders_layout.itemAt(i)
            widget = item.widget()
            if widget:
                # Properly remove and delete the widget
                self.peak_sliders_layout.removeWidget(widget)
                widget.deleteLater()
            else:
                # Handle layouts or spacers if necessary
                layout_item = self.peak_sliders_layout.takeAt(i)
                if layout_item:
                    # Check if it's a spacer item
                    if isinstance(layout_item, QSpacerItem):
                        pass # Spacers don't need explicit deletion
                    # Recursively clear layouts if needed (shouldn't be necessary here)
                    # elif layout_item.layout() is not None: clear_layout(layout_item.layout())
                    del layout_item # Delete the layout item itself

        self.peak_sliders.clear() # Clear the list storing slider references

        if self.profile is None or len(self.profile) == 0: return

        profile_len = len(self.profile)

        # Create sliders based on self.peak_regions
        num_items = len(self.peak_regions)
        # Ensure peaks list is consistent for labels (can happen after manual changes)
        if len(self.peaks) != num_items:
            print(f"Warning: Peak count ({len(self.peaks)}) differs from region count ({num_items}) for slider labels.")


        for i in range(num_items):
            try:
                # Get region boundaries from self.peak_regions
                start_val, end_val = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                # Get peak index for label, handle potential length mismatch safely
                peak_index = int(self.peaks[i]) if i < len(self.peaks) else -1 # Use -1 if no matching peak index
            except (ValueError, TypeError, IndexError) as e:
                print(f"Warning: Invalid data for peak/region at index {i} ({e}). Skipping slider.")
                continue # Skip this iteration if data is corrupt

            peak_group = QGroupBox(f"Peak {i + 1} (Index: {peak_index if peak_index != -1 else 'N/A'})")
            peak_layout = QGridLayout(peak_group)

            # Start slider
            start_slider = QSlider(Qt.Horizontal)
            start_slider.setRange(0, profile_len - 1)
            # Clamp initial value just in case & set slider position
            start_val_clamped = max(0, min(profile_len - 1, start_val))
            start_slider.setValue(start_val_clamped)
            start_slider.valueChanged.connect(self.update_plot) # Update plot when changed by user
            start_label = QLabel(f"Start: {start_slider.value()}") # Show actual slider value
            # Connect to helper that updates self.peak_regions AND the label
            start_slider.valueChanged.connect(lambda val, lbl=start_label, idx=i: self._update_region_from_slider(idx, 'start', val, lbl))
            peak_layout.addWidget(start_label, 0, 0)
            peak_layout.addWidget(start_slider, 0, 1)

            # End slider
            end_slider = QSlider(Qt.Horizontal)
            end_slider.setRange(0, profile_len - 1)
             # Clamp initial value just in case & set slider position
            end_val_clamped = max(0, min(profile_len - 1, end_val))
            # Ensure end slider value is >= start slider value initially
            if end_val_clamped < start_val_clamped: end_val_clamped = start_val_clamped
            end_slider.setValue(end_val_clamped)
            end_slider.valueChanged.connect(self.update_plot) # Update plot when changed by user
            end_label = QLabel(f"End: {end_slider.value()}") # Show actual slider value
            # Connect to helper that updates self.peak_regions AND the label
            end_slider.valueChanged.connect(lambda val, lbl=end_label, idx=i: self._update_region_from_slider(idx, 'end', val, lbl))
            peak_layout.addWidget(end_label, 1, 0)
            peak_layout.addWidget(end_slider, 1, 1)

            self.peak_sliders_layout.addWidget(peak_group)
            self.peak_sliders.append((start_slider, end_slider)) # Store tuple of references

        # Add stretch at the end ONLY if there are sliders, to push them up
        if num_items > 0:
            self.peak_sliders_layout.addStretch(1)

    def _update_region_from_slider(self, index, boundary_type, value, label_widget):
        """Helper to update self.peak_regions when an individual slider is moved."""
        # Update self.peak_regions list first
        if 0 <= index < len(self.peak_regions):
            current_start, current_end = self.peak_regions[index]
            if boundary_type == 'start':
                # Ensure start <= end after update
                new_start = min(value, current_end)
                self.peak_regions[index] = (new_start, current_end)
                label_widget.setText(f"Start: {new_start}") # Update label
                # Update the actual slider value if clamping occurred
                # Use object reference stored in self.peak_sliders
                start_slider_widget, _ = self.peak_sliders[index]
                if start_slider_widget.value() != new_start:
                     # Temporarily block signals to prevent infinite loop
                    start_slider_widget.blockSignals(True)
                    start_slider_widget.setValue(new_start)
                    start_slider_widget.blockSignals(False)
            elif boundary_type == 'end':
                # Ensure start <= end after update
                new_end = max(value, current_start)
                self.peak_regions[index] = (current_start, new_end)
                label_widget.setText(f"End: {new_end}") # Update label
                # Update the actual slider value if clamping occurred
                _, end_slider_widget = self.peak_sliders[index]
                if end_slider_widget.value() != new_end:
                    end_slider_widget.blockSignals(True)
                    end_slider_widget.setValue(new_end)
                    end_slider_widget.blockSignals(False)
        # No need to call update_plot here, it's triggered by the original slider signal

    def update_plot(self):
        """Update the plot based on current settings and peak regions."""
        # --- Guard Clauses ---
        if self.canvas is None: return
        if self.profile is None or len(self.profile) == 0 :
            # Clear plot if no profile
            try:
                self.fig.clf() # Clear figure
                self.ax = self.fig.add_subplot(111) # Add basic axes
                self.ax.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax.transAxes)
                self.canvas.draw_idle()
            except Exception as e:
                print(f"Error clearing plot: {e}")
            return

        # --- Update Instance Attributes from UI (ensure consistency) ---
        # It's good practice to sync attributes before plotting, although signals should handle most cases
        self.method = self.method_combobox.currentText()
        self.rolling_ball_radius = self.rolling_ball_slider.value()
        # No need to update detection factors here unless explicitly requested by user action

        # --- Calculate Rolling Ball Background ---
        try:
            profile_float = self.profile.astype(np.float64)
            # Added check for profile length vs radius
            safe_radius = min(self.rolling_ball_radius, len(profile_float) // 2 -1) # Prevent radius > half length
            if safe_radius < 1: safe_radius = 1 # Ensure radius is at least 1
            if safe_radius != self.rolling_ball_radius:
                 print(f"Adjusted rolling ball radius to {safe_radius} due to profile length.")

            if len(profile_float) > 1 : # rolling_ball needs > 1 point
                background_fwd = rolling_ball(profile_float, radius=safe_radius)
                self.background = np.minimum(background_fwd, profile_float)
            else:
                self.background = profile_float.copy() # Background is just the profile if too short
        except ImportError:
             print("Scikit-image not found. Rolling ball method unavailable.")
             QMessageBox.warning(self, "Dependency Error", "Scikit-image is required for the Rolling Ball method.")
             # Fallback or disable method if needed
             self.background = np.full_like(self.profile, np.min(self.profile))
        except ValueError as ve: # Catch specific skimage value errors (e.g., radius)
            print(f"Error calculating rolling ball background: {ve}. Using min profile value as fallback.")
            self.background = np.full_like(self.profile, np.min(self.profile))
        except Exception as e:
            print(f"Unexpected error calculating rolling ball background: {e}")
            self.background = np.full_like(self.profile, np.min(self.profile))

        # --- Clear Figure and Setup Subplots using GridSpec ---
        self.fig.clf()
        gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig)
        self.ax = self.fig.add_subplot(gs[0])
        ax_image = self.fig.add_subplot(gs[1], sharex=self.ax) # Share X axis

        # --- Calculate Profile Range for text/limit calculations ---
        profile_range = np.ptp(self.profile) if np.ptp(self.profile) > 0 else 1

        # --- Plot Profile and Detected Peak Markers ---
        self.ax.plot(self.profile, label="Smoothed Profile", color="black", linestyle="-", linewidth=1.2)
        if len(self.peaks) > 0:
             # Ensure indices are valid before accessing profile data
             valid_peaks_indices = self.peaks[(self.peaks >= 0) & (self.peaks < len(self.profile))]
             if len(valid_peaks_indices) > 0:
                 peak_y_values = self.profile[valid_peaks_indices]
                 self.ax.scatter(valid_peaks_indices, peak_y_values,
                                 color="red", marker='x', s=50, label="Detected Peaks", zorder=5)

        # --- Process Each Peak Region (using self.peak_regions) ---
        self.peak_areas_rolling_ball.clear()
        self.peak_areas_straight_line.clear()
        self.peak_areas_valley.clear()

        num_items_to_plot = len(self.peak_regions)
        # Check consistency with peak_sliders list (should match)
        if len(self.peak_sliders) != num_items_to_plot:
             print(f"Plot Warning: Slider count ({len(self.peak_sliders)}) differs from region count ({num_items_to_plot}).")
             num_items_to_plot = min(len(self.peak_sliders), num_items_to_plot) # Plot the minimum

        max_text_y_position = np.min(self.profile) # Initialize for y-limit calculation

        for i in range(num_items_to_plot):
            # Get region boundaries from the synchronized list
            start, end = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])

            # --- Skip if region is invalid (start >= end) ---
            if start >= end:
                # Append 0.0 areas for skipped peak - crucial for list consistency
                self.peak_areas_rolling_ball.append(0.0)
                self.peak_areas_straight_line.append(0.0)
                self.peak_areas_valley.append(0.0)
                # Optionally draw markers to show the invalid point
                self.ax.axvline(start, color="grey", linestyle=":", linewidth=1.0, alpha=0.7)
                continue # Skip to the next region

            # --- Define Data Arrays for the Current Region ---
            x_region = np.arange(start, end + 1)
            profile_region = self.profile[start : end + 1]
            # Ensure background region slicing is safe
            if start < len(self.background) and end < len(self.background):
                 background_region = self.background[start : end + 1]
            else: # Fallback if region somehow exceeds background array
                 print(f"Warning: Plot region {i+1} exceeds background bounds.")
                 background_region = np.full_like(profile_region, np.min(self.background))

            # Ensure profile_region and background_region have same shape for subtraction
            if profile_region.shape != background_region.shape:
                print(f"Warning: Shape mismatch between profile ({profile_region.shape}) and background ({background_region.shape}) for peak {i+1}. Skipping area calc.")
                self.peak_areas_rolling_ball.append(0.0)
                self.peak_areas_straight_line.append(0.0)
                self.peak_areas_valley.append(0.0)
                continue

            # --- Calculate Areas Using All Methods ---
            # 1. Rolling Ball Area
            area_rb = np.trapz(profile_region - background_region, x=x_region)
            area_rb = max(0, area_rb)
            self.peak_areas_rolling_ball.append(area_rb)

            # 2. Straight Line Area
            x_baseline_sl = np.array([start, end])
            # Ensure indices are valid for profile access
            if start < len(self.profile) and end < len(self.profile):
                y_baseline_points_sl = np.array([self.profile[start], self.profile[end]])
                y_baseline_interpolated_sl = np.interp(x_region, x_baseline_sl, y_baseline_points_sl)
                area_sl = np.trapz(profile_region - y_baseline_interpolated_sl, x=x_region)
                area_sl = max(0, area_sl)
            else:
                print(f"Warning: Invalid indices for straight line baseline for peak {i+1}.")
                area_sl = 0.0 # Assign 0 area if indices invalid
            self.peak_areas_straight_line.append(area_sl)

            # 3. Valley-to-Valley Area
            area_vv = 0.0 # Default value
            try: # Wrap valley finding in try-except
                search_range = max(15, int((end - start) * 1.5))
                # Find left valley (using robust logic from previous answers)
                valley_start_idx = start
                for idx in range(start - 1, max(-1, start - search_range - 1), -1):
                    if idx < 0: valley_start_idx = 0; break
                    if self.profile[idx] > self.profile[idx + 1]: valley_start_idx = idx + 1; break
                    if idx == max(0, start - search_range):
                        min_idx_in_search = np.argmin(self.profile[max(0, start - search_range):start]) + max(0, start - search_range)
                        valley_start_idx = min_idx_in_search
                        if start > 0 and self.profile[start-1] < self.profile[valley_start_idx]: valley_start_idx = start-1
                        break
                # Find right valley (using robust logic from previous answers)
                valley_end_idx = end
                for idx in range(end + 1, min(len(self.profile), end + search_range + 1)):
                    if idx >= len(self.profile): valley_end_idx = len(self.profile) - 1; break
                    if self.profile[idx] > self.profile[idx - 1]: valley_end_idx = idx - 1; break
                    if idx == min(len(self.profile) - 1, end + search_range):
                        min_idx_in_search = np.argmin(self.profile[end + 1 : min(len(self.profile), end + search_range + 1)]) + end + 1
                        if min_idx_in_search < len(self.profile):
                             valley_end_idx = min_idx_in_search
                             if end < len(self.profile) - 1 and self.profile[end+1] < self.profile[valley_end_idx]: valley_end_idx = end+1
                        else: valley_end_idx = end
                        break
                # Validate and calculate
                valley_start_idx = max(0, valley_start_idx)
                valley_end_idx = min(len(self.profile) - 1, valley_end_idx)
                if valley_start_idx > start: valley_start_idx = start
                if valley_end_idx < end: valley_end_idx = end
                if valley_end_idx <= valley_start_idx:
                    print(f"Warning: Valley detection invalid for Peak {i+1}. Using region boundaries.")
                    valley_start_idx = start
                    valley_end_idx = end

                # Ensure indices valid before accessing profile
                if 0 <= valley_start_idx < len(self.profile) and 0 <= valley_end_idx < len(self.profile):
                    x_baseline_valley = np.array([valley_start_idx, valley_end_idx])
                    y_baseline_points_valley = np.array([self.profile[valley_start_idx], self.profile[valley_end_idx]])
                    y_baseline_interpolated_valley = np.interp(x_region, x_baseline_valley, y_baseline_points_valley)
                    area_vv = np.trapz(profile_region - y_baseline_interpolated_valley, x=x_region)
                    area_vv = max(0, area_vv)
                else:
                    print(f"Warning: Invalid indices for valley baseline for peak {i+1}.")
                    area_vv = 0.0 # Assign 0 area if indices invalid

            except IndexError as ie:
                 print(f"IndexError during valley finding for peak {i+1}: {ie}. Assigning area 0.")
                 area_vv = 0.0
            except Exception as e_vv:
                 print(f"Error during valley calculation for peak {i+1}: {e_vv}. Assigning area 0.")
                 area_vv = 0.0
            self.peak_areas_valley.append(area_vv)


            # --- Plot Baselines and Fills Based on Selected Method ---
            current_area = 0.0 # Area to display in text
            if self.method == "Rolling Ball":
                if i == 0: self.ax.plot(self.background, color="green", linestyle="--", linewidth=1, label="Rolling Ball BG")
                self.ax.fill_between(x_region, background_region, profile_region,
                                     where=profile_region >= background_region,
                                     color="yellow", alpha=0.4, interpolate=True)
                current_area = area_rb
            elif self.method == "Straight Line":
                 self.ax.plot(x_baseline_sl, y_baseline_points_sl, color="purple", linestyle="--", linewidth=1, label="Straight Line BG" if i == 0 else "")
                 self.ax.fill_between(x_region, y_baseline_interpolated_sl, profile_region,
                                      where=profile_region >= y_baseline_interpolated_sl,
                                      color="cyan", alpha=0.4, interpolate=True)
                 current_area = area_sl
            elif self.method == "Valley-to-Valley":
                 # Only plot if baseline points are valid
                 if 0 <= valley_start_idx < len(self.profile) and 0 <= valley_end_idx < len(self.profile):
                     self.ax.plot(x_baseline_valley, y_baseline_points_valley, color="orange", linestyle="--", linewidth=1, label="Valley BG" if i == 0 else "")
                     self.ax.fill_between(x_region, y_baseline_interpolated_valley, profile_region,
                                          where=profile_region >= y_baseline_interpolated_valley,
                                          color="lightblue", alpha=0.4, interpolate=True)
                 current_area = area_vv

            # --- Plot COMBINED Area Text and Region Markers ---
            combined_text = f"Peak {i + 1}\n{current_area:.1f}" # Peak number \n Area

            # Position text slightly above the highest point in the region
            text_y_base = np.max(profile_region)
            # Consider baseline height if it's higher than the profile peak in that region
            if self.method == "Rolling Ball": text_y_base = max(text_y_base, np.max(background_region))
            elif self.method == "Straight Line": text_y_base = max(text_y_base, np.max(y_baseline_interpolated_sl))
            elif self.method == "Valley-to-Valley":
                # Check if valley baseline was valid before using it for positioning
                if 0 <= valley_start_idx < len(self.profile) and 0 <= valley_end_idx < len(self.profile):
                    text_y_base = max(text_y_base, np.max(y_baseline_interpolated_valley))
                # else: use profile_region max as fallback

            text_y_pos = text_y_base + profile_range * 0.06 # Offset slightly above max point

            # Draw the combined text label
            self.ax.text((start + end) / 2, text_y_pos, combined_text,
                         ha="center", va="bottom", fontsize=7, color='black', zorder=6)

            # Update the maximum text position encountered for y-limit calculation
            max_text_y_position = max(max_text_y_position, text_y_pos + profile_range*0.03) # Add buffer for text height


            # Plot vertical markers for region boundaries
            self.ax.axvline(start, color="blue", linestyle=":", linewidth=1.5, alpha=0.8)
            self.ax.axvline(end, color="red", linestyle=":", linewidth=1.5, alpha=0.8)

        # --- Final Plot Configuration (Top Subplot) ---
        self.ax.set_ylabel("Intensity (A.U.)")
        self.ax.legend(fontsize='small', loc='upper right')
        self.ax.set_title("Intensity Profile and Peak Regions")
        plt.setp(self.ax.get_xticklabels(), visible=False) # Hide X labels on top plot
        self.ax.tick_params(axis='x', which='both', bottom=False, top=False) # Remove X ticks

        # Set plot limits for better visualization
        if len(self.profile) > 1:
            self.ax.set_xlim(0, len(self.profile) - 1)
        prof_min, prof_max = np.min(self.profile), np.max(self.profile)
        # Adjust y-limits to give space above the highest point (profile or text label)
        y_max_limit = max(prof_max, max_text_y_position) + profile_range * 0.05 # 5% padding above highest element
        self.ax.set_ylim(prof_min - profile_range * 0.05, y_max_limit)


        # --- Display Cropped Image (Bottom Subplot) ---
        ax_image.clear() # Clear previous image plot
        if self.cropped_image:
            try:
                # Rotate PIL image for horizontal display
                rotated_pil_image = self.cropped_image.rotate(90, expand=True)
                ax_image.imshow(rotated_pil_image, cmap='gray', aspect='auto',
                                extent=[0, len(self.profile)-1, 0, rotated_pil_image.height]) # Match x-axis extent
                ax_image.set_xlabel("Pixel Index Along Profile Axis")
                ax_image.set_yticks([]) # Hide y-axis ticks for the image subplot
                ax_image.set_ylabel("Lane Width", fontsize='small')
            except Exception as img_e:
                print(f"Error displaying cropped image: {img_e}")
                ax_image.text(0.5, 0.5, 'Error loading image', ha='center', va='center', transform=ax_image.transAxes)
                ax_image.set_xticks([]); ax_image.set_yticks([])
        else:
            ax_image.text(0.5, 0.5, 'No Image Available', ha='center', va='center', transform=ax_image.transAxes)
            ax_image.set_xticks([]); ax_image.set_yticks([])

        # Adjust layout tightly
        try:
             self.fig.subplots_adjust(left=0.07, right=0.98, top=0.92, bottom=0.1, hspace=0.05) # Fine-tune margins
        except Exception as layout_e:
            print(f"Error adjusting layout: {layout_e}")

        # --- Draw Plot ---
        try:
            self.canvas.draw_idle() # Use draw_idle for better responsiveness
        except Exception as draw_e:
             print(f"Error drawing canvas: {draw_e}")
             
        plt.close(self.fig)

    def get_final_peak_area(self):
        """Return the list of calculated peak areas based on the selected method."""
        # Determine the number of valid peaks based on the number of regions processed
        num_valid_peaks = len(self.peak_regions)

        areas_to_return = []
        current_area_list = None

        if self.method == "Rolling Ball":
            current_area_list = self.peak_areas_rolling_ball
        elif self.method == "Straight Line":
            current_area_list = self.peak_areas_straight_line
        elif self.method == "Valley-to-Valley":
            current_area_list = self.peak_areas_valley
        else:
            print(f"Warning: Unknown area calculation method '{self.method}'. Returning empty list.")
            return []

        # Ensure consistency between calculated areas and the number of regions plotted
        if len(current_area_list) != num_valid_peaks:
            print(f"Warning: Area list length ({len(current_area_list)}) mismatch for method '{self.method}' "
                  f"(expected {num_valid_peaks}). Returning potentially incomplete list.")
            # Truncate or pad list to match the number of peaks processed
            areas_to_return = current_area_list[:num_valid_peaks]
            # Optional: Pad with zeros if the list is shorter than expected
            # while len(areas_to_return) < num_valid_peaks:
            #     areas_to_return.append(0.0)
        else:
            areas_to_return = current_area_list

        return areas_to_return
    
        
    


class LiveViewLabel(QLabel):
    def __init__(self, font_type, font_size, marker_color, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)  # Enable mouse tracking
        self.preview_marker_enabled = False
        self.preview_marker_text = ""
        self.preview_marker_position = None
        self.marker_font_type = font_type
        self.marker_font_size = font_size
        self.marker_color = marker_color
        self.setFocusPolicy(Qt.StrongFocus)
        self.bounding_box_preview = []
        self.measure_quantity_mode = False
        self.counter = 0
        self.zoom_level = 1.0  # Initial zoom level
        self.pan_start = None  # Start position for panning
        self.pan_offset = QPointF(0, 0)  # Offset for panning
        self.quad_points = []  # Stores 4 points of the quadrilateral
        self.selected_point = -1  # Index of selected corner (-1 = none)
        self.drag_threshold = 10  # Pixel radius for selecting corners
        self.bounding_box_complete = False
        self.mode=None
        # Add rectangle-related attributes
        self.rectangle_start = None  # Start point of the rectangle
        self.rectangle_end = None    # End point of the rectangle
        self.rectangle_points = []   # Stores the rectangle points
        self.drag_start_pos = None  # For tracking drag operations
        self.draw_edges=True

    def mouseMoveEvent(self, event):
        if self.preview_marker_enabled:
            self.preview_marker_position = event.pos()
            self.update()  # Trigger repaint to show the preview
        if self.selected_point != -1 and self.measure_quantity_mode:# and self.mode=="quad":
            # Update dragged corner position
            self.quad_points[self.selected_point] = self.transform_point(event.pos())
            self.update()  # Show the bounding box preview
        if self.selected_point != -1 and self.mode=="move":
            self.drag_start_pos=event.pos()
        
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if self.preview_marker_enabled:
            # Place the marker text permanently at the clicked position
            parent = self.parent()
            parent.place_custom_marker(event, self.preview_marker_text)
            self.update()  # Clear the preview
        if self.measure_quantity_mode and self.mode=="quad":
            # Check if clicking near existing corner
            for i, p in enumerate(self.quad_points):
                if (self.transform_point(event.pos()) - p).manhattanLength() < self.drag_threshold:
                    self.selected_point = i
                    return
    
            # Add new point if < 4 corners
            if len(self.quad_points) < 4:
                self.quad_points.append(self.transform_point(event.pos()))
                self.selected_point = len(self.quad_points) - 1
    
            if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                # self.quad_points = [self.transform_point(p) for p in self.quad_points]
                self.bounding_box_complete = True
            self.update()  # Trigger repaint
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode=="quad":
            self.selected_point = -1
            self.update()
        super().mouseReleaseEvent(event)

    def transform_point(self, point):
        """Transform a point from widget coordinates to image coordinates."""
        if self.zoom_level != 1.0:
            return QPointF(
                (point.x() - self.pan_offset.x()) / self.zoom_level,
                (point.y() - self.pan_offset.y()) / self.zoom_level
            )
        return QPointF(point)

    def zoom_in(self):
        self.zoom_level *= 1.1
        self.update()

    def zoom_out(self):
        self.zoom_level /= 1.1
        if self.zoom_level < 1.0:
            self.zoom_level = 1.0
            self.pan_offset = QPointF(0, 0)  # Reset pan offset when zoom is reset
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)


        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)

        # Draw the preview marker if enabled
        if self.preview_marker_enabled and self.preview_marker_position:
            painter.setOpacity(0.5)  # Semi-transparent preview
            font = QFont(self.marker_font_type)
            font.setPointSize(self.marker_font_size)
            painter.setFont(font)
            painter.setPen(self.marker_color)
            text_width = painter.fontMetrics().horizontalAdvance(self.preview_marker_text)
            text_height = painter.fontMetrics().height()
            # Draw the text at the cursor's position
            x, y = self.preview_marker_position.x(), self.preview_marker_position.y()
            if self.zoom_level != 1.0:
                x = (x - self.pan_offset.x()) / self.zoom_level
                y = (y - self.pan_offset.y()) / self.zoom_level
            painter.drawText(int(x - text_width / 2), int(y + text_height / 4), self.preview_marker_text)
            
        if len(self.quad_points) > 0 and len(self.quad_points) <4 :
            for p in self.quad_points:
                painter.setPen(QPen(Qt.red, 2))
                painter.drawEllipse(p, 1, 1)
    
        if len(self.quad_points) == 4 and self.draw_edges==True:
            painter.setPen(QPen(Qt.red, 2))
            painter.drawPolygon(QPolygonF(self.quad_points))
            # Draw draggable corners
            for p in self.quad_points:
                painter.drawEllipse(p, self.drag_threshold, self.drag_threshold)
    
        # Draw the bounding box preview if it exists
        if self.bounding_box_preview:
            painter.setPen(QPen(Qt.red, 2))  # Use green color for the bounding box
            start_x, start_y, end_x, end_y = self.bounding_box_preview
            rect = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y))
            painter.drawRect(rect)
    
        painter.end()
        self.update()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            if self.preview_marker_enabled:
                self.preview_marker_enabled = False  # Turn off the preview
                self.update()  # Clear the overlay
            self.measure_quantity_mode = False
            self.counter = 0
            self.bounding_box_complete = False
            self.quad_points = []
            self.mode=None
            self.update()
        super().keyPressEvent(event)

class CombinedSDSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
        window_width = int(self.screen_width * 0.5)  # 60% of screen width
        window_height = int(self.screen_height * 0.75)  # 95% of screen height
        self.window_title="IMAGING ASSISTANT V5.2"
        self.setWindowTitle(self.window_title)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.undo_stack = []
        self.redo_stack = []
        self.quantities_peak_area_dict = {}
        self.image_path = None
        self.image = None
        self.image_master= None
        self.image_before_padding = None
        self.image_contrasted=None
        self.image_before_contrast=None
        self.contrast_applied=False
        self.image_padded=False
        self.predict_size=False
        self.warped_image=None
        self.left_markers = []
        self.right_markers = []
        self.top_markers = []
        self.custom_markers=[]
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.font_rotation=-45
        self.image_width=0
        self.image_height=0
        self.new_image_width=0
        self.new_image_height=0
        self.base_name="Image"
        self.image_path=""
        self.x_offset_s=0
        self.y_offset_s=0
        self.peak_area=[]
        self.is_modified=False
        
        self.peak_dialog_settings = {
            'rolling_ball_radius': 50,
            'peak_height_factor': 0.1,
            'peak_distance': 30,
            'peak_prominence_factor': 0.02,
            'peak_spread_pixels': 10,
            'band_estimation_method': "Mean",
            'area_subtraction_method': "Valley-to-Valley"
        }
        self.persist_peak_settings_enabled = True # State of the checkbox
        
        # Variables to store bounding boxes and quantities
        self.bounding_boxes = []
        self.up_bounding_boxes = []
        self.standard_protein_areas=[]
        self.quantities = []
        self.protein_quantities = []
        self.measure_quantity_mode = False
        # Initialize self.marker_values to None initially
        self.marker_values_dict = {
            "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
        }
        self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"]
        self.top_label_dict={"Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"]}
        
        
        self.custom_marker_name = "Custom"
        # Load the config file if exists
        
        self.marker_mode = None
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_slider_range=[-1000,1000]
        self.right_slider_range=[-1000,1000]
        self.top_slider_range=[-1000,1000]
               
        
        self.top_padding = 0
        self.font_color = QColor(0, 0, 0)  # Default to black
        self.custom_marker_color = QColor(0, 0, 0)  # Default to black
        self.font_family = "Arial"  # Default font family
        self.font_size = 12  # Default font size
        self.image_array_backup= None
        self.run_predict_MW=False
        
        self.create_menu_bar()
        
        # Main container widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Upper section (Preview and buttons)
        upper_layout = QHBoxLayout()

        self.label_width=int(self.screen_width * 0.28)
    
        self.live_view_label = LiveViewLabel(
            font_type=QFont("Arial"),
            font_size=int(24),
            marker_color=QColor(0,0,0),
            parent=self,
        )
        # Image display
        self.live_view_label.setStyleSheet("border: 1px solid black;")
        # self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setFixedSize(self.label_width, self.label_width)
        # self.live_view_label.mousePressEvent = self.add_band()
        # self.live_view_label.mousePressEvent = self.add_band
        
        
       

        # Buttons for image loading and saving
        # Load, save, and crop buttons
        buttons_layout = QVBoxLayout()
        load_button = QPushButton("Load Image")
        load_button.setToolTip("Load an image or a previously saved file. Shortcut: Ctrl+O or CMD+O")
        load_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        load_button.clicked.connect(self.load_image)
        buttons_layout.addWidget(load_button)
        
        paste_button = QPushButton('Paste Image')
        paste_button.setToolTip("Paste an image from clipboard or folder. Shortcut: Ctrl+V or CMD+V")
        paste_button.clicked.connect(self.paste_image)  # Connect the button to the paste_image method
        paste_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(paste_button)
    
        
        reset_button = QPushButton("Reset Image")  # Add Reset Image button
        reset_button.setToolTip("Reset all image manipulations and marker placements. Shortcut: Ctrl+R or CMD+R")
        reset_button.clicked.connect(self.reset_image)  # Connect the reset functionality
        reset_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(reset_button)
        
        
        
        copy_button = QPushButton('Copy Image to Clipboard')
        copy_button.setToolTip("Copy the modified image to clipboard. Shortcut: Ctrl+C or CMD+C")
        copy_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        copy_button.clicked.connect(self.copy_to_clipboard)
        buttons_layout.addWidget(copy_button)
        
        save_button = QPushButton("Save Image with Configuration")
        save_button.setToolTip("Save the modified image with the configuration files. Shortcut: Ctrl+S or CMD+S")
        save_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_button.clicked.connect(self.save_image)
        buttons_layout.addWidget(save_button)
        
        #if platform.system() == "Windows": # "Darwin" for MacOS # "Windows" for Windows
        #    copy_svg_button = QPushButton('Copy SVG Image to Clipboard')
        #    copy_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        #    copy_svg_button.clicked.connect(self.copy_to_clipboard_SVG)
        #    buttons_layout.addWidget(copy_svg_button)
            
        
        save_svg_button = QPushButton("Save SVG Image (MS Word Import)")
        save_svg_button.setToolTip("Save the modified image as an SVG file so that it can be modified in MS Word or similar. Shortcut: Ctrl+M or CMD+M")
        save_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_svg_button.clicked.connect(self.save_image_svg)
        buttons_layout.addWidget(save_svg_button)
        
        undo_redo_layout=QHBoxLayout()
        
        undo_button = QPushButton("Undo")
        undo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        undo_button.clicked.connect(self.undo_action)
        undo_button.setToolTip("Undo settings related to image. Cannot Undo Marker Placement. Use remove last option. Shortcut: Ctrl+Z or CMD+Z")
        undo_redo_layout.addWidget(undo_button)
        
        redo_button = QPushButton("Redo")
        redo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        redo_button.clicked.connect(self.redo_action)
        redo_button.setToolTip("Redo settings related to image. Cannot Undo Marker Placement. Use remove last option.Shortcut: Ctrl+Y or CMD+Y")        
        undo_redo_layout.addWidget(redo_button)
        
        buttons_layout.addLayout(undo_redo_layout)
        
        # New Zoom buttons layout
        zoom_layout = QHBoxLayout()
        
        # Zoom In button
        zoom_in_button = QPushButton("Zoom In")
        zoom_in_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_in_button.clicked.connect(self.zoom_in)
        zoom_in_button.setToolTip("Increase zoom level. Click the display window and use arrow keys for moving")
        
        # Zoom Out button
        zoom_out_button = QPushButton("Zoom Out")
        zoom_out_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_out_button.clicked.connect(self.zoom_out)
        zoom_out_button.setToolTip("Decrease zoom level. Click the display window and use arrow keys for moving")
        
        # Add Zoom buttons to the zoom layout
        zoom_layout.addWidget(zoom_in_button)
        zoom_layout.addWidget(zoom_out_button)
        
        # Add Zoom layout below Undo/Redo
        buttons_layout.addLayout(zoom_layout)
        
        # buttons_layout.addStretch()
        
        
        upper_layout.addWidget(self.live_view_label, stretch=1)
        upper_layout.addLayout(buttons_layout, stretch=1)
        layout.addLayout(upper_layout)

        # Lower section (Tabbed interface)
        self.tab_widget = QTabWidget()
        # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
        self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
        self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
        self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
        self.tab_widget.addTab(self.create_markers_tab(), "Markers")
        self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
        self.tab_widget.addTab(self.analysis_tab(), "Analysis")
        
        layout.addWidget(self.tab_widget)
        
        # Example: Shortcut for loading an image (Ctrl + O)
        self.load_shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        self.load_shortcut.activated.connect(self.load_image)
        
        # Example: Shortcut for saving an image (Ctrl + S)
        self.save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        self.save_shortcut.activated.connect(self.save_image)
        
        # Example: Shortcut for resetting the image (Ctrl + R)
        self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.reset_shortcut.activated.connect(self.reset_image)
        
        # Example: Shortcut for copying the image to clipboard (Ctrl + C)
        self.copy_shortcut = QShortcut(QKeySequence("Ctrl+C"), self)
        self.copy_shortcut.activated.connect(self.copy_to_clipboard)
        
        # Example: Shortcut for pasting an image from clipboard (Ctrl + V)
        self.paste_shortcut = QShortcut(QKeySequence("Ctrl+V"), self)
        self.paste_shortcut.activated.connect(self.paste_image)
        
        # Example: Shortcut for predicting molecular weight (Ctrl + P)
        self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        self.predict_shortcut.activated.connect(self.predict_molecular_weight)
        
        # Example: Shortcut for clearing the prediction marker (Ctrl + Shift + P)
        self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)
        
        # Example: Shortcut for saving SVG image (Ctrl + Shift + S)
        self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
        self.save_svg_shortcut.activated.connect(self.save_image_svg)
        
        # Example: Shortcut for enabling left marker mode (Ctrl + L)
        self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
        self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)
        
        # Example: Shortcut for enabling right marker mode (Ctrl + R)
        self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
        self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)
        
        # Example: Shortcut for enabling top marker mode (Ctrl + T)
        self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
        self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
        self.grid_shortcut.activated.connect(lambda: self.show_grid_checkbox.setChecked(not self.show_grid_checkbox.isChecked()))
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.guidelines_shortcut.activated.connect(lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked()))
        
        # Example: Shortcut for increasing grid size (Ctrl + Shift + Up)
        self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
        self.increase_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1))
        
        # Example: Shortcut for decreasing grid size (Ctrl + Shift + Down)
        self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
        self.decrease_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1))
        
        # # Example: Shortcut for increasing font size (Ctrl + Plus)
        # self.increase_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Up"), self)
        # self.increase_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() + 1))
        
        # # Example: Shortcut for decreasing font size (Ctrl + Minus)
        # self.decrease_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Down"), self)
        # self.decrease_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() - 1))
        
        # Example: Shortcut for custom marker left arrow (Ctrl + Left Arrow)
        self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker right arrow (Ctrl + Right Arrow)
        self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker top arrow (Ctrl + Up Arrow)
        self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker bottom arrow (Ctrl + Down Arrow)
        self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
        self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for inverting image (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self.invert_shortcut.activated.connect(self.invert_image)
        
        # Example: Shortcut for converting to grayscale (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
        self.invert_shortcut.activated.connect(self.convert_to_black_and_white)
        
        # Example: Move quickly between tabs (Ctrl + 1,2,3,4)
        self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
        self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
        
        self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
        self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
        
        self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
        self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
        
        self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
        self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
        
        self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
        self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
        
        self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
        self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
        
        self.undo_shortcut = QShortcut(QKeySequence("Ctrl+Z"), self)
        self.undo_shortcut.activated.connect(self.undo_action)
        
        self.redo_shortcut = QShortcut(QKeySequence("Ctrl+Y"), self)
        self.redo_shortcut.activated.connect(self.redo_action)
        self.load_config()
        
    def prompt_save_if_needed(self):
        if not self.is_modified:
            return True # No changes, proceed
        """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
        reply = QMessageBox.question(self, 'Unsaved Changes',
                                     "You have unsaved changes. Do you want to save them?",
                                     QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                     QMessageBox.Save) # Default button is Save

        if reply == QMessageBox.Save:
            return self.save_image() # Returns True if saved, False if save cancelled
        elif reply == QMessageBox.Discard:
            return True # Proceed without saving
        else: # Cancelled
            return False # Abort the current action (close/load)

    def closeEvent(self, event):
        """Overrides the default close event to prompt for saving."""
        if self.prompt_save_if_needed():

            event.accept() # Proceed with closing
        else:
            event.ignore() # Abort closing
        
    def quadrilateral_to_rect(self, image, quad_points):
        image = image.convertToFormat(QImage.Format_RGBA8888)
        # Get the dimensions of the live view label and the actual image
        label_width = self.live_view_label.width()
        label_height = self.live_view_label.height()
        image_width = image.width()
        image_height = image.height()
        
        # Calculate scaling factors and offsets
        scale_x = image_width / label_width
        scale_y = image_height / label_height
        try:
            if len(self.live_view_label.bounding_box_preview) == 4:
                start_x, start_y = self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1]
                end_x, end_y = self.live_view_label.bounding_box_preview[2], self.live_view_label.bounding_box_preview[3]
                
                if self.live_view_label.zoom_level != 1.0:
                    start_x = int((start_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level)
                    start_y = int((start_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level)
                    end_x = int((end_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level)
                    end_y = int((end_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level)
                
                width, height = abs(end_x - start_x), abs(end_y - start_y)
                x, y = min(start_x, end_x), min(start_y, end_y)
                
                # Convert coordinates to image space
                displayed_width = self.live_view_label.width()
                displayed_height = self.live_view_label.height()
                image_width = self.image.width()
                image_height = self.image.height()
                scale = min(displayed_width / image_width, displayed_height / image_height)
                x_offset = (displayed_width - image_width * scale) / 2
                y_offset = (displayed_height - image_height * scale) / 2
                
                x = max(0, min((x - x_offset) / scale, image_width))
                y = max(0, min((y - y_offset) / scale, image_height))
                w = min(width / scale, image_width - x)
                h = min(height / scale, image_height - y)
                
                if x + w > self.image.width():
                    w = self.image.width() - x
                if y + h > self.image.height():
                    h = self.image.height() - y
            
                if w <= 0 or h <= 0:
                    return 0  # Skip invalid regions
            
                cropped = self.image.copy(int(x), int(y), int(w), int(h))
                return cropped
        except:
            QMessageBox.warning(self, "Error", "No bands detected")
            
            
        
        if len(self.live_view_label.quad_points) == 4:
            
            # Transform the quadrilateral points from label coordinates to image coordinates
            src_points = []
            for point in quad_points:
                # Adjust for zoom and pan if applicable
                if self.live_view_label.zoom_level != 1.0:
                    x = (point.x() - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                    y = (point.y() - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
                else:
                    x = point.x()
                    y = point.y()
                
                # Scale the points to match the image dimensions
                x_image = x * scale_x
                y_image = y * scale_y
                src_points.append([x_image, y_image])
            
            src_points = np.array(src_points, dtype=np.float32)
            
            # Calculate the bounding box of the quadrilateral in image coordinates
            min_x = np.min(src_points[:, 0])
            max_x = np.max(src_points[:, 0])
            min_y = np.min(src_points[:, 1])
            max_y = np.max(src_points[:, 1])
            
            # Ensure the bounding box is within the image dimensions
            min_x = np.maximum(0, min_x)
            max_x = np.minimum(image_width, max_x)
            min_y = np.maximum(0, min_y)
            max_y = np.minimum(image_height, max_y)
            
            # Define the destination points (the rectangle to which we want to map the quadrilateral)
            dst_points = np.array([
                [0, 0],
                [max_x - min_x, 0],
                [max_x - min_x, max_y - min_y],
                [0, max_y - min_y]
            ], dtype=np.float32)
            
            # Determine the image format (grayscale vs RGBA)
            if image.format() == image.Format_Grayscale8:
                channels = 1
                format_type = image.Format_Grayscale8
            else:
                channels = 4  # Assume RGBA
                format_type = image.Format_RGBA8888
            
            # Convert QImage to numpy
            ptr = image.constBits()
            ptr.setsize(image.byteCount())
            img_data = np.array(ptr).reshape(image_height, image_width, channels)
            
            if channels == 1:
                img = np.frombuffer(img_data, dtype=np.uint8).reshape((image_height, image_width))
            else:
                img = np.frombuffer(img_data, dtype=np.uint8).reshape((image_height, image_width, channels))
            
            # Compute the perspective transform matrix using numpy
            A = np.zeros((8, 8), dtype=np.float32)
            b = np.zeros((8,), dtype=np.float32)
            
            for i in range(4):
                x, y = src_points[i]
                u, v = dst_points[i]
                A[i*2] = [x, y, 1, 0, 0, 0, -x*u, -y*u]
                A[i*2+1] = [0, 0, 0, x, y, 1, -x*v, -y*v]
                b[i*2] = u
                b[i*2+1] = v
            
            # Solve the system of equations
            h = np.linalg.solve(A, b)
            
            # Correctly create 3x3 transformation matrix
            transform_matrix = np.zeros((3, 3), dtype=np.float32)
            transform_matrix[0, :] = [h[0], h[1], h[2]]
            transform_matrix[1, :] = [h[3], h[4], h[5]]
            transform_matrix[2, :] = [h[6], h[7], 1.0]
            
            # Create output image
            out_width = int(max_x - min_x)
            out_height = int(max_y - min_y)
            
            if channels == 1:
                warped_image = np.zeros((out_height, out_width), dtype=np.uint8)
            else:
                warped_image = np.zeros((out_height, out_width, channels), dtype=np.uint8)
            
            # Create mesh grid for destination coordinates
            y_coords, x_coords = np.mgrid[0:out_height, 0:out_width]
            coords = np.vstack((x_coords.flatten(), y_coords.flatten(), np.ones_like(x_coords.flatten())))
            
            # Apply inverse transformation
            transform_matrix_inv = np.linalg.inv(transform_matrix)
            source_coords = np.dot(transform_matrix_inv, coords)
            
            # Convert to homogeneous coordinates
            source_coords = source_coords / source_coords[2, :]
            source_x = source_coords[0, :].reshape(out_height, out_width)
            source_y = source_coords[1, :].reshape(out_height, out_width)
            
            # Find valid coordinates (within the source image)
            valid = (source_x >= 0) & (source_x < image_width) & (source_y >= 0) & (source_y < image_height)
            
            # Round to nearest pixel
            source_x_int = np.round(source_x).astype(np.int32)
            source_y_int = np.round(source_y).astype(np.int32)
            
            # Apply bounds
            source_x_int = np.clip(source_x_int, 0, image_width - 1)
            source_y_int = np.clip(source_y_int, 0, image_height - 1)
            
            # Copy pixels from source to destination
            if channels == 1:
                warped_image[y_coords[valid], x_coords[valid]] = img[source_y_int[valid], source_x_int[valid]]
            else:
                for i in range(channels):
                    warped_image[y_coords[valid], x_coords[valid], i] = img[source_y_int[valid], source_x_int[valid], i]
            
            # Convert the resulting warped image back to a QImage
            if channels == 1:
                warped_qimage = QImage(warped_image.data, out_width, out_height, out_width, format_type)
            else:
                warped_qimage = QImage(warped_image.data, out_width, out_height, out_width * channels, format_type)
        
        return warped_qimage
    
    

        
    def create_menu_bar(self):
        # Create the menu bar
        menubar = self.menuBar()

        # Create the "File" menu
        file_menu = menubar.addMenu("File")

        # Add "Load Image" action
        load_action = QAction("Load Image", self)
        load_action.triggered.connect(self.load_image)
        file_menu.addAction(load_action)

        # Add "Save Image" action
        save_action = QAction("Save Image", self)
        save_action.triggered.connect(self.save_image)
        file_menu.addAction(save_action)
        
        # Add "Save Image SVG" action
        save_action_svg = QAction("Save Image as SVG", self)
        save_action_svg.triggered.connect(self.save_image_svg)
        file_menu.addAction(save_action_svg)

        # Add "Reset Image" action
        reset_action = QAction("Reset Image", self)
        reset_action.triggered.connect(self.reset_image)
        file_menu.addAction(reset_action)

        # Add a separator
        file_menu.addSeparator()

        # Add "Exit" action
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Create the "About" menu
        about_menu = menubar.addMenu("About")

        # Add "GitHub" action
        github_action = QAction("GitHub", self)
        github_action.triggered.connect(self.open_github)
        about_menu.addAction(github_action)
        
        # self.statusBar().showMessage("Ready")
    def open_github(self):
        # Open the GitHub link in the default web browser
        QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
        
    def zoom_in(self):
        self.live_view_label.zoom_in()
        self.update_live_view()

    def zoom_out(self):
        self.live_view_label.zoom_out()
        self.update_live_view()
    
    def enable_standard_protein_mode(self):
        """"Enable mode to define standard protein amounts for creating a standard curve."""
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        # Assign mouse event handlers for bounding box creation
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_measure_protein_mode(self):
        """Enable mode to measure protein quantity using the standard curve."""
        if len(self.quantities) < 2:
            QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)

    def call_live_view(self):
        self.update_live_view()      

    def analyze_bounding_box(self, image, standard):  
        # Prompt user for quantity input if standard
        if standard:
            quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known protein amount (e.g., 1.5 units):")
            if ok and quantity:
                try:
                    # Extract just the numeric part if unit is included
                    quantity_value = float(quantity.split()[0])
                    # self.quantities.append(quantity_value)
                    # Calculate the area under the peak and subtract the background
                    peak_area = self.calculate_peak_area(image)  
                    self.quantities_peak_area_dict[quantity_value] = round(sum(peak_area),3)
                    self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                except (ValueError, IndexError):
                    QMessageBox.warning(self, "Error", "Please enter a valid number for protein quantity.")
            self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
        else:
            # self.up_bounding_boxes=[]
            # self.up_bounding_boxes.append((image_x, image_y, image_w, image_h))
            self.peak_area = None
            self.peak_area = self.calculate_peak_area(image)
            if len(list(self.quantities_peak_area_dict.keys()))>=2:
                self.calculate_unknown_quantity(list(self.quantities_peak_area_dict.values()), list(self.quantities_peak_area_dict.keys()), self.peak_area)
            try:
                self.target_protein_areas_text.setText(str([round(x, 3) for x in self.peak_area]))
            except:
                pass
        self.update_live_view()
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.bounding_box_start = None
    
    def calculate_peak_area(self, cropped_qimage): # Renamed input for clarity
        """Opens the PeakAreaDialog for interactive adjustment and area calculation."""
        # Convert QImage to PIL Image for the dialog
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        cropped_qimage.save(buffer, "PNG") # Use the passed QImage
        cropped_pil_image = Image.open(io.BytesIO(buffer.data()))
        buffer.close()

        # --- MODIFIED: Pass current settings and persistence state to Dialog ---
        dialog = PeakAreaDialog(
            cropped_image=cropped_pil_image,
            current_settings=self.peak_dialog_settings, # Pass the dict
            persist_checked=self.persist_peak_settings_enabled, # Pass the bool
            parent=self
        )
        # --- END MODIFIED ---

        peak_areas = [] # Default return value
        if dialog.exec_() == QDialog.Accepted:
            peak_areas = dialog.get_final_peak_area()

            # --- ADDED: Retrieve settings from dialog and update main app state ---
            # Check if persistence was enabled *in the dialog* when it closed
            if dialog.should_persist_settings():
                self.peak_dialog_settings = dialog.get_current_settings()
                self.persist_peak_settings_enabled = True
            else:
                self.persist_peak_settings_enabled = False
            # --- END ADDED ---

        return peak_areas # Return calculated areas or empty list if cancelled
    
    
    def calculate_unknown_quantity(self, peak_area_list, known_quantities, peak_area):
        coefficients = np.polyfit(peak_area_list, known_quantities, 1)
        unknown_quantity=[]
        for i in range(len(peak_area)):
            val=np.polyval(coefficients, peak_area[i])
            unknown_quantity.append(round(val,2))
        self.protein_quantities=[]
        QMessageBox.information(self, "Protein Quantification", f"Predicted Quantity: {unknown_quantity} units")

        
    def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
        """Draw quantity text at the correct position."""
        text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
        painter.drawText(text_position, str(quantity))
    
    def update_standard_protein_quantities(self):
        self.standard_protein_values.text()
    
    def move_tab(self,tab):
        self.tab_widget.setCurrentIndex(tab)
        
    def save_state(self):
        """Save the current state of the image, markers, and other relevant data."""
        state = {
            "image": self.image.copy() if self.image else None,
            "left_markers": self.left_markers.copy(),
            "right_markers": self.right_markers.copy(),
            "top_markers": self.top_markers.copy(),
            "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
            "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
            "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
            "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
            "font_family": self.font_family,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "font_rotation": self.font_rotation,
            "left_marker_shift_added": self.left_marker_shift_added,
            "right_marker_shift_added": self.right_marker_shift_added,
            "top_marker_shift_added": self.top_marker_shift_added,
            "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
            "quantities": self.quantities.copy(),
            "protein_quantities": self.protein_quantities.copy(),
            "standard_protein_areas": self.standard_protein_areas.copy(),
        }
        self.undo_stack.append(state)
        self.redo_stack.clear()
    
    def undo_action(self):
        """Undo the last action by restoring the previous state."""
        if self.undo_stack:
            # Save the current state to the redo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.redo_stack.append(current_state)
            
            # Restore the previous state from the undo stack
            previous_state = self.undo_stack.pop()
            self.image = previous_state["image"]
            self.left_markers = previous_state["left_markers"]
            self.right_markers = previous_state["right_markers"]
            self.top_markers = previous_state["top_markers"]
            self.custom_markers = previous_state["custom_markers"]
            self.image_before_padding = previous_state["image_before_padding"]
            self.image_contrasted = previous_state["image_contrasted"]
            self.image_before_contrast = previous_state["image_before_contrast"]
            self.font_family = previous_state["font_family"]
            self.font_size = previous_state["font_size"]
            self.font_color = previous_state["font_color"]
            self.font_rotation = previous_state["font_rotation"]
            self.left_marker_shift_added = previous_state["left_marker_shift_added"]
            self.right_marker_shift_added = previous_state["right_marker_shift_added"]
            self.top_marker_shift_added = previous_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
            self.quantities = previous_state["quantities"]
            self.protein_quantities = previous_state["protein_quantities"]
            self.standard_protein_areas = previous_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    
    def redo_action(self):
        """Redo the last undone action by restoring the next state."""
        if self.redo_stack:
            # Save the current state to the undo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.undo_stack.append(current_state)
            
            # Restore the next state from the redo stack
            next_state = self.redo_stack.pop()
            self.image = next_state["image"]
            self.left_markers = next_state["left_markers"]
            self.right_markers = next_state["right_markers"]
            self.top_markers = next_state["top_markers"]
            self.custom_markers = next_state["custom_markers"]
            self.image_before_padding = next_state["image_before_padding"]
            self.image_contrasted = next_state["image_contrasted"]
            self.image_before_contrast = next_state["image_before_contrast"]
            self.font_family = next_state["font_family"]
            self.font_size = next_state["font_size"]
            self.font_color = next_state["font_color"]
            self.font_rotation = next_state["font_rotation"]
            self.left_marker_shift_added = next_state["left_marker_shift_added"]
            self.right_marker_shift_added = next_state["right_marker_shift_added"]
            self.top_marker_shift_added = next_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
            self.quantities = next_state["quantities"]
            self.protein_quantities = next_state["protein_quantities"]
            self.standard_protein_areas = next_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    def analysis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Increase spacing in this tab

        # --- Molecular Weight Prediction ---
        mw_group = QGroupBox("Molecular Weight Prediction")
        mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        mw_layout = QVBoxLayout(mw_group)
        mw_layout.setSpacing(8)

        self.predict_button = QPushButton("Predict Molecular Weight")
        self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
        self.predict_button.setEnabled(False)  # Initially disabled
        self.predict_button.clicked.connect(self.predict_molecular_weight)
        mw_layout.addWidget(self.predict_button)

        layout.addWidget(mw_group)

        # --- Peak Area / Sample Quantification ---
        quant_group = QGroupBox("Peak Area and Sample Quantification")
        quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        quant_layout = QVBoxLayout(quant_group)
        quant_layout.setSpacing(8)

        # Area Definition Buttons
        area_def_layout=QHBoxLayout()
        self.btn_define_quad = QPushButton("Define Quad Area")
        self.btn_define_quad.setToolTip(
            "Click 4 corner points to define a region. \n"
            "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
            "Results may differ from Rectangle selection due to warping."
        )
        self.btn_define_quad.clicked.connect(self.enable_quad_mode)
        self.btn_define_rec = QPushButton("Define Rectangle Area")
        self.btn_define_rec.setToolTip(
            "Click and drag to define a rectangular region. \n"
            "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
            "Does not correct for skew."
        )
        self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
        self.btn_sel_rec = QPushButton("Move Selected Area")
        self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
        self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
        area_def_layout.addWidget(self.btn_define_quad)
        area_def_layout.addWidget(self.btn_define_rec)
        area_def_layout.addWidget(self.btn_sel_rec)
        quant_layout.addLayout(area_def_layout)

        # Standard Processing
        std_proc_layout = QHBoxLayout()
        self.btn_process_std = QPushButton("Process Standard Bands")
        self.btn_process_std.setToolTip("Analyze the defined area as a standard lane.\nYou will be prompted for the known quantity.")
        self.btn_process_std.clicked.connect(self.process_standard)
        std_proc_layout.addWidget(self.btn_process_std)
        quant_layout.addLayout(std_proc_layout)

        # Standard Info Display (Read-only makes more sense)
        std_info_layout = QGridLayout()
        std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
        self.standard_protein_values = QLineEdit()
        self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
        self.standard_protein_values.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_values, 0, 1)

        std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
        self.standard_protein_areas_text = QLineEdit()
        self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
        self.standard_protein_areas_text.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
        quant_layout.addLayout(std_info_layout)

        quant_layout.addWidget(self.create_separator()) # Add visual separator

        # Sample Processing
        sample_proc_layout = QHBoxLayout()
        self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
        self.btn_analyze_sample.setToolTip("Analyze the defined area as a sample lane using the standard curve.")
        self.btn_analyze_sample.clicked.connect(self.process_sample)
        sample_proc_layout.addWidget(self.btn_analyze_sample)
        quant_layout.addLayout(sample_proc_layout)


        # Sample Info Display
        sample_info_layout = QGridLayout()
        sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
        self.target_protein_areas_text = QLineEdit()
        self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
        self.target_protein_areas_text.setReadOnly(True)
        sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

        # Add Table Export button next to sample areas
        self.table_export_button = QPushButton("Export Results Table")
        self.table_export_button.setToolTip("Export the analysis results (areas, percentages, quantities) to Excel.")
        self.table_export_button.clicked.connect(self.open_table_window)
        sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
        quant_layout.addLayout(sample_info_layout)

        layout.addWidget(quant_group)

        # --- Clear Button ---
        clear_layout = QHBoxLayout() # Layout to center button maybe
        clear_layout.addStretch()
        self.clear_predict_button = QPushButton("Clear Analysis Markers")
        self.clear_predict_button.setToolTip("Clears MW prediction line and analysis regions.\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
        self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
        layout.addWidget(self.clear_predict_button)
        clear_layout.addStretch()
        layout.addLayout(clear_layout)


        layout.addStretch()
        return tab
    
    def enable_move_selection_mode(self):
        """Enable mode to move the selected quadrilateral or rectangle."""
        self.live_view_label.mode = "move"
        self.live_view_label.setCursor(Qt.SizeAllCursor)  # Change cursor to indicate move mode
        self.live_view_label.mousePressEvent = self.start_move_selection
        self.live_view_label.mouseReleaseEvent = self.end_move_selection
        self.update_live_view()
        
    def start_move_selection(self, event):
        """Start moving the selection when the mouse is pressed."""
        if self.live_view_label.mode == "move":
            self.live_view_label.draw_edges=False
            # Check if the mouse is near the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), self.live_view_label.quad_points)
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), [
                    QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1])
                ])
            self.live_view_label.drag_start_pos = event.pos()
            self.update_live_view()
        self.live_view_label.mouseMoveEvent = self.move_selection
    
    def move_selection(self, event):
        """Move the selection while the mouse is being dragged."""
        if self.live_view_label.mode == "move" and self.live_view_label.selected_point is not None:
            # Calculate the offset
            offset = event.pos() - self.live_view_label.drag_start_pos
            self.live_view_label.drag_start_pos = event.pos()
            
            # Move the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                for i in range(len(self.live_view_label.quad_points)):
                    self.live_view_label.quad_points[i] += offset
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.bounding_box_preview = (
                    self.live_view_label.bounding_box_preview[0] + offset.x(),
                    self.live_view_label.bounding_box_preview[1] + offset.y(),
                    self.live_view_label.bounding_box_preview[2] + offset.x(),
                    self.live_view_label.bounding_box_preview[3] + offset.y(),
                )
            self.update_live_view()
    
    def end_move_selection(self, event):
        """End moving the selection when the mouse is released."""
        if self.live_view_label.mode == "move":
            self.live_view_label.selected_point = -1            
            self.update_live_view()
            self.live_view_label.draw_edges=True
        self.live_view_label.mouseMoveEvent = None
        
        
             
    def get_nearest_point(self, mouse_pos, points):
        """Get the nearest point to the mouse position."""
        min_distance = float('inf')
        nearest_point = None
        for point in points:
            distance = (mouse_pos - point).manhattanLength()
            if distance < min_distance:
                min_distance = distance
                nearest_point = point
        return nearest_point
    
    def open_table_window(self):
        # Example list of peak areas
        if self.peak_area==[]:
            peak_areas = [0, 0, 0, 0]
            standard=False
            standard_dictionary={}
        else:
            peak_areas = self.peak_area
            standard_dictionary=self.quantities_peak_area_dict
            if len(self.quantities_peak_area_dict)>=2:
                standard=True
            else:
                standard=False
        
        # Open the table window with the peak areas
        self.table_window = TableWindow(peak_areas, standard_dictionary, standard,self)
        self.table_window.show()
    
    def enable_quad_mode(self):
        """Enable mode to define a quadrilateral area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "quad"
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # # Reset mouse event handlers
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        
        # Update the live view
        self.update_live_view()
    
    def enable_rectangle_mode(self):
        """Enable mode to define a rectangle area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "rectangle"
        self.live_view_label.rectangle_points = []  # Clear previous rectangle points
        self.live_view_label.rectangle_start = None  # Reset rectangle start
        self.live_view_label.rectangle_end = None    # Reset rectangle end
        self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Set mouse event handlers for rectangle mode
        self.live_view_label.mousePressEvent = self.start_rectangle
        self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
        
        # Update the live view
        self.update_live_view()

    
    def start_rectangle(self, event):
        """Record the start position of the rectangle."""
        if self.live_view_label.mode == "rectangle":
            self.live_view_label.rectangle_start = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start]
            self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
            self.update_live_view()
    
    def update_rectangle_preview(self, event):
        """Update the rectangle preview as the mouse moves."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Update bounding_box_preview with the rectangle coordinates
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            self.update_live_view()
    
    def finalize_rectangle(self, event):
        """Finalize the rectangle when the mouse is released."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Save the final bounding box preview
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            
            self.live_view_label.mode = None  # Exit rectangle mode
            self.live_view_label.setCursor(Qt.ArrowCursor)
            self.update_live_view()
        
    
    def process_standard(self,image):
        # if len(self.live_view_label.quad_points) != 4:
        #     QMessageBox.warning(self, "Error", "Please define quadrilateral area first")
        #     return
        try:
            if len(self.live_view_label.quad_points) == 4:
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                self.update_live_view()            
                warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.quad_points)            
                self.analyze_bounding_box(warped_image, standard=True) 
                return
            if len(self.live_view_label.bounding_box_preview) == 4:
                self.live_view_label.pan_offset = QPointF(0, 0)
                self.live_view_label.zoom_level=1.0
                self.update_live_view()            
                warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.rectangle_points)            
                self.analyze_bounding_box(warped_image, standard=True)
                return
        except:
            QMessageBox.warning(self, "Error", "No bands detected")
    
    def process_sample(self):
        # if len(self.live_view_label.quad_points) != 4:
        #     QMessageBox.warning(self, "Error", "Please define quadrilateral area first")
        #     return
        if len(self.live_view_label.quad_points) == 4:
            self.live_view_label.pan_offset = QPointF(0, 0)
            self.live_view_label.zoom_level=1.0
            self.update_live_view()            
            warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.quad_points)            
            self.analyze_bounding_box(warped_image, standard=False) 
            return
        if len(self.live_view_label.bounding_box_preview) == 4:
            self.live_view_label.pan_offset = QPointF(0, 0)
            self.live_view_label.zoom_level=1.0
            self.update_live_view()            
            warped_image=self.quadrilateral_to_rect(self.image,self.live_view_label.rectangle_points)            
            self.analyze_bounding_box(warped_image, standard=False)
            return                
        
        
            
    def combine_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        # Calculate initial slider ranges based on current view size
        render_scale=3
        # Use a sensible default or calculate from screen if view isn't ready
        initial_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 500
        initial_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 500
        render_width = initial_width * render_scale
        render_height = initial_height * render_scale

        # --- Image 1 Group ---
        image1_group = QGroupBox("Image 1 Overlay")
        image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image1_layout = QGridLayout(image1_group) # Use grid for better control layout
        image1_layout.setSpacing(8)

        # Load/Place/Remove buttons
        copy_image1_button = QPushButton("Copy Current Image")
        copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
        copy_image1_button.clicked.connect(self.save_image1) # Keep original name save_image1
        place_image1_button = QPushButton("Place Image 1")
        place_image1_button.setToolTip("Positions Image 1 based on sliders.")
        place_image1_button.clicked.connect(self.place_image1)
        remove_image1_button = QPushButton("Remove Image 1")
        remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
        remove_image1_button.clicked.connect(self.remove_image1)

        image1_layout.addWidget(copy_image1_button, 0, 0)
        image1_layout.addWidget(place_image1_button, 0, 1)
        image1_layout.addWidget(remove_image1_button, 0, 2)

        # Position Sliders
        image1_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image1_left_slider = QSlider(Qt.Horizontal)
        self.image1_left_slider.setRange(-render_width, render_width) # Wider range
        self.image1_left_slider.setValue(0)
        self.image1_left_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2) # Span 2 columns

        image1_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image1_top_slider = QSlider(Qt.Horizontal)
        self.image1_top_slider.setRange(-render_height, render_height) # Wider range
        self.image1_top_slider.setValue(0)
        self.image1_top_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2) # Span 2 columns

        # Resize Slider
        image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image1_resize_slider = QSlider(Qt.Horizontal)
        self.image1_resize_slider.setRange(10, 300)  # Range 10% to 300%
        self.image1_resize_slider.setValue(100)
        self.image1_resize_slider.valueChanged.connect(self.update_live_view)
        self.image1_resize_label = QLabel("100%") # Show current percentage
        self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
        image1_layout.addWidget(self.image1_resize_slider, 3, 1)
        image1_layout.addWidget(self.image1_resize_label, 3, 2)


        layout.addWidget(image1_group)

        # --- Image 2 Group --- (Similar structure to Image 1)
        image2_group = QGroupBox("Image 2 Overlay")
        image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image2_layout = QGridLayout(image2_group)
        image2_layout.setSpacing(8)

        copy_image2_button = QPushButton("Copy Current Image")
        copy_image2_button.clicked.connect(self.save_image2)
        place_image2_button = QPushButton("Place Image 2")
        place_image2_button.clicked.connect(self.place_image2)
        remove_image2_button = QPushButton("Remove Image 2")
        remove_image2_button.clicked.connect(self.remove_image2)
        image2_layout.addWidget(copy_image2_button, 0, 0)
        image2_layout.addWidget(place_image2_button, 0, 1)
        image2_layout.addWidget(remove_image2_button, 0, 2)

        image2_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image2_left_slider = QSlider(Qt.Horizontal)
        self.image2_left_slider.setRange(-render_width, render_width)
        self.image2_left_slider.setValue(0)
        self.image2_left_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)

        image2_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image2_top_slider = QSlider(Qt.Horizontal)
        self.image2_top_slider.setRange(-render_height, render_height)
        self.image2_top_slider.setValue(0)
        self.image2_top_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)

        image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image2_resize_slider = QSlider(Qt.Horizontal)
        self.image2_resize_slider.setRange(10, 300)
        self.image2_resize_slider.setValue(100)
        self.image2_resize_slider.valueChanged.connect(self.update_live_view)
        self.image2_resize_label = QLabel("100%")
        self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
        image2_layout.addWidget(self.image2_resize_slider, 3, 1)
        image2_layout.addWidget(self.image2_resize_label, 3, 2)

        layout.addWidget(image2_group)

        # --- Finalize Button ---
        finalize_layout = QHBoxLayout()
        finalize_layout.addStretch()
        finalize_button = QPushButton("Rasterize Image")
        finalize_button.setToolTip("Permanently merges the placed overlays onto the main image. This action cannot be undone easily.")
        finalize_button.clicked.connect(self.finalize_combined_image)
        layout.addWidget(finalize_button)
        finalize_layout.addStretch()
        layout.addLayout(finalize_layout)

        layout.addStretch() # Push content up
        return tab
    
    def remove_image1(self):
        """Remove Image 1 and reset its sliders."""
        if hasattr(self, 'image1'):
            # del self.image1
            # del self.image1_original
            try:
                del self.image1_position
            except:
                pass
            self.image1_left_slider.setValue(0)
            self.image1_top_slider.setValue(0)
            self.image1_resize_slider.setValue(100)
            self.update_live_view()
    
    def remove_image2(self):
        """Remove Image 2 and reset its sliders."""
        if hasattr(self, 'image2'):
            # del self.image2
            # del self.image2_original
            try:
                del self.image2_position
            except:
                pass
            self.image2_left_slider.setValue(0)
            self.image2_top_slider.setValue(0)
            self.image2_resize_slider.setValue(100)
            self.update_live_view()
    
    def save_image1(self):
        if self.image:
            self.image1 = self.image.copy()
            self.image1_original = self.image1.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 1 copied.")
    
    def save_image2(self):
        if self.image:
            self.image2 = self.image.copy()
            self.image2_original = self.image2.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 2 copied.")
    
    def place_image1(self):
        if hasattr(self, 'image1'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            self.update_live_view()
    
    def place_image2(self):
        if hasattr(self, 'image2'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            self.update_live_view()
    
    # def update_combined_image(self):
    #     if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
    #         return
    
    #     # Create a copy of the original image to avoid modifying it
    #     combined_image = self.image.copy()
    
    #     painter = QPainter(combined_image)
    
    #     # Draw Image 1 if it exists
    #     if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
    #         # Resize Image 1 based on the slider value
    #         scale_factor = self.image1_resize_slider.value() / 100.0
    #         width = int(self.image1_original.width() * scale_factor)
    #         height = int(self.image1_original.height() * scale_factor)
    #         resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image1_position[0], self.image1_position[1], resized_image1)
    
    #     # Draw Image 2 if it exists
    #     if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
    #         # Resize Image 2 based on the slider value
    #         scale_factor = self.image2_resize_slider.value() / 100.0
    #         width = int(self.image2_original.width() * scale_factor)
    #         height = int(self.image2_original.height() * scale_factor)
    #         resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image2_position[0], self.image2_position[1], resized_image2)
    
    #     painter.end()
    
    #     # Update the live view with the combined image
    #     self.live_view_label.setPixmap(QPixmap.fromImage(combined_image))
    
    def finalize_combined_image(self):
        """Overlap image1 and image2 on top of self.image and save the result as self.image."""
        # if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
        #     QMessageBox.warning(self, "Warning", "No images to overlap.")
        #     return
        
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
        # Create a high-resolution canvas for the final image
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Create a blank high-resolution image with a white background
        high_res_canvas = QImage(render_width, render_height, QImage.Format_RGB888)
        high_res_canvas.fill(Qt.white)
    
        # Create a QPainter to draw on the high-resolution canvas
        painter = QPainter(high_res_canvas)
    
        # Draw the base image (self.image) onto the high-resolution canvas
        scaled_base_image = self.image.scaled(render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        painter.drawImage(0, 0, scaled_base_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image1_position[0] + self.x_offset_s),
                int(self.image1_position[1] + self.y_offset_s),
                resized_image1
            )
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image2_position[0] + self.x_offset_s),
                int(self.image2_position[1] + self.y_offset_s),
                resized_image2
            )
    
        # End painting
        painter.end()
        
        
        self.render_image_on_canvas(
                high_res_canvas, scaled_base_image, x_start, y_start, render_scale, draw_guides=False
            )
        
        self.image=high_res_canvas.copy()
        self.image_before_padding=self.image.copy()
        self.image_before_contrast=self.image.copy()
        self.update_live_view()
    
        # Remove Image 1 and Image 2 after finalizing
        self.remove_image1()
        self.remove_image2()
        
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
    
        QMessageBox.information(self, "Success", "The images have been overlapped and saved in memory.")
    
        
    


    def font_and_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Add spacing between groups

        # --- Font Options Group ---
        font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
        font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        font_options_layout = QGridLayout(font_options_group)
        font_options_layout.setSpacing(8)

        # Font type
        font_type_label = QLabel("Font Family:")
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setEditable(False)
        self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
        self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_type_label, 0, 0)
        font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

        # Font size
        font_size_label = QLabel("Font Size:")
        self.font_size_spinner = QSpinBox()
        self.font_size_spinner.setRange(6, 72)  # Adjusted range
        self.font_size_spinner.setValue(self.font_size) # Use initialized value
        self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_size_label, 1, 0)
        font_options_layout.addWidget(self.font_size_spinner, 1, 1)

        # Font color
        self.font_color_button = QPushButton("Font Color")
        self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
        self.font_color_button.clicked.connect(self.select_font_color)
        self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
        font_options_layout.addWidget(self.font_color_button, 1, 2)

        # Font rotation (Top/Bottom)
        font_rotation_label = QLabel("Top Label Rotation:") # Specific label
        self.font_rotation_input = QSpinBox()
        self.font_rotation_input.setRange(-180, 180)
        self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
        self.font_rotation_input.setSuffix(" °") # Add degree symbol
        self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_rotation_label, 2, 0)
        font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

        layout.addWidget(font_options_group)


        # --- Image Adjustments Group ---
        img_adjust_group = QGroupBox("Image Adjustments")
        img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        img_adjust_layout = QGridLayout(img_adjust_group)
        img_adjust_layout.setSpacing(8)

        # Contrast Sliders with Value Labels
        high_contrast_label = QLabel("Brightness:") # Renamed for clarity
        self.high_slider = QSlider(Qt.Horizontal)
        self.high_slider.setRange(0, 200) # Range 0 to 200%
        self.high_slider.setValue(100) # Default 100%
        self.high_slider.valueChanged.connect(self.update_image_contrast)
        self.high_value_label = QLabel("1.00") # Display factor
        self.high_slider.valueChanged.connect(lambda val, lbl=self.high_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(high_contrast_label, 0, 0)
        img_adjust_layout.addWidget(self.high_slider, 0, 1)
        img_adjust_layout.addWidget(self.high_value_label, 0, 2)

        low_contrast_label = QLabel("Contrast:") # Renamed for clarity
        self.low_slider = QSlider(Qt.Horizontal)
        self.low_slider.setRange(0, 200) # Range 0 to 200%
        self.low_slider.setValue(100) # Default 100%
        self.low_slider.valueChanged.connect(self.update_image_contrast)
        self.low_value_label = QLabel("1.00") # Display factor
        self.low_slider.valueChanged.connect(lambda val, lbl=self.low_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(low_contrast_label, 1, 0)
        img_adjust_layout.addWidget(self.low_slider, 1, 1)
        img_adjust_layout.addWidget(self.low_value_label, 1, 2)


        # Gamma Adjustment Slider
        gamma_label = QLabel("Gamma:")
        self.gamma_slider = QSlider(Qt.Horizontal)
        self.gamma_slider.setRange(10, 300)  # Range 0.1 to 3.0
        self.gamma_slider.setValue(100)  # Default 1.0 (100%)
        self.gamma_slider.valueChanged.connect(self.update_image_gamma)
        self.gamma_value_label = QLabel("1.00") # Display factor
        self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(gamma_label, 2, 0)
        img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
        img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


        # Separator
        img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

        # Action Buttons
        btn_layout = QHBoxLayout()
        self.bw_button = QPushButton("Grayscale")
        self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
        self.bw_button.clicked.connect(self.convert_to_black_and_white)
        invert_button = QPushButton("Invert")
        invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
        invert_button.clicked.connect(self.invert_image)
        reset_button = QPushButton("Reset Adjustments")
        reset_button.setToolTip("Reset Brightness, Contrast, and Gamma sliders to default.\n(Does not undo Grayscale or Invert)")
        reset_button.clicked.connect(self.reset_gamma_contrast)
        btn_layout.addWidget(self.bw_button)
        btn_layout.addWidget(invert_button)
        btn_layout.addStretch() # Push reset button to the right
        btn_layout.addWidget(reset_button)

        img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

        layout.addWidget(img_adjust_group)
        layout.addStretch() # Push groups up
        return tab

    
    def _update_color_button_style(self, button, color):
        """ Helper to update button background color preview """
        if color.isValid():
            button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
        else:
            button.setStyleSheet("") # Reset to default stylesheet
    
    def create_separator(self):
        """Creates a horizontal separator line."""
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        return line
    
    
    def create_cropping_tab(self):
        """Create the Cropping tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
    
        # Group Box for Alignment Options
        alignment_params_group = QGroupBox("Alignment Options")
        alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        alignment_layout = QVBoxLayout()
    
        
        #Guide Lines
        self.show_guides_label = QLabel("Show Guide Lines")
        self.show_guides_checkbox = QCheckBox("", self)
        self.show_guides_checkbox.setChecked(False)
        self.show_guides_checkbox.stateChanged.connect(self.update_live_view)
        
        # Rotation Angle 
        rotation_layout = QHBoxLayout()
        self.orientation_label = QLabel("Rotation Angle (0.00°)")
        self.orientation_label.setFixedWidth(150)
        self.orientation_slider = QSlider(Qt.Horizontal)
        self.orientation_slider.setRange(-3600, 3600)  # Scale by 10 to allow decimals
        self.orientation_slider.setValue(0)
        self.orientation_slider.setSingleStep(1)
        self.orientation_slider.valueChanged.connect(self.update_live_view)
        # Align Button
        self.align_button = QPushButton("Apply Rotation")
        self.align_button.clicked.connect(self.align_image)
        
        self.reset_align_button = QPushButton("Reset Rotation")
        self.reset_align_button.clicked.connect(self.reset_align_image)
        
        rotation_layout.addWidget(self.show_guides_label)
        rotation_layout.addWidget(self.show_guides_checkbox)
        rotation_layout.addWidget(self.orientation_label)
        rotation_layout.addWidget(self.orientation_slider)
        rotation_layout.addWidget(self.align_button)
        rotation_layout.addWidget(self.reset_align_button)      
        
        
    
        
        
        # Flip Vertical Button
        self.flip_vertical_button = QPushButton("Flip Vertical")
        self.flip_vertical_button.clicked.connect(self.flip_vertical)
    
        # Flip Horizontal Button
        self.flip_horizontal_button = QPushButton("Flip Horizontal")
        self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
    
        alignment_layout.addLayout(rotation_layout)
        
        alignment_layout.addWidget(self.flip_vertical_button)  
        alignment_layout.addWidget(self.flip_horizontal_button)  
        alignment_params_group.setLayout(alignment_layout)
        
        
       # Add Tapering Skew Fix Group
        taper_skew_group = QGroupBox("Skew Fix")
        taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        taper_skew_layout = QHBoxLayout()
    
        # Taper Skew Slider
        self.taper_skew_label = QLabel("Tapering Skew (0.00)")
        self.taper_skew_label.setFixedWidth(150)
        self.taper_skew_slider = QSlider(Qt.Horizontal)
        self.taper_skew_slider.setRange(-70, 70)  # Adjust as needed
        self.taper_skew_slider.setValue(0)
        self.taper_skew_slider.valueChanged.connect(self.update_live_view)
        
        # Align Button
        self.skew_button = QPushButton("Apply Skew")
        self.skew_button.clicked.connect(self.update_skew)
    
        # Add widgets to taper skew layout
        taper_skew_layout.addWidget(self.taper_skew_label)
        taper_skew_layout.addWidget(self.taper_skew_slider)
        taper_skew_layout.addWidget(self.skew_button)
        taper_skew_group.setLayout(taper_skew_layout)
        
        # Group Box for Cropping Options
        cropping_params_group = QGroupBox("Cropping Options")
        cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        cropping_layout = QVBoxLayout()
    
        # Cropping Sliders
        crop_slider_layout = QGridLayout()
    
        crop_x_start_label = QLabel("Crop Left (%)")
        self.crop_x_start_slider = QSlider(Qt.Horizontal)
        self.crop_x_start_slider.setRange(0, 100)
        self.crop_x_start_slider.setValue(0)
        self.crop_x_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
        crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
    
        crop_x_end_label = QLabel("Crop Right (%)")
        self.crop_x_end_slider = QSlider(Qt.Horizontal)
        self.crop_x_end_slider.setRange(0, 100)
        self.crop_x_end_slider.setValue(100)
        self.crop_x_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_end_label, 0, 2)
        crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 3)
    
        crop_y_start_label = QLabel("Crop Top (%)")
        self.crop_y_start_slider = QSlider(Qt.Horizontal)
        self.crop_y_start_slider.setRange(0, 100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
        crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
    
        crop_y_end_label = QLabel("Crop Bottom (%)")
        self.crop_y_end_slider = QSlider(Qt.Horizontal)
        self.crop_y_end_slider.setRange(0, 100)
        self.crop_y_end_slider.setValue(100)
        self.crop_y_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_end_label, 1, 2)
        crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 3)
    
        cropping_layout.addLayout(crop_slider_layout)
    
        # Crop Update Button
        crop_button = QPushButton("Update Crop")
        crop_button.clicked.connect(self.update_crop)
        cropping_layout.addWidget(crop_button)
    
        cropping_params_group.setLayout(cropping_layout)
    
        # Add both group boxes to the main layout
        layout.addWidget(alignment_params_group)
        layout.addWidget(cropping_params_group)
        layout.addWidget(taper_skew_group)
        layout.addStretch()
    
        return tab
    
    def create_white_space_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # --- Padding Group ---
        padding_params_group = QGroupBox("Add White Space (Padding)")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setSpacing(8)


        # Input fields with validation (optional but good)
        from PyQt5.QtGui import QIntValidator
        int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

        # Left Padding
        left_padding_label = QLabel("Left Padding (px):")
        self.left_padding_input = QLineEdit("100") # Default
        self.left_padding_input.setValidator(int_validator)
        self.left_padding_input.setToolTip("Pixels to add to the left.")
        padding_layout.addWidget(left_padding_label, 0, 0)
        padding_layout.addWidget(self.left_padding_input, 0, 1)

        # Right Padding
        right_padding_label = QLabel("Right Padding (px):")
        self.right_padding_input = QLineEdit("100") # Default
        self.right_padding_input.setValidator(int_validator)
        self.right_padding_input.setToolTip("Pixels to add to the right.")
        padding_layout.addWidget(right_padding_label, 0, 2)
        padding_layout.addWidget(self.right_padding_input, 0, 3)

        # Top Padding
        top_padding_label = QLabel("Top Padding (px):")
        self.top_padding_input = QLineEdit("100") # Default
        self.top_padding_input.setValidator(int_validator)
        self.top_padding_input.setToolTip("Pixels to add to the top.")
        padding_layout.addWidget(top_padding_label, 1, 0)
        padding_layout.addWidget(self.top_padding_input, 1, 1)

        # Bottom Padding
        bottom_padding_label = QLabel("Bottom Padding (px):")
        self.bottom_padding_input = QLineEdit("0") # Default
        self.bottom_padding_input.setValidator(int_validator)
        self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
        padding_layout.addWidget(bottom_padding_label, 1, 2)
        padding_layout.addWidget(self.bottom_padding_input, 1, 3)

        layout.addWidget(padding_params_group)

        # --- Buttons Layout ---
        button_layout = QHBoxLayout()
        self.recommend_button = QPushButton("Set Recommended Values")
        self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
        self.recommend_button.clicked.connect(self.recommended_values)

        self.clear_padding_button = QPushButton("Clear Values")
        self.clear_padding_button.setToolTip("Set all padding values to 0.")
        self.clear_padding_button.clicked.connect(self.clear_padding_values)

        self.finalize_button = QPushButton("Apply Padding")
        self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
        self.finalize_button.clicked.connect(self.finalize_image)

        button_layout.addWidget(self.recommend_button)
        button_layout.addWidget(self.clear_padding_button)
        button_layout.addStretch()
        button_layout.addWidget(self.finalize_button)

        layout.addLayout(button_layout)
        layout.addStretch() # Push content up

        return tab
    def clear_padding_values(self):
        self.bottom_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.left_padding_input.setText("0")
        
    def recommended_values(self):
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        
        
        
    def create_markers_tab(self):
        """Create the Markers tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        marker_options_layout = QHBoxLayout()
    
        # Left/Right Marker Options
        left_right_marker_group = QGroupBox("Left/Right Marker Options")
        left_right_marker_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        left_right_marker_layout = QVBoxLayout()
        
        # Combo box for marker presets or custom option
        self.combo_box = QComboBox(self)
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
        self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
        
        # Textbox to allow modification of marker values (shown when "Custom" is selected)
        self.marker_values_textbox = QLineEdit(self)
        self.marker_values_textbox.setPlaceholderText("Enter custom values as comma-separated list")
        self.marker_values_textbox.setEnabled(False)  # Disable by default
        
        # Rename input for custom option
        self.rename_input = QLineEdit(self)
        self.rename_input.setPlaceholderText("Enter new name for Custom")
        self.rename_input.setEnabled(False)
        
        # Save button for the current configuration
        self.save_button = QPushButton("Save Config", self)
        self.save_button.clicked.connect(self.save_config)
        
        # Delete button 
        self.remove_config_button = QPushButton("Remove Config", self)
        self.remove_config_button.clicked.connect(self.remove_config)
        
        # Add widgets to the Left/Right Marker Options layout
        left_right_marker_layout.addWidget(self.combo_box)
        left_right_marker_layout.addWidget(self.marker_values_textbox)
        left_right_marker_layout.addWidget(self.rename_input)
        left_right_marker_layout.addWidget(self.save_button)
        left_right_marker_layout.addWidget(self.remove_config_button)
        
        # Set layout for the Left/Right Marker Options group
        left_right_marker_group.setLayout(left_right_marker_layout)
        
        # Top Marker Options
        top_marker_group = QGroupBox("Top/Bottom Marker Options")
        top_marker_group.setStyleSheet("QGroupBox { font-weight: bold;}")
        
        # Vertical layout for top marker group
        top_marker_layout = QVBoxLayout()
        
        # Text input for Top Marker Labels (multi-column support)
        self.top_marker_input = QTextEdit(self)
        self.top_marker_input.setText(", ".join(self.top_label))  # Populate with initial values
        self.top_marker_input.setMinimumHeight(50)  # Increase height for better visibility
        self.top_marker_input.setMaximumHeight(120)
        self.top_marker_input.setPlaceholderText("Enter labels for each column, separated by commas. Use new lines for multiple columns.")
        
        # # Button to add a new column
        # self.add_column_button = QPushButton("Add Column")
        # self.add_column_button.clicked.connect(self.add_column)
        
        # # Button to remove the last column
        # self.remove_column_button = QPushButton("Remove Last Column")
        # self.remove_column_button.clicked.connect(self.remove_column)
        
        # # Button to update all labels
        self.update_top_labels_button = QPushButton("Update All Labels")
        self.update_top_labels_button.clicked.connect(self.update_all_labels)
        
        # Layout for column management buttons
        # column_buttons_layout = QHBoxLayout()
        # column_buttons_layout.addWidget(self.add_column_button)
        # column_buttons_layout.addWidget(self.remove_column_button)
        
        # Add widgets to the top marker layout
        top_marker_layout.addWidget(self.top_marker_input)
        # top_marker_layout.addLayout(column_buttons_layout)
        top_marker_layout.addWidget(self.update_top_labels_button)
        
        # Set the layout for the Top Marker Group
        top_marker_group.setLayout(top_marker_layout)
        
        # Add both groups to the horizontal layout
        marker_options_layout.addWidget(left_right_marker_group)
        marker_options_layout.addWidget(top_marker_group)
        
        # Add the horizontal layout to the main layout
        layout.addLayout(marker_options_layout)
    
        # Marker padding sliders - Group box for marker distance adjustment
        padding_params_group = QGroupBox("Marker Placement and Offsets")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        
        # Grid layout for the marker group box
        padding_layout = QGridLayout()
        
        # Left marker: Button, slider, reset, and duplicate in the same row
        left_marker_button = QPushButton("Place Left Markers")
        left_marker_button.setToolTip("Places the left markers at the exact location of the mouse pointer on the left. Shortcut: Ctrl+Shift+L or CMD+Shift+L")
        left_marker_button.clicked.connect(self.enable_left_marker_mode)
        self.left_padding_slider = QSlider(Qt.Horizontal)
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.left_padding_slider.setValue(0)
        self.left_padding_slider.valueChanged.connect(self.update_left_padding)
        
        remove_left_button = QPushButton("Remove Last")
        remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
        
        reset_left_button = QPushButton("Reset")
        reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
        
        duplicate_left_button = QPushButton("Duplicate Right")
        duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
        
        # Add left marker widgets to the grid layout
        padding_layout.addWidget(left_marker_button, 0, 0)
        padding_layout.addWidget(remove_left_button, 0, 1)
        padding_layout.addWidget(reset_left_button, 0, 2)
        padding_layout.addWidget(self.left_padding_slider, 0, 3,1,2)
        padding_layout.addWidget(duplicate_left_button, 0, 5)
        
        
        # Right marker: Button, slider, reset, and duplicate in the same row
        right_marker_button = QPushButton("Place Right Markers")
        right_marker_button.setToolTip("Places the right markers at the exact location of the mouse pointer on the right. Shortcut: Ctrl+Shift+R or CMD+Shift+R")
        right_marker_button.clicked.connect(self.enable_right_marker_mode)
        self.right_padding_slider = QSlider(Qt.Horizontal)
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.right_padding_slider.setValue(0)
        self.right_padding_slider.valueChanged.connect(self.update_right_padding)
        
        remove_right_button = QPushButton("Remove Last")
        remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
        
        reset_right_button = QPushButton("Reset")
        reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
        
        duplicate_right_button = QPushButton("Duplicate Left")
        duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
        
        # Add right marker widgets to the grid layout
        padding_layout.addWidget(right_marker_button, 1, 0)
        padding_layout.addWidget(remove_right_button, 1, 1)
        padding_layout.addWidget(reset_right_button, 1, 2)        
        padding_layout.addWidget(self.right_padding_slider, 1, 3,1,2)
        padding_layout.addWidget(duplicate_right_button, 1, 5)
        
        # Top marker: Button, slider, and reset in the same row
        top_marker_button = QPushButton("Place Top Markers")
        top_marker_button.setToolTip("Places the top markers at the exact location of the mouse pointer on the top. Shortcut: Ctrl+Shift+T or CMD+Shift+T")
        top_marker_button.clicked.connect(self.enable_top_marker_mode)
        self.top_padding_slider = QSlider(Qt.Horizontal)
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.top_padding_slider.setValue(0)
        self.top_padding_slider.valueChanged.connect(self.update_top_padding)
        
        remove_top_button = QPushButton("Remove Last")
        remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
        
        reset_top_button = QPushButton("Reset")
        reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
        
        # Add top marker widgets to the grid layout
        padding_layout.addWidget(top_marker_button, 2, 0)
        padding_layout.addWidget(remove_top_button, 2, 1)
        padding_layout.addWidget(reset_top_button, 2, 2)
        padding_layout.addWidget(self.top_padding_slider, 2, 3, 1, 2)  # Slider spans 2 columns for better alignment
        
        for i in range(6):  # Assuming 6 columns in the grid
            padding_layout.setColumnStretch(i, 1)
        
        # Add button and QLineEdit for the custom marker
        self.custom_marker_button = QPushButton("Place Custom Marker", self)
        self.custom_marker_button.setToolTip("Places custom markers at the middle of the mouse pointer")
    
        self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
        
        self.custom_marker_button_left_arrow = QPushButton("←", self)
        
        self.custom_marker_button_right_arrow = QPushButton("→", self)
        
        self.custom_marker_button_top_arrow = QPushButton("↑", self)
        
        self.custom_marker_button_bottom_arrow = QPushButton("↓", self)
        
        self.custom_marker_text_entry = QLineEdit(self)        
        self.custom_marker_text_entry.setPlaceholderText("Enter custom marker text")
        
        self.remove_custom_marker_button = QPushButton("Remove Last", self)
        self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
        
        self.reset_custom_marker_button = QPushButton("Reset", self)
        self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
        
        # Add color selection button for custom markers
        self.custom_marker_color_button = QPushButton("Custom Marker Color")
        self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
        
        marker_buttons_layout = QHBoxLayout()
        
        # Add the arrow buttons with fixed sizes to the marker buttons layout
        self.custom_marker_button_left_arrow.setFixedSize(30, 30)
    
        marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow)
        
        self.custom_marker_button_right_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
        
        self.custom_marker_button_top_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow)
        
        self.custom_marker_button_bottom_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
        
        #Assign functions to the buttons
        self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))
    
        
        # Create a QWidget to hold the QHBoxLayout
        marker_buttons_widget = QWidget()
        marker_buttons_widget.setLayout(marker_buttons_layout)
        
        # Add the custom marker button
        padding_layout.addWidget(self.custom_marker_button, 3, 0)
        
        # Add the text entry for the custom marker
        padding_layout.addWidget(self.custom_marker_text_entry, 3, 1,1,1)
        
        # Add the marker buttons widget to the layout
        padding_layout.addWidget(marker_buttons_widget, 3, 2) 
        
        # Add the remove button
        padding_layout.addWidget(self.remove_custom_marker_button, 3, 3)
        
        # Add the reset button
        padding_layout.addWidget(self.reset_custom_marker_button, 3, 4)
        
        # Add the color button
        padding_layout.addWidget(self.custom_marker_color_button, 3, 5)
        
        self.custom_font_type_label = QLabel("Custom Marker Font:", self)
        self.custom_font_type_dropdown = QFontComboBox()
        self.custom_font_type_dropdown.setCurrentFont(QFont("Arial"))
        self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
        
        # Font size selector
        self.custom_font_size_label = QLabel("Custom Marker Size:", self)
        self.custom_font_size_spinbox = QSpinBox(self)
        self.custom_font_size_spinbox.setRange(2, 150)  # Allow font sizes from 8 to 72
        self.custom_font_size_spinbox.setValue(12)  # Default font size
        
        # Grid checkbox
        self.show_grid_checkbox = QCheckBox("Show Snap Grid", self)
        self.show_grid_checkbox.setToolTip("Places a snapping grid and the text or marker will be places at the center of the grid. Shortcut: Ctrl+Shift+G. To increase or decrease the grid size: Ctrl+Shift+Up or Down arrow or CMD+Shift+Up or Down arrow ")
        self.show_grid_checkbox.setChecked(False)  # Default: Grid is off
        self.show_grid_checkbox.stateChanged.connect(self.update_live_view)
        
        # Grid size input (optional
        self.grid_size_input = QSpinBox(self)
        self.grid_size_input.setRange(5, 100)  # Grid cell size in pixels
        self.grid_size_input.setValue(20)  # Default grid size
        self.grid_size_input.setPrefix("Grid Size (px): ")
        self.grid_size_input.valueChanged.connect(self.update_live_view)
        
        # Add font type and size widgets to the layout
        padding_layout.addWidget(self.custom_font_type_label, 4, 0,)
        padding_layout.addWidget(self.custom_font_type_dropdown, 4, 1,1,1)  # Span 2 columns
        
        padding_layout.addWidget(self.custom_font_size_label, 4, 2)
        padding_layout.addWidget(self.custom_font_size_spinbox, 4, 3,1,1)
        padding_layout.addWidget(self.show_grid_checkbox,4,4)
        padding_layout.addWidget(self.grid_size_input,4,5)
    
        # Set the layout for the marker group box
        padding_params_group.setLayout(padding_layout)
        
        # Add the font options group box to the main layout
        layout.addWidget(padding_params_group)
        
        layout.addStretch()
        return tab
    
    def add_column(self):
        """Add a new column to the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        if current_text.strip():
            self.top_marker_input.append("")  # Add a new line for a new column
        else:
            self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
    
    def remove_column(self):
        """Remove the last column from the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        lines = current_text.split("\n")
        if len(lines) > 1:
            lines.pop()  # Remove the last line
            self.top_marker_input.setPlainText("\n".join(lines))
        else:
            self.top_marker_input.clear()  # Clear the text if only one line exists

    
    def flip_vertical(self):
        self.save_state()
        """Flip the image vertically."""
        if self.image:
            self.image = self.image.mirrored(vertical=True, horizontal=False)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def flip_horizontal(self):
        self.save_state()
        """Flip the image horizontally."""
        if self.image:
            self.image = self.image.mirrored(vertical=False, horizontal=True)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def convert_to_black_and_white(self):
        self.save_state()
        """Convert the image to black and white."""
        if self.image:
            grayscale_image = self.image.convertToFormat(QImage.Format_Grayscale8)
            self.image = grayscale_image
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()


    def invert_image(self):
        self.save_state()
        if self.image:
            inverted_image = self.image.copy()
            inverted_image.invertPixels()
            self.image = inverted_image
            self.update_live_view()
        self.image_before_contrast=self.image.copy()
        self.image_before_padding=self.image.copy()
        self.image_contrasted=self.image.copy()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if self.live_view_label.preview_marker_enabled:
                self.live_view_label.preview_marker_enabled = False  # Turn off the preview
            self.live_view_label.measure_quantity_mode = False
            self.live_view_label.bounding_box_complete==False
            self.live_view_label.counter=0
            self.live_view_label.quad_points=[]
            self.live_view_label.bounding_box_preview = None
            self.live_view_label.rectangle_points = []
            self.live_view_label.mousePressEvent=None
            self.live_view_label.mode=None

        if self.live_view_label.zoom_level != 1.0:  # Only allow panning when zoomed in
            step = 20  # Adjust the panning step size as needed
            if event.key() == Qt.Key_Left:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - step)
            elif event.key() == Qt.Key_Right:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + step)
            elif event.key() == Qt.Key_Up:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - step)
            elif event.key() == Qt.Key_Down:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + step)
        self.update_live_view()
        super().keyPressEvent(event)
    
    def update_marker_text_font(self, font: QFont):
        """
        Updates the font of self.custom_marker_text_entry based on the selected font
        from self.custom_font_type_dropdown.
    
        :param font: QFont object representing the selected font from the combobox.
        """
        # Get the name of the selected font
        selected_font_name = font.family()
        
        # Set the font of the QLineEdit
        self.custom_marker_text_entry.setFont(QFont(selected_font_name))
    
    def arrow_marker(self, text: str):
        self.custom_marker_text_entry.clear()
        self.custom_marker_text_entry.setText(text)
    
    def enable_custom_marker_mode(self):
        """Enable the custom marker mode and set the mouse event."""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        custom_text = self.custom_marker_text_entry.text().strip()
    
        self.live_view_label.preview_marker_enabled = True
        self.live_view_label.preview_marker_text = custom_text
        self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
        self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
        self.live_view_label.marker_color=self.custom_marker_color
        
        self.live_view_label.setFocus()
        self.live_view_label.update()
        
        self.marker_mode = "custom"  # Indicate custom marker mode
        self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
        
    def remove_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers.pop()  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    def reset_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers=[]  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    
    def place_custom_marker(self, event, custom_text):
        """Place a custom marker at the cursor location."""
        self.save_state()
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            
        # Adjust for snapping to grid
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        image_x = (cursor_x - x_offset) / scale
        image_y = (cursor_y - y_offset) / scale
        
            
        # Store the custom marker's position and text
        self.custom_markers = getattr(self, "custom_markers", [])
        self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value()))
        # print("CUSTOM_MARKER: ",self.custom_markers)
        # Update the live view to render the custom marker
        self.update_live_view()
        
    def select_custom_marker_color(self):
        """Open a color picker dialog to select the color for custom markers."""
        color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
        if color.isValid():
            self.custom_marker_color = color  # Update the custom marker color
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
    
    def update_all_labels(self):
        """Update all labels from the QTextEdit, supporting multiple columns."""
        self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
        input_text = self.top_marker_input.toPlainText()
        
        # Split the text into a list by commas and strip whitespace
        self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
        

        # if len(self.top_label) < len(self.top_markers):
        #     self.top_markers = self.top_markers[:len(self.top_label)]
        for i in range(0, len(self.top_markers)):
            try:
                self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
            except:
                self.top_markers[i] = (self.top_markers[i][0], str(""))

        # if len(self.marker_values) < len(self.left_markers):
        #     self.left_markers = self.left_markers[:len(self.marker_values)]
        for i in range(0, len(self.left_markers)):
            try:
                self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
            except:
                self.left_markers[i] = (self.left_markers[i][0], str(""))
                

        # if len(self.marker_values) < len(self.right_markers):
        #     self.right_markers = self.right_markers[:len(self.marker_values)]
        for i in range(0, len(self.right_markers)):
            try:
                self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
            except:
                self.right_markers[i] = (self.right_markers[i][0], str(""))
                

        
        # Trigger a refresh of the live view
        self.update_live_view()
        
    def reset_marker(self, marker_type, param):       
        if marker_type == 'left':
            if param == 'remove' and len(self.left_markers)!=0:
                self.left_markers.pop()  
                if self.current_left_marker_index > 0:
                    self.current_left_marker_index -= 1
            elif param == 'reset':
                self.left_markers.clear()
                self.current_left_marker_index = 0  

             
        elif marker_type == 'right' and len(self.right_markers)!=0:
            if param == 'remove':
                self.right_markers.pop()  
                if self.current_right_marker_index > 0:
                    self.current_right_marker_index -= 1
            elif param == 'reset':
                self.right_markers.clear()
                self.current_right_marker_index = 0

        elif marker_type == 'top' and len(self.top_markers)!=0:
            if param == 'remove':
                self.top_markers.pop() 
                if self.current_top_label_index > 0:
                    self.current_top_label_index -= 1
            elif param == 'reset':
                self.top_markers.clear()
                self.current_top_label_index = 0
    
        # Call update live view after resetting markers
        self.update_live_view()
        
    def duplicate_marker(self, marker_type):
        if marker_type == 'left' and self.right_markers:
            self.left_markers = self.right_markers.copy()  # Copy right markers to left
        elif marker_type == 'right' and self.left_markers:
            self.right_markers = self.left_markers.copy()  # Copy left markers to right

            # self.right_padding = self.image_width*0.9
    
        # Call update live view after duplicating markers
        self.update_live_view()
        
    def on_combobox_changed(self):
        
        text=self.combo_box.currentText()
        """Handle the ComboBox change event to update marker values."""
        if text == "Custom":
            # Enable the textbox for custom values when "Custom" is selected
            self.marker_values_textbox.setEnabled(True)
            self.rename_input.setEnabled(True)
            
            # self.marker_values_textbox.setText()
        else:
            # Update marker values based on the preset option
            self.marker_values = self.marker_values_dict.get(text, [])
            self.marker_values_textbox.setEnabled(False)  # Disable the textbox
            self.rename_input.setEnabled(False)
            self.top_label = self.top_label_dict.get(text, [])
            self.top_label = [str(item) if not isinstance(item, str) else item for item in self.top_label]
            self.top_marker_input.setText(", ".join(self.top_label))
            try:
                
                # Ensure that the top_markers list only updates the top_label values serially
                for i in range(0, len(self.top_markers)):
                    self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                
                # # If self.top_label has more entries than current top_markers, add them
                # if len(self.top_label) > len(self.top_markers):
                #     additional_markers = [(self.top_markers[-1][0] + 50 * (i + 1), label) 
                #                           for i, label in enumerate(self.top_label[len(self.top_markers):])]
                #     self.top_markers.extend(additional_markers)
                
                # If self.top_label has fewer entries, truncate the list
                if len(self.top_label) < len(self.top_markers):
                    self.top_markers = self.top_markers[:len(self.top_label)]
                    
                
                
            except:
                pass
            try:
                self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
            except:
                pass

    # Functions for updating contrast and gamma
    
    def reset_gamma_contrast(self):
        try:
            if self.image_before_contrast==None:
                self.image_before_contrast=self.image_master.copy()
            self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
            self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
            self.high_slider.setValue(100)  # Reset contrast to default
            self.low_slider.setValue(0)  # Reset contrast to default
            self.gamma_slider.setValue(100)  # Reset gamma to default
            self.update_live_view()
        except:
            pass

    
    def update_image_contrast(self):
        self.save_state()
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
            
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                self.update_live_view()
        except:
            pass
    
    def update_image_gamma(self):
        self.save_state()
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
                
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                self.update_live_view()
        except:
            pass
    
    def apply_contrast_gamma(self, qimage, high, low, gamma):
        """
        Applies contrast and gamma adjustments to a QImage.
        Converts the QImage to RGBA format, performs the adjustments, and returns the modified QImage.
        """
        #Ensure the image is in the correct format (RGBA8888)
        if qimage.format() != QImage.Format_RGBA8888:
            qimage = qimage.convertToFormat(QImage.Format_RGBA8888)

        # Convert the QImage to a NumPy array
        width = qimage.width()
        height = qimage.height()

        # Get the raw byte data from QImage
        ptr = qimage.bits()
        ptr.setsize(height * width * 4)  # 4 bytes per pixel (RGBA)

        # Create a NumPy array from the pointer (for RGBA format)
        img_array = np.array(ptr).reshape(height, width, 4).astype(np.float32)

        # Normalize the image for contrast and gamma adjustments
        img_array = img_array / 255.0

        # Apply contrast adjustment
        img_array = np.clip((img_array - low) / (high - low), 0, 1)

        # Apply gamma correction
        img_array = np.power(img_array, gamma)

        # Scale back to [0, 255] range
        img_array = (img_array * 255).astype(np.uint8)

        # Convert back to QImage
        qimage = QImage(img_array.data, width, height, img_array.strides[0], QImage.Format_RGBA8888)

        return qimage
    

    def save_contrast_options(self):
        if self.image:
            self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
            self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
        else:
            QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

    def remove_config(self):
        try:
            # Get the currently selected marker label
            selected_marker = self.combo_box.currentText()
    
            # Ensure the selected marker is not "Custom" before deleting
            if selected_marker == "Custom":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' marker.")
                return
            
            elif selected_marker == "Precision Plus All Blue/Unstained":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
            
            elif selected_marker == "1 kB Plus":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
    
            # Remove the marker label and top_label if they exist
            if selected_marker in self.marker_values_dict:
                del self.marker_values_dict[selected_marker]
            if selected_marker in self.top_label_dict:
                del self.top_label_dict[selected_marker]
    
            # Save the updated configuration
            with open("Imaging_assistant_config.txt", "w") as f:
                config = {
                    "marker_values": self.marker_values_dict,
                    "top_label": self.top_label_dict
                }
                json.dump(config, f)
    
            # Remove from the ComboBox and reset UI
            self.combo_box.removeItem(self.combo_box.currentIndex())
            self.top_marker_input.clear()
    
            QMessageBox.information(self, "Success", f"Configuration '{selected_marker}' removed.")
    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error removing config: {e}")

    def save_config(self):
        """Rename the 'Custom' marker values option and save the configuration."""
        new_name = self.rename_input.text().strip()
        
        # Ensure that the correct dictionary (self.marker_values_dict) is being used
        if self.rename_input.text() != "Enter new name for Custom" and self.rename_input.text() != "":  # Correct condition
            # Save marker values list
            self.marker_values_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
            
            # Save top_label list under the new_name key
            self.top_label_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.top_marker_input.toPlainText().strip("[]").split(",")]

            try:
                # Save both the marker values and top label (under new_name) to the config file
                with open("Imaging_assistant_config.txt", "w") as f:
                    config = {
                        "marker_values": self.marker_values_dict,
                        "top_label": self.top_label_dict  # Save top_label_dict as a dictionary with new_name as key
                    }
                    json.dump(config, f)  # Save both marker_values and top_label_dict
            except Exception as e:
                print(f"Error saving config: {e}")

        self.combo_box.setCurrentText(new_name)
        self.load_config()  # Reload the configuration after saving
    
    
    def load_config(self):
        """Load the configuration from the file."""
        try:
            with open("Imaging_assistant_config.txt", "r") as f:
                config = json.load(f)
                
                # Load marker values and top label from the file
                self.marker_values_dict = config.get("marker_values", {})
                
                # Retrieve top_label list from the dictionary using the new_name key
                new_name = self.rename_input.text().strip()  # Assuming `new_name` is defined here; otherwise, set it manually
                self.top_label_dict = config.get("top_label", {})  # Default if not found
                # self.top_label = self.top_label_dict.get(new_name, ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])  # Default if not found
                
        except FileNotFoundError:
            # Set default marker values and top_label if config is not found
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }  # Default top labels
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
        except Exception as e:
            print(f"Error loading config: {e}")
            # Fallback to default values if there is an error
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
    
        # Update combo box options with loaded marker values
        self.combo_box.clear()
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
    
    def paste_image(self):
        """Handle pasting image from clipboard."""
        self.load_image_from_clipboard()
        self.update_live_view()
    
    def load_image_from_clipboard(self):
        """Load an image from the clipboard into self.image."""
        self.is_modified=True
        self.reset_image()
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
    
        # Check if the clipboard contains an image
        if mime_data.hasImage():
            image = clipboard.image()  # Get the image directly from the clipboard
            self.image = image  # Store the image in self.image
            self.original_image = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_master = self.image.copy()
            self.image_before_padding = None
            self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}")

    
        # Check if the clipboard contains URLs (file paths)
        if mime_data.hasUrls():
            urls = mime_data.urls()
            if urls:
                file_path = urls[0].toLocalFile()  # Get the first file path
                if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                    # Load the image from the file path
                    self.image = QImage(file_path)
                    self.original_image = self.image.copy()
                    self.image_contrasted = self.image.copy()
                    self.image_master = self.image.copy()
                    self.image_before_contrast = self.image.copy()
                    self.image_before_padding = None
                    
    
                    # Update the window title with the image path
                    self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}:{file_path}")
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
       
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        self.save_state()
        
    def update_font(self):
        """Update the font settings based on UI inputs"""
        # Update font family from the combo box
        self.font_family = self.font_combo_box.currentFont().family()
        
        # Update font size from the spin box
        self.font_size = self.font_size_spinner.value()
        
        self.font_rotation = int(self.font_rotation_input.value())
        
    
        # Once font settings are updated, update the live view immediately
        self.update_live_view()
    
    def select_font_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.font_color = color  # Store the selected color   
            self.update_font()
        self._update_color_button_style(self.font_color_button, self.font_color)
            
    def load_image(self):
        self.is_modified=True
        self.undo_stack = []  # Reset the undo stack
        self.redo_stack = []  # Reset the redo stack
        self.reset_image()
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif)", options=options
        )
        if file_path:
            self.image_path = file_path
            self.original_image = QImage(self.image_path)  # Keep the original image
            self.image = self.original_image.copy()        # Start with a copy of the original
            self.image_master= self.original_image.copy() 
            self.image_before_padding=None
            self.image_before_contrast=self.original_image.copy() 
            self.image_contrasted= self.original_image.copy()  

            self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}:{self.image_path}")
    
            # Determine associated config file
            self.base_name = os.path.splitext(os.path.basename(file_path))[0]
            if self.base_name.endswith("_original"):
                config_name = self.base_name.replace("_original", "_config.txt")
            else:
                config_name = self.base_name + "_config.txt"
            config_path = os.path.join(os.path.dirname(file_path), config_name)
    
            if os.path.exists(config_path):
                try:
                    with open(config_path, "r") as config_file:
                        config_data = json.load(config_file)
                    self.apply_config(config_data)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to load config file: {e}")
        # Preview window
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
        
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        self.save_state()
    
    def apply_config(self, config_data):
        self.left_padding_input.setText(config_data["adding_white_space"]["left"])
        self.right_padding_input.setText(config_data["adding_white_space"]["right"])
        self.top_padding_input.setText(config_data["adding_white_space"]["top"])
        try:
            self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
        except:
            pass
    
        self.crop_x_start_slider.setValue(config_data["cropping_parameters"]["x_start_percent"])
        self.crop_x_end_slider.setValue(config_data["cropping_parameters"]["x_end_percent"])
        self.crop_y_start_slider.setValue(config_data["cropping_parameters"]["y_start_percent"])
        self.crop_y_end_slider.setValue(config_data["cropping_parameters"]["y_end_percent"])
    
        try:
            self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
            self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
            self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
        except (KeyError, ValueError) as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
            pass
        try:
            # print("TOP LABELS: ",config_data["marker_labels"]["top"])
            self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
            self.top_marker_input.setText(", ".join(self.top_label))
        except KeyError as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
            pass
    
        self.font_family = config_data["font_options"]["font_family"]
        self.font_size = config_data["font_options"]["font_size"]
        self.font_rotation = config_data["font_options"]["font_rotation"]
        self.font_color = QColor(config_data["font_options"]["font_color"])
    
        self.top_padding_slider.setValue(config_data["marker_padding"]["top"])
        self.left_padding_slider.setValue(config_data["marker_padding"]["left"])
        self.right_padding_slider.setValue(config_data["marker_padding"]["right"])
        
        try:
            try:
                x = list(map(int, config_data["marker_labels"]["left"]))
                self.marker_values_textbox.setText(str(x))
            except:
                x = list(map(int, config_data["marker_labels"]["right"]))
                self.marker_values_textbox.setText(str(x))
            if self.marker_values_textbox.text!="":
                self.combo_box.setCurrentText("Custom")
                self.marker_values_textbox.setEnabled(True)                
            else:
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.marker_values_textbox.setEnabled(False)
                
        except:
            # print("ERROR IN LEFT/RIGHT MARKER DATA")
            pass
    
        try:
            self.custom_markers = [
                (marker["x"], marker["y"], marker["text"], QColor(marker["color"]), marker["font"], marker["font_size"])
                for marker in config_data.get("custom_markers", [])
            ]
                
                
                
        except:
            pass
        # Set slider ranges from config_data
        try:
            self.left_padding_slider.setRange(
                int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
            )
            self.right_padding_slider.setRange(
                int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
            )
            self.top_padding_slider.setRange(
                int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
            )
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Slider ranges not found in config_data.")
            pass
        
        try:
            self.left_marker_shift_added=int(config_data["added_shift"]["left"])
            self.right_marker_shift_added=int(config_data["added_shift"]["right"])
            self.top_marker_shift_added=int(config_data["added_shift"]["top"])
            
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Added Shift not found in config_data.");
            pass
            
        # Apply font settings

        
        #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
        
        font_size_new=self.font_size
        font_rotation_new=self.font_rotation
        
        self.font_combo_box.setCurrentFont(QFont(self.font_family))
        self.font_size_spinner.setValue(font_size_new)
        self.font_rotation_input.setValue(font_rotation_new)

        self.update_live_view()
        
    def get_current_config(self):
        config = {
            "adding_white_space": {
                "left": self.left_padding_input.text(),
                "right": self.right_padding_input.text(),
                "top": self.top_padding_input.text(),
                "bottom": self.bottom_padding_input.text(),
            },
            "cropping_parameters": {
                "x_start_percent": self.crop_x_start_slider.value(),
                "x_end_percent": self.crop_x_end_slider.value(),
                "y_start_percent": self.crop_y_start_slider.value(),
                "y_end_percent": self.crop_y_end_slider.value(),
            },
            "marker_positions": {
                "left": self.left_markers,
                "right": self.right_markers,
                "top": self.top_markers,
            },
            "marker_labels": {
                "top": self.top_label,
                "left": [marker[1] for marker in self.left_markers],
                "right": [marker[1] for marker in self.right_markers],
            },
            "marker_padding": {
                "top": self.top_padding_slider.value(),
                "left": self.left_padding_slider.value(),
                "right": self.right_padding_slider.value(),
            },
            "font_options": {
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_rotation": self.font_rotation,
                "font_color": self.font_color.name(),
            },
        }
    
        try:
            # Add custom markers with font and font size
            config["custom_markers"] = [
                {"x": x, "y": y, "text": text, "color": color.name(), "font": font, "font_size": font_size}
                for x, y, text, color, font, font_size in self.custom_markers
            ]
        except AttributeError:
            # Handle the case where self.custom_markers is not defined or invalid
            config["custom_markers"] = []
        try:
            config["slider_ranges"] = {
                    "left": self.left_slider_range,
                    "right": self.right_slider_range,
                    "top": self.top_slider_range,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["slider_ranges"] = []
            
        try:
            config["added_shift"] = {
                    "left": self.left_marker_shift_added,
                    "right": self.right_marker_shift_added,
                    "top": self.top_marker_shift_added,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["added_shift"] = []
    
        return config
    
    def add_band(self, event):
        # --- Ensure internal lists match UI input BEFORE adding band ---
        # This call ensures self.marker_values and self.top_label are
        # updated based on the current text in the UI boxes.
        self.update_all_labels()
        # -------------------------------------------------------------

        self.save_state()
        # Ensure there's an image loaded and marker mode is active
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        if not self.image or not self.marker_mode:
            return

        # --- Get Coordinates and Scaling (Same as previous version) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size

        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image.width() > 0 else 1
        image_height = self.image.height() if self.image.height() > 0 else 1
        uniform_scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1
        offset_x = (displayed_width - image_width * uniform_scale) / 2
        offset_y = (displayed_height - image_height * uniform_scale) / 2
        image_x = (cursor_x - offset_x) / uniform_scale if uniform_scale != 0 else 0
        image_y = (cursor_y - offset_y) / uniform_scale if uniform_scale != 0 else 0
        
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)

        # --- Get Render Info for Slider Positioning (Same as previous version) ---
        render_scale = 3
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale

        # --- Validate Coordinates (Same as previous version) ---
        current_image_width = self.image.width()
        current_image_height = self.image.height()
        if not (0 <= image_y <= current_image_height) and self.marker_mode in ["left", "right"]:
             return
        if not (0 <= image_x <= current_image_width) and self.marker_mode == "top":
             return
        # --- End Coordinate/Scaling/Validation ---

        try:
            # --- Left Marker Logic ---
            if self.marker_mode == "left":
                # Determine label based on current count and *now updated* self.marker_values
                current_marker_count = len(self.left_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        print(f"Warning: Adding left marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.left_markers.append((image_y, marker_value_to_add))
                self.current_left_marker_index += 1 # Still increment conceptual index

                # Set slider position only for the *very first* marker placed
                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.left_padding_slider.setValue(0)
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.left_padding_slider.setValue(padding_value)
                    self.left_marker_shift_added = self.left_padding_slider.value()    

            # --- Right Marker Logic ---
            elif self.marker_mode == "right":
                current_marker_count = len(self.right_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        print(f"Warning: Adding right marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.right_markers.append((image_y, marker_value_to_add))
                self.current_right_marker_index += 1

                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.right_padding_slider.setValue(0)
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.right_padding_slider.setValue(padding_value)
                    self.right_marker_shift_added = self.right_padding_slider.value()

            # --- Top Marker Logic ---
            elif self.marker_mode == "top":
                current_marker_count = len(self.top_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.top_label list, which was just updated
                if current_marker_count < len(self.top_label):
                    label_to_add = self.top_label[current_marker_count]
                else:
                    label_to_add = ""
                    if current_marker_count == len(self.top_label): # Print warning only once
                         print(f"Warning: Adding top marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.top_markers.append((image_x, label_to_add))
                self.current_top_label_index += 1

                if is_first_marker:
                    padding_value=int((image_y - y_start) * (render_height / self.image.height()))
                    self.top_padding_slider.setValue(0)
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
                    self.top_padding_slider.setValue(padding_value)
                    self.top_marker_shift_added = self.top_padding_slider.value()

        except Exception as e:
             # Catch other potential errors
             print(f"ERROR ADDING BANDS: An unexpected error occurred - {type(e).__name__}: {e}")
             import traceback
             traceback.print_exc()
             QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

        # Update the live view to render the newly added marker
        self.update_live_view()
        

        
    def enable_left_marker_mode(self):
        self.marker_mode = "left"
        self.current_left_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)

    def enable_right_marker_mode(self):
        self.marker_mode = "right"
        self.current_right_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_top_marker_mode(self):
        self.marker_mode = "top"
        self.current_top_label_index
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
        
    # def remove_padding(self):
    #     if self.image_before_padding!=None:
    #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
    #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
    #     self.image_padded = False  # Reset the padding state
    #     w=self.image.width()
    #     h=self.image.height()
    #     # Preview window
    #     ratio=w/h
    #     self.label_width = 540
    #     label_height=int(self.label_width/ratio)
    #     if label_height>self.label_width:
    #         label_height=540
    #         self.label_width=ratio*label_height
    #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    #     self.update_live_view()
        
    def finalize_image(self):
        self.save_state()
        # Get the padding values from the text inputs
        try:
            padding_left = abs(int(self.left_padding_input.text()))
            padding_right = abs(int(self.right_padding_input.text()))
            padding_top = abs(int(self.top_padding_input.text()))
            padding_bottom = abs(int(self.bottom_padding_input.text()))
        except ValueError:
            # Handle invalid input (non-integer value)
            print("Please enter valid integers for padding.")
            return
    
        # Adjust marker positions based on padding
        self.adjust_markers_for_padding(padding_left, padding_right, padding_top, padding_bottom)
    
        # Ensure self.image_before_padding is initialized
        if self.image_before_padding is None:
            self.image_before_padding = self.image_contrasted.copy()
    
        # Reset to the original image if padding inputs have changed
        if self.image_padded:
            self.image = self.image_before_padding.copy()
            self.image_padded = False
    
        # Calculate the new dimensions with padding
        new_width = self.image.width() + padding_left + padding_right
        new_height = self.image.height() + padding_top + padding_bottom
    
        # Create a new blank image with white background
        padded_image = QImage(new_width, new_height, QImage.Format_RGB888)
        padded_image.fill(QColor(255, 255, 255))  # Fill with white
    
        # Copy the original image onto the padded image
        for y in range(self.image.height()):
            for x in range(self.image.width()):
                color = self.image.pixel(x, y)
                padded_image.setPixel(padding_left + x, padding_top + y, color)
    
        self.image = padded_image
        self.image_padded = True
        self.image_contrasted = self.image.copy()
        w = self.image.width()
        h = self.image.height()
        # Preview window
        ratio = w / h
        self.label_width = int(self.screen_width * 0.28)
        label_height = int(self.label_width / ratio)
        if label_height > self.label_width:
            label_height = self.label_width
            self.label_width = ratio * label_height
        self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    
        # Adjust slider maximum ranges based on the current image width
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range = [-100, int(render_width) + 100]
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.right_slider_range = [-100, int(render_width) + 100]
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.top_slider_range = [-100, int(render_height) + 100]
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.update_live_view()
    
    def adjust_markers_for_padding(self, padding_left, padding_right, padding_top, padding_bottom):
        """Adjust marker positions based on padding."""
        # Adjust left markers
        self.left_markers = [(y + padding_top, label) for y, label in self.left_markers]
        # Adjust right markers
        self.right_markers = [(y + padding_top, label) for y, label in self.right_markers]
        # Adjust top markers
        self.top_markers = [(x + padding_left, label) for x, label in self.top_markers]        
        
    
    def update_left_padding(self):
        # Update left padding when slider value changes
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.update_live_view()

    def update_right_padding(self):
        # Update right padding when slider value changes
        self.right_marker_shift_added = self.right_padding_slider.value()
        self.update_live_view()
        
    def update_top_padding(self):
        # Update top padding when slider value changes
        self.top_marker_shift_added = self.top_padding_slider.value()
        self.update_live_view()

    def update_live_view(self):
        if not self.image:
            return
    
        # Enable the "Predict Molecular Weight" button if markers are present
        if self.left_markers or self.right_markers:
            self.predict_button.setEnabled(True)
        else:
            self.predict_button.setEnabled(False)
    
        # Define a higher resolution for processing (e.g., 2x or 3x label size)
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Calculate scaling factors and offsets
        scale_x = self.image.width() / render_width
        scale_y = self.image.height() / render_height
    
        # Get the crop percentage values from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure the cropping boundaries are valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            self.crop_x_start_slider.setValue(0)
            self.crop_x_end_slider.setValue(100)
            self.crop_y_start_slider.setValue(0)
            self.crop_y_end_slider.setValue(100)
            return
    
        # Crop the image based on the defined boundaries
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
    
        # Get the orientation value from the slider
        orientation = float(self.orientation_slider.value() / 20)  # Orientation slider value
        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")
    
        # Apply the rotation to the cropped image
        rotated_image = cropped_image.transformed(QTransform().rotate(orientation))
    
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1
        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")
    
        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        skewed_image = rotated_image.transformed(transform, Qt.SmoothTransformation)
    
        # Scale the rotated image to the rendering resolution
        scaled_image = skewed_image.scaled(
            render_width,
            render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        # Render on a high-resolution canvas
        canvas = QImage(render_width, render_height, QImage.Format_ARGB32)
        canvas.fill(Qt.transparent)  # Transparent background
    
        # Render the base image and overlays
        self.render_image_on_canvas(canvas, scaled_image, x_start, y_start, render_scale)
    
        # Scale the high-resolution canvas down to the label's size for display
        pixmap = QPixmap.fromImage(canvas).scaled(
            self.live_view_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        painter = QPainter(pixmap)
        pen = QPen(Qt.red)
        pen.setWidth(1)
        painter.setPen(pen)
    
        scale_x = self.live_view_label.width() / self.image.width()
        scale_y = self.live_view_label.height() / self.image.height()
    
        # Draw the quadrilateral or rectangle if in move mode
        if self.live_view_label.mode == "move":
            pen = QPen(Qt.red, 2)
            painter.setPen(pen)
            if self.live_view_label.quad_points:
                # Draw the quadrilateral
                painter.drawPolygon(QPolygonF(self.live_view_label.quad_points))
            elif self.live_view_label.bounding_box_preview:
                # Draw the rectangle
                start_x, start_y, end_x, end_y = self.live_view_label.bounding_box_preview
                rect = QRectF(start_x, start_y, end_x - start_x, end_y - start_y)
                painter.drawRect(rect)
    
        painter.end()
    
        # Apply zoom and pan transformations to the final pixmap
        if self.live_view_label.zoom_level != 1.0:
            # Create a new pixmap to apply zoom and pan
            zoomed_pixmap = QPixmap(pixmap.size())
            zoomed_pixmap.fill(Qt.transparent)
            zoom_painter = QPainter(zoomed_pixmap)
            zoom_painter.translate(self.live_view_label.pan_offset)
            zoom_painter.scale(self.live_view_label.zoom_level, self.live_view_label.zoom_level)
            zoom_painter.drawPixmap(0, 0, pixmap)
            zoom_painter.end()  # Properly end the QPainter
            pixmap = zoomed_pixmap
    
        # Set the final pixmap to the live view label
        self.live_view_label.setPixmap(pixmap)
    
    def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
        painter = QPainter(canvas)
        x_offset = (canvas.width() - scaled_image.width()) // 2
        y_offset = (canvas.height() - scaled_image.height()) // 2
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
    
        # Draw the base image
        painter.drawImage(x_offset, y_offset, scaled_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            # Resize Image 1 based on the slider value
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 1
            image1_x = x_offset + self.image1_position[0]
            image1_y = y_offset + self.image1_position[1]
            painter.drawImage(image1_x, image1_y, resized_image1)
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            # Resize Image 2 based on the slider value
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 2
            image2_x = x_offset + self.image2_position[0]
            image2_y = y_offset + self.image2_position[1]
            painter.drawImage(image2_x, image2_y, resized_image2)
    
        # Get the selected font settings
        font = QFont(self.font_combo_box.currentFont().family(), self.font_size_spinner.value() * render_scale)
        font_color = self.font_color if hasattr(self, 'font_color') else QColor(0, 0, 0)  # Default to black if no color selected
    
        painter.setFont(font)
        painter.setPen(font_color)  # Set the font color
    
        # Measure text height for vertical alignment
        font_metrics = painter.fontMetrics()
        text_height = font_metrics.height()
        line_padding = 5 * render_scale  # Space between text and line
        
        y_offset_global = text_height / 4
    
        # Draw the left markers (aligned right)
        for y_pos, marker_value in self.left_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f"{marker_value} ⎯ "  ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
        
        
        # Draw the right markers (aligned left)
        for y_pos, marker_value in self.right_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f" ⎯ {marker_value}" ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.right_marker_shift_added),# + line_padding),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
                
    
        # Draw the top markers (if needed)
        for x_pos, top_label in self.top_markers:
            x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width())
            if 0 <= x_pos_cropped <= scaled_image.width():
                text = f"{top_label}"
                painter.save()
                text_width = font_metrics.horizontalAdvance(text)
                text_height= font_metrics.height()
                label_x = x_offset + x_pos_cropped 
                label_y = y_offset + self.top_marker_shift + self.top_marker_shift_added 
                painter.translate(int(label_x), int(label_y))
                painter.rotate(self.font_rotation)
                painter.drawText(0,int(y_offset_global), f"{top_label}")
                painter.restore()
    
        # Draw guide lines
        if draw_guides and self.show_guides_checkbox.isChecked():
            pen = QPen(Qt.red, 2 * render_scale)
            painter.setPen(pen)
            center_x = canvas.width() // 2
            center_y = canvas.height() // 2
            painter.drawLine(center_x, 0, center_x, canvas.height())  # Vertical line
            painter.drawLine(0, center_y, canvas.width(), center_y)  # Horizontal line
            
        
        
        
        # Draw the protein location marker (*)
        if hasattr(self, "protein_location") and self.run_predict_MW == False:
            x, y = self.protein_location
            text = "⎯⎯"
            text_width = font_metrics.horizontalAdvance(text)
            text_height= font_metrics.height()
            painter.drawText(
                int(x * render_scale - text_width / 2),  #Currently left edge # FOR Center horizontally use int(x * render_scale - text_width / 2)
                int(y * render_scale + text_height / 4),  # Center vertically
                text
            )
        if hasattr(self, "custom_markers"):
            
            # Get default font type and size
            default_font_type = QFont(self.custom_font_type_dropdown.currentText())
            default_font_size = int(self.custom_font_size_spinbox.value())
        
            for x_pos, y_pos, marker_text, color, *optional in self.custom_markers:
                
                   
                # Use provided font type and size if available, otherwise use defaults
                marker_font_type = optional[0] if len(optional) > 0 else default_font_type
                marker_font_size = optional[1] if len(optional) > 1 else default_font_size
        
                # Ensure marker_font_type is a QFont instance
                font = QFont(marker_font_type) if isinstance(marker_font_type, str) else QFont(marker_font_type)
                font.setPointSize(marker_font_size * render_scale)  # Adjust font size for rendering scale
        
                # Apply the font and pen settings
                painter.setFont(font)
                painter.setPen(color)
        
                # Correct scaling and offsets
                x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width()) 
                y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height()) 
        
                # Only draw markers if they fall within the visible scaled image area
                if 0 <= x_pos_cropped <= scaled_image.width() and 0 <= y_pos_cropped <= scaled_image.height():
                    # Calculate text dimensions for alignment
                    font_metrics = painter.fontMetrics()
                    text_width = font_metrics.horizontalAdvance(marker_text)
                    text_height = font_metrics.height()
        
                    # Draw text centered horizontally and vertically
                    painter.drawText(
                        int(x_pos_cropped + x_offset- text_width / 2),  # Center horizontally
                        int(y_pos_cropped + y_offset+ text_height / 4),  # Center vertically
                        marker_text
                    )
            
            
    

        # Draw the grid (if enabled)
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value() * render_scale
            pen = QPen(Qt.red)
            pen.setStyle(Qt.DashLine)
            painter.setPen(pen)
    
            # Draw vertical grid lines
            for x in range(0, canvas.width(), grid_size):
                painter.drawLine(x, 0, x, canvas.height())
    
            # Draw horizontal grid lines
            for y in range(0, canvas.height(), grid_size):
                painter.drawLine(0, y, canvas.width(), y)
                
        painter.end()


     
    def crop_image(self):
        """Function to crop the current image."""
        if not self.image:
            return None
    
        # Get crop percentage from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate crop boundaries
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure cropping is valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            return None
    
        # Crop the image
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
        return cropped_image
    
    # Modify align_image and update_crop to preserve settings
    def reset_align_image(self):
        self.orientation_slider.setValue(0)
        
    def align_image(self):
        self.save_state()
        """Align the image based on the orientation slider and keep high-resolution updates."""
        if not self.image:
            return
    
        self.draw_guides = False
        self.show_guides_checkbox.setChecked(False)
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
    
        # Get the orientation value from the slider
        angle = float(self.orientation_slider.value()/20)
    
        # Perform rotation
        transform = QTransform()
        transform.translate(self.image.width() / 2, self.image.height() / 2)  # Center rotation
        transform.rotate(angle)
        transform.translate(-self.image.width() / 2, -self.image.height() / 2)
    
        rotated_image = self.image.transformed(transform, Qt.SmoothTransformation)
    
        # Create a white canvas large enough to fit the rotated image
        canvas_width = rotated_image.width()
        canvas_height = rotated_image.height()
        high_res_canvas = QImage(canvas_width, canvas_height, QImage.Format_RGB888)
        high_res_canvas.fill(QColor(255, 255, 255))  # Fill with white background
    
        # Render the high-resolution canvas using `render_image_on_canvas`, without guides
        self.render_image_on_canvas(
            high_res_canvas,
            scaled_image=rotated_image,
            x_start=0,
            y_start=0,
            render_scale=1,  # Adjust scale if needed
            draw_guides=False  # Do not draw guides in this case
        )
    
        # Update the main high-resolution image
        self.image = high_res_canvas
        self.image_before_padding = self.image.copy()
        self.image_contrasted=self.image.copy()
        self.image_before_contrast=self.image.copy()
    
        # Create a low-resolution preview for display in `live_view_label`
        preview = high_res_canvas.scaled(
            self.live_view_label.width(),
            self.live_view_label.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.live_view_label.setPixmap(QPixmap.fromImage(preview))
    
        # Reset the orientation slider
        self.orientation_slider.setValue(0)
    
    
    def update_crop(self):
        self.save_state()
        """Update the image based on current crop sliders. First align, then crop the image."""
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view() # Update once before operations

        # Get crop parameters *before* aligning/cropping
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100

        # Calculate the crop boundaries based on the *current* image dimensions before cropping
        if not self.image:
            return # Should not happen if called from UI, but safety check
        current_width = self.image.width()
        current_height = self.image.height()
        x_start = int(current_width * x_start_percent)
        x_end = int(current_width * x_end_percent)
        y_start = int(current_height * y_start_percent)
        y_end = int(current_height * y_end_percent)

        # Ensure cropping boundaries are valid relative to current image
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values based on current image size.")
            # Optionally reset sliders here if needed
            return


        # Align the image first (rotate it) - align_image modifies self.image
        # We align *before* cropping based on the original logic flow provided.
        # Note: If alignment changes dimensions significantly, this might need rethinking,
        # but typically rotation keeps content centered.
        # self.align_image() # Call align_image *if* it should happen before crop


        # Now apply cropping to the *current* self.image
        cropped_image = self.crop_image() # crop_image uses current self.image state

        if cropped_image:
            # --- Adjust marker positions relative to the crop ---
            new_left_markers = []
            for y, label in self.left_markers:
                if y_start <= y < y_end: # Check if marker was within vertical crop bounds
                    new_y = y - y_start
                    new_left_markers.append((new_y, label))
            self.left_markers = new_left_markers

            new_right_markers = []
            for y, label in self.right_markers:
                if y_start <= y < y_end:
                    new_y = y - y_start
                    new_right_markers.append((new_y, label))
            self.right_markers = new_right_markers

            new_top_markers = []
            for x, label in self.top_markers:
                if x_start <= x < x_end: # Check if marker was within horizontal crop bounds
                    new_x = x - x_start
                    new_top_markers.append((new_x, label))
            self.top_markers = new_top_markers

            new_custom_markers = []
            if hasattr(self, "custom_markers"):
                for x, y, text, color, font, font_size in self.custom_markers:
                    if x_start <= x < x_end and y_start <= y < y_end:
                        new_x = x - x_start
                        new_y = y - y_start
                        new_custom_markers.append((new_x, new_y, text, color, font, font_size))
            self.custom_markers = new_custom_markers
            # -----------------------------------------------------

            # Update the main image and related states
            self.image = cropped_image
            # Ensure these backups reflect the *newly cropped* state
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

        # Reset sliders after applying the crop
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)

        # Update live view label size based on the *new* image dimensions
        try:
            if self.image: # Check if image exists after cropping
                w = self.image.width()
                h = self.image.height()
                # Preview window
                ratio = w / h if h > 0 else 1 # Avoid division by zero
                self.label_width = int(self.screen_width * 0.28)
                label_height = int(self.label_width / ratio)
                if label_height > self.label_width:
                    label_height = self.label_width
                    self.label_width = int(ratio * label_height) # Ensure integer width
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except Exception as e:
            print(f"Error resizing label after crop: {e}")
            # Fallback size?
            self.live_view_label.setFixedSize(int(self.screen_width * 0.28), int(self.screen_width * 0.28))


        self.update_live_view() # Final update with corrected markers and image
        
    def update_skew(self):
        self.save_state()
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

 
        self.image = self.image.transformed(transform, Qt.SmoothTransformation)
        self.taper_skew_slider.setValue(0)

    
        
    def save_image(self):
        if not self.image:
            QMessageBox.warning(self, "Warning", "Please load an image first.")
            return
    
        options = QFileDialog.Options()
        save_path=""
        
        # Set default save directory and file name
        if hasattr(self, "image_path") and hasattr(self, "base_name"):
            default_file_name = os.path.join(self.image_path, f"{self.base_name}.png")
        else:
            default_file_name = ""
        
        base_save_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image", default_file_name, "Image Files (*.png *.jpg *.bmp)", options=options
        )
        if base_save_path:
            # Check if the base name already contains "_original"
            if "_original" not in os.path.splitext(base_save_path)[0]:
                original_save_path = os.path.splitext(base_save_path)[0] + "_original.png"
                modified_save_path = os.path.splitext(base_save_path)[0] + "_modified.png"
                config_save_path = os.path.splitext(base_save_path)[0] + "_config.txt"
                save_path=original_save_path
                
            else:
                original_save_path = base_save_path
                modified_save_path = os.path.splitext(base_save_path)[0].replace("_original", "") + "_modified.png"
                config_save_path = os.path.splitext(base_save_path)[0].replace("_original", "") + "_config.txt"
                save_path=original_save_path
                # Check if "_modified" and "_config" already exist, and overwrite them if so
                if os.path.exists(modified_save_path):
                    os.remove(modified_save_path)
                if os.path.exists(config_save_path):
                    os.remove(config_save_path)
    
            # Save original image (cropped and aligned)
            self.original_image = self.image.copy()  # Save the current (cropped and aligned) image as original
            if not self.original_image.save(original_save_path):
                QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
    
            # Save modified image
            render_scale = 3
            high_res_canvas_width = self.live_view_label.width() * render_scale
            high_res_canvas_height = self.live_view_label.height() * render_scale
            high_res_canvas = QImage(
                high_res_canvas_width, high_res_canvas_height, QImage.Format_RGB888
            )
            high_res_canvas.fill(QColor(255, 255, 255))  # White background
    
            # Define cropping boundaries
            x_start_percent = self.crop_x_start_slider.value() / 100
            y_start_percent = self.crop_y_start_slider.value() / 100
            x_start = int(self.image.width() * x_start_percent)
            y_start = int(self.image.height() * y_start_percent)
    
            # Create a scaled version of the image
            scaled_image = self.image.scaled(
                high_res_canvas_width,
                high_res_canvas_height,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )
            
            self.show_grid_checkbox.setChecked(False)
            self.update_live_view()
            # Render the high-resolution canvas without guides for saving
            self.render_image_on_canvas(
                high_res_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False
            )
    
            if not high_res_canvas.save(modified_save_path):
                QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
    
            # Save configuration to a .txt file
            config_data = self.get_current_config()
            try:
                with open(config_save_path, "w") as config_file:
                    json.dump(config_data, config_file, indent=4)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
    
            QMessageBox.information(self, "Saved", f"Files saved successfully.")
            
            self.setWindowTitle(f"{self.window_title}:{self.image_path}")
            
    def save_image_svg(self):
        """Save the processed image along with markers and labels in SVG format containing EMF data."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image as SVG for MS Word Image Editing", "", "SVG Files (*.svg)", options=options
        )
    
        if not file_path:
            return
    
        if not file_path.endswith(".svg"):
            file_path += ".svg"
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            font_metrics = QFontMetrics(QFont(font, font_size))
            text_width = (font_metrics.horizontalAdvance(text))  # Get text width
            text_height = (font_metrics.height())
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, y-text_height/4),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(self.left_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="end"  # Aligns text to the right
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()

    
            dwg.add(
                dwg.text(
                    f" ⎯ {text}",
                    insert=(self.right_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="start"  # Aligns text to the left
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text}"
            text_width = int(font_metrics.horizontalAdvance(final_text)) # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, self.top_marker_shift_added-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x-text_width/2}, {self.top_marker_shift_added+text_height/4})"
                )
            )
    
        # Save the SVG file
        dwg.save()
    
        QMessageBox.information(self, "Success", f"Image saved as SVG at {file_path}.")
    
    
    def copy_to_clipboard(self):
        """Copy the image from live view label to the clipboard."""
        if not self.image:
            print("No image to copy.")
            return
    
        # Define a high-resolution canvas for clipboard copy
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, QImage.Format_RGB888
        )
        high_res_canvas.fill(QColor(255, 255, 255))  # White background
    
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
    
        # Create a scaled version of the image
        scaled_image = self.image.scaled(
            high_res_canvas_width,
            high_res_canvas_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
        # Render the high-resolution canvas without guides for clipboard
        self.render_image_on_canvas(
            high_res_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False
        )
        
        # Copy the high-resolution image to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setImage(high_res_canvas)  # Copy the rendered image
        
    def copy_to_clipboard_SVG(self):
        """Create a temporary SVG file with EMF data, copy it to clipboard."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
    
        # Get scaling factors to match the live view window
        view_width = self.live_view_label.width()
        view_height = self.live_view_label.height()
        scale_x = self.image.width() / view_width
        scale_y = self.image.height() / view_height
    
        # Create a temporary file to store the SVG
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_file:
            temp_file_path = temp_file.name
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(temp_file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x * scale_x, y * scale_y),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(self.left_marker_shift_added / scale_x, y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=((self.image.width() / scale_x + self.right_marker_shift_added / scale_x), y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x, self.top_marker_shift_added / scale_y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x}, {self.top_marker_shift_added / scale_y})"
                )
            )
    
        # Save the SVG to the temporary file
        dwg.save()
    
        # Read the SVG content
        with open(temp_file_path, "r", encoding="utf-8") as temp_file:
            svg_content = temp_file.read()
    
        # Copy SVG content to clipboard (macOS-compatible approach)
        clipboard = QApplication.clipboard()
        clipboard.setText(svg_content, mode=clipboard.Clipboard)
    
        QMessageBox.information(self, "Success", "SVG content copied to clipboard.")

        
    def clear_predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        if hasattr(self, "protein_location"):
            del self.protein_location  # Clear the protein location marker
        self.predict_size=False
        self.bounding_boxes=[]
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.live_view_label.bounding_box_preview = None
        self.quantities=[]
        self.peak_area_list=[]
        self.protein_quantities=[]
        self.standard_protein_values.setText("")
        self.standard_protein_areas=[]
        self.standard_protein_areas_text.setText("")
        self.live_view_label.quad_points=[]
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.rectangle_points = []
        
        self.update_live_view()  # Update the display
        
    def predict_molecular_weight(self):
        """
        Initiates the molecular weight prediction process.
        Determines the available markers, sorts them by position,
        and sets up the mouse click event to get the target protein location.
        """
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.run_predict_MW = False # Reset prediction flag

        # Determine which markers to use (left or right)
        markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
        if not markers_raw_tuples:
            QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
            self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            return

        # --- Crucial Step: Sort markers by Y-position (migration distance) ---
        # This ensures the order reflects the actual separation on the gel/blot
        try:
            # Ensure marker values are numeric for sorting and processing
            numeric_markers = []
            for pos, val_str in markers_raw_tuples:
                try:
                    numeric_markers.append((float(pos), float(val_str)))
                except (ValueError, TypeError):
                    # Skip markers with non-numeric values for prediction
                    print(f"Warning: Skipping non-numeric marker value '{val_str}' at position {pos}.")
                    continue

            if len(numeric_markers) < 2:
                 QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 return

            sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
            # Separate sorted positions and values
            sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
            sorted_marker_values = np.array([val for pos, val in sorted_markers])

        except Exception as e:
            QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Ensure there are still at least two markers after potential filtering
        if len(sorted_marker_positions) < 2:
            QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Prompt user and set up the click handler
        QMessageBox.information(self, "Instruction",
                                "Click on the target protein location in the preview window.\n"
                                "The closest standard set (Gel or WB) will be used for calculation.")
        # Pass the *sorted* full marker data to the click handler
        self.live_view_label.mousePressEvent = lambda event: self.get_protein_location(
            event, sorted_marker_positions, sorted_marker_values
        )

    def get_protein_location(self, event, all_marker_positions, all_marker_values):
        """
        Handles the mouse click event for protein selection.
        Determines the relevant standard set based on click proximity,
        performs regression on that set, predicts MW, and plots the results.
        """
        # --- 1. Get Protein Click Position (Image Coordinates) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()

        # Account for zoom and pan
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        # Transform cursor position from view coordinates to image coordinates
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image and self.image.width() > 0 else 1
        image_height = self.image.height() if self.image and self.image.height() > 0 else 1

        scale = min(displayed_width / image_width, displayed_height / image_height)
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2

        protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
        # protein_x_image = (cursor_x - x_offset) / scale if scale != 0 else 0 # Keep X if needed later

        # Store the clicked location in *view* coordinates for drawing the marker later
        self.protein_location = (cursor_x, cursor_y) # Use view coordinates for the marker

        # --- 2. Identify Potential Standard Sets (Partitioning) ---
        transition_index = -1
        # Find the first index where the molecular weight INCREASES after initially decreasing
        # (indicates a likely switch from Gel high->low MW to WB high->low MW)
        initial_decrease = False
        for k in range(1, len(all_marker_values)):
             if all_marker_values[k] < all_marker_values[k-1]:
                 initial_decrease = True # Confirm we've started migrating down
             # Check for increase *after* we've seen a decrease
             if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                 transition_index = k
                 break # Found the likely transition

        # --- 3. Select the Active Standard Set based on Click Proximity ---
        active_marker_positions = None
        active_marker_values = None
        set_name = "Full Set" # Default name

        if transition_index != -1:
            # We have two potential sets
            set1_positions = all_marker_positions[:transition_index]
            set1_values = all_marker_values[:transition_index]
            set2_positions = all_marker_positions[transition_index:]
            set2_values = all_marker_values[transition_index:]

            # Check if both sets are valid (at least 2 points)
            valid_set1 = len(set1_positions) >= 2
            valid_set2 = len(set2_positions) >= 2

            if valid_set1 and valid_set2:
                # Calculate the mean Y position for each set
                mean_y_set1 = np.mean(set1_positions)
                mean_y_set2 = np.mean(set2_positions)

                # Assign the click to the set whose mean Y is closer
                if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                    active_marker_positions = set1_positions
                    active_marker_values = set1_values
                    set_name = "Set 1 (Gel?)" # Tentative name
                else:
                    active_marker_positions = set2_positions
                    active_marker_values = set2_values
                    set_name = "Set 2 (WB?)" # Tentative name
            elif valid_set1: # Only set 1 is valid
                 active_marker_positions = set1_positions
                 active_marker_values = set1_values
                 set_name = "Set 1 (Gel?)"
            elif valid_set2: # Only set 2 is valid
                 active_marker_positions = set2_positions
                 active_marker_values = set2_values
                 set_name = "Set 2 (WB?)"
            else: # Neither set is valid after splitting
                 QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return
        else:
            # Only one set detected, use all markers
            if len(all_marker_positions) >= 2:
                active_marker_positions = all_marker_positions
                active_marker_values = all_marker_values
                set_name = "Single Set"
            else: # Should have been caught earlier, but double-check
                 QMessageBox.warning(self, "Error", "Not enough markers in the single set.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return

        # --- 4. Perform Regression on the Active Set ---
        # Normalize distances *within the active set*
        min_pos_active = np.min(active_marker_positions)
        max_pos_active = np.max(active_marker_positions)
        if max_pos_active == min_pos_active: # Avoid division by zero if all points are the same
            QMessageBox.warning(self, "Error", "All markers in the selected set have the same position.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if hasattr(self, "protein_location"): del self.protein_location
            return

        normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)

        # Log transform marker values
        try:
            log_marker_values_active = np.log10(active_marker_values)
        except Exception as e:
             QMessageBox.warning(self, "Error", f"Could not log-transform marker values (are they all positive?): {e}")
             self.live_view_label.setCursor(Qt.ArrowCursor)
             if hasattr(self, "protein_location"): del self.protein_location
             return

        # Perform linear regression (log-linear fit)
        coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, 1)

        # Calculate R-squared for the fit on the active set
        residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1 # Handle case of perfect fit or single point mean

        # --- 5. Predict MW for the Clicked Protein ---
        # Normalize the protein's Y position *using the active set's min/max*
        normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)

        # Predict using the derived coefficients
        predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
        predicted_weight = 10 ** predicted_log10_weight

        # --- 6. Update View and Plot ---
        self.update_live_view() # Draw the '*' marker at self.protein_location

        # Plot the results, passing both active and all data for context
        self.plot_molecular_weight_graph(
            # Active set data for fitting and line plot:
            normalized_distances_active,
            active_marker_values,
            10 ** np.polyval(coefficients, normalized_distances_active), # Fitted values for active set
            # Full set data for context (plotting all points):
            all_marker_positions, # Pass original positions
            all_marker_values,
            min_pos_active,       # Pass min/max of *active* set for normalization context
            max_pos_active,
            # Prediction results:
            normalized_protein_position, # Position relative to active set scale
            predicted_weight,
            r_squared,
            set_name # Name of the set used
        )

        # Reset mouse event handler after prediction
        self.live_view_label.mousePressEvent = None
        self.live_view_label.setCursor(Qt.ArrowCursor)
        self.run_predict_MW = True # Indicate prediction was attempted/completed

        
        
    def plot_molecular_weight_graph(
        self,
        # Data for the active set (used for the fit line and highlighted points)
        active_norm_distances,  # Normalized distances for the active set points
        active_marker_values,   # Original MW values of the active set points
        active_fitted_values,   # Fitted MW values corresponding to active_norm_distances

        # Data for *all* markers (for plotting all points for context)
        all_marker_positions,   # *Original* Y positions of all markers
        all_marker_values,      # *Original* MW values of all markers
        active_min_pos,         # Min Y position of the *active* set (for normalizing all points)
        active_max_pos,         # Max Y position of the *active* set (for normalizing all points)

        # Prediction results
        predicted_norm_position,# Predicted protein position normalized relative to active set scale
        predicted_weight,       # Predicted MW value

        # Fit quality and set info
        r_squared,              # R-squared of the fit on the active set
        set_name                # Name of the set used (e.g., "Set 1", "Set 2", "Single Set")
    ):
        """
        Plots the molecular weight prediction graph, showing all markers,
        highlighting the active set, the fitted line for the active set,
        and the predicted protein position.
        """
        import matplotlib.pyplot as plt
        from io import BytesIO
        from PyQt5.QtGui import QPixmap
        from PyQt5.QtWidgets import QLabel # Make sure QLabel is imported if not already

        # --- Normalize *all* marker positions using the *active* set's scale ---
        # This ensures all points are plotted on the same normalized x-axis
        # defined by the regression range.
        if active_max_pos == active_min_pos: # Avoid division by zero
             print("Warning: Cannot normalize all points - active set min/max are equal.")
             all_norm_distances_for_plot = np.zeros_like(all_marker_positions)
        else:
            all_norm_distances_for_plot = (all_marker_positions - active_min_pos) / (active_max_pos - active_min_pos)

        # --- Create Plot ---
        plt.figure(figsize=(6, 4)) # Slightly larger figure

        # 1. Plot *all* marker points lightly for context
        plt.scatter(all_norm_distances_for_plot, all_marker_values, color="grey", alpha=0.6, label="All Markers", s=30)

        # 2. Plot the *active* marker points prominently
        plt.scatter(active_norm_distances, active_marker_values, color="red", label=f"Active Set ({set_name})", s=50, marker='o')

        # 3. Plot the fitted line for the *active* set
        # Sort points for a smooth line if necessary (linear fit usually doesn't require it)
        sort_indices = np.argsort(active_norm_distances)
        plt.plot(active_norm_distances[sort_indices], active_fitted_values[sort_indices],
                 color="blue", label=f"Fit (R²={r_squared:.3f})")

        # 4. Plot the predicted protein position
        plt.axvline(predicted_norm_position, color="green", linestyle="--",
                    label=f"Target Protein\n({predicted_weight:.2f} units)")

        # --- Configure Plot ---
        plt.xlabel(f"Normalized Distance (Relative to {set_name})")
        plt.ylabel("Molecular Weight (units)")
        plt.yscale("log")  # Log scale for molecular weight is standard
        plt.legend(fontsize='small')
        plt.title(f"Molecular Weight Prediction (Using {set_name})")
        plt.grid(True, which='both', linestyle=':', linewidth=0.5) # Add subtle grid

        # --- Display Plot in QMessageBox ---
        try:
            # Convert the plot to a pixmap
            buffer = BytesIO()
            plt.savefig(buffer, format="png", bbox_inches="tight", dpi=150) # Increase DPI slightly
            buffer.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buffer.read())
            buffer.close()
        except Exception as plot_err:
            print(f"Error generating plot image: {plot_err}")
            QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
            plt.close() # Close the figure even if saving failed
            return
        finally:
             plt.close() # Ensure figure is always closed

        # Display the plot and results in a message box
        message_box = QMessageBox(self)
        message_box.setWindowTitle("Prediction Result")
        message_box.setText(
            f"Used {set_name} for prediction.\n"
            f"The predicted molecular weight is approximately {predicted_weight:.2f} units.\n"
            f"R-squared value of the fit: {r_squared:.3f}"
        )
        # Add plot image below the text
        plot_label = QLabel()
        plot_label.setPixmap(pixmap)
        # Access the grid layout of the QMessageBox to add the widget
        grid_layout = message_box.layout()
        # Add widget in the next available row (row 1), spanning all columns
        grid_layout.addWidget(plot_label, 1, 0, 1, grid_layout.columnCount())
        message_box.exec_()

  
        
    def reset_image(self):
        # Reset the image to original
        if self.image != None:
            self.image = self.image_master.copy()
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy() # Update the contrasted image
            self.image_before_contrast= self.image.copy()
        else:
            self.image_before_padding = None
            self.image_contrasted = None  
            self.image_before_contrast=None
        self.warped_image=None
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
        self.update_live_view()  # Update the live view
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)
        self.orientation_slider.setValue(0)
        self.taper_skew_slider.setValue(0)
        self.high_slider.setValue(100)  # Reset contrast to default
        self.low_slider.setValue(0)  # Reset contrast to default
        self.gamma_slider.setValue(100)  # Reset gamma to default
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_padding_slider.setValue(0)
        self.right_padding_slider.setValue(0)
        self.top_padding_slider.setValue(0)
        self.marker_mode = None
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
        self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
        try:
            w=self.image.width()
            h=self.image.height()
            # Preview window
            ratio=w/h
            self.label_width=int(self.screen_width * 0.28)
            label_height=int(self.label_width/ratio)
            if label_height>self.label_width:
                label_height=self.label_width
                self.label_width=ratio*label_height
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except:
            pass
        self.update_live_view()


if __name__ == "__main__":
    try:
        app = QApplication([])
        app.setStyle("Fusion")
        app.setStyleSheet("""
        QSlider::handle:horizontal {
            width: 100px;
            height: 100px;
            margin: -5px 0;
            background: #FFFFFF;
            border: 2px solid #555;
            border-radius: 30px;
        }
    """)
        window = CombinedSDSApp()
        window.show()
        app.exec_()
    except Exception as e:
        logging.error("Application crashed", exc_info=True)
