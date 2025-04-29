import logging
import traceback
import sys
import svgwrite
import tempfile
from tempfile import NamedTemporaryFile
import base64
from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
from io import BytesIO
import io
from PyQt5.QtWidgets import (
    QDesktopWidget, QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem,QToolBar,QStyle,
    QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy,
    QMainWindow, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit,
    QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog,
    QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox,
    QDialog, QHeaderView, QAbstractItemView, QMenu, QAction, QMenuBar, QFontDialog
)
from PyQt5.QtGui import QPixmap, QIcon, QPalette,QKeySequence, QImage, QPolygonF,QPainter, QBrush, QColor, QFont, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF, QUrl, QSize, QMimeData, QUrl
import json
import os
import numpy as np
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.gridspec import GridSpec
from skimage.restoration import rolling_ball 
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import interp1d # Needed for interpolation
import cv2



# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
script_dir = os.path.dirname(os.path.abspath(__file__))
# Construct the absolute path for the log file
log_file_path = os.path.join(script_dir, "error_log.txt")


# --- Configure logging ---
logging.basicConfig(
    filename=log_file_path, # Use the absolute path
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

try:
    logging.error("--- Logging initialized ---")
except Exception as e:
    print(f"ERROR: Could not write initial log message: {e}") # Print error if immediate logging fails


def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    print("!!! log_exception called !!!") # Add print statement here too
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    # Log the exception
    try:
        logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))
        # Optional: Force flush if you suspect buffering issues, though unlikely for ERROR level
        # logging.getLogger().handlers[0].flush()
    except Exception as log_err:
        print(f"ERROR: Failed to log exception to file: {log_err}") # Print logging specific errors

    # Display a QMessageBox with the error details
    try:
        error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}\n\n(Check error_log.txt for details)"
        QMessageBox.critical(
            None,
            "Unexpected Error",
            error_message,
            QMessageBox.Ok
        )
    except Exception as q_err:
         print(f"ERROR: Failed to show QMessageBox: {q_err}")


# Set the custom exception handler
sys.excepthook = log_exception

def create_text_icon(icon_size: QSize, color: QColor, symbol: str) -> QIcon:
    """Creates a QIcon by drawing text/symbol onto a pixmap."""
    pixmap = QPixmap(icon_size)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)
    painter.setRenderHint(QPainter.TextAntialiasing, True) # Good for text

    # Font settings (adjust as needed)
    font = QFont()
    # Make arrow slightly smaller than +/-, maybe not bold? Experiment.
    font.setPointSize(max(12, int(icon_size.height() * 0.55)))
    # font.setBold(True) # Optional: Make arrows bold or not
    painter.setFont(font)
    painter.setPen(color)

    # Draw the symbol centered
    painter.drawText(pixmap.rect(), Qt.AlignCenter, symbol)
    painter.end()
    return QIcon(pixmap)
    
class ModifyMarkersDialog(QDialog):
    """
    Dialog to view, edit, delete, and reorder custom markers.
    Now includes Bold/Italic controls.
    """
    def __init__(self, markers_list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Modify Custom Markers")
        self.setMinimumSize(750, 450) # Increased width for new columns

        # Store a working copy of the markers, ensuring 8 elements
        self.markers = []
        for marker_data in markers_list:
            if len(marker_data) == 6:
                # Old format: Add default False for bold/italic
                self.markers.append(marker_data + (False, False))
            elif len(marker_data) == 8:
                self.markers.append(marker_data) # Already new format
            else:
                pass
        # self.markers now contains only 8-element tuples

        self._block_signals = False # Flag to prevent recursive signals

        # --- Main Layout ---
        layout = QVBoxLayout(self)

        # --- Table Widget ---
        self.table_widget = QTableWidget()
        # -->> Increased column count to 9 <<--
        self.table_widget.setColumnCount(9)
        self.table_widget.setHorizontalHeaderLabels([
            "Text", "Font", "Size", "Bold", "Italic", "Color", "X Pos", "Y Pos", "Actions"
        ])
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table_widget.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed)
        self.table_widget.setSortingEnabled(True)

        # Connect signals
        self.table_widget.itemChanged.connect(self.handle_item_changed)
        self.table_widget.cellDoubleClicked.connect(self.handle_cell_double_clicked)

        layout.addWidget(self.table_widget)

        # --- Populate Table ---
        self.populate_table()

        # Adjust column widths
        self.table_widget.resizeColumnsToContents()
        self.table_widget.horizontalHeader().setStretchLastSection(False)
        # Adjust widths based on new columns
        self.table_widget.setColumnWidth(0, 150) # Text
        self.table_widget.setColumnWidth(1, 100) # Font
        self.table_widget.setColumnWidth(2, 40)  # Size
        self.table_widget.setColumnWidth(3, 40)  # Bold << NEW
        self.table_widget.setColumnWidth(4, 40)  # Italic << NEW
        self.table_widget.setColumnWidth(5, 80)  # Color (Index 5 now)
        self.table_widget.setColumnWidth(6, 60)  # X Pos (Index 6 now)
        self.table_widget.setColumnWidth(7, 60)  # Y Pos (Index 7 now)
        self.table_widget.setColumnWidth(8, 80)  # Actions (Index 8 now)

        # --- Button Box ---
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        self.setLayout(layout)

    def populate_table(self):
        """Fills the table with the current marker data, including Bold/Italic checkboxes."""
        self._block_signals = True # Block signals during population
        self.table_widget.setRowCount(0) # Clear existing rows
        self.table_widget.setRowCount(len(self.markers))
        self.table_widget.setSortingEnabled(False)

        for row_idx, marker_data in enumerate(self.markers):
            try:
                # -->> Unpack 8 elements <<--
                x, y, text, qcolor, font_family, font_size, is_bold, is_italic = marker_data
            except ValueError as e:
                 self.table_widget.setItem(row_idx, 0, QTableWidgetItem("Error"))
                 continue

            # --- Create Standard Items ---
            text_item = QTableWidgetItem(str(text))
            font_item = QTableWidgetItem(str(font_family))
            size_item = QTableWidgetItem(str(font_size))
            color_item = QTableWidgetItem(qcolor.name())
            x_item = QTableWidgetItem(f"{x:.1f}")
            y_item = QTableWidgetItem(f"{y:.1f}")

            # --- Set Flags for Editability ---
            text_item.setFlags(text_item.flags() | Qt.ItemIsEditable)
            font_item.setFlags(font_item.flags() & ~Qt.ItemIsEditable)
            size_item.setFlags(size_item.flags() | Qt.ItemIsEditable)
            color_item.setFlags(color_item.flags() & ~Qt.ItemIsEditable)
            x_item.setFlags(x_item.flags() | Qt.ItemIsEditable)
            y_item.setFlags(y_item.flags() | Qt.ItemIsEditable)

            # --- Tooltips ---
            color_item.setToolTip("Double-click to change color")
            font_item.setToolTip("Double-click to change font")

            # --- Color Background ---
            color_item.setBackground(QBrush(qcolor))
            text_color = Qt.white if qcolor.lightness() < 128 else Qt.black
            color_item.setForeground(QBrush(text_color))

            # --- Add Standard Items to Table ---
            # Column indices are shifted due to new Bold/Italic columns
            self.table_widget.setItem(row_idx, 0, text_item)   # Col 0: Text
            self.table_widget.setItem(row_idx, 1, font_item)   # Col 1: Font Family
            self.table_widget.setItem(row_idx, 2, size_item)   # Col 2: Font Size
            # Columns 3 and 4 are for checkboxes
            self.table_widget.setItem(row_idx, 5, color_item)  # Col 5: Color
            self.table_widget.setItem(row_idx, 6, x_item)      # Col 6: X Pos
            self.table_widget.setItem(row_idx, 7, y_item)      # Col 7: Y Pos

            # --- Create and Add Checkboxes ---
            # Bold Checkbox (Column 3)
            bold_checkbox = QCheckBox()
            bold_checkbox.setChecked(is_bold)
            bold_checkbox.stateChanged.connect(
                lambda state, r=row_idx: self.handle_style_changed(state, r, "bold")
            )
            # Center the checkbox in the cell
            cell_widget_bold = QWidget()
            layout_bold = QHBoxLayout(cell_widget_bold)
            layout_bold.addWidget(bold_checkbox)
            layout_bold.setAlignment(Qt.AlignCenter)
            layout_bold.setContentsMargins(0,0,0,0)
            cell_widget_bold.setLayout(layout_bold)
            self.table_widget.setCellWidget(row_idx, 3, cell_widget_bold)

            # Italic Checkbox (Column 4)
            italic_checkbox = QCheckBox()
            italic_checkbox.setChecked(is_italic)
            italic_checkbox.stateChanged.connect(
                lambda state, r=row_idx: self.handle_style_changed(state, r, "italic")
            )
            # Center the checkbox
            cell_widget_italic = QWidget()
            layout_italic = QHBoxLayout(cell_widget_italic)
            layout_italic.addWidget(italic_checkbox)
            layout_italic.setAlignment(Qt.AlignCenter)
            layout_italic.setContentsMargins(0,0,0,0)
            cell_widget_italic.setLayout(layout_italic)
            self.table_widget.setCellWidget(row_idx, 4, cell_widget_italic)

            # --- Add Delete Button ---
            delete_button = QPushButton("Delete")
            delete_button.setStyleSheet("QPushButton { padding: 2px 5px; }")
            delete_button.clicked.connect(lambda checked, r=row_idx: self.delete_marker(r))
            self.table_widget.setCellWidget(row_idx, 8, delete_button) # Col 8: Actions

        self.table_widget.setSortingEnabled(True)
        self._block_signals = False # Re-enable signals

    def handle_item_changed(self, item):
        """Update the internal markers list when text, size, or position is edited."""
        if self._block_signals:
            return

        row = item.row()
        col = item.column()
        new_value_str = item.text()

        if 0 <= row < len(self.markers):
            current_marker = list(self.markers[row]) # Now an 8-element list

            try:
                # Check column indices carefully based on the new layout
                if col == 0: # Text column (Index 2 in marker tuple)
                    current_marker[2] = new_value_str

                elif col == 2: # Size column (Index 5 in marker tuple)
                    try:
                        new_size = int(new_value_str)
                        if 2 <= new_size <= 150:
                             current_marker[5] = new_size
                        else: raise ValueError("Size out of range")
                    except ValueError:
                        QMessageBox.warning(self, "Invalid Size", "...")
                        self._block_signals = True
                        item.setText(str(current_marker[5]))
                        self._block_signals = False
                        return

                elif col == 6: # X Position column (Index 0 in marker tuple)
                    try:
                        new_x = float(new_value_str)
                        current_marker[0] = new_x
                    except ValueError:
                        QMessageBox.warning(self, "Invalid Position", "...")
                        self._block_signals = True
                        item.setText(f"{current_marker[0]:.1f}")
                        self._block_signals = False
                        return

                elif col == 7: # Y Position column (Index 1 in marker tuple)
                    try:
                        new_y = float(new_value_str)
                        current_marker[1] = new_y
                    except ValueError:
                        QMessageBox.warning(self, "Invalid Position", "...")
                        self._block_signals = True
                        item.setText(f"{current_marker[1]:.1f}")
                        self._block_signals = False
                        return

                # Update the marker in the internal list
                self.markers[row] = tuple(current_marker)

            except IndexError:
                 print(f"Error: Index mismatch when updating marker at row {row}.")
            except Exception as e:
                 print(f"Error handling item change for row {row}, col {col}: {e}")
        else:
             print(f"Warning: itemChanged signal received for invalid row {row}")

    # --- NEW Method to handle checkbox changes ---
    def handle_style_changed(self, state, row, style_type):
        """Update the bold/italic flag when a checkbox changes."""
        if self._block_signals:
            return

        if 0 <= row < len(self.markers):
            current_marker = list(self.markers[row])
            is_checked = (state == Qt.Checked)

            try:
                if style_type == "bold":
                    current_marker[6] = is_checked # Index 6 is bold
                elif style_type == "italic":
                    current_marker[7] = is_checked # Index 7 is italic

                self.markers[row] = tuple(current_marker)
            except IndexError:
                print(f"Error: Index mismatch updating style for marker at row {row}.")
            except Exception as e:
                print(f"Error handling style change for row {row}, type {style_type}: {e}")
        else:
            print(f"Warning: styleChanged signal received for invalid row {row}")
    # --- END NEW Method ---

    def handle_cell_double_clicked(self, row, column):
        """Handle double-clicks for changing color or font."""
        if not (0 <= row < len(self.markers)): return

        current_marker = list(self.markers[row]) # 8 elements

        # Column indices shifted: Font=1, Color=5
        if column == 1: # Font column
            # Pass existing flags (bold/italic) to font dialog
            initial_qfont = QFont(current_marker[4], current_marker[5])
            initial_qfont.setBold(current_marker[6])
            initial_qfont.setItalic(current_marker[7])

            selected_font, ok = QFontDialog.getFont(initial_qfont, self, "Select Marker Font")

            if ok:
                # Update family, size, AND bold/italic from dialog
                current_marker[4] = selected_font.family()
                current_marker[5] = selected_font.pointSize()
                current_marker[6] = selected_font.bold()    # Get bold state
                current_marker[7] = selected_font.italic()  # Get italic state
                self.markers[row] = tuple(current_marker)

                # Update table view immediately
                self._block_signals = True
                self.table_widget.item(row, 1).setText(selected_font.family())
                self.table_widget.item(row, 2).setText(str(selected_font.pointSize()))
                # Update checkboxes as well
                bold_widget = self.table_widget.cellWidget(row, 3)
                if bold_widget: bold_widget.findChild(QCheckBox).setChecked(selected_font.bold())
                italic_widget = self.table_widget.cellWidget(row, 4)
                if italic_widget: italic_widget.findChild(QCheckBox).setChecked(selected_font.italic())
                self._block_signals = False

        elif column == 5: # Color column (now index 5)
            current_color = current_marker[3] # Color is still index 3 in tuple
            new_color = QColorDialog.getColor(current_color, self, "Select Marker Color")
            if new_color.isValid():
                current_marker[3] = new_color
                self.markers[row] = tuple(current_marker)

                # Update table view immediately
                self._block_signals = True
                color_item = self.table_widget.item(row, 5) # Update correct column index
                color_item.setText(new_color.name())
                color_item.setBackground(QBrush(new_color))
                text_color = Qt.white if new_color.lightness() < 128 else Qt.black
                color_item.setForeground(QBrush(text_color))
                self._block_signals = False

    def delete_marker(self, row_to_delete):
        """Deletes the marker corresponding to the clicked button's row."""
        sort_col = self.table_widget.horizontalHeader().sortIndicatorSection()
        sort_order = self.table_widget.horizontalHeader().sortIndicatorOrder()
        self.table_widget.setSortingEnabled(False)

        if 0 <= row_to_delete < len(self.markers):
            del self.markers[row_to_delete]
            self.table_widget.removeRow(row_to_delete)
            # Reconnect buttons and checkboxes for rows *after* the deleted one
            for current_row in range(row_to_delete, self.table_widget.rowCount()):
                # Delete button
                button_widget = self.table_widget.cellWidget(current_row, 8) # Index 8 now
                if isinstance(button_widget, QPushButton):
                    try: button_widget.clicked.disconnect()
                    except TypeError: pass
                    button_widget.clicked.connect(lambda checked, r=current_row: self.delete_marker(r))
                # Bold checkbox
                bold_cell_widget = self.table_widget.cellWidget(current_row, 3)
                if bold_cell_widget:
                    bold_checkbox = bold_cell_widget.findChild(QCheckBox)
                    if bold_checkbox:
                        try: bold_checkbox.stateChanged.disconnect()
                        except TypeError: pass
                        bold_checkbox.stateChanged.connect(
                            lambda state, r=current_row: self.handle_style_changed(state, r, "bold")
                        )
                # Italic checkbox
                italic_cell_widget = self.table_widget.cellWidget(current_row, 4)
                if italic_cell_widget:
                    italic_checkbox = italic_cell_widget.findChild(QCheckBox)
                    if italic_checkbox:
                        try: italic_checkbox.stateChanged.disconnect()
                        except TypeError: pass
                        italic_checkbox.stateChanged.connect(
                            lambda state, r=current_row: self.handle_style_changed(state, r, "italic")
                        )
        else:
            print(f"Warning: Attempted to delete invalid row index {row_to_delete}")

        self.table_widget.setSortingEnabled(True)
        if sort_col >= 0:
            self.table_widget.sortByColumn(sort_col, sort_order)

    def get_modified_markers(self):
        """Returns the modified list of markers."""
        return self.markers

class TableWindow(QDialog):
    """
    A dialog window to display peak analysis results in a table.
    Includes functionality to copy selected data to the clipboard
    in a format pasteable into Excel (tab-separated).
    Also includes Excel export functionality and standard curve plot.
    """
    # *** MODIFIED: Added standard_dictionary to the constructor ***
    def __init__(self, peak_areas, standard_dictionary, standard, calculated_quantities, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Analysis Results and Standard Curve") # Updated title
        self.setGeometry(100, 100, 700, 650) # Increased height for plot
        self.calculated_quantities = calculated_quantities
        self.standard_dictionary = standard_dictionary # Store standard data

        # --- Main Layout ---
        main_layout = QVBoxLayout(self)

        # --- Standard Curve Plot Area (Conditional) ---
        plot_widget = self._create_standard_curve_plot() # Create plot or placeholder
        if plot_widget:
            plot_group = QGroupBox("Standard Curve")
            plot_layout = QVBoxLayout(plot_group)
            plot_layout.addWidget(plot_widget)
            # Set a fixed or maximum height for the plot group if needed
            plot_group.setMaximumHeight(300) # Example max height
            plot_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            main_layout.addWidget(plot_group) # Add plot at the top


        # --- Table Setup ---
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers) # Data is read-only
        self.table.setSelectionBehavior(QTableWidget.SelectRows) # Select whole rows
        self.table.setSelectionMode(QTableWidget.ContiguousSelection) # Allow selecting multiple adjacent rows

        # --- Populate Table ---
        self.populate_table(peak_areas, standard_dictionary, standard)

        self.scroll_area.setWidget(self.table)
        main_layout.addWidget(self.scroll_area) # Add table below plot


        # --- Buttons ---
        self.copy_button = QPushButton("Copy Selected to Clipboard")
        self.copy_button.setToolTip("Copy selected rows to clipboard (tab-separated)")
        self.copy_button.clicked.connect(self.copy_table_data) # Connect copy function

        self.export_button = QPushButton("Export Table to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        # --- Layout Buttons ---
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.copy_button)
        button_layout.addStretch()
        button_layout.addWidget(self.export_button)

        main_layout.addLayout(button_layout) # Add buttons at the bottom
        self.setLayout(main_layout)

        # Enable keyboard copy (Ctrl+C / Cmd+C)
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.table)
        self.copy_shortcut.activated.connect(self.copy_table_data)
        # Make table focusable to receive shortcut events
        self.table.setFocusPolicy(Qt.StrongFocus)


    def _create_standard_curve_plot(self):
        """Creates and returns the Matplotlib canvas for the standard curve, or None."""
        if not self.standard_dictionary or len(self.standard_dictionary) < 2:
            # Return a placeholder label if no/insufficient standard data
            no_curve_label = QLabel("Standard curve requires at least 2 standard points.")
            no_curve_label.setAlignment(Qt.AlignCenter)
            return no_curve_label

        try:
            # Extract data
            quantities = np.array(list(self.standard_dictionary.keys()), dtype=float)
            areas = np.array(list(self.standard_dictionary.values()), dtype=float)

            # Perform linear regression (Area vs Quantity)
            coeffs = np.polyfit(areas, quantities, 1)
            slope, intercept = coeffs

            # Calculate R-squared
            predicted_quantities = np.polyval(coeffs, areas)
            residuals = quantities - predicted_quantities
            ss_res = np.sum(residuals**2)
            ss_tot = np.sum((quantities - np.mean(quantities))**2)
            r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1.0

            # Create Matplotlib figure and axes
            fig, ax = plt.subplots(figsize=(5, 3)) # Keep adjusted size
            fig.set_dpi(90)

            # Plot data points
            ax.scatter(areas, quantities, label='Standard Points', color='red', zorder=5)

            # Plot regression line
            x_line = np.linspace(min(areas) * 0.95, max(areas) * 1.05, 100)
            y_line = slope * x_line + intercept

            # *** MODIFICATION: Create multi-line label for the legend ***
            fit_label = (
                f'Linear Fit\n'
                f'Qty = {slope}*Area + {intercept}\n'
                f'R² = {r_squared:.3f}'
            )
            ax.plot(x_line, y_line, label=fit_label, color='blue') # Use the combined label

            # *** REMOVED: Separate text annotation is no longer needed ***
            # eq_text = f'Qty = {slope:.2f} * Area + {intercept:.2f}\nR² = {r_squared:.3f}'
            # ax.text(0.05, 0.05, eq_text, transform=ax.transAxes, fontsize=8,
            #         verticalalignment='bottom', # Align text bottom to position
            #         horizontalalignment='left', # Align text left to position
            #         bbox=dict(boxstyle='round,pad=0.3', fc='wheat', alpha=0.5))

            # Customize plot
            ax.set_xlabel('Total Peak Area (Arbitrary Units)', fontsize=9)
            ax.set_ylabel('Known Quantity (Unit)', fontsize=9)
            ax.set_title('Standard Curve', fontsize=10, fontweight='bold')
            # Keep legend font size small and specify location
            ax.legend(fontsize=8, loc='best') # Use 'best' or 'upper right', etc.
            ax.grid(True, linestyle='--', alpha=0.6)
            ax.tick_params(axis='both', which='major', labelsize=8)
            ax.ticklabel_format(style='sci', axis='x', scilimits=(0,0), useMathText=True)

            # Use constrained_layout for potentially better automatic spacing
            try:
                fig.set_constrained_layout(True)
            except AttributeError: # Fallback for older Matplotlib
                plt.tight_layout(pad=0.5)

            # Create canvas
            canvas = FigureCanvas(fig)
            canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            canvas.updateGeometry()
            return canvas

        except Exception as e:
            print(f"Error creating standard curve plot: {e}")
            error_label = QLabel(f"Error generating plot:\n{e}")
            error_label.setAlignment(Qt.AlignCenter)
            error_label.setStyleSheet("color: red;")
            return error_label


    def populate_table(self, peak_areas, standard_dictionary, standard):
        """Populates the table with the provided peak analysis data."""
        total_area = sum(peak_areas) if peak_areas else 0.0 # Handle empty list
        self.table.setRowCount(len(peak_areas))

        for row, area in enumerate(peak_areas):
            band_label = f"Band {row + 1}"
            self.table.setItem(row, 0, QTableWidgetItem(band_label))

            peak_area_rounded = round(area, 3)
            self.table.setItem(row, 1, QTableWidgetItem(str(peak_area_rounded)))

            if total_area != 0:
                percentage = (area / total_area) * 100
                percentage_rounded = round(percentage, 2)
            else:
                percentage_rounded = 0.0
            self.table.setItem(row, 2, QTableWidgetItem(f"{percentage_rounded:.2f}%"))

            # --- Use pre-calculated quantities if available ---
            quantity_str = "" # Default empty string
            if standard and row < len(self.calculated_quantities): # Check if standard mode and index is valid
                quantity_str = f"{self.calculated_quantities[row]:.2f}" # Format as needed
            # Always set the item, even if empty
            self.table.setItem(row, 3, QTableWidgetItem(quantity_str))
            # --- End quantity handling ---

        self.table.resizeColumnsToContents() # Adjust column widths


    def copy_table_data(self):
        """Copy selected table data to the clipboard in a tab-separated format."""
        selected_ranges = self.table.selectedRanges()
        if not selected_ranges:
            return # Nothing selected

        # We copy contiguous blocks as selected. If multiple non-contiguous blocks are selected,
        # this standard implementation will only copy the first one.
        # For pasting into Excel, copying a single contiguous block is usually expected.
        selected_range = selected_ranges[0]
        start_row = selected_range.topRow()
        end_row = selected_range.bottomRow()
        start_col = selected_range.leftColumn()
        end_col = selected_range.rightColumn()

        clipboard_string = ""
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            clipboard_string += "\t".join(row_data) + "\n" # Tab separated, newline at end of row

        # Copy to clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_string.strip()) # Remove trailing newline


    def export_to_excel(self):
        """Export the table data to an Excel file."""
        # Prompt the user to select a save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
        )
        if not file_path:
            return

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Peak Analysis Results" # More descriptive title

        # Write the table headers to the Excel sheet
        headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)  # Make headers bold

        # Write the table data to the Excel sheet
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                value = item.text() if item else ""
                # Attempt to convert numeric strings back to numbers for Excel
                try:
                    if '%' in value: # Handle percentages
                        numeric_value = float(value.replace('%', '')) / 100.0
                        cell = worksheet.cell(row=row + 2, column=col + 1, value=numeric_value)
                        cell.number_format = '0.00%' # Apply percentage format
                    else:
                        numeric_value = float(value)
                        worksheet.cell(row=row + 2, column=col + 1, value=numeric_value)
                except ValueError:
                    # Keep as text if conversion fails
                    worksheet.cell(row=row + 2, column=col + 1, value=value)

        # Auto-adjust column widths (optional, but nice)
        for col_idx, column_letter in enumerate(worksheet.columns, start=1):
            max_length = 0
            column = column_letter[0].column_letter # Get the column letter ('A', 'B', ...)
            for cell in worksheet[column]:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2 # Add padding and factor
            worksheet.column_dimensions[column].width = adjusted_width


        # Save the Excel file
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Success", f"Table data exported to\n{file_path}")
        except Exception as e:
             QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                

class PeakAreaDialog(QDialog):
    """
    Interactive dialog to adjust peak regions and calculate peak areas.
    Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
    Area calculations use original data range (inverted).
    Peak detection uses a 0-255 scaled profile (inverted).
    """

    def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
        """
        Initializes the dialog. Args updated.
        """
        super().__init__(parent)
        self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
        self.setGeometry(100, 100, 1100, 850) # Keep original size for now

        # --- Validate and Store Input Image ---
        # (Keep image validation and processing logic as is)
        if not isinstance(cropped_data, Image.Image):
             raise TypeError("Input 'cropped_data' must be a PIL Image object")
        self.cropped_image_for_display = cropped_data # Keep original PIL for display

        self.original_max_value = 255.0 # Default assumption
        pil_mode = cropped_data.mode

        if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
            self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
            self.original_max_value = 65535.0
        elif pil_mode == 'L':
            self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
            self.original_max_value = 255.0
        elif pil_mode == 'F':
             self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
             max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
             self.original_max_value = 1.0 if max_in_float <= 1.0 and max_in_float > 0 else max_in_float
             try:
                 scaled_for_display = np.clip(self.intensity_array_original_range * 255.0 / self.original_max_value, 0, 255).astype(np.uint8)
                 self.cropped_image_for_display = Image.fromarray(scaled_for_display, mode='L')
             except: pass
        else:
            try:
                gray_img = cropped_data.convert("L")
                self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                self.original_max_value = 255.0
                self.cropped_image_for_display = gray_img
            except Exception as e: raise TypeError(f"Could not convert '{pil_mode}' to 'L': {e}")

        if self.intensity_array_original_range.ndim != 2:
             raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

        self.profile_original_inverted = None
        self.profile = None # Scaled, inverted, SMOOTHED profile for detection
        self.background = None

        # --- Settings and State ---
        self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
        # ** NEW: Add smoothing sigma setting **
        self.smoothing_sigma = current_settings.get('smoothing_sigma', 2.0)
        # ... (rest of settings loading remains the same) ...
        self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
        self.peak_distance = current_settings.get('peak_distance', 30)
        self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.02)
        self.peak_spread_pixels = current_settings.get('peak_spread_pixels', 10)
        self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
        self.area_subtraction_method = current_settings.get('area_subtraction_method', "Valley-to-Valley")
        self.peaks = np.array([])
        self.initial_peak_regions = []
        self.peak_regions = []
        self.peak_areas_rolling_ball = []
        self.peak_areas_straight_line = []
        self.peak_areas_valley = []
        self.peak_sliders = []
        self._final_settings = {}
        self._persist_enabled_on_exit = persist_checked

        # Build UI & Initial Setup
        self._setup_ui(persist_checked)
        self.regenerate_profile_and_detect() # Initial calculation

    def _setup_ui(self, persist_checked_initial):
        """Creates and arranges the UI elements."""
        # --- Main Layout ---
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # --- Matplotlib Plot Canvas ---
        self.fig = plt.figure(figsize=(10, 5))
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.canvas, stretch=3)

        # --- Controls Layout (Horizontal Box for side-by-side groups) ---
        controls_hbox = QHBoxLayout()
        controls_hbox.setSpacing(15)

        # --- Left Controls Column (Global & Detection) ---
        left_controls_vbox = QVBoxLayout()

        # Group 1: Global Settings
        global_settings_group = QGroupBox("Global Settings")
        global_settings_layout = QGridLayout(global_settings_group)
        global_settings_layout.setSpacing(8)

        # Band Estimation Method Dropdown (Row 0)
        self.band_estimation_combobox = QComboBox()
        self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
        self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
        self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
        global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
        global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2) # Span 2

        # Area Subtraction Method Dropdown (Row 1)
        self.method_combobox = QComboBox()
        self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
        self.method_combobox.setCurrentText(self.area_subtraction_method)
        self.method_combobox.currentIndexChanged.connect(self.update_plot) # Only needs plot update
        global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
        global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2) # Span 2

        # Rolling Ball Radius Slider (Row 2) - ** MODIFIED LAYOUT **
        global_settings_layout.addWidget(QLabel("Rolling Ball Radius:"), 2, 0) # Static Label
        self.rolling_ball_slider = QSlider(Qt.Horizontal)
        self.rolling_ball_slider.setRange(1, 500)
        self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
        self.rolling_ball_slider.valueChanged.connect(self.update_plot) # Update plot when radius changes
        # Separate Label for the value, prevents layout jumps
        self.rolling_ball_value_label = QLabel(f"({int(self.rolling_ball_radius)})")
        # Ensure value label has enough space initially
        fm = QFontMetrics(self.rolling_ball_value_label.font())
        self.rolling_ball_value_label.setMinimumWidth(fm.horizontalAdvance("(500) ")) # Width for max value
        self.rolling_ball_value_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter) # Align left
        self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_value_label: lbl.setText(f"({val})")) # Update value label
        global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1) # Slider in middle column
        global_settings_layout.addWidget(self.rolling_ball_value_label, 2, 2) # Value label in last column


        left_controls_vbox.addWidget(global_settings_group)

        # Group 2: Peak Detection Parameters
        peak_detect_group = QGroupBox("Peak Detection Parameters")
        peak_detect_layout = QGridLayout(peak_detect_group)
        peak_detect_layout.setSpacing(8)

        # Manual Peak Number Input & Update Button (Row 0)
        self.peak_number_label = QLabel("Detected Peaks:")
        self.peak_number_input = QLineEdit()
        self.peak_number_input.setPlaceholderText("#")
        self.peak_number_input.setMaximumWidth(60)
        self.update_peak_number_button = QPushButton("Set")
        self.update_peak_number_button.setToolTip("Manually override the number of peaks detected.")
        self.update_peak_number_button.clicked.connect(self.manual_peak_number_update)
        peak_detect_layout.addWidget(self.peak_number_label, 0, 0)
        peak_detect_layout.addWidget(self.peak_number_input, 0, 1)
        peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)

        # ** NEW: Smoothing Sigma Slider (Row 1) **
        self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.smoothing_slider = QSlider(Qt.Horizontal)
        # Range 0-100 represents sigma 0.0 to 10.0
        self.smoothing_slider.setRange(0, 100)
        self.smoothing_slider.setValue(int(self.smoothing_sigma * 10)) # Initial value
        self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})")) # Update label text
        # Changing sigma requires re-smoothing and re-detecting peaks
        self.smoothing_slider.valueChanged.connect(self.regenerate_profile_and_detect) # Connect to regenerate
        peak_detect_layout.addWidget(self.smoothing_label, 1, 0)
        peak_detect_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2) # Span slider


        # Peak Prominence Slider (Row 2) - Shifted down one row
        self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        self.peak_prominence_slider = QSlider(Qt.Horizontal)
        self.peak_prominence_slider.setRange(0, 100) # 0.0 to 1.0 factor
        self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
        # Changing prominence only requires re-detection, not full regeneration
        self.peak_prominence_slider.valueChanged.connect(self.detect_peaks)
        self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f})"))
        peak_detect_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
        peak_detect_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)

        # Peak Height Slider (Row 3) - Shifted down one row
        self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f}) %")
        self.peak_height_slider = QSlider(Qt.Horizontal)
        self.peak_height_slider.setRange(0, 100)
        self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
        self.peak_height_slider.valueChanged.connect(self.detect_peaks)
        self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f}) %"))
        peak_detect_layout.addWidget(self.peak_height_slider_label, 3, 0)
        peak_detect_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)

        # Peak Distance Slider (Row 4) - Shifted down one row
        self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
        self.peak_distance_slider = QSlider(Qt.Horizontal)
        self.peak_distance_slider.setRange(1, 200)
        self.peak_distance_slider.setValue(self.peak_distance)
        self.peak_distance_slider.valueChanged.connect(self.detect_peaks)
        self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
        peak_detect_layout.addWidget(self.peak_distance_slider_label, 4, 0)
        peak_detect_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)


        left_controls_vbox.addWidget(peak_detect_group)
        left_controls_vbox.addStretch(1)

        controls_hbox.addLayout(left_controls_vbox, stretch=1)

        # --- Right Controls Column (Peak Region Adjustments - No changes needed here) ---
        right_controls_vbox = QVBoxLayout()
        peak_spread_group = QGroupBox("Peak Region Adjustments")
        peak_spread_layout = QGridLayout(peak_spread_group)
        peak_spread_layout.setSpacing(8)
        self.peak_spread_label = QLabel(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        self.peak_spread_slider = QSlider(Qt.Horizontal)
        self.peak_spread_slider.setRange(0, 100)
        self.peak_spread_slider.setValue(self.peak_spread_pixels)
        self.peak_spread_slider.setToolTip(
            "Adjusts the width of all detected peak regions simultaneously.\n"
            "Regions expand/contract around the initial detected peak center."
        )
        self.peak_spread_slider.valueChanged.connect(self.apply_peak_spread)
        self.peak_spread_slider.valueChanged.connect(
            lambda value, lbl=self.peak_spread_label: lbl.setText(f"Peak Spread (+/- {value} px)")
        )
        peak_spread_layout.addWidget(self.peak_spread_label, 0, 0)
        peak_spread_layout.addWidget(self.peak_spread_slider, 0, 1)
        right_controls_vbox.addWidget(peak_spread_group)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(250)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.container = QWidget()
        self.peak_sliders_layout = QVBoxLayout(self.container)
        self.peak_sliders_layout.setSpacing(10)
        scroll_area.setWidget(self.container)
        right_controls_vbox.addWidget(scroll_area, stretch=1)
        controls_hbox.addLayout(right_controls_vbox, stretch=2)
        main_layout.addLayout(controls_hbox)

        # --- Bottom Button Layout (No changes needed here) ---
        bottom_button_layout = QHBoxLayout()
        self.persist_settings_checkbox = QCheckBox("Persist Settings")
        self.persist_settings_checkbox.setChecked(persist_checked_initial)
        self.persist_settings_checkbox.setToolTip("Save current detection parameters for the next time this dialog opens during this session.")
        bottom_button_layout.addWidget(self.persist_settings_checkbox)
        bottom_button_layout.addStretch(1)
        self.ok_button = QPushButton("OK")
        self.ok_button.setMinimumWidth(100)
        self.ok_button.setDefault(True)
        self.ok_button.clicked.connect(self.accept_and_close)
        bottom_button_layout.addWidget(self.ok_button)
        main_layout.addLayout(bottom_button_layout)


    # --- Methods for Parent Interaction (No changes needed here) ---
    def accept_and_close(self):
        self._final_settings = {
            'rolling_ball_radius': self.rolling_ball_slider.value(),
            'peak_height_factor': self.peak_height_slider.value() / 100.0,
            'peak_distance': self.peak_distance_slider.value(),
            'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
            'peak_spread_pixels': self.peak_spread_slider.value(),
            'band_estimation_method': self.band_estimation_combobox.currentText(),
            'area_subtraction_method': self.method_combobox.currentText(),
            # ** NEW: Save smoothing sigma **
            'smoothing_sigma': self.smoothing_slider.value() / 10.0,
        }
        self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
        self.accept()

    def get_current_settings(self): return self._final_settings
    def should_persist_settings(self): return self._persist_enabled_on_exit


    # --- Core Logic Methods ---

    def regenerate_profile_and_detect(self):
        """
        Calculates the raw inverted profile, applies smoothing based on the slider,
        stores this as the main profile for plotting and calculation,
        then scales a copy for peak detection.
        """
        self.band_estimation_method = self.band_estimation_combobox.currentText()
        self.area_subtraction_method = self.method_combobox.currentText()
        if hasattr(self, 'smoothing_slider'):
             self.smoothing_sigma = self.smoothing_slider.value() / 10.0
        else:
            print("Warning: Smoothing slider not ready.")

        # --- Calculate profile from ORIGINAL intensity data ---
        # (Calculation logic remains the same)
        profile_temp = None
        # ... (Mean/Percentile calculation as before) ...
        if self.band_estimation_method == "Mean":
            profile_temp = np.mean(self.intensity_array_original_range, axis=1)
        elif self.band_estimation_method.startswith("Percentile"):
            try:
                percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
            except: profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1); print("Warning: Defaulting to 5th percentile.")
        else: profile_temp = np.mean(self.intensity_array_original_range, axis=1)

        if not np.all(np.isfinite(profile_temp)):
            print("Warning: Original profile NaN/Inf. Setting to zero.")
            profile_temp = np.zeros(self.intensity_array_original_range.shape[0])


        # --- Create INVERTED Original Profile ---
        profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
        min_inverted_raw = np.min(profile_original_inv_raw)
        profile_original_inv_raw -= min_inverted_raw # Shift baseline to zero

        # --- Apply Smoothing to the INVERTED ORIGINAL profile ---
        # This smoothed version becomes the primary profile for plotting and calculations
        self.profile_original_inverted = profile_original_inv_raw # Start with raw
        try:
            current_sigma = self.smoothing_sigma
            if current_sigma > 0.1 and len(self.profile_original_inverted) > 6:
                # Smooth the full-range inverted profile
                self.profile_original_inverted = gaussian_filter1d(
                    self.profile_original_inverted, sigma=current_sigma
                )
                print(f"  Applied Gaussian smoothing (sigma={current_sigma:.1f}) to main profile.")
            elif current_sigma <= 0.1:
                 print("  Skipping smoothing on main profile (sigma <= 0.1)")
            else: print("Warning: Main profile too short for smoothing.")
        except Exception as smooth_err:
            print(f"Error smoothing main profile: {smooth_err}")
            # self.profile_original_inverted remains the raw version on error


        # --- Create the SCALED (0-255) version FOR PEAK DETECTION ONLY ---
        # Scale the *already smoothed* self.profile_original_inverted
        prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
        if prof_max_inv > prof_min_inv:
            self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
        else:
            self.profile = np.zeros_like(self.profile_original_inverted) # Handle flat profile

        # No need to smooth self.profile again, it was derived from the smoothed original

        # Detect peaks using the scaled profile
        self.detect_peaks()


    def detect_peaks(self):
        """Detect peaks using the 0-255 SCALED, INVERTED, and **potentially smoothed** self.profile."""
        # self.profile should already be smoothed (or not) by regenerate_profile_and_detect

        if self.profile is None or len(self.profile) == 0:
            print("Profile (for detection) not generated yet.")
            # (Rest of handling for no profile remains the same)
            self.peaks, self.initial_peak_regions, self.peak_regions = np.array([]), [], []
            self.peak_number_input.setText("0"); self.update_sliders(); self.update_plot()
            return

        # --- Update parameters from sliders (excluding smoothing, already done) ---
        # (Parameter update logic remains the same)
        self.peak_height_factor = self.peak_height_slider.value() / 100.0
        self.peak_distance = self.peak_distance_slider.value()
        self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
        self.peak_spread_pixels = self.peak_spread_slider.value()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Update UI Labels ---
        # (Label update logic remains the same)
        # ** NEW: Update smoothing label **
        if hasattr(self, 'smoothing_label'): # Check if UI element exists
            self.smoothing_label.setText(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.peak_height_slider_label.setText(f"Min Height ({self.peak_height_factor:.2f}) %")
        self.peak_distance_slider_label.setText(f"Min Distance ({self.peak_distance}) px")
        self.peak_prominence_slider_label.setText(f"Min Prominence ({self.peak_prominence_factor:.2f})")
        self.peak_spread_label.setText(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        # Rolling ball label is now updated automatically via its connection

        # --- Calculate thresholds ON THE SCALED 0-255 PROFILE ---
        # (Threshold calculation remains the same)
        profile_range = np.ptp(self.profile)
        if profile_range < 1e-6 : profile_range = 1.0
        min_val_profile = np.min(self.profile)
        min_height_abs = min_val_profile + profile_range * self.peak_height_factor
        min_prominence_abs = profile_range * self.peak_prominence_factor
        min_prominence_abs = max(1.0, min_prominence_abs) # Ensure minimum

        # --- Detect peaks using the potentially smoothed self.profile ---
        try:
            # (find_peaks call remains the same)
            peaks_indices, properties = find_peaks(
                self.profile, # Use the potentially smoothed profile
                height=min_height_abs,
                prominence=min_prominence_abs,
                distance=self.peak_distance,
                width=1, rel_height=0.5
            )

            # (Rest of setting initial_peak_regions remains the same)
            left_ips = properties.get('left_ips', [])
            right_ips = properties.get('right_ips', [])
            self.peaks = peaks_indices
            self.initial_peak_regions = []
            profile_len = len(self.profile_original_inverted) # Use original length for bounds

            if len(left_ips) == len(self.peaks) and len(right_ips) == len(self.peaks):
                 for i, peak_idx in enumerate(self.peaks):
                     start = int(np.floor(left_ips[i])); end = int(np.ceil(right_ips[i]))
                     start = max(0, start); end = min(profile_len - 1, end)
                     if start >= end: # Fallback
                         fb_width = max(1, self.peak_distance // 4); start = max(0, peak_idx - fb_width); end = min(profile_len - 1, peak_idx + fb_width)
                         if start >= end: end = min(profile_len - 1, start + 1)
                     self.initial_peak_regions.append((start, end))
            else: # Fallback
                 for i, peak_idx in enumerate(self.peaks):
                     wd_est = self.peak_distance // 2; start = max(0, peak_idx - wd_est); end = min(profile_len - 1, peak_idx + wd_est)
                     if start >= end: start = max(0, peak_idx - 2); end = min(profile_len - 1, peak_idx + 2)
                     if start >= end: end = min(profile_len - 1, start + 1)
                     self.initial_peak_regions.append((start, end))

        except Exception as e:
            QMessageBox.warning(self, "Peak Detection Error", f"Peak detection error:\n{e}")
            self.peaks = np.array([]); self.initial_peak_regions = []

        # (Rest of the function remains the same)
        if not self.peak_number_input.hasFocus() or self.peak_number_input.text() == "":
             self.peak_number_input.setText(str(len(self.peaks)))
        self.apply_peak_spread(self.peak_spread_slider.value()) # Calls update_sliders & update_plot

    # --- apply_peak_spread (No changes needed from previous version) ---
    def apply_peak_spread(self, spread_value):
        """Applies the spread value to the initial peak regions."""
        self.peak_spread_pixels = spread_value
        self.peak_regions = []
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            self.update_sliders(); self.update_plot(); return
        profile_len = len(self.profile_original_inverted)
        num_initial = min(len(self.peaks), len(self.initial_peak_regions))
        if len(self.peaks) != len(self.initial_peak_regions): print(f"Warning: Peak/initial region mismatch.")
        for i in range(num_initial):
            peak_idx = self.peaks[i]; center = peak_idx
            new_start = max(0, int(center - self.peak_spread_pixels))
            new_end = min(profile_len - 1, int(center + self.peak_spread_pixels))
            if new_start > new_end: new_start = new_end
            self.peak_regions.append((new_start, new_end))
        if len(self.peak_regions) != len(self.peaks): self.peak_regions = self.peak_regions[:len(self.peaks)]
        self.update_sliders()
        self.update_plot()

    # --- manual_peak_number_update (No changes needed from previous version) ---
    def manual_peak_number_update(self):
        """Handles manual changes to the number of peaks."""
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            QMessageBox.warning(self, "Error", "Profile must be generated first."); return
        profile_len = len(self.profile_original_inverted)
        try:
            num_peaks_manual = int(self.peak_number_input.text())
            if num_peaks_manual < 0: raise ValueError("Negative number")
            current_num_peaks = len(self.peaks)
            if num_peaks_manual == current_num_peaks: return
            if num_peaks_manual == 0: self.peaks, self.initial_peak_regions, self.peak_regions = np.array([]), [], []
            elif num_peaks_manual < current_num_peaks: self.peaks, self.initial_peak_regions = self.peaks[:num_peaks_manual], self.initial_peak_regions[:num_peaks_manual]
            else: # Add dummies
                num_to_add = num_peaks_manual - current_num_peaks; profile_center = profile_len // 2
                peaks_list = self.peaks.tolist(); initial_regions_list = list(self.initial_peak_regions)
                for i in range(num_to_add):
                    new_peak_pos = max(0, min(profile_len - 1, profile_center + np.random.randint(-50, 50)))
                    peaks_list.append(new_peak_pos)
                    pl_width = 5; initial_start = max(0, new_peak_pos - pl_width); initial_end = min(profile_len - 1, new_peak_pos + pl_width)
                    if initial_start >= initial_end: initial_end = min(profile_len-1, initial_start + 1)
                    initial_regions_list.append((initial_start, initial_end))
                if peaks_list:
                    min_len = min(len(peaks_list), len(initial_regions_list))
                    if len(peaks_list) != len(initial_regions_list): print("Warning: Peak/initial region mismatch manual add."); peaks_list, initial_regions_list = peaks_list[:min_len], initial_regions_list[:min_len]
                    combined = sorted(zip(peaks_list, initial_regions_list), key=lambda pair: pair[0])
                    if combined: sorted_peaks, sorted_initial_regions = zip(*combined); self.peaks, self.initial_peak_regions = np.array(sorted_peaks), list(sorted_initial_regions)
                    else: self.peaks, self.initial_peak_regions = np.array([]), []
                else: self.peaks, self.initial_peak_regions = np.array([]), []
            self.apply_peak_spread(self.peak_spread_slider.value())
        except ValueError: self.peak_number_input.setText(str(len(self.peaks))); QMessageBox.warning(self, "Input Error", "Please enter a valid integer.")
        except Exception as e: print(f"Error manual peak update: {e}"); QMessageBox.critical(self, "Error", f"Manual peak update error:\n{e}"); self.peak_number_input.setText(str(len(self.peaks)))

    # --- update_sliders (No changes needed from previous version) ---
    def update_sliders(self):
        """Update sliders based on self.peak_regions."""
        while self.peak_sliders_layout.count():
            item = self.peak_sliders_layout.takeAt(0); widget = item.widget()
            if widget: widget.deleteLater()
            elif not isinstance(item, QSpacerItem): del item
        self.peak_sliders.clear()
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0: return
        profile_len = len(self.profile_original_inverted)
        num_items = len(self.peak_regions)
        if len(self.peaks) != num_items: print(f"Warning: Peak/region count mismatch sliders.")
        for i in range(num_items):
            try:
                start_val, end_val = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                peak_index = int(self.peaks[i]) if i < len(self.peaks) else -1
            except Exception as e: print(f"Warning: Invalid data slider index {i}: {e}"); continue
            peak_group = QGroupBox(f"Peak {i + 1} (Idx: {peak_index if peak_index != -1 else 'N/A'})")
            peak_layout = QGridLayout(peak_group)
            start_slider = QSlider(Qt.Horizontal); start_slider.setRange(0, profile_len - 1); start_val_clamped = max(0, min(profile_len - 1, start_val)); start_slider.setValue(start_val_clamped)
            start_label = QLabel(f"Start: {start_val_clamped}"); start_slider.valueChanged.connect(lambda val, lbl=start_label, idx=i: self._update_region_from_slider(idx, 'start', val, lbl)); start_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(start_label, 0, 0); peak_layout.addWidget(start_slider, 0, 1)
            end_slider = QSlider(Qt.Horizontal); end_slider.setRange(0, profile_len - 1); end_val_clamped = max(start_val_clamped, min(profile_len - 1, end_val)); end_slider.setValue(end_val_clamped)
            end_label = QLabel(f"End: {end_val_clamped}"); end_slider.valueChanged.connect(lambda val, lbl=end_label, idx=i: self._update_region_from_slider(idx, 'end', val, lbl)); end_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(end_label, 1, 0); peak_layout.addWidget(end_slider, 1, 1)
            self.peak_sliders_layout.addWidget(peak_group); self.peak_sliders.append((start_slider, end_slider))
        if num_items > 0:
            last_item = self.peak_sliders_layout.itemAt(self.peak_sliders_layout.count() - 1)
            if not isinstance(last_item, QSpacerItem): self.peak_sliders_layout.addStretch(1)
        if hasattr(self, 'container') and self.container: self.container.update()

    # --- _update_region_from_slider (No changes needed from previous version) ---
    def _update_region_from_slider(self, index, boundary_type, value, label_widget):
        """Helper to update self.peak_regions."""
        if 0 <= index < len(self.peak_regions):
            current_start, current_end = self.peak_regions[index]
            start_slider_widget, end_slider_widget = self.peak_sliders[index]
            if boundary_type == 'start':
                new_start = min(value, current_end); self.peak_regions[index] = (new_start, current_end)
                label_widget.setText(f"Start: {new_start}")
                if start_slider_widget.value() != new_start: start_slider_widget.blockSignals(True); start_slider_widget.setValue(new_start); start_slider_widget.blockSignals(False)
            elif boundary_type == 'end':
                new_end = max(value, current_start); self.peak_regions[index] = (current_start, new_end)
                label_widget.setText(f"End: {new_end}")
                if end_slider_widget.value() != new_end: end_slider_widget.blockSignals(True); end_slider_widget.setValue(new_end); end_slider_widget.blockSignals(False)


    # --- update_plot (Modified for Correct Profiles) ---
    


    def update_plot(self):
        """
        Update plot using the **SMOOTHED** original inverted profile range for
        line display and area calculations. Positions area text BELOW the baseline.
        """
        if self.canvas is None: return

        profile_to_plot_and_calc = self.profile_original_inverted

        if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
            try:
                self.fig.clf(); self.ax = self.fig.add_subplot(111)
                self.ax.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax.transAxes)
                self.canvas.draw_idle()
            except Exception as e: print(f"Error clearing plot: {e}")
            return

        self.method = self.method_combobox.currentText()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Calculate Rolling Ball Background (on the SMOOTHED profile) ---
        try:
            profile_float = profile_to_plot_and_calc
            safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
            if len(profile_float) > 1 :
                background_smoothed = rolling_ball(profile_float, radius=safe_radius)
                self.background = np.maximum(background_smoothed, 0)
            else: self.background = profile_float.copy()
        except ImportError: self.background = np.zeros_like(profile_to_plot_and_calc); print("Scikit-image needed")
        except Exception as e: print(f"Error rolling ball: {e}."); self.background = np.zeros_like(profile_to_plot_and_calc)

        # --- Setup Plot ---
        self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)

        # --- Plot Smoothed Profile ---
        self.ax.plot(profile_to_plot_and_calc, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.2)

        # --- Plot Detected Peak Markers ---
        if len(self.peaks) > 0:
             valid_peaks = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
             if len(valid_peaks) > 0:
                 peak_y_on_smoothed = profile_to_plot_and_calc[valid_peaks]
                 self.ax.scatter(valid_peaks, peak_y_on_smoothed, color="red", marker='x', s=50, label="Detected Peaks", zorder=5)

        # --- Process Peak Regions ---
        self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
        num_items_to_plot = len(self.peak_regions)
        profile_range_plot = np.ptp(profile_to_plot_and_calc) if np.ptp(profile_to_plot_and_calc) > 0 else 1.0
        # Initialize min_text_y_position to a large value or the profile min
        min_text_y_position = np.min(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 0
        max_text_y_position_above = np.min(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 0 # To keep upper limit ok

        for i in range(num_items_to_plot):
            start, end = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
            if start >= end:
                 self.peak_areas_rolling_ball.append(0.0); self.peak_areas_straight_line.append(0.0); self.peak_areas_valley.append(0.0); continue

            x_region = np.arange(start, end + 1)
            profile_region_smoothed = profile_to_plot_and_calc[start : end + 1]

            # Get background region (interpolation if needed)
            bg_start = max(0, min(start, len(self.background)-1)); bg_end = max(0, min(end + 1, len(self.background)))
            background_region = np.zeros_like(profile_region_smoothed) # Default
            if bg_start < bg_end:
                 raw_bg_region = self.background[bg_start:bg_end]
                 if len(raw_bg_region) == len(profile_region_smoothed):
                     background_region = raw_bg_region
                 elif len(self.background) > 1: # Can we interpolate?
                     try:
                        interp_func_bg = interp1d(np.arange(len(self.background)), self.background, kind='linear', fill_value="extrapolate")
                        background_region = interp_func_bg(x_region)
                     except Exception as interp_err_bg: print(f"Warning: BG interp failed peak {i+1}: {interp_err_bg}")
                 else: print(f"Warning: Cannot get BG region for peak {i+1}")

            # --- Area Calculations (Remain the same) ---
            area_rb = max(0, np.trapz(profile_region_smoothed - background_region, x=x_region))
            self.peak_areas_rolling_ball.append(area_rb)
            area_sl = 0.0; y_baseline_pts_sl = np.array([0,0]); y_baseline_interp_sl = np.zeros_like(x_region) # Defaults
            if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                y_baseline_pts_sl = np.array([profile_to_plot_and_calc[start], profile_to_plot_and_calc[end]])
                y_baseline_interp_sl = np.interp(x_region, [start, end], y_baseline_pts_sl)
                area_sl = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_sl, x=x_region))
            self.peak_areas_straight_line.append(area_sl)
            area_vv = 0.0; valley_start_idx = start; valley_end_idx = end; y_baseline_pts_vv = np.array([0,0]); y_baseline_interp_vv = np.zeros_like(x_region) # Defaults
            try:
                search_range = max(15, int((end - start) * 1.5))
                valley_start_idx, valley_end_idx = self._find_valleys_inverted(profile_to_plot_and_calc, start, end, search_range)
                if 0 <= valley_start_idx < len(profile_to_plot_and_calc) and 0 <= valley_end_idx < len(profile_to_plot_and_calc):
                    y_baseline_pts_vv = np.array([profile_to_plot_and_calc[valley_start_idx], profile_to_plot_and_calc[valley_end_idx]])
                    y_baseline_interp_vv = np.interp(x_region, [valley_start_idx, valley_end_idx], y_baseline_pts_vv)
                    area_vv = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_vv, x=x_region))
                else: print(f"Warning: Invalid indices VV peak {i+1}.")
            except Exception as e_vv: print(f"Error VV calc peak {i+1}: {e_vv}.")
            self.peak_areas_valley.append(area_vv)


            # --- Plot Baselines and Fills (based on smoothed profile) ---
            current_area = 0.0
            baseline_y_at_center = 0.0 # Y-value of the baseline at the text's x position
            text_x_pos = (start + end) / 2.0 # Horizontal center for text

            if self.method == "Rolling Ball":
                if i == 0: self.ax.plot(np.arange(len(self.background)), self.background, color="green", ls="--", lw=1, label="Rolling Ball BG")
                self.ax.fill_between(x_region, background_region, profile_region_smoothed, where=profile_region_smoothed >= background_region, color="yellow", alpha=0.4, interpolate=True)
                current_area = area_rb
                # Get baseline Y at center for text positioning
                if len(self.background) > 1:
                    try:
                        interp_func_bg = interp1d(np.arange(len(self.background)), self.background, kind='linear', fill_value="extrapolate")
                        baseline_y_at_center = interp_func_bg(text_x_pos)
                    except: baseline_y_at_center = np.min(background_region) # Fallback
                else: baseline_y_at_center = np.min(background_region)

            elif self.method == "Straight Line":
                 if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                     self.ax.plot([start, end], y_baseline_pts_sl, color="purple", ls="--", lw=1, label="SL BG" if i == 0 else "")
                     self.ax.fill_between(x_region, y_baseline_interp_sl, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_sl, color="cyan", alpha=0.4, interpolate=True)
                     current_area = area_sl
                     # Get baseline Y at center
                     baseline_y_at_center = np.interp(text_x_pos, [start, end], y_baseline_pts_sl)
                 else: current_area = 0.0; baseline_y_at_center = 0.0

            elif self.method == "Valley-to-Valley":
                 if 0 <= valley_start_idx < len(profile_to_plot_and_calc) and 0 <= valley_end_idx < len(profile_to_plot_and_calc):
                     self.ax.plot([valley_start_idx, valley_end_idx], y_baseline_pts_vv, color="orange", ls="--", lw=1, label="VV BG" if i == 0 else "")
                     self.ax.fill_between(x_region, y_baseline_interp_vv, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_vv, color="lightblue", alpha=0.4, interpolate=True)
                     current_area = area_vv
                     # Get baseline Y at center
                     baseline_y_at_center = np.interp(text_x_pos, [valley_start_idx, valley_end_idx], y_baseline_pts_vv)
                 else: current_area = 0.0; baseline_y_at_center = 0.0


            # --- Plot Area Text BELOW the Baseline ---
            area_text_format = "{:.0f}"
            combined_text = f"Peak {i + 1}\n{area_text_format.format(current_area)}"

            # Calculate Y position BELOW the baseline
            text_y_offset = profile_range_plot * 0.06 # Offset distance from baseline
            text_y_pos = baseline_y_at_center - text_y_offset # Subtract offset to go below

            # Ensure text isn't placed ridiculously low if baseline is near zero
            # Maybe set a minimum baseline value for text placement?
            # text_y_pos = max(text_y_pos, some_absolute_minimum_y) # Optional refinement

            self.ax.text(
                text_x_pos,          # Horizontal center
                text_y_pos,          # Y position below baseline
                combined_text,       # Text content
                ha="center",         # Horizontal alignment: center
                va="top",            # Vertical alignment: TOP of text at text_y_pos
                fontsize=7,
                color='black',
                zorder=6
            )

            # Keep track of the minimum Y position reached by text tops
            min_text_y_position = min(min_text_y_position, text_y_pos) # Track the top Y coord of the text

            # Also track the max Y position above peaks for the upper limit
            max_text_y_position_above = max(max_text_y_position_above, np.max(profile_region_smoothed) + profile_range_plot*0.03)


            self.ax.axvline(start, color="gray", ls=":", lw=1.0, alpha=0.8)
            self.ax.axvline(end, color="gray", ls=":", lw=1.0, alpha=0.8)

        # --- Final Plot Configuration ---
        self.ax.set_ylabel("Intensity (Smoothed, Inverted)")
        self.ax.legend(fontsize='small', loc='upper right')
        self.ax.set_title(f"Smoothed Intensity Profile (σ={self.smoothing_sigma:.1f}) and Peak Regions")
        self.ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

        # Set plot limits based on profile range AND text positions
        if len(profile_to_plot_and_calc) > 1: self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1)
        prof_min_smooth, prof_max_smooth = (np.min(profile_to_plot_and_calc), np.max(profile_to_plot_and_calc)) if len(profile_to_plot_and_calc) > 0 else (0, 1)

        # Calculate plot limits considering text positions
        y_max_limit = max(prof_max_smooth, max_text_y_position_above) + profile_range_plot * 0.05
        # Use min_text_y_position (top of the text) and subtract padding
        y_min_limit = min(prof_min_smooth, min_text_y_position) - profile_range_plot * 0.05 # Ensure space below text

        if y_max_limit <= y_min_limit: y_max_limit = y_min_limit + 1
        self.ax.set_ylim(y_min_limit, y_max_limit)
        if prof_max_smooth > 10000: self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0))

        # --- Display Cropped Image (No changes here) ---
        ax_image.clear()
        if hasattr(self, 'cropped_image_for_display') and isinstance(self.cropped_image_for_display, Image.Image):
             try:
                 rotated_pil_image = self.cropped_image_for_display.rotate(90, expand=True)
                 im_array_disp = np.array(rotated_pil_image)
                 # Determine vmin/vmax for display based on original range, handle float case
                 if self.original_max_value == 1.0 and np.issubdtype(self.intensity_array_original_range.dtype, np.floating):
                      im_vmin, im_vmax = 0.0, 1.0 # Assume float is 0-1 range
                 else:
                      im_vmin, im_vmax = 0, self.original_max_value

                 ax_image.imshow(im_array_disp, cmap='gray', aspect='auto',
                                 extent=[0, len(profile_to_plot_and_calc)-1, 0, rotated_pil_image.height],
                                 vmin=im_vmin, vmax=im_vmax) # Use determined vmin/vmax
                 ax_image.set_xlabel("Pixel Index Along Profile Axis")
                 ax_image.set_yticks([]); ax_image.set_ylabel("Lane Width", fontsize='small')
             except Exception as img_e:
                 print(f"Error displaying cropped image preview: {img_e}")
                 ax_image.text(0.5, 0.5, 'Error loading preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])
        else:
             ax_image.text(0.5, 0.5, 'No Image Preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])


        # --- Adjust and Draw ---
        try: self.fig.subplots_adjust(left=0.08, right=0.98, top=0.92, bottom=0.1, hspace=0.05)
        except Exception as layout_e: print(f"Error adjusting layout: {layout_e}")
        try: self.canvas.draw_idle()
        except Exception as draw_e: print(f"Error drawing canvas: {draw_e}")
        
        plt.close(self.fig)


    # --- NEW HELPER for finding minima on inverted profile ---
    def _find_valleys_inverted(self, profile_data, start, end, search_range):
        """Helper to find valley MINIMA near start and end points on INVERTED profile."""
        profile_len = len(profile_data)

        # Left valley (find minimum)
        valley_start_idx = start
        search_start = max(0, start - search_range)
        if start > search_start : # Check if search range is valid
            # Find index of minimum value within the search window BEFORE the peak start
            min_idx_in_search = np.argmin(profile_data[search_start:start]) + search_start
            valley_start_idx = min_idx_in_search
            # Optional: Check if point immediately before start is even lower
            # if start > 0 and profile_data[start-1] < profile_data[valley_start_idx]:
            #     valley_start_idx = start-1
        # else: valley stays at start

        # Right valley (find minimum)
        valley_end_idx = end
        search_end = min(profile_len, end + search_range + 1)
        slice_start = end + 1
        if slice_start < search_end: # Check if search range is valid
            # Find index of minimum value within the search window AFTER the peak end
            min_idx_in_search = np.argmin(profile_data[slice_start : search_end]) + slice_start
            valley_end_idx = min_idx_in_search
            # Optional: Check if point immediately after end is even lower
            # if end < profile_len - 1 and profile_data[end+1] < profile_data[valley_end_idx]:
            #     valley_end_idx = end+1
        # else: valley stays at end


        # Validate and clamp indices (same as before)
        valley_start_idx = max(0, valley_start_idx)
        valley_end_idx = min(profile_len - 1, valley_end_idx)
        if valley_start_idx > start: valley_start_idx = start # Valley cannot be inside peak region
        if valley_end_idx < end: valley_end_idx = end # Valley cannot be inside peak region
        if valley_end_idx <= valley_start_idx:
            print(f"Warning: Inverted valley detection invalid. Using region boundaries ({start},{end}).")
            valley_start_idx, valley_end_idx = start, end

        return valley_start_idx, valley_end_idx

    # --- get_final_peak_area (Identical to previous versions) ---
    def get_final_peak_area(self):
        """Return the list of calculated peak areas based on the selected method."""
        num_valid_peaks = len(self.peak_regions)
        current_area_list = []
        if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
        elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
        elif self.method == "Valley-to-Valley": current_area_list = self.peak_areas_valley
        else: return []
        if len(current_area_list) != num_valid_peaks:
            print(f"Warning: Area list length mismatch for method '{self.method}'.")
            return current_area_list[:num_valid_peaks]
        else:
            return current_area_list
    
        
    


class LiveViewLabel(QLabel):
    def __init__(self, font_type, font_size, marker_color, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)  # Enable mouse tracking
        self.preview_marker_enabled = False
        self.preview_marker_text = ""
        self.preview_marker_position = None
        self.marker_font_type = font_type
        self.marker_font_size = font_size
        self.marker_color = marker_color
        self.setFocusPolicy(Qt.StrongFocus)
        self.bounding_box_preview = []
        self.measure_quantity_mode = False
        self.counter = 0
        self.zoom_level = 1.0  # Initial zoom level
        self.pan_start = None  # Start position for panning
        self.pan_offset = QPointF(0, 0)  # Offset for panning
        self.quad_points = []  # Stores 4 points of the quadrilateral
        self.selected_point = -1  # Index of selected corner (-1 = none)
        self.drag_threshold = 10  # Pixel radius for selecting corners
        self.bounding_box_complete = False
        self.mode=None
        # Add rectangle-related attributes
        self.rectangle_start = None  # Start point of the rectangle
        self.rectangle_end = None    # End point of the rectangle
        self.rectangle_points = []   # Stores the rectangle points
        self.drag_start_pos = None  # For tracking drag operations
        self.draw_edges=True

    def mouseMoveEvent(self, event):
        if self.preview_marker_enabled:
            self.preview_marker_position = event.pos()
            self.update()  # Trigger repaint to show the preview
        if self.selected_point != -1 and self.measure_quantity_mode:# and self.mode=="quad":
            # Update dragged corner position
            self.quad_points[self.selected_point] = self.transform_point(event.pos())
            self.update()  # Show the bounding box preview
        if self.selected_point != -1 and self.mode=="move":
            self.drag_start_pos=event.pos()
        
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if self.preview_marker_enabled:
            # Place the marker text permanently at the clicked position
            parent = self.parent()
            parent.place_custom_marker(event, self.preview_marker_text)
            self.update()  # Clear the preview
        if self.measure_quantity_mode and self.mode=="quad":
            # Check if clicking near existing corner
            for i, p in enumerate(self.quad_points):
                if (self.transform_point(event.pos()) - p).manhattanLength() < self.drag_threshold:
                    self.selected_point = i
                    return
    
            # Add new point if < 4 corners
            if len(self.quad_points) < 4:
                self.quad_points.append(self.transform_point(event.pos()))
                self.selected_point = len(self.quad_points) - 1
    
            if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                # self.quad_points = [self.transform_point(p) for p in self.quad_points]
                self.bounding_box_complete = True
            self.update()  # Trigger repaint
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode=="quad":
            self.selected_point = -1
            self.update()
        super().mouseReleaseEvent(event)

    def transform_point(self, point):
        """Transform a point from widget coordinates to image coordinates."""
        if self.zoom_level != 1.0:
            return QPointF(
                (point.x() - self.pan_offset.x()) / self.zoom_level,
                (point.y() - self.pan_offset.y()) / self.zoom_level
            )
        return QPointF(point)
    
    
    def zoom_in(self):
        self.zoom_level *= 1.1
        self.update()

    def zoom_out(self):
        self.zoom_level /= 1.1
        if self.zoom_level < 1.0:
            self.zoom_level = 1.0
            self.pan_offset = QPointF(0, 0)  # Reset pan offset when zoom is reset
        self.update()
        

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)


        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)

        # Draw the preview marker if enabled
        if self.preview_marker_enabled and self.preview_marker_position:
            painter.setOpacity(0.5)  # Semi-transparent preview
            font = QFont(self.marker_font_type)
            font.setPointSize(self.marker_font_size)
            painter.setFont(font)
            painter.setPen(self.marker_color)
            text_width = painter.fontMetrics().horizontalAdvance(self.preview_marker_text)
            text_height = painter.fontMetrics().height()
            # Draw the text at the cursor's position
            x, y = self.preview_marker_position.x(), self.preview_marker_position.y()
            if self.zoom_level != 1.0:
                x = (x - self.pan_offset.x()) / self.zoom_level
                y = (y - self.pan_offset.y()) / self.zoom_level
            painter.drawText(int(x - text_width / 2), int(y + text_height / 4), self.preview_marker_text)
            
        if len(self.quad_points) > 0 and len(self.quad_points) <4 :
            for p in self.quad_points:
                painter.setPen(QPen(Qt.red, 2))
                painter.drawEllipse(p, 1, 1)
    
        if len(self.quad_points) == 4 and self.draw_edges==True:
            painter.setPen(QPen(Qt.red, 2))
            painter.drawPolygon(QPolygonF(self.quad_points))
            # Draw draggable corners
            for p in self.quad_points:
                painter.drawEllipse(p, self.drag_threshold, self.drag_threshold)
    
        # Draw the bounding box preview if it exists
        if self.bounding_box_preview:
            painter.setPen(QPen(Qt.red, 2))  # Use green color for the bounding box
            start_x, start_y, end_x, end_y = self.bounding_box_preview
            rect = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y))
            painter.drawRect(rect)
    
        painter.end()
        self.update()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            if self.preview_marker_enabled:
                self.preview_marker_enabled = False  # Turn off the preview
                self.update()  # Clear the overlay
            self.measure_quantity_mode = False
            self.counter = 0
            self.bounding_box_complete = False
            self.quad_points = []
            self.mode=None
            self.update()
        super().keyPressEvent(event)

class CombinedSDSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
        # window_width = int(self.screen_width * 0.5)  # 60% of screen width
        # window_height = int(self.screen_height * 0.75)  # 95% of screen height
        self.preview_label_width_setting = int(self.screen_width * 0.45)
        self.preview_label_max_height_setting = int(self.screen_height * 0.35)
        self.label_size = self.preview_label_width_setting
        self.window_title="IMAGING ASSISTANT V6.0"
        # --- Initialize Status Bar Labels ---
        self.size_label = QLabel("Image Size: N/A")
        self.depth_label = QLabel("Bit Depth: N/A")
        self.location_label = QLabel("Source: N/A")

        # --- Add Labels to Status Bar ---
        statusbar = self.statusBar() # Get the status bar instance
        statusbar.addWidget(self.size_label)
        statusbar.addWidget(self.depth_label)
        # Add location label with stretch factor 1 to push it to the right
        statusbar.addWidget(self.location_label, 1)
        self.setWindowTitle(self.window_title)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.undo_stack = []
        self.redo_stack = []
        self.quantities_peak_area_dict = {}
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []
        self.image_path = None
        self.image = None
        self.image_master= None
        self.image_before_padding = None
        self.image_contrasted=None
        self.image_before_contrast=None
        self.contrast_applied=False
        self.image_padded=False
        self.predict_size=False
        self.warped_image=None
        self.transparency=1
        self.left_markers = []
        self.right_markers = []
        self.top_markers = []
        self.custom_markers=[]
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.font_rotation=-45
        self.image_width=0
        self.image_height=0
        self.new_image_width=0
        self.new_image_height=0
        self.base_name="Image"
        self.image_path=""
        self.x_offset_s=0
        self.y_offset_s=0
        self.peak_area=[]
        self.is_modified=False
        
        self.peak_dialog_settings = {
            'rolling_ball_radius': 50,
            'peak_height_factor': 0.1,
            'peak_distance': 30,
            'peak_prominence_factor': 0.02,
            'peak_spread_pixels': 10,
            'band_estimation_method': "Mean",
            'area_subtraction_method': "Valley-to-Valley"
        }
        self.persist_peak_settings_enabled = True # State of the checkbox
        
        # Variables to store bounding boxes and quantities
        self.bounding_boxes = []
        self.up_bounding_boxes = []
        self.standard_protein_areas=[]
        self.quantities = []
        self.protein_quantities = []
        self.measure_quantity_mode = False
        # Initialize self.marker_values to None initially
        self.marker_values_dict = {
            "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
        }
        self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"]
        self.top_label_dict={"Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"]}
        
        
        self.custom_marker_name = "Custom"
        # Load the config file if exists
        
        self.marker_mode = None
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_slider_range=[-1000,1000]
        self.right_slider_range=[-1000,1000]
        self.top_slider_range=[-1000,1000]
               
        
        self.top_padding = 0
        self.font_color = QColor(0, 0, 0)  # Default to black
        self.custom_marker_color = QColor(0, 0, 0)  # Default to black
        self.font_family = "Arial"  # Default font family
        self.font_size = 12  # Default font size
        self.image_array_backup= None
        self.run_predict_MW=False
        
        
        
        # Main container widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Upper section (Preview and buttons)
        upper_layout = QHBoxLayout()

        self.label_width=int(self.label_size)
    
        self.live_view_label = LiveViewLabel(
            font_type=QFont("Arial"),
            font_size=int(24),
            marker_color=QColor(0,0,0),
            parent=self,
        )
        # Image display
        self.live_view_label.setStyleSheet("background-color: white; border: 1px solid black;")
        
        self._create_actions()
        self.create_menu_bar()
        self.create_tool_bar()
        
        self._update_preview_label_size()
        upper_layout.addWidget(self.live_view_label, stretch=1)
        layout.addLayout(upper_layout)
        
        # Lower section (Tabbed interface)
        self.tab_widget = QTabWidget()
        # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
        self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
        self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
        self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
        self.tab_widget.addTab(self.create_markers_tab(), "Markers")
        self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
        self.tab_widget.addTab(self.analysis_tab(), "Analysis")
        
        layout.addWidget(self.tab_widget)
        
        #self.load_shortcut = QShortcut(QKeySequence.Open, self) # Ctrl+O / Cmd+O
        #self.load_shortcut.activated.connect(self.load_action.trigger) # Trigger the action

        #self.save_shortcut = QShortcut(QKeySequence.Save, self) # Ctrl+S / Cmd+S
        #self.save_shortcut.activated.connect(self.save_action.trigger) # Trigger the action

        #self.copy_shortcut = QShortcut(QKeySequence.Copy, self) # Ctrl+C / Cmd+C
        #self.copy_shortcut.activated.connect(self.copy_action.trigger) # Trigger the action

        #self.paste_shortcut = QShortcut(QKeySequence.Paste, self) # Ctrl+V / Cmd+V
        #self.paste_shortcut.activated.connect(self.paste_action.trigger) # Trigger the action

        #self.undo_shortcut = QShortcut(QKeySequence.Undo, self) # Ctrl+Z / Cmd+Z
        #self.undo_shortcut.activated.connect(self.undo_action_m) # Connect directly to method

        #self.redo_shortcut = QShortcut(QKeySequence.Redo, self) # Ctrl+Y / Cmd+Shift+Z
        #self.redo_shortcut.activated.connect(self.redo_action_m) # Connect directly to method

        #self.zoom_out_shortcut = QShortcut(QKeySequence.ZoomOut, self) # Ctrl+- / Cmd+-
        #self.zoom_out_shortcut.activated.connect(self.zoom_out_action.trigger) # Trigger the action

        # Zoom In - Keep the standard one
        #self.zoom_in_shortcut = QShortcut(QKeySequence.ZoomIn, self) # Attempts Ctrl++ / Cmd++
        #self.zoom_in_shortcut.activated.connect(self.zoom_in_action.trigger) # Trigger the action
        # ======================================================


        # === KEEP QShortcut definitions for NON-STANDARD actions ===
        self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
        self.save_svg_shortcut.activated.connect(self.save_svg_action.trigger)

        self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.reset_shortcut.activated.connect(self.reset_action.trigger)

        self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        self.predict_shortcut.activated.connect(self.predict_molecular_weight)

        self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)

        self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
        self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)

        self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
        self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)

        self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
        self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)

        self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
        self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)

        self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
        self.grid_shortcut.activated.connect(
            lambda: self.show_grid_checkbox.setChecked(not self.show_grid_checkbox.isChecked())
            if hasattr(self, 'show_grid_checkbox') else None
        )

        self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.guidelines_shortcut.activated.connect(
            lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked())
            if hasattr(self, 'show_guides_checkbox') else None
        )

        self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
        self.increase_grid_size_shortcut.activated.connect(
            lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1)
            if hasattr(self, 'grid_size_input') else None
        )
        self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
        self.decrease_grid_size_shortcut.activated.connect(
            lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1)
            if hasattr(self, 'grid_size_input') else None
        )

        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self.invert_shortcut.activated.connect(self.invert_image)

        self.bw_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
        self.bw_shortcut.activated.connect(self.convert_to_black_and_white)

        # Tab movement shortcuts
        self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
        self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
        self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
        self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
        self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
        self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
        self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
        self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
        self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
        self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
        self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
        self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
        
        self.load_config()
    
    def _create_actions(self):
        """Create QAction objects for menus and toolbars."""
        style = self.style() # Still useful for other icons
        icon_size = QSize(24, 24) # Match your toolbar size
        text_color = self.palette().color(QPalette.ButtonText) # Use theme color

        # --- Create Zoom Icons (using previous method) ---
        zoom_in_pixmap = QPixmap(icon_size)
        zoom_in_pixmap.fill(Qt.transparent)
        painter_in = QPainter(zoom_in_pixmap)
        # ... (painter setup as before) ...
        painter_in.drawText(zoom_in_pixmap.rect(), Qt.AlignCenter, "+")
        painter_in.end()
        zoom_in_icon = QIcon(zoom_in_pixmap)

        zoom_out_pixmap = QPixmap(icon_size)
        zoom_out_pixmap.fill(Qt.transparent)
        painter_out = QPainter(zoom_out_pixmap)
        # ... (painter setup as before) ...
        painter_out.drawText(zoom_out_pixmap.rect(), Qt.AlignCenter, "-")
        painter_out.end()
        zoom_out_icon = QIcon(zoom_out_pixmap)
        # --- End Zoom Icons ---


        # --- Create Pan Icons using the helper ---
        pan_up_icon = create_text_icon(icon_size, text_color, "↑") # Unicode Up Arrow
        pan_down_icon = create_text_icon(icon_size, text_color, "↓") # Unicode Down Arrow
        pan_left_icon = create_text_icon(icon_size, text_color, "←") # Unicode Left Arrow
        pan_right_icon = create_text_icon(icon_size, text_color, "→") # Unicode Right Arrow
        # --- End Pan Icons ---


        # --- File Actions ---
        self.load_action = QAction(style.standardIcon(QStyle.SP_DialogOpenButton), "&Load Image...", self)
        self.save_action = QAction(style.standardIcon(QStyle.SP_DialogSaveButton), "&Save with Config", self)
        self.save_svg_action = QAction(style.standardIcon(QStyle.SP_DriveDVDIcon), "Save &SVG...", self)
        self.reset_action = QAction(style.standardIcon(QStyle.SP_DialogDiscardButton), "&Reset Image", self)
        self.exit_action = QAction(style.standardIcon(QStyle.SP_DialogCloseButton), "E&xit", self)

        # --- Edit Actions ---
        self.undo_action = QAction(style.standardIcon(QStyle.SP_ArrowBack), "&Undo", self) # Standard for now
        self.redo_action = QAction(style.standardIcon(QStyle.SP_ArrowForward), "&Redo", self) # Standard for now
        self.copy_action = QAction(style.standardIcon(QStyle.SP_FileDialogContentsView), "&Copy Image", self)
        self.paste_action = QAction(style.standardIcon(QStyle.SP_FileDialogDetailedView), "&Paste Image", self)

        # --- View Actions (using the created icons) ---
        self.zoom_in_action = QAction(zoom_in_icon, "Zoom &In", self)
        self.zoom_out_action = QAction(zoom_out_icon, "Zoom &Out", self)
        self.pan_left_action = QAction(pan_left_icon, "Pan Left", self)
        self.pan_right_action = QAction(pan_right_icon, "Pan Right", self)
        self.pan_up_action = QAction(pan_up_icon, "Pan Up", self)
        self.pan_down_action = QAction(pan_down_icon, "Pan Down", self)
        # --- END: Add Panning Actions ---

        # --- Set Shortcuts ---
        self.load_action.setShortcut(QKeySequence.Open)
        self.save_action.setShortcut(QKeySequence.Save)
        self.copy_action.setShortcut(QKeySequence.Copy)
        self.paste_action.setShortcut(QKeySequence.Paste)
        self.undo_action.setShortcut(QKeySequence.Undo)
        self.redo_action.setShortcut(QKeySequence.Redo)
        self.zoom_in_action.setShortcut(QKeySequence("Ctrl+="))
        self.zoom_out_action.setShortcut(QKeySequence.ZoomOut)
        self.save_svg_action.setShortcut(QKeySequence("Ctrl+M"))
        self.reset_action.setShortcut(QKeySequence("Ctrl+R"))

        # --- Set Tooltips ---
        self.load_action.setToolTip("Load an image file (Ctrl+O)")
        self.save_action.setToolTip("Save image and configuration (Ctrl+S)")
        self.save_svg_action.setToolTip("Save as SVG for Word/vector editing (Ctrl+M)")
        self.reset_action.setToolTip("Reset image and all annotations (Ctrl+R)")
        self.exit_action.setToolTip("Exit the application")
        self.undo_action.setToolTip("Undo last action (Ctrl+Z)")
        self.redo_action.setToolTip("Redo last undone action (Ctrl+Y)")
        self.copy_action.setToolTip("Copy rendered image to clipboard (Ctrl+C)")
        self.paste_action.setToolTip("Paste image from clipboard (Ctrl+V)")
        self.zoom_in_action.setToolTip("Increase zoom level (Ctrl+=)")
        self.zoom_out_action.setToolTip("Decrease zoom level (Ctrl+-)")
        # --- START: Set Tooltips for Panning Actions ---
        self.pan_left_action.setToolTip("Pan the view left (when zoomed) (Arrow key left)")
        self.pan_right_action.setToolTip("Pan the view right (when zoomed) (Arrow key right")
        self.pan_up_action.setToolTip("Pan the view up (when zoomed) (Arrow key up)")
        self.pan_down_action.setToolTip("Pan the view down (when zoomed) (Arrow key down")
        # --- END: Set Tooltips for Panning Actions ---

        # --- Connect signals ---
        self.load_action.triggered.connect(self.load_image)
        self.save_action.triggered.connect(self.save_image)
        self.save_svg_action.triggered.connect(self.save_image_svg)
        self.reset_action.triggered.connect(self.reset_image)
        self.exit_action.triggered.connect(self.close)
        self.undo_action.triggered.connect(self.undo_action_m)
        self.redo_action.triggered.connect(self.redo_action_m)
        self.copy_action.triggered.connect(self.copy_to_clipboard)
        self.paste_action.triggered.connect(self.paste_image)
        self.zoom_in_action.triggered.connect(self.zoom_in)
        self.zoom_out_action.triggered.connect(self.zoom_out)
        # --- START: Connect Panning Action Signals ---
        # Use lambda functions to directly modify the offset
        pan_step = 30 # Define pan step size here or access from elsewhere if needed
        self.pan_right_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_right_action.triggered.connect(self.update_live_view) # Trigger update after offset change

        self.pan_left_action.triggered.connect(lambda: self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_left_action.triggered.connect(self.update_live_view)

        self.pan_down_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_down_action.triggered.connect(self.update_live_view)

        self.pan_up_action.triggered.connect(lambda: self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - pan_step) if self.live_view_label.zoom_level > 1.0 else None)
        self.pan_up_action.triggered.connect(self.update_live_view)
        # --- END: Connect Panning Action Signals ---

        # --- START: Initially Disable Panning Actions ---
        self.pan_left_action.setEnabled(False)
        self.pan_right_action.setEnabled(False)
        self.pan_up_action.setEnabled(False)
        self.pan_down_action.setEnabled(False)
        
    def _update_preview_label_size(self):
        """
        Updates the fixed size of the live_view_label, respecting max width/height
        and maintaining aspect ratio. Prioritizes fixing height, but adjusts if
        width constraint is violated.
        """
        # --- Define Defaults and Minimums ---
        default_max_width = 600  # Fallback if width setting is missing/invalid
        default_max_height = 500 # Fallback if height setting is missing/invalid
        min_dim = 50             # Minimum allowed dimension for the label

        # --- Determine Maximum Constraints (with validation) ---
        max_w = default_max_width
        if hasattr(self, 'preview_label_width_setting'):
            try:
                setting_width = int(self.preview_label_width_setting)
                if setting_width > 0:
                    max_w = setting_width
                else:
                    print("Warning: preview_label_width_setting is not positive, using default.")
            except (TypeError, ValueError):
                print("Warning: preview_label_width_setting is invalid, using default.")
        else:
            print("Warning: preview_label_width_setting attribute not found, using default.")
        max_w = max(min_dim, max_w) # Apply minimum constraint

        max_h = default_max_height
        if hasattr(self, 'preview_label_max_height_setting'):
            try:
                setting_height = int(self.preview_label_max_height_setting)
                if setting_height > 0:
                    max_h = setting_height
                else:
                    print("Warning: preview_label_max_height_setting is not positive, using default.")
            except (TypeError, ValueError):
                print("Warning: preview_label_max_height_setting is invalid, using default.")
        max_h = max(min_dim, max_h) # Apply minimum constraint


        # --- Initialize Final Dimensions ---
        final_w = max_w
        final_h = max_h

        # --- Calculate Dimensions Based on Image ---
        if self.image and not self.image.isNull():
            w = self.image.width()
            h = self.image.height()

            if w > 0 and h > 0:
                img_ratio = w / h

                # Attempt 1: Calculate width based on fixed max height
                calc_w_based_on_h = max_h * img_ratio

                if calc_w_based_on_h <= max_w:
                    # Width is within limits, height is fixed
                    final_w = calc_w_based_on_h
                    final_h = max_h
                else:
                    # Calculated width exceeds max width, so fix width and calculate height
                    final_w = max_w
                    final_h = max_w / img_ratio

                # Ensure calculated dimensions are not below minimum
                final_w = max(min_dim, int(final_w))
                final_h = max(min_dim, int(final_h))

            else:
                # Handle invalid image dimensions (0 width or height)
                final_w = max_w # Already set above
                final_h = max_h # Already set above

        else:
            # No image loaded - Use max constraints as default size
            final_w = max_w # Already set above
            final_h = max_h # Already set above
            # self.live_view_label.clear() # Optionally clear the label

        # --- Set the Calculated Fixed Size ---
        self.live_view_label.setFixedSize(final_w, final_h)
        
    def _update_status_bar(self):
        """Updates the status bar labels with current image information."""
        if self.image and not self.image.isNull():
            w = self.image.width()
            h = self.image.height()
            self.size_label.setText(f"Image Size: {w}x{h}")

            # Determine bit depth/format string
            img_format = self.image.format()
            depth_str = "Unknown"
            if img_format == QImage.Format_Grayscale8:
                depth_str = "8-bit Grayscale"
            elif img_format == QImage.Format_Grayscale16:
                depth_str = "16-bit Grayscale"
            elif img_format == QImage.Format_RGB888:
                 depth_str = "24-bit RGB"
            elif img_format in (QImage.Format_RGB32, QImage.Format_ARGB32,
                                QImage.Format_ARGB32_Premultiplied, QImage.Format_RGBA8888):
                 depth_str = "32-bit (A)RGB"
            elif img_format == QImage.Format_Indexed8:
                 depth_str = "8-bit Indexed"
            elif img_format == QImage.Format_Mono:
                 depth_str = "1-bit Mono"
            else:
                # Fallback to depth() if format is unusual
                depth_val = self.image.depth()
                depth_str = f"{depth_val}-bit Depth"
            self.depth_label.setText(f"Bit Depth: {depth_str}")

            # Update location
            location = "N/A"
            if hasattr(self, 'image_path') and self.image_path:
                # Show only filename if it's a path, otherwise show the source info
                if os.path.exists(self.image_path):
                    location = os.path.basename(self.image_path)
                else: # Likely "Clipboard..." or similar
                    location = self.image_path
            else:
                location = "N/A"
            self.location_label.setText(f"Source: {location}")

        else:
            # No image loaded
            self.size_label.setText("Image Size: N/A")
            self.depth_label.setText("Bit Depth: N/A")
            self.location_label.setText("Source: N/A")
        
    def prompt_save_if_needed(self):
        if not self.is_modified:
            return True # No changes, proceed
        """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
        reply = QMessageBox.question(self, 'Unsaved Changes',
                                     "You have unsaved changes. Do you want to save them?",
                                     QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                     QMessageBox.Save) # Default button is Save

        if reply == QMessageBox.Save:
            return self.save_image() # Returns True if saved, False if save cancelled
        elif reply == QMessageBox.Discard:
            return True # Proceed without saving
        else: # Cancelled
            return False # Abort the current action (close/load)

    def closeEvent(self, event):
        """Overrides the default close event to prompt for saving."""
        if self.prompt_save_if_needed():

            event.accept() # Proceed with closing
        else:
            event.ignore() # Abort closing
            
    def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
        """Converts QImage to NumPy array, preserving format and handling row padding."""
        if qimage.isNull():
            return None
    
        img_format = qimage.format()
        height = qimage.height()
        width = qimage.width()
        bytes_per_line = qimage.bytesPerLine() # Actual bytes per row (includes padding)
    
    
        ptr = qimage.constBits()
        if not ptr:
            return None
    
        # Ensure ptr is treated as a bytes object of the correct size
        # ptr is a sip.voidptr, needs explicit size setting for buffer protocol
        try:
            ptr.setsize(qimage.byteCount())
        except AttributeError: # Handle cases where setsize might not be needed or available depending on Qt/Python version
            pass # Continue, hoping buffer protocol works based on qimage.byteCount()
    
        buffer_data = bytes(ptr) # Create a bytes object from the pointer
        if len(buffer_data) != qimage.byteCount():
             print(f"qimage_to_numpy: Warning - Buffer size mismatch. Expected {qimage.byteCount()}, got {len(buffer_data)}.")
             # Fallback or error? Let's try to continue but warn.
    
        # --- Grayscale Formats ---
        if img_format == QImage.Format_Grayscale16:
            bytes_per_pixel = 2
            dtype = np.uint16
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                # No padding, simple reshape
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy() # Return a copy
            else:
                # Handle padding
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale16): Row data length mismatch for row {y}. Skipping row.") # Error handling
                return arr
    
        elif img_format == QImage.Format_Grayscale8:
            bytes_per_pixel = 1
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy()
            else:
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale8): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Color Formats ---
        # ARGB32 (often BGRA in memory) & RGBA8888 (often RGBA in memory)
        elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied):
            bytes_per_pixel = 4
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # QImage ARGB32 is typically BGRA in memory order for numpy
                # QImage RGBA8888 is typically RGBA in memory order for numpy
                # Let's return the raw order for now. The caller might need to swap if needed.
                # If Format_ARGB32, the array is likely BGRA.
                # If Format_RGBA8888, the array is likely RGBA.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB32 (often BGRX or RGBX in memory) & RGBX8888
        elif img_format in (QImage.Format_RGB32, QImage.Format_RGBX8888):
            bytes_per_pixel = 4 # Stored with an ignored byte
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # Memory order is often BGRX (Blue, Green, Red, Ignored Alpha/Padding) for RGB32
                # Let's return the 4 channels for now.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB888 (Tightly packed RGB)
        elif img_format == QImage.Format_RGB888:
            bytes_per_pixel = 3
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 3)
                # QImage RGB888 is RGB order in memory
                return arr.copy()
            else:
                arr = np.zeros((height, width, 3), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 3)
                    else:
                        print(f"qimage_to_numpy (RGB888): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Fallback / Conversion Attempt ---
        else:
            # For other formats, try converting to a known format first
            try:
                # Convert to ARGB32 as a robust intermediate format
                qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                if qimage_conv.isNull():
                    print("qimage_to_numpy: Fallback conversion to ARGB32 failed.")
                    # Last resort: try Grayscale8
                    qimage_conv_gray = qimage.convertToFormat(QImage.Format_Grayscale8)
                    if qimage_conv_gray.isNull():
                         return None
                    else:
                        return self.qimage_to_numpy(qimage_conv_gray) # Recursive call with Grayscale8
    
                return self.qimage_to_numpy(qimage_conv) # Recursive call with ARGB32
            except Exception as e:
                print(f"qimage_to_numpy: Error during fallback conversion: {e}")
                return None

    def numpy_to_qimage(self, array: np.ndarray) -> QImage:
        """Converts NumPy array to QImage, selecting appropriate format."""
        if array is None:
            return QImage() # Return invalid QImage
    
        if not isinstance(array, np.ndarray):
            return QImage()
    
        try:
            if array.ndim == 2: # Grayscale
                height, width = array.shape
                if array.dtype == np.uint16:
                    bytes_per_line = width * 2
                    # Ensure data is contiguous
                    contiguous_array = np.ascontiguousarray(array)
                    # Create QImage directly from contiguous data buffer
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                    # Important: QImage doesn't own the buffer by default. Return a copy.
                    return qimg.copy()
                elif array.dtype == np.uint8:
                    bytes_per_line = width * 1
                    contiguous_array = np.ascontiguousarray(array)
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                    return qimg.copy()
                elif np.issubdtype(array.dtype, np.floating):
                     # Assume float is in 0-1 range, scale to uint8
                     img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                     bytes_per_line = width * 1
                     contiguous_array = np.ascontiguousarray(img_norm)
                     qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                     return qimg.copy()
                else:
                    raise TypeError(f"Unsupported grayscale NumPy dtype: {array.dtype}")
    
            elif array.ndim == 3: # Color
                height, width, channels = array.shape
                if channels == 3 and array.dtype == np.uint8:
                    # Assume input is BGR (common from OpenCV), convert to RGB for QImage.Format_RGB888
                    # Make a contiguous copy before conversion
                    contiguous_array_bgr = np.ascontiguousarray(array)
                    rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                    # rgb_image is now contiguous RGB
                    bytes_per_line = width * 3
                    qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                    return qimg.copy()
                elif channels == 4 and array.dtype == np.uint8:
                    # Assume input is BGRA (matches typical QImage.Format_ARGB32 memory layout)
                    contiguous_array_bgra = np.ascontiguousarray(array)
                    bytes_per_line = width * 4
                    qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                    return qimg.copy()
                # Add handling for 16-bit color if needed (less common for display)
                elif channels == 3 and array.dtype == np.uint16:
                     # Downscale to 8-bit for display format
                     array_8bit = (array / 257.0).astype(np.uint8)
                     contiguous_array_bgr = np.ascontiguousarray(array_8bit)
                     rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                     bytes_per_line = width * 3
                     qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                     return qimg.copy()
                elif channels == 4 and array.dtype == np.uint16:
                     # Downscale color channels, keep alpha potentially?
                     array_8bit = (array / 257.0).astype(np.uint8) # Downscales all channels including alpha
                     contiguous_array_bgra = np.ascontiguousarray(array_8bit)
                     bytes_per_line = width * 4
                     qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                     return qimg.copy()
                else:
                     raise TypeError(f"Unsupported color NumPy dtype/channel combination: {array.dtype} / {channels} channels")
            else:
                raise ValueError(f"Unsupported array dimension: {array.ndim}")
    
        except Exception as e:
            traceback.print_exc()
            return QImage() # Return invalid QImage on error

    def get_image_format(self, image=None):
        """Helper to safely get the format of self.image or a provided image."""
        img_to_check = image if image is not None else self.image
        if img_to_check and not img_to_check.isNull():
            return img_to_check.format()
        return None

    def get_compatible_grayscale_format(self, image=None):
        """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
        current_format = self.get_image_format(image)
        if current_format == QImage.Format_Grayscale16:
            return QImage.Format_Grayscale16
        else: # Treat 8-bit grayscale or color as needing 8-bit target
            return QImage.Format_Grayscale8
    # --- END: Helper Functions ---
        
    def quadrilateral_to_rect(self, image, quad_points):
        """
        Warps the quadrilateral region defined by quad_points in the input image
        to a rectangular image using perspective transformation.
        Crucially, preserves the original bit depth (e.g., uint16) of the input image.
        """
        # --- Input Validation ---
        if not image or image.isNull():
            QMessageBox.warning(self, "Warp Error", "Invalid input image provided for warping.")
            return None
        if len(quad_points) != 4:
             QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
             return None
        if cv2 is None:
             QMessageBox.critical(self, "Dependency Error", "OpenCV (cv2) is required for quadrilateral warping but is not installed.")
             return None
     
        # --- Convert QImage to NumPy array (Preserves bit depth) ---
        try:
            img_array = self.qimage_to_numpy(image) # Should return uint16 if input is Format_Grayscale16
            if img_array is None: raise ValueError("NumPy Conversion returned None")
            print(f"quad_to_rect: Input image dtype = {img_array.dtype}, shape = {img_array.shape}") # Debug
        except Exception as e:
            QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
            return None
     
        # --- !!! REMOVED INTERNAL GRAYSCALE CONVERSION !!! ---
        # # OLD code that caused the issue if input was color:
        # if img_array.ndim == 3:
        #      print("Warning: Warping input array is 3D. Converting to grayscale.")
        #      # THIS CONVERSION TO UINT8 WAS THE PROBLEM
        #      color_code = cv2.COLOR_BGR2GRAY if img_array.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
        #      img_array = cv2.cvtColor(img_array, color_code)
        #      # img_array = img_array.astype(np.uint8) # Explicit cast not needed, cvtColor usually returns uint8
        # --- End Removed Block ---
        # Now, warpPerspective will operate on the img_array with its original dtype (e.g., uint16 grayscale or color)
     
     
        # --- Transform points from LiveViewLabel space to Image space ---
        # (Coordinate transformation logic remains the same - assumes direct scaling for now)
        label_width = self.live_view_label.width()
        label_height = self.live_view_label.height()
        current_img_width = image.width()
        current_img_height = image.height()
        scale_x_disp = current_img_width / label_width if label_width > 0 else 1
        scale_y_disp = current_img_height / label_height if label_height > 0 else 1
     
        src_points_img = []
        for point in quad_points:
            x_view, y_view = point.x(), point.y()
            # Account for zoom/pan in LiveViewLabel coordinates
            if self.live_view_label.zoom_level != 1.0:
                x_view = (x_view - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
                y_view = (y_view - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            # Scale to image coordinates
            x_image = x_view * scale_x_disp
            y_image = y_view * scale_y_disp
            src_points_img.append([x_image, y_image])
        src_np = np.array(src_points_img, dtype=np.float32)
     
        # --- Define Destination Rectangle ---
        # Calculate dimensions based on the source points
        width_a = np.linalg.norm(src_np[0] - src_np[1])
        width_b = np.linalg.norm(src_np[2] - src_np[3])
        height_a = np.linalg.norm(src_np[0] - src_np[3])
        height_b = np.linalg.norm(src_np[1] - src_np[2])
        max_width = int(max(width_a, width_b))
        max_height = int(max(height_a, height_b))
     
        if max_width <= 0 or max_height <= 0:
             QMessageBox.warning(self, "Warp Error", f"Invalid destination rectangle size ({max_width}x{max_height}).")
             return None
     
        dst_np = np.array([
            [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
        ], dtype=np.float32)
     
        # --- Perform Perspective Warp using OpenCV ---
        try:
            matrix = cv2.getPerspectiveTransform(src_np, dst_np)
            # Determine border color based on input array type
            if img_array.ndim == 3: # Color
                 # Use black (0,0,0) or white (255,255,255) depending on preference. Alpha needs 4 values if present.
                 border_val = (0, 0, 0, 0) if img_array.shape[2] == 4 else (0, 0, 0)
            else: # Grayscale
                 border_val = 0 # Black for grayscale
     
            # Warp the ORIGINAL NumPy array (could be uint8, uint16, color)
            warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                               flags=cv2.INTER_LINEAR, # Linear interpolation is usually good
                                               borderMode=cv2.BORDER_CONSTANT,
                                               borderValue=border_val) # Fill borders appropriately
            print(f"quad_to_rect: Warped array dtype = {warped_array.dtype}, shape = {warped_array.shape}") # Debug
        except Exception as e:
             QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
             traceback.print_exc() # Print full traceback for debugging
             return None
     
        # --- Convert warped NumPy array back to QImage ---
        # numpy_to_qimage should handle uint16 correctly, creating Format_Grayscale16
        try:
            warped_qimage = self.numpy_to_qimage(warped_array) # Handles uint8/uint16/color
            if warped_qimage.isNull(): raise ValueError("numpy_to_qimage conversion failed.")
            print(f"quad_to_rect: Returned QImage format = {warped_qimage.format()}") # Debug
            return warped_qimage
        except Exception as e:
            QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array back to QImage: {e}")
            return None
   # --- END: Modified Warping ---
    
    

        
    def create_menu_bar(self):
        menubar = self.menuBar()

        # --- File Menu ---
        file_menu = menubar.addMenu("&File")
        file_menu.addAction(self.load_action)
        file_menu.addAction(self.save_action)
        file_menu.addAction(self.save_svg_action)
        file_menu.addSeparator()
        file_menu.addAction(self.reset_action)
        file_menu.addSeparator()
        file_menu.addAction(self.exit_action)

        # --- Edit Menu ---
        edit_menu = menubar.addMenu("&Edit")
        edit_menu.addAction(self.undo_action)
        edit_menu.addAction(self.redo_action)
        edit_menu.addSeparator()
        edit_menu.addAction(self.copy_action)
        edit_menu.addAction(self.paste_action)

        # --- View Menu ---
        view_menu = menubar.addMenu("&View")
        view_menu.addAction(self.zoom_in_action)
        view_menu.addAction(self.zoom_out_action)

        # --- About Menu ---
        about_menu = menubar.addMenu("&About")
        # Add "GitHub" action directly here or create it in _create_actions
        github_action = QAction("&GitHub", self)
        github_action.setToolTip("Open the project's GitHub page")
        github_action.triggered.connect(self.open_github)
        about_menu.addAction(github_action)
        
        # self.statusBar().showMessage("Ready")
    def open_github(self):
        # Open the GitHub link in the default web browser
        QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
    
    def create_tool_bar(self):
        """Create the main application toolbar."""
        self.tool_bar = QToolBar("Main Toolbar")
        self.tool_bar.setIconSize(QSize(24, 24))
        self.tool_bar.setMovable(False)
        self.tool_bar.setToolButtonStyle(Qt.ToolButtonIconOnly)

        # Add actions to the toolbar
        self.tool_bar.addAction(self.load_action)
        self.tool_bar.addAction(self.paste_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.save_action)
        self.tool_bar.addAction(self.copy_action)
        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.undo_action)
        self.tool_bar.addAction(self.redo_action)
        self.tool_bar.addSeparator()

        # --- Add Zoom and Pan Buttons ---
        self.tool_bar.addAction(self.zoom_in_action)
        self.tool_bar.addAction(self.zoom_out_action)
        self.tool_bar.addSeparator() # Optional separator

        self.tool_bar.addAction(self.pan_left_action)
        self.tool_bar.addAction(self.pan_up_action)   # Group visually
        self.tool_bar.addAction(self.pan_down_action)
        self.tool_bar.addAction(self.pan_right_action)
        # --- End Adding Pan Buttons ---

        self.tool_bar.addSeparator()
        self.tool_bar.addAction(self.reset_action)

        # Add the toolbar to the main window
        self.addToolBar(Qt.TopToolBarArea, self.tool_bar)
        
    def zoom_in(self):
        self.live_view_label.zoom_in() # LiveViewLabel handles its own update
        # --- START: Update Pan Button State ---
        enable_pan = self.live_view_label.zoom_level > 1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
        self.update_live_view()
        # --- END: Update Pan Button State ---

    def zoom_out(self):
        self.live_view_label.zoom_out() # LiveViewLabel handles its own update
        # --- START: Update Pan Button State ---
        enable_pan = self.live_view_label.zoom_level > 1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
        self.update_live_view()
        # --- END: Update Pan Button State ---
    
    def enable_standard_protein_mode(self):
        """"Enable mode to define standard protein amounts for creating a standard curve."""
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        # Assign mouse event handlers for bounding box creation
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_measure_protein_mode(self):
        """Enable mode to measure protein quantity using the standard curve."""
        if len(self.quantities) < 2:
            QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)

    def call_live_view(self):
        self.update_live_view()      

    def analyze_bounding_box(self, pil_image_for_dialog, standard): # Input is PIL Image
    
        peak_area = None # Initialize
        # Clear previous results before new analysis
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []

        if standard:
            quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
            if ok and quantity:
                try:
                    quantity_value = float(quantity.split()[0])
                    # Calculate peak area using the PIL data
                    peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL
                    if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                        total_area = sum(peak_area)
                        self.quantities_peak_area_dict[quantity_value] = round(total_area, 3)
                        self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                        self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
                        print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                        # Store the areas for potential later export if needed (though usually total area is used for standards)
                        self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store individual areas too
                    else:
                         print("Peak area calculation cancelled or failed for standard.")
                         self.latest_peak_areas = [] # Ensure it's cleared

                except (ValueError, IndexError) as e:
                    QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                    self.latest_peak_areas = []
                except Exception as e:
                     QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                     self.latest_peak_areas = []
            else:
                print("Standard quantity input cancelled.")
                self.latest_peak_areas = []

        else: # Analyze sample
            # Calculate peak area using the PIL data
            peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL

            if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store latest areas
                print(f"Sample Analysis: Calculated Areas = {self.latest_peak_areas}")

                if len(self.quantities_peak_area_dict) >= 2:
                    # Calculate quantities and store them
                    self.latest_calculated_quantities = self.calculate_unknown_quantity(
                        list(self.quantities_peak_area_dict.values()), # Standard Areas (total)
                        list(self.quantities_peak_area_dict.keys()),   # Standard Quantities
                        self.latest_peak_areas                         # Sample Peak Areas (individual)
                    )
                    print(f"Sample Analysis: Calculated Quantities = {self.latest_calculated_quantities}")

                else:
                    self.latest_calculated_quantities = [] # No quantities calculated

                try:
                    # Display the areas in the text box
                    self.target_protein_areas_text.setText(str(self.latest_peak_areas))
                except Exception as e:
                    print(f"Error displaying sample areas: {e}")
                    self.target_protein_areas_text.setText("Error")
            else:
                 print("Peak area calculation cancelled or failed for sample.")
                 self.target_protein_areas_text.setText("N/A")
                 self.latest_peak_areas = []
                 self.latest_calculated_quantities = []


        # --- UI updates after analysis ---
        self.update_live_view() # Update display
    
    def calculate_peak_area(self, pil_image_for_dialog): # Input is EXPECTED to be PIL Image
        """Opens the PeakAreaDialog for interactive adjustment and area calculation."""

        # --- Validate Input ---
        if pil_image_for_dialog is None:
            print("Error: No PIL Image data provided to calculate_peak_area.")
            return None
        if not isinstance(pil_image_for_dialog, Image.Image):
             # This indicates an error in the calling function (process_sample/standard)
             QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
             return None
        # --- End Validation ---


        # --- Call PeakAreaDialog passing the PIL Image with the 'cropped_data' keyword ---
        # Assumes PeakAreaDialog __init__ expects 'cropped_data' as the first arg (PIL Image type)
        dialog = PeakAreaDialog(
            cropped_data=pil_image_for_dialog, # Pass the received PIL Image
            current_settings=self.peak_dialog_settings,
            persist_checked=self.persist_peak_settings_enabled,
            parent=self
        )

        peak_areas = None
        if dialog.exec_() == QDialog.Accepted:
            peak_areas = dialog.get_final_peak_area()
            if dialog.should_persist_settings():
                self.peak_dialog_settings = dialog.get_current_settings()
                self.persist_peak_settings_enabled = True
            else:
                self.persist_peak_settings_enabled = False
        else:
            print("PeakAreaDialog cancelled.")

        return peak_areas if peak_areas is not None else []
    
    
    def calculate_unknown_quantity(self, standard_total_areas, known_quantities, sample_peak_areas):
        """Calculates unknown quantities based on standards and sample areas. Returns a list of quantities."""
        if not standard_total_areas or not known_quantities or len(standard_total_areas) != len(known_quantities):
            print("Error: Invalid standard data for quantity calculation.")
            return []
        if len(standard_total_areas) < 2:
            print("Error: Need at least 2 standards for regression.")
            return []
        if not sample_peak_areas:
             print("Warning: No sample peak areas provided for quantity calculation.")
             return []

        calculated_quantities = []
        try:
            # Use total standard area vs known quantity for the standard curve
            coefficients = np.polyfit(standard_total_areas, known_quantities, 1)

            # Apply the standard curve to *each individual* sample peak area
            for area in sample_peak_areas:
                val = np.polyval(coefficients, area)
                calculated_quantities.append(round(val, 2))

            # Display the results in a message box (optional, but helpful feedback)
            # QMessageBox.information(self, "Protein Quantification", f"Predicted Quantities: {calculated_quantities} units")

        except Exception as e:
             # QMessageBox.warning(self, "Calculation Error", f"Could not calculate quantities: {e}")
             return [] # Return empty list on error

        return calculated_quantities # <-- Return the list

        
    def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
        """Draw quantity text at the correct position."""
        text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
        painter.drawText(text_position, str(quantity))
    
    def update_standard_protein_quantities(self):
        self.standard_protein_values.text()
    
    def move_tab(self,tab):
        self.tab_widget.setCurrentIndex(tab)
        
    def save_state(self):
        """Save the current state of the image, markers, and other relevant data."""
        state = {
            "image": self.image.copy() if self.image else None,
            "left_markers": self.left_markers.copy(),
            "right_markers": self.right_markers.copy(),
            "top_markers": self.top_markers.copy(),
            "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
            "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
            "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
            "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
            "font_family": self.font_family,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "font_rotation": self.font_rotation,
            "left_marker_shift_added": self.left_marker_shift_added,
            "right_marker_shift_added": self.right_marker_shift_added,
            "top_marker_shift_added": self.top_marker_shift_added,
            "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
            "quantities": self.quantities.copy(),
            "protein_quantities": self.protein_quantities.copy(),
            "standard_protein_areas": self.standard_protein_areas.copy(),
        }
        self.undo_stack.append(state)
        self.redo_stack.clear()
    
    def undo_action_m(self):
        """Undo the last action by restoring the previous state."""
        if self.undo_stack:
            # Save the current state to the redo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.redo_stack.append(current_state)
            
            # Restore the previous state from the undo stack
            previous_state = self.undo_stack.pop()
            self.image = previous_state["image"]
            self.left_markers = previous_state["left_markers"]
            self.right_markers = previous_state["right_markers"]
            self.top_markers = previous_state["top_markers"]
            self.custom_markers = previous_state["custom_markers"]
            self.image_before_padding = previous_state["image_before_padding"]
            self.image_contrasted = previous_state["image_contrasted"]
            self.image_before_contrast = previous_state["image_before_contrast"]
            self.font_family = previous_state["font_family"]
            self.font_size = previous_state["font_size"]
            self.font_color = previous_state["font_color"]
            self.font_rotation = previous_state["font_rotation"]
            self.left_marker_shift_added = previous_state["left_marker_shift_added"]
            self.right_marker_shift_added = previous_state["right_marker_shift_added"]
            self.top_marker_shift_added = previous_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
            self.quantities = previous_state["quantities"]
            self.protein_quantities = previous_state["protein_quantities"]
            self.standard_protein_areas = previous_state["standard_protein_areas"]
            try:
                self._update_preview_label_size()
            except:
                pass
            self._update_status_bar()
            self.update_live_view()
            
    
    def redo_action_m(self):
        """Redo the last undone action by restoring the next state."""
        if self.redo_stack:
            # Save the current state to the undo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.undo_stack.append(current_state)
            
            # Restore the next state from the redo stack
            next_state = self.redo_stack.pop()
            self.image = next_state["image"]
            self.left_markers = next_state["left_markers"]
            self.right_markers = next_state["right_markers"]
            self.top_markers = next_state["top_markers"]
            self.custom_markers = next_state["custom_markers"]
            self.image_before_padding = next_state["image_before_padding"]
            self.image_contrasted = next_state["image_contrasted"]
            self.image_before_contrast = next_state["image_before_contrast"]
            self.font_family = next_state["font_family"]
            self.font_size = next_state["font_size"]
            self.font_color = next_state["font_color"]
            self.font_rotation = next_state["font_rotation"]
            self.left_marker_shift_added = next_state["left_marker_shift_added"]
            self.right_marker_shift_added = next_state["right_marker_shift_added"]
            self.top_marker_shift_added = next_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
            self.quantities = next_state["quantities"]
            self.protein_quantities = next_state["protein_quantities"]
            self.standard_protein_areas = next_state["standard_protein_areas"]
            try:
                self._update_preview_label_size()
            except:
                pass
            self._update_status_bar()
            self.update_live_view()
            
    def analysis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        # layout.setSpacing(15) # Increase spacing in this tab

        # --- Molecular Weight Prediction ---
        mw_group = QGroupBox("Molecular Weight Prediction")
        mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        mw_layout = QVBoxLayout(mw_group)
        # mw_layout.setSpacing(8)

        self.predict_button = QPushButton("Predict Molecular Weight")
        self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
        self.predict_button.setEnabled(False)  # Initially disabled
        self.predict_button.clicked.connect(self.predict_molecular_weight)
        mw_layout.addWidget(self.predict_button)

        layout.addWidget(mw_group)

        # --- Peak Area / Sample Quantification ---
        quant_group = QGroupBox("Peak Area and Sample Quantification")
        quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        quant_layout = QVBoxLayout(quant_group)
        # quant_layout.setSpacing(8)

        # Area Definition Buttons
        area_def_layout=QHBoxLayout()
        self.btn_define_quad = QPushButton("Define Quad Area")
        self.btn_define_quad.setToolTip(
            "Click 4 corner points to define a region. \n"
            "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
            "Results may differ from Rectangle selection due to warping."
        )
        self.btn_define_quad.clicked.connect(self.enable_quad_mode)
        self.btn_define_rec = QPushButton("Define Rectangle Area")
        self.btn_define_rec.setToolTip(
            "Click and drag to define a rectangular region. \n"
            "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
            "Does not correct for skew."
        )
        self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
        self.btn_sel_rec = QPushButton("Move Selected Area")
        self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
        self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
        area_def_layout.addWidget(self.btn_define_quad)
        area_def_layout.addWidget(self.btn_define_rec)
        area_def_layout.addWidget(self.btn_sel_rec)
        quant_layout.addLayout(area_def_layout)

        # Standard Processing
        std_proc_layout = QHBoxLayout()
        self.btn_process_std = QPushButton("Process Standard Bands")
        self.btn_process_std.setToolTip("Analyze the defined area as a standard lane.\nYou will be prompted for the known quantity.")
        self.btn_process_std.clicked.connect(self.process_standard)
        std_proc_layout.addWidget(self.btn_process_std)
        quant_layout.addLayout(std_proc_layout)

        # Standard Info Display (Read-only makes more sense)
        std_info_layout = QGridLayout()
        std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
        self.standard_protein_values = QLineEdit()
        self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
        self.standard_protein_values.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_values, 0, 1)

        std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
        self.standard_protein_areas_text = QLineEdit()
        self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
        self.standard_protein_areas_text.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
        quant_layout.addLayout(std_info_layout)

        quant_layout.addWidget(self.create_separator()) # Add visual separator

        # Sample Processing
        sample_proc_layout = QHBoxLayout()
        self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
        self.btn_analyze_sample.setToolTip("Analyze the defined area as a sample lane using the standard curve.")
        self.btn_analyze_sample.clicked.connect(self.process_sample)
        sample_proc_layout.addWidget(self.btn_analyze_sample)
        quant_layout.addLayout(sample_proc_layout)


        # Sample Info Display
        sample_info_layout = QGridLayout()
        sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
        self.target_protein_areas_text = QLineEdit()
        self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
        self.target_protein_areas_text.setReadOnly(True)
        sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

        # Add Table Export button next to sample areas
        self.table_export_button = QPushButton("Export Results Table")
        self.table_export_button.setToolTip("Export the analysis results (areas, percentages, quantities) to Excel.")
        self.table_export_button.clicked.connect(self.open_table_window)
        sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
        quant_layout.addLayout(sample_info_layout)

        layout.addWidget(quant_group)

        # --- Clear Button ---
        clear_layout = QHBoxLayout() # Layout to center button maybe
        clear_layout.addStretch()
        self.clear_predict_button = QPushButton("Clear Analysis Markers")
        self.clear_predict_button.setToolTip("Clears MW prediction line and analysis regions.\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
        self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
        layout.addWidget(self.clear_predict_button)
        # clear_layout.addStretch()
        layout.addLayout(clear_layout)


        layout.addStretch()
        return tab
    
    def enable_move_selection_mode(self):
        """Enable mode to move the selected quadrilateral or rectangle."""
        self.live_view_label.mode = "move"
        self.live_view_label.setCursor(Qt.SizeAllCursor)  # Change cursor to indicate move mode
        self.live_view_label.mousePressEvent = self.start_move_selection
        self.live_view_label.mouseReleaseEvent = self.end_move_selection
        self.update_live_view()
        
    def start_move_selection(self, event):
        """Start moving the selection when the mouse is pressed."""
        if self.live_view_label.mode == "move":
            self.live_view_label.draw_edges=False
            # Check if the mouse is near the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), self.live_view_label.quad_points)
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), [
                    QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1])
                ])
            self.live_view_label.drag_start_pos = event.pos()
            self.update_live_view()
        self.live_view_label.mouseMoveEvent = self.move_selection
    
    def move_selection(self, event):
        """Move the selection while the mouse is being dragged."""
        if self.live_view_label.mode == "move" and self.live_view_label.selected_point is not None:
            # Calculate the offset
            offset = event.pos() - self.live_view_label.drag_start_pos
            self.live_view_label.drag_start_pos = event.pos()
            
            # Move the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                for i in range(len(self.live_view_label.quad_points)):
                    self.live_view_label.quad_points[i] += offset
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.bounding_box_preview = (
                    self.live_view_label.bounding_box_preview[0] + offset.x(),
                    self.live_view_label.bounding_box_preview[1] + offset.y(),
                    self.live_view_label.bounding_box_preview[2] + offset.x(),
                    self.live_view_label.bounding_box_preview[3] + offset.y(),
                )
            self.update_live_view()
    
    def end_move_selection(self, event):
        """End moving the selection when the mouse is released."""
        if self.live_view_label.mode == "move":
            self.live_view_label.selected_point = -1            
            self.update_live_view()
            self.live_view_label.draw_edges=True
        self.live_view_label.mouseMoveEvent = None
        
        
             
    def get_nearest_point(self, mouse_pos, points):
        """Get the nearest point to the mouse position."""
        min_distance = float('inf')
        nearest_point = None
        for point in points:
            distance = (mouse_pos - point).manhattanLength()
            if distance < min_distance:
                min_distance = distance
                nearest_point = point
        return nearest_point
    
    def open_table_window(self):
        # Use the latest calculated peak areas
        peak_areas_to_export = self.latest_peak_areas
        calculated_quantities_to_export = self.latest_calculated_quantities

        if not peak_areas_to_export:
            QMessageBox.information(self, "No Data", "No analysis results available to export.")
            return

        standard_dict = self.quantities_peak_area_dict
        # Determine if quantities should be calculated/displayed based on standards
        standard_flag = len(standard_dict) >= 2

        # Open the table window with the latest data
        # Pass the pre-calculated quantities if available
        self.table_window = TableWindow(
            peak_areas_to_export,
            standard_dict,
            standard_flag,
            calculated_quantities_to_export, # Pass the calculated quantities
            self
        )
        self.table_window.show()
    
    def enable_quad_mode(self):
        """Enable mode to define a quadrilateral area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "quad"
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # # Reset mouse event handlers
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        
        # Update the live view
        self.update_live_view()
    
    def enable_rectangle_mode(self):
        """Enable mode to define a rectangle area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "rectangle"
        self.live_view_label.rectangle_points = []  # Clear previous rectangle points
        self.live_view_label.rectangle_start = None  # Reset rectangle start
        self.live_view_label.rectangle_end = None    # Reset rectangle end
        self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Set mouse event handlers for rectangle mode
        self.live_view_label.mousePressEvent = self.start_rectangle
        self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
        
        # Update the live view
        self.update_live_view()

    
    def start_rectangle(self, event):
        """Record the start position of the rectangle."""
        if self.live_view_label.mode == "rectangle":
            self.live_view_label.rectangle_start = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start]
            self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
            self.update_live_view()
    
    def update_rectangle_preview(self, event):
        """Update the rectangle preview as the mouse moves."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Update bounding_box_preview with the rectangle coordinates
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            self.update_live_view()
    
    def finalize_rectangle(self, event):
        """Finalize the rectangle when the mouse is released."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Save the final bounding box preview
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            
            self.live_view_label.mode = None  # Exit rectangle mode
            self.live_view_label.setCursor(Qt.ArrowCursor)
            self.update_live_view()
        
    
    def process_standard(self):
        """Processes the defined region (quad or rect) as a standard."""
        extracted_qimage = None # Will hold the QImage of the extracted region
        self.live_view_label.pan_offset = QPointF(0, 0)
        self.live_view_label.zoom_level=1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
        self.update_live_view()       

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Standard: Quadrilateral")
            # Pass the *current* self.image (could be color or gray)
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Standard: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS from previous fix)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for standard: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=True)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")
        # No else needed, errors handled above
    
    def process_sample(self):
        self.live_view_label.pan_offset = QPointF(0, 0)
        self.live_view_label.zoom_level=1.0
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
        self.update_live_view()  
        """Processes the defined region (quad or rect) as a sample."""
        extracted_qimage = None # Will hold the QImage of the extracted region

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Sample: Quadrilateral")
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Sample: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for sample: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=False)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")


    def convert_qimage_to_grayscale_pil(self, qimg):
        """
        Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
        for use with PeakAreaDialog. Returns None on failure.
        """
        if not qimg or qimg.isNull():
            return None

        fmt = qimg.format()
        pil_img = None

        try:
            # Already grayscale? Convert directly if possible.
            if fmt == QImage.Format_Grayscale16:
                print("Converting Grayscale16 QImage to PIL 'I;16'")
                np_array = self.qimage_to_numpy(qimg)
                if np_array is not None and np_array.dtype == np.uint16:
                    try: pil_img = Image.fromarray(np_array, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(np_array, mode='I')
                else: raise ValueError("Failed NumPy conversion for Grayscale16")
            elif fmt == QImage.Format_Grayscale8:
                print("Converting Grayscale8 QImage to PIL 'L'")
                # Try direct conversion first
                try:
                    pil_img = ImageQt.fromqimage(qimg).convert('L')
                    if pil_img is None: raise ValueError("Direct QImage->PIL(L) failed.")
                except Exception as e_direct:
                    print(f"Direct QImage->PIL(L) failed ({e_direct}), falling back via NumPy.")
                    np_array = self.qimage_to_numpy(qimg)
                    if np_array is not None and np_array.dtype == np.uint8:
                        pil_img = Image.fromarray(np_array, mode='L')
                    else: raise ValueError("Failed NumPy conversion for Grayscale8")
            else: # Color or other format
                print(f"Converting format {fmt} QImage to Grayscale PIL")
                # Use NumPy for robust conversion to 16-bit grayscale intermediate
                np_img = self.qimage_to_numpy(qimg)
                if np_img is None: raise ValueError("NumPy conversion failed for color.")
                if np_img.ndim == 3:
                    gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) # Assume BGR/BGRA input
                    # Convert to 16-bit PIL
                    gray_np_16bit = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                    try: pil_img = Image.fromarray(gray_np_16bit, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(gray_np_16bit, mode='I')
                elif np_img.ndim == 2: # Should have been caught by Grayscale checks, but handle anyway
                     if np_img.dtype == np.uint16:
                         try: pil_img = Image.fromarray(np_img, mode='I;16')
                         except ValueError: pil_img = Image.fromarray(np_img, mode='I')
                     else: # Assume uint8 or other, convert to L
                         pil_img = Image.fromarray(np_img).convert('L')
                else:
                     raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

            if pil_img is None:
                raise ValueError("PIL Image creation failed.")

            print(f"  Successfully converted to PIL Image: mode={pil_img.mode}, size={pil_img.size}")
            return pil_img

        except Exception as e:
            print(f"Error in convert_qimage_to_grayscale_pil: {e}")
            traceback.print_exc()
            return None            
        
        
            
    def combine_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        # Calculate initial slider ranges based on current view size
        render_scale=3
        # Use a sensible default or calculate from screen if view isn't ready
        initial_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 500
        initial_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 500
        render_width = initial_width * render_scale
        render_height = initial_height * render_scale

        # --- Image 1 Group ---
        image1_group = QGroupBox("Image 1 Overlay")
        image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image1_layout = QGridLayout(image1_group) # Use grid for better control layout
        image1_layout.setSpacing(8)

        # Load/Place/Remove buttons
        copy_image1_button = QPushButton("Copy Current Image")
        copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
        copy_image1_button.clicked.connect(self.save_image1) # Keep original name save_image1
        place_image1_button = QPushButton("Place Image 1")
        place_image1_button.setToolTip("Positions Image 1 based on sliders.")
        place_image1_button.clicked.connect(self.place_image1)
        remove_image1_button = QPushButton("Remove Image 1")
        remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
        remove_image1_button.clicked.connect(self.remove_image1)

        image1_layout.addWidget(copy_image1_button, 0, 0)
        image1_layout.addWidget(place_image1_button, 0, 1)
        image1_layout.addWidget(remove_image1_button, 0, 2)

        # Position Sliders
        image1_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image1_left_slider = QSlider(Qt.Horizontal)
        self.image1_left_slider.setRange(-render_width, render_width) # Wider range
        self.image1_left_slider.setValue(0)
        self.image1_left_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2) # Span 2 columns

        image1_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image1_top_slider = QSlider(Qt.Horizontal)
        self.image1_top_slider.setRange(-render_height, render_height) # Wider range
        self.image1_top_slider.setValue(0)
        self.image1_top_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2) # Span 2 columns

        # Resize Slider
        image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image1_resize_slider = QSlider(Qt.Horizontal)
        self.image1_resize_slider.setRange(10, 300)  # Range 10% to 300%
        self.image1_resize_slider.setValue(100)
        self.image1_resize_slider.valueChanged.connect(self.update_live_view)
        self.image1_resize_label = QLabel("100%") # Show current percentage
        self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
        image1_layout.addWidget(self.image1_resize_slider, 3, 1)
        image1_layout.addWidget(self.image1_resize_label, 3, 2)


        layout.addWidget(image1_group)

        # --- Image 2 Group --- (Similar structure to Image 1)
        image2_group = QGroupBox("Image 2 Overlay")
        image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image2_layout = QGridLayout(image2_group)
        image2_layout.setSpacing(8)

        copy_image2_button = QPushButton("Copy Current Image")
        copy_image2_button.clicked.connect(self.save_image2)
        place_image2_button = QPushButton("Place Image 2")
        place_image2_button.clicked.connect(self.place_image2)
        remove_image2_button = QPushButton("Remove Image 2")
        remove_image2_button.clicked.connect(self.remove_image2)
        image2_layout.addWidget(copy_image2_button, 0, 0)
        image2_layout.addWidget(place_image2_button, 0, 1)
        image2_layout.addWidget(remove_image2_button, 0, 2)

        image2_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image2_left_slider = QSlider(Qt.Horizontal)
        self.image2_left_slider.setRange(-render_width, render_width)
        self.image2_left_slider.setValue(0)
        self.image2_left_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)

        image2_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image2_top_slider = QSlider(Qt.Horizontal)
        self.image2_top_slider.setRange(-render_height, render_height)
        self.image2_top_slider.setValue(0)
        self.image2_top_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)

        image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image2_resize_slider = QSlider(Qt.Horizontal)
        self.image2_resize_slider.setRange(10, 300)
        self.image2_resize_slider.setValue(100)
        self.image2_resize_slider.valueChanged.connect(self.update_live_view)
        self.image2_resize_label = QLabel("100%")
        self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
        image2_layout.addWidget(self.image2_resize_slider, 3, 1)
        image2_layout.addWidget(self.image2_resize_label, 3, 2)

        layout.addWidget(image2_group)

        # --- Finalize Button ---
        finalize_layout = QHBoxLayout()
        finalize_layout.addStretch()
        finalize_button = QPushButton("Rasterize Image")
        finalize_button.setToolTip("Permanently merges the placed overlays onto the main image. This action cannot be undone easily.")
        finalize_button.clicked.connect(self.finalize_combined_image)
        layout.addWidget(finalize_button)
        finalize_layout.addStretch()
        layout.addLayout(finalize_layout)

        layout.addStretch() # Push content up
        return tab
    
    def remove_image1(self):
        """Remove Image 1 and reset its sliders."""
        if hasattr(self, 'image1'):
            # del self.image1
            # del self.image1_original
            try:
                del self.image1_position
            except:
                pass
            self.image1_left_slider.setValue(0)
            self.image1_top_slider.setValue(0)
            self.image1_resize_slider.setValue(100)
            self.update_live_view()
    
    def remove_image2(self):
        """Remove Image 2 and reset its sliders."""
        if hasattr(self, 'image2'):
            # del self.image2
            # del self.image2_original
            try:
                del self.image2_position
            except:
                pass
            self.image2_left_slider.setValue(0)
            self.image2_top_slider.setValue(0)
            self.image2_resize_slider.setValue(100)
            self.update_live_view()
    
    def save_image1(self):
        if self.image:
            self.image1 = self.image.copy()
            self.image1_original = self.image1.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 1 copied.")
    
    def save_image2(self):
        if self.image:
            self.image2 = self.image.copy()
            self.image2_original = self.image2.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 2 copied.")
    
    def place_image1(self):
        if hasattr(self, 'image1'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            self.update_live_view()
    
    def place_image2(self):
        if hasattr(self, 'image2'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            self.update_live_view()
    
    # def update_combined_image(self):
    #     if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
    #         return
    
    #     # Create a copy of the original image to avoid modifying it
    #     combined_image = self.image.copy()
    
    #     painter = QPainter(combined_image)
    
    #     # Draw Image 1 if it exists
    #     if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
    #         # Resize Image 1 based on the slider value
    #         scale_factor = self.image1_resize_slider.value() / 100.0
    #         width = int(self.image1_original.width() * scale_factor)
    #         height = int(self.image1_original.height() * scale_factor)
    #         resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image1_position[0], self.image1_position[1], resized_image1)
    
    #     # Draw Image 2 if it exists
    #     if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
    #         # Resize Image 2 based on the slider value
    #         scale_factor = self.image2_resize_slider.value() / 100.0
    #         width = int(self.image2_original.width() * scale_factor)
    #         height = int(self.image2_original.height() * scale_factor)
    #         resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image2_position[0], self.image2_position[1], resized_image2)
    
    #     painter.end()
    
    #     # Update the live view with the combined image
    #     self.live_view_label.setPixmap(QPixmap.fromImage(combined_image))
    
    def finalize_combined_image(self):
        """Overlap image1 and image2 on top of self.image and save the result as self.image."""
        # if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
        #     QMessageBox.warning(self, "Warning", "No images to overlap.")
        #     return
        
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
        # Create a high-resolution canvas for the final image
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Create a blank high-resolution image with a white background
        high_res_canvas = QImage(render_width, render_height, QImage.Format_RGB888)
        high_res_canvas.fill(Qt.white)
    
        # Create a QPainter to draw on the high-resolution canvas
        painter = QPainter(high_res_canvas)
    
        # Draw the base image (self.image) onto the high-resolution canvas
        scaled_base_image = self.image.scaled(render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        painter.drawImage(0, 0, scaled_base_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image1_position[0] + self.x_offset_s),
                int(self.image1_position[1] + self.y_offset_s),
                resized_image1
            )
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image2_position[0] + self.x_offset_s),
                int(self.image2_position[1] + self.y_offset_s),
                resized_image2
            )
    
        # End painting
        painter.end()
        
        
        self.render_image_on_canvas(
                high_res_canvas, scaled_base_image, x_start, y_start, render_scale, draw_guides=False
            )
        
        self.image=high_res_canvas.copy()
        self.image_before_padding=self.image.copy()
        self.image_before_contrast=self.image.copy()
        self.update_live_view()
    
        # Remove Image 1 and Image 2 after finalizing
        self.remove_image1()
        self.remove_image2()
        
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
    
        QMessageBox.information(self, "Success", "The images have been overlapped and saved in memory.")
    
        
    


    def font_and_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Add spacing between groups

        # --- Font Options Group ---
        font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
        font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        font_options_layout = QGridLayout(font_options_group)
        font_options_layout.setSpacing(8)

        # Font type
        font_type_label = QLabel("Font Family:")
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setEditable(False)
        self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
        self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_type_label, 0, 0)
        font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

        # Font size
        font_size_label = QLabel("Font Size:")
        self.font_size_spinner = QSpinBox()
        self.font_size_spinner.setRange(6, 72)  # Adjusted range
        self.font_size_spinner.setValue(self.font_size) # Use initialized value
        self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_size_label, 1, 0)
        font_options_layout.addWidget(self.font_size_spinner, 1, 1)

        # Font color
        self.font_color_button = QPushButton("Font Color")
        self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
        self.font_color_button.clicked.connect(self.select_font_color)
        self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
        font_options_layout.addWidget(self.font_color_button, 1, 2)

        # Font rotation (Top/Bottom)
        font_rotation_label = QLabel("Top Label Rotation:") # Specific label
        self.font_rotation_input = QSpinBox()
        self.font_rotation_input.setRange(-180, 180)
        self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
        self.font_rotation_input.setSuffix(" °") # Add degree symbol
        self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_rotation_label, 2, 0)
        font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

        layout.addWidget(font_options_group)


        # --- Image Adjustments Group ---
        img_adjust_group = QGroupBox("Image Adjustments")
        img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        img_adjust_layout = QGridLayout(img_adjust_group)
        img_adjust_layout.setSpacing(8)

        # Contrast Sliders with Value Labels
        high_contrast_label = QLabel("Brightness:") # Renamed for clarity
        self.high_slider = QSlider(Qt.Horizontal)
        self.high_slider.setRange(0, 200) # Range 0 to 200%
        self.high_slider.setValue(100) # Default 100%
        self.high_slider.valueChanged.connect(self.update_image_contrast)
        self.high_value_label = QLabel("1.00") # Display factor
        self.high_slider.valueChanged.connect(lambda val, lbl=self.high_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(high_contrast_label, 0, 0)
        img_adjust_layout.addWidget(self.high_slider, 0, 1)
        img_adjust_layout.addWidget(self.high_value_label, 0, 2)

        low_contrast_label = QLabel("Contrast:") # Renamed for clarity
        self.low_slider = QSlider(Qt.Horizontal)
        self.low_slider.setRange(0, 200) # Range 0 to 100%
        self.low_slider.setValue(100) # Default 100%
        self.low_slider.valueChanged.connect(self.update_image_contrast)
        self.low_value_label = QLabel("1.00") # Display factor
        self.low_slider.valueChanged.connect(lambda val, lbl=self.low_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(low_contrast_label, 1, 0)
        img_adjust_layout.addWidget(self.low_slider, 1, 1)
        img_adjust_layout.addWidget(self.low_value_label, 1, 2)


        # Gamma Adjustment Slider
        gamma_label = QLabel("Gamma:")
        self.gamma_slider = QSlider(Qt.Horizontal)
        self.gamma_slider.setRange(10, 300)  # Range 0.1 to 3.0
        self.gamma_slider.setValue(100)  # Default 1.0 (100%)
        self.gamma_slider.valueChanged.connect(self.update_image_gamma)
        self.gamma_value_label = QLabel("1.00") # Display factor
        self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(gamma_label, 2, 0)
        img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
        img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


        # Separator
        img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

        # Action Buttons
        btn_layout = QHBoxLayout()
        self.bw_button = QPushButton("Grayscale")
        self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
        self.bw_button.clicked.connect(self.convert_to_black_and_white)
        invert_button = QPushButton("Invert")
        invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
        invert_button.clicked.connect(self.invert_image)
        reset_button = QPushButton("Reset Adjustments")
        reset_button.setToolTip("Reset Brightness, Contrast, and Gamma sliders to default.\n(Does not undo Grayscale or Invert)")
        reset_button.clicked.connect(self.reset_gamma_contrast)
        btn_layout.addWidget(self.bw_button)
        btn_layout.addWidget(invert_button)
        btn_layout.addStretch() # Push reset button to the right
        btn_layout.addWidget(reset_button)

        img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

        layout.addWidget(img_adjust_group)
        layout.addStretch() # Push groups up
        return tab

    
    def _update_color_button_style(self, button, color):
        """ Helper to update button background color preview """
        if color.isValid():
            button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
        else:
            button.setStyleSheet("") # Reset to default stylesheet
    
    def create_separator(self):
        """Creates a horizontal separator line."""
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        return line
    
    
    def create_cropping_tab(self):
        """Create the Cropping tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
    
        # Group Box for Alignment Options
        alignment_params_group = QGroupBox("Alignment Options")
        alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        alignment_layout = QVBoxLayout()
    
        
        #Guide Lines
        self.show_guides_label = QLabel("Show Guide Lines")
        self.show_guides_checkbox = QCheckBox("", self)
        self.show_guides_checkbox.setChecked(False)
        self.show_guides_checkbox.stateChanged.connect(self.update_live_view)
        
        # Rotation Angle 
        rotation_layout = QHBoxLayout()
        self.orientation_label = QLabel("Rotation Angle (0.00°)")
        self.orientation_label.setFixedWidth(150)
        self.orientation_slider = QSlider(Qt.Horizontal)
        self.orientation_slider.setRange(-3600, 3600)  # Scale by 10 to allow decimals
        self.orientation_slider.setValue(0)
        self.orientation_slider.setSingleStep(1)
        self.orientation_slider.valueChanged.connect(self.update_live_view)
        # Align Button
        self.align_button = QPushButton("Apply Rotation")
        self.align_button.clicked.connect(self.align_image)
        
        self.reset_align_button = QPushButton("Reset Rotation")
        self.reset_align_button.clicked.connect(self.reset_align_image)
        
        rotation_layout.addWidget(self.show_guides_label)
        rotation_layout.addWidget(self.show_guides_checkbox)
        rotation_layout.addWidget(self.orientation_label)
        rotation_layout.addWidget(self.orientation_slider)
        rotation_layout.addWidget(self.align_button)
        rotation_layout.addWidget(self.reset_align_button)      
        
        # Flip Vertical Button
        flip_layout=QHBoxLayout()
        self.flip_vertical_button = QPushButton("Flip Vertical")
        self.flip_vertical_button.clicked.connect(self.flip_vertical)
    
        # Flip Horizontal Button
        self.flip_horizontal_button = QPushButton("Flip Horizontal")
        self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
        
        flip_layout.addWidget(self.flip_vertical_button)
        flip_layout.addWidget(self.flip_horizontal_button)
    
        alignment_layout.addLayout(rotation_layout)
        
        # alignment_layout.addWidget(self.flip_vertical_button)  
        # alignment_layout.addWidget(self.flip_horizontal_button)  
        alignment_layout.addLayout(flip_layout)
        alignment_params_group.setLayout(alignment_layout)
        
        
       # Add Tapering Skew Fix Group
        taper_skew_group = QGroupBox("Skew Fix")
        taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        taper_skew_layout = QHBoxLayout()
    
        # Taper Skew Slider
        self.taper_skew_label = QLabel("Tapering Skew (0.00)")
        self.taper_skew_label.setFixedWidth(150)
        self.taper_skew_slider = QSlider(Qt.Horizontal)
        self.taper_skew_slider.setRange(-70, 70)  # Adjust as needed
        self.taper_skew_slider.setValue(0)
        self.taper_skew_slider.valueChanged.connect(self.update_live_view)
        
        # Align Button
        self.skew_button = QPushButton("Apply Skew")
        self.skew_button.clicked.connect(self.update_skew)
    
        # Add widgets to taper skew layout
        taper_skew_layout.addWidget(self.taper_skew_label)
        taper_skew_layout.addWidget(self.taper_skew_slider)
        taper_skew_layout.addWidget(self.skew_button)
        taper_skew_group.setLayout(taper_skew_layout)
        
        # Group Box for Cropping Options
        cropping_params_group = QGroupBox("Cropping Options")
        cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        cropping_layout = QVBoxLayout()
    
        # Cropping Sliders
        crop_slider_layout = QGridLayout()
    
        crop_x_start_label = QLabel("Crop Left (%)")
        self.crop_x_start_slider = QSlider(Qt.Horizontal)
        self.crop_x_start_slider.setRange(0, 100)
        self.crop_x_start_slider.setValue(0)
        self.crop_x_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
        crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
    
        crop_x_end_label = QLabel("Crop Right (%)")
        self.crop_x_end_slider = QSlider(Qt.Horizontal)
        self.crop_x_end_slider.setRange(0, 100)
        self.crop_x_end_slider.setValue(100)
        self.crop_x_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_end_label, 0, 2)
        crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 3)
    
        crop_y_start_label = QLabel("Crop Top (%)")
        self.crop_y_start_slider = QSlider(Qt.Horizontal)
        self.crop_y_start_slider.setRange(0, 100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
        crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
    
        crop_y_end_label = QLabel("Crop Bottom (%)")
        self.crop_y_end_slider = QSlider(Qt.Horizontal)
        self.crop_y_end_slider.setRange(0, 100)
        self.crop_y_end_slider.setValue(100)
        self.crop_y_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_end_label, 1, 2)
        crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 3)
    
        cropping_layout.addLayout(crop_slider_layout)
    
        # Crop Update Button
        crop_button = QPushButton("Update Crop")
        crop_button.clicked.connect(self.update_crop)
        cropping_layout.addWidget(crop_button)
    
        cropping_params_group.setLayout(cropping_layout)
    
        # Add both group boxes to the main layout
        layout.addWidget(alignment_params_group)
        layout.addWidget(cropping_params_group)
        layout.addWidget(taper_skew_group)
        layout.addStretch()
    
        return tab
    
    def create_white_space_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # --- Padding Group ---
        padding_params_group = QGroupBox("Add White Space (Padding)")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setSpacing(8)


        # Input fields with validation (optional but good)
        from PyQt5.QtGui import QIntValidator
        int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

        # Left Padding
        left_padding_label = QLabel("Left Padding (px):")
        self.left_padding_input = QLineEdit("100") # Default
        self.left_padding_input.setValidator(int_validator)
        self.left_padding_input.setToolTip("Pixels to add to the left.")
        padding_layout.addWidget(left_padding_label, 0, 0)
        padding_layout.addWidget(self.left_padding_input, 0, 1)

        # Right Padding
        right_padding_label = QLabel("Right Padding (px):")
        self.right_padding_input = QLineEdit("100") # Default
        self.right_padding_input.setValidator(int_validator)
        self.right_padding_input.setToolTip("Pixels to add to the right.")
        padding_layout.addWidget(right_padding_label, 0, 2)
        padding_layout.addWidget(self.right_padding_input, 0, 3)

        # Top Padding
        top_padding_label = QLabel("Top Padding (px):")
        self.top_padding_input = QLineEdit("100") # Default
        self.top_padding_input.setValidator(int_validator)
        self.top_padding_input.setToolTip("Pixels to add to the top.")
        padding_layout.addWidget(top_padding_label, 1, 0)
        padding_layout.addWidget(self.top_padding_input, 1, 1)

        # Bottom Padding
        bottom_padding_label = QLabel("Bottom Padding (px):")
        self.bottom_padding_input = QLineEdit("0") # Default
        self.bottom_padding_input.setValidator(int_validator)
        self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
        padding_layout.addWidget(bottom_padding_label, 1, 2)
        padding_layout.addWidget(self.bottom_padding_input, 1, 3)

        layout.addWidget(padding_params_group)

        # --- Buttons Layout ---
        button_layout = QHBoxLayout()
        self.recommend_button = QPushButton("Set Recommended Values")
        self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
        self.recommend_button.clicked.connect(self.recommended_values)

        self.clear_padding_button = QPushButton("Clear Values")
        self.clear_padding_button.setToolTip("Set all padding values to 0.")
        self.clear_padding_button.clicked.connect(self.clear_padding_values)

        self.finalize_button = QPushButton("Apply Padding")
        self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
        self.finalize_button.clicked.connect(self.finalize_image)

        button_layout.addWidget(self.recommend_button)
        button_layout.addWidget(self.clear_padding_button)
        button_layout.addStretch()
        button_layout.addWidget(self.finalize_button)

        layout.addLayout(button_layout)
        layout.addStretch() # Push content up

        return tab
    def clear_padding_values(self):
        self.bottom_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.left_padding_input.setText("0")
        
    def recommended_values(self):
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        
        
        
    def create_markers_tab(self):
        """Create the Markers tab with preset, label, placement, offset, and custom controls."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(8)  # Reduce main vertical spacing

        # --- Combined Group Box for Presets and Labels ---
        presets_labels_group = QGroupBox("Marker Presets and Labels")
        presets_labels_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        presets_labels_layout = QGridLayout(presets_labels_group)
        presets_labels_layout.setSpacing(5) # Reduce spacing within the grid

        # -- Left/Right Marker Options (Columns 0 & 1) --
        presets_labels_layout.addWidget(QLabel("Preset:"), 0, 0) # Label for combo
        self.combo_box = QComboBox(self)
        if not hasattr(self, 'marker_values_dict'):
             self.load_config() # Load config attempts to initialize it
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained") # Default or load last used?
        self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
        presets_labels_layout.addWidget(self.combo_box, 0, 1) # Combo box in col 1

        self.marker_values_textbox = QLineEdit(self)
        self.marker_values_textbox.setPlaceholderText("Custom values (comma-separated)") # Shorter text
        self.marker_values_textbox.setEnabled(False)
        presets_labels_layout.addWidget(self.marker_values_textbox, 1, 0, 1, 2) # Span text box across cols 0 & 1

        self.rename_input = QLineEdit(self)
        self.rename_input.setPlaceholderText("New name for Custom")
        self.rename_input.setEnabled(False)
        presets_labels_layout.addWidget(self.rename_input, 2, 0, 1, 2) # Span rename input

        preset_buttons_layout = QHBoxLayout()
        self.save_button = QPushButton("Save Config", self)
        self.save_button.clicked.connect(self.save_config)
        self.remove_config_button = QPushButton("Remove Config", self)
        self.remove_config_button.clicked.connect(self.remove_config)
        preset_buttons_layout.addWidget(self.save_button)
        preset_buttons_layout.addWidget(self.remove_config_button)
        preset_buttons_layout.addStretch() # Push buttons left
        presets_labels_layout.addLayout(preset_buttons_layout, 3, 0, 1, 2) # Add button layout, spanning

        # -- Top Marker Options (Columns 2 & 3) --
        presets_labels_layout.addWidget(QLabel("Top Labels:"), 0, 2) # Label for text edit
        self.top_marker_input = QTextEdit(self)
        if not hasattr(self, 'top_label'):
            self.top_label = ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"] # Default
        self.top_marker_input.setText(", ".join(map(str, self.top_label))) # Ensure all elements are strings for join
        self.top_marker_input.setMinimumHeight(40)
        self.top_marker_input.setMaximumHeight(80) # Reduced max height
        self.top_marker_input.setPlaceholderText("Labels (comma-separated)") # Shorter text
        presets_labels_layout.addWidget(self.top_marker_input, 0, 3, 3, 1) # Row 0, Col 3, Span 3 rows, 1 col

        self.update_top_labels_button = QPushButton("Update All Labels")
        self.update_top_labels_button.clicked.connect(self.update_all_labels)
        presets_labels_layout.addWidget(self.update_top_labels_button, 3, 3) # Place below text edit in col 3

        presets_labels_layout.setColumnStretch(1, 1)
        presets_labels_layout.setColumnStretch(3, 1)

        layout.addWidget(presets_labels_group)

        # --- Marker Placement and Offsets Group ---
        padding_params_group = QGroupBox("Marker Placement and Offsets")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setVerticalSpacing(5)
        padding_layout.setHorizontalSpacing(8)

        # Initialize slider ranges if not already done (e.g., by loading config)
        if not hasattr(self, 'left_slider_range'): self.left_slider_range = [-100, 1000]
        if not hasattr(self, 'right_slider_range'): self.right_slider_range = [-100, 1000]
        if not hasattr(self, 'top_slider_range'): self.top_slider_range = [-100, 1000]
        # Initialize shift values if not already done
        if not hasattr(self, 'left_marker_shift_added'): self.left_marker_shift_added = 0
        if not hasattr(self, 'right_marker_shift_added'): self.right_marker_shift_added = 0
        if not hasattr(self, 'top_marker_shift_added'): self.top_marker_shift_added = 0

        # --- Row 0: Left Marker ---
        left_marker_button = QPushButton("Place Left")
        left_marker_button.setToolTip("Place left markers. Shortcut: Ctrl+Shift+L")
        left_marker_button.clicked.connect(self.enable_left_marker_mode)
        self.left_padding_slider = QSlider(Qt.Horizontal)
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.left_padding_slider.setValue(self.left_marker_shift_added) # Use initialized/loaded value
        self.left_padding_slider.valueChanged.connect(self.update_left_padding)
        remove_left_button = QPushButton("Remove Last")
        remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
        reset_left_button = QPushButton("Reset")
        reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
        duplicate_left_button = QPushButton("Copy Right")
        duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
        padding_layout.addWidget(QLabel("Offset Left:"), 0, 3)
        padding_layout.addWidget(left_marker_button, 0, 0)
        padding_layout.addWidget(remove_left_button, 0, 1)
        padding_layout.addWidget(reset_left_button, 0, 2)
        padding_layout.addWidget(self.left_padding_slider, 0, 4, 1, 3) # Span 3 columns
        padding_layout.addWidget(duplicate_left_button, 0, 7)

        # --- Row 1: Right Marker ---
        right_marker_button = QPushButton("Place Right")
        right_marker_button.setToolTip("Place right markers. Shortcut: Ctrl+Shift+R")
        right_marker_button.clicked.connect(self.enable_right_marker_mode)
        self.right_padding_slider = QSlider(Qt.Horizontal)
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.right_padding_slider.setValue(self.right_marker_shift_added)
        self.right_padding_slider.valueChanged.connect(self.update_right_padding)
        remove_right_button = QPushButton("Remove Last")
        remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
        reset_right_button = QPushButton("Reset")
        reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
        duplicate_right_button = QPushButton("Copy Left")
        duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
        padding_layout.addWidget(QLabel("Offset Right:"), 1, 3)
        padding_layout.addWidget(right_marker_button, 1, 0)
        padding_layout.addWidget(remove_right_button, 1, 1)
        padding_layout.addWidget(reset_right_button, 1, 2)
        padding_layout.addWidget(self.right_padding_slider, 1, 4, 1, 3) # Span 3 columns
        padding_layout.addWidget(duplicate_right_button, 1, 7)

        # --- Row 2: Top Marker ---
        top_marker_button = QPushButton("Place Top")
        top_marker_button.setToolTip("Place top markers. Shortcut: Ctrl+Shift+T")
        top_marker_button.clicked.connect(self.enable_top_marker_mode)
        self.top_padding_slider = QSlider(Qt.Horizontal)
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.top_padding_slider.setValue(self.top_marker_shift_added)
        self.top_padding_slider.valueChanged.connect(self.update_top_padding)
        remove_top_button = QPushButton("Remove Last")
        remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
        reset_top_button = QPushButton("Reset")
        reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
        padding_layout.addWidget(QLabel("Offset Top:"), 2, 3)
        padding_layout.addWidget(top_marker_button, 2, 0)
        padding_layout.addWidget(remove_top_button, 2, 1)
        padding_layout.addWidget(reset_top_button, 2, 2)
        padding_layout.addWidget(self.top_padding_slider, 2, 4, 1, 4)  # Span 4 columns (takes space of duplicate btn too)

        # --- Row 3: Custom Marker Main Controls ---
        self.custom_marker_button = QPushButton("Place Custom", self)
        self.custom_marker_button.setToolTip("Click to activate, then click on image to place the text/arrow")
        self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
        self.custom_marker_text_entry = QLineEdit(self)
        self.custom_marker_text_entry.setPlaceholderText("Custom text")

        marker_buttons_layout = QHBoxLayout()
        marker_buttons_layout.setContentsMargins(0, 0, 0, 0)
        marker_buttons_layout.setSpacing(2)
        self.custom_marker_button_left_arrow = QPushButton("←", self)
        self.custom_marker_button_right_arrow = QPushButton("→", self)
        self.custom_marker_button_top_arrow = QPushButton("↑", self)
        self.custom_marker_button_bottom_arrow = QPushButton("↓", self)
        arrow_size = 25
        self.custom_marker_button_left_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_right_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_top_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_bottom_arrow.setFixedSize(arrow_size, arrow_size)
        self.custom_marker_button_left_arrow.setToolTip("Ctrl+Left: Add ← and activate placement")
        self.custom_marker_button_right_arrow.setToolTip("Ctrl+Right: Add → and activate placement")
        self.custom_marker_button_top_arrow.setToolTip("Ctrl+Up: Add ↑ and activate placement")
        self.custom_marker_button_bottom_arrow.setToolTip("Ctrl+Down: Add ↓ and activate placement")
        marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow)
        marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
        marker_buttons_layout.addStretch()
        self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))

        self.remove_custom_marker_button = QPushButton("Remove Last", self)
        self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
        self.reset_custom_marker_button = QPushButton("Reset", self)
        self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
        self.modify_custom_marker_button = QPushButton("Modify", self) # << THE NEW BUTTON
        self.modify_custom_marker_button.setToolTip("Open a dialog to manage custom markers")
        self.modify_custom_marker_button.clicked.connect(self.open_modify_markers_dialog) # << CONNECTED
        self.custom_marker_color_button = QPushButton("Color")
        self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
        if not hasattr(self, 'custom_marker_color'): self.custom_marker_color = QColor(0,0,0)
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)

        padding_layout.addWidget(self.custom_marker_button, 3, 0)
        padding_layout.addWidget(self.custom_marker_text_entry, 3, 1, 1, 2)
        padding_layout.addLayout(marker_buttons_layout, 3, 3)
        padding_layout.addWidget(self.remove_custom_marker_button, 3, 4)
        padding_layout.addWidget(self.reset_custom_marker_button, 3, 5)
        padding_layout.addWidget(self.modify_custom_marker_button, 3, 6) # Added Modify Button
        padding_layout.addWidget(self.custom_marker_color_button, 3, 7)   # Color Button shifted

        # --- Row 4: Custom Marker Font and Grid ---
        self.custom_font_type_label = QLabel("Custom Font:", self)
        self.custom_font_type_dropdown = QFontComboBox()
        # Initialize font: Check if 'Arial' exists, otherwise use first available
        initial_font = QFont("Arial")
        # if initial_font.family() not in [f.family() for f in self.custom_font_type_dropdown.availableFonts()]:
        #      initial_font = self.custom_font_type_dropdown.availableFonts()[0] if self.custom_font_type_dropdown.availableFonts() else QFont()
        self.custom_font_type_dropdown.setCurrentFont(initial_font)
        self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)

        self.custom_font_size_label = QLabel("Size:", self)
        self.custom_font_size_spinbox = QSpinBox(self)
        self.custom_font_size_spinbox.setRange(2, 150)
        self.custom_font_size_spinbox.setValue(12) # Default

        self.show_grid_checkbox = QCheckBox("Snap Grid", self)
        self.show_grid_checkbox.setToolTip("Places a snapping grid. Shortcut: Ctrl+Shift+G")
        self.show_grid_checkbox.setChecked(False)
        self.show_grid_checkbox.stateChanged.connect(self.update_live_view)

        self.grid_size_input = QSpinBox(self)
        self.grid_size_input.setRange(5, 100)
        self.grid_size_input.setValue(20)
        self.grid_size_input.setPrefix("Grid (px): ")
        self.grid_size_input.valueChanged.connect(self.update_live_view)

        padding_layout.addWidget(self.custom_font_type_label, 4, 0)
        padding_layout.addWidget(self.custom_font_type_dropdown, 4, 1)
        padding_layout.addWidget(self.custom_font_size_label, 4, 2)
        padding_layout.addWidget(self.custom_font_size_spinbox, 4, 3)
        padding_layout.addWidget(self.show_grid_checkbox, 4, 4)
        padding_layout.addWidget(self.grid_size_input, 4, 5, 1, 3) # Span 3 columns

        # Set column stretches - Now 8 columns (0-7)
        padding_layout.setColumnStretch(0, 1) # Place buttons
        padding_layout.setColumnStretch(1, 1) # Remove/Custom Text/Font Combo
        padding_layout.setColumnStretch(2, 1) # Reset/Custom Text/Font Size Label
        padding_layout.setColumnStretch(3, 1) # Offset Label/Arrows/Font Size Spin
        padding_layout.setColumnStretch(4, 2) # Slider/Remove/Grid Check
        padding_layout.setColumnStretch(5, 1) # Slider/Reset/Grid Size
        padding_layout.setColumnStretch(6, 1) # Slider/Modify/Grid Size
        padding_layout.setColumnStretch(7, 1) # Duplicate/Color/Grid Size

        layout.addWidget(padding_params_group)
        layout.addStretch()

        return tab

    def open_modify_markers_dialog(self):
        """Opens the dialog to modify custom markers."""
        if not hasattr(self, "custom_markers") or not isinstance(self.custom_markers, list):
            self.custom_markers = []

        if not self.custom_markers:
            QMessageBox.information(self, "No Markers", "There are no custom markers to modify.")
            return

        dialog = ModifyMarkersDialog(list(self.custom_markers), self)

        if dialog.exec_() == QDialog.Accepted:
            modified_markers = dialog.get_modified_markers()
            if modified_markers != self.custom_markers:
                 self.save_state()
                 self.custom_markers = modified_markers
                 self.is_modified = True
                 self.update_live_view()
    
    def add_column(self):
        """Add a new column to the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        if current_text.strip():
            self.top_marker_input.append("")  # Add a new line for a new column
        else:
            self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
    
    def remove_column(self):
        """Remove the last column from the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        lines = current_text.split("\n")
        if len(lines) > 1:
            lines.pop()  # Remove the last line
            self.top_marker_input.setPlainText("\n".join(lines))
        else:
            self.top_marker_input.clear()  # Clear the text if only one line exists

    
    def flip_vertical(self):
        self.save_state()
        """Flip the image vertically."""
        if self.image:
            self.image = self.image.mirrored(vertical=True, horizontal=False)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def flip_horizontal(self):
        self.save_state()
        """Flip the image horizontally."""
        if self.image:
            self.image = self.image.mirrored(vertical=False, horizontal=True)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def convert_to_black_and_white(self):
        """Explicitly converts the current self.image to grayscale."""
        self.save_state()
        if not self.image or self.image.isNull():
             QMessageBox.warning(self, "Error", "No image loaded.")
             return

        current_format = self.image.format()

        # Check if already grayscale
        if current_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16]:
            QMessageBox.information(self, "Info", "Image is already grayscale.")
            return

        # Preferred target format for grayscale conversion
        # Let's aim for 16-bit if the source might have high dynamic range (like color)
        # If source was already 8-bit (though format check above prevents this), stick to 8-bit.
        target_format = QImage.Format_Grayscale16 # Default target
        converted_image = None

        try:
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")

            if np_img.ndim == 3: # Color image
                # Use weighted average for luminance (standard RGB/BGR -> Gray)
                # OpenCV handles BGR/BGRA automatically in cvtColor
                # color_code = cv2.COLOR_BGR2GRAY if np_img.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
                # gray_np = cv2.cvtColor(np_img, color_code)
                # Use Pillow for potentially better color space handling? Let's stick to OpenCV for now.
                gray_np = cv2.cvtColor(np_img[..., :3], cv2.COLOR_BGR2GRAY) # Use first 3 channels

                # Scale to target bit depth (16-bit)
                gray_np_target = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                converted_image = self.numpy_to_qimage(gray_np_target)
            elif np_img.ndim == 2: # Already grayscale (but maybe different format code)
                 # Try to convert to the target format
                 if target_format == QImage.Format_Grayscale16 and np_img.dtype == np.uint8:
                     converted_image = self.numpy_to_qimage((np_img * 257.0).astype(np.uint16)) # Scale up
                 elif target_format == QImage.Format_Grayscale8 and np_img.dtype == np.uint16:
                     converted_image = self.numpy_to_qimage((np_img / 257.0).astype(np.uint8)) # Scale down
                 else:
                     converted_image = self.numpy_to_qimage(np_img.astype(np.uint16 if target_format == QImage.Format_Grayscale16 else np.uint8)) # Just ensure dtype
            else:
                raise ValueError(f"Unsupported NumPy array dimension: {np_img.ndim}")

            if converted_image is None or converted_image.isNull():
                raise ValueError("Grayscale conversion via NumPy failed.")

        except Exception as e:
            # Fallback to simple QImage conversion (likely Grayscale8)
            converted_image = self.image.convertToFormat(QImage.Format_Grayscale8) # Fallback target

        if not converted_image.isNull():
             self.image = converted_image
             # Update backups consistently
             self.image_before_contrast = self.image.copy()
             self.image_contrasted = self.image.copy()
             # Reset padding state if format changes implicitly? Let's assume padding needs re-application.
             if self.image_padded:
                 self.image_before_padding = None # Invalidate padding backup
                 self.image_padded = False
             else:
                 self.image_before_padding = self.image.copy()

             # Reset contrast/gamma sliders as appearance changed significantly
             self.reset_gamma_contrast() # Resets sliders and updates view
             self.update_live_view() # Ensure view updates even if reset_gamma_contrast fails
        else:
              QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


    def invert_image(self):
        self.save_state()
        if self.image:
            inverted_image = self.image.copy()
            inverted_image.invertPixels()
            self.image = inverted_image
            self.update_live_view()
        self.image_before_contrast=self.image.copy()
        self.image_before_padding=self.image.copy()
        self.image_contrasted=self.image.copy()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if self.live_view_label.preview_marker_enabled:
                self.live_view_label.preview_marker_enabled = False  # Turn off the preview
            self.live_view_label.measure_quantity_mode = False
            self.live_view_label.bounding_box_complete==False
            self.live_view_label.counter=0
            self.live_view_label.quad_points=[]
            self.live_view_label.bounding_box_preview = None
            self.live_view_label.rectangle_points = []
            self.live_view_label.mousePressEvent=None
            self.live_view_label.mode=None

        if self.live_view_label.zoom_level != 1.0:  # Only allow panning when zoomed in
            step = 20  # Adjust the panning step size as needed
            if event.key() == Qt.Key_Left:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - step)
            elif event.key() == Qt.Key_Right:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + step)
            elif event.key() == Qt.Key_Up:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - step)
            elif event.key() == Qt.Key_Down:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + step)
        self.update_live_view()
        super().keyPressEvent(event)
    
    def update_marker_text_font(self, font: QFont):
        """
        Updates the font of self.custom_marker_text_entry based on the selected font
        from self.custom_font_type_dropdown.
    
        :param font: QFont object representing the selected font from the combobox.
        """
        # Get the name of the selected font
        selected_font_name = font.family()
        
        # Set the font of the QLineEdit
        self.custom_marker_text_entry.setFont(QFont(selected_font_name))
    
    def arrow_marker(self, text: str):
        self.custom_marker_text_entry.clear()
        self.custom_marker_text_entry.setText(text)
    
    def enable_custom_marker_mode(self):
        """Enable the custom marker mode and set the mouse event."""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        custom_text = self.custom_marker_text_entry.text().strip()
    
        self.live_view_label.preview_marker_enabled = True
        self.live_view_label.preview_marker_text = custom_text
        self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
        self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
        self.live_view_label.marker_color=self.custom_marker_color
        
        self.live_view_label.setFocus()
        self.live_view_label.update()
        
        self.marker_mode = "custom"  # Indicate custom marker mode
        self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
        
    def remove_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers.pop()  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    def reset_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers=[]  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    
    def place_custom_marker(self, event, custom_text):
        """Place a custom marker at the cursor location."""
        self.save_state()
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            
        # Adjust for snapping to grid
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        image_x = (cursor_x - x_offset) / scale
        image_y = (cursor_y - y_offset) / scale
        
            
        # Store the custom marker's position and text
        self.custom_markers = getattr(self, "custom_markers", [])
        self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value(),False,False))
        self.update_live_view()
        
    def select_custom_marker_color(self):
        """Open a color picker dialog to select the color for custom markers."""
        color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
        if color.isValid():
            self.custom_marker_color = color  # Update the custom marker color
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
    
    def update_all_labels(self):
        """Update all labels from the QTextEdit, supporting multiple columns."""
        self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
        input_text = self.top_marker_input.toPlainText()
        
        # Split the text into a list by commas and strip whitespace
        self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
        

        # if len(self.top_label) < len(self.top_markers):
        #     self.top_markers = self.top_markers[:len(self.top_label)]
        for i in range(0, len(self.top_markers)):
            try:
                self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
            except:
                self.top_markers[i] = (self.top_markers[i][0], str(""))

        # if len(self.marker_values) < len(self.left_markers):
        #     self.left_markers = self.left_markers[:len(self.marker_values)]
        for i in range(0, len(self.left_markers)):
            try:
                self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
            except:
                self.left_markers[i] = (self.left_markers[i][0], str(""))
                

        # if len(self.marker_values) < len(self.right_markers):
        #     self.right_markers = self.right_markers[:len(self.marker_values)]
        for i in range(0, len(self.right_markers)):
            try:
                self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
            except:
                self.right_markers[i] = (self.right_markers[i][0], str(""))
                

        
        # Trigger a refresh of the live view
        self.update_live_view()
        
    def reset_marker(self, marker_type, param):    
        self.save_state()
        if marker_type == 'left':
            if param == 'remove' and len(self.left_markers)!=0:
                self.left_markers.pop()  
                if self.current_left_marker_index > 0:
                    self.current_left_marker_index -= 1
            elif param == 'reset':
                self.left_markers.clear()
                self.current_left_marker_index = 0  

             
        elif marker_type == 'right' and len(self.right_markers)!=0:
            if param == 'remove':
                self.right_markers.pop()  
                if self.current_right_marker_index > 0:
                    self.current_right_marker_index -= 1
            elif param == 'reset':
                self.right_markers.clear()
                self.current_right_marker_index = 0

        elif marker_type == 'top' and len(self.top_markers)!=0:
            if param == 'remove':
                self.top_markers.pop() 
                if self.current_top_label_index > 0:
                    self.current_top_label_index -= 1
            elif param == 'reset':
                self.top_markers.clear()
                self.current_top_label_index = 0
    
        # Call update live view after resetting markers
        self.update_live_view()
        
    def duplicate_marker(self, marker_type):
        self.save_state()
        if marker_type == 'left' and self.right_markers:
            self.left_markers = self.right_markers.copy()  # Copy right markers to left
        elif marker_type == 'right' and self.left_markers:
            self.right_markers = self.left_markers.copy()  # Copy left markers to right

            # self.right_padding = self.image_width*0.9
    
        # Call update live view after duplicating markers
        self.update_live_view()
        
    def on_combobox_changed(self):
        
        text=self.combo_box.currentText()
        """Handle the ComboBox change event to update marker values."""
        if text == "Custom":
            # Enable the textbox for custom values when "Custom" is selected
            self.marker_values_textbox.setEnabled(True)
            self.rename_input.setEnabled(True)
            
            # self.marker_values_textbox.setText()
        else:
            # Update marker values based on the preset option
            self.marker_values = self.marker_values_dict.get(text, [])
            self.marker_values_textbox.setEnabled(False)  # Disable the textbox
            self.rename_input.setEnabled(False)
            self.top_label = self.top_label_dict.get(text, [])
            self.top_label = [str(item) if not isinstance(item, str) else item for item in self.top_label]
            self.top_marker_input.setText(", ".join(self.top_label))
            try:
                
                # Ensure that the top_markers list only updates the top_label values serially
                for i in range(0, len(self.top_markers)):
                    self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                
                # # If self.top_label has more entries than current top_markers, add them
                # if len(self.top_label) > len(self.top_markers):
                #     additional_markers = [(self.top_markers[-1][0] + 50 * (i + 1), label) 
                #                           for i, label in enumerate(self.top_label[len(self.top_markers):])]
                #     self.top_markers.extend(additional_markers)
                
                # If self.top_label has fewer entries, truncate the list
                if len(self.top_label) < len(self.top_markers):
                    self.top_markers = self.top_markers[:len(self.top_label)]
                    
                
                
            except:
                pass
            try:
                self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
            except:
                pass

    # Functions for updating contrast and gamma
    
    def reset_gamma_contrast(self):
        try:
            if self.image_before_contrast==None:
                self.image_before_contrast=self.image_master.copy()
            self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
            self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
            self.high_slider.setValue(100)  # Reset contrast to default
            self.low_slider.setValue(100)  # Reset contrast to default
            self.gamma_slider.setValue(100)  # Reset gamma to default
            self.update_live_view()
        except:
            pass

    
    def update_image_contrast(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
            
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                self.update_live_view()
        except:
            pass
    
    def update_image_gamma(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
                
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                self.update_live_view()
        except:
            pass
    
    def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
        """
        Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
        preserving the original format (including color and bit depth) where possible.
        Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
        """
        if not qimage or qimage.isNull():
            return qimage

        original_format = qimage.format()
        try:
            img_array = self.qimage_to_numpy(qimage)
            if img_array is None: raise ValueError("NumPy conversion failed.")

            # Work with float64 for calculations to avoid precision issues
            img_array_float = img_array.astype(np.float64)

            # Determine max value based on original data type
            if img_array.dtype == np.uint16:
                max_val = 65535.0
            elif img_array.dtype == np.uint8:
                max_val = 255.0
            else: # Default for unexpected types (e.g., float input?)
                 max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0

            # --- Apply adjustments ---
            if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                num_channels = img_array.shape[2]
                adjusted_channels = []

                # Process only the color channels (first 3 usually)
                channels_to_process = min(num_channels, 3)
                for i in range(channels_to_process):
                    channel = img_array_float[:, :, i]
                    # Normalize to 0-1 range
                    channel_norm = channel / max_val

                    # Apply brightness (high_factor): Multiply
                    channel_norm = channel_norm * high_factor

                    # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                    mid_grey = 0.5
                    contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                    channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                    # Clip to 0-1 range after contrast/brightness
                    channel_norm = np.clip(channel_norm, 0.0, 1.0)

                    # Apply gamma correction
                    safe_gamma = max(0.01, gamma)
                    channel_norm = np.power(channel_norm, safe_gamma)

                    # Clip again after gamma
                    channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                    # Scale back to original range
                    adjusted_channels.append(channel_norm_clipped * max_val)

                # Reconstruct the image array
                img_array_final_float = np.stack(adjusted_channels, axis=2)

                # Keep the alpha channel (if present) untouched
                if num_channels == 4:
                    alpha_channel = img_array_float[:, :, 3] # Get original alpha
                    img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

            elif img_array.ndim == 2: # Grayscale Image
                # Normalize to 0-1 range
                img_array_norm = img_array_float / max_val

                # Apply brightness
                img_array_norm = img_array_norm * high_factor

                # Apply contrast
                mid_grey = 0.5
                contrast_factor = max(0.01, low_factor)
                img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                # Clip
                img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                # Apply gamma
                safe_gamma = max(0.01, gamma)
                img_array_norm = np.power(img_array_norm, safe_gamma)

                # Clip again
                img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                # Scale back
                img_array_final_float = img_array_norm_clipped * max_val
            else:
                return qimage # Return original if unsupported dimensions

            # Convert back to original data type
            img_array_final = img_array_final_float.astype(img_array.dtype)

            # Convert back to QImage using the helper function
            result_qimage = self.numpy_to_qimage(img_array_final)
            if result_qimage.isNull():
                raise ValueError("Conversion back to QImage failed.")

            # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
            return result_qimage

        except Exception as e:
            traceback.print_exc() # Print detailed traceback
            return qimage # Return original QImage on error
    

    def save_contrast_options(self):
        if self.image:
            self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
            self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
        else:
            QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

    def remove_config(self):
        try:
            # Get the currently selected marker label
            selected_marker = self.combo_box.currentText()
    
            # Ensure the selected marker is not "Custom" before deleting
            if selected_marker == "Custom":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' marker.")
                return
            
            elif selected_marker == "Precision Plus All Blue/Unstained":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
            
            elif selected_marker == "1 kB Plus":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
    
            # Remove the marker label and top_label if they exist
            if selected_marker in self.marker_values_dict:
                del self.marker_values_dict[selected_marker]
            if selected_marker in self.top_label_dict:
                del self.top_label_dict[selected_marker]
    
            # Save the updated configuration
            with open("Imaging_assistant_config.txt", "w") as f:
                config = {
                    "marker_values": self.marker_values_dict,
                    "top_label": self.top_label_dict
                }
                json.dump(config, f)
    
            # Remove from the ComboBox and reset UI
            self.combo_box.removeItem(self.combo_box.currentIndex())
            self.top_marker_input.clear()
    
            QMessageBox.information(self, "Success", f"Configuration '{selected_marker}' removed.")
    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error removing config: {e}")

    def save_config(self):
        """Rename the 'Custom' marker values option and save the configuration."""
        new_name = self.rename_input.text().strip()
        
        # Ensure that the correct dictionary (self.marker_values_dict) is being used
        if self.rename_input.text() != "Enter new name for Custom" and self.rename_input.text() != "":  # Correct condition
            # Save marker values list
            self.marker_values_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
            
            # Save top_label list under the new_name key
            self.top_label_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.top_marker_input.toPlainText().strip("[]").split(",")]

            try:
                # Save both the marker values and top label (under new_name) to the config file
                with open("Imaging_assistant_config.txt", "w") as f:
                    config = {
                        "marker_values": self.marker_values_dict,
                        "top_label": self.top_label_dict  # Save top_label_dict as a dictionary with new_name as key
                    }
                    json.dump(config, f)  # Save both marker_values and top_label_dict
            except Exception as e:
                pass

        self.combo_box.setCurrentText(new_name)
        self.load_config()  # Reload the configuration after saving
    
    
    def load_config(self):
        """Load the configuration from the file."""
        try:
            with open("Imaging_assistant_config.txt", "r") as f:
                config = json.load(f)
                
                # Load marker values and top label from the file
                self.marker_values_dict = config.get("marker_values", {})
                
                # Retrieve top_label list from the dictionary using the new_name key
                new_name = self.rename_input.text().strip()  # Assuming `new_name` is defined here; otherwise, set it manually
                self.top_label_dict = config.get("top_label", {})  # Default if not found
                # self.top_label = self.top_label_dict.get(new_name, ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])  # Default if not found
                
        except FileNotFoundError:
            # Set default marker values and top_label if config is not found
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }  # Default top labels
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
        except Exception as e:
            # Fallback to default values if there is an error
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
    
        # Update combo box options with loaded marker values
        self.combo_box.clear()
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
    
    def paste_image(self):
        """Handle pasting image from clipboard."""
        self.is_modified = True
        self.reset_image() # Clear previous state first
        self.load_image_from_clipboard()
        # UI updates (label size, sliders) should happen within load_image_from_clipboard
        self.update_live_view()
        self.save_state() # Save state after pasting
    
    def load_image_from_clipboard(self):
        """
        Load an image from the clipboard into self.image, preserving format.
        Prioritizes file paths over raw image data to handle macOS file copying correctly.
        If loaded from a file path, attempts to load an associated config file.
        """
        # Check for unsaved changes before proceeding
        # if not self.prompt_save_if_needed(): # Optional: uncomment if needed
        #     return # Abort paste if user cancels save

        self.reset_image() # Clear previous state first

        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
        loaded_image = None
        source_info = "Clipboard" # Default source
        config_loaded_from_paste = False # Flag to track if config was loaded

        # --- PRIORITY 1: Check for File URLs ---
        if mime_data.hasUrls():
            urls = mime_data.urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if os.path.isfile(file_path) and file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                    loaded_image = QImage(file_path)
                    source_info = file_path

                    if loaded_image.isNull():
                        try:
                            pil_image = Image.open(file_path)
                            loaded_image = ImageQt.ImageQt(pil_image)
                            if loaded_image.isNull():
                                loaded_image = None
                            else:
                                source_info = f"{file_path} (Pillow)"
                        except Exception as e:
                            QMessageBox.warning(self, "File Load Error", f"Could not load image from file '{os.path.basename(file_path)}':\n{e}")
                            loaded_image = None

                    # --- *** ADDED: CONFIG FILE LOADING FOR PASTED FILE *** ---
                    if loaded_image and not loaded_image.isNull():
                        self.image_path = file_path # Store the path early for config lookup
                        base_name_no_ext = os.path.splitext(os.path.basename(file_path))[0]
                        # Construct potential config file name (remove _original/_modified if present)
                        config_base = base_name_no_ext.replace("_original", "").replace("_modified", "")
                        config_path = os.path.join(os.path.dirname(file_path), f"{config_base}_config.txt")

                        if os.path.exists(config_path):
                            try:
                                with open(config_path, "r") as config_file:
                                    config_data = json.load(config_file)
                                self.apply_config(config_data) # Apply loaded settings
                                config_loaded_from_paste = True # Set flag
                            except Exception as e:
                                QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply associated config file '{os.path.basename(config_path)}': {e}")

                    # --- *** END OF ADDED CONFIG LOADING *** ---


        # --- PRIORITY 2: Check for Raw Image Data (if no valid file was loaded) ---
        if loaded_image is None and mime_data.hasImage():
            loaded_image = clipboard.image()
            if not loaded_image.isNull():
                source_info = "Clipboard Image Data"
            else:
                try:
                    pil_image = ImageGrab.grabclipboard()
                    if isinstance(pil_image, Image.Image):
                        loaded_image = ImageQt.ImageQt(pil_image)
                        if loaded_image.isNull():
                             loaded_image = None
                        else:
                            source_info = "Clipboard Image Data (Pillow Grab)"
                    else:
                        loaded_image = None
                except Exception: # Keep quiet on Pillow grab errors
                    loaded_image = None

        # --- Process the successfully loaded image ---
        if loaded_image and not loaded_image.isNull():
            self.is_modified = True
            self.image = loaded_image

            # Initialize backups
            self.original_image = self.image.copy()
            self.image_master = self.image.copy()
            self.image_before_padding = None
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_padded = False
            # self.image_path is set earlier if loaded from file

            # --- Update UI Elements ---
            try:
                w = self.image.width()
                h = self.image.height()
                if w <= 0 or h <= 0: raise ValueError("Invalid image dimensions after load.")

                # Update preview label size
                ratio=w/h if h > 0 else 1
                target_width = int(self.label_size) # Use current label_size setting
                target_height = int(target_width / ratio)
                # Optional: Apply max height constraint
                # if hasattr(self, 'preview_label_max_height_setting') and target_height > self.preview_label_max_height_setting:
                #     target_height = self.preview_label_max_height_setting
                #     target_width = int(target_height * ratio)
                self._update_preview_label_size()

                # Update slider ranges based on *new* label size
                render_scale = 3
                render_width = self.live_view_label.width() * render_scale
                render_height = self.live_view_label.height() * render_scale
                self.left_slider_range = [-100, int(render_width) + 100]
                self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
                self.right_slider_range = [-100, int(render_width) + 100]
                self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
                self.top_slider_range = [-100, int(render_height) + 100]
                self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])

                # Set recommended padding only if config wasn't loaded
                if not config_loaded_from_paste:
                    self.left_padding_input.setText(str(int(w * 0.1)))
                    self.right_padding_input.setText(str(int(w * 0.1)))
                    self.top_padding_input.setText(str(int(h * 0.15)))
                    self.bottom_padding_input.setText("0")

                # Update window title and status bar
                display_source = os.path.basename(source_info.split(" (")[0]) # Show filename or simplified source
                title_suffix = " (+ Config)" if config_loaded_from_paste else ""
                self.setWindowTitle(f"{self.window_title}::{display_source}{title_suffix}")
                self._update_status_bar()

            except Exception as e:
                QMessageBox.warning(self, "UI Update Error", f"Could not update UI elements after pasting: {e}")
                self._update_preview_label_size()
                
            enable_pan = self.live_view_label.zoom_level > 1.0
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
            self.update_live_view()
            self.save_state()

        else:
            # If no image was successfully loaded
            QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
            # Clean up state
            self.image = None
            self.original_image = None
            self.image_master = None
            self.image_path = None
            self.live_view_label.clear()
            self._update_preview_label_size()
            self.setWindowTitle(self.window_title)
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
            self._update_status_bar()
            self.update_live_view()
        
    def update_font(self):
        """Update the font settings based on UI inputs"""
        # Update font family from the combo box
        self.font_family = self.font_combo_box.currentFont().family()
        
        # Update font size from the spin box
        self.font_size = self.font_size_spinner.value()
        
        self.font_rotation = int(self.font_rotation_input.value())
        
    
        # Once font settings are updated, update the live view immediately
        self.update_live_view()
    
    def select_font_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.font_color = color  # Store the selected color   
            self.update_font()
        self._update_color_button_style(self.font_color_button, self.font_color)
            
    def load_image(self):
        self.is_modified = True # Mark as modified when loading new image
        self.undo_stack = []
        self.redo_stack = []
        self.reset_image() # Clear previous state

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
        )
        if file_path:
            self.image_path = file_path
            loaded_image = QImage(self.image_path)

            if loaded_image.isNull():
                # Try loading with Pillow as fallback
                try:
                    pil_image = Image.open(self.image_path)
                    # Use ImageQt for reliable conversion, preserves most formats
                    loaded_image = ImageQt.ImageQt(pil_image)
                    if loaded_image.isNull():
                        raise ValueError("Pillow could not convert to QImage.")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}': {e}")
                    self.image_path = None
                    return

            # --- Keep the loaded image format ---
            self.image = loaded_image

            # --- Initialize backups with the loaded format ---
            if not self.image.isNull():
                self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                self.image_master = self.image.copy()   # Master copy for resets
                self.image_before_padding = None        # Reset padding state
                self.image_contrasted = self.image.copy() # Backup for contrast
                self.image_before_contrast = self.image.copy() # Backup for contrast
                self.image_padded = False               # Reset flag

                self.setWindowTitle(f"{self.window_title}::{self.image_path}")

                # --- Load Associated Config File (Logic remains the same) ---
                self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                config_name = ""
                if self.base_name.endswith("_original"):
                    config_name = self.base_name.replace("_original", "_config.txt")
                else:
                    config_name = self.base_name + "_config.txt"

                config_path = os.path.join(os.path.dirname(file_path), config_name)

                if os.path.exists(config_path):
                    try:
                        with open(config_path, "r") as config_file:
                            config_data = json.load(config_file)
                        self.apply_config(config_data) # Apply loaded settings
                    except Exception as e:
                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                # --- End Config File Loading ---

            else:
                 QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                 return

            # --- Update UI Elements (Label size, sliders) ---
            if self.image and not self.image.isNull():
                try:
                    # (UI update logic remains the same as before)
                    self._update_preview_label_size()

                    render_scale = 3
                    render_width = self.live_view_label.width() * render_scale
                    render_height = self.live_view_label.height() * render_scale
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])

                    if not os.path.exists(config_path):
                        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                        self.bottom_padding_input.setText("0")

                except Exception as e:
                    pass
            # --- End UI Element Update ---
            
            
            self._update_status_bar() # <--- Add this
            enable_pan = self.live_view_label.zoom_level > 1.0
            if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(enable_pan)
            if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(enable_pan)
            self.update_live_view() # Render the loaded image
            self.save_state() # Save initial loaded state
    
    def apply_config(self, config_data):
        self.left_padding_input.setText(config_data["adding_white_space"]["left"])
        self.right_padding_input.setText(config_data["adding_white_space"]["right"])
        self.top_padding_input.setText(config_data["adding_white_space"]["top"])
        try:
            self.transparency=int(config_data["adding_white_space"]["transparency"])
        except:
            pass
        
            
        try:
            self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
        except:
            pass
    
        self.crop_x_start_slider.setValue(config_data["cropping_parameters"]["x_start_percent"])
        self.crop_x_end_slider.setValue(config_data["cropping_parameters"]["x_end_percent"])
        self.crop_y_start_slider.setValue(config_data["cropping_parameters"]["y_start_percent"])
        self.crop_y_end_slider.setValue(config_data["cropping_parameters"]["y_end_percent"])
    
        try:
            self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
            self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
            self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
        except (KeyError, ValueError) as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
            pass
        try:
            self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
            self.top_marker_input.setText(", ".join(self.top_label))
        except KeyError as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
            pass
    
        self.font_family = config_data["font_options"]["font_family"]
        self.font_size = config_data["font_options"]["font_size"]
        self.font_rotation = config_data["font_options"]["font_rotation"]
        self.font_color = QColor(config_data["font_options"]["font_color"])
    
        self.top_padding_slider.setValue(config_data["marker_padding"]["top"])
        self.left_padding_slider.setValue(config_data["marker_padding"]["left"])
        self.right_padding_slider.setValue(config_data["marker_padding"]["right"])
        
        try:
            try:
                x = list(map(int, config_data["marker_labels"]["left"]))
                self.marker_values_textbox.setText(str(x))
            except:
                x = list(map(int, config_data["marker_labels"]["right"]))
                self.marker_values_textbox.setText(str(x))
            if self.marker_values_textbox.text!="":
                self.combo_box.setCurrentText("Custom")
                self.marker_values_textbox.setEnabled(True)                
            else:
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.marker_values_textbox.setEnabled(False)
                
        except:
            pass
    
        try:
            loaded_custom_markers = []
            for marker_dict in config_data.get("custom_markers", []):
                try:
                    x = float(marker_dict["x"])
                    y = float(marker_dict["y"])
                    text = str(marker_dict["text"])
                    color = QColor(marker_dict["color"])
                    font = str(marker_dict["font"])
                    font_size = int(marker_dict["font_size"])
                    # Check for bold/italic, default to False if missing (backward compatible)
                    is_bold = bool(marker_dict.get("bold", False))
                    is_italic = bool(marker_dict.get("italic", False))

                    loaded_custom_markers.append((
                        x, y, text, color, font, font_size, is_bold, is_italic
                    ))
                except (KeyError, ValueError, TypeError) as e:
                    pass
                    
            self.custom_markers = loaded_custom_markers
                
                
                
        except:
            pass
        # Set slider ranges from config_data
        try:
            self.left_padding_slider.setRange(
                int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
            )
            self.right_padding_slider.setRange(
                int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
            )
            self.top_padding_slider.setRange(
                int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
            )
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Slider ranges not found in config_data.")
            pass
        
        try:
            self.left_marker_shift_added=int(config_data["added_shift"]["left"])
            self.right_marker_shift_added=int(config_data["added_shift"]["right"])
            self.top_marker_shift_added=int(config_data["added_shift"]["top"])
            
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Added Shift not found in config_data.");
            pass
            
        # Apply font settings

        
        #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
        
        font_size_new=self.font_size
        font_rotation_new=self.font_rotation
        
        self.font_combo_box.setCurrentFont(QFont(self.font_family))
        self.font_size_spinner.setValue(font_size_new)
        self.font_rotation_input.setValue(font_rotation_new)

        self.update_live_view()
        
    def get_current_config(self):
        """Gathers the current application state into a dictionary for saving."""
        config = {
            "adding_white_space": {
                "left": self.left_padding_input.text(),
                "right": self.right_padding_input.text(),
                "top": self.top_padding_input.text(),
                "bottom": self.bottom_padding_input.text(),
                "transparency": self.transparency, # Assuming self.transparency exists
            },
            "cropping_parameters": {
                "x_start_percent": self.crop_x_start_slider.value(),
                "x_end_percent": self.crop_x_end_slider.value(),
                "y_start_percent": self.crop_y_start_slider.value(),
                "y_end_percent": self.crop_y_end_slider.value(),
            },
            "marker_positions": {
                # Store positions only for standard markers
                "left": [(pos, label) for pos, label in getattr(self, 'left_markers', [])],
                "right": [(pos, label) for pos, label in getattr(self, 'right_markers', [])],
                "top": [(pos, label) for pos, label in getattr(self, 'top_markers', [])],
            },
            "marker_labels": { # Storing labels separately might be redundant if already in positions
                "top": getattr(self, 'top_label', []),
                "left": [marker[1] for marker in getattr(self, 'left_markers', [])],
                "right": [marker[1] for marker in getattr(self, 'right_markers', [])],
            },
            "marker_padding": { # Offsets for standard markers
                "top": self.top_padding_slider.value(),
                "left": self.left_padding_slider.value(),
                "right": self.right_padding_slider.value(),
            },
            "font_options": { # Default font options for standard markers
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_rotation": self.font_rotation,
                "font_color": self.font_color.name(),
            },
            "slider_ranges": { # Store current slider ranges
                 "left": getattr(self, 'left_slider_range', [-100, 1000]),
                 "right": getattr(self, 'right_slider_range', [-100, 1000]),
                 "top": getattr(self, 'top_slider_range', [-100, 1000]),
             },
            "added_shift": { # Store current added shifts
                 "left": getattr(self, 'left_marker_shift_added', 0),
                 "right": getattr(self, 'right_marker_shift_added', 0),
                 "top": getattr(self, 'top_marker_shift_added', 0),
            }
        }

        # --- Updated Custom Markers Section ---
        custom_markers_data = []
        # Use getattr for safety, default to empty list
        for marker_tuple in getattr(self, "custom_markers", []):
            try:
                # Unpack 8 elements
                x, y, text, color, font, font_size, is_bold, is_italic = marker_tuple
                custom_markers_data.append({
                    "x": x,
                    "y": y,
                    "text": text,
                    "color": color.name(), # Save color name/hex
                    "font": font,
                    "font_size": font_size,
                    "bold": is_bold,       # Save bold flag
                    "italic": is_italic    # Save italic flag
                })
            except (ValueError, TypeError, IndexError) as e:
                pass
        config["custom_markers"] = custom_markers_data
        # --- End Updated Custom Markers ---

        return config
    
    def add_band(self, event):
        # --- Ensure internal lists match UI input BEFORE adding band ---
        # This call ensures self.marker_values and self.top_label are
        # updated based on the current text in the UI boxes.
        self.update_all_labels()
        # -------------------------------------------------------------

        self.save_state()
        # Ensure there's an image loaded and marker mode is active
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        if not self.image or not self.marker_mode:
            return

        # --- Get Coordinates and Scaling (Same as previous version) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size

        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image.width() > 0 else 1
        image_height = self.image.height() if self.image.height() > 0 else 1
        uniform_scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1
        offset_x = (displayed_width - image_width * uniform_scale) / 2
        offset_y = (displayed_height - image_height * uniform_scale) / 2
        image_x = (cursor_x - offset_x) / uniform_scale if uniform_scale != 0 else 0
        image_y = (cursor_y - offset_y) / uniform_scale if uniform_scale != 0 else 0
        
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)

        # --- Get Render Info for Slider Positioning (Same as previous version) ---
        render_scale = 3
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale

        # --- Validate Coordinates (Same as previous version) ---
        current_image_width = self.image.width()
        current_image_height = self.image.height()
        if not (0 <= image_y <= current_image_height) and self.marker_mode in ["left", "right"]:
             return
        if not (0 <= image_x <= current_image_width) and self.marker_mode == "top":
             return
        # --- End Coordinate/Scaling/Validation ---

        try:
            # --- Left Marker Logic ---
            if self.marker_mode == "left":
                # Determine label based on current count and *now updated* self.marker_values
                current_marker_count = len(self.left_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        pass

                    
                self.left_markers.append((image_y, marker_value_to_add))
                self.current_left_marker_index += 1 # Still increment conceptual index

                # Set slider position only for the *very first* marker placed
                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.left_padding_slider.setValue(0)
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.left_padding_slider.setValue(padding_value)
                    self.left_marker_shift_added = self.left_padding_slider.value()    

            # --- Right Marker Logic ---
            elif self.marker_mode == "right":
                current_marker_count = len(self.right_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        pass

                self.right_markers.append((image_y, marker_value_to_add))
                self.current_right_marker_index += 1

                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.right_padding_slider.setValue(0)
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.right_padding_slider.setValue(padding_value)
                    self.right_marker_shift_added = self.right_padding_slider.value()

            # --- Top Marker Logic ---
            elif self.marker_mode == "top":
                current_marker_count = len(self.top_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.top_label list, which was just updated
                if current_marker_count < len(self.top_label):
                    label_to_add = self.top_label[current_marker_count]
                else:
                    label_to_add = ""
                    if current_marker_count == len(self.top_label): # Print warning only once
                        pass     

                self.top_markers.append((image_x, label_to_add))
                self.current_top_label_index += 1

                if is_first_marker:
                    padding_value=int((image_y - y_start) * (render_height / self.image.height()))
                    self.top_padding_slider.setValue(0)
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
                    self.top_padding_slider.setValue(padding_value)
                    self.top_marker_shift_added = self.top_padding_slider.value()

        except Exception as e:
             # Catch other potential errors
             import traceback
             traceback.print_exc()
             QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

        # Update the live view to render the newly added marker
        self.update_live_view()
        

        
    def enable_left_marker_mode(self):
        self.marker_mode = "left"
        self.current_left_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)

    def enable_right_marker_mode(self):
        self.marker_mode = "right"
        self.current_right_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_top_marker_mode(self):
        self.marker_mode = "top"
        self.current_top_label_index
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
        
    # def remove_padding(self):
    #     if self.image_before_padding!=None:
    #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
    #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
    #     self.image_padded = False  # Reset the padding state
    #     w=self.image.width()
    #     h=self.image.height()
    #     # Preview window
    #     ratio=w/h
    #     self.label_width = 540
    #     label_height=int(self.label_width/ratio)
    #     if label_height>self.label_width:
    #         label_height=540
    #         self.label_width=ratio*label_height
    #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    #     self.update_live_view()
        
    def finalize_image(self): # Padding
        self.save_state()
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
            return

        try:
            # Ensure padding values are non-negative
            padding_left = max(0, int(self.left_padding_input.text()))
            padding_right = max(0, int(self.right_padding_input.text()))
            padding_top = max(0, int(self.top_padding_input.text()))
            padding_bottom = max(0, int(self.bottom_padding_input.text()))
            # Update input fields in case negative values were entered
            self.left_padding_input.setText(str(padding_left))
            self.right_padding_input.setText(str(padding_right))
            self.top_padding_input.setText(str(padding_top))
            self.bottom_padding_input.setText(str(padding_bottom))

        except ValueError:
            QMessageBox.warning(self, "Error", "Please enter valid non-negative integers for padding.")
            return

        # --- Check if padding is actually needed ---
        if padding_left == 0 and padding_right == 0 and padding_top == 0 and padding_bottom == 0:
            QMessageBox.information(self, "Info", "No padding specified. Image remains unchanged.")
            return # Nothing to do

        try:
            # --- Determine if the source image has transparency ---
            
            source_has_alpha = self.image.hasAlphaChannel()

            # --- Convert source QImage to NumPy array ---
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")
            
            fill_value = None
            # Check if NumPy array indicates alpha (4 channels) even if QImage didn't report it explicitly
            numpy_indicates_alpha = (np_img.ndim == 3 and np_img.shape[2] == 4)
            has_alpha = source_has_alpha or numpy_indicates_alpha # Combine checks

            if has_alpha:
                # Pad images WITH existing alpha using TRANSPARENT black
                # OpenCV expects (B, G, R, Alpha) or just Alpha if grayscale+alpha
                # Provide 4 values for safety, assuming color/grayscale + alpha format from qimage_to_numpy
                fill_value = (0, 0, 0, 0) # Transparent Black for NumPy/OpenCV
            elif np_img.ndim == 3: # Opaque Color (BGR)
                # Pad opaque color images with WHITE
                fill_value = (255, 255, 255) # White for BGR format expected by OpenCV
            elif np_img.ndim == 2: # Opaque Grayscale
                # Pad opaque grayscale images with WHITE
                # Determine white value based on dtype
                if np_img.dtype == np.uint16:
                    fill_value = 65535 # White for 16-bit grayscale
                else: # Assume uint8 or other standard grayscale
                    fill_value = 255 # White for 8-bit grayscale
            else:
                 raise ValueError(f"Unsupported image dimensions for padding: {np_img.ndim}")

            # --- Adjust markers BEFORE creating the new padded image ---
            self.adjust_markers_for_padding(padding_left, padding_right, padding_top, padding_bottom)

            # --- Pad using OpenCV's copyMakeBorder ---
            padded_np = cv2.copyMakeBorder(np_img, padding_top, padding_bottom,
                                           padding_left, padding_right,
                                           cv2.BORDER_CONSTANT, value=fill_value)

            # --- Convert padded NumPy array back to QImage ---
            # numpy_to_qimage should automatically select a format supporting alpha if padded_np has 4 channels
            padded_image = self.numpy_to_qimage(padded_np)
            if padded_image.isNull():
                 raise ValueError("Conversion back to QImage failed after padding.")


            # --- Update main image and backups ---
            # Store the image *before* this padding operation, unless padding was already applied
            if not self.image_padded:
                self.image_before_padding = self.image.copy()
            # Else: keep the existing image_before_padding from the *previous* unpadded state

            self.image = padded_image
            self.image_padded = True # Indicate padding has been applied

            # Update backups consistently to reflect the new padded state
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy() # Contrast baseline resets after padding

            # --- Update UI (label size, slider ranges, live view) ---
            # (Keep the existing UI update logic here)
 
            self._update_preview_label_size()

            render_scale = 3
            render_width = self.live_view_label.width() * render_scale
            render_height = self.live_view_label.height() * render_scale
            self.left_slider_range = [-100, int(render_width) + 100]
            self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
            self.right_slider_range = [-100, int(render_width) + 100]
            self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
            self.top_slider_range = [-100, int(render_height) + 100]
            self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
            # --- End UI Update ---

            self.update_live_view() # Refresh display
            self._update_status_bar() # Update size/depth info

        except Exception as e:
            QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
            traceback.print_exc()
    # --- END: Modified Image Operations ---
    
    def adjust_markers_for_padding(self, padding_left, padding_right, padding_top, padding_bottom):
        """Adjust marker positions based on padding."""
        # Adjust left markers
        self.left_markers = [(y + padding_top, label) for y, label in self.left_markers]
        # Adjust right markers
        self.right_markers = [(y + padding_top, label) for y, label in self.right_markers]
        # Adjust top markers
        self.top_markers = [(x + padding_left, label) for x, label in self.top_markers]        
        
    
    def update_left_padding(self):
        # Update left padding when slider value changes
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.update_live_view()

    def update_right_padding(self):
        # Update right padding when slider value changes
        self.right_marker_shift_added = self.right_padding_slider.value()
        self.update_live_view()
        
    def update_top_padding(self):
        # Update top padding when slider value changes
        self.top_marker_shift_added = self.top_padding_slider.value()
        self.update_live_view()

    def update_live_view(self):
        if not self.image:
            return
    
        # Enable the "Predict Molecular Weight" button if markers are present
        if self.left_markers or self.right_markers:
            self.predict_button.setEnabled(True)
        else:
            self.predict_button.setEnabled(False)
    
        # Define a higher resolution for processing (e.g., 2x or 3x label size)
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Calculate scaling factors and offsets
        scale_x = self.image.width() / render_width
        scale_y = self.image.height() / render_height
    
        # Get the crop percentage values from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure the cropping boundaries are valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            self.crop_x_start_slider.setValue(0)
            self.crop_x_end_slider.setValue(100)
            self.crop_y_start_slider.setValue(0)
            self.crop_y_end_slider.setValue(100)
            return
    
        # Crop the image based on the defined boundaries
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
    
        # Get the orientation value from the slider
        orientation = float(self.orientation_slider.value() / 20)  # Orientation slider value
        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")
    
        # Apply the rotation to the cropped image
        rotated_image = cropped_image.transformed(QTransform().rotate(orientation))
    
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1
        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")
    
        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            return
    
        # Apply the transformation
        skewed_image = rotated_image.transformed(transform, Qt.SmoothTransformation)
    
        # Scale the rotated image to the rendering resolution
        scaled_image = skewed_image.scaled(
            render_width,
            render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        # Render on a high-resolution canvas
        canvas = QImage(render_width, render_height, QImage.Format_ARGB32)
        canvas.fill(Qt.transparent)  # Transparent background
    
        # Render the base image and overlays
        self.render_image_on_canvas(canvas, scaled_image, x_start, y_start, render_scale)
    
        # Scale the high-resolution canvas down to the label's size for display
        pixmap = QPixmap.fromImage(canvas).scaled(
            self.live_view_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        painter = QPainter(pixmap)
        pen = QPen(Qt.red)
        pen.setWidth(1)
        painter.setPen(pen)
    
        scale_x = self.live_view_label.width() / self.image.width()
        scale_y = self.live_view_label.height() / self.image.height()
    
        # Draw the quadrilateral or rectangle if in move mode
        if self.live_view_label.mode == "move":
            pen = QPen(Qt.red, 2)
            painter.setPen(pen)
            if self.live_view_label.quad_points:
                # Draw the quadrilateral
                painter.drawPolygon(QPolygonF(self.live_view_label.quad_points))
            elif self.live_view_label.bounding_box_preview:
                # Draw the rectangle
                start_x, start_y, end_x, end_y = self.live_view_label.bounding_box_preview
                rect = QRectF(start_x, start_y, end_x - start_x, end_y - start_y)
                painter.drawRect(rect)
    
        painter.end()
    
        # Apply zoom and pan transformations to the final pixmap
        if self.live_view_label.zoom_level != 1.0:
            # Create a new pixmap to apply zoom and pan
            zoomed_pixmap = QPixmap(pixmap.size())
            zoomed_pixmap.fill(Qt.transparent)
            zoom_painter = QPainter(zoomed_pixmap)
            zoom_painter.translate(self.live_view_label.pan_offset)
            zoom_painter.scale(self.live_view_label.zoom_level, self.live_view_label.zoom_level)
            zoom_painter.drawPixmap(0, 0, pixmap)
            zoom_painter.end()  # Properly end the QPainter
            pixmap = zoomed_pixmap
    
        # Set the final pixmap to the live view label
        self.live_view_label.setPixmap(pixmap)
    
    def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
        painter = QPainter(canvas)
        x_offset = (canvas.width() - scaled_image.width()) // 2
        y_offset = (canvas.height() - scaled_image.height()) // 2
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
    
        # Draw the base image
        painter.drawImage(x_offset, y_offset, scaled_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            # Resize Image 1 based on the slider value
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 1
            image1_x = x_offset + self.image1_position[0]
            image1_y = y_offset + self.image1_position[1]
            painter.drawImage(image1_x, image1_y, resized_image1)
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            # Resize Image 2 based on the slider value
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 2
            image2_x = x_offset + self.image2_position[0]
            image2_y = y_offset + self.image2_position[1]
            painter.drawImage(image2_x, image2_y, resized_image2)
    
        # Get the selected font settings
        font = QFont(self.font_combo_box.currentFont().family(), self.font_size_spinner.value() * render_scale)
        font_color = self.font_color if hasattr(self, 'font_color') else QColor(0, 0, 0)  # Default to black if no color selected
    
        painter.setFont(font)
        painter.setPen(font_color)  # Set the font color
    
        # Measure text height for vertical alignment
        font_metrics = painter.fontMetrics()
        text_height = font_metrics.height()
        line_padding = 5 * render_scale  # Space between text and line
        
        y_offset_global = text_height / 4
    
        # Draw the left markers (aligned right)
        for y_pos, marker_value in self.left_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f"{marker_value} ⎯ "  ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
        
        
        # Draw the right markers (aligned left)
        for y_pos, marker_value in self.right_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f" ⎯ {marker_value}" ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.right_marker_shift_added),# + line_padding),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
                
    
        # Draw the top markers (if needed)
        for x_pos, top_label in self.top_markers:
            x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width())
            if 0 <= x_pos_cropped <= scaled_image.width():
                text = f"{top_label}"
                painter.save()
                text_width = font_metrics.horizontalAdvance(text)
                text_height= font_metrics.height()
                label_x = x_offset + x_pos_cropped 
                label_y = y_offset + self.top_marker_shift + self.top_marker_shift_added 
                painter.translate(int(label_x), int(label_y))
                painter.rotate(self.font_rotation)
                painter.drawText(0,int(y_offset_global), f"{top_label}")
                painter.restore()
    
        # Draw guide lines
        if draw_guides and self.show_guides_checkbox.isChecked():
            pen = QPen(Qt.red, 2 * render_scale)
            painter.setPen(pen)
            center_x = canvas.width() // 2
            center_y = canvas.height() // 2
            painter.drawLine(center_x, 0, center_x, canvas.height())  # Vertical line
            painter.drawLine(0, center_y, canvas.width(), center_y)  # Horizontal line
            
        
        
        
        # Draw the protein location marker (*)
        if hasattr(self, "protein_location") and self.run_predict_MW == False:
            x, y = self.protein_location
            text = "⎯⎯"
            text_width = font_metrics.horizontalAdvance(text)
            text_height= font_metrics.height()
            painter.drawText(
                int(x * render_scale - text_width / 2),  #Currently left edge # FOR Center horizontally use int(x * render_scale - text_width / 2)
                int(y * render_scale + text_height / 4),  # Center vertically
                text
            )
        for marker_tuple in getattr(self, "custom_markers", []):
            try:
                # -->> Unpack all 8 elements using the correct order <<--
                # The tuple structure is (x, y, text, color, font_family, font_size, is_bold, is_italic)
                x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple

                # Calculate position on canvas, scaling relative to the CROPPED image extent
                # Ensure original image dimensions are valid for scaling calculation
                orig_img_w = self.image.width() if self.image and self.image.width() > 0 else 1
                orig_img_h = self.image.height() if self.image and self.image.height() > 0 else 1
                # Dimensions of the image drawn on the canvas (after potential cropping/scaling)
                img_w_on_canvas = scaled_image.width()
                img_h_on_canvas = scaled_image.height()

                # Calculate the marker's position relative to the canvas origin (x_offset, y_offset)
                # Scale based on how the original coordinates map to the scaled_image dimensions
                x_pos_on_canvas = x_offset + (x_pos - x_start) * (img_w_on_canvas / orig_img_w)
                y_pos_on_canvas = y_offset + (y_pos - y_start) * (img_h_on_canvas / orig_img_h)

                # Only draw if within the visible scaled image bounds on the canvas
                # Check against the area occupied by scaled_image on the canvas
                if (x_offset <= x_pos_on_canvas <= x_offset + img_w_on_canvas and
                    y_offset <= y_pos_on_canvas <= y_offset + img_h_on_canvas):

                    # -->> Create and configure font using unpacked values <<--
                    # Ensure font_size is a valid integer before scaling
                    try:
                        # Use int() directly on font_size which should already be an int/float
                        scaled_font_size = int(int(font_size) * render_scale)
                        if scaled_font_size < 1: scaled_font_size = 1 # Ensure minimum size 1
                    except (ValueError, TypeError):
                        scaled_font_size = int(12 * render_scale) # Default scaled size (e.g., 12pt scaled)

                    # Ensure font_family is a string
                    marker_font = QFont(str(font_family), scaled_font_size)
                    # Ensure bold/italic are booleans and apply them
                    marker_font.setBold(bool(is_bold))
                    marker_font.setItalic(bool(is_italic))

                    # Set the configured font on the painter
                    painter.setFont(marker_font)

                    # -->> Set the color, ensuring it's a QColor <<--
                    if isinstance(color, str):
                        current_color = QColor(color) # Create QColor from name/hex string
                    elif isinstance(color, QColor):
                        current_color = color # It's already a QColor
                    else:
                         current_color = Qt.black # Fallback color
                    painter.setPen(current_color)

                    # Calculate text dimensions for alignment using the specific marker's font
                    font_metrics_custom = painter.fontMetrics()
                    # Ensure marker_text is a string
                    text_to_draw = str(marker_text)
                    text_width = font_metrics_custom.horizontalAdvance(text_to_draw)
                    text_height = font_metrics_custom.height()
                    # Approximate vertical offset to center text around the point's Y coordinate
                    y_offset_global_custom = text_height / 4

                    # Draw text centered horizontally at its calculated canvas position
                    painter.drawText(
                        int(x_pos_on_canvas - text_width / 2),          # Center horizontally
                        int(y_pos_on_canvas + y_offset_global_custom),  # Center vertically around baseline
                        text_to_draw                                    # Draw the text
                    )
            except (ValueError, TypeError, IndexError) as e:
                # Catch errors during unpacking or processing of a single marker tuple
                import traceback
                traceback.print_exc() # Print stack trace for debugging tuple issues
            
            
    

        # Draw the grid (if enabled)
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value() * render_scale
            pen = QPen(Qt.red)
            pen.setStyle(Qt.DashLine)
            painter.setPen(pen)
    
            # Draw vertical grid lines
            for x in range(0, canvas.width(), grid_size):
                painter.drawLine(x, 0, x, canvas.height())
    
            # Draw horizontal grid lines
            for y in range(0, canvas.height(), grid_size):
                painter.drawLine(0, y, canvas.width(), y)
                
        painter.end()


     
    def crop_image(self):
        """Function to crop the current image."""
        if not self.image:
            return None
    
        # Get crop percentage from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate crop boundaries
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure cropping is valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            return None
    
        # Crop the image
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
        return cropped_image
    
    # Modify align_image and update_crop to preserve settings
    def reset_align_image(self):
        self.orientation_slider.setValue(0)
        try:
            if self.image: # Check if image exists after cropping
                self._update_preview_label_size()
        except Exception as e:
            # Fallback size?
            self._update_preview_label_size()


        self.update_live_view() # Final update with corrected markers and image
        
        
    def align_image(self):
        self.save_state()
        """Align the image based on the orientation slider and keep high-resolution updates."""
        if not self.image:
            return
    
        self.draw_guides = False
        self.show_guides_checkbox.setChecked(False)
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
    
        # Get the orientation value from the slider
        angle = float(self.orientation_slider.value()/20)
    
        # Perform rotation
        transform = QTransform()
        transform.translate(self.image.width() / 2, self.image.height() / 2)  # Center rotation
        transform.rotate(angle)
        transform.translate(-self.image.width() / 2, -self.image.height() / 2)
    
        rotated_image = self.image.transformed(transform, Qt.SmoothTransformation)
    
        # Create a white canvas large enough to fit the rotated image
        canvas_width = rotated_image.width()
        canvas_height = rotated_image.height()
        high_res_canvas = QImage(canvas_width, canvas_height, QImage.Format_ARGB32_Premultiplied)
        high_res_canvas.fill(Qt.transparent)  # Fill with white background
    
        # Render the high-resolution canvas using `render_image_on_canvas`, without guides
        self.render_image_on_canvas(
            high_res_canvas,
            scaled_image=rotated_image,
            x_start=0,
            y_start=0,
            render_scale=1,  # Adjust scale if needed
            draw_guides=False  # Do not draw guides in this case
        )
    
        # Update the main high-resolution image
        self.image = high_res_canvas
        self.image_before_padding = self.image.copy()
        self.image_contrasted=self.image.copy()
        self.image_before_contrast=self.image.copy()
    
        # Create a low-resolution preview for display in `live_view_label`
        preview = high_res_canvas.scaled(
            self.live_view_label.width(),
            self.live_view_label.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.live_view_label.setPixmap(QPixmap.fromImage(preview))
    
        # Reset the orientation slider
        self.orientation_slider.setValue(0)
        try:
            if self.image: # Check if image exists after cropping
                self._update_preview_label_size()
        except Exception as e:
            # Fallback size?
            self._update_preview_label_size()


        self.update_live_view() # Final update with corrected markers and image
    
    
    def update_crop(self):
        self.save_state()
        """Update the image based on current crop sliders. First align, then crop the image."""
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view() # Update once before operations

        # Get crop parameters *before* aligning/cropping
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100

        # Calculate the crop boundaries based on the *current* image dimensions before cropping
        if not self.image:
            return # Should not happen if called from UI, but safety check
        current_width = self.image.width()
        current_height = self.image.height()
        x_start = int(current_width * x_start_percent)
        x_end = int(current_width * x_end_percent)
        y_start = int(current_height * y_start_percent)
        y_end = int(current_height * y_end_percent)

        # Ensure cropping boundaries are valid relative to current image
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values based on current image size.")
            # Optionally reset sliders here if needed
            return


        # Align the image first (rotate it) - align_image modifies self.image
        # We align *before* cropping based on the original logic flow provided.
        # Note: If alignment changes dimensions significantly, this might need rethinking,
        # but typically rotation keeps content centered.
        # self.align_image() # Call align_image *if* it should happen before crop


        # Now apply cropping to the *current* self.image
        cropped_image = self.crop_image() # crop_image uses current self.image state

        if cropped_image:
            # --- Adjust marker positions relative to the crop ---
            new_left_markers = []
            for y, label in self.left_markers:
                if y_start <= y < y_end: # Check if marker was within vertical crop bounds
                    new_y = y - y_start
                    new_left_markers.append((new_y, label))
            self.left_markers = new_left_markers

            new_right_markers = []
            for y, label in self.right_markers:
                if y_start <= y < y_end:
                    new_y = y - y_start
                    new_right_markers.append((new_y, label))
            self.right_markers = new_right_markers

            new_top_markers = []
            for x, label in self.top_markers:
                if x_start <= x < x_end: # Check if marker was within horizontal crop bounds
                    new_x = x - x_start
                    new_top_markers.append((new_x, label))
            self.top_markers = new_top_markers

            new_custom_markers = []
            if hasattr(self, "custom_markers"):
                for x, y, text, color, font, font_size,*optional in self.custom_markers:
                    if x_start <= x < x_end and y_start <= y < y_end:
                        new_x = x - x_start
                        new_y = y - y_start
                        new_custom_markers.append((new_x, new_y, text, color, font, font_size,*optional))
            self.custom_markers = new_custom_markers
            # -----------------------------------------------------

            # Update the main image and related states
            self.image = cropped_image
            # Ensure these backups reflect the *newly cropped* state
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

        # Reset sliders after applying the crop
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)

        # Update live view label size based on the *new* image dimensions
        try:
            self._update_preview_label_size()
        except Exception as e:
            # Fallback size?
            self._update_preview_label_size()


        self.update_live_view() # Final update with corrected markers and image
        
    def update_skew(self):
        self.save_state()
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            return
    
        # Apply the transformation
        # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

 
        self.image = self.image.transformed(transform, Qt.SmoothTransformation)
        self.taper_skew_slider.setValue(0)

    
        
    def save_image(self):
        self.draw_guides = False
        self.show_guides_checkbox.setChecked(False)
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
        """Saves the original image, the modified image (rendered view), and configuration."""
        if not self.image_master: # Check if an initial image was ever loaded/pasted
             QMessageBox.warning(self, "Error", "No image data to save.")
             return False # Indicate save failed or was aborted

        self.is_modified = False # Mark as saved

        options = QFileDialog.Options()
        # Suggest a filename based on the loaded image or a default
        suggested_name = ""
        if self.image_path:
            base = os.path.splitext(os.path.basename(self.image_path))[0]
            # Remove common suffixes if they exist from previous saves
            base = base.replace("_original", "").replace("_modified", "")
            suggested_name = f"{base}" # Suggest PNG for modified view
        elif hasattr(self, 'base_name') and self.base_name:
            suggested_name = f"{base}"
        else:
            suggested_name = "untitled_image"

        save_dir = os.path.dirname(self.image_path) if self.image_path else "" # Suggest save in original dir

        # --- Get the base save path from the user ---
        base_save_path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save Image Base Name", os.path.join(save_dir, suggested_name),
            "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
            options=options
        )
        # --- End getting base save path ---

        if not base_save_path: # User cancelled
            self.is_modified = True # Revert saved status if cancelled
            return False # Indicate save cancelled

        # Determine paths based on the base path
        base_name_nosuffix = (os.path.splitext(base_save_path)[0]).replace("_original", "").replace("_modified", "")
        # Determine the desired suffix based on the selected filter or default to .png
        if "png" in selected_filter.lower():
             suffix = ".png"
        elif "tif" in selected_filter.lower():
             suffix = ".tif"
        elif "jpg" in selected_filter.lower() or "jpeg" in selected_filter.lower():
             suffix = ".jpg"
        elif "bmp" in selected_filter.lower():
             suffix = ".bmp"
        else: # Default or All Files
             suffix = os.path.splitext(base_save_path)[1] # Keep user's suffix if provided
             if not suffix: suffix = ".png" # Default to png if no suffix given

        original_save_path = f"{base_name_nosuffix}_original{suffix}"
        modified_save_path = f"{base_name_nosuffix}_modified{suffix}"
        config_save_path = f"{base_name_nosuffix}_config.txt"

        # --- Save original image (using self.image_master) ---
        img_to_save_orig = self.image 
        if img_to_save_orig and not img_to_save_orig.isNull():
            save_format_orig = suffix.replace(".", "").upper() # Determine format from suffix
            if save_format_orig == "TIF": save_format_orig = "TIFF" # Use standard TIFF identifier
            elif save_format_orig == "JPEG": save_format_orig = "JPG"

            # QImage.save attempts to match format, quality is for lossy formats
            quality = 95 if save_format_orig in ["JPG", "JPEG"] else -1 # Default quality (-1) for lossless

            if not img_to_save_orig.save(original_save_path, format=save_format_orig if save_format_orig else None, quality=quality):
                QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "Original master image data is missing.")

        # --- Save modified image (Rendered view - likely RGB) ---
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        # Use ARGB32 for rendering to support potential transparency (good for PNG)
        # Use RGB888 if saving as JPG/BMP which don't support alpha well.
        save_format_mod = suffix.replace(".", "").upper()
        if save_format_mod in ["JPG", "JPEG", "BMP"]:
            canvas_format = QImage.Format_ARGB32_Premultiplied
            fill_color = Qt.transparent # Use white background for opaque formats
        else: # PNG, TIFF
            canvas_format = QImage.Format_ARGB32_Premultiplied # Good for rendering quality with alpha
            fill_color = Qt.transparent # Use transparent background

        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, canvas_format
        )
        high_res_canvas.fill(fill_color)

        # Use current self.image for rendering the modified view
        if self.image and not self.image.isNull():
             scaled_image_mod = self.image.scaled(
                 high_res_canvas_width, high_res_canvas_height,
                 Qt.KeepAspectRatio, Qt.SmoothTransformation)

             # Render onto the high-res canvas (render_image_on_canvas draws markers etc.)
             # Pass draw_guides=False to avoid saving them
             self.render_image_on_canvas(
                 high_res_canvas, scaled_image_mod,
                 x_start=0, y_start=0, # Rendering is relative to current self.image extent
                 render_scale=render_scale, draw_guides=False
             )

             # Save the rendered canvas
             quality_mod = 95 if save_format_mod in ["JPG", "JPEG"] else -1
             if not high_res_canvas.save(modified_save_path, format=save_format_mod if save_format_mod else None, quality=quality_mod):
                 QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "No current image to render for saving modified view.")


        # --- Save configuration ---
        config_data = self.get_current_config()
        try:
            with open(config_save_path, "w") as config_file:
                json.dump(config_data, config_file, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
            return False # Indicate save failure


        QMessageBox.information(self, "Saved", f"Files saved successfully:\n- {os.path.basename(original_save_path)}\n- {os.path.basename(modified_save_path)}\n- {os.path.basename(config_save_path)}")

        # Update window title (optional, remove base_name if confusing)
        self.setWindowTitle(f"{self.window_title}::{base_name_nosuffix}")
        return True # Indicate successful save
            
    def save_image_svg(self):
        """Save the processed image along with markers and labels in SVG format containing EMF data."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image as SVG for MS Word Image Editing", "", "SVG Files (*.svg)", options=options
        )
    
        if not file_path:
            return
    
        if not file_path.endswith(".svg"):
            file_path += ".svg"
            
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
        
        x_offset = 0
        y_offset = 0
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
        line_padding = 5 * render_scale
        
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(file_path, profile='tiny', size=(high_res_canvas_width, high_res_canvas_height))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
        
    
    
        # Add custom markers to the SVG
        for marker_tuple in getattr(self, "custom_markers", []):
                try:
                    # Default values for optional elements
                    is_bold = False
                    is_italic = False

                    # Unpack based on length for backward compatibility
                    if len(marker_tuple) == 8:
                        x_pos, y_pos, marker_text, color, font_family, font_size, is_bold, is_italic = marker_tuple
                    elif len(marker_tuple) == 6:
                        x_pos, y_pos, marker_text, color, font_family, font_size = marker_tuple
                        # is_bold and is_italic remain False (default)
                    else:
                        continue # Skip this marker

                    # --- Prepare SVG attributes ---
                    # Text content
                    text_content = str(marker_text)

                    # Color: Convert QColor to hex string if needed
                    if isinstance(color, QColor):
                        fill_color = color.name() # Gets hex #RRGGBB or #AARRGGBB
                    elif isinstance(color, str):
                         # Basic validation if it's already a hex string
                         if color.startswith('#') and len(color) in [7, 9]:
                             fill_color = color
                         else: # Try converting string name to QColor, then get hex
                             temp_color = QColor(color)
                             fill_color = temp_color.name() if temp_color.isValid() else "#000000" # Fallback
                    else:
                        fill_color = "#000000" # Fallback to black

                    # Font attributes
                    font_family_svg = str(font_family)
                    try:
                        font_size_svg = f"{int(font_size)}px" # Ensure integer and add px
                    except (ValueError, TypeError):
                         font_size_svg = "12px" # Default size

                    # SVG font weight and style based on boolean flags
                    font_weight_svg = "bold" if bool(is_bold) else "normal"
                    font_style_svg = "italic" if bool(is_italic) else "normal"
                    x_pos_cropped = (x_pos - x_start) * (high_res_canvas_width / self.image.width())
                    y_pos_cropped = (y_pos - y_start) * (high_res_canvas_height / self.image.height())

                    # --- Add SVG text element ---
                    # Use text-anchor and dominant-baseline for centering at (x_pos, y_pos)
                    font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
                    text_width = int(font_metrics.horizontalAdvance(marker_text))  # Get text width
                    text_height = font_metrics.height()
                    dwg.add(
                        dwg.text(
                            text_content,
                            insert=(x_pos_cropped, y_pos_cropped),       # Position text anchor at the coordinate
                            fill=fill_color,
                            font_family=font_family_svg,
                            font_size=font_size_svg,
                            font_weight=font_weight_svg, # Apply bold/normal
                            font_style=font_style_svg   # Apply italic/normal
                        )
                    )
                except Exception as e:
                     # Catch errors during processing a single marker tuple
                     import traceback
                     traceback.print_exc() # Print full traceback for debugging
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            y_pos_cropped = (y - y_start) * (high_res_canvas_height / self.image.height())
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width), int(y_offset + y_pos_cropped)),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="end"  # Aligns text to the right
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            y_pos_cropped = (y - y_start) * (high_res_canvas_height / self.image.height())
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()

    
            dwg.add(
                dwg.text(
                    f" ⎯ {text}",
                    insert=(int(x_offset + self.right_marker_shift_added), int(y_offset + y_pos_cropped )),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="start"  # Aligns text to the left
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            x_pos_cropped = (x - x_start) * (high_res_canvas_width / self.image.width())
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text}"
            text_width = int(font_metrics.horizontalAdvance(final_text)) # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x_offset + x_pos_cropped, y_offset + self.top_marker_shift + self.top_marker_shift_added),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x_offset + x_pos_cropped}, {y_offset + self.top_marker_shift + self.top_marker_shift_added})"
                )
            )
    
        # Save the SVG file
        dwg.save()
    
        QMessageBox.information(self, "Success", f"Image saved as SVG at {file_path}.")
    
    
    def copy_to_clipboard(self):
        """Copy the image via a temporary file, cleaning up on exit."""
        # --- Initial Check ---
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Warning", "No image to copy.")
            return
    
        # --- Create the high-resolution rendered canvas ---
        # (Canvas creation, rendering, straight alpha conversion - KEEP AS IS)
        render_scale = 3
        try:
            label_width = self.live_view_label.width(); label_height = self.live_view_label.height()
            if label_width <= 0: label_width = 500 # Fallback
            if label_height <= 0: label_height = 500 # Fallback
            high_res_canvas_width = label_width * render_scale
            high_res_canvas_height = label_height * render_scale
            if high_res_canvas_width <= 0 or high_res_canvas_height <= 0: raise ValueError("Invalid canvas size")
        except Exception as e:
            QMessageBox.critical(self, "Internal Error", f"Could not calculate render dimensions: {e}")
            return
    
        render_canvas = QImage(high_res_canvas_width, high_res_canvas_height, QImage.Format_ARGB32_Premultiplied)
        render_canvas.fill(Qt.transparent)
    
        if self.image and not self.image.isNull():
            x_start, y_start = 0, 0 # Assuming self.image is pre-cropped
            # *** Robust Check ***
            if not self.image or self.image.isNull():
                 QMessageBox.critical(self, "Internal Error", "Image became invalid before scaling.")
                 self.cleanup_temp_clipboard_file() # Attempt cleanup
                 return
            try:
                scaled_image = self.image.scaled(high_res_canvas_width, high_res_canvas_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                if scaled_image.isNull(): raise ValueError("Scaling failed.")
            except Exception as e:
                QMessageBox.critical(self, "Copy Error", f"Failed to scale image: {e}")
                self.cleanup_temp_clipboard_file()
                return
    
            self.render_image_on_canvas(render_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False)
    
            final_canvas_for_clipboard = render_canvas
            if render_canvas.hasAlphaChannel():
                straight_alpha_canvas = render_canvas.convertToFormat(QImage.Format_ARGB32)
                if not straight_alpha_canvas.isNull():
                    final_canvas_for_clipboard = straight_alpha_canvas
            # --- End Rendering ---
    
            # --- Save to Temporary PNG File ---
            try:
                self.cleanup_temp_clipboard_file() # Clean up previous one first
    
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False, mode='wb') as temp_file:
                    self.temp_clipboard_file_path = temp_file.name
                    png_save_buffer = QBuffer()
                    png_save_buffer.open(QBuffer.WriteOnly)
                    if not final_canvas_for_clipboard.save(png_save_buffer, "PNG"):
                         raise IOError("Failed to save canvas to PNG buffer.")
                    temp_file.write(png_save_buffer.data())
                    png_save_buffer.close()
    
                if not os.path.exists(self.temp_clipboard_file_path):
                    raise IOError("Temporary file not created.")
    
            except Exception as e:
                QMessageBox.critical(self, "Copy Error", f"Failed to create/save temporary file: {e}")
                self.temp_clipboard_file_path = None
                return
            # --- End Save to Temp File ---
    
    
            # --- Prepare QMimeData ---
            clipboard = QApplication.clipboard()
            mime_data = QMimeData()
    
            # 1. Set File URL
            try:
                file_url = QUrl.fromLocalFile(self.temp_clipboard_file_path)
                if not file_url.isValid() or not file_url.isLocalFile(): raise ValueError("Invalid URL")
                mime_data.setUrls([file_url])
            except Exception as e:
                QMessageBox.warning(self, "Copy Warning", f"Could not set file URL: {e}")
    
            # 2. Set PNG Data fallback
            png_buffer_clip = QBuffer()
            png_buffer_clip.open(QBuffer.WriteOnly)
            if final_canvas_for_clipboard.save(png_buffer_clip, "PNG"):
                mime_data.setData("image/png", png_buffer_clip.data())
            png_buffer_clip.close()
    
            # 3. Set Standard Image Data fallback
            mime_data.setImageData(final_canvas_for_clipboard)
    
            # --- Set clipboard ---
            clipboard.setMimeData(mime_data)
    
            # --- REMOVED QTimer.singleShot ---
    
        else:
            QMessageBox.warning(self, "Warning", "No valid image data to render for copying.")
        
    def cleanup_temp_clipboard_file(self):
        """Deletes the temporary clipboard file if it exists."""
        path_to_delete = getattr(self, 'temp_clipboard_file_path', None)
        if path_to_delete:
            # print(f"Attempting to clean up temp file on exit: {path_to_delete}") # Debug
            # No need to clear self.temp_clipboard_file_path immediately here,
            # as the object is likely being destroyed soon anyway.
            if os.path.exists(path_to_delete):
                try:
                    os.remove(path_to_delete)
                    # print(f"Cleaned up temp file: {path_to_delete}") # Debug
                # Don't retry with QTimer here, the app is quitting.
                except OSError as e:
                    # Log it, but can't do much else at this point.
                    print(f"Warning: Could not delete temp clipboard file {path_to_delete} on exit: {e}")
        # Clear the attribute *after* attempting deletion (optional on exit)
        self.temp_clipboard_file_path = None
        
    def copy_to_clipboard_SVG(self):
        """Create a temporary SVG file with EMF data, copy it to clipboard."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
    
        # Get scaling factors to match the live view window
        view_width = self.live_view_label.width()
        view_height = self.live_view_label.height()
        scale_x = self.image.width() / view_width
        scale_y = self.image.height() / view_height
    
        # Create a temporary file to store the SVG
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_file:
            temp_file_path = temp_file.name
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(temp_file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x * scale_x, y * scale_y),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(self.left_marker_shift_added / scale_x, y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=((self.image.width() / scale_x + self.right_marker_shift_added / scale_x), y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x, self.top_marker_shift_added / scale_y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x}, {self.top_marker_shift_added / scale_y})"
                )
            )
    
        # Save the SVG to the temporary file
        dwg.save()
    
        # Read the SVG content
        with open(temp_file_path, "r", encoding="utf-8") as temp_file:
            svg_content = temp_file.read()
    
        # Copy SVG content to clipboard (macOS-compatible approach)
        clipboard = QApplication.clipboard()
        clipboard.setText(svg_content, mode=clipboard.Clipboard)
    
        QMessageBox.information(self, "Success", "SVG content copied to clipboard.")

        
    def clear_predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        if hasattr(self, "protein_location"):
            del self.protein_location  # Clear the protein location marker
        self.predict_size=False
        self.bounding_boxes=[]
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.live_view_label.bounding_box_preview = None
        self.quantities=[]
        self.peak_area_list=[]
        self.protein_quantities=[]
        self.standard_protein_values.setText("")
        self.standard_protein_areas=[]
        self.standard_protein_areas_text.setText("")
        self.live_view_label.quad_points=[]
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.rectangle_points = []
        self.latest_calculated_quantities = []
        self.quantities_peak_area_dict={}
        
        self.update_live_view()  # Update the display
        
    def predict_molecular_weight(self):
        """
        Initiates the molecular weight prediction process.
        Determines the available markers, sorts them by position,
        and sets up the mouse click event to get the target protein location.
        """
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.run_predict_MW = False # Reset prediction flag

        # Determine which markers to use (left or right)
        markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
        if not markers_raw_tuples:
            QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
            self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            return

        # --- Crucial Step: Sort markers by Y-position (migration distance) ---
        # This ensures the order reflects the actual separation on the gel/blot
        try:
            # Ensure marker values are numeric for sorting and processing
            numeric_markers = []
            for pos, val_str in markers_raw_tuples:
                try:
                    numeric_markers.append((float(pos), float(val_str)))
                except (ValueError, TypeError):
                    # Skip markers with non-numeric values for prediction
                    continue

            if len(numeric_markers) < 2:
                 QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 return

            sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
            # Separate sorted positions and values
            sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
            sorted_marker_values = np.array([val for pos, val in sorted_markers])

        except Exception as e:
            QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Ensure there are still at least two markers after potential filtering
        if len(sorted_marker_positions) < 2:
            QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Prompt user and set up the click handler
        QMessageBox.information(self, "Instruction",
                                "Click on the target protein location in the preview window.\n"
                                "The closest standard set (Gel or WB) will be used for calculation.")
        # Pass the *sorted* full marker data to the click handler
        self.live_view_label.mousePressEvent = lambda event: self.get_protein_location(
            event, sorted_marker_positions, sorted_marker_values
        )

    def get_protein_location(self, event, all_marker_positions, all_marker_values):
        """
        Handles the mouse click event for protein selection.
        Determines the relevant standard set based on click proximity,
        performs regression on that set, predicts MW, and plots the results.
        """
        # --- 1. Get Protein Click Position (Image Coordinates) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()

        # Account for zoom and pan
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        # Transform cursor position from view coordinates to image coordinates
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image and self.image.width() > 0 else 1
        image_height = self.image.height() if self.image and self.image.height() > 0 else 1

        scale = min(displayed_width / image_width, displayed_height / image_height)
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2

        protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
        # protein_x_image = (cursor_x - x_offset) / scale if scale != 0 else 0 # Keep X if needed later

        # Store the clicked location in *view* coordinates for drawing the marker later
        self.protein_location = (cursor_x, cursor_y) # Use view coordinates for the marker

        # --- 2. Identify Potential Standard Sets (Partitioning) ---
        transition_index = -1
        # Find the first index where the molecular weight INCREASES after initially decreasing
        # (indicates a likely switch from Gel high->low MW to WB high->low MW)
        initial_decrease = False
        for k in range(1, len(all_marker_values)):
             if all_marker_values[k] < all_marker_values[k-1]:
                 initial_decrease = True # Confirm we've started migrating down
             # Check for increase *after* we've seen a decrease
             if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                 transition_index = k
                 break # Found the likely transition

        # --- 3. Select the Active Standard Set based on Click Proximity ---
        active_marker_positions = None
        active_marker_values = None
        set_name = "Full Set" # Default name

        if transition_index != -1:
            # We have two potential sets
            set1_positions = all_marker_positions[:transition_index]
            set1_values = all_marker_values[:transition_index]
            set2_positions = all_marker_positions[transition_index:]
            set2_values = all_marker_values[transition_index:]

            # Check if both sets are valid (at least 2 points)
            valid_set1 = len(set1_positions) >= 2
            valid_set2 = len(set2_positions) >= 2

            if valid_set1 and valid_set2:
                # Calculate the mean Y position for each set
                mean_y_set1 = np.mean(set1_positions)
                mean_y_set2 = np.mean(set2_positions)

                # Assign the click to the set whose mean Y is closer
                if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                    active_marker_positions = set1_positions
                    active_marker_values = set1_values
                    set_name = "Set 1 (Gel?)" # Tentative name
                else:
                    active_marker_positions = set2_positions
                    active_marker_values = set2_values
                    set_name = "Set 2 (WB?)" # Tentative name
            elif valid_set1: # Only set 1 is valid
                 active_marker_positions = set1_positions
                 active_marker_values = set1_values
                 set_name = "Set 1 (Gel?)"
            elif valid_set2: # Only set 2 is valid
                 active_marker_positions = set2_positions
                 active_marker_values = set2_values
                 set_name = "Set 2 (WB?)"
            else: # Neither set is valid after splitting
                 QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return
        else:
            # Only one set detected, use all markers
            if len(all_marker_positions) >= 2:
                active_marker_positions = all_marker_positions
                active_marker_values = all_marker_values
                set_name = "Single Set"
            else: # Should have been caught earlier, but double-check
                 QMessageBox.warning(self, "Error", "Not enough markers in the single set.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return

        # --- 4. Perform Regression on the Active Set ---
        # Normalize distances *within the active set*
        min_pos_active = np.min(active_marker_positions)
        max_pos_active = np.max(active_marker_positions)
        if max_pos_active == min_pos_active: # Avoid division by zero if all points are the same
            QMessageBox.warning(self, "Error", "All markers in the selected set have the same position.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if hasattr(self, "protein_location"): del self.protein_location
            return

        normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)

        # Log transform marker values
        try:
            log_marker_values_active = np.log10(active_marker_values)
        except Exception as e:
             QMessageBox.warning(self, "Error", f"Could not log-transform marker values (are they all positive?): {e}")
             self.live_view_label.setCursor(Qt.ArrowCursor)
             if hasattr(self, "protein_location"): del self.protein_location
             return

        # Perform linear regression (log-linear fit)
        coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, 1)

        # Calculate R-squared for the fit on the active set
        residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1 # Handle case of perfect fit or single point mean

        # --- 5. Predict MW for the Clicked Protein ---
        # Normalize the protein's Y position *using the active set's min/max*
        normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)

        # Predict using the derived coefficients
        predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
        predicted_weight = 10 ** predicted_log10_weight

        # --- 6. Update View and Plot ---
        self.update_live_view() # Draw the '*' marker at self.protein_location

        # Plot the results, passing both active and all data for context
        self.plot_molecular_weight_graph(
            # Active set data for fitting and line plot:
            normalized_distances_active,
            active_marker_values,
            10 ** np.polyval(coefficients, normalized_distances_active), # Fitted values for active set
            # Full set data for context (plotting all points):
            all_marker_positions, # Pass original positions
            all_marker_values,
            min_pos_active,       # Pass min/max of *active* set for normalization context
            max_pos_active,
            # Prediction results:
            normalized_protein_position, # Position relative to active set scale
            predicted_weight,
            r_squared,
            set_name # Name of the set used
        )

        # Reset mouse event handler after prediction
        self.live_view_label.mousePressEvent = None
        self.live_view_label.setCursor(Qt.ArrowCursor)
        self.run_predict_MW = True # Indicate prediction was attempted/completed

        
        
    def plot_molecular_weight_graph(
        self,
        # Data for the active set (used for the fit line and highlighted points)
        active_norm_distances,  # Normalized distances for the active set points
        active_marker_values,   # Original MW values of the active set points
        active_fitted_values,   # Fitted MW values corresponding to active_norm_distances

        # Data for *all* markers (for plotting all points for context)
        all_marker_positions,   # *Original* Y positions of all markers
        all_marker_values,      # *Original* MW values of all markers
        active_min_pos,         # Min Y position of the *active* set (for normalizing all points)
        active_max_pos,         # Max Y position of the *active* set (for normalizing all points)

        # Prediction results
        predicted_norm_position,# Predicted protein position normalized relative to active set scale
        predicted_weight,       # Predicted MW value

        # Fit quality and set info
        r_squared,              # R-squared of the fit on the active set
        set_name                # Name of the set used (e.g., "Set 1", "Set 2", "Single Set")
    ):
        """
        Plots the molecular weight prediction graph and displays it in a custom dialog.
        """
        # --- Normalize *all* marker positions using the *active* set's scale ---
        if active_max_pos == active_min_pos: # Avoid division by zero
             all_norm_distances_for_plot = np.zeros_like(all_marker_positions)
        else:
            all_norm_distances_for_plot = (all_marker_positions - active_min_pos) / (active_max_pos - active_min_pos)

        # --- Create Plot ---
        # Adjust figsize and DPI for a more compact plot suitable for a dialog
        fig, ax = plt.subplots(figsize=(4.5, 3.5)) # Smaller figure size

        # 1. Plot *all* marker points lightly for context
        ax.scatter(all_norm_distances_for_plot, all_marker_values, color="grey", alpha=0.5, label="All Markers", s=25) # Slightly smaller

        # 2. Plot the *active* marker points prominently
        ax.scatter(active_norm_distances, active_marker_values, color="red", label=f"Active Set ({set_name})", s=40, marker='o') # Slightly smaller

        # 3. Plot the fitted line for the *active* set
        sort_indices = np.argsort(active_norm_distances)
        # Move R^2 out of the legend
        ax.plot(active_norm_distances[sort_indices], active_fitted_values[sort_indices],
                 color="blue", label="Fit Line", linewidth=1.5)

        # 4. Plot the predicted protein position
        # Move predicted value out of the legend
        ax.axvline(predicted_norm_position, color="green", linestyle="--",
                    label="Target Protein", linewidth=1.5)

        # --- Configure Plot ---
        ax.set_xlabel(f"Normalized Distance (Relative to {set_name})", fontsize=9)
        ax.set_ylabel("Molecular Weight (units)", fontsize=9)
        ax.set_yscale("log")
        # Smaller legend font size
        ax.legend(fontsize='x-small', loc='best') # Use 'best' location
        ax.set_title(f"Molecular Weight Prediction", fontsize=10) # Simpler title
        ax.grid(True, which='both', linestyle=':', linewidth=0.5)
        ax.tick_params(axis='both', which='major', labelsize=8) # Smaller tick labels

        plt.tight_layout(pad=0.5) # Adjust layout to prevent labels overlapping

        # --- Save Plot to Buffer ---
        pixmap = None
        try:
            buffer = BytesIO()
            # Use a moderate DPI suitable for screen display in a dialog
            plt.savefig(buffer, format="png", bbox_inches="tight", dpi=100)
            buffer.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buffer.read())
            buffer.close()
        except Exception as plot_err:
            QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
            # pixmap remains None
        finally:
             plt.close(fig) # Ensure figure is always closed using the fig object

        # --- Display Results in a Custom Dialog ---
        if pixmap:
            # Create a custom dialog instead of modifying QMessageBox layout
            dialog = QDialog(self)
            dialog.setWindowTitle("Prediction Result")

            # Layout for the dialog
            layout = QVBoxLayout(dialog)

            # Text label for results
            results_text = (
                f"<b>Prediction using {set_name}:</b><br>"
                f"Predicted MW: <b>{predicted_weight:.2f}</b> units<br>"
                f"Fit R²: {r_squared:.3f}"
            )
            info_label = QLabel(results_text)
            info_label.setTextFormat(Qt.RichText) # Allow basic HTML like <b>
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            # Label to display the plot
            plot_label = QLabel()
            plot_label.setPixmap(pixmap)
            plot_label.setAlignment(Qt.AlignCenter) # Center the plot
            layout.addWidget(plot_label)

            # OK button
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)

            dialog.setLayout(layout)
            dialog.exec_() # Show the custom dialog modally
        else:
             # If plot generation failed, show a simpler message box
             QMessageBox.information(self, "Prediction Result (No Plot)",
                f"Used {set_name} for prediction.\n"
                f"Predicted MW: {predicted_weight:.2f} units\n"
                f"Fit R²: {r_squared:.3f}"
             )

  
        
    def reset_image(self):
        # Reset the image to original master (which keeps original format)
        if hasattr(self, 'image_master') and self.image_master and not self.image_master.isNull():
            self.image = self.image_master.copy()
            self.image_before_padding = None # Padding is removed
            self.image_contrasted = self.image.copy() # Reset contrast state
            self.image_before_contrast = self.image.copy()
            self.image_padded = False # Reset padding flag
        else:
            # No master image loaded, clear everything
            self.image = None
            self.original_image = None
            self.image_master = None
            self.image_before_padding = None
            self.image_contrasted = None
            self.image_before_contrast = None
            self.image_padded = False

        # (Rest of reset logic for markers, UI elements, etc. remains the same)
        self.warped_image=None
        self.left_markers.clear()
        self.right_markers.clear()
        self.top_markers.clear()
        self.custom_markers.clear()
        self.clear_predict_molecular_weight() # Clears analysis state too

        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)
        self.orientation_slider.setValue(0)
        self.taper_skew_slider.setValue(0)
        self.high_slider.setValue(100)
        self.low_slider.setValue(100)
        self.gamma_slider.setValue(100)
        self.left_padding_slider.setValue(0)
        self.right_padding_slider.setValue(0)
        self.top_padding_slider.setValue(0)
        self.left_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.bottom_padding_input.setText("0")

        self.marker_mode = None
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.left_marker_shift_added = 0
        self.right_marker_shift_added = 0
        self.top_marker_shift_added = 0
        self.live_view_label.mode = None
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.setCursor(Qt.ArrowCursor)

        try:
            self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
            self.on_combobox_changed()
        except Exception as e:
            pass

        if self.image and not self.image.isNull():
            try:
                self._update_preview_label_size()
                self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                self.top_padding_input.setText(str(int(self.image.height()*0.15)))
            except Exception as e:
                pass
        else:
            self.live_view_label.clear()
            self._update_preview_label_size()
        self.live_view_label.zoom_level = 1.0
        self.live_view_label.pan_offset = QPointF(0, 0)
        if hasattr(self, 'pan_left_action'): self.pan_left_action.setEnabled(False)
        if hasattr(self, 'pan_right_action'): self.pan_right_action.setEnabled(False)
        if hasattr(self, 'pan_up_action'): self.pan_up_action.setEnabled(False)
        if hasattr(self, 'pan_down_action'): self.pan_down_action.setEnabled(False)
        self.update_live_view()
        self._update_status_bar() # <--- Add this


if __name__ == "__main__":
    try:
        app = QApplication([])
        app.setStyle("Fusion")
        
        app.setStyleSheet("""
        QSlider::handle:horizontal {
            width: 100px;
            height: 100px;
            margin: -5px 0;
            background: #FFFFFF;
            border: 2px solid #555;
            border-radius: 30px;
        }
        QStatusBar QLabel {
            margin-left: 2px;
            margin-right: 5px;
            padding: 0px 0px;
            border: none;
        }
    """)
        window = CombinedSDSApp()
        window.show()
        app.aboutToQuit.connect(window.cleanup_temp_clipboard_file)
        app.exec_()
    except Exception as e:
        # This top-level except might catch errors *before* the app runs,
        # so log_exception might not be called. Log directly here as a fallback.
        print(f"FATAL: Application failed to start: {e}")
        try:
            # Try logging one last time if basicConfig worked at all
            logging.critical("Application failed to start", exc_info=True)
        except:
            pass # Ignore errors during this final logging attemp